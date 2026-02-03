#!/usr/bin/env python3
"""
Convert the official CCM Excel file to a CISO Assistant Excel workbook that follows
the v2 format (library_meta, framework_meta/_content, implementation_groups, answers).
"""

import argparse
import re
import sys
from typing import Dict, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

LIBRARY_NAME = "CCM Controls v4"
LIBRARY_SLUG = "ccm-controls-v4"
LIBRARY_REF_ID = "CCM-Controls-v4"
LIBRARY_DESCRIPTION = "CCM Controls v4"
LIBRARY_VERSION = "2"
DEFAULT_OUTPUT = "ccm-controls-v4-v2.xlsx"

FRAMEWORK_BASE_NAME = "framework"
IMPLEMENTATION_GROUPS_BASE_NAME = "implementation_groups"
ANSWERS_BASE_NAME = "answers"


def slugify_packager(packager_name: str) -> str:
    """Lowercase the packager name and keep only safe characters for URNs."""
    slug = re.sub(r"[^a-z0-9]+", "-", packager_name.lower()).strip("-")
    return slug or "packager"


def prettify_content(content) -> str:
    """Format multi-line cell values similarly to the original helper."""
    if content is None:
        return ""
    res = None
    stop_join = False
    for raw_line in str(content).splitlines():
        line = raw_line.rstrip()
        if not line:
            continue
        res = (
            f"{res}\n{line}"
            if stop_join and res
            else (line if res is None else f"{res} {line}")
        )
        if line.endswith(":"):
            stop_join = True
    return res or ""


def parse_ccm_sheet(sheet) -> Tuple[List[dict], str, Dict[str, dict]]:
    """Parse the CCM worksheet and return framework rows, copyright, and control lookup."""
    framework_rows: List[dict] = []
    control_lookup: Dict[str, dict] = {}
    library_copyright = ""
    eos = False

    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if idx < 4:
            continue
        domain, title, control_id, specification, lite = row[:5]

        if eos:
            if domain:
                library_copyright = str(domain).strip()
            continue

        if lite:
            ref_id = str(control_id).strip()
            impl_groups = "lite,full" if str(lite).strip().lower() == "yes" else "full"
            entry = {
                "assessable": "x",
                "depth": 2,
                "ref_id": ref_id,
                "name": str(title).strip() if title else ref_id,
                "description": prettify_content(specification),
                "implementation_groups": impl_groups,
                "questions": "",
                "answer": "",
            }
            framework_rows.append(entry)
            control_lookup[ref_id] = entry
        else:
            domain_value = str(domain).strip() if domain else ""
            if "End of Standard" in domain_value:
                eos = True
                continue
            if " - " not in domain_value:
                continue
            name, ref = domain_value.split(" - ", 1)
            entry = {
                "assessable": "",
                "depth": 1,
                "ref_id": ref.strip(),
                "name": name.strip(),
                "description": "",
                "implementation_groups": "",
                "questions": "",
                "answer": "",
            }
            framework_rows.append(entry)

    return framework_rows, library_copyright, control_lookup


def attach_caiq_questions(sheet, control_lookup: Dict[str, dict]) -> None:
    """Attach CAIQ questions to the appropriate control rows."""
    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if idx < 4:
            continue
        question_id, question = row[4:6]
        if not question_id:
            continue
        match = re.match(r"(.*)\.\d+$", str(question_id).strip())
        if not match:
            continue
        control_id = match.group(1)
        control = control_lookup.get(control_id)
        if not control:
            continue
        question_text = prettify_content(question)
        if not question_text:
            continue
        existing = control.get("questions", "")
        control["questions"] = (
            question_text if not existing else f"{existing}\n{question_text}"
        )
        control["answer"] = "A1"


def build_v2_workbook(
    packager: str,
    framework_rows: List[dict],
    library_copyright: str,
    output_filename: str,
) -> None:
    packager_slug = slugify_packager(packager)
    library_urn = f"urn:{packager_slug}:risk:library:{LIBRARY_SLUG}"
    framework_urn = f"urn:{packager_slug}:risk:framework:{LIBRARY_SLUG}"
    framework_base_urn = f"urn:{packager_slug}:risk:req_node:{LIBRARY_SLUG}"

    wb = Workbook()

    # library_meta
    ws_library = wb.active
    ws_library.title = "library_meta"
    ws_library.append(["type", "library"])
    ws_library.append(["urn", library_urn])
    ws_library.append(["version", LIBRARY_VERSION])
    ws_library.append(["locale", "en"])
    ws_library.append(["ref_id", LIBRARY_REF_ID])
    ws_library.append(["name", LIBRARY_NAME])
    ws_library.append(["description", LIBRARY_DESCRIPTION])
    ws_library.append(["copyright", library_copyright])
    ws_library.append(["provider", "CSA"])
    ws_library.append(["packager", packager])

    # framework_meta
    ws_framework_meta = wb.create_sheet(f"{FRAMEWORK_BASE_NAME}_meta")
    ws_framework_meta.append(["type", "framework"])
    ws_framework_meta.append(["base_urn", framework_base_urn])
    ws_framework_meta.append(["urn", framework_urn])
    ws_framework_meta.append(["ref_id", LIBRARY_REF_ID])
    ws_framework_meta.append(["name", LIBRARY_NAME])
    ws_framework_meta.append(["description", LIBRARY_DESCRIPTION])
    ws_framework_meta.append(
        ["implementation_groups_definition", IMPLEMENTATION_GROUPS_BASE_NAME]
    )
    ws_framework_meta.append(["answers_definition", ANSWERS_BASE_NAME])

    # framework_content
    ws_framework_content = wb.create_sheet(f"{FRAMEWORK_BASE_NAME}_content")
    framework_header = [
        "assessable",
        "depth",
        "ref_id",
        "name",
        "description",
        "implementation_groups",
        "questions",
        "answer",
    ]
    ws_framework_content.append(framework_header)
    for row in framework_rows:
        ws_framework_content.append(
            [
                row.get("assessable", ""),
                row.get("depth", ""),
                row.get("ref_id", ""),
                row.get("name", ""),
                row.get("description", ""),
                row.get("implementation_groups", ""),
                row.get("questions", ""),
                row.get("answer", ""),
            ]
        )

    # implementation_groups meta/content
    ws_ig_meta = wb.create_sheet(f"{IMPLEMENTATION_GROUPS_BASE_NAME}_meta")
    ws_ig_meta.append(["type", "implementation_groups"])
    ws_ig_meta.append(["name", IMPLEMENTATION_GROUPS_BASE_NAME])

    ws_ig_content = wb.create_sheet(f"{IMPLEMENTATION_GROUPS_BASE_NAME}_content")
    ws_ig_content.append(["ref_id", "name", "description"])
    ws_ig_content.append(
        [
            "lite",
            "Foundational",
            "Foundational controls that should be implemented by any organization, regardless of their budget, maturity and risk profile.",
        ]
    )
    ws_ig_content.append(
        [
            "full",
            "Systematic",
            "Systematic assessment of a cloud implementation.",
        ]
    )

    # answers meta/content
    ws_answers_meta = wb.create_sheet(f"{ANSWERS_BASE_NAME}_meta")
    ws_answers_meta.append(["type", "answers"])
    ws_answers_meta.append(["name", ANSWERS_BASE_NAME])

    ws_answers_content = wb.create_sheet(f"{ANSWERS_BASE_NAME}_content")
    ws_answers_content.append(["id", "question_type", "question_choices"])
    ws_answers_content.append(["A1", "unique_choice", "Yes\nNo\nNA"])

    wb.save(output_filename)
    print(f'✅ Excel file saved successfully: "{output_filename}"')


def main():
    parser = argparse.ArgumentParser(
        prog="convert_ccm_v2",
        description="Convert CCM official Excel file to a CISO Assistant v2 Excel file.",
    )
    parser.add_argument("filename", help="Path to the CCM Excel file")
    parser.add_argument("--packager", help="Name of the packager entity")
    parser.add_argument(
        "-o",
        "--output",
        default=DEFAULT_OUTPUT,
        help=f"Output Excel file name (default: {DEFAULT_OUTPUT})",
    )
    args = parser.parse_args()

    print(f'⌛ Parsing "{args.filename}"...')
    try:
        workbook = load_workbook(args.filename)
        print(f'✅ Excel file loaded successfully: "{args.filename}"')
    except FileNotFoundError:
        print(f'❌ [ERROR] File not found: "{args.filename}"')
        sys.exit(1)
    except PermissionError:
        print(f'❌ [ERROR] Permission denied while accessing "{args.filename}"')
        sys.exit(1)
    except InvalidFileException:
        print(f'❌ [ERROR] The file is not a valid Excel file: "{args.filename}"')
        sys.exit(1)
    except Exception as exc:  # noqa: BLE001
        print(f"❌ [ERROR] Unexpected error while loading Excel file: {exc}")
        sys.exit(1)

    framework_rows: List[dict] = []
    control_lookup: Dict[str, dict] = {}
    library_copyright = ""

    for sheet in workbook.worksheets:
        print(f'⌛ Parsing tab "{sheet.title}"...')
        if sheet.title == "CCM":
            rows, copyright_text, lookup = parse_ccm_sheet(sheet)
            framework_rows.extend(rows)
            control_lookup.update(lookup)
            library_copyright = copyright_text or library_copyright
        elif sheet.title == "CAIQ":
            attach_caiq_questions(sheet, control_lookup)
        else:
            print(f'⏩ Ignored tab: "{sheet.title}"')

    if not framework_rows:
        print("❌ [ERROR] No controls were extracted from the workbook.")
        sys.exit(1)

    if not library_copyright:
        library_copyright = "See CCM documentation for copyright details."

    try:
        build_v2_workbook(args.packager, framework_rows, library_copyright, args.output)
    except PermissionError:
        print(
            f'❌ [ERROR] Permission denied. The file may be open or locked: "{args.output}"'
        )
        sys.exit(1)
    except OSError as exc:  # noqa: BLE001
        print(f'❌ [ERROR] OS error while saving the file: "{exc}"')
        sys.exit(1)
    except Exception as exc:  # noqa: BLE001
        print(f'❌ [ERROR] Unexpected error while saving Excel file: "{exc}"')
        sys.exit(1)


if __name__ == "__main__":
    main()
