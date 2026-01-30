"""
simple script to transform the official CCM Excel file to another Excel file for CISO assistant framework conversion tool
"""

import openpyxl
import re
import sys
import argparse
import io  # Added for MIME optimization
import zipfile  # Added for MIME optimization
from openpyxl.utils.exceptions import InvalidFileException


def pretify_content(content):
    res = None
    stop_join = False
    for line in content.splitlines():
        if stop_join:
            res = res + "\n" + line if res else line
        else:
            res = res + " " + line if res else line
        if line[-1] == ":":
            stop_join = True
    return res


parser = argparse.ArgumentParser(
    prog="convert_ccm",
    description="convert CCM controls offical Excel file to CISO Assistant Excel file",
)
parser.add_argument("filename", help="name of CCM controls Excel file")
parser.add_argument("packager", help="name of packager entity")

args = parser.parse_args()
input_file_name = args.filename
packager = args.packager
output_file_name = "ccm-controls-v4.xlsx"

print(f'⌛ Parsing "{input_file_name}"...')

# Define variable to load the dataframe
try:
    dataframe = openpyxl.load_workbook(input_file_name)
    print(f'✅ Excel file loaded successfully: "{input_file_name}"')

except FileNotFoundError:
    print(f'❌ [ERROR] File not found: "{input_file_name}"')
    sys.exit(1)
except PermissionError:
    print(f'❌ [ERROR] Permission denied while accessing "{input_file_name}"')
    sys.exit(1)
except InvalidFileException:
    print(f'❌ [ERROR] The file is not a valid Excel file: "{input_file_name}"')
    sys.exit(1)
except Exception as e:
    print(f"❌ [ERROR] Unexpected error while loading Excel file: {e}")
    sys.exit(1)


output_table = []
library_copyright = ""

for tab in dataframe:
    print("⌛ Parsing tab", tab.title)
    title = tab.title
    if title == "CCM":
        line = 0
        eos = False
        for row in tab:
            line += 1
            if line < 4:
                continue
            (domain, title, id, specification, lite) = (r.value for r in row[0:5])
            if eos:
                library_copyright = domain  # last line after end of standard
            elif lite:
                output_table.append(
                    (
                        "x",
                        2,
                        id,
                        title,
                        pretify_content(specification),
                        "lite,full" if lite == "Yes" else "full",
                        "",
                    )
                )
            else:
                if domain and "End of Standard" in domain:
                    eos = True
                elif domain:
                    parts = domain.split(" - ")
                    if len(parts) == 2:
                        (d, id) = parts
                        output_table.append(("", 1, id, d, "", None, None))
    elif title == "CAIQ":
        line = 0
        eos = False
        for row in tab:
            line += 1
            if line < 4:
                continue
            (question_id, question) = (r.value for r in row[4:6])
            if question_id:
                q = re.match(r"(.*)\.\d+$", question_id)
                if q:
                    id = q.group(1)
                    for i in range(len(output_table)):
                        if output_table[i][2] == id:
                            a = output_table[i][6]
                            b = pretify_content(question)
                            c = b if a == "" else a + "\n" + b
                            output_table[i] = output_table[i][0:6] + (c,) + ("A1",)
    else:
        print(f"⏩ Ignored tab: {title}")


print(f'⌛ Generating "{output_file_name}"...')
wb_output = openpyxl.Workbook()
ws = wb_output.active
ws.title = "library_content"
ws.append(["library_urn", f"urn:{packager.lower()}:risk:library:ccm-controls-v4"])
ws.append(["library_version", "2"])
ws.append(["library_locale", "en"])
ws.append(["library_ref_id", "CCM-Controls-v4"])
ws.append(["library_name", "CCM Controls v4"])
ws.append(["library_description", "CCM Controls v4"])
ws.append(["library_copyright", library_copyright])
ws.append(["library_provider", "CSA"])
ws.append(["library_packager", packager])
ws.append(["framework_urn", f"urn:{packager.lower()}:risk:framework:ccm-controls-v4"])
ws.append(["framework_ref_id", "CCM-Controls-v4"])
ws.append(["framework_name", "CCM Controls v4"])
ws.append(["framework_description", "CCM Controls v4"])
ws.append(["tab", "controls", "requirements"])
ws.append(["tab", "implementation_groups", "implementation_groups"])
ws.append(["tab", "answers", "answers"])

ws1 = wb_output.create_sheet("controls")
ws1.append(
    [
        "assessable",
        "depth",
        "ref_id",
        "name",
        "description",
        "implementation_groups",
        "questions",
        "answer",
    ]
)
for row in output_table:
    ws1.append(row)
ws2 = wb_output.create_sheet("implementation_groups")
ws2.append(["ref_id", "name", "description"])
ws2.append(
    [
        "lite",
        "foundational",
        "foundational controls that should be implemented by any organization, regardless of their budget, maturity and risk profile",
    ]
)
ws2.append(["full", "systematic ", "systematic assessment of a cloud implementation"])
ws3 = wb_output.create_sheet("answers")
ws3.append(["id", "question_type", "question_choices"])
ws3.append(
    [
        "A1",
        "unique_choice",
        "Yes\nNo\nNA",
    ]
)

try:
    # --- MIME Type Optimization Block ---
    # We save to memory first, then repack into the final file to ensure
    # [Content_Types].xml is the FIRST file in the zip archive.

    print("⌛ Optimizing file structure for MIME detection...")

    buffer = io.BytesIO()
    wb_output.save(buffer)
    buffer.seek(0)

    with zipfile.ZipFile(buffer, "r") as z_in:
        with zipfile.ZipFile(
            output_file_name, "w", compression=zipfile.ZIP_DEFLATED
        ) as z_out:
            # Files required by libmagic/file-command to identify 'application/vnd.openxml...'
            # reliably within the first 2KB of the file.
            priority_files = [
                "[Content_Types].xml",
                "_rels/.rels",
                "docProps/app.xml",
                "docProps/core.xml",
            ]

            # 1. Write priority files first
            for filename in priority_files:
                if filename in z_in.namelist():
                    z_out.writestr(filename, z_in.read(filename))

            # 2. Write the rest of the files (worksheets, styles, etc.)
            for filename in z_in.namelist():
                if filename not in priority_files:
                    z_out.writestr(filename, z_in.read(filename))

    print(f'✅ Excel file saved successfully: "{output_file_name}"')
    sys.exit(0)

except PermissionError:
    print(
        f'❌ [ERROR] Permission denied. The file may be open or locked: "{output_file_name}"'
    )
    sys.exit(1)
except FileNotFoundError:
    print(f'❌ [ERROR] Invalid path. Cannot save to: "{output_file_name}"')
    sys.exit(1)
except OSError as e:
    print(f'❌ [ERROR] OS error while saving the file: "{e}"')
    sys.exit(1)
except Exception as e:
    print(f'❌ [ERROR] Unexpected error while saving Excel file: "{e}"')
    sys.exit(1)
