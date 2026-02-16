#!/usr/bin/env python3
"""
Convert the official CIS Controls Excel workbook into a CISO Assistant
framework workbook that follows the v2 Excel format.

The resulting file contains the following sheets:
- library_meta
- framework_meta / framework_content
- implementation_groups_meta / implementation_groups_content
- reference_controls_meta / reference_controls_content
- urn_prefix_meta / urn_prefix_content
"""

import argparse
import re
import sys
from typing import List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

LIBRARY_NAME = "CIS Controls v8"
LIBRARY_SLUG = "cis-controls-v8"
LIBRARY_REF_ID = "CIS-Controls-v8"
LIBRARY_DESCRIPTION = "CIS Controls v8"
DEFAULT_OUTPUT = "cis-controls-v8-v2.xlsx"

FRAMEWORK_SHEET_BASE = "framework"
IMPLEMENTATION_GROUPS_SHEET_BASE = "implementation_groups"
REFERENCE_CONTROLS_SHEET_BASE = "reference_controls"
URN_PREFIX_SHEET_BASE = "urn_prefix"

REFERENCE_CONTROL_PREFIX_ID = "cis"

CIS_FRAMEWORK_SHEET_NAMES = ("contrôles v8", "controls v8")
CIS_LANGUAGES = {
    "contrôles v8": "fr",
    "controls v8": "en"
}

CIS_COPYRIGHT_SHEET_NAMES = ("License for Use", "License dutilisation", "License d'utilisation")

IG_DESCRIPTIONS = {
    "IG1": {
        "en": "Minimum standard of information security for all enterprises.",
        "fr": "Standard minimum de sécurité de l'information pour toutes les entreprises.",
    },
    "IG2": {
        "en": "For enterprises managing IT infrastructure of multiple departments with differing risk profiles.",
        "fr": "Pour les entreprises gérant une infrastructure IT pour plusieurs départements avec des profils de risque différents.",
    },
    "IG3": {
        "en": "To secure sensitive and confidential data.",
        "fr": "Pour sécuriser les données sensibles et confidentielles.",
    },
}

CSF_FUNCTION_TRANSLATIONS = {
    "govern":  ["gouverner"],
    "identify":["identifier"],
    "protect": ["protéger"],
    "detect":  ["détecter"],
    "respond": ["répondre"],
    "recover": ["rétablir"],
}

# Default language
current_cis_language = "en"


def slugify_packager(packager_name: str) -> str:
    """Lowercase the packager name and keep only safe characters for URNs."""
    slug = re.sub(r"[^a-z0-9]+", "-", packager_name.lower()).strip("-")
    return slug or "packager"


def safe_strip(value, lower: bool = False) -> str:
    """Return a stripped string (optionally lower-cased) or an empty string."""
    if value is None:
        return ""
    text = str(value).strip()
    return text.lower() if lower else text


def normalize_csf_function(value: str, lang: str = "en") -> str:
    """
    Normalize the CSF function to its English keyword.
    - If lang == "en": returns the lowercased value as-is.
    - Otherwise: maps localized values back to English using CSF_FUNCTION_TRANSLATIONS.
    """
    raw = safe_strip(value, lower=True)
    if not raw:
        return ""

    # If already one of the official English keywords, keep it
    if raw in CSF_FUNCTION_TRANSLATIONS:
        return raw

    if lang == "en":
        return raw

    # Reverse-lookup: localized -> English
    for en_key, translations in CSF_FUNCTION_TRANSLATIONS.items():
        if raw == en_key:
            return en_key
        if raw in (t.lower() for t in translations):
            return en_key

    # Unknown value: keep raw (or return "" if you prefer strictness)
    return raw


def parse_controls_sheet(sheet, lang: str = "en") -> Tuple[List[dict], List[dict]]:
    """
    Extract framework rows and reference controls from the CIS Controls worksheet.
    """
    framework_rows: List[dict] = []
    reference_controls: List[dict] = []
    safeguard_index = 0

    for row in sheet.iter_rows(values_only=True):
        try:
            control, safeguard, _asset_type, sf, title, description, ig1, ig2, ig3 = (
                row[:9]
            )
        except ValueError:
            continue

        control_str = safe_strip(control)
        if not control_str or not re.match(r"\d+", control_str):
            continue

        if not safe_strip(safeguard):
            safeguard_index = 0
            framework_rows.append(
                {
                    "assessable": "",
                    "depth": 1,
                    "ref_id": control_str,
                    "name": safe_strip(title) or control_str,
                    "description": safe_strip(description),
                }
            )
            continue

        safeguard_index += 1
        safeguard_ref = f"{control_str}.{safeguard_index}"
        implementation_groups = "IG1,IG2,IG3" if ig1 else "IG2,IG3" if ig2 else "IG3"

        framework_rows.append(
            {
                "assessable": "x",
                "depth": 2,
                "ref_id": safeguard_ref,
                "name": safe_strip(title) or safeguard_ref,
                "description": safe_strip(description),
                "implementation_groups": implementation_groups,
                "reference_controls": f"{REFERENCE_CONTROL_PREFIX_ID}:{safeguard_ref}",
            }
        )
        reference_controls.append(
            {
                "ref_id": safeguard_ref,
                "name": safe_strip(title) or safeguard_ref,
                "csf_function": normalize_csf_function(sf, lang=lang),
                "description": safe_strip(description),
            }
        )

    return framework_rows, reference_controls


def extract_library_copyright(sheet) -> str:
    """Build the copyright string from the license sheet."""
    lines = [safe_strip(sheet["B11"].value), safe_strip(sheet["B13"].value)]
    combined = "\n".join([line for line in lines if line])
    return combined or "See CIS Controls official license for usage details."


def build_v2_workbook(
    packager: str,
    framework_rows: List[dict],
    reference_controls: List[dict],
    library_copyright: str,
    output_filename: str,
    cis_language: str = "en",
) -> None:
    
    lang = cis_language if cis_language in CIS_LANGUAGES.values() else "en"

    packager_slug = slugify_packager(packager)
    library_urn = f"urn:{packager_slug}:risk:library:{LIBRARY_SLUG}"
    framework_urn = f"urn:{packager_slug}:risk:framework:{LIBRARY_SLUG}"
    framework_base_urn = f"urn:{packager_slug}:risk:req_node:{LIBRARY_SLUG}"
    reference_control_base_urn = f"urn:{packager_slug}:risk:function:{LIBRARY_SLUG}"

    wb = Workbook()

    # library_meta sheet
    ws_library = wb.active
    ws_library.title = "library_meta"
    ws_library.append(["type", "library"])
    ws_library.append(["urn", library_urn])
    ws_library.append(["version", "1"])
    ws_library.append(["locale", cis_language])
    ws_library.append(["ref_id", LIBRARY_REF_ID])
    ws_library.append(["name", LIBRARY_NAME])
    ws_library.append(["description", LIBRARY_DESCRIPTION])
    ws_library.append(["copyright", library_copyright])
    ws_library.append(["provider", "CIS"])
    ws_library.append(["packager", packager])

    # framework_meta sheet
    ws_framework_meta = wb.create_sheet(f"{FRAMEWORK_SHEET_BASE}_meta")
    ws_framework_meta.append(["type", "framework"])
    ws_framework_meta.append(["base_urn", framework_base_urn])
    ws_framework_meta.append(["urn", framework_urn])
    ws_framework_meta.append(["ref_id", LIBRARY_REF_ID])
    ws_framework_meta.append(["name", LIBRARY_NAME])
    ws_framework_meta.append(["description", LIBRARY_DESCRIPTION])
    ws_framework_meta.append(
        ["implementation_groups_definition", IMPLEMENTATION_GROUPS_SHEET_BASE]
    )

    # framework_content sheet
    ws_framework_content = wb.create_sheet(f"{FRAMEWORK_SHEET_BASE}_content")
    framework_header = [
        "assessable",
        "depth",
        "ref_id",
        "name",
        "description",
        "implementation_groups",
        "reference_controls",
    ]
    ws_framework_content.append(framework_header)
    for row in framework_rows:
        ws_framework_content.append(
            [
                row.get("assessable", ""),
                row.get("depth", ""),
                row.get("ref_id", ""),
                row.get("name", ""),
                row.get("description", ""),
                row.get("implementation_groups", ""),
                row.get("reference_controls", ""),
            ]
        )

    # implementation_groups sheets
    ws_impl_meta = wb.create_sheet(f"{IMPLEMENTATION_GROUPS_SHEET_BASE}_meta")
    ws_impl_meta.append(["type", "implementation_groups"])
    ws_impl_meta.append(["name", IMPLEMENTATION_GROUPS_SHEET_BASE])

    ws_impl_content = wb.create_sheet(f"{IMPLEMENTATION_GROUPS_SHEET_BASE}_content")
    ws_impl_content.append(["ref_id", "name", "description"])
    ws_impl_content.append(["IG1", "IG1", IG_DESCRIPTIONS["IG1"][lang]])
    ws_impl_content.append(["IG2", "IG2", IG_DESCRIPTIONS["IG2"][lang]])
    ws_impl_content.append(["IG3", "IG3", IG_DESCRIPTIONS["IG3"][lang]])

    # reference_controls sheets
    ws_ref_meta = wb.create_sheet(f"{REFERENCE_CONTROLS_SHEET_BASE}_meta")
    ws_ref_meta.append(["type", "reference_controls"])
    ws_ref_meta.append(["base_urn", reference_control_base_urn])

    ws_ref_content = wb.create_sheet(f"{REFERENCE_CONTROLS_SHEET_BASE}_content")
    ws_ref_content.append(["ref_id", "name", "csf_function", "description"])
    for ctrl in reference_controls:
        ws_ref_content.append(
            [
                ctrl.get("ref_id", ""),
                ctrl.get("name", ""),
                ctrl.get("csf_function", ""),
                ctrl.get("description", ""),
            ]
        )

    # urn_prefix sheets
    ws_prefix_meta = wb.create_sheet(f"{URN_PREFIX_SHEET_BASE}_meta")
    ws_prefix_meta.append(["type", "urn_prefix"])

    ws_prefix_content = wb.create_sheet(f"{URN_PREFIX_SHEET_BASE}_content")
    ws_prefix_content.append(["prefix_id", "prefix_value"])
    ws_prefix_content.append([REFERENCE_CONTROL_PREFIX_ID, reference_control_base_urn])

    wb.save(output_filename)
    print(f'✅ Excel file saved successfully: "{output_filename}"')


def main():
    parser = argparse.ArgumentParser(
        prog="prep_cis_v2",
        description="Convert CIS Controls official Excel file to a CISO Assistant v2 Excel file.",
    )
    parser.add_argument("filename", help="Path to the CIS Controls Excel file")
    parser.add_argument("-p", "--packager", required=True, help="Name of the packager entity")
    parser.add_argument(
        "-o",
        "--output",
        default=DEFAULT_OUTPUT,
        help=f"Output Excel file name (default: {DEFAULT_OUTPUT})",
    )
    args = parser.parse_args()

    print(f'⌛ Parsing "{args.filename}"...')
    try:
        workbook = load_workbook(args.filename)
        print(f'✅ Excel file loaded successfully: "{args.filename}"')
    except FileNotFoundError:
        print(f'❌ [ERROR] File not found: "{args.filename}"')
        sys.exit(1)
    except PermissionError:
        print(f'❌ [ERROR] Permission denied while accessing "{args.filename}"')
        sys.exit(1)
    except InvalidFileException:
        print(f'❌ [ERROR] The file is not a valid Excel file: "{args.filename}"')
        sys.exit(1)
    except Exception as exc:  # noqa: BLE001
        print(f"❌ [ERROR] Unexpected error while loading Excel file: {exc}")
        sys.exit(1)

    framework_rows: List[dict] = []
    reference_controls: List[dict] = []
    library_copyright = ""

    for sheet in workbook.worksheets:

        title = sheet.title
        print(f'⌛ Parsing tab "{title}"...')

        if title in CIS_COPYRIGHT_SHEET_NAMES:
            library_copyright = extract_library_copyright(sheet)

        else:
            # In order to retrieve the matched element in order to set  right language for IGs
            matched_prefix = next(
                (prefix for prefix in CIS_FRAMEWORK_SHEET_NAMES if title.strip().lower().startswith(prefix)),
                None,
            )

            if matched_prefix:
                current_cis_language = CIS_LANGUAGES.get(matched_prefix, "en")
                fw_rows, ref_ctrls = parse_controls_sheet(sheet, current_cis_language)
                framework_rows.extend(fw_rows)
                reference_controls.extend(ref_ctrls)
            else:
                print(f'⏩ Ignored tab: "{title}"')

    if not framework_rows:
        print("❌ [ERROR] No controls were extracted from the workbook.")
        sys.exit(1)

    if not library_copyright:
        print(
            "⚠️  [WARNING] Could not extract copyright notice from the workbook. "
            "A generic message will be used."
        )
        library_copyright = "See CIS Controls official license for usage details."

    try:
        build_v2_workbook(
            args.packager,
            framework_rows,
            reference_controls,
            library_copyright,
            args.output,
            current_cis_language
        )
    except PermissionError:
        print(
            f'❌ [ERROR] Permission denied. The file may be open or locked: "{args.output}"'
        )
        sys.exit(1)
    except OSError as exc:  # noqa: BLE001
        print(f'❌ [ERROR] OS error while saving the file: "{exc}"')
        sys.exit(1)
    except Exception as exc:  # noqa: BLE001
        print(f'❌ [ERROR] Unexpected error while saving Excel file: "{exc}"')
        sys.exit(1)


if __name__ == "__main__":
    main()
