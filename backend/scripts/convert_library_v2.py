"""
Convert Library v2 - Build a CISO Assistant YAML library from a v2 Excel file

This script processes an Excel file in v2 format (with *_meta and *_content tabs),
extracts all declared objects, and generates a fully structured YAML library.

Usage:
    python convert_library_v2.py path/to/library.xlsx [--compat MODE] [--output out.yaml] [--verbose]
    python convert_library_v2.py path/to/folder --bulk [--compat MODE] [--output-dir out_folder] [--verbose]

Arguments:
    --compat       Specify compatibility mode number.
                   Refer to the COMPATIBILITY_MODES dictionary in the script for available modes and their descriptions.
    --verbose      Enable verbose output. Verbose messages start with a 💬 (speech bubble) emoji.
    --output       Custom output file name (only for single file mode). Adds '.yaml' if missing.
    --bulk         Enable bulk mode to process all .xlsx files in a directory.
    --output-dir   Destination directory for YAML files (only valid with --bulk).

Note: the "node_id" column can be defined to force an urn suffix. This can be useful to fix "ref_id" errors
"""

import sys
import re
import yaml
import datetime
import argparse
import unicodedata
import openpyxl
from typing import Any, Dict, List
from pathlib import Path
from collections import Counter

SCRIPT_VERSION = "2.1"

# Maximum length for 'name' fields as enforced by the CISO Assistant database.
NAME_MAX_LENGTH = 200

# --- Compatibility modes definition ------------------------------------------
# NOTE: No compatibility mode includes another (unless otherwise stated)
# NOTE: So far, no compatibility mode has an impact on the mapping creation process.

COMPATIBILITY_MODES = {
    0: f"[v{SCRIPT_VERSION}] (DEFAULT) Don't use any Compatibility Mode",
    1: "[< v2] Use legacy URN fallback logic (for requirements without ref_id)",
    2: "[v2] Don't clean the URNs before saving it into the YAML file (Only spaces ' ' are replaced with hyphen '-' and the URN is lower-cased)",
    3: '[< v2++] Updated version of "[< v2]". Handling of the new "fix_count" column in order to ADD or SUBTRACT from the counter (replace "skip_count"). Fixed the URN writing issue when "skip_count" was true and a "ref_id" was defined.',
    # Future modes can be added here with an integer key and description
}

# --- Logging helpers ----------------------------------------------------------


def print_error(message: str) -> None:
    """Ensure errors are emitted on stderr and consistently prefixed."""
    print(f"❌ [ERROR] {message}", file=sys.stderr)


# --- Translation helpers ------------------------------------------------------


def extract_translations_from_row(header, row):
    translations = {}
    for i, col_name in enumerate(header):
        match = re.match(r"(\w+)\[(\w+)\]", col_name)
        if match and i < len(row):
            base_key, lang = match.groups()
            value = row[i].value
            if value:
                translations.setdefault(lang, {})[base_key] = str(value).strip()
    return translations


def extract_translations_from_metadata(meta_dict, prefix):
    translations = {}
    pattern = re.compile(r"(\w+)\[(\w+)\]")
    for key, value in meta_dict.items():
        match = pattern.match(key)
        if match and value:
            base_key, lang = match.groups()
            translations.setdefault(lang, {})[base_key] = str(value).strip()
    return translations


# --- Sheet parsing ------------------------------------------------------------


def parse_key_value_sheet(sheet):
    result = {}
    for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
        key = str(row[0]).strip().lower() if row[0] else None
        val = str(row[1]).strip() if row[1] is not None else None
        if key:
            result[key] = val
    return result


def parse_content_rows(content_ws):
    """Parse a content worksheet into a header list and list of (raw_row, data_dict) pairs.

    Returns:
        tuple: (header, rows_with_data) where:
            - header: list[str] of normalized column names
            - rows_with_data: list[tuple[Row, dict]] of (original_row, data_dict) pairs,
              skipping empty rows. The original row is preserved for cases that need
              cell-level access (translations, colors, row numbers).
        Returns ([], []) if the worksheet has no rows.
    """
    rows = list(content_ws.iter_rows())
    if not rows:
        return [], []
    header = [str(cell.value).strip().lower() if cell.value else "" for cell in rows[0]]
    rows_with_data = []
    for row in rows[1:]:
        if not any(cell.value for cell in row):
            continue
        data = {header[i]: row[i].value for i in range(len(header)) if i < len(row)}
        rows_with_data.append((row, data))
    return header, rows_with_data


# --- URN cleaning utility -----------------------------------------------------
def clean_urn_suffix(value: str, compat_mode: int = 0):
    """
    Cleans a URN suffix by:
        - Removing accents using Unicode normalization,
        - Converting to lowercase and replacing spaces with hyphens,
        - Replacing any disallowed characters with an underscore.

    Compatibility modes :
        - 0 : Default behavior
        - 1,2 : Replace only spaces " " with hyphens "-" sets text in lower case.
    """
    # [+] Compat check
    if 1 <= compat_mode <= 2:
        value = value.lower().replace(" ", "-")
    else:  # Default behavior
        # Normalize to separate characters from their accents (e.g., é → e + ́)
        value = (
            unicodedata.normalize("NFKD", value)
            .encode("ascii", "ignore")
            .decode("ascii")
        )
        value = value.lower().replace(" ", "-")
        # Keep only allowed characters for a URN suffix
        value = re.sub(r"[^a-z0-9_\-\.\[\]\(\):]", "_", value)

    return value


# --- Common helpers -----------------------------------------------------------


def set_optional_fields(entry: dict, data: dict, fields: list) -> None:
    """Copy non-empty string fields from data to entry."""
    for field in fields:
        val = data.get(field)
        if val:
            entry[field] = str(val).strip()


def extend_or_set(objects: dict, key: str, items: list) -> None:
    """Extend existing list in objects[key], or set it if absent."""
    if objects.get(key):
        objects[key].extend(items)
    else:
        objects[key] = items


def clean_ref_id_for_urn(
    ref_id_raw: str, compat_mode: int, verbose: bool, context: str
) -> tuple:
    """Clean a ref_id for URN usage, returning (display_ref_id, urn_ref_id).

    In compat modes 1-2, only lowercases. Otherwise, applies full URN cleaning.
    """
    ref_id_raw = ref_id_raw.strip()
    if 1 <= compat_mode <= 2:
        return ref_id_raw, ref_id_raw.lower()
    ref_id_clean = clean_urn_suffix(ref_id_raw, compat_mode=0)
    if verbose and ref_id_raw != ref_id_clean:
        print(
            f"💬 ⚠️  [WARNING] ({context}) Cleaned ref_id (for use in URN) '{ref_id_raw}' → '{ref_id_clean}'"
        )
    return ref_id_raw, ref_id_clean


def attach_translations_from_row(entry: dict, header: list, row) -> None:
    """Attach translations to entry if any exist in the row."""
    translations = extract_translations_from_row(header, row)
    if translations:
        entry["translations"] = translations


# --- urn expansion ------------------------------------------------------------


def expand_urns_from_prefixed_list(
    field_value: str, prefix_to_urn: str, compat_mode: int = 0, verbose: bool = False
):
    """
    Convert a prefixed list like 'abc:xyz, def:uvw' into a list of full URNs,
    using a prefix mapping. Fully qualified URNs (starting with 'urn:') are left untouched.

    Args:
        field_value (str): Raw string field value from the sheet (e.g., 'abc:id1, urn:...').
        prefix_to_urn (dict): Mapping from prefix (e.g., 'abc') to full base URN.
        verbose (bool): If True, prints a warning when a suffix is modified.

    Returns:
        list[str]: Fully qualified URNs
    """
    result = []
    elements = re.split(r"[\s,]+", str(field_value))
    for element in elements:
        element = element.strip()
        if not element:
            continue

        if element.startswith("urn:"):
            result.append(element)
            continue

        parts = element.split(":")
        if len(parts) >= 2:
            prefix = parts[0]
            raw_suffix = ":".join(parts[1:]).strip()
            # [+] Compat check
            cleaned_suffix = clean_urn_suffix(raw_suffix, compat_mode=compat_mode)
            base = prefix_to_urn.get(prefix)

            if base:
                # Show a warning if the suffix had to be cleaned/renamed
                if raw_suffix != cleaned_suffix and verbose:
                    print(
                        f"💬 ⚠️  [WARNING] (expand_urns_from_prefixed_list) URN suffix renamed for '{prefix}:{raw_suffix}' → '{base}:{cleaned_suffix}'"
                    )
                result.append(f"{base}:{cleaned_suffix}")
            else:
                print(
                    f"⚠️  [WARNING] (expand_urns_from_prefixed_list) Unknown prefix '{prefix}' for element '{element}' — skipped"
                )
        else:
            print(
                f"⚠️  [WARNING] (expand_urns_from_prefixed_list) Malformed element '{element}' — skipped"
            )
    return result


# --- question management ------------------------------------------------------------


def inject_questions_into_node(
    qa_data: dict[str, Any], node: Dict[str, Any], answers_dict: dict
) -> None:
    """
    Injects parsed questions and their metadata into a requirement node.
    Ensures that question type is valid and handles multiple questions/answers.

    :param qa_data: Allows to retrieve the string from the following columns: "questions", "answer" and "depends_on"
    :param node: The requirement node (dict)
    :param answers_dict: Dictionary of all available answers (from answer sheet)
    """

    raw_question_str = qa_data.get("questions")
    raw_answer_str = qa_data.get("answer")
    raw_depends_on_str = qa_data.get("depends_on")
    raw_condition_str = qa_data.get("condition")

    if not raw_question_str:
        return

    allowed_types = {"unique_choice", "multiple_choice", "text", "date"}

    question_lines = [q.strip() for q in str(raw_question_str).split("\n") if q.strip()]

    depends_on_lines = None
    if raw_depends_on_str:
        depends_on_lines = [
            dep.strip() for dep in str(raw_depends_on_str).split("\n") if dep.strip()
        ]
        depends_on_lines = [
            None if dep.lower() == "/" else dep for dep in depends_on_lines
        ]

        if not node.get("urn"):
            raise ValueError(
                f"(inject_questions_into_node) Missing Framework URNs to compute the 'depends_on' column"
            )

    condition_lines = None
    if raw_condition_str:
        condition_lines = [
            cond.strip() for cond in str(raw_condition_str).split("\n") if cond.strip()
        ]
        condition_lines = [
            None if cond.lower() == "/" else cond for cond in condition_lines
        ]
    # If NO condition [AND]
    elif not raw_condition_str and depends_on_lines:
        raise ValueError(
            f"(inject_questions_into_node) Missing value 'condition' to compute 'depends_on' column"
        )

    answer_ids = (
        [a.strip() for a in str(raw_answer_str).split("\n") if raw_answer_str]
        if raw_answer_str
        else []
    )

    if len(answer_ids) != 1 and len(answer_ids) != len(question_lines):
        raise ValueError(
            f"Mismatch between questions and 'answers' for node {node.get('urn')}"
        )

    if (
        depends_on_lines
        and len(depends_on_lines) != 1
        and len(depends_on_lines) != len(question_lines)
    ):
        raise ValueError(
            f"Mismatch between questions and 'depends_on' for node {node.get('urn')}"
        )

    if (
        condition_lines
        and len(condition_lines) != 1
        and len(condition_lines) != len(question_lines)
    ):
        raise ValueError(
            f"Mismatch between questions and 'condition' for node {node.get('urn')}"
        )

    question_block = {}

    for idx, question_text in enumerate(question_lines):
        answer_id = answer_ids[0] if len(answer_ids) == 1 else answer_ids[idx]
        answer_meta = answers_dict.get(answer_id)

        if not answer_meta:
            raise ValueError(
                f"Unknown answer ID: {answer_id} for node {node.get('urn')}"
            )
        qtype = answer_meta.get("type")
        if qtype not in allowed_types:
            raise ValueError(
                f"Invalid question type '{qtype}' for answer ID: {answer_id}"
            )
        q_urn = f"{node['urn']}:question:{idx + 1}"
        question_entry = {
            "type": qtype,
        }

        question_entry["text"] = question_text

        # Optional: depends_on
        depends_on_block = {}
        depends_on_question_urn = ""
        depends_on_question_answers = []

        if depends_on_lines:
            dependency = (
                depends_on_lines[0]
                if len(depends_on_lines) == 1
                else depends_on_lines[idx]
            )

            # Skip if dependency is not defined [AND]
            if dependency is not None:
                condition = None
                if len(condition_lines) == 1:
                    condition = condition_lines[0]
                else:
                    condition = condition_lines[idx]

                # Check string format (Format == "x:y,z,a,b,c,..." with x,y,a,b,c,... >= 1)
                if not re.fullmatch(r"[1-9]\d*:(?:[1-9]\d*)(?:,[1-9]\d*)*", dependency):
                    raise ValueError(
                        f"Invalid 'depends_on' format for question #{str(idx + 1)} for node {node.get('urn')}: '{dependency}'. Expected 'x:y[,z,...]' with x,y,z >= 1"
                    )

                dep_split = dependency.split(":")
                dependency_question = int(dep_split[0])
                dependency_question_answers = [int(c) for c in dep_split[1].split(",")]

                # If the question DOESN'T want to depend on itself
                if dependency_question != idx + 1:
                    ##### Recreate correct URN for Question & Answers (Choices) #####
                    # SOON: Checks for this will be processed in ./check_library_v2.py

                    # Get URN of the line we refer to
                    depends_on_urn = node.get("urn")
                    # URN of the question
                    depends_on_question_urn = (
                        depends_on_urn + f":question:{str(dependency_question)}"
                    )
                    # URN of choices
                    for choice in dependency_question_answers:
                        depends_on_question_answers.append(
                            depends_on_question_urn + f":choice:{str(choice)}"
                        )

                    if not condition:
                        raise ValueError(
                            f"Missing 'condition' for question #{str(idx + 1)} for node {node.get('urn')}. Must be 'any' or 'all' "
                        )

                    if condition not in ["any", "all"]:
                        raise ValueError(
                            f"Invalid 'condition' for question #{str(idx + 1)} for node {node.get('urn')}: '{condition}'. Must be 'any' or 'all', '/' (= 'undefined') or empty cell"
                        )

                    depends_on_block["condition"] = condition
                    depends_on_block["question"] = depends_on_question_urn
                    depends_on_block["answers"] = depends_on_question_answers

                    question_entry["depends_on"] = depends_on_block

        if qtype in {"unique_choice", "multiple_choice"}:
            choices = []
            for j, choice in enumerate(answer_meta["choices"]):
                # make a shallow copy so we don't mutate the original dict
                entry = choice.copy()
                # overwrite / add the per-question urn
                entry["urn"] = f"{q_urn}:choice:{j + 1}"
                choices.append(entry)
            question_entry["choices"] = choices

        question_block[q_urn] = question_entry
    if question_block:
        node["questions"] = question_block


# --- risk matrix management ------------------------------------------------------------

# https://gist.github.com/Mike-Honey/b36e651e9a7f1d2e1d60ce1c63b9b633
from colorsys import rgb_to_hls, hls_to_rgb

RGBMAX = 0xFF  # Corresponds to 255
HLSMAX = 240  # MS excel's tint function expects that HLS is base 240. see:
# https://social.msdn.microsoft.com/Forums/en-US/e9d8c136-6d62-4098-9b1b-dac786149f43/excel-color-tint-algorithm-incorrect?forum=os_binaryfile#d3c2ac95-52e0-476b-86f1-e2a697f24969


def rgb_to_ms_hls(red, green=None, blue=None):
    """Converts rgb values in range (0,1) or a hex string of the form '[#aa]rrggbb' to HLSMAX based HLS, (alpha values are ignored)"""
    if green is None:
        if isinstance(red, str):
            if len(red) > 6:
                red = red[-6:]  # Ignore preceding '#' and alpha values
            blue = int(red[4:], 16) / RGBMAX
            green = int(red[2:4], 16) / RGBMAX
            red = int(red[0:2], 16) / RGBMAX
        else:
            red, green, blue = red
    h, l, s = rgb_to_hls(red, green, blue)
    return (int(round(h * HLSMAX)), int(round(l * HLSMAX)), int(round(s * HLSMAX)))


def ms_hls_to_rgb(hue, lightness=None, saturation=None):
    """Converts HLSMAX based HLS values to rgb values in the range (0,1)"""
    if lightness is None:
        hue, lightness, saturation = hue
    return hls_to_rgb(hue / HLSMAX, lightness / HLSMAX, saturation / HLSMAX)


def rgb_to_hex(red, green=None, blue=None):
    """Converts (0,1) based RGB values to a hex string 'rrggbb'"""
    if green is None:
        red, green, blue = red
    return (
        "%02x%02x%02x"
        % (
            int(round(red * RGBMAX)),
            int(round(green * RGBMAX)),
            int(round(blue * RGBMAX)),
        )
    ).upper()


def get_theme_colors(wb):
    """Gets theme colors from the workbook"""
    # see: https://groups.google.com/forum/#!topic/openpyxl-users/I0k3TfqNLrc
    from openpyxl.xml.functions import QName, fromstring

    xlmns = "http://schemas.openxmlformats.org/drawingml/2006/main"
    root = fromstring(wb.loaded_theme)
    themeEl = root.find(QName(xlmns, "themeElements").text)
    colorSchemes = themeEl.findall(QName(xlmns, "clrScheme").text)
    firstColorScheme = colorSchemes[0]

    colors = []

    for c in [
        "lt1",
        "dk1",
        "lt2",
        "dk2",
        "accent1",
        "accent2",
        "accent3",
        "accent4",
        "accent5",
        "accent6",
    ]:
        accent = firstColorScheme.find(QName(xlmns, c).text)
        for i in list(accent):  # walk all child nodes, rather than assuming [0]
            if "window" in i.attrib["val"]:
                colors.append(i.attrib["lastClr"])
            else:
                colors.append(i.attrib["val"])

    return colors


def tint_luminance(tint, lum):
    """Tints a HLSMAX based luminance"""
    # See: http://ciintelligence.blogspot.co.uk/2012/02/converting-excel-theme-color-and-tint.html
    if tint < 0:
        return int(round(lum * (1.0 + tint)))
    else:
        return int(round(lum * (1.0 - tint) + (HLSMAX - HLSMAX * (1.0 - tint))))


def theme_and_tint_to_rgb(wb, theme, tint):
    """Given a workbook, a theme number and a tint return a hex based rgb"""
    rgb = get_theme_colors(wb)[theme]
    h, l, s = rgb_to_ms_hls(rgb)
    return rgb_to_hex(ms_hls_to_rgb(h, tint_luminance(tint, l), s))


def get_color(wb, cell):
    """get cell color; None for no fill"""
    if not cell.fill.patternType:
        return None
    if isinstance(cell.fill.fgColor.rgb, str):
        return "#" + cell.fill.fgColor.rgb[2:]
    theme = cell.fill.start_color.theme
    tint = cell.fill.start_color.tint
    color = theme_and_tint_to_rgb(wb, theme, tint)
    return "#" + color


def parse_risk_matrix(meta, content_ws, wb):
    """Parse a risk_matrix block into a structured dict."""
    rows = list(content_ws.iter_rows())
    if not rows:
        return None
    header = {}
    grid_count = 0
    for i, cell in enumerate(rows[0]):
        col = str(cell.value).strip().lower() if cell.value else ""
        if col == "grid":
            col = f"grid{grid_count}"
            grid_count += 1
        header[col] = i

    assert all(
        k in header
        for k in ["type", "id", "color", "abbreviation", "name", "description", "grid0"]
    )
    index = {k: i for i, k in enumerate(header)}
    size_grid = len([h for h in header if h.startswith("grid")])

    risk_matrix = {
        "urn": meta.get("urn"),
        "ref_id": meta.get("ref_id"),
        "name": meta.get("name"),
        "description": meta.get("description"),
        "probability": [],
        "impact": [],
        "risk": [],
        "grid": [],
    }

    translations = extract_translations_from_metadata(meta, "risk_matrix")
    if translations:
        risk_matrix["translations"] = translations

    grid = {}
    grid_color = {}

    for row in rows[1:]:
        if not any(c.value for c in row):
            continue

        ctype = str(row[index["type"]].value).strip().lower()
        assert ctype in ("probability", "impact", "risk")

        rid = int(row[index["id"]].value)
        abbrev = str(row[index["abbreviation"]].value or "").strip()
        name = str(row[index["name"]].value or "").strip()
        desc = str(row[index["description"]].value or "").strip()
        color_cell = row[index["color"]]

        obj_data = {
            "id": rid,
            "abbreviation": abbrev,
            "name": name,
            "description": desc,
        }

        translations = extract_translations_from_row(header, row)
        if translations:
            obj_data["translations"] = translations

        if color := get_color(wb, color_cell):
            obj_data["hexcolor"] = color

        risk_matrix[ctype].append(obj_data)

        if ctype == "probability":
            grid[rid] = []
            grid_color[rid] = []
            for i in range(size_grid):
                cell = row[index[f"grid{i}"]]
                grid[rid].append(int(cell.value))
                grid_color[rid].append(get_color(wb, cell))

    # Build grid from sorted probability IDs
    sorted_ids = sorted(grid.keys())
    risk_matrix["grid"] = [grid[rid] for rid in sorted_ids]

    # Ensure order
    for k in ("probability", "impact", "risk"):
        risk_matrix[k].sort(key=lambda x: x["id"])

    return risk_matrix


# --- Mapping logic ------------------------------------------------------------


def revert_relationship(relation: str):
    if relation == "subset":
        return "superset"
    elif relation == "superset":
        return "subset"
    else:
        return relation


# --- Extract optional columns for answers -------------------------------------


def _per_choice_lines(data: dict, col: str, n_choices: int, answer_id: str):
    raw = str(data.get(col, "") or "").strip()

    if not raw:
        return None

    lines: List[str] = []

    # If col == "description", take into consideration the multiline description logic
    if col.lower() == "description":
        for desc in raw.split("\n"):
            desc = desc.strip()

            if not desc:
                continue

            # if line starts with "|", concatenate the whole line with the previous description (without the "|")
            if desc.startswith("|") and lines:
                # Multi-line value, append to previous
                lines[-1] += "\n" + desc[1:].strip()
            else:
                lines.append(desc)
    else:
        lines = [line.strip() for line in raw.split("\n")]

    if len(lines) == 1:
        lines *= n_choices

    if len(lines) != n_choices:
        raise ValueError(
            f"(answers_definition) Invalid {col} count for answer ID '{answer_id}': "
            f"{len(lines)} values for {n_choices} choices."
        )

    return lines


# --- Object type handlers -----------------------------------------------------

_SKIPPED_TYPES = {"answers", "implementation_groups", "scores", "urn_prefix"}


def _handle_reference_controls(obj, library, compat_mode, verbose):
    """Process a reference_controls object block into the library."""
    controls = []
    base_urn = obj["meta"].get("base_urn")
    header, rows_with_data = parse_content_rows(obj["content_sheet"])
    if not header:
        return

    for row, data in rows_with_data:
        ref_id_raw = str(data.get("ref_id", "")).strip()
        default_ref_id, ref_id_for_urn = clean_ref_id_for_urn(
            ref_id_raw, compat_mode, verbose, "reference_controls"
        )

        if not default_ref_id:
            continue

        # node_id override: if a node_id column exists and cell is not empty, use it for the URN suffix
        node_id_raw = data.get("node_id")
        if node_id_raw and str(node_id_raw).strip():
            node_id_val = str(node_id_raw).strip()
            if 1 <= compat_mode <= 2:
                ref_id_for_urn = node_id_val.lower()
            else:
                ref_id_for_urn = clean_urn_suffix(node_id_val, compat_mode=0)
                if verbose and node_id_val != ref_id_for_urn:
                    print(
                        f"💬 ⚠️  [WARNING] (reference_controls) Cleaned node_id (for use in URN) '{node_id_val}' → '{ref_id_for_urn}'"
                    )

        entry = {
            "urn": f"{base_urn}:{ref_id_for_urn}",
            "ref_id": default_ref_id,
        }

        set_optional_fields(
            entry,
            data,
            ["name", "category", "csf_function", "description", "annotation"],
        )
        attach_translations_from_row(entry, header, row)
        controls.append(entry)

    extend_or_set(library["objects"], "reference_controls", controls)


def _handle_threats(obj, library, compat_mode, verbose):
    """Process a threats object block into the library."""
    threats = []
    base_urn = obj["meta"].get("base_urn")
    header, rows_with_data = parse_content_rows(obj["content_sheet"])
    if not header:
        return

    for row, data in rows_with_data:
        ref_id = str(data.get("ref_id", "")).strip()
        if not ref_id:
            continue

        # node_id override: if a node_id column exists and cell is not empty, use it for the URN suffix
        node_id_raw = data.get("node_id")
        if node_id_raw and str(node_id_raw).strip():
            urn_suffix = str(node_id_raw).strip().lower()
        else:
            urn_suffix = ref_id.lower()

        entry = {"urn": f"{base_urn}:{urn_suffix}", "ref_id": ref_id}
        set_optional_fields(entry, data, ["name", "description"])
        attach_translations_from_row(entry, header, row)
        threats.append(entry)

    extend_or_set(library["objects"], "threats", threats)


def _handle_framework(obj, library, object_blocks, prefix_to_urn, compat_mode, verbose):
    """Process a framework object block into the library."""
    meta = obj["meta"]
    content_ws = obj["content_sheet"]
    base_urn = obj["meta"].get("base_urn")

    # --- Retrieve answers block if declared ---
    answers_dict = {}
    answers_block_name = meta.get("answers_definition")
    answer_sheet = None

    if answers_block_name:
        if answers_block_name not in object_blocks:
            raise ValueError(f"Missing answers sheet: '{answers_block_name}'")

        answers_sheet = object_blocks[answers_block_name]["content_sheet"]
        rows = list(answers_sheet.iter_rows())
        if rows:
            header = [str(c.value).strip().lower() if c.value else "" for c in rows[0]]

            for row in rows[1:]:
                data = {
                    header[i]: row[i].value for i in range(len(header)) if i < len(row)
                }
                answer_id = str(data.get("id", "")).strip()
                answer_type = str(data.get("question_type", "")).strip()
                choices_raw = str(data.get("question_choices", "")).strip()

                if not answer_id or not answer_type or not choices_raw:
                    continue  # Incomplete row

                choices = []
                for line in choices_raw.split("\n"):
                    line = line.strip()
                    if not line:
                        continue
                    if line.startswith("|") and choices:
                        # Multi-line value, append to previous
                        choices[-1]["value"] += "\n" + line[1:].strip()
                    else:
                        choices.append({"urn": "", "value": line})

                # --- Optional: description ---------------------------
                description_lines = _per_choice_lines(
                    data, "description", len(choices), answer_id
                )
                if description_lines:
                    for i, desc in enumerate(description_lines):
                        if desc and desc != "/":
                            choices[i]["description"] = desc

                # --- Optional: compute_result -----------------------------------------
                compute_lines = _per_choice_lines(
                    data, "compute_result", len(choices), answer_id
                )
                if compute_lines:
                    for i, val in enumerate(compute_lines):
                        v = val.lower()
                        if v not in ("true", "false", "/", ""):
                            raise ValueError(
                                f"(answers_definition) Invalid compute_result value '{val}' "
                                f"for answer ID '{answer_id}', choice #{i + 1}. Must be 'true', 'false', '/' (= 'undefined') or empty."
                            )

                        # Use Boolean instead of string
                        if v == "/" or v == "":
                            v = None
                        elif v == "true":
                            v = True
                        elif v == "false":
                            v = False

                        if v is not None:
                            choices[i]["compute_result"] = v

                # --- Optional: add_score ----------------------------------------------
                score_lines = _per_choice_lines(
                    data, "add_score", len(choices), answer_id
                )
                if score_lines:
                    for i, val in enumerate(score_lines):
                        if val:
                            try:
                                score_to_add = int(val)
                                choices[i]["add_score"] = score_to_add
                            except (TypeError, ValueError):
                                raise ValueError(
                                    f"(answers_definition) Invalid add_score value '{val}' "
                                    f"for answer ID '{answer_id}', choice #{i + 1}. Must be an integer"
                                )

                # --- Optional: select_implementation_groups ---------------------------
                sig_lines = _per_choice_lines(
                    data,
                    "select_implementation_groups",
                    len(choices),
                    answer_id,
                )
                if sig_lines:
                    for i, val in enumerate(sig_lines):
                        if val:
                            groups = [s.strip() for s in val.split(",") if s.strip()]

                            # If IG for choice == "/undefined", continue
                            if len(groups) == 1 and groups[0].lower() == "/":
                                continue

                            if groups:
                                choices[i]["select_implementation_groups"] = groups

                # --- Optional: color ---------------------------------------------------
                color_lines = _per_choice_lines(data, "color", len(choices), answer_id)
                if color_lines:
                    for i, val in enumerate(color_lines):
                        if val:
                            if (
                                not re.fullmatch(r"#([0-9a-fA-F]{6})", val)
                                and val.lower() != "/"
                            ):
                                raise ValueError(
                                    f"(answers_definition) Invalid color value '{val}' "
                                    f"for answer ID '{answer_id}', choice #{i + 1}. Must match #RRGGBB, be '/' (= 'undefined') or the cell must be empty."
                                )

                            if val.lower() != "/":
                                choices[i]["color"] = val.upper()

                answers_dict[answer_id] = {
                    "type": answer_type,
                    "choices": choices,
                }
    else:
        if verbose:
            print(f'💬 ℹ️  No "Answers Definition" found')

    framework = {
        "urn": meta.get("urn"),
        "ref_id": meta.get("ref_id"),
        "name": meta.get("name"),
        "description": meta.get("description"),
    }

    translations = extract_translations_from_metadata(meta, "framework")
    if translations:
        framework["translations"] = translations
    else:
        if verbose:
            print(f'💬 ℹ️  No "Translation" found')

    if "min_score" in meta:
        framework["min_score"] = int(meta["min_score"])
    if "max_score" in meta:
        framework["max_score"] = int(meta["max_score"])

    score_name = meta.get("scores_definition")
    if score_name and score_name in object_blocks:
        score_header, score_rows_data = parse_content_rows(
            object_blocks[score_name]["content_sheet"]
        )
        score_defs = []
        for row, data in score_rows_data:
            score_entry = {
                "score": int(data.get("score")),
                "name": str(data.get("name", "")).strip(),
                "description": (
                    str(data.get("description", "")).strip()
                    if data.get("description", "") is not None
                    else None
                ),
            }
            if "description_doc" in data and data["description_doc"]:
                score_entry["description_doc"] = str(data["description_doc"]).strip()
            attach_translations_from_row(score_entry, score_header, row)
            score_defs.append(score_entry)
        framework["scores_definition"] = score_defs
    else:
        if verbose:
            print(f'💬 ℹ️  No "Score Definition" found')

    # [CONTENT] Implementation Groups
    ig_name = meta.get("implementation_groups_definition")
    if ig_name and ig_name in object_blocks:
        ig_header, ig_rows_data = parse_content_rows(
            object_blocks[ig_name]["content_sheet"]
        )
        ig_defs = []
        for row, data in ig_rows_data:
            ig_entry = {
                "ref_id": str(data.get("ref_id", "")).strip(),
                "name": str(data.get("name", "")).strip(),
                "description": str(data.get("description", "")).strip()
                if data.get("description")
                else None,
            }

            if data.get("default_selected") is not None:
                ig_entry["default_selected"] = bool(data.get("default_selected"))

            attach_translations_from_row(ig_entry, ig_header, row)
            ig_defs.append(ig_entry)

        framework["implementation_groups_definition"] = ig_defs

    # NOTE: The requirement_nodes loop counts ALL rows (including empty) for its
    # counter logic, so we cannot use parse_content_rows here.
    rows = list(content_ws.iter_rows())
    if rows:
        header = [str(c.value).strip().lower() if c.value else "" for c in rows[0]]
        parent_for_depth = {}
        count_for_depth = {}
        previous_node_urn = None
        previous_depth = 0
        counter = 0
        counter_fix = -1
        requirement_nodes = []
        all_urns = set()  # to detect duplicates
        for row in rows[1:]:
            counter += 1
            data = {header[i]: row[i].value for i in range(len(header)) if i < len(row)}
            if all(value is None or value == "" for value in data.values()):
                print(f"empty line {counter}")
                continue
            depth = int(data.get("depth", 1))
            ref_id = data.get("ref_id")
            ref_id = str(ref_id).strip() if ref_id is not None else None
            name = data.get("name")
            name = str(name).strip() if name is not None else None

            if depth == previous_depth + 1:
                parent_for_depth[depth] = previous_node_urn
                count_for_depth[depth] = 1
            elif depth <= previous_depth:
                pass
            else:
                raise ValueError(
                    f"Invalid depth jump from {previous_depth} to {depth} at row {counter}"
                )

            # calculate urn
            if (
                compat_mode == 1
            ):  # Use legacy URN fallback logic (for requirements without ref_id)
                skip_count = str(data.get("skip_count", "")).strip().lower() in (
                    "1",
                    "true",
                    "yes",
                    "x",
                )
                if skip_count:
                    counter_fix += 1
                    ref_id_urn = f"node{counter - counter_fix}-{counter_fix + 1}"
                else:
                    # Adds the ability to use the "node_id" column despite compatibility mode set to "1"
                    if data.get("node_id") and data.get("node_id").strip():
                        ref_id_urn = data.get("node_id").strip()
                    else:
                        ref_id_urn = (
                            ref_id.lower().replace(" ", "-")
                            if ref_id
                            else f"node{counter - counter_fix}"
                        )

                urn = f"{base_urn}:{ref_id_urn}"
            elif (
                compat_mode == 3
            ):  # Updated version of "[< v2]" (Compat Mode 1). Handling of the new "fix_count" column in order to ADD or SUBTRACT from the counter (replace "skip_count").
                # Fixed the URN writing issue when "skip_count" was true and a "ref_id" was defined.

                try:
                    fix_count = int(data.get("fix_count", ""))
                except Exception as e:
                    fix_count = None

                if fix_count:
                    counter_fix += fix_count

                    # If "ref_id" is already defined, use the defined "ref_id"
                    if data.get("node_id") and data.get("node_id").strip():
                        ref_id_urn = data.get("node_id").strip()
                    # Else if no "ref_id", use the custom node version
                    else:
                        ref_id_urn = f"node{counter + counter_fix}"

                else:
                    # Adds the ability to use the "node_id" column despite compatibility mode set to "3"
                    if data.get("node_id") and data.get("node_id").strip():
                        ref_id_urn = data.get("node_id").strip()
                    else:
                        ref_id_urn = (
                            ref_id.lower().replace(" ", "-")
                            if ref_id
                            else f"node{counter + counter_fix}"
                        )

                urn = f"{base_urn}:{ref_id_urn}"

            else:  # If compat mode = {0,2}
                if data.get("node_id") and data.get("node_id").strip():
                    urn = f"{base_urn}:{data.get('node_id').strip()}"
                elif ref_id:
                    # [+] Compat check
                    ref_id_clean = clean_urn_suffix(ref_id, compat_mode=compat_mode)
                    if verbose and ref_id != ref_id_clean:
                        print(
                            f"💬 ⚠️  [WARNING] (calculate urn [ref_id]) Cleaned ref_id (for use in URN) '{ref_id}' → '{ref_id_clean}'"
                        )
                    urn = f"{base_urn}:{ref_id_clean}"
                else:
                    p = parent_for_depth.get(depth)
                    c = count_for_depth.get(depth, 1)
                    if p:
                        urn = f"{p}:{c}" if p else f"{base_urn}:{c}"
                    elif name:
                        # [+] Compat check
                        name_clean = clean_urn_suffix(name, compat_mode=compat_mode)
                        if verbose and name != name_clean:
                            print(
                                f"💬 ⚠️  [WARNING] (calculate urn [name]) Cleaned name (for use in URN) '{name}' → '{name_clean}'"
                            )
                        urn = f"{base_urn}:{name_clean}"
                    else:
                        urn = f"{base_urn}:node{c}"
                    count_for_depth[depth] = c + 1
            previous_node_urn = urn
            previous_depth = depth
            parent_urn = parent_for_depth.get(depth)
            node = {
                "urn": urn,
                "assessable": bool(data.get("assessable")),
                "depth": depth,
            }
            if parent_urn:
                node["parent_urn"] = parent_urn
            set_optional_fields(
                node,
                data,
                ["ref_id", "name", "description", "annotation", "typical_evidence"],
            )
            # Optional: importance: mandatory/recommended/nice_to_have or empty (= undefined)
            if "importance" in data and data["importance"]:
                importance = str(data["importance"]).strip().lower()

                if importance not in [
                    "mandatory",
                    "recommended",
                    "nice_to_have",
                ]:
                    raise ValueError(
                        f'(framework_content) Invalid "importance" at row #{row[0].row}: "{data["importance"]}". Must be "mandatory"/"recommended"/"nice_to_have" or empty cell (= "undefined").'
                    )

                if importance != "undefined":
                    node["importance"] = importance

            # Optional: weight (integer)
            if (
                "weight" in data
                and data["weight"] is not None
                and str(data["weight"]).strip() != ""
            ):
                try:
                    if (w := int(data["weight"])) <= 0:
                        raise ValueError
                    node["weight"] = w
                except (TypeError, ValueError):
                    raise ValueError(
                        f"(framework) Invalid weight at row #{row[0].row}: {data['weight']}. Must be a strictly positive integer."
                    )
            if "implementation_groups" in data and data["implementation_groups"]:
                node["implementation_groups"] = [
                    s.strip() for s in str(data["implementation_groups"]).split(",")
                ]
            if "threats" in data and data["threats"]:
                threats = expand_urns_from_prefixed_list(
                    data["threats"],
                    prefix_to_urn,
                    compat_mode=compat_mode,
                    verbose=verbose,
                )
                if threats:
                    node["threats"] = threats
            if "reference_controls" in data and data["reference_controls"]:
                rc = expand_urns_from_prefixed_list(
                    data["reference_controls"],
                    prefix_to_urn,
                    compat_mode=compat_mode,
                    verbose=verbose,
                )
                if rc:
                    node["reference_controls"] = rc
            if "questions" in data and data["questions"]:
                inject_questions_into_node(
                    data,
                    node,
                    answers_dict,
                )
            attach_translations_from_row(node, header, row)
            if node.get("urn") in all_urns:
                raise ValueError(f"urn already used: {node.get('urn')}")
            all_urns.add(node.get("urn"))
            requirement_nodes.append(node)

        framework["requirement_nodes"] = requirement_nodes

    library["objects"]["framework"] = framework


def _handle_metric_definitions(obj, library, compat_mode, verbose):
    """Process a metric_definitions object block into the library."""
    metric_definitions = []
    base_urn = obj["meta"].get("base_urn")
    header, rows_with_data = parse_content_rows(obj["content_sheet"])
    if not header:
        return

    for row, data in rows_with_data:
        ref_id = str(data.get("ref_id", "")).strip()
        if not ref_id:
            continue

        # Clean ref_id for URN
        ref_id_for_urn = clean_urn_suffix(ref_id, compat_mode=compat_mode)
        if verbose and ref_id != ref_id_for_urn:
            print(
                f"💬 ⚠️  [WARNING] (metric_definitions) Cleaned ref_id (for use in URN) '{ref_id}' → '{ref_id_for_urn}'"
            )

        entry = {
            "urn": f"{base_urn}:{ref_id_for_urn}",
            "ref_id": ref_id,
        }

        set_optional_fields(entry, data, ["name", "description"])
        if "category" in data and data["category"]:
            category = str(data["category"]).strip().lower()
            if category not in ("quantitative", "qualitative"):
                raise ValueError(
                    f"(metric_definitions) Invalid category '{category}' at row #{row[0].row}. Must be 'quantitative' or 'qualitative'."
                )
            entry["category"] = category
        else:
            entry["category"] = "quantitative"  # default

        if "unit" in data and data["unit"]:
            entry["unit"] = str(data["unit"]).strip()

        # higher_is_better: default True, accept true/false/1/0/yes/no
        if "higher_is_better" in data and data["higher_is_better"] is not None:
            hib = str(data["higher_is_better"]).strip().lower()
            entry["higher_is_better"] = hib in ("1", "true", "yes", "x")
        else:
            entry["higher_is_better"] = True

        # default_target: numeric value
        if "default_target" in data and data["default_target"] is not None:
            try:
                entry["default_target"] = float(data["default_target"])
            except (TypeError, ValueError):
                raise ValueError(
                    f"(metric_definitions) Invalid default_target '{data['default_target']}' at row #{row[0].row}. Must be a number."
                )

        # choices_definition for qualitative metrics
        if entry["category"] == "qualitative":
            if "choices_definition" in data and data["choices_definition"]:
                choices = []
                choices_raw = str(data["choices_definition"]).strip()
                for line in choices_raw.split("\n"):
                    line = line.strip()
                    if not line:
                        continue
                    # Format: "name|description" or just "name"
                    if "|" in line:
                        parts = line.split("|", 1)
                        choices.append(
                            {
                                "name": parts[0].strip(),
                                "description": parts[1].strip()
                                if len(parts) > 1
                                else "",
                            }
                        )
                    else:
                        choices.append({"name": line})
                if choices:
                    entry["choices_definition"] = choices
            else:
                raise ValueError(
                    f"(metric_definitions) Qualitative metric at row #{row[0].row} requires 'choices_definition' column."
                )

        attach_translations_from_row(entry, header, row)
        metric_definitions.append(entry)

    extend_or_set(library["objects"], "metric_definitions", metric_definitions)


def _handle_risk_matrix(obj, library, wb):
    """Process a risk_matrix object block into the library."""
    matrix = parse_risk_matrix(obj["meta"], obj["content_sheet"], wb)
    if matrix:
        if "risk_matrix" not in library["objects"]:
            library["objects"]["risk_matrix"] = []
        library["objects"]["risk_matrix"].append(matrix)


def _handle_requirement_mapping_set(obj, library, wb, sheets, compat_mode, verbose):
    """Process a requirement_mapping_set object block into the library."""
    meta = obj["meta"]
    content_ws = obj["content_sheet"]
    source_node_base_urn = obj["meta"].get("source_node_base_urn")
    target_node_base_urn = obj["meta"].get("target_node_base_urn")

    requirement_mapping_set = {
        "urn": meta.get("urn"),
        "ref_id": meta.get("ref_id"),
        "name": meta.get("name"),
        "description": meta.get("description"),
        "source_framework_urn": meta.get("source_framework_urn"),
        "target_framework_urn": meta.get("target_framework_urn"),
    }
    requirement_mapping_set_revert = {
        "urn": meta.get("urn") + "-revert",
        "ref_id": meta.get("ref_id") + "-revert",
        "name": meta.get("name"),
        "description": meta.get("description"),
        "target_framework_urn": meta.get("source_framework_urn"),
        "source_framework_urn": meta.get("target_framework_urn"),
    }

    translations = extract_translations_from_metadata(meta, "requirement_mapping_set")
    if translations:
        requirement_mapping_set["translations"] = translations
        requirement_mapping_set_revert["translations"] = translations

    header, rows_with_data = parse_content_rows(content_ws)
    if not header:
        return
    requirement_mappings = []
    requirement_mappings_revert = []

    for row, data in rows_with_data:
        source_node_id_raw = str(data.get("source_node_id", "")).strip()
        target_node_id_raw = str(data.get("target_node_id", "")).strip()

        # Checks the required fields, cleaning the values first
        required_fields = ["source_node_id", "target_node_id", "relationship"]
        missing_fields = []

        for field in required_fields:
            value = data.get(field)
            if value is None or str(value).strip() == "":
                missing_fields.append(field)

        if missing_fields:
            quoted_fields = [f'"{field}"' for field in missing_fields]
            raise ValueError(
                f"(requirement_mapping_set) Missing or empty required field{'s' if len(missing_fields) > 1 else ''} in row #{row[0].row}: {', '.join(quoted_fields)}"
            )

        # [+] Compat mode set to "2" so it can be compatible with every version of mappings >=v2 without having to choose a specific compat mode
        source_node_id = clean_urn_suffix(source_node_id_raw, compat_mode=2)
        target_node_id = clean_urn_suffix(target_node_id_raw, compat_mode=2)
        if verbose and source_node_id_raw != source_node_id:
            print(
                f"💬 ⚠️  [WARNING] (requirement_mapping_set) Cleaned source_node_id '{source_node_id_raw}' → '{source_node_id}'"
            )
        if verbose and target_node_id_raw != target_node_id:
            print(
                f"💬 ⚠️  [WARNING] (requirement_mapping_set) Cleaned target_node_id '{target_node_id_raw}' → '{target_node_id}'"
            )
        entry = {
            "source_requirement_urn": source_node_base_urn + ":" + source_node_id,
            "target_requirement_urn": target_node_base_urn + ":" + target_node_id,
            "relationship": data.get("relationship").strip(),
        }
        entry_revert = {
            "source_requirement_urn": target_node_base_urn + ":" + target_node_id,
            "target_requirement_urn": source_node_base_urn + ":" + source_node_id,
            "relationship": revert_relationship(data.get("relationship").strip()),
        }
        if "rationale" in data and data["rationale"]:
            entry["rationale"] = data.get("rationale").strip()
            entry_revert["rationale"] = data.get("rationale").strip()
        if "strength_of_relationship" in data and data["strength_of_relationship"]:
            entry["strength_of_relationship"] = int(
                data.get("strength_of_relationship")
            )
            entry_revert["strength_of_relationship"] = int(
                data.get("strength_of_relationship")
            )
        requirement_mappings.append(entry)
        requirement_mappings_revert.append(entry_revert)
    requirement_mapping_set["requirement_mappings"] = requirement_mappings
    requirement_mapping_set_revert["requirement_mappings"] = requirement_mappings_revert
    library["objects"]["requirement_mapping_sets"] = [
        requirement_mapping_set,
        requirement_mapping_set_revert,
    ]

    # --- Validate source_node_id and target_node_id from mappings against Excel sheets ---
    # NOTE: This code simply checks the validity of the "source_node_id" and "target_node_id".
    #       It doesn't remove them from the created mapping if they aren't found in the Excel file.

    source_ids = set()
    target_ids = set()
    source_sheet_available = "source" in sheets
    target_sheet_available = "target" in sheets

    if source_sheet_available:
        source_sheet = wb["source"]
        source_header = [cell.value for cell in source_sheet[1]]
        if "node_id" in source_header:
            idx = source_header.index("node_id")
            for row in source_sheet.iter_rows(min_row=2):
                if idx < len(row) and row[idx].value:
                    source_ids.add(str(row[idx].value).strip())
        else:
            source_sheet_available = False
            if verbose:
                print('💬 ℹ️  "node_id" in "source" sheet header not found')
    else:
        if verbose:
            print('💬 ℹ️  Sheet "source" not found')

    if target_sheet_available:
        target_sheet = wb["target"]
        target_header = [cell.value for cell in target_sheet[1]]
        if "node_id" in target_header:
            idx = target_header.index("node_id")
            for row in target_sheet.iter_rows(min_row=2):
                if idx < len(row) and row[idx].value:
                    target_ids.add(str(row[idx].value).strip())
        else:
            target_sheet_available = False
            if verbose:
                print('💬 ℹ️  "node_id" in "target" sheet header not found')

    else:
        if verbose:
            print('💬 ℹ️  Sheet "target" not found')

    # Collect all used source/target IDs from mappings (with duplicates)
    used_source_ids = [
        m["source_requirement_urn"].split(":")[-1] for m in requirement_mappings
    ]
    used_target_ids = [
        m["target_requirement_urn"].split(":")[-1] for m in requirement_mappings
    ]

    source_missing_counts = Counter(
        id for id in used_source_ids if source_sheet_available and id not in source_ids
    )
    target_missing_counts = Counter(
        id for id in used_target_ids if target_sheet_available and id not in target_ids
    )

    # Print all warnings first (one per ID)
    if source_sheet_available:
        for sid in source_missing_counts:
            print(f'⚠️  [WARNING] source_node_id "{sid}" not found in sheet "source"')
    if target_sheet_available:
        for tid in target_missing_counts:
            print(f'⚠️  [WARNING] target_node_id "{tid}" not found in sheet "target"')

    # Check for duplicate (source, target) pairs in mappings
    mapping_pairs = [
        (
            m["source_requirement_urn"].split(":")[-1],
            m["target_requirement_urn"].split(":")[-1],
        )
        for m in requirement_mappings
    ]
    duplicate_pairs = Counter(mapping_pairs)
    duplicates_found = {
        pair: count for pair, count in duplicate_pairs.items() if count > 1
    }
    if duplicates_found:
        for (sid, tid), count in duplicates_found.items():
            print(f'🔁 [DUPLICATE] mapping "{sid}" -> "{tid}" appears {count} times')

    # Print info about missing IDs that are referenced multiple times
    if source_sheet_available:
        for sid, count in source_missing_counts.items():
            if count > 1:
                print(
                    f'ℹ️  [INFO] missing source_node_id "{sid}" is referenced {count} times in mappings'
                )
    if target_sheet_available:
        for tid, count in target_missing_counts.items():
            if count > 1:
                print(
                    f'ℹ️  [INFO] missing target_node_id "{tid}" is referenced {count} times in mappings'
                )

    # Final summary
    total_missing_sources = sum(source_missing_counts.values())
    total_missing_targets = sum(target_missing_counts.values())
    if total_missing_sources or total_missing_targets:
        print(
            f"⚠️  [SUMMARY] Missing usage count - Source: {total_missing_sources}, Target: {total_missing_targets}"
        )
        print(
            f"ℹ️  Please note that these node IDs have been added to the mapping anyway.\
            \n   If you want to correct them, please do so in your Excel file."
        )


# --- Validation ---------------------------------------------------------------


def validate_name_lengths(library: dict) -> list:
    """Check that all 'name' fields in the library respect NAME_MAX_LENGTH.

    Returns a list of (object_type, urn_or_ref, name_length) tuples for violations.
    """
    violations = []

    # Check top-level library name
    if library.get("name") and len(library["name"]) > NAME_MAX_LENGTH:
        violations.append(("library", library.get("urn", "?"), len(library["name"])))

    objects = library.get("objects", {})

    # Check framework
    framework = objects.get("framework")
    if framework:
        if framework.get("name") and len(framework["name"]) > NAME_MAX_LENGTH:
            violations.append(
                ("framework", framework.get("urn", "?"), len(framework["name"]))
            )
        for node in framework.get("requirement_nodes", []):
            if node.get("name") and len(node["name"]) > NAME_MAX_LENGTH:
                violations.append(
                    (
                        "requirement_node",
                        node.get("urn") or node.get("ref_id", "?"),
                        len(node["name"]),
                    )
                )

    # Check list-based object types
    for obj_type in (
        "threats",
        "reference_controls",
        "risk_matrix",
        "metric_definitions",
        "requirement_mapping_sets",
    ):
        for item in objects.get(obj_type, []):
            if item.get("name") and len(item["name"]) > NAME_MAX_LENGTH:
                violations.append(
                    (
                        obj_type,
                        item.get("urn") or item.get("ref_id", "?"),
                        len(item["name"]),
                    )
                )

    return violations


# --- Main logic ---------------------------------------------------------------


def create_library(
    input_file: str, output_file: str, compat_mode: int = 0, verbose: bool = False
):
    wb = openpyxl.load_workbook(input_file)
    sheets = wb.sheetnames
    object_blocks = {}

    # Step 1: Load library_meta
    if "library_meta" not in sheets:
        raise ValueError("Missing 'library_meta' sheet.")

    library_meta_sheet = wb["library_meta"]
    library_meta = parse_key_value_sheet(library_meta_sheet)

    if library_meta.get("type") != "library":
        raise ValueError("'library_meta' must declare type = library")

    # Step 2: Format publication_date
    lib_date = library_meta.get("publication_date", datetime.datetime.now())
    if isinstance(lib_date, str):
        try:
            lib_date = datetime.datetime.fromisoformat(lib_date).date()
        except ValueError:
            lib_date = datetime.datetime.now().date()
    elif isinstance(lib_date, datetime.datetime):
        lib_date = lib_date.date()

    # Step 3: Build library with fixed key order
    library_urn = None

    # [+] Compat check
    if 1 <= compat_mode <= 2:
        library_urn = library_meta.get("urn")
    else:  # Default behavior
        # Clean URN
        library_urn_raw = library_meta.get("urn")
        library_urn_clean = clean_urn_suffix(library_urn_raw, compat_mode=0)
        if library_urn_raw != library_urn_clean:
            print(
                f"⚠️  [WARNING] (create_library) Cleaned library URN '{library_urn_raw}' → '{library_urn_clean}'"
            )

        library_urn = library_urn_clean

    convert_library_version = f"v2 ; Compat Mode: [{compat_mode}] {'{' + COMPATIBILITY_MODES[compat_mode] + '}'}"

    library = {
        "convert_library_version": convert_library_version,
        "urn": library_urn,
        "locale": library_meta.get("locale"),
        "ref_id": library_meta.get("ref_id"),
        "name": library_meta.get("name"),
        "description": library_meta.get("description"),
        "copyright": library_meta.get("copyright"),
        "version": int(library_meta["version"]),
        "publication_date": lib_date,
        "provider": library_meta.get("provider"),
        "packager": library_meta.get("packager"),
    }

    # Labels Addition
    labels_cell = library_meta.get("labels")
    if labels_cell:
        library["labels"] = list(
            set(
                [
                    label.upper()
                    for label in re.split(r"[\s,\n]+", labels_cell.strip())
                    if label
                ]
            )
        )

    translations = extract_translations_from_metadata(library_meta, "library")
    if translations:
        library["translations"] = translations
    if "dependencies" in library_meta:
        dependencies = [
            x.strip()
            for x in re.split(r"[\s,]+", library_meta["dependencies"])
            if x.strip()
        ]
        if dependencies:
            library["dependencies"] = dependencies

    library["objects"] = {}
    prefix_to_urn = {}

    # Step 4: Detect all object blocks
    for sheet_name in sheets:
        if sheet_name.endswith("_meta") and sheet_name != "library_meta":
            prefix = sheet_name[:-5]
            content_name = f"{prefix}_content"
            if content_name not in sheets:
                print(f"⚠️  [WARNING] Missing content for {prefix}")
                continue

            meta_sheet = wb[sheet_name]
            meta_data = parse_key_value_sheet(meta_sheet)
            object_type = meta_data.get("type")
            object_name = meta_data.get("name", prefix)

            # Capture prefix mappings if this is a urn_prefix block
            if object_type == "urn_prefix" and f"{prefix}_content" in wb.sheetnames:
                ws_prefix = wb[f"{prefix}_content"]
                rows = list(ws_prefix.iter_rows(values_only=True))
                if rows:
                    header = [
                        str(cell).strip().lower() if cell else "" for cell in rows[0]
                    ]
                    idx_id = (
                        header.index("prefix_id") if "prefix_id" in header else None
                    )
                    idx_val = (
                        header.index("prefix_value")
                        if "prefix_value" in header
                        else None
                    )
                    if idx_id is not None and idx_val is not None:
                        for row in rows[1:]:
                            prefix_id = (
                                str(row[idx_id]).strip() if row[idx_id] else None
                            )
                            prefix_val = (
                                str(row[idx_val]).strip() if row[idx_val] else None
                            )
                            if prefix_id and prefix_val:
                                prefix_to_urn[prefix_id] = prefix_val

            object_blocks[object_name] = {
                "type": object_type,
                "meta": meta_data,
                "content_sheet": wb[content_name],
            }

    print(f"📦 Found {len(object_blocks)} objects.")

    # Step 5: Ordered object insertion (reference_controls before threats)
    priority_order = ["reference_controls", "threats"]

    sorted_object_names = sorted(
        object_blocks.keys(),
        key=lambda k: (
            priority_order.index(object_blocks[k]["type"])
            if object_blocks[k]["type"] in priority_order
            else len(priority_order)
        ),
    )

    # Dispatch table for object type handlers
    handler_map = {
        "reference_controls": lambda obj: _handle_reference_controls(
            obj, library, compat_mode, verbose
        ),
        "threats": lambda obj: _handle_threats(obj, library, compat_mode, verbose),
        "framework": lambda obj: _handle_framework(
            obj, library, object_blocks, prefix_to_urn, compat_mode, verbose
        ),
        "metric_definitions": lambda obj: _handle_metric_definitions(
            obj, library, compat_mode, verbose
        ),
        "risk_matrix": lambda obj: _handle_risk_matrix(obj, library, wb),
        "requirement_mapping_set": lambda obj: _handle_requirement_mapping_set(
            obj, library, wb, sheets, compat_mode, verbose
        ),
    }

    for name in sorted_object_names:
        obj = object_blocks[name]
        obj_type = obj["type"]
        print(f"→ Handling {obj_type}: {name}")

        handler = handler_map.get(obj_type)
        if handler:
            handler(obj)
        elif obj_type not in _SKIPPED_TYPES:
            print("type not handled:", obj_type)

    # Step 6: Validate name lengths
    violations = validate_name_lengths(library)
    if violations:
        print(
            f"\n❌ [ERROR] {len(violations)} object(s) have a 'name' exceeding {NAME_MAX_LENGTH} characters:"
        )
        for obj_type, identifier, length in violations:
            print(f"   - [{obj_type}] {identifier} ({length} chars)")
        raise ValueError(
            f"{len(violations)} object(s) have a 'name' exceeding the {NAME_MAX_LENGTH}-character limit. "
            f"Please shorten them in the source file."
        )

    # Step 7: Export to YAML
    print(f'✅ YAML saved as: "{output_file}"')
    if not verbose:
        print(
            '💡 Tip: Use "--verbose" to display hidden messages. This can help to understand certain behaviors.'
        )

    with open(output_file, "w", encoding="utf-8") as f:
        yaml.dump(library, f, sort_keys=False, allow_unicode=False)


# --- CLI interface ------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(
        description="Create a YAML library from a v2 Excel file.",
        formatter_class=argparse.RawTextHelpFormatter,
    )
    parser.add_argument("input_file", type=str, help="Input Excel file (v2 format)")
    parser.add_argument(
        "-c",
        "--compat",
        type=int,
        choices=COMPATIBILITY_MODES.keys(),
        help="Compatibility mode number:\n    "
        + "\n    ".join(f"{k} = {v}" for k, v in COMPATIBILITY_MODES.items()),
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Enable verbose output. Verbose messages start with a 💬 (speech bubble) emoji.",
    )
    parser.add_argument(
        "--output",
        type=str,
        help="Custom output file name (only used in non-bulk mode).",
    )
    parser.add_argument(
        "--bulk",
        action="store_true",
        help="Enable bulk mode to process all .xlsx files in a directory.",
    )
    parser.add_argument(
        "--output-dir",
        type=str,
        help="Output directory for YAML files (only used with --bulk mode).",
    )
    args = parser.parse_args()

    input_path = Path(args.input_file)
    if not input_path.exists():
        print_error(f"File not found: {input_path}")
        sys.exit(1)

    # Determine compatibility mode (default to 0 if not provided)
    compat_mode = args.compat if args.compat else 0
    if compat_mode not in COMPATIBILITY_MODES:
        print_error(
            f"Invalid compatibility mode: {compat_mode}. Allowed modes: {list(COMPATIBILITY_MODES.keys())}"
        )
        sys.exit(1)

    # --- BULK MODE ------------------------------------------------------------
    if args.bulk:
        if args.output:
            print_error('The option "--output" cannot be used with "--bulk" mode.')
            sys.exit(1)

        if not input_path.is_dir():
            print_error("Bulk mode requires a directory as input")
            sys.exit(1)

        # Validate output directory and create it if needed
        if args.output_dir:
            output_dir = Path(args.output_dir)
            try:
                output_dir.mkdir(parents=True, exist_ok=True)
            except Exception as e:
                print_error(f'Cannot create output directory: "{output_dir}": {e}')
                sys.exit(1)
        else:
            output_dir = Path.cwd()  # Use current working directory as default

        error_files = []  # Collect names of files that failed
        # Find all .xlsx files in the input directory (temp Excel files starting with "~$" are excluded)
        xlsx_files = [
            f for f in input_path.glob("*.xlsx") if not f.name.startswith("~$")
        ]

        if not xlsx_files:
            print_error(f'No .xlsx files found in directory: "{input_path}". Abort...')
            sys.exit(1)

        for i, file in enumerate(xlsx_files):
            output_path = output_dir / (file.stem + ".yaml")  # Output file path
            try:
                print(f'\n▶️  Processing file [{i + 1}/{len(xlsx_files)}]: "{file}"')
                create_library(
                    str(file),
                    str(output_path),
                    compat_mode=compat_mode,
                    verbose=args.verbose,
                )
            except Exception as e:
                print_error(f'Failed to process "{file}": {e}')
                error_files.append(file.name)

        # Summary at the end of bulk processing
        print("\n📋 Bulk mode completed!")

        if error_files:
            print(
                f"❌ The following file{'s' if len(error_files) > 1 else ''} failed to process:"
            )
            for f in error_files:
                print(f"   - {f}")
                if not args.verbose:
                    print(
                        '💡 Tip: Use "--verbose" to display hidden messages. This can help to understand certain errors.'
                    )
            sys.exit(1)
        else:
            print("✅ All files processed successfully!")
            sys.exit(0)

    # --- SINGLE FILE MODE -----------------------------------------------------
    else:
        if args.output_dir:
            print_error(
                'The option "--output-dir" can only be used with "--bulk" mode.'
            )
            sys.exit(1)

        # Determine output file path (add .yaml if missing)
        if args.output:
            output_path = Path(args.output)
            if output_path.suffix != ".yaml":
                output_path = output_path.name + ".yaml"
        else:
            output_path = Path(input_path.stem + ".yaml")

        try:
            create_library(
                str(input_path),
                str(output_path),
                compat_mode=compat_mode,
                verbose=args.verbose,
            )
        except Exception as e:
            print_error(str(e))
            if not args.verbose:
                print(
                    '💡 Tip: Use "--verbose" to display hidden messages. This can help to understand certain errors.'
                )
            sys.exit(1)


if __name__ == "__main__":
    main()
