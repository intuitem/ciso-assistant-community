"""
POA&M Export Service

Enhanced POA&M export capabilities inspired by RampControl patterns:
- FedRAMP POA&M Excel export (Appendix A format)
- CSV export for bulk operations
- OSCAL POA&M export
- Deviation tracking reports
- Milestone tracking reports
"""

import json
import logging
import uuid
from typing import Dict, List, Any, Optional
from dataclasses import dataclass
from datetime import datetime, date
from io import BytesIO

logger = logging.getLogger(__name__)


@dataclass
class ExportResult:
    """Result of an export operation"""
    success: bool
    content: bytes
    filename: str
    content_type: str
    errors: List[str]


class DeviationType:
    """FedRAMP deviation types"""
    FUNCTIONAL = "functional"  # Cannot implement as written
    OPERATIONAL = "operational"  # Operational constraints
    RISK_ASSESSMENT = "risk_assessment"  # Risk-based decision


class POAMExportService:
    """
    Service for exporting POA&M data in various formats.

    Supports:
    - FedRAMP Appendix A Excel format
    - Standard CSV export
    - OSCAL POA&M JSON/YAML
    - Custom deviation reports
    - Milestone tracking reports
    """

    # FedRAMP POA&M column headers
    FEDRAMP_COLUMNS = [
        "POA&M ID",
        "Controls",
        "Weakness Name",
        "Weakness Description",
        "Weakness Detector Source",
        "Source Identifying Vulnerability",
        "Asset Identifier",
        "Point of Contact",
        "Resources Required",
        "Overall Remediation Plan",
        "Original Detection Date",
        "Scheduled Completion Date",
        "Planned Milestones",
        "Milestone Changes",
        "Status",
        "Vendor Dependency",
        "Last Vendor Check-in Date",
        "Vendor Dependent Product Name",
        "Original Risk Rating",
        "Adjusted Risk Rating",
        "Risk Adjustment",
        "False Positive",
        "Operational Requirement",
        "Deviation Rationale",
        "Supporting Documents",
        "Comments",
    ]

    def export_fedramp_xlsx(
        self,
        poam_items: List[Dict[str, Any]],
        system_info: Dict[str, Any]
    ) -> ExportResult:
        """
        Export POA&M items to FedRAMP Appendix A Excel format.

        Args:
            poam_items: List of POA&M items
            system_info: System information for header

        Returns:
            ExportResult with Excel content
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter

            wb = Workbook()

            # Create POA&M sheet
            ws = wb.active
            ws.title = "POA&M"

            # Styling
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=10)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Write headers
            for col_idx, header in enumerate(self.FEDRAMP_COLUMNS, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border

                # Set column widths
                ws.column_dimensions[get_column_letter(col_idx)].width = 15

            # Write data rows
            for row_idx, item in enumerate(poam_items, start=2):
                row_data = self._map_to_fedramp_row(item)
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.border = border
                    cell.alignment = Alignment(vertical='top', wrap_text=True)

            # Freeze header row
            ws.freeze_panes = 'A2'

            # Create system info sheet
            ws_info = wb.create_sheet("System Information")
            self._write_system_info_sheet(ws_info, system_info)

            # Create summary sheet
            ws_summary = wb.create_sheet("Summary")
            self._write_summary_sheet(ws_summary, poam_items)

            # Save to bytes
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            return ExportResult(
                success=True,
                content=buffer.getvalue(),
                filename=f"fedramp_poam_{system_info.get('name', 'system')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                errors=[]
            )

        except ImportError:
            return ExportResult(
                success=False,
                content=b'',
                filename='',
                content_type='',
                errors=['openpyxl library not installed. Run: pip install openpyxl']
            )
        except Exception as e:
            logger.error(f"Error exporting FedRAMP POA&M: {e}")
            return ExportResult(
                success=False,
                content=b'',
                filename='',
                content_type='',
                errors=[str(e)]
            )

    def _map_to_fedramp_row(self, item: Dict[str, Any]) -> List[Any]:
        """Map POA&M item to FedRAMP column order"""
        # Format milestones
        milestones = item.get('milestones', [])
        milestone_text = "\n".join([
            f"{m.get('description', '')} ({m.get('target_date', '')})"
            for m in milestones
        ]) if milestones else ""

        # Format milestone changes
        milestone_changes = []
        for m in milestones:
            if m.get('updated_at'):
                milestone_changes.append(
                    f"{m.get('description', '')}: Status changed to {m.get('status', '')} on {m.get('updated_at', '')}"
                )
        milestone_changes_text = "\n".join(milestone_changes)

        # Determine if there's a deviation
        has_deviation = item.get('has_deviation', False)
        deviation_type = item.get('deviation_type', '')

        return [
            item.get('weakness_id', ''),  # POA&M ID
            item.get('control_id', ''),  # Controls
            item.get('title', ''),  # Weakness Name
            item.get('description', ''),  # Weakness Description
            item.get('source_type', ''),  # Weakness Detector Source
            item.get('source_reference', ''),  # Source Identifying Vulnerability
            item.get('asset_identifier', ''),  # Asset Identifier
            item.get('point_of_contact', ''),  # Point of Contact
            item.get('resources_required', ''),  # Resources Required
            item.get('remediation_plan', ''),  # Overall Remediation Plan
            item.get('identified_date', ''),  # Original Detection Date
            item.get('estimated_completion_date', ''),  # Scheduled Completion Date
            milestone_text,  # Planned Milestones
            milestone_changes_text,  # Milestone Changes
            item.get('status', ''),  # Status
            'Yes' if item.get('vendor_dependent') else 'No',  # Vendor Dependency
            item.get('last_vendor_checkin', ''),  # Last Vendor Check-in Date
            item.get('vendor_product_name', ''),  # Vendor Dependent Product Name
            item.get('risk_level', ''),  # Original Risk Rating
            item.get('adjusted_risk_level', item.get('risk_level', '')),  # Adjusted Risk Rating
            item.get('risk_adjustment_reason', ''),  # Risk Adjustment
            'Yes' if item.get('is_false_positive') else 'No',  # False Positive
            'Yes' if deviation_type == DeviationType.OPERATIONAL else 'No',  # Operational Requirement
            item.get('deviation_justification', '') if has_deviation else '',  # Deviation Rationale
            ', '.join(item.get('supporting_documents', [])) if item.get('supporting_documents') else '',  # Supporting Documents
            item.get('comments', ''),  # Comments
        ]

    def _write_system_info_sheet(self, ws, system_info: Dict[str, Any]):
        """Write system information sheet"""
        from openpyxl.styles import Font

        info_rows = [
            ("System Name:", system_info.get('name', '')),
            ("System ID:", system_info.get('id', '')),
            ("Authorization Type:", system_info.get('authorization_type', 'FedRAMP')),
            ("Impact Level:", system_info.get('impact_level', '')),
            ("ISSO:", system_info.get('isso', '')),
            ("System Owner:", system_info.get('system_owner', '')),
            ("Authorizing Official:", system_info.get('ao', '')),
            ("Report Generated:", datetime.now().strftime('%Y-%m-%d %H:%M')),
        ]

        for row_idx, (label, value) in enumerate(info_rows, start=1):
            ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=value)

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50

    def _write_summary_sheet(self, ws, poam_items: List[Dict[str, Any]]):
        """Write summary sheet with statistics"""
        from openpyxl.styles import Font, PatternFill

        # Calculate statistics
        total_items = len(poam_items)
        by_status = {}
        by_risk = {}
        overdue = 0
        with_deviation = 0

        today = date.today()

        for item in poam_items:
            status = item.get('status', 'unknown')
            by_status[status] = by_status.get(status, 0) + 1

            risk = item.get('risk_level', 'unknown')
            by_risk[risk] = by_risk.get(risk, 0) + 1

            # Check if overdue
            completion_date = item.get('estimated_completion_date')
            if completion_date and status not in ['completed', 'cancelled']:
                try:
                    if isinstance(completion_date, str):
                        comp_date = date.fromisoformat(completion_date)
                    else:
                        comp_date = completion_date
                    if comp_date < today:
                        overdue += 1
                except (ValueError, TypeError):
                    pass

            if item.get('has_deviation'):
                with_deviation += 1

        # Write summary
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        ws.cell(row=1, column=1, value="POA&M Summary").font = Font(bold=True, size=14)

        # Overall stats
        stats_start = 3
        ws.cell(row=stats_start, column=1, value="Overall Statistics").font = Font(bold=True, size=12)

        stats = [
            ("Total POA&M Items:", total_items),
            ("Overdue Items:", overdue),
            ("Items with Deviations:", with_deviation),
        ]

        for idx, (label, value) in enumerate(stats, start=stats_start + 1):
            ws.cell(row=idx, column=1, value=label).font = Font(bold=True)
            ws.cell(row=idx, column=2, value=value)

        # By Status
        status_start = stats_start + len(stats) + 3
        ws.cell(row=status_start, column=1, value="By Status").font = Font(bold=True, size=12)

        for idx, (status, count) in enumerate(sorted(by_status.items()), start=status_start + 1):
            ws.cell(row=idx, column=1, value=status.replace('_', ' ').title())
            ws.cell(row=idx, column=2, value=count)

        # By Risk Level
        risk_start = status_start + len(by_status) + 3
        ws.cell(row=risk_start, column=1, value="By Risk Level").font = Font(bold=True, size=12)

        risk_order = ['very_high', 'high', 'moderate', 'low', 'very_low']
        for idx, risk in enumerate(risk_order, start=risk_start + 1):
            if risk in by_risk:
                ws.cell(row=idx, column=1, value=risk.replace('_', ' ').title())
                ws.cell(row=idx, column=2, value=by_risk[risk])

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15

    def export_csv(
        self,
        poam_items: List[Dict[str, Any]],
        columns: Optional[List[str]] = None
    ) -> ExportResult:
        """
        Export POA&M items to CSV format.

        Args:
            poam_items: List of POA&M items
            columns: Optional list of columns to include

        Returns:
            ExportResult with CSV content
        """
        import csv
        from io import StringIO

        columns = columns or [
            'weakness_id', 'title', 'description', 'control_id',
            'status', 'risk_level', 'identified_date',
            'estimated_completion_date', 'actual_completion_date',
            'point_of_contact', 'has_deviation'
        ]

        try:
            output = StringIO()
            writer = csv.DictWriter(output, fieldnames=columns, extrasaction='ignore')
            writer.writeheader()

            for item in poam_items:
                writer.writerow(item)

            csv_content = output.getvalue().encode('utf-8')

            return ExportResult(
                success=True,
                content=csv_content,
                filename=f"poam_export_{datetime.now().strftime('%Y%m%d')}.csv",
                content_type='text/csv',
                errors=[]
            )

        except Exception as e:
            logger.error(f"Error exporting CSV: {e}")
            return ExportResult(
                success=False,
                content=b'',
                filename='',
                content_type='',
                errors=[str(e)]
            )

    def export_oscal_poam(
        self,
        poam_items: List[Dict[str, Any]],
        system_info: Dict[str, Any]
    ) -> ExportResult:
        """
        Export POA&M items to OSCAL POA&M format.

        Args:
            poam_items: List of POA&M items
            system_info: System information

        Returns:
            ExportResult with OSCAL JSON content
        """
        try:
            # Build OSCAL POA&M structure
            oscal_poam = {
                "plan-of-action-and-milestones": {
                    "uuid": str(uuid.uuid4()),
                    "metadata": {
                        "title": f"POA&M for {system_info.get('name', 'System')}",
                        "last-modified": datetime.now().isoformat(),
                        "version": "1.0",
                        "oscal-version": "1.1.2"
                    },
                    "system-id": {
                        "identifier-type": "https://fedramp.gov",
                        "id": system_info.get('id', str(uuid.uuid4()))
                    },
                    "poam-items": []
                }
            }

            # Convert each POA&M item
            for item in poam_items:
                oscal_item = self._convert_to_oscal_poam_item(item)
                oscal_poam["plan-of-action-and-milestones"]["poam-items"].append(oscal_item)

            json_content = json.dumps(oscal_poam, indent=2, default=str)

            return ExportResult(
                success=True,
                content=json_content.encode('utf-8'),
                filename=f"poam_oscal_{system_info.get('name', 'system')}_{datetime.now().strftime('%Y%m%d')}.json",
                content_type='application/json',
                errors=[]
            )

        except Exception as e:
            logger.error(f"Error exporting OSCAL POA&M: {e}")
            return ExportResult(
                success=False,
                content=b'',
                filename='',
                content_type='',
                errors=[str(e)]
            )

    def _convert_to_oscal_poam_item(self, item: Dict[str, Any]) -> Dict[str, Any]:
        """Convert POA&M item to OSCAL format"""
        oscal_item = {
            "uuid": str(item.get('id', uuid.uuid4())),
            "title": item.get('title', ''),
            "description": item.get('description', ''),
            "props": [
                {"name": "weakness-id", "value": item.get('weakness_id', '')},
                {"name": "status", "value": item.get('status', 'open')},
                {"name": "risk-level", "value": item.get('risk_level', 'moderate')},
            ]
        }

        # Add related observations/risks
        if item.get('control_id'):
            oscal_item["related-findings"] = [{
                "finding-uuid": str(uuid.uuid4()),
                "control-id": item.get('control_id')
            }]

        # Add milestones
        milestones = item.get('milestones', [])
        if milestones:
            oscal_item["milestones"] = [
                {
                    "uuid": str(m.get('id', uuid.uuid4())),
                    "title": m.get('description', ''),
                    "due-date": m.get('target_date', ''),
                    "props": [
                        {"name": "status", "value": m.get('status', 'pending')}
                    ]
                }
                for m in milestones
            ]

        # Add deviation info as remarks
        if item.get('has_deviation'):
            oscal_item["remarks"] = f"Deviation: {item.get('deviation_justification', '')}"

        return oscal_item

    def generate_deviation_report(
        self,
        poam_items: List[Dict[str, Any]],
        system_info: Dict[str, Any]
    ) -> ExportResult:
        """
        Generate a deviation tracking report.

        Args:
            poam_items: List of POA&M items
            system_info: System information

        Returns:
            ExportResult with report content
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Border, Side

            # Filter items with deviations
            deviation_items = [item for item in poam_items if item.get('has_deviation')]

            wb = Workbook()
            ws = wb.active
            ws.title = "Deviation Report"

            # Styling
            header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Headers
            headers = [
                "POA&M ID",
                "Weakness Name",
                "Control ID",
                "Deviation Type",
                "Justification",
                "Approval Status",
                "Approval Date",
                "Risk Level",
                "Compensating Controls"
            ]

            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border

            # Data rows
            for row_idx, item in enumerate(deviation_items, start=2):
                row_data = [
                    item.get('weakness_id', ''),
                    item.get('title', ''),
                    item.get('control_id', ''),
                    item.get('deviation_type', ''),
                    item.get('deviation_justification', ''),
                    'Approved' if item.get('deviation_approved') else 'Pending',
                    item.get('deviation_approval_date', ''),
                    item.get('risk_level', ''),
                    item.get('compensating_controls', ''),
                ]

                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.border = border

            # Set column widths
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 50
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 30

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            return ExportResult(
                success=True,
                content=buffer.getvalue(),
                filename=f"deviation_report_{datetime.now().strftime('%Y%m%d')}.xlsx",
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                errors=[]
            )

        except Exception as e:
            logger.error(f"Error generating deviation report: {e}")
            return ExportResult(
                success=False,
                content=b'',
                filename='',
                content_type='',
                errors=[str(e)]
            )

    def generate_milestone_report(
        self,
        poam_items: List[Dict[str, Any]],
        include_completed: bool = False
    ) -> ExportResult:
        """
        Generate a milestone tracking report.

        Args:
            poam_items: List of POA&M items
            include_completed: Whether to include completed milestones

        Returns:
            ExportResult with report content
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "Milestone Report"

            # Styling
            header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
            overdue_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Headers
            headers = [
                "POA&M ID",
                "Weakness Name",
                "Milestone Description",
                "Target Date",
                "Status",
                "Days Remaining/Overdue",
                "Last Updated"
            ]

            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border

            # Collect all milestones
            today = date.today()
            row_idx = 2

            for item in poam_items:
                milestones = item.get('milestones', [])

                for milestone in milestones:
                    status = milestone.get('status', 'pending')

                    if not include_completed and status == 'completed':
                        continue

                    # Calculate days
                    target_date = milestone.get('target_date')
                    days_remaining = ''
                    is_overdue = False

                    if target_date:
                        try:
                            if isinstance(target_date, str):
                                target = date.fromisoformat(target_date)
                            else:
                                target = target_date

                            delta = (target - today).days
                            if delta < 0:
                                days_remaining = f"{abs(delta)} days overdue"
                                is_overdue = True
                            elif delta == 0:
                                days_remaining = "Due today"
                            else:
                                days_remaining = f"{delta} days remaining"
                        except (ValueError, TypeError):
                            pass

                    row_data = [
                        item.get('weakness_id', ''),
                        item.get('title', ''),
                        milestone.get('description', ''),
                        target_date or '',
                        status,
                        days_remaining,
                        milestone.get('updated_at', ''),
                    ]

                    for col_idx, value in enumerate(row_data, start=1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.value = value
                        cell.border = border

                        if is_overdue:
                            cell.fill = overdue_fill

                    row_idx += 1

            # Set column widths
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 15

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            return ExportResult(
                success=True,
                content=buffer.getvalue(),
                filename=f"milestone_report_{datetime.now().strftime('%Y%m%d')}.xlsx",
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                errors=[]
            )

        except Exception as e:
            logger.error(f"Error generating milestone report: {e}")
            return ExportResult(
                success=False,
                content=b'',
                filename='',
                content_type='',
                errors=[str(e)]
            )
