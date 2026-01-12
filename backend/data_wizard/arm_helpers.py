"""
ARM Format Helper for EBIOS RM Study Import

This module handles the import of EBIOS RM studies from ARM Excel format.
ARM exports studies with 91 sheets organized by workshop.

Supports both French and English ARM exports with variant sheet/column names.
"""

import logging
from io import BytesIO

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


# =============================================================================
# Sheet and Column Name Variants (FR/EN support)
# =============================================================================

# Sheet name variants - maps logical name to list of possible sheet names
SHEET_VARIANTS = {
    # Workshop 1 - Scope & Security Baseline
    "business_values": [
        "1 - Base de valeurs métier",
        "1 - Business values",
        "1 - Primary assets",
    ],
    "missions": [
        "1 - Base de missions",
        "1 - Missions",
    ],
    "feared_events": [
        "1 - Base d'évènements redoutés",
        "1 - Feared events",
    ],
    "supporting_assets": [
        "1 - Biens supports",
        "1 - Supporting assets",
    ],
    "impact_level_scale": [
        "1 - Échelle de niveau d'impact",
        "1 - Impact level scale",
    ],
    "complementary_measures": [
        "1 - Mesures complémentaires",
        "1 - Complementary measures",
        "1 - Applied controls",
    ],
    # Workshop 2 - Risk Origins
    "roto_couples": [
        "2 - Base de couples SR OV",
        "2 - Couples SR OV",
        "2 - Couples SR OV 1",
        "2 - RO TO pairs",
        "2 - RO TO pairs 1",
    ],
    # Workshop 3 - Stakeholders & Strategic Scenarios
    "stakeholder_categories": [
        "3 - Catégories de partie prena",
        "3 - Stakeholder types",
    ],
    "stakeholders": [
        "3 - Base de parties prenantes",
        "3 - Stakeholders",
    ],
    "stakeholder_danger_levels": [
        "3 - Niveau de danger des parti",
        "3 - Stakeholders danger level",
    ],
    "strategic_scenarios": [
        "3 - Synthèse des scénarios str",
        "3 - Strategic scenarios synthe",
    ],
    # Workshop 4 - Operational Scenarios
    "elementary_actions": [
        "4 - Actions élémentaires",
        "4 - Elementary actions",
    ],
}

# Column name variants - maps logical name to list of possible column names
COLUMN_VARIANTS = {
    # Common columns
    "name": ["Nom", "Name"],
    "description": ["Description"],
    "ref_id": ["Réf.", "Ref.", "Ref", "Reference"],
    "abbreviation": ["Abrév.", "Abbrev.", "Abbreviation"],
    "justification": ["Justification"],
    "selected_check": ["✓/❒", "✓", "Selected"],
    # Workshop 1
    "mission": ["Mission"],
    "gravity_level": ["Niveau de gravité", "Gravity level", "Severity level"],
    "level": ["Niveau", "Level"],
    "supporting_assets": ["Biens supports", "Supporting assets"],
    "parent_asset": ["Bien support parent", "Parent supporting asset", "Parent asset"],
    "feared_event": ["Évènement redouté", "Feared event"],
    "retained_gravity": ["Gravité retenue", "Retained gravity", "Selected gravity"],
    "business_values": ["Valeurs métier", "Business values", "Primary assets"],
    "impacts": ["Impacts"],
    # Workshop 2 - RoTo couples
    "risk_origin": ["Source de risque", "Risk origin", "RO", "Risk Origin"],
    "target_objective": ["Objectif visé", "Target objective", "TO", "Target Objective"],
    "motivation": ["Motivation"],
    "resources": ["Ressources", "Resources"],
    "activity": ["Activité", "Activity"],
    "retained": ["Retenu", "Selected", "Retained"],
    # Workshop 3 - Stakeholders
    "category": ["Catégorie", "Category", "Type"],
    "risk_origins": ["Sources de risque", "Risk origins"],
    "short_name": ["Abrév.", "Short name"],
    "dependence": ["Dépendance", "Dependence", "Dependency"],
    "penetration": ["Pénétration", "Penetration"],
    "cyber_maturity": ["Maturité cyber", "Cyber maturity", "Maturity"],
    "trust": ["Confiance", "Trust"],
    "exposure": ["Exposition", "Exposure"],
    "reliability": ["Fiabilité", "Reliability"],
    "danger_level": ["Niveau de menace", "Danger level", "Threat level"],
    # Workshop 3 - Strategic scenarios
    "scenario_name": ["Scénario stratégique", "Strategic scenario", "Scenario"],
    "attack_path": ["Chemin d'attaque", "Attack path"],
    "attack_path_ref": [
        "Réf. chemin d'attaque",
        "Attack path ref",
        "Attack path reference",
    ],
    # Workshop 4
    "action_name": ["Action élémentaire", "Elementary action", "Action"],
}


def find_sheet(workbook, sheet_key: str):
    """
    Find a sheet in the workbook by trying all variant names.

    Args:
        workbook: openpyxl Workbook object
        sheet_key: Logical sheet key from SHEET_VARIANTS

    Returns:
        Tuple of (sheet, sheet_name) or (None, None) if not found
    """
    variants = SHEET_VARIANTS.get(sheet_key, [])
    if not variants:
        logger.warning(f"[ARM] Unknown sheet key: {sheet_key}")
        return None, None

    for variant in variants:
        if variant in workbook.sheetnames:
            logger.debug(f"[ARM] Found sheet '{variant}' for key '{sheet_key}'")
            return workbook[variant], variant

    logger.warning(
        f"[ARM] No sheet found for key '{sheet_key}'. "
        f"Tried: {variants}. Available: {workbook.sheetnames}"
    )
    return None, None


def get_col(row: dict, col_key: str, default=None):
    """
    Get a column value from a row by trying all variant column names.

    Args:
        row: Dictionary of column name -> value
        col_key: Logical column key from COLUMN_VARIANTS
        default: Default value if not found

    Returns:
        The column value or default
    """
    variants = COLUMN_VARIANTS.get(col_key, [col_key])  # Fallback to key itself

    for variant in variants:
        val = row.get(variant)
        if val is not None and val != "":
            return val

    return default


# Legacy class for backwards compatibility - now uses SHEET_VARIANTS
class ARMSheets:
    """ARM Excel sheet names organized by workshop (legacy, use SHEET_VARIANTS instead)."""

    # Workshop 1 - Scope & Security Baseline
    BUSINESS_VALUES = SHEET_VARIANTS["business_values"][0]
    MISSIONS = SHEET_VARIANTS["missions"][0]
    FEARED_EVENTS = SHEET_VARIANTS["feared_events"][0]
    SUPPORTING_ASSETS = SHEET_VARIANTS["supporting_assets"][0]
    IMPACT_LEVEL_SCALE = SHEET_VARIANTS["impact_level_scale"][0]
    COMPLEMENTARY_MEASURES = SHEET_VARIANTS["complementary_measures"][0]

    # Workshop 2 - Risk Origins
    ROTO_COUPLES_VARIANTS = SHEET_VARIANTS["roto_couples"]

    # Workshop 3 - Stakeholders & Strategic Scenarios
    STAKEHOLDER_CATEGORIES = SHEET_VARIANTS["stakeholder_categories"][0]
    STAKEHOLDERS = SHEET_VARIANTS["stakeholders"][0]
    STAKEHOLDER_DANGER_LEVELS = SHEET_VARIANTS["stakeholder_danger_levels"][0]
    STRATEGIC_SCENARIOS = SHEET_VARIANTS["strategic_scenarios"][0]

    # Workshop 4 - Operational Scenarios
    ELEMENTARY_ACTIONS = SHEET_VARIANTS["elementary_actions"][0]


# =============================================================================
# Helper functions
# =============================================================================


def normalize_category_name(text: str) -> str:
    """
    Normalize a category name for matching.

    ARM Excel uses French plural forms (e.g., 'Clients') that need to be
    matched to our internal singular forms (e.g., 'client').

    Args:
        text: The raw category name from Excel

    Returns:
        Normalized name (lowercase, trailing 's' stripped)
    """
    if not text:
        return ""

    normalized = text.strip().lower()

    # Strip trailing 's' for plural -> singular conversion
    if normalized.endswith("s") and len(normalized) > 1:
        normalized = normalized[:-1]

    return normalized


def parse_plus_signs(text: str) -> int:
    """
    Parse '+' signs from ARM Excel format to numeric value.

    ARM uses '+', '++', '+++', '++++' to represent levels 1-4.
    Empty or invalid values return 0 (undefined).

    Args:
        text: The raw text from the Excel cell

    Returns:
        Numeric value (0-4)
    """
    if not text:
        return 0

    text = str(text).strip()
    # Count the '+' signs
    plus_count = text.count("+")

    # Validate it's only '+' characters
    if plus_count > 0 and text == "+" * plus_count:
        return min(plus_count, 4)  # Cap at 4

    return 0


def parse_bullet_list(text: str) -> list[str]:
    """
    Parse a bullet-pointed list from ARM Excel format.

    ARM uses '•' as bullet points with '_x000D_\\n' or '\\n' as line separators.

    Args:
        text: The raw text from the Excel cell

    Returns:
        List of cleaned items without bullet points
    """
    if not text:
        return []

    # Replace ARM line breaks with standard newlines
    text = text.replace("_x000D_\n", "\n").replace("_x000D_", "\n")

    items = []
    for line in text.split("\n"):
        line = line.strip()
        # Remove bullet point prefix
        if line.startswith("•"):
            line = line[1:].strip()
        if line:
            items.append(line)

    return items


def get_sheet_rows(workbook, sheet_key: str, skip_header_rows: int = 2) -> list[dict]:
    """Extract rows from sheet by logical key (tries all FR/EN variants)."""
    sheet, sheet_name = find_sheet(workbook, sheet_key)
    if sheet is None:
        return []

    headers = [cell.value for cell in sheet[1]]
    rows = []
    for row in sheet.iter_rows(min_row=skip_header_rows + 1):
        row_data = {
            headers[i]: cell.value
            for i, cell in enumerate(row)
            if i < len(headers) and headers[i]
        }
        if any(row_data.values()):
            rows.append(row_data)
    return rows


def get_sheet_data(workbook, sheet_name: str, skip_header_rows: int = 2) -> list[dict]:
    """
    Extract data rows from an ARM Excel sheet.

    ARM sheets typically have:
    - Row 1: Main headers
    - Row 2: Sub-headers (for merged columns)
    - Row 3+: Data

    Args:
        workbook: openpyxl Workbook object
        sheet_name: Name of the sheet to read
        skip_header_rows: Number of header rows to skip (default 2)

    Returns:
        List of dictionaries with column name -> value mapping
    """
    if sheet_name not in workbook.sheetnames:
        logger.warning(f"Sheet '{sheet_name}' not found in workbook")
        return []

    sheet = workbook[sheet_name]

    # Get headers from first row
    headers = [cell.value for cell in sheet[1]]

    rows = []
    for row_idx, row in enumerate(
        sheet.iter_rows(min_row=skip_header_rows + 1), start=skip_header_rows + 1
    ):
        row_data = {}
        for i, cell in enumerate(row):
            if i < len(headers) and headers[i]:
                row_data[headers[i]] = cell.value

        # Skip empty rows
        if any(v for v in row_data.values() if v):
            rows.append(row_data)

    return rows


# =============================================================================
# Workshop 1 Processing
# =============================================================================


def build_gravity_scale_mapping(workbook) -> dict[str, int]:
    """Build mapping from gravity level names to numeric values (0-indexed)."""
    mapping = {}
    for row in get_sheet_rows(workbook, "impact_level_scale"):
        gravity_name = get_col(row, "gravity_level")
        niveau = get_col(row, "level")
        if gravity_name and niveau is not None:
            try:
                mapping[gravity_name.lower().strip()] = int(niveau) - 1
            except (ValueError, TypeError):
                logger.warning(f"Invalid Niveau value: {niveau}")
    return mapping


def extract_missions(workbook) -> str:
    """Extract missions from ARM file to use as study description."""
    missions = []
    for row in get_sheet_rows(workbook, "missions"):
        mission = get_col(row, "mission")
        if mission:
            missions.append(mission.strip())
    return "\n".join(missions)


def extract_primary_assets(workbook) -> list[dict]:
    """Extract primary assets from business values sheet."""
    assets = []
    for row in get_sheet_rows(workbook, "business_values"):
        name = get_col(row, "name")
        if not name:
            continue
        assets.append(
            {
                "name": name.strip(),
                "description": (get_col(row, "description") or "").strip(),
                "type": "PR",
                "supporting_asset_names": parse_bullet_list(
                    get_col(row, "supporting_assets") or ""
                ),
            }
        )
    return assets


def extract_supporting_assets(workbook) -> list[dict]:
    """Extract supporting assets from supporting assets sheet."""
    assets = []
    for row in get_sheet_rows(workbook, "supporting_assets"):
        name = get_col(row, "name")
        if not name:
            continue
        parent_name = (get_col(row, "parent_asset") or "").strip()
        assets.append(
            {
                "name": name.strip(),
                "description": (get_col(row, "description") or "").strip(),
                "type": "SP",
                "parent_name": parent_name if parent_name else None,
            }
        )
    return assets


def extract_feared_events(workbook, gravity_mapping: dict[str, int]) -> list[dict]:
    """Extract feared events from feared events sheet."""
    feared_events = []
    for row in get_sheet_rows(workbook, "feared_events"):
        name = get_col(row, "feared_event")
        if not name:
            continue

        gravity_text = get_col(row, "retained_gravity") or ""
        gravity = gravity_mapping.get(gravity_text.lower().strip(), -1)
        asset_names = parse_bullet_list(get_col(row, "business_values") or "")
        selected_value = get_col(row, "selected_check") or ""
        is_selected = "✔" in str(selected_value) or "✓" in str(selected_value)

        feared_events.append(
            {
                "name": name.strip(),
                "justification": (get_col(row, "impacts") or "").strip(),
                "gravity": gravity,
                "asset_names": asset_names,
                "is_selected": is_selected,
            }
        )
    return feared_events


def extract_applied_controls(workbook) -> list[dict]:
    """Extract applied controls from complementary measures sheet."""
    controls = []
    for row in get_sheet_rows(workbook, "complementary_measures"):
        name = get_col(row, "name")
        if not name:
            continue
        controls.append(
            {
                "name": name.strip(),
                "description": (get_col(row, "description") or "").strip(),
                "ref_id": (get_col(row, "ref_id") or "").strip(),
            }
        )
    return controls


# =============================================================================
# Workshop 2 Processing
# =============================================================================


def get_roto_sheet_data(workbook) -> list[dict]:
    """Extract data from RoTo couples sheet (uses row 2 sub-headers for merged columns)."""
    sheet, sheet_name = find_sheet(workbook, "roto_couples")
    if sheet is None:
        return []

    logger.info(f"[RoTo] Sheet '{sheet_name}' found, dimensions: {sheet.dimensions}")

    # RoTo sheet uses row 2 sub-headers (row 1 has merged headers)
    sub_headers = [cell.value for cell in sheet[2]]
    logger.info(f"[RoTo] Sub-headers: {sub_headers}")

    rows = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=3), start=3):
        row_data = {
            sub_headers[i]: cell.value
            for i, cell in enumerate(row)
            if i < len(sub_headers) and sub_headers[i]
        }
        if any(row_data.values()):
            rows.append(row_data)

    logger.info(f"[RoTo] Extracted {len(rows)} rows")
    return rows


def extract_roto_couples(workbook) -> list[dict]:
    """Extract RoTo (Risk Origin / Target Objective) couples from ARM file."""
    rows = get_roto_sheet_data(workbook)
    logger.info(f"[RoTo Extract] get_roto_sheet_data returned {len(rows)} rows")

    roto_couples = []
    for row_idx, row in enumerate(rows):
        logger.debug(f"[RoTo Extract] Row {row_idx}: keys={list(row.keys())}")

        risk_origin = get_col(row, "risk_origin")
        target_objective = get_col(row, "target_objective")

        if not risk_origin or not target_objective:
            logger.debug(
                f"[RoTo Extract] Row {row_idx}: SKIPPING - missing risk_origin or target_objective"
            )
            continue

        motivation = parse_plus_signs(get_col(row, "motivation") or "")
        resources = parse_plus_signs(get_col(row, "resources") or "")
        activity = parse_plus_signs(get_col(row, "activity") or "")

        retenu = get_col(row, "retained") or ""
        is_selected = "✔" in str(retenu) or "✓" in str(retenu)

        roto_couples.append(
            {
                "risk_origin": risk_origin.strip(),
                "target_objective": target_objective.strip(),
                "motivation": motivation,
                "resources": resources,
                "activity": activity,
                "is_selected": is_selected,
                "justification": (get_col(row, "justification") or "").strip(),
            }
        )
        logger.debug(
            f"[RoTo Extract] Row {row_idx}: ADDED '{risk_origin.strip()}' - '{target_objective.strip()}'"
        )

    logger.info(f"[RoTo Extract] Extracted {len(roto_couples)} RoTo couples")
    return roto_couples


# =============================================================================
# Workshop 3 Processing
# =============================================================================


def extract_stakeholder_categories(workbook) -> list[dict]:
    """Extract stakeholder categories from ARM file (maps to Terminology entries)."""
    categories = []
    for row in get_sheet_rows(workbook, "stakeholder_categories"):
        name = get_col(row, "name")
        if not name:
            continue
        categories.append(
            {
                "name": name.strip(),
                "normalized_name": normalize_category_name(name),
                "description": (get_col(row, "description") or "").strip(),
            }
        )
    logger.info(f"Extracted {len(categories)} stakeholder categories")
    return categories


def get_stakeholder_danger_data(workbook) -> dict[str, dict]:
    """Extract stakeholder danger/assessment data (uses row 2 sub-headers)."""
    sheet, sheet_name = find_sheet(workbook, "stakeholder_danger_levels")
    if sheet is None:
        return {}

    sub_headers = [cell.value for cell in sheet[2]]
    header_to_col = {h.strip(): i for i, h in enumerate(sub_headers) if h}

    # Column name variants for stakeholder name
    stakeholder_col_variants = ["Partie prenante", "Stakeholder", "Name"]

    danger_data = {}
    for row in sheet.iter_rows(min_row=3):
        row_values = [cell.value for cell in row]
        if not any(row_values):
            continue

        # Find stakeholder name
        stakeholder_name = None
        for variant in stakeholder_col_variants:
            if variant in header_to_col and header_to_col[variant] < len(row_values):
                stakeholder_name = row_values[header_to_col[variant]]
                if stakeholder_name:
                    break
        if not stakeholder_name and len(row_values) > 2:
            stakeholder_name = row_values[2]  # Fallback to column C
        if not stakeholder_name:
            continue

        assessment = {"dependency": 0, "penetration": 0, "maturity": 1, "trust": 1}

        # Extract assessment values (check FR and EN header variants)
        for header, col_idx in header_to_col.items():
            if col_idx >= len(row_values):
                continue
            value = row_values[col_idx]
            header_lower = header.lower()

            if any(
                v in header_lower for v in ["dépendance", "dependence", "dependency"]
            ):
                assessment["dependency"] = int(value) if value else 0
            elif any(v in header_lower for v in ["pénétration", "penetration"]):
                assessment["penetration"] = int(value) if value else 0
            elif any(v in header_lower for v in ["maturité", "maturity"]):
                assessment["maturity"] = max(1, int(value) if value else 1)
            elif any(v in header_lower for v in ["confiance", "trust"]):
                assessment["trust"] = max(1, int(value) if value else 1)

        danger_data[stakeholder_name.strip().lower()] = assessment

    logger.info(f"Extracted danger data for {len(danger_data)} stakeholders")
    return danger_data


def extract_stakeholders(workbook) -> list[dict]:
    """Extract stakeholders (entities) from ARM file."""
    danger_data = get_stakeholder_danger_data(workbook)

    stakeholders = []
    for row in get_sheet_rows(workbook, "stakeholders"):
        name = get_col(row, "name")
        if not name:
            continue

        name = name.strip()
        category_raw = get_col(row, "category") or ""
        assessment = danger_data.get(
            name.lower(), {"dependency": 0, "penetration": 0, "maturity": 1, "trust": 1}
        )
        risk_origins = parse_bullet_list(get_col(row, "risk_origins") or "")
        selected_value = get_col(row, "selected_check") or ""
        is_selected = "✔" in str(selected_value) or "✓" in str(selected_value)

        stakeholders.append(
            {
                "name": name,
                "description": (get_col(row, "description") or "").strip(),
                "category": normalize_category_name(category_raw),
                "category_raw": category_raw.strip(),
                "risk_origins": risk_origins,
                "is_selected": is_selected,
                "current_dependency": assessment["dependency"],
                "current_penetration": assessment["penetration"],
                "current_maturity": assessment["maturity"],
                "current_trust": assessment["trust"],
            }
        )

    logger.info(f"Extracted {len(stakeholders)} stakeholders")
    return stakeholders


def extract_strategic_scenarios(workbook) -> list[dict]:
    """Extract strategic scenarios from ARM file (uses row 2 sub-headers)."""
    sheet, sheet_name = find_sheet(workbook, "strategic_scenarios")
    if sheet is None:
        return []

    logger.info(f"[Strategic Scenarios] Sheet '{sheet_name}' found")

    headers_row1 = [cell.value for cell in sheet[1]]
    sub_headers = [cell.value for cell in sheet[2]]

    # Find column indices - check for FR and EN variants
    def find_col(variants):
        for i, h in enumerate(sub_headers):
            if h and any(v.lower() in h.lower() for v in variants):
                return i
        return None

    risk_origin_col = find_col(["Source de risque", "Risk origin"])
    target_objective_col = find_col(["Objectif visé", "Target objective"])
    attack_path_ref_col = find_col(["Réf.", "Ref.", "Short name"])

    logger.info(f"[Strategic Scenarios] Row 1: {headers_row1}")
    logger.info(f"[Strategic Scenarios] Row 2: {sub_headers}")
    logger.info(
        f"[Strategic Scenarios] Columns: risk_origin={risk_origin_col}, target_objective={target_objective_col}"
    )
    attack_path_name_col = None
    scenario_name_col = None
    scenario_ref_col = None

    # Find "Nom" column (appears multiple times - scenario name and attack path name)
    for i, h in enumerate(sub_headers):
        if h and h.strip().lower() == "nom":
            if attack_path_name_col is None:
                attack_path_name_col = i

    # Check row 1 for scenario name/ref columns
    for i, h in enumerate(headers_row1):
        if h:
            if "nom" in str(h).lower() or "name" in str(h).lower():
                scenario_name_col = i
            elif "abrév" in str(h).lower() or "abbrev" in str(h).lower():
                scenario_ref_col = i

    if scenario_name_col is None:
        scenario_name_col = 2  # Fallback to column C

    scenarios = []
    for row_num, row in enumerate(sheet.iter_rows(min_row=3), start=3):
        row_values = [cell.value for cell in row]
        if not any(row_values):
            continue

        def get_val(col_idx):
            return (
                row_values[col_idx]
                if col_idx is not None and col_idx < len(row_values)
                else None
            )

        scenario_name = get_val(scenario_name_col)
        if not scenario_name:
            continue

        scenarios.append(
            {
                "name": str(scenario_name).strip(),
                "ref_id": (str(get_val(scenario_ref_col) or "")).strip(),
                "risk_origin": (str(get_val(risk_origin_col) or "")).strip(),
                "target_objective": (str(get_val(target_objective_col) or "")).strip(),
                "attack_path_ref_id": (str(get_val(attack_path_ref_col) or "")).strip(),
                "attack_path_name": (str(get_val(attack_path_name_col) or "")).strip(),
            }
        )

    logger.info(f"[Strategic Scenarios] Extracted {len(scenarios)} scenarios")
    return scenarios


# =============================================================================
# Workshop 4 Processing
# =============================================================================


def extract_elementary_actions(workbook) -> list[dict]:
    """Extract elementary actions from ARM file."""
    elementary_actions = []
    for row in get_sheet_rows(workbook, "elementary_actions"):
        name = get_col(row, "name")
        if not name:
            continue
        elementary_actions.append(
            {
                "name": name.strip(),
                "description": (get_col(row, "description") or "").strip(),
                "ref_id": (get_col(row, "abbreviation") or "").strip(),
            }
        )
    logger.info(f"Extracted {len(elementary_actions)} elementary actions")
    return elementary_actions


# =============================================================================
# Main processing function
# =============================================================================


def process_arm_file(file_content: bytes) -> dict:
    """
    Process an ARM Excel file and extract all relevant data.

    Args:
        file_content: Raw bytes of the Excel file

    Returns:
        Dictionary with extracted data organized by type
    """
    workbook = load_workbook(BytesIO(file_content), data_only=True)

    # Build gravity mapping first (needed for feared events)
    gravity_mapping = build_gravity_scale_mapping(workbook)
    logger.info(
        f"Built gravity mapping with {len(gravity_mapping)} levels: {gravity_mapping}"
    )

    # Extract Workshop 1 data
    result = {
        "study_description": extract_missions(workbook),
        "primary_assets": extract_primary_assets(workbook),
        "supporting_assets": extract_supporting_assets(workbook),
        "feared_events": extract_feared_events(workbook, gravity_mapping),
        "applied_controls": extract_applied_controls(workbook),
        "gravity_mapping": gravity_mapping,
    }

    # Extract Workshop 2 data
    result["roto_couples"] = extract_roto_couples(workbook)

    # Extract Workshop 3 data
    result["stakeholder_categories"] = extract_stakeholder_categories(workbook)
    result["stakeholders"] = extract_stakeholders(workbook)
    result["strategic_scenarios"] = extract_strategic_scenarios(workbook)

    # Extract Workshop 4 data
    result["elementary_actions"] = extract_elementary_actions(workbook)

    logger.info(
        f"Extracted from ARM file: "
        f"{len(result['primary_assets'])} primary assets, "
        f"{len(result['supporting_assets'])} supporting assets, "
        f"{len(result['feared_events'])} feared events, "
        f"{len(result['applied_controls'])} applied controls, "
        f"{len(result['roto_couples'])} RoTo couples, "
        f"{len(result['stakeholders'])} stakeholders, "
        f"{len(result['strategic_scenarios'])} strategic scenarios, "
        f"{len(result['elementary_actions'])} elementary actions"
    )

    return result
