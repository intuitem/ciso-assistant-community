"""
Regression tests for the ARM importer against the sheet naming used by ARM
exports without the "Base de" prefix (SUP-1624).
"""

import io

import pytest
from openpyxl import Workbook

from data_wizard.arm_helpers import find_measure_sheets, find_sheet, process_arm_file

FR_SHEETS = {
    "1 - Missions": [
        ["", "✓/❒", "Mission", "Valeurs métier", "Paramètres", "Description"],
        [],
        ["1", "✔", "Fabriquer des vaccins", "", "", ""],
    ],
    "1 - Valeurs métier": [
        ["", "✓/❒", "Nom", "Abrév.", "Nature", "Biens supports", "Description"],
        [],
        ["1", "✔", "R&D", "VM01", "Processus", "• Serveurs", "Recherche"],
    ],
    "1 - Biens supports": [
        ["", "✓/❒", "Nom", "Abrév.", "Bien support parent", "Description"],
        [],
        ["1", "✔", "Serveurs", "BS01", "", "Serveurs bureautiques"],
    ],
    "1 - Échelle de gravité": [
        ["", "✓/❒", "Niveau", "Nom", "Description"],
        [],
        ["1", "✔", "1", "Mineure", ""],
        ["2", "✔", "2", "Significative", ""],
        ["3", "✔", "3", "Grave", ""],
    ],
    "1 - Évènements redoutés": [
        [
            "",
            "✓/❒",
            "Valeurs métier",
            "Évènement redouté",
            "Abrév.",
            "Impacts",
            "Gravité retenue",
        ],
        [],
        [
            "1",
            "✔",
            "• R&D",
            "Perte des données",
            "ER01",
            "• Impacts juridiques",
            "Grave",
        ],
    ],
    "1 - Référentiel de sécurité X": [
        [
            "",
            "✓/❒",
            "Réf.",
            "Nom",
            "Type",
            "Réduction de la gravité",
            "Description",
            "Statut",
        ],
        [],
        ["1", "✔", "M1", "Sensibilisation", "Gouvernance", "0 %", "", "En cours"],
    ],
}

EN_SHEETS = {
    "1 - Missions": [
        ["", "✓/❒", "Mission", "Business assets", "Parameters", "Description"],
        [],
        ["1", "✔", "Manufacture vaccines", "", "", ""],
    ],
    "1 - Business assets": [
        ["", "✓/❒", "Name", "Short name", "Nature", "Supporting assets", "Description"],
        [],
        ["1", "✔", "R&D", "BA01", "Process", "• Servers", "Research"],
    ],
    "1 - Supporting assets": [
        ["", "✓/❒", "Name", "Short name", "Parent supporting asset", "Description"],
        [],
        ["1", "✔", "Servers", "SA01", "", "Desktop servers"],
    ],
    "1 - Severity scale": [
        ["", "✓/❒", "Level", "Name", "Description"],
        [],
        ["1", "✔", "1", "Minor", ""],
        ["2", "✔", "2", "Significant", ""],
        ["3", "✔", "3", "Serious", ""],
    ],
    "1 - Feared events": [
        [
            "",
            "✓/❒",
            "Business assets",
            "Feared event",
            "Short name",
            "Impacts",
            "Retained severity",
        ],
        [],
        ["1", "✔", "• R&D", "Loss of data", "FE01", "• Legal impacts", "Serious"],
    ],
    "1 - X reference standard": [
        [
            "",
            "✓/❒",
            "Ref.",
            "Name",
            "Type",
            "Severity reduction",
            "Description",
            "Status",
        ],
        [],
        ["1", "✔", "M1", "Awareness", "Governance", "0 %", "", "In progress"],
    ],
}


def build_workbook(sheets: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(name)
        for row in rows:
            ws.append(row or [None])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


@pytest.fixture(params=["fr", "en"])
def arm_data(request):
    sheets = FR_SHEETS if request.param == "fr" else EN_SHEETS
    return process_arm_file(build_workbook(sheets))


class TestSheetResolution:
    @pytest.mark.parametrize(
        "sheets,key,expected",
        [
            (FR_SHEETS, "business_values", "1 - Valeurs métier"),
            (FR_SHEETS, "feared_events", "1 - Évènements redoutés"),
            (FR_SHEETS, "impact_level_scale", "1 - Échelle de gravité"),
            (EN_SHEETS, "business_values", "1 - Business assets"),
            (EN_SHEETS, "feared_events", "1 - Feared events"),
            (EN_SHEETS, "impact_level_scale", "1 - Severity scale"),
        ],
    )
    def test_sheet_found(self, sheets, key, expected):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(build_workbook(sheets)))
        assert find_sheet(wb, key)[1] == expected

    def test_truncated_sheet_name_is_matched(self):
        sheets = dict(FR_SHEETS)
        sheets["1 - Échelle de niveau d'impac"] = sheets.pop("1 - Échelle de gravité")
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(build_workbook(sheets)))
        assert (
            find_sheet(wb, "impact_level_scale")[1] == "1 - Échelle de niveau d'impac"
        )

    @pytest.mark.parametrize(
        "sheets,expected",
        [
            (FR_SHEETS, ["1 - Référentiel de sécurité X"]),
            (EN_SHEETS, ["1 - X reference standard"]),
        ],
    )
    def test_measure_sheets_detected_by_headers(self, sheets, expected):
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(build_workbook(sheets)))
        assert find_measure_sheets(wb) == expected


class TestExtraction:
    def test_primary_assets(self, arm_data):
        assert len(arm_data["primary_assets"]) == 1
        assert arm_data["primary_assets"][0]["supporting_asset_names"] != []

    def test_feared_events_linked_and_graded(self, arm_data):
        assert len(arm_data["feared_events"]) == 1
        event = arm_data["feared_events"][0]
        assert event["gravity"] == 2
        assert event["asset_names"] == ["R&D"]

    def test_applied_controls(self, arm_data):
        assert [c["ref_id"] for c in arm_data["applied_controls"]] == ["M1"]

    def test_workshop_1_sheets_all_resolved(self, arm_data):
        workshop_1 = {
            "business_values",
            "missions",
            "feared_events",
            "supporting_assets",
            "impact_level_scale",
            "security_measures",
        }
        assert workshop_1.isdisjoint(arm_data["missing_sheets"])


class TestMissingSheetReporting:
    def test_absent_sheets_are_reported(self):
        sheets = {k: v for k, v in FR_SHEETS.items() if k == "1 - Missions"}
        result = process_arm_file(build_workbook(sheets))
        assert "feared_events" in result["missing_sheets"]
        assert "security_measures" in result["missing_sheets"]
