"""
Helper functions for importing EBIOS RM studies from the Excel export format.
Sheet names are prefixed with workshop.activity numbers (e.g., "1.1 Study", "1.3 Feared Events").
"""

import re
from openpyxl import load_workbook
from io import BytesIO

# Sheet name patterns - match by prefix
SHEET_PREFIXES = {
    "study": "1.1",
    "assets": "1.2",
    "feared_events": "1.3",
    "compliance": "1.4",
    "ro_to": "2.1",
    "stakeholders": "3.1",
    "strategic_scenarios": "3.2.1",
    "attack_paths": "3.2.2",
    "stakeholder_controls": "3.3",
    "elementary_actions": "4.0",
    "operational_scenarios": "4.1.1",
    "operating_modes": "4.1.2",
}


def find_sheet_by_prefix(workbook, prefix: str):
    """Find a sheet by its prefix (e.g., '1.1' matches '1.1 Study')."""
    for sheet_name in workbook.sheetnames:
        if sheet_name.startswith(prefix):
            return workbook[sheet_name]
    return None


def find_sheets_by_prefix(workbook, prefix: str):
    """Find all sheets that start with a prefix (e.g., '1.4' for compliance assessments)."""
    sheets = []
    for sheet_name in workbook.sheetnames:
        if sheet_name.startswith(prefix):
            sheets.append((sheet_name, workbook[sheet_name]))
    return sheets


def get_sheet_rows(sheet):
    """Extract rows from a sheet as list of dicts with header keys."""
    if sheet is None:
        return []

    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    headers = [
        str(h).strip().lower().replace(" ", "_") if h else f"col_{i}"
        for i, h in enumerate(rows[0])
    ]
    result = []
    for row in rows[1:]:
        if all(cell is None for cell in row):
            continue
        row_dict = {}
        for i, cell in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = cell
        result.append(row_dict)
    return result


def parse_multiline(value):
    """Parse a multiline cell value into a list."""
    if not value:
        return []
    if isinstance(value, str):
        return [line.strip() for line in value.split("\n") if line.strip()]
    return [str(value)]


def extract_study(workbook) -> dict:
    """Extract study metadata from 1.1 sheet."""
    sheet = find_sheet_by_prefix(workbook, SHEET_PREFIXES["study"])
    rows = get_sheet_rows(sheet)
    if not rows:
        return {}
    row = rows[0]
    return {
        "ref_id": row.get("ref_id") or "",
        "name": row.get("name") or "",
        "description": row.get("description") or "",
        "version": row.get("version") or "",
        "status": row.get("status") or "",
        "observation": row.get("observation") or "",
    }


def extract_assets(workbook) -> list:
    """Extract assets from 1.2 sheet."""
    sheet = find_sheet_by_prefix(workbook, SHEET_PREFIXES["assets"])
    rows = get_sheet_rows(sheet)
    assets = []
    for row in rows:
        assets.append(
            {
                "name": row.get("name") or "",
                "description": row.get("description") or "",
                "type": row.get("type") or "",
                "parent_asset": row.get("parent_asset") or "",
            }
        )
    return assets


def extract_feared_events(workbook) -> list:
    """Extract feared events from 1.3 sheet."""
    sheet = find_sheet_by_prefix(workbook, SHEET_PREFIXES["feared_events"])
    rows = get_sheet_rows(sheet)
    feared_events = []
    for row in rows:
        feared_events.append(
            {
                "ref_id": row.get("ref_id") or "",
                "name": row.get("name") or "",
                "description": row.get("description") or "",
                "gravity": row.get("gravity"),
                "is_selected": row.get("is_selected") or False,
                "justification": row.get("justification") or "",
                "assets": parse_multiline(row.get("assets") or ""),
            }
        )
    return feared_events


def extract_compliance_assessments(workbook) -> list:
    """Extract compliance assessments from 1.4.x sheets."""
    sheets = find_sheets_by_prefix(workbook, "1.4")
    assessments = []
    for sheet_name, sheet in sheets:
        # Extract audit name from sheet name (e.g., "1.4.1 AuditName" -> "AuditName")
        match = re.match(r"1\.4\.\d+\s+(.+)", sheet_name)
        audit_name = match.group(1) if match else sheet_name
        rows = get_sheet_rows(sheet)
        requirements = []
        for row in rows:
            requirements.append(
                {
                    "urn": row.get("urn") or "",
                    "ref_id": row.get("ref_id") or "",
                    "name": row.get("name") or "",
                    "description": row.get("description") or "",
                    "result": row.get("result") or "",
                    "observation": row.get("observation") or "",
                    "applied_controls": parse_multiline(
                        row.get("applied_controls") or ""
                    ),
                }
            )
        assessments.append(
            {
                "name": audit_name,
                "requirements": requirements,
            }
        )
    return assessments


def extract_ro_to_couples(workbook) -> list:
    """Extract RO/TO couples from 2.1 sheet."""
    sheet = find_sheet_by_prefix(workbook, SHEET_PREFIXES["ro_to"])
    rows = get_sheet_rows(sheet)
    couples = []
    for row in rows:
        couples.append(
            {
                "risk_origin": row.get("risk_origin") or "",
                "target_objective": row.get("target_objective") or "",
                "motivation": row.get("motivation") or "",
                "resources": row.get("resources") or "",
                "activity": row.get("activity") or "",
                "pertinence": row.get("pertinence") or "",
                "is_selected": row.get("is_selected") or False,
                "justification": row.get("justification") or "",
                "feared_events": parse_multiline(row.get("feared_events") or ""),
            }
        )
    return couples


def extract_stakeholders(workbook) -> list:
    """Extract stakeholders from 3.1 sheet."""
    sheet = find_sheet_by_prefix(workbook, SHEET_PREFIXES["stakeholders"])
    rows = get_sheet_rows(sheet)
    stakeholders = []
    for row in rows:
        stakeholders.append(
            {
                "entity": row.get("entity") or "",
                "category": row.get("category") or "",
                "current_dependency": row.get("current_dependency"),
                "current_penetration": row.get("current_penetration"),
                "current_maturity": row.get("current_maturity"),
                "current_trust": row.get("current_trust"),
                "current_criticality": row.get("current_criticality") or "",
                "residual_dependency": row.get("residual_dependency"),
                "residual_penetration": row.get("residual_penetration"),
                "residual_maturity": row.get("residual_maturity"),
                "residual_trust": row.get("residual_trust"),
                "residual_criticality": row.get("residual_criticality") or "",
                "is_selected": row.get("is_selected") or False,
                "justification": row.get("justification") or "",
            }
        )
    return stakeholders


def extract_strategic_scenarios(workbook) -> list:
    """Extract strategic scenarios from 3.2.1 sheet."""
    sheet = find_sheet_by_prefix(workbook, SHEET_PREFIXES["strategic_scenarios"])
    rows = get_sheet_rows(sheet)
    scenarios = []
    for row in rows:
        scenarios.append(
            {
                "ref_id": row.get("ref_id") or "",
                "name": row.get("name") or "",
                "description": row.get("description") or "",
                "risk_origin": row.get("risk_origin") or "",
                "target_objective": row.get("target_objective") or "",
                "gravity": row.get("gravity") or "",
            }
        )
    return scenarios


def extract_attack_paths(workbook) -> list:
    """Extract attack paths from 3.2.2 sheet."""
    sheet = find_sheet_by_prefix(workbook, SHEET_PREFIXES["attack_paths"])
    rows = get_sheet_rows(sheet)
    paths = []
    for row in rows:
        paths.append(
            {
                "ref_id": row.get("ref_id") or "",
                "name": row.get("name") or "",
                "description": row.get("description") or "",
                "strategic_scenario": row.get("strategic_scenario") or "",
                "stakeholders": parse_multiline(row.get("stakeholders") or ""),
                "is_selected": row.get("is_selected") or False,
                "justification": row.get("justification") or "",
            }
        )
    return paths


def extract_stakeholder_controls(workbook) -> list:
    """Extract stakeholder controls from 3.3 sheet."""
    sheet = find_sheet_by_prefix(workbook, SHEET_PREFIXES["stakeholder_controls"])
    rows = get_sheet_rows(sheet)
    controls = []
    for row in rows:
        controls.append(
            {
                "ref_id": row.get("ref_id") or "",
                "name": row.get("name") or "",
                "description": row.get("description") or "",
                "status": row.get("status") or "",
                "stakeholders": parse_multiline(row.get("stakeholders") or ""),
            }
        )
    return controls


def extract_elementary_actions(workbook) -> list:
    """Extract elementary actions from 4.0 sheet."""
    sheet = find_sheet_by_prefix(workbook, SHEET_PREFIXES["elementary_actions"])
    rows = get_sheet_rows(sheet)
    actions = []
    for row in rows:
        actions.append(
            {
                "ref_id": row.get("ref_id") or "",
                "name": row.get("name") or "",
                "description": row.get("description") or "",
                "attack_stage": row.get("attack_stage") or "",
                "icon": row.get("icon") or "",
            }
        )
    return actions


def extract_operational_scenarios(workbook) -> list:
    """Extract operational scenarios from 4.1.1 sheet."""
    sheet = find_sheet_by_prefix(workbook, SHEET_PREFIXES["operational_scenarios"])
    rows = get_sheet_rows(sheet)
    scenarios = []
    for row in rows:
        scenarios.append(
            {
                "ref_id": row.get("ref_id") or "",
                "name": row.get("name") or "",
                "attack_path": row.get("attack_path") or "",
                "likelihood": row.get("likelihood") or "",
                "gravity": row.get("gravity") or "",
                "risk_level": row.get("risk_level") or "",
                "operating_modes_description": row.get("operating_modes_description")
                or "",
                "is_selected": row.get("is_selected") or False,
                "justification": row.get("justification") or "",
            }
        )
    return scenarios


def extract_operating_modes(workbook) -> list:
    """Extract operating modes from 4.1.2 sheet."""
    sheet = find_sheet_by_prefix(workbook, SHEET_PREFIXES["operating_modes"])
    rows = get_sheet_rows(sheet)
    modes = []
    for row in rows:
        modes.append(
            {
                "ref_id": row.get("ref_id") or "",
                "name": row.get("name") or "",
                "description": row.get("description") or "",
                "operational_scenario": row.get("operational_scenario") or "",
                "likelihood": row.get("likelihood") or "",
                "elementary_actions": parse_multiline(
                    row.get("elementary_actions") or ""
                ),
            }
        )
    return modes


def process_excel_file(file_content: bytes) -> dict:
    """
    Process an EBIOS RM Excel export file and extract all data.
    Returns a dictionary with all extracted data organized by section.
    """
    workbook = load_workbook(filename=BytesIO(file_content), data_only=True)

    return {
        "study": extract_study(workbook),
        "assets": extract_assets(workbook),
        "feared_events": extract_feared_events(workbook),
        "compliance_assessments": extract_compliance_assessments(workbook),
        "ro_to_couples": extract_ro_to_couples(workbook),
        "stakeholders": extract_stakeholders(workbook),
        "strategic_scenarios": extract_strategic_scenarios(workbook),
        "attack_paths": extract_attack_paths(workbook),
        "stakeholder_controls": extract_stakeholder_controls(workbook),
        "elementary_actions": extract_elementary_actions(workbook),
        "operational_scenarios": extract_operational_scenarios(workbook),
        "operating_modes": extract_operating_modes(workbook),
    }
