import csv
import io
import json

RESULT_VALUES = {"pass", "fail", "not_applicable", "error", "not_checked"}

# OCSF Compliance Finding (class_uid 2003) status mapping.
# Warning = needs human review, Unknown = not measured -> both land on not_checked.
OCSF_STATUS_MAP = {
    "pass": "pass",
    "fail": "fail",
    "warning": "not_checked",
    "unknown": "not_checked",
}
OCSF_STATUS_ID_MAP = {1: "pass", 2: "not_checked", 3: "fail", 99: "not_checked"}


class ImportError_(Exception):
    pass


def _entry(row, extras):
    ref_id = (row.get("ref_id") or "").strip()
    result = (row.get("result") or "").strip().lower()
    if not ref_id or result not in RESULT_VALUES:
        extras["parse_errors"].append(row)
        return None
    return {
        "ref_id": ref_id,
        "result": result,
        "actual": (row.get("actual") or "").strip(),
        "expected": (row.get("expected") or "").strip(),
        "message": (row.get("message") or "").strip(),
    }


def parse_csv(file):
    extras = {"parse_errors": []}
    try:
        text = file.read().decode("utf-8-sig")
    except UnicodeDecodeError as e:
        raise ImportError_("file is not valid UTF-8 text") from e
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames or "ref_id" not in reader.fieldnames:
        raise ImportError_("missing header row with at least ref_id,result columns")
    entries = [
        entry
        for row in reader
        if any((value or "").strip() for value in row.values())
        and (entry := _entry(row, extras))
    ]
    return entries, extras


def parse_xlsx(file):
    from openpyxl import load_workbook

    extras = {"parse_errors": []}
    try:
        workbook = load_workbook(file, read_only=True, data_only=True)
    except Exception as e:
        raise ImportError_("file is not a valid xlsx workbook") from e
    sheet = workbook.worksheets[0]
    rows = sheet.iter_rows(values_only=True)
    try:
        header = [str(cell).strip().lower() if cell else "" for cell in next(rows)]
    except StopIteration:
        raise ImportError_("empty worksheet") from None
    if "ref_id" not in header:
        raise ImportError_("missing header row with at least ref_id,result columns")
    entries = []
    for values in rows:
        row = {
            header[i]: str(values[i]) if values[i] is not None else ""
            for i in range(min(len(header), len(values)))
        }
        if not any(value.strip() for value in row.values()):
            continue
        entry = _entry(row, extras)
        if entry:
            entries.append(entry)
    return entries, extras


def parse_ocsf(file):
    extras = {"parse_errors": [], "skipped_suppressed": 0, "skipped_other_class": 0}
    try:
        data = json.load(file)
    except (json.JSONDecodeError, UnicodeDecodeError) as e:
        raise ImportError_("file is not valid JSON") from e
    events = data if isinstance(data, list) else data.get("events")
    if not isinstance(events, list):
        raise ImportError_(
            "expected an array of OCSF events or an object with an 'events' array"
        )
    entries = []
    tool = ""
    for event in events:
        if not isinstance(event, dict):
            extras["parse_errors"].append(event)
            continue
        if event.get("class_uid") != 2003:
            extras["skipped_other_class"] += 1
            continue
        if str(event.get("status", "")).lower() == "suppressed":
            extras["skipped_suppressed"] += 1
            continue
        compliance = event.get("compliance") or {}
        status = str(compliance.get("status", "")).lower()
        result = OCSF_STATUS_MAP.get(status) or OCSF_STATUS_ID_MAP.get(
            compliance.get("status_id")
        )
        if result is None:
            extras["parse_errors"].append({"status": compliance.get("status")})
            continue
        if not tool:
            product = (event.get("metadata") or {}).get("product") or {}
            tool = " ".join(
                filter(None, [product.get("name"), product.get("version")])
            )[:100]
        message = compliance.get("status_detail") or event.get("message") or ""
        for ref in compliance.get("requirements") or []:
            entries.append(
                {
                    "ref_id": str(ref).strip(),
                    "result": result,
                    "actual": "",
                    "expected": "",
                    "message": str(message)[:2000],
                }
            )
    if tool:
        extras["tool"] = tool
    return entries, extras


PARSERS = {".csv": parse_csv, ".xlsx": parse_xlsx, ".json": parse_ocsf}


def parse_file(file):
    name = (file.name or "").lower()
    for suffix, parser in PARSERS.items():
        if name.endswith(suffix):
            return parser(file)
    raise ImportError_("unsupported file type (expected .csv, .xlsx or .json)")
