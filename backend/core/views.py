from django.db.utils import OperationalError, ProgrammingError
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import csv
import json
import mimetypes
import re
from django_filters.filterset import filterset_factory
from django_filters.utils import try_dbfield
import regex
import os
import uuid
import zipfile
import tempfile
from datetime import date, datetime, timedelta
from typing import Dict, Any, List, Tuple
import time
from django.db.models import (
    F,
    Count,
    Q,
    ExpressionWrapper,
    FloatField,
    IntegerField,
    Value,
    Min,
    Subquery,
    OuterRef,
    When,
    Case,
    Exists,
    ForeignKey,
    OneToOneField,
    ManyToManyField,
    QuerySet,
)
from django.db.models.functions import Greatest, Coalesce

from collections import defaultdict
import pytz
from uuid import UUID
from itertools import chain, cycle
import django_filters as df
from ciso_assistant.settings import (
    EMAIL_HOST,
    EMAIL_HOST_RESCUE,
)

import shutil
from pathlib import Path
import humanize

from wsgiref.util import FileWrapper

import pandas as pd
import io

import random
from django.db.models.functions import Lower

from docxtpl import DocxTemplate
from integrations.models import SyncMapping
from integrations.tasks import sync_object_to_integrations
from integrations.registry import IntegrationRegistry
from library.serializers import StoredLibrarySerializer
from webhooks.service import dispatch_webhook_event
from .generators import gen_audit_context

from django.utils import timezone
from django.utils.text import slugify
from django.utils.decorators import method_decorator
from django.views.decorators.cache import cache_page
from django.views.decorators.vary import vary_on_cookie
from django.core.cache import cache


from django.apps import apps
from django.contrib.auth.models import Permission
from django.conf import settings
from django.core.files.storage import default_storage

from django.db import models, transaction
from django.forms import ValidationError
from django.http import FileResponse, HttpResponse, StreamingHttpResponse
from django.middleware import csrf
from django.template.loader import render_to_string
from django.utils.functional import Promise
from django.shortcuts import get_object_or_404
from django_filters.rest_framework import DjangoFilterBackend
from iam.models import Folder, RoleAssignment, User, UserGroup
from rest_framework import filters, generics, permissions, status, viewsets
from django.utils.translation import gettext_lazy as _
from rest_framework.decorators import (
    action,
    api_view,
    permission_classes,
    renderer_classes,
)
from rest_framework.parsers import (
    FileUploadParser,
)
from rest_framework.renderers import JSONRenderer
from rest_framework.request import Request
from rest_framework.response import Response
from rest_framework.utils.serializer_helpers import ReturnDict
from rest_framework.views import APIView
from rest_framework.exceptions import NotFound, PermissionDenied


from weasyprint import HTML

from core.helpers import *
from core.models import (
    AppliedControl,
    ComplianceAssessment,
    RequirementMappingSet,
    RiskAssessment,
    RiskMatrix,
    RiskScenario,
    AssetClass,
)
from core.serializers import ComplianceAssessmentReadSerializer
from core.utils import (
    compare_schema_versions,
    _generate_occurrences,
    _create_task_dict,
)
from dateutil import relativedelta as rd

from ebios_rm.models import (
    EbiosRMStudy,
    FearedEvent,
    RoTo,
    StrategicScenario,
    Stakeholder,
    AttackPath,
)

from tprm.models import Entity

from .models import *
from .serializers import *

from .models import Severity
from . import dora

from serdes.utils import (
    get_domain_export_objects,
    import_export_serializer_class,
    topological_sort,
    build_dependency_graph,
    get_self_referencing_field,
    sort_objects_by_self_reference,
)
from serdes.serializers import ExportSerializer
from django.contrib.admin.utils import NestedObjects
from django.db import router
from global_settings.utils import ff_is_enabled

import structlog

logger = structlog.get_logger(__name__)

SHORT_CACHE_TTL = 2  # mn
MED_CACHE_TTL = 5  # mn
LONG_CACHE_TTL = 60  # mn


MAPPING_MAX_DEPTH = 3

SETTINGS_MODULE = __import__(os.environ.get("DJANGO_SETTINGS_MODULE"))
MODULE_PATHS = SETTINGS_MODULE.settings.MODULE_PATHS


class NullableChoiceFilter(df.MultipleChoiceFilter):
    """
    A filter that supports filtering for null values using '--' as a special value.
    When '--' is in the filter values, it matches records where the field is null.
    """

    def __init__(self, *args, **kwargs):
        # Add "--" to the choices
        if "choices" in kwargs:
            original_choices = kwargs["choices"]
            # Add ("--", "--") to the beginning of choices
            kwargs["choices"] = [("--", "--")] + list(original_choices)
        super().__init__(*args, **kwargs)

    def filter(self, qs, value):
        if not value:
            return qs

        # Convert single value to list if needed
        if not isinstance(value, list):
            value = [value]

        # Separate "--" (null) from other actual values
        has_null = "--" in value
        real_values = [v for v in value if v != "--"]

        if has_null and real_values:
            # Both null and specific values: OR condition
            return qs.filter(
                Q(**{f"{self.field_name}__isnull": True})
                | Q(**{f"{self.field_name}__in": real_values})
            )
        elif has_null:
            # Only null values
            return qs.filter(**{f"{self.field_name}__isnull": True})
        elif real_values:
            # Only real values - filter by those values
            return qs.filter(**{f"{self.field_name}__in": real_values})
        else:
            # No valid values, return empty queryset
            return qs.none()


def add_unset_option(choices):
    """Add '--' (unset) option to choices dictionary or list"""
    # Handle both dict and list of tuples format
    if isinstance(choices, dict):
        # For dict format {value: label}, prepend with "--": "--"
        return {"--": "--", **choices}
    else:
        # For list of tuples like [(value, label), ...]
        return [("--", "--")] + list(choices)


def get_mapping_max_depth():
    """Get mapping max depth from general settings at runtime; safe during migrations."""
    try:
        gs = GlobalSettings.objects.filter(name="general").only("value").first()
        if not gs or not isinstance(getattr(gs, "value", None), dict):
            return MAPPING_MAX_DEPTH
        raw = gs.value.get("mapping_max_depth", MAPPING_MAX_DEPTH)
        try:
            val = int(raw)
        except (TypeError, ValueError):
            return MAPPING_MAX_DEPTH
        # Clamp to UI constraints
        return max(2, min(5, val))
    except (OperationalError, ProgrammingError):
        # DB not ready (e.g., migrate, makemigrations)
        return MAPPING_MAX_DEPTH


def escape_excel_formula(value):
    """
    Escape Excel formula injection by prefixing dangerous characters.
    Prevents CSV/Formula injection (OWASP) when values start with =+-@
    """
    if value is None:
        return ""
    s = str(value)
    return "'" + s if s and s[0] in "=" else s


def create_xlsx_response(entries, filename, wrap_columns=None):
    """
    DRY helper to create XLSX response with consistent formatting.

    Args:
        entries: List of dictionaries with data
        filename: Output filename
        wrap_columns: List of column names to wrap text (default: ["name", "description"])

    Returns:
        HttpResponse with XLSX file
    """
    if wrap_columns is None:
        wrap_columns = ["name", "description"]

    df = pd.DataFrame(entries)
    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)
        worksheet = writer.sheets["Sheet1"]

        wrap_indices = [
            df.columns.get_loc(col) + 1 for col in wrap_columns if col in df.columns
        ]

        for col_idx in wrap_indices:
            for row_idx in range(2, len(df) + 2):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(wrap_text=True)

        for idx, col in enumerate(df.columns):
            column_width = 20
            if col in wrap_columns:
                column_width = 40
            worksheet.column_dimensions[
                worksheet.cell(row=1, column=idx + 1).column_letter
            ].width = column_width

    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


class ExportMixin:
    """
    Generic export mixin for CSV/XLSX exports.
    ViewSets define export_config with fields, formatting, and query optimization hints.
    """

    export_config = None

    def _get_export_queryset(self):
        """Get filtered queryset with permissions applied."""
        (viewable_ids, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), self.request.user, self.model
        )
        queryset = self.model.objects.filter(id__in=viewable_ids)

        if self.export_config:
            if self.export_config.get("select_related"):
                queryset = queryset.select_related(
                    *self.export_config["select_related"]
                )
            if self.export_config.get("prefetch_related"):
                queryset = queryset.prefetch_related(
                    *self.export_config["prefetch_related"]
                )

        return queryset

    def _resolve_field_value(self, obj, field_config):
        """Resolve field value from object using dot notation or callable."""
        source = field_config.get("source")
        if not source:
            return ""

        parts = source.split(".")
        value = obj
        for part in parts:
            if value is None:
                return ""
            value = getattr(value, part, "")

        # Call methods but keep managers for format functions
        if callable(value):
            # Check if it's a Django manager/queryset (has .all() method)
            is_manager = hasattr(value, "all")
            # Call methods, but pass managers as-is
            if not is_manager:
                try:
                    value = value()
                except TypeError:
                    pass

        # Always apply if formatter exists, let formatter handle empties
        format_func = field_config.get("format")
        if format_func:
            value = format_func(value)

        if field_config.get("escape") and value:
            value = escape_excel_formula(value)

        return value if value is not None else ""

    @action(detail=False, name="Export as CSV")
    def export_csv(self, request):
        if not self.export_config:
            return HttpResponse(
                status=501, content="Export not configured for this model"
            )

        try:
            queryset = self._get_export_queryset()
            response = HttpResponse(content_type="text/csv")
            filename = f"{self.export_config.get('filename', 'export')}.csv"
            response["Content-Disposition"] = f'attachment; filename="{filename}"'

            writer = csv.writer(response, delimiter=";")
            fields = self.export_config["fields"]
            writer.writerow([f.get("label", name) for name, f in fields.items()])

            has_prefetch = bool(self.export_config.get("prefetch_related"))
            if has_prefetch:
                for obj in queryset:
                    row = [
                        self._resolve_field_value(obj, field_config)
                        for field_config in fields.values()
                    ]
                    writer.writerow(row)
            else:
                for obj in queryset.iterator():
                    row = [
                        self._resolve_field_value(obj, field_config)
                        for field_config in fields.values()
                    ]
                    writer.writerow(row)

            return response

        except Exception as e:
            logger.error(f"Error exporting {self.model.__name__} to CSV: {str(e)}")
            return HttpResponse(
                status=500, content="An error occurred while generating the CSV export."
            )

    @action(detail=False, name="Export as XLSX")
    def export_xlsx(self, request):
        if not self.export_config:
            return HttpResponse(
                status=501, content="Export not configured for this model"
            )

        queryset = self._get_export_queryset()
        entries = []
        fields = self.export_config["fields"]

        for obj in queryset:
            entry = {}
            for field_name, field_config in fields.items():
                entry[field_config.get("label", field_name)] = (
                    self._resolve_field_value(obj, field_config)
                )
            entries.append(entry)

        filename = f"{self.export_config.get('filename', 'export')}.xlsx"
        wrap_columns = self.export_config.get("wrap_columns", ["name", "description"])
        return create_xlsx_response(entries, filename, wrap_columns)

    def _create_multi_sheet_xlsx(self, queryset, main_fields, detail_config):
        """
        Create multi-sheet XLSX with main summary sheet and per-object detail sheets.

        Args:
            queryset: Main objects queryset
            main_fields: Fields config for summary sheet
            detail_config: Dict with 'fields', 'related_queryset_method', 'sheet_name_field'

        Returns:
            HttpResponse with XLSX file
        """
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create summary sheet
        summary_ws = wb.create_sheet(title="Summary")
        main_headers = [f.get("label", name) for name, f in main_fields.items()]
        summary_ws.append(main_headers)

        # Wrap text alignment for summary sheet
        wrap_alignment = Alignment(wrap_text=True, vertical="top")
        wrap_columns = self.export_config.get("wrap_columns", ["name", "description"])

        for col_idx, header in enumerate(main_headers, 1):
            if header.lower() in wrap_columns:
                for row_idx in range(1, summary_ws.max_row + 1):
                    summary_ws.cell(
                        row=row_idx, column=col_idx
                    ).alignment = wrap_alignment

        # Track sheet names to avoid duplicates
        used_sheet_names = {"Summary"}

        for obj in queryset:
            # Add row to summary sheet
            row = [
                self._resolve_field_value(obj, field_config)
                for field_config in main_fields.values()
            ]
            summary_ws.append(row)

            # Get related objects for detail sheet
            related_getter = detail_config.get("related_queryset_method")
            if related_getter:
                related_objects = getattr(obj, related_getter)()

                # Skip empty detail sheets if configured
                if not related_objects and detail_config.get("skip_empty", True):
                    continue

                # Create unique sheet name
                sheet_name_field = detail_config.get("sheet_name_field", "name")
                base_name = str(
                    self._resolve_field_value(obj, {"source": sheet_name_field})
                )
                if not base_name:
                    base_name = f"Object_{obj.pk}"

                # Truncate and ensure uniqueness
                sheet_name = base_name[:31]
                counter = 1
                while sheet_name in used_sheet_names:
                    suffix = f" ({counter})"
                    sheet_name = base_name[: 31 - len(suffix)] + suffix
                    counter += 1
                used_sheet_names.add(sheet_name)

                # Create detail sheet
                detail_ws = wb.create_sheet(title=sheet_name)
                detail_fields = detail_config.get("fields", {})
                detail_headers = [
                    f.get("label", name) for name, f in detail_fields.items()
                ]
                detail_ws.append(detail_headers)

                # Add related objects
                for related_obj in related_objects:
                    detail_row = [
                        self._resolve_field_value(related_obj, field_config)
                        for field_config in detail_fields.values()
                    ]
                    detail_ws.append(detail_row)

                # Apply wrap text to detail sheet
                for col_idx, header in enumerate(detail_headers, 1):
                    if header.lower() in wrap_columns:
                        for row_idx in range(1, detail_ws.max_row + 1):
                            detail_ws.cell(
                                row=row_idx, column=col_idx
                            ).alignment = wrap_alignment

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"{self.export_config.get('filename', 'export')}_multi_sheet.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class GenericFilterSet(df.FilterSet):
    @classmethod
    def filter_for_lookup(cls, field, lookup_type):
        DEFAULTS = dict(cls.FILTER_DEFAULTS)
        if hasattr(cls, "_meta"):
            DEFAULTS.update(cls._meta.filter_overrides)

        data = try_dbfield(DEFAULTS.get, field.__class__) or {}
        filter_class = data.get("filter_class")
        params = data.get("extra", lambda field: {})(field)

        # if there is no filter class, exit early
        if not filter_class:
            return None, {}

        # perform lookup specific checks
        if lookup_type == "exact" and getattr(field, "choices", None):
            return df.MultipleChoiceFilter, {"choices": field.choices, **params}

        if lookup_type == "isnull":
            data = try_dbfield(DEFAULTS.get, models.BooleanField)

            filter_class = data.get("filter_class")
            params = data.get("extra", lambda field: {})(field)
            return filter_class, params

        if lookup_type == "in":

            class ConcreteInFilter(df.BaseInFilter, filter_class):
                pass

            ConcreteInFilter.__name__ = cls._csv_filter_class_name(
                filter_class, lookup_type
            )

            return ConcreteInFilter, params

        if lookup_type == "range":

            class ConcreteRangeFilter(df.BaseRangeFilter, filter_class):
                pass

            ConcreteRangeFilter.__name__ = cls._csv_filter_class_name(
                filter_class, lookup_type
            )

            return ConcreteRangeFilter, params

        return filter_class, params

    class Meta:
        model = None  # This will be set dynamically via filterset_factory.
        filter_overrides = {
            models.ForeignKey: {
                "filter_class": df.ModelMultipleChoiceFilter,
                "extra": lambda f: {
                    "queryset": f.remote_field.model.objects.all(),
                },
            },
            models.ManyToManyField: {
                "filter_class": df.ModelMultipleChoiceFilter,
                "extra": lambda f: {
                    "queryset": f.remote_field.model.objects.all(),
                },
            },
        }


class BaseModelViewSet(viewsets.ModelViewSet):
    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    ordering = ["created_at"]
    ordering_fields = "__all__"
    search_fields = ["name", "description"]
    filterset_fields = []
    model = None

    serializers_module = "core.serializers"

    @property
    def filterset_class(self):
        # If you have defined filterset_fields, build the FilterSet on the fly.
        if self.filterset_fields:
            return filterset_factory(
                model=self.model,
                filterset=GenericFilterSet,
                fields=self.filterset_fields,
            )
        return None

    def get_queryset(self) -> models.query.QuerySet:
        """the scope_folder_id query_param allows scoping the objects to retrieve"""
        if not self.model:
            return None
        object_ids_view = None
        if self.request.method == "GET":
            if q := re.match(
                r"/api/[\w-]+/([\w-]+/)?([0-9a-fA-F]{8}(-[0-9a-fA-F]{4}){3}-[0-9a-fA-F]{12}(,[0-9a-fA-F]{8}(-[0-9a-fA-F]{4}){3}-[0-9a-fA-F]{12})+)",
                self.request.path,
            ):
                """"get_queryset is called by Django even for an individual object via get_object
                https://stackoverflow.com/questions/74048193/why-does-a-retrieve-request-end-up-calling-get-queryset"""
                id = UUID(q.group(1))
                if RoleAssignment.is_object_readable(self.request.user, self.model, id):
                    object_ids_view = [id]

        if not object_ids_view:
            scope_folder_id = self.request.query_params.get("scope_folder_id")
            scope_folder = (
                get_object_or_404(Folder, id=scope_folder_id)
                if scope_folder_id
                else Folder.get_root_folder()
            )
            object_ids_view = RoleAssignment.get_accessible_object_ids(
                scope_folder, self.request.user, self.model
            )[0]

        queryset = self.model.objects.filter(id__in=object_ids_view)
        return queryset

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context["action"] = self.action
        return context

    def get_serializer_class(self, **kwargs):
        serializer_factory = SerializerFactory(
            self.serializers_module, MODULE_PATHS.get("serializers", [])
        )
        serializer_class = serializer_factory.get_serializer(
            self.model.__name__, kwargs.get("action", self.action)
        )
        logger.debug(
            "Serializer class",
            serializer_class=serializer_class,
            action=kwargs.get("action", self.action),
            viewset=self,
            module_paths=MODULE_PATHS,
        )

        return serializer_class

    COMMA_SEPARATED_UUIDS_REGEX = r"^[0-9a-fA-F]{8}(-[0-9a-fA-F]{4}){3}-[0-9a-fA-F]{12}(,[0-9a-fA-F]{8}(-[0-9a-fA-F]{4}){3}-[0-9a-fA-F]{12})*$"

    def _process_request_data(self, request: Request) -> None:
        """
        Process the request data to split comma-separated UUIDs into a list
        and handle empty list scenarios.
        """
        for field in request.data:
            # NOTE: This is due to sveltekit-superforms not coercing the value into a list when
            # the form's dataType is "form", rather than "json".
            # Typically, dataType is "form" when the form contains a file input (e.g. for evidence attachments).
            # I am not ruling out the possibility that I am doing something wrong in the frontend. (Nassim)
            # TODO: Come back to this once superForms v2 is out of alpha. https://github.com/ciscoheat/sveltekit-superforms/releases
            if isinstance(request.data[field], list) and len(request.data[field]) == 1:
                if isinstance(request.data[field][0], str) and re.match(
                    self.COMMA_SEPARATED_UUIDS_REGEX, request.data[field][0]
                ):
                    request.data[field] = request.data[field][0].split(",")
                elif not request.data[field][0]:
                    request.data[field] = []

    def _process_labels(self, labels):
        """
        Creates a FilteringLabel and replaces the value with the ID of the newly created label.
        """
        new_labels = []
        for label in labels:
            try:
                uuid.UUID(label, version=4)
                new_labels.append(label)
            except ValueError:
                new_label = FilteringLabel(label=label)
                new_label.full_clean()
                new_label.save()
                new_labels.append(str(new_label.id))
        return new_labels

    def _process_evidences(self, evidences, folder):
        new_evidences = []
        for value in evidences or []:
            try:
                uuid.UUID(str(value), version=4)
                new_evidences.append(str(value))
            except ValueError:
                new_evidence = Evidence(name=value, folder=folder)
                new_evidence.full_clean()
                new_evidence.save()
                new_evidences.append(str(new_evidence.id))
        return new_evidences

    def list(self, request, *args, **kwargs):
        """
        Override the list method to inject optimized data into the serializer context.
        """
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        objects = page if page is not None else queryset
        # 1. Perform the bulk calculation for the current page (or entire set if not paginated)
        optimized_data = self._get_optimized_object_data(objects)
        # 2. Pass the data to the serializer via context
        context = self.get_serializer_context()
        context["optimized_data"] = optimized_data
        if page is not None:
            serializer = self.get_serializer(page, many=True, context=context)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True, context=context)
        return Response(serializer.data)

    def perform_create(self, serializer):
        instance = serializer.save()
        dispatch_webhook_event(instance, "created", serializer=serializer)
        return instance

    def perform_update(self, serializer):
        instance = serializer.save()
        dispatch_webhook_event(instance, "updated", serializer=serializer)
        return instance

    def create(self, request: Request, *args, **kwargs) -> Response:
        self._process_request_data(request)
        if request.data.get("filtering_labels"):
            request.data["filtering_labels"] = self._process_labels(
                request.data["filtering_labels"]
            )
        # Experimental: process evidences on TaskTemplate creation
        if request.data.get("evidences") and self.model == TaskTemplate:
            folder = Folder.objects.get(id=request.data.get("folder"))
            request.data["evidences"] = self._process_evidences(
                request.data.get("evidences"), folder=folder
            )
        return super().create(request, *args, **kwargs)

    def update(self, request: Request, *args, **kwargs) -> Response:
        # Experimental: process evidences on TaskTemplate update
        if request.data.get("evidences") and self.model == TaskTemplate:
            folder = Folder.objects.get(id=request.data.get("folder"))
            request.data["evidences"] = self._process_evidences(
                request.data["evidences"], folder=folder
            )

        # NOTE: Handle filtering_labels field - SvelteKit SuperForms behavior inconsistency:
        # Forms with file inputs (like Evidence attachments) use dataType="form" and omit empty fields
        # Forms without file inputs use dataType="json" and send empty arrays []
        # When the field is missing, we need to explicitly clear the labels by passing empty list to serializer
        if hasattr(self.model, "_meta") and "filtering_labels" in [
            f.name for f in self.model._meta.get_fields()
        ]:
            if "filtering_labels" in request.data:
                labels = request.data.get("filtering_labels")
                if labels:
                    # Make request.data mutable if needed (e.g., for multipart/form-data)
                    if hasattr(request.data, "_mutable"):
                        request.data._mutable = True
                    request.data["filtering_labels"] = self._process_labels(labels)
            else:
                # Field is missing entirely - add empty list to clear labels
                # Make request.data mutable if needed (e.g., for multipart/form-data)
                if hasattr(request.data, "_mutable"):
                    request.data._mutable = True
                # Use setlist() for QueryDict to properly set an empty list
                if hasattr(request.data, "setlist"):
                    request.data.setlist("filtering_labels", [])
                else:
                    request.data["filtering_labels"] = []

        return super().update(request, *args, **kwargs)

    def partial_update(self, request: Request, *args, **kwargs) -> Response:
        self._process_request_data(request)
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request: Request, *args, **kwargs) -> Response:
        self._process_request_data(request)
        instance = self.get_object()
        try:
            dispatch_webhook_event(instance, "deleted")
        except Exception:
            logger.error("Webhook dispatch failed on delete", exc_info=True)
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=["get"], url_path="cascade-info")
    def cascade_info(self, request, pk=None):
        """
        Cascade preview:
        - deleted: objects actually deleted by cascade
        - affected: objects not deleted but whose relationships will be removed (through rows, SET_NULL, local links)
        """
        instance = self.get_object()
        collector = NestedObjects(using=router.db_for_write(instance))
        collector.collect([instance])

        skip_model_names = {
            "Token",
            "AuthToken",
            "EmailAddress",
            "Session",
            "RequirementAssessment",
            "LogEntry",
            "Folder",
            "Permission",
            "LibraryFilteringLabel",
        }
        skip_model_classes = set()

        def is_hidden_model(model):
            return (skip_model_classes and model in skip_model_classes) or (
                not skip_model_classes and model.__name__ in skip_model_names
            )

        # Build index of concrete objects that will be deleted
        deleted_index = set()
        for model, objs in collector.model_objs.items():
            if model is type(instance):
                continue
            if getattr(model._meta, "auto_created", False):
                continue  # skip through rows here; we will bubble endpoints separately
            for o in objs:
                deleted_index.add((model.__name__, str(getattr(o, "pk", ""))))

        def add_grouped(bucket, obj):
            model = obj.__class__
            if is_hidden_model(model):
                return
            key = model.__name__
            pk = str(getattr(obj, "pk", "")) or ""
            if (key, pk) in bucket["_seen"]:
                return

            if key == "RoleAssignment":
                role_name = getattr(getattr(obj, "role", None), "name", "") or ""
                if getattr(obj, "user", None):
                    display = f"{role_name} → {getattr(obj.user, 'email', '')}"
                elif getattr(obj, "user_group", None):
                    display = f"{role_name} → {getattr(obj.user_group, 'name', '')}"
                else:
                    display = role_name
            else:
                display = (
                    getattr(obj, "name", None)
                    or getattr(obj, "email", None)
                    or str(obj)
                )

            group = bucket["by_model"].setdefault(
                key,
                {
                    "model": key,
                    "verbose_name": str(model._meta.verbose_name),
                    "objects": [],
                },
            )
            group["objects"].append({"id": pk, "name": display})
            bucket["_seen"].add((key, pk))

        def finalize(bucket):
            groups = list(bucket["by_model"].values())
            for g in groups:
                g["objects"].sort(key=lambda o: (o["name"] or "").lower())
            groups.sort(key=lambda g: (g.get("verbose_name") or g["model"]).lower())
            count = sum(len(g["objects"]) for g in groups)
            return groups, count

        deleted_bucket = {"by_model": {}, "_seen": set()}
        affected_bucket = {"by_model": {}, "_seen": set()}

        # 1) Concrete deletions
        through_rows = []
        for model, instances in collector.model_objs.items():
            if model is type(instance):
                continue
            if getattr(model._meta, "auto_created", False):
                through_rows.append((model, list(instances)))
                continue
            if is_hidden_model(model):
                continue
            for obj in instances:
                add_grouped(deleted_bucket, obj)

        # 2) Bubble endpoints from through rows (M2M join tables)
        for through_model, t_instances in through_rows:
            # Join tables have ≥2 FKs; skip others
            fk_fields = [
                f
                for f in through_model._meta.fields
                if isinstance(f, ForeignKey) and getattr(f, "remote_field", None)
            ]
            if len(fk_fields) < 2:
                continue

            for t in t_instances:
                endpoints = []
                for fk in fk_fields:
                    try:
                        rel_obj = getattr(t, fk.name, None)
                    except Exception:
                        rel_obj = None
                    if not rel_obj:
                        continue
                    if isinstance(rel_obj, type(instance)):
                        continue
                    if getattr(rel_obj._meta, "auto_created", False):
                        continue
                    endpoints.append(rel_obj)

                for ep in endpoints:
                    key = ep.__class__.__name__
                    pk = str(getattr(ep, "pk", "")) or ""
                    if (key, pk) in deleted_index:
                        add_grouped(deleted_bucket, ep)
                    else:
                        add_grouped(affected_bucket, ep)

        # 2b) Incoming updates (SET_NULL/DEFAULT) from Django's plan
        updates = getattr(collector, "field_updates", {})

        def _iter_objs(maybe_iter):
            if maybe_iter is None:
                return
            if isinstance(maybe_iter, QuerySet):
                for x in maybe_iter.iterator():
                    yield x
            else:
                for x in maybe_iter:
                    yield x

        for key, ops in updates.items():
            # Shape A: (field, value) -> [objs, ...]
            if isinstance(key, tuple) and len(key) == 2:
                field, value = key
                for objs in ops or []:
                    for o in _iter_objs(objs):
                        model = field.model  # model whose rows will be updated
                        pair = (model.__name__, str(getattr(o, "pk", "")) or "")
                        if pair not in deleted_index:
                            # If you want only SET_NULL cases: if value is None: ...
                            add_grouped(affected_bucket, o)
            # Shape B: model -> [(field, value, objs) ...]
            elif isinstance(ops, (list, tuple)) and ops and isinstance(ops[0], tuple):
                for item in ops:
                    if len(item) == 3:
                        field, value, objs = item
                    elif len(item) == 2:
                        field, objs = item
                        value = None
                    else:
                        continue
                    for o in _iter_objs(objs):
                        model = field.model
                        pair = (model.__name__, str(getattr(o, "pk", "")) or "")
                        if pair not in deleted_index:
                            add_grouped(affected_bucket, o)
            else:
                # Unknown shape; ignore safely
                pass

        # 2c) Outgoing links from the instance (FK/O2O/M2M)
        for f in instance._meta.get_fields():
            if getattr(f, "auto_created", False):
                continue  # skip reverse relations

            # Local FK/O2O targets: will remain but lose the link to this instance.
            if isinstance(f, (ForeignKey, OneToOneField)):
                try:
                    rel_obj = getattr(instance, f.name, None)
                except Exception:
                    rel_obj = None
                if rel_obj:
                    key = rel_obj.__class__.__name__
                    pk = str(getattr(rel_obj, "pk", "")) or ""
                    if (key, pk) not in deleted_index:
                        add_grouped(affected_bucket, rel_obj)

            # Local M2M targets: will remain; join rows are removed.
            elif isinstance(f, ManyToManyField):
                try:
                    manager = getattr(instance, f.name)
                    for rel_obj in manager.all():
                        key = rel_obj.__class__.__name__
                        pk = str(getattr(rel_obj, "pk", "")) or ""
                        if (key, pk) not in deleted_index:
                            add_grouped(affected_bucket, rel_obj)
                except Exception:
                    pass

        # 3) Sort and respond
        deleted_groups, deleted_count = finalize(deleted_bucket)
        affected_groups, affected_count = finalize(affected_bucket)

        def flatten(groups):
            return [
                {
                    "model": g["verbose_name"] or g["model"],
                    "name": o["name"],
                    "id": o["id"],
                }
                for g in groups
                for o in g["objects"]
            ]

        payload = {
            "deleted": {
                "count": deleted_count,
                "grouped_objects": deleted_groups,
                "related_objects": flatten(deleted_groups),
                "message": "The following objects will be permanently deleted.",
                "level": "warning",
            },
            "affected": {
                "count": affected_count,
                "grouped_objects": affected_groups,
                "related_objects": flatten(affected_groups),
                "message": "These objects will NOT be deleted but will lose one or more relationships.",
                "level": "info",
            },
        }
        return Response(payload)

    def _get_optimized_object_data(self, queryset):
        """
        Calculate folder full paths for objects in the queryset in 1 DB request.
        """
        initial_objects = list(queryset)
        if not initial_objects:
            return {}

        path_results = {}
        folders = {f.id: f for f in Folder.objects.all()}
        for obj in initial_objects:
            path = []
            if hasattr(obj, "folder"):
                queue = deque([obj.folder.id])
            elif hasattr(obj, "parent_folder") and obj.parent_folder:
                queue = deque([obj.parent_folder.id])
            else:
                continue
            while queue:
                folder_id = queue.popleft()
                folder = folders[folder_id]
                if folder.parent_folder:
                    path.append(
                        {
                            "str": str(folder),
                            "id": folder.id,
                            "parent_id": folder.parent_folder.id,
                        }
                    )
                    queue.append(folder.parent_folder.id)
            path_results[obj.id] = path[::-1]  # Reverse to get root to leaf order

        return {
            "paths": path_results,
        }

    class Meta:
        abstract = True

    @action(detail=True, name="Get write data")
    def object(self, request, pk):
        serializer_class = self.get_serializer_class(action="update")

        return Response(serializer_class(super().get_object()).data)


# Content types


class ContentTypeListView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    @method_decorator(cache_page(60 * MED_CACHE_TTL))
    def get(self, request, format=None):
        content_types = []
        for model in apps.get_models():
            content_types.append(
                {"label": model.__name__, "value": model._meta.model_name}
            )
        content_types.sort(key=lambda x: x["label"].lower())
        return Response(content_types)


# Risk Assessment


class PerimeterFilter(GenericFilterSet):
    folder = df.ModelMultipleChoiceFilter(
        queryset=Folder.objects.all(),
    )
    lc_status = df.MultipleChoiceFilter(
        choices=Perimeter.PRJ_LC_STATUS, lookup_expr="icontains"
    )

    class Meta:
        model = Perimeter
        fields = ["name", "folder", "lc_status", "campaigns"]


class PerimeterViewSet(BaseModelViewSet):
    """
    API endpoint that allows perimeters to be viewed or edited.
    """

    model = Perimeter
    filterset_class = PerimeterFilter
    search_fields = ["name", "ref_id", "description"]
    filterset_fields = ["name", "folder", "campaigns"]

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get status choices")
    def lc_status(self, request):
        return Response(dict(Perimeter.PRJ_LC_STATUS))

    @action(detail=False, methods=["get"])
    def names(self, request):
        uuid_list = request.query_params.getlist("id[]", [])
        queryset = Perimeter.objects.filter(id__in=uuid_list)

        return Response({str(perimeter.id): perimeter.name for perimeter in queryset})

    @action(detail=False, methods=["get"])
    def quality_check(self, request):
        """
        Returns the quality check of the perimeters
        """
        (viewable_objects, _, _) = RoleAssignment.get_accessible_object_ids(
            folder=Folder.get_root_folder(), user=request.user, object_type=Perimeter
        )
        perimeters = Perimeter.objects.filter(id__in=viewable_objects)
        res = {
            str(p.id): {
                "perimeter": PerimeterReadSerializer(p).data,
                "compliance_assessments": {"objects": {}},
                "risk_assessments": {"objects": {}},
            }
            for p in perimeters
        }
        for compliance_assessment in ComplianceAssessment.objects.filter(
            perimeter__in=perimeters
        ):
            res[str(compliance_assessment.perimeter.id)]["compliance_assessments"][
                "objects"
            ][str(compliance_assessment.id)] = {
                "object": ComplianceAssessmentReadSerializer(
                    compliance_assessment
                ).data,
                "quality_check": compliance_assessment.quality_check(),
            }
        for risk_assessment in RiskAssessment.objects.filter(perimeter__in=perimeters):
            res[str(risk_assessment.perimeter.id)]["risk_assessments"]["objects"][
                str(risk_assessment.id)
            ] = {
                "object": RiskAssessmentReadSerializer(risk_assessment).data,
                "quality_check": risk_assessment.quality_check(),
            }
        return Response({"results": res})

    @action(detail=True, methods=["get"], url_path="quality_check")
    def quality_check_detail(self, request, pk):
        """
        Returns the quality check of the perimeter
        """
        (viewable_objects, _, _) = RoleAssignment.get_accessible_object_ids(
            folder=Folder.get_root_folder(), user=request.user, object_type=Perimeter
        )
        if UUID(pk) in viewable_objects:
            perimeter = self.get_object()
            res = {
                "perimeter": PerimeterReadSerializer(perimeter).data,
                "compliance_assessments": {"objects": {}},
                "risk_assessments": {"objects": {}},
            }
            for compliance_assessment in ComplianceAssessment.objects.filter(
                perimeter=perimeter
            ):
                res["compliance_assessments"]["objects"][
                    str(compliance_assessment.id)
                ] = {
                    "object": ComplianceAssessmentReadSerializer(
                        compliance_assessment
                    ).data,
                    "quality_check": compliance_assessment.quality_check(),
                }
            for risk_assessment in RiskAssessment.objects.filter(perimeter=perimeter):
                res["risk_assessments"]["objects"][str(risk_assessment.id)] = {
                    "object": RiskAssessmentReadSerializer(risk_assessment).data,
                    "quality_check": risk_assessment.quality_check(),
                }
            return Response(res)
        else:
            return Response(status=status.HTTP_403_FORBIDDEN)

    @action(detail=False, methods=["get"])
    def ids(self, request):
        my_map = dict()

        (viewable_items, _, _) = RoleAssignment.get_accessible_object_ids(
            folder=Folder.get_root_folder(),
            user=request.user,
            object_type=Perimeter,
        )
        for item in Perimeter.objects.filter(id__in=viewable_items):
            if my_map.get(item.folder.name) is None:
                my_map[item.folder.name] = {}
            my_map[item.folder.name].update({item.name: item.id})

        return Response(my_map)


class ThreatViewSet(BaseModelViewSet):
    """
    API endpoint that allows threats to be viewed or edited.
    """

    model = Threat
    filterset_fields = [
        "folder",
        "provider",
        "library",
        "risk_scenarios",
        "filtering_labels",
    ]
    search_fields = ["name", "provider", "description"]

    def get_queryset(self):
        return (
            super()
            .get_queryset()
            .select_related(
                "folder",
                "folder__parent_folder",  # For get_folder_full_path() optimization
                "library",  # FieldsRelatedField includes library
            )
            .prefetch_related(
                "filtering_labels__folder",  # FieldsRelatedField includes folder
            )
        )

    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)

    @action(detail=False, name="Get provider choices")
    def provider(self, request):
        providers = set(
            Threat.objects.filter(provider__isnull=False).values_list(
                "provider", flat=True
            )
        )
        return Response({p: p for p in providers})

    @action(detail=False, name="Get threats count")
    def threats_count(self, request):
        folder_id = request.query_params.get("folder", None)
        return Response({"results": threats_count_per_name(request.user, folder_id)})

    @action(detail=False, methods=["get"])
    def ids(self, request):
        my_map = dict()

        (viewable_items, _, _) = RoleAssignment.get_accessible_object_ids(
            folder=Folder.get_root_folder(),
            user=request.user,
            object_type=Threat,
        )
        for item in Threat.objects.filter(id__in=viewable_items):
            if my_map.get(item.folder.name) is None:
                my_map[item.folder.name] = {}
            my_map[item.folder.name].update({item.name: item.id})
        return Response(my_map)


class AssetFilter(GenericFilterSet):
    folder = df.ModelMultipleChoiceFilter(queryset=Folder.objects.all())
    asset_class = df.ModelMultipleChoiceFilter(queryset=AssetClass.objects.all())

    exclude_children = df.ModelChoiceFilter(
        queryset=Asset.objects.all(),
        method="filter_exclude_children",
        label="Exclude children",
    )
    exclude_parents = df.ModelChoiceFilter(
        queryset=Asset.objects.all(),
        method="filter_exclude_parents",
        label="Exclude parents",
    )

    def filter_exclude_children(self, queryset, name, value):
        descendants = value.get_descendants()
        return queryset.exclude(id__in=[descendant.id for descendant in descendants])

    def filter_exclude_parents(self, queryset, name, value):
        ancestors = value.ancestors_plus_self()
        return queryset.exclude(id__in=[ancestor.id for ancestor in ancestors])

    class Meta:
        model = Asset
        fields = [
            "name",
            "type",
            "parent_assets",
            "exclude_children",
            "exclude_parents",
            "ebios_rm_studies",
            "risk_scenarios",
            "security_exceptions",
            "applied_controls",
            "filtering_labels",
            "personal_data",
            "is_business_function",
            "dora_licenced_activity",
            "dora_criticality_assessment",
            "dora_discontinuing_impact",
        ]


class AssetCapabilityViewSet(BaseModelViewSet):
    model = AssetCapability
    search_fields = ["name"]


class AssetViewSet(ExportMixin, BaseModelViewSet):
    """
    API endpoint that allows assets to be viewed or edited.
    """

    model = Asset
    filterset_class = AssetFilter
    search_fields = ["name", "description", "ref_id"]
    ordering = ["folder__name", "name"]

    def get_queryset(self) -> models.query.QuerySet:
        return (
            super()
            .get_queryset()
            .select_related("asset_class", "folder")
            .prefetch_related(
                "parent_assets",
                "child_assets",
                "owner",
                "security_exceptions",
                "filtering_labels",
                "personal_data",
                "overridden_children_capabilities",
            )
        )

    def _get_optimized_object_data(self, queryset):
        """
        Extends the base optimization to add asset-specific data for objectives
        and descendants, ensuring maximum query efficiency.
        """
        optimized_data = super()._get_optimized_object_data(queryset)
        initial_assets = list(queryset)
        if not initial_assets:
            return optimized_data

        graph_data = Asset._prefetch_graph_data(initial_assets)
        child_to_parents = graph_data["child_to_parents"]
        parent_to_children = graph_data["parent_to_children"]

        scale = Asset._get_security_objective_scale()
        sec_obj_results = {}
        dro_obj_results = {}
        sec_cap_results = {}
        rec_cap_results = {}
        descendant_results = {}

        for asset in initial_assets:
            # Calculate and store descendants.
            descendants = Asset._get_all_descendants(asset, parent_to_children)
            descendant_results[asset.id] = [
                {"id": str(d.id), "str": str(d)} for d in descendants
            ]

            # Calculate and store objectives.
            if asset.is_primary:
                sec_obj = asset.security_objectives.get("objectives", {})
                dro_obj = asset.disaster_recovery_objectives.get("objectives", {})

                # For primary assets, aggregate capabilities from supporting descendants
                supporting_descendants = {d for d in descendants if not d.is_primary}
                sec_cap = Asset._aggregate_security_capabilities(
                    supporting_descendants, asset
                )
                rec_cap = Asset._aggregate_recovery_capabilities(
                    supporting_descendants, asset
                )
            else:
                ancestors = Asset._get_all_ancestors(asset, child_to_parents)
                primary_ancestors = {anc for anc in ancestors if anc.is_primary}
                sec_obj = Asset._aggregate_security_objectives(primary_ancestors)
                dro_obj = Asset._aggregate_dro_objectives(primary_ancestors)

                # For supporting assets, use stored capabilities
                sec_cap = asset.security_capabilities.get("objectives", {})
                rec_cap = asset.recovery_capabilities.get("objectives", {})

            sec_obj_results[asset.id] = self._format_security_objectives(sec_obj, scale)
            dro_obj_results[asset.id] = self._format_disaster_recovery_objectives(
                dro_obj
            )
            sec_cap_results[asset.id] = self._format_security_objectives(sec_cap, scale)
            rec_cap_results[asset.id] = self._format_disaster_recovery_objectives(
                rec_cap
            )

        optimized_data.update(
            {
                "security_objectives": sec_obj_results,
                "disaster_recovery_objectives": dro_obj_results,
                "security_capabilities": sec_cap_results,
                "recovery_capabilities": rec_cap_results,
                "descendants": descendant_results,
            }
        )
        return optimized_data

    def _format_security_objectives(self, objectives, scale):
        if not objectives:
            return []
        return [
            {key: Asset.SECURITY_OBJECTIVES_SCALES[scale][content.get("value", 0)]}
            for key, content in sorted(
                objectives.items(),
                key=lambda x: Asset.DEFAULT_SECURITY_OBJECTIVES.index(x[0])
                if x[0] in Asset.DEFAULT_SECURITY_OBJECTIVES
                else -1,
            )
            if content.get("is_enabled", False)
            and content.get("value", -1) in range(0, 5)
        ]

    def _format_disaster_recovery_objectives(self, objectives):
        if not objectives:
            return []

        def format_seconds(s: int) -> str:
            if not isinstance(s, int) or s < 0:
                return "0s"
            h, r = divmod(s, 3600)
            m, s_rem = divmod(r, 60)
            parts = []
            if h > 0:
                parts.append(f"{h}h")
            if m > 0:
                parts.append(f"{m:02d}m")
            if s_rem > 0 or not parts:
                parts.append(f"{s_rem:02d}s")
            return "".join(parts)

        return [
            {"str": f"{key.upper()}: {format_seconds(content.get('value', 0))}"}
            for key, content in sorted(
                objectives.items(),
                key=lambda x: Asset.DEFAULT_DISASTER_RECOVERY_OBJECTIVES.index(x[0])
                if x[0] in Asset.DEFAULT_DISASTER_RECOVERY_OBJECTIVES
                else -1,
            )
            if content.get("value") is not None and content.get("value") > 0
        ]

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get type choices")
    def type(self, request):
        return Response(dict(Asset.Type.choices))

    @action(detail=False, name="Get asset class choices")
    def asset_class(self, request):
        # This endpoint is exclusively for filters.
        return Response(
            [
                {"id": ac.id, "name": ac.full_path}
                for ac in AssetClass.objects.filter(assets__isnull=False).distinct()
            ]
        )

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get DORA licensed activity choices")
    def dora_licenced_activity(self, request):
        return Response(dict(dora.DORA_LICENSED_ACTIVITY_CHOICES))

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get DORA criticality assessment choices")
    def dora_criticality_assessment(self, request):
        return Response(dict(dora.DORA_FUNCTION_CRITICALITY_CHOICES))

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get DORA discontinuing impact choices")
    def dora_discontinuing_impact(self, request):
        return Response(dict(dora.DORA_DISCONTINUING_IMPACT_CHOICES))

    @action(detail=True, name="Get asset write data")
    def object(self, request, pk):
        serializer_class = self.get_serializer_class(action="update")
        asset_data = serializer_class(super().get_object()).data

        general_settings = GlobalSettings.objects.filter(name="general").first()
        scale_key = (
            general_settings.value.get("security_objective_scale", "1-4")
            if general_settings
            else "1-4"
        )
        objective_scale = Asset.SECURITY_OBJECTIVES_SCALES[scale_key]

        objectives = asset_data["security_objectives"].get("objectives", {})
        security_capabilities = asset_data["security_capabilities"].get(
            "objectives", {}
        )
        reduced_objective_scale: dict[str, int] = {
            value: index
            for index, value in enumerate(objective_scale)
            if value not in objective_scale[:index]
        }

        for objective_dict in [objectives, security_capabilities]:
            for objective_data in objective_dict.values():
                if (objective_value := objective_data.get("value")) is None:
                    continue
                objective_label = objective_scale[objective_value]
                reduced_value = reduced_objective_scale[objective_label]
                objective_data["value"] = reduced_value

        return Response(asset_data)

    @action(detail=False, name="Get assets graph")
    def graph(self, request):
        nodes = []
        links = []
        nodes_idx = dict()
        categories = []
        N = 0
        hide_domains = (
            request.query_params.get("hide_domains", "false").lower() == "true"
        )

        (viewable_folders, _, _) = RoleAssignment.get_accessible_object_ids(
            folder=Folder.get_root_folder(),
            user=request.user,
            object_type=Folder,
        )
        (viewable_assets, _, _) = RoleAssignment.get_accessible_object_ids(
            folder=Folder.get_root_folder(),
            user=request.user,
            object_type=Asset,
        )
        # Build category index mapping first (key by UUID to avoid name collisions)
        domain_to_category = {}
        for domain in Folder.objects.filter(id__in=viewable_folders):
            categories.append({"name": domain.name})
            domain_to_category[domain.id] = len(categories) - 1

            if not hide_domains:
                nodes_idx[domain.id] = N
                nodes.append(
                    {
                        "name": domain.name,
                        "category": domain_to_category[domain.id],
                        "symbol": "roundRect",
                        "symbolSize": 30,
                        "value": "Domain",
                        "pk": str(domain.id),
                    }
                )
                N += 1
        for asset in Asset.objects.filter(id__in=viewable_assets):
            # Only include assets whose folders are also viewable to avoid KeyError
            if asset.folder.id not in viewable_folders:
                continue

            symbol = "circle"
            if asset.type == "PR":
                symbol = "diamond"
            # Use domain.name/asset.name as unique key to handle duplicate asset names across domains
            asset_key = f"{asset.folder.name}/{asset.name}"
            nodes.append(
                {
                    "name": asset.name,
                    "symbol": symbol,
                    "symbolSize": 25,
                    "category": domain_to_category[asset.folder.id],
                    "value": "Primary" if asset.type == "PR" else "Support",
                    "pk": str(asset.id),
                }
            )
            nodes_idx[asset_key] = N
            N += 1

        # Add links between domains (folders) based on parent-child relationships
        if not hide_domains:
            for domain in Folder.objects.filter(id__in=viewable_folders):
                if domain.parent_folder and domain.parent_folder.id in viewable_folders:
                    links.append(
                        {
                            "source": nodes_idx[domain.parent_folder.id],
                            "target": nodes_idx[domain.id],
                            "value": "contains",
                        }
                    )

        # Add links between assets and their domains
        if not hide_domains:
            for asset in Asset.objects.filter(id__in=viewable_assets):
                # Only include assets whose folders are also viewable to avoid KeyError
                if asset.folder.id not in viewable_folders:
                    continue

                asset_key = f"{asset.folder.name}/{asset.name}"
                links.append(
                    {
                        "source": nodes_idx[asset.folder.id],
                        "target": nodes_idx[asset_key],
                        "value": "contains",
                    }
                )

        # Add links between assets (existing relationships)
        for asset in Asset.objects.filter(id__in=viewable_assets):
            # Only include assets whose folders are also viewable to avoid KeyError
            if asset.folder.id not in viewable_folders:
                continue

            asset_key = f"{asset.folder.name}/{asset.name}"
            for relationship in asset.parent_assets.all():
                # Only include relationship if both assets and their folders are viewable
                if (
                    relationship.id not in viewable_assets
                    or relationship.folder.id not in viewable_folders
                ):
                    continue

                relationship_key = f"{relationship.folder.name}/{relationship.name}"
                links.append(
                    {
                        "source": nodes_idx[relationship_key],
                        "target": nodes_idx[asset_key],
                        "value": "depends on",
                    }
                )
        meta = {"display_name": "Assets Explorer"}
        return Response(
            {"nodes": nodes, "links": links, "categories": categories, "meta": meta}
        )

    @action(detail=False, methods=["get"])
    def ids(self, request):
        my_map = dict()

        (viewable_items, _, _) = RoleAssignment.get_accessible_object_ids(
            folder=Folder.get_root_folder(),
            user=request.user,
            object_type=Asset,
        )
        for item in Asset.objects.filter(id__in=viewable_items):
            if my_map.get(item.folder.name) is None:
                my_map[item.folder.name] = {}
            my_map[item.folder.name].update({item.name: item.id})
        return Response(my_map)

    @action(detail=False, name="Get security objectives")
    def security_objectives(self, request):
        return Response({"results": Asset.DEFAULT_SECURITY_OBJECTIVES})

    @action(detail=False, name="Get disaster recovery objectives")
    def disaster_recovery_objectives(self, request):
        return Response({"results": Asset.DEFAULT_DISASTER_RECOVERY_OBJECTIVES})

    export_config = {
        "fields": {
            "internal_id": {"source": "id", "label": "internal_id"},
            "name": {"source": "name", "label": "name", "escape": True},
            "description": {
                "source": "description",
                "label": "description",
                "escape": True,
            },
            "type": {"source": "type", "label": "type"},
            "folder": {"source": "folder.name", "label": "folder", "escape": True},
            "security_objectives": {
                "source": "get_security_objectives_display",
                "label": "security_objectives",
                "format": lambda objs: ",".join(
                    f"{k}: {v}" for obj in objs for k, v in obj.items()
                ),
                "escape": True,
            },
            "disaster_recovery_objectives": {
                "source": "get_disaster_recovery_objectives_display",
                "label": "disaster_recovery_objectives",
                "format": lambda objs: ",".join(i["str"] for i in objs),
                "escape": True,
            },
            "link": {"source": "reference_link", "label": "link", "escape": True},
            "owners": {
                "source": "owner",
                "label": "owners",
                "format": lambda qs: ",".join(
                    escape_excel_formula(o.email) for o in qs.all()
                ),
            },
            "parent_assets": {
                "source": "parent_assets",
                "label": "parent_assets",
                "format": lambda qs: ",".join(
                    escape_excel_formula(o.name) for o in qs.all()
                ),
            },
            "labels": {
                "source": "filtering_labels",
                "label": "labels",
                "format": lambda qs: ",".join(
                    escape_excel_formula(o.label) for o in qs.all()
                ),
            },
        },
        "filename": "assets_export",
        "select_related": ["folder"],
        "prefetch_related": ["owner", "parent_assets", "filtering_labels"],
    }

    @action(detail=False, methods=["post"], url_path="batch-create")
    def batch_create(self, request):
        """
        Batch create multiple assets from a text list with parent-child relationships.
        Expected format:
        {
            "assets_text": "PR:Parent Asset\\n  SP:Child Asset 1\\n  SP:Child Asset 2",
            "folder": "folder-uuid"
        }
        Lines with leading spaces (2+) are children of the previous non-indented line.
        """
        try:
            assets_text = request.data.get("assets_text", "")
            folder_id = request.data.get("folder")

            if not assets_text:
                return Response(
                    {"error": "assets_text is required"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            if not folder_id:
                return Response(
                    {"error": "folder is required"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Verify folder exists and user has access
            if not RoleAssignment.is_object_readable(request.user, Folder, folder_id):
                return Response(
                    {"error": "Folder not found"},
                    status=status.HTTP_404_NOT_FOUND,
                )
            folder = Folder.objects.get(id=folder_id)

            # Parse the assets text with indentation (2 spaces per level)
            lines = [line.rstrip() for line in assets_text.split("\n") if line.strip()]
            created_assets = []
            reused_assets = []
            errors = []
            depth_stack = []  # Stack of assets at each depth level

            for line in lines:
                # Count leading spaces and calculate depth (2 spaces = 1 level)
                leading_spaces = len(line) - len(line.lstrip())
                depth = leading_spaces // 2
                line_content = line.strip()

                # Check for type prefix (SP: or PR:)
                asset_type = Asset.Type.SUPPORT  # default
                asset_name = line_content

                if line_content.upper().startswith("SP:"):
                    asset_type = Asset.Type.SUPPORT
                    asset_name = line_content[3:].strip()
                elif line_content.upper().startswith("PR:"):
                    asset_type = Asset.Type.PRIMARY
                    asset_name = line_content[3:].strip()

                if not asset_name:
                    errors.append({"line": line_content, "error": "Empty asset name"})
                    continue

                # Trim stack to current depth
                depth_stack = depth_stack[:depth]

                # Parent is the asset at the previous depth level (skip None entries from errors)
                parent_asset = None
                if depth_stack:
                    # Find the last non-None entry in the stack
                    for i in range(len(depth_stack) - 1, -1, -1):
                        if depth_stack[i] is not None:
                            parent_asset = depth_stack[i]
                            break

                # Check if asset already exists in the folder
                existing_asset = Asset.objects.filter(
                    name=asset_name, folder=folder_id
                ).first()

                if existing_asset:
                    # Reuse existing asset
                    asset = existing_asset

                    # Update parent relationship if needed and parent exists
                    if parent_asset and parent_asset not in asset.parent_assets.all():
                        asset.parent_assets.add(parent_asset)

                    # Add to stack for potential children
                    depth_stack.append(asset)

                    reused_assets.append(
                        {
                            "id": str(asset.id),
                            "name": asset.name,
                            "type": asset.get_type_display(),
                            "parent": parent_asset.name if parent_asset else None,
                            "depth": depth,
                        }
                    )
                else:
                    # Create new asset using the serializer to respect IAM
                    asset_data = {
                        "name": asset_name,
                        "type": asset_type,
                        "folder": folder_id,
                    }

                    # Add parent relationship if exists
                    if parent_asset:
                        asset_data["parent_assets"] = [parent_asset.id]

                    serializer = AssetWriteSerializer(
                        data=asset_data, context={"request": request}
                    )

                    if serializer.is_valid():
                        asset = serializer.save()

                        # Add to stack for potential children
                        depth_stack.append(asset)

                        created_assets.append(
                            {
                                "id": str(asset.id),
                                "name": asset.name,
                                "type": asset.get_type_display(),
                                "parent": parent_asset.name if parent_asset else None,
                                "depth": depth,
                            }
                        )
                    else:
                        # Error creating asset - add None to stack to maintain depth
                        depth_stack.append(None)
                        errors.append(
                            {
                                "line": line_content,
                                "errors": serializer.errors,
                            }
                        )

            return Response(
                {
                    "created": len(created_assets),
                    "reused": len(reused_assets),
                    "assets": created_assets,
                    "reused_assets": reused_assets,
                    "errors": errors,
                },
                status=status.HTTP_201_CREATED
                if created_assets or reused_assets
                else status.HTTP_400_BAD_REQUEST,
            )

        except Exception as e:
            logger.error(f"Error in batch asset creation: {str(e)}")
            return Response(
                {"error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class AssetClassViewSet(BaseModelViewSet):
    model = AssetClass
    filterset_fields = ["parent"]

    ordering = ["parent", "name"]
    search_fields = ["name", "description"]

    @action(detail=False, name="Get Asset Class Tree")
    def tree(self, request):
        return Response(AssetClass.build_tree())


class ReferenceControlViewSet(BaseModelViewSet):
    """
    API endpoint that allows reference controls to be viewed or edited.
    """

    model = ReferenceControl
    filterset_fields = [
        "folder",
        "category",
        "csf_function",
        "provider",
        "findings",
        "filtering_labels",
    ]
    search_fields = ["name", "description", "provider", "ref_id"]

    @action(detail=False, name="Get provider choices")
    def provider(self, request):
        providers = set(
            ReferenceControl.objects.filter(provider__isnull=False).values_list(
                "provider", flat=True
            )
        )
        return Response({p: p for p in providers})

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get category choices")
    def category(self, request):
        return Response(dict(ReferenceControl.CATEGORY))

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get function choices")
    def csf_function(self, request):
        return Response(dict(ReferenceControl.CSF_FUNCTION))


class RiskMatrixViewSet(BaseModelViewSet):
    """
    API endpoint that allows risk matrices to be viewed or edited.
    """

    model = RiskMatrix
    filterset_fields = ["folder", "is_enabled", "provider"]

    @action(detail=False)  # Set a name there
    def colors(self, request):
        return Response({"results": get_risk_color_ordered_list(request.user)})

    @action(detail=False, name="Get provider choices")
    def provider(self, request):
        providers = set(
            RiskMatrix.objects.filter(provider__isnull=False).values_list(
                "provider", flat=True
            )
        )
        return Response({p: p for p in providers})

    @action(detail=False, name="Get risk level choices")
    def risk(self, request):
        viewable_matrices: list[RiskMatrix] = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, RiskMatrix
        )[0]
        undefined = {-1: "--"}
        options = undefined
        for matrix in RiskMatrix.objects.filter(id__in=viewable_matrices):
            _choices = {
                i: name
                for i, name in enumerate(
                    x["name"] for x in matrix.json_definition["risk"]
                )
            }
            options = options | _choices
        res = [{"value": k, "label": v} for k, v in options.items()]
        return Response(res)

    @action(detail=False, name="Get impact choices")
    def impact(self, request):
        viewable_matrices: list[RiskMatrix] = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, RiskMatrix
        )[0]
        undefined = {-1: "--"}
        options = []
        for matrix in RiskMatrix.objects.filter(id__in=viewable_matrices):
            _choices = {
                i: name
                for i, name in enumerate(
                    x["name"] for x in matrix.json_definition["impact"]
                )
            }
            choices = undefined | _choices
            options = options | choices.items()
        return Response(
            [{k: v for k, v in zip(("value", "label"), o)} for o in options]
        )

    @action(detail=False, name="Get probability choices")
    def probability(self, request):
        viewable_matrices: list[RiskMatrix] = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, RiskMatrix
        )[0]
        undefined = {-1: "--"}
        options = []
        for matrix in RiskMatrix.objects.filter(id__in=viewable_matrices):
            _choices = {
                i: name
                for i, name in enumerate(
                    x["name"] for x in matrix.json_definition["probability"]
                )
            }
            choices = undefined | _choices
            options = options | choices.items()
        return Response(
            [{k: v for k, v in zip(("value", "label"), o)} for o in options]
        )

    @action(detail=False, name="Get used risk matrices")
    def used(self, request):
        viewable_matrices = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, RiskMatrix
        )[0]
        viewable_assessments = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, RiskAssessment
        )[0]
        _used_matrices = (
            RiskMatrix.objects.filter(riskassessment__isnull=False)
            .filter(id__in=viewable_matrices)
            .filter(riskassessment__id__in=viewable_assessments)
            .distinct()
        )
        used_matrices = _used_matrices.values("id", "name")
        for i in range(len(used_matrices)):
            used_matrices[i]["risk_assessments_count"] = (
                RiskAssessment.objects.filter(risk_matrix=_used_matrices[i].id)
                .filter(id__in=viewable_assessments)
                .count()
            )
        return Response({"results": used_matrices})

    @action(detail=False, methods=["get"])
    def ids(self, request):
        my_map = dict()

        (viewable_items, _, _) = RoleAssignment.get_accessible_object_ids(
            folder=Folder.get_root_folder(),
            user=request.user,
            object_type=RiskMatrix,
        )
        for item in RiskMatrix.objects.filter(id__in=viewable_items):
            if my_map.get(item.folder.name) is None:
                my_map[item.folder.name] = {}
            my_map[item.folder.name].update({item.name: item.id})

        return Response(my_map)


class VulnerabilityViewSet(BaseModelViewSet):
    """
    API endpoint that allows vulnerabilities to be viewed or edited.
    """

    model = Vulnerability
    filterset_fields = [
        "folder",
        "assets",
        "status",
        "severity",
        "risk_scenarios",
        "applied_controls",
        "security_exceptions",
        "filtering_labels",
        "findings",
    ]
    search_fields = ["name", "description", "ref_id"]

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get status choices")
    def status(self, request):
        return Response(dict(Vulnerability.Status.choices))

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get severity choices")
    def severity(self, request):
        return Response(dict(Severity.choices))

    @action(detail=False, methods=["get"], name="Get sankey data")
    def sankey_data(self, request):
        """
        Returns vulnerability data structured for Sankey diagram:
        Folders -> Severity -> Status (as links)
        """
        folder_id = request.query_params.get("folder", None)

        # Get viewable vulnerabilities
        scoped_folder = (
            Folder.objects.get(id=folder_id) if folder_id else Folder.get_root_folder()
        )
        (object_ids, _, _) = RoleAssignment.get_accessible_object_ids(
            scoped_folder, request.user, Vulnerability
        )

        vulnerabilities = Vulnerability.objects.filter(
            id__in=object_ids
        ).select_related("folder")

        # Build link structure for Sankey: folder -> severity -> status
        links = []
        folder_severity_counts = {}
        severity_status_counts = {}

        for vuln in vulnerabilities:
            # Get folder name
            if vuln.folder:
                folder_name = vuln.folder.name
            else:
                folder_name = "No Folder"

            # Get severity label
            severity_value = vuln.severity
            severity_label = dict(Severity.choices).get(severity_value, "undefined")

            # Get status label
            status_value = vuln.status
            status_label = status_value if status_value else "--"

            # Count folder -> severity links
            folder_severity_key = f"{folder_name}||{severity_label}"
            folder_severity_counts[folder_severity_key] = (
                folder_severity_counts.get(folder_severity_key, 0) + 1
            )

            # Count severity -> status links
            severity_status_key = f"{severity_label}||{status_label}"
            severity_status_counts[severity_status_key] = (
                severity_status_counts.get(severity_status_key, 0) + 1
            )

        # Convert to Sankey link format
        for key, value in folder_severity_counts.items():
            folder, severity = key.split("||")
            links.append({"source": folder, "target": severity, "value": value})

        for key, value in severity_status_counts.items():
            severity, status_label = key.split("||")
            links.append({"source": severity, "target": status_label, "value": value})

        return Response(links)

    @action(detail=False, methods=["get"], name="Get treemap data")
    def treemap_data(self, request):
        """
        Returns vulnerability data structured for treemap visualization:
        Folders -> Severity -> Status
        """
        folder_id = request.query_params.get("folder", None)

        # Get viewable vulnerabilities
        scoped_folder = (
            Folder.objects.get(id=folder_id) if folder_id else Folder.get_root_folder()
        )
        (object_ids, _, _) = RoleAssignment.get_accessible_object_ids(
            scoped_folder, request.user, Vulnerability
        )

        vulnerabilities = Vulnerability.objects.filter(
            id__in=object_ids
        ).select_related("folder")

        # Build hierarchical structure: folder -> severity -> status
        folder_data = {}

        for vuln in vulnerabilities:
            # Get folder name
            if vuln.folder:
                folder_name = vuln.folder.name
                folder_id = str(vuln.folder.id)
            else:
                folder_name = "No Folder"
                folder_id = "no-folder"

            # Get severity label (lowercase for frontend translation)
            severity_value = vuln.severity
            severity_label = dict(Severity.choices).get(severity_value, "undefined")

            # Get status label (from the choice value, not display)
            status_value = vuln.status
            status_label = status_value if status_value else "--"

            # Initialize folder structure if needed
            if folder_name not in folder_data:
                folder_data[folder_name] = {"folder_id": folder_id, "severities": {}}

            # Initialize severity structure if needed
            if severity_label not in folder_data[folder_name]["severities"]:
                folder_data[folder_name]["severities"][severity_label] = {}

            # Count status occurrences
            if (
                status_label
                not in folder_data[folder_name]["severities"][severity_label]
            ):
                folder_data[folder_name]["severities"][severity_label][status_label] = 0

            folder_data[folder_name]["severities"][severity_label][status_label] += 1

        # Convert to treemap structure
        treemap_data = []
        for folder_name, folder_info in folder_data.items():
            folder_node = {
                "name": folder_name,
                "id": folder_info["folder_id"],
                "children": [],
            }

            for severity_name, statuses in folder_info["severities"].items():
                severity_node = {
                    "name": severity_name,
                    "id": f"{folder_info['folder_id']}-{severity_name.lower()}",
                    "children": [],
                }

                for status_name, count in statuses.items():
                    status_node = {
                        "name": status_name,
                        "id": f"{folder_info['folder_id']}-{severity_name.lower()}-{status_name.lower().replace(' ', '-')}",
                        "value": count,
                    }
                    severity_node["children"].append(status_node)

                folder_node["children"].append(severity_node)

            treemap_data.append(folder_node)

        return Response(treemap_data, status=status.HTTP_200_OK)


class FilteringLabelViewSet(BaseModelViewSet):
    """
    API endpoint that allows labels to be viewed or edited.
    """

    model = FilteringLabel
    filterset_fields = ["folder"]
    search_fields = ["label"]
    ordering = ["label"]


class LibraryFilteringLabelViewSet(BaseModelViewSet):
    """
    API endpoint that allows library labels to be viewed or edited.
    """

    model = LibraryFilteringLabel
    filterset_fields = ["folder"]
    search_fields = ["label"]
    ordering = ["label"]


class RiskAssessmentFilterSet(GenericFilterSet):
    folder = df.ModelMultipleChoiceFilter(queryset=Folder.objects.all())
    perimeter = df.ModelMultipleChoiceFilter(queryset=Perimeter.objects.all())
    risk_matrix = df.ModelMultipleChoiceFilter(queryset=RiskMatrix.objects.all())
    ebios_rm_study = df.ModelMultipleChoiceFilter(queryset=EbiosRMStudy.objects.all())

    status = df.MultipleChoiceFilter(
        choices=RiskAssessment.get_status_choices(), method="filter_status"
    )

    class Meta:
        model = RiskAssessment
        fields = {
            "name": ["exact"],
            "ref_id": ["exact"],
            "authors": ["exact"],
            "reviewers": ["exact"],
            "genericcollection": ["exact"],
        }

    def filter_status(self, queryset, name, value):
        ra_undefined_status = queryset.filter(status__isnull=True)
        if "--" in value:
            return queryset.filter(status__in=value) | ra_undefined_status
        return queryset.filter(status__in=value)


class RiskAssessmentViewSet(BaseModelViewSet):
    """
    API endpoint that allows risk assessments to be viewed or edited.
    """

    model = RiskAssessment
    filterset_class = RiskAssessmentFilterSet

    def get_queryset(self):
        queryset = super().get_queryset()
        return queryset.select_related(
            "folder",
            "perimeter",
            "perimeter__folder",
            "risk_matrix",
            "ebios_rm_study",
        ).prefetch_related(
            "authors",
            "reviewers",
            "risk_scenarios",
        )

    def perform_create(self, serializer):
        instance: RiskAssessment = serializer.save()
        if instance.ebios_rm_study:
            # Unlink all previous risk assessments from this EBIOS RM study
            RiskAssessment.objects.filter(
                ebios_rm_study=instance.ebios_rm_study
            ).exclude(id=instance.id).update(ebios_rm_study=None)

            instance.risk_matrix = instance.ebios_rm_study.risk_matrix
            ebios_rm_study = EbiosRMStudy.objects.get(id=instance.ebios_rm_study.id)
            for operational_scenario in [
                operational_scenario
                for operational_scenario in ebios_rm_study.operational_scenarios.all()
                if operational_scenario.is_selected
            ]:
                # Build comprehensive description from EBIOS RM components
                description_parts = []

                # Feared events
                ro_to = operational_scenario.ro_to
                feared_events = ro_to.feared_events.filter(is_selected=True)
                if feared_events.exists():
                    feared_events_list = []
                    for fe in feared_events:
                        gravity_display = fe.get_gravity_display()
                        gravity_text = (
                            f" [{_('Gravity').capitalize()}: {gravity_display['name']}]"
                            if gravity_display["value"] >= 0
                            else ""
                        )
                        fe_text = f"- {fe.name}{gravity_text}"
                        if fe.description:
                            fe_text += f": {fe.description}"
                        feared_events_list.append(fe_text)
                    feared_events_text = (
                        f"**{_('Feared events').capitalize()}:**\n"
                        + "\n".join(feared_events_list)
                    )
                    description_parts.append(feared_events_text)

                # Risk origin and target objective
                risk_origin_name = (
                    ro_to.risk_origin.get_name_translated
                    if hasattr(ro_to.risk_origin, "get_name_translated")
                    else str(ro_to.risk_origin)
                )
                ro_to_text = f"**{_('Risk origin').capitalize()}:** {risk_origin_name}\n**{_('Target objective').capitalize()}:** {ro_to.target_objective}"
                description_parts.append(ro_to_text)

                # Strategic scenario
                strategic_scenario = operational_scenario.attack_path.strategic_scenario
                if strategic_scenario.description:
                    description_parts.append(
                        f"**{_('Strategic Scenario').capitalize()}:** {strategic_scenario.name}\n{strategic_scenario.description}"
                    )
                elif strategic_scenario.name:
                    description_parts.append(
                        f"**{_('Strategic Scenario').capitalize()}:** {strategic_scenario.name}"
                    )

                # Attack path
                attack_path = operational_scenario.attack_path
                if attack_path.description:
                    description_parts.append(
                        f"**{_('Attack path').capitalize()}:** {attack_path.name}\n{attack_path.description}"
                    )
                elif attack_path.name:
                    description_parts.append(
                        f"**{_('Attack path').capitalize()}:** {attack_path.name}"
                    )

                # Operating modes
                operating_modes = operational_scenario.operating_modes.all()
                if operating_modes.exists():
                    operating_modes_list = []
                    for om in operating_modes:
                        likelihood_display = om.get_likelihood_display()
                        likelihood_text = (
                            f" [{_('Likelihood').capitalize()}: {likelihood_display['name']}]"
                            if likelihood_display["value"] >= 0
                            else ""
                        )
                        om_text = f"- {om.name}{likelihood_text}"
                        if om.description:
                            om_text += f": {om.description}"
                        operating_modes_list.append(om_text)
                    operating_modes_text = (
                        f"**{_('operating modes').capitalize()}:**\n"
                        + "\n".join(operating_modes_list)
                    )
                    description_parts.append(operating_modes_text)
                elif operational_scenario.operating_modes_description:
                    description_parts.append(
                        f"**{_('operating modes').capitalize()}:**\n{operational_scenario.operating_modes_description}"
                    )

                risk_scenario = RiskScenario(
                    risk_assessment=instance,
                    operational_scenario=operational_scenario,
                    name=operational_scenario.name,
                    ref_id=operational_scenario.ref_id
                    if operational_scenario.ref_id
                    else RiskScenario.get_default_ref_id(instance),
                    description="\n\n".join(description_parts),
                )
                if ff_is_enabled("inherent_risk"):
                    risk_scenario.inherent_proba = operational_scenario.likelihood
                    risk_scenario.inherent_impact = operational_scenario.gravity
                else:
                    risk_scenario.current_proba = operational_scenario.likelihood
                    risk_scenario.current_impact = operational_scenario.gravity
                risk_scenario.save()

                risk_scenario.assets.set(operational_scenario.get_assets())
                risk_scenario.threats.set(operational_scenario.threats.all())
                risk_scenario.existing_applied_controls.set(
                    operational_scenario.get_applied_controls()
                )
        instance.save()
        return super().perform_create(serializer)

    @action(detail=True, methods=["post"], name="Sync with EBIOS RM")
    def sync_from_ebios_rm(self, request, pk=None):
        """
        Synchronize an existing risk assessment with its linked EBIOS RM study.
        Updates existing risk scenarios, adds new ones, and archives outdated ones.
        """
        risk_assessment = self.get_object()

        if not risk_assessment.ebios_rm_study:
            return Response(
                {"error": "Risk assessment is not linked to an EBIOS RM study"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ebios_rm_study = risk_assessment.ebios_rm_study
        selected_operational_scenarios = [
            os for os in ebios_rm_study.operational_scenarios.all() if os.is_selected
        ]

        # Track which operational scenarios we've processed
        processed_os_ids = set()
        updated_count = 0
        created_count = 0
        archived_count = 0

        # Helper function to build description (DRY)
        def build_description(operational_scenario):
            description_parts = []

            # Feared events
            ro_to = operational_scenario.ro_to
            feared_events = ro_to.feared_events.filter(is_selected=True)
            if feared_events.exists():
                feared_events_list = []
                for fe in feared_events:
                    gravity_display = fe.get_gravity_display()
                    gravity_text = (
                        f" [{_('Gravity').capitalize()}: {gravity_display['name']}]"
                        if gravity_display["value"] >= 0
                        else ""
                    )
                    fe_text = f"- {fe.name}{gravity_text}"
                    if fe.description:
                        fe_text += f": {fe.description}"
                    feared_events_list.append(fe_text)
                feared_events_text = (
                    f"**{_('Feared events').capitalize()}:**\n"
                    + "\n".join(feared_events_list)
                )
                description_parts.append(feared_events_text)

            # Risk origin and target objective
            risk_origin_name = (
                ro_to.risk_origin.get_name_translated
                if hasattr(ro_to.risk_origin, "get_name_translated")
                else str(ro_to.risk_origin)
            )
            ro_to_text = f"**{_('Risk origin').capitalize()}:** {risk_origin_name}\n**{_('Target objective').capitalize()}:** {ro_to.target_objective}"
            description_parts.append(ro_to_text)

            # Strategic scenario
            strategic_scenario = operational_scenario.attack_path.strategic_scenario
            if strategic_scenario.description:
                description_parts.append(
                    f"**{_('Strategic Scenario').capitalize()}:** {strategic_scenario.name}\n{strategic_scenario.description}"
                )
            elif strategic_scenario.name:
                description_parts.append(
                    f"**{_('Strategic Scenario').capitalize()}:** {strategic_scenario.name}"
                )

            # Attack path
            attack_path = operational_scenario.attack_path
            if attack_path.description:
                description_parts.append(
                    f"**{_('Attack path').capitalize()}:** {attack_path.name}\n{attack_path.description}"
                )
            elif attack_path.name:
                description_parts.append(
                    f"**{_('Attack path').capitalize()}:** {attack_path.name}"
                )

            # Operating modes
            operating_modes = operational_scenario.operating_modes.all()
            if operating_modes.exists():
                operating_modes_list = []
                for om in operating_modes:
                    likelihood_display = om.get_likelihood_display()
                    likelihood_text = (
                        f" [{_('Likelihood').capitalize()}: {likelihood_display['name']}]"
                        if likelihood_display["value"] >= 0
                        else ""
                    )
                    om_text = f"- {om.name}{likelihood_text}"
                    if om.description:
                        om_text += f": {om.description}"
                    operating_modes_list.append(om_text)
                operating_modes_text = (
                    f"**{_('operating modes').capitalize()}:**\n"
                    + "\n".join(operating_modes_list)
                )
                description_parts.append(operating_modes_text)
            elif operational_scenario.operating_modes_description:
                description_parts.append(
                    f"**{_('operating modes').capitalize()}:**\n{operational_scenario.operating_modes_description}"
                )

            return "\n\n".join(description_parts)

        # Update or create risk scenarios for selected operational scenarios
        for operational_scenario in selected_operational_scenarios:
            processed_os_ids.add(operational_scenario.id)

            # Try to find existing risk scenario for this operational scenario
            risk_scenario = None

            # First, try to find by operational_scenario link
            try:
                risk_scenario = risk_assessment.risk_scenarios.get(
                    operational_scenario=operational_scenario
                )
            except RiskScenario.DoesNotExist:
                # Fallback: try to match by name (for backward compatibility with old data)
                # Remove [ARCHIVED] prefix when matching
                clean_name = operational_scenario.name.replace("[ARCHIVED] ", "")
                try:
                    risk_scenario = risk_assessment.risk_scenarios.get(
                        operational_scenario__isnull=True, name=clean_name
                    )
                    # Establish the link for this old risk scenario
                    risk_scenario.operational_scenario = operational_scenario
                except RiskScenario.DoesNotExist:
                    # Also try matching with [ARCHIVED] prefix
                    try:
                        risk_scenario = risk_assessment.risk_scenarios.get(
                            operational_scenario__isnull=True,
                            name=f"[ARCHIVED] {clean_name}",
                        )
                        # Establish the link for this old risk scenario
                        risk_scenario.operational_scenario = operational_scenario
                    except RiskScenario.DoesNotExist:
                        pass

            if risk_scenario:
                # Update existing risk scenario - remove [ARCHIVED] prefix if present
                risk_scenario.name = operational_scenario.name.replace(
                    "[ARCHIVED] ", ""
                )
                risk_scenario.description = build_description(operational_scenario)

                risk_scenario.risk_origin = operational_scenario.ro_to.risk_origin

                # Update inherent or current probability/impact based on feature flag
                if ff_is_enabled("inherent_risk"):
                    risk_scenario.inherent_proba = operational_scenario.likelihood
                    risk_scenario.inherent_impact = operational_scenario.gravity
                else:
                    risk_scenario.current_proba = operational_scenario.likelihood
                    risk_scenario.current_impact = operational_scenario.gravity

                risk_scenario.save()

                # Update relationships
                risk_scenario.assets.set(operational_scenario.get_assets())
                risk_scenario.threats.set(operational_scenario.threats.all())

                # Merge existing controls: keep user-added ones + update from EBIOS RM
                # Get current existing controls
                current_existing_controls = set(
                    risk_scenario.existing_applied_controls.all()
                )
                # Get controls from EBIOS RM
                ebios_controls = set(operational_scenario.get_applied_controls())
                # Merge them (union)
                merged_controls = current_existing_controls | ebios_controls
                risk_scenario.existing_applied_controls.set(merged_controls)

                updated_count += 1

            else:
                # Create new risk scenario (no existing match found)
                risk_scenario = RiskScenario(
                    risk_assessment=risk_assessment,
                    operational_scenario=operational_scenario,
                    name=operational_scenario.name,
                    ref_id=operational_scenario.ref_id
                    if operational_scenario.ref_id
                    else RiskScenario.get_default_ref_id(risk_assessment),
                    description=build_description(operational_scenario),
                    risk_origin=operational_scenario.ro_to.risk_origin,
                )
                if ff_is_enabled("inherent_risk"):
                    risk_scenario.inherent_proba = operational_scenario.likelihood
                    risk_scenario.inherent_impact = operational_scenario.gravity
                else:
                    risk_scenario.current_proba = operational_scenario.likelihood
                    risk_scenario.current_impact = operational_scenario.gravity
                risk_scenario.save()

                risk_scenario.assets.set(operational_scenario.get_assets())
                risk_scenario.threats.set(operational_scenario.threats.all())
                risk_scenario.existing_applied_controls.set(
                    operational_scenario.get_applied_controls()
                )

                created_count += 1

        # Archive risk scenarios that are no longer selected
        for risk_scenario in risk_assessment.risk_scenarios.filter(
            operational_scenario__isnull=False
        ):
            if risk_scenario.operational_scenario_id not in processed_os_ids:
                if not risk_scenario.name.startswith("[ARCHIVED] "):
                    risk_scenario.name = f"[ARCHIVED] {risk_scenario.name}"
                    risk_scenario.save()
                    archived_count += 1

        return Response(
            {
                "success": True,
                "updated": updated_count,
                "created": created_count,
                "archived": archived_count,
                "total_scenarios": risk_assessment.risk_scenarios.count(),
            }
        )

    @action(detail=False, name="Risk assessments per status")
    def per_status(self, request):
        data = assessment_per_status(request.user, RiskAssessment)
        return Response({"results": data})

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get status choices")
    def status(self, request):
        return Response(dict(RiskAssessment.Status.choices))

    @action(detail=False, name="Get quality check")
    def quality_check(self, request):
        """
        Returns the quality check of the risk assessments
        """
        (viewable_objects, _, _) = RoleAssignment.get_accessible_object_ids(
            folder=Folder.get_root_folder(),
            user=request.user,
            object_type=RiskAssessment,
        )
        risk_assessments = RiskAssessment.objects.filter(id__in=viewable_objects)
        res = [
            {"id": a.id, "name": a.name, "quality_check": a.quality_check()}
            for a in risk_assessments
        ]
        return Response({"results": res})

    @action(detail=True, methods=["get"], url_path="quality_check")
    def quality_check_detail(self, request, pk):
        """
        Returns the quality check of the risk_assessment
        """
        (viewable_objects, _, _) = RoleAssignment.get_accessible_object_ids(
            folder=Folder.get_root_folder(),
            user=request.user,
            object_type=RiskAssessment,
        )
        if UUID(pk) in viewable_objects:
            risk_assessment = self.get_object()
            return Response(risk_assessment.quality_check())
        else:
            return Response(status=status.HTTP_403_FORBIDDEN)

    @action(detail=True, name="Get action plan Excel")
    def action_plan_excel(self, request, pk):
        (object_ids_view, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, RiskAssessment
        )
        if UUID(pk) not in object_ids_view:
            return Response(
                {"error": "Permission denied"}, status=status.HTTP_403_FORBIDDEN
            )

        risk_assessment = RiskAssessment.objects.get(id=pk)
        risk_scenarios = risk_assessment.risk_scenarios.all()
        queryset = AppliedControl.objects.filter(
            risk_scenarios__in=risk_scenarios
        ).distinct()

        serializer = RiskAssessmentActionPlanSerializer(
            queryset, many=True, context={"pk": pk}
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Action Plan"

        headers = [
            "Name",
            "Description",
            "Domain",
            "Category",
            "CSF Function",
            "Priority",
            "Status",
            "ETA",
            "Expiry date",
            "Effort",
            "Impact",
            "Cost",
            "Assigned to",
            "Covered scenarios",
        ]
        ws.append(headers)

        for item in serializer.data:
            row = [
                item.get("name"),
                item.get("description"),
                item.get("folder").get("str"),
                item.get("category"),
                item.get("csf_function"),
                item.get("priority"),
                item.get("status"),
                item.get("eta"),
                item.get("expiry_date"),
                item.get("effort"),
                item.get("impact"),
                item.get("annual_cost"),
                "\n".join([ra.get("str") for ra in item.get("owner")]),
                "\n".join([ra.get("str") for ra in item.get("risk_scenarios")]),
            ]
            ws.append(row)

        for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):  # Skip header row
            max_lines = 1
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    line_count = cell.value.count("\n") + 1
                    max_lines = max(max_lines, line_count)
            ws.row_dimensions[row_idx].height = max(15, max_lines * 15)

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception as e:
                    logger.error(f"Error processing cell value: {e}")
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        last_col_letter = get_column_letter(len(headers))
        for cell in ws[last_col_letter]:
            cell.alignment = Alignment(wrap_text=True)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="action_plan_{pk}.xlsx"'
        )
        wb.save(response)

        return response

    @action(detail=True, name="Get risk assessment CSV")
    def risk_assessment_csv(self, request, pk):
        (object_ids_view, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, RiskAssessment
        )
        if UUID(pk) in object_ids_view:
            risk_assessment = self.get_object()

            response = HttpResponse(content_type="text/csv")

            writer = csv.writer(response, delimiter=";")
            columns = [
                "ref_id",
                "assets",
                "threats",
                "name",
                "description",
                "existing_applied_controls",
                "current_impact",
                "current_proba",
                "current_risk",
                "additional_controls",
                "residual_impact",
                "residual_proba",
                "residual_risk",
                "treatment",
                "strength_of_knowledge",
            ]
            if ff_is_enabled("inherent_risk"):
                # insert inherent_risk just before existing_applied_controls
                columns.insert(
                    columns.index("existing_applied_controls"), "inherent_impact"
                )
                columns.insert(
                    columns.index("existing_applied_controls"), "inherent_proba"
                )
                columns.insert(
                    columns.index("existing_applied_controls"), "inherent_level"
                )
            writer.writerow(columns)

            for scenario in risk_assessment.risk_scenarios.all().order_by("ref_id"):
                additional_controls = ",".join(
                    [m.name for m in scenario.applied_controls.all()]
                )
                existing_controls = ",".join(
                    [m.name for m in scenario.existing_applied_controls.all()]
                )

                threats = ",".join([t.name for t in scenario.threats.all()])
                assets = ",".join([t.name for t in scenario.assets.all()])

                row = [
                    scenario.ref_id,
                    assets,
                    threats,
                    scenario.name,
                    scenario.description,
                    existing_controls,
                    scenario.get_current_impact()["name"],
                    scenario.get_current_proba()["name"],
                    scenario.get_current_risk()["name"],
                    additional_controls,
                    scenario.get_residual_impact()["name"],
                    scenario.get_residual_proba()["name"],
                    scenario.get_residual_risk()["name"],
                    scenario.treatment,
                    RiskScenario.DEFAULT_SOK_OPTIONS[scenario.strength_of_knowledge][
                        "name"
                    ],
                ]
                if ff_is_enabled("inherent_risk"):
                    row.insert(
                        columns.index("inherent_impact"),
                        scenario.get_inherent_impact()["name"],
                    )
                    row.insert(
                        columns.index("inherent_proba"),
                        scenario.get_inherent_proba()["name"],
                    )
                    row.insert(
                        columns.index("inherent_level"),
                        scenario.get_inherent_risk()["name"],
                    )
                writer.writerow(row)

            return response
        else:
            return Response(
                {"error": "Permission denied"}, status=status.HTTP_403_FORBIDDEN
            )

    @action(detail=True, name="Get risk assessment XLSX")
    def risk_assessment_xlsx(self, request, pk):
        (object_ids_view, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, RiskAssessment
        )
        if UUID(pk) not in object_ids_view:
            return Response(
                {"error": "Permission denied"}, status=status.HTTP_403_FORBIDDEN
            )

        risk_assessment = self.get_object()
        entries = []

        columns = [
            "ref_id",
            "assets",
            "threats",
            "name",
            "description",
            "existing_applied_controls",
            "current_impact",
            "current_proba",
            "current_risk",
            "additional_controls",
            "residual_impact",
            "residual_proba",
            "residual_risk",
            "treatment",
            "strength_of_knowledge",
        ]

        if ff_is_enabled("inherent_risk"):
            # insert inherent_risk columns just before existing_applied_controls
            columns.insert(
                columns.index("existing_applied_controls"), "inherent_impact"
            )
            columns.insert(columns.index("existing_applied_controls"), "inherent_proba")
            columns.insert(columns.index("existing_applied_controls"), "inherent_level")

        scenarios = risk_assessment.risk_scenarios.prefetch_related(
            "applied_controls",
            "existing_applied_controls",
            "threats",
            "assets",
        ).order_by("ref_id")

        for scenario in scenarios:
            additional_controls = ", ".join(
                escape_excel_formula(m.name) for m in scenario.applied_controls.all()
            )
            existing_controls = ", ".join(
                escape_excel_formula(m.name)
                for m in scenario.existing_applied_controls.all()
            )
            threats = ", ".join(
                escape_excel_formula(t.name) for t in scenario.threats.all()
            )
            assets = ", ".join(
                escape_excel_formula(t.name) for t in scenario.assets.all()
            )

            entry = {
                "ref_id": escape_excel_formula(scenario.ref_id),
                "assets": assets,
                "threats": threats,
                "name": escape_excel_formula(scenario.name),
                "description": escape_excel_formula(scenario.description),
                "existing_applied_controls": existing_controls,
                "current_impact": scenario.get_current_impact()["name"],
                "current_proba": scenario.get_current_proba()["name"],
                "current_risk": scenario.get_current_risk()["name"],
                "additional_controls": additional_controls,
                "residual_impact": scenario.get_residual_impact()["name"],
                "residual_proba": scenario.get_residual_proba()["name"],
                "residual_risk": scenario.get_residual_risk()["name"],
                "treatment": scenario.treatment,
                "strength_of_knowledge": RiskScenario.DEFAULT_SOK_OPTIONS[
                    scenario.strength_of_knowledge
                ]["name"],
            }

            if ff_is_enabled("inherent_risk"):
                entry["inherent_impact"] = scenario.get_inherent_impact()["name"]
                entry["inherent_proba"] = scenario.get_inherent_proba()["name"]
                entry["inherent_level"] = scenario.get_inherent_risk()["name"]

            entries.append(entry)

        df = pd.DataFrame(entries, columns=columns)

        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Risk Assessment")
            worksheet = writer.sheets["Risk Assessment"]

            from openpyxl.styles import Alignment

            # Apply text wrapping to description and control columns
            wrap_columns = [
                "name",
                "description",
                "existing_applied_controls",
                "additional_controls",
            ]
            wrap_indices = [
                df.columns.get_loc(col) + 1 for col in wrap_columns if col in df.columns
            ]

            for col_idx in wrap_indices:
                for row_idx in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.alignment = Alignment(wrap_text=True)

            # Adjust column widths
            for idx, col in enumerate(df.columns, 1):
                column_letter = worksheet.cell(row=1, column=idx).column_letter
                if col in wrap_columns:
                    worksheet.column_dimensions[column_letter].width = 40
                else:
                    worksheet.column_dimensions[column_letter].width = 20

        buffer.seek(0)
        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="risk_assessment_{pk}.xlsx"'
        )

        return response

    @action(detail=True, name="Get risk assessment PDF")
    def risk_assessment_pdf(self, request, pk):
        (object_ids_view, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, RiskAssessment
        )
        if UUID(pk) in object_ids_view:
            risk_assessment = self.get_object()
            context = RiskScenario.objects.filter(
                risk_assessment=risk_assessment
            ).order_by("ref_id")
            for scenario in context:
                scenario.strength_of_knowledge = RiskScenario.DEFAULT_SOK_OPTIONS[
                    scenario.strength_of_knowledge
                ]["name"]
            general_settings = GlobalSettings.objects.filter(name="general").first()
            swap_axes = general_settings.value.get("risk_matrix_swap_axes", False)
            flip_vertical = general_settings.value.get(
                "risk_matrix_flip_vertical", False
            )
            matrix_settings = {
                "swap_axes": "_swapaxes" if swap_axes else "",
                "flip_vertical": "_vflip" if flip_vertical else "",
            }
            ff_settings = GlobalSettings.objects.filter(
                name=GlobalSettings.Names.FEATURE_FLAGS
            ).first()
            if ff_settings is None:
                feature_flags = {}
            else:
                feature_flags = ff_settings.value
            data = {
                "context": context,
                "risk_assessment": risk_assessment,
                "ri_clusters": build_scenario_clusters(
                    risk_assessment,
                    include_inherent=feature_flags.get("inherent_risk", False),
                ),
                "risk_matrix": risk_assessment.risk_matrix,
                "settings": matrix_settings,
                "feature_flags": feature_flags,
            }
            html = render_to_string("core/ra_pdf.html", data)
            pdf_file = HTML(string=html).write_pdf()
            response = HttpResponse(pdf_file, content_type="application/pdf")
            return response
        else:
            return Response({"error": "Permission denied"})

    @action(detail=True, name="Get action plan PDF")
    def action_plan_pdf(self, request, pk):
        (object_ids_view, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, RiskAssessment
        )
        if UUID(pk) in object_ids_view:
            context = {
                "to_do": list(),
                "in_progress": list(),
                "on_hold": list(),
                "active": list(),
                "deprecated": list(),
                "--": list(),
            }
            color_map = {
                "to_do": "#FFF8F0",
                "in_progress": "#392F5A",
                "on_hold": "#F4D06F",
                "active": "#9DD9D2",
                "deprecated": "#ff8811",
                "--": "#e5e7eb",
            }
            status = AppliedControl.Status.choices
            risk_assessment_object: RiskAssessment = self.get_object()
            risk_scenarios_objects = risk_assessment_object.risk_scenarios.all()
            applied_controls = (
                AppliedControl.objects.filter(risk_scenarios__in=risk_scenarios_objects)
                .distinct()
                .order_by("eta")
            )
            for applied_control in applied_controls:
                context[applied_control.status].append(
                    applied_control
                ) if applied_control.status else context["--"].append(applied_control)
            data = {
                "status_text": status,
                "color_map": color_map,
                "context": context,
                "risk_assessment": risk_assessment_object,
            }
            html = render_to_string("core/risk_action_plan_pdf.html", data)
            pdf_file = HTML(string=html).write_pdf()
            response = HttpResponse(pdf_file, content_type="application/pdf")
            return response
        else:
            return Response({"error": "Permission denied"})

    @action(
        detail=True,
        name="Duplicate risk assessment",
        methods=["post"],
        serializer_class=RiskAssessmentDuplicateSerializer,
    )
    def duplicate(self, request, pk):
        (object_ids_view, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, RiskAssessment
        )

        if UUID(pk) in object_ids_view:
            risk_assessment = self.get_object()
            data = request.data

            duplicate_risk_assessment = RiskAssessment.objects.create(
                name=data.get("name"),
                description=data.get("description"),
                perimeter=Perimeter.objects.get(id=data.get("perimeter")),
                version=data.get("version"),
                risk_matrix=risk_assessment.risk_matrix,
                ref_id=data.get("ref_id"),
                eta=risk_assessment.eta,
                due_date=risk_assessment.due_date,
                status=risk_assessment.status,
            )

            duplicate_risk_assessment.authors.set(risk_assessment.authors.all())
            duplicate_risk_assessment.reviewers.set(risk_assessment.reviewers.all())

            for scenario in risk_assessment.risk_scenarios.all():
                duplicate_scenario = RiskScenario.objects.create(
                    risk_assessment=duplicate_risk_assessment,
                    name=scenario.name,
                    description=scenario.description,
                    treatment=scenario.treatment,
                    current_proba=scenario.current_proba,
                    current_impact=scenario.current_impact,
                    residual_proba=scenario.residual_proba,
                    residual_impact=scenario.residual_impact,
                    strength_of_knowledge=scenario.strength_of_knowledge,
                    justification=scenario.justification,
                    ref_id=scenario.ref_id,
                )

                duplicate_scenario.qualifications.set(scenario.qualifications.all())

                for field in [
                    "applied_controls",
                    "threats",
                    "assets",
                    "existing_applied_controls",
                ]:
                    duplicate_related_objects(
                        scenario,
                        duplicate_scenario,
                        duplicate_risk_assessment.folder,
                        field,
                    )

                if duplicate_risk_assessment.folder in [risk_assessment.folder] + [
                    folder for folder in risk_assessment.folder.get_sub_folders()
                ]:
                    duplicate_scenario.owner.set(scenario.owner.all())

                duplicate_scenario.save()

            duplicate_risk_assessment.save()
            return Response({"results": "risk assessment duplicated"})

    @action(
        detail=True,
        methods=["post"],
        url_path="sync-to-actions",
    )
    def sync_to_applied_controls(self, request, pk):
        dry_run = request.query_params.get("dry_run", True)
        if dry_run == "false":
            dry_run = False
        reset_residual = request.data.get("reset_residual", False)
        risk_assessment = RiskAssessment.objects.get(id=pk)

        if not RoleAssignment.is_access_allowed(
            user=request.user,
            perm=Permission.objects.get(codename="change_riskassessment"),
            folder=Folder.get_folder(risk_assessment),
        ):
            return Response(status=status.HTTP_403_FORBIDDEN)

        changes = risk_assessment.sync_to_applied_controls(
            reset_residual=reset_residual, dry_run=dry_run
        )
        return Response(
            {"changes": RiskScenarioReadSerializer(changes, many=True).data}
        )

    @action(
        detail=True,
        methods=["post"],
        url_path="convert_to_quantitative",
    )
    def convert_to_quantitative(self, request, pk):
        """
        Convert a qualitative risk assessment to a quantitative risk study.

        Expected payload:
        {
            "probability_anchors": [{"index": 0, "value": 0.05}, ...],
            "impact_anchors": [{"index": 0, "central_value": 25000}, ...],
            "loss_threshold": 100000
        }
        """
        from crq.models import (
            QuantitativeRiskStudy,
            QuantitativeRiskScenario,
            QuantitativeRiskHypothesis,
        )

        risk_assessment: RiskAssessment = self.get_object()

        # Check permissions
        if not RoleAssignment.is_access_allowed(
            user=request.user,
            perm=Permission.objects.get(codename="add_quantitativeriskstudy"),
            folder=risk_assessment.folder,
        ):
            return Response(
                {
                    "detail": "You do not have permission to create quantitative risk studies in this domain."
                },
                status=status.HTTP_403_FORBIDDEN,
            )

        # Validate payload
        probability_anchors = request.data.get("probability_anchors", [])
        impact_anchors = request.data.get("impact_anchors", [])
        loss_threshold = request.data.get("loss_threshold")

        if not probability_anchors or not impact_anchors or loss_threshold is None:
            return Response(
                {
                    "detail": "Missing required fields: probability_anchors, impact_anchors, and loss_threshold"
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Validate loss_threshold is a positive number
        try:
            loss_threshold = float(loss_threshold)
            if loss_threshold <= 0:
                return Response(
                    {"detail": "loss_threshold must be greater than 0"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        except (TypeError, ValueError):
            return Response(
                {"detail": "loss_threshold must be a valid number"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Validate and build probability mapping
        probability_map = {}
        try:
            for item in probability_anchors:
                idx = item["index"]
                value = float(item["value"])
                if not (0 <= value <= 1):
                    return Response(
                        {
                            "detail": f"Probability value must be between 0 and 1, got {value} for index {idx}"
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                probability_map[idx] = value
        except (KeyError, TypeError, ValueError) as e:
            logger.warning(f"Invalid probability anchor format: {e}")
            return Response(
                {
                    "detail": "Invalid probability anchor format. Each anchor must have 'index' and 'value' fields, with value between 0 and 1."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Validate and build impact mapping
        # Using a narrow fixed ratio: roughly ±30% around central value
        # LB = central_value * 0.7, UB = central_value * 1.43
        # This gives a spread factor of ~2x which minimizes overlap between adjacent levels
        impact_map = {}
        try:
            for item in impact_anchors:
                idx = item["index"]
                central_value = float(item["central_value"])
                if central_value <= 0:
                    return Response(
                        {
                            "detail": f"Impact central value must be positive, got {central_value} for index {idx}"
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                impact_map[idx] = {
                    "lb": central_value * 0.7,
                    "ub": central_value * 1.43,
                }
        except (KeyError, TypeError, ValueError) as e:
            logger.warning(f"Invalid impact anchor format: {e}")
            return Response(
                {
                    "detail": "Invalid impact anchor format. Each anchor must have 'index' and 'central_value' fields, with central_value greater than 0."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Generate ref_id for the quantitative study
        # If source has ref_id, append _QUANT, otherwise leave blank
        quant_ref_id = (
            f"{risk_assessment.ref_id}_QUANT" if risk_assessment.ref_id else ""
        )

        # Create the quantitative risk study with [QUANT] prefix
        quantitative_study = QuantitativeRiskStudy.objects.create(
            name=f"[QUANT] {risk_assessment.name}",
            description=f"Converted from qualitative risk assessment: {risk_assessment.name}\n\n{risk_assessment.description or ''}",
            folder=risk_assessment.folder,
            ref_id=quant_ref_id,
            status=risk_assessment.status if risk_assessment.status else "planned",
            loss_threshold=loss_threshold,
            distribution_model="lognormal_ci90",
        )

        # Copy authors and reviewers
        quantitative_study.authors.set(risk_assessment.authors.all())
        quantitative_study.reviewers.set(risk_assessment.reviewers.all())

        scenarios_converted = 0
        scenarios_skipped = 0

        # Convert each risk scenario to a quantitative risk scenario
        for scenario in risk_assessment.risk_scenarios.all():
            # Skip scenarios without threats or assets
            if not scenario.threats.exists() and not scenario.assets.exists():
                logger.info(
                    f"Skipping scenario '{scenario.name}' - no threats or assets associated"
                )
                scenarios_skipped += 1
                continue

            # Check if we have current probability and impact
            current_proba = scenario.current_proba
            current_impact = scenario.current_impact
            has_current = (
                current_proba is not None
                and current_impact is not None
                and current_proba in probability_map
                and current_impact in impact_map
            )

            # Check if we have residual probability and impact
            residual_proba = scenario.residual_proba
            residual_impact = scenario.residual_impact
            has_residual = (
                residual_proba is not None
                and residual_impact is not None
                and residual_proba in probability_map
                and residual_impact in impact_map
            )

            # Skip if we don't have at least current values
            if not has_current and not has_residual:
                logger.info(
                    f"Skipping scenario '{scenario.name}' - no valid probability/impact values"
                )
                scenarios_skipped += 1
                continue

            # Create quantitative risk scenario with same name and ref_id
            quant_scenario = QuantitativeRiskScenario.objects.create(
                quantitative_risk_study=quantitative_study,
                name=scenario.name,
                description=scenario.description or "",
                ref_id=scenario.ref_id or "",
                folder=risk_assessment.folder,
            )

            # Copy threats, assets, and owners
            quant_scenario.threats.set(scenario.threats.all())
            quant_scenario.assets.set(scenario.assets.all())
            quant_scenario.vulnerabilities.set(scenario.vulnerabilities.all())
            quant_scenario.owner.set(scenario.owner.all())

            # Delete the auto-created "baseline" current hypothesis
            # (we'll create our own if needed)
            QuantitativeRiskHypothesis.objects.filter(
                quantitative_risk_scenario=quant_scenario, risk_stage="current"
            ).delete()

            # Create current hypothesis if current values are set
            if has_current:
                probability_value = probability_map[current_proba]
                impact_value = impact_map[current_impact]

                current_hypothesis = QuantitativeRiskHypothesis.objects.create(
                    quantitative_risk_scenario=quant_scenario,
                    name="current",
                    risk_stage="current",
                    ref_id=scenario.ref_id or "",
                    folder=risk_assessment.folder,
                    parameters={
                        "probability": probability_value,
                        "impact": {
                            "distribution": "LOGNORMAL-CI90",
                            "lb": impact_value["lb"],
                            "ub": impact_value["ub"],
                        },
                    },
                )

                # Link to existing applied controls
                if scenario.existing_applied_controls.exists():
                    current_hypothesis.existing_applied_controls.set(
                        scenario.existing_applied_controls.all()
                    )
                current_hypothesis.save()

                # Run simulation for the current hypothesis
                try:
                    current_hypothesis.run_simulation()
                    logger.info(
                        f"Simulation completed for current hypothesis: {scenario.name}"
                    )
                except Exception as e:
                    logger.error(
                        f"Failed to run simulation for scenario {scenario.name}: {str(e)}"
                    )

            # Create residual hypothesis if residual values are set
            if has_residual:
                residual_probability_value = probability_map[residual_proba]
                residual_impact_value = impact_map[residual_impact]

                residual_hypothesis = QuantitativeRiskHypothesis.objects.create(
                    quantitative_risk_scenario=quant_scenario,
                    name="residual",
                    risk_stage="residual",
                    is_selected=True,
                    ref_id=f"{scenario.ref_id}_RES" if scenario.ref_id else "",
                    folder=risk_assessment.folder,
                    parameters={
                        "probability": residual_probability_value,
                        "impact": {
                            "distribution": "LOGNORMAL-CI90",
                            "lb": residual_impact_value["lb"],
                            "ub": residual_impact_value["ub"],
                        },
                    },
                )

                # Link to existing controls and added controls
                if scenario.existing_applied_controls.exists():
                    residual_hypothesis.existing_applied_controls.set(
                        scenario.existing_applied_controls.all()
                    )
                if scenario.applied_controls.exists():
                    residual_hypothesis.added_applied_controls.set(
                        scenario.applied_controls.all()
                    )
                residual_hypothesis.save()

                # Run simulation for the residual hypothesis
                try:
                    residual_hypothesis.run_simulation()
                    logger.info(
                        f"Simulation completed for residual hypothesis: {scenario.name}"
                    )
                except Exception as e:
                    logger.error(
                        f"Failed to run simulation for residual scenario {scenario.name}: {str(e)}"
                    )

            scenarios_converted += 1

        logger.info(
            f"Conversion completed: {scenarios_converted} scenarios converted, {scenarios_skipped} skipped"
        )

        return Response(
            {
                "message": "Conversion successful",
                "quantitative_risk_study_id": str(quantitative_study.id),
                "scenarios_converted": scenarios_converted,
                "scenarios_skipped": scenarios_skipped,
            },
            status=status.HTTP_201_CREATED,
        )


def convert_date_to_timestamp(date):
    """
    Converts a date object (datetime.date) to a Linux timestamp.
    It creates a datetime object for the date at midnight and makes it timezone-aware.
    """
    if date:
        date_as_datetime = datetime.combine(date, datetime.min.time())
        aware_datetime = pytz.UTC.localize(date_as_datetime)
        return int(time.mktime(aware_datetime.timetuple())) * 1000
    return None


class AppliedControlFilterSet(GenericFilterSet):
    folder = df.ModelMultipleChoiceFilter(queryset=Folder.objects.all())
    reference_control = df.ModelMultipleChoiceFilter(
        queryset=ReferenceControl.objects.all()
    )

    todo = df.BooleanFilter(method="filter_todo")
    to_review = df.BooleanFilter(method="filter_to_review")
    compliance_assessments = df.ModelMultipleChoiceFilter(
        method="filter_compliance_assessments",
        queryset=ComplianceAssessment.objects.all(),
    )
    risk_assessments = df.ModelMultipleChoiceFilter(
        method="filter_risk_assessments",
        queryset=RiskAssessment.objects.all(),
    )
    findings_assessments = df.ModelMultipleChoiceFilter(
        method="filter_findings_assessments",
        queryset=FindingsAssessment.objects.all(),
    )
    status = df.MultipleChoiceFilter(
        choices=AppliedControl.Status.choices, lookup_expr="icontains"
    )
    is_assigned = df.BooleanFilter(method="filter_is_assigned")
    # Nullable choice filters that support filtering for unset values using "--"
    category = NullableChoiceFilter(choices=AppliedControl.CATEGORY)
    csf_function = NullableChoiceFilter(choices=AppliedControl.CSF_FUNCTION)
    priority = NullableChoiceFilter(choices=AppliedControl.PRIORITY)
    effort = NullableChoiceFilter(choices=AppliedControl.EFFORT)
    control_impact = NullableChoiceFilter(choices=AppliedControl.IMPACT)

    def filter_findings_assessments(self, queryset, name, value):
        if value:
            findings_assessments = FindingsAssessment.objects.filter(
                id__in=[x.id for x in value]
            )
            if len(findings_assessments) == 0:
                return queryset
            findings = chain.from_iterable(
                [fa.findings.all() for fa in findings_assessments]
            )
            return queryset.filter(findings__in=findings).distinct()
        return queryset

    def filter_risk_assessments(self, queryset, name, value):
        if value:
            risk_assessments = RiskAssessment.objects.filter(
                id__in=[x.id for x in value]
            )
            if len(risk_assessments) == 0:
                return queryset
            risk_scenarios = chain.from_iterable(
                [ra.risk_scenarios.all() for ra in risk_assessments]
            )
            return queryset.filter(risk_scenarios__in=risk_scenarios).distinct()
        return queryset

    def filter_compliance_assessments(self, queryset, name, value):
        if value:
            compliance_assessments = ComplianceAssessment.objects.filter(
                id__in=[x.id for x in value]
            )
            if len(compliance_assessments) == 0:
                return queryset
            requirement_assessments = chain.from_iterable(
                [ca.requirement_assessments.all() for ca in compliance_assessments]
            )
            return queryset.filter(
                requirement_assessments__in=requirement_assessments
            ).distinct()
        return queryset

    def filter_todo(self, queryset, name, value):
        if value:
            return (
                queryset.filter(eta__lte=date.today() + timedelta(days=30))
                .exclude(status="active")
                .order_by("eta")
            )

        return queryset

    def filter_to_review(self, queryset, name, value):
        if value:
            return queryset.filter(
                expiry_date__lte=date.today() + timedelta(days=30)
            ).order_by("expiry_date")
        return queryset

    def filter_is_assigned(self, queryset, name, value):
        if value:
            return queryset.filter(owner__isnull=False).distinct()
        else:
            return queryset.filter(owner__isnull=True)

    class Meta:
        model = AppliedControl
        fields = {
            "name": ["exact"],
            "category": ["exact"],
            "csf_function": ["exact"],
            "priority": ["exact"],
            "effort": ["exact"],
            "control_impact": ["exact"],
            # category, csf_function, priority, effort, control_impact use NullableChoiceFilter (defined above)
            "filtering_labels": ["exact"],
            "risk_scenarios": ["exact"],
            "risk_scenarios_e": ["exact"],
            "requirement_assessments": ["exact"],
            "evidences": ["exact"],
            "objectives": ["exact"],
            "assets": ["exact"],
            "stakeholders": ["exact"],
            "progress_field": ["exact"],
            "security_exceptions": ["exact"],
            "owner": ["exact"],
            "findings": ["exact"],
            "eta": ["exact", "lte", "gte", "lt", "gt", "month", "year"],
            "ref_id": ["exact"],
            "processings": ["exact"],
            "genericcollection": ["exact"],
        }


class AppliedControlViewSet(ExportMixin, BaseModelViewSet):
    """
    API endpoint that allows applied controls to be viewed or edited.
    """

    model = AppliedControl
    filterset_class = AppliedControlFilterSet
    search_fields = ["name", "description", "ref_id"]

    @staticmethod
    def _extract_cost_field(control, *path):
        """Helper to safely extract nested cost fields."""
        value = control.cost
        if not value:
            return ""
        for key in path:
            value = value.get(key, "")
            if not value:
                return ""
        return value

    export_config = {
        "fields": {
            "internal_id": {"source": "id", "label": "internal_id"},
            "ref_id": {"source": "ref_id", "label": "ref_id", "escape": True},
            "name": {"source": "name", "label": "name", "escape": True},
            "description": {
                "source": "description",
                "label": "description",
                "escape": True,
            },
            "reference_control_name": {
                "source": "reference_control.name",
                "label": "reference_control_name",
                "escape": True,
            },
            "reference_control_ref_id": {
                "source": "reference_control.ref_id",
                "label": "reference_control_ref_id",
                "escape": True,
            },
            "category": {"source": "category", "label": "category"},
            "csf_function": {"source": "csf_function", "label": "csf_function"},
            "folder": {"source": "folder.name", "label": "folder"},
            "status": {"source": "status", "label": "status"},
            "start_date": {"source": "start_date", "label": "start_date"},
            "eta": {"source": "eta", "label": "eta"},
            "expiry_date": {"source": "expiry_date", "label": "expiry_date"},
            "priority": {"source": "priority", "label": "priority"},
            "effort": {"source": "get_effort_display", "label": "effort"},
            "impact": {"source": "get_control_impact_display", "label": "impact"},
            "cost_currency": {
                "source": "cost",
                "label": "cost_currency",
                "format": lambda cost: cost.get("currency", "") if cost else "",
            },
            "cost_build_fixed": {
                "source": "cost",
                "label": "cost_build_fixed",
                "format": lambda cost: cost.get("build", {}).get("fixed_cost", "")
                if cost
                else "",
            },
            "cost_build_people_days": {
                "source": "cost",
                "label": "cost_build_people_days",
                "format": lambda cost: cost.get("build", {}).get("people_days", "")
                if cost
                else "",
            },
            "cost_run_fixed": {
                "source": "cost",
                "label": "cost_run_fixed",
                "format": lambda cost: cost.get("run", {}).get("fixed_cost", "")
                if cost
                else "",
            },
            "cost_run_people_days": {
                "source": "cost",
                "label": "cost_run_people_days",
                "format": lambda cost: cost.get("run", {}).get("people_days", "")
                if cost
                else "",
            },
            "cost_amortization_period": {
                "source": "cost",
                "label": "cost_amortization_period",
                "format": lambda cost: cost.get("amortization_period", "")
                if cost
                else "",
            },
            "owner": {
                "source": "owner",
                "label": "owner",
                "format": lambda qs: ",".join(o.email for o in qs.all())
                if qs.exists()
                else "",
            },
            "labels": {
                "source": "filtering_labels",
                "label": "labels",
                "format": lambda qs: ",".join(lbl.label for lbl in qs.all()),
            },
        },
        "filename": "audit_export",
        "select_related": ["reference_control", "folder"],
        "prefetch_related": ["owner", "filtering_labels"],
    }

    def get_queryset(self):
        """Optimize queries by prefetching related objects used in the table view and serializer"""
        return (
            super()
            .get_queryset()
            .select_related(
                "folder",
                "folder__parent_folder",  # For get_folder_full_path() optimization
                "reference_control",
            )
            .prefetch_related(
                "owner",
                "filtering_labels__folder",  # FieldsRelatedField includes folder
                "findings",  # Used for findings_count
                "evidences",  # Serialized as FieldsRelatedField
                "objectives",  # ManyToManyField to OrganisationObjective
                "assets",  # ManyToManyField used in table
                "security_exceptions",  # Serialized as FieldsRelatedField
            )
        )

    def perform_create(self, serializer):
        create_remote_object = serializer.validated_data.pop(
            "create_remote_object", False
        )
        integration_config = serializer.validated_data.pop("integration_config", None)
        serializer.validated_data.pop("remote_object_id", None)  # Remove if present

        # Create the local object first
        super().perform_create(serializer)
        instance = serializer.instance

        if create_remote_object and integration_config:
            try:
                logger.info(
                    "Creating remote object for Applied Control",
                    applied_control_id=instance.id,
                    integration_config_id=integration_config.id,
                )
                sync_object_to_integrations.schedule(
                    args=(
                        ContentType.objects.get_for_model(self.model),
                        serializer.instance.id,
                        [integration_config.id],
                        ["name", "description", "status", "priority"],
                    ),
                    delay=1,  # Small delay to ensure transaction is committed
                )

            except Exception:
                logger.error(
                    "Error creating remote object for Applied Control",
                    applied_control_id=instance.id,
                    exc_info=True,
                )

    def perform_update(self, serializer):
        integration_config = serializer.validated_data.pop("integration_config", None)
        remote_object_id = serializer.validated_data.pop("remote_object_id", None)
        serializer.validated_data.pop("create_remote_object", None)  # Remove if present

        super().perform_update(serializer)
        if not integration_config or not remote_object_id:
            return
        try:
            if remote_object_id:
                logger.info(
                    "Attaching applied control to external object",
                    applied_control_id=serializer.instance.id,
                    remote_id=remote_object_id,
                )
                sync_mapping = SyncMapping.objects.create(
                    configuration=integration_config,
                    content_type=ContentType.objects.get_for_model(self.model),
                    local_object_id=serializer.instance.id,
                    remote_id=remote_object_id,
                    sync_status=SyncMapping.SyncStatus.PENDING,
                )
                sync_object_to_integrations.schedule(
                    args=(
                        sync_mapping.content_type,
                        serializer.instance.id,
                        [integration_config.id],
                        ["status"],
                    ),
                    delay=1,  # Small delay to ensure transaction is committed
                )

        except Exception:
            logger.error("Error creating SyncMapping", exc_info=True)

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get status choices")
    def status(self, request):
        return Response(dict(AppliedControl.Status.choices))

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get category choices")
    def category(self, request):
        return Response(add_unset_option(dict(AppliedControl.CATEGORY)))

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get csf_function choices")
    def csf_function(self, request):
        return Response(add_unset_option(dict(AppliedControl.CSF_FUNCTION)))

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get priority choices")
    def priority(self, request):
        return Response(add_unset_option(dict(AppliedControl.PRIORITY)))

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get effort choices")
    def effort(self, request):
        return Response(add_unset_option(dict(AppliedControl.EFFORT)))

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get impact choices")
    def control_impact(self, request):
        return Response(add_unset_option(dict(AppliedControl.IMPACT)))

    @action(detail=False, name="Get all applied controls owners")
    def owner(self, request):
        return Response(
            UserReadSerializer(
                User.objects.filter(applied_controls__isnull=False).distinct(),
                many=True,
            ).data
        )

    @action(detail=False, name="Get updatable measures")
    def updatables(self, request):
        (_, object_ids_change, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, AppliedControl
        )

        return Response({"results": object_ids_change})

    @action(
        detail=False, name="Something"
    )  # Write a good name for the "name" keyword argument
    def per_status(self, request):
        data = applied_control_per_status(request.user)
        return Response({"results": data})

    @action(detail=False, name="Get the ordered todo applied controls")
    def todo(self, request):
        (object_ids_view, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, AppliedControl
        )

        measures = sorted(
            AppliedControl.objects.filter(id__in=object_ids_view)
            .filter(eta__lte=date.today() + timedelta(days=30))
            .exclude(status="active")
            .order_by("eta"),
            key=lambda mtg: mtg.get_ranking_score(),
            reverse=True,
        )

        ranking_scores = {str(mtg.id): mtg.get_ranking_score() for mtg in measures}

        measures = [AppliedControlReadSerializer(mtg).data for mtg in measures]

        # How to add ranking_score directly in the serializer ?

        for i in range(len(measures)):
            measures[i]["ranking_score"] = ranking_scores[measures[i]["id"]]

        """
        The serializer of AppliedControl isn't applied automatically for this function
        """

        return Response({"results": measures})

    @action(detail=False, name="Get the secuity measures to review")
    def to_review(self, request):
        measures = measures_to_review(request.user)

        measures = [AppliedControlReadSerializer(mtg).data for mtg in measures]

        """
        The serializer of AppliedControl isn't applied automatically for this function
        """

        return Response({"results": measures})

    @action(detail=False, methods=["get"])
    def get_controls_info(self, request):
        nodes = list()
        links = list()
        for ac in AppliedControl.objects.all():
            related_items_count = 0
            for ca in ComplianceAssessment.objects.filter(
                requirement_assessments__applied_controls=ac
            ).distinct():
                audit_coverage = (
                    RequirementAssessment.objects.filter(compliance_assessment=ca)
                    .filter(applied_controls=ac)
                    .count()
                )
                related_items_count += audit_coverage
                links.append(
                    {
                        "source": ca.id,
                        "target": ac.id,
                        "coverage": audit_coverage,
                    }
                )
            for ra in RiskAssessment.objects.filter(
                risk_scenarios__applied_controls=ac
            ).distinct():
                risk_coverage = (
                    RiskScenario.objects.filter(risk_assessment=ra)
                    .filter(applied_controls=ac)
                    .count()
                )
                related_items_count += risk_coverage
                links.append(
                    {
                        "source": ra.id,
                        "target": ac.id,
                        "coverage": risk_coverage,
                    }
                )
            nodes.append(
                {
                    "id": ac.id,
                    "label": ac.name,
                    "shape": "hexagon",
                    "counter": related_items_count,
                    "color": "#47e845",
                }
            )
        for audit in ComplianceAssessment.objects.all():
            nodes.append(
                {
                    "id": audit.id,
                    "label": audit.name,
                    "shape": "circle",
                    "color": "#5D4595",
                }
            )
        for ra in RiskAssessment.objects.all():
            nodes.append(
                {
                    "id": ra.id,
                    "label": ra.name,
                    "shape": "square",
                    "color": "#E6499F",
                }
            )
        return Response(
            {
                "nodes": nodes,
                "links": links,
            }
        )

    @action(detail=False, name="Get priority chart data")
    def priority_chart_data(self, request):
        qs = AppliedControl.objects.exclude(status="active")

        data = {
            "--": [],
            "to_do": [],
            "in_progress": [],
            "on_hold": [],
            "deprecated": [],
        }
        angle_offsets = {"4": 0, "3": 90, "1": 180, "2": 270}
        status_offset = {
            "--": 4,
            "to_do": 12,
            "in_progress": 20,
            "on_hold": 28,
            "deprecated": 36,
        }

        not_displayed_cnt = 0

        p_dict = qs.aggregate(
            p1=Count("priority", filter=Q(priority=1)),
            p2=Count("priority", filter=Q(priority=2)),
            p3=Count("priority", filter=Q(priority=3)),
            p4=Count("priority", filter=Q(priority=4)),
        )
        for ac in qs:
            if ac.priority:
                if ac.eta:
                    days_countdown = min(100, ac.days_until_eta)
                    # how many days until the ETA
                else:
                    days_countdown = 100
                impact_factor = 5 + ac.links_count

                # angle = angle_offsets[str(ac.priority)]+ (next(offsets) % 80) + random.randint(1,4)
                angle = (
                    angle_offsets[str(ac.priority)]
                    + status_offset[ac.status]
                    + random.randint(1, 40)
                )
                # angle = angle_offsets[str(ac.priority)] + next(offsets)

                vector = [
                    days_countdown,
                    angle,
                    impact_factor,
                    f"[{ac.priority}] {str(ac)}",
                    ac.status,
                    ac.id,
                ]
                if ac.status:
                    data[ac.status].append(vector)
                else:
                    data["unclassified"].append(vector)
            else:
                print("priority unset - add it to triage lot")
                not_displayed_cnt += 1

        data["not_displayed"] = not_displayed_cnt
        data["priority_cnt"] = p_dict

        return Response(data)

    @action(detail=False, methods=["get"])
    def get_gantt_data(self, request):
        def format_date(input):
            return datetime.strftime(input, "%Y-%m-%d")

        entries = []
        (viewable_controls_ids, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, AppliedControl
        )

        applied_controls = AppliedControl.objects.filter(
            id__in=viewable_controls_ids
        ).select_related("folder")

        for ac in applied_controls:
            if ac.eta:
                endDate = format_date(ac.eta)
                startDate = format_date(ac.start_date) if ac.start_date else endDate
                if ac.start_date:
                    startDate = format_date(ac.start_date)
                else:
                    startDate = format_date(ac.eta - timedelta(days=1))
                entries.append(
                    {
                        "id": ac.id,
                        "start": startDate,
                        "end": endDate,
                        "name": ac.name,
                        "progress": ac.progress_field,
                        "description": ac.description
                        if ac.description
                        else "(no description)",
                        "domain": ac.folder.name,
                    }
                )
        return Response(entries)

    @action(detail=False, methods=["get"])
    def impact_effort(self, request):
        # TODO consider the case of passing the domain as a filter
        (viewable_controls_ids, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, AppliedControl
        )
        viewable_controls_set = set(viewable_controls_ids)

        output = [
            [[], [], [], [], []],
            [[], [], [], [], []],
            [[], [], [], [], []],
            [[], [], [], [], []],
            [[], [], [], [], []],
        ]

        assessed_controls = AppliedControl.objects.filter(
            id__in=viewable_controls_set,
            control_impact__isnull=False,
            effort__isnull=False,
        ).values_list(
            "id",
            "name",
            "control_impact",
            "effort",
        )

        # MAP_EFFORT for converting effort strings to integers
        MAP_EFFORT = {None: -1, "XS": 1, "S": 2, "M": 3, "L": 4, "XL": 5}

        for control_id, name, impact, effort_str in assessed_controls:
            # Convert effort string to integer
            effort_int = MAP_EFFORT.get(effort_str, -1)

            # Convert to 0-based indices
            impact_idx = 5 - impact
            effort_idx = effort_int - 1

            # Add to matrix if both values are valid (1-5)
            if 0 <= impact_idx < 5 and 0 <= effort_idx < 5:
                output[impact_idx][effort_idx].append(
                    {
                        "id": control_id,
                        "impact": impact,
                        "effort": effort_int,
                        "name": name,
                    }
                )

        # Get viewable but not assessed controls
        all_assessed_ids = set(control_id for control_id, _, _, _ in assessed_controls)
        viewable_not_assessed = viewable_controls_set - all_assessed_ids

        print("Matrix populated with assessed controls")
        print(f"Viewable but not assessed: {len(viewable_not_assessed)} controls")
        return Response(output)

    @action(detail=False, methods=["get"])
    def get_timeline_info(self, request):
        entries = []
        COLORS_PALETTE = [
            "#F72585",
            "#7209B7",
            "#3A0CA3",
            "#4361EE",
            "#4CC9F0",
            "#A698DC",
        ]
        colorMap = {}
        (viewable_controls_ids, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, AppliedControl
        )

        applied_controls = AppliedControl.objects.filter(
            id__in=viewable_controls_ids
        ).select_related("folder")

        for ac in applied_controls:
            if ac.eta:
                endDate = convert_date_to_timestamp(ac.eta)
                startDate = (
                    convert_date_to_timestamp(ac.start_date)
                    if ac.start_date
                    else endDate
                )
                entries.append(
                    {
                        "startDate": startDate,
                        "endDate": endDate,
                        "name": ac.name,
                        "description": ac.description
                        if ac.description
                        else "(no description)",
                        "domain": ac.folder.name,
                    }
                )
        color_cycle = cycle(COLORS_PALETTE)
        for domain in Folder.objects.all():
            colorMap[domain.name] = next(color_cycle)
        return Response({"entries": entries, "colorMap": colorMap})

    @action(
        detail=True,
        name="Duplicate applied control",
        methods=["post"],
        serializer_class=AppliedControlDuplicateSerializer,
    )
    def duplicate(self, request, pk):
        (object_ids_view, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, AppliedControl
        )
        if UUID(pk) not in object_ids_view:
            return Response(
                {"results": "applied control duplicated"},
                status=status.HTTP_404_NOT_FOUND,
            )

        applied_control = self.get_object()
        data = request.data
        new_folder = Folder.objects.get(id=data["folder"])
        duplicate_applied_control = AppliedControl.objects.create(
            reference_control=applied_control.reference_control,
            name=data["name"],
            description=data["description"],
            folder=new_folder,
            ref_id=applied_control.ref_id,
            category=applied_control.category,
            csf_function=applied_control.csf_function,
            priority=applied_control.priority,
            status=applied_control.status,
            start_date=applied_control.start_date,
            eta=applied_control.eta,
            expiry_date=applied_control.expiry_date,
            link=applied_control.link,
            effort=applied_control.effort,
            cost=applied_control.cost,
            progress_field=applied_control.progress_field,
        )
        duplicate_applied_control.owner.set(applied_control.owner.all())
        duplicate_applied_control.filtering_labels.set(
            applied_control.filtering_labels.all()
        )
        if data["duplicate_evidences"]:
            duplicate_related_objects(
                applied_control, duplicate_applied_control, new_folder, "evidences"
            )
            duplicate_applied_control.save()

        return Response(
            {"results": AppliedControlReadSerializer(duplicate_applied_control).data}
        )

    @action(detail=False, methods=["get"])
    def ids(self, request):
        my_map = dict()

        (viewable_items, _, _) = RoleAssignment.get_accessible_object_ids(
            folder=Folder.get_root_folder(),
            user=request.user,
            object_type=AppliedControl,
        )
        for item in AppliedControl.objects.filter(id__in=viewable_items):
            if my_map.get(item.folder.name) is None:
                my_map[item.folder.name] = {}
            my_map[item.folder.name].update({item.name: item.id})

        return Response(my_map)

    @action(detail=False, name="Generate data for applied controls impact graph")
    def impact_graph(self, request):
        (viewable_controls_ids, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, AppliedControl
        )
        csf_functions_map = dict()
        categories = [{"name": "--"}]
        for i, option in enumerate(ReferenceControl.CSF_FUNCTION, 1):
            csf_functions_map[option[0]] = i
            categories.append({"name": option[1]})
        categories.append({"name": "requirements"})  # 7
        categories.append({"name": "scenarios"})  # 9
        categories.append({"name": "audits"})  # 8
        categories.append({"name": "risk assessments"})

        nodes = list()
        links = list()
        indexes = dict()
        idx_cnt = 0
        for ac in AppliedControl.objects.filter(id__in=viewable_controls_ids):
            ac_key = f"ac-{ac.id}"
            nodes.append(
                {
                    "name": ac.name,
                    "value": ac.name,
                    "category": csf_functions_map.get(ac.csf_function, 0),
                }
            )
            indexes[ac_key] = idx_cnt
            idx_cnt += 1
            # attached requirement_assessments
            for req in RequirementAssessment.objects.filter(applied_controls__id=ac.id):
                req_key = f"req-{req.id}"
                if req_key not in indexes:
                    # Add RequirementAssessment node only if it doesn't exist
                    nodes.append(
                        {
                            "name": req.requirement.ref_id
                            or req.requirement.safe_display_str,
                            "value": req.requirement.description
                            or req.requirement.safe_display_str,
                            "category": 7,
                            "symbol": "triangle",
                        }
                    )
                    indexes[req_key] = idx_cnt
                    idx_cnt += 1

                audit = req.compliance_assessment
                audit_key = f"audit-{audit.id}"
                if audit_key not in indexes:
                    nodes.append(
                        {
                            "name": audit.name,
                            "value": audit.framework.name,
                            "category": 9,
                            "symbol": "rect",
                        }
                    )
                    indexes[audit_key] = idx_cnt
                    idx_cnt += 1
                links.append({"source": indexes[audit_key], "target": indexes[req_key]})
                links.append({"source": indexes[ac_key], "target": indexes[req_key]})

            for sc in RiskScenario.objects.filter(applied_controls__id=ac.id):
                sc_key = f"sc-{sc.id}"
                if sc_key not in indexes:
                    nodes.append(
                        {
                            "name": sc.ref_id,
                            "value": sc.name,
                            "category": 8,
                            "symbol": "diamond",
                        }
                    )
                    indexes[sc_key] = idx_cnt
                    idx_cnt += 1

                ra = sc.risk_assessment
                ra_key = f"ra-{ra.id}"
                if ra_key not in indexes:
                    nodes.append(
                        {
                            "name": ra.name,
                            "value": ra.name,
                            "category": 10,
                            "symbol": "rect",
                        }
                    )
                    indexes[ra_key] = idx_cnt
                    idx_cnt += 1
                links.append({"source": indexes[ra_key], "target": indexes[sc_key]})
                links.append({"source": indexes[ac_key], "target": indexes[sc_key]})

        return Response({"nodes": nodes, "categories": categories, "links": links})

    @action(detail=False, name="Get applied controls sunburst data")
    def sunburst_data(self, request):
        (viewable_objects, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, AppliedControl
        )
        queryset = AppliedControl.objects.filter(id__in=viewable_objects)

        # Build hierarchical structure: csf_function -> category -> priority -> status
        hierarchy = {}

        for control in queryset.select_related("folder"):
            # Level 1: CSF Function
            csf_function = (
                dict(AppliedControl.CSF_FUNCTION).get(
                    control.csf_function, "No CSF Function"
                )
                if control.csf_function
                else "No CSF Function"
            )
            if csf_function not in hierarchy:
                hierarchy[csf_function] = {}

            # Level 2: Category
            category = (
                dict(AppliedControl.CATEGORY).get(control.category, "No Category")
                if control.category
                else "No Category"
            )
            if category not in hierarchy[csf_function]:
                hierarchy[csf_function][category] = {}

            # Level 3: Priority
            priority = (
                dict(AppliedControl.PRIORITY).get(control.priority, "No Priority")
                if control.priority
                else "No Priority"
            )
            if priority not in hierarchy[csf_function][category]:
                hierarchy[csf_function][category][priority] = {}

            # Level 4: Status
            status = (
                dict(AppliedControl.Status.choices).get(control.status, "No Status")
                if control.status
                else "No Status"
            )
            if status not in hierarchy[csf_function][category][priority]:
                hierarchy[csf_function][category][priority][status] = 0
            hierarchy[csf_function][category][priority][status] += 1

        # CSF Function color mapping (matching NightingaleChart palette)
        csf_color_map = {
            "(undefined)": "#505372",
            "Govern": "#FAE482",
            "Identify": "#85C4EA",
            "Protect": "#B29BBA",
            "Detect": "#FAB647",
            "Respond": "#E47677",
            "Recover": "#8ACB93",
            "No CSF Function": "#505372",
        }

        # Convert to sunburst format
        def build_sunburst_data(data_dict, name="Root", level=0, parent_color=None):
            if isinstance(data_dict, int):
                result = {"name": name, "value": data_dict}
                if parent_color:
                    result["itemStyle"] = {"color": parent_color}
                return result

            children = []
            total_value = 0

            # Determine color for this level
            current_color = parent_color
            if level == 1 and name in csf_color_map:
                current_color = csf_color_map[name]

            for key, value in data_dict.items():
                child = build_sunburst_data(value, key, level + 1, current_color)
                children.append(child)
                total_value += child.get("value", 0)

            result = {"name": name, "value": total_value}
            if children:
                result["children"] = children

            # Apply color at all levels if we have a color
            if current_color and level >= 1:
                result["itemStyle"] = {"color": current_color}

            return result

        sunburst_data = []
        for csf_function, function_data in hierarchy.items():
            function_node = build_sunburst_data(function_data, csf_function, level=1)
            sunburst_data.append(function_node)

        return Response({"results": sunburst_data})


class ActionPlanList(generics.ListAPIView):
    filterset_fields = {
        "folder": ["exact"],
        "status": ["exact"],
        "category": ["exact"],
        "csf_function": ["exact"],
        "priority": ["exact"],
        "reference_control": ["exact"],
        "effort": ["exact"],
        "control_impact": ["exact"],
        "filtering_labels": ["exact"],
        "risk_scenarios": ["exact"],
        "risk_scenarios_e": ["exact"],
        "requirement_assessments": ["exact"],
        "evidences": ["exact"],
        "objectives": ["exact"],
        "assets": ["exact"],
        "stakeholders": ["exact"],
        "progress_field": ["exact"],
        "security_exceptions": ["exact"],
        "owner": ["exact"],
        "findings": ["exact"],
        "eta": ["exact", "lte", "gte", "lt", "gt"],
    }
    search_fields = ["name", "description", "ref_id"]

    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    ordering_fields = "__all__"
    ordering = ["eta"]

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context.update({"pk": self.kwargs["pk"]})
        return context


class UserRolesOnFolderList(generics.ListAPIView):
    filterset_fields = {}
    search_fields = ["email"]
    serializer_class = UserRolesOnFolderSerializer

    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]
    ordering_fields = "__all__"
    ordering = ["email"]

    _user_roles_map = None  # cached variable

    def get_queryset(self):
        folder = get_object_or_404(Folder, id=self.kwargs["pk"])

        # authorize
        (viewable_ids, _updatable_ids, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), self.request.user, Folder
        )
        if folder.id not in viewable_ids:
            raise PermissionDenied()

        # visibility
        visible_ids = set(
            User.visible_users(self.request.user, view_all_users=True).values_list(
                "id", flat=True
            )
        )

        # roles per user (no role filtering)
        raw_map = folder.get_user_roles()  # {user_id: [Role, ...]}

        # keep users that are visible AND have at least one role in raw_map
        self._user_roles_map = {
            uid: roles
            for uid, roles in raw_map.items()
            if uid in visible_ids and roles  # roles non-empty in raw_map
        }

        return User.objects.filter(id__in=self._user_roles_map.keys())

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx.update(
            {
                "pk": self.kwargs["pk"],
                "user_roles_map": self._user_roles_map or {},
            }
        )
        return ctx

    def list(self, request, *args, **kwargs):
        if not RoleAssignment.is_access_allowed(
            user=self.request.user,
            perm=Permission.objects.get(codename="change_folder"),
            folder=get_object_or_404(Folder, id=self.kwargs["pk"]),
        ):
            raise PermissionDenied()

        return super().list(request, *args, **kwargs)


class ComplianceAssessmentActionPlanList(ActionPlanList):
    serializer_class = ComplianceAssessmentActionPlanSerializer

    def get_queryset(self):
        """RBAC not automatic as we don't inherit from BaseModelViewSet -> enforce it explicitly"""
        compliance_id = self.kwargs["pk"]

        if not RoleAssignment.is_object_readable(
            self.request.user,
            ComplianceAssessment,
            compliance_id,
        ):
            raise PermissionDenied()

        assessment = ComplianceAssessment.objects.get(id=compliance_id)
        requirement_assessments = assessment.get_requirement_assessments(
            include_non_assessable=True
        )

        qs = AppliedControl.objects.filter(
            requirement_assessments__in=requirement_assessments
        ).distinct()

        viewable_controls, _, _ = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(),
            self.request.user,
            AppliedControl,
        )
        return qs.filter(id__in=viewable_controls)


class ComplianceAssessmentEvidenceList(generics.ListAPIView):
    serializer_class = ComplianceAssessmentEvidenceSerializer
    filterset_fields = {
        "folder": ["exact"],
        "status": ["exact"],
        "owner": ["exact"],
        "name": ["icontains"],
        "expiry_date": ["exact", "lte", "gte"],
        "created_at": ["exact", "lte", "gte"],
        "updated_at": ["exact", "lte", "gte"],
    }
    search_fields = ["name", "description"]
    ordering_fields = ["name", "status", "updated_at", "expiry_date"]
    ordering = ["name"]

    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context.update({"pk": self.kwargs["pk"]})
        return context

    def get_queryset(self):
        """RBAC not automatic as we don't inherit from BaseModelViewSet -> enforce it explicitly"""
        compliance_id = self.kwargs["pk"]

        if not RoleAssignment.is_object_readable(
            self.request.user,
            ComplianceAssessment,
            compliance_id,
        ):
            raise PermissionDenied()

        compliance_assessment = ComplianceAssessment.objects.get(id=compliance_id)

        # Get all requirement assessments for this compliance assessment
        requirement_assessments = RequirementAssessment.objects.filter(
            compliance_assessment=compliance_assessment
        ).prefetch_related("evidences", "applied_controls__evidences")

        # Get visible evidences to filter result
        viewable_evidences, _, _ = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), self.request.user, Evidence
        )

        # Collect evidence IDs from both direct and indirect relationships
        evidence_ids = set()
        for req_assessment in requirement_assessments:
            for evidence in req_assessment.evidences.all():
                if evidence.id in viewable_evidences:
                    evidence_ids.add(evidence.id)
            for applied_control in req_assessment.applied_controls.all():
                for evidence in applied_control.evidences.all():
                    if evidence.id in viewable_evidences:
                        evidence_ids.add(evidence.id)

        return Evidence.objects.filter(id__in=evidence_ids).distinct()


class RiskAssessmentActionPlanList(ActionPlanList):
    serializer_class = RiskAssessmentActionPlanSerializer

    def get_queryset(self):
        """RBAC not automatic as we don't inherit from BaseModelViewSet -> enforce it explicitly"""
        risk_id = self.kwargs["pk"]

        if not RoleAssignment.is_object_readable(
            self.request.user,
            RiskAssessment,
            risk_id,
        ):
            raise PermissionDenied()

        assessment = RiskAssessment.objects.get(id=risk_id)
        risk_scenarios = assessment.risk_scenarios.all()

        qs = AppliedControl.objects.filter(
            Q(risk_scenarios__in=risk_scenarios)
            | Q(risk_scenarios_e__in=risk_scenarios)
        ).distinct()

        viewable_controls, _, _ = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(),
            self.request.user,
            AppliedControl,
        )
        return qs.filter(id__in=viewable_controls)


class PolicyViewSet(AppliedControlViewSet):
    model = Policy
    filterset_fields = [
        "folder",
        "csf_function",
        "status",
        "reference_control",
        "effort",
        "risk_scenarios",
        "requirement_assessments",
        "evidences",
        "genericcollection",
    ]
    search_fields = ["name", "description", "ref_id"]

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get csf_function choices")
    def csf_function(self, request):
        return Response(dict(AppliedControl.CSF_FUNCTION))


class RiskScenarioFilter(GenericFilterSet):
    risk_assessment = df.ModelMultipleChoiceFilter(
        queryset=RiskAssessment.objects.all()
    )
    # Aliased filters for user-friendly query params
    folder = df.UUIDFilter(
        field_name="risk_assessment__perimeter__folder", label="Folder ID"
    )
    perimeter = df.UUIDFilter(
        field_name="risk_assessment__perimeter", label="Perimeter ID"
    )
    within_tolerance = df.ChoiceFilter(
        choices=[("YES", "YES"), ("NO", "NO"), ("--", "--")],
        method="filter_within_tolerance",
    )
    applied_controls = df.ModelMultipleChoiceFilter(
        method="filter_applied_controls",
        queryset=AppliedControl.objects.all(),
    )
    exclude = df.UUIDFilter(
        method="filter_exclude",
        label="Exclude scenario",
    )

    def filter_within_tolerance(self, queryset, name, value):
        if value == "YES":
            return queryset.filter(
                risk_assessment__risk_tolerance__gte=0,
                current_level__lte=models.F("risk_assessment__risk_tolerance"),
            )
        elif value == "NO":
            return queryset.filter(
                risk_assessment__risk_tolerance__gte=0,
                current_level__gt=models.F("risk_assessment__risk_tolerance"),
            )
        elif value == "--":
            return queryset.filter(risk_assessment__risk_tolerance__lt=0)
        return queryset

    def filter_applied_controls(self, queryset, name, value):
        """Filter by both extra controls (applied_controls) and existing controls (existing_applied_controls)"""
        if value:
            return queryset.filter(
                Q(applied_controls__in=value) | Q(existing_applied_controls__in=value)
            ).distinct()
        return queryset

    def filter_exclude(self, queryset, name, value):
        """Exclude a specific scenario from the queryset"""
        if value:
            return queryset.exclude(id=value)
        return queryset

    class Meta:
        model = RiskScenario
        # Only include actual model fields here
        fields = {
            "name": ["exact"],
            "current_impact": ["exact"],
            "current_proba": ["exact"],
            "current_level": ["exact"],
            "residual_impact": ["exact"],
            "residual_proba": ["exact"],
            "residual_level": ["exact"],
            "treatment": ["exact"],
            "threats": ["exact"],
            "assets": ["exact"],
            "existing_applied_controls": ["exact"],
            "security_exceptions": ["exact"],
            "owner": ["exact"],
        }


class RiskScenarioViewSet(ExportMixin, BaseModelViewSet):
    """
    API endpoint that allows risk scenarios to be viewed or edited.
    """

    model = RiskScenario
    filterset_class = RiskScenarioFilter
    ordering = ["ref_id"]
    search_fields = ["name", "description", "ref_id"]

    export_config = {
        "fields": {
            "internal_id": {"source": "id", "label": "internal_id"},
            "ref_id": {"source": "ref_id", "label": "ref_id", "escape": True},
            "name": {"source": "name", "label": "name", "escape": True},
            "description": {
                "source": "description",
                "label": "description",
                "escape": True,
            },
            "risk_assessment": {
                "source": "risk_assessment.name",
                "label": "risk_assessment",
                "escape": True,
            },
            "treatment": {"source": "get_treatment_display", "label": "treatment"},
            "inherent_probability": {
                "source": "get_inherent_proba",
                "label": "inherent_probability",
                "format": lambda v: v.get("name", "--"),
            },
            "inherent_impact": {
                "source": "get_inherent_impact",
                "label": "inherent_impact",
                "format": lambda v: v.get("name", "--"),
            },
            "inherent_level": {
                "source": "get_inherent_risk",
                "label": "inherent_level",
                "format": lambda v: v.get("name", "--"),
            },
            "current_probability": {
                "source": "get_current_proba",
                "label": "current_probability",
                "format": lambda v: v.get("name", "--"),
            },
            "current_impact": {
                "source": "get_current_impact",
                "label": "current_impact",
                "format": lambda v: v.get("name", "--"),
            },
            "current_level": {
                "source": "get_current_risk",
                "label": "current_level",
                "format": lambda v: v.get("name", "--"),
            },
            "residual_probability": {
                "source": "get_residual_proba",
                "label": "residual_probability",
                "format": lambda v: v.get("name", "--"),
            },
            "residual_impact": {
                "source": "get_residual_impact",
                "label": "residual_impact",
                "format": lambda v: v.get("name", "--"),
            },
            "residual_level": {
                "source": "get_residual_risk",
                "label": "residual_level",
                "format": lambda v: v.get("name", "--"),
            },
            "owners": {
                "source": "owner",
                "label": "owners",
                "format": lambda qs: ",".join(
                    escape_excel_formula(o.email) for o in qs.all()
                ),
            },
            "threats": {
                "source": "threats",
                "label": "threats",
                "format": lambda qs: ",".join(
                    escape_excel_formula(t.name) for t in qs.all()
                ),
            },
            "assets": {
                "source": "assets",
                "label": "assets",
                "format": lambda qs: ",".join(
                    escape_excel_formula(a.name) for a in qs.all()
                ),
            },
            "vulnerabilities": {
                "source": "vulnerabilities",
                "label": "vulnerabilities",
                "format": lambda qs: ",".join(
                    escape_excel_formula(v.name) for v in qs.all()
                ),
            },
            "applied_controls": {
                "source": "applied_controls",
                "label": "applied_controls",
                "format": lambda qs: ",".join(
                    escape_excel_formula(c.name) for c in qs.all()
                ),
            },
            "existing_applied_controls": {
                "source": "existing_applied_controls",
                "label": "existing_applied_controls",
                "format": lambda qs: ",".join(
                    escape_excel_formula(c.name) for c in qs.all()
                ),
            },
            "qualifications": {
                "source": "qualifications",
                "label": "qualifications",
                "format": lambda qs: ",".join(
                    escape_excel_formula(q.name) for q in qs.all()
                ),
            },
        },
        "filename": "risk_scenarios_export",
        "select_related": ["risk_assessment"],
        "prefetch_related": [
            "owner",
            "threats",
            "assets",
            "vulnerabilities",
            "applied_controls",
            "existing_applied_controls",
            "qualifications",
        ],
    }

    def get_queryset(self):
        queryset = super().get_queryset()
        return queryset.select_related(
            "risk_assessment",
            "risk_assessment__risk_matrix",
            "risk_assessment__perimeter",
            "risk_assessment__perimeter__folder",
        ).prefetch_related(
            "threats",
            "assets",
            "applied_controls",
            "existing_applied_controls",
            "owner",
            "security_exceptions",
        )

    def _perform_write(self, serializer):
        if not serializer.validated_data.get(
            "ref_id"
        ) and serializer.validated_data.get("risk_assessment"):
            risk_assessment = serializer.validated_data["risk_assessment"]
            ref_id = RiskScenario.get_default_ref_id(risk_assessment)
            serializer.validated_data["ref_id"] = ref_id
        serializer.save()

    def perform_create(self, serializer):
        return self._perform_write(serializer)

    def perform_update(self, serializer):
        return self._perform_write(serializer)

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get treatment choices")
    def treatment(self, request):
        return Response(dict(RiskScenario.TREATMENT_OPTIONS))

    @action(detail=False, name="Get qualifications count")
    def qualifications_count(self, request):
        folder_id = request.query_params.get("folder", None)
        return Response(
            {"results": qualifications_count_per_name(request.user, folder_id)}
        )

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=True, name="Get probability choices")
    def probability(self, request, pk):
        undefined = {-1: "--"}
        _choices = {
            i: name
            for i, name in enumerate(
                x["name"] for x in self.get_object().get_matrix()["probability"]
            )
        }
        choices = undefined | _choices
        return Response(choices)

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=True, name="Get impact choices")
    def impact(self, request, pk):
        undefined = dict([(-1, "--")])
        _choices = dict(
            zip(
                list(range(0, 64)),
                [x["name"] for x in self.get_object().get_matrix()["impact"]],
            )
        )
        choices = undefined | _choices
        return Response(choices)

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=True, name="Get strength of knowledge choices")
    def strength_of_knowledge(self, request, pk):
        undefined = {-1: RiskScenario.DEFAULT_SOK_OPTIONS[-1]}
        _sok_choices = self.get_object().get_matrix().get("strength_of_knowledge")
        if _sok_choices is not None:
            sok_choices = dict(
                enumerate(
                    {
                        "name": x["name"],
                        "description": x.get("description"),
                        "symbol": x.get("symbol"),
                    }
                    for x in _sok_choices
                )
            )
        else:
            sok_choices = RiskScenario.DEFAULT_SOK_OPTIONS
        choices = undefined | sok_choices
        return Response(choices)

    @action(detail=False, name="Get risk count per level")
    def count_per_level(self, request):
        folder_id = request.query_params.get("folder", None)
        return Response(
            {
                "results": risks_count_per_level(
                    request.user,
                    None,
                    folder_id,
                    include_inherent=ff_is_enabled("inherent_risk"),
                )
            }
        )

    @action(detail=False, name="Get risk scenarios count per status")
    def per_status(self, request):
        return Response({"results": risk_per_status(request.user)})

    @action(
        detail=False, methods=["get"], permission_classes=[permissions.IsAuthenticated]
    )
    def default_ref_id(self, request):
        risk_assessment_id = request.query_params.get("risk_assessment")
        if not risk_assessment_id:
            return Response(
                {"error": "Missing 'risk_assessment' parameter."}, status=400
            )
        try:
            risk_assessment = RiskAssessment.objects.get(pk=risk_assessment_id)

            # Use the class method to compute the default ref_id
            default_ref_id = RiskScenario.get_default_ref_id(risk_assessment)
            return Response({"results": default_ref_id})
        except Exception as e:
            logger.error("Error in default_ref_id: %s", str(e))
            return Response(
                {"error": "Error in default_ref_id has occurred."}, status=400
            )

    @action(
        detail=True,
        methods=["post"],
        url_path="sync-to-actions",
    )
    def sync_to_applied_controls(self, request, pk):
        dry_run = request.query_params.get("dry_run", True)
        if dry_run == "false":
            dry_run = False
        reset_residual = request.data.get("reset_residual", False)
        risk_scenario = RiskScenario.objects.get(id=pk)

        if not RoleAssignment.is_access_allowed(
            user=request.user,
            perm=Permission.objects.get(codename="change_riskscenario"),
            folder=Folder.get_folder(risk_scenario),
        ):
            return Response(status=status.HTTP_403_FORBIDDEN)

        changes = risk_scenario.sync_to_applied_controls(
            reset_residual=reset_residual, dry_run=dry_run
        )
        return Response(
            {"changes": AppliedControlReadSerializer(changes, many=True).data}
        )


class RiskAcceptanceFilterSet(GenericFilterSet):
    folder = df.ModelMultipleChoiceFilter(queryset=Folder.objects.all())
    approver = df.ModelMultipleChoiceFilter(queryset=User.objects.all())

    to_review = df.BooleanFilter(method="filter_to_review")

    def filter_to_review(self, queryset, name, value):
        if value:
            return (
                queryset.filter(expiry_date__lte=date.today() + timedelta(days=30))
                .filter(state__in=["submitted", "accepted"])
                .order_by("expiry_date")
            )

        return queryset

    class Meta:
        model = RiskAcceptance
        fields = {
            "state": ["exact"],
            "risk_scenarios": ["exact"],
            "expiry_date": ["exact", "lte", "gte", "lt", "gt", "month", "year"],
        }


class RiskAcceptanceViewSet(BaseModelViewSet):
    """
    API endpoint that allows risk acceptance to be viewed or edited.
    """

    permission_overrides = {
        "accept": "approve_riskacceptance",
        "reject": "approve_riskacceptance",
        "revoke": "approve_riskacceptance",
    }

    model = RiskAcceptance
    serializer_class = RiskAcceptanceWriteSerializer
    filterset_class = RiskAcceptanceFilterSet
    search_fields = ["name", "description", "justification"]

    def update(self, request, *args, **kwargs):
        initial_data = self.get_object()
        updated_data = request.data
        if (
            updated_data.get("justification") != initial_data.justification
            and request.user != initial_data.approver
        ):
            _data = {
                "non_field_errors": "The justification can only be edited by the approver"
            }
            return Response(data=_data, status=status.HTTP_400_BAD_REQUEST)
        else:
            return super().update(request, *args, **kwargs)

    @action(detail=False, name="Get acceptances to review")
    def to_review(self, request):
        acceptances = acceptances_to_review(request.user)

        acceptances = [
            RiskAcceptanceReadSerializer(acceptance).data for acceptance in acceptances
        ]

        """
        The serializer of AppliedControl isn't applied automatically for this function
        """

        return Response({"results": acceptances})

    @action(detail=True, methods=["post"], name="Submit risk acceptance")
    def submit(self, request, pk):
        if self.get_object().approver:
            self.get_object().set_state("submitted")
            return Response({"results": "state updated to submitted"})
        else:
            return Response(
                {"error": "Missing 'approver' field"}, status=status.HTTP_403_FORBIDDEN
            )

    # This set back risk acceptance to "Created"
    @action(detail=True, methods=["post"], name="Draft risk acceptance")
    def draft(self, request, pk):
        self.get_object().set_state("created")
        return Response({"results": "state updated back to created"})

    @action(detail=True, methods=["post"], name="Accept risk acceptance")
    def accept(self, request, pk):
        if request.user != self.get_object().approver:
            logger.error(
                "Only the approver can accept the risk acceptance",
                user=request.user,
                approver=self.get_object().approver,
            )
            raise PermissionDenied(
                {"error": "Only the approver can accept the risk acceptance"}
            )
        self.get_object().set_state("accepted")
        return Response({"results": "state updated to accepted"})

    @action(detail=True, methods=["post"], name="Reject risk acceptance")
    def reject(self, request, pk):
        if request.user != self.get_object().approver:
            logger.error(
                "Only the approver can reject the risk acceptance",
                user=request.user,
                approver=self.get_object().approver,
            )
            raise PermissionDenied(
                {"error": "Only the approver can reject the risk acceptance"}
            )
        self.get_object().set_state("rejected")
        return Response({"results": "state updated to rejected"})

    @action(detail=True, methods=["post"], name="Revoke risk acceptance")
    def revoke(self, request, pk):
        if request.user != self.get_object().approver:
            logger.error(
                "Only the approver can revoke the risk acceptance",
                user=request.user,
                approver=self.get_object().approver,
            )
            raise PermissionDenied(
                {"error": "Only the approver can revoke the risk acceptance"}
            )
        self.get_object().set_state("revoked")
        return Response({"results": "state updated to revoked"})

    @action(detail=False, methods=["get"], name="Get waiting risk acceptances")
    def waiting(self, request):
        acceptance_count = RiskAcceptance.objects.filter(
            approver=request.user, state="submitted"
        ).count()
        return Response({"count": acceptance_count})

    def perform_update(self, serializer):
        risk_acceptance = serializer.validated_data

        if risk_acceptance.get("approver"):
            for scenario in risk_acceptance.get("risk_scenarios"):
                if not RoleAssignment.is_access_allowed(
                    risk_acceptance.get("approver"),
                    Permission.objects.get(codename="approve_riskacceptance"),
                    scenario.risk_assessment.perimeter.folder,
                ):
                    raise ValidationError(
                        "The approver is not allowed to approve this risk acceptance"
                    )
        risk_acceptance = serializer.save()
        dispatch_webhook_event(risk_acceptance, "updated", serializer)

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get state choices")
    def state(self, request):
        return Response(dict(RiskAcceptance.ACCEPTANCE_STATE))


class UserFilter(GenericFilterSet):
    is_approver = df.BooleanFilter(method="filter_approver", label="Approver")
    is_applied_control_owner = df.BooleanFilter(
        method="filter_applied_control_owner", label="Applied control owner"
    )
    exclude_current = df.BooleanFilter(
        method="filter_exclude_current", label="Exclude current user"
    )

    def filter_approver(self, queryset, name, value):
        """we don't know yet which folders will be used, so filter on any folder"""
        approvers_id = []
        for candidate in User.objects.all():
            if "approve_riskacceptance" in candidate.permissions:
                approvers_id.append(candidate.id)
        if value:
            return queryset.filter(id__in=approvers_id)
        return queryset.exclude(id__in=approvers_id)

    def filter_applied_control_owner(self, queryset, name, value):
        return queryset.filter(applied_controls__isnull=not value)

    def filter_exclude_current(self, queryset, name, value):
        """Exclude the current user from the queryset"""
        if value and self.request and self.request.user:
            return queryset.exclude(id=self.request.user.id)
        return queryset

    class Meta:
        model = User
        fields = [
            "email",
            "first_name",
            "last_name",
            "is_active",
            "keep_local_login",
            "is_approver",
            "is_third_party",
            "expiry_date",
            "user_groups",
            "exclude_current",
        ]


class ValidationFlowFilterSet(GenericFilterSet):
    folder = df.ModelMultipleChoiceFilter(queryset=Folder.objects.all())
    requester = df.ModelMultipleChoiceFilter(queryset=User.objects.all())
    approver = df.ModelMultipleChoiceFilter(queryset=User.objects.all())

    linked_models = df.CharFilter(method="filter_linked_models", label="Linked models")

    def filter_linked_models(self, queryset, name, value):
        """
        Filter validation flows by linked model types.
        Usage: ?linked_models=risk_assessments,evidences
        """
        if not value:
            return queryset

        model_types = [m.strip() for m in value.split(",")]
        filtered_qs = queryset

        for model_type in model_types:
            # Use the field name with __isnull=False to check if objects are linked
            filter_kwargs = {f"{model_type}__isnull": False}
            filtered_qs = filtered_qs.filter(**filter_kwargs).distinct()

        return filtered_qs

    class Meta:
        model = ValidationFlow
        fields = [
            "status",
            "filtering_labels",
            "compliance_assessments",
            "risk_assessments",
            "crq_studies",
            "ebios_studies",
            "entity_assessments",
            "findings_assessments",
            "evidences",
            "security_exceptions",
            "policies",
        ]


class ValidationFlowViewSet(BaseModelViewSet):
    """
    API endpoint that allows validation flows to be viewed or edited.
    """

    model = ValidationFlow
    serializer_class = ValidationFlowWriteSerializer
    filterset_class = ValidationFlowFilterSet
    search_fields = ["ref_id", "request_notes"]

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get status choices")
    def status(self, request):
        return Response(dict(ValidationFlow.Status.choices))

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get linked models choices")
    def linked_models(self, request):
        """Return available model types that can be linked to validation flows"""
        model_types = {
            "compliance_assessments": "Compliance Assessments",
            "risk_assessments": "Risk Assessments",
            "business_impact_analysis": "Business Impact Analysis",
            "crq_studies": "Quantitative Risk Studies",
            "ebios_studies": "EBIOS RM Studies",
            "entity_assessments": "Entity Assessments",
            "findings_assessments": "Findings Assessments",
            "evidences": "Evidences",
            "security_exceptions": "Security Exceptions",
            "policies": "Policies",
        }
        return Response(model_types)

    @action(
        detail=False, methods=["get"], permission_classes=[permissions.IsAuthenticated]
    )
    def default_ref_id(self, request):
        try:
            # Use the class method to compute the default ref_id (globally unique)
            default_ref_id = ValidationFlow.get_default_ref_id()
            return Response({"results": default_ref_id})
        except Exception as e:
            logger.error("Error in default_ref_id: %s", str(e))
            return Response(
                {"error": "Error in default_ref_id has occurred."}, status=400
            )


class UserViewSet(BaseModelViewSet):
    """
    API endpoint that allows users to be viewed or edited
    """

    model = User
    ordering = ["-is_active", "-is_superuser", "email", "id"]
    filterset_class = UserFilter
    search_fields = ["email", "first_name", "last_name"]

    def get_queryset(self):
        return User.visible_users(self.request.user, view_all_users=True)

    def update(self, request: Request, *args, **kwargs) -> Response:
        user = self.get_object()
        if user.is_admin():
            number_of_admin_users = User.get_admin_users().count()
            admin_group = UserGroup.objects.get(name="BI-UG-ADM")
            if number_of_admin_users == 1:
                new_user_groups = set(request.data["user_groups"])
                if str(admin_group.pk) not in new_user_groups:
                    return Response(
                        {"error": "attemptToRemoveOnlyAdminUserGroup"},
                        status=status.HTTP_403_FORBIDDEN,
                    )

        return super().update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        user = self.get_object()
        if user.is_admin():
            number_of_admin_users = User.get_admin_users().count()
            if number_of_admin_users == 1:
                return Response(
                    {"error": "attemptToDeleteOnlyAdminAccountError"},
                    status=status.HTTP_403_FORBIDDEN,
                )

        return super().destroy(request, *args, **kwargs)


class UserGroupOrderingFilter(filters.OrderingFilter):
    """
    Custom ordering filter:
    - Performs in-memory (Python) sorting only for `localization_dict`.
    - The sort key is a tuple: (folder full_path OR folder.name, object.name).
    - Supports `-localization_dict` for descending order.
    - For all other fields, it falls back to standard SQL ordering.
    """

    def filter_queryset(self, request, queryset, view):
        ordering = self.get_ordering(request, queryset, view)
        if not ordering:
            return queryset

        # Special case: in-memory sorting for `localization_dict`
        if len(ordering) == 1 and ordering[0].lstrip("-") == "localization_dict":
            desc = ordering[0].startswith("-")

            # Optimize DB access: fetch the related folder in one query
            queryset = queryset.select_related("folder")

            # Materialize queryset into a list to sort in Python
            data = list(queryset)

            def full_path_or_name(folder):
                """
                Build a string key from the folder:
                - Prefer the full path (names of all parent folders + current).
                - Fall back to the folder's name if no path is available.
                """
                if folder is None:
                    return ""

                path_list = getattr(folder, "get_folder_full_path", None)
                if callable(path_list):
                    items = folder.get_folder_full_path(include_root=False)
                    names = [getattr(f, "name", "") or "" for f in items]
                    if names:
                        return "/".join(names)

                # Fallback: just the folder name
                return getattr(folder, "name", "") or ""

            def key_func(obj):
                # Get the folder from the object
                folder = getattr(obj, "folder", None)
                # If you want to be more robust, you could use:
                # from yourapp.models import Folder
                # folder = Folder.get_folder(obj)

                path_key = full_path_or_name(folder).casefold()
                name_key = (getattr(obj, "name", "") or "").casefold()
                return (path_key, name_key)

            # Perform stable sort, reverse if `-localization_dict`
            data.sort(key=key_func, reverse=desc)
            return data

        # Default case: fall back to SQL ordering
        return super().filter_queryset(request, queryset, view)


class UserGroupViewSet(BaseModelViewSet):
    """
    API endpoint that allows user groups to be viewed or edited
    """

    model = UserGroup
    ordering = ["builtin", "name"]
    ordering_fields = ["localization_dict"]
    filterset_fields = ["folder"]
    search_fields = [
        "folder__name"
    ]  # temporary hack, filters only by folder name, not role name
    filter_backends = [
        DjangoFilterBackend,
        UserGroupOrderingFilter,
        filters.SearchFilter,
    ]


class RoleAssignmentViewSet(BaseModelViewSet):
    """
    API endpoint that allows role assignments to be viewed or edited.
    """

    model = RoleAssignment
    ordering = ["builtin", "folder"]
    filterset_fields = ["folder"]


class FolderFilter(GenericFilterSet):
    parent_folder = df.ModelMultipleChoiceFilter(queryset=Folder.objects.all())

    owned = df.BooleanFilter(method="get_owned_folders", label="owned")
    content_type = df.MultipleChoiceFilter(
        choices=Folder.ContentType, lookup_expr="icontains"
    )

    def get_owned_folders(self, queryset, name, value):
        owned_folders_id = []
        for folder in Folder.objects.all():
            if folder.owner.all().first():
                owned_folders_id.append(folder.id)
        if value:
            return queryset.filter(id__in=owned_folders_id)
        return queryset.exclude(id__in=owned_folders_id)

    class Meta:
        model = Folder
        fields = [
            "name",
            "content_type",
            "owner",
            "owned",
            "filtering_labels",
        ]


class FolderViewSet(BaseModelViewSet):
    """
    API endpoint that allows folders to be viewed or edited.
    """

    model = Folder
    filterset_class = FolderFilter
    search_fields = ["name"]
    batch_size = 100  # Configurable batch size for processing domain import

    def perform_create(self, serializer):
        """
        Create the default user groups after domain creation
        """
        serializer.save()
        folder = Folder.objects.get(id=serializer.data["id"])
        Folder.create_default_ug_and_ra(folder)

    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    def retrieve(self, request, *args, **kwargs):
        return super().retrieve(request, *args, **kwargs)

    @action(detail=False, methods=["get"])
    def org_tree(self, request):
        """
        Returns the tree of domains and perimeters
        """
        include_perimeters = request.query_params.get(
            "include_perimeters", "True"
        ).lower() in ["true", "1", "yes"]

        (viewable_objects, _, _) = RoleAssignment.get_accessible_object_ids(
            folder=Folder.get_root_folder(),
            user=request.user,
            object_type=Folder,
        )

        # Add ancestors so viewable folders aren't orphaned
        needed_folders = set(viewable_objects)

        for folder_id in viewable_objects:
            current = Folder.objects.get(pk=folder_id)
            while current and current.parent_folder_id:
                needed_folders.add(current.parent_folder_id)
                current = Folder.objects.get(pk=current.parent_folder_id)

        folders_list = []
        for folder in (
            Folder.objects.exclude(content_type="GL")
            .filter(id__in=needed_folders, parent_folder=Folder.get_root_folder())
            .distinct()
        ):
            entry = {
                "name": folder.name,
                "uuid": folder.id,
                "viewable": folder.id in viewable_objects,
            }
            folder_content = get_folder_content(
                folder,
                include_perimeters=include_perimeters,
                viewable_objects=viewable_objects,
                needed_folders=needed_folders,
            )
            if len(folder_content) > 0:
                entry.update({"children": folder_content})
            folders_list.append(entry)

        return Response({"name": "Global", "children": folders_list})

    @action(detail=False, methods=["get"])
    def ids(self, request):
        my_map = dict()

        (viewable_items, _, _) = RoleAssignment.get_accessible_object_ids(
            folder=Folder.get_root_folder(),
            user=request.user,
            object_type=Folder,
        )
        for item in Folder.objects.filter(id__in=viewable_items):
            my_map[item.name] = item.id
        return Response(my_map)

    @action(detail=False, methods=["get"])
    def my_assignments(self, request):
        risk_assessments = RiskAssessment.objects.filter(
            Q(authors=request.user) | Q(reviewers=request.user)
        ).distinct()

        audits = (
            ComplianceAssessment.objects.filter(
                Q(authors=request.user) | Q(reviewers=request.user)
            )
            .order_by(F("eta").asc(nulls_last=True))
            .distinct()
        )

        sum = 0
        avg_progress = 0
        audits_count = audits.count()
        if audits_count > 0:
            for audit in audits:
                sum += audit.get_progress()
            avg_progress = int(sum / audits.count())

        controls = (
            AppliedControl.objects.filter(owner=request.user)
            .order_by(F("eta").asc(nulls_last=True))
            .distinct()
        )
        non_active_controls = controls.exclude(status="active")
        risk_scenarios = RiskScenario.objects.filter(owner=request.user).distinct()
        controls_progress = 0
        evidences_progress = 0
        tot_ac = controls.count()
        if tot_ac > 0:
            alive_ac = controls.filter(status="active").count()
            controls_progress = int((alive_ac / tot_ac) * 100)

            with_evidences = 0
            for ctl in controls:
                with_evidences += 1 if ctl.has_evidences() else 0

            evidences_progress = int((with_evidences / tot_ac) * 100)

        RA_serializer = RiskAssessmentReadSerializer(risk_assessments[:10], many=True)
        CA_serializer = ComplianceAssessmentReadSerializer(audits[:6], many=True)
        AC_serializer = AppliedControlReadSerializer(
            non_active_controls[:10], many=True
        )
        RS_serializer = RiskScenarioReadSerializer(risk_scenarios[:10], many=True)

        return Response(
            {
                "risk_assessments": RA_serializer.data,
                "audits": CA_serializer.data,
                "controls": AC_serializer.data,
                "risk_scenarios": RS_serializer.data,
                "metrics": {
                    "progress": {
                        "audits": avg_progress,
                        "controls": controls_progress,
                        "evidences": evidences_progress,
                    }
                },
            }
        )

    @action(detail=True, methods=["get"])
    def export(self, request, pk):
        include_attachments = True
        instance = self.get_object()

        logger.info(
            "Starting domain export",
            domain_id=instance.id,
            domain_name=instance.name,
            include_attachments=include_attachments,
            user=request.user.username,
        )

        objects = get_domain_export_objects(instance)

        for model in objects.keys():
            if not RoleAssignment.is_access_allowed(
                user=request.user,
                perm=Permission.objects.get(codename=f"view_{model}"),
                folder=instance,
            ):
                logger.error(
                    "User does not have permission to export object",
                    user=request.user,
                    model=model,
                )
                raise PermissionDenied(
                    {"error": "userDoesNotHavePermissionToExportDomain"}
                )

        logger.debug(
            "Retrieved domain objects for export",
            object_types=list(objects.keys()),
            total_objects=sum(len(queryset) for queryset in objects.values()),
            objects_per_model={
                model: len(queryset) for model, queryset in objects.items()
            },
        )

        # Create in-memory zip file
        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
            if include_attachments:
                revisions = objects.get(
                    "evidencerevision", EvidenceRevision.objects.none()
                ).filter(attachment__isnull=False)
                logger.info(
                    "Processing evidence attachments",
                    total_revisions=revisions.count(),
                    domain_id=instance.id,
                )

                for revision in revisions:
                    if revision.attachment and default_storage.exists(
                        revision.attachment.name
                    ):
                        # Read file directly into memory
                        with default_storage.open(revision.attachment.name) as file:
                            file_content = file.read()
                            # Write the file content directly to the zip
                            zipf.writestr(
                                os.path.join(
                                    "attachments",
                                    "evidence-revisions",
                                    f"{revision.evidence_id}_v{revision.version}_"
                                    f"{os.path.basename(revision.attachment.name)}",
                                ),
                                file_content,
                            )

            # Add the JSON dump to the zip file
            dumpfile_name = (
                f"ciso-assistant-{slugify(instance.name)}-domain-{timezone.now()}"
            )
            dump_data = ExportSerializer.dump_data(scope=[*objects.values()])

            logger.debug(
                "Adding JSON dump to zip",
                json_size=len(json.dumps(dump_data).encode("utf-8")),
                filename=f"{dumpfile_name}.json",
            )

            zipf.writestr("data.json", json.dumps(dump_data).encode("utf-8"))

        # Reset buffer position to the start
        zip_buffer.seek(0)
        final_size = len(zip_buffer.getvalue())

        # Create the response with the in-memory zip file
        response = HttpResponse(zip_buffer.getvalue(), content_type="application/zip")
        response["Content-Disposition"] = f'attachment; filename="{dumpfile_name}.zip"'

        logger.info(
            "Domain export completed successfully",
            domain_id=instance.id,
            domain_name=instance.name,
            zip_size=final_size,
            filename=f"{dumpfile_name}.zip",
        )

        return response

    @action(
        detail=False,
        methods=["post"],
        url_path="import",
        parser_classes=(FileUploadParser,),
    )
    def import_domain(self, request):
        """Handle file upload and initiate import process."""
        load_missing_libraries = (
            request.query_params.get("load_missing_libraries", "false").lower()
            == "true"
        )
        try:
            if not RoleAssignment.is_access_allowed(
                user=request.user,
                perm=Permission.objects.get(codename="add_folder"),
                folder=Folder.get_root_folder(),
            ):
                raise PermissionDenied()
            domain_name = request.headers.get(
                "X-CISOAssistantDomainName", str(uuid.uuid4())
            )
            parsed_data = self._process_uploaded_file(request.data["file"])
            result = self._import_objects(
                parsed_data, domain_name, load_missing_libraries, user=request.user
            )
            return Response(result, status=status.HTTP_200_OK)

        except PermissionDenied:
            logger.error(
                "User does not have permission to import domain",
                user=request.user,
                exc_info=True,
            )
            return Response(
                {"error": "userDoesNotHavePermissionToImportDomain"},
                status=status.HTTP_403_FORBIDDEN,
            )

        except KeyError:
            logger.error("No file provided in the request", exc_info=True)
            return Response(
                {"errors": ["No file provided"]}, status=status.HTTP_400_BAD_REQUEST
            )

        except json.JSONDecodeError:
            logger.error("Invalid JSON format in uploaded file", exc_info=True)
            return Response(
                {"errors": ["Invalid JSON format"]}, status=status.HTTP_400_BAD_REQUEST
            )

    @action(
        detail=False,
        methods=["post"],
        url_path="import-dummy",
    )
    def import_dummy_domain(self, request):
        domain_name = "DEMO"
        try:
            PROJECT_DIR = Path(__file__).resolve().parent.parent
            dummy_fixture_path = PROJECT_DIR / "fixtures" / "dummy-domain.bak"
            if not dummy_fixture_path.exists():
                logger.error("Dummy domain fixture not found", path=dummy_fixture_path)
                return Response(
                    {"error": "dummyDomainFixtureNotFound"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )
            parsed_data = self._process_uploaded_file(dummy_fixture_path)
            result = self._import_objects(
                parsed_data, domain_name, load_missing_libraries=True, user=request.user
            )
            logger.info("Dummy domain imported successfully", domain_name=domain_name)
            return Response(result, status=status.HTTP_200_OK)

        except json.JSONDecodeError:
            logger.error("Invalid JSON format in dummy fixture file")
            return Response(
                {"errors": ["Invalid JSON format"]},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
        except Exception:
            logger.error("Error importing dummy domain")
            return Response(
                {"error": "failedToImportDummyDomain"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    def _process_uploaded_file(self, dump_file: str | Path) -> Any:
        """Process the uploaded file and return parsed data."""
        if not zipfile.is_zipfile(dump_file):
            logger.error("Invalid ZIP file format")
            raise ValidationError({"file": "invalidZipFileFormat"})

        with zipfile.ZipFile(dump_file, mode="r") as zipf:
            if "data.json" not in zipf.namelist():
                logger.error(
                    "No data.json file found in uploaded file", files=zipf.namelist()
                )
                raise ValidationError({"file": "noDataJsonFileFound"})
            infolist = zipf.infolist()
            directories = list(set([Path(f.filename).parent.name for f in infolist]))
            decompressed_data = zipf.read("data.json")
            # Decode bytes to string if necessary
            if isinstance(decompressed_data, bytes):
                decompressed_data = decompressed_data.decode("utf-8")
            try:
                json_dump = json.loads(decompressed_data)
                import_version = json_dump["meta"]["media_version"]
                schema_version = json_dump["meta"].get("schema_version")
            except json.JSONDecodeError:
                logger.error("Invalid JSON format in uploaded file", exc_info=True)
                raise
            if "objects" not in json_dump:
                raise ValidationError("badly formatted json")

            # Check backup and local version

            try:
                schema_version_int = int(schema_version)
            except (ValueError, TypeError) as e:
                logger.error(
                    "Invalid schema version format",
                    schema_version=schema_version,
                    exc_info=e,
                )
                raise ValidationError({"error": "invalidSchemaVersionFormat"})
            compare_schema_versions(schema_version_int, import_version)

            if "attachments" in directories:
                attachments = {
                    f for f in infolist if Path(f.filename).parent.name == "attachments"
                }
                logger.info(
                    "Attachments found in uploaded file",
                    attachments_count=len(attachments),
                )
                for attachment in attachments:
                    try:
                        content = zipf.read(attachment)
                        current_name = Path(attachment.filename).name
                        new_name = default_storage.save(
                            current_name, io.BytesIO(content)
                        )
                        if new_name != current_name:
                            for x in json_dump["objects"]:
                                if (
                                    x["model"] == "core.evidence"
                                    and x["fields"]["attachment"] == current_name
                                ):
                                    x["fields"]["attachment"] = new_name

                    except Exception:
                        logger.error("Error extracting attachment", exc_info=True)

        return json_dump

    def _get_models_map(self, objects):
        """Build a map of model names to model classes."""
        model_names = {obj["model"] for obj in objects}
        return {name: apps.get_model(name) for name in model_names}

    def _resolve_dependencies(self, all_models):
        """Resolve model dependencies and detect cycles."""
        logger.debug("Resolving model dependencies", all_models=all_models)

        graph = build_dependency_graph(all_models)

        logger.debug("Dependency graph", graph=graph)

        try:
            return topological_sort(graph)
        except ValueError as e:
            logger.error("Cyclic dependency detected", error=str(e))
            raise ValidationError({"error": "Cyclic dependency detected"})

    def _import_objects(
        self, parsed_data: dict, domain_name: str, load_missing_libraries: bool, user
    ):
        """
        Import and validate objects using appropriate serializers.
        Handles both validation and creation in separate phases within a transaction.
        """
        validation_errors = []
        required_libraries = []
        missing_libraries = []
        link_dump_database_ids = {}

        # First check if objects exist
        objects = parsed_data.get("objects")
        if not objects:
            logger.error("No objects found in the dump")
            raise ValidationError({"error": "No objects found in the dump"})

        try:
            # Validate models and check for domain
            models_map = self._get_models_map(objects)
            if Folder in models_map.values():
                logger.error("Dump contains a domain")
                raise ValidationError({"error": "Dump contains a domain"})

            # check that user has permission to create all objects to import
            error_dict = {}
            for model in filter(
                lambda x: x not in [RequirementAssessment], models_map.values()
            ):
                if not RoleAssignment.is_access_allowed(
                    user=user,
                    perm=Permission.objects.get(
                        codename=f"add_{model._meta.model_name}"
                    ),
                    folder=Folder.get_root_folder(),
                ):
                    error_dict[model._meta.model_name] = "permission_denied"
            if error_dict:
                logger.error(
                    "User does not have permission to import objects",
                    error_dict=error_dict,
                )
                raise PermissionDenied()

            # Validation phase (outside transaction since it doesn't modify database)
            creation_order = self._resolve_dependencies(list(models_map.values()))

            logger.debug("Resolved creation order", creation_order=creation_order)
            logger.debug("Starting objects validation", objects_count=len(objects))

            # Validate all objects first
            for model in creation_order:
                self._validate_model_objects(
                    model=model,
                    objects=objects,
                    validation_errors=validation_errors,
                    required_libraries=required_libraries,
                )

            logger.debug("required_libraries", required_libraries=required_libraries)

            # If validation errors exist, raise them immediately
            if validation_errors:
                logger.error(
                    "Failed to validate objects", validation_errors=validation_errors
                )
                raise ValidationError({"validation_errors": validation_errors})

            # Creation phase - wrap in transaction
            with transaction.atomic():
                # Create base folder and store its ID
                base_folder = Folder.objects.create(
                    name=domain_name, content_type=Folder.ContentType.DOMAIN
                )
                link_dump_database_ids["base_folder"] = base_folder
                Folder.create_default_ug_and_ra(base_folder)

                # Check for missing libraries after folder creation
                for library in required_libraries:
                    if not LoadedLibrary.objects.filter(
                        urn=library["urn"], version=library["version"]
                    ).exists():
                        if (
                            StoredLibrary.objects.filter(
                                urn=library["urn"], version__gte=library["version"]
                            ).exists()
                            and load_missing_libraries
                        ):
                            StoredLibrary.objects.get(
                                urn=library["urn"], version__gte=library["version"]
                            ).load()
                        else:
                            missing_libraries.append(library)

                logger.debug("missing_libraries", missing_libraries=missing_libraries)

                # If missing libraries exist, raise specific error
                if missing_libraries:
                    logger.warning(f"Missing libraries: {missing_libraries}")
                    raise ValidationError({"missing_libraries": missing_libraries})

                logger.info(
                    "Starting objects creation",
                    objects_count=len(objects),
                    creation_order=creation_order,
                )

                # Create all objects within the transaction
                for model in creation_order:
                    self._create_model_objects(
                        model=model,
                        objects=objects,
                        link_dump_database_ids=link_dump_database_ids,
                    )

            return {"message": "Import successful"}

        except ValidationError as e:
            logger.error(f"error: {e}")
            raise
        except Exception as e:
            # Handle unexpected errors with a generic message
            logger.exception(f"Failed to import objects: {str(e)}")
            raise ValidationError({"non_field_errors": "errorOccuredDuringImport"})

    def _validate_model_objects(
        self, model, objects, validation_errors, required_libraries
    ):
        """Validate all objects for a model before creation."""
        model_name = f"{model._meta.app_label}.{model._meta.model_name}"
        model_objects = [obj for obj in objects if obj["model"] == model_name]

        if not model_objects:
            return

        # Process validation in batches
        for i in range(0, len(model_objects), self.batch_size):
            batch = model_objects[i : i + self.batch_size]
            self._validate_batch(
                model=model,
                batch=batch,
                validation_errors=validation_errors,
                required_libraries=required_libraries,
            )

    def _validate_batch(self, model, batch, validation_errors, required_libraries):
        """Validate a batch of objects."""
        model_name = f"{model._meta.app_label}.{model._meta.model_name}"

        for obj in batch:
            obj_id = obj.get("id")
            fields = obj.get("fields", {}).copy()

            try:
                # Handle library objects
                if model == LoadedLibrary:
                    required_libraries.append(
                        {"urn": fields["urn"], "version": fields["version"]}
                    )
                    logger.info(
                        "Adding library to required libraries", urn=fields["urn"]
                    )
                    continue
                if fields.get("library"):
                    continue

                # Validate using serializer
                serializer_class = import_export_serializer_class(model)
                serializer = serializer_class(data=fields)

                if not serializer.is_valid():
                    validation_errors.append(
                        {
                            "model": model_name,
                            "id": obj_id,
                            "errors": serializer.errors,
                        }
                    )

            except Exception as e:
                logger.error(
                    f"Error validating object {obj_id} in {model_name}: {str(e)}",
                    exc_info=True,
                )
                validation_errors.append(
                    {
                        "model": model_name,
                        "id": obj_id,
                        "errors": [str(e)],
                    }
                )

    def _create_model_objects(self, model, objects, link_dump_database_ids):
        """Create all objects for a model after validation."""
        logger.debug("Creating objects for model", model=model)

        model_name = f"{model._meta.app_label}.{model._meta.model_name}"
        model_objects = [obj for obj in objects if obj["model"] == model_name]

        logger.debug("Model objects", model=model, count=len(model_objects))

        if not model_objects:
            return

        # Handle self-referencing dependencies
        self_ref_field = get_self_referencing_field(model)
        if self_ref_field:
            try:
                model_objects = sort_objects_by_self_reference(
                    model_objects, self_ref_field
                )
            except ValueError as e:
                logger.error(f"Cyclic dependency detected in {model_name}: {str(e)}")
                raise ValidationError(
                    {"error": f"Cyclic dependency detected in {model_name}"}
                )

        # Process creation in batches
        for i in range(0, len(model_objects), self.batch_size):
            batch = model_objects[i : i + self.batch_size]
            self._create_batch(
                model=model,
                batch=batch,
                link_dump_database_ids=link_dump_database_ids,
            )

    def _create_batch(self, model, batch, link_dump_database_ids):
        """Create a batch of objects with proper relationship handling."""
        # Create all objects in the batch within a single transaction
        with transaction.atomic():
            for obj in batch:
                obj_id = obj.get("id")
                fields = obj.get("fields", {}).copy()

                try:
                    # Handle library objects
                    if fields.get("library") or model == LoadedLibrary:
                        logger.info(f"Skipping creation of library object {obj_id}")
                        link_dump_database_ids[obj_id] = fields.get("urn")
                        continue

                    # Handle folder reference
                    if fields.get("folder"):
                        fields["folder"] = link_dump_database_ids.get("base_folder")

                    # Process model-specific relationships
                    many_to_many_map_ids = {}
                    fields = self._process_model_relationships(
                        model=model,
                        fields=fields,
                        link_dump_database_ids=link_dump_database_ids,
                        many_to_many_map_ids=many_to_many_map_ids,
                    )

                    try:
                        # Run clean to validate unique constraints
                        model(**fields).clean()
                    except ValidationError as e:
                        for field, error in e.error_dict.items():
                            fields[field] = f"{fields[field]} {uuid.uuid4()}"

                    logger.debug("Creating object", fields=fields)

                    # Create the object
                    obj_created = model.objects.create(**fields)
                    link_dump_database_ids[obj_id] = obj_created.id

                    # Handle many-to-many relationships
                    self._set_many_to_many_relations(
                        model=model,
                        obj=obj_created,
                        many_to_many_map_ids=many_to_many_map_ids,
                    )

                except Exception as e:
                    logger.error("Error creating object", obj=obj, exc_info=True)
                    # This will trigger a rollback of the entire batch
                    raise ValidationError(
                        f"Error creating {model._meta.model_name}: {str(e)}"
                    )

    def _process_model_relationships(
        self,
        model,
        fields,
        link_dump_database_ids,
        many_to_many_map_ids,
    ):
        """Process model-specific relationships."""

        def get_mapped_ids(
            ids: List[str], link_dump_database_ids: Dict[str, str]
        ) -> List[str]:
            return [
                link_dump_database_ids[id] for id in ids if id in link_dump_database_ids
            ]

        model_name = model._meta.model_name
        _fields = fields.copy()

        logger.debug(
            "Processing model relationships", model=model_name, _fields=_fields
        )

        match model_name:
            case "asset":
                many_to_many_map_ids["parent_ids"] = get_mapped_ids(
                    _fields.pop("parent_assets", []), link_dump_database_ids
                )

            case "riskassessment":
                _fields["perimeter"] = Perimeter.objects.get(
                    id=link_dump_database_ids.get(_fields["perimeter"])
                )
                _fields["risk_matrix"] = RiskMatrix.objects.get(
                    urn=_fields.get("risk_matrix")
                )
                _fields["ebios_rm_study"] = (
                    EbiosRMStudy.objects.get(
                        id=link_dump_database_ids.get(_fields["ebios_rm_study"])
                    )
                    if _fields.get("ebios_rm_study")
                    else None
                )

            case "complianceassessment":
                _fields["perimeter"] = Perimeter.objects.get(
                    id=link_dump_database_ids.get(_fields["perimeter"])
                )
                _fields["framework"] = Framework.objects.get(urn=_fields["framework"])

            case "appliedcontrol":
                many_to_many_map_ids["evidence_ids"] = get_mapped_ids(
                    _fields.pop("evidences", []), link_dump_database_ids
                )
                many_to_many_map_ids["objective_ids"] = get_mapped_ids(
                    _fields.pop("objectives", []), link_dump_database_ids
                )
                ref_control_id = link_dump_database_ids.get(
                    _fields["reference_control"]
                )
                _fields["reference_control"] = ReferenceControl.objects.filter(
                    urn=ref_control_id
                ).first()

            case "evidence":
                many_to_many_map_ids["owner_ids"] = get_mapped_ids(
                    _fields.pop("owner", []), link_dump_database_ids
                )

            case "evidencerevision":
                _fields.pop("size", None)
                _fields.pop("attachment_hash", None)
                _fields["evidence"] = Evidence.objects.get(
                    id=link_dump_database_ids.get(_fields["evidence"])
                )

            case "requirementassessment":
                logger.debug("Looking for requirement", urn=_fields.get("requirement"))
                _fields["requirement"] = RequirementNode.objects.get(
                    urn=_fields.get("requirement")
                )
                _fields["compliance_assessment"] = ComplianceAssessment.objects.get(
                    id=link_dump_database_ids.get(_fields["compliance_assessment"])
                )
                many_to_many_map_ids.update(
                    {
                        "applied_controls": get_mapped_ids(
                            _fields.pop("applied_controls", []), link_dump_database_ids
                        ),
                        "evidence_ids": get_mapped_ids(
                            _fields.pop("evidences", []), link_dump_database_ids
                        ),
                    }
                )

            case "vulnerability":
                many_to_many_map_ids["applied_controls"] = get_mapped_ids(
                    _fields.pop("applied_controls", []), link_dump_database_ids
                )

            case "riskscenario":
                _fields["risk_assessment"] = RiskAssessment.objects.get(
                    id=link_dump_database_ids.get(_fields["risk_assessment"])
                )
                # Process all related _fields at once
                related__fields = [
                    "threats",
                    "vulnerabilities",
                    "assets",
                    "applied_controls",
                    "existing_applied_controls",
                    "qualifications",
                ]
                for field in related__fields:
                    map_key = (
                        f"{field.rstrip('s')}_ids"
                        if not field.endswith("controls")
                        else f"{field}_ids"
                    )
                    if field == "qualifications":
                        many_to_many_map_ids[map_key] = _fields.pop(field, [])
                    else:
                        many_to_many_map_ids[map_key] = get_mapped_ids(
                            _fields.pop(field, []), link_dump_database_ids
                        )

            case "entity":
                _fields.pop("owned_folders", None)

            case "ebiosrmstudy":
                _fields.update(
                    {
                        "risk_matrix": RiskMatrix.objects.get(
                            urn=_fields.get("risk_matrix")
                        ),
                        "reference_entity": Entity.objects.get(
                            id=link_dump_database_ids.get(_fields["reference_entity"])
                        ),
                    }
                )
                many_to_many_map_ids.update(
                    {
                        "asset_ids": get_mapped_ids(
                            _fields.pop("assets", []), link_dump_database_ids
                        ),
                        "compliance_assessment_ids": get_mapped_ids(
                            _fields.pop("compliance_assessments", []),
                            link_dump_database_ids,
                        ),
                    }
                )

            case "fearedevent":
                _fields["ebios_rm_study"] = EbiosRMStudy.objects.get(
                    id=link_dump_database_ids.get(_fields["ebios_rm_study"])
                )
                many_to_many_map_ids.update(
                    {
                        "qualification_ids": _fields.pop("qualifications", []),
                        "asset_ids": get_mapped_ids(
                            _fields.pop("assets", []), link_dump_database_ids
                        ),
                    }
                )

            case "roto":
                _fields["ebios_rm_study"] = EbiosRMStudy.objects.get(
                    id=link_dump_database_ids.get(_fields["ebios_rm_study"])
                )
                many_to_many_map_ids["feared_event_ids"] = get_mapped_ids(
                    _fields.pop("feared_events", []), link_dump_database_ids
                )
                _fields["risk_origin"], _ = Terminology.objects.get_or_create(
                    name=_fields["risk_origin"],
                    is_visible=True,
                    field_path=Terminology.FieldPath.ROTO_RISK_ORIGIN,
                )

            case "stakeholder":
                _fields.update(
                    {
                        "ebios_rm_study": EbiosRMStudy.objects.get(
                            id=link_dump_database_ids.get(_fields["ebios_rm_study"])
                        ),
                        "entity": Entity.objects.get(
                            id=link_dump_database_ids.get(_fields["entity"])
                        ),
                    }
                )
                many_to_many_map_ids["applied_controls"] = get_mapped_ids(
                    _fields.pop("applied_controls", []), link_dump_database_ids
                )

            case "strategicscenario":
                _fields.update(
                    {
                        "ebios_rm_study": EbiosRMStudy.objects.get(
                            id=link_dump_database_ids.get(_fields["ebios_rm_study"])
                        ),
                        "ro_to_couple": RoTo.objects.get(
                            id=link_dump_database_ids.get(_fields["ro_to_couple"])
                        ),
                    }
                )

            case "attackpath":
                _fields.update(
                    {
                        "ebios_rm_study": EbiosRMStudy.objects.get(
                            id=link_dump_database_ids.get(_fields["ebios_rm_study"])
                        ),
                        "strategic_scenario": StrategicScenario.objects.get(
                            id=link_dump_database_ids.get(_fields["strategic_scenario"])
                        ),
                    }
                )
                many_to_many_map_ids["stakeholder_ids"] = get_mapped_ids(
                    _fields.pop("stakeholders", []), link_dump_database_ids
                )

            case "operationalscenario":
                _fields.update(
                    {
                        "ebios_rm_study": EbiosRMStudy.objects.get(
                            id=link_dump_database_ids.get(_fields["ebios_rm_study"])
                        ),
                        "attack_path": AttackPath.objects.get(
                            id=link_dump_database_ids.get(_fields["attack_path"])
                        ),
                    }
                )
                many_to_many_map_ids["threat_ids"] = get_mapped_ids(
                    _fields.pop("threats", []), link_dump_database_ids
                )

        return _fields

    def _set_many_to_many_relations(self, model, obj, many_to_many_map_ids):
        """Set many-to-many relationships after object creation."""
        model_name = model._meta.model_name

        match model_name:
            case "asset":
                if parent_ids := many_to_many_map_ids.get("parent_ids"):
                    logger.debug(
                        "Setting parent assets", asset=obj, parent_ids=parent_ids
                    )
                    obj.parent_assets.set(Asset.objects.filter(id__in=parent_ids))

            case "appliedcontrol":
                if evidence_ids := many_to_many_map_ids.get("evidence_ids"):
                    obj.evidences.set(Evidence.objects.filter(id__in=evidence_ids))

                if objectives_ids := many_to_many_map_ids.get("objective_ids"):
                    obj.objectives.set(
                        OrganisationObjective.objects.filter(id__in=objectives_ids)
                    )

            case "requirementassessment":
                if applied_control_ids := many_to_many_map_ids.get("applied_controls"):
                    obj.applied_controls.set(
                        AppliedControl.objects.filter(id__in=applied_control_ids)
                    )
                if evidence_ids := many_to_many_map_ids.get("evidence_ids"):
                    obj.evidences.set(Evidence.objects.filter(id__in=evidence_ids))

            case "vulnerability":
                if applied_control_ids := many_to_many_map_ids.get("applied_controls"):
                    obj.applied_controls.set(
                        AppliedControl.objects.filter(id__in=applied_control_ids)
                    )

            case "riskscenario":
                if threat_ids := many_to_many_map_ids.get("threat_ids"):
                    uuids, urns = self._split_uuids_urns(threat_ids)
                    obj.threats.set(
                        Threat.objects.filter(Q(id__in=uuids) | Q(urn__in=urns))
                    )

                if qualification_ids := many_to_many_map_ids.get("qualification_ids"):
                    # Get existing qualifications
                    existing_qualifications = Terminology.objects.filter(
                        name__in=qualification_ids
                    )
                    existing_names = set(
                        existing_qualifications.values_list("name", flat=True)
                    )

                    # Find missing names
                    missing_names = set(qualification_ids) - existing_names

                    # Create missing qualifications
                    if missing_names:
                        Terminology.objects.bulk_create(
                            [
                                Terminology(
                                    name=name,
                                    is_visible=True,
                                    field_path=Terminology.FieldPath.QUALIFICATIONS,
                                )
                                for name in missing_names
                            ],
                            ignore_conflicts=True,
                        )

                    # Now set all qualifications
                    obj.qualifications.set(
                        Terminology.objects.filter(name__in=qualification_ids)
                    )

                for field, model_class in {
                    "vulnerability_ids": (Vulnerability, "vulnerabilities"),
                    "asset_ids": (Asset, "assets"),
                    "applied_control_ids": (AppliedControl, "applied_controls"),
                    "existing_applied_control_ids": (
                        AppliedControl,
                        "existing_applied_controls",
                    ),
                }.items():
                    if ids := many_to_many_map_ids.get(field):
                        getattr(obj, model_class[1]).set(
                            model_class[0].objects.filter(id__in=ids)
                        )

            case "ebiosrmstudy":
                if asset_ids := many_to_many_map_ids.get("asset_ids"):
                    obj.assets.set(Asset.objects.filter(id__in=asset_ids))
                if compliance_assessment_ids := many_to_many_map_ids.get(
                    "compliance_assessment_ids"
                ):
                    obj.compliance_assessments.set(
                        ComplianceAssessment.objects.filter(
                            id__in=compliance_assessment_ids
                        )
                    )

            case "fearedevent":
                if qualification_ids := many_to_many_map_ids.get("qualification_ids"):
                    # Get existing qualifications
                    existing_qualifications = Terminology.objects.filter(
                        name__in=qualification_ids
                    )
                    existing_names = set(
                        existing_qualifications.values_list("name", flat=True)
                    )

                    # Find missing names
                    missing_names = set(qualification_ids) - existing_names

                    # Create missing qualifications
                    if missing_names:
                        Terminology.objects.bulk_create(
                            [
                                Terminology(
                                    name=name,
                                    is_visible=True,
                                    field_path=Terminology.FieldPath.QUALIFICATIONS,
                                )
                                for name in missing_names
                            ],
                            ignore_conflicts=True,
                        )

                    # Now set all qualifications
                    obj.qualifications.set(
                        Terminology.objects.filter(name__in=qualification_ids)
                    )
                if asset_ids := many_to_many_map_ids.get("asset_ids"):
                    obj.assets.set(Asset.objects.filter(id__in=asset_ids))

            case "roto":
                if feared_event_ids := many_to_many_map_ids.get("feared_event_ids"):
                    obj.feared_events.set(
                        FearedEvent.objects.filter(id__in=feared_event_ids)
                    )

            case "stakeholder":
                if applied_control_ids := many_to_many_map_ids.get("applied_controls"):
                    obj.applied_controls.set(
                        AppliedControl.objects.filter(id__in=applied_control_ids)
                    )

            case "attackpath":
                if stakeholder_ids := many_to_many_map_ids.get("stakeholder_ids"):
                    obj.stakeholders.set(
                        Stakeholder.objects.filter(id__in=stakeholder_ids)
                    )

            case "operationalscenario":
                if threat_ids := many_to_many_map_ids.get("threat_ids"):
                    uuids, urns = self._split_uuids_urns(threat_ids)
                    obj.threats.set(
                        Threat.objects.filter(Q(id__in=uuids) | Q(urn__in=urns))
                    )

    def _split_uuids_urns(self, ids: List[str]) -> Tuple[List[str], List[str]]:
        """Split a list of strings into UUIDs and URNs."""
        uuids = []
        urns = []
        for item in ids:
            try:
                uuid = UUID(str(item))
                uuids.append(uuid)
            except ValueError:
                urns.append(item)
        return uuids, urns

    @action(detail=False, methods=["get"])
    def get_accessible_objects(self, request):
        """
        Return the list of folders, perimeters, frameworks and risk matrices
        the current user can access, based on their role assignments.
        """
        (viewable_folders_ids, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, Folder
        )
        (viewable_perimeters_ids, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, Perimeter
        )
        (viewable_frameworks_ids, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, Framework
        )
        (viewable_risk_matrices_ids, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, RiskMatrix
        )
        res = {
            "folders": [
                {"name": str(f), "id": f.id}
                for f in Folder.objects.filter(id__in=viewable_folders_ids).order_by(
                    Lower("name")
                )
            ],
            "perimeters": [
                {"name": str(p), "id": p.id}
                for p in Perimeter.objects.filter(
                    id__in=viewable_perimeters_ids
                ).order_by(Lower("name"))
            ],
            "frameworks": [
                {"name": f.name, "id": f.id}
                for f in Framework.objects.filter(
                    id__in=viewable_frameworks_ids
                ).order_by(Lower("name"))
            ],
            "risk_matrices": [
                {"name": rm.name, "id": rm.id}
                for rm in RiskMatrix.objects.filter(
                    id__in=viewable_risk_matrices_ids
                ).order_by(Lower("name"))
            ],
        }
        return Response(res)


class UserPreferencesView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request) -> Response:
        return Response(request.user.preferences, status=status.HTTP_200_OK)

    def patch(self, request) -> Response:
        new_language = request.data.get("lang")
        if new_language is None or new_language not in (
            lang[0] for lang in settings.LANGUAGES
        ):
            logger.error(
                f"Error in UserPreferencesView: new_language={new_language} available languages={[lang[0] for lang in settings.LANGUAGES]}"
            )
            return Response(
                {"error": "This language doesn't exist."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        request.user.preferences["lang"] = new_language
        request.user.save()
        return Response({}, status=status.HTTP_200_OK)


@cache_page(60 * SHORT_CACHE_TTL)
@vary_on_cookie
@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated])
def get_counters_view(request):
    """
    API endpoint that returns the counters
    """
    return Response({"results": get_counters(request.user)})


@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated])
def get_metrics_view(request):
    """
    API endpoint that returns the counters
    """
    folder_id = request.query_params.get("folder", None)
    return Response({"results": get_metrics(request.user, folder_id)})


@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated])
def get_combined_assessments_status_view(request):
    """
    API endpoint that returns combined assessment counts per status
    for RiskAssessment, ComplianceAssessment, and FindingsAssessment
    """
    return Response({"results": combined_assessments_per_status(request.user)})


@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated])
def get_governance_calendar_data_view(request):
    """
    API endpoint that returns governance activity calendar data
    Aggregates TaskNode due dates, AppliedControl ETAs, and RiskAcceptance expiry dates
    """
    year = request.query_params.get("year", None)
    if year:
        year = int(year)
    return Response({"results": get_governance_calendar_data(request.user, year)})


# TODO: Add all the proper docstrings for the following list of functions


@cache_page(60 * SHORT_CACHE_TTL)
@vary_on_cookie
@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated])
def get_agg_data(request):
    viewable_risk_assessments = RoleAssignment.get_accessible_object_ids(
        Folder.get_root_folder(), request.user, RiskAssessment
    )[0]
    data = risk_status(
        request.user, RiskAssessment.objects.filter(id__in=viewable_risk_assessments)
    )

    return Response({"results": data})


def serialize_nested(data: Any) -> dict:
    if isinstance(data, (list, tuple)):
        return [serialize_nested(i) for i in data]
    elif isinstance(data, dict):
        return {key: serialize_nested(value) for key, value in data.items()}
    elif isinstance(data, set):
        return {serialize_nested(i) for i in data}
    elif isinstance(data, ReturnDict):
        return dict(data)
    elif isinstance(data, models.query.QuerySet):
        return serialize_nested(list(data))
    elif isinstance(data, RiskAssessment):
        return RiskAssessmentReadSerializer(data).data
    elif isinstance(data, RiskScenario):
        return RiskScenarioReadSerializer(data).data
    elif isinstance(data, uuid.UUID):
        return str(data)
    elif isinstance(data, Promise):
        str_attr = {attr for attr in dir(str) if not attr.startswith("_")}
        proxy_attr = {attr for attr in dir(data) if not attr.startswith("_")}
        if len(str_attr - proxy_attr) == 0:
            return str(data)
    return data


@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated])
def get_composer_data(request):
    risk_assessments = request.GET.get("risk_assessment")
    if risk_assessments is None:
        return Response(
            {"error": "This endpoint requires the 'risk_assessment' query parameter"},
            status=400,
        )

    risk_assessments = risk_assessments.split(",")
    if not all(
        re.fullmatch(
            # UUID REGEX
            r"([0-9a-f]{8}\-[0-9a-f]{4}\-[0-9a-f]{4}\-[0-9a-f]{4}\-[0-9a-f]{12})",
            risk_assessment,
        )
        for risk_assessment in risk_assessments
    ):
        return Response({"error": "Invalid UUID list"}, status=400)

    data = compile_risk_assessment_for_composer(request.user, risk_assessments)
    for _data in data["risk_assessment_objects"]:
        quality_check = serialize_nested(_data["risk_assessment"].quality_check())
        _data["risk_assessment"] = RiskAssessmentReadSerializer(
            _data["risk_assessment"]
        ).data
        _data["risk_assessment"]["quality_check"] = quality_check

    data = serialize_nested(data)
    return Response({"result": data})


# Compliance Assessment


class FrameworkFilter(GenericFilterSet):
    folder = df.ModelMultipleChoiceFilter(queryset=Folder.objects.all())

    baseline = df.ModelChoiceFilter(
        queryset=ComplianceAssessment.objects.all(),
        method="filter_framework",
        label="Baseline",
    )

    def filter_framework(self, queryset, name, value):
        if not value:
            return queryset
        source_framework = value.framework
        target_framework_ids = list(
            RequirementMappingSet.objects.filter(
                source_framework=source_framework
            ).values_list("target_framework__id", flat=True)
        )
        target_framework_ids.append(source_framework.id)
        return queryset.filter(id__in=target_framework_ids)

    class Meta:
        model = Framework
        fields = ["provider"]


class FrameworkViewSet(BaseModelViewSet):
    """
    API endpoint that allows frameworks to be viewed or edited.
    """

    model = Framework
    filterset_class = FrameworkFilter
    search_fields = ["name", "description"]

    def get_queryset(self):
        qs = super().get_queryset().prefetch_related("requirement_nodes")

        # Annotate if the framework is dynamic (any question uses implementation groups)
        qs = qs.annotate(
            is_dynamic=Exists(
                RequirementNode.objects.filter(
                    framework=OuterRef("pk"),
                    questions__icontains="select_implementation_groups",
                )
            )
        )

        return qs

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @method_decorator(vary_on_cookie)
    @action(detail=False, methods=["get"])
    def names(self, request):
        uuid_list = request.query_params.getlist("id[]", [])
        queryset = Framework.objects.filter(id__in=uuid_list)

        return Response(
            {
                str(framework.id): framework.get_name_translated()
                for framework in queryset
            }
        )

    @action(detail=True, methods=["get"])
    def tree(self, request, pk):
        _framework = Framework.objects.get(id=pk)
        return Response(
            get_sorted_requirement_nodes(
                RequirementNode.objects.filter(framework=_framework).all(),
                None,
                _framework.max_score,
            )
        )

    @action(detail=False, name="Get used frameworks")
    def used(self, request):
        viewable_framework = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, Framework
        )[0]
        viewable_assessments = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, ComplianceAssessment
        )[0]
        _used_frameworks = (
            Framework.objects.filter(complianceassessment__isnull=False)
            .filter(id__in=viewable_framework)
            .filter(complianceassessment__id__in=viewable_assessments)
            .distinct()
        )
        used_frameworks = _used_frameworks.values("id", "name")
        for i in range(len(used_frameworks)):
            used_frameworks[i]["compliance_assessments_count"] = (
                ComplianceAssessment.objects.filter(framework=_used_frameworks[i].id)
                .filter(id__in=viewable_assessments)
                .count()
            )
        return Response({"results": used_frameworks})

    @action(
        detail=True, methods=["get"], name="Get framework coverage data from mappings"
    )
    def mapping_stats(self, request, pk):
        from core.mappings.engine import engine

        framework_urn = Framework.objects.filter(id=pk).values_list("urn")[0][0]
        res = engine.paths_and_coverages(framework_urn)
        return Response({"response": res})

    @action(detail=True, methods=["get"], name="Get target frameworks from mappings")
    def mappings(self, request, pk):
        framework = self.get_object()
        available_target_frameworks_objects = [framework]
        mappings = RequirementMappingSet.objects.filter(source_framework=framework)
        for mapping in mappings:
            available_target_frameworks_objects.append(mapping.target_framework)
        available_target_frameworks = FrameworkReadSerializer(
            available_target_frameworks_objects, many=True
        ).data
        return Response({"results": available_target_frameworks})

    @action(detail=False, name="Get provider choices")
    def provider(self, request):
        providers = set(
            Framework.objects.filter(provider__isnull=False).values_list(
                "provider", flat=True
            )
        )
        return Response({p: p for p in providers})

    @action(detail=True, methods=["get"], name="Framework as an Excel template")
    def excel_template(self, request, pk):
        fwk = Framework.objects.get(id=pk)
        req_nodes = RequirementNode.objects.filter(framework=fwk).order_by("urn")
        entries = []
        for rn in req_nodes:
            entry = {
                "urn": rn.urn,
                "assessable": rn.assessable,
                "ref_id": rn.ref_id,
                "name": rn.get_name_translated,
                "description": rn.get_description_translated,
                "compliance_result": "",
                "requirement_progress": "",
                "score": "",
                "observations": "",
            }
            entries.append(entry)

        # Create DataFrame from entries
        df = pd.DataFrame(entries)

        # Create BytesIO object
        buffer = io.BytesIO()

        # Create ExcelWriter with openpyxl engine
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df.to_excel(writer, index=False)

            # Get the worksheet
            worksheet = writer.sheets["Sheet1"]

            # For text wrapping, we need to define which columns need wrapping
            # Assuming 'description' and 'observations' columns need text wrapping
            wrap_columns = ["name", "description", "observations"]

            # Find the indices of the columns that need wrapping
            wrap_indices = [
                df.columns.get_loc(col) + 1 for col in wrap_columns if col in df.columns
            ]

            # Apply text wrapping to those columns
            from openpyxl.styles import Alignment

            for col_idx in wrap_indices:
                for row_idx in range(
                    2, len(df) + 2
                ):  # +2 because of header row and 1-indexing
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.alignment = Alignment(wrap_text=True)

            # Adjust column widths for better readability
            for idx, col in enumerate(df.columns):
                column_width = 40  # default width
                if col in wrap_columns:
                    column_width = 60  # wider for wrapped text columns
                worksheet.column_dimensions[
                    worksheet.cell(row=1, column=idx + 1).column_letter
                ].width = column_width

        # Get the value of the buffer and return as response
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{fwk.name}_template.xlsx"'
        )

        return response


class RequirementNodeViewSet(BaseModelViewSet):
    """
    API endpoint that allows requirement groups to be viewed or edited.
    """

    model = RequirementNode
    filterset_fields = ["framework", "urn"]
    search_fields = ["name", "description"]

    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)


class RequirementViewSet(BaseModelViewSet):
    """
    API endpoint that allows requirements to be viewed or edited.
    """

    model = RequirementNode
    filterset_fields = ["framework", "urn"]
    search_fields = ["name"]

    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    @action(detail=True, methods=["get"], name="Inspect specific requirements")
    def inspect_requirement(self, request, pk):
        requirement = RequirementNode.objects.get(id=pk)
        (viewable_objects, _, _) = RoleAssignment.get_accessible_object_ids(
            folder=Folder.get_root_folder(),
            user=request.user,
            object_type=RequirementAssessment,
        )
        requirement_assessments = RequirementAssessment.objects.filter(
            requirement=requirement, id__in=viewable_objects
        ).prefetch_related("folder", "compliance_assessment__perimeter")
        serialized_requirement_assessments = RequirementAssessmentReadSerializer(
            requirement_assessments, many=True
        ).data

        # Group by Domain and Perimeter
        grouped_data = requirement_assessments.values(
            "folder__name", "compliance_assessment__perimeter__name"
        ).annotate(
            compliant_count=Count(
                "id", filter=Q(result=RequirementAssessment.Result.COMPLIANT)
            ),
            total_count=Count("id"),
            assessed_count=Count(
                "id", filter=~Q(status=RequirementAssessment.Status.TODO)
            ),
            assessment_completion_rate=ExpressionWrapper(
                Count("id", filter=~Q(status=RequirementAssessment.Status.TODO))
                * 100.0
                / Count("id"),
                output_field=FloatField(),
            ),
        )

        # Collect all data by domain for calculations
        domain_data_collector = defaultdict(
            lambda: {
                "compliant_count": 0,
                "total_count": 0,
                "assessed_count": 0,
                "perimeter_data": [],
            }
        )

        for item in grouped_data:
            domain_name = item["folder__name"]
            domain_data_collector[domain_name]["compliant_count"] += item[
                "compliant_count"
            ]
            domain_data_collector[domain_name]["total_count"] += item["total_count"]
            domain_data_collector[domain_name]["assessed_count"] += item[
                "assessed_count"
            ]
            domain_data_collector[domain_name]["perimeter_data"].append(item)

        # Organize data by domain and perimeter using the collected data
        domain_dict = defaultdict(
            lambda: {
                "name": "",
                "compliance_result": {
                    "compliant_count": 0,
                    "total_count": 0,
                    "compliance_percentage": 0,
                },
                "assessment_progress": {
                    "assessed_count": 0,
                    "total_count": 0,
                    "assessment_completion_rate": 0,
                },
                "perimeters": [],
            }
        )

        # Structure the final response
        for domain_name, collector_data in domain_data_collector.items():
            domain_dict[domain_name]["name"] = domain_name

            # Calculate domain-level metrics
            domain_total_count = collector_data["total_count"]
            domain_compliant_count = collector_data["compliant_count"]
            domain_assessed_count = collector_data["assessed_count"]

            # Avoid division by zero
            domain_compliance_percentage = 0
            if domain_total_count > 0:
                domain_compliance_percentage = int(
                    (domain_compliant_count / domain_total_count) * 100
                )

            domain_assessment_completion_rate = 0
            if domain_total_count > 0:
                domain_assessment_completion_rate = int(
                    (domain_assessed_count / domain_total_count) * 100
                )

            # Set domain metrics
            domain_dict[domain_name]["compliance_result"] = {
                "compliant_count": domain_compliant_count,
                "total_count": domain_total_count,
                "compliance_percentage": domain_compliance_percentage,
            }

            domain_dict[domain_name]["assessment_progress"] = {
                "assessed_count": domain_assessed_count,
                "total_count": domain_total_count,
                "assessment_completion_rate": domain_assessment_completion_rate,
            }

            # Process perimeter data
            for item in collector_data["perimeter_data"]:
                perimeter_name = item["compliance_assessment__perimeter__name"]

                perimeter_entry = {
                    "name": perimeter_name,
                    "compliance_assessments": [],
                }

                # Add compliance assessments to the perimeter entry
                compliance_assessments = (
                    requirement_assessments.filter(
                        folder__name=domain_name,
                        compliance_assessment__perimeter__name=perimeter_name,
                    )
                    .select_related("compliance_assessment")
                    .values(
                        "compliance_assessment__id",
                        "compliance_assessment__name",
                        "compliance_assessment__version",
                        "compliance_assessment__show_documentation_score",
                        "compliance_assessment__max_score",
                    )
                    .distinct()
                )

                for ca in compliance_assessments:
                    perimeter_entry["compliance_assessments"].append(
                        {
                            "id": ca["compliance_assessment__id"],
                            "name": ca["compliance_assessment__name"],
                            "version": ca["compliance_assessment__version"],
                            "show_documentation_score": ca[
                                "compliance_assessment__show_documentation_score"
                            ],
                            "max_score": ca["compliance_assessment__max_score"],
                        }
                    )

                domain_dict[domain_name]["perimeters"].append(perimeter_entry)

        # Convert defaultdict to list for the response
        data_by_domain = [
            domain_data for domain_data in domain_dict.values() if domain_data["name"]
        ]

        return Response(
            {
                "requirement_assessments": serialized_requirement_assessments,
                "metrics": data_by_domain,
            }
        )


class EvidenceViewSet(BaseModelViewSet):
    """
    API endpoint that allows evidences to be viewed or edited.
    """

    model = Evidence
    filterset_fields = [
        "folder",
        "applied_controls",
        "requirement_assessments",
        "name",
        "timeline_entries",
        "filtering_labels",
        "findings",
        "findings_assessments",
        "genericcollection",
        "owner",
        "status",
        "expiry_date",
        "contracts",
        "processings",
    ]

    @action(detail=False, name="Get all evidences owners")
    def owner(self, request):
        return Response(
            UserReadSerializer(
                User.objects.filter(evidences__isnull=False).distinct(),
                many=True,
            ).data
        )

    @action(methods=["get"], detail=True)
    def attachment(self, request, pk):
        (
            object_ids_view,
            _,
            _,
        ) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, Evidence
        )
        response = Response(status=status.HTTP_403_FORBIDDEN)
        if UUID(pk) in object_ids_view:
            evidence = self.get_object()
            if (
                not evidence.last_revision.attachment
                or not evidence.last_revision.attachment.storage.exists(
                    evidence.last_revision.attachment.name
                )
            ):
                return Response(status=status.HTTP_404_NOT_FOUND)
            if request.method == "GET":
                content_type = mimetypes.guess_type(evidence.last_revision.filename())[
                    0
                ]
                response = HttpResponse(
                    evidence.last_revision.attachment,
                    content_type=content_type,
                    headers={
                        "Content-Disposition": f"attachment; filename={evidence.last_revision.filename()}"
                    },
                    status=status.HTTP_200_OK,
                )
        return response

    @action(detail=False, name="Get status choices")
    def status(self, request):
        return Response(dict(Evidence.Status.choices))


class EvidenceRevisionViewSet(BaseModelViewSet):
    """
    API endpoint that allows evidence revisions to be viewed or edited.
    """

    model = EvidenceRevision
    filterset_fields = ["evidence"]
    ordering = ["-version"]

    @action(methods=["get"], detail=True)
    def attachment(self, request, pk):
        (
            object_ids_view,
            _,
            _,
        ) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, EvidenceRevision
        )
        response = Response(status=status.HTTP_403_FORBIDDEN)
        if UUID(pk) in object_ids_view:
            evidence = self.get_object()
            if not evidence.attachment or not evidence.attachment.storage.exists(
                evidence.attachment.name
            ):
                return Response(status=status.HTTP_404_NOT_FOUND)
            if request.method == "GET":
                content_type = mimetypes.guess_type(evidence.filename())[0]
                response = HttpResponse(
                    evidence.attachment,
                    content_type=content_type,
                    headers={
                        "Content-Disposition": f"attachment; filename={evidence.filename()}"
                    },
                    status=status.HTTP_200_OK,
                )
        return response

    @action(methods=["post"], detail=True)
    def delete_attachment(self, request, pk):
        (
            _,
            _,
            object_ids_delete,
        ) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, EvidenceRevision
        )
        response = Response(status=status.HTTP_403_FORBIDDEN)
        if UUID(pk) in object_ids_delete:
            evidence = self.get_object()
            if evidence.attachment:
                evidence.attachment.delete()
                response = Response(status=status.HTTP_200_OK)
        return response


class UploadAttachmentView(APIView):
    parser_classes = (FileUploadParser,)
    serializer_class = AttachmentUploadSerializer

    def post(self, request, *args, **kwargs):
        pk = kwargs["pk"]
        revision = None
        evidence = None

        try:
            revision = EvidenceRevision.objects.get(pk=pk)
            evidence = revision.evidence
        except EvidenceRevision.DoesNotExist:
            try:
                evidence = Evidence.objects.get(pk=pk)
            except Evidence.DoesNotExist:
                return Response(
                    {"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND
                )

        if revision is None:
            revision = evidence.revisions.order_by(
                "-version"
            ).first() or EvidenceRevision.objects.create(evidence=evidence)

        attachment = request.FILES.get("file")
        if attachment and attachment.name != "undefined":
            if not revision.attachment or revision.attachment != attachment:
                if revision.attachment:
                    revision.attachment.delete()
                revision.attachment = attachment
                revision.save()

        return Response(status=status.HTTP_200_OK)


class QuickStartView(APIView):
    serializer_class = QuickStartSerializer

    def post(self, request):
        serializer = self.serializer_class(
            data=request.data, context={"request": request}
        )
        if not serializer.is_valid():
            raise ValidationError(serializer.errors)
        try:
            objects = serializer.save()
        except Exception as e:
            logger.error(f"Error in QuickStartView: {e}")
            raise
        else:
            return Response(objects, status=status.HTTP_201_CREATED)


class OrganisationObjectiveViewSet(BaseModelViewSet):
    model = OrganisationObjective

    filterset_fields = ["folder", "status", "health", "issues", "assigned_to"]
    search_fields = ["name", "description"]

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get status choices")
    def status(self, request):
        return Response(dict(OrganisationObjective.Status.choices))

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get health choices")
    def health(self, request):
        return Response(dict(OrganisationObjective.Health.choices))


class OrganisationIssueViewSet(BaseModelViewSet):
    model = OrganisationIssue

    filterset_fields = ["folder", "category", "origin"]
    search_fields = ["name", "description"]

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get category choices")
    def category(self, request):
        return Response(dict(OrganisationIssue.Category.choices))

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get origin choices")
    def origin(self, request):
        return Response(dict(OrganisationIssue.Origin.choices))


class CampaignViewSet(BaseModelViewSet):
    model = Campaign

    filterset_fields = ["folder", "frameworks", "perimeters", "status"]
    search_fields = ["name", "description"]

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get status choices")
    def status(self, request):
        return Response(dict(Campaign.Status.choices))

    @action(detail=True, name="Get campaign metrics")
    def metrics(self, request, pk):
        (viewable_objects, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, Campaign
        )
        if UUID(pk) not in viewable_objects:
            return Response(
                {"error": "Permission denied"}, status=status.HTTP_403_FORBIDDEN
            )
        campaign = self.get_object()
        return Response(campaign.metrics())

    def perform_create(self, serializer):
        super().perform_create(serializer)
        campaign = serializer.instance
        frameworks = serializer.instance.frameworks.all()
        for perimeter in campaign.perimeters.all():
            for framework in frameworks:
                framework_implementation_groups = None
                if campaign.selected_implementation_groups:
                    framework_implementation_groups = [
                        group["value"]
                        for group in campaign.selected_implementation_groups
                        if group["framework"] == str(framework.id)
                    ]
                compliance_assessment = ComplianceAssessment.objects.create(
                    name=f"{campaign.name} - {perimeter.name} - {framework.name}",
                    campaign=campaign,
                    perimeter=perimeter,
                    framework=framework,
                    folder=perimeter.folder,
                    selected_implementation_groups=framework_implementation_groups
                    if framework_implementation_groups
                    else None,
                )
                compliance_assessment.create_requirement_assessments()


class ComplianceAssessmentViewSet(BaseModelViewSet):
    """
    API endpoint that allows compliance assessments to be viewed or edited.
    """

    model = ComplianceAssessment
    filterset_fields = [
        "name",
        "ref_id",
        "folder",
        "framework",
        "perimeter",
        "campaign",
        "status",
        "ebios_rm_studies",
        "assets",
        "evidences",
        "authors",
        "reviewers",
        "genericcollection",
    ]
    search_fields = ["name", "description", "ref_id", "framework__name"]

    def get_queryset(self):
        """Optimize queries for table view and serializer, with conditional annotations for sorting"""
        qs = (
            super()
            .get_queryset()
            .select_related(
                "folder",
                "folder__parent_folder",  # For get_folder_full_path() optimization
                "framework",  # Displayed in table
                "perimeter",  # Displayed in table
                "perimeter__folder",  # FieldsRelatedField(["id", "folder"]) optimization
                "campaign",  # Serialized as FieldsRelatedField
            )
            .prefetch_related(
                "assets",  # ManyToManyField serialized as FieldsRelatedField
                "evidences",  # ManyToManyField serialized as FieldsRelatedField
                "authors",  # ManyToManyField from Assessment parent class
                "reviewers",  # ManyToManyField from Assessment parent class
            )
        )

        qs = qs.annotate(
            total_requirements=Count(
                "requirement_assessments",
                filter=Q(requirement_assessments__requirement__assessable=True),
                distinct=True,
            ),
            assessed_requirements=Count(
                "requirement_assessments",
                filter=Q(
                    ~Q(
                        requirement_assessments__result=RequirementAssessment.Result.NOT_ASSESSED
                    ),
                    requirement_assessments__requirement__assessable=True,
                ),
                distinct=True,
            ),
            progress=ExpressionWrapper(
                F("assessed_requirements")
                * 100
                / Greatest(Coalesce(F("total_requirements"), Value(0)), Value(1)),
                output_field=IntegerField(),
            ),
        )

        return qs

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get status choices")
    def status(self, request):
        return Response(dict(ComplianceAssessment.Status.choices))

    @action(
        detail=True,
        name="Get target frameworks mapping options with compliance distribution",
    )
    def frameworks(self, request, pk):
        audit = ComplianceAssessment.objects.get(id=pk)
        from core.mappings.engine import engine

        audit_from_results = engine.load_audit_fields(audit)
        max_depth = get_mapping_max_depth()
        data = []
        for dest_urn in sorted(
            [
                p[-1]
                for p in engine.all_paths_from(
                    source_urn=audit.framework.urn, max_depth=max_depth
                )
            ]
        ):
            best_results, _ = engine.best_mapping_inferences(
                audit_from_results,
                audit.framework.urn,
                dest_urn,
                max_depth=max_depth,
            )
            if best_results:
                framework = Framework.objects.filter(urn=dest_urn).first()
                if framework and str(framework) not in str(data):
                    assessable_urns = set(
                        framework.requirement_nodes.filter(assessable=True).values_list(
                            "urn", flat=True
                        )
                    )
                    data.append(
                        {
                            "id": framework.id,
                            "str": str(framework),
                            "results": engine.summary_results(
                                best_results, filter_urns=assessable_urns
                            ),
                            "assessable_requirements_count": len(assessable_urns),
                        }
                    )
        return Response(data, status=status.HTTP_200_OK)

    @action(detail=True, name="Get compliance assessment (audit) CSV")
    def compliance_assessment_csv(self, request, pk):
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="audit_export.csv"'

        (viewable_objects, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, ComplianceAssessment
        )

        if UUID(pk) in viewable_objects:
            writer = csv.writer(response, delimiter=";")
            columns = [
                "urn",
                "ref_id",
                "name",
                "description",
                "compliance_result",
                "requirement_progress",
                "score",
                "observations",
            ]
            writer.writerow(columns)

            compliance_assessment = ComplianceAssessment.objects.get(id=pk)
            reqs = list(
                compliance_assessment.get_requirement_assessments(
                    include_non_assessable=True
                )
            )
            req_nodes = RequirementNode.objects.in_bulk(
                [ra.requirement_id for ra in reqs]
            )
            for req in reqs:
                req_node = req_nodes.get(req.requirement_id)
                row = [
                    req_node.urn,
                    req_node.ref_id,
                    req_node.get_name_translated,
                    req_node.get_description_translated,
                ]
                if req_node.assessable:
                    row += [
                        req.result,
                        req.status,
                        req.score,
                        req.observation,
                    ]
                else:
                    row += ["", "", "", ""]
                writer.writerow(row)

            return response
        else:
            return Response(
                {"error": "Permission denied"}, status=status.HTTP_403_FORBIDDEN
            )

    @action(detail=True, methods=["get"], name="Audit as an Excel")
    def xlsx(self, request, pk):
        (viewable_objects, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, ComplianceAssessment
        )
        if UUID(pk) not in viewable_objects:
            return Response(
                {"error": "Permission denied"}, status=status.HTTP_403_FORBIDDEN
            )
        audit = ComplianceAssessment.objects.get(id=pk)
        entries = []
        show_documentation_score = audit.show_documentation_score
        requirement_assessments = audit.get_requirement_assessments(
            include_non_assessable=True
        )
        req_node_ids = [req.requirement.id for req in requirement_assessments]
        req_nodes = {
            node.id: node
            for node in RequirementNode.objects.filter(id__in=req_node_ids)
        }

        for req in requirement_assessments:
            req_node = req_nodes.get(req.requirement.id)
            if not req_node:
                continue

            entry = {
                "urn": escape_excel_formula(req_node.urn),
                "assessable": req_node.assessable,
                "ref_id": escape_excel_formula(req_node.ref_id),
                "name": escape_excel_formula(req_node.get_name_translated),
                "description": escape_excel_formula(
                    req_node.get_description_translated
                ),
                "compliance_result": req.result,
                "requirement_progress": req.status,
                "observations": escape_excel_formula(req.observation),
            }
            if show_documentation_score:
                entry["implementation_score"] = req.score
                entry["documentation_score"] = req.documentation_score
            else:
                entry["score"] = req.score
            entries.append(entry)

        df = pd.DataFrame(entries)
        buffer = io.BytesIO()

        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df.to_excel(writer, index=False)
            worksheet = writer.sheets["Sheet1"]

            wrap_columns = ["name", "description", "observations"]
            wrap_indices = [
                df.columns.get_loc(col) + 1 for col in wrap_columns if col in df.columns
            ]

            for col_idx in wrap_indices:
                for row_idx in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.alignment = Alignment(wrap_text=True)

            for idx, col in enumerate(df.columns):
                column_width = 40
                if col in wrap_columns:
                    column_width = 60  # wider for compliance assessments
                worksheet.column_dimensions[
                    worksheet.cell(row=1, column=idx + 1).column_letter
                ].width = column_width

        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{audit.name}.xlsx"'

        return response

    @action(detail=True, methods=["get"])
    def word_report(self, request, pk):
        """
        Word report generation (Exec)
        """
        lang = "en"
        if request.user.preferences.get("lang") is not None:
            lang = request.user.preferences.get("lang")
            if lang not in ["fr", "en"]:
                lang = "en"
        template_path = (
            Path(settings.BASE_DIR)
            / "core"
            / "templates"
            / "core"
            / f"audit_report_template_{lang}.docx"
        )
        doc = DocxTemplate(template_path)
        _framework = self.get_object().framework
        tree = get_sorted_requirement_nodes(
            RequirementNode.objects.filter(framework=_framework).all(),
            RequirementAssessment.objects.filter(
                compliance_assessment=self.get_object()
            ).all(),
            _framework.max_score,
        )
        implementation_groups = self.get_object().selected_implementation_groups
        filter_graph_by_implementation_groups(tree, implementation_groups)
        context = gen_audit_context(pk, doc, tree, lang)
        doc.render(context)
        buffer_doc = io.BytesIO()
        doc.save(buffer_doc)
        buffer_doc.seek(0)

        response = StreamingHttpResponse(
            FileWrapper(buffer_doc),
            content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        )
        response["Content-Disposition"] = "attachment; filename=exec_report.docx"

        return response

    @action(detail=True, name="Get action plan CSV")
    def action_plan_csv(self, request, pk):
        (object_ids_view, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, ComplianceAssessment
        )
        if UUID(pk) not in object_ids_view:
            return Response(
                {"error": "Permission denied"}, status=status.HTTP_403_FORBIDDEN
            )
        compliance_assessment = ComplianceAssessment.objects.get(id=pk)
        requirement_assessments = compliance_assessment.get_requirement_assessments(
            include_non_assessable=False
        )
        queryset = AppliedControl.objects.filter(
            requirement_assessments__in=requirement_assessments
        ).distinct()

        # Use the same serializer to maintain consistency - to review
        serializer = ComplianceAssessmentActionPlanSerializer(
            queryset, many=True, context={"pk": pk}
        )

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="action_plan_{pk}.csv"'

        writer = csv.writer(response)

        writer.writerow(
            [
                "Name",
                "Description",
                "Category",
                "CSF Function",
                "Priority",
                "Status",
                "ETA",
                "Expiry date",
                "Effort",
                "Impact",
                "Cost",
                "Covered requirements",
            ]
        )

        for item in serializer.data:
            writer.writerow(
                [
                    item.get("name"),
                    item.get("description"),
                    item.get("category"),
                    item.get("csf_function"),
                    item.get("priority"),
                    item.get("status"),
                    item.get("eta"),
                    item.get("expiry_date"),
                    item.get("effort"),
                    item.get("impact"),
                    item.get("annual_cost"),
                    "\n".join(
                        [ra.get("str") for ra in item.get("requirement_assessments")]
                    ),
                ]
            )

        return response

    @action(detail=True, name="Get action plan XLSX")
    def action_plan_xlsx(self, request, pk):
        (object_ids_view, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, ComplianceAssessment
        )
        if UUID(pk) not in object_ids_view:
            return Response(
                {"error": "Permission denied"}, status=status.HTTP_403_FORBIDDEN
            )
        compliance_assessment = ComplianceAssessment.objects.get(id=pk)
        requirement_assessments = compliance_assessment.get_requirement_assessments(
            include_non_assessable=False
        )
        queryset = AppliedControl.objects.filter(
            requirement_assessments__in=requirement_assessments
        ).distinct()

        serializer = ComplianceAssessmentActionPlanSerializer(
            queryset, many=True, context={"pk": pk}
        )

        entries = []
        for item in serializer.data:
            entry = {
                "name": item.get("name"),
                "description": item.get("description"),
                "category": item.get("category"),
                "csf_function": item.get("csf_function"),
                "priority": item.get("priority"),
                "status": item.get("status"),
                "eta": item.get("eta"),
                "expiry_date": item.get("expiry_date"),
                "effort": item.get("effort"),
                "impact": item.get("impact"),
                "cost": item.get("annual_cost"),
                "covered_requirements": "\n".join(
                    [ra.get("str") for ra in item.get("requirement_assessments")]
                ),
            }
            entries.append(entry)

        df = pd.DataFrame(entries)
        buffer = io.BytesIO()

        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df.to_excel(writer, index=False)
            worksheet = writer.sheets["Sheet1"]

            wrap_columns = ["name", "description", "covered_requirements"]
            wrap_indices = [
                df.columns.get_loc(col) + 1 for col in wrap_columns if col in df.columns
            ]

            for col_idx in wrap_indices:
                for row_idx in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.alignment = Alignment(wrap_text=True)

            for idx, col in enumerate(df.columns):
                column_width = 20
                if col in wrap_columns:
                    column_width = 40
                worksheet.column_dimensions[
                    worksheet.cell(row=1, column=idx + 1).column_letter
                ].width = column_width

        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="action_plan_{pk}.xlsx"'
        )
        return response

    @action(detail=True, name="Get action plan PDF")
    def action_plan_pdf(self, request, pk):
        (object_ids_view, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, ComplianceAssessment
        )
        if UUID(pk) in object_ids_view:
            context = {
                "to_do": list(),
                "in_progress": list(),
                "on_hold": list(),
                "active": list(),
                "deprecated": list(),
                "--": list(),
            }
            color_map = {
                "to_do": "#FFF8F0",
                "in_progress": "#392F5A",
                "on_hold": "#F4D06F",
                "active": "#9DD9D2",
                "deprecated": "#ff8811",
                "--": "#e5e7eb",
            }
            status = AppliedControl.Status.choices
            compliance_assessment_object: ComplianceAssessment = self.get_object()
            requirement_assessments_objects = (
                compliance_assessment_object.get_requirement_assessments(
                    include_non_assessable=True
                )
            )
            applied_controls = (
                AppliedControl.objects.filter(
                    requirement_assessments__in=requirement_assessments_objects
                )
                .distinct()
                .order_by("eta")
            )
            for applied_control in applied_controls:
                context[applied_control.status].append(
                    applied_control
                ) if applied_control.status else context["--"].append(applied_control)
            data = {
                "status_text": status,
                "color_map": color_map,
                "context": context,
                "compliance_assessment": compliance_assessment_object,
            }
            html = render_to_string("core/action_plan_pdf.html", data)
            pdf_file = HTML(string=html).write_pdf()
            response = HttpResponse(pdf_file, content_type="application/pdf")
            return response
        else:
            return Response({"error": "Permission denied"})

    @action(
        detail=True,
        methods=["post"],
        name="Send compliance assessment by mail to authors",
    )
    def mailing(self, request, pk):
        instance = self.get_object()
        if EMAIL_HOST or EMAIL_HOST_RESCUE:
            for author in instance.authors.all():
                try:
                    author.mailing(
                        email_template_name="tprm/third_party_email.html",
                        subject=_(
                            "CISO Assistant: A questionnaire has been assigned to you"
                        ),
                        object="compliance-assessments",
                        object_id=instance.id,
                    )
                except Exception as primary_exception:
                    logger.error(
                        f"Failed to send email to {author}: {primary_exception}"
                    )
                    raise ValidationError(
                        {"error": ["An error occurred while sending the email"]}
                    )
            return Response({"results": "mail sent"})
        raise ValidationError({"warning": ["noMailerConfigured"]})

    @action(detail=True, methods=["post"])
    def update_requirement(self, request, pk):
        compliance_assessment = get_object_or_404(self.get_queryset(), pk=pk)

        viewable_objects, _, _ = RoleAssignment.get_accessible_object_ids(
            folder=Folder.get_root_folder(),
            user=request.user,
            object_type=ComplianceAssessment,
        )
        if compliance_assessment.id not in viewable_objects:
            return Response(status=status.HTTP_403_FORBIDDEN)
        try:
            ref_id = request.data.get("ref_id")
            result = request.data.get("result")
            observation = request.data.get("observation")
            score = request.data.get("score")
            status_value = request.data.get("status")

            if not all([ref_id, result]):
                return Response(
                    {"error": "ref_id and result are required fields"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            # validate if result value is valid choice
            valid_results = [
                choice[0] for choice in RequirementAssessment.Result.choices
            ]
            if result not in valid_results:
                return Response(
                    {"error": f"invalid result value. Must be one of: {valid_results}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # validate if status value is valid choice
            if status_value is not None:
                valid_statuses = [
                    choice[0] for choice in RequirementAssessment.Status.choices
                ]
                if status_value not in valid_statuses:
                    return Response(
                        {
                            "error": f"invalid status value. Must be one of: {valid_statuses}"
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )

            # validate if score value is within allowed range
            if score is not None:
                try:
                    score = int(score)
                    # Only validate range if min_score and max_score are defined
                    if (
                        compliance_assessment.min_score is not None
                        and compliance_assessment.max_score is not None
                    ):
                        if (
                            score < compliance_assessment.min_score
                            or score > compliance_assessment.max_score
                        ):
                            return Response(
                                {
                                    "error": f"Score must be between {compliance_assessment.min_score} and {compliance_assessment.max_score}"
                                },
                                status=status.HTTP_400_BAD_REQUEST,
                            )
                except (ValueError, TypeError):
                    return Response(
                        {"error": "Score must be a valid integer"},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

            # Find the requirement assessment to update
            requirement_assessment = RequirementAssessment.objects.filter(
                compliance_assessment=compliance_assessment, requirement__ref_id=ref_id
            ).first()

            if not requirement_assessment:
                return Response(
                    {"error": f"Requirement with ref_id {ref_id} not found"},
                    status=status.HTTP_404_NOT_FOUND,
                )

            # Update the requirement assessment
            requirement_assessment.result = result
            requirement_assessment.observation = observation

            # Update status if provided
            if status_value is not None:
                requirement_assessment.status = status_value

            # Update score and toggle is_scored accordingly
            if score is not None:
                requirement_assessment.score = score
                requirement_assessment.is_scored = True
            elif score is None and "score" in request.data:
                # Explicitly setting score to null/empty
                requirement_assessment.score = None
                requirement_assessment.is_scored = False

            requirement_assessment.save()

            response_data = {
                "message": "Requirement updated successfully",
                "ref_id": ref_id,
                "result": result,
            }

            # Include status in response if it was updated
            if status_value is not None:
                response_data["status"] = status_value

            # Include score and is_scored in response if score was updated
            if score is not None:
                response_data["score"] = score
                response_data["is_scored"] = True
            elif score is None and "score" in request.data:
                response_data["score"] = None
                response_data["is_scored"] = False

            return Response(response_data, status=status.HTTP_200_OK)

        except RequirementAssessment.DoesNotExist:
            return Response(
                {"error": f"Requirement with ref_id {ref_id} not found"},
                status=status.HTTP_404_NOT_FOUND,
            )
        except ValidationError:
            return Response(
                {"error": "invalid input provided"}, status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            logger.error(f"Unexpected error in update_requirement: {str(e)}")
            return Response(
                {"error": "An unexpected error occurred"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    def perform_create(self, serializer):
        """
        Create RequirementAssessment objects for the newly created ComplianceAssessment
        with optimized database operations for different scenarios.
        """
        baseline = serializer.validated_data.pop("baseline", None)
        create_applied_controls = serializer.validated_data.pop(
            "create_applied_controls_from_suggestions", False
        )
        from core.mappings.engine import engine

        with transaction.atomic():
            instance: ComplianceAssessment = serializer.save()
            instance.create_requirement_assessments(baseline)

            if baseline and baseline.framework != instance.framework:
                source_urn = baseline.framework.urn
                audit_from_results = engine.load_audit_fields(baseline)
                dest_urn = serializer.validated_data["framework"].urn
                max_depth = get_mapping_max_depth()

                best_results, _ = engine.best_mapping_inferences(
                    audit_from_results, source_urn, dest_urn, max_depth
                )

                requirement_assessments_to_update: list[RequirementAssessment] = []

                target_requirement_assessments = RequirementAssessment.objects.filter(
                    compliance_assessment=instance,
                    requirement__urn__in=best_results["requirement_assessments"],
                )

                for req in target_requirement_assessments:
                    source = best_results["requirement_assessments"][
                        req.requirement.urn
                    ]
                    for field in [
                        "result",
                        "status",
                        "observation",
                        "score",
                        "is_scored",
                        "documentation_score",
                    ]:
                        if field in source and source[field] is not None:
                            setattr(req, field, source[field])
                        requirement_assessments_to_update.append(req)

                RequirementAssessment.objects.bulk_update(
                    requirement_assessments_to_update,
                    [
                        "result",
                        "status",
                        "observation",
                        "score",
                        "is_scored",
                        "documentation_score",
                    ],
                    batch_size=500,
                )

                for ra in requirement_assessments_to_update:
                    if best_results["requirement_assessments"][ra.requirement.urn].get(
                        "applied_controls"
                    ):
                        ra.applied_controls.add(
                            *[
                                control
                                for control in best_results["requirement_assessments"][
                                    ra.requirement.urn
                                ]["applied_controls"]
                            ]
                        )
                    if best_results["requirement_assessments"][ra.requirement.urn].get(
                        "evidences"
                    ):
                        ra.evidences.add(
                            *[
                                evidence
                                for evidence in best_results["requirement_assessments"][
                                    ra.requirement.urn
                                ]["evidences"]
                            ]
                        )
                    if best_results["requirement_assessments"][ra.requirement.urn].get(
                        "security_exceptions"
                    ):
                        ra.security_exceptions.add(
                            *[
                                exception
                                for exception in best_results[
                                    "requirement_assessments"
                                ][ra.requirement.urn]["security_exceptions"]
                            ]
                        )

            # Handle applied controls creation
            if create_applied_controls:
                # Prefetch all requirement assessments with their suggestions
                assessments = instance.requirement_assessments.all().prefetch_related(
                    "requirement__reference_controls"
                )

                # Create applied controls in bulk for each assessment
                for requirement_assessment in assessments:
                    requirement_assessment.create_applied_controls_from_suggestions()

    def perform_update(self, serializer):
        compliance_assessment = serializer.save()
        if compliance_assessment.show_documentation_score:
            ra_null_documentation_score = RequirementAssessment.objects.filter(
                compliance_assessment=compliance_assessment,
                is_scored=True,
                documentation_score__isnull=True,
            )
            ra_null_documentation_score.update(
                documentation_score=compliance_assessment.min_score
            )

    @action(detail=False, name="Compliance assessments per status")
    def per_status(self, request):
        data = assessment_per_status(request.user, ComplianceAssessment)
        return Response({"results": data})

    @action(detail=False, methods=["get"])
    def quality_check(self, request):
        """
        Returns the quality check of every compliance assessment
        """
        (viewable_objects, _, _) = RoleAssignment.get_accessible_object_ids(
            folder=Folder.get_root_folder(),
            user=request.user,
            object_type=ComplianceAssessment,
        )
        compliance_assessments = ComplianceAssessment.objects.filter(
            id__in=viewable_objects
        )
        res = [
            {"id": a.id, "name": a.name, "quality_check": a.quality_check()}
            for a in compliance_assessments
        ]
        return Response({"results": res})

    @action(detail=True, methods=["get"])
    def global_score(self, request, pk):
        """Returns the global score of the compliance assessment"""
        compliance_assessment = self.get_object()
        return Response(
            {
                "score": compliance_assessment.get_global_score(),
                "max_score": compliance_assessment.max_score,
                "min_score": compliance_assessment.min_score,
                "scores_definition": get_referential_translation(
                    compliance_assessment.framework, "scores_definition", get_language()
                ),
                "show_documentation_score": compliance_assessment.show_documentation_score,
            }
        )

    @action(detail=True, methods=["get"], url_path="quality_check")
    def quality_check_detail(self, request, pk):
        """
        Returns the quality check of a specific assessment
        """
        (viewable_objects, _, _) = RoleAssignment.get_accessible_object_ids(
            folder=Folder.get_root_folder(), user=request.user, object_type=Assessment
        )
        if UUID(pk) in viewable_objects:
            compliance_assessment = self.get_object()
            return Response(compliance_assessment.quality_check())
        else:
            return Response(status=status.HTTP_403_FORBIDDEN)

    @action(detail=True, methods=["get"])
    def tree(self, request, pk):
        compliance_assessment = self.get_object()
        _framework = compliance_assessment.framework
        tree = get_sorted_requirement_nodes(
            RequirementNode.objects.filter(framework=_framework)
            .select_related("framework")
            .all(),
            compliance_assessment.requirement_assessments.select_related(
                "requirement"
            ).all(),
            _framework.max_score,
        )
        implementation_groups = compliance_assessment.selected_implementation_groups
        return Response(
            filter_graph_by_implementation_groups(tree, implementation_groups)
        )

    @action(detail=True, methods=["get"])
    def requirements_list(self, request, pk):
        """Returns the list of requirement assessments for the different audit modes"""
        assessable = str(
            self.request.query_params.get("assessable", "false")
        ).lower() in {"true", "1", "yes"}
        compliance_assessment = self.get_object()
        requirement_assessments_objects = (
            compliance_assessment.get_requirement_assessments(
                include_non_assessable=not assessable
            )
        )
        requirements_objects = RequirementNode.objects.filter(
            framework=compliance_assessment.framework
        ).select_related("framework")
        requirement_assessments = RequirementAssessmentReadSerializer(
            requirement_assessments_objects, many=True
        ).data
        requirements = RequirementNodeReadSerializer(
            requirements_objects, many=True
        ).data
        requirements_list = {
            "requirements": requirements,
            "requirement_assessments": requirement_assessments,
        }
        return Response(requirements_list, status=status.HTTP_200_OK)

    @action(detail=True)
    def export(self, request, pk):
        def sanitize_filename(name):
            return regex.sub(r"[^\p{L}\p{N}\p{M}\-_.]+", "_", name)

        (object_ids_view, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, ComplianceAssessment
        )
        if UUID(pk) in object_ids_view:
            compliance_assessment = self.get_object()
            (index_content, evidences) = generate_html(compliance_assessment)
            zip_name = f"{sanitize_filename(compliance_assessment.name)}-{sanitize_filename(compliance_assessment.framework.name)}-{datetime.now():%Y-%m-%d-%H-%M}.zip"

            # Create temporary file that will be automatically deleted
            temp_file = tempfile.NamedTemporaryFile(delete=True, suffix=".zip")

            try:
                with zipfile.ZipFile(temp_file, "w") as zipf:
                    for evidence in evidences:
                        if (
                            evidence.last_revision
                            and evidence.last_revision.attachment
                            and default_storage.exists(
                                evidence.last_revision.attachment.name
                            )
                        ):
                            with default_storage.open(
                                evidence.last_revision.attachment.name
                            ) as attachment_file:
                                zipf.writestr(
                                    os.path.join(
                                        "evidences",
                                        os.path.basename(
                                            evidence.last_revision.attachment.name
                                        ),
                                    ),
                                    attachment_file.read(),
                                )
                    zipf.writestr("index.html", index_content)

                # Seek to beginning for reading
                temp_file.seek(0)

                # Create response - FileResponse will handle closing the temp_file
                response = FileResponse(
                    temp_file, as_attachment=True, filename=zip_name
                )
                return response

            except Exception:
                # Clean up on error
                temp_file.close()
                raise

        else:
            return Response({"error": "Permission denied"})

    @action(detail=True, methods=["get"])
    def donut_data(self, request, pk):
        compliance_assessment = ComplianceAssessment.objects.get(id=pk)
        return Response(compliance_assessment.donut_render())

    @action(detail=True, methods=["get"])
    def comparable_audits(self, request, pk):
        """
        Get list of compliance assessments that can be compared with this one
        (same framework, user has view permission, excludes current audit)
        """
        try:
            base_audit = self.get_object()
        except ComplianceAssessment.DoesNotExist:
            return Response(
                {"error": "Base audit not found"}, status=status.HTTP_404_NOT_FOUND
            )

        # Get viewable objects for permission checking
        (viewable_objects, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, ComplianceAssessment
        )

        # Filter audits: same framework, viewable, exclude current
        comparable_audits = (
            ComplianceAssessment.objects.filter(
                framework=base_audit.framework, id__in=viewable_objects
            )
            .exclude(id=UUID(pk))
            .select_related("folder", "framework", "perimeter")
            .order_by("-created_at")
        )

        # Build response with prioritization for same perimeter
        results = []
        for audit in comparable_audits:
            result = {
                "id": str(audit.id),
                "name": audit.name,
                "ref_id": audit.ref_id,
                "version": audit.version,
                "status": audit.status,
                "perimeter": {
                    "id": str(audit.perimeter.id),
                    "str": audit.perimeter.name,
                }
                if audit.perimeter
                else None,
                "folder": {"id": str(audit.folder.id), "str": audit.folder.name},
                "created_at": audit.created_at,
                # Flag for prioritization in frontend
                "same_perimeter": (
                    base_audit.perimeter
                    and audit.perimeter
                    and audit.perimeter.id == base_audit.perimeter.id
                ),
            }
            results.append(result)

        # Sort: same perimeter first, then by creation date
        results.sort(
            key=lambda x: (not x["same_perimeter"], x["created_at"]), reverse=True
        )

        return Response({"results": results})

    @action(detail=True, methods=["get"])
    def compare(self, request, pk):
        """
        Compare two compliance assessments that use the same framework
        """
        compare_id = request.query_params.get("compare_id")

        if not compare_id:
            return Response(
                {"error": "compare_id parameter is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Get viewable objects for permission checking
        (viewable_objects, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, ComplianceAssessment
        )

        # Check permissions for base audit
        if UUID(pk) not in viewable_objects:
            return Response(
                {"error": "Permission denied for base audit"},
                status=status.HTTP_403_FORBIDDEN,
            )

        # Check permissions for comparison audit
        if UUID(compare_id) not in viewable_objects:
            return Response(
                {"error": "Permission denied for comparison audit"},
                status=status.HTTP_403_FORBIDDEN,
            )

        try:
            base_audit = ComplianceAssessment.objects.get(id=pk)
            compare_audit = ComplianceAssessment.objects.get(id=compare_id)
        except ComplianceAssessment.DoesNotExist:
            return Response(
                {"error": "One or both audits not found"},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Validate same framework
        if base_audit.framework.id != compare_audit.framework.id:
            return Response(
                {"error": "Audits must use the same framework for comparison"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Helper function to aggregate data by top-level requirements
        def aggregate_by_top_level(audit):
            from core.helpers import get_referential_translation

            requirement_nodes = list(
                RequirementNode.objects.filter(framework=audit.framework).all()
            )
            requirement_assessments = list(
                audit.requirement_assessments.select_related("requirement").all()
            )

            # Build mapping of requirement_id to assessment
            req_assessment_map = {
                str(ra.requirement_id): ra for ra in requirement_assessments
            }

            # Build children dictionary for quick lookup
            children_dict = {}
            for rn in requirement_nodes:
                parent = rn.parent_urn or "root"
                if parent not in children_dict:
                    children_dict[parent] = []
                children_dict[parent].append(rn)

            # Get top-level nodes (no parent)
            top_level_nodes = children_dict.get("root", [])

            # Sort by order_id if available
            for node in top_level_nodes:
                if node.order_id is None:
                    node.order_id = node.created_at.timestamp()
            top_level_nodes.sort(key=lambda x: x.order_id)

            radar_data = {
                "labels": [],
                "compliance_percentages": [],
                "maturity_scores": [],
            }

            # Collect all assessable descendants recursively
            def collect_assessable_descendants(node_urn):
                assessable = []
                node_children = children_dict.get(node_urn, [])

                for child in node_children:
                    # If this child is assessable, add its assessment
                    if child.assessable:
                        ra = req_assessment_map.get(str(child.id))
                        if ra:
                            assessable.append(ra)

                    # Recursively collect from this child's descendants
                    assessable.extend(collect_assessable_descendants(child.urn))

                return assessable

            for node in top_level_nodes:
                # Try multiple ways to get a meaningful name
                node_name = (
                    get_referential_translation(node, "name")
                    or node.name
                    or node.ref_id
                    or f"Node {node.id}"
                )
                radar_data["labels"].append(node_name)

                # Check if the node itself is assessable
                assessable_list = []
                if node.assessable:
                    ra = req_assessment_map.get(str(node.id))
                    if ra:
                        assessable_list.append(ra)

                # Add all assessable descendants
                assessable_list.extend(collect_assessable_descendants(node.urn))

                # Calculate compliance percentage (compliant or partially_compliant assessments)
                if assessable_list:
                    compliant = sum(
                        1
                        for ra in assessable_list
                        if ra.result in ["compliant", "partially_compliant"]
                    )
                    compliance_percentage = (compliant / len(assessable_list)) * 100
                else:
                    compliance_percentage = 0

                # Calculate maturity score (average score, not percentage)
                scored_list = [
                    ra
                    for ra in assessable_list
                    if ra.is_scored and ra.result != "not_applicable"
                ]
                if scored_list:
                    total_score = sum(ra.score or 0 for ra in scored_list)
                    # Calculate mean score (same as frontend nodeScore function)
                    maturity_score = total_score / len(scored_list)
                else:
                    maturity_score = 0

                radar_data["compliance_percentages"].append(
                    round(compliance_percentage, 1)
                )
                radar_data["maturity_scores"].append(round(maturity_score, 1))

            return radar_data

        # Build comparison data
        comparison_data = {
            "framework": {
                "id": str(base_audit.framework.id),
                "str": base_audit.framework.name,
            },
            "base": {
                "id": str(base_audit.id),
                "name": base_audit.name,
                "version": base_audit.version,
                "status": base_audit.status,
                "perimeter": {
                    "id": str(base_audit.perimeter.id),
                    "str": base_audit.perimeter.name,
                }
                if base_audit.perimeter
                else None,
                "selected_implementation_groups": base_audit.get_selected_implementation_groups(),
                "created_at": base_audit.created_at,
                "updated_at": base_audit.updated_at,
                "observation": base_audit.observation,
                "global_score": base_audit.get_global_score(),
                "max_score": base_audit.max_score,
                "donut_data": base_audit.donut_render(),
                "radar_data": aggregate_by_top_level(base_audit),
            },
            "compare": {
                "id": str(compare_audit.id),
                "name": compare_audit.name,
                "version": compare_audit.version,
                "status": compare_audit.status,
                "perimeter": {
                    "id": str(compare_audit.perimeter.id),
                    "str": compare_audit.perimeter.name,
                }
                if compare_audit.perimeter
                else None,
                "selected_implementation_groups": compare_audit.get_selected_implementation_groups(),
                "created_at": compare_audit.created_at,
                "updated_at": compare_audit.updated_at,
                "observation": compare_audit.observation,
                "global_score": compare_audit.get_global_score(),
                "max_score": compare_audit.max_score,
                "donut_data": compare_audit.donut_render(),
                "radar_data": aggregate_by_top_level(compare_audit),
            },
        }

        # Build differences list
        differences = []

        # Get all requirement assessments from both audits
        base_ras = base_audit.requirement_assessments.select_related(
            "requirement"
        ).all()
        compare_ras_dict = {
            ra.requirement_id: ra
            for ra in compare_audit.requirement_assessments.select_related(
                "requirement"
            ).all()
        }

        for base_ra in base_ras:
            compare_ra = compare_ras_dict.get(base_ra.requirement_id)

            # Skip if no matching requirement in compare audit
            if not compare_ra:
                continue

            # Check if result or score is different
            result_different = base_ra.result != compare_ra.result
            score_different = base_ra.score != compare_ra.score

            if result_different or score_different:
                # Use safe_display_str for smart fallback: name → ref_id → description → URN
                req_name = base_ra.requirement.safe_display_str

                differences.append(
                    {
                        "requirement": {
                            "id": str(base_ra.requirement.id),
                            "ref_id": base_ra.requirement.ref_id,
                            "name": req_name,
                            "description": base_ra.requirement.description,
                        },
                        "base": {
                            "id": str(base_ra.id),
                            "result": base_ra.result,
                            "status": base_ra.status,
                            "score": base_ra.score,
                        },
                        "compare": {
                            "id": str(compare_ra.id),
                            "result": compare_ra.result,
                            "status": compare_ra.status,
                            "score": compare_ra.score,
                        },
                    }
                )

        comparison_data["differences"] = differences

        return Response(comparison_data)

    @staticmethod
    @api_view(["POST"])
    @renderer_classes([JSONRenderer])
    def create_suggested_applied_controls(request, pk):
        compliance_assessment = ComplianceAssessment.objects.get(id=pk)
        if not RoleAssignment.is_access_allowed(
            user=request.user,
            perm=Permission.objects.get(codename="add_appliedcontrol"),
            folder=compliance_assessment.folder,
        ):
            return Response(status=status.HTTP_403_FORBIDDEN)
        requirement_assessments = compliance_assessment.requirement_assessments.all()
        controls = []
        for requirement_assessment in requirement_assessments:
            controls.append(
                requirement_assessment.create_applied_controls_from_suggestions()
            )
        return Response(
            AppliedControlReadSerializer(chain.from_iterable(controls), many=True).data,
            status=status.HTTP_200_OK,
        )

    @action(
        detail=True,
        methods=["get", "post"],
        url_path="syncToActions",
    )
    def sync_to_applied_controls(self, request, pk):
        dry_run = request.query_params.get("dry_run", True)
        if dry_run == "false":
            dry_run = False
        compliance_assessment = ComplianceAssessment.objects.get(id=pk)

        if not RoleAssignment.is_access_allowed(
            user=request.user,
            perm=Permission.objects.get(codename="change_requirementassessment"),
            folder=compliance_assessment.folder,
        ):
            return Response(status=status.HTTP_403_FORBIDDEN)

        changes = compliance_assessment.sync_to_applied_controls(dry_run=dry_run)
        return Response({"changes": changes})

    @action(detail=True, methods=["get"], url_path="progress_ts")
    def progress_ts(self, request, pk):
        try:
            raw = (
                HistoricalMetric.objects.filter(
                    model="ComplianceAssessment", object_id=pk
                )
                .annotate(progress=F("data__reqs__progress_perc"))
                .values("date", "progress")
                .order_by("date")
            )

            # Transform the data into the required format
            formatted_data = [
                [entry["date"].isoformat(), entry["progress"]] for entry in raw
            ]

            return Response({"data": formatted_data})

        except HistoricalMetric.DoesNotExist:
            return Response(
                {"error": "No metrics found for this assessment"},
                status=status.HTTP_404_NOT_FOUND,
            )

    @action(detail=True, methods=["get"])
    def threats_metrics(self, request, pk=None):
        compliance_assessment = self.get_object()
        self.check_object_permissions(request, compliance_assessment)

        threat_metrics = compliance_assessment.get_threats_metrics()
        if threat_metrics.get("total_unique_threats") == 0:
            return Response(threat_metrics, status=status.HTTP_200_OK)
        children = []
        for th in threat_metrics["threats"]:
            children.append(
                {
                    "name": th["name"],
                    "children": [
                        ra["requirement_name"] for ra in th["requirement_assessments"]
                    ],
                    "value": len(th["requirement_assessments"]),
                }
            )
        tree = {"name": "Threats", "children": children}
        nodes = []
        for th in threat_metrics["threats"]:
            nodes.append(
                {
                    "name": th["name"],
                    "value": len(th["requirement_assessments"]),
                    "items": [
                        f"{ra['requirement_name']} ({ra['result']})"
                        for ra in th["requirement_assessments"]
                    ],
                }
            )
        threat_metrics.update({"tree": tree, "graph": {"nodes": nodes}})

        return Response(threat_metrics, status=status.HTTP_200_OK)

    @action(detail=False, methods=["get"], name="Get compliance analytics")
    def analytics(self, request):
        """
        Returns compliance analytics data grouped by framework and domain
        """
        folder_id = request.query_params.get("folder", None)
        analytics_data = get_compliance_analytics(request.user, folder_id)
        return Response(analytics_data, status=status.HTTP_200_OK)


class RequirementAssessmentViewSet(BaseModelViewSet):
    """
    API endpoint that allows requirement assessments to be viewed or edited.
    """

    model = RequirementAssessment
    filterset_fields = [
        "folder",
        "folder__name",
        "evidences",
        "compliance_assessment",
        "applied_controls",
        "security_exceptions",
        "requirement__ref_id",
        "result",
        "compliance_assessment__ref_id",
        "compliance_assessment__perimeter",
        "compliance_assessment__perimeter__name",
        "compliance_assessment__assets__ref_id",
        "requirement__assessable",
    ]
    search_fields = [
        "requirement__name",
        "requirement__description",
        "requirement__ref_id",
    ]

    def get_queryset(self):
        """Optimize queries for table view and serializer - high-impact due to many nested relationships"""
        return (
            super()
            .get_queryset()
            .select_related(
                "folder",
                "folder__parent_folder",  # For get_folder_full_path() optimization
                "compliance_assessment",  # Displayed in table and serialized
                "compliance_assessment__perimeter",  # perimeter field uses compliance_assessment.perimeter
                "compliance_assessment__perimeter__folder",  # Nested FieldsRelatedField optimization
                "requirement",  # Used for name (__str__), description, assessable in table
            )
            .prefetch_related(
                "evidences",  # ManyToManyField serialized as FieldsRelatedField
                "applied_controls",  # ManyToManyField to AppliedControl
                "security_exceptions",  # ManyToManyField serialized as FieldsRelatedField
            )
        )

    def update(self, request, *args, **kwargs):
        response = super().update(request, *args, **kwargs)
        cache.clear()
        return response

    @action(detail=False, name="Get updatable measures")
    def updatables(self, request):
        (_, object_ids_change, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, AppliedControl
        )

        return Response({"results": object_ids_change})

    @action(
        detail=False, name="Something"
    )  # Write a good name for the "name" keyword argument
    def per_status(self, request):
        data = applied_control_per_status(request.user)
        return Response({"results": data})

    @action(detail=False, name="Get the ordered todo applied controls")
    def todo(self, request):
        (object_ids_view, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, AppliedControl
        )

        measures = sorted(
            AppliedControl.objects.filter(id__in=object_ids_view)
            .exclude(status="done")
            .order_by("eta"),
            key=lambda mtg: mtg.get_ranking_score(),
            reverse=True,
        )

        """measures = [{
            key: getattr(mtg,key)
            for key in [
                "id","folder","reference_control","type","status","effort","cost","name","description","eta","link","created_at","updated_at"
            ]
        } for mtg in measures]
        for i in range(len(measures)) :
            measures[i]["id"] = str(measures[i]["id"])
            measures[i]["folder"] = str(measures[i]["folder"].name)
            for key in ["created_at","updated_at","eta"] :
                measures[i][key] = str(measures[i][key])"""

        ranking_scores = {str(mtg.id): mtg.get_ranking_score() for mtg in measures}

        measures = [AppliedControlReadSerializer(mtg).data for mtg in measures]

        for i in range(len(measures)):
            measures[i]["ranking_score"] = ranking_scores[measures[i]["id"]]

        """
        The serializer of AppliedControl isn't applied automatically for this function
        """

        return Response({"results": measures})

    @action(detail=False, name="Get the secuity measures to review")
    def to_review(self, request):
        measures = measures_to_review(request.user)
        measures = [AppliedControlReadSerializer(mtg).data for mtg in measures]

        """
        The serializer of AppliedControl isn't applied automatically for this function
        """

        return Response({"results": measures})

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get status choices")
    def status(self, request):
        return Response(dict(RequirementAssessment.Status.choices))

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get result choices")
    def result(self, request):
        return Response(dict(RequirementAssessment.Result.choices))

    @staticmethod
    @api_view(["POST"])
    @renderer_classes([JSONRenderer])
    def create_suggested_applied_controls(request, pk):
        requirement_assessment = RequirementAssessment.objects.get(id=pk)
        if not RoleAssignment.is_access_allowed(
            user=request.user,
            perm=Permission.objects.get(codename="add_appliedcontrol"),
            folder=requirement_assessment.folder,
        ):
            return Response(status=status.HTTP_403_FORBIDDEN)
        controls = requirement_assessment.create_applied_controls_from_suggestions()
        return Response(
            AppliedControlReadSerializer(controls, many=True).data,
            status=status.HTTP_200_OK,
        )


class RequirementMappingSetViewSet(BaseModelViewSet):
    model = StoredLibrary

    filterset_fields = [
        "provider",
    ]

    def get_serializer_class(self, **kwargs):
        return RequirementMappingSetReadSerializer

    def get_queryset(self):
        return (
            super()
            .get_queryset()
            .filter(
                Q(content__requirement_mapping_set__isnull=False)
                | Q(content__requirement_mapping_sets__isnull=False)
            )
        )

    @action(detail=False, name="Get provider choices")
    def provider(self, request):
        providers = set(
            StoredLibrary.objects.filter(
                provider__isnull=False,
                objects_meta__requirement_mapping_sets__isnull=False,
            ).values_list("provider", flat=True)
        )
        return Response({p: p for p in providers})

    @action(detail=False, methods=["get"], url_path="graph-data")
    def graph_data_list(self, request):
        from core.mappings.engine import engine

        max_depth = get_mapping_max_depth()

        all_paths = engine.get_mapping_graph(max_depth=max_depth)

        all_framework_urns = set()
        for path in all_paths:
            all_framework_urns.update(path)

        framework_details = {}

        framework_libs_qs = StoredLibrary.objects.filter(
            content__framework__urn__in=all_framework_urns,
            content__framework__isnull=False,
            content__requirement_mapping_set__isnull=True,
            content__requirement_mapping_sets__isnull=True,
        )

        for lib in framework_libs_qs:
            try:
                fw_content = lib.content["framework"]
                urn = fw_content.get("urn")

                if not urn:
                    continue

                display_name = fw_content.get("name", urn)

                display_value = fw_content.get("description", display_name)

                framework_details[urn] = {
                    "name": display_name,
                    "value": display_value,
                    "pk": lib.pk,
                }
            except (KeyError, TypeError):
                # Skip this library if content is malformed
                continue

        nodes = []
        urn_to_index = {}  # Map URNs to their index in the `nodes` list

        # Sort for a consistent node order
        sorted_urns = sorted(list(all_framework_urns))

        for i, urn in enumerate(sorted_urns):
            urn_to_index[urn] = i

            details = framework_details.get(urn, {"name": urn, "value": urn})

            nodes.append(
                {
                    "name": details["name"],  # Display name for the node
                    "category": 0
                    if urn in engine.frameworks.keys()
                    else 1,  # All nodes are in the "Frameworks" category
                    "value": details["value"],  # Tooltip content
                    "urn": urn,  # Pass URN as extra data
                    "pk": details.get("pk"),
                    # `symbolSize` will be added below after degrees are calculated
                }
            )

        links = []
        link_set = set()  # Use a set to prevent duplicate links

        node_degrees = [0] * len(nodes)

        for path in all_paths:
            # A path [A, B, C] creates links (A, B) and (B, C)
            for i in range(len(path) - 1):
                source_urn = path[i]
                target_urn = path[i + 1]
                link_tuple = (source_urn, target_urn)

                if link_tuple not in link_set:
                    link_set.add(link_tuple)

                    # Get the integer index for source and target
                    source_index = urn_to_index.get(source_urn)
                    target_index = urn_to_index.get(target_urn)

                    if source_index is not None and target_index is not None:
                        links.append(
                            {
                                "source": source_index,
                                "target": target_index,
                                "value": "maps to",
                            }
                        )

                        node_degrees[source_index] += 1
                        node_degrees[target_index] += 1

        min_size = 10
        max_size = 30

        non_zero_degrees = [d for d in node_degrees if d > 0]

        if not non_zero_degrees:
            min_degree = 0
            max_degree = 0
        else:
            min_degree = min(non_zero_degrees)
            max_degree = max(node_degrees)

        degree_range = max_degree - min_degree
        size_range = max_size - min_size

        for i, node in enumerate(nodes):
            degree = node_degrees[i]

            if degree == 0:
                # Assign min size for nodes with no connections (if any)
                node["symbolSize"] = min_size
            elif degree_range == 0:
                # All nodes have the same degree, assign min_size
                node["symbolSize"] = min_size
            else:
                # Scale the size proportionally
                normalized_degree = (degree - min_degree) / degree_range
                scaled_size = min_size + (normalized_degree * size_range)
                node["symbolSize"] = int(round(scaled_size))

        categories = [{"name": "importedFrameworks"}, {"name": "notImportedFrameworks"}]

        return Response({"nodes": nodes, "links": links, "categories": categories})

    @action(detail=True, methods=["get"], url_path="graph_data")
    def graph_data(self, request, pk=None):
        obj = StoredLibrary.objects.get(id=pk)

        mapping_set = obj.content.get(
            "requirement_mapping_sets",
            [obj.content.get("requirement_mapping_set", {})],
        )[0]

        source_framework_lib = StoredLibrary.objects.filter(
            content__framework__urn=mapping_set["source_framework_urn"],
            content__framework__isnull=False,
            content__requirement_mapping_set__isnull=True,
            content__requirement_mapping_sets__isnull=True,
        ).first()
        if not source_framework_lib:
            raise NotFound("Source framework library not found")

        target_framework_lib = StoredLibrary.objects.filter(
            content__framework__urn=mapping_set["target_framework_urn"],
            content__framework__isnull=False,
            content__requirement_mapping_set__isnull=True,
            content__requirement_mapping_sets__isnull=True,
        ).first()
        if not target_framework_lib:
            raise NotFound("Target framework library not found")

        source_framework = source_framework_lib.content["framework"]
        target_framework = target_framework_lib.content["framework"]

        source_nodes_dict = {
            n.get("urn"): n for n in source_framework["requirement_nodes"]
        }
        target_nodes_dict = {
            n.get("urn"): n for n in target_framework["requirement_nodes"]
        }

        nodes = []
        links = []
        snodes_idx = dict()
        tnodes_idx = dict()
        categories = [
            {
                "name": source_framework["name"],
            },
            {
                "name": target_framework["name"],
            },
        ]
        N = 0
        for req in source_framework["requirement_nodes"]:
            if not req.get("assessable", False):
                continue
            nodes.append(
                {
                    "name": req.get("ref_id"),
                    "category": 0,
                    "value": req.get("name", req.get("description")),
                }
            )
            snodes_idx[req.get("urn")] = N
            N += 1

        for req in target_framework["requirement_nodes"]:
            if not req.get("assessable", False):
                continue
            nodes.append(
                {
                    "name": req.get("ref_id"),
                    "category": 1,
                    "value": req.get("name", req.get("description")),
                }
            )
            tnodes_idx[req.get("urn")] = N
            N += 1
        req_mappings = mapping_set.get("requirement_mappings", [])
        for mapping in req_mappings:
            source_urn = mapping.get("source_requirement_urn")
            target_urn = mapping.get("target_requirement_urn")
            source_node = source_nodes_dict.get(source_urn)
            target_node = target_nodes_dict.get(target_urn)

            if (
                mapping.get("source_requirement_urn") not in snodes_idx
                or mapping.get("target_requirement_urn") not in tnodes_idx
                or not source_node
                or not target_node
            ):
                continue

            links.append(
                {
                    "source": snodes_idx[source_node.get("urn")],
                    "target": tnodes_idx[target_node.get("urn")],
                    "value": mapping.get("relationship"),
                }
            )

        # Calculate coverage in both directions
        source_assessable_count = len(snodes_idx)
        target_assessable_count = len(tnodes_idx)

        linked_source_urns = set(
            mapping.get("source_requirement_urn")
            for mapping in req_mappings
            if mapping.get("source_requirement_urn") in snodes_idx
        )
        linked_target_urns = set(
            mapping.get("target_requirement_urn")
            for mapping in req_mappings
            if mapping.get("target_requirement_urn") in tnodes_idx
        )

        source_coverage = (
            round(len(linked_source_urns) / source_assessable_count * 100)
            if source_assessable_count > 0
            else 0
        )
        target_coverage = (
            round(len(linked_target_urns) / target_assessable_count * 100)
            if target_assessable_count > 0
            else 0
        )

        meta = {
            "display_name": f"{source_framework['name']} ➜ {target_framework['name']}",
            "source_framework": source_framework["name"],
            "target_framework": target_framework["name"],
            "source_coverage": source_coverage,
            "target_coverage": target_coverage,
            "source_total": source_assessable_count,
            "source_linked": len(linked_source_urns),
            "target_total": target_assessable_count,
            "target_linked": len(linked_target_urns),
        }

        return Response(
            {"nodes": nodes, "links": links, "categories": categories, "meta": meta}
        )


@api_view(["GET"])
@permission_classes([permissions.AllowAny])
def get_csrf_token(request):
    """
    API endpoint that returns the CSRF token.
    """
    return Response({"csrfToken": csrf.get_token(request)})


def get_disk_usage():
    try:
        path = Path(settings.BASE_DIR) / "db"
        usage = shutil.disk_usage(path)
        return usage
    except PermissionError:
        logger.error(
            "Permission issue: cannot access the path to retrieve the disk_usage info"
        )
        return None
    except FileNotFoundError:
        logger.error(
            "Path issue: cannot access the path to retrieve the disk_usage info"
        )
        return None


@api_view(["GET"])
@permission_classes([permissions.AllowAny])
def healthcheck(request):
    return Response({"status": "ok"})


@api_view(["GET"])
def get_build(request):
    """
    API endpoint that returns the build version of the application.
    """
    BUILD = settings.BUILD
    VERSION = settings.VERSION
    default_db_engine = settings.DATABASES["default"]["ENGINE"]
    if "postgresql" in default_db_engine:
        database_type = "P-FS"
    elif "sqlite" in default_db_engine:
        database_type = "S-FS"
    else:
        database_type = "Unknown"

    disk_info = get_disk_usage()

    if disk_info:
        total, used, free = disk_info
        disk_response = {
            "diskSpace": f"{humanize.naturalsize(total)}",
            "diskUsed": f"{humanize.naturalsize(used)} ({int((used / total) * 100)} %)",
        }
    else:
        disk_response = {
            "diskSpace": "Unable to retrieve disk usage",
        }

    return Response(
        {
            "version": VERSION,
            "build": BUILD,
            "infrastructure": database_type,
            **disk_response,
        }
    )


# NOTE: Important functions/classes from old views.py, to be reviewed


def generate_html(
    compliance_assessment: ComplianceAssessment,
) -> Tuple[str, list[Evidence]]:
    selected_evidences = []

    requirement_nodes = RequirementNode.objects.filter(
        framework=compliance_assessment.framework
    )

    assessments = RequirementAssessment.objects.filter(
        compliance_assessment=compliance_assessment,
    ).all()

    implementation_groups = compliance_assessment.selected_implementation_groups
    graph = get_sorted_requirement_nodes(
        list(requirement_nodes),
        list(assessments),
        compliance_assessment.framework.max_score,
    )
    graph = filter_graph_by_implementation_groups(graph, implementation_groups)
    flattened_graph = flatten_dict(graph)

    requirement_nodes = requirement_nodes.filter(urn__in=flattened_graph.values())
    assessments = assessments.filter(requirement__urn__in=flattened_graph.values())

    node_per_urn = {r.urn: r for r in requirement_nodes}
    ancestors = {}
    for a in assessments:
        ancestors[a] = set()
        req = a.requirement
        while req:
            ancestors[a].add(req)
            p = req.parent_urn
            req = None if not (p) else node_per_urn[p]

    def generate_data_rec(requirement_node: RequirementNode):
        selected_evidences = []
        children_nodes = [
            req for req in requirement_nodes if req.parent_urn == requirement_node.urn
        ]

        node_data = {
            "requirement_node": requirement_node,
            "children": [],
            "assessments": None,
            "bar_graph": None,
            "direct_evidences": [],
            "applied_controls": [],
            "result": "",
            "status": "",
            "color_class": "",
        }

        node_data["bar_graph"] = True if children_nodes else False

        if requirement_node.assessable:
            assessment = RequirementAssessment.objects.filter(
                requirement__urn=requirement_node.urn,
                compliance_assessment=compliance_assessment,
            ).first()

            if assessment:
                node_data["assessments"] = assessment
                node_data["result"] = assessment.get_result_display()
                node_data["status"] = assessment.get_status_display()
                node_data["result_color_class"] = color_css_class(assessment.result)
                node_data["status_color_class"] = color_css_class(assessment.status)
                direct_evidences = assessment.evidences.all()
                if direct_evidences:
                    selected_evidences += direct_evidences
                    node_data["direct_evidences"] = direct_evidences

                measures = assessment.applied_controls.all()
                if measures:
                    applied_controls = []
                    for measure in measures:
                        evidences = measure.evidences.all()
                        applied_controls.append(
                            {
                                "measure": measure,
                                "evidences": evidences,
                            }
                        )
                        selected_evidences += evidences
                    node_data["applied_controls"] = applied_controls

        for child_node in children_nodes:
            child_data, child_evidences = generate_data_rec(child_node)
            node_data["children"].append(child_data)
            selected_evidences += child_evidences

        return node_data, selected_evidences

    top_level_nodes = [req for req in requirement_nodes if not req.parent_urn]
    update_translations_in_object(top_level_nodes)
    top_level_nodes_data = []
    for requirement_node in top_level_nodes:
        node_data, node_evidences = generate_data_rec(requirement_node)
        top_level_nodes_data.append(node_data)
        selected_evidences += node_evidences

    data = {
        "compliance_assessment": compliance_assessment,
        "top_level_nodes": top_level_nodes_data,
        "assessments": assessments,
        "ancestors": ancestors,
    }

    return render_to_string("core/audit_report.html", data), list(
        set(selected_evidences)
    )


def export_mp_csv(request):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="MP.csv"'

    writer = csv.writer(response, delimiter=";")
    columns = [
        "measure_id",
        "measure_name",
        "measure_desc",
        "category",
        "csf_function",
        "reference_control",
        "eta",
        "priority",
        "effort",
        "control_impact",
        "annual_cost",
        "link",
        "status",
    ]
    writer.writerow(columns)
    (
        object_ids_view,
        object_ids_change,
        object_ids_delete,
    ) = RoleAssignment.get_accessible_object_ids(
        Folder.get_root_folder(), request.user, AppliedControl
    )
    for mtg in AppliedControl.objects.filter(id__in=object_ids_view):
        row = [
            mtg.id,
            mtg.name,
            mtg.description,
            mtg.category,
            mtg.csf_function,
            mtg.priority,
            mtg.reference_control,
            mtg.eta,
            mtg.effort,
            mtg.control_impact,
            mtg.annual_cost,
            mtg.link,
            mtg.status,
        ]
        writer.writerow(row)

    return response


class SecurityExceptionViewSet(ExportMixin, BaseModelViewSet):
    """
    API endpoint that allows security exceptions to be viewed or edited.
    """

    model = SecurityException
    filterset_fields = [
        "name",
        "requirement_assessments",
        "risk_scenarios",
        "owners",
        "approver",
        "folder",
        "severity",
        "status",
        "genericcollection",
    ]
    search_fields = ["name", "description", "ref_id"]

    export_config = {
        "fields": {
            "internal_id": {"source": "id", "label": "internal_id"},
            "ref_id": {"source": "ref_id", "label": "ref_id", "escape": True},
            "name": {"source": "name", "label": "name", "escape": True},
            "description": {
                "source": "description",
                "label": "description",
                "escape": True,
            },
            "severity": {"source": "get_severity_display", "label": "severity"},
            "status": {"source": "get_status_display", "label": "status"},
            "expiration_date": {
                "source": "expiration_date",
                "label": "expiration_date",
            },
            "owners": {
                "source": "owners",
                "label": "owners",
                "format": lambda qs: ",".join(
                    escape_excel_formula(o.email) for o in qs.all()
                ),
            },
            "approver": {
                "source": "approver.email",
                "label": "approver",
                "escape": True,
            },
            "folder": {
                "source": "folder.name",
                "label": "folder",
                "escape": True,
            },
        },
        "filename": "security_exceptions_export",
        "select_related": ["folder", "approver"],
        "prefetch_related": ["owners"],
    }

    @action(detail=False, name="Get severity choices")
    def severity(self, request):
        return Response(dict(Severity.choices))

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get status choices")
    def status(self, request):
        return Response(dict(SecurityException.Status.choices))

    def get_queryset(self):
        return (
            super()
            .get_queryset()
            .select_related("folder")
            .prefetch_related(
                "assets",
                "applied_controls",
                "vulnerabilities",
                "risk_scenarios",
                "requirement_assessments",
                "owners",
            )
        )

    @action(detail=False, name="Get security exception Sankey data")
    def sankey_data(self, request):
        folder_id = request.query_params.get("folder", None)
        (viewable_objects, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, SecurityException
        )
        queryset = SecurityException.objects.filter(id__in=viewable_objects)

        if folder_id:
            folder = Folder.objects.get(id=folder_id)
            queryset = queryset.filter(folder=folder)

        # Get severity and status combinations
        from django.db.models import Count

        combinations = (
            queryset.values("severity", "status")
            .annotate(count=Count("id"))
            .filter(count__gt=0, status__isnull=False)
            .exclude(status="")
        )

        # Build Sankey data structure
        nodes = []
        links = []
        node_names = set()

        # Create maps for severity and status labels
        severity_choice_map = {choice[0]: choice[1] for choice in Severity.choices}
        status_choice_map = {
            choice[0]: choice[1] for choice in SecurityException.Status.choices
        }

        # Create severity and status nodes only for data that exists
        severity_map = {}
        status_map = {}

        for combo in combinations:
            # Create severity node if not already created
            if combo["severity"] not in severity_map:
                severity_label = f"Severity: {severity_choice_map.get(combo['severity'], 'Unknown').title()}"
                severity_map[combo["severity"]] = severity_label
                node_names.add(severity_label)

            # Create status node if not already created
            if combo["status"] not in status_map:
                status_label = f"Status: {status_choice_map.get(combo['status'], 'Unknown').title()}"
                status_map[combo["status"]] = status_label
                node_names.add(status_label)

        # Convert node names to indexed list
        nodes = [{"name": name} for name in sorted(node_names)]
        node_index = {name: i for i, name in enumerate(sorted(node_names))}

        # Create links
        for combo in combinations:
            severity_name = severity_map[combo["severity"]]
            status_name = status_map[combo["status"]]

            links.append(
                {
                    "source": node_index[severity_name],
                    "target": node_index[status_name],
                    "value": combo["count"],
                }
            )

        return Response({"results": {"nodes": nodes, "links": links}})


class FindingsAssessmentViewSet(BaseModelViewSet):
    model = FindingsAssessment
    filterset_fields = [
        "name",
        "ref_id",
        "owner",
        "category",
        "perimeter",
        "folder",
        "authors",
        "status",
        "evidences",
        "genericcollection",
    ]
    search_fields = ["name", "description", "ref_id"]

    def get_queryset(self):
        return (
            super()
            .get_queryset()
            .select_related("folder", "perimeter")
            .prefetch_related(
                "evidences",
                "authors",
                "owner",
            )
        )

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get status choices")
    def status(self, request):
        return Response(dict(FindingsAssessment.Status.choices))

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get category choices")
    def category(self, request):
        return Response(dict(FindingsAssessment.Category.choices))

    @action(detail=True, name="Get Follow up metrics")
    def metrics(self, request, pk=None):
        assessment = self.get_object()
        raw_metrics = assessment.get_findings_metrics()

        def format_severity_data(metrics):
            severity_colors = {
                "info": "#3B82F6",
                "low": "#59BBB2",
                "medium": "#F5C481",
                "high": "#E6686D",
                "critical": "#C71E1D",
                "undefined": "#CCCCCC",
            }

            severity_chart_data = []
            for severity, count in metrics["severity_distribution"].items():
                severity_chart_data.append(
                    {
                        "name": severity.capitalize(),
                        "value": count,
                        "color": severity_colors.get(severity, "#CCCCCC"),
                    }
                )

            return severity_chart_data

        def format_status_data(metrics):
            status_mapping = {
                "identified": {"localName": "identified", "color": "#F5C481"},
                "confirmed": {"localName": "confirmed", "color": "#E6686D"},
                "assigned": {"localName": "assigned", "color": "#fab998"},
                "in_progress": {"localName": "inProgress", "color": "#fac858"},
                "mitigated": {
                    "localName": "mitigated",
                    "color": "hsl(80deg, 80%, 60%)",
                },
                "resolved": {"localName": "resolved", "color": "hsl(120deg, 80%, 45%)"},
                "dismissed": {"localName": "dismissed", "color": "#5470c6"},
                "deprecated": {"localName": "deprecated", "color": "#91cc75"},
                "--": {"localName": "undefined", "color": "#CCCCCC"},
            }

            grouped_status_counts = {}

            for status, count in metrics["status_distribution"].items():
                mapping_info = status_mapping.get(
                    status, {"localName": "other", "color": "#CCCCCC"}
                )
                local_name = mapping_info["localName"]
                color = mapping_info["color"]

                if local_name in grouped_status_counts:
                    grouped_status_counts[local_name]["value"] += count
                else:
                    grouped_status_counts[local_name] = {
                        "value": count,
                        "localName": local_name,
                        "itemStyle": {"color": color},
                    }

            status_values = list(grouped_status_counts.values())

            expected_statuses = ["open", "mitigate", "accept", "avoid", "transfer"]
            for status in expected_statuses:
                if not any(item["localName"] == status for item in status_values):
                    status_values.append(
                        {
                            "value": 0,
                            "localName": status,
                            "itemStyle": {"color": "#CCCCCC"},
                        }
                    )

            return {"values": status_values}

        formatted_metrics = {
            "raw_metrics": raw_metrics,
            "severity_chart_data": format_severity_data(raw_metrics),
            "status_chart_data": format_status_data(raw_metrics),
        }

        return Response(formatted_metrics)

    @action(detail=True, methods=["get"], name="Findings Assessment as Excel")
    def xlsx(self, request, pk):
        (viewable_objects, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, FindingsAssessment
        )
        if UUID(pk) not in viewable_objects:
            return Response(
                {"error": "Permission denied"}, status=status.HTTP_403_FORBIDDEN
            )

        findings_assessment = FindingsAssessment.objects.get(id=pk)
        findings = Finding.objects.filter(findings_assessment=pk).order_by("ref_id")

        # Prepare data for Excel export
        entries = []
        for finding in findings:
            entry = {
                "ref_id": finding.ref_id,
                "name": finding.name,
                "description": finding.description,
                "status": finding.get_status_display(),
                "severity": finding.get_severity_display(),
                "folder": finding.folder.name if finding.folder else "",
                "owner": ", ".join([user.email for user in finding.owner.all()]),
                "applied_controls": "\n".join(
                    [
                        f"{ac.name} [{ac.get_status_display().lower()}]"
                        for ac in finding.applied_controls.all()
                    ]
                ),
                "evidences": "\n".join([ev.name for ev in finding.evidences.all()]),
                "created_at": finding.created_at.strftime("%Y-%m-%d %H:%M:%S")
                if finding.created_at
                else "",
                "eta": finding.eta.strftime("%Y-%m-%d") if finding.eta else "",
                "due_date": finding.due_date.strftime("%Y-%m-%d")
                if finding.due_date
                else "",
            }
            entries.append(entry)

        df = pd.DataFrame(entries)
        buffer = io.BytesIO()

        # Create ExcelWriter with openpyxl engine
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Findings")

            # Get the worksheet
            worksheet = writer.sheets["Findings"]

            # Apply text wrapping to columns with line breaks
            wrap_columns = ["name", "description", "applied_controls", "evidences"]
            wrap_indices = [
                df.columns.get_loc(col) + 1 for col in wrap_columns if col in df.columns
            ]

            for col_idx in wrap_indices:
                for row_idx in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.alignment = Alignment(wrap_text=True)

            # Adjust column widths
            for idx, col in enumerate(df.columns):
                column_width = 40 if col in wrap_columns else 20
                worksheet.column_dimensions[
                    worksheet.cell(row=1, column=idx + 1).column_letter
                ].width = column_width

        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{findings_assessment.name}_findings.xlsx"'
        )
        return response

    @action(detail=True, methods=["get"], name="Findings Assessment as Markdown")
    def md(self, request, pk):
        (viewable_objects, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, FindingsAssessment
        )
        if UUID(pk) not in viewable_objects:
            return Response(
                {"error": "Permission denied"}, status=status.HTTP_403_FORBIDDEN
            )

        findings_assessment = (
            FindingsAssessment.objects.select_related("folder")
            .prefetch_related("owner", "authors", "reviewers")
            .get(id=pk)
        )
        findings = (
            Finding.objects.filter(findings_assessment_id=pk)
            .select_related("folder")
            .prefetch_related("owner", "applied_controls", "evidences")
            .order_by("ref_id")
        )

        # Calculate closed and open findings counts
        closed_statuses = [
            Finding.Status.DISMISSED,
            Finding.Status.MITIGATED,
            Finding.Status.RESOLVED,
            Finding.Status.CLOSED,
            Finding.Status.DEPRECATED,
        ]
        open_statuses = [
            Finding.Status.UNDEFINED,
            Finding.Status.IDENTIFIED,
            Finding.Status.CONFIRMED,
            Finding.Status.ASSIGNED,
            Finding.Status.IN_PROGRESS,
        ]

        closed_findings_count = findings.filter(status__in=closed_statuses).count()
        open_findings_count = findings.filter(status__in=open_statuses).count()

        # Generate Markdown content
        md_content = f"# {findings_assessment.name}\n\n"

        # Assessment metadata
        md_content += "## Assessment Information\n\n"
        md_content += f"- **Name**: {findings_assessment.name}\n"
        md_content += f"- **Reference ID**: {findings_assessment.ref_id or 'N/A'}\n"
        md_content += f"- **Description**: {findings_assessment.description or 'N/A'}\n"
        md_content += f"- **Category**: {findings_assessment.get_category_display()}\n"
        md_content += f"- **Status**: {findings_assessment.get_status_display()}\n"
        md_content += f"- **Folder**: {findings_assessment.folder.name if findings_assessment.folder else 'N/A'}\n"
        if findings_assessment.owner.exists():
            md_content += f"- **Owners**: {', '.join([user.email for user in findings_assessment.owner.all()])}\n"
        if findings_assessment.authors.exists():
            md_content += f"- **Authors**: {', '.join([user.email for user in findings_assessment.authors.all()])}\n"
        if findings_assessment.reviewers.exists():
            md_content += f"- **Reviewers**: {', '.join([user.email for user in findings_assessment.reviewers.all()])}\n"
        md_content += f"- **Created**: {findings_assessment.created_at.strftime('%Y-%m-%d %H:%M:%S') if findings_assessment.created_at else 'N/A'}\n\n"

        # Metrics summary
        metrics = findings_assessment.get_findings_metrics()
        md_content += "## Summary\n\n"
        md_content += "| Metric | Count |\n"
        md_content += "|--------|-------|\n"
        md_content += f"| Total Findings | {metrics['total_count']} |\n"
        md_content += f"| Closed Findings | {closed_findings_count} |\n"
        md_content += f"| Open Findings | {open_findings_count} |\n\n"

        # Severity distribution
        if metrics["severity_distribution"]:
            md_content += "### Severity Distribution\n\n"
            md_content += "| Severity | Count |\n"
            md_content += "|----------|-------|\n"
            for severity, count in metrics["severity_distribution"].items():
                if count > 0:
                    md_content += f"| {severity.capitalize()} | {count} |\n"
            md_content += "\n"

        # Status distribution
        if metrics["status_distribution"]:
            md_content += "### Status Distribution\n\n"
            md_content += "| Status | Count |\n"
            md_content += "|--------|-------|\n"
            status_choices = dict(Finding.Status.choices)
            for status, count in metrics["status_distribution"].items():
                if count > 0:
                    status_display = status_choices.get(
                        status, status.replace("_", " ").title()
                    )
                    md_content += f"| {status_display} | {count} |\n"
            md_content += "\n"

        # Findings details
        md_content += "## Findings\n\n"
        for finding in findings:
            md_content += f"### {finding.ref_id or 'N/A'} - {finding.name}\n\n"
            md_content += f"- **Status**: {finding.get_status_display()}\n"
            md_content += f"- **Severity**: {finding.get_severity_display()}\n"
            md_content += f"- **Description**: {finding.description or 'N/A'}\n"
            md_content += f"- **Observation**: {finding.observation or 'N/A'}\n"
            if finding.owner.exists():
                md_content += f"- **Owner**: {', '.join([user.email for user in finding.owner.all()])}\n"
            if finding.applied_controls.exists():
                md_content += "- **Applied Controls**:\n"
                for ac in finding.applied_controls.all():
                    md_content += f"  - {ac.name} [{ac.get_status_display().lower()}]\n"
            if finding.evidences.exists():
                md_content += f"- **Evidences**: {', '.join([ev.name for ev in finding.evidences.all()])}\n"
            if finding.eta:
                md_content += f"- **ETA**: {finding.eta.strftime('%Y-%m-%d')}\n"
            if finding.due_date:
                md_content += (
                    f"- **Due Date**: {finding.due_date.strftime('%Y-%m-%d')}\n"
                )
            md_content += "\n"

        response = HttpResponse(md_content, content_type="text/markdown")
        safe_name = slugify(findings_assessment.name) or "findings_assessment"
        response["Content-Disposition"] = (
            f'attachment; filename="{safe_name}_findings.md"'
        )
        return response

    @action(detail=True, methods=["get"], name="Findings Assessment as PDF")
    def pdf(self, request, pk):
        (viewable_objects, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, FindingsAssessment
        )
        if UUID(pk) not in viewable_objects:
            return Response(
                {"error": "Permission denied"}, status=status.HTTP_403_FORBIDDEN
            )

        findings_assessment = (
            FindingsAssessment.objects.select_related("folder")
            .prefetch_related("owner", "authors", "reviewers")
            .get(id=pk)
        )
        findings = (
            Finding.objects.filter(findings_assessment_id=pk)
            .select_related("folder")
            .prefetch_related("owner", "applied_controls", "evidences")
            .order_by("ref_id")
        )
        metrics = findings_assessment.get_findings_metrics()

        # Calculate closed and open findings counts
        closed_statuses = [
            Finding.Status.DISMISSED,
            Finding.Status.MITIGATED,
            Finding.Status.RESOLVED,
            Finding.Status.CLOSED,
            Finding.Status.DEPRECATED,
        ]
        open_statuses = [
            Finding.Status.UNDEFINED,
            Finding.Status.IDENTIFIED,
            Finding.Status.CONFIRMED,
            Finding.Status.ASSIGNED,
            Finding.Status.IN_PROGRESS,
        ]

        closed_findings_count = findings.filter(status__in=closed_statuses).count()
        open_findings_count = findings.filter(status__in=open_statuses).count()

        # Process status distribution with display names
        status_choices = dict(Finding.Status.choices)
        processed_status_distribution = []
        for status, count in metrics["status_distribution"].items():
            if count > 0:
                display_name = status_choices.get(
                    status, status.replace("_", " ").title()
                )
                processed_status_distribution.append(
                    {"status": status, "display_name": display_name, "count": count}
                )

        context = {
            "findings_assessment": findings_assessment,
            "findings": findings,
            "metrics": metrics,
            "total_findings": metrics["total_count"],
            "closed_findings_count": closed_findings_count,
            "open_findings_count": open_findings_count,
            "unresolved_important": metrics["unresolved_important_count"],
            "severity_distribution": metrics["severity_distribution"],
            "status_distribution": metrics["status_distribution"],
            "processed_status_distribution": processed_status_distribution,
            "finding_status_choices": dict(Finding.Status.choices),
        }

        html = render_to_string("core/findings_assessment_pdf.html", context)
        pdf_file = HTML(string=html).write_pdf()
        response = HttpResponse(pdf_file, content_type="application/pdf")
        safe_name = slugify(findings_assessment.name) or "findings_assessment"
        response["Content-Disposition"] = (
            f'attachment; filename="{safe_name}_findings.pdf"'
        )
        return response

    @action(detail=False, methods=["get"], name="Get sunburst data")
    def sunburst_data(self, request):
        """
        Returns FindingsAssessment data structured for sunburst visualization:
        Categories (pentest, audit, self-identified) -> Status
        """
        folder_id = request.query_params.get("folder", None)

        # Get viewable findings assessments
        scoped_folder = (
            Folder.objects.get(id=folder_id) if folder_id else Folder.get_root_folder()
        )
        (object_ids, _, _) = RoleAssignment.get_accessible_object_ids(
            scoped_folder, request.user, FindingsAssessment
        )

        findings_assessments = FindingsAssessment.objects.filter(id__in=object_ids)

        # Color mapping for statuses
        status_colors = {
            "planned": "#BFDBFE",
            "in_progress": "#5470c6",
            "in_review": "#BBF7D0",
            "done": "#46D39A",
            "deprecated": "#E55759",
        }

        # Build hierarchical structure: category -> status
        category_data = {}

        for fa in findings_assessments:
            category = fa.category if fa.category else "--"
            status = fa.status if fa.status else "planned"

            if category not in category_data:
                category_data[category] = {}

            if status not in category_data[category]:
                category_data[category][status] = 0

            category_data[category][status] += 1

        # Convert to sunburst format
        sunburst_data = []
        for category, statuses in category_data.items():
            children = []
            for status, count in statuses.items():
                if count > 0:
                    children.append(
                        {
                            "name": status,
                            "value": count,
                            "itemStyle": {"color": status_colors.get(status, "#CCC")},
                        }
                    )

            if children:
                sunburst_data.append({"name": category, "children": children})

        return Response(sunburst_data)


class FindingViewSet(BaseModelViewSet):
    model = Finding
    filterset_fields = [
        "name",
        "owner",
        "folder",
        "status",
        "severity",
        "priority",
        "findings_assessment",
        "filtering_labels",
        "applied_controls",
        "evidences",
    ]
    ordering = ["ref_id"]

    def get_queryset(self) -> models.query.QuerySet:
        return (
            super()
            .get_queryset()
            .select_related("folder", "findings_assessment")
            .prefetch_related("filtering_labels", "applied_controls", "evidences")
        )

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get status choices")
    def status(self, request):
        return Response(dict(Finding.Status.choices))

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get severity choices")
    def severity(self, request):
        return Response(dict(Severity.choices))

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get priority choices")
    def priority(self, request):
        return Response(dict(Finding.PRIORITY))

    @action(detail=False, name="Get all findings owners")
    def owner(self, request):
        return Response(
            UserReadSerializer(
                User.objects.filter(findings__isnull=False).distinct(),
                many=True,
            ).data
        )

    @action(detail=False, name="Get findings sankey data")
    def sankey_data(self, request):
        """
        Returns findings data structured for Sankey diagram:
        Category -> Severity -> Status
        """
        folder_id = request.query_params.get("folder", None)
        (viewable_objects, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, Finding
        )
        queryset = Finding.objects.filter(id__in=viewable_objects).select_related(
            "findings_assessment"
        )

        if folder_id:
            folder = Folder.objects.get(id=folder_id)
            queryset = queryset.filter(folder=folder)

        # Build Sankey data structure

        # Get category -> severity -> status combinations
        nodes = []
        links = []
        node_names = set()

        # Create maps for choices
        category_choice_map = {
            choice[0]: choice[1] for choice in FindingsAssessment.Category.choices
        }
        severity_choice_map = {choice[0]: choice[1] for choice in Severity.choices}
        status_choice_map = {choice[0]: choice[1] for choice in Finding.Status.choices}

        # Track category -> severity links
        category_severity_counts = {}
        # Track severity -> status links
        severity_status_counts = {}

        for finding in queryset:
            # Get category from parent findings assessment
            category_value = (
                finding.findings_assessment.category
                if finding.findings_assessment
                else "--"
            )
            category_label = f"Category: {category_choice_map.get(category_value, 'Unknown').title()}"

            # Get severity
            severity_value = finding.severity if finding.severity else "--"
            severity_label = f"Severity: {severity_choice_map.get(severity_value, 'Unknown').title()}"

            # Get status
            status_value = finding.status if finding.status else "--"
            status_label = (
                f"Status: {status_choice_map.get(status_value, 'Unknown').title()}"
            )

            # Add to node sets
            node_names.add(category_label)
            node_names.add(severity_label)
            node_names.add(status_label)

            # Track links
            cat_sev_key = f"{category_label}||{severity_label}"
            category_severity_counts[cat_sev_key] = (
                category_severity_counts.get(cat_sev_key, 0) + 1
            )

            sev_stat_key = f"{severity_label}||{status_label}"
            severity_status_counts[sev_stat_key] = (
                severity_status_counts.get(sev_stat_key, 0) + 1
            )

        # Convert node names to indexed list
        nodes = [{"name": name} for name in sorted(node_names)]
        node_index = {name: i for i, name in enumerate(sorted(node_names))}

        # Create links for category -> severity
        for key, value in category_severity_counts.items():
            category, severity = key.split("||")
            links.append(
                {
                    "source": node_index[category],
                    "target": node_index[severity],
                    "value": value,
                }
            )

        # Create links for severity -> status
        for key, value in severity_status_counts.items():
            severity, status = key.split("||")
            links.append(
                {
                    "source": node_index[severity],
                    "target": node_index[status],
                    "value": value,
                }
            )

        return Response({"results": {"nodes": nodes, "links": links}})


class IncidentViewSet(ExportMixin, BaseModelViewSet):
    model = Incident
    search_fields = ["name", "description", "ref_id"]
    filterset_fields = [
        "folder",
        "status",
        "severity",
        "qualifications",
        "detection",
        "owners",
        "entities",
        "assets",
    ]

    export_config = {
        "fields": {
            "internal_id": {"source": "id", "label": "internal_id"},
            "ref_id": {"source": "ref_id", "label": "ref_id", "escape": True},
            "name": {"source": "name", "label": "name", "escape": True},
            "description": {
                "source": "description",
                "label": "description",
                "escape": True,
            },
            "status": {"source": "get_status_display", "label": "status"},
            "severity": {"source": "get_severity_display", "label": "severity"},
            "detection": {"source": "get_detection_display", "label": "detection"},
            "reported_at": {
                "source": "reported_at",
                "label": "reported_at",
                "format": lambda dt: dt.replace(tzinfo=None)
                if dt and hasattr(dt, "tzinfo") and dt.tzinfo
                else dt,
            },
            "owners": {
                "source": "owners",
                "label": "owners",
                "format": lambda qs: ",".join(
                    escape_excel_formula(o.email) for o in qs.all()
                ),
            },
            "folder": {"source": "folder.name", "label": "folder", "escape": True},
            "qualifications": {
                "source": "qualifications",
                "label": "qualifications",
                "format": lambda qs: ",".join(
                    escape_excel_formula(q.name) for q in qs.all()
                ),
            },
            "threats": {
                "source": "threats",
                "label": "threats",
                "format": lambda qs: ",".join(
                    escape_excel_formula(t.name) for t in qs.all()
                ),
            },
            "assets": {
                "source": "assets",
                "label": "assets",
                "format": lambda qs: ",".join(
                    escape_excel_formula(a.name) for a in qs.all()
                ),
            },
            "entities": {
                "source": "entities",
                "label": "entities",
                "format": lambda qs: ",".join(
                    escape_excel_formula(e.name) for e in qs.all()
                ),
            },
            "link": {"source": "link", "label": "link", "escape": True},
        },
        "filename": "incidents_export",
        "select_related": ["folder"],
        "prefetch_related": [
            "owners",
            "qualifications",
            "threats",
            "assets",
            "entities",
        ],
    }

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get status choices")
    def status(self, request):
        return Response(dict(Incident.Status.choices))

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get severity choices")
    def severity(self, request):
        return Response(dict(Incident.Severity.choices))

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get detection channel choices")
    def detection(self, request):
        return Response(dict(Incident.Detection.choices))

    def perform_update(self, serializer):
        previous_instance = self.get_object()
        previous_status = previous_instance.status
        previous_severity = previous_instance.severity

        instance = serializer.save()

        if previous_status != instance.status and previous_status is not None:
            TimelineEntry.objects.create(
                incident=instance,
                entry=f"{previous_instance.get_status_display()}->{instance.get_status_display()}",
                entry_type=TimelineEntry.EntryType.STATUS_CHANGED,
                author=self.request.user,
                timestamp=now(),
            )

        if previous_severity != instance.severity and previous_severity is not None:
            TimelineEntry.objects.create(
                incident=instance,
                entry=f"{previous_instance.get_severity_display()}->{instance.get_severity_display()}",
                entry_type=TimelineEntry.EntryType.SEVERITY_CHANGED,
                author=self.request.user,
                timestamp=now(),
            )

        return super().perform_update(serializer)

    @action(detail=True, methods=["get"], name="Incident as PDF")
    def pdf(self, request, pk):
        (viewable_objects, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, Incident
        )
        if UUID(pk) not in viewable_objects:
            return Response(
                {"error": "Permission denied"}, status=status.HTTP_403_FORBIDDEN
            )

        incident = (
            Incident.objects.select_related("folder")
            .prefetch_related("owners", "entities", "assets", "threats")
            .get(id=pk)
        )

        timeline_entries = (
            TimelineEntry.objects.filter(incident_id=pk)
            .select_related("author")
            .prefetch_related("evidences")
            .order_by("timestamp")
        )

        # Count timeline entry types
        detection_count = timeline_entries.filter(
            entry_type=TimelineEntry.EntryType.DETECTION
        ).count()
        mitigation_count = timeline_entries.filter(
            entry_type=TimelineEntry.EntryType.MITIGATION
        ).count()

        context = {
            "incident": incident,
            "timeline_entries": timeline_entries,
            "detection_count": detection_count,
            "mitigation_count": mitigation_count,
        }

        html = render_to_string("core/incident_pdf.html", context)
        pdf_file = HTML(string=html).write_pdf()
        response = HttpResponse(pdf_file, content_type="application/pdf")
        safe_name = slugify(incident.name) or "incident"
        response["Content-Disposition"] = (
            f'attachment; filename="{safe_name}_report.pdf"'
        )
        return response

    @action(detail=True, methods=["get"], name="Incident as Markdown")
    def md(self, request, pk):
        (viewable_objects, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, Incident
        )
        if UUID(pk) not in viewable_objects:
            return Response(
                {"error": "Permission denied"}, status=status.HTTP_403_FORBIDDEN
            )

        incident = (
            Incident.objects.select_related("folder")
            .prefetch_related("owners", "entities", "assets", "threats")
            .get(id=pk)
        )

        timeline_entries = (
            TimelineEntry.objects.filter(incident_id=pk)
            .select_related("author")
            .prefetch_related("evidences")
            .order_by("timestamp")
        )

        # Generate Markdown content
        md_content = f"# {incident.name}\n\n"

        # Incident metadata
        md_content += "## Incident Information\n\n"
        md_content += f"- **Name**: {incident.name}\n"
        md_content += f"- **Reference ID**: {incident.ref_id or 'N/A'}\n"
        md_content += f"- **Description**: {incident.description or 'N/A'}\n"
        md_content += f"- **Status**: {incident.get_status_display()}\n"
        md_content += f"- **Severity**: {incident.get_severity_display()}\n"
        md_content += f"- **Detection**: {incident.get_detection_display() or 'N/A'}\n"
        md_content += (
            f"- **Domain**: {incident.folder.name if incident.folder else 'N/A'}\n"
        )

        if incident.owners.exists():
            md_content += f"- **Owners**: {', '.join([user.email for user in incident.owners.all()])}\n"

        if incident.entities.exists():
            md_content += f"- **Related Entities**: {', '.join([entity.name for entity in incident.entities.all()])}\n"

        md_content += f"- **Created**: {incident.created_at.strftime('%Y-%m-%d %H:%M:%S') if incident.created_at else 'N/A'}\n"
        md_content += f"- **Last Updated**: {incident.updated_at.strftime('%Y-%m-%d %H:%M:%S') if incident.updated_at else 'N/A'}\n\n"

        # Affected Assets
        if incident.assets.exists():
            md_content += "## Affected Assets\n\n"
            md_content += "| Name | Type |\n"
            md_content += "|------|------|\n"
            for asset in incident.assets.all():
                md_content += f"| {asset.name} | {asset.get_type_display()} |\n"
            md_content += "\n"

        # Related Threats
        if incident.threats.exists():
            md_content += "## Related Threats\n\n"
            md_content += "| Name | Reference ID |\n"
            md_content += "|------|------|\n"
            for threat in incident.threats.all():
                md_content += f"| {threat.name} | {threat.ref_id or 'N/A'} |\n"
            md_content += "\n"

        # Timeline
        if timeline_entries:
            md_content += "## Timeline\n\n"
            md_content += "*Events are listed in chronological order*\n\n"

            for entry in timeline_entries:
                md_content += f"### {entry.timestamp.strftime('%Y-%m-%d %H:%M:%S')} - {entry.entry}\n\n"
                md_content += f"**Type**: {entry.get_entry_type_display()}\n\n"

                if entry.author:
                    md_content += f"**Author**: {entry.author.email}\n\n"

                if entry.observation:
                    md_content += f"**Observation**: {entry.observation}\n\n"

                if entry.evidences.exists():
                    md_content += "**Associated Evidence**:\n"
                    for evidence in entry.evidences.all():
                        md_content += f"- {evidence.name}\n"
                    md_content += "\n"

                md_content += "---\n\n"
        else:
            md_content += "## Timeline\n\n"
            md_content += "*No timeline events found for this incident.*\n\n"

        response = HttpResponse(md_content, content_type="text/markdown")
        safe_name = slugify(incident.name) or "incident"
        response["Content-Disposition"] = (
            f'attachment; filename="{safe_name}_report.md"'
        )
        return response

    @action(detail=False, name="Get incident detection breakdown")
    def detection_breakdown(self, request):
        folder_id = request.query_params.get("folder", None)
        (viewable_objects, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, Incident
        )
        queryset = Incident.objects.filter(id__in=viewable_objects)

        if folder_id:
            folder = Folder.objects.get(id=folder_id)
            queryset = queryset.filter(folder=folder)

        detection_stats = []
        for detection_choice in Incident.Detection.choices:
            count = queryset.filter(detection=detection_choice[0]).count()
            detection_stats.append(
                {
                    "name": detection_choice[1],
                    "value": count,
                    "itemStyle": {
                        "color": "#3B82F6"
                        if detection_choice[0] == "internally_detected"
                        else "#EF4444"
                    },
                }
            )

        return Response({"results": detection_stats})

    @action(detail=False, name="Get monthly incident metrics")
    def monthly_metrics(self, request):
        folder_id = request.query_params.get("folder", None)
        (viewable_objects, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, Incident
        )
        queryset = Incident.objects.filter(id__in=viewable_objects)

        if folder_id:
            folder = Folder.objects.get(id=folder_id)
            queryset = queryset.filter(folder=folder)

        # Get incidents with reported_at dates
        incidents_with_dates = queryset.filter(reported_at__isnull=False).order_by(
            "reported_at"
        )

        if not incidents_with_dates.exists():
            return Response(
                {
                    "results": {
                        "months": [],
                        "monthly_counts": [],
                        "cumulative_counts": [],
                    }
                }
            )

        # Group by month
        from collections import defaultdict
        from datetime import datetime

        monthly_counts = defaultdict(int)

        for incident in incidents_with_dates:
            month_key = incident.reported_at.strftime("%Y-%m")
            monthly_counts[month_key] += 1

        # Sort months and calculate cumulative
        sorted_months = sorted(monthly_counts.keys())
        cumulative_count = 0
        cumulative_counts = []

        for month in sorted_months:
            cumulative_count += monthly_counts[month]
            cumulative_counts.append(cumulative_count)

        # Format months for display
        formatted_months = []
        for month in sorted_months:
            date_obj = datetime.strptime(month, "%Y-%m")
            formatted_months.append(date_obj.strftime("%b %Y"))

        return Response(
            {
                "results": {
                    "months": formatted_months,
                    "monthly_counts": [
                        monthly_counts[month] for month in sorted_months
                    ],
                    "cumulative_counts": cumulative_counts,
                }
            }
        )

    @action(detail=False, name="Get incident summary statistics")
    def summary_stats(self, request):
        folder_id = request.query_params.get("folder", None)
        (viewable_objects, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, Incident
        )
        queryset = Incident.objects.filter(id__in=viewable_objects)

        if folder_id:
            folder = Folder.objects.get(id=folder_id)
            queryset = queryset.filter(folder=folder)

        # Total incidents
        total_incidents = queryset.count()

        # Incidents this month
        from datetime import date
        import calendar

        today = date.today()
        first_day = today.replace(day=1)
        last_day = today.replace(day=calendar.monthrange(today.year, today.month)[1])

        incidents_this_month = queryset.filter(
            reported_at__date__gte=first_day, reported_at__date__lte=last_day
        ).count()

        # Currently open incidents (not closed or dismissed)
        open_incidents = queryset.exclude(
            status__in=[Incident.Status.CLOSED, Incident.Status.DISMISSED]
        ).count()

        return Response(
            {
                "results": {
                    "total_incidents": total_incidents,
                    "incidents_this_month": incidents_this_month,
                    "open_incidents": open_incidents,
                }
            }
        )

    @action(detail=False, name="Get incident severity breakdown")
    def severity_breakdown(self, request):
        folder_id = request.query_params.get("folder", None)
        (viewable_objects, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, Incident
        )
        queryset = Incident.objects.filter(id__in=viewable_objects)

        if folder_id:
            folder = Folder.objects.get(id=folder_id)
            queryset = queryset.filter(folder=folder)

        # Define severity colors
        severity_colors = {
            1: "#DC2626",  # Critical - Red
            2: "#EA580C",  # Major - Orange
            3: "#D97706",  # Moderate - Amber
            4: "#65A30D",  # Minor - Lime
            5: "#16A34A",  # Low - Green
            6: "#6B7280",  # Unknown - Gray
        }

        severity_stats = []
        for severity_choice in Incident.Severity.choices:
            count = queryset.filter(severity=severity_choice[0]).count()
            severity_stats.append(
                {
                    "name": severity_choice[1],
                    "value": count,
                    "itemStyle": {
                        "color": severity_colors.get(severity_choice[0], "#6B7280")
                    },
                }
            )

        return Response({"results": severity_stats})

    @action(detail=False, name="Get incident qualifications breakdown")
    def qualifications_breakdown(self, request):
        folder_id = request.query_params.get("folder", None)
        (viewable_objects, _, _) = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, Incident
        )
        queryset = Incident.objects.filter(id__in=viewable_objects)

        if folder_id:
            folder = Folder.objects.get(id=folder_id)
            queryset = queryset.filter(folder=folder)

        # Get all unique qualifications used in incidents
        from django.db.models import Count

        qualifications_stats = []

        # Get qualification counts
        qualification_counts = (
            queryset.values("qualifications__name")
            .annotate(count=Count("id", distinct=True))
            .filter(qualifications__name__isnull=False)
            .order_by("-count")
        )

        # Format for radar chart
        labels = []
        values = []
        for item in qualification_counts:
            if item["qualifications__name"]:
                values.append(item["count"])

        # Create labels with proper format for radar chart
        max_offset = max(values, default=0)
        for item in qualification_counts:
            if item["qualifications__name"]:
                labels.append({"name": item["qualifications__name"], "max": max_offset})

        return Response({"results": {"labels": labels, "values": values}})


class TimelineEntryViewSet(BaseModelViewSet):
    model = TimelineEntry
    filterset_fields = ["incident"]
    search_fields = ["entry", "entry_type"]
    ordering = ["-timestamp"]

    def get_queryset(self):
        return (
            super()
            .get_queryset()
            .select_related("folder", "incident", "author")
            .prefetch_related("evidences")
        )

    @action(detail=False, name="Get entry type choices")
    def entry_type(self, request):
        return Response(dict(TimelineEntry.EntryType.get_manual_entry_types()))

    def perform_create(self, serializer):
        instance = serializer.save(author=self.request.user)
        dispatch_webhook_event(instance, "created", serializer)
        return instance

    def perform_destroy(self, instance):
        if instance.entry_type in [
            TimelineEntry.EntryType.SEVERITY_CHANGED,
            TimelineEntry.EntryType.STATUS_CHANGED,
        ]:
            raise ValidationError({"error": "cannotDeleteAutoTimelineEntry"})
        return super().perform_destroy(instance)


class TaskTemplateFilter(GenericFilterSet):
    folder = df.ModelMultipleChoiceFilter(queryset=Folder.objects.all())

    next_occurrence_status = df.MultipleChoiceFilter(
        choices=TaskNode.TASK_STATUS_CHOICES, method="filter_next_occurrence_status"
    )
    last_occurrence_status = df.MultipleChoiceFilter(
        choices=TaskNode.TASK_STATUS_CHOICES, method="filter_last_occurrence_status"
    )

    class Meta:
        model = TaskTemplate
        fields = [
            "name",
            "assigned_to",
            "is_recurrent",
            "applied_controls",
            "last_occurrence_status",
            "next_occurrence_status",
            "evidences",
        ]

    def filter_last_occurrence_status(self, queryset, name, values):
        start = timezone.now().date()
        status_subquery = (
            TaskNode.objects.filter(task_template=OuterRef("pk"), due_date__lt=start)
            .order_by("-due_date")
            .values("status")[:1]
        )

        return queryset.annotate(last_status=Subquery(status_subquery)).filter(
            last_status__in=values
        )

    def filter_next_occurrence_status(self, queryset, name, values):
        start = timezone.now().date()
        status_subquery = (
            TaskNode.objects.filter(task_template=OuterRef("pk"), due_date__gte=start)
            .order_by("due_date")
            .values("status")[:1]
        )

        return queryset.annotate(next_status=Subquery(status_subquery)).filter(
            next_status__in=values
        )


class TaskTemplateViewSet(ExportMixin, BaseModelViewSet):
    model = TaskTemplate
    filterset_fields = [
        "assigned_to",
        "is_recurrent",
        "folder",
        "applied_controls",
        "evidences",
    ]
    search_fields = ["ref_id", "name"]
    filterset_class = TaskTemplateFilter

    export_config = {
        "filename": "task_templates",
        "select_related": ["folder"],
        "prefetch_related": [
            "assigned_to",
            "evidences",
            "applied_controls",
            "compliance_assessments",
            "risk_assessments",
            "assets",
            "findings_assessment",
        ],
        "fields": {
            "ref_id": {"source": "ref_id", "label": "ref_id", "escape": True},
            "name": {"source": "name", "label": "name", "escape": True},
            "description": {
                "source": "description",
                "label": "description",
                "escape": True,
            },
            "folder": {"source": "folder.name", "label": "folder", "escape": True},
            "is_recurrent": {
                "source": "is_recurrent",
                "label": "is_recurrent",
                "format": lambda x: "Yes" if x else "No",
            },
            "task_date": {
                "source": "task_date",
                "label": "task_date",
                "format": lambda x: x.strftime("%Y-%m-%d") if x else "",
            },
            "schedule_frequency": {
                "source": "schedule",
                "label": "schedule_frequency",
                "format": lambda s: s.get("frequency", "") if s else "",
            },
            "schedule_interval": {
                "source": "schedule",
                "label": "schedule_interval",
                "format": lambda s: str(s.get("interval", "")) if s else "",
            },
            "schedule_end_date": {
                "source": "schedule",
                "label": "schedule_end_date",
                "format": lambda s: s.get("end_date", "") if s else "",
            },
            "next_occurrence": {
                "source": "get_next_occurrence",
                "label": "next_occurrence",
                "format": lambda x: x.strftime("%Y-%m-%d") if x else "",
            },
            "next_occurrence_status": {
                "source": "get_next_occurrence_status",
                "label": "next_occurrence_status",
            },
            "enabled": {
                "source": "enabled",
                "label": "enabled",
                "format": lambda x: "Yes" if x else "No",
            },
            "assigned_to": {
                "source": "assigned_to.all",
                "label": "assigned_to",
                "format": lambda users: ", ".join([u.email for u in users])
                if users
                else "",
            },
            "assets": {
                "source": "assets.all",
                "label": "assets",
                "format": lambda items: ", ".join(
                    escape_excel_formula(str(item)) for item in items
                )
                if items
                else "",
            },
            "applied_controls": {
                "source": "applied_controls.all",
                "label": "applied_controls",
                "format": lambda items: ", ".join(
                    escape_excel_formula(str(item)) for item in items
                )
                if items
                else "",
            },
            "evidences": {
                "source": "evidences.all",
                "label": "evidences",
                "format": lambda items: ", ".join(
                    escape_excel_formula(item.name) for item in items
                )
                if items
                else "",
            },
            "compliance_assessments": {
                "source": "compliance_assessments.all",
                "label": "compliance_assessments",
                "format": lambda items: ", ".join(
                    escape_excel_formula(str(item)) for item in items
                )
                if items
                else "",
            },
            "risk_assessments": {
                "source": "risk_assessments.all",
                "label": "risk_assessments",
                "format": lambda items: ", ".join(
                    escape_excel_formula(str(item)) for item in items
                )
                if items
                else "",
            },
            "findings_assessment": {
                "source": "findings_assessment.all",
                "label": "findings_assessment",
                "format": lambda items: ", ".join(
                    escape_excel_formula(str(item)) for item in items
                )
                if items
                else "",
            },
            "link": {"source": "link", "label": "link", "escape": True},
        },
        "wrap_columns": ["name", "description"],
        "task_node_fields": {
            "due_date": {
                "source": "due_date",
                "label": "due_date",
                "format": lambda x: x.strftime("%Y-%m-%d") if x else "",
            },
            "status": {"source": "status", "label": "status"},
            "observation": {
                "source": "observation",
                "label": "observation",
                "escape": True,
            },
            "assigned_to": {
                "source": "task_template.assigned_to.all",
                "label": "assigned_to",
                "format": lambda users: ", ".join([u.email for u in users])
                if users
                else "",
            },
            "expected_evidence": {
                "source": "task_template.evidences.all",
                "label": "expected_evidence",
                # Note: Formatting with done/pending status is handled specially in export_tasks_xlsx
                "format": lambda items: ", ".join([item.name for item in items])
                if items
                else "",
            },
            # "evidences": {
            #     "source": "evidences.all",
            #     "label": "evidences",
            #     "format": lambda items: ", ".join([item.name for item in items])
            #     if items
            #     else "",
            # },
            "assets": {
                "source": "task_template.assets.all",
                "label": "assets",
                "format": lambda items: ", ".join(
                    escape_excel_formula(str(item)) for item in items
                )
                if items
                else "",
            },
            "applied_controls": {
                "source": "task_template.applied_controls.all",
                "label": "applied_controls",
                "format": lambda items: ", ".join(
                    escape_excel_formula(str(item)) for item in items
                )
                if items
                else "",
            },
            "compliance_assessments": {
                "source": "task_template.compliance_assessments.all",
                "label": "compliance_assessments",
                "format": lambda items: ", ".join(
                    escape_excel_formula(str(item)) for item in items
                )
                if items
                else "",
            },
            "risk_assessments": {
                "source": "task_template.risk_assessments.all",
                "label": "risk_assessments",
                "format": lambda items: ", ".join(
                    escape_excel_formula(str(item)) for item in items
                )
                if items
                else "",
            },
            "findings_assessment": {
                "source": "task_template.findings_assessment.all",
                "label": "findings_assessment",
                "format": lambda items: ", ".join(
                    escape_excel_formula(str(item)) for item in items
                )
                if items
                else "",
            },
        },
    }

    def get_queryset(self):
        qs = super().get_queryset()
        ordering = self.request.query_params.get("ordering", "")

        if any(
            f in ordering
            for f in (
                "next_occurrence",
                "last_occurrence_status",
                "next_occurrence_status",
            )
        ):
            today = timezone.localdate()
            qs = qs.annotate(
                next_occurrence=Case(
                    When(is_recurrent=False, then=Min("tasknode__due_date")),
                    When(
                        is_recurrent=True,
                        then=Min(
                            "tasknode__due_date",
                            filter=Q(tasknode__due_date__gte=today),
                        ),
                    ),
                    default=Value(None),
                    output_field=models.DateField(),
                ),
                last_occurrence_status=Subquery(
                    TaskNode.objects.filter(
                        task_template=OuterRef("pk"), due_date__lt=today
                    )
                    .order_by("-due_date")
                    .values("status")[:1]
                ),
                next_occurrence_status=Case(
                    When(
                        is_recurrent=False,
                        then=Subquery(
                            TaskNode.objects.filter(task_template=OuterRef("pk"))
                            .order_by("due_date")
                            .values("status")[:1]
                        ),
                    ),
                    When(
                        is_recurrent=True,
                        then=Subquery(
                            TaskNode.objects.filter(
                                task_template=OuterRef("pk"),
                                due_date__gte=today,
                            )
                            .order_by("due_date")
                            .values("status")[:1]
                        ),
                    ),
                    default=Value(None),
                    output_field=models.CharField(),
                ),
            )

        return qs

    def task_calendar(self, task_templates, start=None, end=None):
        """Generate calendar of tasks for the given templates."""
        tasks_list = []
        for template in task_templates:
            if not template.is_recurrent:
                tasks_list.append(_create_task_dict(template, template.task_date))
                continue

            start_date_param = start or template.task_date or datetime.now().date()
            end_date_param = end or template.schedule.get("end_date")

            if not end_date_param:
                start_date = datetime.strptime(str(start_date_param), "%Y-%m-%d").date()
                end_date_param = (start_date + rd.relativedelta(months=1)).strftime(
                    "%Y-%m-%d"
                )

            try:
                start_date = datetime.strptime(str(start_date_param), "%Y-%m-%d").date()
                end_date = datetime.strptime(str(end_date_param), "%Y-%m-%d").date()
            except ValueError:
                return {"error": "Invalid date format. Use YYYY-MM-DD"}

            tasks = _generate_occurrences(template, start_date, end_date)
            tasks_list.extend(tasks)

        processed_tasks_identifiers = set()  # Track tasks we've already processed

        # Sort tasks by due date
        sorted_tasks = sorted(
            [t for t in tasks_list if t.get("due_date")], key=lambda x: x["due_date"]
        )

        # Process all past tasks and next 10 upcoming tasks
        current_date = datetime.now().date()

        # First separate past and future tasks
        past_tasks = [task for task in sorted_tasks if task["due_date"] <= current_date]
        next_tasks = [task for task in sorted_tasks if task["due_date"] > current_date]

        # Combined list of tasks to process (past and next 10)
        tasks_to_process = past_tasks + next_tasks

        # Directly modify tasks in the original tasks_list
        for i in range(len(tasks_list)):
            task = tasks_list[i]
            task_date = task["due_date"]
            task_template_id = task["task_template"]

            # Create a unique identifier for this task to avoid duplication
            task_identifier = (task_template_id, task_date)

            # Skip if we've already processed this task
            if task_identifier in processed_tasks_identifiers:
                continue

            # Check if this task should be processed (is in past or next 10)
            if task in tasks_to_process:
                processed_tasks_identifiers.add(task_identifier)

                # Get or create the TaskNode
                task_template = TaskTemplate.objects.get(id=task_template_id)
                task_node, created = TaskNode.objects.get_or_create(
                    task_template=task_template,
                    due_date=task_date,
                    defaults={
                        "status": "pending",
                        "folder": task_template.folder,
                    },
                )
                task_node.to_delete = False
                task_node.save(update_fields=["to_delete"])
                # Replace the task dictionary with the actual TaskNode in the original list
                tasks_list[i] = TaskNodeReadSerializer(task_node).data

        return tasks_list

    def _sync_task_nodes(self, task_template: TaskTemplate):
        if task_template.is_recurrent:
            with transaction.atomic():
                # Soft-delete all existing TaskNode instances associated with this TaskTemplate
                TaskNode.objects.filter(
                    task_template=task_template, due_date__gt=date.today()
                ).update(to_delete=True)
                # Determine the end date based on the frequency
                start_date = task_template.task_date
                if task_template.is_recurrent:
                    if task_template.schedule["frequency"] == "DAILY":
                        delta = rd.relativedelta(months=2)
                    elif task_template.schedule["frequency"] == "WEEKLY":
                        delta = rd.relativedelta(months=4)
                    elif task_template.schedule["frequency"] == "MONTHLY":
                        delta = rd.relativedelta(years=1)
                    elif task_template.schedule["frequency"] == "YEARLY":
                        delta = rd.relativedelta(years=5)

                    end_date_param = task_template.schedule.get("end_date")
                    if end_date_param:
                        end_date = datetime.strptime(end_date_param, "%Y-%m-%d").date()
                    else:
                        end_date = datetime.now().date() + delta
                    # Ensure end_date is not before the calculated delta
                    if end_date < datetime.now().date() + delta:
                        end_date = datetime.now().date() + delta
                else:
                    end_date = start_date
                # Generate the task nodes
                self.task_calendar(
                    task_templates=TaskTemplate.objects.filter(id=task_template.id),
                    start=start_date,
                    end=end_date,
                )

                # garbage-collect
                TaskNode.objects.filter(to_delete=True).delete()

    @action(
        detail=False,
        name="Get tasks for the calendar",
        url_path="calendar/(?P<start>.+)/(?P<end>.+)",
    )
    def calendar(
        self,
        request,
        start=None,
        end=None,
    ):
        if start is None:
            start = timezone.now().date()
        if end is None:
            end = timezone.now().date() + relativedelta.relativedelta(months=1)
        return Response(
            self.task_calendar(
                task_templates=TaskTemplate.objects.filter(enabled=True),
                start=start,
                end=end,
            )
        )

    def perform_update(self, serializer):
        task_template = serializer.save()
        self._sync_task_nodes(task_template)

    @action(detail=False, name="Export Tasks with Nodes as Multi-Sheet XLSX")
    def export_xlsx(self, request):
        """
        Export task templates with a summary sheet and individual sheets for each template's task nodes.
        """
        if not self.export_config:
            return HttpResponse(
                status=501, content="Export not configured for this model"
            )

        queryset = self._get_export_queryset()

        wb = Workbook()
        wb.remove(wb.active)

        summary_ws = wb.create_sheet(title="Summary")
        main_fields = self.export_config["fields"]
        main_headers = [f.get("label", name) for name, f in main_fields.items()]
        summary_ws.append(main_headers)

        wrap_alignment = Alignment(wrap_text=True, vertical="top")
        wrap_columns = self.export_config.get("wrap_columns", ["name", "description"])

        used_sheet_names = {"Summary"}
        task_node_fields = self.export_config.get("task_node_fields", {})

        template_counter = 0
        for obj in queryset:
            row = [
                self._resolve_field_value(obj, field_config)
                for field_config in main_fields.values()
            ]
            summary_ws.append(row)

            # Only include past task nodes due_date <= today
            today = timezone.localdate()

            task_nodes = (
                TaskNode.objects.filter(task_template=obj, due_date__lte=today)
                .select_related("task_template", "task_template__folder")
                .prefetch_related(
                    "evidences",
                    "task_template__assigned_to",
                    "task_template__evidences__revisions",  # Prefetch revisions for expected evidences
                    "task_template__assets",
                    "task_template__applied_controls",
                    "task_template__compliance_assessments",
                    "task_template__risk_assessments",
                    "task_template__findings_assessment",
                )
                .order_by("due_date")
            )

            if not task_nodes.exists():
                continue

            template_counter += 1
            task_name = obj.name or f"Template_{obj.pk}"
            # Format: "1-task_name", "2-task_name", etc.
            base_name = f"{template_counter}-{task_name}"

            # Excel sheet names have a 31 character limit
            sheet_name = base_name[:31]
            counter = 1
            while sheet_name in used_sheet_names:
                suffix = f" ({counter})"
                sheet_name = base_name[: 31 - len(suffix)] + suffix
                counter += 1
            used_sheet_names.add(sheet_name)

            detail_ws = wb.create_sheet(title=sheet_name)
            detail_headers = [
                f.get("label", name) for name, f in task_node_fields.items()
            ]
            detail_ws.append(detail_headers)

            for task_node in task_nodes:
                detail_row = []
                for field_name, field_config in task_node_fields.items():
                    if field_name == "expected_evidence":
                        # Special handling for expected evidence with done/pending status
                        evidences = task_node.task_template.evidences.all()
                        evidence_statuses = []
                        for evidence in evidences:
                            last_revision = evidence.last_revision
                            if last_revision and last_revision.task_node == task_node:
                                status = "done"
                            else:
                                status = "pending"
                            evidence_statuses.append(f"{evidence.name} ({status})")
                        value = (
                            ", ".join(evidence_statuses) if evidence_statuses else ""
                        )
                    else:
                        value = self._resolve_field_value(task_node, field_config)
                    detail_row.append(value)
                detail_ws.append(detail_row)

            for col_idx, header in enumerate(detail_headers, 1):
                if header.lower() in wrap_columns:
                    for row_idx in range(1, detail_ws.max_row + 1):
                        detail_ws.cell(
                            row=row_idx, column=col_idx
                        ).alignment = wrap_alignment

        for col_idx, header in enumerate(main_headers, 1):
            if header.lower() in wrap_columns:
                for row_idx in range(1, summary_ws.max_row + 1):
                    summary_ws.cell(
                        row=row_idx, column=col_idx
                    ).alignment = wrap_alignment

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"{self.export_config.get('filename', 'export')}_multi_sheet.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def perform_create(self, serializer):
        super().perform_create(serializer)
        self._sync_task_nodes(serializer.instance)

    @action(detail=True, name="Get write data")
    def object(self, request, pk):
        serializer_class = self.get_serializer_class(action="update")
        self._sync_task_nodes(
            self.get_object()
        )  # Synchronize task nodes when fetching a task template
        return Response(serializer_class(super().get_object()).data)

    @action(detail=False, name="Get all task template assigned_to users")
    def assigned_to(self, request):
        return Response(
            UserReadSerializer(
                User.objects.filter(task_templates__isnull=False).distinct(),
                many=True,
            ).data
        )

    @action(detail=False, name="Yearly tasks review")
    def yearly_review(self, request):
        """Get recurrent task templates grouped by folder for yearly review."""
        # Get date range from query params or use current year
        current_year = timezone.now().year

        try:
            start_month = int(request.query_params.get("start_month", 1))
            start_year = int(request.query_params.get("start_year", current_year))
            end_month = int(request.query_params.get("end_month", 12))
            end_year = int(request.query_params.get("end_year", current_year))
        except ValueError:
            # Default to current year if any parsing fails
            start_month, start_year = 1, current_year
            end_month, end_year = 12, current_year

        # Validate month values
        start_month = max(1, min(12, start_month))
        end_month = max(1, min(12, end_month))

        year_start = date(start_year, start_month, 1)
        # Get last day of end_month
        if end_month == 12:
            year_end = date(end_year, 12, 31)
        else:
            year_end = date(end_year, end_month + 1, 1) - timedelta(days=1)

        # Get folder filter from query params
        folder_id = request.query_params.get("folder")

        # Get all viewable recurrent task templates
        queryset = self.get_queryset().filter(
            is_recurrent=True,
            enabled=True,
        )

        # Apply folder filter if provided
        if folder_id:
            queryset = queryset.filter(folder_id=folder_id)

        task_templates = queryset.select_related("folder").prefetch_related(
            "assigned_to"
        )

        # Group by folder
        folders_dict = defaultdict(list)
        for template in task_templates:
            folder_id = str(template.folder.id)
            folder_name = str(template.folder)

            # Use TaskTemplateReadSerializer for proper serialization
            template_data = TaskTemplateReadSerializer(template).data
            # Add schedule field (excluded by default in read serializer)
            template_data["schedule"] = template.schedule

            # Get TaskNodes for this template in the current year
            task_nodes = TaskNode.objects.filter(
                task_template=template, due_date__gte=year_start, due_date__lte=year_end
            ).values("due_date", "status")

            # Group by month and determine status
            monthly_status = {}
            # Generate list of (year, month) tuples for the date range
            months_in_range = []
            current = date(start_year, start_month, 1)
            end = date(end_year, end_month, 1)
            while current <= end:
                months_in_range.append((current.year, current.month))
                if current.month == 12:
                    current = date(current.year + 1, 1, 1)
                else:
                    current = date(current.year, current.month + 1, 1)

            for year, month in months_in_range:
                # Create a unique key for year-month combination
                key = f"{year}-{month:02d}"
                month_nodes = [
                    node
                    for node in task_nodes
                    if node["due_date"].year == year and node["due_date"].month == month
                ]

                if not month_nodes:
                    monthly_status[key] = None
                else:
                    # Simple aggregation logic
                    statuses = [node["status"] for node in month_nodes]
                    # If all completed, show completed (green)
                    if all(s == "completed" for s in statuses):
                        monthly_status[key] = "completed"
                    # If any in_progress or mix of statuses, show in_progress (orange)
                    elif "in_progress" in statuses or "completed" in statuses:
                        monthly_status[key] = "in_progress"
                    # If all pending, show pending (red)
                    elif "pending" in statuses:
                        monthly_status[key] = "pending"
                    else:
                        monthly_status[key] = statuses[0] if statuses else None

            template_data["monthly_status"] = monthly_status

            if folder_id not in folders_dict:
                folders_dict[folder_id] = {
                    "folder_id": folder_id,
                    "folder_name": folder_name,
                    "tasks": [],
                }
            folders_dict[folder_id]["tasks"].append(template_data)

        # Convert to list and sort by folder name
        result = list(folders_dict.values())
        result.sort(key=lambda x: x["folder_name"])

        # Include date range metadata in response
        return Response(
            {
                "folders": result,
                "start_month": start_month,
                "start_year": start_year,
                "end_month": end_month,
                "end_year": end_year,
            }
        )

    @action(detail=False, name="Task templates per status")
    def per_status(self, request):
        data = task_template_per_status(request.user)
        return Response({"results": data})

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get Task Node status choices")
    def status(srlf, request):
        return Response(dict(TaskNode.TASK_STATUS_CHOICES))


class TaskNodeFilterSet(GenericFilterSet):
    past = df.BooleanFilter(method="filter_past")

    def filter_past(self, queryset, name, value):
        if value is True:
            return queryset.filter(due_date__lt=date.today()).order_by("due_date")
        elif value is False:
            return queryset.filter(due_date__gte=date.today()).order_by("due_date")
        return queryset

    class Meta:
        model = TaskNode
        fields = {
            "status": ["exact"],
            "task_template": ["exact"],
            "due_date": ["exact", "lte", "gte", "lt", "gt", "month", "year"],
        }


class TaskNodeViewSet(BaseModelViewSet):
    model = TaskNode
    filterset_class = TaskNodeFilterSet
    search_fields = ["observation"]
    ordering = ["due_date"]

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get Task Node status choices")
    def status(srlf, request):
        return Response(dict(TaskNode.TASK_STATUS_CHOICES))

    @action(
        detail=True, name="Remove/Move evidence to expected evidence", methods=["post"]
    )
    def remove_evidence(self, request, pk):
        task_node = TaskNode.objects.get(id=pk)
        evidence_id = request.data.get("evidence_id")
        to_move = request.data.get("move", False)
        evidence = Evidence.objects.get(id=evidence_id)
        task_node.evidences.remove(evidence)
        if to_move:
            task_node.task_template.evidences.add(evidence)
        return Response(status=status.HTTP_200_OK)

    def perform_create(self, serializer):
        instance: TaskNode = serializer.save()
        instance.save()
        return super().perform_create(serializer)


class TaskNodeEvidenceList(generics.ListAPIView):
    serializer_class = EvidenceReadSerializer
    filterset_fields = {
        "folder": ["exact"],
        "status": ["exact"],
        "owner": ["exact"],
        "name": ["icontains"],
        "expiry_date": ["exact", "lte", "gte"],
        "created_at": ["exact", "lte", "gte"],
        "updated_at": ["exact", "lte", "gte"],
    }
    search_fields = ["name", "description"]
    ordering_fields = ["name", "status", "updated_at", "expiry_date"]
    ordering = ["name"]

    filter_backends = [
        DjangoFilterBackend,
        filters.SearchFilter,
        filters.OrderingFilter,
    ]

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context.update({"pk": self.kwargs["pk"]})
        return context

    def get_queryset(self):
        """RBAC not automatic as we don't inherit from BaseModelViewSet -> enforce it explicitly"""
        task_node_id = self.kwargs["pk"]

        if not RoleAssignment.is_object_readable(
            self.request.user,
            TaskNode,
            task_node_id,
        ):
            raise PermissionDenied()

        task_node = TaskNode.objects.get(id=task_node_id)

        # Get task node's template
        task_template = task_node.task_template

        # Get visible evidences to filter result
        viewable_evidences, _, _ = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), self.request.user, Evidence
        )
        return Evidence.objects.filter(
            id__in=task_template.evidences.filter(
                id__in=viewable_evidences
            ).values_list("id", flat=True)
        )


class TerminologyViewSet(BaseModelViewSet):
    model = Terminology
    filterset_fields = ["field_path", "folder", "is_visible", "builtin"]
    search_fields = ["name", "description"]
    ordering = ["field_path", "name"]

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get class name choices")
    def field_path(self, request):
        return Response(dict(Terminology.FieldPath.choices))
