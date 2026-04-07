"""
Regression tests for custom library import (YAML and Excel).

These tests verify the end-to-end flow of uploading custom libraries
through the API, covering both YAML and Excel formats.

Test coverage:
- YAML: framework, risk matrix, threats, reference controls
- Excel: v2 format with library_meta + object sheets
- Dry-run mode for both formats
- Error handling for invalid files
- Library re-upload / duplicate detection
"""

import io
import json
from unittest.mock import patch, MagicMock

import openpyxl
import pytest
from django.urls import reverse
from knox.models import AuthToken
from rest_framework import status
from rest_framework.test import APIClient

from core.apps import startup
from core.models import (
    Framework,
    LoadedLibrary,
    ReferenceControl,
    RequirementNode,
    RiskMatrix,
    StoredLibrary,
    Threat,
)
from iam.models import Folder, User, UserGroup

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _upload_yaml(client, url, filename, content: bytes):
    """Upload a YAML file using the FileUploadParser-compatible format.

    The StoredLibraryViewSet uses FileUploadParser, which expects the raw
    file content in the request body with a Content-Disposition header.
    """
    return client.post(
        url,
        data=content,
        content_type="application/yaml",
        HTTP_CONTENT_DISPOSITION=f"attachment; filename={filename}",
    )


def _upload_excel(client, url, filename, content: bytes):
    """Upload an Excel file using the FileUploadParser-compatible format."""
    return client.post(
        url,
        data=content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        HTTP_CONTENT_DISPOSITION=f"attachment; filename={filename}",
    )


# Sample YAML libraries

SAMPLE_FRAMEWORK_YAML = """
urn: urn:intuitem:test:library:sample-framework-regression
locale: en
ref_id: SAMPLE-FW-REG
name: Sample Framework for Regression Test
description: A minimal framework library used for regression testing
copyright: Test
version: 1
publication_date: 2025-01-01
provider: test-provider
packager: test-packager
objects:
  framework:
    urn: urn:intuitem:test:framework:sample-framework-regression
    ref_id: SAMPLE-FW-REG
    name: Sample Framework for Regression Test
    description: A sample framework
    requirement_nodes:
    - urn: urn:intuitem:test:req_node:sample-fw-reg:cat-1
      assessable: false
      depth: 1
      ref_id: CAT-1
      name: Category 1
      description: A sample category
    - urn: urn:intuitem:test:req_node:sample-fw-reg:cat-1.1
      assessable: true
      depth: 2
      ref_id: CAT-1.1
      parent_urn: urn:intuitem:test:req_node:sample-fw-reg:cat-1
      name: Requirement 1.1
      description: A sample assessable requirement
    - urn: urn:intuitem:test:req_node:sample-fw-reg:cat-1.2
      assessable: true
      depth: 2
      ref_id: CAT-1.2
      parent_urn: urn:intuitem:test:req_node:sample-fw-reg:cat-1
      name: Requirement 1.2
      description: Another sample assessable requirement
""".lstrip()

SAMPLE_RISK_MATRIX_YAML = """
urn: urn:intuitem:test:library:sample-risk-matrix-regression
locale: en
ref_id: SAMPLE-RM-REG
name: Sample Risk Matrix for Regression Test
description: A minimal risk matrix library used for regression testing
copyright: Test
version: 1
publication_date: 2025-01-01
provider: test-provider
packager: test-packager
objects:
  risk_matrix:
    - urn: urn:intuitem:test:matrix:sample-risk-matrix-regression
      ref_id: SAMPLE-RM-REG
      name: Sample Risk Matrix 2x2
      description: A minimal 2x2 risk matrix
      probability:
        - abbreviation: L
          name: Low
          description: Low probability
          hexcolor: '#00FF00'
        - abbreviation: H
          name: High
          description: High probability
          hexcolor: '#FF0000'
      impact:
        - abbreviation: L
          name: Low
          description: Low impact
          hexcolor: '#00FF00'
        - abbreviation: H
          name: High
          description: High impact
          hexcolor: '#FF0000'
      risk:
        - abbreviation: L
          name: Low
          description: Low risk
          hexcolor: '#00FF00'
        - abbreviation: H
          name: High
          description: High risk
          hexcolor: '#FF0000'
      grid:
        - - 0
          - 1
        - - 1
          - 1
""".lstrip()

SAMPLE_THREATS_AND_CONTROLS_YAML = """
urn: urn:intuitem:test:library:sample-threats-controls-regression
locale: en
ref_id: SAMPLE-TC-REG
name: Sample Threats and Controls for Regression Test
description: A library with threats and reference controls for regression testing
copyright: Test
version: 1
publication_date: 2025-01-01
provider: test-provider
packager: test-packager
objects:
  threats:
    - urn: urn:intuitem:test:threat:sample-threat-1
      ref_id: THREAT-1
      name: Sample Threat 1
      description: A sample threat for testing
      annotation: test annotation
    - urn: urn:intuitem:test:threat:sample-threat-2
      ref_id: THREAT-2
      name: Sample Threat 2
      description: Another sample threat for testing
  reference_controls:
    - urn: urn:intuitem:test:reference_control:sample-rc-1
      ref_id: RC-1
      name: Sample Reference Control 1
      description: A sample reference control for testing
      category: policy
    - urn: urn:intuitem:test:reference_control:sample-rc-2
      ref_id: RC-2
      name: Sample Reference Control 2
      description: Another sample reference control for testing
      category: process
""".lstrip()

SAMPLE_INVALID_YAML = """
this is not: a valid library
missing: required fields
""".lstrip()

SAMPLE_MALFORMED_YAML = """
urn: urn:intuitem:test:library:malformed
version: not_an_integer
  - broken indentation
""".lstrip()


# Helpers


def _create_excel_framework_v2() -> bytes:
    """Create a minimal valid v2 Excel file with a framework.

    The v2 format requires:
    - library_meta sheet: key-value pairs for library metadata
    - <prefix>_meta sheet: key-value pairs for each object block
    - <prefix>_content sheet: tabular data for each object block
    """
    wb = openpyxl.Workbook()

    # --- library_meta sheet ---
    ws_meta = wb.active
    ws_meta.title = "library_meta"
    meta_entries = [
        ("type", "library"),
        ("urn", "urn:intuitem:test:library:excel-framework-regression"),
        ("locale", "en"),
        ("ref_id", "EXCEL-FW-REG"),
        ("name", "Excel Framework for Regression Test"),
        ("description", "A framework imported from Excel v2 format"),
        ("copyright", "Test"),
        ("version", "1"),
        ("publication_date", "2025-01-01"),
        ("provider", "test-provider"),
        ("packager", "test-packager"),
    ]
    for row_idx, (key, value) in enumerate(meta_entries, start=1):
        ws_meta.cell(row=row_idx, column=1, value=key)
        ws_meta.cell(row=row_idx, column=2, value=value)

    # --- framework_meta sheet ---
    ws_fw_meta = wb.create_sheet("framework_meta")
    fw_meta_entries = [
        ("type", "framework"),
        ("name", "Excel Test Framework"),
        ("urn", "urn:intuitem:test:framework:excel-framework-regression"),
        ("ref_id", "EXCEL-FW-REG"),
        ("description", "A framework from Excel"),
    ]
    for row_idx, (key, value) in enumerate(fw_meta_entries, start=1):
        ws_fw_meta.cell(row=row_idx, column=1, value=key)
        ws_fw_meta.cell(row=row_idx, column=2, value=value)

    # --- framework_content sheet ---
    ws_fw_content = wb.create_sheet("framework_content")
    headers = ["assessable", "depth", "ref_id", "name", "description"]
    for col_idx, header in enumerate(headers, start=1):
        ws_fw_content.cell(row=1, column=col_idx, value=header)

    rows = [
        (False, 1, "CAT-1", "Category 1", "A sample category"),
        (True, 2, "REQ-1.1", "Requirement 1.1", "A sample requirement"),
        (True, 2, "REQ-1.2", "Requirement 1.2", "Another sample requirement"),
    ]
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws_fw_content.cell(row=row_idx, column=col_idx, value=value)

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def _create_excel_risk_matrix_v2() -> bytes:
    """Create a minimal valid v2 Excel file with a risk matrix."""
    wb = openpyxl.Workbook()

    # --- library_meta sheet ---
    ws_meta = wb.active
    ws_meta.title = "library_meta"
    meta_entries = [
        ("type", "library"),
        ("urn", "urn:intuitem:test:library:excel-risk-matrix-regression"),
        ("locale", "en"),
        ("ref_id", "EXCEL-RM-REG"),
        ("name", "Excel Risk Matrix for Regression Test"),
        ("description", "A risk matrix imported from Excel v2 format"),
        ("copyright", "Test"),
        ("version", "1"),
        ("publication_date", "2025-01-01"),
        ("provider", "test-provider"),
        ("packager", "test-packager"),
    ]
    for row_idx, (key, value) in enumerate(meta_entries, start=1):
        ws_meta.cell(row=row_idx, column=1, value=key)
        ws_meta.cell(row=row_idx, column=2, value=value)

    # --- risk_matrix_meta sheet ---
    ws_rm_meta = wb.create_sheet("risk_matrix_meta")
    rm_meta_entries = [
        ("type", "risk_matrix"),
        ("name", "Excel Test Risk Matrix 2x2"),
        ("urn", "urn:intuitem:test:matrix:excel-risk-matrix-regression"),
        ("ref_id", "EXCEL-RM-REG"),
        ("description", "A 2x2 risk matrix from Excel"),
    ]
    for row_idx, (key, value) in enumerate(rm_meta_entries, start=1):
        ws_rm_meta.cell(row=row_idx, column=1, value=key)
        ws_rm_meta.cell(row=row_idx, column=2, value=value)

    # --- risk_matrix_content sheet ---
    ws_rm_content = wb.create_sheet("risk_matrix_content")
    headers = ["type", "abbreviation", "name", "description", "hexcolor", "grid"]
    for col_idx, header in enumerate(headers, start=1):
        ws_rm_content.cell(row=1, column=col_idx, value=header)

    rows = [
        ("probability", "L", "Low", "Low probability", "#00FF00", None),
        ("probability", "H", "High", "High probability", "#FF0000", None),
        ("impact", "L", "Low", "Low impact", "#00FF00", None),
        ("impact", "H", "High", "High impact", "#FF0000", None),
        ("risk", "L", "Low", "Low risk", "#00FF00", None),
        ("risk", "H", "High", "High risk", "#FF0000", None),
        ("grid", None, None, None, None, "0,1\n1,1"),
    ]
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            if value is not None:
                ws_rm_content.cell(row=row_idx, column=col_idx, value=value)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


# Fixtures


@pytest.fixture
def app_config():
    """Initialize the application (creates built-in groups, etc.)."""
    startup(sender=None, **{})


@pytest.fixture
def admin_client(app_config):
    """Return an authenticated API client with admin privileges."""
    admin = User.objects.create_superuser(
        "admin@regression-tests.com", is_published=True
    )
    admin_group = UserGroup.objects.get(name="BI-UG-ADM")
    admin.folder = admin_group.folder
    admin.save()
    admin_group.user_set.add(admin)
    client = APIClient()
    _auth_token = AuthToken.objects.create(user=admin)
    auth_token = _auth_token[1]
    client.credentials(HTTP_AUTHORIZATION=f"Token {auth_token}")
    return client


@pytest.fixture
def upload_url():
    """Return the URL for the library upload endpoint."""
    return reverse("stored-libraries-upload-library")


# YAML Import Tests


@pytest.mark.django_db
class TestCustomLibraryImportYAML:
    """Regression tests for custom library import via YAML files."""

    def test_upload_framework_yaml(self, admin_client, upload_url):
        """Test uploading a valid YAML framework library creates all expected objects."""
        response = _upload_yaml(
            admin_client,
            upload_url,
            "sample_framework.yaml",
            SAMPLE_FRAMEWORK_YAML.encode("utf-8"),
        )

        assert response.status_code == status.HTTP_201_CREATED, (
            f"Framework YAML upload failed: {response.content}"
        )

        stored = StoredLibrary.objects.filter(
            urn="urn:intuitem:test:library:sample-framework-regression"
        )
        assert stored.exists(), "StoredLibrary not created"
        assert stored.first().is_loaded is True

        loaded = LoadedLibrary.objects.filter(
            urn="urn:intuitem:test:library:sample-framework-regression"
        )
        assert loaded.exists(), "LoadedLibrary not created"
        assert loaded.first().name == "Sample Framework for Regression Test"
        assert loaded.first().provider == "test-provider"

        fw = Framework.objects.filter(
            urn="urn:intuitem:test:framework:sample-framework-regression"
        )
        assert fw.exists(), "Framework not created"
        assert fw.first().name == "Sample Framework for Regression Test"

        nodes = RequirementNode.objects.filter(framework=fw.first())
        assert nodes.count() == 3, f"Expected 3 requirement nodes, got {nodes.count()}"

        cat_node = nodes.get(ref_id="CAT-1")
        assert cat_node.assessable is False
        child_nodes = nodes.filter(
            parent_urn="urn:intuitem:test:req_node:sample-fw-reg:cat-1"
        )
        assert child_nodes.count() == 2

    def test_upload_risk_matrix_yaml(self, admin_client, upload_url):
        """Test uploading a valid YAML risk matrix library."""
        response = _upload_yaml(
            admin_client,
            upload_url,
            "sample_risk_matrix.yaml",
            SAMPLE_RISK_MATRIX_YAML.encode("utf-8"),
        )

        assert response.status_code == status.HTTP_201_CREATED, (
            f"Risk matrix YAML upload failed: {response.content}"
        )

        rm = RiskMatrix.objects.filter(
            urn="urn:intuitem:test:matrix:sample-risk-matrix-regression"
        )
        assert rm.exists(), "RiskMatrix not created"
        assert rm.first().name == "Sample Risk Matrix 2x2"

        json_def = rm.first().json_definition
        assert len(json_def["probability"]) == 2
        assert len(json_def["impact"]) == 2
        assert len(json_def["risk"]) == 2
        assert json_def["grid"] == [[0, 1], [1, 1]]

    def test_upload_threats_and_controls_yaml(self, admin_client, upload_url):
        """Test uploading a YAML library with threats and reference controls."""
        response = _upload_yaml(
            admin_client,
            upload_url,
            "sample_threats_controls.yaml",
            SAMPLE_THREATS_AND_CONTROLS_YAML.encode("utf-8"),
        )

        assert response.status_code == status.HTTP_201_CREATED, (
            f"Threats/controls YAML upload failed: {response.content}"
        )

        threats = Threat.objects.filter(
            urn__startswith="urn:intuitem:test:threat:sample-threat-"
        )
        assert threats.count() == 2, f"Expected 2 threats, got {threats.count()}"
        assert threats.filter(ref_id="THREAT-1").exists()
        assert threats.filter(ref_id="THREAT-2").exists()

        rcs = ReferenceControl.objects.filter(
            urn__startswith="urn:intuitem:test:reference_control:sample-rc-"
        )
        assert rcs.count() == 2, f"Expected 2 reference controls, got {rcs.count()}"
        assert rcs.filter(ref_id="RC-1").exists()
        assert rcs.filter(ref_id="RC-2").exists()

    def test_upload_yaml_dry_run(self, admin_client, upload_url):
        """Test dry-run mode returns library metadata without persisting."""
        response = _upload_yaml(
            admin_client,
            f"{upload_url}?dry_run=true",
            "sample_framework.yaml",
            SAMPLE_FRAMEWORK_YAML.encode("utf-8"),
        )

        assert response.status_code == status.HTTP_200_OK, (
            f"Dry-run failed: {response.content}"
        )

        data = response.json()
        assert data["urn"] == "urn:intuitem:test:library:sample-framework-regression"
        assert data["version"] == 1
        assert "objects_meta" in data
        assert data["objects_meta"]["framework"] == 1

        assert not StoredLibrary.objects.filter(
            urn="urn:intuitem:test:library:sample-framework-regression"
        ).exists(), "Dry-run should not persist StoredLibrary"
        assert not LoadedLibrary.objects.filter(
            urn="urn:intuitem:test:library:sample-framework-regression"
        ).exists(), "Dry-run should not persist LoadedLibrary"

    def test_upload_duplicate_yaml_rejected(self, admin_client, upload_url):
        """Test that uploading the same library twice is rejected."""
        content = SAMPLE_FRAMEWORK_YAML.encode("utf-8")

        response1 = _upload_yaml(
            admin_client, upload_url, "sample_framework.yaml", content
        )
        assert response1.status_code == status.HTTP_201_CREATED

        response2 = _upload_yaml(
            admin_client, upload_url, "sample_framework.yaml", content
        )
        assert response2.status_code in (
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            status.HTTP_400_BAD_REQUEST,
        ), f"Duplicate upload should be rejected, got {response2.status_code}"

        data = json.loads(response2.content)
        assert "error" in data

    def test_upload_invalid_yaml_rejected(self, admin_client, upload_url):
        """Test that uploading YAML missing required fields is rejected."""
        response = _upload_yaml(
            admin_client,
            upload_url,
            "invalid.yaml",
            SAMPLE_INVALID_YAML.encode("utf-8"),
        )

        assert response.status_code in (
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            status.HTTP_400_BAD_REQUEST,
        ), (
            f"Invalid YAML should be rejected, got {response.status_code}: {response.content}"
        )

    def test_upload_malformed_yaml_rejected(self, admin_client, upload_url):
        """Test that uploading malformed YAML is rejected."""
        response = _upload_yaml(
            admin_client,
            upload_url,
            "malformed.yaml",
            SAMPLE_MALFORMED_YAML.encode("utf-8"),
        )

        assert response.status_code in (
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            status.HTTP_400_BAD_REQUEST,
        ), (
            f"Malformed YAML should be rejected, got {response.status_code}: {response.content}"
        )

    def test_upload_empty_file_rejected(self, admin_client, upload_url):
        """Test that uploading an empty file is rejected.

        There is no explicit empty-file guard in the view. Rejection happens
        because magic.from_buffer(b"") returns a MIME type such as
        'application/x-empty' that is not in the allowed list.
        """
        response = _upload_yaml(admin_client, upload_url, "empty.yaml", b"")

        assert response.status_code == status.HTTP_400_BAD_REQUEST, (
            f"Empty file should be rejected, got {response.status_code}"
        )

    def test_upload_wrong_extension_rejected(self, admin_client, upload_url):
        """Test that uploading a file with wrong extension is rejected."""
        response = admin_client.post(
            upload_url,
            data=SAMPLE_FRAMEWORK_YAML.encode("utf-8"),
            content_type="text/plain",
            HTTP_CONTENT_DISPOSITION="attachment; filename=library.txt",
        )

        assert response.status_code == status.HTTP_400_BAD_REQUEST, (
            f"Wrong extension should be rejected, got {response.status_code}: {response.content}"
        )

    def test_upload_unauthenticated_rejected(self, upload_url):
        """Test that unauthenticated upload is rejected."""
        client = APIClient()
        response = _upload_yaml(
            client,
            upload_url,
            "sample_framework.yaml",
            SAMPLE_FRAMEWORK_YAML.encode("utf-8"),
        )

        assert response.status_code == status.HTTP_401_UNAUTHORIZED

    def test_upload_newer_version_stored_for_update(self, admin_client, upload_url):
        """Test that uploading a newer version stores it for update without reloading.

        When a LoadedLibrary already exists for a given URN, the upload endpoint
        stores the new StoredLibrary (v2) but does NOT reload it automatically.
        It returns HTTP 201 with a 'libraryStoredForUpdate' warning.
        The actual update must be triggered separately via the _update endpoint.
        """
        response1 = _upload_yaml(
            admin_client,
            upload_url,
            "sample_framework_v1.yaml",
            SAMPLE_FRAMEWORK_YAML.encode("utf-8"),
        )
        assert response1.status_code == status.HTTP_201_CREATED

        v2_yaml = SAMPLE_FRAMEWORK_YAML.replace("version: 1", "version: 2").replace(
            "name: Sample Framework for Regression Test",
            "name: Sample Framework for Regression Test v2",
        )
        response2 = _upload_yaml(
            admin_client,
            upload_url,
            "sample_framework_v2.yaml",
            v2_yaml.encode("utf-8"),
        )

        assert response2.status_code == status.HTTP_201_CREATED, (
            f"Version 2 upload failed: {response2.content}"
        )
        data2 = response2.json()
        assert data2.get("warning") == "libraryStoredForUpdate", (
            f"Expected 'libraryStoredForUpdate' warning, got: {data2}"
        )

        assert StoredLibrary.objects.filter(
            urn="urn:intuitem:test:library:sample-framework-regression",
            version=2,
        ).exists(), "StoredLibrary v2 should be stored"

        loaded = LoadedLibrary.objects.filter(
            urn="urn:intuitem:test:library:sample-framework-regression"
        ).first()
        assert loaded is not None
        assert loaded.version == 1, (
            f"LoadedLibrary should still be v1 (update not triggered), got v{loaded.version}"
        )


# Excel Import Tests


@pytest.mark.django_db
class TestCustomLibraryImportExcel:
    """Regression tests for custom library import via Excel files.

    Since the Excel flow relies on an external conversion script
    (convert_library_v2.py) run inside a sandbox, these tests mock the
    sandbox layer and verify the integration between the upload endpoint
    and the store/load pipeline.
    """

    @patch("library.views.ExcelUploadHandler")
    @patch("library.views.magic.from_buffer", return_value=XLSX_MIME)
    def test_upload_framework_excel(
        self, _mock_magic, mock_excel_handler, admin_client, upload_url
    ):
        """Test uploading a valid Excel framework file.

        The sandbox is mocked to return known YAML output, then we verify
        the full store + load pipeline works.
        """
        excel_content = _create_excel_framework_v2()

        # The Excel upload path uses ExcelUploadHandler which calls the
        # sandbox with the conversion script.  We mock the handler to
        # return valid YAML so the rest of the pipeline (store + load)
        # is exercised for real.
        yaml_output = (
            SAMPLE_FRAMEWORK_YAML.replace(
                "urn: urn:intuitem:test:library:sample-framework-regression",
                "urn: urn:intuitem:test:library:excel-framework-regression",
            )
            .replace(
                "urn: urn:intuitem:test:framework:sample-framework-regression",
                "urn: urn:intuitem:test:framework:excel-framework-regression",
            )
            .replace(
                "ref_id: SAMPLE-FW-REG",
                "ref_id: EXCEL-FW-REG",
            )
            .replace(
                "name: Sample Framework for Regression Test",
                "name: Excel Framework for Regression Test",
            )
        )

        mock_instance = MagicMock()
        mock_instance.process_upload.return_value = {
            "yaml": yaml_output,
            "status": 200,
        }
        mock_excel_handler.return_value = mock_instance

        response = _upload_excel(
            admin_client,
            upload_url,
            "test_framework.xlsx",
            excel_content,
        )

        assert response.status_code == status.HTTP_201_CREATED, (
            f"Excel framework upload failed: {response.content}"
        )

        stored = StoredLibrary.objects.filter(
            urn="urn:intuitem:test:library:excel-framework-regression"
        )
        assert stored.exists(), "StoredLibrary not created from Excel upload"

        loaded = LoadedLibrary.objects.filter(
            urn="urn:intuitem:test:library:excel-framework-regression"
        )
        assert loaded.exists(), "LoadedLibrary not created from Excel upload"

        fw = Framework.objects.filter(
            urn="urn:intuitem:test:framework:excel-framework-regression"
        )
        assert fw.exists(), "Framework not created from Excel upload"

        nodes = RequirementNode.objects.filter(framework=fw.first())
        assert nodes.count() == 3, (
            f"Expected 3 requirement nodes from Excel, got {nodes.count()}"
        )

    @patch("library.views.ExcelUploadHandler")
    @patch("library.views.magic.from_buffer", return_value=XLSX_MIME)
    def test_upload_risk_matrix_excel(
        self, _mock_magic, mock_excel_handler, admin_client, upload_url
    ):
        """Test uploading a valid Excel risk matrix file."""
        excel_content = _create_excel_risk_matrix_v2()

        yaml_output = (
            SAMPLE_RISK_MATRIX_YAML.replace(
                "urn: urn:intuitem:test:library:sample-risk-matrix-regression",
                "urn: urn:intuitem:test:library:excel-risk-matrix-regression",
            )
            .replace(
                "urn: urn:intuitem:test:matrix:sample-risk-matrix-regression",
                "urn: urn:intuitem:test:matrix:excel-risk-matrix-regression",
            )
            .replace(
                "ref_id: SAMPLE-RM-REG",
                "ref_id: EXCEL-RM-REG",
            )
            .replace(
                "name: Sample Risk Matrix for Regression Test",
                "name: Excel Risk Matrix for Regression Test",
            )
            .replace(
                "name: Sample Risk Matrix 2x2",
                "name: Excel Test Risk Matrix 2x2",
            )
        )

        mock_instance = MagicMock()
        mock_instance.process_upload.return_value = {
            "yaml": yaml_output,
            "status": 200,
        }
        mock_excel_handler.return_value = mock_instance

        response = _upload_excel(
            admin_client,
            upload_url,
            "test_risk_matrix.xlsx",
            excel_content,
        )

        assert response.status_code == status.HTTP_201_CREATED, (
            f"Excel risk matrix upload failed: {response.content}"
        )

        rm = RiskMatrix.objects.filter(
            urn="urn:intuitem:test:matrix:excel-risk-matrix-regression"
        )
        assert rm.exists(), "RiskMatrix not created from Excel upload"

    @patch("library.views.ExcelUploadHandler")
    @patch("library.views.magic.from_buffer", return_value=XLSX_MIME)
    def test_upload_excel_dry_run(
        self, _mock_magic, mock_excel_handler, admin_client, upload_url
    ):
        """Test dry-run mode with Excel file."""
        excel_content = _create_excel_framework_v2()

        yaml_output = SAMPLE_FRAMEWORK_YAML.replace(
            "urn: urn:intuitem:test:library:sample-framework-regression",
            "urn: urn:intuitem:test:library:excel-dry-run-regression",
        ).replace(
            "urn: urn:intuitem:test:framework:sample-framework-regression",
            "urn: urn:intuitem:test:framework:excel-dry-run-regression",
        )

        mock_instance = MagicMock()
        mock_instance.process_upload.return_value = {
            "yaml": yaml_output,
            "status": 200,
        }
        mock_excel_handler.return_value = mock_instance

        response = _upload_excel(
            admin_client,
            f"{upload_url}?dry_run=true",
            "test_framework.xlsx",
            excel_content,
        )

        assert response.status_code == status.HTTP_200_OK, (
            f"Excel dry-run failed: {response.content}"
        )

        data = response.json()
        assert data["urn"] == "urn:intuitem:test:library:excel-dry-run-regression"
        assert "objects_meta" in data

        assert not StoredLibrary.objects.filter(
            urn="urn:intuitem:test:library:excel-dry-run-regression"
        ).exists()

    @patch("library.views.ExcelUploadHandler")
    @patch("library.views.magic.from_buffer", return_value=XLSX_MIME)
    def test_upload_excel_conversion_failure(
        self, _mock_magic, mock_excel_handler, admin_client, upload_url
    ):
        """Test that Excel conversion failure returns appropriate error."""
        excel_content = _create_excel_framework_v2()

        mock_instance = MagicMock()
        mock_instance.process_upload.return_value = {
            "error": "invalidExcelFile",
            "status": 400,
        }
        mock_excel_handler.return_value = mock_instance

        response = _upload_excel(
            admin_client,
            upload_url,
            "bad_framework.xlsx",
            excel_content,
        )

        assert response.status_code == status.HTTP_400_BAD_REQUEST, (
            f"Excel conversion failure should return 400, got {response.status_code}: {response.content}"
        )

    @patch("library.views.ExcelUploadHandler")
    @patch("library.views.magic.from_buffer", return_value=XLSX_MIME)
    def test_upload_excel_timeout(
        self, _mock_magic, mock_excel_handler, admin_client, upload_url
    ):
        """Test that Excel conversion timeout returns appropriate error."""
        excel_content = _create_excel_framework_v2()

        mock_instance = MagicMock()
        mock_instance.process_upload.return_value = {
            "error": "processingTimeout",
            "status": 504,
        }
        mock_excel_handler.return_value = mock_instance

        response = _upload_excel(
            admin_client,
            upload_url,
            "slow_framework.xlsx",
            excel_content,
        )

        assert response.status_code == status.HTTP_504_GATEWAY_TIMEOUT, (
            f"Timeout should return 504, got {response.status_code}"
        )

    def test_upload_non_excel_as_xlsx_rejected(self, admin_client, upload_url):
        """Test that a non-Excel file with .xlsx extension is rejected."""
        response = _upload_excel(
            admin_client,
            upload_url,
            "fake.xlsx",
            b"This is not an Excel file",
        )

        # Should fail MIME validation or Excel validation
        assert response.status_code in (
            status.HTTP_400_BAD_REQUEST,
            status.HTTP_422_UNPROCESSABLE_ENTITY,
        ), (
            f"Fake Excel should be rejected, got {response.status_code}: {response.content}"
        )


# Store Library Content Unit Tests (direct model-level)


@pytest.mark.django_db
class TestStoreLibraryContentRegression:
    """Regression tests at the model level for store_library_content."""

    def test_store_framework_library(self):
        """Test storing a framework library directly."""
        stored, error = StoredLibrary.store_library_content(
            SAMPLE_FRAMEWORK_YAML.encode("utf-8")
        )
        assert error is None, f"Unexpected error: {error}"
        assert stored is not None
        assert stored.urn == "urn:intuitem:test:library:sample-framework-regression"
        assert stored.version == 1
        assert stored.provider == "test-provider"
        assert stored.objects_meta["framework"] == 1

    def test_store_risk_matrix_library(self):
        """Test storing a risk matrix library directly."""
        stored, error = StoredLibrary.store_library_content(
            SAMPLE_RISK_MATRIX_YAML.encode("utf-8")
        )
        assert error is None, f"Unexpected error: {error}"
        assert stored is not None
        assert stored.urn == "urn:intuitem:test:library:sample-risk-matrix-regression"
        assert "risk_matrix" in stored.objects_meta

    def test_store_threats_and_controls_library(self):
        """Test storing a threats and reference controls library directly."""
        stored, error = StoredLibrary.store_library_content(
            SAMPLE_THREATS_AND_CONTROLS_YAML.encode("utf-8")
        )
        assert error is None, f"Unexpected error: {error}"
        assert stored is not None
        assert stored.objects_meta["threats"] == 2
        assert stored.objects_meta["reference_controls"] == 2

    def test_store_and_load_framework(self):
        """Test storing then loading a framework library creates DB objects."""
        stored, error = StoredLibrary.store_library_content(
            SAMPLE_FRAMEWORK_YAML.encode("utf-8")
        )
        assert error is None
        assert stored is not None

        load_error = stored.load()
        assert load_error is None, f"Load error: {load_error}"

        assert Framework.objects.filter(
            urn="urn:intuitem:test:framework:sample-framework-regression"
        ).exists()
        assert (
            RequirementNode.objects.filter(
                urn__startswith="urn:intuitem:test:req_node:sample-fw-reg:"
            ).count()
            == 3
        )

    def test_store_and_load_risk_matrix(self):
        """Test storing then loading a risk matrix library creates DB objects."""
        stored, error = StoredLibrary.store_library_content(
            SAMPLE_RISK_MATRIX_YAML.encode("utf-8")
        )
        assert error is None
        assert stored is not None

        load_error = stored.load()
        assert load_error is None, f"Load error: {load_error}"

        rm = RiskMatrix.objects.filter(
            urn="urn:intuitem:test:matrix:sample-risk-matrix-regression"
        )
        assert rm.exists()
        assert rm.first().json_definition["grid"] == [[0, 1], [1, 1]]

    def test_store_and_load_threats_controls(self):
        """Test storing then loading a threats/controls library creates DB objects."""
        stored, error = StoredLibrary.store_library_content(
            SAMPLE_THREATS_AND_CONTROLS_YAML.encode("utf-8")
        )
        assert error is None
        assert stored is not None

        load_error = stored.load()
        assert load_error is None, f"Load error: {load_error}"

        assert Threat.objects.filter(ref_id="THREAT-1").exists()
        assert Threat.objects.filter(ref_id="THREAT-2").exists()
        assert ReferenceControl.objects.filter(ref_id="RC-1").exists()
        assert ReferenceControl.objects.filter(ref_id="RC-2").exists()

    def test_store_missing_required_fields_raises(self):
        """Test that storing YAML with missing required fields raises ValueError."""
        with pytest.raises(ValueError, match="following fields are missing"):
            StoredLibrary.store_library_content(SAMPLE_INVALID_YAML.encode("utf-8"))

    def test_store_same_version_rejected(self):
        """Test that storing the same content twice returns a duplicate error.

        Note: StoredLibrary.HASH_CHECKSUM_SET is only populated by management
        commands (storelibraries / autoloadlibraries), so the hash-based shortcut
        never fires in tests. Duplication is detected via the same-version DB
        query (same urn + locale + version), which still returns the correct
        'libraryAlreadyLoadedError'.
        """
        content = SAMPLE_FRAMEWORK_YAML.encode("utf-8")
        stored1, error1 = StoredLibrary.store_library_content(content)
        assert error1 is None
        assert stored1 is not None

        stored2, error2 = StoredLibrary.store_library_content(content)
        assert stored2 is None
        assert error2 == "libraryAlreadyLoadedError"

    def test_store_outdated_version_rejected(self):
        """Test that storing an older version after a newer one is rejected."""

        v2_yaml = SAMPLE_RISK_MATRIX_YAML.replace("version: 1", "version: 2")
        _, error_v2 = StoredLibrary.store_library_content(v2_yaml.encode("utf-8"))
        assert error_v2 is None

        # Try storing version 1 — should be rejected
        stored_v1, error_v1 = StoredLibrary.store_library_content(
            SAMPLE_RISK_MATRIX_YAML.encode("utf-8")
        )
        assert stored_v1 is None
        assert error_v1 == "libraryOutdatedError"

    def test_dry_run_does_not_persist(self):
        """Test dry-run mode returns metadata without creating DB records."""
        data, error = StoredLibrary.store_library_content(
            SAMPLE_FRAMEWORK_YAML.encode("utf-8"), dry_run=True
        )
        assert error is None
        assert isinstance(data, dict)
        assert data["urn"] == "urn:intuitem:test:library:sample-framework-regression"
        assert data["version"] == 1
        assert not StoredLibrary.objects.filter(
            urn="urn:intuitem:test:library:sample-framework-regression"
        ).exists()
