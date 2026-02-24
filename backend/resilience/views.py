import io
import re

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment

from core.views import (
    BaseModelViewSet as AbstractBaseModelViewSet,
    ExportMixin,
    escape_excel_formula,
)
from rest_framework.decorators import action
from rest_framework.response import Response
from core.serializers import RiskMatrixReadSerializer
from django.utils.decorators import method_decorator
from django.views.decorators.cache import cache_page

from iam.models import RoleAssignment, Folder
from core.models import Asset
from .models import BusinessImpactAnalysis, AssetAssessment, EscalationThreshold

SHORT_CACHE_TTL = 2  # mn
MED_CACHE_TTL = 5  # mn
LONG_CACHE_TTL = 60  # mn


class BaseModelViewSet(AbstractBaseModelViewSet):
    serializers_module = "resilience.serializers"


class BusinessImpactAnalysisViewSet(BaseModelViewSet, ExportMixin):
    model = BusinessImpactAnalysis
    filterset_fields = [
        "folder",
        "perimeter",
        "perimeter__folder",
        "authors",
        "risk_matrix",
        "status",
    ]

    export_config = {
        "fields": {
            "internal_id": {
                "source": "id",
                "label": "internal_id",
                "format": lambda v: str(v),
            },
            "name": {"source": "name", "label": "name", "escape": True},
            "description": {
                "source": "description",
                "label": "description",
                "escape": True,
            },
            "perimeter": {
                "source": "perimeter.name",
                "label": "perimeter",
                "escape": True,
            },
            "perimeter_ref_id": {
                "source": "perimeter.ref_id",
                "label": "perimeter_ref_id",
                "escape": True,
            },
            "risk_matrix": {
                "source": "risk_matrix.name",
                "label": "risk_matrix",
                "escape": True,
            },
            "risk_matrix_ref_id": {
                "source": "risk_matrix.ref_id",
                "label": "risk_matrix_ref_id",
                "escape": True,
            },
            "folder": {
                "source": "folder.name",
                "label": "folder",
                "escape": True,
            },
            "version": {
                "source": "version",
                "label": "version",
            },
            "status": {
                "source": "get_status_display",
                "label": "status",
            },
            "eta": {
                "source": "eta",
                "label": "eta",
            },
            "due_date": {
                "source": "due_date",
                "label": "due_date",
            },
            "observation": {
                "source": "observation",
                "label": "observation",
                "escape": True,
            },
            "authors": {
                "source": "authors",
                "label": "authors",
                "format": lambda qs: ",".join(
                    escape_excel_formula(str(a)) for a in qs.all()
                ),
            },
            "reviewers": {
                "source": "reviewers",
                "label": "reviewers",
                "format": lambda qs: ",".join(
                    escape_excel_formula(str(r)) for r in qs.all()
                ),
            },
        },
        "asset_assessment_fields": {
            "bia_name": {
                "source": "bia.name",
                "label": "bia_name",
                "escape": True,
            },
            "asset": {
                "source": "asset.name",
                "label": "asset",
                "escape": True,
            },
            "asset_ref_id": {
                "source": "asset.ref_id",
                "label": "asset_ref_id",
                "escape": True,
            },
            "recovery_documented": {
                "source": "recovery_documented",
                "label": "recovery_documented",
            },
            "recovery_tested": {
                "source": "recovery_tested",
                "label": "recovery_tested",
            },
            "recovery_targets_met": {
                "source": "recovery_targets_met",
                "label": "recovery_targets_met",
            },
            "dependencies": {
                "source": "dependencies",
                "label": "dependencies",
                "format": lambda qs: ",".join(
                    escape_excel_formula(o.name) for o in qs.all()
                ),
            },
            "associated_controls": {
                "source": "associated_controls",
                "label": "associated_controls",
                "format": lambda qs: ",".join(
                    escape_excel_formula(o.name) for o in qs.all()
                ),
            },
            "evidences": {
                "source": "evidences",
                "label": "evidences",
                "format": lambda qs: ",".join(
                    escape_excel_formula(o.name) for o in qs.all()
                ),
            },
            "observation": {
                "source": "observation",
                "label": "observation",
                "escape": True,
            },
        },
        "threshold_fields": {
            "bia_name": {
                "source": "asset_assessment.bia.name",
                "label": "bia_name",
                "escape": True,
            },
            "asset": {
                "source": "asset_assessment.asset.name",
                "label": "asset",
                "escape": True,
            },
            "asset_ref_id": {
                "source": "asset_assessment.asset.ref_id",
                "label": "asset_ref_id",
                "escape": True,
            },
            "point_in_time": {
                "source": "point_in_time",
                "label": "point_in_time",
            },
            "quali_impact": {
                "source": "quali_impact",
                "label": "quali_impact",
            },
            "quanti_impact": {
                "source": "quanti_impact",
                "label": "quanti_impact",
            },
            "quanti_impact_unit": {
                "source": "quanti_impact_unit",
                "label": "quanti_impact_unit",
            },
            "qualifications": {
                "source": "qualifications",
                "label": "qualifications",
                "format": lambda qs: ",".join(
                    escape_excel_formula(o.name) for o in qs.all()
                ),
            },
            "justification": {
                "source": "justification",
                "label": "justification",
                "escape": True,
            },
        },
        "wrap_columns": ["name", "description", "observation"],
        "filename": "business_impact_analysis_export",
        "select_related": ["folder", "perimeter", "risk_matrix"],
        "prefetch_related": ["authors", "reviewers"],
    }

    def _build_bia_workbook(self, queryset) -> Workbook:
        wb = Workbook()
        wb.remove(wb.active)

        summary_ws = wb.create_sheet(title="Summary")
        main_fields = self.export_config["fields"]
        main_headers = [f.get("label", name) for name, f in main_fields.items()]
        summary_ws.append(main_headers)

        wrap_alignment = Alignment(wrap_text=True, vertical="top")
        wrap_columns = self.export_config.get("wrap_columns", ["name", "description"])

        used_sheet_names = {"Summary"}
        assessment_fields = self.export_config.get("asset_assessment_fields", {})
        threshold_fields = self.export_config.get("threshold_fields", {})

        def _make_sheet_name(base_name: str, suffix: str = "") -> str:
            invalid_chars = r"[\\/\?\*\[\]:]"
            base_name = re.sub(invalid_chars, "", base_name)
            max_len = max(1, 31 - len(suffix))
            sheet_name = f"{base_name[:max_len]}{suffix}"[:31]
            counter = 1
            while sheet_name in used_sheet_names:
                counter_suffix = f" ({counter})"
                max_len = max(1, 31 - len(suffix) - len(counter_suffix))
                sheet_name = (f"{base_name[:max_len]}{suffix}{counter_suffix}")[:31]
                counter += 1
            used_sheet_names.add(sheet_name)
            return sheet_name

        for bia in queryset:
            row = [
                self._resolve_field_value(bia, field_config)
                for field_config in main_fields.values()
            ]
            summary_ws.append(row)

            asset_assessments = (
                AssetAssessment.objects.filter(bia=bia)
                .select_related("asset", "asset__folder", "folder", "bia")
                .prefetch_related("dependencies", "associated_controls", "evidences")
                .order_by("asset__name")
            )

            if asset_assessments.exists():
                base_name = bia.name or f"BIA_{bia.pk}"
                assessment_sheet_name = _make_sheet_name(base_name)

                assessment_ws = wb.create_sheet(title=assessment_sheet_name)
                assessment_headers = [
                    f.get("label", name) for name, f in assessment_fields.items()
                ]
                assessment_ws.append(assessment_headers)

                for assessment in asset_assessments:
                    assessment_row = [
                        self._resolve_field_value(assessment, field_config)
                        for field_config in assessment_fields.values()
                    ]
                    assessment_ws.append(assessment_row)

                for col_idx, header in enumerate(assessment_headers, 1):
                    if header.lower() in wrap_columns:
                        for row_idx in range(1, assessment_ws.max_row + 1):
                            assessment_ws.cell(
                                row=row_idx, column=col_idx
                            ).alignment = wrap_alignment

            thresholds = (
                EscalationThreshold.objects.filter(asset_assessment__bia=bia)
                .select_related(
                    "asset_assessment",
                    "asset_assessment__asset",
                    "asset_assessment__bia",
                )
                .prefetch_related("qualifications")
                .order_by("asset_assessment__asset__name", "point_in_time")
            )

            if thresholds.exists():
                base_name = bia.name or f"BIA_{bia.pk}"
                threshold_sheet_name = _make_sheet_name(base_name, " - thresholds")

                threshold_ws = wb.create_sheet(title=threshold_sheet_name)
                threshold_headers = [
                    f.get("label", name) for name, f in threshold_fields.items()
                ]
                threshold_ws.append(threshold_headers)

                for threshold in thresholds:
                    threshold_row = [
                        self._resolve_field_value(threshold, field_config)
                        for field_config in threshold_fields.values()
                    ]
                    threshold_ws.append(threshold_row)

                for col_idx, header in enumerate(threshold_headers, 1):
                    if header.lower() in wrap_columns:
                        for row_idx in range(1, threshold_ws.max_row + 1):
                            threshold_ws.cell(
                                row=row_idx, column=col_idx
                            ).alignment = wrap_alignment

        for col_idx, header in enumerate(main_headers, 1):
            if header.lower() in wrap_columns:
                for row_idx in range(1, summary_ws.max_row + 1):
                    summary_ws.cell(
                        row=row_idx, column=col_idx
                    ).alignment = wrap_alignment

        return wb

    @action(detail=False, name="Export BIAs with assessments and thresholds")
    def export_xlsx(self, request):
        if not self.export_config:
            return HttpResponse(
                status=501, content="Export not configured for this model"
            )

        wb = self._build_bia_workbook(self._get_export_queryset())

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"{self.export_config.get('filename', 'export')}_multi_sheet.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    @action(detail=True, methods=["get"], name="Export BIA as Excel")
    def xlsx(self, request, pk):
        if not self.export_config:
            return HttpResponse(
                status=501, content="Export not configured for this model"
            )

        bia = self.get_object()
        wb = self._build_bia_workbook(BusinessImpactAnalysis.objects.filter(pk=bia.pk))

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"bia-{pk}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get status choices")
    def status(self, request):
        return Response(dict(BusinessImpactAnalysis.Status.choices))

    @action(detail=True, name="Get the BIA metrics")
    def metrics(self, request, pk):
        bia = self.get_object()
        return Response(bia.metrics())

    @action(detail=True, name="Build qualitative table", url_path="build-table")
    def impact_table(self, request, pk):
        bia = self.get_object()
        table = bia.build_table()
        return Response(table)


class AssetAssessmentViewSet(BaseModelViewSet):
    model = AssetAssessment
    filterset_fields = ["bia", "asset"]
    search_fields = ["bia__name", "asset__name"]
    ordering = ["asset"]

    def _get_asset_verdict(self, asset):
        """
        Calculate verdict based on security and recovery objectives vs capabilities.
        Returns False if any objective is not met, True if all are met, None otherwise.
        """
        # Get comparisons from the asset's methods
        security_comparison = asset.get_security_objectives_comparison()
        recovery_comparison = asset.get_recovery_objectives_comparison()

        all_comparisons = []

        if security_comparison:
            all_comparisons.extend(security_comparison)
        if recovery_comparison:
            all_comparisons.extend(recovery_comparison)

        if not all_comparisons:
            return None

        # Check if any comparison has failed (verdict is False)
        if any(comp.get("verdict") is False for comp in all_comparisons):
            return False

        # Check if all comparisons have passed (verdict is True)
        if all(comp.get("verdict") is True for comp in all_comparisons):
            return True

        return None  # Some are indeterminate

    @action(detail=True, name="Get risk matrix", url_path="risk-matrix")
    def risk_matrix(self, request, pk=None):
        aa = self.get_object()
        return Response(RiskMatrixReadSerializer(aa.bia.risk_matrix).data)

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=True, name="Get impact choices")
    def quali_impact(self, request, pk):
        aa = self.get_object()
        undefined = dict([(-1, "--")])
        _choices = dict(
            zip(
                list(range(0, 64)),
                [x["name"] for x in aa.bia.parsed_matrix["impact"]],
            )
        )
        choices = undefined | _choices
        return Response(choices)

    @action(detail=True, name="Get the asset assessment details")
    def metrics(self, request, pk):
        res = self.get_object().metrics()
        return Response(res)

    @action(detail=True, name="Get dependency graph", url_path="dependency-graph")
    def dependency_graph(self, request, pk):
        """
        Returns graph data for visualizing asset dependencies.
        Includes nodes grouped by folder and links between parent-child assets.
        Only includes assets that are viewable by the current user.
        """

        aa = self.get_object()
        asset = aa.asset

        # Get all descendants of the asset
        descendants = asset.get_descendants()
        all_assets = {asset} | descendants
        all_asset_ids = {a.id for a in all_assets}

        # Get viewable asset IDs using the standard pattern
        viewable_asset_ids = RoleAssignment.get_accessible_object_ids(
            folder=Folder.get_root_folder(), user=request.user, object_type=Asset
        )[0]

        # Filter to only assets that are both in our tree and viewable
        accessible_asset_ids = all_asset_ids & set(viewable_asset_ids)
        viewable_assets = Asset.objects.filter(id__in=accessible_asset_ids)

        # Build nodes grouped by folder
        folder_groups = {}
        nodes = []

        for a in viewable_assets:
            folder_name = a.folder.name if a.folder else "No Folder"
            folder_id = str(a.folder.id) if a.folder else "no-folder"

            if folder_id not in folder_groups:
                folder_groups[folder_id] = {
                    "id": folder_id,
                    "name": folder_name,
                    "nodes": [],
                }

            # Calculate verdict based on objectives vs capabilities
            verdict = self._get_asset_verdict(a)

            node = {
                "id": str(a.id),
                "label": a.name,
                "folder": folder_id,
                "verdict": verdict,
            }
            nodes.append(node)
            folder_groups[folder_id]["nodes"].append(str(a.id))

        # Build links - invert direction to show children supporting parents
        links = []

        for a in viewable_assets:
            # Get direct children that are also viewable
            children = a.child_assets.filter(id__in=accessible_asset_ids)
            for child in children:
                # Invert: child (source) -> parent (target) to show support relationship
                links.append(
                    {
                        "source": str(child.id),
                        "target": str(a.id),
                    }
                )

        return Response(
            {
                "nodes": nodes,
                "links": links,
                "groups": list(folder_groups.values()),
            }
        )


class EscalationThresholdViewSet(BaseModelViewSet):
    model = EscalationThreshold
    filterset_fields = ["asset_assessment", "quali_impact"]
    ordering = ["point_in_time"]

    @action(detail=True, name="Get risk matrix", url_path="risk-matrix")
    def risk_matrix(self, request, pk=None):
        et = self.get_object()
        return Response(RiskMatrixReadSerializer(et.risk_matrix).data)

    @method_decorator(cache_page(60 * LONG_CACHE_TTL))
    @action(detail=False, name="Get quantification units")
    def quant_unit(self, request):
        return Response(dict(EscalationThreshold.QUANT_IMPACT_UNIT))

    @action(detail=True, name="Get impact choices")
    def quali_impact(self, request, pk):
        escalation_threshold = self.get_object()
        undefined_choice = {-1: "--"}
        impact_choices = dict(
            enumerate(
                impact["name"]
                for impact in escalation_threshold.parsed_matrix["impact"]
            )
        )
        choices = undefined_choice | impact_choices
        return Response(choices)
