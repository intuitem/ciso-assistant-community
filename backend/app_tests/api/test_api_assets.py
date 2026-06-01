import uuid

import pytest
from knox.models import AuthToken
from rest_framework import status
from rest_framework.test import APIClient
from core.models import Asset
from core.utils import RoleCodename
from iam.models import Folder, Role, RoleAssignment, User

from test_utils import EndpointTestsQueries

# Generic asset data for tests
ASSET_NAME = "Test Asset"
ASSET_DESCRIPTION = "Test Description"
ASSET_TYPE = ("PR", "Primary")
ASSET_TYPE2 = ("SP", "Support")
ASSET_PARENT_ASSETS = []


@pytest.mark.django_db
class TestAssetsUnauthenticated:
    """Perform tests on Assets API endpoint without authentication"""

    client = APIClient()

    def test_get_assets(self):
        """test to get assets from the API without authentication"""

        EndpointTestsQueries.get_object(
            self.client,
            "Assets",
            Asset,
            {
                "name": ASSET_NAME,
                "description": ASSET_DESCRIPTION,
                "folder": Folder.objects.create(name="test"),
            },
        )

    def test_create_assets(self):
        """test to create assets with the API without authentication"""

        EndpointTestsQueries.create_object(
            self.client,
            "Assets",
            Asset,
            {
                "name": ASSET_NAME,
                "description": ASSET_DESCRIPTION,
                "folder": Folder.objects.create(name="test").id,
            },
        )

    def test_update_assets(self):
        """test to update assets with the API without authentication"""

        EndpointTestsQueries.update_object(
            self.client,
            "Assets",
            Asset,
            {
                "name": ASSET_NAME,
                "description": ASSET_DESCRIPTION,
                "folder": Folder.objects.create(name="test"),
            },
            {
                "name": "new " + ASSET_NAME,
                "description": "new " + ASSET_DESCRIPTION,
            },
        )

    def test_delete_assets(self):
        """test to delete assets with the API without authentication"""

        EndpointTestsQueries.delete_object(
            self.client,
            "Assets",
            Asset,
            {"name": ASSET_NAME, "folder": Folder.objects.create(name="test")},
        )


@pytest.mark.django_db
class TestAssetsAuthenticated:
    """Perform tests on Assets API endpoint with authentication"""

    def test_get_assets(self, test):
        """test to get assets from the API with authentication"""

        EndpointTestsQueries.Auth.get_object(
            test.client,
            "Assets",
            Asset,
            {
                "name": ASSET_NAME,
                "description": ASSET_DESCRIPTION,
                "type": ASSET_TYPE[0],
                "folder": test.folder,
            },
            {
                "folder": {"id": str(test.folder.id), "str": test.folder.name},
                "type": ASSET_TYPE[1],
            },
            user_group=test.user_group,
        )

    def test_create_assets(self, test):
        """test to create assets without a parent asset the API with authentication"""

        EndpointTestsQueries.Auth.create_object(
            test.client,
            "Assets",
            Asset,
            {
                "name": ASSET_NAME,
                "description": ASSET_DESCRIPTION,
                "type": ASSET_TYPE[0],
                "parent_assets": [],
                "folder": str(test.folder.id),
            },
            {
                "folder": {"id": str(test.folder.id), "str": test.folder.name},
                "type": ASSET_TYPE[1],
            },
            user_group=test.user_group,
            scope=str(test.folder),
        )

    def test_create_assets_with_parent(self, test):
        """test to create assets with a parent asset with the API with authentication"""

        root_asset = Asset.objects.create(
            name="root",
            description=ASSET_DESCRIPTION,
            type=ASSET_TYPE[0],
            folder=test.folder,
        )

        EndpointTestsQueries.Auth.create_object(
            test.client,
            "Assets",
            Asset,
            {
                "name": ASSET_NAME,
                "description": ASSET_DESCRIPTION,
                "type": ASSET_TYPE2[0],
                "parent_assets": [str(root_asset.id)],
                "folder": str(test.folder.id),
            },
            {
                "folder": {"id": str(test.folder.id), "str": test.folder.name},
                "type": ASSET_TYPE2[1],
                "parent_assets": [{"id": str(root_asset.id), "str": root_asset.name}],
            },
            base_count=1,
            item_search_field="name",
            user_group=test.user_group,
            scope=str(test.folder),
        )

    def test_update_assets(self, test):
        """test to update assets with the API with authentication"""

        folder = Folder.objects.create(name="test2")

        EndpointTestsQueries.Auth.update_object(
            test.client,
            "Assets",
            Asset,
            {
                "name": ASSET_NAME,
                "description": ASSET_DESCRIPTION,
                "type": ASSET_TYPE[0],
                "folder": test.folder,
            },
            {
                "name": "new " + ASSET_NAME,
                "description": "new " + ASSET_DESCRIPTION,
                "type": ASSET_TYPE2[0],
                "folder": str(folder.id),
            },
            {
                "folder": {"id": str(test.folder.id), "str": test.folder.name},
                "type": ASSET_TYPE[1],
            },
            user_group=test.user_group,
        )

    def test_delete_assets(self, test):
        """test to delete assets with the API with authentication"""

        EndpointTestsQueries.Auth.delete_object(
            test.client,
            "Assets",
            Asset,
            {"name": ASSET_NAME, "folder": test.folder},
            user_group=test.user_group,
        )

    def test_get_type_choices(self, test):
        """test to get type choices from the API with authentication"""

        EndpointTestsQueries.Auth.get_object_options(
            test.client, "Assets", "type", Asset.Type.choices
        )


# ---------------------------------------------------------------------------
# IAM-scoped visibility on /api/assets/.
#
# Locks down the queryset-level filter (`BaseModelViewSet.get_queryset`
# materialising `RoleAssignment.get_accessible_object_ids`) for a reader
# scoped to a single domain folder: an asset in a sibling folder must
# not appear in their list at all (and a fortiori not as a masked
# placeholder — masking is for related-field references, not list rows).
# ---------------------------------------------------------------------------


def _client_for(user):
    client = APIClient()
    _, token = AuthToken.objects.create(user=user)
    client.credentials(HTTP_AUTHORIZATION=f"Token {token}")
    return client


def _make_scoped_reader(folder):
    user = User.objects.create_user(
        f"reader-{uuid.uuid4().hex[:6]}@perf.test", is_published=True
    )
    role = Role.objects.get(name=RoleCodename.READER.value)
    ra = RoleAssignment.objects.create(
        user=user,
        role=role,
        folder=Folder.get_root_folder(),
        is_recursive=True,
    )
    ra.perimeter_folders.add(folder)
    return user


@pytest.mark.django_db
class TestAssetListIAMScope:
    def test_scoped_reader_does_not_see_cross_folder_asset(self, authenticated_client):
        """Reader scoped to folder A must not see an asset in folder B
        in `/api/assets/` results — that asset is outside their scope at
        the queryset level (handled by `get_accessible_object_ids`,
        before any post-filter masking)."""
        root = Folder.get_root_folder()
        folder_a = Folder.objects.create(
            name=f"perf-test-A-{uuid.uuid4().hex[:6]}",
            parent_folder=root,
            content_type=Folder.ContentType.DOMAIN,
        )
        folder_b = Folder.objects.create(
            name=f"perf-test-B-{uuid.uuid4().hex[:6]}",
            parent_folder=root,
            content_type=Folder.ContentType.DOMAIN,
        )
        asset_in_b = Asset.objects.create(
            folder=folder_b,
            name=f"asset-B-{uuid.uuid4().hex[:6]}",
            type=Asset.Type.PRIMARY,
        )

        client = _client_for(_make_scoped_reader(folder_a))
        r = client.get("/api/assets/")
        assert r.status_code == status.HTTP_200_OK, r.content
        body = r.json()
        results = body.get("results", body) if isinstance(body, dict) else body
        ids = {it["id"] for it in results}
        assert str(asset_in_b.id) not in ids, (
            f"scoped reader saw asset {asset_in_b.id} from another folder; "
            f"results: {ids}"
        )

    def _build_export_fixture(self):
        """Two folders × two asset types — A is in-scope, B is out-of-scope.
        Returns (client, assets) for the scoped reader of folder A."""
        root = Folder.get_root_folder()
        folder_a = Folder.objects.create(
            name=f"perf-test-A-{uuid.uuid4().hex[:6]}",
            parent_folder=root,
            content_type=Folder.ContentType.DOMAIN,
        )
        folder_b = Folder.objects.create(
            name=f"perf-test-B-{uuid.uuid4().hex[:6]}",
            parent_folder=root,
            content_type=Folder.ContentType.DOMAIN,
        )
        assets = {
            "a_pr": Asset.objects.create(
                folder=folder_a,
                name=f"asset-A-PR-{uuid.uuid4().hex[:6]}",
                type=Asset.Type.PRIMARY,
            ),
            "a_sp": Asset.objects.create(
                folder=folder_a,
                name=f"asset-A-SP-{uuid.uuid4().hex[:6]}",
                type=Asset.Type.SUPPORT,
            ),
            "b_pr": Asset.objects.create(
                folder=folder_b,
                name=f"asset-B-PR-{uuid.uuid4().hex[:6]}",
                type=Asset.Type.PRIMARY,
            ),
            "b_sp": Asset.objects.create(
                folder=folder_b,
                name=f"asset-B-SP-{uuid.uuid4().hex[:6]}",
                type=Asset.Type.SUPPORT,
            ),
        }
        return _client_for(_make_scoped_reader(folder_a)), assets

    def test_scoped_reader_export_csv_respects_scope_and_filter(self):
        """CSV export must honor both IAM scope and list filters: the
        scoped reader hitting /api/assets/export_csv/?type=PR sees only
        their folder's PR asset."""
        client, assets = self._build_export_fixture()
        r = client.get("/api/assets/export_csv/?type=PR")
        assert r.status_code == status.HTTP_200_OK, r.content
        assert r["Content-Type"] == "text/csv"
        content = r.content.decode("utf-8")

        assert assets["a_pr"].name in content, (
            f"scoped reader should see PR asset {assets['a_pr'].name} from their folder"
        )
        assert assets["a_sp"].name not in content, (
            f"scoped reader should not see SP asset {assets['a_sp'].name} (filtered by type)"
        )
        assert assets["b_pr"].name not in content, (
            f"scoped reader should not see PR asset {assets['b_pr'].name} from another folder"
        )
        assert assets["b_sp"].name not in content, (
            f"scoped reader should not see SP asset {assets['b_sp'].name} from another folder"
        )

    def test_scoped_reader_export_xlsx_respects_scope_and_filter(self):
        """XLSX export shares ExportMixin._get_export_queryset with CSV; this
        regression-guards against a future override that re-skips
        filter_queryset on the xlsx path."""
        import io
        from openpyxl import load_workbook

        client, assets = self._build_export_fixture()
        r = client.get("/api/assets/export_xlsx/?type=PR")
        assert r.status_code == status.HTTP_200_OK, r.content
        assert (
            r["Content-Type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        wb = load_workbook(io.BytesIO(r.content), read_only=True)
        cells = {
            str(c.value)
            for ws in wb.worksheets
            for row in ws.iter_rows()
            for c in row
            if c.value is not None
        }

        assert assets["a_pr"].name in cells
        assert assets["a_sp"].name not in cells
        assert assets["b_pr"].name not in cells
        assert assets["b_sp"].name not in cells
