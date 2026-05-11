"""Questionnaire Autopilot: parser, value mapping, agent loop, refiner.

Extracted from chat/tasks.py to keep questionnaire-specific code (parsing
xlsx, vocabulary detection, the agentic prefill loop, citation-based verdict
refinement, retry / suggest-control flows) separate from the general chat
RAG/indexing tasks. Shared text-building helpers live in ``chat/text.py``,
imported at module level — chat/tasks.py also re-exports the questionnaire
public callables, so a third module is what untangles what would otherwise
be a circular import.
"""

import structlog
import uuid

from django.core.exceptions import FieldError
from django.utils import timezone
from huey.contrib.djhuey import db_task

from .constants import Verdict
from .text import (
    _build_object_text,
    _normalize_model_name,
    _text_for_applied_control,
)

logger = structlog.get_logger(__name__)


VALUE_MAPPING_PROMPT = """A customer security questionnaire uses a controlled \
vocabulary in its answer column: {candidates}.

Map each of these three internal status labels to whichever customer value \
best expresses the same idea:
- "yes": we fully meet the requirement
- "partial": we partially meet the requirement
- "no": we do not meet the requirement

IMPORTANT: do NOT include "needs_info" in your mapping. When the agent cannot \
answer, the cell will be left blank for human review — never auto-mapped to \
"N/A" or "Not Applicable" (which is a legitimate compliance answer meaning \
the requirement does not apply, not "I don't know").

Each mapped value MUST be exactly one of the strings from the customer list \
(verbatim). If no customer value reasonably expresses one of the labels, \
reuse the closest one.

Reply with a single JSON object, no prose:
{{"yes": "...", "partial": "...", "no": "..."}}"""


@db_task()
def suggest_value_mapping(run_id: str):
    """Compute per-question value mappings for a run.

    Walks the run's questions, groups them by their detected
    ``answer_candidates`` signature, runs the LLM once per distinct
    vocabulary, and persists each question's ``answer_mapping``. Also fills
    ``QuestionnaireRun.value_mapping`` with a summary (the dominant
    vocabulary, or a flag indicating multiple) so the UI can hint at what
    the export will write.
    """
    from .models import QuestionnaireRun, QuestionnaireQuestion
    from .providers import get_llm

    try:
        run = QuestionnaireRun.objects.get(id=run_id)
    except QuestionnaireRun.DoesNotExist:
        logger.error("QuestionnaireRun %s not found", run_id)
        return

    questions = list(QuestionnaireQuestion.objects.filter(questionnaire_run=run))
    if not questions:
        logger.info(
            "suggest_value_mapping: no questions for run %s yet — nothing to map",
            run_id,
        )
        return

    # Group questions by their candidate vocabulary. Empty candidates means
    # free-text answer cell — those questions just use internal labels.
    groups: dict[tuple[str, ...], list[QuestionnaireQuestion]] = {}
    for q in questions:
        sig = tuple(q.answer_candidates or [])
        groups.setdefault(sig, []).append(q)

    try:
        llm = get_llm()
    except Exception as e:
        logger.warning("LLM unavailable for value mapping: %s", e)
        llm = None

    summary_mappings: list[dict] = []
    for sig, group_questions in groups.items():
        if not sig:
            # Free-text questions — use fallback internal labels
            mapping_for_group = {
                "yes": "Yes",
                "partial": "Partial",
                "no": "No",
                "candidates": [],
                "source": "fallback",
            }
        else:
            candidates = list(sig)
            mapping_for_group = _llm_map_candidates(llm, candidates)

        # Persist on each question in the group
        ids = [q.id for q in group_questions]
        QuestionnaireQuestion.objects.filter(id__in=ids).update(
            answer_mapping=mapping_for_group
        )
        summary_mappings.append(
            {**mapping_for_group, "question_count": len(group_questions)}
        )

    # Build the run-level summary. Most common vocab wins for the UI hint.
    summary_mappings.sort(key=lambda m: m.get("question_count", 0), reverse=True)
    primary = summary_mappings[0] if summary_mappings else {}

    distinct_non_fallback = sum(
        1 for m in summary_mappings if m.get("source") not in (None, "fallback")
    )
    has_multiple_distinct_vocabs = (
        sum(1 for m in summary_mappings if m.get("candidates")) > 1
    )

    run_summary = {
        "yes": primary.get("yes", "Yes"),
        "partial": primary.get("partial", "Partial"),
        "no": primary.get("no", "No"),
        "candidates": primary.get("candidates", []),
        "source": primary.get("source", "fallback"),
        "vocab_count": len(summary_mappings),
        "has_multiple_vocabs": has_multiple_distinct_vocabs,
    }
    run.value_mapping = run_summary
    run.save(update_fields=["value_mapping", "updated_at"])

    logger.info(
        "Value mapping done for run %s: %d question(s) across %d vocab(s) "
        "(%d non-fallback)",
        run_id,
        len(questions),
        len(summary_mappings),
        distinct_non_fallback,
    )


def _llm_map_candidates(llm, candidates: list[str]) -> dict:
    """LLM-driven mapping of yes/partial/no onto a candidate list, with retries.

    Returns the mapping dict to store on a question (or all questions sharing
    this vocabulary). Falls back to internal labels with ``source='fallback'``
    if the LLM is unavailable or doesn't produce a valid mapping after retries.
    """
    import json
    import time

    if llm is None or not candidates:
        return {
            "yes": "Yes",
            "partial": "Partial",
            "no": "No",
            "candidates": candidates,
            "source": "fallback",
        }

    MAX_ATTEMPTS = 3
    BACKOFF_SECONDS = (0, 2, 5)
    candidate_set = {c.lower(): c for c in candidates}

    def pick(parsed: dict, label: str) -> str | None:
        v = parsed.get(label)
        if not isinstance(v, str):
            return None
        if v in candidates:
            return v
        return candidate_set.get(v.lower())

    last_raw = ""
    for attempt in range(MAX_ATTEMPTS):
        if attempt > 0:
            time.sleep(BACKOFF_SECONDS[attempt])

        prompt = VALUE_MAPPING_PROMPT.format(
            candidates=json.dumps(candidates, ensure_ascii=False)
        )
        try:
            raw = llm.generate(prompt=prompt, context="", history=[])
        except Exception as e:
            logger.warning(
                "LLM call failed for value mapping (attempt %d/%d): %s",
                attempt + 1,
                MAX_ATTEMPTS,
                e,
            )
            continue

        last_raw = raw or ""
        parsed = _parse_json_response(raw) or {}
        yes = pick(parsed, "yes")
        partial = pick(parsed, "partial")
        no = pick(parsed, "no")
        if all([yes, partial, no]):
            # We deliberately do NOT include needs_info — export leaves those
            # cells blank rather than risk auto-mapping to "N/A" (which is a
            # legitimate "Not Applicable" answer, not "I don't know").
            return {
                "yes": yes,
                "partial": partial,
                "no": no,
                "candidates": candidates,
                "source": "data_validation",
            }
        logger.info(
            "LLM produced incomplete mapping (attempt %d/%d, raw=%r) — retrying",
            attempt + 1,
            MAX_ATTEMPTS,
            (raw or "")[:200],
        )

    logger.warning(
        "Value mapping fallback after %d attempts (last raw=%r)",
        MAX_ATTEMPTS,
        last_raw[:200],
    )
    return {
        "yes": "Yes",
        "partial": "Partial",
        "no": "No",
        "candidates": candidates,
        "source": "fallback",
    }


def extract_questions_from_sheet(run, sheet: dict, mapping: dict) -> int:
    """Materialize QuestionnaireQuestion rows from a sheet preview + mapping.

    Synchronous helper, called from the API. Walks the workbook to capture
    every question-bearing row. For each row, also detects the answer cell's
    data-validation vocabulary so a per-question mapping can be computed
    (some questionnaires use different dropdowns per row range).
    Returns the count of created questions.
    """
    import io
    import openpyxl

    from .models import QuestionnaireQuestion

    q_col = mapping["question_col"]
    answer_col = mapping.get("answer_col")
    section_col = mapping.get("section_col")
    sheet_name = sheet["name"]
    header_row = sheet["header_row"]

    with run.file.open("rb") as fp:
        content = fp.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb[sheet_name]
    header_count = len(sheet["headers"])

    rows_to_create = []
    ord_idx = 0
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx <= header_row:
            continue
        cells = list(row[:header_count])
        if all(c in (None, "") for c in cells):
            continue
        text = cells[q_col] if q_col < len(cells) else None
        if text in (None, ""):
            ord_idx += 1
            continue
        text = str(text).strip()
        if not text:
            ord_idx += 1
            continue
        section = ""
        if section_col is not None and section_col < len(cells):
            section_val = cells[section_col]
            if section_val not in (None, ""):
                section = str(section_val).strip()[:200]

        ref_id = ""

        candidates: list[str] = []
        if answer_col is not None:
            # openpyxl is 1-indexed; idx is 0-indexed.
            candidates = _detect_cell_vocabulary(ws, idx + 1, answer_col)

        rows_to_create.append(
            QuestionnaireQuestion(
                questionnaire_run=run,
                ord=ord_idx,
                ref_id=ref_id,
                section=section,
                text=text,
                answer_candidates=candidates,
            )
        )
        ord_idx += 1

    QuestionnaireQuestion.objects.bulk_create(rows_to_create)
    return len(rows_to_create)


def _detect_cell_vocabulary(ws, excel_row: int, col_idx: int) -> list[str]:
    """Return the data-validation list values that apply to one cell, or [].

    Scans ``ws.data_validations`` for a list-type validation whose ``sqref``
    covers the cell at (excel_row, col_idx). Returns the verbatim values from
    the validation's formula1 (inline or range-referenced).
    """
    import openpyxl.utils

    col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
    coord = f"{col_letter}{excel_row}"
    try:
        for dv in ws.data_validations.dataValidation:
            if getattr(dv, "type", None) != "list":
                continue
            sqref = getattr(dv, "sqref", None)
            try:
                covers = bool(sqref) and coord in sqref
            except Exception:
                covers = False
            if not covers:
                continue
            values = _parse_data_validation_list(ws, dv.formula1)
            if values:
                return values
    except Exception as e:
        logger.debug("cell_vocabulary_detection_failed %s: %s", coord, e)
    return []


@db_task()
def parse_questionnaire(run_id: str):
    """Open the uploaded xlsx, capture sheet/header structure for review.

    No interpretation of which column is question/answer — that's the user's
    call in the mapping UI. We only surface what the file actually contains.
    """
    import io
    import openpyxl

    from .models import QuestionnaireRun

    try:
        run = QuestionnaireRun.objects.get(id=run_id)
    except QuestionnaireRun.DoesNotExist:
        logger.error("QuestionnaireRun %s not found", run_id)
        return

    run.status = QuestionnaireRun.Status.PARSING
    run.save(update_fields=["status"])

    try:
        with run.file.open("rb") as fp:
            content = fp.read()

        # Not read_only: we need access to data_validations (Excel dropdowns)
        # which are only exposed on a fully-loaded workbook.
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

        sheets = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_row_idx, headers = _detect_headers(ws)
            row_count, rows_preview = _collect_rows(ws, header_row_idx, len(headers))
            column_candidates = _detect_all_column_candidates(
                ws, headers, header_row_idx
            )
            sheets.append(
                {
                    "name": sheet_name,
                    "header_row": header_row_idx,
                    "headers": headers,
                    "row_count": row_count,
                    "rows_preview": rows_preview,
                    "column_candidates": column_candidates,
                }
            )

        active_sheet = next(
            (s["name"] for s in sheets if s["headers"] and s["row_count"] > 0),
            sheets[0]["name"] if sheets else "",
        )

        run.parsed_data = {"sheets": sheets, "active_sheet": active_sheet}
        run.status = QuestionnaireRun.Status.PARSED
        run.save(update_fields=["parsed_data", "status"])
        logger.info("Parsed questionnaire %s: %d sheet(s)", run.filename, len(sheets))

    except Exception as e:
        logger.error("Failed to parse questionnaire %s: %s", run_id, e)
        run.status = QuestionnaireRun.Status.FAILED
        run.error_message = str(e)
        run.save(update_fields=["status", "error_message"])


def _detect_headers(ws):
    """First row with at least 2 non-empty cells is treated as the header row.

    Scans up to the first 20 rows; returns (row_index_0_based, headers_list).
    Empty sheet -> (0, []).
    """
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True)):
        non_empty = [c for c in row if c not in (None, "")]
        if len(non_empty) >= 2:
            headers = [
                str(c).strip() if c not in (None, "") else f"(col {i + 1})"
                for i, c in enumerate(row)
            ]
            while headers and headers[-1].startswith("(col "):
                headers.pop()
            return idx, headers
    return 0, []


def _collect_rows(ws, header_row_idx: int, header_count: int, preview_size: int = 20):
    """Return (total_data_row_count, first_N_rows_as_list_of_lists)."""
    if header_count == 0:
        return 0, []
    preview = []
    total = 0
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx <= header_row_idx:
            continue
        cells = list(row[:header_count])
        if all(c in (None, "") for c in cells):
            continue
        total += 1
        if len(preview) < preview_size:
            preview.append(["" if c is None else str(c) for c in cells])
    return total, preview


def _detect_all_column_candidates(ws, headers: list, header_row_idx: int) -> dict:
    """Per column, find the controlled vocabulary the customer expects.

    Returns {col_idx: {"values": [...], "source": "data_validation"|"distinct_values"}}.
    Columns with no controlled vocabulary (free text, single value, too many
    distinct values) are simply absent from the dict — that's the signal to
    fall back to internal labels later.
    """
    out: dict = {}
    if not headers:
        return out
    for col_idx in range(len(headers)):
        candidates = _detect_column_candidates(ws, col_idx, header_row_idx)
        if candidates:
            out[col_idx] = candidates
    return out


def _detect_column_candidates(
    ws, col_idx: int, header_row_idx: int, max_distinct: int = 20
) -> dict | None:
    """Return {"values": [...], "source": "..."} for a single column, or None."""
    import openpyxl.utils

    target_col_letter = openpyxl.utils.get_column_letter(col_idx + 1)

    # 1. Excel data validation list constraint — the customer's authored vocab.
    try:
        for dv in ws.data_validations.dataValidation:
            if getattr(dv, "type", None) != "list":
                continue
            # Probe two cells inside this column to see if the validation applies
            probe_cells = [
                f"{target_col_letter}{header_row_idx + 2}",
                f"{target_col_letter}{header_row_idx + 3}",
            ]
            sqref = getattr(dv, "sqref", None)
            try:
                covers = any(c in sqref for c in probe_cells) if sqref else False
            except Exception:
                covers = False
            if not covers:
                continue
            values = _parse_data_validation_list(ws, dv.formula1)
            if values:
                return {"values": values, "source": "data_validation"}
    except Exception as e:
        logger.debug("data_validation_check_failed col=%s: %s", col_idx, e)

    # 2. Distinct existing values in the column.
    distinct: list[str] = []
    seen: set[str] = set()
    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx <= header_row_idx:
            continue
        if col_idx >= len(row):
            continue
        cell = row[col_idx]
        if cell in (None, ""):
            continue
        s = str(cell).strip()
        if not s:
            continue
        if s in seen:
            continue
        seen.add(s)
        distinct.append(s)
        if len(distinct) > max_distinct:
            return None  # likely free text — not a controlled vocab
    if 2 <= len(distinct) <= max_distinct:
        return {"values": distinct, "source": "distinct_values"}
    return None


def _parse_data_validation_list(ws, formula1) -> list:
    """Best-effort parsing of the formula1 of a list-type data validation.

    Inline:    '"Yes,No,Partial"' → ['Yes', 'No', 'Partial']
    Range ref: 'Sheet1!$A$1:$A$5' → values from that range
    """
    if formula1 is None:
        return []
    s = str(formula1).strip()
    if not s:
        return []
    # Inline quoted list
    if s.startswith('"') and s.endswith('"'):
        inner = s[1:-1]
        return [v.strip() for v in inner.split(",") if v.strip()]
    # Range reference, optionally cross-sheet
    try:
        if "!" in s:
            sheet_name, range_str = s.split("!", 1)
            sheet_name = sheet_name.strip().strip("'")
            target_ws = ws.parent[sheet_name]
        else:
            target_ws = ws
            range_str = s
        range_str = range_str.replace("$", "")
        cells = target_ws[range_str]
        values: list[str] = []
        # cells can be a single cell, a tuple (1D range), or tuple-of-tuples (2D)
        if not isinstance(cells, tuple):
            cells = (cells,)
        for row in cells:
            if isinstance(row, tuple):
                for c in row:
                    v = getattr(c, "value", None)
                    if v not in (None, ""):
                        values.append(str(v).strip())
            else:
                v = getattr(row, "value", None)
                if v not in (None, ""):
                    values.append(str(v).strip())
        return values
    except Exception as e:
        logger.debug("data_validation_range_parse_failed formula=%s: %s", s, e)
        return []


ANSWER_PROMPT = """You are answering a customer security questionnaire on \
behalf of an organization. Use ONLY the provided context to write your answer. \
If the context does not support an answer, set status to "needs_info".

What counts as evidence (in order of priority):

1. **Applied controls** — the verdict is in their **Status** + **Observation**. \
Other fields (name, reference, description) describe what the control is, not \
whether we do it.
2. **Requirement assessments** — the verdict is in their **Result** + \
**Observation**. The requirement ref / requirement text / framework name \
attached to an assessment describe what's being assessed (the question, in \
effect); they are NOT evidence we comply. Cite an RA only when its Result \
genuinely supports the answer.

How to map evidence to verdict:

- **Status: Active** (applied control) or **Result: Compliant** (requirement \
assessment) — real, currently-in-place evidence; can support "yes".
- **Status: In progress** or **Result: Partially compliant** — these support \
**"partial"**, not "needs_info". If you cite an in-progress / partially \
compliant item, the answer is at minimum "partial" — explain in the comment \
what's in place and what's missing rather than hedging to needs_info.
- **Status: To do**, **Status: On hold**, **Result: Non compliant**, \
**Result: Not assessed** — control is planned, paused, or unverified; do NOT \
justify "yes". Use "partial" only if there is also active/compliant evidence; \
otherwise prefer "needs_info" or "no".
- **Status: Deprecated** — ignore as evidence.
- A compliance assessment with **Status: In progress** or **In review** means \
we are actively working through that framework — supports "partial" for \
"do you have a process/policy" questions; for "do you HOLD certificate X" \
questions, it supports "no" (we're working toward it but don't hold it yet).
- "needs_info" is correct when **no** relevant evidence is cited, or when the \
question asks for a specific value (e.g. exact RTO, exact frequency) that the \
cited evidence doesn't pin down — but if any cited evidence is at least \
in-progress / partially compliant, prefer "partial" + explain the gap in the \
comment.

Question:
{question}

{section_hint}Context (numbered passages):
{context}

Reply with a single JSON object, no prose around it:
{{
  "status": "yes" | "no" | "partial" | "needs_info",
  "comment": "<one or two sentences answering the question, with citation \
markers like [1] [2] referring to the numbered context passages>",
  "citation_indices": [<1-based indices of passages used>]
}}"""

CRITIC_PROMPT = """You are reviewing a draft answer to a customer security \
questionnaire. Score how well the cited context supports the answer (0.0 = \
contradicted or unsupported, 1.0 = fully supported and on-topic).

Pay particular attention to the status of each cited passage:

- A "yes" verdict requires at least one cited passage with **Status: Active** \
or **Result: Compliant**. If the draft says "yes" while citing only "To do", \
"On hold", "In progress", "Non compliant", "Not assessed" or "Deprecated" \
items, score below 0.4 and explain that planned/non-active items do not \
support "yes".
- A "partial" verdict needs evidence that something is actually in place \
(active/compliant or in-progress) — it should not be used to dress up a "no".
- "needs_info" is correct when no relevant active/compliant evidence is \
cited; do not penalize it for that reason alone.

Question:
{question}

Draft answer (status={status}):
{comment}

Cited context:
{cited_context}

Reply with a single JSON object, no prose around it:
{{
  "score": <float between 0.0 and 1.0>,
  "issue": "<one short sentence describing the main weakness if score < 0.8, otherwise empty>"
}}"""

# Tunable thresholds — defaults live in ciso_assistant/settings.py and can be
# overridden via env (QUESTIONNAIRE_RETRY_THRESHOLD, etc.).
from django.conf import settings as _django_settings  # noqa: E402

# Score below this triggers a retry in Thorough mode.
RETRY_THRESHOLD = float(getattr(_django_settings, "QUESTIONNAIRE_RETRY_THRESHOLD", 0.7))
# Critic-less floor: assume mid confidence so review UI orders sensibly.
FAST_MODE_DEFAULT_CONFIDENCE = float(
    getattr(_django_settings, "QUESTIONNAIRE_FAST_MODE_DEFAULT_CONFIDENCE", 0.5)
)
# Per-question wall-clock cap so the loop can't hang on one item.
PER_QUESTION_TIMEOUT_SEC = int(
    getattr(_django_settings, "QUESTIONNAIRE_PER_QUESTION_TIMEOUT_SEC", 90)
)


def _extract_status_marker(text: str) -> str:
    """Pull the verdict-bearing field out of an indexed passage's text body.

    The indexer (see _text_for_applied_control / _text_for_requirement_assessment)
    writes lines like "Status: To do" or "Result: Partially compliant".
    Surfacing this in the context label makes it visible at a glance to the
    LLM rather than buried mid-paragraph. Returns e.g. "status=to_do" or
    "result=partially_compliant", or "" when neither marker is present.
    """
    if not text:
        return ""
    # Result wins over Status when both appear (RA carries both; Result is
    # the verdict-bearing one).
    result_value = ""
    status_value = ""
    for line in text.split("\n"):
        stripped = line.strip()
        lower = stripped.lower()
        if not result_value and lower.startswith("result:"):
            result_value = stripped.split(":", 1)[1].strip()
        elif not status_value and lower.startswith("status:"):
            status_value = stripped.split(":", 1)[1].strip()
    chosen = result_value or status_value
    if not chosen:
        return ""
    key = "result" if result_value else "status"
    normalized = chosen.lower().replace(" ", "_").replace("-", "_")
    return f"{key}={normalized}"


def _search_folder_evidence(query: str, folder_id: str, top_k: int = 6) -> list[dict]:
    """Folder-scoped, library-excluded RAG search for the questionnaire agent.

    Differences vs ``chat.rag.search``:
      - Hard-scoped to ONE folder (the questionnaire's). The general search
        scopes by the *user's accessible perimeter* — that pulls in evidence
        from sibling folders (e.g. Starter MFA assessments leaking into a
        DEMO run). For prefill we want only the questionnaire's folder.
      - Excludes the library partition. RequirementNodes describe what
        *should* be done, not what we *do*; they're metadata, not verdict
        evidence. Including them in context invites the LLM to over-weight
        framework text it can't use to judge compliance.

    Returns the same flat-dict shape as ``chat.rag.search`` so existing
    callers (``_build_context_block``) consume it unchanged.
    """
    from qdrant_client.models import FieldCondition, Filter, MatchValue

    from .providers import get_embedder
    from .rag import COLLECTION_NAME, _get_reranker, get_qdrant_client

    try:
        client = get_qdrant_client()
        embedder = get_embedder()
    except Exception as e:
        logger.warning("folder_evidence_search: infra unavailable (%s)", e)
        return []

    folder_id_str = str(folder_id)
    fetch_limit = max(top_k * 3, top_k)

    try:
        query_vector = embedder.embed_query(query)
    except Exception as e:
        logger.warning("folder_evidence_search: embed_query failed (%s)", e)
        return []

    qfilter = Filter(
        must=[FieldCondition(key="folder_id", match=MatchValue(value=folder_id_str))]
    )

    try:
        result = client.query_points(
            collection_name=COLLECTION_NAME,
            query=query_vector,
            limit=fetch_limit,
            query_filter=qfilter,
        )
        candidates = list(result.points)
    except Exception as e:
        logger.warning("folder_evidence_search: query failed (%s)", e)
        return []

    # Cross-encoder rerank when we have more than we need.
    reranker = _get_reranker()
    if reranker and len(candidates) > top_k:
        try:
            pairs = [
                (query, (c.payload or {}).get("text", "")[:512]) for c in candidates
            ]
            scores = reranker.predict(pairs)
            candidates = [
                p
                for _, p in sorted(
                    zip(scores, candidates), key=lambda x: x[0], reverse=True
                )[:top_k]
            ]
        except Exception as e:
            logger.warning("folder_evidence_search: rerank failed (%s)", e)
            candidates = candidates[:top_k]
    else:
        candidates = candidates[:top_k]

    return [
        {
            "id": str(c.id),
            "score": getattr(c, "score", 0.0),
            "text": (c.payload or {}).get("text", ""),
            "source_type": (c.payload or {}).get("source_type", ""),
            "object_type": (c.payload or {}).get("object_type", ""),
            "object_id": (c.payload or {}).get("object_id"),
            "name": (c.payload or {}).get("name", ""),
            "ref_id": (c.payload or {}).get("ref_id", ""),
            "framework": (c.payload or {}).get("framework", ""),
            "urn": (c.payload or {}).get("urn", ""),
        }
        for c in candidates
    ]


def _build_context_block(results: list[dict]) -> tuple[str, list[dict]]:
    """Format retrieval results as numbered passages + return source_refs.

    `chat.rag.search` returns flat dicts with fields at the top level
    (id, score, text, source_type, object_type, name, ref_id, …).
    Don't try to read them from a nested "payload" key — there isn't one.

    The status / result of status-bearing models (AppliedControl,
    RequirementAssessment) is lifted into the visible label so the LLM
    sees `status=to_do` right next to the passage marker, not mid-text.
    """
    lines = []
    refs = []
    for i, r in enumerate(results, start=1):
        score = r.get("score")
        text = r.get("text", "") or ""
        snippet = text.strip().replace("\n", " ")
        if len(snippet) > 400:
            snippet = snippet[:400] + "…"
        name = r.get("name") or ""
        object_type = r.get("object_type") or ""
        source_type = r.get("source_type") or ""
        ref_id = r.get("ref_id") or ""
        status_marker = _extract_status_marker(text)
        label_bits = []
        if object_type:
            label_bits.append(object_type)
        if ref_id:
            label_bits.append(ref_id)
        elif name:
            label_bits.append(name[:60])
        if status_marker:
            label_bits.append(status_marker)
        label = " · ".join(label_bits) or "passage"
        lines.append(f"[{i}] ({label}) {snippet}")
        # ``source_type`` is the broad category ("model" / "document"); for
        # model rows the concrete kind (applied_control, evidence, …) lives
        # in ``object_type``. The refiner + frontend citation linker key on
        # the concrete kind, so prefer object_type when source_type is
        # "model" and fall back to source_type otherwise (e.g. "document").
        ref_kind = (
            object_type if source_type == "model" else (source_type or object_type)
        )
        refs.append(
            {
                "index": i,
                "kind": ref_kind or "",
                "id": str(r.get("object_id") or r.get("id") or ""),
                "name": name or object_type or "",
                "ref_id": ref_id,
                "score": float(score) if score is not None else None,
                "snippet": snippet,
                "status_marker": status_marker,
            }
        )
    return "\n".join(lines), refs


def _parse_json_response(text: str) -> dict | None:
    """Pull a JSON object out of an LLM response. Tolerant of code fences and prose."""
    import json
    import re

    if not text:
        return None
    # Strip code fences
    fence = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", text, re.DOTALL)
    if fence:
        try:
            return json.loads(fence.group(1))
        except json.JSONDecodeError:
            # Fenced block was malformed — fall through to the broader
            # top-level {...} regex below.
            pass
    # First top-level {...} block
    match = re.search(r"\{.*\}", text, re.DOTALL)
    if match:
        try:
            return json.loads(match.group(0))
        except Exception:
            return None
    return None


def _heartbeat(run, label: str | None = None) -> bool:
    """Update heartbeat. Returns False if the run was cancelled — caller should bail."""
    from .models import AgentRun

    fresh = AgentRun.objects.only("status").get(id=run.id)
    if fresh.status == AgentRun.Status.CANCELLED:
        return False
    update_fields = ["last_heartbeat_at"]
    run.last_heartbeat_at = timezone.now()
    if label is not None:
        run.current_step_label = label[:300]
        update_fields.append("current_step_label")
    run.save(update_fields=update_fields + ["updated_at"])
    return True


def refresh_folder_index(folder_id: str) -> dict:
    """Prune stale model points + (re)index live model objects in a folder.

    Run as the first step of a questionnaire prefill so the LLM only sees
    citations that resolve to real, current objects. Idempotent —
    embed+upsert overwrites by deterministic point id.

    Returns ``{"pruned": <count>, "indexed": <count>}``.
    """
    import uuid as _uuid

    from django.apps import apps
    from qdrant_client.models import (
        FieldCondition,
        Filter,
        MatchValue,
        PointStruct,
    )

    from .providers import get_embedder
    from .rag import COLLECTION_NAME, get_qdrant_client
    from .signals import INDEXED_MODELS

    folder_id_str = str(folder_id)

    try:
        client = get_qdrant_client()
    except Exception as e:
        logger.warning("refresh_folder_index: qdrant unavailable (%s)", e)
        return {"pruned": 0, "indexed": 0, "skipped": True}

    scroll_filter = Filter(
        must=[
            FieldCondition(key="folder_id", match=MatchValue(value=folder_id_str)),
            FieldCondition(key="source_type", match=MatchValue(value="model")),
        ]
    )

    # 1. Walk every existing model point in the folder so we know which IDs
    # the index currently holds. Tuples of (qdrant_point_id, object_id).
    points_in_folder: list[tuple] = []
    next_offset = None
    while True:
        try:
            batch, next_offset = client.scroll(
                collection_name=COLLECTION_NAME,
                scroll_filter=scroll_filter,
                limit=500,
                offset=next_offset,
                with_payload=True,
                with_vectors=False,
            )
        except Exception as e:
            logger.warning("refresh_folder_index: scroll failed (%s)", e)
            break
        for p in batch or []:
            payload = getattr(p, "payload", {}) or {}
            obj_id = payload.get("object_id")
            if obj_id:
                points_in_folder.append((p.id, str(obj_id)))
        if not next_offset:
            break

    # 2. Walk the indexed model classes for live objects scoped to the folder.
    live_object_ids: set = set()
    indexable_rows: list[tuple] = []  # (app_label, model_name, obj)
    for model_path in INDEXED_MODELS:
        try:
            app_label, model_name = model_path.split(".")
            model_class = apps.get_model(app_label, model_name)
        except (LookupError, ValueError):
            continue
        try:
            qs = model_class.objects.filter(folder_id=folder_id_str)
        except FieldError:
            # Model isn't folder-scoped this way — skip; refresh covers
            # FolderMixin'd models which is what we actually need here.
            continue
        for obj in qs.iterator():
            live_object_ids.add(str(obj.id))
            indexable_rows.append((app_label, model_name, obj))

    # 3. Drop points whose object_id is no longer in the live set.
    stale_point_ids = [
        pid for (pid, oid) in points_in_folder if oid not in live_object_ids
    ]
    if stale_point_ids:
        try:
            client.delete(
                collection_name=COLLECTION_NAME,
                points_selector=stale_point_ids,
            )
        except Exception as e:
            logger.warning("refresh_folder_index: delete failed (%s)", e)

    # 4. Re-embed and upsert live objects in batches.
    if not indexable_rows:
        return {"pruned": len(stale_point_ids), "indexed": 0}

    try:
        embedder = get_embedder()
    except Exception as e:
        logger.warning("refresh_folder_index: embedder unavailable (%s)", e)
        return {"pruned": len(stale_point_ids), "indexed": 0, "skipped": True}

    BATCH_SIZE = 100
    indexed_count = 0
    pending_text: list[str] = []
    pending_meta: list[tuple] = []  # (app_label, model_name, obj)

    def _flush() -> int:
        if not pending_text:
            return 0
        embeddings = embedder.embed(pending_text)
        points = []
        for (app_label, model_name, obj), vector, text in zip(
            pending_meta, embeddings, pending_text
        ):
            point_id = str(
                _uuid.uuid5(
                    _uuid.NAMESPACE_URL,
                    f"{app_label}.{model_name}:{obj.id}",
                )
            )
            points.append(
                PointStruct(
                    id=point_id,
                    vector=vector,
                    payload={
                        "text": text,
                        "folder_id": folder_id_str,
                        "source_type": "model",
                        "object_type": _normalize_model_name(model_name),
                        "object_id": str(obj.id),
                        "name": str(obj),
                        "ref_id": getattr(obj, "ref_id", "") or "",
                    },
                )
            )
        try:
            client.upsert(collection_name=COLLECTION_NAME, points=points)
        except Exception as e:
            logger.warning("refresh_folder_index: upsert failed (%s)", e)
            pending_text.clear()
            pending_meta.clear()
            return 0
        n = len(points)
        pending_text.clear()
        pending_meta.clear()
        return n

    for app_label, model_name, obj in indexable_rows:
        text = _build_object_text(obj, model_name)
        if not text:
            continue
        pending_text.append(text)
        pending_meta.append((app_label, model_name, obj))
        if len(pending_text) >= BATCH_SIZE:
            indexed_count += _flush()
    indexed_count += _flush()

    return {"pruned": len(stale_point_ids), "indexed": indexed_count}


@db_task()
def run_questionnaire_prefill(agent_run_id: str):
    """The agentic loop. Per question:
        retrieve → answer → (critic + retry on Thorough) → record AgentAction.
    Heartbeats between each question; checks cancellation each iteration.
    """
    import time

    from django.contrib.contenttypes.models import ContentType

    from .models import AgentRun, AgentAction, QuestionnaireQuestion
    from .providers import get_llm, get_chat_settings
    from .rag import search as rag_search
    from .tokens import count_tokens

    try:
        run = AgentRun.objects.select_related("owner", "folder").get(id=agent_run_id)
    except AgentRun.DoesNotExist:
        logger.error("AgentRun %s not found", agent_run_id)
        return

    if run.status != AgentRun.Status.QUEUED:
        logger.warning("AgentRun %s is %s; skipping", agent_run_id, run.status)
        return

    settings_dict = get_chat_settings()
    provider = settings_dict.get("llm_provider", "ollama")
    if provider == "openai_compatible":
        model_name = settings_dict.get("openai_model") or "openai_compatible"
    else:
        model_name = settings_dict.get("ollama_model") or "ollama"

    run.status = AgentRun.Status.RUNNING
    run.started_at = timezone.now()
    run.last_heartbeat_at = timezone.now()
    run.model_used = model_name
    run.error_message = ""
    run.save(
        update_fields=[
            "status",
            "started_at",
            "last_heartbeat_at",
            "model_used",
            "error_message",
            "updated_at",
        ]
    )

    # Refresh the folder's vector index before answering. Two reasons:
    # (1) Prune stale points whose underlying object was deleted — the LLM
    # would otherwise cite ghost evidence the post-processor can't verify;
    # (2) Pick up controls/assessments created since the last index pass.
    run.current_step_label = "Refreshing folder index…"
    run.last_heartbeat_at = timezone.now()
    run.save(update_fields=["current_step_label", "last_heartbeat_at", "updated_at"])
    try:
        refresh_stats = refresh_folder_index(str(run.folder_id))
        logger.info(
            "Folder index refreshed for run %s: pruned=%d indexed=%d",
            run.id,
            refresh_stats.get("pruned", 0),
            refresh_stats.get("indexed", 0),
        )
    except Exception as e:
        logger.warning("Folder index refresh failed for run %s: %s", run.id, e)

    questions = list(
        QuestionnaireQuestion.objects.filter(
            questionnaire_run_id=run.target_object_id
        ).order_by("ord")
    )
    run.total_steps = len(questions)
    run.completed_steps = 0
    run.save(update_fields=["total_steps", "completed_steps", "updated_at"])

    qq_ct = ContentType.objects.get_for_model(QuestionnaireQuestion)
    is_thorough = run.strictness == AgentRun.Strictness.THOROUGH
    max_retries = 1 if is_thorough else 0

    try:
        llm = get_llm()
    except Exception as e:
        run.status = AgentRun.Status.FAILED
        run.error_message = f"LLM unavailable: {e}"
        run.finished_at = timezone.now()
        run.save(update_fields=["status", "error_message", "finished_at", "updated_at"])
        return

    # Tolerate the occasional flaky question, but bail loudly when the
    # whole infrastructure is broken (e.g. Qdrant dies mid-run). Without
    # this guard, every question would stamp a fake `needs_info` proposal
    # and the user would see a complete-looking run full of wrong answers.
    _CONSECUTIVE_FAILURE_LIMIT = 3
    consecutive_failures = 0
    last_failure_message = ""

    try:
        for question in questions:
            label = f"{run.completed_steps + 1}/{run.total_steps}: {question.text[:80]}"
            if not _heartbeat(run, label):
                # Cancelled mid-flight
                return

            t0 = time.time()
            try:
                _process_question(
                    run=run,
                    question=question,
                    qq_ct=qq_ct,
                    llm=llm,
                    max_retries=max_retries,
                    rag_search=rag_search,
                    count_tokens=count_tokens,
                )
                consecutive_failures = 0
            except Exception as e:
                consecutive_failures += 1
                last_failure_message = str(e)
                logger.error(
                    "Failed processing question %s in run %s (consecutive=%d): %s",
                    question.id,
                    run.id,
                    consecutive_failures,
                    e,
                )
                # Record the failure as an AgentAction so it shows in the UI;
                # don't kill the whole run for one bad question.
                AgentAction.objects.create(
                    agent_run=run,
                    kind=AgentAction.Kind.PROPOSE_ANSWER,
                    target_content_type=qq_ct,
                    target_object_id=question.id,
                    payload={"status": Verdict.NEEDS_INFO, "comment": ""},
                    rationale=f"Error: {e}",
                    source_refs=[],
                    confidence=0.0,
                    state=AgentAction.State.PROPOSED,
                    iteration=0,
                    duration_ms=int((time.time() - t0) * 1000),
                )
                if consecutive_failures >= _CONSECUTIVE_FAILURE_LIMIT:
                    run.status = AgentRun.Status.FAILED
                    run.finished_at = timezone.now()
                    run.error_message = (
                        f"Aborting after {consecutive_failures} consecutive "
                        f"question failures. Last error: {last_failure_message}"
                    )
                    run.current_step_label = ""
                    run.save(
                        update_fields=[
                            "status",
                            "finished_at",
                            "error_message",
                            "current_step_label",
                            "updated_at",
                        ]
                    )
                    logger.error(
                        "AgentRun %s aborted: %s",
                        run.id,
                        run.error_message,
                    )
                    return

            # Soft per-question budget guard
            if time.time() - t0 > PER_QUESTION_TIMEOUT_SEC:
                logger.warning(
                    "Question %s exceeded soft budget (%ss)",
                    question.id,
                    PER_QUESTION_TIMEOUT_SEC,
                )

            run.completed_steps += 1
            run.last_heartbeat_at = timezone.now()
            run.save(
                update_fields=["completed_steps", "last_heartbeat_at", "updated_at"]
            )

        run.status = AgentRun.Status.SUCCEEDED
        run.finished_at = timezone.now()
        run.current_step_label = ""
        run.save(
            update_fields=[
                "status",
                "finished_at",
                "current_step_label",
                "updated_at",
            ]
        )
        logger.info(
            "AgentRun %s completed: %d/%d", run.id, run.completed_steps, run.total_steps
        )

    except Exception as e:
        logger.error("AgentRun %s failed: %s", run.id, e)
        run.status = AgentRun.Status.FAILED
        run.error_message = str(e)
        run.finished_at = timezone.now()
        run.save(update_fields=["status", "error_message", "finished_at", "updated_at"])


def _process_question(
    *,
    run,
    question,
    qq_ct,
    llm,
    max_retries: int,
    rag_search,
    count_tokens,
):
    """Single-question pipeline: retrieve → answer → optional critic + retry."""
    import time

    from .models import AgentAction

    # --- Retrieve ---
    t0 = time.time()
    query = question.text
    try:
        results = _search_folder_evidence(query, run.folder_id, top_k=6)
    except Exception as e:
        logger.error("Folder-scoped search failed for q %s: %s", question.id, e)
        results = []
    duration_ms = int((time.time() - t0) * 1000)

    context_block, source_refs = _build_context_block(results)

    AgentAction.objects.create(
        agent_run=run,
        kind=AgentAction.Kind.RETRIEVE,
        target_content_type=qq_ct,
        target_object_id=question.id,
        payload={"top_k": len(results)},
        rationale=f"RAG search returned {len(results)} passages.",
        source_refs=source_refs,
        state=AgentAction.State.PROPOSED,
        duration_ms=duration_ms,
    )

    # --- Answer (iteration 0) ---
    proposed = _answer_iteration(
        run=run,
        question=question,
        qq_ct=qq_ct,
        llm=llm,
        context_block=context_block,
        source_refs=source_refs,
        critic_hint="",
        iteration=0,
        count_tokens=count_tokens,
    )

    confidence = proposed["confidence"]

    # --- Critic + retry (Thorough only) ---
    if max_retries > 0:
        critic = _critique(
            run=run,
            question=question,
            qq_ct=qq_ct,
            llm=llm,
            answer_payload=proposed["payload"],
            context_block=context_block,
            cited_indices=proposed["cited_indices"],
            iteration=0,
            count_tokens=count_tokens,
        )
        confidence = critic["score"]

        if confidence < RETRY_THRESHOLD:
            retry = _answer_iteration(
                run=run,
                question=question,
                qq_ct=qq_ct,
                llm=llm,
                context_block=context_block,
                source_refs=source_refs,
                critic_hint=critic["issue"],
                iteration=1,
                count_tokens=count_tokens,
            )
            critic2 = _critique(
                run=run,
                question=question,
                qq_ct=qq_ct,
                llm=llm,
                answer_payload=retry["payload"],
                context_block=context_block,
                cited_indices=retry["cited_indices"],
                iteration=1,
                count_tokens=count_tokens,
            )
            # Use whichever iteration scored higher. Expire the loser so the
            # review UI (which picks the latest non-expired iteration) shows
            # the better answer — not the most recent one.
            loser_iteration = 0 if critic2["score"] >= confidence else 1
            if critic2["score"] >= confidence:
                proposed = retry
                confidence = critic2["score"]
            AgentAction.objects.filter(
                agent_run=run,
                kind=AgentAction.Kind.PROPOSE_ANSWER,
                target_content_type=qq_ct,
                target_object_id=question.id,
                iteration=loser_iteration,
            ).update(state=AgentAction.State.EXPIRED)

    # Update final proposal with critic-derived confidence
    AgentAction.objects.filter(id=proposed["action_id"]).update(confidence=confidence)


SUGGEST_CONTROL_PROMPT = """A customer security questionnaire is asking us \
this question:

"{question_text}"

We do not currently have an applied control documenting how we address it. \
Draft a realistic applied control that would let us answer it — something \
concrete and well-known in security practice.

Reply with a single JSON object, no prose:
{{
  "name": "<short, capitalized name (max 80 chars)>",
  "description": "<one sentence summarizing what this control is>",
  "observation": "<two or three sentences a security team would write \
describing how it's actually implemented and verified — be concrete with \
technologies, frequencies, owners>",
  "status": "to_do",
  "category": "policy" | "process" | "technical" | "physical" | "procedure",
  "csf_function": "govern" | "identify" | "protect" | "detect" | "respond" | "recover"
}}"""


# AppliedControl.category accepts a closed list — values that don't match
# fall back to empty string at create time.
_VALID_AC_CATEGORIES = {"policy", "process", "technical", "physical", "procedure"}
_VALID_AC_CSF_FUNCTIONS = {
    "govern",
    "identify",
    "protect",
    "detect",
    "respond",
    "recover",
}
_VALID_AC_STATUSES = {
    "to_do",
    "in_progress",
    "on_hold",
    "active",
    "degraded",
    "deprecated",
    "--",
}


def draft_applied_control_for_question(question_text: str) -> dict:
    """Synchronous LLM call producing a draft AppliedControl for the question.

    Returns ``{name, description, observation, status, category, csf_function}``.
    Falls back to a minimal stub if the LLM fails — the user can still edit
    and create from there.
    """
    from .providers import get_llm

    fallback = {
        "name": "",
        "description": "",
        "observation": "",
        "status": "to_do",
        "category": "",
        "csf_function": "",
    }
    try:
        llm = get_llm()
    except Exception as e:
        logger.warning("LLM unavailable for control suggestion: %s", e)
        return fallback

    prompt = SUGGEST_CONTROL_PROMPT.format(question_text=question_text)
    try:
        raw = llm.generate(prompt=prompt, context="", history=[])
    except Exception as e:
        logger.warning("LLM call failed for control suggestion: %s", e)
        return fallback

    parsed = _parse_json_response(raw) or {}
    name = (parsed.get("name") or "").strip()[:200]
    description = (parsed.get("description") or "").strip()
    observation = (parsed.get("observation") or "").strip()
    status_raw = (parsed.get("status") or "to_do").strip().lower()
    category_raw = (parsed.get("category") or "").strip().lower()
    csf_raw = (parsed.get("csf_function") or "").strip().lower()

    return {
        "name": name,
        "description": description,
        "observation": observation,
        "status": status_raw if status_raw in _VALID_AC_STATUSES else "to_do",
        "category": category_raw if category_raw in _VALID_AC_CATEGORIES else "",
        "csf_function": csf_raw if csf_raw in _VALID_AC_CSF_FUNCTIONS else "",
    }


def retry_question_with_hints(
    *,
    question_id: str,
    hint_applied_control_ids: list[str],
) -> dict:
    """Synchronous re-run of one question's answer pipeline with hint controls.

    Used by the review UI when a user attaches an existing AppliedControl as
    priority context for an unanswered or low-confidence question. Hint
    controls are formatted as the top-priority context passages (numbered
    [1], [2], …) and the LLM is instructed via critic_hint to use them.

    Marks the previous `propose_answer` for this question as expired and
    creates a fresh one (iteration += 1). Returns ``{action_id, confidence,
    iteration}`` so the caller can update the UI immediately.
    """
    from django.contrib.contenttypes.models import ContentType

    from core.models import AppliedControl

    from .models import (
        AgentAction,
        AgentRun,
        QuestionnaireQuestion,
        QuestionnaireRun,
    )
    from .providers import get_llm
    from .rag import search as rag_search
    from .tokens import count_tokens

    question = QuestionnaireQuestion.objects.get(id=question_id)
    qq_ct = ContentType.objects.get_for_model(QuestionnaireQuestion)
    qr_ct = ContentType.objects.get_for_model(QuestionnaireRun)

    run = (
        AgentRun.objects.filter(
            target_content_type=qr_ct,
            target_object_id=question.questionnaire_run_id,
        )
        .order_by("-created_at")
        .first()
    )
    if not run:
        raise ValueError("No agent run found for this questionnaire.")

    # Resolve hint controls; skip silently any that no longer exist or are
    # outside the run's folder (defense in depth — view does the perm check).
    hint_controls = list(
        AppliedControl.objects.filter(
            id__in=hint_applied_control_ids, folder=run.folder
        )
    )
    hint_dicts = [_applied_control_to_rag_dict(ac) for ac in hint_controls]

    try:
        rag_results = _search_folder_evidence(question.text, run.folder_id, top_k=4)
    except Exception as e:
        logger.error(
            "Folder-scoped search failed during retry for q %s: %s",
            question.id,
            e,
        )
        rag_results = []

    combined = hint_dicts + list(rag_results)
    context_block, source_refs = _build_context_block(combined)

    llm = get_llm()

    # Highest current iteration on this question; the new attempt sits above it.
    from django.db.models import Max

    max_iter = AgentAction.objects.filter(
        agent_run=run,
        kind=AgentAction.Kind.PROPOSE_ANSWER,
        target_content_type=qq_ct,
        target_object_id=question.id,
    ).aggregate(m=Max("iteration"))["m"]
    next_iter = (max_iter or 0) + 1

    hint_summary = (
        ", ".join(f"{ac.ref_id or ac.name[:40]}" for ac in hint_controls if ac)
        or "(no hint controls resolved)"
    )
    proposed = _answer_iteration(
        run=run,
        question=question,
        qq_ct=qq_ct,
        llm=llm,
        context_block=context_block,
        source_refs=source_refs,
        critic_hint=(
            f"Reviewer attached priority control(s): {hint_summary}. "
            "Lean on these passages first; only fall back to other context if "
            "they don't address the question."
        ),
        iteration=next_iter,
        count_tokens=count_tokens,
    )

    critic = _critique(
        run=run,
        question=question,
        qq_ct=qq_ct,
        llm=llm,
        answer_payload=proposed["payload"],
        context_block=context_block,
        cited_indices=proposed["cited_indices"],
        iteration=next_iter,
        count_tokens=count_tokens,
    )

    AgentAction.objects.filter(id=proposed["action_id"]).update(
        confidence=critic["score"]
    )

    # Expire any earlier proposed/active answers for this question — review UI
    # only shows the latest non-expired one.
    AgentAction.objects.filter(
        agent_run=run,
        kind=AgentAction.Kind.PROPOSE_ANSWER,
        target_content_type=qq_ct,
        target_object_id=question.id,
        state=AgentAction.State.PROPOSED,
    ).exclude(id=proposed["action_id"]).update(state=AgentAction.State.EXPIRED)

    return {
        "action_id": str(proposed["action_id"]),
        "confidence": critic["score"],
        "iteration": next_iter,
    }


def _applied_control_to_rag_dict(ac) -> dict:
    """Format an AppliedControl as a dict matching ``rag.search`` results.

    Lets us prepend hand-picked controls onto a RAG result list and run them
    through the same context-formatting code, so citation numbering stays
    consistent.
    """
    return {
        "id": str(ac.id),
        "score": 1.0,
        "text": _text_for_applied_control(ac),
        "source_type": "model",
        "object_type": "applied_control",
        "object_id": str(ac.id),
        "name": ac.name or "",
        "ref_id": ac.ref_id or "",
        "framework": "",
        "urn": "",
    }


# Citation kinds that describe what *should* be done (framework material,
# library reference controls, library threats). These don't justify a "yes"
# on their own — internal records of practice do.
#
# Two flavors covered: ``"library"`` is the indexer's ``source_type`` for
# items pulled from YAML libraries (matched first by _build_context_block
# when source_type != "model"), and the per-model ``object_type`` slugs
# cover the case where a library row was indexed model-style.
_LIBRARY_REF_KINDS = frozenset(
    {
        "library",
        "requirement_node",
        "reference_control",
        "library_threat",
        "framework",
        "threat",
    }
)


def _refine_verdict_against_citations(
    answer_status: str,
    cited_indices: list,
    source_refs: list[dict],
) -> tuple[str, str]:
    """Deterministic post-processing: downgrade verdicts that aren't supported
    by their citations. Belt-and-braces guard for cases where the LLM cites
    only planned/unverified evidence yet still claims yes/partial.

    Returns ``(refined_status, downgrade_note)``. When the verdict is left
    untouched, the note is empty.

    Rules (mirror the answer prompt):
      - 'yes' requires at least one cited model item with status=active or
        result=compliant.
      - 'partial' requires at least one with status in {active, in_progress}
        or result in {compliant, partially_compliant}.
      - Citations to non-status-bearing items (document chunks, library
        requirement nodes) are treated as neutral and don't trigger a
        downgrade — those can legitimately support a yes.
    """
    from core.models import (
        AppliedControl,
        ComplianceAssessment,
        RequirementAssessment,
    )

    if answer_status not in (Verdict.YES, Verdict.PARTIAL):
        return answer_status, ""
    if not cited_indices:
        return answer_status, ""

    has_active = False
    has_partial = False
    has_status_bearing_only = False
    has_document_evidence = False  # uploaded evidence chunks — real internal evidence
    has_library_only = False  # library requirement nodes etc. — reference, not evidence

    # Collect the citations we actually need to resolve so we can do exactly
    # three bulk queries (AC / RA / CA) instead of N×3 single-row lookups.
    cited_index_set = set(cited_indices)
    refs_to_resolve = []  # list of (ref_id, ref_kind)
    for ref in source_refs:
        if ref.get("index") not in cited_index_set:
            continue
        ref_id = ref.get("id") or ""
        ref_kind = (ref.get("kind") or "").lower()
        if not ref_id:
            if ref_kind == "document":
                has_document_evidence = True
            else:
                has_library_only = True
            continue
        refs_to_resolve.append((ref_id, ref_kind))

    if refs_to_resolve:
        ids = [r[0] for r in refs_to_resolve]
        # Three bulk lookups by id — Django dedupes the IN list. .values()
        # avoids hydrating full model instances when we only need one field.
        ac_status_by_id = dict(
            AppliedControl.objects.filter(id__in=ids).values_list("id", "status")
        )
        ra_result_by_id = dict(
            RequirementAssessment.objects.filter(id__in=ids).values_list("id", "result")
        )
        ca_status_by_id = dict(
            ComplianceAssessment.objects.filter(id__in=ids).values_list("id", "status")
        )
        # IDs come from JSONField source_refs as strings; cast Django UUIDs
        # to str so the lookups match.
        ac_status_by_id = {str(k): v for k, v in ac_status_by_id.items()}
        ra_result_by_id = {str(k): v for k, v in ra_result_by_id.items()}
        ca_status_by_id = {str(k): v for k, v in ca_status_by_id.items()}

        for ref_id, ref_kind in refs_to_resolve:
            if ref_id in ac_status_by_id:
                ac_status = ac_status_by_id[ref_id]
                if ac_status == "active":
                    has_active = True
                elif ac_status == "in_progress":
                    has_partial = True
                else:
                    has_status_bearing_only = True
                continue
            if ref_id in ra_result_by_id:
                ra_result = ra_result_by_id[ref_id]
                if ra_result == "compliant":
                    has_active = True
                elif ra_result == "partially_compliant":
                    has_partial = True
                else:
                    has_status_bearing_only = True
                continue
            if ref_id in ca_status_by_id:
                # CA represents the audit itself, not the org's compliance.
                # `done` ≈ assessment finished; `in_progress`/`in_review` ≈
                # engaged. Neither directly equals "we're compliant" — both
                # are at-best partial signals. `planned`/`deprecated` carry
                # no support.
                if ca_status_by_id[ref_id] in ("in_progress", "in_review", "done"):
                    has_partial = True
                else:
                    has_status_bearing_only = True
                continue
            # Not AC, not RA, not CA. Three buckets:
            #   - "document": uploaded evidence — supports the verdict.
            #   - library/reference (requirement_node, reference_control,
            #     framework, library_*): describes what *should* be done,
            #     not what we do. Pure library citations should not justify
            #     "yes" on their own.
            #   - everything else (evidence, incident, vulnerability, asset,
            #     risk_scenario, …): real internal records — neutral here.
            #     Don't downgrade on these.
            if ref_kind == "document":
                has_document_evidence = True
            elif ref_kind in _LIBRARY_REF_KINDS:
                has_library_only = True

    # Active/compliant evidence or uploaded documents support the verdict.
    if has_active or has_document_evidence:
        return answer_status, ""

    if answer_status == Verdict.YES:
        if has_partial:
            return (
                Verdict.PARTIAL,
                "Auto-downgraded yes→partial: cited evidence is in progress / "
                "partially compliant, not active.",
            )
        if has_status_bearing_only or has_library_only:
            return (
                Verdict.NEEDS_INFO,
                "Auto-downgraded yes→needs_info: cited evidence is planned, "
                "unverified, or only references the requirement itself "
                "without evidence of practice.",
            )
    elif answer_status == Verdict.PARTIAL:
        if has_partial:
            return answer_status, ""  # in_progress/partially_compliant supports partial
        if has_status_bearing_only or has_library_only:
            return (
                Verdict.NEEDS_INFO,
                "Auto-downgraded partial→needs_info: cited evidence is "
                "planned, unverified, or only references the requirement "
                "itself without evidence of practice.",
            )

    return answer_status, ""


def _answer_iteration(
    *,
    run,
    question,
    qq_ct,
    llm,
    context_block: str,
    source_refs: list[dict],
    critic_hint: str,
    iteration: int,
    count_tokens,
):
    """One answer attempt. Records and returns the AgentAction id + parsed payload."""
    import time

    from .models import AgentAction

    section_hint = (
        f"Question section: {question.section}\n\n" if question.section else ""
    )
    prompt = ANSWER_PROMPT.format(
        question=question.text,
        section_hint=section_hint,
        context=context_block or "(no relevant context found)",
    )
    if critic_hint:
        prompt += f"\n\nReviewer feedback to address: {critic_hint}"

    t0 = time.time()
    try:
        raw = llm.generate(prompt=prompt, context="", history=[])
    except Exception as e:
        logger.error("LLM generate failed for q %s: %s", question.id, e)
        raw = ""
    duration_ms = int((time.time() - t0) * 1000)

    parsed = _parse_json_response(raw) or {}
    answer_status = (
        parsed.get("status")
        if Verdict.is_valid(parsed.get("status") or "")
        else Verdict.NEEDS_INFO
    )
    comment = (parsed.get("comment") or "").strip()
    cited_indices = parsed.get("citation_indices") or []
    if not isinstance(cited_indices, list):
        cited_indices = []

    # Deterministic guard: if the cited evidence is only planned / unverified,
    # downgrade the verdict regardless of what the LLM said.
    refined_status, downgrade_note = _refine_verdict_against_citations(
        answer_status, cited_indices, source_refs
    )
    if refined_status != answer_status:
        logger.info(
            "verdict_downgraded q=%s %s→%s (%s)",
            question.id,
            answer_status,
            refined_status,
            downgrade_note,
        )
        answer_status = refined_status
        if downgrade_note:
            comment = (comment + ("\n\n" if comment else "") + downgrade_note).strip()

    used_refs = [r for r in source_refs if r["index"] in set(cited_indices)]

    tokens = 0
    try:
        tokens = count_tokens(prompt) + count_tokens(raw)
    except Exception:
        # Token accounting is best-effort: the answer/critic record is far
        # more important than its token tally. Log at debug for triage.
        logger.debug("count_tokens failed", exc_info=True)

    action = AgentAction.objects.create(
        agent_run=run,
        kind=AgentAction.Kind.PROPOSE_ANSWER,
        target_content_type=qq_ct,
        target_object_id=question.id,
        payload={"status": answer_status, "comment": comment},
        rationale=raw[:2000],
        source_refs=used_refs,
        confidence=FAST_MODE_DEFAULT_CONFIDENCE,
        state=AgentAction.State.PROPOSED,
        iteration=iteration,
        tokens=tokens,
        duration_ms=duration_ms,
    )

    run.total_tokens = (run.total_tokens or 0) + tokens
    run.save(update_fields=["total_tokens", "updated_at"])

    return {
        "action_id": action.id,
        "payload": {"status": answer_status, "comment": comment},
        "cited_indices": cited_indices,
        "confidence": FAST_MODE_DEFAULT_CONFIDENCE,
    }


def _critique(
    *,
    run,
    question,
    qq_ct,
    llm,
    answer_payload: dict,
    context_block: str,
    cited_indices: list[int],
    iteration: int,
    count_tokens,
):
    """Critic LLM call. Returns dict with score (float) and issue (string)."""
    import time

    from .models import AgentAction

    cited_lines = []
    for line in context_block.split("\n"):
        for idx in cited_indices:
            if line.startswith(f"[{idx}]"):
                cited_lines.append(line)
                break
    cited_text = "\n".join(cited_lines) if cited_lines else "(answer cited no passages)"

    prompt = CRITIC_PROMPT.format(
        question=question.text,
        status=answer_payload.get("status", ""),
        comment=answer_payload.get("comment", ""),
        cited_context=cited_text,
    )

    t0 = time.time()
    try:
        raw = llm.generate(prompt=prompt, context="", history=[])
    except Exception as e:
        logger.error("Critic LLM failed for q %s: %s", question.id, e)
        raw = ""
    duration_ms = int((time.time() - t0) * 1000)

    parsed = _parse_json_response(raw) or {}
    try:
        score = float(parsed.get("score", 0.0))
    except (TypeError, ValueError):
        score = 0.0
    score = max(0.0, min(1.0, score))
    issue = (parsed.get("issue") or "").strip()

    tokens = 0
    try:
        tokens = count_tokens(prompt) + count_tokens(raw)
    except Exception:
        # Token accounting is best-effort: the answer/critic record is far
        # more important than its token tally. Log at debug for triage.
        logger.debug("count_tokens failed", exc_info=True)

    AgentAction.objects.create(
        agent_run=run,
        kind=AgentAction.Kind.CRITIQUE,
        target_content_type=qq_ct,
        target_object_id=question.id,
        payload={"score": score, "issue": issue},
        rationale=raw[:2000],
        source_refs=[],
        confidence=score,
        state=AgentAction.State.PROPOSED,
        iteration=iteration,
        tokens=tokens,
        duration_ms=duration_ms,
    )

    run.total_tokens = (run.total_tokens or 0) + tokens
    run.save(update_fields=["total_tokens", "updated_at"])

    return {"score": score, "issue": issue}
