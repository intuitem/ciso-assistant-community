import json
import re
import structlog
import time

from django.apps import apps
from django.contrib.contenttypes.models import ContentType
from django.http import StreamingHttpResponse
from django.utils import timezone
from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.throttling import UserRateThrottle


class ChatMessageThrottle(UserRateThrottle):
    rate = "120/hour"


from core.views import BaseModelViewSet as AbstractBaseModelViewSet
from .constants import (
    CHAT_SESSION_SUMMARY_ASYNC,
    LANG_MAP,
    MODEL_CONTEXT_TOKENS,
    RAG_CONTEXT_TOKENS,
    VERBATIM_WINDOW_TOKENS,
    Verdict,
)
from .context import ContextBuilder
from .memory import (
    build_replay_payload,
    inject_summary,
    inject_tool_replays,
    pack_verbatim_window,
    update_summary_for_session,
)
from .metrics import build_turn_metrics, record_metric
from .models import (
    ChatSession,
    ChatMessage,
    IndexedDocument,
    QuestionnaireRun,
    QuestionnaireQuestion,
    AgentRun,
    AgentAction,
)
from .page_context import parse_page_context
from .providers import get_llm, is_ollama_available
from .tokens import count_tokens
from .serializers import (
    ChatSessionListSerializer,
    SendMessageSerializer,
)

logger = structlog.get_logger(__name__)

# Patterns that attempt to impersonate system/assistant roles or override instructions
_INJECTION_PATTERNS = re.compile(
    r"(?:"
    r"\[/?(?:SYSTEM|CONTEXT|INST)\]"  # Fake delimiter tags
    r"|<\|(?:im_start|im_end|system)\|>"  # ChatML role markers
    r"|```\s*(?:system|tool_call)"  # Fenced role blocks
    r")",
    re.IGNORECASE,
)


def _sanitize_user_input(text: str) -> str:
    """
    Neutralize prompt injection patterns in user messages.
    Strips characters that could be interpreted as role markers or
    delimiter tags by the LLM, while preserving the user's intent.
    """
    # Remove injection patterns
    text = _INJECTION_PATTERNS.sub("", text)
    # Strip null bytes and other control characters (keep newlines/tabs)
    text = "".join(
        c for c in text if c == "\n" or c == "\t" or (c >= " " and c <= "\U0010ffff")
    )
    return text.strip()


class BaseModelViewSet(AbstractBaseModelViewSet):
    serializers_module = "chat.serializers"


class ChatSessionViewSet(BaseModelViewSet):
    """ViewSet for chat sessions with streaming message endpoint."""

    model = ChatSession

    def get_queryset(self):
        # Users can only see their own sessions
        return ChatSession.objects.filter(owner=self.request.user)

    def perform_create(self, serializer):
        serializer.save(owner=self.request.user)

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        serializer = ChatSessionListSerializer(queryset, many=True)
        return Response(serializer.data)

    @action(
        detail=True,
        methods=["post"],
        url_path="message",
        throttle_classes=[ChatMessageThrottle],
    )
    def send_message(self, request, pk=None):
        """
        Send a message and get a streaming SSE response.
        The response includes RAG retrieval + LLM generation.
        """
        session = self.get_object()

        serializer = SendMessageSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user_content = _sanitize_user_input(serializer.validated_data["content"])
        page_context = serializer.validated_data.get("page_context", {})

        # Save user message
        ChatMessage.objects.create(
            session=session,
            role=ChatMessage.Role.USER,
            content=user_content,
        )

        # Auto-set title from first message
        if not session.title:
            session.title = user_content[:100]
            session.save(update_fields=["title"])

        # Try tool calling first, fall back to RAG semantic search
        from .rag import (
            search,
            graph_expand,
            format_context,
            build_context_refs,
            get_accessible_folder_ids,
        )
        from .orm_query import format_query_result
        from .tools import (
            get_tools,
            dispatch_tool_call,
            PARENT_CHILD_MAP,
            ATTACHABLE_RELATIONS,
            MODEL_MAP,
        )

        accessible_folders = get_accessible_folder_ids(request.user)
        context_refs = []
        context = ""
        query_result = None

        # Parse page context into structured reference
        parsed_context = parse_page_context(page_context) if page_context else None

        # Step 1: Route to tool or workflow
        from .workflows import get_workflow_tools, get_workflow_by_tool_name
        from .workflows.base import WorkflowContext

        llm = get_llm()
        logger.info(
            "llm_provider_selected",
            provider=type(llm).__name__,
            page_context=page_context,
            parsed_context=str(parsed_context) if parsed_context else None,
        )

        # Step 1a: Deterministic pre-routing for workflows.
        # Small LLMs are unreliable at tool selection with 5+ tools,
        # so we match workflows by context + keywords first.
        pre_routed_workflow = _match_workflow(user_content, parsed_context)
        if pre_routed_workflow:
            logger.info("workflow_pre_routed", workflow=pre_routed_workflow.name)
            tool_response = {"name": f"workflow_{pre_routed_workflow.name}"}
        else:
            # Step 1b: LLM-based tool selection
            # Build page context prefix for the LLM
            page_context_prefix = _build_context_prompt(page_context, parsed_context)

            # Inject previous query metadata so the LLM can handle follow-ups
            # like "give me more", "next page", "show me page 3"
            last_query_meta = _get_last_query_meta(session)
            user_lang = request.META.get("HTTP_ACCEPT_LANGUAGE", "en")[:2]
            lang_name = LANG_MAP.get(user_lang, "English")
            tool_prompt = f"LANGUAGE: Generate all names and descriptions in {lang_name}.\n\n{user_content}"
            if page_context_prefix or last_query_meta:
                tool_prompt = f"LANGUAGE: Generate all names and descriptions in {lang_name}.\n\n{page_context_prefix}"
                if last_query_meta:
                    tool_prompt += (
                        f"Previous query context: model={last_query_meta.get('model')}, "
                        f"domain={last_query_meta.get('domain', 'all')}, "
                        f"page={last_query_meta.get('page', 1)}, "
                        f"total_pages={last_query_meta.get('total_pages', 1)}, "
                        f"total_count={last_query_meta.get('total_count', 0)}. "
                        f"If the user asks for 'more', 'next', 'next page', etc., "
                        f"repeat the same query with page={last_query_meta.get('page', 1) + 1}.\n\n"
                    )
                tool_prompt += f"User message: {user_content}"

            # Combine standard tools + context-aware workflow tools
            all_tools = get_tools() + get_workflow_tools(parsed_context)

            history_for_tool = list(
                session.messages.order_by("created_at").values("role", "content")
            )
            # The current user message is passed separately via tool_prompt
            if (
                history_for_tool
                and history_for_tool[-1]["role"] == "user"
                and history_for_tool[-1]["content"] == user_content
            ):
                history_for_tool = history_for_tool[:-1]
            history_for_tool = pack_verbatim_window(
                history_for_tool, VERBATIM_WINDOW_TOKENS
            )
            # Tool routing gets the summary but no tool replays (lean prompt)
            history_for_tool = inject_summary(history_for_tool, session.summary)

            t0 = time.time()
            tool_response = llm.tool_call(
                tool_prompt,
                all_tools,
                history=history_for_tool,
            )
            logger.info(
                "tool_selection_complete",
                duration=round(time.time() - t0, 2),
                tool=tool_response.get("name") if tool_response else "no tool",
            )

        # Check if LLM selected a workflow (or pre-routed)
        if tool_response and tool_response.get("name", "").startswith("workflow_"):
            workflow = get_workflow_by_tool_name(tool_response["name"])
            if workflow:
                logger.info("workflow_selected", workflow=workflow.name)

                # Build conversation history (exclude the message we just saved)
                wf_history = list(
                    session.messages.order_by("created_at").values("role", "content")
                )
                if (
                    wf_history
                    and wf_history[-1]["role"] == "user"
                    and wf_history[-1]["content"] == user_content
                ):
                    wf_history = wf_history[:-1]
                wf_history = pack_verbatim_window(wf_history, VERBATIM_WINDOW_TOKENS)

                wf_ctx = WorkflowContext(
                    user_message=user_content,
                    parsed_context=parsed_context,
                    accessible_folder_ids=accessible_folders,
                    llm=llm,
                    history=wf_history,
                    user_lang=request.META.get("HTTP_ACCEPT_LANGUAGE", "en")[:2],
                    session=session,
                )

                def stream_workflow():
                    full_response = ""
                    wf_context_refs = []
                    wf_start = time.time()
                    try:
                        for event in workflow.run(wf_ctx):
                            if event.type == "token":
                                full_response += (
                                    event.content
                                    if isinstance(event.content, str)
                                    else ""
                                )
                            if event.type == "pending_action" and isinstance(
                                event.content, dict
                            ):
                                wf_context_refs.append(
                                    {"source": "pending_action", **event.content}
                                )
                            yield event.encode()

                        # Strip the JSON recommendation block before saving
                        saved_response = re.sub(
                            r"\s*```json\s*\{[^}]*\"recommended\"[^}]*\}\s*```\s*",
                            "",
                            full_response,
                        ).rstrip()
                        ChatMessage.objects.create(
                            session=session,
                            role=ChatMessage.Role.ASSISTANT,
                            content=saved_response or full_response,
                            context_refs=wf_context_refs,
                        )
                        logger.info(
                            "workflow_complete",
                            workflow=workflow.name,
                            duration=round(time.time() - wf_start, 2),
                        )
                        done_data = json.dumps(
                            {"type": "done", "context_refs": wf_context_refs}
                        )
                        yield f"data: {done_data}\n\n"

                    except Exception as e:
                        logger.error(
                            "Workflow '%s' error after %.2fs: %s",
                            workflow.name,
                            time.time() - wf_start,
                            e,
                        )
                        error_msg = "Sorry, I encountered an error. Please try again."
                        ChatMessage.objects.create(
                            session=session,
                            role=ChatMessage.Role.ASSISTANT,
                            content=error_msg,
                        )
                        error_data = json.dumps({"type": "error", "content": error_msg})
                        yield f"data: {error_data}\n\n"

                response = StreamingHttpResponse(
                    stream_workflow(), content_type="text/event-stream"
                )
                response["Cache-Control"] = "no-cache"
                response["X-Accel-Buffering"] = "no"
                return response

        tool_observation_payload: dict | None = None

        if tool_response and tool_response.get("name"):
            logger.info(
                "Tool call: %s(%s)",
                tool_response["name"],
                tool_response.get("arguments", {}),
            )
            t1 = time.time()
            query_result = dispatch_tool_call(
                tool_response["name"],
                tool_response.get("arguments", {}),
                accessible_folders,
                parsed_context,
                user_message=user_content,
            )
            logger.info("tool_dispatch_complete", duration=round(time.time() - t1, 2))

            # Only query_objects/search_library have the right shape;
            # other tools fall through to build_replay_payload returning None
            if query_result:
                tool_name = tool_response["name"]
                raw_result_text = ""
                if tool_name == "search_library":
                    raw_result_text = query_result.get("text", "") or ""
                elif tool_name == "query_objects" and "total_count" in query_result:
                    raw_result_text = format_query_result(query_result)
                tool_observation_payload = build_replay_payload(
                    tool_name,
                    tool_response.get("arguments", {}),
                    raw_result_text,
                )

        # Track if this is a creation/attach proposal (different SSE flow)
        creation_proposal = None

        if query_result and query_result.get("type") == "propose_create":
            # Creation proposal — don't execute, let the user confirm via UI
            creation_proposal = query_result
            context = (
                f"The system is proposing to create {len(query_result.get('items', []))} "
                f"{query_result['display_name']}. "
                "Interactive confirmation cards are already displayed in the UI.\n\n"
                "YOUR RESPONSE MUST:\n"
                "- Tell the user you're proposing to create these items (1 sentence).\n"
                "- Tell them to review and use the Confirm/Cancel buttons below.\n\n"
                "YOUR RESPONSE MUST NOT:\n"
                "- List the items (the UI already shows them).\n"
                "- Say you have created them.\n"
                "- Generate links or URLs.\n"
                "Keep it to 2-3 sentences maximum."
            )
            context_refs.append(
                {
                    "source": "pending_action",
                    "action": "create",
                    "model_key": query_result["model_key"],
                    "url_slug": query_result["url_slug"],
                    "display_name": query_result["display_name"],
                    "items": query_result["items"],
                }
            )
        elif query_result and query_result.get("type") == "propose_attach":
            # Attach proposal — show existing objects to link
            creation_proposal = query_result
            context = (
                f"The system found {len(query_result.get('items', []))} existing "
                f"{query_result['related_display']} that can be attached to the current "
                f"{query_result['parent_display']}. "
                "Interactive confirmation cards are already displayed in the UI.\n\n"
                "YOUR RESPONSE MUST:\n"
                "- Briefly explain why these items were suggested (1 sentence).\n"
                "- Tell the user to review and use the Confirm/Cancel buttons below.\n\n"
                "YOUR RESPONSE MUST NOT:\n"
                "- List the items (the UI already shows them).\n"
                "- Generate links, URLs, or markdown links.\n"
                "- Include IDs, UUIDs, or technical references.\n"
                "- Describe how to use tools.\n"
                "Keep it to 2-3 sentences maximum."
            )
            context_refs.append(
                {
                    "source": "pending_action",
                    "action": "attach",
                    "parent_model_key": query_result["parent_model_key"],
                    "parent_id": query_result["parent_id"],
                    "parent_url_slug": query_result["parent_url_slug"],
                    "m2m_field": query_result["m2m_field"],
                    "related_model_key": query_result["related_model_key"],
                    "related_display": query_result["related_display"],
                    "items": query_result["items"],
                }
            )
        elif query_result and query_result.get("type") == "multi_query":
            parts = [
                "INSTRUCTIONS: Multiple queries were executed to answer your question. "
                "Present ALL results to the user in a clear, structured way. "
                "Do NOT hallucinate additional information.\n"
            ]
            for i, sub in enumerate(query_result.get("results", []), 1):
                parts.append(f"--- Query {i}: {sub.get('display_name', '')} ---")
                parts.append(format_query_result(sub))
            context = "\n\n".join(parts)
        elif query_result and query_result.get("type") == "search_library":
            context = (
                "INSTRUCTIONS: The following data comes from the frameworks knowledge base. "
                "Present this information to the user in a clear, structured way. "
                "When citing requirements, always include the framework name and reference ID. "
                "Do NOT hallucinate additional information beyond what is provided.\n\n"
                + query_result.get("text", "No results found.")
            )
        elif query_result:
            context = (
                "INSTRUCTIONS: The following data comes from a database query. "
                "Present ONLY this data to the user. Use the total_count as the authoritative count. "
                "The listed items are one page of results. Do NOT add explanations, "
                "framework references, or commentary beyond what is in the data. "
                "Do NOT hallucinate additional information. "
                "IMPORTANT: Items include markdown links like [Name](/path/id) — "
                "you MUST preserve these links exactly as-is in your response so the user "
                "can click them. Do NOT rewrite, shorten, or remove the links.\n\n"
                + format_query_result(query_result)
            )
            url_slug = query_result.get("url_slug", "")
            for obj in query_result.get("objects", []):
                context_refs.append(
                    {
                        "type": query_result["model_name"],
                        "id": obj.get("id", ""),
                        "name": obj.get("name", ""),
                        "url": f"/{url_slug}/{obj.get('id', '')}",
                        "source": "orm_query",
                    }
                )
            # Save query metadata for follow-up (next page, etc.)
            context_refs.append(
                {
                    "source": "query_meta",
                    "model": tool_response.get("arguments", {}).get("model"),
                    "domain": tool_response.get("arguments", {}).get("domain"),
                    "page": query_result.get("page", 1),
                    "total_pages": query_result.get("total_pages", 1),
                    "total_count": query_result.get("total_count", 0),
                }
            )
        else:
            # Step 2: Knowledge graph + Semantic RAG search
            # Check the knowledge graph for framework context, but only for
            # general questions — skip when on action pages (detail/edit)
            # where workflows and attach proposals are the expected path.
            graph_context = ""
            if not (parsed_context and parsed_context.object_id):
                graph_context = _get_graph_context(user_content)

            results = search(user_content, request.user, top_k=10)

            structured = [r for r in results if r.get("source_type") == "model"]
            expanded = (
                graph_expand(structured, accessible_folders) if structured else []
            )

            context = ""
            if graph_context:
                context = "FRAMEWORK KNOWLEDGE BASE:\n" + graph_context + "\n\n"
            context += format_context(results, expanded)
            context_refs = build_context_refs(results, expanded)

            # Step 3: Auto-suggest attachable objects when on a contextual page
            # If the LLM didn't call a tool but we're on a detail/edit page
            # with attachable relations, try attaching relevant objects automatically.
            if parsed_context and parsed_context.object_id:
                from .tools import ATTACHABLE_RELATIONS, _build_attach_proposal

                relations = ATTACHABLE_RELATIONS.get(parsed_context.model_key, [])
                if relations:
                    # Try the first attachable relation (usually applied_controls)
                    first_rel_key = relations[0][0]
                    attach_result = _build_attach_proposal(
                        {"related_model": first_rel_key},
                        accessible_folders,
                        parsed_context,
                    )
                    if attach_result and attach_result.get("items"):
                        creation_proposal = attach_result
                        item_names = [i["name"] for i in attach_result["items"]]
                        # Replace the RAG context entirely — the attach cards ARE the response
                        context = (
                            f"The system found {len(item_names)} existing "
                            f"{attach_result['related_display']} that may be relevant. "
                            "Interactive confirmation cards are already displayed in the UI "
                            "showing each item with a Confirm/Cancel button.\n\n"
                            "YOUR RESPONSE MUST:\n"
                            "- Briefly explain why these items were suggested (1-2 sentences).\n"
                            "- Tell the user to review and use the buttons below to attach them.\n\n"
                            "YOUR RESPONSE MUST NOT:\n"
                            "- List the items (the UI already shows them).\n"
                            "- Generate links, URLs, or markdown links.\n"
                            "- Include IDs, UUIDs, or technical references.\n"
                            "- Describe how to use tools or the attach_existing tool.\n"
                            "- Use emojis.\n"
                            "Keep it to 2-3 sentences maximum."
                        )
                        context_refs.append(
                            {
                                "source": "pending_action",
                                "action": "attach",
                                "parent_model_key": attach_result["parent_model_key"],
                                "parent_id": attach_result["parent_id"],
                                "parent_url_slug": attach_result["parent_url_slug"],
                                "m2m_field": attach_result["m2m_field"],
                                "related_model_key": attach_result["related_model_key"],
                                "related_display": attach_result["related_display"],
                                "items": attach_result["items"],
                            }
                        )

        # Build conversation history for the LLM (exclude the message we just saved)
        history_messages = list(
            session.messages.order_by("created_at").values(
                "role", "content", "tool_observation"
            )
        )
        # Remove the last entry (the user message we just created) — it's passed separately
        if (
            history_messages
            and history_messages[-1]["role"] == "user"
            and history_messages[-1]["content"] == user_content
        ):
            history_messages = history_messages[:-1]
        history_messages = pack_verbatim_window(
            history_messages, VERBATIM_WINDOW_TOKENS
        )
        # Replays before summary so notes attach to the right messages
        history_messages = inject_tool_replays(history_messages)
        history_messages = inject_summary(history_messages, session.summary)

        # Assemble context with structured priorities
        user_lang = request.META.get("HTTP_ACCEPT_LANGUAGE", "en")[:2]
        lang_name = LANG_MAP.get(user_lang, "English")

        ctx_builder = ContextBuilder(max_tokens=RAG_CONTEXT_TOKENS)
        ctx_builder.add(
            "language",
            f"LANGUAGE: You MUST respond in {lang_name}.",
            priority=10,
        )
        if page_context_prefix and not query_result:
            ctx_builder.add("page_context", page_context_prefix, priority=8)
        ctx_builder.add("main_context", context, priority=9)

        # Enrich context with domain objects when on a detail page
        if parsed_context and parsed_context.object_id:
            enrichment = _enrich_context(parsed_context, accessible_folders)
            if enrichment:
                ctx_builder.add("enrichment", enrichment, priority=5)

        context = ctx_builder.build()

        system_prompt_text = llm.system_prompt if hasattr(llm, "system_prompt") else ""
        system_prompt_chars = len(system_prompt_text)
        system_prompt_tokens = count_tokens(system_prompt_text)
        context_chars = len(context)
        context_tokens = count_tokens(context)
        user_chars = len(user_content)
        user_tokens = count_tokens(user_content)
        history_chars = sum(len(m.get("content", "")) for m in history_messages)
        history_tokens = sum(
            count_tokens(m.get("content", "")) for m in history_messages
        )
        total_prompt_chars = (
            system_prompt_chars + context_chars + user_chars + history_chars
        )
        total_prompt_tokens = (
            system_prompt_tokens + context_tokens + user_tokens + history_tokens
        )

        # One dict, three sinks: structlog event, JSONL file, ChatMessage.metrics
        summary_tokens_now = count_tokens(session.summary or "")
        turn_metrics = build_turn_metrics(
            prompt_tokens=total_prompt_tokens,
            model_context_tokens=MODEL_CONTEXT_TOKENS,
            system_prompt_tokens=system_prompt_tokens,
            context_tokens=context_tokens,
            history_tokens=history_tokens,
            user_tokens=user_tokens,
            summary_tokens=summary_tokens_now,
            history_messages=len(history_messages),
            section_names=ctx_builder.section_names(),
        )

        logger.info(
            "llm_context_size",
            system_prompt_chars=system_prompt_chars,
            context_chars=context_chars,
            sections=turn_metrics["sections"],
            history_messages=turn_metrics["history_messages"],
            total_prompt_chars=total_prompt_chars,
            system_prompt_tokens=system_prompt_tokens,
            context_tokens=context_tokens,
            history_tokens=history_tokens,
            user_tokens=user_tokens,
            summary_tokens=summary_tokens_now,
            total_prompt_tokens=total_prompt_tokens,
            model_context_tokens=MODEL_CONTEXT_TOKENS,
            over_budget=turn_metrics["over_budget"],
        )

        record_metric(
            "llm_context_size",
            session_id=str(session.pk),
            **turn_metrics,
        )

        # Warn at 80% — provider may silently truncate before our `over_budget`
        # fires if MODEL_CONTEXT_TOKENS exceeds the real provider limit
        if turn_metrics["high_watermark"]:
            logger.warning(
                "llm_context_high_watermark",
                total_prompt_tokens=total_prompt_tokens,
                model_context_tokens=MODEL_CONTEXT_TOKENS,
                utilization_pct=turn_metrics["utilization_pct"],
            )

        def stream_response():
            """Generator for SSE streaming."""
            full_response = ""

            try:
                # Emit pending_action event before streaming text
                if creation_proposal:
                    if creation_proposal.get("type") == "propose_attach":
                        action_data = json.dumps(
                            {
                                "type": "pending_action",
                                "action": "attach",
                                "parent_model_key": creation_proposal[
                                    "parent_model_key"
                                ],
                                "parent_id": creation_proposal["parent_id"],
                                "parent_url_slug": creation_proposal["parent_url_slug"],
                                "m2m_field": creation_proposal["m2m_field"],
                                "related_model_key": creation_proposal[
                                    "related_model_key"
                                ],
                                "related_display": creation_proposal["related_display"],
                                "items": creation_proposal["items"],
                            }
                        )
                    else:
                        action_data = json.dumps(
                            {
                                "type": "pending_action",
                                "action": "create",
                                "model_key": creation_proposal["model_key"],
                                "url_slug": creation_proposal["url_slug"],
                                "display_name": creation_proposal["display_name"],
                                "items": creation_proposal["items"],
                                "folder_id": creation_proposal.get("folder_id"),
                                "folder_name": creation_proposal.get("folder_name", ""),
                                "available_folders": creation_proposal.get(
                                    "available_folders", []
                                ),
                            }
                        )
                    yield f"data: {action_data}\n\n"

                t_stream = time.time()
                t_first_token = None
                t_first_content = None
                thinking_count = 0
                token_count = 0
                for token_type, token in llm.stream(
                    user_content, context, history=history_messages
                ):
                    if t_first_token is None:
                        t_first_token = time.time()
                    if token_type == "token":
                        if t_first_content is None:
                            t_first_content = time.time()
                        full_response += token
                        token_count += 1
                    else:
                        thinking_count += 1
                    # SSE format — "thinking" tokens go to a collapsible block in the UI
                    data = json.dumps({"type": token_type, "content": token})
                    yield f"data: {data}\n\n"
                t_end = time.time()
                ttft = round(t_first_token - t_stream, 2) if t_first_token else None
                thinking_duration = (
                    round(t_first_content - t_first_token, 2)
                    if t_first_content and t_first_token
                    else None
                )
                content_elapsed = t_end - (t_first_content or t_end)
                tps = (
                    round(token_count / content_elapsed, 1)
                    if content_elapsed > 0
                    else None
                )
                logger.info(
                    "llm_stream_complete",
                    duration=round(t_end - t_stream, 2),
                    ttft_s=ttft,
                    thinking_s=thinking_duration,
                    thinking_tokens=thinking_count,
                    tokens=token_count,
                    tokens_per_s=tps,
                )

                ChatMessage.objects.create(
                    session=session,
                    role=ChatMessage.Role.ASSISTANT,
                    content=full_response,
                    context_refs=context_refs,
                    tool_observation=tool_observation_payload,
                    metrics=turn_metrics,
                )

                # Send completion event with context refs
                done_data = json.dumps(
                    {
                        "type": "done",
                        "context_refs": context_refs,
                    }
                )
                yield f"data: {done_data}\n\n"

                # Compaction runs after 'done' flushed; failures retried next turn
                try:
                    if CHAT_SESSION_SUMMARY_ASYNC:
                        from .tasks import update_session_summary

                        update_session_summary(str(session.pk))
                    else:
                        update_summary_for_session(session, llm)
                except Exception as e:
                    logger.warning(
                        "summary_compaction_failed",
                        session_id=str(session.pk),
                        error=str(e),
                    )

            except Exception as e:
                logger.error("Chat stream error: %s", e)
                # Save whatever was generated before the failure
                saved_content = full_response.strip()
                if saved_content:
                    saved_content += "\n\n---\n*Response interrupted due to an error.*"
                else:
                    saved_content = (
                        "Sorry, I encountered an error generating a response. "
                        "Please check that the LLM service is configured and running."
                    )
                ChatMessage.objects.create(
                    session=session,
                    role=ChatMessage.Role.ASSISTANT,
                    content=saved_content,
                )
                error_data = json.dumps({"type": "error", "content": saved_content})
                yield f"data: {error_data}\n\n"

        response = StreamingHttpResponse(
            stream_response(),
            content_type="text/event-stream",
        )
        response["Cache-Control"] = "no-cache"
        response["X-Accel-Buffering"] = "no"
        return response

    @action(detail=True, methods=["post"], url_path="upload")
    def upload_document(self, request, pk=None):
        """Upload a document to be indexed for RAG in this session's folder context."""
        session = self.get_object()
        file = request.FILES.get("file")

        if not file:
            return Response(
                {"detail": "No file provided."}, status=status.HTTP_400_BAD_REQUEST
            )

        doc = IndexedDocument.objects.create(
            folder=session.folder,
            file=file,
            filename=file.name,
            content_type=file.content_type or "application/octet-stream",
            source_type=IndexedDocument.SourceType.CHAT,
            source_content_type=ContentType.objects.get_for_model(ChatSession),
            source_object_id=session.id,
        )

        # Queue async indexing
        from .tasks import ingest_document

        ingest_document(str(doc.id))

        return Response(
            {"id": str(doc.id), "filename": doc.filename, "status": doc.status},
            status=status.HTTP_201_CREATED,
        )


class IndexedDocumentViewSet(BaseModelViewSet):
    """ViewSet for managing indexed documents."""

    model = IndexedDocument
    filterset_fields = ["folder", "source_type", "status"]


def _excel_safe(value):
    """Defang Excel formula injection: prefix a single quote to any string
    starting with =, +, -, @ so Excel treats the cell as text rather than
    evaluating it as a formula. The leading quote is invisible to humans
    but the literal value passes through unchanged when copy-pasted.
    """
    if not isinstance(value, str):
        return value
    if value and value[0] in ("=", "+", "-", "@"):
        return "'" + value
    return value


class QuestionnaireRunViewSet(BaseModelViewSet):
    """Experimental: questionnaire prefill runs."""

    model = QuestionnaireRun
    filterset_fields = ["folder", "status"]
    search_fields = ["title", "filename"]

    def perform_create(self, serializer):
        serializer.save(owner=self.request.user)

    @action(detail=False, methods=["post"], url_path="upload")
    def upload(self, request):
        """Multipart upload entry point: file + folder + optional title.

        Creates the run in PENDING and queues the parse task. Frontend then
        polls the detail endpoint for status to flip to PARSED.
        """
        file = request.FILES.get("file")
        folder_id = request.data.get("folder")
        title = (request.data.get("title") or "").strip()

        if not file:
            return Response(
                {"detail": "No file provided."}, status=status.HTTP_400_BAD_REQUEST
            )
        if not folder_id:
            return Response(
                {"detail": "folder is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Hardened content check — magic bytes + zip structure + macro reject
        # + zip-bomb guard. Catches renamed/forged/malicious uploads that the
        # filename + size validators alone would pass.
        from django.core.exceptions import ValidationError as DjValidationError

        from .upload_validation import validate_questionnaire_upload

        try:
            validate_questionnaire_upload(file)
        except DjValidationError as e:
            # ``validate_questionnaire_upload`` raises ValidationError with
            # curated, user-safe messages (e.g. "Not a valid .xlsx file"). If
            # ``messages`` is empty (shouldn't happen — defence in depth),
            # fall back to a static message and log the real exception
            # server-side rather than echoing arbitrary text to the client.
            if e.messages:
                detail = e.messages[0]
            else:
                logger.warning("upload_validation_unexpected", exc_info=True)
                detail = "Uploaded file failed validation."
            return Response(
                {"detail": detail},
                status=status.HTTP_400_BAD_REQUEST,
            )

        from django.contrib.auth.models import Permission

        from iam.models import Folder, RoleAssignment

        try:
            folder = Folder.objects.get(id=folder_id)
        except (Folder.DoesNotExist, ValueError):
            return Response(
                {"detail": "Folder not found."}, status=status.HTTP_404_NOT_FOUND
            )

        # Folder-scoped check: the user must hold add_questionnairerun in
        # this specific folder. Holding the perm globally is not enough —
        # the codebase's permission model is per-folder via role assignment.
        try:
            add_qr_perm = Permission.objects.get(codename="add_questionnairerun")
        except Permission.DoesNotExist:
            return Response(
                {"detail": "Server permission state is inconsistent."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
        if not RoleAssignment.is_access_allowed(
            user=request.user, perm=add_qr_perm, folder=folder
        ):
            return Response(
                {
                    "detail": "You do not have permission to upload a questionnaire to this folder."
                },
                status=status.HTTP_403_FORBIDDEN,
            )

        if not title:
            from core.utils import generate_friendly_name

            title = generate_friendly_name()

        run = QuestionnaireRun.objects.create(
            folder=folder,
            owner=request.user,
            title=title,
            file=file,
            filename=file.name,
        )

        from .tasks import parse_questionnaire

        parse_questionnaire(str(run.id))

        return Response(
            {"id": str(run.id), "status": run.status},
            status=status.HTTP_201_CREATED,
        )

    @action(detail=True, methods=["post"], url_path="extract-questions")
    def extract_questions(self, request, pk=None):
        """Materialize QuestionnaireQuestion rows from parsed_data + column_mapping.

        Idempotent: re-running deletes prior extracted rows. Refuses to run if
        any agent run already targets this questionnaire (would orphan
        AgentActions referencing deleted question rows).
        """
        run = self.get_object()
        if run.status != QuestionnaireRun.Status.PARSED:
            return Response(
                {"detail": f"Run must be parsed; currently {run.status}."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        mapping = run.column_mapping or {}
        if "sheet" not in mapping or "question_col" not in mapping:
            return Response(
                {"detail": "Column mapping not yet saved."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        existing_agent_runs = AgentRun.objects.filter(
            target_content_type=ContentType.objects.get_for_model(QuestionnaireRun),
            target_object_id=run.id,
        )
        if existing_agent_runs.exists():
            return Response(
                {
                    "detail": "Cannot re-extract: one or more agent runs already "
                    "reference this questionnaire."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        sheet_name = mapping["sheet"]
        sheet = next(
            (s for s in run.parsed_data.get("sheets", []) if s["name"] == sheet_name),
            None,
        )
        if not sheet:
            return Response(
                {"detail": f"Sheet '{sheet_name}' not found in parsed data."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        from django.db import transaction

        from .tasks import extract_questions_from_sheet, suggest_value_mapping

        # Atomic: if extraction blows up (parser bug, malformed sheet, transient
        # DB error), the prior questions stay intact so the run isn't stranded.
        with transaction.atomic():
            QuestionnaireQuestion.objects.filter(questionnaire_run=run).delete()
            created = extract_questions_from_sheet(run, sheet, mapping)

        # Each question now carries its detected answer_candidates. Group +
        # LLM-map them in the background so per-question mappings are ready
        # by the time the user clicks Start prefill / Download.
        suggest_value_mapping(str(run.id))

        return Response(
            {"questionnaire_run": str(run.id), "extracted": created},
            status=status.HTTP_201_CREATED,
        )

    @action(detail=True, methods=["patch"], url_path="mapping")
    def set_mapping(self, request, pk=None):
        """Persist the user's column-mapping choice for this run."""
        from .serializers import QuestionnaireRunMappingSerializer

        run = self.get_object()
        if run.status != QuestionnaireRun.Status.PARSED:
            return Response(
                {"detail": f"Run is {run.status}; mapping requires 'parsed'."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        serializer = QuestionnaireRunMappingSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        mapping = {k: v for k, v in serializer.validated_data.items() if v is not None}

        sheet_names = [s["name"] for s in run.parsed_data.get("sheets", [])]
        if mapping["sheet"] not in sheet_names:
            return Response(
                {"detail": f"Unknown sheet '{mapping['sheet']}'."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        run.column_mapping = mapping
        run.save(update_fields=["column_mapping", "updated_at"])

        # Note: value mapping is now triggered post-extract (it depends on
        # per-question answer_candidates which only exist once questions
        # have been materialized).
        return Response({"column_mapping": run.column_mapping})

    @action(detail=True, methods=["get"], url_path="export")
    def export_filled(self, request, pk=None):
        """Stream a copy of the original xlsx with response/comment columns
        filled from the latest non-expired propose_answer AgentActions.

        Walks the original sheet using the same row-skip logic as the extract
        step, so the answer/comment cells line up exactly with the extracted
        QuestionnaireQuestions. Original formatting, cover sheet, and any
        unmapped columns are preserved untouched.
        """
        import io
        import openpyxl

        from django.http import HttpResponse

        from .models import QuestionnaireQuestion, AgentRun, AgentAction

        run = self.get_object()
        mapping = run.column_mapping or {}
        if "sheet" not in mapping or "answer_col" not in mapping:
            return Response(
                {
                    "detail": "Mapping incomplete — at least an answer column "
                    "must be mapped to export."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        questions_by_ord = {
            q.ord: q
            for q in QuestionnaireQuestion.objects.filter(questionnaire_run=run)
        }
        if not questions_by_ord:
            return Response(
                {"detail": "No questions extracted yet."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Latest agent run that has produced answers (any terminal state).
        agent_run = (
            AgentRun.objects.filter(
                target_content_type=ContentType.objects.get_for_model(QuestionnaireRun),
                target_object_id=run.id,
            )
            .order_by("-created_at")
            .first()
        )
        if not agent_run:
            return Response(
                {"detail": "No agent run found for this questionnaire."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        qq_ct = ContentType.objects.get_for_model(QuestionnaireQuestion)
        actions_qs = (
            AgentAction.objects.filter(
                agent_run=agent_run,
                kind=AgentAction.Kind.PROPOSE_ANSWER,
                target_content_type=qq_ct,
            )
            .exclude(state=AgentAction.State.EXPIRED)
            .order_by("target_object_id", "-iteration")
        )

        # Latest iteration per question
        action_by_question: dict = {}
        for a in actions_qs:
            qid = str(a.target_object_id)
            if qid not in action_by_question:
                action_by_question[qid] = a

        # Honor the customer's vocabulary if we computed one; otherwise fall
        # back to internal labels for yes/partial/no.
        #
        # Two cases where we deliberately leave the answer cell blank:
        # 1. status == needs_info — never auto-write anything. Mapping it to
        #    "N/A" / "Not Applicable" is a real semantic mistake (those mean
        #    the requirement does not apply, not "I don't know"). Reviewer
        #    fills the cell manually.
        # 2. The column has a customer dropdown (candidates non-empty) but
        #    suggest_value_mapping is in fallback state — writing internal
        #    labels would violate the dropdown.
        # In both cases the comment cell still fills with the agent's text.
        DEFAULT_LABELS = {
            Verdict.YES: "Yes",
            Verdict.PARTIAL: "Partial",
            Verdict.NO: "No",
        }

        sheet_meta = next(
            (
                s
                for s in run.parsed_data.get("sheets", [])
                if s["name"] == mapping["sheet"]
            ),
            None,
        )
        if not sheet_meta:
            return Response(
                {"detail": f"Sheet '{mapping['sheet']}' not found."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        header_row = sheet_meta["header_row"]
        header_count = len(sheet_meta["headers"])
        q_col = mapping["question_col"]
        a_col = mapping["answer_col"]
        c_col = mapping.get("comment_col")

        with run.file.open("rb") as fp:
            content = fp.read()
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb[mapping["sheet"]]

        # Walk the same way extract does, in parallel with the in-memory ord index.
        ord_idx = 0
        for excel_row_idx_zero_based, row in enumerate(
            ws.iter_rows(values_only=True, max_row=ws.max_row)
        ):
            if excel_row_idx_zero_based <= header_row:
                continue
            cells = list(row[:header_count])
            if all(c in (None, "") for c in cells):
                continue
            text = cells[q_col] if q_col < len(cells) else None
            if text in (None, ""):
                ord_idx += 1
                continue
            text = str(text).strip()
            if not text:
                ord_idx += 1
                continue

            question = questions_by_ord.get(ord_idx)
            ord_idx += 1
            if question is None:
                continue
            action = action_by_question.get(str(question.id))
            if action is None:
                continue

            payload = action.payload or {}
            status_key = (payload.get("status") or Verdict.NEEDS_INFO).lower()
            comment = (payload.get("comment") or "").strip()
            # openpyxl is 1-indexed; column indices in mapping are 0-indexed
            excel_row = excel_row_idx_zero_based + 1

            qm = question.answer_mapping or {}
            qm_source = qm.get("source")
            if qm_source and qm_source != "fallback":
                cell_status_labels = {
                    Verdict.YES: qm.get(Verdict.YES),
                    Verdict.PARTIAL: qm.get(Verdict.PARTIAL),
                    Verdict.NO: qm.get(Verdict.NO),
                }
                skip_answer_cell = False
            elif question.answer_candidates:
                # Dropdown present but no clean mapping yet — don't pollute it.
                cell_status_labels = {}
                skip_answer_cell = True
            else:
                cell_status_labels = dict(DEFAULT_LABELS)
                skip_answer_cell = False

            should_write_answer = (
                not skip_answer_cell
                and status_key in cell_status_labels
                and bool(cell_status_labels[status_key])
            )
            if should_write_answer:
                ws.cell(
                    row=excel_row,
                    column=a_col + 1,
                    value=_excel_safe(cell_status_labels[status_key]),
                )
            if c_col is not None and comment:
                ws.cell(row=excel_row, column=c_col + 1, value=_excel_safe(comment))

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        # run.filename came from the upload's File.name — client-supplied.
        # A " or CRLF in there would break the Content-Disposition header
        # (response-splitting territory). Strip to a safe ASCII fallback for
        # the legacy `filename` token, then attach the RFC 5987 `filename*`
        # form so non-ASCII titles still render cleanly on modern clients.
        import re as _re
        import urllib.parse as _urlparse

        original_name = run.filename or "questionnaire.xlsx"
        stem = (
            original_name[:-5]
            if original_name.lower().endswith(".xlsx")
            else original_name
        )
        ascii_stem = (
            _re.sub(r"[^A-Za-z0-9._-]+", "_", stem).strip("._-") or "questionnaire"
        )
        ascii_name = f"{ascii_stem[:120]}__prefilled.xlsx"
        utf8_name = _urlparse.quote(f"{stem}__prefilled.xlsx", safe="")

        response = HttpResponse(
            out.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )
        response["Content-Disposition"] = (
            f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{utf8_name}"
        )
        return response


class QuestionnaireQuestionViewSet(BaseModelViewSet):
    model = QuestionnaireQuestion
    filterset_fields = ["questionnaire_run"]
    search_fields = ["text", "ref_id", "section"]

    @action(detail=True, methods=["post"], url_path="retry-with-control")
    def retry_with_control(self, request, pk=None):
        """Re-run a single question's answer pipeline with a chosen
        AppliedControl as priority context.

        Body: ``{applied_control_id}`` (a single id is the common case;
        accepts ``applied_control_ids`` list too).
        Synchronous — caller is the user clicking a button on the review row.
        """
        question = self.get_object()
        ids = request.data.get("applied_control_ids")
        single = request.data.get("applied_control_id")
        if not ids and single:
            ids = [single]
        if not ids or not isinstance(ids, list):
            return Response(
                {"detail": "applied_control_id (or applied_control_ids) required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        from core.models import AppliedControl
        from iam.models import Folder, RoleAssignment

        # The question's questionnaire run carries the folder; only allow
        # controls in that same folder, and only those the user can read.
        run_folder = question.questionnaire_run.folder
        readable_ac_ids, _, _ = RoleAssignment.get_accessible_object_ids(
            Folder.get_root_folder(), request.user, AppliedControl
        )
        ids_set = set(str(i) for i in ids) & set(str(i) for i in readable_ac_ids)
        valid_acs = AppliedControl.objects.filter(id__in=ids_set, folder=run_folder)
        valid_ids = [str(a.id) for a in valid_acs]
        if not valid_ids:
            return Response(
                {
                    "detail": "No accessible AppliedControl in this folder matched "
                    "the supplied id(s)."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        from .tasks import retry_question_with_hints

        try:
            result = retry_question_with_hints(
                question_id=str(question.id),
                hint_applied_control_ids=valid_ids,
            )
        except ValueError:
            # ValueErrors here are controlled, user-safe ("No agent run found
            # for this questionnaire.") — but we don't echo them to avoid
            # leaking future raises that might carry internal state. Log the
            # original on the server, return a static message.
            logger.warning(
                "retry_with_control_invalid_state",
                question_id=str(question.id),
                exc_info=True,
            )
            return Response(
                {"detail": "Cannot retry this question in its current state."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        except Exception:
            logger.error(
                "retry_with_control_failed", question_id=str(question.id), exc_info=True
            )
            return Response(
                {"detail": "Retry failed. Check server logs."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        return Response(result, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="suggest-control")
    def suggest_control(self, request, pk=None):
        """Synchronous LLM call: draft an AppliedControl that would let us
        answer this question. Returns the draft as JSON for the user to edit
        before creating.
        """
        question = self.get_object()
        from .tasks import draft_applied_control_for_question

        draft = draft_applied_control_for_question(question.text)
        return Response(draft, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], url_path="create-and-retry")
    def create_and_retry(self, request, pk=None):
        """Create an AppliedControl in the question's folder from a (possibly
        user-edited) draft, then immediately retry the question with the new
        control as priority context.
        """
        from django.contrib.auth.models import Permission

        from core.models import AppliedControl
        from iam.models import RoleAssignment

        question = self.get_object()
        folder = question.questionnaire_run.folder

        # User must hold add_appliedcontrol in this folder. View access alone
        # is not enough — Approver / Auditor roles can read a folder's
        # controls but must not be able to write through this path.
        try:
            add_ac_perm = Permission.objects.get(codename="add_appliedcontrol")
        except Permission.DoesNotExist:
            return Response(
                {"detail": "Server permission state is inconsistent."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
        if not RoleAssignment.is_access_allowed(
            user=request.user, perm=add_ac_perm, folder=folder
        ):
            return Response(
                {
                    "detail": "You do not have permission to add an applied "
                    "control in this folder."
                },
                status=status.HTTP_403_FORBIDDEN,
            )

        name = (request.data.get("name") or "").strip()
        if not name:
            return Response(
                {"detail": "name is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        description = (request.data.get("description") or "").strip()
        observation = (request.data.get("observation") or "").strip()
        ref_id = (request.data.get("ref_id") or "").strip()[:100]
        status_value = (request.data.get("status") or "to_do").strip().lower()
        category = (request.data.get("category") or "").strip().lower()
        csf_function = (request.data.get("csf_function") or "").strip().lower()

        from .tasks import (
            _VALID_AC_CATEGORIES,
            _VALID_AC_CSF_FUNCTIONS,
            _VALID_AC_STATUSES,
        )

        if status_value not in _VALID_AC_STATUSES:
            status_value = "to_do"
        if category and category not in _VALID_AC_CATEGORIES:
            category = ""
        if csf_function and csf_function not in _VALID_AC_CSF_FUNCTIONS:
            csf_function = ""

        ac_kwargs = {
            "folder": folder,
            "name": name[:200],
            "description": description,
            "observation": observation,
            "status": status_value,
        }
        if ref_id:
            ac_kwargs["ref_id"] = ref_id
        if category:
            ac_kwargs["category"] = category
        if csf_function:
            ac_kwargs["csf_function"] = csf_function

        ac = AppliedControl.objects.create(**ac_kwargs)
        logger.info(
            "Created AppliedControl %s from question %s suggestion",
            ac.id,
            question.id,
        )

        from .tasks import retry_question_with_hints

        try:
            result = retry_question_with_hints(
                question_id=str(question.id),
                hint_applied_control_ids=[str(ac.id)],
            )
        except Exception:
            logger.error(
                "create_and_retry_retry_failed",
                question_id=str(question.id),
                applied_control_id=str(ac.id),
                exc_info=True,
            )
            # The control was created successfully; retry can be re-issued
            # via the existing retry-with-control endpoint. Suppress the
            # exception detail to avoid leaking internals — the control id
            # is enough for the client to retry.
            return Response(
                {
                    "applied_control_id": str(ac.id),
                    "retry_failed": True,
                    "detail": "Control created, but the retry step failed. "
                    "Use the existing-control retry to try again.",
                },
                status=status.HTTP_207_MULTI_STATUS,
            )

        return Response(
            {"applied_control_id": str(ac.id), **result},
            status=status.HTTP_201_CREATED,
        )


_STUCK_RUN_HEARTBEAT_GRACE_SECONDS = 5 * 60


def _check_agent_dependencies() -> str:
    """Cheap preflight before enqueueing an agent run.

    Probes the three pieces of infrastructure the worker can't run without:
    LLM provider, embedder, and Qdrant (collection present + reachable).
    Returns an empty string on success, or a user-facing message describing
    the first failing dependency. Catches the common "Qdrant isn't running"
    / "Ollama not configured" failure modes that would otherwise produce a
    run full of silent ``needs_info`` defaults from the worker's per-question
    exception handler.

    Exception details are logged server-side; the returned message is a
    static label so we don't echo stack-trace fragments (provider URLs,
    credentials, internal paths) to API clients.
    """
    from .providers import get_llm, get_embedder
    from .rag import COLLECTION_NAME, get_qdrant_client

    try:
        get_llm()
    except Exception:
        logger.exception("agent_dependency_check_failed: llm_provider")
        return "LLM provider not available. Check server logs for details."
    try:
        get_embedder()
    except Exception:
        logger.exception("agent_dependency_check_failed: embedder")
        return "Embedder not available. Check server logs for details."
    try:
        client = get_qdrant_client()
        collection_names = {c.name for c in client.get_collections().collections}
    except Exception:
        logger.exception("agent_dependency_check_failed: qdrant")
        return "Vector store (Qdrant) unreachable. Check server logs for details."
    if COLLECTION_NAME not in collection_names:
        return (
            f"Vector store collection '{COLLECTION_NAME}' is missing. "
            "Run `manage.py init_qdrant` first."
        )
    return ""


def _heal_if_stuck(run):
    """Lazy stuck-run detector.

    The agent worker heartbeats before each question (`_heartbeat()` in
    chat.questionnaire). If an AgentRun is RUNNING but its last_heartbeat is
    older than the grace period, the worker has almost certainly died —
    flip the run to FAILED so the UI stops spinning forever.
    Called from the polling/retrieve paths so the heal is opportunistic.
    """
    if not run or run.status != AgentRun.Status.RUNNING:
        return
    if not run.last_heartbeat_at:
        return
    age = (timezone.now() - run.last_heartbeat_at).total_seconds()
    if age <= _STUCK_RUN_HEARTBEAT_GRACE_SECONDS:
        return
    run.status = AgentRun.Status.FAILED
    run.finished_at = timezone.now()
    msg = (
        f"Worker heartbeat stale for {int(age)} s — marking the run as "
        "failed. Restart Huey and re-run if you want a clean retry."
    )
    run.error_message = (
        f"{run.error_message}\n{msg}".strip() if run.error_message else msg
    )
    run.save(
        update_fields=[
            "status",
            "finished_at",
            "error_message",
            "updated_at",
        ]
    )
    logger.warning(
        "AgentRun %s healed from stuck-running (heartbeat %ds old)",
        run.id,
        int(age),
    )


class AgentRunViewSet(BaseModelViewSet):
    model = AgentRun
    filterset_fields = ["folder", "kind", "status", "target_object_id"]
    search_fields = ["current_step_label"]
    # Agent runs are server-spawned: generic POST/PUT/PATCH/DELETE would let
    # a client point a run at any ContentType+UUID and bypass the per-folder
    # add_agentrun + per-target read checks done in start_questionnaire_prefill.
    # PUT/PATCH/DELETE dropped entirely; POST stays only for @action routes.
    http_method_names = ["get", "head", "options", "post"]

    def create(self, request, *args, **kwargs):
        return Response(
            {
                "detail": "Use /chat/agent-runs/start-questionnaire-prefill/ "
                "to start an agent run.",
            },
            status=status.HTTP_405_METHOD_NOT_ALLOWED,
        )

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        _heal_if_stuck(instance)
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @action(detail=False, methods=["post"], url_path="start-questionnaire-prefill")
    def start_questionnaire_prefill(self, request):
        """Create + enqueue an AgentRun for prefilling a questionnaire.

        Body: {questionnaire_run, strictness}.
        """
        from .serializers import StartQuestionnairePrefillSerializer

        serializer = StartQuestionnairePrefillSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        try:
            qr = QuestionnaireRun.objects.get(id=data["questionnaire_run"])
        except QuestionnaireRun.DoesNotExist:
            return Response(
                {"detail": "Questionnaire run not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        from django.contrib.auth.models import Permission

        from iam.models import RoleAssignment

        # Two-layer check: the user must be able to read the questionnaire run
        # itself, AND must be able to add an AgentRun in this folder. View
        # access alone shouldn't license LLM-driven writes.
        if not RoleAssignment.is_object_readable(request.user, QuestionnaireRun, qr.id):
            return Response(
                {"detail": "You do not have access to this questionnaire."},
                status=status.HTTP_403_FORBIDDEN,
            )
        try:
            add_agentrun_perm = Permission.objects.get(codename="add_agentrun")
        except Permission.DoesNotExist:
            return Response(
                {"detail": "Server permission state is inconsistent."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
        if not RoleAssignment.is_access_allowed(
            user=request.user, perm=add_agentrun_perm, folder=qr.folder
        ):
            return Response(
                {
                    "detail": "You do not have permission to start an agent "
                    "run in this folder."
                },
                status=status.HTTP_403_FORBIDDEN,
            )

        question_count = qr.questions.count()
        if question_count == 0:
            return Response(
                {"detail": "No questions extracted yet — extract first."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Fail fast if the worker's required infrastructure is unreachable.
        # Without this, a Qdrant outage would still let the run start and
        # then silently produce a needs_info for every question via the
        # per-question exception handler in chat.questionnaire.
        dep_error = _check_agent_dependencies()
        if dep_error:
            return Response(
                {"detail": dep_error},
                status=status.HTTP_503_SERVICE_UNAVAILABLE,
            )

        # Lock the parent QuestionnaireRun row so two concurrent Start clicks
        # can't both pass the active-run check and each insert a new AgentRun.
        # The check + create are now atomic relative to that lock.
        from django.db import transaction

        with transaction.atomic():
            QuestionnaireRun.objects.select_for_update().get(id=qr.id)
            active_exists = AgentRun.objects.filter(
                target_content_type=ContentType.objects.get_for_model(QuestionnaireRun),
                target_object_id=qr.id,
                status__in=[AgentRun.Status.QUEUED, AgentRun.Status.RUNNING],
            ).exists()
            if active_exists:
                return Response(
                    {
                        "detail": "An agent run is already in progress for this questionnaire."
                    },
                    status=status.HTTP_409_CONFLICT,
                )

            agent_run = AgentRun.objects.create(
                owner=request.user,
                folder=qr.folder,
                kind=AgentRun.Kind.QUESTIONNAIRE_PREFILL,
                strictness=data["strictness"],
                target_content_type=ContentType.objects.get_for_model(QuestionnaireRun),
                target_object_id=qr.id,
                total_steps=question_count,
            )

        from .tasks import run_questionnaire_prefill

        run_questionnaire_prefill(str(agent_run.id))

        return Response(
            {"id": str(agent_run.id), "status": agent_run.status},
            status=status.HTTP_201_CREATED,
        )

    @action(detail=True, methods=["post"], url_path="cancel")
    def cancel(self, request, pk=None):
        """Mark an active run as cancelled. The worker checks before each step."""
        run = self.get_object()
        if run.status not in (AgentRun.Status.QUEUED, AgentRun.Status.RUNNING):
            return Response(
                {"detail": f"Run is already {run.status}; cannot cancel."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        run.status = AgentRun.Status.CANCELLED
        run.finished_at = timezone.now()
        run.save(update_fields=["status", "finished_at", "updated_at"])
        return Response({"status": run.status})


class AgentActionViewSet(BaseModelViewSet):
    model = AgentAction
    filterset_fields = [
        "agent_run",
        "kind",
        "state",
        "target_content_type",
        "target_object_id",
    ]
    # Agent actions are the AI audit trail. The worker creates them; users
    # only transition state via approve/reject. Block generic mutation paths.
    http_method_names = ["get", "head", "options", "post"]

    def create(self, request, *args, **kwargs):
        return Response(
            {
                "detail": "Agent actions are created by the agent, not via this endpoint.",
            },
            status=status.HTTP_405_METHOD_NOT_ALLOWED,
        )

    # Approve / reject are intentionally not exercised by the questionnaire
    # autopilot UI — that flow uses confidence-banded review without explicit
    # per-action approval. They're reserved for the upcoming audit-prefill
    # page where a human signs off each RequirementAssessment proposal before
    # it lands on the audit. State transitions stay PROPOSED → APPROVED /
    # REJECTED; the worker also writes the terminal EXPIRED state during
    # retry iterations (see _process_question in chat/questionnaire.py).
    @action(detail=True, methods=["post"], url_path="approve")
    def approve(self, request, pk=None):
        action_obj = self.get_object()
        if action_obj.state != AgentAction.State.PROPOSED:
            return Response(
                {"detail": f"Action is {action_obj.state}; cannot approve."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        action_obj.state = AgentAction.State.APPROVED
        action_obj.approved_by = request.user
        action_obj.approved_at = timezone.now()
        action_obj.save(
            update_fields=["state", "approved_by", "approved_at", "updated_at"]
        )
        return Response({"state": action_obj.state})

    @action(detail=True, methods=["post"], url_path="reject")
    def reject(self, request, pk=None):
        action_obj = self.get_object()
        if action_obj.state != AgentAction.State.PROPOSED:
            return Response(
                {"detail": f"Action is {action_obj.state}; cannot reject."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        action_obj.state = AgentAction.State.REJECTED
        action_obj.save(update_fields=["state", "updated_at"])
        return Response({"state": action_obj.state})


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def chat_status(request):
    """Check chat service health: Ollama availability, index status."""
    from .providers import get_chat_settings

    settings = get_chat_settings()
    ollama_ok = is_ollama_available()

    return Response(
        {
            "llm_provider": settings.get("llm_provider", "ollama"),
            "ollama_available": ollama_ok,
            "ollama_url": settings["ollama_base_url"],
            "ollama_model": settings["ollama_model"],
            "embedding_backend": settings["embedding_backend"],
        }
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def ollama_models(request):
    """List available models from the Ollama server."""
    from .providers import get_chat_settings

    settings = get_chat_settings()
    base_url = settings["ollama_base_url"]

    try:
        import httpx

        resp = httpx.get(f"{base_url}/api/tags", timeout=5)
        resp.raise_for_status()
        data = resp.json()
        models = [
            {
                "name": m["name"],
                "size": m.get("size"),
                "modified_at": m.get("modified_at"),
            }
            for m in data.get("models", [])
        ]
        return Response({"models": models})
    except Exception as e:
        logger.warning("Failed to fetch Ollama models: %s", e)
        return Response(
            {"models": [], "error": "Unable to connect to Ollama service."},
            status=status.HTTP_503_SERVICE_UNAVAILABLE,
        )


def _get_last_query_meta(session) -> dict | None:
    """
    Find the most recent query metadata from the session's assistant messages.
    Used to support follow-up queries like 'next page', 'give me more'.
    """
    last_msgs = session.messages.filter(role=ChatMessage.Role.ASSISTANT).order_by(
        "-created_at"
    )[:3]
    for msg in last_msgs:
        if not msg.context_refs:
            continue
        for ref in msg.context_refs:
            if isinstance(ref, dict) and ref.get("source") == "query_meta":
                return ref
    return None


def _build_context_prompt(page_context: dict, parsed_context) -> str:
    """
    Build a rich context prompt for the LLM based on the user's current page.
    Includes structured hints about available contextual actions.
    """
    if not page_context:
        return ""

    from .tools import PARENT_CHILD_MAP, ATTACHABLE_RELATIONS, MODEL_MAP

    parts = []
    page_path = page_context.get("path", "")
    page_model = page_context.get("model", "")
    page_title = page_context.get("title", "")

    info_parts = []
    if page_path:
        info_parts.append(f"path={page_path}")
    if page_model:
        info_parts.append(f"model={page_model}")
    if page_title:
        info_parts.append(f"title={page_title}")

    if not info_parts:
        return ""

    parts.append(
        f"The user is currently viewing: {', '.join(info_parts)}.\n"
        "NOTE: The page title and path above are for context only.\n"
        "- Do NOT use them as search filters or domain filters in query_objects.\n"
        "- When querying child objects of this page, rely on automatic parent scoping (no search/domain needed).\n"
        "- For follow-up questions ('dessus', 'on it', 'about that'), refer to the CONVERSATION HISTORY, "
        "not the current page."
    )

    # Add contextual action hints when on a detail/edit page
    if parsed_context and parsed_context.object_id:
        actions = []

        # Child creation hints
        for child_key, fk_field in PARENT_CHILD_MAP.get(parsed_context.model_key, []):
            if child_key in MODEL_MAP:
                child_display = MODEL_MAP[child_key][2]
                actions.append(
                    f"- Create {child_display} linked to this {page_model or parsed_context.model_key}"
                )

        # Attachment hints
        for rel_key, m2m_field in ATTACHABLE_RELATIONS.get(
            parsed_context.model_key, []
        ):
            if rel_key in MODEL_MAP:
                rel_display = MODEL_MAP[rel_key][2]
                actions.append(
                    f"- Attach existing {rel_display} to this {page_model or parsed_context.model_key} "
                    f"(use attach_existing tool with related_model='{rel_key}')"
                )

        if actions:
            parts.append("\nContextual actions available on this page:")
            parts.extend(actions)
            parts.append(
                "\nIMPORTANT: When the user asks about implementing, complying, "
                "what controls to add, or what to do for a requirement — "
                "DO NOT just describe actions in text. Instead, USE the attach_existing "
                "or propose_create tools to provide actionable confirmation cards. "
                "The user expects interactive buttons, not instructions."
            )

    parts.append(
        "\nUse this context if the user refers to 'this', 'these', 'here', "
        "or asks about items on their current page.\n"
    )

    return "\n".join(parts) + "\n"


def _enrich_context(parsed_context, accessible_folder_ids: list[str]) -> str:
    """
    Enrich the LLM context with domain objects relevant to the current page.
    For example, when on a risk assessment page, include the domain's assets
    so the LLM can reason about additional risks.
    """
    from .tools import MODEL_MAP

    # Determine the parent object's folder
    parent_info = MODEL_MAP.get(parsed_context.model_key)
    if not parent_info:
        return ""

    try:
        parent_model = apps.get_model(parent_info[0], parent_info[1])
        parent_obj = parent_model.objects.filter(id=parsed_context.object_id).first()
        if not parent_obj or not hasattr(parent_obj, "folder_id"):
            return ""
        folder_id = str(parent_obj.folder_id)
    except Exception:
        return ""

    parts = []

    # For risk assessments: include assets from the same domain
    if parsed_context.model_key == "risk_assessment":
        Asset = apps.get_model("core", "Asset")
        assets = Asset.objects.filter(
            folder_id=folder_id,
            folder_id__in=accessible_folder_ids,
        )[:30]
        if assets:
            parts.append("ASSETS IN THIS DOMAIN:")
            for asset in assets:
                line = f"  - {asset.name}"
                extras = []
                if hasattr(asset, "type") and asset.type:
                    extras.append(
                        asset.get_type_display()
                        if hasattr(asset, "get_type_display")
                        else asset.type
                    )
                if hasattr(asset, "business_value") and asset.business_value:
                    extras.append(f"business_value={asset.business_value}")
                if extras:
                    line += f" ({', '.join(extras)})"
                if asset.description:
                    line += f" — {asset.description[:150]}"
                parts.append(line)

        # Also include existing risk scenarios for awareness
        RiskScenario = apps.get_model("core", "RiskScenario")
        scenarios = RiskScenario.objects.filter(
            risk_assessment_id=parsed_context.object_id,
        )[:20]
        if scenarios:
            parts.append("\nEXISTING RISK SCENARIOS (already identified):")
            for s in scenarios:
                line = f"  - {s.name}"
                if hasattr(s, "treatment") and s.treatment:
                    line += f" (treatment={s.get_treatment_display()})"
                parts.append(line)

    return "\n".join(parts)


def _get_graph_context(user_message: str) -> str:
    """
    Query the knowledge graph for framework-related context.
    Extracts potential framework references from the user message and
    returns structured graph data if matches are found.

    This runs on every RAG fallback to ensure framework knowledge is always
    available, regardless of whether the LLM selected the search_library tool.
    """
    from .knowledge_graph import (
        find_frameworks,
        get_framework_detail,
        format_graph_result,
    )

    # Try finding frameworks mentioned in the query
    # Split on common separators and try each token + the full query
    tokens = set()
    for sep in [" et ", " and ", " vs ", " versus ", " ou ", " or ", ",", "/"]:
        if sep in user_message.lower():
            for part in user_message.lower().split(sep):
                cleaned = part.strip().strip("?!.\"'")
                if len(cleaned) >= 2:
                    tokens.add(cleaned)

    # Also try individual words (3+ chars) for framework name matching
    for word in user_message.split():
        cleaned = word.strip("?!.,\"'()").strip()
        if len(cleaned) >= 3:
            tokens.add(cleaned.lower())

    # Search the graph for each token
    found_frameworks = {}
    for token in tokens:
        results = find_frameworks(query=token)
        for fw in results:
            if fw["urn"] not in found_frameworks:
                found_frameworks[fw["urn"]] = fw

    if not found_frameworks:
        return ""

    # Get details for found frameworks (cap at 5 to avoid context overflow)
    parts = []
    for urn in list(found_frameworks)[:5]:
        detail = get_framework_detail(found_frameworks[urn]["name"])
        if detail:
            parts.append(format_graph_result(detail))

    if not parts:
        return ""

    return "\n\n---\n\n".join(parts)


# ---------------------------------------------------------------------------
# Deterministic workflow pre-routing
# ---------------------------------------------------------------------------

# Maps (model_key, keyword_set) → workflow_name.
# If the user is on a matching page and the message contains any keyword,
# the workflow is selected without asking the LLM.
_WORKFLOW_ROUTES: list[tuple[str, set[str], str]] = [
    (
        "requirement_assessment",
        {
            # EN
            "control",
            "controls",
            "implement",
            "comply",
            "compliance",
            "what to do",
            "what should",
            "how to",
            "suggest",
            "recommend",
            "measure",
            "measures",
            # FR
            "contrôle",
            "contrôles",
            "mesure",
            "mesures",
            "implémenter",
            "conformité",
            "conformer",
            "que faire",
            "quoi faire",
            "comment",
            "suggérer",
            "recommander",
        },
        "suggest_controls",
    ),
    (
        "risk_scenario",
        {
            # EN
            "treat",
            "treatment",
            "mitigat",
            "control",
            "controls",
            "reduce",
            "accept",
            "avoid",
            "transfer",
            "what to do",
            "how to",
            "suggest",
            "recommend",
            # FR
            "traitement",
            "traiter",
            "atténuer",
            "contrôle",
            "contrôles",
            "réduire",
            "accepter",
            "éviter",
            "transférer",
            "que faire",
            "comment",
            "suggérer",
            "recommander",
        },
        "risk_treatment",
    ),
    (
        "applied_control",
        {
            # EN
            "evidence",
            "evidences",
            "proof",
            "prove",
            "document",
            "what evidence",
            "how to demonstrate",
            "audit",
            # FR
            "preuve",
            "preuves",
            "documenter",
            "démontrer",
            "audit",
        },
        "evidence_guidance",
    ),
    (
        "ebios_rm_study",
        {
            # EN
            "conduct",
            "assist",
            "help",
            "study",
            "workshop",
            "generate",
            "fill",
            "draft",
            # FR
            "mener",
            "conduire",
            "assister",
            "aider",
            "étude",
            "atelier",
            "générer",
            "remplir",
            "ébauche",
        },
        "ebios_rm_assist",
    ),
    (
        "folder",
        {
            # EN
            "ebios",
            "ebios rm",
            "study",
            "risk study",
            # FR
            "étude",
            "etude",
            "analyse de risque",
            "analyse des risques",
        },
        "ebios_rm_assist",
    ),
]


def _match_workflow(user_message: str, parsed_context) -> "Workflow | None":
    """
    Match a workflow based on page context + message keywords.
    Returns the workflow instance if matched, None otherwise.
    """
    msg_lower = user_message.lower()

    # Context-specific routes (require being on a matching detail page)
    if parsed_context and parsed_context.object_id:
        for model_key, keywords, workflow_name in _WORKFLOW_ROUTES:
            if parsed_context.model_key != model_key:
                continue
            if any(kw in msg_lower for kw in keywords):
                from .workflows import get_workflow_by_tool_name

                return get_workflow_by_tool_name(f"workflow_{workflow_name}")

    # Global routes — match from ANY page when keywords are strong enough
    if any(kw in msg_lower for kw in ("ebios", "ebios rm")):
        from .workflows import get_workflow_by_tool_name

        return get_workflow_by_tool_name("workflow_ebios_rm_assist")
