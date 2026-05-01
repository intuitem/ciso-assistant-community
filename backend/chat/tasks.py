"""
Huey background tasks for document ingestion and object indexing.
"""

import structlog
import uuid

from django.utils import timezone
from huey.contrib.djhuey import db_task

logger = structlog.get_logger(__name__)


VALUE_MAPPING_PROMPT = """A customer security questionnaire uses a controlled \
vocabulary in its answer column: {candidates}.

Map each of these three internal status labels to whichever customer value \
best expresses the same idea:
- "yes": we fully meet the requirement
- "partial": we partially meet the requirement
- "no": we do not meet the requirement

IMPORTANT: do NOT include "needs_info" in your mapping. When the agent cannot \
answer, the cell will be left blank for human review — never auto-mapped to \
"N/A" or "Not Applicable" (which is a legitimate compliance answer meaning \
the requirement does not apply, not "I don't know").

Each mapped value MUST be exactly one of the strings from the customer list \
(verbatim). If no customer value reasonably expresses one of the labels, \
reuse the closest one.

Reply with a single JSON object, no prose:
{{"yes": "...", "partial": "...", "no": "..."}}"""


@db_task()
def suggest_value_mapping(run_id: str):
    """Compute per-question value mappings for a run.

    Walks the run's questions, groups them by their detected
    ``answer_candidates`` signature, runs the LLM once per distinct
    vocabulary, and persists each question's ``answer_mapping``. Also fills
    ``QuestionnaireRun.value_mapping`` with a summary (the dominant
    vocabulary, or a flag indicating multiple) so the UI can hint at what
    the export will write.
    """
    from .models import QuestionnaireRun, QuestionnaireQuestion
    from .providers import get_llm

    try:
        run = QuestionnaireRun.objects.get(id=run_id)
    except QuestionnaireRun.DoesNotExist:
        logger.error("QuestionnaireRun %s not found", run_id)
        return

    questions = list(QuestionnaireQuestion.objects.filter(questionnaire_run=run))
    if not questions:
        logger.info(
            "suggest_value_mapping: no questions for run %s yet — nothing to map",
            run_id,
        )
        return

    # Group questions by their candidate vocabulary. Empty candidates means
    # free-text answer cell — those questions just use internal labels.
    groups: dict[tuple[str, ...], list[QuestionnaireQuestion]] = {}
    for q in questions:
        sig = tuple(q.answer_candidates or [])
        groups.setdefault(sig, []).append(q)

    try:
        llm = get_llm()
    except Exception as e:
        logger.warning("LLM unavailable for value mapping: %s", e)
        llm = None

    summary_mappings: list[dict] = []
    for sig, group_questions in groups.items():
        if not sig:
            # Free-text questions — use fallback internal labels
            mapping_for_group = {
                "yes": "Yes",
                "partial": "Partial",
                "no": "No",
                "candidates": [],
                "source": "fallback",
            }
        else:
            candidates = list(sig)
            mapping_for_group = _llm_map_candidates(llm, candidates)

        # Persist on each question in the group
        ids = [q.id for q in group_questions]
        QuestionnaireQuestion.objects.filter(id__in=ids).update(
            answer_mapping=mapping_for_group
        )
        summary_mappings.append(
            {**mapping_for_group, "question_count": len(group_questions)}
        )

    # Build the run-level summary. Most common vocab wins for the UI hint.
    summary_mappings.sort(key=lambda m: m.get("question_count", 0), reverse=True)
    primary = summary_mappings[0] if summary_mappings else {}

    distinct_non_fallback = sum(
        1 for m in summary_mappings if m.get("source") not in (None, "fallback")
    )
    has_multiple_distinct_vocabs = (
        sum(1 for m in summary_mappings if m.get("candidates")) > 1
    )

    run_summary = {
        "yes": primary.get("yes", "Yes"),
        "partial": primary.get("partial", "Partial"),
        "no": primary.get("no", "No"),
        "candidates": primary.get("candidates", []),
        "source": primary.get("source", "fallback"),
        "vocab_count": len(summary_mappings),
        "has_multiple_vocabs": has_multiple_distinct_vocabs,
    }
    run.value_mapping = run_summary
    run.save(update_fields=["value_mapping", "updated_at"])

    logger.info(
        "Value mapping done for run %s: %d question(s) across %d vocab(s) "
        "(%d non-fallback)",
        run_id,
        len(questions),
        len(summary_mappings),
        distinct_non_fallback,
    )


def _llm_map_candidates(llm, candidates: list[str]) -> dict:
    """LLM-driven mapping of yes/partial/no onto a candidate list, with retries.

    Returns the mapping dict to store on a question (or all questions sharing
    this vocabulary). Falls back to internal labels with ``source='fallback'``
    if the LLM is unavailable or doesn't produce a valid mapping after retries.
    """
    import time

    if llm is None or not candidates:
        return {
            "yes": "Yes",
            "partial": "Partial",
            "no": "No",
            "candidates": candidates,
            "source": "fallback",
        }

    MAX_ATTEMPTS = 3
    BACKOFF_SECONDS = (0, 2, 5)
    candidate_set = {c.lower(): c for c in candidates}

    def pick(parsed: dict, label: str) -> str | None:
        v = parsed.get(label)
        if not isinstance(v, str):
            return None
        if v in candidates:
            return v
        return candidate_set.get(v.lower())

    last_raw = ""
    for attempt in range(MAX_ATTEMPTS):
        if attempt > 0:
            time.sleep(BACKOFF_SECONDS[attempt])

        prompt = VALUE_MAPPING_PROMPT.format(candidates=json_dumps(candidates))
        try:
            raw = llm.generate(prompt=prompt, context="", history=[])
        except Exception as e:
            logger.warning(
                "LLM call failed for value mapping (attempt %d/%d): %s",
                attempt + 1,
                MAX_ATTEMPTS,
                e,
            )
            continue

        last_raw = raw or ""
        parsed = _parse_json_response(raw) or {}
        yes = pick(parsed, "yes")
        partial = pick(parsed, "partial")
        no = pick(parsed, "no")
        if all([yes, partial, no]):
            # We deliberately do NOT include needs_info — export leaves those
            # cells blank rather than risk auto-mapping to "N/A" (which is a
            # legitimate "Not Applicable" answer, not "I don't know").
            return {
                "yes": yes,
                "partial": partial,
                "no": no,
                "candidates": candidates,
                # Vocabulary came from a non-empty answer cell (data validation
                # or distinct existing values); export honors it.
                "source": "data_validation",
            }
        logger.info(
            "LLM produced incomplete mapping (attempt %d/%d, raw=%r) — retrying",
            attempt + 1,
            MAX_ATTEMPTS,
            (raw or "")[:200],
        )

    logger.warning(
        "Value mapping fallback after %d attempts (last raw=%r)",
        MAX_ATTEMPTS,
        last_raw[:200],
    )
    return {
        "yes": "Yes",
        "partial": "Partial",
        "no": "No",
        "candidates": candidates,
        "source": "fallback",
    }


def json_dumps(obj):
    """Stable JSON for prompt embedding."""
    import json

    return json.dumps(obj, ensure_ascii=False)


def extract_questions_from_sheet(run, sheet: dict, mapping: dict) -> int:
    """Materialize QuestionnaireQuestion rows from a sheet preview + mapping.

    Synchronous helper, called from the API. Walks the workbook to capture
    every question-bearing row. For each row, also detects the answer cell's
    data-validation vocabulary so a per-question mapping can be computed
    (some questionnaires use different dropdowns per row range).
    Returns the count of created questions.
    """
    import io
    import openpyxl

    from .models import QuestionnaireQuestion

    q_col = mapping["question_col"]
    answer_col = mapping.get("answer_col")
    section_col = mapping.get("section_col")
    sheet_name = sheet["name"]
    header_row = sheet["header_row"]

    with run.file.open("rb") as fp:
        content = fp.read()
    # Not read_only: ws.data_validations is only accessible on a fully
    # loaded workbook, and we need it for per-cell vocabulary detection.
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb[sheet_name]
    header_count = len(sheet["headers"])

    rows_to_create = []
    ord_idx = 0
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx <= header_row:
            continue
        cells = list(row[:header_count])
        if all(c in (None, "") for c in cells):
            continue
        text = cells[q_col] if q_col < len(cells) else None
        if text in (None, ""):
            ord_idx += 1
            continue
        text = str(text).strip()
        if not text:
            ord_idx += 1
            continue
        section = ""
        if section_col is not None and section_col < len(cells):
            section_val = cells[section_col]
            if section_val not in (None, ""):
                section = str(section_val).strip()[:200]

        ref_id = ""

        # Per-cell vocabulary: openpyxl uses 1-indexed Excel rows.
        candidates: list[str] = []
        if answer_col is not None:
            candidates = _detect_cell_vocabulary(ws, idx + 1, answer_col)

        rows_to_create.append(
            QuestionnaireQuestion(
                questionnaire_run=run,
                ord=ord_idx,
                ref_id=ref_id,
                section=section,
                text=text,
                answer_candidates=candidates,
            )
        )
        ord_idx += 1

    QuestionnaireQuestion.objects.bulk_create(rows_to_create)
    return len(rows_to_create)


def _detect_cell_vocabulary(ws, excel_row: int, col_idx: int) -> list[str]:
    """Return the data-validation list values that apply to one cell, or [].

    Scans ``ws.data_validations`` for a list-type validation whose ``sqref``
    covers the cell at (excel_row, col_idx). Returns the verbatim values from
    the validation's formula1 (inline or range-referenced).
    """
    import openpyxl.utils

    col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
    coord = f"{col_letter}{excel_row}"
    try:
        for dv in ws.data_validations.dataValidation:
            if getattr(dv, "type", None) != "list":
                continue
            sqref = getattr(dv, "sqref", None)
            try:
                covers = bool(sqref) and coord in sqref
            except Exception:
                covers = False
            if not covers:
                continue
            values = _parse_data_validation_list(ws, dv.formula1)
            if values:
                return values
    except Exception as e:
        logger.debug("cell_vocabulary_detection_failed %s: %s", coord, e)
    return []


@db_task()
def parse_questionnaire(run_id: str):
    """Open the uploaded xlsx, capture sheet/header structure for review.

    No interpretation of which column is question/answer — that's the user's
    call in the mapping UI. We only surface what the file actually contains.
    """
    import io
    import openpyxl

    from .models import QuestionnaireRun

    try:
        run = QuestionnaireRun.objects.get(id=run_id)
    except QuestionnaireRun.DoesNotExist:
        logger.error("QuestionnaireRun %s not found", run_id)
        return

    run.status = QuestionnaireRun.Status.PARSING
    run.save(update_fields=["status"])

    try:
        with run.file.open("rb") as fp:
            content = fp.read()

        # Not read_only: we need access to data_validations (Excel dropdowns)
        # which are only exposed on a fully-loaded workbook.
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

        sheets = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_row_idx, headers = _detect_headers(ws)
            row_count, rows_preview = _collect_rows(ws, header_row_idx, len(headers))
            column_candidates = _detect_all_column_candidates(
                ws, headers, header_row_idx
            )
            sheets.append(
                {
                    "name": sheet_name,
                    "header_row": header_row_idx,
                    "headers": headers,
                    "row_count": row_count,
                    "rows_preview": rows_preview,
                    "column_candidates": column_candidates,
                }
            )

        active_sheet = next(
            (s["name"] for s in sheets if s["headers"] and s["row_count"] > 0),
            sheets[0]["name"] if sheets else "",
        )

        run.parsed_data = {"sheets": sheets, "active_sheet": active_sheet}
        run.status = QuestionnaireRun.Status.PARSED
        run.save(update_fields=["parsed_data", "status"])
        logger.info("Parsed questionnaire %s: %d sheet(s)", run.filename, len(sheets))

    except Exception as e:
        logger.error("Failed to parse questionnaire %s: %s", run_id, e)
        run.status = QuestionnaireRun.Status.FAILED
        run.error_message = str(e)
        run.save(update_fields=["status", "error_message"])


def _detect_headers(ws):
    """First row with at least 2 non-empty cells is treated as the header row.

    Scans up to the first 20 rows; returns (row_index_0_based, headers_list).
    Empty sheet -> (0, []).
    """
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True)):
        non_empty = [c for c in row if c not in (None, "")]
        if len(non_empty) >= 2:
            headers = [
                str(c).strip() if c not in (None, "") else f"(col {i + 1})"
                for i, c in enumerate(row)
            ]
            while headers and headers[-1].startswith("(col "):
                headers.pop()
            return idx, headers
    return 0, []


def _collect_rows(ws, header_row_idx: int, header_count: int, preview_size: int = 20):
    """Return (total_data_row_count, first_N_rows_as_list_of_lists)."""
    if header_count == 0:
        return 0, []
    preview = []
    total = 0
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx <= header_row_idx:
            continue
        cells = list(row[:header_count])
        if all(c in (None, "") for c in cells):
            continue
        total += 1
        if len(preview) < preview_size:
            preview.append(["" if c is None else str(c) for c in cells])
    return total, preview


def _detect_all_column_candidates(ws, headers: list, header_row_idx: int) -> dict:
    """Per column, find the controlled vocabulary the customer expects.

    Returns {col_idx: {"values": [...], "source": "data_validation"|"distinct_values"}}.
    Columns with no controlled vocabulary (free text, single value, too many
    distinct values) are simply absent from the dict — that's the signal to
    fall back to internal labels later.
    """
    out: dict = {}
    if not headers:
        return out
    for col_idx in range(len(headers)):
        candidates = _detect_column_candidates(ws, col_idx, header_row_idx)
        if candidates:
            out[col_idx] = candidates
    return out


def _detect_column_candidates(
    ws, col_idx: int, header_row_idx: int, max_distinct: int = 20
) -> dict | None:
    """Return {"values": [...], "source": "..."} for a single column, or None."""
    import openpyxl.utils

    target_col_letter = openpyxl.utils.get_column_letter(col_idx + 1)

    # 1. Excel data validation list constraint — the customer's authored vocab.
    try:
        for dv in ws.data_validations.dataValidation:
            if getattr(dv, "type", None) != "list":
                continue
            # Probe two cells inside this column to see if the validation applies
            probe_cells = [
                f"{target_col_letter}{header_row_idx + 2}",
                f"{target_col_letter}{header_row_idx + 3}",
            ]
            sqref = getattr(dv, "sqref", None)
            try:
                covers = any(c in sqref for c in probe_cells) if sqref else False
            except Exception:
                covers = False
            if not covers:
                continue
            values = _parse_data_validation_list(ws, dv.formula1)
            if values:
                return {"values": values, "source": "data_validation"}
    except Exception as e:
        logger.debug("data_validation_check_failed col=%s: %s", col_idx, e)

    # 2. Distinct existing values in the column.
    distinct: list[str] = []
    seen: set[str] = set()
    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx <= header_row_idx:
            continue
        if col_idx >= len(row):
            continue
        cell = row[col_idx]
        if cell in (None, ""):
            continue
        s = str(cell).strip()
        if not s:
            continue
        if s in seen:
            continue
        seen.add(s)
        distinct.append(s)
        if len(distinct) > max_distinct:
            return None  # likely free text — not a controlled vocab
    if 2 <= len(distinct) <= max_distinct:
        return {"values": distinct, "source": "distinct_values"}
    return None


def _parse_data_validation_list(ws, formula1) -> list:
    """Best-effort parsing of the formula1 of a list-type data validation.

    Inline:    '"Yes,No,Partial"' → ['Yes', 'No', 'Partial']
    Range ref: 'Sheet1!$A$1:$A$5' → values from that range
    """
    if formula1 is None:
        return []
    s = str(formula1).strip()
    if not s:
        return []
    # Inline quoted list
    if s.startswith('"') and s.endswith('"'):
        inner = s[1:-1]
        return [v.strip() for v in inner.split(",") if v.strip()]
    # Range reference, optionally cross-sheet
    try:
        if "!" in s:
            sheet_name, range_str = s.split("!", 1)
            sheet_name = sheet_name.strip().strip("'")
            target_ws = ws.parent[sheet_name]
        else:
            target_ws = ws
            range_str = s
        range_str = range_str.replace("$", "")
        cells = target_ws[range_str]
        values: list[str] = []
        # cells can be a single cell, a tuple (1D range), or tuple-of-tuples (2D)
        if not isinstance(cells, tuple):
            cells = (cells,)
        for row in cells:
            if isinstance(row, tuple):
                for c in row:
                    v = getattr(c, "value", None)
                    if v not in (None, ""):
                        values.append(str(v).strip())
            else:
                v = getattr(row, "value", None)
                if v not in (None, ""):
                    values.append(str(v).strip())
        return values
    except Exception as e:
        logger.debug("data_validation_range_parse_failed formula=%s: %s", s, e)
        return []


ANSWER_PROMPT = """You are answering a customer security questionnaire on \
behalf of an organization. Use ONLY the provided context to write your answer. \
If the context does not support an answer, set status to "needs_info".

Question:
{question}

{section_hint}Context (numbered passages):
{context}

Reply with a single JSON object, no prose around it:
{{
  "status": "yes" | "no" | "partial" | "needs_info",
  "comment": "<one or two sentences answering the question, with citation \
markers like [1] [2] referring to the numbered context passages>",
  "citation_indices": [<1-based indices of passages used>]
}}"""

CRITIC_PROMPT = """You are reviewing a draft answer to a customer security \
questionnaire. Score how well the cited context supports the answer (0.0 = \
contradicted or unsupported, 1.0 = fully supported and on-topic).

Question:
{question}

Draft answer (status={status}):
{comment}

Cited context:
{cited_context}

Reply with a single JSON object, no prose around it:
{{
  "score": <float between 0.0 and 1.0>,
  "issue": "<one short sentence describing the main weakness if score < 0.8, otherwise empty>"
}}"""

# Score below this triggers a retry in Thorough mode.
RETRY_THRESHOLD = 0.7
# Critic-less floor: assume mid confidence so review UI orders sensibly.
FAST_MODE_DEFAULT_CONFIDENCE = 0.5
# Per-question wall-clock cap so the loop can't hang on one item.
PER_QUESTION_TIMEOUT_SEC = 90


def _build_context_block(results: list[dict]) -> tuple[str, list[dict]]:
    """Format retrieval results as numbered passages + return source_refs.

    `chat.rag.search` returns flat dicts with fields at the top level
    (id, score, text, source_type, object_type, name, ref_id, …).
    Don't try to read them from a nested "payload" key — there isn't one.
    """
    lines = []
    refs = []
    for i, r in enumerate(results, start=1):
        score = r.get("score")
        text = r.get("text", "") or ""
        snippet = text.strip().replace("\n", " ")
        if len(snippet) > 400:
            snippet = snippet[:400] + "…"
        name = r.get("name") or ""
        object_type = r.get("object_type") or ""
        ref_id = r.get("ref_id") or ""
        label_bits = []
        if object_type:
            label_bits.append(object_type)
        if ref_id:
            label_bits.append(ref_id)
        elif name:
            label_bits.append(name[:60])
        label = " · ".join(label_bits) or "passage"
        lines.append(f"[{i}] ({label}) {snippet}")
        refs.append(
            {
                "index": i,
                "kind": r.get("source_type") or object_type or "",
                "id": str(r.get("object_id") or r.get("id") or ""),
                "name": name or object_type or "",
                "ref_id": ref_id,
                "score": float(score) if score is not None else None,
                "snippet": snippet,
            }
        )
    return "\n".join(lines), refs


def _parse_json_response(text: str) -> dict | None:
    """Pull a JSON object out of an LLM response. Tolerant of code fences and prose."""
    import json
    import re

    if not text:
        return None
    # Strip code fences
    fence = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", text, re.DOTALL)
    if fence:
        try:
            return json.loads(fence.group(1))
        except Exception:
            pass
    # First top-level {...} block
    match = re.search(r"\{.*\}", text, re.DOTALL)
    if match:
        try:
            return json.loads(match.group(0))
        except Exception:
            return None
    return None


def _heartbeat(run, label: str | None = None) -> bool:
    """Update heartbeat. Returns False if the run was cancelled — caller should bail."""
    from .models import AgentRun

    fresh = AgentRun.objects.only("status").get(id=run.id)
    if fresh.status == AgentRun.Status.CANCELLED:
        return False
    update_fields = ["last_heartbeat_at"]
    run.last_heartbeat_at = timezone.now()
    if label is not None:
        run.current_step_label = label[:300]
        update_fields.append("current_step_label")
    run.save(update_fields=update_fields + ["updated_at"])
    return True


@db_task()
def run_questionnaire_prefill(agent_run_id: str):
    """The agentic loop. Per question:
        retrieve → answer → (critic + retry on Thorough) → record AgentAction.
    Heartbeats between each question; checks cancellation each iteration.
    """
    import time

    from django.contrib.contenttypes.models import ContentType

    from .models import AgentRun, AgentAction, QuestionnaireQuestion
    from .providers import get_llm, get_chat_settings
    from .rag import search as rag_search
    from .tokens import count_tokens

    try:
        run = AgentRun.objects.select_related("owner", "folder").get(id=agent_run_id)
    except AgentRun.DoesNotExist:
        logger.error("AgentRun %s not found", agent_run_id)
        return

    if run.status != AgentRun.Status.QUEUED:
        logger.warning("AgentRun %s is %s; skipping", agent_run_id, run.status)
        return

    settings_dict = get_chat_settings()
    provider = settings_dict.get("llm_provider", "ollama")
    if provider == "openai_compatible":
        model_name = settings_dict.get("openai_model") or "openai_compatible"
    else:
        model_name = settings_dict.get("ollama_model") or "ollama"

    run.status = AgentRun.Status.RUNNING
    run.started_at = timezone.now()
    run.last_heartbeat_at = timezone.now()
    run.model_used = model_name
    run.error_message = ""
    run.save(
        update_fields=[
            "status",
            "started_at",
            "last_heartbeat_at",
            "model_used",
            "error_message",
            "updated_at",
        ]
    )

    questions = list(
        QuestionnaireQuestion.objects.filter(
            questionnaire_run_id=run.target_object_id
        ).order_by("ord")
    )
    run.total_steps = len(questions)
    run.completed_steps = 0
    run.save(update_fields=["total_steps", "completed_steps", "updated_at"])

    qq_ct = ContentType.objects.get_for_model(QuestionnaireQuestion)
    is_thorough = run.strictness == AgentRun.Strictness.THOROUGH
    max_retries = 1 if is_thorough else 0

    try:
        llm = get_llm()
    except Exception as e:
        run.status = AgentRun.Status.FAILED
        run.error_message = f"LLM unavailable: {e}"
        run.finished_at = timezone.now()
        run.save(update_fields=["status", "error_message", "finished_at", "updated_at"])
        return

    try:
        for question in questions:
            label = f"{run.completed_steps + 1}/{run.total_steps}: {question.text[:80]}"
            if not _heartbeat(run, label):
                # Cancelled mid-flight
                return

            t0 = time.time()
            try:
                _process_question(
                    run=run,
                    question=question,
                    qq_ct=qq_ct,
                    llm=llm,
                    max_retries=max_retries,
                    rag_search=rag_search,
                    count_tokens=count_tokens,
                )
            except Exception as e:
                logger.error(
                    "Failed processing question %s in run %s: %s",
                    question.id,
                    run.id,
                    e,
                )
                # Record the failure as an AgentAction so it shows in the UI;
                # don't kill the whole run for one bad question.
                AgentAction.objects.create(
                    agent_run=run,
                    kind=AgentAction.Kind.PROPOSE_ANSWER,
                    target_content_type=qq_ct,
                    target_object_id=question.id,
                    payload={"status": "needs_info", "comment": ""},
                    rationale=f"Error: {e}",
                    source_refs=[],
                    confidence=0.0,
                    state=AgentAction.State.PROPOSED,
                    iteration=0,
                    duration_ms=int((time.time() - t0) * 1000),
                )

            # Soft per-question budget guard
            if time.time() - t0 > PER_QUESTION_TIMEOUT_SEC:
                logger.warning(
                    "Question %s exceeded soft budget (%ss)",
                    question.id,
                    PER_QUESTION_TIMEOUT_SEC,
                )

            run.completed_steps += 1
            run.last_heartbeat_at = timezone.now()
            run.save(
                update_fields=["completed_steps", "last_heartbeat_at", "updated_at"]
            )

        run.status = AgentRun.Status.SUCCEEDED
        run.finished_at = timezone.now()
        run.current_step_label = ""
        run.save(
            update_fields=[
                "status",
                "finished_at",
                "current_step_label",
                "updated_at",
            ]
        )
        logger.info(
            "AgentRun %s completed: %d/%d", run.id, run.completed_steps, run.total_steps
        )

    except Exception as e:
        logger.error("AgentRun %s failed: %s", run.id, e)
        run.status = AgentRun.Status.FAILED
        run.error_message = str(e)
        run.finished_at = timezone.now()
        run.save(update_fields=["status", "error_message", "finished_at", "updated_at"])


def _process_question(
    *,
    run,
    question,
    qq_ct,
    llm,
    max_retries: int,
    rag_search,
    count_tokens,
):
    """Single-question pipeline: retrieve → answer → optional critic + retry."""
    import time

    from .models import AgentAction

    # --- Retrieve ---
    t0 = time.time()
    query = question.text
    try:
        results = rag_search(query, run.owner, top_k=6)
    except Exception as e:
        logger.error("RAG search failed for q %s: %s", question.id, e)
        results = []
    duration_ms = int((time.time() - t0) * 1000)

    context_block, source_refs = _build_context_block(results)

    AgentAction.objects.create(
        agent_run=run,
        kind=AgentAction.Kind.RETRIEVE,
        target_content_type=qq_ct,
        target_object_id=question.id,
        payload={"top_k": len(results)},
        rationale=f"RAG search returned {len(results)} passages.",
        source_refs=source_refs,
        state=AgentAction.State.PROPOSED,
        duration_ms=duration_ms,
    )

    # --- Answer (iteration 0) ---
    proposed = _answer_iteration(
        run=run,
        question=question,
        qq_ct=qq_ct,
        llm=llm,
        context_block=context_block,
        source_refs=source_refs,
        critic_hint="",
        iteration=0,
        count_tokens=count_tokens,
    )

    confidence = proposed["confidence"]

    # --- Critic + retry (Thorough only) ---
    if max_retries > 0:
        critic = _critique(
            run=run,
            question=question,
            qq_ct=qq_ct,
            llm=llm,
            answer_payload=proposed["payload"],
            context_block=context_block,
            cited_indices=proposed["cited_indices"],
            iteration=0,
            count_tokens=count_tokens,
        )
        confidence = critic["score"]

        if confidence < RETRY_THRESHOLD:
            retry = _answer_iteration(
                run=run,
                question=question,
                qq_ct=qq_ct,
                llm=llm,
                context_block=context_block,
                source_refs=source_refs,
                critic_hint=critic["issue"],
                iteration=1,
                count_tokens=count_tokens,
            )
            critic2 = _critique(
                run=run,
                question=question,
                qq_ct=qq_ct,
                llm=llm,
                answer_payload=retry["payload"],
                context_block=context_block,
                cited_indices=retry["cited_indices"],
                iteration=1,
                count_tokens=count_tokens,
            )
            # Use whichever iteration scored higher
            if critic2["score"] >= confidence:
                proposed = retry
                confidence = critic2["score"]
                # Mark earlier proposal as expired so review UI shows the latest
                AgentAction.objects.filter(
                    agent_run=run,
                    kind=AgentAction.Kind.PROPOSE_ANSWER,
                    target_content_type=qq_ct,
                    target_object_id=question.id,
                    iteration=0,
                ).update(state=AgentAction.State.EXPIRED)

    # Update final proposal with critic-derived confidence
    AgentAction.objects.filter(id=proposed["action_id"]).update(confidence=confidence)


def _answer_iteration(
    *,
    run,
    question,
    qq_ct,
    llm,
    context_block: str,
    source_refs: list[dict],
    critic_hint: str,
    iteration: int,
    count_tokens,
):
    """One answer attempt. Records and returns the AgentAction id + parsed payload."""
    import time

    from .models import AgentAction

    section_hint = (
        f"Question section: {question.section}\n\n" if question.section else ""
    )
    prompt = ANSWER_PROMPT.format(
        question=question.text,
        section_hint=section_hint,
        context=context_block or "(no relevant context found)",
    )
    if critic_hint:
        prompt += f"\n\nReviewer feedback to address: {critic_hint}"

    t0 = time.time()
    try:
        raw = llm.generate(prompt=prompt, context="", history=[])
    except Exception as e:
        logger.error("LLM generate failed for q %s: %s", question.id, e)
        raw = ""
    duration_ms = int((time.time() - t0) * 1000)

    parsed = _parse_json_response(raw) or {}
    answer_status = (
        parsed.get("status")
        if parsed.get("status")
        in {
            "yes",
            "no",
            "partial",
            "needs_info",
        }
        else "needs_info"
    )
    comment = (parsed.get("comment") or "").strip()
    cited_indices = parsed.get("citation_indices") or []
    if not isinstance(cited_indices, list):
        cited_indices = []

    used_refs = [r for r in source_refs if r["index"] in set(cited_indices)]

    tokens = 0
    try:
        tokens = count_tokens(prompt) + count_tokens(raw)
    except Exception:
        pass

    action = AgentAction.objects.create(
        agent_run=run,
        kind=AgentAction.Kind.PROPOSE_ANSWER,
        target_content_type=qq_ct,
        target_object_id=question.id,
        payload={"status": answer_status, "comment": comment},
        rationale=raw[:2000],
        source_refs=used_refs,
        confidence=FAST_MODE_DEFAULT_CONFIDENCE,
        state=AgentAction.State.PROPOSED,
        iteration=iteration,
        tokens=tokens,
        duration_ms=duration_ms,
    )

    run.total_tokens = (run.total_tokens or 0) + tokens
    run.save(update_fields=["total_tokens", "updated_at"])

    return {
        "action_id": action.id,
        "payload": {"status": answer_status, "comment": comment},
        "cited_indices": cited_indices,
        "confidence": FAST_MODE_DEFAULT_CONFIDENCE,
    }


def _critique(
    *,
    run,
    question,
    qq_ct,
    llm,
    answer_payload: dict,
    context_block: str,
    cited_indices: list[int],
    iteration: int,
    count_tokens,
):
    """Critic LLM call. Returns dict with score (float) and issue (string)."""
    import time

    from .models import AgentAction

    cited_lines = []
    for line in context_block.split("\n"):
        for idx in cited_indices:
            if line.startswith(f"[{idx}]"):
                cited_lines.append(line)
                break
    cited_text = "\n".join(cited_lines) if cited_lines else "(answer cited no passages)"

    prompt = CRITIC_PROMPT.format(
        question=question.text,
        status=answer_payload.get("status", ""),
        comment=answer_payload.get("comment", ""),
        cited_context=cited_text,
    )

    t0 = time.time()
    try:
        raw = llm.generate(prompt=prompt, context="", history=[])
    except Exception as e:
        logger.error("Critic LLM failed for q %s: %s", question.id, e)
        raw = ""
    duration_ms = int((time.time() - t0) * 1000)

    parsed = _parse_json_response(raw) or {}
    try:
        score = float(parsed.get("score", 0.0))
    except (TypeError, ValueError):
        score = 0.0
    score = max(0.0, min(1.0, score))
    issue = (parsed.get("issue") or "").strip()

    tokens = 0
    try:
        tokens = count_tokens(prompt) + count_tokens(raw)
    except Exception:
        pass

    AgentAction.objects.create(
        agent_run=run,
        kind=AgentAction.Kind.CRITIQUE,
        target_content_type=qq_ct,
        target_object_id=question.id,
        payload={"score": score, "issue": issue},
        rationale=raw[:2000],
        source_refs=[],
        confidence=score,
        state=AgentAction.State.PROPOSED,
        iteration=iteration,
        tokens=tokens,
        duration_ms=duration_ms,
    )

    run.total_tokens = (run.total_tokens or 0) + tokens
    run.save(update_fields=["total_tokens", "updated_at"])

    return {"score": score, "issue": issue}


@db_task()
def ingest_document(document_id: str):
    """
    Async task: extract text from a document, chunk it, embed it, store in Qdrant.
    """
    from .models import IndexedDocument
    from .extractors import get_extractor
    from .providers import get_embedder
    from .rag import COLLECTION_NAME, get_qdrant_client

    try:
        doc = IndexedDocument.objects.get(id=document_id)
    except IndexedDocument.DoesNotExist:
        logger.error("IndexedDocument %s not found", document_id)
        return

    doc.status = IndexedDocument.Status.PROCESSING
    doc.save(update_fields=["status"])

    try:
        # Get extractor for file type
        extractor = get_extractor(doc.content_type)
        if not extractor:
            raise ValueError(f"No extractor for content type: {doc.content_type}")

        # Extract chunks
        chunks = extractor(doc.file)
        if not chunks:
            raise ValueError("No content extracted from document")

        # Embed chunks
        embedder = get_embedder()
        texts = [chunk.text for chunk in chunks]
        embeddings = embedder.embed(texts)

        # Store in Qdrant
        from qdrant_client.models import PointStruct

        client = get_qdrant_client()
        points = [
            PointStruct(
                id=str(uuid.uuid5(uuid.NAMESPACE_URL, f"{doc.id}:{chunk.index}")),
                vector=embedding,
                payload={
                    "text": chunk.text,
                    "folder_id": str(doc.folder_id),
                    "source_type": "document",
                    "object_type": "document_chunk",
                    "object_id": str(doc.id),
                    "document_id": str(doc.id),
                    "chunk_index": chunk.index,
                    "filename": doc.filename,
                    **chunk.metadata,
                },
            )
            for chunk, embedding in zip(chunks, embeddings)
        ]

        client.upsert(collection_name=COLLECTION_NAME, points=points)

        doc.status = IndexedDocument.Status.INDEXED
        doc.chunk_count = len(chunks)
        doc.indexed_at = timezone.now()
        doc.save(update_fields=["status", "chunk_count", "indexed_at"])

        logger.info("Indexed document %s: %d chunks", doc.filename, len(chunks))

    except Exception as e:
        logger.error("Failed to ingest document %s: %s", document_id, e)
        doc.status = IndexedDocument.Status.FAILED
        doc.error_message = str(e)
        doc.save(update_fields=["status", "error_message"])


@db_task()
def index_model_object(app_label: str, model_name: str, object_id: str):
    """
    Async task: index or re-index a single Django model object into Qdrant.
    """
    from django.apps import apps

    from .providers import get_embedder
    from .rag import COLLECTION_NAME, get_qdrant_client

    try:
        model_class = apps.get_model(app_label, model_name)
        obj = model_class.objects.get(id=object_id)
    except Exception as e:
        logger.error("Cannot load %s.%s/%s: %s", app_label, model_name, object_id, e)
        return

    # Build text representation
    text = _build_object_text(obj, model_name)
    if not text:
        return

    folder_id = _resolve_folder_id(obj)
    if not folder_id:
        return

    try:
        embedder = get_embedder()
        vector = embedder.embed_query(text)

        from qdrant_client.models import PointStruct

        client = get_qdrant_client()
        point = PointStruct(
            id=str(
                uuid.uuid5(uuid.NAMESPACE_URL, f"{app_label}.{model_name}:{object_id}")
            ),
            vector=vector,
            payload={
                "text": text,
                "folder_id": folder_id,
                "source_type": "model",
                "object_type": _normalize_model_name(model_name),
                "object_id": object_id,
                "name": str(obj),
                "ref_id": getattr(obj, "ref_id", "") or "",
            },
        )

        client.upsert(collection_name=COLLECTION_NAME, points=[point])
        logger.debug("Indexed %s.%s/%s", app_label, model_name, object_id)

    except Exception as e:
        logger.error(
            "Failed to index %s.%s/%s: %s", app_label, model_name, object_id, e
        )


@db_task()
def remove_model_object(app_label: str, model_name: str, object_id: str):
    """Remove a deleted object from the vector store."""
    from .rag import COLLECTION_NAME, get_qdrant_client
    from qdrant_client.models import Filter, FieldCondition, MatchValue

    try:
        client = get_qdrant_client()
        client.delete(
            collection_name=COLLECTION_NAME,
            points_selector=Filter(
                must=[
                    FieldCondition(key="object_id", match=MatchValue(value=object_id)),
                    FieldCondition(key="source_type", match=MatchValue(value="model")),
                ]
            ),
        )
        logger.debug("Removed %s.%s/%s from index", app_label, model_name, object_id)
    except Exception as e:
        logger.error(
            "Failed to remove %s.%s/%s: %s", app_label, model_name, object_id, e
        )


def _resolve_folder_id(obj) -> str:
    """Walk the FK chain to find a folder_id for a model object.

    Most indexed models have FolderMixin (direct folder_id). RequirementAssessment
    has it too via FolderMixin, but we keep this helper future-proof for parent-FK
    resolutions (e.g. children of RequirementAssessment if we index those later).
    """
    direct = getattr(obj, "folder_id", None)
    if direct:
        return str(direct)
    for parent_attr in (
        "compliance_assessment",
        "risk_assessment",
        "questionnaire_run",
    ):
        parent = getattr(obj, parent_attr, None)
        if parent and getattr(parent, "folder_id", None):
            return str(parent.folder_id)
    return ""


def _build_object_text(obj, model_name: str) -> str:
    """Build a searchable text representation of a model object.

    For the questionnaire-prefill use case, RequirementAssessment and
    AppliedControl get a deliberately narrow shape — only the fields that
    bear on the verdict. Other indexed models keep the broader, generic
    representation so existing chat features aren't disturbed.
    """
    if model_name == "RequirementAssessment":
        return _text_for_requirement_assessment(obj)
    if model_name == "AppliedControl":
        return _text_for_applied_control(obj)
    return _text_for_generic_model(obj, model_name)


def _text_for_requirement_assessment(obj) -> str:
    """Identity (so retrieval can match by question wording) + verdict + narrative.

    Per first-user feedback, the verdict signal lives in `result` and the
    narrative in `observation`; everything else (descriptions, categories,
    linked controls, etc.) is noise for the questionnaire-answering task.
    """
    parts = ["Type: Requirement assessment"]

    requirement = getattr(obj, "requirement", None)
    if requirement is not None:
        req_ref = (getattr(requirement, "ref_id", "") or "").strip()
        req_name = (getattr(requirement, "name", "") or "").strip()
        req_desc = (getattr(requirement, "description", "") or "").strip()
        if req_ref:
            parts.append(f"Requirement ref: {req_ref}")
        if req_name and req_name != req_ref:
            parts.append(f"Requirement: {req_name}")
        if req_desc:
            parts.append(f"Requirement text: {req_desc}")
        framework = getattr(requirement, "framework", None)
        framework_name = getattr(framework, "name", "") if framework else ""
        if framework_name:
            parts.append(f"Framework: {framework_name}")

    result = getattr(obj, "result", None)
    if result:
        result_display = (
            obj.get_result_display() if hasattr(obj, "get_result_display") else result
        )
        parts.append(f"Result: {result_display}")

    observation = (getattr(obj, "observation", None) or "").strip()
    if observation:
        parts.append(f"Observation: {observation}")

    return "\n".join(parts)


def _text_for_applied_control(obj) -> str:
    """Identity + verdict (status) + narrative (observation).

    Same rationale as RequirementAssessment — keep retrieval focused on
    fields that decide whether the control answers the question.
    """
    parts = ["Type: Applied control"]

    name = (getattr(obj, "name", "") or "").strip()
    if name:
        parts.append(f"Name: {name}")
    ref_id = (getattr(obj, "ref_id", "") or "").strip()
    if ref_id:
        parts.append(f"Reference: {ref_id}")

    status = getattr(obj, "status", None)
    if status:
        status_display = (
            obj.get_status_display() if hasattr(obj, "get_status_display") else status
        )
        parts.append(f"Status: {status_display}")

    observation = (getattr(obj, "observation", None) or "").strip()
    if observation:
        parts.append(f"Observation: {observation}")

    return "\n".join(parts)


def _text_for_generic_model(obj, model_name: str) -> str:
    """Broader representation for other indexed models (RiskScenario, Asset,
    Threat, ComplianceAssessment, RiskAssessment).
    """
    parts = [f"Type: {model_name.replace('_', ' ').title()}"]

    name = getattr(obj, "name", None)
    if name:
        parts.append(f"Name: {name}")

    ref_id = getattr(obj, "ref_id", None)
    if ref_id:
        parts.append(f"Reference: {ref_id}")

    description = getattr(obj, "description", None)
    if description:
        parts.append(f"Description: {description}")

    observation = getattr(obj, "observation", None)
    if observation:
        parts.append(f"Observation: {observation}")

    if hasattr(obj, "current_level"):
        parts.append(
            f"Current risk level: {obj.get_current_level_display() if hasattr(obj, 'get_current_level_display') else obj.current_level}"
        )

    if hasattr(obj, "treatment"):
        parts.append(
            f"Treatment: {obj.get_treatment_display() if hasattr(obj, 'get_treatment_display') else obj.treatment}"
        )

    if hasattr(obj, "status"):
        status_display = (
            obj.get_status_display()
            if hasattr(obj, "get_status_display")
            else obj.status
        )
        parts.append(f"Status: {status_display}")

    if hasattr(obj, "category"):
        cat_display = (
            obj.get_category_display()
            if hasattr(obj, "get_category_display")
            else obj.category
        )
        if cat_display:
            parts.append(f"Category: {cat_display}")

    if hasattr(obj, "business_value"):
        bv = (
            obj.get_business_value_display()
            if hasattr(obj, "get_business_value_display")
            else obj.business_value
        )
        if bv:
            parts.append(f"Business value: {bv}")

    return "\n".join(parts)


def _normalize_model_name(model_name: str) -> str:
    """Convert model class name to snake_case identifier."""
    import re

    s = re.sub(r"(?<=[a-z])(?=[A-Z])", "_", model_name)
    return s.lower()


# ---------------------------------------------------------------------------
# Library knowledge base indexing
# ---------------------------------------------------------------------------

LIBRARY_DIR = None  # Resolved lazily


def _get_library_dir():
    """Get the path to the YAML library directory."""
    global LIBRARY_DIR
    if LIBRARY_DIR is None:
        from pathlib import Path

        LIBRARY_DIR = Path(__file__).resolve().parent.parent / "library" / "libraries"
    return LIBRARY_DIR


def _parse_library_yaml(filepath) -> list[dict]:
    """
    Parse a YAML library file and extract indexable entries.
    Returns a list of dicts with: urn, ref_id, name, description, etc.
    """
    import yaml

    try:
        with open(filepath, "r", encoding="utf-8") as f:
            data = yaml.safe_load(f)
    except Exception as e:
        logger.warning("failed_to_parse_yaml", file=str(filepath), error=e)
        return []

    if not isinstance(data, dict):
        return []

    library_name = data.get("name", "")
    library_ref_id = data.get("ref_id", "")
    locale = data.get("locale", "en")
    provider = data.get("provider", "")

    entries = []
    objects = data.get("objects", {})

    # Index the library itself as a summary entry
    if library_name:
        lib_description = data.get("description", "")
        entries.append(
            {
                "urn": data.get("urn", ""),
                "ref_id": library_ref_id,
                "name": library_name,
                "description": lib_description,
                "annotation": "",
                "framework": library_name,
                "framework_ref_id": library_ref_id,
                "provider": provider,
                "locale": locale,
                "object_type": "framework",
            }
        )

    # Extract framework requirement nodes
    framework = objects.get("framework", {})
    if isinstance(framework, dict):
        framework_name = framework.get("name", library_name)
        framework_ref_id = framework.get("ref_id", library_ref_id)
        for node in framework.get("requirement_nodes", []):
            # Skip empty section headers
            if not node.get("name") and not node.get("description"):
                continue
            entries.append(
                {
                    "urn": node.get("urn", ""),
                    "ref_id": node.get("ref_id", ""),
                    "name": node.get("name", ""),
                    "description": node.get("description", ""),
                    "annotation": node.get("annotation", ""),
                    "framework": framework_name,
                    "framework_ref_id": framework_ref_id,
                    "provider": provider,
                    "locale": locale,
                    "object_type": "requirement_node",
                }
            )

    # Extract threats
    for threat in objects.get("threats", []):
        entries.append(
            {
                "urn": threat.get("urn", ""),
                "ref_id": threat.get("ref_id", ""),
                "name": threat.get("name", ""),
                "description": threat.get("description", ""),
                "annotation": "",
                "framework": library_name,
                "framework_ref_id": library_ref_id,
                "provider": provider,
                "locale": locale,
                "object_type": "library_threat",
            }
        )

    # Extract reference controls
    for ctrl in objects.get("reference_controls", []):
        entries.append(
            {
                "urn": ctrl.get("urn", ""),
                "ref_id": ctrl.get("ref_id", ""),
                "name": ctrl.get("name", ""),
                "description": ctrl.get("description", ""),
                "annotation": ctrl.get("annotation", ""),
                "framework": library_name,
                "framework_ref_id": library_ref_id,
                "provider": provider,
                "locale": locale,
                "object_type": "reference_control",
            }
        )

    return entries


def _build_library_entry_text(entry: dict) -> str:
    """Build searchable text for a library entry."""
    parts = [f"Framework: {entry['framework']}"]
    if entry.get("ref_id"):
        parts.append(f"Reference: {entry['ref_id']}")
    if entry.get("name"):
        parts.append(f"Name: {entry['name']}")
    if entry.get("description"):
        parts.append(f"Description: {entry['description']}")
    if entry.get("annotation"):
        parts.append(f"Guidance: {entry['annotation']}")
    return "\n".join(parts)


@db_task()
def index_library_knowledge_base():
    """
    Async task: parse all YAML library files and index requirement nodes,
    threats, and reference controls into Qdrant as shared knowledge.

    Uses source_type="library" — accessible to all authenticated users
    without folder-based permission filtering.
    """
    import time

    from .providers import get_embedder
    from .rag import COLLECTION_NAME, get_qdrant_client

    t0 = time.time()
    library_dir = _get_library_dir()
    if not library_dir.exists():
        logger.warning("library_dir_not_found", path=str(library_dir))
        return

    yaml_files = sorted(library_dir.glob("*.yaml"))
    logger.info("library_indexing_started", file_count=len(yaml_files))

    # Collect all entries from all YAML files
    all_entries = []
    for filepath in yaml_files:
        entries = _parse_library_yaml(filepath)
        all_entries.extend(entries)

    if not all_entries:
        logger.info("library_indexing_no_entries")
        return

    logger.info(
        "library_entries_parsed",
        count=len(all_entries),
        duration=round(time.time() - t0, 2),
    )

    # Embed in batches
    embedder = get_embedder()
    client = get_qdrant_client()

    from qdrant_client.models import PointStruct

    BATCH_SIZE = 100
    total_indexed = 0

    for batch_start in range(0, len(all_entries), BATCH_SIZE):
        batch = all_entries[batch_start : batch_start + BATCH_SIZE]
        texts = [_build_library_entry_text(e) for e in batch]

        try:
            embeddings = embedder.embed(texts)
        except Exception as e:
            logger.error(
                "library_embedding_failed",
                batch_start=batch_start,
                error=e,
            )
            continue

        points = []
        for entry, embedding in zip(batch, embeddings):
            point_id = str(uuid.uuid5(uuid.NAMESPACE_URL, f"library:{entry['urn']}"))
            points.append(
                PointStruct(
                    id=point_id,
                    vector=embedding,
                    payload={
                        "text": _build_library_entry_text(entry),
                        "source_type": "library",
                        "object_type": entry["object_type"],
                        "urn": entry["urn"],
                        "ref_id": entry.get("ref_id", ""),
                        "name": entry.get("name", ""),
                        "framework": entry["framework"],
                        "framework_ref_id": entry.get("framework_ref_id", ""),
                        "provider": entry.get("provider", ""),
                        "locale": entry.get("locale", "en"),
                    },
                )
            )

        try:
            client.upsert(collection_name=COLLECTION_NAME, points=points)
            total_indexed += len(points)
        except Exception as e:
            logger.error(
                "library_upsert_failed",
                batch_start=batch_start,
                error=e,
            )

    logger.info(
        "library_indexing_complete",
        total_indexed=total_indexed,
        duration=round(time.time() - t0, 2),
    )


@db_task()
def update_session_summary(session_id: str):
    """Async wrapper around memory.update_summary_for_session."""
    from .memory import update_summary_for_session
    from .models import ChatSession
    from .providers import get_llm

    try:
        session = ChatSession.objects.get(pk=session_id)
    except ChatSession.DoesNotExist:
        return
    update_summary_for_session(session, get_llm())
