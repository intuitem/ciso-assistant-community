# simple script to transform an Excel file to a json library for threats

import openpyxl
import sys
import re
import json

if len(sys.argv) <= 1:
    print("missing input file parameter")
    exit()
input_file_name = sys.argv[1]
name = re.sub(r"\.\w+$", "", input_file_name)
output_file_name = name + ".json"

print("parsing", input_file_name)

output = {"name": name, "description": "threats from " + name, "format_version": "1.0", "objects": []}

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook(input_file_name)
 
# Define variable to read sheet
dataframe1 = dataframe.active
 
# Iterate the loop to read the cell values
for row in range(0, dataframe1.max_row):
    id = next(dataframe1.iter_cols(1, dataframe1.max_column))[row].value
    if id and re.match(r"M", id):
        description = next(dataframe1.iter_cols(2, dataframe1.max_column))[row].value
        output["objects"].append(
            {
                "type": "threat", 
                "fields": {
                    "name": f"{id}", 
                    "description": description
                }
            }
        )

print("writing", output_file_name)
out_file = open(output_file_name, "w")
json.dump(output, out_file, indent = 6)
out_file.close()
