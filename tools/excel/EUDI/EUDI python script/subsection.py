from openpyxl import load_workbook

input_file = "Excel_version1_generated.xlsx"
output_file = "Excel_version1_merged.xlsx"

# 1) Workbook avec les formules
wb_formula = load_workbook(input_file, data_only=False)

# 2) Workbook avec les valeurs calculées
wb_values = load_workbook(input_file, data_only=True)

ws_formula = wb_formula["Sheet1"]
ws_values = wb_values["Sheet1"]

# Parcourir toutes les lignes
for row in range(1, ws_formula.max_row + 1):
    cell_f_formula = ws_formula[f"F{row}"].value
    cell_f_value = ws_values[f"F{row}"].value

    # Si F contient une formule ET qu'elle retourne une valeur,
    # on remplace E sur la même ligne
    if isinstance(cell_f_formula, str) and cell_f_formula.startswith("=") and cell_f_value not in (None, ""):
        ws_formula[f"E{row}"] = cell_f_value

# Optionnel : supprimer la colonne F
# ws_formula.delete_cols(6)

wb_formula.save(output_file)
print(f"Fichier créé : {output_file}")
