#!/usr/bin/env python3
"""
Converter for E-ITS Excel file to CISO Assistant format.

Output columns:
- assessable: boolean (True for measures, False for categories)
- depth: integer (hierarchy level)
- ref_id: identifier
- name: name/title
- description: content
- annotation: additional notes
- implementation_groups: P/S/K (first letter of measure level)
"""

import openpyxl
from openpyxl import Workbook


def extract_ref_from_module(module_text):
    """Extract ref_id from module text like 'ISMS.1: Turbehaldus' -> 'ISMS.1'"""
    if not module_text or module_text == "-":
        return None
    if ":" in module_text:
        return module_text.split(":")[0].strip()
    return module_text.strip()


def extract_name_from_module(module_text):
    """Extract name from module text like 'ISMS.1: Turbehaldus' -> 'Turbehaldus'"""
    if not module_text or module_text == "-":
        return None
    if ":" in module_text:
        return module_text.split(":", 1)[1].strip()
    return module_text.strip()


def get_implementation_group(measure_level):
    """Get first letter of measure level: Põhimeede->P, Standardmeede->S, Kõrgmeede->K"""
    if not measure_level:
        return ""
    first_char = measure_level[0].upper()
    if first_char in ("P", "S", "K"):
        return first_char
    return ""


def is_empty_level(value):
    """Check if a level value is empty or just a dash"""
    return not value or value.strip() == "-" or value.strip() == ""


def convert_eits(input_file, output_file):
    """Convert E-ITS Excel to CISO Assistant format"""

    wb_in = openpyxl.load_workbook(input_file)
    ws_in = wb_in.active

    # Create output workbook
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "e-its"

    # Write headers
    headers = [
        "assessable",
        "depth",
        "ref_id",
        "name",
        "description",
        "annotation",
        "implementation_groups",
    ]
    ws_out.append(headers)

    # Track what we've already written to avoid duplicates
    written_refs = set()

    # Store rows to write (we need to process all data first to build hierarchy)
    output_rows = []

    # Column indices (0-based, from row 4 headers)
    # A=0: #, B=1: version, C=2: module_group, D=3: level1, E=4: level2, F=5: level3
    # G=6: owner, H=7: lifecycle, I=8: measure_level, J=9: protection_need
    # K=10: measure_id, L=11: measure_name, M=12: co_responsible, N=13: measure_content

    COL_MODULE_GROUP = 2  # C
    COL_LEVEL1 = 3  # D
    COL_LEVEL2 = 4  # E
    COL_LEVEL3 = 5  # F
    COL_MEASURE_LEVEL = 8  # I (Meetme tase)
    COL_MEASURE_ID = 10  # K (Meetme tunnus)
    COL_MEASURE_NAME = 11  # L (Meetme nimetus)
    COL_MEASURE_CONTENT = 13  # N (Meetme sisu)

    # Process data rows (starting from row 5, index 4)
    for row_idx, row in enumerate(
        ws_in.iter_rows(min_row=5, values_only=True), start=5
    ):
        module_group = row[COL_MODULE_GROUP]
        level1 = row[COL_LEVEL1]
        level2 = row[COL_LEVEL2]
        level3 = row[COL_LEVEL3]
        measure_level = row[COL_MEASURE_LEVEL]
        measure_id = row[COL_MEASURE_ID]
        measure_name = row[COL_MEASURE_NAME]
        measure_content = row[COL_MEASURE_CONTENT]

        if not module_group:
            continue

        current_depth = 1

        # 1. Module Group (depth 1)
        group_ref = extract_ref_from_module(module_group)
        group_name = extract_name_from_module(module_group)
        if group_ref and group_ref not in written_refs:
            output_rows.append(
                {
                    "assessable": "",  # categories are not assessable
                    "depth": 1,
                    "ref_id": group_ref,
                    "name": group_name,
                    "description": "",
                    "annotation": "",
                    "implementation_groups": "",
                }
            )
            written_refs.add(group_ref)
        current_depth = 1

        # 2. Level 1 (depth 2)
        if not is_empty_level(level1):
            level1_ref = extract_ref_from_module(level1)
            level1_name = extract_name_from_module(level1)
            if level1_ref and level1_ref not in written_refs:
                output_rows.append(
                    {
                        "assessable": "",  # categories are not assessable
                        "depth": 2,
                        "ref_id": level1_ref,
                        "name": level1_name,
                        "description": "",
                        "annotation": "",
                        "implementation_groups": "",
                    }
                )
                written_refs.add(level1_ref)
            current_depth = 2

        # 3. Level 2 (depth 3)
        if not is_empty_level(level2):
            level2_ref = extract_ref_from_module(level2)
            level2_name = extract_name_from_module(level2)
            if level2_ref and level2_ref not in written_refs:
                output_rows.append(
                    {
                        "assessable": "",  # categories are not assessable
                        "depth": 3,
                        "ref_id": level2_ref,
                        "name": level2_name,
                        "description": "",
                        "annotation": "",
                        "implementation_groups": "",
                    }
                )
                written_refs.add(level2_ref)
            current_depth = 3

        # 4. Level 3 (depth 4)
        if not is_empty_level(level3):
            level3_ref = extract_ref_from_module(level3)
            level3_name = extract_name_from_module(level3)
            if level3_ref and level3_ref not in written_refs:
                output_rows.append(
                    {
                        "assessable": "",  # categories are not assessable
                        "depth": 4,
                        "ref_id": level3_ref,
                        "name": level3_name,
                        "description": "",
                        "annotation": "",
                        "implementation_groups": "",
                    }
                )
                written_refs.add(level3_ref)
            current_depth = 4

        # 5. Measure (depth = current_depth + 1)
        if measure_id and measure_id not in written_refs:
            impl_group = get_implementation_group(measure_level)
            output_rows.append(
                {
                    "assessable": "x",  # measures are assessable
                    "depth": current_depth + 1,
                    "ref_id": measure_id,
                    "name": measure_name or "",
                    "description": measure_content or "",
                    "annotation": "",
                    "implementation_groups": impl_group,
                }
            )
            written_refs.add(measure_id)

    # Write rows in original order (no sorting)
    for row_data in output_rows:
        ws_out.append(
            [
                row_data["assessable"],
                row_data["depth"],
                row_data["ref_id"],
                row_data["name"],
                row_data["description"],
                row_data["annotation"],
                row_data["implementation_groups"],
            ]
        )

    # Save output
    wb_out.save(output_file)
    print(f"Converted {len(output_rows)} rows to {output_file}")

    # Print summary
    depth_counts = {}
    impl_group_counts = {}
    assessable_count = 0
    for row in output_rows:
        d = row["depth"]
        depth_counts[d] = depth_counts.get(d, 0) + 1
        if row["assessable"] == "x":
            assessable_count += 1
        if row["implementation_groups"]:
            ig = row["implementation_groups"]
            impl_group_counts[ig] = impl_group_counts.get(ig, 0) + 1

    print("\n=== SUMMARY ===")
    print(f"Total rows: {len(output_rows)}")
    print(f"Assessable (measures): {assessable_count}")
    print(f"Categories: {len(output_rows) - assessable_count}")
    print("\nBy depth:")
    for d in sorted(depth_counts.keys()):
        print(f"  Depth {d}: {depth_counts[d]} rows")
    print("\nBy implementation group (measures only):")
    for ig in sorted(impl_group_counts.keys()):
        label = {
            "P": "Põhimeede (Basic)",
            "S": "Standardmeede (Standard)",
            "K": "Kõrgmeede (High)",
        }.get(ig, ig)
        print(f"  {ig}: {impl_group_counts[ig]} - {label}")


if __name__ == "__main__":
    convert_eits("e-its-2024-source.xlsx", "e-its-2024-converted.xlsx")
