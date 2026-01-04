import re
from openpyxl import load_workbook, Workbook

# === Utility functions ===
def has_text(cell):
    return isinstance(cell, str) and cell.strip() != ""

def build_annotation(ref, fi):
    parts = []
    if has_text(ref):
        parts.append(f"Reference: {ref.strip()}")
    if has_text(fi):
        parts.append(fi.strip())
    return "\n\n".join(parts)

# === Load source Excel ===
source_file = "ENX_VCSA_1_1_EN.xlsx"  # Must exist in working directory
source_wb = load_workbook(filename=source_file, data_only=True)
source_ws = source_wb["Vehicle CyberSecurity"]

# === Header indexing ===
header_row_index = 3
headers = [cell.value for cell in source_ws[header_row_index]]
col_map = {header: idx for idx, header in enumerate(headers)}

# === Extract vcsa_content rows ===
vcsa_content_rows = []
seen_leaf_ids = {}

for row in source_ws.iter_rows(min_row=header_row_index + 1, values_only=True):
    vcs = str(row[col_map.get("VCS", 0)] or "").strip()
    control_question = str(row[col_map.get("Control question", 0)] or "").strip()
    objective = row[col_map.get("Objective", 0)] or ""

    if re.fullmatch(r"\d+", vcs):
        vcsa_content_rows.append(["", 1, vcs, control_question or f"Section {vcs}", "", "", "", ""])
    elif re.fullmatch(r"\d+\.\d+", vcs):
        vcsa_content_rows.append(["", 2, vcs, control_question, str(objective).strip(), "", "", ""])
    elif re.fullmatch(r"\d+\.\d+\.[A-Za-z]+", vcs):
        if vcs == "3.1.i":
            count = seen_leaf_ids.get(vcs, 0)
            if count == 1:
                vcs = "3.1.ii"
            seen_leaf_ids["3.1.i"] = count + 1

        x_y_z = vcs
        must_text = row[col_map.get("Requirements (must)", 0)] or ""
        must_ref = row[col_map.get("Reference (must)", 0)] or ""
        should_text = row[col_map.get("Requirements (should)", 0)] or ""
        should_ref = row[col_map.get("Reference (should)", 0)] or ""
        further_info = row[col_map.get("Further information", 0)] or ""
        possible_evidence = row[col_map.get("Possible evidence (not mandatory)", 0)] or ""

        if has_text(must_text):
            vcsa_content_rows.append([
                "x", 3, f"{x_y_z}-must", "", must_text.strip(),
                "must",
                build_annotation(must_ref, further_info),
                possible_evidence.strip() if has_text(possible_evidence) else ""
            ])
        if has_text(should_text):
            vcsa_content_rows.append([
                "x", 3, f"{x_y_z}-should", "", should_text.strip(),
                "should",
                build_annotation(should_ref, further_info),
                possible_evidence.strip() if has_text(possible_evidence) else ""
            ])

# === Create new workbook ===
wb = Workbook()
wb.remove(wb.active)

# === Meta sheets ===
meta_data = {
    "library_meta": [
        ["type", "library"],
        ["urn", "urn:intuitem:risk:library:vcsa-v1.1"],
        ["version", "1"],
        ["locale", "en"],
        ["ref_id", "vcsa-v1.1"],
        ["name", "Vehicle CyberSecurity Audit (VCSA) v1.1"],
        ["description", """The VCSA serves as the basis for 
- a self assessment to determine the state of vehicle cybersecurity within the organization (e.g. company)
- audits performed by internal departments (e.g. Internal Audit, Quality Management, Information Security, Cybersecurity)
- an audit in accordance with ENX 3rd party audit management framework
Source: https://enx.com/en-US/VCS/downloads/"""],
        ["copyright", """© 2023 ENX Association
Contact: vcs@enx.com +49 69 9866927-71
www.enx.com
This work has been licensed under the Creative Commons Attribution - NoDerivs 4.0 International Public License. In addition, You are granted the right to distribute derivatives under certain terms. The complete and valid text of the license is to be found in line 17ff."""],
        ["provider", "ENX"],
        ["packager", "intuitem"]
    ],
    "vcsa_meta": [
        ["type", "framework"],
        ["base_urn", "urn:intuitem:risk:req_node:vcsa-v1.1"],
        ["urn", "urn:intuitem:risk:framework:vcsa-v1.1"],
        ["ref_id", "vcsa-v1.1"],
        ["name", "Vehicle CyberSecurity Audit (VCSA) v1.1"],
        ["description", """The VCSA serves as the basis for 
- a self assessment to determine the state of vehicle cybersecurity within the organization (e.g. company)
- audits performed by internal departments (e.g. Internal Audit, Quality Management, Information Security, Cybersecurity)
- an audit in accordance with ENX 3rd party audit management framework
Source: https://enx.com/en-US/VCS/downloads/"""],
        ["implementation_groups_definition", "implementation_groups"]
    ],
    "implementation_groups_meta": [
        ["type", "implementation_groups"],
        ["name", "implementation_groups"]
    ]
}
for sheet_name, rows in meta_data.items():
    ws = wb.create_sheet(sheet_name)
    for row in rows:
        ws.append(row)

# === implementation_groups_content ===
ws_impl = wb.create_sheet("implementation_groups_content")
ws_impl.append(["ref_id", "name", "description"])
ws_impl.append([
    "must", "Requirements (must)",
    "The requirements indicated in this column are strict requirements without any exemptions. They are defined abstractly enough to encompass all VCS supplier types"
])
ws_impl.append([
    "should", "Requirements (should)",
    "The requirements indicated in this column are principally to be implemented by the organization. These requirements go into granular detail. However, for certain supplier types, there may be a valid justification for non-compliance with these requirements. In case of any deviation, its effects must be understood by the supplier organization and it must be plausibly justified"
])

# === vcsa_content ===
ws_vcsa = wb.create_sheet("vcsa_content")
ws_vcsa.append([
    "assessable", "depth", "ref_id", "name", "description",
    "implementation_groups", "annotation", "typical_evidence"
])
for row in vcsa_content_rows:
    ws_vcsa.append(row)

# === Save final Excel ===
wb.save("vcsa-v1.1.xlsx")
print("✅ Export complete: vcsa-v1.1.xlsx")
