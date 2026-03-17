#!/usr/bin/env python3
"""Build a PCI DSS library workbook from the PCI SSC prioritized approach workbook.

Scope intentionally limited to:
- source sheet: "Prioritized Approach Milestones"
- target sheet: "PCI DSS_content"
"""

import argparse
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook

SOURCE_SHEET = "Prioritized Approach Milestones"
TARGET_CONTENT_SHEET = "PCI DSS_content"
DEFAULT_SOURCE_FR = Path("i18n/Prioritized-Approach-Tool-For-PCI-DSS-v4_0_1-FR.xlsx")

PCI_DSS_VERSION_REF = "PCI DSS 4.0.1"
PCI_DSS_TITLE_EN = "Payment Card Industry Data Security Standard"
PCI_DSS_TITLE_FR = (
    "Payment Card Industry - Standard de Sécurité des Données (PCI-DSS) v4.0.1"
)
PCI_DSS_DESCRIPTION_URL = "https://www.pcisecuritystandards.org/"
PCI_DSS_DESCRIPTION_EN = (
    "Payment Card Industry Data Security Standard: Requirements and Testing Procedures, v4.0.1 (June 2024)\n"
    f"{PCI_DSS_DESCRIPTION_URL}"
)
PCI_DSS_DESCRIPTION_FR = (
    "Payment Card Industry - Standard de Sécurité des Données : Exigences et Procédures de Test, v4.0.1 (Juin 2024)\n"
    f"{PCI_DSS_DESCRIPTION_URL}"
)

LIBRARY_META_ROWS = [
    ("type", "library"),
    ("urn", "urn:intuitem:risk:library:pcidss-4_0"),
    ("version", "3"),
    ("locale", "en"),
    ("ref_id", PCI_DSS_VERSION_REF),
    ("name", PCI_DSS_TITLE_EN),
    ("description", PCI_DSS_DESCRIPTION_EN),
    (
        "copyright",
        "©2006 - 2024 PCI Security Standards Council, LLC. All Rights Reserved.",
    ),
    ("provider", "PCI Security Standards Council"),
    ("packager", "intuitem"),
    ("name[fr]", PCI_DSS_TITLE_FR),
    ("description[fr]", PCI_DSS_DESCRIPTION_FR),
]

FRAMEWORK_META_ROWS = [
    ("type", "framework"),
    ("base_urn", "urn:intuitem:risk:req_node:pcidss-4_0"),
    ("urn", "urn:intuitem:risk:framework:pcidss-4_0"),
    ("ref_id", PCI_DSS_VERSION_REF),
    ("name", f"{PCI_DSS_TITLE_EN} (PCI-DSS) 4.0.1"),
    ("description", PCI_DSS_DESCRIPTION_EN),
    ("name[fr]", PCI_DSS_TITLE_FR),
    ("description[fr]", PCI_DSS_DESCRIPTION_FR),
]

GOAL_BY_REQUIREMENT = {
    "1": "Build and Maintain a Secure Network and Systems",
    "3": "Protect Account Data",
    "5": "Maintain a Vulnerability Management Program",
    "7": "Implement Strong Access Control Measures",
    "10": "Regularly Monitor and Test Networks",
    "12": "Maintain an Information Security Policy",
}

GOAL_BY_REQUIREMENT_FR = {
    "1": "Créer et Maintenir un Réseau et des Systèmes Sécurisés",
    "3": "Protéger les Données de Carte",
    "5": "Maintenir un Programme de Gestion des Vulnérabilités",
    "7": "Mettre en œuvre des Mesures Robustes de Contrôle D'accès",
    "10": "Surveiller et Tester Régulièrement les Réseaux",
    "12": "Maintenir une Politique de Sécurité de L’information",
}

SYNTHETIC_REQUIREMENTS = {
    # Not present as a standalone row in the source sheet.
    "A2.1": "POI terminals using SSL and/or early TLS are confirmed as not susceptible to known SSL/TLS exploits.",
}
SYNTHETIC_REQUIREMENTS_FR = {
    "A2.1": "Il est confirmé que les terminaux POI utilisant SSL et/ou TLS obsolète ne sont pas vulnérables aux failles connues SSL/TLS.",
}

# Manual node_id remapping for specific PCI DSS 4.0.1 controls.
NODE_ID_BY_REF_ID = {
    "3.5.1.1": "3.5.2",
    "3.5.1.2": "3.5.3",
    "3.5.1.3": "3.5.4",
}

RE_REQUIREMENT = re.compile(
    r"^Requirement\s+(\d+)\s*:\s*(.+)$", re.IGNORECASE | re.DOTALL
)
RE_REQUIREMENT_FR = re.compile(
    r"^Exigence\s+(\d+)\s*:\s*(.+)$", re.IGNORECASE | re.DOTALL
)
# Captures appendix headers and keeps the appendix code (A1, A2, ...).
RE_APPENDIX = re.compile(r"^Appendix\s+(A\d+)\s*:\s*(.+)$", re.IGNORECASE | re.DOTALL)
RE_APPENDIX_FR = re.compile(r"^Annexe\s+(A\d+)\s*:\s*(.+)$", re.IGNORECASE | re.DOTALL)
# Handles OCR-ish refs such as: "7. 2 ...", "8. 1Processes ...", "11.2 2 ..."
RE_REF_LINE = re.compile(r"^((?:A\s*)?\d+(?:(?:\s*\.\s*|\s+)\d+)+)\s*(.*)$", re.DOTALL)


@dataclass
class ContentRow:
    assessable: str | None
    depth: int
    node_id: str | None
    ref_id: str | None
    name: str | None
    description: str | None
    name_fr: str | None
    description_fr: str | None


@dataclass
class ParsedItem:
    kind: str
    ref_id: str
    text: str


def normalize_newlines(value: str) -> str:
    return value.replace("\r\n", "\n").replace("\r", "\n")


def normalize_ref(raw_ref: str) -> str:
    # Canonicalize broken OCR numbering into a stable dotted ref_id.
    ref = re.sub(r"\s*\.\s*", ".", raw_ref.strip())
    ref = re.sub(r"\s+", ".", ref)
    ref = ref.replace("A.", "A")
    return ref


def clean_title(value: str) -> str:
    # Titles keep one-line text only and drop translated note blocks.
    value = normalize_newlines(value)
    value = re.split(
        r"\n\s*(?:Note|Remarque)\s*:\s*", value, maxsplit=1, flags=re.IGNORECASE
    )[0]
    value = re.sub(r"\n+", " ", value)
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def clean_control_text(value: str) -> str:
    # Keep control statement, remove applicability note section in EN/FR.
    value = normalize_newlines(value)
    value = re.split(
        r"\n\s*(?:Applicability Notes|Notes d'Applicabilité)\s*\n",
        value,
        maxsplit=1,
        flags=re.IGNORECASE,
    )[0]
    value = re.sub(r"\n[ \t]+", "\n", value)
    value = re.sub(r"[ \t]+\n", "\n", value)
    value = re.sub(r"\n{3,}", "\n\n", value)
    value = re.sub(r"[ \t]{2,}", " ", value)
    return value.strip()


def resolve_source_sheet(workbook) -> str:
    # Prefer the canonical EN sheet name, then fallback to EN/FR localized names.
    if SOURCE_SHEET in workbook.sheetnames:
        return SOURCE_SHEET

    for sheet_name in workbook.sheetnames:
        low = sheet_name.lower()
        if "prioritized" in low and "milestone" in low:
            return sheet_name
        if "jalon" in low and "approche" in low:
            return sheet_name

    raise ValueError(
        f"Unable to resolve source sheet from workbook sheets: {workbook.sheetnames}"
    )


def parse_source_items(source_xlsx: Path) -> list[ParsedItem]:
    # Parse a flat list of logical nodes from column A (requirements, appendices, controls).
    wb = load_workbook(source_xlsx, data_only=False)
    ws = wb[resolve_source_sheet(wb)]
    parsed: list[ParsedItem] = []

    for row in range(3, ws.max_row + 1):
        raw = ws.cell(row=row, column=1).value
        if not isinstance(raw, str):
            continue
        text = raw.strip()
        if not text:
            continue

        req_match = RE_REQUIREMENT.match(text) or RE_REQUIREMENT_FR.match(text)
        if req_match:
            req = req_match.group(1)
            title = clean_title(req_match.group(2))
            parsed.append(ParsedItem(kind="requirement", ref_id=req, text=title))
            continue

        appendix_match = RE_APPENDIX.match(text) or RE_APPENDIX_FR.match(text)
        if appendix_match:
            appendix_ref = appendix_match.group(1).upper()
            title = clean_title(appendix_match.group(2))
            parsed.append(ParsedItem(kind="appendix", ref_id=appendix_ref, text=title))
            continue

        ref_match = RE_REF_LINE.match(text)
        if ref_match:
            ref = normalize_ref(ref_match.group(1))
            body = clean_control_text(ref_match.group(2))
            parsed.append(ParsedItem(kind="control", ref_id=ref, text=body))

    return parsed


def infer_depth(ref_id: str) -> int:
    # Depth model expected by the library format (A is a special top appendix node).
    if ref_id.startswith("A") and "." not in ref_id:
        return 2
    return 2 + ref_id.count(".")


def objective_for_requirement(requirement_ref: str) -> str | None:
    return GOAL_BY_REQUIREMENT.get(requirement_ref)


def objective_for_requirement_fr(requirement_ref: str) -> str | None:
    return GOAL_BY_REQUIREMENT_FR.get(requirement_ref)


def build_ref_text_lookup(items: Iterable[ParsedItem]) -> dict[str, str]:
    # Keep first occurrence for duplicated refs in source files.
    lookup: dict[str, str] = {}
    for item in items:
        lookup.setdefault(item.ref_id, item.text)
    return lookup


def to_content_rows(
    items_en: Iterable[ParsedItem], fr_lookup: dict[str, str]
) -> list[ContentRow]:
    # Build final hierarchical rows in EN and enrich FR text by ref_id alignment.
    rows: list[ContentRow] = []
    seen_refs: set[str] = set()
    last_goal: str | None = None
    appendix_root_added = False

    for item in items_en:
        if item.kind == "requirement":
            goal = objective_for_requirement(item.ref_id)
            goal_fr = objective_for_requirement_fr(item.ref_id)
            if goal and goal != last_goal:
                rows.append(
                    ContentRow(
                        assessable=None,
                        depth=1,
                        node_id=None,
                        ref_id=None,
                        name=goal,
                        description=None,
                        name_fr=goal_fr,
                        description_fr=None,
                    )
                )
                last_goal = goal

            rows.append(
                ContentRow(
                    assessable=None,
                    depth=2,
                    node_id=None,
                    ref_id=item.ref_id,
                    name=f"Requirement {item.ref_id}",
                    description=item.text,
                    name_fr=f"Exigence {item.ref_id}",
                    description_fr=fr_lookup.get(item.ref_id),
                )
            )
            seen_refs.add(item.ref_id)
            continue

        if item.kind == "appendix":
            if not appendix_root_added:
                rows.append(
                    ContentRow(
                        assessable=None,
                        depth=1,
                        node_id=None,
                        ref_id="A",
                        name="Appendix A",
                        description="Additional PCI DSS Requirements",
                        name_fr="Annexe A",
                        description_fr="Exigences PCI DSS supplémentaires",
                    )
                )
                appendix_root_added = True

            rows.append(
                ContentRow(
                    assessable=None,
                    depth=2,
                    node_id=None,
                    ref_id=item.ref_id,
                    name=f"Appendix {item.ref_id}",
                    description=item.text,
                    name_fr=f"Annexe {item.ref_id}",
                    description_fr=fr_lookup.get(item.ref_id),
                )
            )
            seen_refs.add(item.ref_id)
            continue

        # control rows
        if item.ref_id.startswith("A2.1.") and "A2.1" not in seen_refs:
            rows.append(
                ContentRow(
                    assessable="x",
                    depth=3,
                    node_id=None,
                    ref_id="A2.1",
                    name=None,
                    description=SYNTHETIC_REQUIREMENTS["A2.1"],
                    name_fr=None,
                    description_fr=SYNTHETIC_REQUIREMENTS_FR["A2.1"],
                )
            )
            seen_refs.add("A2.1")

        depth = infer_depth(item.ref_id)
        rows.append(
            ContentRow(
                assessable="x" if depth >= 3 else None,
                depth=depth,
                node_id=NODE_ID_BY_REF_ID.get(item.ref_id),
                ref_id=item.ref_id,
                name=None,
                description=item.text,
                name_fr=None,
                description_fr=fr_lookup.get(item.ref_id),
            )
        )
        seen_refs.add(item.ref_id)

    return rows


def write_meta_sheet(ws, rows: list[tuple[str, str]]) -> None:
    for i, (key, value) in enumerate(rows, start=1):
        ws.cell(row=i, column=1, value=key)
        ws.cell(row=i, column=2, value=value)


def write_content_sheet(ws, rows: list[ContentRow]) -> None:
    headers = [
        "assessable",
        "depth",
        "node_id",
        "ref_id",
        "name",
        "description",
        "name[fr]",
        "description[fr]",
    ]
    ws.append(headers)

    for row in rows:
        ws.append(
            [
                row.assessable,
                row.depth,
                row.node_id,
                row.ref_id,
                row.name,
                row.description,
                row.name_fr,
                row.description_fr,
            ]
        )


def build_workbook(source_xlsx: Path, source_xlsx_fr: Path, output_xlsx: Path) -> int:
    items_en = parse_source_items(source_xlsx)
    items_fr = parse_source_items(source_xlsx_fr)
    fr_lookup = build_ref_text_lookup(items_fr)
    content_rows = to_content_rows(items_en, fr_lookup)

    wb = Workbook()
    ws_library = wb.active
    ws_library.title = "library_meta"
    ws_framework = wb.create_sheet("PCI DSS_meta")
    ws_content = wb.create_sheet(TARGET_CONTENT_SHEET)

    write_meta_sheet(ws_library, LIBRARY_META_ROWS)
    write_meta_sheet(ws_framework, FRAMEWORK_META_ROWS)
    write_content_sheet(ws_content, content_rows)

    wb.save(output_xlsx)
    return len(content_rows)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate a PCI DSS workbook from Prioritized Approach Milestones."
    )
    parser.add_argument(
        "--source",
        type=Path,
        default=Path("Prioritized-Approach-Tool-For-PCI-DSS-v4_0_1.xlsx"),
        help="Source workbook path.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("pcidss.generated.xlsx"),
        help="Output workbook path.",
    )
    parser.add_argument(
        "--source-fr",
        type=Path,
        default=DEFAULT_SOURCE_FR,
        help="French source workbook path used to fill [fr] columns.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    rows_count = build_workbook(args.source, args.source_fr, args.output)
    print(f"Generated {args.output} with {rows_count} rows in {TARGET_CONTENT_SHEET}.")


if __name__ == "__main__":
    main()
