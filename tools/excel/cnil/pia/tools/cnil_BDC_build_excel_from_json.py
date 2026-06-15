#!/usr/bin/env python3
"""
CNIL PIA Knowledge Base - Excel Framework Builder
=================================================

Builds a v2 CISO Assistant framework Excel workbook from `bdc_repo.json`.

Workbook includes:
- library metadata
- framework metadata and requirements
- implementation groups
- reference controls
- URN prefixes

Run
---
python cnil_BDC_build_excel_from_json.py --json bdc_repo.json --output cnil-pia-bdc.xlsx
"""

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook


DEFAULT_LOCALE = "fr"

FRAMEWORK_HEADERS_BASE = [
    "assessable",
    "depth",
    "ref_id",
    "name",
    "description",
    "annotation",
    "implementation_groups",
    "reference_controls",
]

IMP_GRP_HEADERS_BASE = ["ref_id", "name"]
REF_CTRL_HEADERS_BASE = [
    "ref_id",
    "name",
    "description",
    "annotation",
]

URN_PREF_HEADERS = ["prefix_id", "prefix_value"]
URN_PREF_ROWS = [
    {"prefix_id": "1", "prefix_value": "urn:intuitem:risk:function:cnil-bdc-pia"},
]

LIBRARY_NAME = "CNIL - Base de Connaissance PIA"
LIBRARY_REF_ID = "CNIL-BDC-PIA"
LIBRARY_URN = "urn:intuitem:risk:library:cnil-bdc-pia"
FRAMEWORK_URN = "urn:intuitem:risk:framework:cnil-bdc-pia"
REQ_BASE_URN = "urn:intuitem:risk:req_node:cnil-bdc-pia"
REF_CTRL_BASE_URN = "urn:intuitem:risk:function:cnil-bdc-pia"


def as_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def load_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        payload = json.load(f)
    if not isinstance(payload, dict):
        raise ValueError("The input JSON root must be an object")
    return payload


def translation_locales(payload: dict[str, Any]) -> list[str]:
    locales = payload.get("locales", [])
    if not isinstance(locales, list):
        return []
    return sorted(str(locale) for locale in locales if str(locale) != DEFAULT_LOCALE)


def headers_with_translations(base_headers: list[str], locales: list[str]) -> list[str]:
    headers = list(base_headers)
    for locale in locales:
        for field in ("name", "description", "annotation"):
            if field in base_headers:
                headers.append(f"{field}[{locale}]")
    return headers


def translated_value(
    translations: dict[str, Any],
    field: str,
    locale: str = DEFAULT_LOCALE,
) -> str:
    localized = translations.get(locale)
    if isinstance(localized, dict):
        return as_text(localized.get(field))
    return ""


def source_description(payload: dict[str, Any]) -> str:
    source = payload.get("source", {})
    if not isinstance(source, dict):
        return ""
    repo = as_text(source.get("repository"))
    kb = as_text(source.get("knowledge_base"))
    i18n = as_text(source.get("i18n"))
    parts = [part for part in (repo, kb, i18n) if part]
    if not parts:
        return ""
    return "Sources :\n" + "\n".join(f"- {part}" for part in parts)


def metadata_names(payload: dict[str, Any]) -> dict[str, str]:
    names = payload.get("framework_names", {})
    return names if isinstance(names, dict) else {}


def build_library_meta_rows(payload: dict[str, Any], locales: list[str]) -> list[tuple[str, str]]:
    names = metadata_names(payload)
    descriptions = payload.get("framework_descriptions", {})
    if not isinstance(descriptions, dict):
        descriptions = {}
    name = as_text(names.get(DEFAULT_LOCALE)) or LIBRARY_NAME
    description = as_text(descriptions.get(DEFAULT_LOCALE)) or source_description(payload)

    rows: list[tuple[str, str]] = [
        ("type", "library"),
        ("urn", LIBRARY_URN),
        ("version", "1"),
        ("locale", DEFAULT_LOCALE),
        ("ref_id", LIBRARY_REF_ID),
        ("name", name),
        ("description", description),
        (
            "copyright",
            "Copyright CNIL / LINC. See the source repository license and notices: https://github.com/LINCnil/pia",
        ),
        ("provider", "CNIL"),
        ("packager", "intuitem"),
    ]
    for locale in locales:
        if names.get(locale):
            rows.append((f"name[{locale}]", as_text(names.get(locale))))
        if descriptions.get(locale):
            rows.append((f"description[{locale}]", as_text(descriptions.get(locale))))
    return rows


def build_framework_meta_rows(payload: dict[str, Any], locales: list[str]) -> list[tuple[str, str]]:
    names = metadata_names(payload)
    descriptions = payload.get("framework_descriptions", {})
    if not isinstance(descriptions, dict):
        descriptions = {}
    name = as_text(names.get(DEFAULT_LOCALE)) or LIBRARY_NAME
    description = as_text(descriptions.get(DEFAULT_LOCALE)) or source_description(payload)

    rows: list[tuple[str, str]] = [
        ("type", "framework"),
        ("base_urn", REQ_BASE_URN),
        ("urn", FRAMEWORK_URN),
        ("ref_id", LIBRARY_REF_ID),
        ("name", name),
        ("description", description),
        ("implementation_groups_definition", "imp_grp"),
    ]
    for locale in locales:
        if names.get(locale):
            rows.append((f"name[{locale}]", as_text(names.get(locale))))
        if descriptions.get(locale):
            rows.append((f"description[{locale}]", as_text(descriptions.get(locale))))
    return rows


def build_imp_grp_rows(payload: dict[str, Any], locales: list[str]) -> list[dict[str, str]]:
    categories = payload.get("categories", {})
    if not isinstance(categories, dict):
        return []

    rows: list[dict[str, str]] = []
    for ref_id, category in sorted(categories.items()):
        if not isinstance(category, dict):
            continue
        translations = category.get("translations", {})
        if not isinstance(translations, dict):
            translations = {}

        row = {
            "ref_id": as_text(category.get("ref_id") or ref_id),
            "name": as_text(translations.get(DEFAULT_LOCALE) or ref_id),
        }
        for locale in locales:
            if translations.get(locale):
                row[f"name[{locale}]"] = as_text(translations.get(locale))
        rows.append(row)
    return rows


def knowledge_translations(entry: dict[str, Any]) -> dict[str, Any]:
    translations = entry.get("translations", {})
    return translations if isinstance(translations, dict) else {}


def build_framework_rows(payload: dict[str, Any], locales: list[str]) -> list[dict[str, str]]:
    entries = payload.get("knowledge_base", [])
    if not isinstance(entries, list):
        return []

    rows: list[dict[str, str]] = []
    for entry in entries:
        if not isinstance(entry, dict):
            continue
        slug = as_text(entry.get("slug")).strip()
        if not slug:
            continue
        translations = knowledge_translations(entry)
        row = {
            "assessable": "x",
            "depth": "1",
            "ref_id": slug,
            "name": translated_value(translations, "name"),
            "description": translated_value(translations, "description"),
            "annotation": translated_value(translations, "annotation"),
            "implementation_groups": as_text(entry.get("category")),
            "reference_controls": f"1:{slug}",
        }
        for locale in locales:
            row[f"name[{locale}]"] = translated_value(translations, "name", locale)
            row[f"description[{locale}]"] = translated_value(
                translations, "description", locale
            )
            row[f"annotation[{locale}]"] = translated_value(
                translations, "annotation", locale
            )
        rows.append(row)
    return rows


def build_ref_ctrl_rows(payload: dict[str, Any], locales: list[str]) -> list[dict[str, str]]:
    entries = payload.get("knowledge_base", [])
    if not isinstance(entries, list):
        return []

    rows: list[dict[str, str]] = []
    for entry in entries:
        if not isinstance(entry, dict):
            continue
        slug = as_text(entry.get("slug")).strip()
        if not slug:
            continue
        translations = knowledge_translations(entry)
        row = {
            "ref_id": slug,
            "name": translated_value(translations, "name"),
            "description": translated_value(translations, "description"),
            "annotation": translated_value(translations, "annotation"),
        }
        for locale in locales:
            row[f"name[{locale}]"] = translated_value(translations, "name", locale)
            row[f"description[{locale}]"] = translated_value(
                translations, "description", locale
            )
            row[f"annotation[{locale}]"] = translated_value(
                translations, "annotation", locale
            )
        rows.append(row)
    return rows


def write_sheet(
    wb: Workbook, title: str, headers: list[str], rows: list[dict[str, str]]
) -> None:
    ws = wb.create_sheet(title)
    ws.append(headers)
    for line in rows:
        ws.append([as_text(line.get(header, "")) for header in headers])

    for row in ws.iter_rows(
        min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)
    ):
        for cell in row:
            cell.value = as_text(cell.value)
            cell.number_format = "@"


def write_kv_sheet(wb: Workbook, title: str, rows: list[tuple[str, str]]) -> None:
    ws = wb.create_sheet(title)
    for key, value in rows:
        ws.append([as_text(key), as_text(value)])

    max_row = ws.max_row if ws.max_row > 0 else 1
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=2):
        for cell in row:
            cell.value = as_text(cell.value)
            cell.number_format = "@"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build a CISO Assistant XLSX file from bdc_repo.json"
    )
    parser.add_argument("--json", default="bdc_repo.json", help="Path to JSON input")
    parser.add_argument(
        "--output", default="cnil-pia-bdc.xlsx", help="Output XLSX file"
    )
    return parser.parse_args()


def build_excel_from_json(json_path: Path, output_path: Path) -> dict[str, int]:
    payload = load_json(json_path)
    locales = translation_locales(payload)

    fwk_headers = headers_with_translations(FRAMEWORK_HEADERS_BASE, locales)
    imp_grp_headers = headers_with_translations(IMP_GRP_HEADERS_BASE, locales)
    ref_ctrl_headers = headers_with_translations(REF_CTRL_HEADERS_BASE, locales)

    fwk_rows = build_framework_rows(payload, locales)
    imp_grp_rows = build_imp_grp_rows(payload, locales)
    ref_ctrl_rows = build_ref_ctrl_rows(payload, locales)

    wb = Workbook()
    wb.remove(wb.active)

    write_kv_sheet(wb, "library_meta", build_library_meta_rows(payload, locales))
    write_kv_sheet(wb, "fwk_meta", build_framework_meta_rows(payload, locales))
    write_sheet(wb, "fwk_content", fwk_headers, fwk_rows)
    write_kv_sheet(
        wb,
        "imp_grp_meta",
        [("type", "implementation_groups"), ("name", "imp_grp")],
    )
    write_sheet(wb, "imp_grp_content", imp_grp_headers, imp_grp_rows)
    write_kv_sheet(
        wb,
        "ref_ctrl_meta",
        [("type", "reference_controls"), ("base_urn", REF_CTRL_BASE_URN)],
    )
    write_sheet(wb, "ref_ctrl_content", ref_ctrl_headers, ref_ctrl_rows)
    write_kv_sheet(wb, "urn_pref_meta", [("type", "urn_prefix")])
    write_sheet(wb, "urn_pref_content", URN_PREF_HEADERS, URN_PREF_ROWS)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return {
        "fwk_content": len(fwk_rows),
        "imp_grp_content": len(imp_grp_rows),
        "ref_ctrl_content": len(ref_ctrl_rows),
        "locales": len(locales) + 1,
    }


def main() -> None:
    args = parse_args()
    stats = build_excel_from_json(Path(args.json), Path(args.output))
    print(f"Excel generated: {args.output}")
    print(f"- fwk_content: {stats['fwk_content']} row(s)")
    print(f"- imp_grp_content: {stats['imp_grp_content']} row(s)")
    print(f"- ref_ctrl_content: {stats['ref_ctrl_content']} row(s)")
    print(f"- locales: {stats['locales']}")


if __name__ == "__main__":
    main()
