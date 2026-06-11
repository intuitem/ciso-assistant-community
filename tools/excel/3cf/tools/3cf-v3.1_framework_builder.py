#!/usr/bin/env python3
"""Build a 3CF workbook directly from a PDF export.

The script creates the workbook structure on its own and builds the
`3cf_content` sheet from the PDF content between section 3 and section 8.2.
"""

from __future__ import annotations

import argparse
import bisect
import re
import subprocess
import tempfile
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook

DEFAULT_START_REF = "3"
DEFAULT_END_REF = "8.2"
DEFAULT_FILENAME = "3cf-v3.1.xlsx"

CONTENT_HEADERS = [
    "assessable",
    "depth",
    "implementation_groups",
    "ref_id",
    "name",
    "description",
    "skip_count",
]

IMP_GRP_META_ROWS = [
    ("type", "implementation_groups"),
    ("name", "imp_grp"),
]

IMP_GRP_HEADERS = ["ref_id", "name", "description"]

IMP_GRP_CONTENT_ROWS = [
    (
        "sec",
        "Sécurité aérienne",
        "État dans lequel les risques liés aux activités aéronautiques concernant, "
        "ou appuyant directement, l’exploitation des aéronefs sont réduits et "
        "maîtrisés à un niveau acceptable",
    ),
    (
        "sur",
        "Sûreté aérienne",
        "Combinaison des mesures ainsi que des moyens humains et matériels visant à "
        "protéger l’aviation civile contre les actes d’interventions illicites.",
    ),
]

CONTENT_COLUMN_WIDTHS = {
    "A": 9.0,
    "B": 5.83203125,
    "C": 20.5,
    "D": 10.5,
    "E": 55.83203125,
    "F": 140.1640625,
}

IMP_GRP_COLUMN_WIDTHS = {
    "A": 10.83203125,
    "B": 14.1640625,
    "C": 69.83203125,
}

PAGE_RE = re.compile(r"Page\s*:\s*(\d+)\s*/\s*(\d+)")
HEADING_RE = re.compile(r"^\s*(\d+(?:\.\d+)*)\.\s+(.+?)\s*$")
BULLET_RE = re.compile(r"^(?P<indent>\s*)(?P<bullet>[-o▪•])\s+(?P<text>.*\S)?\s*$")
REF_START_RE = re.compile(
    r"^\s*(IS\.I/D\.OR|Règlement \(UE\)|Code des transports\s*:\s*L|AIM\b)"
)
GOOD_PRACTICE_RE = re.compile(r"Bonne\s+pratique\s+n°", re.IGNORECASE)
DURATION_START_RE = re.compile(r"^(Au moins|La\s+durée|Tant\s+que)\b")
VERSION_RE = re.compile(r"\bVersion(?:\s*n[°o])?\s*([0-9]+(?:\.[0-9]+)*)\b", re.IGNORECASE)
DATE_RE = re.compile(r"\bdu\s+([0-9]{1,2}\s+[A-Za-zéûôàèùîïÉÛÔÀÈÙÎÏ]+\s+[0-9]{4})\b")
FILENAME_VERSION_RE = re.compile(r"version\s*([0-9]+(?:\.[0-9]+)*)", re.IGNORECASE)
LEADER_DOTS_RE = re.compile(r"\.{3,}")
CHAPTER_TOKEN_RE = re.compile(
    r"^(?:3CFv[0-9.]+|[0-9]+(?:\.[0-9]+)*(?:\s*(?:,|et)\s*[0-9]+(?:\.[0-9]+)*)*|et\s+[0-9]+(?:\.[0-9]+)*)$"
)

GUIDANCE_PREFIXES = (
    "Pour la suite,",
    "Néanmoins,",
    "Il est recommandé",
    "L’organisme peut ",
    "L'organisme peut ",
    "L’organisme peut également",
    "L'organisme peut egalement",
    "Uniquement à destination",
    "Uniquement a destination",
    "Note :",
    "Note:",
)


@dataclass
class ContentRow:
    assessable: str | None
    depth: int
    implementation_groups: str | None
    ref_id: str | None
    name: str | None
    description: str | None
    skip_count: str | None = None


@dataclass
class PendingItem:
    indent: int
    is_bullet: bool
    depth: int
    assessable: str | None
    text: str


@dataclass
class TableRow:
    doc_lines: list[str]
    duration_lines: list[str]


@dataclass
class ChapterTableEntry:
    doc_lines: list[str]
    chapter_lines: list[str]
    is_bullet: bool


@dataclass
class DocumentMetadata:
    slug: str
    version: str
    ref_id: str
    name: str
    description: str


def normalize_spaces(value: str) -> str:
    value = value.replace("\r\n", "\n").replace("\r", "\n")
    value = re.sub(r"(?<=[A-Za-zÀ-ÿ)])\d+(?=\s)", "", value)
    value = re.sub(r"\s+", " ", value)
    value = value.replace(" ,", ",").replace(" .", ".").replace(" ;", ";")
    value = value.replace(" :", ":")
    return value.strip()


def infer_assessable(text: str) -> str | None:
    normalized = normalize_spaces(text)
    for prefix in GUIDANCE_PREFIXES:
        if normalized.startswith(prefix):
            return None
    return "x"


def heading_depth(ref_id: str) -> int:
    return ref_id.count(".") + 1


def slugify(value: str) -> str:
    slug = value.lower()
    slug = re.sub(r"[^a-z0-9.]+", "-", slug)
    slug = re.sub(r"-{2,}", "-", slug).strip("-")
    return slug or "3cf"


def parse_heading_line(line: str) -> tuple[str, str] | None:
    match = HEADING_RE.match(line)
    if not match:
        return None

    title = normalize_spaces(match.group(2))
    if LEADER_DOTS_RE.search(title):
        return None
    if re.search(r"\s\d+\s*$", title) and LEADER_DOTS_RE.search(line):
        return None

    return match.group(1), title


def infer_metadata(pdf_path: Path, extracted_text: str) -> DocumentMetadata:
    intro_text = extracted_text[:12000]

    version_match = VERSION_RE.search(intro_text) or FILENAME_VERSION_RE.search(pdf_path.stem)
    version = version_match.group(1) if version_match else "1"

    date_match = DATE_RE.search(intro_text)
    date_text = date_match.group(1) if date_match else None

    ref_id = f"3CF-v{version}"
    slug = f"3cf-v{slugify(version)}"
    name = f"Cadre de Conformité Cyber France (3CF) pour l'aviation civile - v{version}"

    description = (
        "Ce document, établi par la direction de la sécurité de l’aviation civile "
        "(DSAC) et l’OSAC, présente le Cadre de Conformité Cyber France (3CF) "
        "pour l’aviation civile."
    )
    if date_text:
        description = f"{description}\nVersion {version} du {date_text}"
    else:
        description = f"{description}\nVersion {version}"

    return DocumentMetadata(
        slug=slug,
        version=version,
        ref_id=ref_id,
        name=name,
        description=description,
    )


def build_library_meta_rows(metadata: DocumentMetadata) -> list[tuple[str, str]]:
    return [
        ("type", "library"),
        ("urn", f"urn:intuitem:risk:library:{metadata.slug}"),
        ("version", "1"),
        ("locale", "fr"),
        ("ref_id", metadata.ref_id),
        ("name", metadata.name),
        ("description", metadata.description),
        ("provider", "DSAC / OSAC"),
        ("packager", "intuitem"),
    ]


def build_framework_meta_rows(metadata: DocumentMetadata) -> list[tuple[str, str]]:
    return [
        ("type", "framework"),
        ("base_urn", f"urn:intuitem:risk:req_node:{metadata.slug}"),
        ("urn", f"urn:intuitem:risk:framework:{metadata.slug}"),
        ("ref_id", metadata.ref_id),
        ("name", metadata.name),
        ("description", metadata.description),
        ("implementation_groups_definition", "imp_grp"),
    ]


def is_reference_line(line: str) -> bool:
    return bool(REF_START_RE.match(line))


def implementation_groups_from_refs(ref_lines: list[str]) -> str | None:
    has_sec = any(line.strip().startswith("IS.I/D.OR") for line in ref_lines)
    has_sur = any(
        any(marker in line for marker in ("AIM", "Règlement (UE)", "Code des transports"))
        for line in ref_lines
    )
    if has_sec and has_sur:
        return "sec,sur"
    if has_sec:
        return "sec"
    if has_sur:
        return "sur"
    return None


def merge_implementation_groups(
    current_value: str | None, new_value: str | None
) -> str | None:
    if not current_value:
        return new_value
    if not new_value:
        return current_value

    values = []
    for value in [*current_value.split(","), *new_value.split(",")]:
        if value not in values:
            values.append(value)
    return ",".join(values)


def apply_pending_implementation_groups(
    current_value: str | None,
    ref_lines: list[str],
) -> str | None:
    new_value = implementation_groups_from_refs(ref_lines)
    return new_value or current_value


def extract_pdf_text(pdf_path: Path) -> str:
    with tempfile.TemporaryDirectory(prefix="3cf_pdf_") as tmp_dir:
        out_path = Path(tmp_dir) / "3cf.txt"
        cmd = [
            "gs",
            "-q",
            "-dNOPAUSE",
            "-dBATCH",
            "-sDEVICE=txtwrite",
            f"-sOutputFile={out_path}",
            str(pdf_path),
        ]
        result = subprocess.run(cmd, capture_output=True, text=True, check=False)
        if result.returncode != 0:
            raise RuntimeError(
                "Ghostscript text extraction failed:\n"
                f"stdout:\n{result.stdout}\n"
                f"stderr:\n{result.stderr}"
            )
        return out_path.read_text(encoding="utf-8")


def split_pages(extracted_text: str) -> dict[int, list[str]]:
    pages: dict[int, list[str]] = {}
    current_lines: list[str] = []

    for line in extracted_text.splitlines():
        page_match = PAGE_RE.search(line)
        if page_match:
            page_number = int(page_match.group(1))
            pages[page_number] = current_lines
            current_lines = []
            continue
        current_lines.append(line.rstrip())

    return pages


def clean_page_lines(lines: list[str]) -> list[str]:
    cleaned: list[str] = []
    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue
        if "TLP:CLEAR" in stripped:
            continue
        if stripped.startswith("Cadre de Conformité Cyber France"):
            continue
        if stripped.startswith("Version "):
            continue
        if stripped.startswith("du ") and re.search(r"\d{4}", stripped):
            continue
        if "Version " in stripped and re.search(r"\d{4}", stripped):
            continue
        if PAGE_RE.search(stripped):
            continue
        if re.fullmatch(r"\d+", stripped):
            continue
        cleaned.append(line.rstrip())
    return cleaned


def split_table_columns(line: str) -> tuple[str, str | None]:
    trimmed = line.rstrip()
    matches = [match for match in re.finditer(r"\s{50,}", trimmed) if match.start() > 0]
    if not matches:
        return trimmed, None

    longest = max(matches, key=lambda match: match.end() - match.start())
    left = trimmed[: longest.start()].rstrip()
    right = trimmed[longest.end() :].strip()
    if left.strip() and right:
        return left, right
    return trimmed, None


def line_needs_continuation(text: str) -> bool:
    normalized = normalize_spaces(text).lower()
    if not normalized:
        return True
    if normalized.endswith((":", ";", ",", "/", "(")):
        return True

    last_word = re.split(r"\s+", normalized)[-1]
    return last_word in {
        "le",
        "la",
        "les",
        "de",
        "du",
        "des",
        "pour",
        "par",
        "au",
        "aux",
        "et",
        "ou",
        "un",
        "une",
        "celle",
        "validité",
        "modification",
        "création",
        "travaille",
        "après",
    }


def table_row_is_complete(row: TableRow) -> bool:
    if not row.doc_lines or not row.duration_lines:
        return False
    return not line_needs_continuation(row.doc_lines[-1]) and not line_needs_continuation(
        row.duration_lines[-1]
    )


def parse_archive_table(lines: list[str]) -> list[TableRow]:
    rows: list[TableRow] = []
    current = TableRow(doc_lines=[], duration_lines=[])

    def flush_current() -> None:
        nonlocal current
        if current.doc_lines or current.duration_lines:
            rows.append(current)
        current = TableRow(doc_lines=[], duration_lines=[])

    for raw_line in lines:
        stripped = raw_line.strip()
        if not stripped:
            continue
        if stripped == "Documents" or stripped == "Durée d’archivage":
            continue

        indent = len(raw_line) - len(raw_line.lstrip(" "))
        left, right = split_table_columns(raw_line)
        left_text = left.strip()

        if right and DURATION_START_RE.match(normalize_spaces(right)):
            if current.doc_lines and not current.duration_lines:
                if left_text:
                    current.doc_lines.append(left)
                current.duration_lines.append(right)
            else:
                flush_current()
                if left_text:
                    current.doc_lines.append(left)
                current.duration_lines.append(right)
            continue

        if right is not None:
            if left_text:
                current.doc_lines.append(left)
            if right.strip():
                current.duration_lines.append(right)
            continue

        if current.duration_lines and indent >= 300:
            current.duration_lines.append(stripped)
            continue

        if current.duration_lines and indent <= 150 and table_row_is_complete(current):
            flush_current()

        if left_text:
            current.doc_lines.append(left)

    flush_current()
    return rows


def is_chapter_value(text: str) -> bool:
    normalized = normalize_spaces(text)
    return bool(CHAPTER_TOKEN_RE.match(normalized))


def is_protected_chapterless_entry(text: str) -> bool:
    normalized = normalize_spaces(text)
    return normalized.startswith("Tous les documents")


def should_start_new_chapter_entry(
    current: ChapterTableEntry | None,
    left_text: str,
    right_text: str | None,
) -> bool:
    if current is None:
        return True
    if not left_text:
        return False

    bullet_match = BULLET_RE.match(left_text)
    if bullet_match:
        return True

    if right_text and current.doc_lines and not line_needs_continuation(current.doc_lines[-1]):
        return True

    if current.is_bullet:
        return False

    if current.chapter_lines:
        if line_needs_continuation(current.doc_lines[-1]):
            return False
        if re.match(r"^(et|ou|de|du|des|aux|au)\b", left_text.lower()):
            return False
        return True

    return False


def propagate_chapters(entries: list[ChapterTableEntry]) -> None:
    for index, entry in enumerate(entries):
        if entry.is_bullet or not entry.chapter_lines:
            continue

        chapter_lines = list(entry.chapter_lines)
        assigned_backward = False

        previous = index - 1
        while previous >= 0 and not entries[previous].is_bullet:
            target = entries[previous]
            if target.chapter_lines:
                break
            if is_protected_chapterless_entry(" ".join(target.doc_lines)):
                break
            target.chapter_lines = list(chapter_lines)
            assigned_backward = True
            previous -= 1

        if assigned_backward:
            following = index + 1
            while following < len(entries) and not entries[following].is_bullet:
                target = entries[following]
                if target.chapter_lines:
                    break
                if is_protected_chapterless_entry(" ".join(target.doc_lines)):
                    break
                target.chapter_lines = list(chapter_lines)
                following += 1


def parse_chapter_table(lines: list[str]) -> tuple[list[ChapterTableEntry], int]:
    entries: list[ChapterTableEntry] = []
    current: ChapterTableEntry | None = None
    consumed_count = 0

    def flush_current() -> None:
        nonlocal current
        if current and current.doc_lines:
            entries.append(current)
        current = None

    for raw_line in lines:
        stripped = raw_line.strip()
        indent = len(raw_line) - len(raw_line.lstrip(" "))
        if not stripped:
            consumed_count += 1
            continue
        if stripped in {"Documents", "Chapitres"}:
            consumed_count += 1
            continue
        if stripped.startswith("3CFv"):
            consumed_count += 1
            continue
        if GOOD_PRACTICE_RE.search(stripped):
            break
        if parse_heading_line(stripped):
            break

        if is_chapter_value(stripped) and current is not None and indent >= 200:
            current.chapter_lines.append(stripped)
            consumed_count += 1
            continue

        left, right = split_table_columns(raw_line)
        left_text = left.strip()
        right_text = right.strip() if right else None

        if not left_text and is_chapter_value(stripped) and current is not None:
            current.chapter_lines.append(stripped)
            consumed_count += 1
            continue

        if should_start_new_chapter_entry(current, left_text, right_text):
            flush_current()
            current = ChapterTableEntry(
                doc_lines=[],
                chapter_lines=[],
                is_bullet=bool(BULLET_RE.match(left_text)),
            )

        if current is None:
            current = ChapterTableEntry(doc_lines=[], chapter_lines=[], is_bullet=False)

        if left_text:
            current.doc_lines.append(left)
        if right_text and is_chapter_value(right_text):
            current.chapter_lines.append(right_text)
        consumed_count += 1

    flush_current()
    propagate_chapters(entries)
    return entries, consumed_count


def chapter_table_rows_to_content(
    entries: list[ChapterTableEntry],
    heading_depth_value: int,
    implementation_groups: str | None,
) -> list[ContentRow]:
    rows: list[ContentRow] = []
    for entry in entries:
        doc_rows = parse_doc_lines(
            entry.doc_lines,
            heading_depth_value,
            implementation_groups,
        )
        rows.extend(doc_rows)
    return rows


def append_archive_duration(
    rows: list[ContentRow],
    duration_lines: list[str],
) -> list[ContentRow]:
    if not duration_lines:
        return rows

    duration_text = normalize_spaces(" ".join(duration_lines))
    if not duration_text:
        return rows

    for row in rows:
        if row.description:
            row.description = (
                f"{row.description}\nDurée d’archivage : {duration_text}"
            )
    return rows


def build_documents_parent_row(
    heading_depth_value: int,
    implementation_groups: str | None,
) -> ContentRow:
    return ContentRow(
        assessable=None,
        depth=heading_depth_value + 1,
        implementation_groups=implementation_groups,
        ref_id=None,
        name=None,
        description="Documents :",
    )


def add_indent_level(levels: list[int], indent: int) -> int:
    if indent not in levels:
        bisect.insort(levels, indent)
    return levels.index(indent)


def should_append_to_current(current_item: PendingItem | None, indent: int, text: str) -> bool:
    if current_item is None:
        return False

    if current_item.is_bullet:
        return indent > current_item.indent

    if indent > current_item.indent:
        return True

    if indent < current_item.indent:
        return False

    if text and text[0].islower():
        return True

    if not re.search(r"[.;:!?)]$", current_item.text):
        return True

    return False


def emit_pending_item(
    rows: list[ContentRow],
    pending_item: PendingItem | None,
    implementation_groups: str | None,
) -> None:
    if pending_item is None:
        return
    description = normalize_spaces(pending_item.text)
    assessable = pending_item.assessable
    if description.endswith(":"):
        assessable = None
    rows.append(
        ContentRow(
            assessable=assessable,
            depth=pending_item.depth,
            implementation_groups=implementation_groups,
            ref_id=None,
            name=None,
            description=description,
        )
    )


def parse_doc_lines(
    doc_lines: list[str],
    heading_depth_value: int,
    implementation_groups: str | None,
) -> list[ContentRow]:
    rows: list[ContentRow] = []
    pending_item: PendingItem | None = None
    indent_levels: list[int] = []

    for raw_line in doc_lines:
        line = raw_line.rstrip()
        if not line.strip():
            continue

        bullet_match = BULLET_RE.match(line)
        if bullet_match:
            emit_pending_item(rows, pending_item, implementation_groups)
            indent = len(bullet_match.group("indent"))
            level = add_indent_level(indent_levels, indent)
            pending_item = PendingItem(
                indent=indent,
                is_bullet=True,
                depth=heading_depth_value + 1 + level,
                assessable="x",
                text=bullet_match.group("text") or "",
            )
            continue

        indent = len(line) - len(line.lstrip(" "))
        text = line.strip()
        if should_append_to_current(pending_item, indent, text):
            pending_item.text = f"{pending_item.text} {text}"
            continue

        emit_pending_item(rows, pending_item, implementation_groups)
        level = add_indent_level(indent_levels, indent)
        pending_item = PendingItem(
            indent=indent,
            is_bullet=False,
            depth=heading_depth_value + 1 + level,
            assessable=infer_assessable(text),
            text=text,
        )

    emit_pending_item(rows, pending_item, implementation_groups)
    return rows


def parse_relevant_content(
    pages: dict[int, list[str]],
    start_ref: str,
    end_ref: str,
) -> list[ContentRow]:
    rows: list[ContentRow] = []
    end_major = int(end_ref.split(".")[0])
    capture_started = False
    finished = False
    current_heading_ref: str | None = None
    current_heading_depth = 0
    current_implementation_groups: str | None = None
    pending_refs: list[str] = []
    pending_item: PendingItem | None = None
    indent_levels: list[int] = []

    for page_number in sorted(pages):
        if finished:
            break
        page_lines = clean_page_lines(pages.get(page_number, []))
        at_page_start = True
        index = 0

        while index < len(page_lines):
            if finished:
                break
            line = page_lines[index]
            stripped = line.strip()

            heading = parse_heading_line(stripped)
            if heading:
                emit_pending_item(rows, pending_item, current_implementation_groups)
                pending_item = None
                pending_refs = []
                indent_levels = []
                current_implementation_groups = None

                current_heading_ref, title = heading
                current_major = int(current_heading_ref.split(".")[0])

                if not capture_started:
                    if current_heading_ref != start_ref:
                        current_heading_ref = None
                        index += 1
                        continue
                    capture_started = True

                if current_major > end_major:
                    finished = True
                    break

                current_heading_depth = heading_depth(current_heading_ref)
                rows.append(
                    ContentRow(
                        assessable=None,
                        depth=current_heading_depth,
                        implementation_groups=None,
                        ref_id=current_heading_ref,
                        name=title,
                        description=None,
                    )
                )
                at_page_start = False
                index += 1
                continue

            if not capture_started or current_heading_ref is None:
                index += 1
                continue

            if GOOD_PRACTICE_RE.search(stripped):
                emit_pending_item(rows, pending_item, current_implementation_groups)
                pending_item = None
                indent_levels = []
                index += 1
                continue

            if is_reference_line(line):
                emit_pending_item(rows, pending_item, current_implementation_groups)
                pending_item = None
                pending_refs.append(stripped)
                index += 1
                continue

            if current_heading_ref == "7.2" and stripped.startswith("Documents"):
                emit_pending_item(rows, pending_item, current_implementation_groups)
                pending_item = None
                if pending_refs:
                    current_implementation_groups = apply_pending_implementation_groups(
                        current_implementation_groups,
                        pending_refs,
                    )
                    pending_refs = []
                rows.append(
                    build_documents_parent_row(
                        current_heading_depth,
                        current_implementation_groups,
                    )
                )
                for table_row in parse_archive_table(page_lines[index + 1 :]):
                    doc_rows = parse_doc_lines(
                        table_row.doc_lines,
                        current_heading_depth + 1,
                        current_implementation_groups,
                    )
                    rows.extend(
                        append_archive_duration(doc_rows, table_row.duration_lines)
                    )
                if current_heading_ref == end_ref:
                    finished = True
                index = len(page_lines)
                continue

            if current_heading_ref in {"8.1", "8.2"} and stripped.startswith("Documents"):
                emit_pending_item(rows, pending_item, current_implementation_groups)
                pending_item = None
                if pending_refs:
                    current_implementation_groups = apply_pending_implementation_groups(
                        current_implementation_groups,
                        pending_refs,
                    )
                    pending_refs = []
                rows.append(
                    build_documents_parent_row(
                        current_heading_depth,
                        current_implementation_groups,
                    )
                )

                chapter_entries, consumed_count = parse_chapter_table(page_lines[index + 1 :])
                rows.extend(
                    chapter_table_rows_to_content(
                        entries=chapter_entries,
                        heading_depth_value=current_heading_depth + 1,
                        implementation_groups=current_implementation_groups,
                    )
                )
                if current_heading_ref == end_ref:
                    finished = True
                else:
                    index += 1 + consumed_count
                continue

            if pending_refs:
                current_implementation_groups = apply_pending_implementation_groups(
                    current_implementation_groups,
                    pending_refs,
                )
                if implementation_groups_from_refs(pending_refs):
                    indent_levels = []
                pending_refs = []

            bullet_match = BULLET_RE.match(line)
            if bullet_match:
                emit_pending_item(rows, pending_item, current_implementation_groups)
                indent = len(bullet_match.group("indent"))
                level = add_indent_level(indent_levels, indent)
                pending_item = PendingItem(
                    indent=indent,
                    is_bullet=True,
                    depth=current_heading_depth + 1 + level,
                    assessable="x",
                    text=bullet_match.group("text") or "",
                )
                at_page_start = False
                index += 1
                continue

            indent = len(line) - len(line.lstrip(" "))
            text = stripped
            if should_append_to_current(pending_item, indent, text):
                pending_item.text = f"{pending_item.text} {text}"
            else:
                emit_pending_item(rows, pending_item, current_implementation_groups)
                level = add_indent_level(indent_levels, indent)
                pending_item = PendingItem(
                    indent=indent,
                    is_bullet=False,
                    depth=current_heading_depth + 1 + level,
                    assessable=infer_assessable(text),
                    text=text,
                )
            at_page_start = False
            index += 1

        emit_pending_item(rows, pending_item, current_implementation_groups)
        pending_item = None

    return rows


def replace_two_column_sheet(sheet, rows: list[tuple[str, str]]) -> None:
    if sheet.max_row:
        sheet.delete_rows(1, sheet.max_row)
    for row_index, (key, value) in enumerate(rows, start=1):
        sheet.cell(row=row_index, column=1, value=key)
        sheet.cell(row=row_index, column=2, value=value)


def replace_content_sheet(sheet, rows: list[ContentRow]) -> None:
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)

    for row_index, row in enumerate(rows, start=2):
        values = [
            row.assessable,
            row.depth,
            row.implementation_groups,
            row.ref_id,
            row.name,
            row.description,
            row.skip_count,
        ]
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column_index, value=value)


def replace_implementation_groups_content_sheet(sheet) -> None:
    if sheet.max_row:
        sheet.delete_rows(1, sheet.max_row)

    for column_index, header in enumerate(IMP_GRP_HEADERS, start=1):
        sheet.cell(row=1, column=column_index, value=header)

    for row_index, row in enumerate(IMP_GRP_CONTENT_ROWS, start=2):
        for column_index, value in enumerate(row, start=1):
            sheet.cell(row=row_index, column=column_index, value=value)


def create_workbook() -> Workbook:
    workbook = Workbook()

    library_meta = workbook.active
    library_meta.title = "library_meta"
    workbook.create_sheet("3cf_meta")
    content_sheet = workbook.create_sheet("3cf_content")
    workbook.create_sheet("imp_grp_meta")
    implementation_groups_content = workbook.create_sheet("imp_grp_content")

    for column_index, header in enumerate(CONTENT_HEADERS, start=1):
        content_sheet.cell(row=1, column=column_index, value=header)
    for column_name, width in CONTENT_COLUMN_WIDTHS.items():
        content_sheet.column_dimensions[column_name].width = width

    for column_index, header in enumerate(IMP_GRP_HEADERS, start=1):
        implementation_groups_content.cell(row=1, column=column_index, value=header)
    for column_name, width in IMP_GRP_COLUMN_WIDTHS.items():
        implementation_groups_content.column_dimensions[column_name].width = width

    return workbook


def build_workbook(
    pdf_path: Path,
    output_path: Path,
    start_ref: str,
    end_ref: str,
) -> None:
    extracted_text = extract_pdf_text(pdf_path)
    metadata = infer_metadata(pdf_path, extracted_text)
    pages = split_pages(extracted_text)
    content_rows = parse_relevant_content(
        pages=pages,
        start_ref=start_ref,
        end_ref=end_ref,
    )

    workbook = create_workbook()

    replace_two_column_sheet(workbook["library_meta"], build_library_meta_rows(metadata))
    replace_two_column_sheet(workbook["3cf_meta"], build_framework_meta_rows(metadata))
    replace_content_sheet(workbook["3cf_content"], content_rows)
    replace_two_column_sheet(
        workbook["imp_grp_meta"], IMP_GRP_META_ROWS
    )
    replace_implementation_groups_content_sheet(workbook["imp_grp_content"])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build a 3CF XLSX workbook from a PDF export."
    )
    parser.add_argument("pdf", type=Path, help="Source PDF path")
    parser.add_argument(
        "--output",
        default=DEFAULT_FILENAME,
        type=Path,
        help="Output XLSX path (defaults to the PDF path with a .xlsx extension)",
    )
    parser.add_argument(
        "--start-ref",
        default=DEFAULT_START_REF,
        help="First numbered section to include (default: 3)",
    )
    parser.add_argument(
        "--end-ref",
        default=DEFAULT_END_REF,
        help="Last numbered section to include (default: 8.2)",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output_path = args.output.resolve() if args.output else args.pdf.with_suffix(".xlsx").resolve()
    build_workbook(
        pdf_path=args.pdf.resolve(),
        output_path=output_path,
        start_ref=args.start_ref,
        end_ref=args.end_ref,
    )
    print(f"Workbook generated: {output_path}")


if __name__ == "__main__":
    main()
