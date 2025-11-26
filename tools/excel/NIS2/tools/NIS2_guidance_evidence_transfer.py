"""
This script copies the columns "annotation" and "typical_evidence" from a source Excel file to a destination Excel file
by matching rows based on the "ref_id" column, inserting the new columns after "description" in the "req_content" sheet.
"""

import argparse
import shutil
from pathlib import Path
from openpyxl import load_workbook

def transfer_columns(source_path, dest_path):
    # Load the source Excel file
    src_wb = load_workbook(source_path)
    src_ws = src_wb.active

    # Read headers from the source file
    src_headers = [cell.value for cell in next(src_ws.iter_rows(min_row=1, max_row=1))]

    # Get indexes of required columns
    try:
        src_ref_idx = src_headers.index("ref_id")
        src_annotation_idx = src_headers.index("annotation")
        src_evidence_idx = src_headers.index("typical_evidence")
    except ValueError as e:
        raise Exception(f"Missing column in source file: {e}")

    # Build a mapping from ref_id to (annotation, typical_evidence)
    ref_mapping = {}
    for row in src_ws.iter_rows(min_row=2, values_only=True):
        ref_id = str(row[src_ref_idx]).strip()
        if ref_id in ref_mapping:
            raise Exception(f"Duplicate 'ref_id' found in source file: {ref_id}")
        ref_mapping[ref_id] = (
            row[src_annotation_idx],
            row[src_evidence_idx]
        )

    # Load the destination Excel file
    dest_wb = load_workbook(dest_path)
    if "req_content" not in dest_wb.sheetnames:
        raise Exception("Sheet 'req_content' is missing in the destination file.")
    dest_ws = dest_wb["req_content"]

    # Read headers from the destination file
    dest_headers = [cell.value for cell in next(dest_ws.iter_rows(min_row=1, max_row=1))]

    try:
        ref_col_idx = dest_headers.index("ref_id")
        desc_col_idx = dest_headers.index("description")
    except ValueError as e:
        raise Exception(f"Missing column in destination file: {e}")

    # Insert new columns after 'description'
    annotation_col = desc_col_idx + 2
    evidence_col = desc_col_idx + 3
    dest_ws.insert_cols(annotation_col, amount=2)

    # Write headers for new columns
    dest_ws.cell(row=1, column=annotation_col, value="annotation")
    dest_ws.cell(row=1, column=evidence_col, value="typical_evidence")

    # Populate new columns using matching ref_id
    for row in dest_ws.iter_rows(min_row=2):
        ref_value = str(row[ref_col_idx].value).strip()
        if ref_value in ref_mapping:
            row[annotation_col - 1].value = ref_mapping[ref_value][0]
            row[evidence_col - 1].value = ref_mapping[ref_value][1]

    # Save the updated destination file
    dest_wb.save(dest_path)
    print(f"Transfer completed successfully. Output saved to '{dest_path}'.")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Transfer 'annotation' and 'typical_evidence' columns based on matching 'ref_id'")
    parser.add_argument("source", help="Path to the source Excel file")
    parser.add_argument("destination", nargs="?", help="Path to the destination Excel file (optional). If not provided, 'output.xlsx' will be created.")
    args = parser.parse_args()

    source_path = Path(args.source)
    if args.destination:
        destination_path = Path(args.destination)
    else:
        destination_path = Path("output.xlsx")
        shutil.copy(source_path, destination_path)
        print(f"No destination provided. Using '{destination_path}' as the output file.")

    transfer_columns(str(source_path), str(destination_path))
