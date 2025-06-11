"""
Simple PDF text extractor and structurer for FINMA 2023/01.

This script reads a PDF file page by page, extracts text blocks,
and attempts to detect hierarchical structures like chapters,
subchapters, sections, and paragraphs.

It also tries to find and link annotations (footnotes or references)
to corresponding paragraph markers using regex.

⚠️ WARNING: This code performs a partial automated extraction.
It does not guarantee perfect accuracy, especially for complex layouts,
multi-line titles, or accented characters at all.

Final review and manual adjustments are strongly recommended
to ensure data quality and correct structure before any further use.
"""


import fitz
import re
from openpyxl import Workbook

REF_ID_PREFIX = 'Rz'

def is_chapter(line): return re.match(r"^[IVX]+\.$", line.strip())
def is_subchapter(line): return re.match(r"^[A-Z]\.$", line.strip())
def is_section(line): return re.match(r"^[a-z]\)$", line.strip())
def is_paragraph_start(line): return re.match(r"^[a-z]\.", line.strip()) or line.strip().startswith("•")
def clean_text(text): return ' '.join(text.strip().split())

def extract_annotations(text):
    annotations = {}
    lines = text.strip().split('\n')
    current_num = None
    current_lines = []

    for line in lines:
        line = line.strip()

        if re.match(r"^\d+/\d+$", line):  # Stop on page number
            if current_num and current_lines:
                annotations[current_num] = clean_text(" ".join(current_lines))
                current_num = None
                current_lines = []
            continue

        match = re.match(r"^(\d+)\s+(.+)", line)
        if match:
            if current_num and current_lines:
                annotations[current_num] = clean_text(" ".join(current_lines))
            current_num = match.group(1)
            current_lines = [match.group(2)]
        elif current_num:
            current_lines.append(line)

    if current_num and current_lines:
        annotations[current_num] = clean_text(" ".join(current_lines))

    return annotations

def join_lines_with_hyphen(lines):
    result = ""
    for i, line in enumerate(lines):
        line = line.strip()
        if line.endswith("-"):
            result += line[:-1]  # on enlève le tiret, pas d’espace ajoutée
        else:
            result += line
            # ajouter un espace *seulement* s’il reste une ligne suivante
            if i < len(lines) - 1:
                result += " "
    return result

# Exceptions pour titres multi-lignes
title_exceptions = {
    "chapter": {"VI.": 3, "VII.": 1},
    "subchapter": {"VII.B.": 3},
    "section": {
        "IV.B.a)": 1,
        "IV.B.b)": 1,
    }
}

merge_page_paragraphs = {13+1}

wb = Workbook()
ws = wb.active
ws.title = "finma_framework"
ws.append(["assessable", "depth", "ref_id", "name", "description", "annotation"])

context = {"chapter": None, "subchapter": None, "section": None}
current_depth = 0
in_chapter_IV = False
paragraph_counter = 1
previous_paragraph_row = None

pdf_path = "finma_rs_2023_01_20221207_de.pdf"
doc = fitz.open(pdf_path)
full_text = "\n".join(page.get_text() for page in doc)
annotations = extract_annotations(full_text)

for i in range(2, 15):
    page = doc.load_page(i)
    blocks = page.get_text("blocks")
    blocks.sort(key=lambda b: (round(b[1]), round(b[0])))

    lines = []
    for b in blocks:
        text = b[4]
        if text.strip():
            lines.extend(text.strip().split('\n'))

    j = 0
    paragraph_lines = []
    is_merge_page = i in merge_page_paragraphs
    first_paragraph_handled = False

    while j < len(lines):
        line = clean_text(lines[j])
        if not line or re.match(r"^\d+/\d+$", line):
            j += 1
            continue

        # Chapitre
        if is_chapter(line):
            title_lines = [line]
            key = line
            num_lines = title_exceptions["chapter"].get(key, 2)
            for _ in range(num_lines - 1):
                j += 1
                if j < len(lines): title_lines.append(clean_text(lines[j]))
            full_title = join_lines_with_hyphen(title_lines)
            context = {"chapter": full_title, "subchapter": None, "section": None}
            current_depth = 1
            in_chapter_IV = full_title.startswith("IV.")
            ws.append(["", current_depth, "", full_title, "", ""])
            j += 1
            continue

        elif is_subchapter(line):
            title_lines = [line]
            parent_key = context["chapter"].split()[0] + "." + line
            num_lines = title_exceptions["subchapter"].get(parent_key, 2)
            for _ in range(num_lines - 1):
                j += 1
                if j < len(lines): title_lines.append(clean_text(lines[j]))
            full_title = join_lines_with_hyphen(title_lines)
            context["subchapter"] = full_title
            context["section"] = None
            current_depth = 2
            ws.append(["", current_depth, "", full_title, "", ""])
            j += 1
            continue

        elif is_section(line):
            title_lines = [line]
            key = line
            num_lines = title_exceptions["section"].get(key, 2)
            for _ in range(num_lines - 1):
                j += 1
                if j < len(lines): title_lines.append(clean_text(lines[j]))
            full_title = join_lines_with_hyphen(title_lines)
            context["section"] = full_title
            current_depth = 3
            ws.append(["", current_depth, "", full_title, "", ""])
            j += 1
            continue

        # Paragraphe spécial (puce ou a.)
        if is_paragraph_start(line):
            para_lines = [line]
            j += 1
            while j < len(lines):
                next_line = clean_text(lines[j])
                if (
                    is_chapter(next_line)
                    or is_subchapter(next_line)
                    or is_section(next_line)
                    or is_paragraph_start(next_line)
                    or not next_line
                ):
                    break
                para_lines.append(next_line)
                j += 1

            full_paragraph = join_lines_with_hyphen(para_lines)
            assessable = "x" if in_chapter_IV else ""
            ref_id = f"{REF_ID_PREFIX} {paragraph_counter}"

            annotation_set = set()
            for word, num in re.findall(r"([^\W\d_]+)(\d+)(?!\w)", full_paragraph, re.UNICODE):
                if num in annotations:
                    annotation_set.add(f"[{num}] {annotations[num]}")
            annotation_str = "\n".join(sorted(annotation_set))

            ws.append([assessable, current_depth + 1, "", ref_id, full_paragraph, annotation_str])
            previous_paragraph_row = ws.max_row
            paragraph_counter += 1
            continue

        # Paragraphe normal
        paragraph_lines.append(line)
        if line.endswith((".", "!", "?", ";")):
            full_paragraph = join_lines_with_hyphen(paragraph_lines)
            assessable = "x" if in_chapter_IV else ""
            ref_id = f"{REF_ID_PREFIX} {paragraph_counter}"

            annotation_set = set()
            for word, num in re.findall(r"([^\W\d_]+)(\d+)(?!\w)", full_paragraph, re.UNICODE):
                if num in annotations:
                    annotation_set.add(f"[{num}] {annotations[num]}")
            annotation_str = "\n".join(sorted(annotation_set))

            if is_merge_page and not first_paragraph_handled and previous_paragraph_row:
                previous_cell = ws.cell(row=previous_paragraph_row, column=5)
                previous_cell.value = previous_cell.value.strip() + " " + full_paragraph
                if annotation_str:
                    ann_cell = ws.cell(row=previous_paragraph_row, column=6)
                    existing = ann_cell.value.strip().split("\n") if ann_cell.value else []
                    for new_ann in sorted(annotation_set):
                        if new_ann not in existing:
                            existing.append(new_ann)
                    ann_cell.value = "\n".join(existing)
            else:
                ws.append([assessable, current_depth + 1, "", ref_id, full_paragraph, annotation_str])
                previous_paragraph_row = ws.max_row
                paragraph_counter += 1

            paragraph_lines = []
            first_paragraph_handled = True
        j += 1

# Dernier paragraphe
if paragraph_lines:
    full_paragraph = join_lines_with_hyphen(paragraph_lines)
    assessable = "x" if in_chapter_IV else ""
    ref_id = f"{REF_ID_PREFIX} {paragraph_counter}"
    annotation_set = set()
    for word, num in re.findall(r"([^\W\d_]+)(\d+)(?!\w)", full_paragraph, re.UNICODE):
        if num in annotations:
            annotation_set.add(f"[{num}] {annotations[num]}")
    annotation_str = "\n".join(sorted(annotation_set))
    ws.append([assessable, current_depth + 1, "", ref_id, full_paragraph, annotation_str])

wb.save("finma_framework_final_DE.xlsx")
