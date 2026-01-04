import openpyxl
import re
from openpyxl import Workbook
from openpyxl.styles import NamedStyle

# Cleans a string: lowercase, alphanumeric only, replaces other characters with "_"
def clean_string(s):
    return re.sub(r'[^a-zA-Z0-9]', '_', s.strip().lower()) if s else ""

# Forces text format for all cells in a worksheet
def force_text_format(ws):
    text_style = NamedStyle(name="text")
    text_style.number_format = "@"
    for col in ws.iter_cols():
        for cell in col:
            cell.style = text_style

# Reads a cell only if it's not part of a merged cell
def read_merged_cell(sheet, row, col):
    for merged_range in sheet.merged_cells.ranges:
        if sheet.cell(row=row, column=col).coordinate in merged_range:
            return None
    return sheet.cell(row=row, column=col).value

# Main transformation function
def transform_excel(source_path, target_path):
    # Load the source workbook
    wb_src = openpyxl.load_workbook(source_path, data_only=True)
    sheet_rfi = wb_src["Control Narrative RFI"]
    sheet_ccf = wb_src["CCF V3"]

    # Create the target workbook and required sheets
    wb_tgt = Workbook()
    ws_main = wb_tgt.active
    ws_main.title = "cisco_ccf_framework"
    ws_impl = wb_tgt.create_sheet("impl_groups")

    # Headers for the target sheets
    headers_main = ["assessable", "depth", "ref_id", "name", "description", "annotation", "typical_evidence", "implementation_groups"]
    headers_impl = ["ref_id", "name", "description"]
    ws_main.append(headers_main)
    ws_impl.append(headers_impl)

    # Track known domain titles and control types
    seen_domains = set()
    seen_control_types = {}

    # Get column indices based on the header row (row 3)
    rfi_header = [cell.value for cell in sheet_rfi[3]]
    idx_domain = rfi_header.index("Domain Title") + 1
    idx_ref = rfi_header.index("Control Reference") + 1
    idx_wording = rfi_header.index("Control Wording") + 1
    idx_artifact = rfi_header.index("Control Supporting Audit Artifacts/ RFI") + 1
    idx_narrative = rfi_header.index("Control Narrative") + 1

    # Process the "Control Narrative RFI" sheet
    row = 4
    domain_depth = 1
    control_depth = 2

    while row <= sheet_rfi.max_row:
        # Handle the special "IRAP Unique Controls" merged row
        cell_val = sheet_rfi.cell(row=row, column=1).value
        if cell_val and isinstance(cell_val, str) and "irap unique controls" in cell_val.lower():
            cleaned = clean_string(cell_val)
            ws_main.append(["", "1", "", cell_val.strip(), "", "", "", ""])
            seen_domains = set()     # Reset known domains
            domain_depth = 2         # Adjust depth for new section
            control_depth = 3
            row += 1
            continue

        # Add new domain title if not already seen
        domain = read_merged_cell(sheet_rfi, row, idx_domain)
        if domain and domain not in seen_domains:
            seen_domains.add(domain)
            suffix = "2" if domain_depth == 2 else ""
            ws_main.append([
                "", str(domain_depth), "", domain, "", "", "", ""
            ])

        # Add control reference row if present
        ref_id = read_merged_cell(sheet_rfi, row, idx_ref)
        if ref_id:
            wording = read_merged_cell(sheet_rfi, row, idx_wording) or ""
            narrative = read_merged_cell(sheet_rfi, row, idx_narrative) or ""
            artifact = read_merged_cell(sheet_rfi, row, idx_artifact) or ""
            ws_main.append([
                "x", str(control_depth), ref_id, "", wording, narrative, artifact, ""
            ])

        row += 1

    # Get column indices for "CCF V3" sheet (header row = row 5)
    ccf_header = [cell.value for cell in sheet_ccf[5]]
    idx_ccf_ref = ccf_header.index("Control Reference") + 1
    idx_ccf_title = ccf_header.index("Control Title") + 1
    idx_ccf_type = ccf_header.index("Control Type") + 1

    # Process "CCF V3" sheet starting at row 6
    row = 6
    while row <= sheet_ccf.max_row:
        ref_id = read_merged_cell(sheet_ccf, row, idx_ccf_ref)
        if not ref_id:
            row += 1
            continue

        title = read_merged_cell(sheet_ccf, row, idx_ccf_title) or ""
        ctrl_type = read_merged_cell(sheet_ccf, row, idx_ccf_type) or ""
        impl_group = clean_string(ctrl_type)

        # Update existing row in main sheet with matching ref_id
        for row_tgt in ws_main.iter_rows(min_row=2, values_only=False):
            if row_tgt[2].value == ref_id:
                row_tgt[3].value = title
                row_tgt[7].value = impl_group
                break

        # Add new implementation group if not already seen
        if ctrl_type and ctrl_type not in seen_control_types:
            seen_control_types[ctrl_type] = impl_group
            ws_impl.append([impl_group, ctrl_type, ""])

        row += 1

    # Apply text format to all target sheets
    force_text_format(ws_main)
    force_text_format(ws_impl)

    # Save the transformed workbook
    wb_tgt.save(target_path)
    
    # Notify the user
    print(f"✅ Framework successfully converted and exported to \"{target_path}\"")

# Example usage
transform_excel("Cisco-CCFv3-Public.xlsx", "conv_Cisco-CCFv3-Public.xlsx")
