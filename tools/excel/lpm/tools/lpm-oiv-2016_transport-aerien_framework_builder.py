#!/usr/bin/env python3
"""Build the full LPM OIV 2016 Transport aerien workbook from Legifrance."""


import argparse
import sys
from copy import copy
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook, load_workbook

from legifrance_web_scraper import (
    DEFAULT_SHEET_NAME as CONTENT_SHEET_NAME,
    main as build_annex_workbook,
)

# Import Excel to YAML converter with the same pattern used by other framework builders
REPO_ROOT = Path(__file__).resolve().parents[4]
BACKEND_SCRIPTS_DIR = REPO_ROOT / "backend" / "scripts"
sys.path.insert(0, str(BACKEND_SCRIPTS_DIR))
from convert_library_v2 import create_library as convert_excel_to_yaml


SOURCE_URL = "https://www.legifrance.gouv.fr/jorf/id/JORFTEXT000033063127"
FRAMEWORK_SLUG = "lpm-oiv-2016_transport-aerien"
FRAMEWORK_REF_ID = "LPM-OIV-2016_Transport-aerien"
FRAMEWORK_NAME = "Règles OIV - Secteur « Transport aérien » (2016)"
FRAMEWORK_DESCRIPTION = """
RÈGLES DE SÉCURITÉ RELATIVES AU SOUS-SECTEUR D'ACTIVITÉS D'IMPORTANCE VITALE « TRANSPORT AÉRIEN »

Arrêté du 11 août 2016 fixant les règles de sécurité et les modalités de déclaration des systèmes d'information d'importance vitale et des incidents de sécurité relatives au sous-secteur d'activités d'importance vitale « Transport aérien » et pris en application des articles R. 1332-41-1, R. 1332-41-2 et R. 1332-41-10 du code de la défense.

Source : https://www.legifrance.gouv.fr/jorf/id/JORFTEXT000033063127
""".strip()
DEFAULT_OUTPUT_PATH = (
    Path(__file__).resolve().parent.parent / "lpm-oiv-2016_transport-aerien.xlsx"
)
DEFAULT_YAML_PATH = (
    Path(__file__).resolve().parent.parent / "lpm-oiv-2016_transport-aerien.yaml"
)

LIBRARY_META_ROWS = [
    ("type", "library"),
    ("urn", f"urn:intuitem:risk:library:{FRAMEWORK_SLUG}"),
    ("version", "1"),
    ("locale", "fr"),
    ("ref_id", FRAMEWORK_REF_ID),
    ("name", FRAMEWORK_NAME),
    ("description", FRAMEWORK_DESCRIPTION),
    ("copyright", "Loi française"),
    ("provider", "Gouvernement français"),
    ("packager", "intuitem"),
]

FRAMEWORK_META_ROWS = [
    ("type", "framework"),
    ("base_urn", f"urn:intuitem:risk:req_node:{FRAMEWORK_SLUG}"),
    ("urn", f"urn:intuitem:risk:framework:{FRAMEWORK_SLUG}"),
    ("ref_id", FRAMEWORK_REF_ID),
    ("name", FRAMEWORK_NAME),
    ("description", FRAMEWORK_DESCRIPTION),
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Build the full LPM OIV 2016 Transport aerien workbook "
            "from the Legifrance Annex page."
        )
    )
    parser.add_argument(
        "--url",
        default=SOURCE_URL,
        help="Legifrance URL to scrape.",
    )
    parser.add_argument(
        "--output",
        default=str(DEFAULT_OUTPUT_PATH),
        help="Output workbook path.",
    )
    parser.add_argument(
        "--yaml-output",
        default=str(DEFAULT_YAML_PATH),
        help="Output YAML path.",
    )
    return parser.parse_args()


def print_step_banner(step_number: int, title: str) -> None:
    # Print one readable step banner with spacing around it.
    message = f"##### [STEP {step_number}] {title} #####"
    line = "#" * len(message)
    print(f"\n{line}")
    print(message)
    print(f"{line}\n")


def write_meta_sheet(worksheet, rows: list[tuple[str, str]]) -> None:
    # Write key/value metadata rows in the target worksheet
    for key, value in rows:
        worksheet.append((key, value))


def copy_sheet_contents(source_sheet, target_sheet) -> None:
    # Copy cell values, simple formatting and sheet layout from one workbook to another
    for row in source_sheet.iter_rows():
        for cell in row:
            target_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                target_cell.font = copy(cell.font)
                target_cell.fill = copy(cell.fill)
                target_cell.border = copy(cell.border)
                target_cell.alignment = copy(cell.alignment)
                target_cell.number_format = cell.number_format
                target_cell.protection = copy(cell.protection)

    for column_key, dimension in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[column_key].width = dimension.width

    target_sheet.freeze_panes = source_sheet.freeze_panes


def convert_workbook_to_yaml(excel_path: str | Path, yaml_path: str | Path) -> None:
    # Convert the generated Excel workbook into an YAML library
    resolved_excel_path = Path(excel_path).expanduser()
    resolved_yaml_path = Path(yaml_path).expanduser()
    resolved_yaml_path.parent.mkdir(parents=True, exist_ok=True)

    print_step_banner(3, "Convert Excel To YAML")
    print(f"Source Excel : \"{resolved_excel_path}\"")
    print(f"Target YAML  : \"{resolved_yaml_path}\"\n")
    print("⌛ Converting Excel workbook to YAML...")
    convert_excel_to_yaml(
        input_file=str(resolved_excel_path),
        output_file=str(resolved_yaml_path),
    )
    print(f"✅ YAML conversion completed: \"{resolved_yaml_path}\"")


def build_workbook(url: str, output_path: str | Path) -> Path:
    # Build the final workbook by combining meta sheets with the Annex content sheet
    resolved_output_path = Path(output_path).expanduser()

    with TemporaryDirectory() as temp_dir:
        temp_annex_path = Path(temp_dir) / "annex_only.xlsx"

        print_step_banner(1, "Extract Annex Content")
        print(f"Source URL          : \"{url}\"")
        print(f"Temporary workbook  : \"{temp_annex_path}\"\n")
        print("⌛ Building Annex workbook...")
        build_annex_workbook(url=url, output_path=temp_annex_path)

        print_step_banner(2, "Build Framework Workbook")
        print(f"Source Annex Excel  : \"{temp_annex_path}\"")
        print(f"Target workbook     : \"{resolved_output_path}\"\n")
        print("⌛ Building final framework workbook...")
        annex_workbook = load_workbook(temp_annex_path)
        annex_sheet = annex_workbook[CONTENT_SHEET_NAME]

        workbook = Workbook()
        library_sheet = workbook.active
        library_sheet.title = "library_meta"
        framework_sheet = workbook.create_sheet("oiv_meta")
        content_sheet = workbook.create_sheet(CONTENT_SHEET_NAME)

        write_meta_sheet(library_sheet, LIBRARY_META_ROWS)
        write_meta_sheet(framework_sheet, FRAMEWORK_META_ROWS)
        copy_sheet_contents(annex_sheet, content_sheet)

        resolved_output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(resolved_output_path)

    print(f"✅ Final framework workbook written to \"{resolved_output_path}\"")
    return resolved_output_path


def main(
    url: str = SOURCE_URL,
    output_path: str | Path = DEFAULT_OUTPUT_PATH,
    yaml_output_path: str | Path = DEFAULT_YAML_PATH,
) -> None:
    workbook_path = build_workbook(url, output_path)
    convert_workbook_to_yaml(workbook_path, yaml_output_path)
    print("\n✅ Framework build pipeline completed.")


if __name__ == "__main__":
    args = parse_args()
    main(url=args.url, output_path=args.output, yaml_output_path=args.yaml_output)
