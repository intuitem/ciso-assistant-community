#!/usr/bin/env python3
"""Repair confirmed CyFun 2025 review issues in a copied workbook.

The input workbook is never modified. The Dutch function names and the French
PR.AT-01.1 and PR.PS-01.1 annotations are checked against the official
localized PDFs. Only cells whose value differs from those sources are replaced
and highlighted in opaque yellow.
"""

from __future__ import annotations

import argparse
from copy import copy
from pathlib import Path
import re
import sys

import fitz
import openpyxl
from openpyxl.styles import PatternFill

import cyfun2025_framework_FR as french
import cyfun2025_framework_NL as base


SCRIPT_DIR = Path(__file__).resolve().parent
DATA_DIR = SCRIPT_DIR.parent
DEFAULT_DUTCH_PDF = DATA_DIR / "CyFun2025_Booklet-ESSENTIAL_N.pdf"
DEFAULT_FRENCH_PDF = DATA_DIR / "CyFun2025_Booklet-ESSENTIAL_F_pr3.pdf"
TARGET_REFS = ("PR.AT-01.1", "PR.PS-01.1")
CONTENT_SHEETS = ("requirements_content", "controls_content")
YELLOW_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFFFFF00",
    bgColor="FFFFFF00",
)


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Copy a CyFun 2025 workbook, repair confirmed Dutch function-name "
            "and French PR.AT-01.1/PR.PS-01.1 issues, and highlight changed "
            "cells."
        )
    )
    parser.add_argument("--input", type=Path, required=True)
    parser.add_argument(
        "--output",
        type=Path,
        help="Output workbook. Defaults to <input_stem>_review_fixed.xlsx.",
    )
    parser.add_argument("--dutch-pdf", type=Path, default=DEFAULT_DUTCH_PDF)
    parser.add_argument("--french-pdf", type=Path, default=DEFAULT_FRENCH_PDF)
    parser.add_argument(
        "--force",
        action="store_true",
        help="Allow replacing an existing output workbook.",
    )
    args = parser.parse_args()
    if args.output is None:
        args.output = args.input.with_name(
            f"{args.input.stem}_review_fixed{args.input.suffix}"
        )
    return args


def header_map(sheet) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in sheet[1]
        if cell.value not in (None, "")
    }


def validate_dutch_function_names(pdf_path: Path) -> dict[str, str]:
    expected = dict(base.DUTCH_FUNCTION_NAMES)
    with fitz.open(pdf_path) as document:
        complete_text = "\n".join(page.get_text("text") for page in document).upper()
    missing = [name for name in expected.values() if name not in complete_text]
    if missing:
        raise ValueError(
            "Dutch function names not found in the official PDF: "
            + ", ".join(missing)
        )
    return expected


def extract_dutch_texts(
    workbook_path: Path, pdf_path: Path
) -> base.ExtractedTexts:
    rows = base.load_workbook_rows(workbook_path)
    expected_functions = {
        row.ref_id
        for row in rows
        if row.depth == 1 and base.FUNCTION_RE.match(row.ref_id)
    }
    expected_categories = {
        row.ref_id
        for row in rows
        if row.depth == 2 and base.CATEGORY_RE.match(row.ref_id)
    }
    expected_body_refs = [row.ref_id for row in rows if row.depth in (3, 4)]
    extracted = base.extract_pdf(
        pdf_path,
        expected_functions,
        expected_categories,
        expected_body_refs,
    )
    base.validate_dutch_coverage(rows, extracted)
    return extracted


def extract_french_target_annotations(
    workbook_path: Path, pdf_path: Path
) -> dict[str, str]:
    french.configure_french_extraction()
    rows = base.load_workbook_rows(workbook_path)
    expected_functions = {
        row.ref_id
        for row in rows
        if row.depth == 1 and base.FUNCTION_RE.match(row.ref_id)
    }
    expected_categories = {
        row.ref_id
        for row in rows
        if row.depth == 2 and base.CATEGORY_RE.match(row.ref_id)
    }
    expected_body_refs = [row.ref_id for row in rows if row.depth in (3, 4)]
    extracted = base.extract_pdf(
        pdf_path,
        expected_functions,
        expected_categories,
        expected_body_refs,
    )
    missing = [ref_id for ref_id in TARGET_REFS if ref_id not in extracted.annotations]
    if missing:
        raise ValueError(
            "Could not extract complete French annotations for: "
            + ", ".join(missing)
        )

    annotations = {
        ref_id: extracted.annotations[ref_id] for ref_id in TARGET_REFS
    }
    pr_at_bullets = len(re.findall(r"(?m)^-\s+", annotations["PR.AT-01.1"]))
    if pr_at_bullets != 8:
        raise ValueError(
            f"Expected 8 French PR.AT-01.1 bullets, got {pr_at_bullets}."
        )
    pr_ps = annotations["PR.PS-01.1"]
    required_pr_ps_texts = (
        "Les systèmes pourraient être surveillés en continu",
        "Mettre à jour les bases de référence",
        "Systèmes gérés par le cloud et les fournisseurs",
    )
    missing_pr_ps_texts = [text for text in required_pr_ps_texts if text not in pr_ps]
    if missing_pr_ps_texts:
        raise ValueError(
            "Incomplete French PR.PS-01.1 annotation; missing: "
            + ", ".join(missing_pr_ps_texts)
        )
    return annotations


def repair_workbook(
    workbook,
    dutch_names: dict[str, str],
    dutch_texts: base.ExtractedTexts,
    french_annotations: dict[str, str],
) -> set[tuple[str, str]]:
    corrected: set[tuple[str, str]] = set()

    requirements = workbook["requirements_content"]
    headers = header_map(requirements)
    required_headers = {
        "depth",
        "ref_id",
        "name[nl]",
        "description[nl]",
        "annotation[nl]",
        "annotation[fr]",
    }
    missing_headers = required_headers - headers.keys()
    if missing_headers:
        raise ValueError(
            "Missing requirements_content headers: "
            + ", ".join(sorted(missing_headers))
        )

    found_functions: set[str] = set()
    found_targets: set[str] = set()
    for row_number in range(2, requirements.max_row + 1):
        ref_id = requirements.cell(row_number, headers["ref_id"]).value
        depth = requirements.cell(row_number, headers["depth"]).value
        for field_name, expected_values in (
            ("name[nl]", dutch_texts.names),
            ("description[nl]", dutch_texts.descriptions),
            ("annotation[nl]", dutch_texts.annotations),
        ):
            if ref_id not in expected_values:
                continue
            cell = requirements.cell(row_number, headers[field_name])
            if cell.value != expected_values[ref_id]:
                cell.value = expected_values[ref_id]
                cell.fill = copy(YELLOW_FILL)
                corrected.add((requirements.title, cell.coordinate))
        if depth == 1 and ref_id in dutch_names:
            if ref_id in found_functions:
                raise ValueError(f"Duplicate depth-1 function row: {ref_id}.")
            found_functions.add(ref_id)
            cell = requirements.cell(row_number, headers["name[nl]"])
            if cell.value != dutch_names[ref_id]:
                cell.value = dutch_names[ref_id]
                cell.fill = copy(YELLOW_FILL)
                corrected.add((requirements.title, cell.coordinate))
        if ref_id in french_annotations:
            found_targets.add(ref_id)
            cell = requirements.cell(row_number, headers["annotation[fr]"])
            if cell.value != french_annotations[ref_id]:
                cell.value = french_annotations[ref_id]
                cell.fill = copy(YELLOW_FILL)
                corrected.add((requirements.title, cell.coordinate))

    if found_functions != dutch_names.keys():
        raise ValueError(
            "Missing depth-1 Dutch function rows: "
            + ", ".join(sorted(dutch_names.keys() - found_functions))
        )
    if found_targets != set(TARGET_REFS):
        raise ValueError(
            "Missing target refs in requirements_content: "
            + ", ".join(sorted(set(TARGET_REFS) - found_targets))
        )

    controls = workbook["controls_content"]
    control_headers = header_map(controls)
    for required in (
        "ref_id",
        "description[nl]",
        "annotation[nl]",
        "annotation[fr]",
    ):
        if required not in control_headers:
            raise ValueError(f"Missing controls_content header: {required}.")
    for row_number in range(2, controls.max_row + 1):
        ref_id = controls.cell(row_number, control_headers["ref_id"]).value
        for field_name, expected_values in (
            ("description[nl]", dutch_texts.descriptions),
            ("annotation[nl]", dutch_texts.annotations),
        ):
            if ref_id not in expected_values:
                continue
            cell = controls.cell(row_number, control_headers[field_name])
            if cell.value != expected_values[ref_id]:
                cell.value = expected_values[ref_id]
                cell.fill = copy(YELLOW_FILL)
                corrected.add((controls.title, cell.coordinate))
    for ref_id, annotation in french_annotations.items():
        control_target_rows = [
            row_number
            for row_number in range(2, controls.max_row + 1)
            if controls.cell(row_number, control_headers["ref_id"]).value == ref_id
        ]
        if len(control_target_rows) != 1:
            raise ValueError(
                f"Expected one {ref_id} row in controls_content, "
                f"got {len(control_target_rows)}."
            )
        control_cell = controls.cell(
            control_target_rows[0], control_headers["annotation[fr]"]
        )
        if control_cell.value != annotation:
            control_cell.value = annotation
            control_cell.fill = copy(YELLOW_FILL)
            corrected.add((controls.title, control_cell.coordinate))

    return corrected


def verify_output(
    input_path: Path,
    output_path: Path,
    corrected_cells: set[tuple[str, str]],
    dutch_names: dict[str, str],
    dutch_texts: base.ExtractedTexts,
    french_annotations: dict[str, str],
) -> None:
    source = openpyxl.load_workbook(input_path, data_only=False)
    output = openpyxl.load_workbook(output_path, data_only=False)
    try:
        if source.sheetnames != output.sheetnames:
            raise ValueError("The output workbook changed the sheet list.")

        actual_differences: set[tuple[str, str]] = set()
        formula_changes = []
        unexpected_style_changes = []
        for sheet_name in source.sheetnames:
            left = source[sheet_name]
            right = output[sheet_name]
            if left.max_row != right.max_row or left.max_column != right.max_column:
                raise ValueError(f"Sheet dimensions changed: {sheet_name}.")
            if list(left.merged_cells.ranges) != list(right.merged_cells.ranges):
                raise ValueError(f"Merged cells changed: {sheet_name}.")
            for row in left.iter_rows():
                for source_cell in row:
                    output_cell = right[source_cell.coordinate]
                    key = (sheet_name, source_cell.coordinate)
                    if source_cell.value != output_cell.value:
                        actual_differences.add(key)
                    if (
                        isinstance(source_cell.value, str)
                        and source_cell.value.startswith("=")
                        and source_cell.value != output_cell.value
                    ):
                        formula_changes.append(f"{sheet_name}!{source_cell.coordinate}")
                    if key in corrected_cells:
                        left_style = copy(source_cell._style)
                        right_style = copy(output_cell._style)
                        left_style.fillId = right_style.fillId
                        if left_style != right_style:
                            unexpected_style_changes.append(
                                f"{sheet_name}!{source_cell.coordinate}"
                            )
                        if (
                            output_cell.fill.fill_type != "solid"
                            or output_cell.fill.fgColor.rgb != "FFFFFF00"
                        ):
                            raise ValueError(
                                "Corrected cell is not opaque yellow: "
                                f"{sheet_name}!{source_cell.coordinate}."
                            )
                    elif source_cell._style != output_cell._style:
                        unexpected_style_changes.append(
                            f"{sheet_name}!{source_cell.coordinate}"
                        )

        if actual_differences != corrected_cells:
            raise ValueError(
                "Unexpected value differences: "
                + ", ".join(
                    f"{sheet}!{cell}"
                    for sheet, cell in sorted(actual_differences ^ corrected_cells)
                )
            )
        if formula_changes:
            raise ValueError("Formula changes: " + ", ".join(formula_changes))
        if unexpected_style_changes:
            raise ValueError(
                "Unexpected style changes: "
                + ", ".join(unexpected_style_changes[:20])
            )

        requirements = output["requirements_content"]
        headers = header_map(requirements)
        actual_names = {
            requirements.cell(row, headers["ref_id"]).value: requirements.cell(
                row, headers["name[nl]"]
            ).value
            for row in range(2, requirements.max_row + 1)
            if requirements.cell(row, headers["depth"]).value == 1
        }
        if actual_names != dutch_names:
            raise ValueError(f"Unexpected Dutch function names: {actual_names}.")

        for sheet_name in CONTENT_SHEETS:
            sheet = output[sheet_name]
            headers = header_map(sheet)
            expected_fields = (
                (("name[nl]", dutch_texts.names),)
                if sheet_name == "requirements_content"
                else ()
            ) + (
                ("description[nl]", dutch_texts.descriptions),
                ("annotation[nl]", dutch_texts.annotations),
            )
            for field_name, expected_values in expected_fields:
                for row in range(2, sheet.max_row + 1):
                    ref_id = sheet.cell(row, headers["ref_id"]).value
                    if ref_id not in expected_values:
                        continue
                    if sheet.cell(row, headers[field_name]).value != expected_values[ref_id]:
                        raise ValueError(
                            f"Unexpected Dutch {field_name} for {ref_id} "
                            f"in {sheet_name}."
                        )

        for sheet_name in CONTENT_SHEETS:
            sheet = output[sheet_name]
            headers = header_map(sheet)
            for ref_id, annotation in french_annotations.items():
                target_values = [
                    sheet.cell(row, headers["annotation[fr]"]).value
                    for row in range(2, sheet.max_row + 1)
                    if sheet.cell(row, headers["ref_id"]).value == ref_id
                ]
                if target_values != [annotation]:
                    raise ValueError(
                        f"Unexpected French {ref_id} annotation in {sheet_name}."
                    )

        error_values = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"}
        formula_errors = [
            f"{sheet.title}!{cell.coordinate}"
            for sheet in output.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if cell.data_type == "e" or cell.value in error_values
        ]
        if formula_errors:
            raise ValueError(
                "Formula errors found: " + ", ".join(formula_errors[:20])
            )
    finally:
        output.close()
        source.close()


def main() -> int:
    args = parse_arguments()
    for path, label in (
        (args.input, "workbook"),
        (args.dutch_pdf, "Dutch PDF"),
        (args.french_pdf, "French PDF"),
    ):
        if not path.is_file():
            raise FileNotFoundError(f"{label} not found: {path}")
    if args.output.resolve() == args.input.resolve():
        raise ValueError("The output path must differ from the input path.")
    if args.output.exists() and not args.force:
        raise FileExistsError(
            f"Output already exists: {args.output}. Use --force to replace it."
        )

    dutch_names = validate_dutch_function_names(args.dutch_pdf)
    dutch_texts = extract_dutch_texts(args.input, args.dutch_pdf)
    french_annotations = extract_french_target_annotations(
        args.input, args.french_pdf
    )

    workbook = openpyxl.load_workbook(args.input)
    try:
        corrected_cells = repair_workbook(
            workbook, dutch_names, dutch_texts, french_annotations
        )
        args.output.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(args.output)
    finally:
        workbook.close()

    verify_output(
        args.input,
        args.output,
        corrected_cells,
        dutch_names,
        dutch_texts,
        french_annotations,
    )
    print(f"Verification passed: {len(corrected_cells)} corrected cells.")
    for sheet_name, coordinate in sorted(corrected_cells):
        print(f"Corrected and highlighted: {sheet_name}!{coordinate}")
    print(f"Created: {args.output.resolve()}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (FileNotFoundError, FileExistsError, KeyError, ValueError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1) from exc
