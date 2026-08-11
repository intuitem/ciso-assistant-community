"""Generate the official CyFun 2025 mapping to ISO 27001/27002.

The mapping is read from columns E, I, O, U and V of the official CCB
workbook. Before updating the destination workbook, the script creates a
timestamped backup next to it.
"""

from __future__ import annotations

import argparse
import os
import re
import shutil
import tempfile
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Font, PatternFill


SCRIPT_DIR = Path(__file__).resolve().parent
CCB_DIR = SCRIPT_DIR.parent
DEFAULT_SOURCE = CCB_DIR / "CyFun2025_mapping_2026-05-WIP.xlsx"
DEFAULT_TARGET = (
    CCB_DIR
    / "mappings"
    / "mapping-ccb-cyfun2025-and-iso27001-2022.xlsx"
)

SOURCE_SHEET = "CyFun® 2025"
MAPPING_SHEET = "mappings_content"
SOURCE_REFERENCE_SHEET = "source"
TARGET_REFERENCE_SHEET = "target"

CYFUN_COLUMNS = (5, 9, 15)  # E, I and O
ISO_27001_COLUMN = 21  # U
ISO_27002_COLUMN = 22  # V
MAPPING_HEADERS = (
    "source_node_id",
    "target_node_id",
    "relationship",
    "rationale",
    "strength_of_relationship",
)

ERROR_FILL = PatternFill(fill_type="solid", fgColor="8B0000")
ERROR_FONT = Font(color="FFFFFF", bold=True)
CORRECTION_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
EXPANSION_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")

SOURCE_NODE_ID_CORRECTIONS = {
    "de.cm-03.2": "de.cm-03-2",
    "gv.rr-03.1": "gv.rr-03-1",
    "gv.rr-03.2": "gv.rr-03-2",
    "id.am-03.2": "id.am-03-2",
    "id.am-03.3": "id.am-03-3",
    "id.am-08.10": "d.am-08.10",
}


@dataclass
class MappingRow:
    source_node_id: str
    target_node_id: str
    source_fill: PatternFill
    origin: str
    source_row: int
    source_corrected: bool = False
    target_expanded: bool = False
    source_invalid: bool = False
    target_invalid: bool = False

    @property
    def invalid(self) -> bool:
        return self.source_invalid or self.target_invalid


def clean_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def normalize_id(value: object) -> str:
    return clean_text(value).lower()


def create_backup(path: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S-%f")
    backup = path.with_name(f"{path.stem}.backup-{timestamp}{path.suffix}")
    shutil.copy2(path, backup)
    return backup


def read_reference_ids(sheet) -> set[str]:
    return {
        normalize_id(row[0])
        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True)
        if row[0] is not None
    }


def parse_iso27001_targets(value: object) -> list[str]:
    targets: list[str] = []
    for line in str(value or "").splitlines():
        token = clean_text(line)
        if not token:
            continue
        token = re.sub(r"^clause\s+", "", token, flags=re.IGNORECASE)
        targets.append(token.lower())
    return targets


def parse_iso27002_targets(value: object) -> list[str]:
    targets: list[str] = []
    for line in str(value or "").splitlines():
        token = clean_text(line)
        if not token:
            continue
        token = re.sub(r"^con(?:t)?rol\s+", "", token, flags=re.IGNORECASE)
        token = token.replace(" ", "").lower()

        range_match = re.fullmatch(r"(\d+)\.(\d+)-(\d+)\.(\d+)", token)
        if range_match:
            start_group, start_number, end_group, end_number = map(
                int, range_match.groups()
            )
            if start_group == end_group and start_number <= end_number:
                targets.extend(
                    f"a.{start_group}.{number}"
                    for number in range(start_number, end_number + 1)
                )
                continue

        targets.append("a." + token)
    return targets


def collect_mappings(source_sheet) -> list[MappingRow]:
    mappings: dict[tuple[str, str], MappingRow] = {}

    for row_number in range(2, source_sheet.max_row + 1):
        clause_targets = parse_iso27001_targets(
            source_sheet.cell(row_number, ISO_27001_COLUMN).value
        )
        control_targets = parse_iso27002_targets(
            source_sheet.cell(row_number, ISO_27002_COLUMN).value
        )
        if not clause_targets and not control_targets:
            continue

        for column in CYFUN_COLUMNS:
            source_cell: Cell = source_sheet.cell(row_number, column)
            if source_cell.value is None:
                continue

            original_source_id = normalize_id(source_cell.value)
            source_id = SOURCE_NODE_ID_CORRECTIONS.get(
                original_source_id, original_source_id
            )
            source_corrected = source_id != original_source_id
            source_fill = copy(source_cell.fill)
            for target_id in clause_targets:
                key = (source_id, target_id)
                mappings.setdefault(
                    key,
                    MappingRow(
                        source_node_id=source_id,
                        target_node_id=target_id,
                        source_fill=source_fill,
                        origin="ISO 27001 column U",
                        source_row=row_number,
                        source_corrected=source_corrected,
                    ),
                )
            for target_id in control_targets:
                key = (source_id, target_id)
                mappings.setdefault(
                    key,
                    MappingRow(
                        source_node_id=source_id,
                        target_node_id=target_id,
                        source_fill=source_fill,
                        origin="ISO 27002 column V",
                        source_row=row_number,
                        source_corrected=source_corrected,
                    ),
                )

    return list(mappings.values())


def find_assessable_descendants(target_sheet) -> dict[str, list[str]]:
    assessable_targets: list[tuple[str, str]] = []
    for node_id, assessable, _urn, ref_id, *_ in target_sheet.iter_rows(
        min_row=2, values_only=True
    ):
        if node_id is None or ref_id is None or not assessable:
            continue
        assessable_targets.append((normalize_id(ref_id), normalize_id(node_id)))

    descendants: dict[str, list[str]] = {}
    for ref_id, node_id in assessable_targets:
        segments = ref_id.split(".")
        for segment_count in range(1, len(segments)):
            parent = ".".join(segments[:segment_count])
            descendants.setdefault(parent, []).append(node_id)
    return descendants


def add_parent_expansions(mappings: list[MappingRow], target_sheet) -> list[MappingRow]:
    descendants = find_assessable_descendants(target_sheet)
    original_keys = {
        (mapping.source_node_id, mapping.target_node_id) for mapping in mappings
    }
    emitted_keys = set(original_keys)
    result: list[MappingRow] = []

    for mapping in mappings:
        descendant_ids = descendants.get(mapping.target_node_id, [])
        if not descendant_ids:
            result.append(mapping)
            continue

        for descendant_id in descendant_ids:
            key = (mapping.source_node_id, descendant_id)
            if key in emitted_keys:
                continue
            emitted_keys.add(key)
            result.append(
                MappingRow(
                    source_node_id=mapping.source_node_id,
                    target_node_id=descendant_id,
                    source_fill=copy(mapping.source_fill),
                    origin=f"Expanded from {mapping.target_node_id}",
                    source_row=mapping.source_row,
                    source_corrected=mapping.source_corrected,
                    target_expanded=True,
                )
            )

    return result


def validate_mappings(
    mappings: list[MappingRow], source_ids: set[str], target_ids: set[str]
) -> None:
    for mapping in mappings:
        mapping.source_invalid = mapping.source_node_id not in source_ids
        mapping.target_invalid = mapping.target_node_id not in target_ids


def check_destination_workbook(workbook) -> None:
    required_sheets = {
        MAPPING_SHEET,
        SOURCE_REFERENCE_SHEET,
        TARGET_REFERENCE_SHEET,
    }
    missing = sorted(required_sheets - set(workbook.sheetnames))
    if missing:
        raise ValueError(f"Missing destination sheet(s): {', '.join(missing)}")

    mapping_sheet = workbook[MAPPING_SHEET]
    actual_headers = tuple(
        clean_text(mapping_sheet.cell(1, column).value or "")
        for column in range(1, len(MAPPING_HEADERS) + 1)
    )
    if actual_headers != MAPPING_HEADERS:
        raise ValueError(
            f"Unexpected headers in {MAPPING_SHEET}: {actual_headers!r}; "
            f"expected {MAPPING_HEADERS!r}"
        )


def write_mappings(mapping_sheet, mappings: list[MappingRow]) -> None:
    if mapping_sheet.max_row > 1:
        mapping_sheet.delete_rows(2, mapping_sheet.max_row - 1)

    for output_row, mapping in enumerate(mappings, start=2):
        source_cell = mapping_sheet.cell(output_row, 1, mapping.source_node_id)
        target_cell = mapping_sheet.cell(output_row, 2, mapping.target_node_id)
        mapping_sheet.cell(output_row, 3, "intersect")
        mapping_sheet.cell(output_row, 4, None)
        mapping_sheet.cell(output_row, 5, None)

        relationship_cell = mapping_sheet.cell(output_row, 3)
        relationship_cell.fill = copy(mapping.source_fill)

        source_cell.fill = copy(
            ERROR_FILL
            if mapping.source_invalid
            else EXPANSION_FILL
            if mapping.target_expanded
            else CORRECTION_FILL
            if mapping.source_corrected
            else mapping.source_fill
        )
        target_cell.fill = copy(
            ERROR_FILL
            if mapping.target_invalid
            else EXPANSION_FILL
            if mapping.target_expanded
            else mapping.source_fill
        )
        if mapping.source_invalid:
            source_cell.font = copy(ERROR_FONT)
        if mapping.target_invalid:
            target_cell.font = copy(ERROR_FONT)


def save_atomically(workbook, destination: Path) -> None:
    temporary_file = tempfile.NamedTemporaryFile(
        prefix=destination.stem + ".tmp-",
        suffix=destination.suffix,
        dir=destination.parent,
        delete=False,
    )
    temporary_path = Path(temporary_file.name)
    temporary_file.close()
    try:
        workbook.save(temporary_path)
        os.replace(temporary_path, destination)
    finally:
        if temporary_path.exists():
            temporary_path.unlink()


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--target", type=Path, default=DEFAULT_TARGET)
    args = parser.parse_args()

    source_path = args.source.resolve()
    target_path = args.target.resolve()
    if not source_path.is_file():
        raise FileNotFoundError(f"Source workbook not found: {source_path}")
    if not target_path.is_file():
        raise FileNotFoundError(f"Destination workbook not found: {target_path}")

    source_workbook = load_workbook(source_path, data_only=False)
    if SOURCE_SHEET not in source_workbook.sheetnames:
        raise ValueError(f"Missing source sheet: {SOURCE_SHEET}")

    destination_workbook = load_workbook(target_path, data_only=False)
    check_destination_workbook(destination_workbook)

    mappings = collect_mappings(source_workbook[SOURCE_SHEET])
    mappings = add_parent_expansions(
        mappings, destination_workbook[TARGET_REFERENCE_SHEET]
    )
    source_ids = read_reference_ids(destination_workbook[SOURCE_REFERENCE_SHEET])
    target_ids = read_reference_ids(destination_workbook[TARGET_REFERENCE_SHEET])
    validate_mappings(mappings, source_ids, target_ids)

    backup_path = create_backup(target_path)
    write_mappings(destination_workbook[MAPPING_SHEET], mappings)
    save_atomically(destination_workbook, target_path)
    source_workbook.close()
    destination_workbook.close()

    invalid_mappings = [mapping for mapping in mappings if mapping.invalid]
    corrected_mappings = [mapping for mapping in mappings if mapping.source_corrected]
    expanded_mappings = [mapping for mapping in mappings if mapping.target_expanded]
    print(f"Backup created: {backup_path}")
    print(f"Mappings written: {len(mappings)}")
    print(f"Mappings requiring review: {len(invalid_mappings)}")
    print(f"Mappings using a corrected source node_id: {len(corrected_mappings)}")
    print(f"Mappings added by parent expansion: {len(expanded_mappings)}")
    for mapping in invalid_mappings:
        invalid_cells = []
        if mapping.source_invalid:
            invalid_cells.append("source")
        if mapping.target_invalid:
            invalid_cells.append("target")
        print(
            f"  source row {mapping.source_row}: "
            f"{mapping.source_node_id} -> {mapping.target_node_id} "
            f"({mapping.origin}; invalid: {', '.join(invalid_cells)})"
        )
    print(f"Destination updated: {target_path}")


if __name__ == "__main__":
    main()
