#!/usr/bin/env python3
"""Fix known CyFun PDF extraction overflows in a copied localized workbook.

The input workbook is never modified. The script re-extracts the official Dutch
and French annotations with the corrected parser, replaces every known target
cell whose content differs from the PDF, and highlights those corrected cells
in yellow. An optional final pass replaces every remaining Greek bullet
character with an ASCII hyphen.
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

import cyfun2025_framework_FR as french
import cyfun2025_framework_NL as base


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_DUTCH_PDF = BASE_DIR / "CyFun2025_Booklet-ESSENTIAL_N.pdf"
DEFAULT_FRENCH_PDF = BASE_DIR / "CyFun2025_Booklet-ESSENTIAL_F_pr3.pdf"
CONTENT_SHEETS = ("requirements_content", "controls_content")
YELLOW_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFFFFF00",
    bgColor="FFFFFF00",
)

EXPECTED_ERROR_REFS = {
    "nl": {"RC.CO-04.3"},
    "fr": {
        "GV.OC-05.1",
        "GV.RM-05.1",
        "GV.RR-04.2",
        "GV.PO-01.2",
        "GV.OV-03.1",
        "ID.RA-08.2",
        "RC.CO-04.3",
    },
}


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Copy a localized CyFun workbook, replace known target cells when "
            "they differ from the PDFs, and highlight repaired cells."
        )
    )
    parser.add_argument(
        "--input",
        type=Path,
        required=True,
        help="Localized NL+FR workbook to copy and correct.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        help="Output workbook. Defaults to <input_stem>_corrected.xlsx.",
    )
    parser.add_argument("--dutch-pdf", type=Path, default=DEFAULT_DUTCH_PDF)
    parser.add_argument("--french-pdf", type=Path, default=DEFAULT_FRENCH_PDF)
    parser.add_argument(
        "--replace-bullets",
        action="store_true",
        help="Replace every PDF bullet character (·, · or •) with an ASCII hyphen -.",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Allow replacing an existing output workbook.",
    )
    args = parser.parse_args()
    if args.output is None:
        args.output = args.input.with_name(
            f"{args.input.stem}_corrected{args.input.suffix}"
        )
    return args


def expected_reference_sets(rows: list[base.WorkbookRow]):
    functions = {
        row.ref_id
        for row in rows
        if row.depth == 1 and base.FUNCTION_RE.match(row.ref_id)
    }
    categories = {
        row.ref_id
        for row in rows
        if row.depth == 2 and base.CATEGORY_RE.match(row.ref_id)
    }
    body_refs = [row.ref_id for row in rows if row.depth in (3, 4)]
    return functions, categories, body_refs


def extract_annotations(
    input_path: Path, dutch_pdf: Path, french_pdf: Path
) -> dict[str, dict[str, str]]:
    rows = base.load_workbook_rows(input_path)
    functions, categories, body_refs = expected_reference_sets(rows)

    dutch = base.extract_pdf(dutch_pdf, functions, categories, body_refs)
    base.validate_dutch_coverage(rows, dutch)

    french.configure_french_extraction()
    extracted_french = base.extract_pdf(
        french_pdf, functions, categories, body_refs
    )
    base.validate_dutch_coverage(rows, extracted_french)
    return {
        "nl": dutch.annotations,
        "fr": extracted_french.annotations,
    }


def header_map(sheet) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in sheet[1]
        if cell.value not in (None, "")
    }


def preserve_bullet_style(expected: str, current: str) -> str:
    """Keep the existing PDF bullet style when replacement is disabled."""
    source_bullet = next(
        (char for char in base.PDF_BULLET_CHARS if char in current), None
    )
    if source_bullet is None:
        return expected
    return re.sub(
        r"(?m)^(\s*)- (?=\S)",
        lambda match: f"{match.group(1)}{source_bullet} ",
        expected,
    )


def normalize_pdf_bullets(text: str) -> str:
    for bullet in base.PDF_BULLET_CHARS:
        text = text.replace(bullet, "-")
    return text


def repair_cells(
    workbook,
    annotations: dict[str, dict[str, str]],
    replace_bullets: bool,
) -> set[tuple[str, str]]:
    corrected: set[tuple[str, str]] = set()
    found_refs = {language: set() for language in annotations}

    for sheet_name in CONTENT_SHEETS:
        sheet = workbook[sheet_name]
        headers = header_map(sheet)
        if "ref_id" not in headers:
            raise ValueError(f"Missing ref_id column in {sheet_name}.")
        for language, expected_by_ref in annotations.items():
            annotation_header = f"annotation[{language}]"
            if annotation_header not in headers:
                raise ValueError(
                    f"Missing {annotation_header} column in {sheet_name}."
                )
            for row_number in range(2, sheet.max_row + 1):
                ref_id = sheet.cell(row_number, headers["ref_id"]).value
                if ref_id not in EXPECTED_ERROR_REFS[language]:
                    continue
                expected = expected_by_ref.get(ref_id)
                if not expected:
                    raise ValueError(
                        f"No extracted {language} annotation for {ref_id}."
                    )
                cell = sheet.cell(row_number, headers[annotation_header])
                current = cell.value
                current_text = current if isinstance(current, str) else ""
                found_refs[language].add(ref_id)

                # Bullet style is an output choice, not a content difference.
                # All other differences are resolved in favour of the PDF.
                if normalize_pdf_bullets(current_text) == expected:
                    continue
                cell.value = (
                    expected
                    if replace_bullets
                    else preserve_bullet_style(expected, current_text)
                )
                cell.fill = YELLOW_FILL
                corrected.add((sheet_name, cell.coordinate))

    for language, expected_refs in EXPECTED_ERROR_REFS.items():
        if found_refs[language] != expected_refs:
            raise ValueError(
                f"Unexpected repaired refs for {language}: "
                f"{sorted(found_refs[language])}; expected {sorted(expected_refs)}."
            )
    return corrected


def replace_pdf_bullets(workbook) -> tuple[int, int]:
    changed_cells = 0
    replaced_characters = 0
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                replacements = sum(
                    cell.value.count(bullet) for bullet in base.PDF_BULLET_CHARS
                )
                if replacements == 0:
                    continue
                replaced_characters += replacements
                cell.value = normalize_pdf_bullets(cell.value)
                changed_cells += 1
    return changed_cells, replaced_characters


def verify_output(
    input_path: Path,
    output_path: Path,
    corrected_cells: set[tuple[str, str]],
    replace_bullets: bool,
) -> None:
    source = openpyxl.load_workbook(input_path, data_only=False)
    output = openpyxl.load_workbook(output_path, data_only=False)
    try:
        if output.sheetnames != source.sheetnames:
            raise ValueError("The output workbook changed the sheet list.")

        remaining_bullets = []
        remaining_bleeds = []
        formula_errors = []
        error_values = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"}
        for sheet in output.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if (
                        replace_bullets
                        and isinstance(value, str)
                        and any(bullet in value for bullet in base.PDF_BULLET_CHARS)
                    ):
                        remaining_bullets.append(f"{sheet.title}!{cell.coordinate}")
                    if (
                        isinstance(value, str)
                        and cell.column > 1
                        and base.SINGLE_LETTER_BLEED_RE.search(value)
                    ):
                        remaining_bleeds.append(f"{sheet.title}!{cell.coordinate}")
                    if cell.data_type == "e" or value in error_values:
                        formula_errors.append(f"{sheet.title}!{cell.coordinate}")
        if remaining_bullets:
            raise ValueError(
                "PDF bullet characters remain: "
                + ", ".join(remaining_bullets[:20])
            )
        if remaining_bleeds:
            raise ValueError(
                "Single-letter PDF bleed remains: "
                + ", ".join(remaining_bleeds[:20])
            )
        if formula_errors:
            raise ValueError(
                "Formula errors found: " + ", ".join(formula_errors[:20])
            )

        for sheet_name, coordinate in corrected_cells:
            cell = output[sheet_name][coordinate]
            fill_rgb = cell.fill.fgColor.rgb
            if cell.fill.fill_type != "solid" or fill_rgb != "FFFFFF00":
                raise ValueError(
                    f"Corrected cell is not highlighted yellow: "
                    f"{sheet_name}!{coordinate}."
                )
        print(
            f"Verification passed: {len(corrected_cells)} corrected cells are "
            "yellow and no extraction bleed remains"
            + ("; no PDF bullet characters remain." if replace_bullets else ".")
        )
    finally:
        output.close()
        source.close()


def main() -> int:
    args = parse_arguments()
    if not args.input.is_file():
        raise FileNotFoundError(f"Workbook not found: {args.input}")
    if args.output.resolve() == args.input.resolve():
        raise ValueError("The output path must differ from the input path.")
    if args.output.exists() and not args.force:
        raise FileExistsError(
            f"Output already exists: {args.output}. Use --force to replace it."
        )

    annotations = extract_annotations(
        args.input, args.dutch_pdf, args.french_pdf
    )
    workbook = openpyxl.load_workbook(args.input)
    try:
        corrected_cells = repair_cells(
            workbook, annotations, replace_bullets=args.replace_bullets
        )
        if args.replace_bullets:
            bullet_cells, bullet_characters = replace_pdf_bullets(workbook)
        else:
            bullet_cells, bullet_characters = 0, 0
        args.output.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(args.output)
    finally:
        workbook.close()

    verify_output(
        args.input,
        args.output,
        corrected_cells,
        replace_bullets=args.replace_bullets,
    )
    for sheet_name, coordinate in sorted(corrected_cells):
        print(f"Corrected and highlighted: {sheet_name}!{coordinate}")
    if args.replace_bullets:
        print(
            f"Replaced {bullet_characters} PDF bullet characters in "
            f"{bullet_cells} cells."
        )
    else:
        print("Bullet replacement disabled; existing Greek bullets were preserved.")
    print(f"Created: {args.output.resolve()}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (FileNotFoundError, FileExistsError, ValueError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1) from exc
