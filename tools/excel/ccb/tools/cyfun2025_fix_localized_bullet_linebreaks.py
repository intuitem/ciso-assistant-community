#!/usr/bin/env python3
"""Restore missing title/body line breaks in localized CyFun bullet lists.

The Dutch and French PDFs visually distinguish a bullet title (medium/bold
font) from its explanatory paragraph (light/regular font).  The original PDF
parser flattens every line inside a bullet, so that typographic boundary is
lost in the localized annotations.

This script never modifies the input workbook.  It first copies it, then only
inserts a newline after PDF-confirmed bullet titles whose explanatory text
starts on the next visual PDF line in ``annotation[nl]`` and ``annotation[fr]``
on ``requirements_content`` and ``controls_content``.
The optional ``--highlight`` flag gives corrected cells a yellow fill.  All
other text, bullet levels, styles, formulas, and cells are kept.
"""

from __future__ import annotations

import argparse
from copy import copy
import re
import shutil
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path

import fitz
import openpyxl
from openpyxl.styles import PatternFill

import cyfun2025_framework_FR as french
import cyfun2025_framework_NL as base


SCRIPT_DIR = Path(__file__).resolve().parent
CCB_DIR = SCRIPT_DIR.parent
DEFAULT_DUTCH_PDF = CCB_DIR / "CyFun2025_Booklet-ESSENTIAL_N.pdf"
DEFAULT_FRENCH_PDF = CCB_DIR / "CyFun2025_Booklet-ESSENTIAL_F_pr3.pdf"
CONTENT_SHEETS = ("requirements_content", "controls_content")
LANGUAGE_COLUMNS = {"nl": "annotation[nl]", "fr": "annotation[fr]"}
HEADING_FONT_MARKERS = ("-Medium", "-Semibold", "-Bold", "-Heavy")
YELLOW_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFFFFF00",
    bgColor="FFFFFF00",
)


@dataclass(frozen=True)
class CellChange:
    sheet: str
    coordinate: str
    ref_id: str
    language: str
    headings: tuple[str, ...]
    before: str
    after: str


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Copy a localized CyFun workbook and restore only the missing "
            "line breaks between formatted bullet titles and their text."
        )
    )
    parser.add_argument(
        "--input",
        type=Path,
        required=True,
        help="Localized NL+FR workbook to copy and correct.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        help="Output workbook. Defaults to <input_stem>_linebreaks_fixed.xlsx.",
    )
    parser.add_argument("--dutch-pdf", type=Path, default=DEFAULT_DUTCH_PDF)
    parser.add_argument("--french-pdf", type=Path, default=DEFAULT_FRENCH_PDF)
    parser.add_argument(
        "--force",
        action="store_true",
        help="Allow replacing an existing output workbook.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Report the cells that would change without creating a copy.",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="List every corrected cell in addition to the summary.",
    )
    parser.add_argument(
        "--highlight",
        action="store_true",
        help="Highlight every corrected cell with a solid yellow fill.",
    )
    args = parser.parse_args()
    if args.output is None:
        output_suffix = (
            "_linebreaks_fixed_highlighted"
            if args.highlight
            else "_linebreaks_fixed"
        )
        args.output = args.input.with_name(
            f"{args.input.stem}{output_suffix}{args.input.suffix}"
        )
    return args


def header_map(sheet) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in sheet[1]
        if cell.value not in (None, "")
    }


def expected_reference_sets(rows: list[base.WorkbookRow]):
    functions = {
        row.ref_id
        for row in rows
        if row.depth == 1 and base.FUNCTION_RE.match(row.ref_id)
    }
    categories = {
        row.ref_id
        for row in rows
        if row.depth == 2 and base.CATEGORY_RE.match(row.ref_id)
    }
    body_refs = [row.ref_id for row in rows if row.depth in (3, 4)]
    return functions, categories, body_refs


def extract_localized_annotations(
    input_path: Path, dutch_pdf: Path, french_pdf: Path
) -> dict[str, dict[str, str]]:
    rows = base.load_workbook_rows(input_path)
    functions, categories, body_refs = expected_reference_sets(rows)

    dutch = base.extract_pdf(dutch_pdf, functions, categories, body_refs)
    base.validate_dutch_coverage(rows, dutch)

    french.configure_french_extraction()
    extracted_french = base.extract_pdf(
        french_pdf, functions, categories, body_refs
    )
    base.validate_dutch_coverage(rows, extracted_french)
    return {"nl": dutch.annotations, "fr": extracted_french.annotations}


def is_pdf_bullet_span(span: dict) -> bool:
    return span.get("text", "").strip() in base.PDF_BULLET_CHARS


def is_heading_font(font_name: str) -> bool:
    return any(marker in font_name for marker in HEADING_FONT_MARKERS)


def clean_heading(parts: list[str]) -> str:
    cleaned = base.clean_content_block("\n".join(parts))
    return cleaned.removesuffix(base.WORD_JOIN_MARKER).strip()


def extract_pdf_bullet_headings_by_ref(
    pdf_path: Path, expected_body_refs: list[str]
) -> dict[str, set[str]]:
    """Return PDF-confirmed bullet titles scoped to their own ref_id."""
    headings_by_ref: dict[str, set[str]] = defaultdict(set)
    expected_positions = {
        ref_id: index for index, ref_id in enumerate(expected_body_refs)
    }
    expected_index = 0
    current_ref: str | None = None
    document = fitz.open(pdf_path)
    try:
        for page_number, page in enumerate(document, start=1):
            pending_parts: list[str] = []
            waiting_after_bullet = False
            pending_ref: str | None = None
            parsed_blocks = {
                (block.x0, block.y0, block.x1, block.y1): block
                for block in base.iter_pdf_blocks(page, page_number)
            }
            raw_blocks = sorted(
                (
                    block
                    for block in page.get_text("dict")["blocks"]
                    if block.get("type") == 0
                    and 84 <= block["bbox"][0] <= 190
                    and block["bbox"][2] <= 570
                    and block["bbox"][3] < 790
                ),
                key=lambda block: (block["bbox"][1], block["bbox"][0]),
            )
            for block in raw_blocks:
                parsed_block = parsed_blocks.get(tuple(block["bbox"]))
                if parsed_block is None:
                    continue
                reference_match = base.REFERENCE_RE.match(parsed_block.text)
                if reference_match and (
                    parsed_block.first_font.endswith("-Bold")
                    or parsed_block.first_font.endswith("-Semibold")
                ):
                    ref_id = base.normalized_pdf_reference(reference_match)
                    position = expected_positions.get(ref_id)
                    if position is not None and position == expected_index:
                        current_ref = ref_id
                        expected_index += 1
                        pending_parts = []
                        waiting_after_bullet = False
                        pending_ref = None

                for line in block.get("lines", []):
                    spans = [
                        span
                        for span in line.get("spans", [])
                        if span.get("text", "").strip()
                    ]
                    if not spans:
                        continue

                    bullet_indexes = [
                        index
                        for index, span in enumerate(spans)
                        if is_pdf_bullet_span(span)
                    ]
                    if bullet_indexes:
                        pending_parts = []
                        waiting_after_bullet = True
                        pending_ref = current_ref
                        spans = spans[bullet_indexes[-1] + 1 :]
                        if not spans:
                            continue

                    if not waiting_after_bullet:
                        continue

                    leading_heading_spans: list[dict] = []
                    remaining_spans: list[dict] = []
                    body_started = False
                    for span in spans:
                        if not body_started and is_heading_font(
                            span.get("font", "")
                        ):
                            leading_heading_spans.append(span)
                        else:
                            body_started = True
                            remaining_spans.append(span)

                    if leading_heading_spans:
                        pending_parts.append(
                            "".join(
                                span.get("text", "")
                                for span in leading_heading_spans
                            )
                        )

                    if remaining_spans or (not leading_heading_spans and spans):
                        heading = clean_heading(pending_parts)
                        # A font change within one visual PDF line is emphasis,
                        # not a line-break boundary. Only keep the heading when
                        # the normal-weight body starts on a later PDF line.
                        transition_is_on_same_line = bool(
                            leading_heading_spans and remaining_spans
                        )
                        if (
                            heading
                            and pending_ref is not None
                            and not transition_is_on_same_line
                        ):
                            headings_by_ref[pending_ref].add(heading)
                        pending_parts = []
                        waiting_after_bullet = False
                        pending_ref = None
    finally:
        document.close()
    if expected_index != len(expected_body_refs):
        missing = expected_body_refs[expected_index : expected_index + 10]
        raise ValueError(
            f"Could not map PDF bullet headings after {expected_index} of "
            f"{len(expected_body_refs)} ref_id values in {pdf_path.name}. "
            f"First missing refs: {missing}."
        )
    return dict(headings_by_ref)


def map_headings_to_annotations(
    headings_by_ref: dict[str, set[str]], annotations: dict[str, str]
) -> dict[str, tuple[str, ...]]:
    """Keep only PDF headings found at the start of an extracted bullet."""
    by_ref: dict[str, set[str]] = defaultdict(set)
    for ref_id, annotation in annotations.items():
        headings = headings_by_ref.get(ref_id, set())
        if not headings:
            continue
        bullet_texts = []
        for line in annotation.splitlines():
            match = re.match(r"^\s*-\s{1,2}(\S.*)$", line)
            if match:
                bullet_texts.append(match.group(1).strip())
        for bullet in bullet_texts:
            matching_headings = [
                heading
                for heading in headings
                if bullet == heading or bullet.startswith(f"{heading} ")
            ]
            if matching_headings:
                by_ref[ref_id].add(max(matching_headings, key=len))
    return {
        ref_id: tuple(sorted(ref_headings, key=lambda value: (-len(value), value)))
        for ref_id, ref_headings in by_ref.items()
    }


def validate_known_unsplit_bullets(
    headings_by_language: dict[str, dict[str, tuple[str, ...]]]
) -> None:
    forbidden_boundaries = {
        ("fr", "GV.OC-04.3", "Tests réguliers"),
        ("nl", "GV.OC-04.3", "Regelmatige tests"),
        ("fr", "GV.SC-01.1", "Identifier les objectifs:"),
    }
    unexpected = [
        (language, ref_id, heading)
        for language, ref_id, heading in forbidden_boundaries
        if heading in headings_by_language[language].get(ref_id, ())
    ]
    if unexpected:
        raise ValueError(
            "A known non-heading bullet prefix would be split: "
            f"{unexpected}."
        )


def insert_missing_breaks(
    text: str, headings: tuple[str, ...]
) -> tuple[str, tuple[str, ...]]:
    changed_headings: list[str] = []
    output_lines: list[str] = []
    bullet_pattern = re.compile(
        r"^(?P<indent> *)(?P<marker>-[ \t]{1,2})(?P<content>\S.*)$"
    )
    for line in text.split("\n"):
        match = bullet_pattern.match(line)
        if match is None:
            output_lines.append(line)
            continue

        content = match.group("content")
        matching_headings = [
            heading
            for heading in headings
            if content.startswith(f"{heading} ")
        ]
        if not matching_headings:
            output_lines.append(line)
            continue

        heading = max(matching_headings, key=len)
        body = content[len(heading) :].lstrip(" \t")
        continuation_indent = " " * (len(match.group("indent")) + 2)
        output_lines.extend(
            (
                f"{match.group('indent')}{match.group('marker')}{heading}",
                f"{continuation_indent}{body}",
            )
        )
        changed_headings.append(heading)
    return "\n".join(output_lines), tuple(changed_headings)


def plan_changes(
    workbook, headings_by_language: dict[str, dict[str, tuple[str, ...]]]
) -> list[CellChange]:
    changes: list[CellChange] = []
    for sheet_name in CONTENT_SHEETS:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Missing sheet: {sheet_name}.")
        sheet = workbook[sheet_name]
        headers = header_map(sheet)
        required_headers = {"ref_id", *LANGUAGE_COLUMNS.values()}
        missing_headers = required_headers - headers.keys()
        if missing_headers:
            raise ValueError(
                f"Missing columns in {sheet_name}: {sorted(missing_headers)}."
            )

        for row_number in range(2, sheet.max_row + 1):
            ref_id = sheet.cell(row_number, headers["ref_id"]).value
            if not isinstance(ref_id, str):
                continue
            for language, column_name in LANGUAGE_COLUMNS.items():
                headings = headings_by_language[language].get(ref_id, ())
                if not headings:
                    continue
                cell = sheet.cell(row_number, headers[column_name])
                if not isinstance(cell.value, str):
                    continue
                after, changed_headings = insert_missing_breaks(
                    cell.value, headings
                )
                if after == cell.value:
                    continue
                changes.append(
                    CellChange(
                        sheet=sheet_name,
                        coordinate=cell.coordinate,
                        ref_id=ref_id,
                        language=language,
                        headings=changed_headings,
                        before=cell.value,
                        after=after,
                    )
                )
    return changes


def apply_changes(
    workbook, changes: list[CellChange], highlight: bool
) -> None:
    for change in changes:
        cell = workbook[change.sheet][change.coordinate]
        cell.value = change.after
        if highlight:
            cell.fill = copy(YELLOW_FILL)


def non_fill_style_signature(cell) -> tuple:
    return (
        copy(cell.font),
        copy(cell.border),
        copy(cell.alignment),
        cell.number_format,
        copy(cell.protection),
        cell.quotePrefix,
        cell.pivotButton,
    )


def has_yellow_fill(cell) -> bool:
    return (
        cell.fill.fill_type == "solid"
        and cell.fill.fgColor.type == "rgb"
        and cell.fill.fgColor.rgb == "FFFFFF00"
    )


def verify_output(
    input_path: Path,
    output_path: Path,
    planned_changes: list[CellChange],
    highlight: bool,
) -> None:
    source = openpyxl.load_workbook(input_path, data_only=False)
    output = openpyxl.load_workbook(output_path, data_only=False)
    try:
        if source.sheetnames != output.sheetnames:
            raise ValueError("The output workbook changed the sheet list.")

        expected = {
            (change.sheet, change.coordinate): change for change in planned_changes
        }
        unexpected_value_changes: list[str] = []
        unexpected_style_changes: list[str] = []
        formula_errors: list[str] = []
        error_values = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"}

        for source_sheet in source.worksheets:
            output_sheet = output[source_sheet.title]
            if (
                source_sheet.max_row != output_sheet.max_row
                or source_sheet.max_column != output_sheet.max_column
            ):
                raise ValueError(
                    f"The used range changed in {source_sheet.title}."
                )
            for source_row, output_row in zip(
                source_sheet.iter_rows(), output_sheet.iter_rows()
            ):
                for source_cell, output_cell in zip(source_row, output_row):
                    key = (source_sheet.title, source_cell.coordinate)
                    planned = expected.get(key)
                    expected_value = planned.after if planned else source_cell.value
                    if output_cell.value != expected_value:
                        unexpected_value_changes.append(
                            f"{source_sheet.title}!{source_cell.coordinate}"
                        )
                    if planned and highlight:
                        if (
                            non_fill_style_signature(source_cell)
                            != non_fill_style_signature(output_cell)
                            or not has_yellow_fill(output_cell)
                        ):
                            unexpected_style_changes.append(
                                f"{source_sheet.title}!{source_cell.coordinate}"
                            )
                    elif source_cell.style_id != output_cell.style_id:
                        unexpected_style_changes.append(
                            f"{source_sheet.title}!{source_cell.coordinate}"
                        )
                    if (
                        output_cell.data_type == "e"
                        or output_cell.value in error_values
                    ):
                        formula_errors.append(
                            f"{source_sheet.title}!{source_cell.coordinate}"
                        )

        if unexpected_value_changes:
            raise ValueError(
                "Unexpected value changes: "
                + ", ".join(unexpected_value_changes[:20])
            )
        if unexpected_style_changes:
            raise ValueError(
                "Unexpected style changes: "
                + ", ".join(unexpected_style_changes[:20])
            )
        if formula_errors:
            raise ValueError(
                "Formula errors found: " + ", ".join(formula_errors[:20])
            )

        for change in planned_changes:
            round_trip, reapplied = insert_missing_breaks(
                output[change.sheet][change.coordinate].value,
                change.headings,
            )
            if reapplied or round_trip != change.after:
                raise ValueError(
                    f"Unresolved line break in {change.sheet}!"
                    f"{change.coordinate}."
                )

        for ref_id in {
            change.ref_id for change in planned_changes
        }:
            requirement_values = {}
            control_values = {}
            for sheet_name, target in (
                ("requirements_content", requirement_values),
                ("controls_content", control_values),
            ):
                sheet = output[sheet_name]
                headers = header_map(sheet)
                for row_number in range(2, sheet.max_row + 1):
                    if sheet.cell(row_number, headers["ref_id"]).value == ref_id:
                        for language, column_name in LANGUAGE_COLUMNS.items():
                            target[language] = sheet.cell(
                                row_number, headers[column_name]
                            ).value
            if requirement_values != control_values:
                raise ValueError(
                    f"requirements/controls mismatch after repair for {ref_id}."
                )
    finally:
        output.close()
        source.close()


def main() -> int:
    args = parse_arguments()
    if not args.input.is_file():
        raise FileNotFoundError(f"Workbook not found: {args.input}")
    for pdf_path in (args.dutch_pdf, args.french_pdf):
        if not pdf_path.is_file():
            raise FileNotFoundError(f"PDF not found: {pdf_path}")
    if args.output.resolve() == args.input.resolve():
        raise ValueError("The output path must differ from the input path.")
    if args.output.exists() and not args.force and not args.dry_run:
        raise FileExistsError(
            f"Output already exists: {args.output}. Use --force to replace it."
        )

    annotations = extract_localized_annotations(
        args.input, args.dutch_pdf, args.french_pdf
    )
    rows = base.load_workbook_rows(args.input)
    _, _, body_refs = expected_reference_sets(rows)
    pdf_headings = {
        "nl": extract_pdf_bullet_headings_by_ref(
            args.dutch_pdf, body_refs
        ),
        "fr": extract_pdf_bullet_headings_by_ref(
            args.french_pdf, body_refs
        ),
    }
    headings_by_language = {
        language: map_headings_to_annotations(
            pdf_headings[language], annotations[language]
        )
        for language in LANGUAGE_COLUMNS
    }
    validate_known_unsplit_bullets(headings_by_language)

    workbook = openpyxl.load_workbook(args.input)
    try:
        changes = plan_changes(workbook, headings_by_language)
    finally:
        workbook.close()

    if args.dry_run:
        print(f"Dry run: {len(changes)} cells would be corrected.")
    else:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(args.input, args.output)
        output_workbook = openpyxl.load_workbook(args.output)
        try:
            apply_changes(
                output_workbook, changes, highlight=args.highlight
            )
            output_workbook.save(args.output)
        finally:
            output_workbook.close()
        verify_output(
            args.input,
            args.output,
            changes,
            highlight=args.highlight,
        )
        print(f"Verification passed for {len(changes)} corrected cells.")

    counts = Counter(
        (change.sheet, change.language) for change in changes
    )
    total_breaks = sum(len(change.headings) for change in changes)
    unique_targets = {
        (change.ref_id, change.language) for change in changes
    }
    print(
        f"Summary: {len(changes)} cells, {len(unique_targets)} unique "
        f"ref/language targets, {total_breaks} inserted line breaks."
    )
    print(
        "  Yellow highlighting: "
        + ("enabled" if args.highlight else "disabled")
    )
    for (sheet_name, language), count in sorted(counts.items()):
        print(f"  {sheet_name} annotation[{language}]: {count} cells")
    if args.verbose:
        for change in changes:
            print(
                f"{change.sheet}!{change.coordinate} | {change.ref_id} | "
                f"annotation[{change.language}] | "
                f"{len(change.headings)} break(s)"
            )
    if not args.dry_run:
        print(f"Created: {args.output.resolve()}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (FileNotFoundError, FileExistsError, ValueError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1) from exc
