"""Add official Dutch CyFun 2025 texts to a copied workbook.

The script keeps the source workbook unchanged and saves a separate output
workbook. In that copy, the Dutch metadata and content columns are added directly
to the existing sheets when their source content exists in the PDF. The scoring
scale is deliberately left unchanged because it is not part of the official
booklet.

Before writing the Dutch content, the same extraction is run on the official
English PDF and compared with the existing English workbook fields. This makes
the shared ``ref_id`` values, rather than page or row numbers, the source of the
mapping.
"""

from __future__ import annotations

import argparse
import re
import statistics
import sys
import unicodedata
from copy import copy
from dataclasses import dataclass, field
from difflib import SequenceMatcher
from pathlib import Path
from typing import Iterable

try:
    import fitz  # PyMuPDF
    import openpyxl
except ImportError as exc:  # pragma: no cover - only used for a clear CLI error
    raise SystemExit(
        "Missing dependency. Install PyMuPDF and openpyxl in the active Python "
        "environment before running this script."
    ) from exc


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR.parent
DEFAULT_WORKBOOK = DATA_DIR / "cyfun2025.xlsx"
DEFAULT_ENGLISH_PDF = DATA_DIR / "CyFun2025_Booklet_ESSENTIAL_E.pdf"
DEFAULT_DUTCH_PDF = DATA_DIR / "CyFun2025_Booklet-ESSENTIAL_N.pdf"
DEFAULT_OUTPUT = DATA_DIR / "cyfun2025_nl.xlsx"
LANGUAGE_CODE = "nl"
LANGUAGE_NAME = "Dutch"
INTRODUCTION_HEADING = "INLEIDING"

SOURCE_SHEET_NAME = "requirements_content"
TRANSLATED_COLUMNS = tuple(
    f"{field_name}[{LANGUAGE_CODE}]"
    for field_name in ("name", "description", "annotation")
)
CONTROLS_SOURCE_SHEET_NAME = "controls_content"
IG_SOURCE_SHEET_NAME = "IG_content"
SCORES_SOURCE_SHEET_NAME = "scores_content"
META_SHEET_NAMES = ("library_meta", "requirements_meta")
FORBIDDEN_LOCALIZED_SHEET_NAMES = {
    "requirements_content_nl",
    "controls_content_nl",
    "IG_content_nl",
    "scores_content_nl",
}
ENGLISH_PDF_ALLOWED_MISSING_REFS = {"GV.SC-10"}

IG_DUTCH_NAMES = {
    "B": "basic",
    "I": "important",
    "E": "essential",
    "BK": "basic - kernmaatregelen",
    "IK": "important - kernmaatregelen",
    "EK": "essential - kernmaatregelen",
    "BG": "basic - managementaspecten",
    "IG": "important - managementaspecten",
    "EG": "essential - managementaspecten",
}
IG_REQUIRED_TERMS = (
    "basic",
    "important",
    "essential",
    "kernmaatregelen",
    "managementaspecten",
)

DUTCH_FUNCTION_NAMES = {
    "GV": "BEHEREN",
    "ID": "IDENTIFICEREN",
    "PR": "BESCHERMEN",
    "DE": "DETECTEREN",
    "RS": "REAGEREN",
    "RC": "HERSTELLEN",
}

REFERENCE_RE = re.compile(r"^\s*([A-Z]{2}\.[A-Z]{2}-\d{2}(?:[.-]\d+)?)\b")
CATEGORY_RE = re.compile(r"^[A-Z]{2}\.[A-Z]{2}$")
FUNCTION_RE = re.compile(r"^[A-Z]{2}$")
ILLEGAL_XML_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
WORD_JOIN_MARKER = "\ufff0"
BULLET_LEVEL_MARKER = "\ufff1"
BULLET_SPLIT_MARKER = "\ufff2"
BULLET_CHAR = "-"
PDF_BULLET_CHARS = ("\u0387", "\u00b7", "\u2022")
BULLET_LEFT_X = 87.9
BULLET_INDENT_STEP_X = 10.65
BULLET_INDENT_SPACES = 4
EXPECTED_TARGET_BULLET_INDENTATIONS = {0, 4, 8}
BODY_TEXT_MAX_SIZE = 10.1

GUIDANCE_LABELS = ("Implementation guidance", "Implementatierichtlijnen")
REFERENCES_LABELS = ("References", "Referenties")
REFERENCE_TEXT_REPLACEMENTS: dict[str, str] = {}
SINGLE_LETTER_BLEED_RE = re.compile(
    r"(?:^|\n\n)(?:[^\W\d_]\n\n){3,}", re.UNICODE
)
BACK_MATTER_MARKERS = (
    "ANNEX A",
    "ANNEXE A",
    "BIJLAGE A",
    "Legal deposit",
    "Dépôt légal",
    "Juridisch depot",
)


@dataclass
class PdfBlock:
    page_number: int
    x0: float
    y0: float
    x1: float
    y1: float
    text: str
    first_font: str
    first_size: float


@dataclass
class ExtractedTexts:
    names: dict[str, str] = field(default_factory=dict)
    descriptions: dict[str, str] = field(default_factory=dict)
    annotations: dict[str, str] = field(default_factory=dict)


@dataclass
class WorkbookRow:
    row_number: int
    ref_id: str
    depth: int
    name: str | None
    description: str | None
    annotation: str | None


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Extract official Dutch CyFun 2025 content and add it to copied "
            "content sheets."
        )
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--english-pdf", type=Path, default=DEFAULT_ENGLISH_PDF)
    parser.add_argument("--dutch-pdf", type=Path, default=DEFAULT_DUTCH_PDF)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument(
        "--force",
        action="store_true",
        help="Allow replacing an existing output workbook.",
    )
    return parser.parse_args()


def localized_column(field_name: str) -> str:
    return f"{field_name}[{LANGUAGE_CODE}]"


def clean_inline_text(text: str) -> str:
    """Remove PDF artefacts and flatten wrapped lines into one readable value."""
    text = re.sub(r"\u00ad[ \t]*\n[ \t]*", "", text)
    text = text.replace("\u00ad", "")
    text = text.replace("\u200b", "").replace("\ufeff", "")
    text = text.replace("\x02", "").replace("\x07", "")
    text = text.replace("\u00a0", " ").replace("\u2003", " ")
    text = ILLEGAL_XML_RE.sub("", text)
    text = re.sub(r"-\s*\n\s*(?=\w)", "", text)
    text = re.sub(r"\s*\n\s*", " ", text)
    return re.sub(r"[ \t]+", " ", text).strip()


def clean_intro_text(text: str) -> str:
    """Clean an introduction block and normalise its footnote typography."""
    text = clean_inline_text(text)
    text = re.sub(r"\s+1\s*,", "1,", text)
    text = re.sub(r"\bTR\s+103\s+305-1\b", "TR 103305-1", text)
    return text


def extract_intro_description(pdf_path: Path, heading: str) -> str:
    """Extract the three introduction blocks used by the English metadata."""
    document = fitz.open(pdf_path)
    try:
        intro_blocks: list[PdfBlock] | None = None
        for page_number, page in enumerate(document, start=1):
            blocks = list(iter_pdf_blocks(page, page_number))
            heading_block = next(
                (
                    block
                    for block in blocks
                    if clean_inline_text(block.text).upper() == heading.upper()
                    and block.y0 < 130
                ),
                None,
            )
            if heading_block is None:
                continue
            intro_blocks = sorted(
                (
                    block
                    for block in blocks
                    if 80 <= block.x0 <= 100
                    and block.x1 >= 300
                    and block.y0 > heading_block.y1
                    and block.y1 < 400
                ),
                key=lambda block: (block.y0, block.x0),
            )
            break
        if intro_blocks is None:
            raise ValueError(f"No {heading!r} page found in {pdf_path.name}.")
        if len(intro_blocks) < 4:
            raise ValueError(
                f"Unexpected introduction layout in {pdf_path.name}: "
                f"found {len(intro_blocks)} body blocks."
            )

        opening = clean_intro_text(intro_blocks[0].text)
        bullets = clean_content_block(intro_blocks[1].text)
        bullet_lines = []
        for line in bullets.splitlines():
            line = line.strip()
            if not line:
                continue
            line = re.sub(
                rf"^[{''.join(re.escape(char) for char in (BULLET_CHAR, *PDF_BULLET_CHARS))}]\s*",
                "· ",
                line,
            )
            bullet_lines.append(clean_intro_text(line))
        sources = clean_intro_text(intro_blocks[2].text)
        if len(bullet_lines) != 3 or not all(
            line.startswith("· ") for line in bullet_lines
        ):
            raise ValueError(
                f"Unexpected introduction bullet list in {pdf_path.name}: "
                f"{bullet_lines}"
            )
        return "\n".join((opening, *bullet_lines, sources))
    finally:
        document.close()


def build_localized_meta_description(
    input_path: Path, english_pdf: Path, localized_pdf: Path
) -> str:
    """Validate the English baseline and build the localized metadata value."""
    english_intro = extract_intro_description(english_pdf, "INTRODUCTION")
    localized_intro = extract_intro_description(localized_pdf, INTRODUCTION_HEADING)
    workbook = openpyxl.load_workbook(input_path, read_only=True, data_only=False)
    try:
        suffixes: set[str] = set()
        for sheet_name in META_SHEET_NAMES:
            sheet = workbook[sheet_name]
            values = {
                str(row[0]).strip(): row[1]
                for row in sheet.iter_rows(values_only=True)
                if row and row[0] not in (None, "")
            }
            description = values.get("description")
            if not isinstance(description, str):
                raise ValueError(f"Missing English description in {sheet_name}.")
            description_lines = description.splitlines()
            suffix_lines = [
                line.strip()
                for line in description_lines
                if line.strip().startswith(("http://", "https://", "Version "))
            ]
            if len(suffix_lines) != 2:
                raise ValueError(
                    f"Expected URL and version in {sheet_name}:description."
                )
            suffixes.add("\n".join(suffix_lines))
            english_core = "\n".join(
                line for line in description_lines if line.strip() not in suffix_lines
            ).strip()
            score = similarity(english_core, english_intro)
            if score < 0.97:
                raise ValueError(
                    f"English introduction mismatch in {sheet_name}: "
                    f"similarity={score:.3f}."
                )
        if len(suffixes) != 1:
            raise ValueError("The English metadata URL/version suffix is inconsistent.")
        suffix = suffixes.pop()
        return f"{localized_intro}\n{suffix}"
    finally:
        workbook.close()


def clean_content_block(text: str) -> str:
    """Clean a body block while retaining bullet separators and nesting."""
    ends_with_soft_hyphen = bool(re.search(r"\u00ad\s*$", text))
    text = re.sub(r"\u00ad[ \t]*\n[ \t]*", "", text)
    text = text.replace("\u00ad", "")
    text = text.replace("\u200b", "").replace("\ufeff", "")
    text = text.replace("\x02", "").replace("\x07", "")
    text = text.replace("\u00a0", " ").replace("\u2003", " ")
    text = ILLEGAL_XML_RE.sub("", text)
    text = text.replace("\r", "")

    # A normal hyphen at the end of a PDF line is a typesetting hyphen. A soft
    # hyphen has already been removed, so genuine compounds such as ICT/OT-
    # systems retain their hard hyphen before the former soft-hyphen position.
    text = re.sub(r"-\s*\n\s*(?=\w)", "", text)
    text = re.sub(
        rf"[ \t]*{re.escape(BULLET_LEVEL_MARKER)}(\d+)[ \t]*\n?",
        lambda match: f"\n{BULLET_SPLIT_MARKER}{match.group(1)}:",
        text,
    )
    text = re.sub(
        rf"[ \t]*[{''.join(re.escape(char) for char in PDF_BULLET_CHARS)}][ \t]*\n?",
        f"\n{BULLET_SPLIT_MARKER}0:",
        text,
    )

    split_parts = re.split(
        rf"{re.escape(BULLET_SPLIT_MARKER)}(\d+):", text
    )

    def flatten(raw_part: str) -> str:
        flattened = re.sub(r"\s*\n\s*", " ", raw_part)
        return re.sub(r"[ \t]+", " ", flattened).strip()

    parts: list[str] = []
    introductory_text = flatten(split_parts[0])
    if introductory_text:
        parts.append(introductory_text)
    for index in range(1, len(split_parts), 2):
        level = int(split_parts[index])
        bullet_text = flatten(split_parts[index + 1])
        if not bullet_text:
            continue
        indentation = " " * (level * BULLET_INDENT_SPACES)
        parts.append(f"{indentation}{BULLET_CHAR}  {bullet_text}")

    if not parts:
        return ""
    cleaned = "\n".join(parts)
    return cleaned + WORD_JOIN_MARKER if ends_with_soft_hyphen else cleaned


def join_description_blocks(blocks: Iterable[str]) -> str:
    output = ""
    for block in blocks:
        block = block.strip()
        if not block:
            continue
        if output.endswith(WORD_JOIN_MARKER):
            output = output[: -len(WORD_JOIN_MARKER)] + block.lstrip()
        else:
            output = block if not output else f"{output} {block}"
    output = output.replace(WORD_JOIN_MARKER, "")
    return re.sub(r"\s+", " ", output).strip()


def join_content_blocks(blocks: Iterable[str]) -> str:
    output = ""
    previous_block: str | None = None
    for block in blocks:
        block = block.strip("\r\n")
        if not block.strip():
            continue
        if block == previous_block:
            continue
        previous_block = block
        if output.endswith(WORD_JOIN_MARKER):
            output = output[: -len(WORD_JOIN_MARKER)] + block.lstrip()
            continue
        starts_with_bullet = bool(
            re.match(rf" *{re.escape(BULLET_CHAR)}  ", block)
        )
        separator = "\n" if starts_with_bullet else "\n\n"
        output = block if not output else output.rstrip() + separator + block
    return output.replace(WORD_JOIN_MARKER, "").strip("\r\n")


def normalized_pdf_reference(match: re.Match[str]) -> str:
    """Normalize two known English-PDF refs such as GV.RR-03-1 to dotted form."""
    ref_id = match.group(1)
    if ref_id.count(".") == 1:
        prefix, final_number = ref_id.rsplit("-", 1)
        if final_number.isdigit() and prefix.rsplit("-", 1)[-1].isdigit():
            return f"{prefix}.{final_number}"
    return ref_id


def remove_reference_prefix(text: str, ref_id: str) -> str:
    match = REFERENCE_RE.match(text)
    if not match or normalized_pdf_reference(match) != ref_id:
        return clean_content_block(text)
    return clean_content_block(text[match.end() :])


def iter_pdf_blocks(page: fitz.Page, page_number: int) -> Iterable[PdfBlock]:
    for raw_block in page.get_text("dict")["blocks"]:
        if raw_block.get("type") != 0:
            continue

        lines = raw_block.get("lines", [])
        spans = [span for line in lines for span in line.get("spans", [])]
        if not spans:
            continue

        line_texts: list[str] = []
        for line in lines:
            line_text = "".join(
                span.get("text", "") for span in line.get("spans", [])
            )
            if line_text.strip() in PDF_BULLET_CHARS:
                line_x0 = float(line.get("bbox", (BULLET_LEFT_X,))[0])
                level = round(
                    (line_x0 - BULLET_LEFT_X) / BULLET_INDENT_STEP_X
                )
                level = max(0, min(4, level))
                line_text = f"{BULLET_LEVEL_MARKER}{level}"
            line_texts.append(line_text)
        text = "\n".join(line_texts)
        # Some localized PDFs place control characters before a reference id.
        # Remove them before matching; the cleaning functions already discard
        # the same non-content characters from extracted cell values.
        text = text.replace("\x02", "").replace("\x07", "")
        for incorrect_reference, corrected_reference in REFERENCE_TEXT_REPLACEMENTS.items():
            text = text.replace(incorrect_reference, corrected_reference)
        first_span = next(
            (span for span in spans if span.get("text", "").strip()), spans[0]
        )
        x0, y0, x1, y1 = raw_block["bbox"]
        yield PdfBlock(
            page_number=page_number,
            x0=x0,
            y0=y0,
            x1=x1,
            y1=y1,
            text=text,
            first_font=first_span.get("font", ""),
            first_size=float(first_span.get("size", 0)),
        )


def is_body_candidate(block: PdfBlock) -> bool:
    return 84 <= block.x0 <= 190 and block.x1 <= 570 and block.y1 < 790


def extract_sidebar_content(
    document: fitz.Document,
    expected_categories: set[str],
    expected_functions: set[str],
) -> tuple[dict[str, str], dict[str, str], dict[str, str]]:
    missing_function_labels = expected_functions - DUTCH_FUNCTION_NAMES.keys()
    if missing_function_labels:
        raise ValueError(
            "Missing Dutch function labels: "
            + ", ".join(sorted(missing_function_labels))
        )
    category_names: dict[str, str] = {}
    category_descriptions: dict[str, str] = {}
    function_names = {
        ref_id: DUTCH_FUNCTION_NAMES[ref_id] for ref_id in expected_functions
    }

    for page_number, page in enumerate(document, start=1):
        blocks = list(iter_pdf_blocks(page, page_number))
        for category_ref in expected_categories - category_names.keys():
            marker = f"[{category_ref}]"
            category_block = next(
                (
                    block
                    for block in blocks
                    if marker in block.text
                    and 190 <= block.x0 <= 280
                    and 70 <= block.y0 <= 250
                ),
                None,
            )
            if category_block is None:
                continue

            raw_name, inline_description = category_block.text.split(marker, 1)
            category_name = clean_inline_text(raw_name)
            if category_name.isupper():
                category_name = category_name.capitalize()
            category_names[category_ref] = category_name

            inline_description = clean_inline_text(inline_description)
            if inline_description:
                category_descriptions[category_ref] = inline_description
            else:
                description_candidates = [
                    block
                    for block in blocks
                    if 280 <= block.x0 <= 380
                    and category_block.y0 - 35 <= block.y0 <= category_block.y1 + 25
                    and block.x1 <= 570
                ]
                if not description_candidates:
                    raise ValueError(
                        f"No sidebar description found for {category_ref} on PDF page "
                        f"{page_number}."
                    )
                description_block = max(
                    description_candidates, key=lambda block: len(block.text)
                )
                category_descriptions[category_ref] = clean_inline_text(
                    description_block.text
                )

    return function_names, category_names, category_descriptions


def extract_body_content(
    document: fitz.Document,
    expected_body_refs: list[str],
    allowed_missing_refs: set[str] | None = None,
) -> tuple[dict[str, str], dict[str, str]]:
    allowed_missing_refs = allowed_missing_refs or set()
    description_parts: dict[str, list[str]] = {
        ref_id: [] for ref_id in expected_body_refs
    }
    annotation_parts: dict[str, list[str]] = {
        ref_id: [] for ref_id in expected_body_refs if ref_id.count(".") == 2
    }

    expected_index = 0
    current_ref: str | None = None
    current_section = "description"
    skip_references = False

    def expected_position(ref_id: str) -> int | None:
        try:
            position = expected_body_refs.index(ref_id, expected_index)
        except ValueError:
            return None
        skipped = expected_body_refs[expected_index:position]
        return position if all(ref in allowed_missing_refs for ref in skipped) else None

    for page_number, page in enumerate(document, start=1):
        all_blocks = list(iter_pdf_blocks(page, page_number))
        blocks = sorted(
            (block for block in all_blocks if is_body_candidate(block)),
            key=lambda block: (block.y0, block.x0),
        )

        # Once the last expected control has been found, a blank separator page
        # marks the end of the framework. Without this stop, annexes and legal
        # notices would be appended to the final annotation.
        if (
            expected_index >= len(expected_body_refs)
            and current_ref is not None
            and not blocks
        ):
            break

        category_marker_on_page = any(
            re.search(r"[\[(][A-Z]{2}\.[A-Z]{2}[\])]", block.text)
            and 80 <= block.x0 <= 300
            for block in all_blocks
        )
        diagram_letter_count = sum(
            len(clean_inline_text(block.text)) == 1
            and clean_inline_text(block.text).isalpha()
            for block in all_blocks
        )
        sidebar_on_page = category_marker_on_page or diagram_letter_count >= 8
        first_new_ref_position: int | None = None
        for index, block in enumerate(blocks):
            match = REFERENCE_RE.match(block.text)
            if not match or expected_index >= len(expected_body_refs):
                continue
            position = expected_position(normalized_pdf_reference(match))
            if position is None:
                continue
            if not (
                block.first_font.endswith("-Bold")
                or block.first_font.endswith("-Semibold")
            ):
                continue
            first_new_ref_position = index
            break

        for index, block in enumerate(blocks):
            stripped_text = block.text.strip()
            if not stripped_text:
                continue

            match = REFERENCE_RE.match(block.text)
            new_ref_position = (
                expected_position(normalized_pdf_reference(match))
                if match is not None and expected_index < len(expected_body_refs)
                else None
            )
            is_expected_new_ref = new_ref_position is not None and (
                block.first_font.endswith("-Bold")
                or block.first_font.endswith("-Semibold")
            )
            if is_expected_new_ref:
                current_ref = expected_body_refs[new_ref_position]
                expected_index = new_ref_position + 1
                current_section = "description"
                skip_references = False
                remainder = remove_reference_prefix(block.text, current_ref)
                if remainder:
                    description_parts[current_ref].append(remainder)
                continue

            guidance_label = next(
                (
                    label
                    for label in GUIDANCE_LABELS
                    if stripped_text.casefold().startswith(label.casefold())
                ),
                None,
            )
            if guidance_label is not None:
                if current_ref is not None and current_ref.count(".") == 2:
                    current_section = "annotation"
                    skip_references = False
                    remainder = clean_content_block(
                        stripped_text[len(guidance_label) :]
                    )
                    if remainder:
                        annotation_parts[current_ref].append(remainder)
                continue

            if any(
                stripped_text.casefold().startswith(label.casefold())
                for label in REFERENCES_LABELS
            ):
                skip_references = True
                continue

            if current_ref is None or skip_references:
                continue
            if (
                sidebar_on_page
                and first_new_ref_position is not None
                and index < first_new_ref_position
            ):
                continue

            # Large display text belongs to divider pages, not to requirement
            # content. Header continuations use the same bold/semibold font as
            # the reference line and remain included.
            is_header_continuation = (
                current_section == "description"
                and (
                    block.first_font.endswith("-Bold")
                    or block.first_font.endswith("-Semibold")
                )
                and block.first_size <= 13.5
            )
            if (
                current_section == "description"
                and current_ref.count(".") == 2
                and not is_header_continuation
                and block.first_size <= BODY_TEXT_MAX_SIZE
            ):
                # Two controls in the official PDFs omit the explicit guidance
                # label. The change from semibold requirement text to regular
                # body text is the same visual boundary.
                current_section = "annotation"
            if block.first_size > BODY_TEXT_MAX_SIZE and not is_header_continuation:
                continue
            if stripped_text.startswith("Version 2025-"):
                continue

            cleaned = clean_content_block(block.text)
            if not cleaned:
                continue
            if current_section == "annotation" and current_ref in annotation_parts:
                annotation_parts[current_ref].append(cleaned)
            else:
                description_parts[current_ref].append(cleaned)

    missing_refs = [
        ref_id
        for ref_id in expected_body_refs[expected_index:]
        if ref_id not in allowed_missing_refs
    ]
    if missing_refs:
        preview = ", ".join(missing_refs[:10])
        raise ValueError(
            f"The PDF parser did not find {len(missing_refs)} expected ref_id values. "
            f"First missing values: {preview}"
        )

    descriptions = {
        ref_id: join_description_blocks(parts)
        for ref_id, parts in description_parts.items()
    }
    annotations = {
        ref_id: join_content_blocks(parts)
        for ref_id, parts in annotation_parts.items()
    }
    return descriptions, annotations


def extract_pdf(
    pdf_path: Path,
    expected_functions: set[str],
    expected_categories: set[str],
    expected_body_refs: list[str],
    allowed_missing_body_refs: set[str] | None = None,
    include_sidebar: bool = True,
) -> ExtractedTexts:
    if not pdf_path.is_file():
        raise FileNotFoundError(f"PDF not found: {pdf_path}")

    document = fitz.open(pdf_path)
    try:
        if include_sidebar:
            function_names, category_names, category_descriptions = extract_sidebar_content(
                document, expected_categories, expected_functions
            )
        else:
            function_names, category_names, category_descriptions = {}, {}, {}
        body_descriptions, annotations = extract_body_content(
            document, expected_body_refs, allowed_missing_body_refs
        )
    finally:
        document.close()

    missing_functions = expected_functions - function_names.keys()
    missing_categories = expected_categories - category_names.keys()
    missing_category_descriptions = expected_categories - category_descriptions.keys()
    if include_sidebar and (
        missing_functions or missing_categories or missing_category_descriptions
    ):
        raise ValueError(
            "Incomplete sidebar extraction: "
            f"functions={sorted(missing_functions)}, "
            f"category names={sorted(missing_categories)}, "
            f"category descriptions={sorted(missing_category_descriptions)}"
        )

    names = {**function_names, **category_names}
    descriptions = {**category_descriptions, **body_descriptions}
    return ExtractedTexts(
        names=names,
        descriptions=descriptions,
        annotations=annotations,
    )


def load_workbook_rows(workbook_path: Path) -> list[WorkbookRow]:
    if not workbook_path.is_file():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        if SOURCE_SHEET_NAME not in workbook.sheetnames:
            raise ValueError(
                f"Sheet {SOURCE_SHEET_NAME!r} not found in {workbook_path.name}."
            )
        sheet = workbook[SOURCE_SHEET_NAME]
        headers = {
            cell.value: column_number
            for column_number, cell in enumerate(sheet[1], start=1)
        }
        required_headers = {
            "depth",
            "ref_id",
            "name",
            "description",
            "annotation",
        }
        missing_headers = required_headers - headers.keys()
        if missing_headers:
            raise ValueError(f"Missing workbook headers: {sorted(missing_headers)}")

        rows: list[WorkbookRow] = []
        for row_number in range(2, sheet.max_row + 1):
            ref_id = sheet.cell(row_number, headers["ref_id"]).value
            depth = sheet.cell(row_number, headers["depth"]).value
            if not isinstance(ref_id, str) or not isinstance(depth, int):
                raise ValueError(
                    f"Invalid ref_id/depth at {SOURCE_SHEET_NAME}!{row_number}."
                )
            rows.append(
                WorkbookRow(
                    row_number=row_number,
                    ref_id=ref_id.strip(),
                    depth=depth,
                    name=sheet.cell(row_number, headers["name"]).value,
                    description=sheet.cell(
                        row_number, headers["description"]
                    ).value,
                    annotation=sheet.cell(
                        row_number, headers["annotation"]
                    ).value,
                )
            )
        return rows
    finally:
        workbook.close()


def comparison_text(value: str | None) -> str:
    if not value:
        return ""
    value = unicodedata.normalize("NFKC", value).casefold()
    value = value.replace("organisation", "organization")
    value = re.sub(r"[^\w]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def similarity(left: str | None, right: str | None) -> float:
    return SequenceMatcher(
        None, comparison_text(left), comparison_text(right), autojunk=False
    ).ratio()


def pdf_search_text(pdf_path: Path) -> str:
    document = fitz.open(pdf_path)
    try:
        text = "\n".join(page.get_text("text") for page in document)
    finally:
        document.close()
    return comparison_text(text)


def inspect_optional_pdf_sources(
    workbook_path: Path, english_pdf: Path, dutch_pdf: Path
) -> dict[str, str] | None:
    """Confirm which optional sheet values are actually sourced by the PDFs."""
    english_text = pdf_search_text(english_pdf)
    dutch_text = pdf_search_text(dutch_pdf)
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=False)
    try:
        controls_sheet = workbook[CONTROLS_SOURCE_SHEET_NAME]
        control_headers = {
            cell.value: index for index, cell in enumerate(controls_sheet[1], start=1)
        }
        control_names = [
            controls_sheet.cell(row_number, control_headers["name"]).value
            for row_number in range(2, controls_sheet.max_row + 1)
        ]
        control_name_matches = sum(
            bool(name) and comparison_text(name) in english_text
            for name in control_names
        )

        scores_sheet = workbook[SCORES_SOURCE_SHEET_NAME]
        score_headers = {
            cell.value: index for index, cell in enumerate(scores_sheet[1], start=1)
        }
        score_source_texts = [
            scores_sheet.cell(row_number, score_headers[field_name]).value
            for row_number in range(2, scores_sheet.max_row + 1)
            for field_name in ("description", "description_doc")
        ]
        score_matches = sum(
            bool(value) and comparison_text(value) in english_text
            for value in score_source_texts
        )
    finally:
        workbook.close()

    print(
        "Controls name source check: "
        f"{control_name_matches}/{len(control_names)} short names found in the "
        f"English PDF; {localized_column('name')} will remain empty."
    )
    print(
        "Scores source check: "
        f"{score_matches}/{len(score_source_texts)} score descriptions found in "
        "the English PDF; scores_content will not be copied or translated."
    )

    required_ig_terms = set(IG_REQUIRED_TERMS)
    missing_ig_terms = {
        term for term in required_ig_terms if comparison_text(term) not in dutch_text
    }
    if missing_ig_terms:
        print(
            f"IG source check: missing official {LANGUAGE_NAME} terms "
            f"{sorted(missing_ig_terms)}; IG_content will not be translated."
        )
        return None

    print(
        f"IG source check passed: {sorted(required_ig_terms)} are present in "
        f"the {LANGUAGE_NAME} PDF."
    )
    return IG_DUTCH_NAMES


def validate_english_extraction(
    rows: list[WorkbookRow],
    extracted: ExtractedTexts,
    allowed_missing_refs: set[str] | None = None,
) -> None:
    allowed_missing_refs = allowed_missing_refs or set()
    comparisons: list[tuple[float, str, str]] = []
    missing: list[str] = []

    for row in rows:
        if row.depth < 3:
            continue
        for field_name in ("name", "description", "annotation"):
            workbook_value = getattr(row, field_name)
            if not workbook_value:
                continue
            extracted_value = getattr(extracted, f"{field_name}s").get(row.ref_id)
            if not extracted_value:
                if row.ref_id in allowed_missing_refs:
                    continue
                missing.append(f"{row.ref_id}:{field_name}")
                continue
            comparisons.append(
                (
                    similarity(workbook_value, extracted_value),
                    row.ref_id,
                    field_name,
                )
            )

    if missing:
        raise ValueError(
            "English extraction is incomplete. Missing fields: "
            + ", ".join(missing[:20])
        )
    if not comparisons:
        raise ValueError("No English workbook fields were available for validation.")

    low_matches = [comparison for comparison in comparisons if comparison[0] < 0.72]
    median_score = statistics.median(score for score, _, _ in comparisons)
    if median_score < 0.92 or len(low_matches) > 5:
        worst = ", ".join(
            f"{ref_id}:{field_name}={score:.3f}"
            for score, ref_id, field_name in sorted(comparisons)[:10]
        )
        raise ValueError(
            "English PDF validation failed: "
            f"median similarity={median_score:.3f}, "
            f"fields below 0.72={len(low_matches)}. Worst matches: {worst}"
        )

    print(
        "English validation passed: "
        f"{len(comparisons)} fields, median similarity {median_score:.3f}, "
        f"{len(low_matches)} fields below 0.72."
    )


def validate_dutch_coverage(
    rows: list[WorkbookRow], extracted: ExtractedTexts
) -> None:
    missing: list[str] = []
    for row in rows:
        for field_name in ("name", "description", "annotation"):
            if not getattr(row, field_name):
                continue
            if not getattr(extracted, f"{field_name}s").get(row.ref_id):
                missing.append(f"{row.ref_id}:{field_name}")
    if missing:
        raise ValueError(
            f"{LANGUAGE_NAME} extraction is incomplete. Missing fields: "
            + ", ".join(missing[:20])
        )

    if LANGUAGE_CODE == "nl":
        actual_function_names = {
            ref_id: extracted.names.get(ref_id)
            for ref_id in DUTCH_FUNCTION_NAMES
        }
        if actual_function_names != DUTCH_FUNCTION_NAMES:
            raise ValueError(
                "Unexpected Dutch function names: "
                f"{actual_function_names}; expected {DUTCH_FUNCTION_NAMES}."
            )

    if LANGUAGE_CODE == "fr" and "PR.AT-01.1" in extracted.annotations:
        pr_at_bullet_count = len(
            re.findall(r"(?m)^-\s+", extracted.annotations["PR.AT-01.1"])
        )
        if pr_at_bullet_count != 8:
            raise ValueError(
                "French PR.AT-01.1 annotation must contain 8 bullets, "
                f"got {pr_at_bullet_count}."
            )

    descriptions_with_line_breaks = [
        ref_id
        for ref_id, description in extracted.descriptions.items()
        if "\n" in description or "\r" in description
    ]
    if descriptions_with_line_breaks:
        raise ValueError(
            f"Unexpected line breaks in {LANGUAGE_NAME} descriptions: "
            + ", ".join(descriptions_with_line_breaks[:20])
        )

    unresolved_markers = [
        f"{field_name}:{ref_id}"
        for field_name in ("names", "descriptions", "annotations")
        for ref_id, value in getattr(extracted, field_name).items()
        if BULLET_LEVEL_MARKER in value or BULLET_SPLIT_MARKER in value
    ]
    if unresolved_markers:
        raise ValueError(
            f"Unresolved bullet markers in {LANGUAGE_NAME} extraction: "
            + ", ".join(unresolved_markers[:20])
        )

    diagram_bleed_refs = [
        ref_id
        for ref_id, annotation in extracted.annotations.items()
        if SINGLE_LETTER_BLEED_RE.search(annotation)
    ]
    if diagram_bleed_refs:
        raise ValueError(
            f"Diagram text leaked into {LANGUAGE_NAME} annotations: "
            + ", ".join(diagram_bleed_refs[:20])
        )

    back_matter_refs = [
        ref_id
        for ref_id, annotation in extracted.annotations.items()
        if any(marker.casefold() in annotation.casefold() for marker in BACK_MATTER_MARKERS)
    ]
    if back_matter_refs:
        raise ValueError(
            f"Back matter leaked into {LANGUAGE_NAME} annotations: "
            + ", ".join(back_matter_refs[:20])
        )

    source_bullet_refs = [
        ref_id
        for ref_id, annotation in extracted.annotations.items()
        if any(char in annotation for char in PDF_BULLET_CHARS)
    ]
    if source_bullet_refs:
        raise ValueError(
            f"PDF bullet characters remain in {LANGUAGE_NAME} annotations: "
            + ", ".join(source_bullet_refs[:20])
        )

    invalid_bullet_indents: list[str] = []
    for ref_id, annotation in extracted.annotations.items():
        for line in annotation.splitlines():
            stripped = line.lstrip(" ")
            if not stripped.startswith(f"{BULLET_CHAR}  "):
                continue
            indentation = len(line) - len(stripped)
            if indentation % BULLET_INDENT_SPACES:
                invalid_bullet_indents.append(ref_id)
                break
    if invalid_bullet_indents:
        raise ValueError(
            f"Invalid {LANGUAGE_NAME} bullet indentation: "
            + ", ".join(invalid_bullet_indents[:20])
        )

    if LANGUAGE_CODE == "nl":
        required_annotation_texts = {
            "GV.SC-03.1": (
                "Definieer criteria voor materialiteit",
                "Zorg voor een formeel escalatiepad",
            ),
            "PR.AA-03.2": (
                "Als individuele accounts niet haalbaar zijn",
                "Veilige externe toegang tot OT-systemen",
            ),
            "PR.AA-05.7": (
                "In OT-omgevingen zou geprivilegieerde toegang",
                "Afstemmen op ENISA-richtlijnen",
            ),
            "PR.AA-06.3": (
                "Kritieke zones zouden geïdentificeerd kunnen worden",
            ),
        }
        missing_required_texts = [
            f"{ref_id}:{required_text}"
            for ref_id, required_texts in required_annotation_texts.items()
            for required_text in required_texts
            if required_text not in extracted.annotations.get(ref_id, "")
        ]
        if missing_required_texts:
            raise ValueError(
                "Confirmed Dutch annotation text is missing: "
                + ", ".join(missing_required_texts)
            )

    target_annotation = extracted.annotations.get("DE.CM-01.2", "")
    target_indentations = {
        len(line) - len(line.lstrip(" "))
        for line in target_annotation.splitlines()
        if line.lstrip(" ").startswith(f"{BULLET_CHAR}  ")
    }
    expected_indentations = EXPECTED_TARGET_BULLET_INDENTATIONS
    if not expected_indentations.issubset(target_indentations):
        raise ValueError(
            "DE.CM-01.2 bullet hierarchy was not preserved. "
            f"Expected {sorted(expected_indentations)}, "
            f"found {sorted(target_indentations)}."
        )


def copy_column_format(source_sheet, target_sheet, source_column: int, target_column: int) -> None:
    source_letter = openpyxl.utils.get_column_letter(source_column)
    target_letter = openpyxl.utils.get_column_letter(target_column)
    source_dimension = source_sheet.column_dimensions[source_letter]
    target_dimension = target_sheet.column_dimensions[target_letter]
    target_dimension.width = source_dimension.width
    target_dimension.hidden = source_dimension.hidden
    target_dimension.bestFit = source_dimension.bestFit
    target_dimension.outlineLevel = source_dimension.outlineLevel

    for row_number in range(1, target_sheet.max_row + 1):
        source_cell = source_sheet.cell(row_number, source_column)
        target_cell = target_sheet.cell(row_number, target_column)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.protection = copy(source_cell.protection)


def copy_cell_format(source_cell, target_cell) -> None:
    if source_cell.has_style:
        target_cell._style = copy(source_cell._style)
    if source_cell.number_format:
        target_cell.number_format = source_cell.number_format
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.protection = copy(source_cell.protection)


def add_localized_meta_rows(workbook, sheet_name: str, description: str) -> None:
    sheet = workbook[sheet_name]
    key_rows = {
        str(sheet.cell(row_number, 1).value).strip(): row_number
        for row_number in range(1, sheet.max_row + 1)
        if sheet.cell(row_number, 1).value not in (None, "")
    }
    missing = {"name", "description"} - key_rows.keys()
    if missing:
        raise ValueError(f"Missing metadata keys in {sheet_name}: {sorted(missing)}")

    localized_values = {
        localized_column("name"): sheet.cell(key_rows["name"], 2).value,
        localized_column("description"): description,
    }
    existing = set(localized_values) & key_rows.keys()
    if existing:
        raise ValueError(
            f"{sheet_name} already contains {LANGUAGE_NAME} metadata: "
            f"{sorted(existing)}"
        )

    for base_key, (localized_key, localized_value) in zip(
        ("name", "description"), localized_values.items()
    ):
        source_row = key_rows[base_key]
        target_row = sheet.max_row + 1
        for column_number in range(1, sheet.max_column + 1):
            copy_cell_format(
                sheet.cell(source_row, column_number),
                sheet.cell(target_row, column_number),
            )
        sheet.cell(target_row, 1).value = localized_key
        sheet.cell(target_row, 2).value = localized_value
        source_height = sheet.row_dimensions[source_row].height
        if source_height is not None:
            sheet.row_dimensions[target_row].height = source_height


def add_localized_columns(
    workbook, sheet_name: str, base_fields: tuple[str, ...]
):
    sheet = workbook[sheet_name]
    headers = {
        cell.value: column_number
        for column_number, cell in enumerate(sheet[1], start=1)
    }
    missing_fields = set(base_fields) - headers.keys()
    if missing_fields:
        raise ValueError(
            f"Missing fields in {sheet_name}: {sorted(missing_fields)}"
        )

    translated_headers = tuple(localized_column(field_name) for field_name in base_fields)
    existing_translated = set(translated_headers) & headers.keys()
    if existing_translated:
        raise ValueError(
            f"{sheet_name} already contains {LANGUAGE_NAME} columns: "
            f"{sorted(existing_translated)}"
        )

    translated_columns: dict[str, int] = {}
    first_target_column = sheet.max_column + 1
    for offset, (base_field, translated_header) in enumerate(
        zip(base_fields, translated_headers)
    ):
        target_column = first_target_column + offset
        copy_column_format(sheet, sheet, headers[base_field], target_column)
        sheet.cell(1, target_column).value = translated_header
        translated_columns[base_field] = target_column

    if sheet.auto_filter.ref:
        sheet.auto_filter.ref = (
            f"A1:{openpyxl.utils.get_column_letter(sheet.max_column)}"
            f"{sheet.max_row}"
        )
    return sheet, headers, translated_columns


def write_output_workbook(
    input_path: Path,
    output_path: Path,
    rows: list[WorkbookRow],
    dutch: ExtractedTexts,
    ig_dutch_names: dict[str, str] | None,
    meta_description: str,
    force: bool,
) -> None:
    if output_path.resolve() == input_path.resolve():
        raise ValueError("The output path must differ from the source workbook path.")
    if output_path.exists() and not force:
        raise FileExistsError(
            f"Output already exists: {output_path}. Use --force to replace it."
        )

    workbook = openpyxl.load_workbook(input_path)
    try:
        for meta_sheet_name in META_SHEET_NAMES:
            add_localized_meta_rows(
                workbook, meta_sheet_name, meta_description
            )

        target_sheet, _, translated_columns = add_localized_columns(
            workbook,
            SOURCE_SHEET_NAME,
            ("name", "description", "annotation"),
        )

        for row in rows:
            target_sheet.cell(row.row_number, translated_columns["name"]).value = (
                dutch.names.get(row.ref_id) if row.name else None
            )
            target_sheet.cell(
                row.row_number, translated_columns["description"]
            ).value = (
                dutch.descriptions.get(row.ref_id) if row.description else None
            )
            target_sheet.cell(
                row.row_number, translated_columns["annotation"]
            ).value = (
                dutch.annotations.get(row.ref_id) if row.annotation else None
            )

        controls_target, controls_headers, controls_translated_columns = (
            add_localized_columns(
                workbook,
                CONTROLS_SOURCE_SHEET_NAME,
                ("name", "description", "annotation"),
            )
        )
        for row_number in range(2, controls_target.max_row + 1):
            ref_id = controls_target.cell(
                row_number, controls_headers["ref_id"]
            ).value
            if not isinstance(ref_id, str):
                raise ValueError(
                    f"Invalid ref_id in {CONTROLS_SOURCE_SHEET_NAME}!{row_number}."
                )
            ref_id = ref_id.strip()
            controls_target.cell(
                row_number, controls_translated_columns["name"]
            ).value = None
            controls_target.cell(
                row_number, controls_translated_columns["description"]
            ).value = dutch.descriptions.get(ref_id)
            controls_target.cell(
                row_number, controls_translated_columns["annotation"]
            ).value = dutch.annotations.get(ref_id)

        if ig_dutch_names is not None:
            ig_target, ig_headers, ig_translated_columns = (
                add_localized_columns(
                    workbook,
                    IG_SOURCE_SHEET_NAME,
                    ("name", "description"),
                )
            )
            for row_number in range(2, ig_target.max_row + 1):
                ref_id = ig_target.cell(row_number, ig_headers["ref_id"]).value
                if ref_id not in ig_dutch_names:
                    raise ValueError(
                        f"No {LANGUAGE_NAME} IG label available for ref_id {ref_id!r}."
                    )
                ig_target.cell(
                    row_number, ig_translated_columns["name"]
                ).value = ig_dutch_names[ref_id]
                ig_target.cell(
                    row_number, ig_translated_columns["description"]
                ).value = None

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
    finally:
        workbook.close()


def verify_saved_workbook(
    input_path: Path,
    output_path: Path,
    rows: list[WorkbookRow],
    dutch: ExtractedTexts,
    ig_dutch_names: dict[str, str] | None,
    meta_description: str,
) -> None:
    original = openpyxl.load_workbook(input_path, read_only=False, data_only=False)
    workbook = openpyxl.load_workbook(output_path, read_only=False, data_only=False)
    try:
        if workbook.sheetnames != original.sheetnames:
            raise ValueError(
                "The copied workbook must preserve the original sheet list. "
                f"Expected {original.sheetnames}, got {workbook.sheetnames}."
            )
        unexpected_sheets = FORBIDDEN_LOCALIZED_SHEET_NAMES & set(
            workbook.sheetnames
        )
        if unexpected_sheets:
            raise ValueError(
                f"Unexpected localized sheets: {sorted(unexpected_sheets)}"
            )

        modified_sheets = {
            SOURCE_SHEET_NAME,
            CONTROLS_SOURCE_SHEET_NAME,
        }
        if ig_dutch_names is not None:
            modified_sheets.add(IG_SOURCE_SHEET_NAME)

        for sheet_name in workbook.sheetnames:
            original_sheet = original[sheet_name]
            output_sheet = workbook[sheet_name]
            expected_row_count = original_sheet.max_row + (
                2 if sheet_name in META_SHEET_NAMES else 0
            )
            if output_sheet.max_row != expected_row_count:
                raise ValueError(
                    f"Unexpected row count in {sheet_name}: "
                    f"expected {expected_row_count}, got {output_sheet.max_row}."
                )
            if sheet_name in META_SHEET_NAMES:
                for row_number in range(1, original_sheet.max_row + 1):
                    for column_number in range(1, original_sheet.max_column + 1):
                        if original_sheet.cell(
                            row_number, column_number
                        ).value != output_sheet.cell(row_number, column_number).value:
                            raise ValueError(
                                f"Original value changed in {sheet_name}!"
                                f"{output_sheet.cell(row_number, column_number).coordinate}."
                            )
                continue
            if sheet_name not in modified_sheets:
                original_values = list(original_sheet.iter_rows(values_only=True))
                output_values = list(output_sheet.iter_rows(values_only=True))
                if output_values != original_values:
                    raise ValueError(f"Unexpected changes in {sheet_name}.")
                continue
            for row_number in range(1, original_sheet.max_row + 1):
                for column_number in range(1, original_sheet.max_column + 1):
                    if original_sheet.cell(row_number, column_number).value != output_sheet.cell(
                        row_number, column_number
                    ).value:
                        raise ValueError(
                            f"Original value changed in {sheet_name}!"
                            f"{output_sheet.cell(row_number, column_number).coordinate}."
                        )

        for meta_sheet_name in META_SHEET_NAMES:
            original_meta = original[meta_sheet_name]
            output_meta = workbook[meta_sheet_name]
            original_key_rows = {
                str(original_meta.cell(row_number, 1).value).strip(): row_number
                for row_number in range(1, original_meta.max_row + 1)
                if original_meta.cell(row_number, 1).value not in (None, "")
            }
            output_key_rows = {
                str(output_meta.cell(row_number, 1).value).strip(): row_number
                for row_number in range(1, output_meta.max_row + 1)
                if output_meta.cell(row_number, 1).value not in (None, "")
            }
            expected_meta = {
                localized_column("name"): original_meta.cell(
                    original_key_rows["name"], 2
                ).value,
                localized_column("description"): meta_description,
            }
            if [
                output_meta.cell(output_meta.max_row - 1, 1).value,
                output_meta.cell(output_meta.max_row, 1).value,
            ] != list(expected_meta):
                raise ValueError(
                    f"Unexpected {LANGUAGE_NAME} metadata order in {meta_sheet_name}."
                )
            for localized_key, expected_value in expected_meta.items():
                target_row = output_key_rows.get(localized_key)
                if target_row is None or output_meta.cell(target_row, 2).value != expected_value:
                    raise ValueError(
                        f"Unexpected value for {meta_sheet_name}:{localized_key}."
                    )
                base_key = localized_key.split("[", 1)[0]
                source_row = original_key_rows[base_key]
                for column_number in range(1, original_meta.max_column + 1):
                    if original_meta.cell(source_row, column_number).style_id != output_meta.cell(
                        target_row, column_number
                    ).style_id:
                        raise ValueError(
                            f"Metadata style mismatch in {meta_sheet_name}:{localized_key}."
                        )
            print(f"Metadata verification passed: {meta_sheet_name}.")

        requirements_sheet = workbook[SOURCE_SHEET_NAME]
        requirements_headers = [cell.value for cell in requirements_sheet[1]]
        if requirements_headers[-3:] != list(TRANSLATED_COLUMNS):
            raise ValueError(
                f"Unexpected {LANGUAGE_NAME} columns in requirements_content: "
                f"{requirements_headers[-3:]}"
            )
        requirements_columns = {
            cell.value: index
            for index, cell in enumerate(requirements_sheet[1], start=1)
        }
        requirements_populated = {
            header: sum(
                requirements_sheet.cell(
                    row.row_number, requirements_columns[header]
                ).value
                not in (None, "")
                for row in rows
            )
            for header in TRANSLATED_COLUMNS
        }
        requirements_expected = {
            localized_column("name"): sum(bool(row.name) for row in rows),
            localized_column("description"): sum(bool(row.description) for row in rows),
            localized_column("annotation"): sum(bool(row.annotation) for row in rows),
        }
        if requirements_populated != requirements_expected:
            raise ValueError(
                f"Unexpected requirements counts: {requirements_populated}; "
                f"expected {requirements_expected}."
            )
        for row in rows:
            expected_values = {
                localized_column("name"): dutch.names.get(row.ref_id) if row.name else None,
                localized_column("description"): (
                    dutch.descriptions.get(row.ref_id) if row.description else None
                ),
                localized_column("annotation"): (
                    dutch.annotations.get(row.ref_id) if row.annotation else None
                ),
            }
            for header, expected_value in expected_values.items():
                if requirements_sheet.cell(
                    row.row_number, requirements_columns[header]
                ).value != expected_value:
                    raise ValueError(
                        f"Unexpected value for {row.ref_id}:{header}."
                    )
        print(f"Requirements verification passed: {requirements_populated}.")

        controls_sheet = workbook[CONTROLS_SOURCE_SHEET_NAME]
        controls_headers = [cell.value for cell in controls_sheet[1]]
        if controls_headers[-3:] != list(TRANSLATED_COLUMNS):
            raise ValueError(
                f"Unexpected {LANGUAGE_NAME} columns in controls_content: "
                f"{controls_headers[-3:]}"
            )
        controls_columns = {
            cell.value: index
            for index, cell in enumerate(controls_sheet[1], start=1)
        }
        controls_populated = {
            header: sum(
                controls_sheet.cell(row_number, controls_columns[header]).value
                not in (None, "")
                for row_number in range(2, controls_sheet.max_row + 1)
            )
            for header in TRANSLATED_COLUMNS
        }
        controls_expected = {
            localized_column("name"): 0,
            localized_column("description"): sum(
                bool(controls_sheet.cell(row_number, controls_columns["description"]).value)
                for row_number in range(2, controls_sheet.max_row + 1)
            ),
            localized_column("annotation"): sum(
                bool(controls_sheet.cell(row_number, controls_columns["annotation"]).value)
                for row_number in range(2, controls_sheet.max_row + 1)
            ),
        }
        if controls_populated != controls_expected:
            raise ValueError(
                f"Unexpected controls counts: {controls_populated}; "
                f"expected {controls_expected}."
            )
        for row_number in range(2, controls_sheet.max_row + 1):
            ref_id = controls_sheet.cell(
                row_number, controls_columns["ref_id"]
            ).value
            if not isinstance(ref_id, str):
                raise ValueError(
                    f"Invalid ref_id in {CONTROLS_SOURCE_SHEET_NAME}!{row_number}."
                )
            ref_id = ref_id.strip()
            if controls_sheet.cell(
                row_number, controls_columns[localized_column("description")]
            ).value != dutch.descriptions.get(ref_id):
                raise ValueError(
                    f"Unexpected controls {localized_column('description')} for {ref_id}."
                )
            if controls_sheet.cell(
                row_number, controls_columns[localized_column("annotation")]
            ).value != dutch.annotations.get(ref_id):
                raise ValueError(
                    f"Unexpected controls {localized_column('annotation')} for {ref_id}."
                )
        print(f"Controls verification passed: {controls_populated}.")

        if ig_dutch_names is not None:
            ig_sheet = workbook[IG_SOURCE_SHEET_NAME]
            ig_headers = [cell.value for cell in ig_sheet[1]]
            ig_translated_headers = [
                localized_column("name"),
                localized_column("description"),
            ]
            if ig_headers[-2:] != ig_translated_headers:
                raise ValueError(
                    f"Unexpected {LANGUAGE_NAME} columns in IG_content: {ig_headers[-2:]}"
                )
            ig_columns = {
                cell.value: index
                for index, cell in enumerate(ig_sheet[1], start=1)
            }
            ig_populated = {
                header: sum(
                    ig_sheet.cell(row_number, ig_columns[header]).value
                    not in (None, "")
                    for row_number in range(2, ig_sheet.max_row + 1)
                )
                for header in ig_translated_headers
            }
            ig_expected = {
                localized_column("name"): ig_sheet.max_row - 1,
                localized_column("description"): 0,
            }
            if ig_populated != ig_expected:
                raise ValueError(
                    f"Unexpected IG counts: {ig_populated}; expected {ig_expected}."
                )
            for row_number in range(2, ig_sheet.max_row + 1):
                ref_id = ig_sheet.cell(row_number, ig_columns["ref_id"]).value
                actual_name = ig_sheet.cell(
                    row_number, ig_columns[localized_column("name")]
                ).value
                if actual_name != ig_dutch_names.get(ref_id):
                    raise ValueError(
                        f"Unexpected IG {localized_column('name')} for {ref_id!r}."
                    )
            print(f"IG verification passed: {ig_populated}.")

        localized_style_checks = {
            SOURCE_SHEET_NAME: ("name", "description", "annotation"),
            CONTROLS_SOURCE_SHEET_NAME: ("name", "description", "annotation"),
        }
        if ig_dutch_names is not None:
            localized_style_checks[IG_SOURCE_SHEET_NAME] = ("name", "description")
        for sheet_name, base_fields in localized_style_checks.items():
            sheet = workbook[sheet_name]
            headers = {
                cell.value: index for index, cell in enumerate(sheet[1], start=1)
            }
            for base_field in base_fields:
                translated_header = localized_column(base_field)
                source_column = headers[base_field]
                target_column = headers[translated_header]
                if sheet.cell(1, source_column).style_id != sheet.cell(
                    1, target_column
                ).style_id:
                    raise ValueError(
                        f"Header style mismatch in {sheet_name}:{translated_header}."
                    )
                source_letter = openpyxl.utils.get_column_letter(source_column)
                target_letter = openpyxl.utils.get_column_letter(target_column)
                if (
                    sheet.column_dimensions[source_letter].width
                    != sheet.column_dimensions[target_letter].width
                ):
                    raise ValueError(
                        f"Column width mismatch in {sheet_name}:{translated_header}."
                    )

        formula_errors = []
        error_values = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A"}
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == "e" or cell.value in error_values:
                        formula_errors.append(f"{sheet.title}!{cell.coordinate}")
        if formula_errors:
            raise ValueError(
                "Formula errors found in the saved workbook: "
                + ", ".join(formula_errors[:20])
            )
        print("Sheet-list, formatting and formula-error verification passed.")
    finally:
        workbook.close()
        original.close()


def main() -> int:
    args = parse_arguments()
    rows = load_workbook_rows(args.input)

    expected_functions = {
        row.ref_id for row in rows if row.depth == 1 and FUNCTION_RE.match(row.ref_id)
    }
    expected_categories = {
        row.ref_id for row in rows if row.depth == 2 and CATEGORY_RE.match(row.ref_id)
    }
    expected_body_refs = [
        row.ref_id for row in rows if row.depth in (3, 4)
    ]

    print(f"Extracting and validating {args.english_pdf.name} ...")
    english = extract_pdf(
        args.english_pdf,
        expected_functions,
        expected_categories,
        expected_body_refs,
        ENGLISH_PDF_ALLOWED_MISSING_REFS,
        include_sidebar=False,
    )
    validate_english_extraction(
        rows, english, ENGLISH_PDF_ALLOWED_MISSING_REFS
    )

    print(f"Extracting {args.dutch_pdf.name} ...")
    dutch = extract_pdf(
        args.dutch_pdf,
        expected_functions,
        expected_categories,
        expected_body_refs,
    )
    validate_dutch_coverage(rows, dutch)
    meta_description = build_localized_meta_description(
        args.input, args.english_pdf, args.dutch_pdf
    )
    print(
        f"Introduction metadata extraction passed for {LANGUAGE_NAME}."
    )

    ig_dutch_names = inspect_optional_pdf_sources(
        args.input, args.english_pdf, args.dutch_pdf
    )
    write_output_workbook(
        args.input,
        args.output,
        rows,
        dutch,
        ig_dutch_names,
        meta_description,
        args.force,
    )
    verify_saved_workbook(
        args.input,
        args.output,
        rows,
        dutch,
        ig_dutch_names,
        meta_description,
    )
    print(f"Created: {args.output.resolve()}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (FileNotFoundError, FileExistsError, ValueError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1) from exc
