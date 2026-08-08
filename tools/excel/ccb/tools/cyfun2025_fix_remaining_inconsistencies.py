#!/usr/bin/env python3
"""Fix the final audited CyFun 2025 workbook inconsistencies.

The script never modifies its input workbook. It plans a small, explicit set of
audited replacements, copies the input, applies only those replacements, and
then verifies that every other cell and style stayed unchanged.
"""

from __future__ import annotations

import argparse
from collections import Counter
from copy import copy
from dataclasses import dataclass
import hashlib
from pathlib import Path
import shutil
import sys

import openpyxl
from openpyxl.styles import PatternFill


CONTENT_SHEETS = ("requirements_content", "controls_content")
ORANGE_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFFFA500",
    bgColor="FFFFA500",
)


@dataclass(frozen=True)
class TextPatch:
    ref_id: str
    column: str
    replacements: tuple[tuple[str, str], ...]


@dataclass(frozen=True)
class RefPatch:
    sheet: str
    ref_id: str
    old_value: str


@dataclass(frozen=True)
class CellChange:
    sheet: str
    coordinate: str
    ref_id: str
    field: str
    before: str
    after: str


TEXT_PATCHES = (
    TextPatch(
        "PR.AA-05.3",
        "annotation[fr]",
        (
            (
                "\nConsidérations spécifiques à l’OT Dans les environnements OT,",
                "\n-  Considérations spécifiques à l’OT\n"
                "  Dans les environnements OT,",
            ),
            ("voir PR. AA-01.1", "voir PR.AA-01.1"),
        ),
    ),
    TextPatch(
        "PR.PS-01.1",
        "annotation[fr]",
        (
            (
                "-  Systèmes gérés par le cloud et les fournisseurs Pour les "
                "systèmes gérés par des tiers,",
                "-  Systèmes gérés par le cloud et les fournisseurs\n"
                "  Pour les systèmes gérés par des tiers,",
            ),
        ),
    ),
    TextPatch(
        "PR.PS-01.1",
        "annotation[nl]",
        (("configuratieen wijzigingsbeheer", "configuratie- en wijzigingsbeheer"),),
    ),
    TextPatch(
        "GV.SC-01.1",
        "annotation[nl]",
        (
            (
                "Opleiding en bewustmaking Medewerkers opleiden:",
                "Opleiding en bewustmaking\n-  Medewerkers opleiden:",
            ),
        ),
    ),
    TextPatch(
        "GV.SC-01.1",
        "annotation",
        (
            (
                "Training and Awareness \nEducate Employees:",
                "Training and Awareness \n-  Educate Employees:",
            ),
        ),
    ),
    TextPatch(
        "PR.IR-03.1",
        "annotation[fr]",
        (("GV. OC-04.3", "GV.OC-04.3"),),
    ),
    TextPatch(
        "ID.IM-04.1",
        "annotation",
        (("_x0008_ These should address", "These should address"),),
    ),
)


REF_PATCHES = (
    RefPatch("requirements_content", "ID.AM-08.7", "ID.AM-08.7  "),
    RefPatch("requirements_content", "ID.AM-08.10", "ID.AM-08.10  "),
    RefPatch("requirements_content", "ID.AM-08.13", "ID.AM-08.13 "),
    RefPatch("requirements_content", "RS.MI", " RS.MI"),
    RefPatch("controls_content", "ID.AM-08.7", "ID.AM-08.7  "),
    RefPatch("controls_content", "ID.AM-08.10", "ID.AM-08.10  "),
    RefPatch("controls_content", "ID.AM-08.13", "ID.AM-08.13 "),
)


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Copy a CyFun 2025 workbook and fix only the final audited "
            "text, line-break, and ref_id inconsistencies."
        )
    )
    parser.add_argument(
        "--input",
        type=Path,
        required=True,
        help="Workbook to copy and correct.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        help="Output workbook. Defaults to <input_stem>_remaining_fixed.xlsx.",
    )
    parser.add_argument(
        "--highlight",
        action="store_true",
        help="Highlight every newly corrected cell with a solid orange fill.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Report planned changes without creating a copy.",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Allow replacing an existing output workbook.",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="List every changed cell.",
    )
    args = parser.parse_args()
    if args.output is None:
        args.output = args.input.with_name(
            f"{args.input.stem}_remaining_fixed{args.input.suffix}"
        )
    return args


def file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def header_map(sheet) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in sheet[1]
        if cell.value not in (None, "")
    }


def find_ref_row(sheet, ref_column: int, ref_id: str) -> int:
    matches = [
        row
        for row in range(2, sheet.max_row + 1)
        if isinstance(sheet.cell(row, ref_column).value, str)
        and sheet.cell(row, ref_column).value.strip() == ref_id
    ]
    if len(matches) != 1:
        raise ValueError(
            f"Expected one {ref_id} row in {sheet.title}, found {matches}."
        )
    return matches[0]


def apply_expected_replacements(
    text: str,
    replacements: tuple[tuple[str, str], ...],
    location: str,
) -> str:
    result = text
    for old, new in replacements:
        old_count = result.count(old)
        if old_count == 1:
            result = result.replace(old, new, 1)
        elif old_count > 1:
            raise ValueError(
                f"Ambiguous replacement in {location}: {old!r} occurs "
                f"{old_count} times."
            )
        elif new not in result:
            raise ValueError(
                f"Unexpected content in {location}: neither the audited old "
                f"nor corrected text was found for {old!r}."
            )
    return result


def plan_changes(workbook) -> list[CellChange]:
    changes: list[CellChange] = []
    for sheet_name in CONTENT_SHEETS:
        sheet = workbook[sheet_name]
        headers = header_map(sheet)
        ref_column = headers["ref_id"]
        for patch in TEXT_PATCHES:
            if patch.column not in headers:
                raise ValueError(
                    f"Missing column {patch.column!r} in {sheet_name}."
                )
            row = find_ref_row(sheet, ref_column, patch.ref_id)
            cell = sheet.cell(row, headers[patch.column])
            if not isinstance(cell.value, str):
                raise ValueError(
                    f"Expected text in {sheet_name}!{cell.coordinate}."
                )
            corrected = apply_expected_replacements(
                cell.value,
                patch.replacements,
                f"{sheet_name}!{cell.coordinate}",
            )
            if corrected != cell.value:
                changes.append(
                    CellChange(
                        sheet_name,
                        cell.coordinate,
                        patch.ref_id,
                        patch.column,
                        cell.value,
                        corrected,
                    )
                )

    for patch in REF_PATCHES:
        sheet = workbook[patch.sheet]
        headers = header_map(sheet)
        ref_column = headers["ref_id"]
        row = find_ref_row(sheet, ref_column, patch.ref_id)
        cell = sheet.cell(row, ref_column)
        if cell.value == patch.old_value:
            changes.append(
                CellChange(
                    patch.sheet,
                    cell.coordinate,
                    patch.ref_id,
                    "ref_id",
                    patch.old_value,
                    patch.ref_id,
                )
            )
        elif cell.value != patch.ref_id:
            raise ValueError(
                f"Unexpected ref_id in {patch.sheet}!{cell.coordinate}: "
                f"{cell.value!r}."
            )

    coordinates = [(change.sheet, change.coordinate) for change in changes]
    if len(coordinates) != len(set(coordinates)):
        raise ValueError("Multiple change plans target the same cell.")
    return changes


def apply_changes(workbook, changes: list[CellChange], highlight: bool) -> None:
    for change in changes:
        cell = workbook[change.sheet][change.coordinate]
        if cell.value != change.before:
            raise ValueError(
                f"Content changed before application in {change.sheet}!"
                f"{change.coordinate}."
            )
        cell.value = change.after
        if highlight:
            cell.fill = ORANGE_FILL


def color_value(color) -> tuple[str | None, str | None, int | None]:
    return (color.type, color.rgb, color.indexed)


def validate_target_style(source_cell, output_cell, highlight: bool) -> None:
    for attribute in (
        "font",
        "border",
        "alignment",
        "protection",
    ):
        if copy(getattr(source_cell, attribute)) != copy(
            getattr(output_cell, attribute)
        ):
            raise ValueError(
                f"Unexpected {attribute} change in "
                f"{output_cell.parent.title}!{output_cell.coordinate}."
            )
    if source_cell.number_format != output_cell.number_format:
        raise ValueError(
            f"Unexpected number_format change in "
            f"{output_cell.parent.title}!{output_cell.coordinate}."
        )
    if highlight:
        if output_cell.fill.fill_type != "solid" or color_value(
            output_cell.fill.fgColor
        )[1] not in {"FFFFA500", "00FFA500"}:
            raise ValueError(
                f"Missing orange highlight in {output_cell.parent.title}!"
                f"{output_cell.coordinate}."
            )
    elif source_cell.fill != output_cell.fill:
        raise ValueError(
            f"Unexpected fill change in {output_cell.parent.title}!"
            f"{output_cell.coordinate}."
        )


def verify_output(
    input_path: Path,
    output_path: Path,
    changes: list[CellChange],
    highlight: bool,
) -> None:
    source = openpyxl.load_workbook(input_path, data_only=False)
    output = openpyxl.load_workbook(output_path, data_only=False)
    planned = {
        (change.sheet, change.coordinate): change for change in changes
    }
    try:
        if source.sheetnames != output.sheetnames:
            raise ValueError("Worksheet order or names changed in the output.")
        for sheet_name in source.sheetnames:
            source_sheet = source[sheet_name]
            output_sheet = output[sheet_name]
            if (
                source_sheet.max_row != output_sheet.max_row
                or source_sheet.max_column != output_sheet.max_column
            ):
                raise ValueError(f"Used range changed in {sheet_name}.")
            if set(source_sheet.merged_cells.ranges) != set(
                output_sheet.merged_cells.ranges
            ):
                raise ValueError(f"Merged ranges changed in {sheet_name}.")
            for row in range(1, source_sheet.max_row + 1):
                for column in range(1, source_sheet.max_column + 1):
                    source_cell = source_sheet.cell(row, column)
                    output_cell = output_sheet.cell(row, column)
                    key = (sheet_name, output_cell.coordinate)
                    change = planned.get(key)
                    if change is None:
                        if source_cell.value != output_cell.value:
                            raise ValueError(
                                f"Unexpected value change in {sheet_name}!"
                                f"{output_cell.coordinate}."
                            )
                        if source_cell._style != output_cell._style:
                            raise ValueError(
                                f"Unexpected style change in {sheet_name}!"
                                f"{output_cell.coordinate}."
                            )
                    else:
                        if source_cell.value != change.before:
                            raise ValueError(
                                f"Source drift in {sheet_name}!{output_cell.coordinate}."
                            )
                        if output_cell.value != change.after:
                            raise ValueError(
                                f"Correction missing in {sheet_name}!"
                                f"{output_cell.coordinate}."
                            )
                        validate_target_style(
                            source_cell, output_cell, highlight=highlight
                        )
                    if source_cell.comment != output_cell.comment:
                        raise ValueError(
                            f"Comment changed in {sheet_name}!"
                            f"{output_cell.coordinate}."
                        )
                    if source_cell.hyperlink != output_cell.hyperlink:
                        raise ValueError(
                            f"Hyperlink changed in {sheet_name}!"
                            f"{output_cell.coordinate}."
                        )

        for patch in TEXT_PATCHES:
            values = []
            for sheet_name in CONTENT_SHEETS:
                sheet = output[sheet_name]
                headers = header_map(sheet)
                row = find_ref_row(sheet, headers["ref_id"], patch.ref_id)
                values.append(sheet.cell(row, headers[patch.column]).value)
            if values[0] != values[1]:
                raise ValueError(
                    f"requirements/controls mismatch after correction for "
                    f"{patch.ref_id} {patch.column}."
                )
    finally:
        output.close()
        source.close()


def main() -> int:
    args = parse_arguments()
    if not args.input.is_file():
        raise FileNotFoundError(f"Workbook not found: {args.input}")
    if args.output.resolve() == args.input.resolve():
        raise ValueError("The output path must differ from the input path.")
    if args.output.exists() and not args.force and not args.dry_run:
        raise FileExistsError(
            f"Output already exists: {args.output}. Use --force to replace it."
        )

    input_hash = file_hash(args.input)
    workbook = openpyxl.load_workbook(args.input, data_only=False)
    try:
        changes = plan_changes(workbook)
    finally:
        workbook.close()

    if args.dry_run:
        print(f"Dry run: {len(changes)} cells would be corrected.")
    else:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(args.input, args.output)
        output_workbook = openpyxl.load_workbook(args.output, data_only=False)
        try:
            apply_changes(output_workbook, changes, highlight=args.highlight)
            output_workbook.save(args.output)
        finally:
            output_workbook.close()
        if file_hash(args.input) != input_hash:
            raise ValueError("The input workbook changed during execution.")
        verify_output(
            args.input,
            args.output,
            changes,
            highlight=args.highlight,
        )
        print(f"Verification passed for {len(changes)} corrected cells.")

    counts = Counter(change.field for change in changes)
    print(f"Summary: {len(changes)} corrected cells.")
    print(
        "  Orange highlighting: "
        + ("enabled" if args.highlight else "disabled")
    )
    for field, count in sorted(counts.items()):
        print(f"  {field}: {count}")
    if args.verbose:
        for change in changes:
            print(
                f"{change.sheet}!{change.coordinate} | {change.ref_id} | "
                f"{change.field}"
            )
    if not args.dry_run:
        print(f"Created: {args.output.resolve()}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (FileNotFoundError, FileExistsError, ValueError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1) from exc
