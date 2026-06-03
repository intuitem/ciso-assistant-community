from __future__ import annotations

import argparse
import re
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT = SCRIPT_DIR / "ABRO2026 Dutch and English.xlsx"
DEFAULT_OUTPUT = SCRIPT_DIR / "abro_2026.xlsx"

FRAMEWORK_ID = "abro_2026"
FRAMEWORK_REF_ID = "ABRO_2026"
FRAMEWORK_NAME_NL = "Algemene Beveiligingseisen voor Rijksoverheidsopdrachten (ABRO) 2026"
FRAMEWORK_NAME_EN = "General Security Requirements for Central Government Contracts (ABRO) 2026"
FRAMEWORK_DESCRIPTION_NL = """## VOORWOORD

Geopolitieke ontwikkelingen en gebeurtenissen zoals sabotageacties in Europa en spionage door statelijke actoren onderstrepen de noodzaak van het beschermen van nationale veiligheidsbe- langen. De kabinetsbrede aanpak van economische veiligheid heeft deze aandacht verder versterkt. Wanneer de Rijksoverheid en politie producten en/of diensten inkopen bij een leverancier, kunnen risico's voor de nationale veiligheid ontstaan. Om deze risico's te beperken, zijn de Algemene Beveiligingseisen voor Rijksoverheidsopdrachten (ABRO) 2026 opgesteld.

De Rijksoverheid en politie beschikken over informatie, systemen, materieel en objecten, ook wel Te Beschermen Belangen. In veel gevallen zijn deze van belang voor onze nationale veiligheid en moet te allen tijde worden voorkomen dat kwaadwillenden toegang krijgen of hiervan kennis kunnen nemen. U kunt als leverancier van een opdracht voor de Rijksoverheid of politie toegang krijgen tot een Te Beschermen Belang dat raakt aan de nationale veiligheid. Dan is het belangrijk dat er voldoende waarborgen zijn om het beveiligingsniveau van het Te Beschermen Belang te garanderen. In deze gevallen zult u als leverancier moeten voldoen aan de ABRO 2026.

De ABRO 2026 is een doorontwikkeling van de ABDO 2019, de Algemene Beveiligingseisen voor Defensieopdrachten. Met het in werking treden van de ABRO 2026 gebruiken de gehele Rijksoverheid en de politie dezelfde beveiligingseisen als leveranciers worden ingeschakeld bij opdrachten waarbij Te Beschermen Belangen zijn betrokken. Voor bestaande contracten van Defensie waarop de ABDO van toepassing zijn, blijven de ABDO gelden voor de looptijd van het contract.

Vastgesteld ter invulling van artikel 2 van het Kaderbesluit ABRO rijksdienst, zoals gepubliceerd in de Staatscourant, te Den Haag op 3 december 2025.

Bron :
* https://open.overheid.nl/documenten/a13699d9-fcf4-43ef-b282-ece9519e9b81/file
* https://open.overheid.nl/documenten/8a2c1e3a-f62c-4cd6-9827-3d4ccc9162ba/file"""
FRAMEWORK_DESCRIPTION_EN = """## INTRODUCTION

Geopolitical developments and events such as acts of sabotage in Europe and espionage by state actors stress the need to protect national security interests. The government-wide approach to economic security has further reinforced the attention for this issue. When the Central Government and the police procure products and/or services from a supplier, this may trigger risks to national security. The General Security Requirements for Central Government Contracts (ABRO) 2026 have been prepared to mitigate these risks.

The Central Government and the police are in the possession of information, systems, equipment and objects, also known as Interests to be Protected. In many cases they are important to our national security and malicious parties must be prevented from accessing them or taking cognisance of them at all times. If you are a supplier for the Central Government or the police and you gain access to an Interest to be Protected that affects national security, it is important that sufficient safeguards are in place to guarantee the security level of the Interest to be Protected. In these cases, you must comply with the ABRO 2026 in your role of supplier.

The ABRO 2026 represents the continued development of ABDO 2019, the General Security Requirements for Defence Contracts. With the coming into force of the ABRO 2026, the entire Central Government and the police subject suppliers to the same security requirements where contracts involving Interests to be Protected are concerned. Existing Defence contracts to which ABDO applies, are still subject to ABDO for the term of the contract. This document is a translation of the original Dutch ABRO 2026. In the event of any conflict or inconsistency between this English translation and the original Dutch text, the Dutch text shall prevail.

ABRO 2026 is adopted to implement article 2 of the 'Kaderbesluit ABRO Rijksdienst', as published in the Government Gazette in The Hague on December 3, 2025.

Sources :
* https://open.overheid.nl/documenten/a13699d9-fcf4-43ef-b282-ece9519e9b81/file
* https://open.overheid.nl/documenten/8a2c1e3a-f62c-4cd6-9827-3d4ccc9162ba/file"""
COPYRIGHT = """
Government of the Netherlands / Open Overheid (CC0)
https://www.open-overheid.nl/service/copyright"""
FRAMEWORK_SHEET_BASE = "ABRO"
IMPLEMENTATION_GROUPS_SHEET_BASE = "imp_grp"

BULLET = "\u2022"

SOURCE_SHEETS = ("H1", "H2", "H3", "H4", "H5")
SOURCE_COLUMNS = {
    "ref_id": "ABRO eis nr.",
    "name": "ABRO eis",
    "name_en": "ABRO Requirement (ENG)",
}

CHAPTER_TITLES = {
    "H1": {
        "ref_id": "1",
        "name": "Hoofdstuk 1: Bestuur en organisatie",
        "name[en]": "Chapter 1: Management and organisation",
    },
    "H2": {
        "ref_id": "2",
        "name": "Hoofdstuk 2: Personeel",
        "name[en]": "Chapter 2: Personnel",
    },
    "H3": {
        "ref_id": "3",
        "name": "Hoofdstuk 3: Fysiek",
        "name[en]": "Chapter 3: Physical",
    },
    "H4": {
        "ref_id": "4",
        "name": "Hoofdstuk 4: Cyber",
        "name[en]": "Chapter 4: Cyber",
    },
    "H5": {
        "ref_id": "5",
        "name": "Hoofdstuk 5: Cloud",
        "name[en]": "Chapter 5: Cloud",
    },
}

IMPLEMENTATION_GROUPS = (
    {
        "ref_id": "1",
        "source_column": "TBB 1 / Stg. ZG",
        "name": "TBB 1 / Stg. ZG",
        "name[en]": "ITBP 1 / NLD TS",
    },
    {
        "ref_id": "2",
        "source_column": "TBB 2 / Stg. G",
        "name": "TBB 2 / Stg. G",
        "name[en]": "ITBP 2 / NLD S",
    },
    {
        "ref_id": "3",
        "source_column": "TBB 3 / Stg. C",
        "name": "TBB 3 / Stg. C",
        "name[en]": "ITBP 3 / NLD C",
    },
    {
        "ref_id": "4",
        "source_column": "TBB 4 / DV",
        "name": "TBB 4 / DV",
        "name[en]": "ITBP 4 / NLD RES",
    },
)

ABRO_CONTENT_HEADERS = [
    "assessable",
    "depth",
    "ref_id",
    "name",
    "description",
    "implementation_groups",
    "name[en]",
    "description[en]",
]

IMPLEMENTATION_GROUP_HEADERS = [
    "ref_id",
    "name",
    "name[en]",
]

SECTION_REF_RE = re.compile(r"^\d+\.\d+$")
REQUIREMENT_REF_RE = re.compile(r"^\d+\.\d+\.\d+$")
SECTION_REF_PARTS_RE = re.compile(r"^(\d+)\.(\d+)$")


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    lines = [re.sub(r"[ \t]+", " ", line).strip() for line in text.splitlines()]
    return "\n".join(line for line in lines if line)


def decimal_places_from_number_format(number_format: str) -> int:
    main_format = str(number_format or "").split(";")[0]
    if "." not in main_format:
        return 0
    decimal_part = main_format.split(".", 1)[1]
    return decimal_part.count("0")


def clean_ref_id(cell_or_value: Any) -> str:
    value = getattr(cell_or_value, "value", cell_or_value)
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, (int, float)):
        decimal_places = decimal_places_from_number_format(
            getattr(cell_or_value, "number_format", "")
        )
        if decimal_places:
            return f"{value:.{decimal_places}f}"
    return clean_text(value)


def normalize_section_ref_id(
    ref_id: str, last_section_minor_by_major: dict[int, int]
) -> str:
    match = SECTION_REF_PARTS_RE.match(ref_id)
    if not match:
        return ref_id

    major = int(match.group(1))
    minor = int(match.group(2))
    last_minor = last_section_minor_by_major.get(major, 0)

    if minor <= last_minor:
        minor = last_minor + 1

    last_section_minor_by_major[major] = minor
    return f"{major}.{minor}"


def append_key_value_rows(ws: Worksheet, rows: list[tuple[str, Any]]) -> None:
    for key, value in rows:
        ws.append([key, value])


def build_library_meta(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "library_meta"
    append_key_value_rows(
        ws,
        [
            ("type", "library"),
            ("urn", f"urn:intuitem:risk:library:{FRAMEWORK_ID}"),
            ("version", 1),
            ("locale", "nl"),
            ("ref_id", FRAMEWORK_REF_ID),
            ("name", FRAMEWORK_NAME_NL),
            ("description", FRAMEWORK_DESCRIPTION_NL),
            ("copyright", COPYRIGHT),
            ("provider", "Government of the Netherlands"),
            ("packager", "intuitem"),
            ("name[en]", FRAMEWORK_NAME_EN),
            ("description[en]", FRAMEWORK_DESCRIPTION_EN),
        ],
    )


def build_framework_meta(wb: Workbook) -> None:
    ws = wb.create_sheet(f"{FRAMEWORK_SHEET_BASE}_meta")
    append_key_value_rows(
        ws,
        [
            ("type", "framework"),
            ("base_urn", f"urn:intuitem:risk:req_node:{FRAMEWORK_ID}"),
            ("urn", f"urn:intuitem:risk:framework:{FRAMEWORK_ID}"),
            ("ref_id", FRAMEWORK_REF_ID),
            ("name", FRAMEWORK_NAME_NL),
            ("description", FRAMEWORK_DESCRIPTION_NL),
            ("implementation_groups_definition", IMPLEMENTATION_GROUPS_SHEET_BASE),
            ("name[en]", FRAMEWORK_NAME_EN),
            ("description[en]", FRAMEWORK_DESCRIPTION_EN),
        ],
    )


def header_map(ws: Worksheet) -> dict[str, int]:
    headers = {clean_text(cell.value): idx for idx, cell in enumerate(ws[2])}
    required_headers = {
        *SOURCE_COLUMNS.values(),
        *(group["source_column"] for group in IMPLEMENTATION_GROUPS),
    }
    missing = sorted(header for header in required_headers if header not in headers)
    if missing:
        raise ValueError(f'Sheet "{ws.title}" is missing expected columns: {missing}')
    return headers


def implementation_groups_for_row(row: tuple[Any, ...], headers: dict[str, int]) -> str:
    selected_groups = []
    for group in IMPLEMENTATION_GROUPS:
        cell_value = clean_text(row[headers[group["source_column"]]].value)
        if cell_value == BULLET:
            selected_groups.append(group["ref_id"])
    return ",".join(selected_groups)


def unique_ref_id(ref_id: str, ref_id_counts: dict[str, int]) -> str:
    if not ref_id:
        return ref_id
    ref_id_counts[ref_id] = ref_id_counts.get(ref_id, 0) + 1
    if ref_id_counts[ref_id] == 1:
        return ref_id
    return f"{ref_id}-{ref_id_counts[ref_id]}"


def append_content_row(
    ws: Worksheet,
    *,
    assessable: str | None = None,
    depth: int,
    implementation_groups: str | None = None,
    ref_id: str | None = None,
    name: str,
    description: str | None = None,
    name_en: str,
    description_en: str | None = None,
) -> None:
    ws.append(
        [
            assessable,
            depth,
            ref_id,
            name,
            description,
            implementation_groups,
            name_en,
            description_en,
        ]
    )


def build_abro_content(wb_out: Workbook, source_wb) -> int:
    ws_out = wb_out.create_sheet(f"{FRAMEWORK_SHEET_BASE}_content")
    ws_out.append(ABRO_CONTENT_HEADERS)
    rows_written = 0
    ref_id_counts: dict[str, int] = {}
    last_section_minor_by_major: dict[int, int] = {}

    for sheet_name in SOURCE_SHEETS:
        ws_source = source_wb[sheet_name]
        headers = header_map(ws_source)
        chapter = CHAPTER_TITLES[sheet_name]

        append_content_row(
            ws_out,
            depth=1,
            ref_id=unique_ref_id(chapter["ref_id"], ref_id_counts),
            name=chapter["name"],
            name_en=chapter["name[en]"],
        )
        rows_written += 1

        for row in ws_source.iter_rows(min_row=3):
            ref_id = clean_ref_id(row[headers[SOURCE_COLUMNS["ref_id"]]])
            if not ref_id:
                continue

            name = clean_text(row[headers[SOURCE_COLUMNS["name"]]].value)
            name_en = clean_text(row[headers[SOURCE_COLUMNS["name_en"]]].value)
            if not name and not name_en:
                continue

            if SECTION_REF_RE.match(ref_id):
                ref_id = normalize_section_ref_id(ref_id, last_section_minor_by_major)
                append_content_row(
                    ws_out,
                    depth=2,
                    ref_id=unique_ref_id(ref_id, ref_id_counts),
                    name=name,
                    name_en=name_en,
                )
                rows_written += 1
                continue

            if REQUIREMENT_REF_RE.match(ref_id):
                append_content_row(
                    ws_out,
                    assessable="x",
                    depth=3,
                    implementation_groups=implementation_groups_for_row(row, headers),
                    ref_id=unique_ref_id(ref_id, ref_id_counts),
                    name="",
                    description=name,
                    name_en="",
                    description_en=name_en,
                )
                rows_written += 1

    return rows_written


def build_implementation_groups(wb: Workbook) -> None:
    ws_meta = wb.create_sheet(f"{IMPLEMENTATION_GROUPS_SHEET_BASE}_meta")
    append_key_value_rows(
        ws_meta,
        [
            ("type", "implementation_groups"),
            ("name", IMPLEMENTATION_GROUPS_SHEET_BASE),
        ],
    )

    ws_content = wb.create_sheet(f"{IMPLEMENTATION_GROUPS_SHEET_BASE}_content")
    ws_content.append(IMPLEMENTATION_GROUP_HEADERS)
    for group in IMPLEMENTATION_GROUPS:
        ws_content.append(
            [
                group["ref_id"],
                group["name"],
                group["name[en]"],
            ]
        )


def style_workbook(wb: Workbook) -> None:
    for ws in wb.worksheets:
        for column_cells in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                value = clean_text(cell.value)
                line_lengths = [len(line) for line in value.splitlines()]
                max_length = max(max_length, max(line_lengths, default=0))
            ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 80)

        for row in ws.iter_rows():
            for cell in row:
                alignment = copy(cell.alignment)
                alignment.wrap_text = True
                alignment.vertical = "top"
                cell.alignment = alignment


def build_workbook(input_path: Path, output_path: Path) -> int:
    source_wb = load_workbook(input_path, read_only=True, data_only=True)
    missing_sheets = [sheet for sheet in SOURCE_SHEETS if sheet not in source_wb.sheetnames]
    if missing_sheets:
        raise ValueError(f"Missing expected source sheets: {missing_sheets}")

    wb_out = Workbook()
    build_library_meta(wb_out)
    build_framework_meta(wb_out)
    rows_written = build_abro_content(wb_out, source_wb)
    build_implementation_groups(wb_out)
    style_workbook(wb_out)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(output_path)
    return rows_written


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert ABRO 2026 Dutch and English source workbook to CISO Assistant framework format."
    )
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="Source ABRO workbook.")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Generated framework workbook.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    rows_written = build_workbook(args.input, args.output)
    print(f"Wrote {rows_written} ABRO content rows to {args.output}")


if __name__ == "__main__":
    main()
