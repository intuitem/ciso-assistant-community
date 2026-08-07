from __future__ import annotations

import argparse
import re
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT = SCRIPT_DIR.parent / "NIS2 CBB NL + ENG (+mappings).xlsx"
DEFAULT_OUTPUT = SCRIPT_DIR.parent / "cbw_nis2.xlsx"

FRAMEWORK_ID = "cbw_nis2_control_framework"
FRAMEWORK_REF_ID = "Cbw_NIS2_Control_Framework"
FRAMEWORK_VERSION = "1"
FRAMEWORK_NAME_NL = "Cbw (NIS2) Control Framework (v1.2)"
FRAMEWORK_NAME_EN = "Cbw (NIS2) Control Framework (v1.2)"
FRAMEWORK_DESCRIPTION_NL = """# Introductie

De evaluatietool is opgesteld als praktisch hulpmiddel om organisaties te ondersteunen bij het op een effectieve wijze verkrijgen van inzicht in de mate waarin zij voldoen aan de Cyberbeveiligingswet (Cbw) en het onderliggende Cyberbeveiligingsbesluit (Cbb).

De evaluatietool is ontwikkeld om inzicht te geven in de huidige staat van cyberweerbaarheid en om verbeterpotentieel te identificeren, gebaseerd op de eisen zoals vastgelegd in de Cbw en Cbb. Het Cbw (NIS2) Control Framework vervangt de geldende wet- en regelgeving niet, maar maakt deze inzichtelijk en hanteerbaar - zie ook het studierapport vanaf p. 16. De Cbw en de Cbb dagen organisaties uit om middels risicoanalyse en inzicht in de eigen organisatie de Cbw en de Cbb te implementeren. Medewerkers van organisaties zijn zelf verantwoordelijk voor het kennen en naleven van de relevante wet- en regelgeving. Bovendien dient bij het gebruik ervan altijd rekening te worden gehouden met de specifieke context en kenmerken van de organisatie en haar omgeving."""
FRAMEWORK_DESCRIPTION_EN = """# Introduction

The evaluation tool has been developed as a practical tool to support organizations in effectively gaining insight into their compliance with the Cybersecurity Act (Cbw) and the underlying Cybersecurity Decree (Cbb). The evaluation tool was developed to provide insight into the current state of cyber resilience and to identify potential for improvement, based on the requirements laid out in the Cbw and Cbb. The Cbw (NIS2) Control Framework does not replace current laws and regulations, but rather makes them transparent and manageable – see also the study report from p. 16 onwards. The Cbw and Cbb challenge organizations to implement the Cbw and Cbb through risk analysis and insight into their own organization. Employees of organizations are responsible for knowing and complying with the relevant laws and regulations. Furthermore, the specific context and characteristics of the organization and its environment should always be taken into account when using the tool."""
COPYRIGHT = """Unless otherwise noted, everything in this work is licensed under a Creative Commons Attribution 4.0 License.
If you wish to use this work, please use the following attribution method: ADR & NOREA, Cbw (NIS2) Control Framework (2025), CC-BY 4.0 licensed
The full license text can be read at: https://creativecommons.org/licenses/by/4.0/."""
PROVIDER = "ADR (Auditdienst Rijk) & NOREA"
PACKAGER = "intuitem"

NL_SHEET = "Cyberbeveiligingswet (NL)"
EN_SHEET = "Cyber Security Act (ENG)"

NL_HEADERS = {
    "theme": "Thema",
    "domain": "Domein",
    "ref_id": "Control ID",
    "description": "Beheersmaatregel",
    "annotation": "Toelichting",
}
EN_HEADERS = {
    "theme": "Subject",
    "domain": "Domain",
    "ref_id": "Control ID",
    "description": "Control measure",
    "annotation": "Explanation",
}

CONTENT_HEADERS = [
    "assessable",
    "depth",
    "ref_id",
    "name",
    "description",
    "annotation",
    "name[en]",
    "description[en]",
    "annotation[en]",
]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    lines = [re.sub(r"[ \t]+", " ", line).strip() for line in text.splitlines()]
    return "\n".join(line for line in lines if line)


def append_key_value_rows(ws: Worksheet, rows: list[tuple[str, Any]]) -> None:
    for key, value in rows:
        ws.append([key, value])


def build_library_meta(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "library_meta"
    append_key_value_rows(
        ws,
        [
            ("type", "library"),
            ("urn", f"urn:intuitem:risk:library:{FRAMEWORK_ID}"),
            ("version", FRAMEWORK_VERSION),
            ("locale", "nl"),
            ("ref_id", FRAMEWORK_REF_ID),
            ("name", FRAMEWORK_NAME_NL),
            ("description", FRAMEWORK_DESCRIPTION_NL),
            ("copyright", COPYRIGHT),
            ("provider", PROVIDER),
            ("packager", PACKAGER),
            ("name[en]", FRAMEWORK_NAME_EN),
            ("description[en]", FRAMEWORK_DESCRIPTION_EN),
        ],
    )


def build_framework_meta(wb: Workbook) -> None:
    ws = wb.create_sheet("fwk_meta")
    append_key_value_rows(
        ws,
        [
            ("type", "framework"),
            ("base_urn", f"urn:intuitem:risk:req_node:{FRAMEWORK_ID}"),
            ("urn", f"urn:intuitem:risk:framework:{FRAMEWORK_ID}"),
            ("ref_id", FRAMEWORK_REF_ID),
            ("name", FRAMEWORK_NAME_NL),
            ("description", FRAMEWORK_DESCRIPTION_NL),
            ("name[en]", FRAMEWORK_NAME_EN),
            ("description[en]", FRAMEWORK_DESCRIPTION_EN),
        ],
    )


def find_header_columns(
    ws: Worksheet, expected_headers: dict[str, str], header_rows: range
) -> dict[str, int]:
    found: dict[str, int] = {}
    expected_by_label = {label: key for key, label in expected_headers.items()}

    for row in ws.iter_rows(
        min_row=header_rows.start,
        max_row=header_rows.stop - 1,
    ):
        for cell in row:
            label = clean_text(cell.value)
            key = expected_by_label.get(label)
            if key is not None:
                found[key] = cell.column

    missing = [
        label for key, label in expected_headers.items() if key not in found
    ]
    if missing:
        raise ValueError(f'Sheet "{ws.title}" is missing expected headers: {missing}')
    return found


def merged_values(ws: Worksheet, columns: set[int]) -> dict[tuple[int, int], Any]:
    values: dict[tuple[int, int], Any] = {}
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_col != merged_range.max_col:
            continue
        if merged_range.min_col not in columns:
            continue

        value = ws.cell(merged_range.min_row, merged_range.min_col).value
        for row_number in range(merged_range.min_row, merged_range.max_row + 1):
            values[(row_number, merged_range.min_col)] = value
    return values


def cell_text(
    ws: Worksheet,
    row_number: int,
    column_number: int,
    merged_lookup: dict[tuple[int, int], Any],
) -> str:
    value = ws.cell(row_number, column_number).value
    if value is None:
        value = merged_lookup.get((row_number, column_number))
    return clean_text(value)


def extract_controls(
    ws: Worksheet,
    headers: dict[str, int],
    first_data_row: int,
) -> list[dict[str, str]]:
    hierarchy_columns = {headers["theme"], headers["domain"]}
    merged_lookup = merged_values(ws, hierarchy_columns)
    controls: list[dict[str, str]] = []
    seen_ref_ids: set[str] = set()

    for row_number in range(first_data_row, ws.max_row + 1):
        ref_id = clean_text(ws.cell(row_number, headers["ref_id"]).value)
        if not ref_id:
            continue
        if ref_id in seen_ref_ids:
            raise ValueError(
                f'Duplicate Control ID "{ref_id}" in sheet "{ws.title}".'
            )

        seen_ref_ids.add(ref_id)
        controls.append(
            {
                "theme": cell_text(
                    ws, row_number, headers["theme"], merged_lookup
                ),
                "domain": cell_text(
                    ws, row_number, headers["domain"], merged_lookup
                ),
                "ref_id": ref_id,
                "description": clean_text(
                    ws.cell(row_number, headers["description"]).value
                ),
                "annotation": clean_text(
                    ws.cell(row_number, headers["annotation"]).value
                ),
            }
        )

    return controls


def append_content_row(
    ws: Worksheet,
    *,
    assessable: str | None = None,
    depth: int,
    ref_id: str | None = None,
    name: str | None = None,
    description: str | None = None,
    annotation: str | None = None,
    name_en: str | None = None,
    description_en: str | None = None,
    annotation_en: str | None = None,
) -> None:
    ws.append(
        [
            assessable,
            depth,
            ref_id,
            name,
            description,
            annotation,
            name_en,
            description_en,
            annotation_en,
        ]
    )


def build_content_sheet(
    wb_out: Workbook,
    nl_controls: list[dict[str, str]],
    en_controls: list[dict[str, str]],
) -> tuple[int, list[str], list[str]]:
    ws = wb_out.create_sheet("fwk_content")
    ws.append(CONTENT_HEADERS)

    en_by_ref_id = {control["ref_id"]: control for control in en_controls}
    nl_ref_ids = {control["ref_id"] for control in nl_controls}
    missing_in_en = [
        control["ref_id"]
        for control in nl_controls
        if control["ref_id"] not in en_by_ref_id
    ]
    missing_in_nl = [
        control["ref_id"]
        for control in en_controls
        if control["ref_id"] not in nl_ref_ids
    ]

    current_theme: str | None = None
    current_domain: str | None = None
    rows_written = 0

    for nl_control in nl_controls:
        en_control = en_by_ref_id.get(nl_control["ref_id"], {})
        theme = nl_control["theme"]
        domain = nl_control["domain"]

        if not theme or not domain:
            raise ValueError(
                f'Control "{nl_control["ref_id"]}" has no Dutch theme or domain.'
            )

        if theme != current_theme:
            append_content_row(
                ws,
                depth=1,
                name=theme,
                name_en=en_control.get("theme") or None,
            )
            rows_written += 1
            current_theme = theme
            current_domain = None

        if domain != current_domain:
            append_content_row(
                ws,
                depth=2,
                name=domain,
                name_en=en_control.get("domain") or None,
            )
            rows_written += 1
            current_domain = domain

        append_content_row(
            ws,
            assessable="x",
            depth=3,
            ref_id=nl_control["ref_id"],
            description=nl_control["description"] or None,
            annotation=nl_control["annotation"] or None,
            description_en=en_control.get("description") or None,
            annotation_en=en_control.get("annotation") or None,
        )
        rows_written += 1

    return rows_written, missing_in_en, missing_in_nl


def style_workbook(wb: Workbook) -> None:
    for ws in wb.worksheets:
        if ws.title == "fwk_content":
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

        for column_cells in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                value = clean_text(cell.value)
                max_length = max(
                    max_length,
                    max((len(line) for line in value.splitlines()), default=0),
                )
            ws.column_dimensions[column_letter].width = min(
                max(max_length + 2, 10), 80
            )

        for row in ws.iter_rows():
            for cell in row:
                alignment = copy(cell.alignment)
                alignment.wrap_text = True
                alignment.vertical = "top"
                cell.alignment = alignment


def build_workbook(input_path: Path, output_path: Path) -> tuple[int, int, list[str], list[str]]:
    source_wb = load_workbook(input_path, data_only=True)
    missing_sheets = [
        sheet_name
        for sheet_name in (NL_SHEET, EN_SHEET)
        if sheet_name not in source_wb.sheetnames
    ]
    if missing_sheets:
        raise ValueError(f"Missing expected source sheets: {missing_sheets}")

    nl_ws = source_wb[NL_SHEET]
    en_ws = source_wb[EN_SHEET]
    nl_columns = find_header_columns(nl_ws, NL_HEADERS, range(4, 7))
    en_columns = find_header_columns(en_ws, EN_HEADERS, range(5, 8))
    nl_controls = extract_controls(nl_ws, nl_columns, first_data_row=7)
    en_controls = extract_controls(en_ws, en_columns, first_data_row=8)
    source_wb.close()

    wb_out = Workbook()
    build_library_meta(wb_out)
    build_framework_meta(wb_out)
    rows_written, missing_in_en, missing_in_nl = build_content_sheet(
        wb_out, nl_controls, en_controls
    )
    style_workbook(wb_out)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(output_path)
    return rows_written, len(nl_controls), missing_in_en, missing_in_nl


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Convert the Dutch and English CBW/NIS2 workbook to a CISO "
            "Assistant fwk_content sheet."
        )
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT,
        help="Source CBW/NIS2 workbook.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help="Generated framework workbook.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    rows_written, control_count, missing_in_en, missing_in_nl = build_workbook(
        args.input, args.output
    )
    print(
        f"Wrote {rows_written} fwk_content rows ({control_count} controls) "
        f"to {args.output}"
    )
    if missing_in_en:
        print(
            "Warning: Control IDs missing from the English sheet: "
            + ", ".join(missing_in_en)
        )
    if missing_in_nl:
        print(
            "Warning: Control IDs missing from the Dutch sheet: "
            + ", ".join(missing_in_nl)
        )


if __name__ == "__main__":
    main()
