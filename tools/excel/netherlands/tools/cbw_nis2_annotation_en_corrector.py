from __future__ import annotations

import argparse
import re
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT = SCRIPT_DIR.parent / "cbw_nis2.xlsx"
DEFAULT_OUTPUT = SCRIPT_DIR.parent / "cbw_nis2_annotations_corrected.xlsx"

CONTENT_SHEET = "fwk_content"
ANNOTATION_EN_HEADER = "annotation[en]"
REF_ID_HEADER = "ref_id"

YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFFF00")
LETTER_MARKER_RE = re.compile(r"(?<![A-Za-z])([a-j])\.\s+")
COMPACT_HYPHEN_RE = re.compile(r"[ \t]+-[ \t]+")
MISSING_MARKER_SPACE_RE = re.compile(r"\b([a-j])\.([a-z])(?=\s)")

PARAGRAPH_STARTS_BY_REF_ID = {
    "14.1": (
        "The training certificate shall, at a minimum, include:",
        "The training certificate shall be available in Dutch or English.",
    ),
    "15.1": (
        "Voluntary reporting obligation:",
    ),
    "16.1": (
        "In the interim update, the entity shall provide",
        "The final report shall include",
    ),
}


def header_columns(ws: Worksheet) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in ws[1]
        if cell.value is not None
    }


def correct_compact_lists(value: str, ref_id: str) -> str:
    corrected = value

    hyphen_markers = COMPACT_HYPHEN_RE.findall(corrected)
    if len(hyphen_markers) >= 2:
        corrected = COMPACT_HYPHEN_RE.sub("\n- ", corrected)

    corrected = MISSING_MARKER_SPACE_RE.sub(r"\1. \2", corrected)
    letter_markers = LETTER_MARKER_RE.findall(corrected)
    if "a" in letter_markers and "b" in letter_markers:
        corrected = LETTER_MARKER_RE.sub(r"\n\1. ", corrected)

    for paragraph_start in PARAGRAPH_STARTS_BY_REF_ID.get(ref_id, ()):
        corrected = re.sub(
            rf"\s*{re.escape(paragraph_start)}",
            f"\n\n{paragraph_start}",
            corrected,
        )

    corrected = re.sub(r"[ \t]+\n", "\n", corrected)
    corrected = re.sub(r"\n[ \t]+", "\n", corrected)
    corrected = re.sub(r"\n{3,}", "\n\n", corrected)
    return corrected


def correct_workbook(input_path: Path, output_path: Path) -> list[str]:
    if input_path.resolve() == output_path.resolve():
        raise ValueError("The output path must differ from the input path.")

    wb = load_workbook(input_path)
    if CONTENT_SHEET not in wb.sheetnames:
        raise ValueError(f'Missing expected sheet: "{CONTENT_SHEET}"')

    ws = wb[CONTENT_SHEET]
    headers = header_columns(ws)
    missing_headers = [
        header
        for header in (REF_ID_HEADER, ANNOTATION_EN_HEADER)
        if header not in headers
    ]
    if missing_headers:
        raise ValueError(
            f'Sheet "{CONTENT_SHEET}" is missing expected headers: {missing_headers}'
        )

    ref_id_column = headers[REF_ID_HEADER]
    annotation_column = headers[ANNOTATION_EN_HEADER]
    changed_ref_ids: list[str] = []

    for row_number in range(2, ws.max_row + 1):
        cell = ws.cell(row_number, annotation_column)
        if not isinstance(cell.value, str) or not cell.value.strip():
            continue

        ref_id_value = ws.cell(row_number, ref_id_column).value
        ref_id = str(ref_id_value) if ref_id_value is not None else ""
        corrected = correct_compact_lists(cell.value, ref_id)
        if corrected == cell.value:
            continue

        cell.value = corrected
        cell.fill = copy(YELLOW_FILL)
        alignment = copy(cell.alignment)
        alignment.wrap_text = True
        alignment.vertical = "top"
        cell.alignment = alignment

        changed_ref_ids.append(ref_id or f"row {row_number}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()
    return changed_ref_ids


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Create a copy of a CBW/NIS2 workbook, put compact English "
            "annotation lists on separate lines, and highlight changed cells."
        )
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_INPUT,
        help="Source CBW/NIS2 workbook.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help="Corrected workbook copy.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    changed_ref_ids = correct_workbook(args.input, args.output)
    print(
        f"Corrected {len(changed_ref_ids)} annotation[en] cells in {args.output}"
    )
    if changed_ref_ids:
        print("Changed Control IDs: " + ", ".join(changed_ref_ids))


if __name__ == "__main__":
    main()
