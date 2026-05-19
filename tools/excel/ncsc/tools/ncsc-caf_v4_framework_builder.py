#!/usr/bin/env python3
"""Build the NCSC CAF 4.0 workbook from the official PDF.

The PDF uses a stable document structure:
- objective / principle / contributing-outcome headings are normal left-column text,
- IGP tables use either 2 columns (Not Achieved / Achieved) or 3 columns
  (Not Achieved / Partially Achieved / Achieved).

The CISO Assistant library only keeps the "Achieved" IGP statements as assessable
requirements, following the existing CAF 3.2 workbook structure.
"""

from __future__ import annotations

import argparse
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

try:
    import fitz
except ImportError as exc:  # pragma: no cover - dependency guard for local runs
    raise SystemExit(
        "PyMuPDF is required to extract this PDF. Install it in the virtualenv "
        "(`pip install pymupdf`) or run the script from the repo .venv."
    ) from exc

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet


FRAMEWORK_VERSION = "4.0"
FRAMEWORK_SLUG = "ncsc-caf-4.0"
PDF_FILENAME = "NCSC-Cyber-Assessment-Framework-4.0.pdf"
OUTPUT_FILENAME = f"{FRAMEWORK_SLUG}.xlsx"

LIBRARY_SHEET = "library_meta"
FRAMEWORK_SHEET = "caf_meta"
CONTENT_SHEET = "caf_content"

NCSC_COLLECTION_URL = "https://www.ncsc.gov.uk/collection/cyber-assessment-framework"
FRAMEWORK_NAME = "NCSC - Cyber Assessment Framework (CAF) v4.0"
FRAMEWORK_DESCRIPTION = (
    "National Cyber Security Centre - Cyber Assessment Framework\n"
    f"{NCSC_COLLECTION_URL}"
)
COPYRIGHT = f"NCSC {NCSC_COLLECTION_URL}"
PROVIDER = "NCSC"
PACKAGER = "intuitem"

LIBRARY_META_ROWS = [
    ("type", "library"),
    ("urn", f"urn:intuitem:risk:library:{FRAMEWORK_SLUG}"),
    ("version", "1"),
    ("locale", "en"),
    ("ref_id", FRAMEWORK_SLUG),
    ("name", FRAMEWORK_NAME),
    ("description", FRAMEWORK_DESCRIPTION),
    ("copyright", COPYRIGHT),
    ("provider", PROVIDER),
    ("packager", PACKAGER),
]

FRAMEWORK_META_ROWS = [
    ("type", "framework"),
    ("base_urn", f"urn:intuitem:risk:req_node:{FRAMEWORK_SLUG}"),
    ("urn", f"urn:intuitem:risk:framework:{FRAMEWORK_SLUG}"),
    ("ref_id", FRAMEWORK_SLUG),
    ("name", FRAMEWORK_NAME),
    ("description", FRAMEWORK_DESCRIPTION),
]

CONTENT_HEADERS = ["assessable", "depth", "ref_id", "name", "description"]
CONTENT_COLUMN_WIDTHS = {
    "A": 8.33203125,
    "B": 5.83203125,
    "C": 8.33203125,
    "D": 43.6640625,
    "E": 255.83203125,
}

EXPECTED_OBJECTIVES = 4
EXPECTED_PRINCIPLES = 14
EXPECTED_OUTCOMES = 41
EXPECTED_ACHIEVED_STATEMENTS = 225

OBJECTIVE_RE = re.compile(
    r"^CAF\s*-\s*Objective\s+([A-D])\s*[-–]\s*(.+)$", re.IGNORECASE | re.DOTALL
)
PRINCIPLE_RE = re.compile(r"^Principle\s+([A-D]\d+)\s+(.+)$", re.IGNORECASE | re.DOTALL)
OUTCOME_RE = re.compile(
    r"^([A-D]\d+\.[a-z])\.?\s+(.+)$", re.IGNORECASE | re.DOTALL
)
TABLE_HEADER_RE = re.compile(r"\bNot\s+Achieved\b.*\bAchieved\b", re.IGNORECASE | re.DOTALL)

LEFT_TEXT_MAX_X = 90.0
TWO_COLUMN_ACHIEVED_MIN_X = 270.0
THREE_COLUMN_ACHIEVED_MIN_X = 380.0
TABLE_BOTTOM_Y = 775.0

STRUCTURAL_PREFIXES = (
    "The Cyber Assessment Framework",
    "CAF - Objective",
    "Principle ",
)

CONTINUATION_TERMINALS = {
    "a",
    "an",
    "and",
    "as",
    "at",
    "by",
    "for",
    "from",
    "in",
    "including",
    "into",
    "of",
    "on",
    "or",
    "the",
    "their",
    "to",
    "using",
    "with",
    "within",
    "your",
}


@dataclass
class PdfBlock:
    page: int
    x0: float
    y0: float
    x1: float
    y1: float
    text: str


@dataclass
class ContentRow:
    assessable: str | None
    depth: int
    ref_id: str | None
    name: str | None
    description: str | None


@dataclass
class AchievedSegment:
    page: int
    x0: float
    y0: float
    y1: float
    text: str


@dataclass
class ParseState:
    rows: list[ContentRow] = field(default_factory=list)
    current_description_row: int | None = None
    current_outcome_ref: str | None = None
    current_segments: list[AchievedSegment] = field(default_factory=list)
    table_active: bool = False
    achieved_min_x: float | None = None


def default_pdf_path() -> Path:
    return Path.cwd() / PDF_FILENAME


def default_output_path() -> Path:
    return Path.cwd() / OUTPUT_FILENAME


def normalize_text(value: str) -> str:
    value = value.replace("\r\n", "\n").replace("\r", "\n")
    value = value.replace("\u2028", "\n")
    value = value.replace("’", "'")
    value = re.sub(r"[ \t]+\n", "\n", value)
    value = re.sub(r"\n[ \t]+", "\n", value)
    value = re.sub(r"[ \t]{2,}", " ", value)
    value = re.sub(r"\n{3,}", "\n\n", value)
    return value.strip()


def one_line(value: str) -> str:
    value = normalize_text(value)
    value = re.sub(r"\s+", " ", value)
    value = re.sub(r"(?<=\w)-\s+(?=\w)", "-", value)
    value = value.replace(" ,", ",").replace(" .", ".").replace(" ;", ";")
    value = value.replace(" :", ":")
    return value.strip()


def is_noise_block(block: PdfBlock) -> bool:
    text = one_line(block.text)
    if not text:
        return True
    if text == "National Cyber Security Centre":
        return True
    if text.startswith("© Crown copyright"):
        return True
    if text.startswith("third parties and are not available for re-use"):
        return True
    if text.isdigit() and block.y0 > TABLE_BOTTOM_Y:
        return True
    if block.y0 > TABLE_BOTTOM_Y:
        return True
    return False


def is_table_instruction(text: str) -> bool:
    normalized = one_line(text)
    return normalized in {
        "At least one of the following statements is true",
        "All the following statements",
        "All the following statements are true",
        "All of the following statements",
        "All of the following statements are true",
        "are true",
        "is true",
        "true",
    }


def iter_pdf_blocks(pdf_path: Path) -> Iterable[PdfBlock]:
    with fitz.open(pdf_path) as document:
        for page_index, page in enumerate(document):
            blocks = sorted(page.get_text("blocks"), key=lambda block: (block[1], block[0]))
            for block in blocks:
                x0, y0, x1, y1, text = block[:5]
                pdf_block = PdfBlock(
                    page=page_index + 1,
                    x0=float(x0),
                    y0=float(y0),
                    x1=float(x1),
                    y1=float(y1),
                    text=normalize_text(str(text)),
                )
                if not is_noise_block(pdf_block):
                    yield pdf_block


def add_content_row(
    state: ParseState,
    assessable: str | None,
    depth: int,
    ref_id: str | None,
    name: str | None,
    description: str | None = None,
) -> None:
    state.rows.append(ContentRow(assessable, depth, ref_id, name, description))
    state.current_description_row = len(state.rows) - 1


def append_description(state: ParseState, text: str) -> None:
    if state.current_description_row is None:
        return
    cleaned = one_line(text)
    if not cleaned:
        return
    row = state.rows[state.current_description_row]

    if row.depth == 3 and row.name and has_unclosed_parenthesis(row.name):
        row.name, cleaned = merge_heading_continuation(row.name, cleaned)
        if not cleaned:
            return

    if row.description:
        row.description = f"{row.description} {cleaned}"
    else:
        row.description = cleaned


def has_unclosed_parenthesis(value: str) -> bool:
    return value.count("(") > value.count(")")


def merge_heading_continuation(name: str, text: str) -> tuple[str, str]:
    if ")" not in text:
        return f"{name} {text}".strip(), ""

    continuation, remainder = text.split(")", maxsplit=1)
    merged_name = f"{name} {continuation})"
    return one_line(merged_name), remainder.strip()


def starts_new_structure(text: str) -> bool:
    normalized = one_line(text)
    if any(normalized.startswith(prefix) for prefix in STRUCTURAL_PREFIXES):
        return True
    return bool(OUTCOME_RE.match(normalized))


def finalize_outcome_segments(state: ParseState) -> None:
    if not state.current_outcome_ref:
        state.current_segments.clear()
        return

    for index, statement in enumerate(split_statements(state.current_segments), start=1):
        state.rows.append(
            ContentRow(
                assessable="x",
                depth=4,
                ref_id=f"{state.current_outcome_ref}.{index}",
                name=None,
                description=statement,
            )
        )

    state.current_segments.clear()


def parse_structure_block(state: ParseState, text: str) -> bool:
    normalized = one_line(text)

    objective_match = OBJECTIVE_RE.match(normalized)
    if objective_match:
        finalize_outcome_segments(state)
        objective_ref = objective_match.group(1).upper()
        add_content_row(
            state,
            assessable=None,
            depth=1,
            ref_id=objective_ref,
            name=objective_match.group(2).strip(),
        )
        state.current_outcome_ref = None
        state.table_active = False
        state.achieved_min_x = None
        return True

    principle_match = PRINCIPLE_RE.match(normalized)
    if principle_match:
        finalize_outcome_segments(state)
        add_content_row(
            state,
            assessable=None,
            depth=2,
            ref_id=principle_match.group(1).upper(),
            name=principle_match.group(2).strip(),
        )
        state.current_outcome_ref = None
        state.table_active = False
        state.achieved_min_x = None
        return True

    outcome_match = OUTCOME_RE.match(normalized)
    if outcome_match:
        finalize_outcome_segments(state)
        ref_id = outcome_match.group(1)
        add_content_row(
            state,
            assessable=None,
            depth=3,
            ref_id=ref_id,
            name=outcome_match.group(2).strip(),
        )
        state.current_outcome_ref = ref_id
        state.table_active = False
        state.achieved_min_x = None
        return True

    return False


def activate_table(state: ParseState, text: str) -> bool:
    if not TABLE_HEADER_RE.search(text):
        return False

    state.table_active = True
    state.current_description_row = None
    if "Partially Achieved" in text:
        state.achieved_min_x = THREE_COLUMN_ACHIEVED_MIN_X
    else:
        state.achieved_min_x = TWO_COLUMN_ACHIEVED_MIN_X
    return True


def should_collect_achieved_segment(state: ParseState, block: PdfBlock) -> bool:
    if not state.table_active or state.achieved_min_x is None:
        return False
    if not state.current_outcome_ref:
        return False
    if block.x0 < state.achieved_min_x:
        return False
    if is_table_instruction(block.text):
        return False
    if starts_new_structure(block.text):
        return False
    return bool(one_line(block.text))


def parse_pdf(pdf_path: Path) -> list[ContentRow]:
    state = ParseState()
    inside_framework = False

    for block in iter_pdf_blocks(pdf_path):
        text = block.text
        normalized = one_line(text)

        if normalized == "The Cyber Assessment Framework":
            inside_framework = True
            continue
        if not inside_framework:
            continue

        if parse_structure_block(state, text):
            continue

        if activate_table(state, text):
            continue

        if should_collect_achieved_segment(state, block):
            state.current_segments.append(
                AchievedSegment(
                    page=block.page,
                    x0=block.x0,
                    y0=block.y0,
                    y1=block.y1,
                    text=block.text,
                )
            )
            continue

        if not state.table_active and block.x0 <= LEFT_TEXT_MAX_X:
            append_description(state, text)

    finalize_outcome_segments(state)
    return state.rows


def split_statements(segments: list[AchievedSegment]) -> list[str]:
    statements: list[str] = []
    current = ""

    for segment in segments:
        text = one_line(segment.text)
        if not text:
            continue

        if not current:
            current = text
            continue

        if is_statement_continuation(current, text):
            current = f"{current} {text}"
        else:
            statements.append(current)
            current = text

    if current:
        statements.append(current)

    return [cleanup_statement(statement) for statement in statements if statement.strip()]


def is_statement_continuation(current: str, next_text: str) -> bool:
    current = current.rstrip()
    if not current:
        return True

    if next_text[:1].islower():
        return True

    if current.endswith((".", "!", "?")):
        return False

    if current.endswith((",", ";", ":", "/", "-", "(", "[")):
        return True

    last_word_match = re.search(r"([A-Za-z]+)$", current)
    if last_word_match and last_word_match.group(1).lower() in CONTINUATION_TERMINALS:
        return True

    # Some CAF statements intentionally end without a full stop. If the next block
    # starts like a new sentence, treat it as the next IGP rather than merging it.
    if next_text[:1].isupper():
        return False

    return True


def cleanup_statement(statement: str) -> str:
    statement = one_line(statement)
    statement = re.sub(
        r"^All\s+(?:of\s+)?the\s+following\s+statements\s+are\s+true\s+",
        "",
        statement,
        flags=re.IGNORECASE,
    )
    statement = statement.replace("organisation's", "organisation's")
    statement = re.sub(r"\s+", " ", statement)
    return statement.strip()


def validate_rows(rows: list[ContentRow]) -> None:
    objective_count = sum(1 for row in rows if row.depth == 1)
    principle_count = sum(1 for row in rows if row.depth == 2)
    outcome_count = sum(1 for row in rows if row.depth == 3)
    assessable_count = sum(1 for row in rows if row.assessable == "x")

    expected = {
        "objectives": (objective_count, EXPECTED_OBJECTIVES),
        "principles": (principle_count, EXPECTED_PRINCIPLES),
        "outcomes": (outcome_count, EXPECTED_OUTCOMES),
        "achieved statements": (assessable_count, EXPECTED_ACHIEVED_STATEMENTS),
    }
    mismatches = [
        f"{name}: got {got}, expected {want}"
        for name, (got, want) in expected.items()
        if got != want
    ]
    if mismatches:
        raise ValueError("Unexpected parse result: " + "; ".join(mismatches))

    outcome_refs = {row.ref_id for row in rows if row.depth == 3 and row.ref_id}
    assessable_refs = {row.ref_id for row in rows if row.assessable == "x" and row.ref_id}
    outcomes_without_children = [
        ref_id
        for ref_id in sorted(outcome_refs)
        if not any(child_ref.startswith(f"{ref_id}.") for child_ref in assessable_refs)
    ]
    if outcomes_without_children:
        raise ValueError(
            "Unexpected parse result: outcomes without achieved statements: "
            + ", ".join(outcomes_without_children)
        )

    if assessable_count < outcome_count:
        raise ValueError(
            f"Unexpected parse result: got only {assessable_count} assessable "
            f"statements for {outcome_count} outcomes"
        )


def write_rows(ws: Worksheet, rows: list[tuple[str | int | None, ...]]) -> None:
    for row in rows:
        ws.append(row)


def style_meta_sheet(ws: Worksheet) -> None:
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 120
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def style_content_sheet(ws: Worksheet) -> None:
    for column, width in CONTENT_COLUMN_WIDTHS.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_workbook(rows: list[ContentRow]) -> Workbook:
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    library_ws = wb.create_sheet(LIBRARY_SHEET)
    write_rows(library_ws, LIBRARY_META_ROWS)
    style_meta_sheet(library_ws)

    framework_ws = wb.create_sheet(FRAMEWORK_SHEET)
    write_rows(framework_ws, FRAMEWORK_META_ROWS)
    style_meta_sheet(framework_ws)

    content_ws = wb.create_sheet(CONTENT_SHEET)
    content_ws.append(CONTENT_HEADERS)
    for row in rows:
        content_ws.append(
            [row.assessable, row.depth, row.ref_id, row.name, row.description]
        )
    style_content_sheet(content_ws)

    return wb


def build_caf_workbook(pdf_path: Path, output_path: Path) -> list[ContentRow]:
    rows = parse_pdf(pdf_path)
    validate_rows(rows)
    workbook = build_workbook(rows)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return rows


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build the NCSC CAF 4.0 CISO Assistant workbook from the PDF."
    )
    parser.add_argument(
        "-p", "--pdf",
        type=Path,
        default=default_pdf_path(),
        help=f"Source PDF path. Defaults to {PDF_FILENAME}",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=default_output_path(),
        help=f"Output workbook path. Defaults to {OUTPUT_FILENAME}",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    rows = build_caf_workbook(args.pdf, args.output)

    objective_count = sum(1 for row in rows if row.depth == 1)
    principle_count = sum(1 for row in rows if row.depth == 2)
    outcome_count = sum(1 for row in rows if row.depth == 3)
    assessable_count = sum(1 for row in rows if row.assessable == "x")

    print(f"Built {args.output}")
    print(
        "Parsed "
        f"{objective_count} objectives, {principle_count} principles, "
        f"{outcome_count} outcomes, {assessable_count} achieved statements."
    )


if __name__ == "__main__":
    main()
