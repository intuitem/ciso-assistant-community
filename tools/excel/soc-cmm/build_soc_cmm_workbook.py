"""
Build a CISO Assistant v2-format framework Excel file for SOC-CMM 2.4.2 (Basic),
from:
- the source scheme JSON (control titles EN/FR + aspect/domain names EN/FR)
- the official SOC-CMM Excel self-assessment tool (per-question maturity
  guidance text, found in the "_Guidance" sheet)

The resulting .xlsx is meant to be processed by tools/convert_library_v2.py
to produce the final YAML library.

Usage:
    python build_soc_cmm_workbook.py

Reads (relative to repo root):
    tracking/Librairies/scheme-2024-advanced.json
    tracking/Librairies/62-soc-cmm-242-basic.xlsx

Writes:
    tools/excel/soc-cmm/soc-cmm-2.4.2-basic.xlsx
"""

import json
import warnings
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore")

REPO_ROOT = Path(__file__).resolve().parents[3]
SCHEME_PATH = REPO_ROOT / "tracking" / "Librairies" / "scheme-2024-advanced.json"
SOURCE_XLSX = REPO_ROOT / "tracking" / "Librairies" / "62-soc-cmm-242-basic.xlsx"
OUTPUT_XLSX = Path(__file__).resolve().parent / "soc-cmm-2.4.2-basic.xlsx"

LIBRARY_SLUG = "soc-cmm-2.4.2-basic"
PACKAGER = "Nadia Qoudhadh"
PROVIDER = "SOC-CMM (Rob van Os)"
LIBRARY_NAME = "SOC-CMM 2.4.2 - Basic Self-Assessment"
LIBRARY_NAME_FR = "SOC-CMM 2.4.2 - Auto-évaluation de base"
LIBRARY_DESCRIPTION = (
    "SOC-CMM is a capability maturity model used to perform a self-assessment of a "
    "Security Operations Center (SOC). It evaluates 5 domains (Business, People, "
    "Process, Technology, Services) using a continuous 0-5 maturity scale, with no "
    "prerequisites between levels: every element adds individually to the maturity "
    "score.\nSource: https://www.soc-cmm.com/ (Rob van Os)"
)
LIBRARY_DESCRIPTION_FR = (
    "SOC-CMM est un modèle de maturité des capacités utilisé pour réaliser "
    "l'auto-évaluation d'un centre des opérations de sécurité (SOC). Il évalue 5 "
    "domaines (Business, People, Process, Technology, Services) sur une échelle de "
    "maturité continue de 0 à 5, sans prérequis entre les niveaux : chaque élément "
    "contribue individuellement au score de maturité.\nSource : https://www.soc-cmm.com/ (Rob van Os)"
)
COPYRIGHT = (
    "Copyright (C) 2025 - SOC-CMM. Licensed under CC BY-SA 4.0 "
    "(https://creativecommons.org/licenses/by-sa/4.0/)"
)

# Prefixes used as row keys in the official "_Guidance" sheet, mapped to the
# domain names used in control_schemas keys (verified against the source file:
# Business -> "B", People -> "P", Process -> "M", Services -> "S".
# Technology has no per-question guidance in the source file.
GUIDANCE_PREFIX_TO_DOMAIN = {
    "B": "Business",
    "P": "People",
    "M": "Process",
    "S": "Services",
}

MATURITY_SCALE = [
    (
        0,
        "Non-existent",
        "The process is not implemented or not recognized within the SOC.",
        "Inexistant",
        "Le processus n'est pas mis en œuvre ou pas reconnu au sein du SOC.",
    ),
    (
        1,
        "Initial",
        "The process is performed, but in an ad hoc and inconsistent manner.",
        "Initial",
        "Le processus est réalisé, mais de manière ad hoc et non systématique.",
    ),
    (
        2,
        "Managed",
        "The process is planned and executed according to a defined approach; "
        "performance is managed but not yet standardized.",
        "Géré",
        "Le processus est planifié et exécuté selon une approche définie ; la "
        "performance est gérée mais pas encore standardisée.",
    ),
    (
        3,
        "Defined",
        "The process is well characterized, documented, and standardized across the SOC.",
        "Défini",
        "Le processus est bien caractérisé, documenté et standardisé au sein du SOC.",
    ),
    (
        4,
        "Quantitatively managed",
        "The process is measured and controlled using quantitative techniques.",
        "Géré quantitativement",
        "Le processus est mesuré et piloté à l'aide de techniques quantitatives.",
    ),
    (
        5,
        "Optimizing",
        "The process is continuously improved based on quantitative feedback and "
        "innovative practices.",
        "Optimisé",
        "Le processus est amélioré en continu sur la base de retours quantitatifs "
        "et de pratiques innovantes.",
    ),
]


def parse_guidance(ws):
    """Parse the "_Guidance" sheet into {control_schema_key: [(level, text), ...]}.

    Layout: a header row holds the key (e.g. "B 1.1") in column A; the following
    rows hold the per-level guidance with column A empty, the level (1-5) in
    column B, and the guidance text in column C.
    """
    guidance = {}
    current_key = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        a = row[0] if len(row) > 0 else None
        b = row[1] if len(row) > 1 else None
        c = row[2] if len(row) > 2 else None
        if isinstance(a, str) and a.strip():
            text = a.strip()
            if text.startswith("SOC-CMM") or " - " in text:
                current_key = None
                continue
            parts = text.split(" ", 1)
            if len(parts) == 2 and parts[0] in GUIDANCE_PREFIX_TO_DOMAIN:
                domain = GUIDANCE_PREFIX_TO_DOMAIN[parts[0]]
                current_key = f"{domain}.{parts[1].strip()}"
                guidance[current_key] = []
            else:
                current_key = None
        elif (
            not a
            and isinstance(b, (int, float))
            and current_key
            and 1 <= b <= 5
            and isinstance(c, str)
        ):
            guidance[current_key].append((int(b), c.strip()))
    return guidance


def parent_key(key: str) -> str:
    return key.rsplit(".", 1)[0]


def build_rows(scheme: dict, guidance: dict) -> list[dict]:
    """
    Note: the source scheme has gaps in the dotted numbering (e.g. "Process.4.1.1"
    exists but its would-be parent "Process.4.1" is never itself defined as a
    control_schema entry). Depth is therefore derived purely from the number of
    dot-separated segments relative to the aspect key, and any missing
    intermediate ancestor is synthesized as an empty, non-assessable container
    row so the depth sequence stays contiguous (required for depth-based tree
    inference in convert_library_v2.py).
    """
    aspects = scheme["aspects"]
    aspects_fr = scheme["aspects_fr"]
    control_schemas = scheme["control_schemas"]

    rows: list[dict] = []
    emitted: set[str] = set()

    def emit_leaf(key: str, depth: int):
        row = {
            "assessable": "",
            "depth": depth,
            "ref_id": key,
            "name": "",
            "description": "",
            "name[fr]": "",
            "description[fr]": "",
        }
        meta = control_schemas.get(key)
        if meta is None:
            # synthesized placeholder for a gap in the source numbering
            row["name"] = f"[{key}]"
            row["name[fr]"] = f"[{key}]"
            rows.append(row)
            emitted.add(key)
            return
        control_type = meta["control_type"]
        if control_type == "Any":
            emitted.add(key)
            return  # excluded: free-text rationale/comment field, not a requirement
        assessable = control_type in ("Detailed", "Bool")
        title = meta["title"]
        title_fr = meta.get("title_fr", "")
        if assessable:
            desc = title
            desc_fr = title_fr
            if key in guidance:
                lines = "\n".join(f"{lvl} - {txt}" for lvl, txt in guidance[key])
                guidance_block = f"\n\nMaturity guidance:\n{lines}"
                desc += guidance_block
                desc_fr += guidance_block
            row["assessable"] = "x"
            row["description"] = desc
            row["description[fr]"] = desc_fr
        else:
            row["name"] = title
            row["name[fr]"] = title_fr
        rows.append(row)
        emitted.add(key)

    def ensure_ancestors_and_emit(key: str, aspect_ref: str):
        chain = []
        k = key
        while k != aspect_ref:
            chain.append(k)
            k = parent_key(k)
        chain.reverse()
        for ancestor in chain:
            if ancestor in emitted:
                continue
            depth = 2 + (ancestor.count(".") - aspect_ref.count("."))
            emit_leaf(ancestor, depth)

    for domain, aspect_list in aspects.items():
        rows.append(
            {
                "assessable": "",
                "depth": 1,
                "ref_id": domain,
                "name": domain,
                "description": "",
                # the source scheme has no translated domain-level names (only
                # aspect names are translated in aspects_fr); fall back to the
                # English domain name so the FR locale isn't left blank
                "name[fr]": domain,
                "description[fr]": "",
            }
        )
        aspect_list_fr = aspects_fr[domain]
        for idx, aspect_name in enumerate(aspect_list, start=1):
            aspect_ref = f"{domain}.{idx}"
            rows.append(
                {
                    "assessable": "",
                    "depth": 2,
                    "ref_id": aspect_ref,
                    "name": aspect_name,
                    "description": "",
                    "name[fr]": aspect_list_fr[idx - 1],
                    "description[fr]": "",
                }
            )
            emitted.add(aspect_ref)
            for key in control_schemas:
                if key == aspect_ref or not key.startswith(aspect_ref + "."):
                    continue
                if key in emitted:
                    continue
                ensure_ancestors_and_emit(key, aspect_ref)

    return rows


def main():
    with open(SCHEME_PATH, encoding="utf-8") as f:
        scheme = json.load(f)

    src_wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True)
    guidance = parse_guidance(src_wb["_Guidance"])

    rows = build_rows(scheme, guidance)
    print(f"Built {len(rows)} requirement rows")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("library_meta")
    for r in [
        ("type", "library"),
        ("urn", f"urn:intuitem:risk:library:{LIBRARY_SLUG}"),
        ("version", "1"),
        ("locale", "en"),
        ("ref_id", LIBRARY_SLUG),
        ("name", LIBRARY_NAME),
        ("description", LIBRARY_DESCRIPTION),
        ("copyright", COPYRIGHT),
        ("provider", PROVIDER),
        ("packager", PACKAGER),
        ("name[fr]", LIBRARY_NAME_FR),
        ("description[fr]", LIBRARY_DESCRIPTION_FR),
        ("copyright[fr]", COPYRIGHT),
    ]:
        ws.append(list(r))

    ws = wb.create_sheet("fwk_meta")
    for r in [
        ("type", "framework"),
        ("base_urn", f"urn:intuitem:risk:req_node:{LIBRARY_SLUG}"),
        ("urn", f"urn:intuitem:risk:framework:{LIBRARY_SLUG}"),
        ("ref_id", LIBRARY_SLUG),
        ("name", LIBRARY_NAME),
        ("description", LIBRARY_DESCRIPTION),
        ("scores_definition", "soccmm_scores"),
        ("min_score", "0"),
        ("max_score", "5"),
        ("name[fr]", LIBRARY_NAME_FR),
        ("description[fr]", LIBRARY_DESCRIPTION_FR),
    ]:
        ws.append(list(r))

    ws = wb.create_sheet("fwk_content")
    ws.append(["assessable", "depth", "ref_id", "name", "description", "name[fr]", "description[fr]"])
    for row in rows:
        ws.append(
            [
                row["assessable"],
                row["depth"],
                row["ref_id"],
                row["name"],
                row["description"],
                row["name[fr]"],
                row["description[fr]"],
            ]
        )

    ws = wb.create_sheet("scr_meta")
    for r in [("type", "scores"), ("name", "soccmm_scores")]:
        ws.append(list(r))

    ws = wb.create_sheet("scr_content")
    ws.append(["score", "name", "description", "name[fr]", "description[fr]"])
    for score, name, desc, name_fr, desc_fr in MATURITY_SCALE:
        ws.append([score, name, desc, name_fr, desc_fr])

    wb.save(OUTPUT_XLSX)
    print(f"Wrote {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
