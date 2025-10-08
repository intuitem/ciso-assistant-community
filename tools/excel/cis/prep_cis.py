"""
simple script to transform the official CIS Excel file to another Excel file for CISO assistant framework conversion tool
"""

import sys
import re
import openpyxl
import argparse
from openpyxl.utils.exceptions import InvalidFileException

parser = argparse.ArgumentParser(
    prog="convert_cis",
    description="convert CIS controls offical Excel file to CISO Assistant Excel file",
)
parser.add_argument("filename", help="name of CIS controls Excel file")
parser.add_argument("packager", help="name of packager entity")

args = parser.parse_args()
input_file_name = args.filename
packager = args.packager
output_file_name = "cis-controls-v8.xlsx"

print(f'⌛ Parsing "{input_file_name}"...')

# Define variable to load the dataframe
try:
    dataframe = openpyxl.load_workbook(input_file_name)
    print(f'✅ Excel file loaded successfully: "{input_file_name}"')

except FileNotFoundError:
    print(f'❌ [ERROR] File not found: "{input_file_name}"')
    sys.exit(1)
except PermissionError:
    print(f'❌ [ERROR] Permission denied while accessing "{input_file_name}"')
    sys.exit(1)
except InvalidFileException:
    print(f'❌ [ERROR] The file is not a valid Excel file: "{input_file_name}"')
    sys.exit(1)
except Exception as e:
    print(f'❌ [ERROR] Unexpected error while loading Excel file: {e}')
    sys.exit(1)
    
output_table = []
output_table_ref_ctrl = []

for tab in dataframe:
    print(f'⌛ Parsing tab "{tab.title}"...')
    title = tab.title
    if title == "License for Use":
        library_copyright = tab["B11"].value + "\n" + tab["B13"].value
    # The script previously ignored the controls sheet because the title wasn't an exact match 
    # in the latest version of the CIS Controls .xlsx file (e.g., "Controls V8.1.2" instead of "Controls V8").
    # Using startswith() allows the script to work with all version changes going forward.
    elif title.lower().startswith("controls v"):
        for row in tab:
            (control, safeguard, asset_type, sf, title, description, ig1, ig2, ig3) = (
                r.value for r in row
            )
            control = str(control).strip()
            if re.match(r"\d+", control):
                if not safeguard:
                    safeguard_index = 0
                    output_table.append(("", 1, control, title, description))
                else:
                    safeguard_index += 1
                    safeguard = f"{control},{safeguard_index}"
                    implementation_groups = (
                        "IG1,IG2,IG3" if ig1 else "IG2,IG3" if ig2 else "IG3"
                    )
                    
                    # "," replace by "." because "," is used as a separator in the "reference_controls" column
                    output_table.append(
                        ("x", 2, safeguard.replace(",", "."), title, description, implementation_groups, "1:"+safeguard.replace(",", "."))
                    )
                    output_table_ref_ctrl.append(
                        (safeguard.replace(",", "."), title, sf.strip().lower(), description)
                    )
    else:
        print(f'⏩ Ignored tab: "{title}"')


print(f'⌛ Generating "{output_file_name}"...')

wb_output = openpyxl.Workbook()
ws = wb_output.active

# Library & Framework Metadata
ws.title = "library_content"
ws.append(["library_urn", f"urn:{packager.lower()}:risk:library:cis-controls-v8"])
ws.append(["library_version", "1"])
ws.append(["library_locale", "en"])
ws.append(["library_ref_id", "CIS-Controls-v8"])
ws.append(["library_name", "CIS Controls v8"])
ws.append(["library_description", "CIS Controls v8"])
ws.append(["library_copyright", library_copyright])
ws.append(["library_provider", "CIS"])
ws.append(["library_packager", packager])
ws.append(["framework_urn", f"urn:{packager.lower()}:risk:framework:cis-controls-v8"])
ws.append(["framework_ref_id", "CIS-Controls-v8"])
ws.append(["framework_name", "CIS Controls v8"])
ws.append(["framework_description", "CIS Controls v8"])
ws.append(["reference_control_base_urn", "urn:intuitem:risk:function:cis-controls-v8", "1"])
ws.append(["tab", "controls", "requirements"])
ws.append(["tab", "imp_grp", "implementation_groups"])
ws.append(["tab", "ref_ctrl", "reference_controls"])

# Framework
ws1 = wb_output.create_sheet("controls")
ws1.append(
    ["assessable", "depth", "ref_id", "name", "description", "implementation_groups", "reference_controls"]
)
for row in output_table:
    ws1.append(row)

# Implementation Groups
ws2 = wb_output.create_sheet("imp_grp")
ws2.append(["ref_id", "name", "description"])
ws2.append(
    [
        "IG1",
        "IG1",
        "Minimum standard of information security for all enterprises.",
    ]
)
ws2.append(
    [
        "IG2",
        "IG2",
        "For enterprises managing IT infrastructure of multiple departments with differing risk profiles.",
    ]
)
ws2.append(["IG3", "IG3", "To secure sensitive and confidential data."])


# Reference Controls
ws3 = wb_output.create_sheet("ref_ctrl")
ws3.append(
    ["ref_id", "name", "csf_function", "description"]
)
for row in output_table_ref_ctrl:
    ws3.append(row)


try:
    wb_output.save(output_file_name)
    print(f'✅ Excel file saved successfully: "{output_file_name}"')
except PermissionError:
    print(f'❌ [ERROR] Permission denied. The file may be open or locked: "{output_file_name}"')
    sys.exit(1)
except FileNotFoundError:
    print(f'❌ [ERROR] Invalid path. Cannot save to: "{output_file_name}"')
    sys.exit(1)
except OSError as e:
    print(f'❌ [ERROR] OS error while saving the file: "{e}"')
    sys.exit(1)
except Exception as e:
    print(f'❌ [ERROR] Unexpected error while saving Excel file: "{e}"')
    sys.exit(1)