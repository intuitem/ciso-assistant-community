"""
Simple script to convert NIST SP-800-66 excel in a CISO Assistant Excel file
Source;  https://csrc.nist.gov/Projects/cprt/catalog#/cprt/framework/version/SP800_66_2_0_0/home
"""

import openpyxl
import re
import argparse

parser = argparse.ArgumentParser(
    prog="convert_nist-sp-800-66",
    description="convert NIST SP-800-66 controls official Excel file to CISO Assistant Excel file",
)

parser.add_argument("filename", help="name of official NIST SP-800-66 Excel file")
args = parser.parse_args()
input_file_name = args.filename
output_file_name = "nist-sp-800-66-rev2.xlsx"

library_copyright = """With the exception of material marked as copyrighted, information presented on NIST sites are considered public information and may be distributed or copied."""
packager = "intuitem"

library_description = """Implementing the Health Insurance Portability and Accountability Act (HIPAA) Security Rule: A Cybersecurity Resource Guide, 2.0.0
Source: https://csrc.nist.gov/Projects/cprt/catalog#/cprt/framework/version/SP800_66_2_0_0/home
"""

print("parsing", input_file_name)

# Load Excel input
dataframe = openpyxl.load_workbook(input_file_name)
output_table = []

for tab in dataframe:
    print("parsing tab", tab.title)
    title = tab.title
    if title == "NIST SP 800-66":
        line = 0
        for row in tab:
            line += 1
            if line > 1:
                (
                    security_rule_id,
                    security_rule,
                    std_id,
                    std,
                    key_activity,
                    description,
                    sample_questions,
                ) = (r.value for r in row)

                # Add main lines
                if security_rule_id:
                    output_table.append(["", 1, security_rule_id, None, security_rule])
                if std_id:
                    output_table.append(["", 2, std_id, None, std])

                # Beginning of description line
                current_row = ["x", 3, None, key_activity, description]

                # Cleaning and adding questions
                if sample_questions and isinstance(sample_questions, str):
                    # Add line breaks after each "?" followed by text
                    sample_questions = re.sub(r'\?(?=\s*\S)', '?\n', sample_questions)

                    # Cleaning and cutting questions
                    cleaned_lines = [line.strip() for line in sample_questions.splitlines() if line.strip()]
                    cleaned_questions = "\n".join(cleaned_lines)
                    current_row.append(cleaned_questions)

                    # Generation of the "answer" field line by line according to the interrogative word
                    answers = []
                    for row_num in cleaned_lines:
                        lowered = row_num.lower()

                        if re.match(r"^(when|since when|as of when)\b", lowered):
                            answers.append("DATE")
                        elif re.match(r"^(how|what|which|who|whose|explain|describe|provide|detail)\b", lowered):
                            answers.append("TEXT")
                        elif re.match(r"^(is|are|do|does|can|could|should|will|would|has|have|had)\b", lowered):
                            answers.append("YNNA")
                        else:
                            answers.append("TEXT")  # By Default

                    current_row.append("\n".join(answers))
                else:
                    current_row.append(None)
                    current_row.append(None)

                # Add "implementation_groups" column
                if key_activity:
                    if "(Required)" in key_activity:
                        current_row.append("R")
                    elif "(Addressable)" in key_activity:
                        current_row.append("A")
                    else:
                        current_row.append("N")
                else:
                    current_row.append(None)

                output_table.append(current_row)

print("generating", output_file_name)
wb_output = openpyxl.Workbook()

# ===== Sheet: library_content =====
ws = wb_output.active
ws.title = "library_content"
ws.append(["library_urn", f"urn:{packager.lower()}:risk:library:nist-sp-800-66-rev2"])
ws.append(["library_version", "2"])
ws.append(["library_locale", "en"])
ws.append(["library_ref_id", "NIST-SP-800-66-rev2"])
ws.append(["library_name", "NIST SP-800-66 rev2 (HIPAA)"])
ws.append(["library_description", library_description])
ws.append(["library_copyright", library_copyright])
ws.append(["library_provider", "NIST"])
ws.append(["library_packager", packager])
ws.append(
    ["framework_urn", f"urn:{packager.lower()}:risk:framework:nist-sp-800-66-rev2"]
)
ws.append(["framework_ref_id", "nist-sp-800-66-rev2"])
ws.append(["framework_name", "NIST SP-800-66 rev2 (HIPAA)"])
ws.append(["framework_description", library_description])
ws.append(["tab", "controls", "requirements"])
ws.append(["tab", "answers", "answers"])
ws.append(["tab", "imp_grp", "implementation_groups"])

# ===== Sheet: controls =====
ws1 = wb_output.create_sheet("controls")
ws1.append(["assessable", "depth", "ref_id", "name", "description", "questions", "answer", "implementation_groups"])

for row in output_table:
    ws1.append(row)

# ===== Sheet: answers =====
ws2 = wb_output.create_sheet("answers")
ws2.append(["id", "question_type", "question_choices"])
ws2.append(["YNNA", "unique_choice", "Yes\nNo\nN/A"])
ws2.append(["TEXT", "text", None])
ws2.append(["DATE", "date", None])

# ===== Sheet: imp_grp =====
ws3 = wb_output.create_sheet("imp_grp")
ws3.append(["ref_id", "name", "description"])
ws3.append(["R", "Required", ""])
ws3.append(["A", "Addressable", ""])
ws3.append(["N", "Neutral", ""])

print("generate", output_file_name)
wb_output.save(output_file_name)
