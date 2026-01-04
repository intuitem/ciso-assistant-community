"""
Script that numbers the requirement names correctly in the Excel framework file.
"""


from openpyxl import load_workbook

# Load the Excel file
wb = load_workbook("nist-sp-800-66-rev2.xlsx")  # Replace with your actual file name
ws = wb.active

# Column E = column index 5
col_index = 5
counter = 1
reset = False

# Start from row 2 to skip the header
for row in range(2, ws.max_row + 1):
    cell = ws.cell(row=row, column=col_index)
    value = str(cell.value).strip() if cell.value is not None else ""

    if value == "":
        # Empty cell → reset counter for next valid cell
        reset = True
        continue

    if "(" in value and ")" in value:
        # Cell contains parentheses → skip numbering, but don't reset counter
        continue

    if reset:
        # Restart numbering after an empty cell
        counter = 1
        reset = False

    # Add numbering to the beginning of the cell content
    cell.value = f"{counter}. {value}"
    counter += 1

# Save the modified file
wb.save("nist-sp-800-66-rev2_v2.xlsx")

