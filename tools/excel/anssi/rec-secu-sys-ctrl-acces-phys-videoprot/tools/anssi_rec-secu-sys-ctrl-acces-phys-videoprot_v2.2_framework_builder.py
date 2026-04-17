#!/usr/bin/env python3
"""
Build an XLSX extraction of the ANSSI v2.2 PDF recommendations with a
self-contained hybrid parser:

- Ghostscript remains the source of truth for section hierarchy, recommendation
  ordering, and bibliography links.
- pypdf is used to extract recommendation text blocks with better word spacing.

This script is intentionally independent from the legacy builder.
"""

from __future__ import annotations

import argparse
import html
import re
import subprocess
from dataclasses import dataclass
from pathlib import Path
from textwrap import dedent
from typing import Iterable

from openpyxl import Workbook
from pypdf import PdfReader


SCRIPT_DIR = Path(__file__).resolve().parent
FRAMEWORK_DIR = SCRIPT_DIR.parent
DEFAULT_PDF = (
    FRAMEWORK_DIR
    / "recommandations-sur-la-sécurisation-des-systèmes-de-contrôle-d_acces-physique-et-vidéoprotection-v2.2.pdf"
)
DEFAULT_OUTPUT = (
    FRAMEWORK_DIR
    / "anssi_rec-secu-sys-ctrl-acces-phys-videoprot_v2.2.xlsx"
)
DEFAULT_GS_TEXT_FORMAT = 3


# ============================================================================
# Workbook and sheet metadata constants
# ============================================================================

SHEET_NAME = "reco_content"
COLUMNS = [
    "assessable",
    "depth",
    "node_id",
    "ref_id",
    "name",
    "description",
    "annotation",
    "implementation_groups",
]
IMP_GRP_COLUMNS = ["ref_id", "name", "description"]
URN_ROOT = "anssi_rec-secu-sys-ctrl-acces-phys-videoprot_v2.2"
FRAMEWORK_REF_ID = "ANSSI_rec-secu-sys-ctrl-acces-phys-videoprot_v2.2"
FRAMEWORK_NAME = (
    "ANSSI - Sécurisation des systèmes de contrôle d’accès physique et vidéoprotection "
    "(v2.2)"
)
FRAMEWORK_DESCRIPTION = dedent(
    """
    Ce guide développe les aspects d’architecture et de sécurité logique propres aux systèmes de contrôle d’accès utilisant des technologies sans contact, et aux systèmes de vidéoprotection. Comme lors de la rédaction du premier guide, l’ANSSI s’est associée au CNPP pour mener une réflexion intégrant l’ensemble des éléments qui composent ces systèmes. Ce guide est ainsi complémentaire aux référentiels CNPP intitulés :
      * « Référentiel APSAD D83 – Contrôle d’accès – Document technique pour la conception et l’installation » ;
      * « Référentiel APSAD R82 – Vidéosurveillance – Règle d’installation » ;
      * « Référentiel APSAD D32 – Cybersécurité – Document technique pour l’installation de systèmes de sécurité ou de sûreté sur un réseau informatique ».

    Ce guide s’adresse :
      * aux chefs de projet ou personnes en charge de la mise en place d’un système de contrôle d’accès sans contact ou de vidéoprotection, que ce soit dans une entreprise privée ou un organisme public ;
      * aux acheteurs, qui pourront imposer dans leurs appels d’offres les exigences détaillées en annexe afin de les rendre contraignantes pour le fournisseur ;
      * aux installateurs ou intégrateurs, qui pourront tenir compte du contenu de ce guide afin de proposer des services adaptés ;
      * aux exploitants, qui pourront s’intéresser aux aspects liés à l’exploitation et la maintenance du système.

    Le guide est accompagné d'un outil d'aide et de suivi d'implémentation de la sécurité (OASIS) sous forme de tableur. Ce document aide les entités à donner des priorités aux recommandations, à les classer selon une approche graduelle, puis à faire le suivi de leur mise en œuvre. Chaque entité peut adapter cette approche à ses enjeux, moyens, compétences et SI existants.

    Source : https://messervices.cyber.gouv.fr/guides/securisation-des-systemes-de-controle-dacces-physique-et-videoprotection
    """
).strip()
LIBRARY_META_ROWS: tuple[tuple[str, str], ...] = (
    ("type", "library"),
    ("urn", f"urn:intuitem:risk:library:{URN_ROOT}"),
    ("version", "1"),
    ("locale", "fr"),
    ("ref_id", FRAMEWORK_REF_ID),
    ("name", FRAMEWORK_NAME),
    ("description", FRAMEWORK_DESCRIPTION),
    ("copyright", "Licence Ouverte v2.0 - Etalab"),
    ("provider", "ANSSI"),
    ("packager", "intuitem"),
)
RECO_META_ROWS: tuple[tuple[str, str], ...] = (
    ("type", "framework"),
    ("base_urn", f"urn:intuitem:risk:req_node:{URN_ROOT}"),
    ("urn", f"urn:intuitem:risk:framework:{URN_ROOT}"),
    ("ref_id", FRAMEWORK_REF_ID),
    ("name", FRAMEWORK_NAME),
    ("description", FRAMEWORK_DESCRIPTION),
    ("implementation_groups_definition", "imp_grp"),
)
IMP_GRP_META_ROWS: tuple[tuple[str, str], ...] = (
    ("type", "implementation_groups"),
    ("name", "imp_grp"),
)
IMP_GRP_CONTENT_ROWS: tuple[tuple[str, str, str], ...] = (
    (
        "R",
        "[R] Recommandation à l'état de l'art",
        "Cette recommandation permet de mettre en œuvre un niveau de sécurité à l'état de l'art.",
    ),
    (
        "R-",
        "[R-] Recommandation alternative de premier niveau",
        "Cette recommandation permet de mettre en œuvre une première alternative, d'un niveau de sécurité moindre que la recommandation R.",
    ),
    (
        "R+",
        "[R+] Recommandation renforcée",
        "Cette recommandation permet de mettre en œuvre un niveau de sécurité renforcé. Elle est destinée aux entités qui sont matures en sécurité des systèmes d'information.",
    ),
)


# ============================================================================
# PDF extraction and parsing constants
# ============================================================================

SECTION_RE = re.compile(r"^\s*(\d+(?:\.\d+)*)\s+(.+?)\s*$")
PURE_NUMBER_RE = re.compile(r"^\s*(\d+)\s*$")
RECOMMENDATION_RE = re.compile(r"^\s*(R\d+[+-]?)\s*$")
PAGE_HEADER_RE = re.compile(r"RECOMMANDATIONSSURLASÉCURISATION", re.IGNORECASE)
TOC_PAGE_NUMBER_RE = re.compile(r"\s+\d+\s*$")
TOC_DOTS_RE = re.compile(r"\s*\.\s*(?:\.\s*)+\d+\s*$")
MAX_HEADING_INDENT = 40
BULLET_LINE_RE = re.compile(r"^(?:n|>|•|▪|■|-)\s+(.+)$")
ANNOTATION_LABELS = {"Attention", "Information"}
BIBLIOGRAPHY_ENTRY_RE = re.compile(r"^\s*\[(\d+)\]\s*(.*)$")
REFERENCE_MARKER_RE = re.compile(r"\[(\d+)\]")
URL_START_RE = re.compile(r"(https?://\S+|www\.\S+)")
KNOWN_TEXT_FIXES: tuple[tuple[str, str], ...] = (
    ("gestioncommun", "gestion commun"),
    ("administrationdu", "administration du"),
)

PYPDF_RECOMMENDATION_RE = re.compile(r"^\s*R\s*((?:\d+\s*)+)\s*([+-])?\s*$")
PYPDF_HEADER_RE = re.compile(
    r"RECOMMANDATIONS\s+SUR\s+LA\s+S[ÉE]CURISATION", re.IGNORECASE
)
PYPDF_FOOTER_RE = re.compile(
    r"^\s*\d+\s*[–-]\s*RECOMMANDATIONS\s+SUR\s+LA\s+S[ÉE]CURISATION",
    re.IGNORECASE,
)
# Footnotes at the start of a line, e.g. "23. Délibération ..."
PYPDF_FOOTNOTE_START_RE = re.compile(r"^\s*(\d{1,2})\.\s+(.+)$")
# Inline footnote calls, excluding bibliography references like "[n]".
INLINE_FOOTNOTE_RE = re.compile(
    r"(?<!\[)(?<![\w])(\d{1,2})(?=(?:\s*[.,;:!?])|[A-Za-zÀ-ÿ])"
)


@dataclass
class Section:
    node_id: str
    depth: int
    name: str


@dataclass
class Recommendation:
    ref_id: str
    name: str
    description: str
    annotation: str
    section_id: str

    @property
    def node_id(self) -> str:
        match = re.fullmatch(r"R(\d+)([+-]?)", self.ref_id)
        if not match:
            raise ValueError(f"Unsupported recommendation id: {self.ref_id}")
        number, suffix = match.groups()
        suffix_map = {"+": "p", "-": "m", "": ""}
        return f"r{number}{suffix_map[suffix]}"

    @property
    def implementation_groups(self) -> str:
        if self.ref_id.endswith("+"):
            return "R+"
        if self.ref_id.endswith("-"):
            return "R-"
        return "R"


@dataclass
class PypdfRecommendationBlock:
    recommendation: Recommendation
    pages: tuple[int, ...]


# ============================================================================
# CLI
# ============================================================================

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--pdf", type=Path, default=DEFAULT_PDF)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    return parser.parse_args()


# ============================================================================
# PDF extraction and parsing functions
# ============================================================================

def extract_pdf_text(pdf_path: Path) -> str:
    if not pdf_path.exists():
        raise FileNotFoundError(f"PDF not found: {pdf_path}")

    # Ghostscript txtwrite keeps indentation cues that are useful to recover
    # section and recommendation boundaries.
    command = [
        "gs",
        "-q",
        "-dNOPAUSE",
        "-dBATCH",
        "-sDEVICE=txtwrite",
        f"-dTextFormat={DEFAULT_GS_TEXT_FORMAT}",
        "-sOutputFile=-",
        str(pdf_path),
    ]
    result = subprocess.run(command, capture_output=True, text=True, check=True)
    return result.stdout.replace("\r\n", "\n")


def normalize_line(line: str) -> str:
    return (
        line.replace("\xa0", " ")
        .replace("’", "'")
        .replace("–", "-")
        .replace("œ", "oe")
        .rstrip()
    )


def cleanup_text(text: str) -> str:
    # Drop inline footnote markers glued to lowercase words while keeping
    # meaningful identifiers like R42 or AVA_VAN.5 intact.
    text = re.sub(r"(?<=[a-zà-ÿ)])\d+(?=\s)", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    text = re.sub(r"\s+([,.;:!?])", r"\1", text)
    text = re.sub(r"\(\s+", "(", text)
    text = re.sub(r"\s+\)", ")", text)
    return text


def is_page_noise(line: str) -> bool:
    stripped = normalize_line(line).strip()
    if not stripped:
        return False
    if PAGE_HEADER_RE.search(stripped):
        return True
    if stripped.startswith("FIGURE"):
        return True
    if re.fullmatch(r"\d+\s*-\s*RECOMMANDATIONSSURLASÉCURISATION.*", stripped):
        return True
    if re.fullmatch(r"RECOMMANDATIONSSURLASÉCURISATION.*-\s*\d+", stripped):
        return True
    return False


def indentation(line: str) -> int:
    return len(line) - len(line.lstrip(" "))


def merge_hyphenated_lines(prefix: str, line: str) -> str:
    stripped = normalize_line(line).strip()
    if not stripped:
        return prefix
    first_word, separator, remainder = stripped.partition(" ")
    merged = prefix + first_word
    if separator and remainder:
        merged += " " + remainder
    return merged


def join_wrapped_lines(lines: Iterable[str]) -> str:
    text = ""
    for raw_line in lines:
        line = normalize_line(raw_line).strip()
        if not line:
            continue
        if not text:
            text = line
            continue
        # Rejoin PDF line wraps while keeping punctuation and hyphenated words
        # readable in the final cells.
        if text.endswith("-"):
            text = merge_hyphenated_lines(text[:-1], line)
        elif line.startswith((".", ",", ";", ":", ")", "]")):
            text += line
        else:
            text += " " + line
    return cleanup_text(text)


def format_content_lines(lines: Iterable[str]) -> str:
    blocks: list[tuple[str, list[str]]] = []
    current_kind: str | None = None
    current_lines: list[str] = []

    def flush_current() -> None:
        nonlocal current_kind, current_lines
        if current_kind and current_lines:
            blocks.append((current_kind, current_lines[:]))
        current_kind = None
        current_lines = []

    for raw_line in lines:
        stripped = normalize_line(raw_line).strip()
        if not stripped:
            flush_current()
            continue

        bullet_match = BULLET_LINE_RE.match(stripped)
        if bullet_match:
            flush_current()
            current_kind = "bullet"
            current_lines = [bullet_match.group(1)]
            continue

        if current_kind is None:
            current_kind = "paragraph"
            current_lines = [stripped]
            continue

        current_lines.append(stripped)

    flush_current()

    rendered: list[tuple[str, str]] = []
    for kind, block_lines in blocks:
        text = join_wrapped_lines(block_lines)
        if not text:
            continue
        if kind == "bullet":
            rendered.append((kind, f"* {text}"))
        else:
            rendered.append((kind, text))

    result = ""
    previous_kind: str | None = None
    for kind, text in rendered:
        if not result:
            result = text
        elif kind == "bullet" or previous_kind == "bullet":
            result += "\n" + text
        else:
            result += " " + text
        previous_kind = kind

    return result


def normalize_inline_spacing(text: str) -> str:
    if not text:
        return text

    def fix_text(value: str) -> str:
        for source, target in KNOWN_TEXT_FIXES:
            value = value.replace(source, target)
        return value

    normalized_lines: list[str] = []
    for raw_line in text.splitlines():
        if not raw_line.strip():
            normalized_lines.append("")
            continue
        bullet_prefix_match = re.match(r"^\t?([*-])\s+", raw_line)
        if bullet_prefix_match:
            bullet_prefix = bullet_prefix_match.group(0)
            bullet_content = raw_line[len(bullet_prefix) :]
            bullet_content = re.sub(r"(?<=[,.;:!?])(?=[A-Za-zÀ-ÿ])", " ", bullet_content)
            bullet_content = re.sub(r"(?<=[»\]])(?=[A-Za-zÀ-ÿ])", " ", bullet_content)
            normalized_lines.append(bullet_prefix + fix_text(cleanup_text(bullet_content)))
            continue
        line = re.sub(r"(?<=[,.;:!?])(?=[A-Za-zÀ-ÿ])", " ", raw_line)
        line = re.sub(r"(?<=[»\]])(?=[A-Za-zÀ-ÿ])", " ", line)
        normalized_lines.append(fix_text(cleanup_text(line)))
    return "\n".join(normalized_lines)


def load_lines(text: str) -> list[str]:
    return [normalize_line(line) for line in text.splitlines()]


def extract_toc_lines(lines: list[str]) -> list[str]:
    start = None
    end = None
    for index, line in enumerate(lines):
        if "Table des matières" in line:
            start = index + 1
            continue
        # The first bibliography heading after the table of contents is used
        # as the TOC closing marker in this PDF.
        if start is not None and "Bibliographie" in line:
            end = index
            break
    if start is None or end is None:
        raise ValueError("Unable to isolate the table of contents in the PDF extraction")
    return lines[start:end]


def extract_bibliography_reference_links(lines: list[str]) -> dict[str, str]:
    start = None
    for index, line in enumerate(lines):
        if "Bibliographie" in line:
            start = index + 1
            break

    if start is None:
        return {}

    entries: dict[str, list[str]] = {}
    current_ref: str | None = None
    current_lines: list[str] = []

    for index in range(start, len(lines)):
        line = lines[index]
        if is_page_noise(line):
            continue
        stripped = line.strip()
        if not stripped:
            continue
        if stripped.startswith("Version 2.2"):
            break

        match = BIBLIOGRAPHY_ENTRY_RE.match(line)
        if match:
            if current_ref is not None:
                entries[current_ref] = current_lines[:]
            current_ref = match.group(1)
            current_lines = [match.group(2).strip()] if match.group(2).strip() else []
            continue

        if current_ref is not None:
            current_lines.append(stripped)

    if current_ref is not None:
        entries[current_ref] = current_lines[:]

    reference_links: dict[str, str] = {}
    for ref_id, entry_lines in entries.items():
        # Only bibliography markers like [12] are converted to markdown links.
        url = extract_url_from_bibliography_entry(entry_lines)
        if url:
            reference_links[ref_id] = url
    return reference_links


def extract_url_from_bibliography_entry(entry_lines: list[str]) -> str | None:
    collecting = False
    url_parts: list[str] = []

    for line in entry_lines:
        match = URL_START_RE.search(line)
        if match:
            collecting = True
            url_parts.append(match.group(1))
            continue
        if collecting:
            compact = line.strip()
            if compact:
                url_parts.append(compact)

    if not url_parts:
        return None

    url = "".join(url_parts).replace(" ", "")
    while url.endswith((".", ",", ";", ")")):
        url = url[:-1]
    if url.startswith("www."):
        url = "https://" + url
    return url


def clean_toc_title_piece(text: str) -> str:
    text = TOC_DOTS_RE.sub("", text)
    text = TOC_PAGE_NUMBER_RE.sub("", text)
    return cleanup_text(text)


def linkify_reference_markers(text: str, reference_links: dict[str, str]) -> str:
    if not text:
        return text

    def replace(match: re.Match[str]) -> str:
        ref_id = match.group(1)
        url = reference_links.get(ref_id)
        if not url:
            return match.group(0)
        return f"[[{ref_id}]]({url})"

    return REFERENCE_MARKER_RE.sub(replace, text)


def parse_sections_from_toc(lines: list[str]) -> list[Section]:
    sections: list[Section] = []
    index = 0
    while index < len(lines):
        line = lines[index]
        if is_page_noise(line):
            index += 1
            continue
        match = SECTION_RE.match(line)
        if not match:
            index += 1
            continue

        node_id, title = match.groups()
        # TOC titles are often wrapped on multiple lines, so keep consuming
        # text until the next numbered heading.
        parts = [clean_toc_title_piece(title)]
        index += 1
        while index < len(lines):
            next_line = lines[index]
            if is_page_noise(next_line):
                index += 1
                continue
            if SECTION_RE.match(next_line) or next_line.strip().startswith("Annexe"):
                break
            stripped = next_line.strip()
            if stripped:
                parts.append(clean_toc_title_piece(stripped))
            index += 1

        full_title = cleanup_text(" ".join(part for part in parts if part))
        if node_id == "1" or node_id.startswith("Annexe"):
            continue
        sections.append(
            Section(node_id=node_id, depth=node_id.count(".") + 1, name=full_title)
        )

    return [
        section
        for section in sections
        if section.node_id == "2" or int(section.node_id.split(".")[0]) >= 2
    ]


def section_ids(sections: list[Section]) -> set[str]:
    return {section.node_id for section in sections}


def body_slice(lines: list[str]) -> tuple[int, int]:
    start = None
    end = None
    after_toc = False
    for index, line in enumerate(lines):
        if "Bibliographie" in line:
            after_toc = True
            continue
        # Start at the first standalone "2" after the front matter.
        if (
            start is None
            and after_toc
            and PURE_NUMBER_RE.match(line)
            and line.strip() == "2"
        ):
            start = index
        # Stop before Annex A, which is out of scope for the framework.
        if start is not None and "AnnexeA" in line.replace(" ", ""):
            end = index
            break
    if start is None or end is None:
        raise ValueError("Unable to isolate the PDF body from chapter 2 to Annex A")
    return max(start, 0), end


def detect_heading_occurrences(
    lines: list[str], sections: list[Section], start: int, end: int
) -> list[tuple[int, str]]:
    known_ids = section_ids(sections)
    occurrences: list[tuple[int, str]] = []

    for index in range(start, end):
        line = lines[index]
        stripped = line.strip()
        if not stripped or is_page_noise(line):
            continue

        match = SECTION_RE.match(line)
        if match and indentation(line) <= MAX_HEADING_INDENT:
            node_id = match.group(1)
            if node_id in known_ids:
                occurrences.append((index, node_id))

    # Deduplicate repeated Ghostscript echoes so recommendation-to-section
    # lookup stays stable.
    deduped: list[tuple[int, str]] = []
    last_id = None
    for occurrence in sorted(occurrences):
        if occurrence[1] != last_id:
            deduped.append(occurrence)
            last_id = occurrence[1]
    return deduped


def is_recommendation_line(line: str) -> bool:
    return bool(RECOMMENDATION_RE.match(line))


def is_section_heading_line(line: str, known_ids: set[str]) -> bool:
    match = SECTION_RE.match(line)
    return bool(
        match and indentation(line) <= MAX_HEADING_INDENT and match.group(1) in known_ids
    )


def is_footnote_or_marker(line: str) -> bool:
    stripped = cleanup_text(line)
    return bool(re.fullmatch(r"\d+", stripped) or re.fullmatch(r"\d+\..*", stripped))


def next_boundary_index(
    lines: list[str], start_index: int, end_index: int, known_ids: set[str]
) -> int:
    index = start_index + 1
    while index < end_index:
        line = lines[index]
        if is_recommendation_line(line) or is_section_heading_line(line, known_ids):
            return index
        index += 1
    return end_index


def detect_content_indent(lines: list[str], start_index: int, end_index: int) -> int | None:
    for probe in range(start_index, end_index):
        line = lines[probe]
        stripped = line.strip()
        if not stripped or is_page_noise(line) or is_footnote_or_marker(line):
            continue
        return indentation(line)
    return None


def looks_like_title_continuation(line: str) -> bool:
    stripped = cleanup_text(line)
    if not stripped:
        return False
    first_token = stripped.split()[0]
    acronym_like = bool(re.fullmatch(r"[A-Z0-9][A-Z0-9-]{1,3}", first_token))
    return stripped[:1].islower() or acronym_like


def find_active_section(
    line_index: int, heading_occurrences: list[tuple[int, str]]
) -> str | None:
    current_section = None
    for heading_line, section_id in heading_occurrences:
        if heading_line >= line_index:
            break
        current_section = section_id
    return current_section


def collect_title_prefix(
    lines: list[str], rec_line_index: int, content_indent: int, known_ids: set[str]
) -> list[str]:
    prefix: list[str] = []
    index = rec_line_index - 1
    max_prefix_lines = 2
    # In Ghostscript, the recommendation title may sit just above the "Rxx"
    # line with the same indentation as the content block.
    while index >= 0:
        line = lines[index]
        stripped = line.strip()
        if (
            not stripped
            or is_page_noise(line)
            or is_recommendation_line(line)
            or is_section_heading_line(line, known_ids)
        ):
            break
        if indentation(line) < content_indent:
            break
        prefix.insert(0, stripped)
        if len(prefix) >= max_prefix_lines:
            break
        index -= 1
    while len(prefix) > 1 and prefix[0].endswith((".", ";", ":")):
        prefix.pop(0)
    return prefix


def split_description_and_annotation(
    block_lines: list[str],
) -> tuple[list[str], list[tuple[str, list[str]]]]:
    description_lines: list[str] = []
    annotation_blocks: list[tuple[str, list[str]]] = []
    current_annotation_label: str | None = None
    current_annotation_lines: list[str] = []

    for raw_line in block_lines:
        stripped = raw_line.strip()
        if not stripped or is_page_noise(raw_line):
            continue
        # "Attention" and "Information" open dedicated annotation blocks.
        if stripped in ANNOTATION_LABELS:
            if current_annotation_label is not None:
                annotation_blocks.append(
                    (current_annotation_label, current_annotation_lines[:])
                )
            current_annotation_label = stripped
            current_annotation_lines = []
            continue
        if current_annotation_label is not None:
            current_annotation_lines.append(stripped)
        else:
            description_lines.append(stripped)

    if current_annotation_label is not None:
        annotation_blocks.append((current_annotation_label, current_annotation_lines[:]))

    return description_lines, annotation_blocks


def format_annotation_blocks(blocks: list[tuple[str, list[str]]]) -> str:
    rendered_blocks: list[str] = []
    for label, lines in blocks:
        content = format_content_lines(lines)
        if not content:
            continue
        rendered_blocks.append(f"## {label}\n{content}")
    return "\n\n".join(rendered_blocks)


def normalize_pypdf_line(line: str) -> str:
    line = normalize_line(line)
    # pypdf sometimes leaves stray spaces in apostrophes and ids.
    line = re.sub(r"(?<=\b[A-Za-zÀ-ÿ])\s+'", "'", line)
    line = re.sub(r"'\s+(?=[A-Za-zÀ-ÿ])", "'", line)
    line = re.sub(r"\[\s+(\d+)\s*\]", r"[\1]", line)
    line = re.sub(r"\b(\d+)\s+\.\s+(\d+)\b", r"\1.\2", line)
    line = re.sub(r"(?<=[a-zà-ÿ])(?=R\d)", " ", line)
    line = re.sub(r"\bR\s+(\d)", r"R\1", line)
    line = re.sub(r"\s+([+-])$", r"\1", line)
    return line.rstrip()


def normalize_pypdf_references(text: str) -> str:
    if not text:
        return text
    return re.sub(r"\[\s+(\d+)\s*\]", r"[\1]", text)


def prepare_section_text_pypdf(text: str, reference_links: dict[str, str]) -> str:
    text = normalize_pypdf_references(text)
    text = normalize_inline_spacing(text)
    text = re.sub(r"(?<=\b[A-Za-zÀ-ÿ])\s+'", "'", text)
    text = re.sub(r"'\s+(?=[A-Za-zÀ-ÿ])", "'", text)
    return linkify_reference_markers(text, reference_links)


def extract_pypdf_page_texts(pdf_path: Path) -> list[tuple[int, str]]:
    reader = PdfReader(str(pdf_path))
    return [
        (page_number, (page.extract_text() or "").replace("\r\n", "\n"))
        for page_number, page in enumerate(reader.pages, start=1)
    ]


def extract_clean_top_level_section_titles_from_pypdf(
    pdf_path: Path, sections: list[Section]
) -> dict[str, str]:
    top_level_ids = {section.node_id for section in sections if section.depth == 1}
    titles: dict[str, str] = {}

    for _, page_text in extract_pypdf_page_texts(pdf_path):
        lines = [
            normalize_pypdf_line(line).strip()
            for line in page_text.splitlines()
            if line.strip()
        ]
        filtered_lines = [
            line for line in lines if line and not is_pypdf_header_or_footer(line)
        ]
        if not filtered_lines:
            continue

        # Top-level chapter headings appear at the top of their page in the body.
        for index, line in enumerate(filtered_lines[:3]):
            if line not in top_level_ids or line in titles:
                continue

            title_lines: list[str] = []
            for candidate in filtered_lines[index + 1 : index + 6]:
                if (
                    not candidate
                    or SECTION_RE.match(candidate)
                    or RECOMMENDATION_RE.match(candidate)
                    or candidate.startswith("n ")
                    or len(candidate) > 70
                ):
                    break
                title_lines.append(candidate)

            if title_lines:
                titles[line] = join_wrapped_lines(title_lines)
            break

    return titles


def merge_clean_section_titles(
    sections: list[Section], clean_titles: dict[str, str]
) -> list[Section]:
    merged_sections: list[Section] = []
    for section in sections:
        merged_sections.append(
            Section(
                node_id=section.node_id,
                depth=section.depth,
                name=clean_titles.get(section.node_id, section.name),
            )
        )
    return merged_sections


def is_pypdf_page_noise(line: str) -> bool:
    stripped = line.strip()
    if not stripped:
        return False
    if stripped.startswith("FIGURE"):
        return True
    if PYPDF_HEADER_RE.search(stripped):
        return True
    if PYPDF_FOOTER_RE.match(stripped):
        return True
    return False


def is_pypdf_header_or_footer(line: str) -> bool:
    stripped = line.strip()
    if not stripped:
        return False
    if PYPDF_HEADER_RE.search(stripped):
        return True
    if PYPDF_FOOTER_RE.match(stripped):
        return True
    return False


def split_pypdf_page(page_text: str) -> tuple[list[str], dict[int, str]]:
    lines = [normalize_pypdf_line(line) for line in page_text.splitlines()]

    footer_index = len(lines)
    for index, line in enumerate(lines):
        # Keep figure captions in the body: they can appear before a recommendation
        # on the same page and should not truncate the rest of the content.
        if is_pypdf_header_or_footer(line):
            footer_index = index
            break

    content_lines = [line for line in lines[:footer_index] if line.strip()]
    footnote_starts = [
        index
        for index, line in enumerate(content_lines)
        if PYPDF_FOOTNOTE_START_RE.match(line)
    ]

    if not footnote_starts:
        return content_lines, {}

    # Split the page into two blocks:
    # - the body used for recommendation parsing
    # - the footnotes, stored separately for <abbr> output
    first_footnote_index = footnote_starts[0]
    body_lines = content_lines[:first_footnote_index]
    footnote_lines = content_lines[first_footnote_index:]
    relative_starts = [
        index - first_footnote_index
        for index in footnote_starts
        if index >= first_footnote_index
    ]

    footnotes: dict[int, str] = {}
    for offset, start_index in enumerate(relative_starts):
        end_index = (
            relative_starts[offset + 1]
            if offset + 1 < len(relative_starts)
            else len(footnote_lines)
        )
        chunk = footnote_lines[start_index:end_index]
        if not chunk:
            continue

        match = PYPDF_FOOTNOTE_START_RE.match(chunk[0])
        if not match:
            continue

        number = int(match.group(1))
        content_lines = [match.group(2), *chunk[1:]]
        footnotes[number] = join_wrapped_lines(content_lines)

    return body_lines, footnotes


def extract_pypdf_pages(pdf_path: Path) -> tuple[list[tuple[int, str]], dict[int, dict[int, str]]]:
    page_lines: list[tuple[int, str]] = []
    page_footnotes: dict[int, dict[int, str]] = {}

    for page_number, page_text in extract_pypdf_page_texts(pdf_path):
        # Keep the page number on each line so footnotes can later be mapped
        # back to the right recommendation block.
        body_lines, footnotes = split_pypdf_page(page_text)
        page_footnotes[page_number] = footnotes
        for line in body_lines:
            page_lines.append((page_number, line))

    return page_lines, page_footnotes


def parse_pypdf_ref_id(line: str) -> str | None:
    match = PYPDF_RECOMMENDATION_RE.match(line)
    if not match:
        return None

    digits = re.sub(r"\s+", "", match.group(1))
    suffix = match.group(2) or ""
    return f"R{digits}{suffix}"


def is_pypdf_section_boundary(line: str, known_ids: set[str]) -> bool:
    stripped = line.strip()
    if not stripped:
        return False
    if stripped in known_ids:
        return True

    match = SECTION_RE.match(stripped)
    return bool(match and match.group(1) in known_ids)


def parse_pypdf_recommendations(
    page_lines: list[tuple[int, str]],
    sections: list[Section],
    start: int,
    end: int,
    section_map: dict[str, str],
) -> list[PypdfRecommendationBlock]:
    known_ids = section_ids(sections)
    blocks: list[PypdfRecommendationBlock] = []
    index = start

    while index < end:
        ref_id = parse_pypdf_ref_id(page_lines[index][1])
        if ref_id is None:
            index += 1
            continue

        # Collect the block until the next recommendation or section.
        boundary = index + 1
        block_lines: list[str] = []
        block_pages: set[int] = {page_lines[index][0]}
        while boundary < end:
            page_number, line = page_lines[boundary]
            if parse_pypdf_ref_id(line) or is_pypdf_section_boundary(line, known_ids):
                break
            stripped = line.strip()
            if stripped and not is_footnote_or_marker(stripped):
                block_lines.append(stripped)
                block_pages.add(page_number)
            boundary += 1

        # The first lines of the block are treated as the title until the text
        # looks like regular body content.
        title_lines: list[str] = []
        while block_lines:
            title_lines.append(block_lines.pop(0))
            if not block_lines or not looks_like_title_continuation(block_lines[0]):
                break

        description_lines, annotation_blocks = split_description_and_annotation(
            block_lines
        )
        name = join_wrapped_lines(title_lines)
        description = format_content_lines(description_lines)
        annotation = format_annotation_blocks(annotation_blocks)

        if name and description and ref_id in section_map:
            blocks.append(
                PypdfRecommendationBlock(
                    recommendation=Recommendation(
                        ref_id=ref_id,
                        name=name,
                        description=description,
                        annotation=annotation,
                        section_id=section_map[ref_id],
                    ),
                    pages=tuple(sorted(block_pages)),
                )
            )

        index = boundary

    return blocks


def collect_ghostscript_post_lines(
    lines: list[str], start_index: int, end_index: int, content_indent: int
) -> list[str]:
    post_lines: list[str] = []
    inside_footnote = False

    for probe in range(start_index, end_index):
        raw_line = lines[probe]
        if is_page_noise(raw_line):
            inside_footnote = False
            continue

        stripped = raw_line.strip()
        if not stripped:
            continue

        # Ghostscript sometimes splits a recommendation with a footnote right
        # across a page break. Skip the footnote, then resume if indentation
        # returns to the recommendation block.
        if inside_footnote:
            if is_footnote_or_marker(raw_line):
                continue
            if indentation(raw_line) >= content_indent:
                inside_footnote = False
            else:
                continue

        if is_footnote_or_marker(raw_line):
            inside_footnote = True
            continue

        if indentation(raw_line) < content_indent:
            break

        post_lines.append(stripped)

    return post_lines


def parse_ghostscript_recommendations_precise(
    lines: list[str], sections: list[Section], start: int, end: int
) -> list[Recommendation]:
    known_ids = section_ids(sections)
    heading_occurrences = detect_heading_occurrences(lines, sections, start, end)
    recommendations: list[Recommendation] = []

    for index in range(start, end):
        match = RECOMMENDATION_RE.match(lines[index])
        if not match:
            continue

        ref_id = match.group(1)
        boundary = next_boundary_index(lines, index, end, known_ids)

        content_indent = detect_content_indent(lines, index + 1, boundary)
        if content_indent is None:
            content_indent = indentation(lines[index])

        title_prefix = collect_title_prefix(lines, index, content_indent, known_ids)
        # Replace the default collection logic with a version that is more
        # robust to footnotes and page breaks.
        post_lines = collect_ghostscript_post_lines(
            lines, index + 1, boundary, content_indent
        )

        if boundary < end and is_recommendation_line(lines[boundary]):
            next_content_indent = detect_content_indent(lines, boundary + 1, end)
            if next_content_indent is not None:
                next_prefix = collect_title_prefix(
                    lines, boundary, next_content_indent, known_ids
                )
                if next_prefix:
                    trimmed = post_lines[:]
                    remaining_prefix = [cleanup_text(value) for value in next_prefix]
                    for expected in reversed(remaining_prefix):
                        if trimmed and cleanup_text(trimmed[-1]) == expected:
                            trimmed.pop()
                    post_lines = trimmed

        title_suffix: list[str] = []
        while post_lines and looks_like_title_continuation(post_lines[0]):
            title_suffix.append(post_lines.pop(0))

        description_lines, annotation_blocks = split_description_and_annotation(post_lines)
        name = join_wrapped_lines([*title_prefix, *title_suffix])
        description = format_content_lines(description_lines)
        annotation = format_annotation_blocks(annotation_blocks)

        if not name or not description:
            raise ValueError(
                f"Incomplete ghostscript recommendation block parsed for {ref_id}"
            )

        active_section = find_active_section(index, heading_occurrences)
        if active_section is None:
            raise ValueError(f"No active section found for recommendation {ref_id}")

        recommendations.append(
            Recommendation(
                ref_id=ref_id,
                name=name,
                description=description,
                annotation=annotation,
                section_id=active_section,
            )
        )

    return recommendations


def merge_recommendations(
    ghostscript_recommendations: list[Recommendation],
    pypdf_blocks: list[PypdfRecommendationBlock],
) -> list[Recommendation]:
    pypdf_by_ref = {
        block.recommendation.ref_id: block.recommendation for block in pypdf_blocks
    }
    merged: list[Recommendation] = []

    for recommendation in ghostscript_recommendations:
        pypdf_version = pypdf_by_ref.get(recommendation.ref_id)
        if pypdf_version is None:
            merged.append(recommendation)
            continue

        # Keep Ghostscript as the structural source of truth, but replace the
        # text with the cleaner pypdf version when available.
        merged.append(
            Recommendation(
                ref_id=recommendation.ref_id,
                name=pypdf_version.name,
                description=trim_text_with_ghostscript_boundary(
                    pypdf_version.description,
                    recommendation.description,
                ),
                annotation=trim_text_with_ghostscript_boundary(
                    pypdf_version.annotation or recommendation.annotation,
                    recommendation.annotation,
                ),
                section_id=recommendation.section_id,
            )
        )

    return merged


def normalize_for_boundary_match(text: str) -> tuple[str, list[int]]:
    # Coarse normalization to compare Ghostscript text and pypdf text despite
    # spacing and hyphenation differences.
    normalized_chars: list[str] = []
    positions: list[int] = []
    for index, char in enumerate(text):
        if char.isalnum():
            normalized_chars.append(char.lower())
            positions.append(index)
    return "".join(normalized_chars), positions


def build_boundary_candidates(text: str) -> list[str]:
    candidates: list[str] = []

    # Try several plausible endings of the Ghostscript block so the same
    # ending can be found in the cleaner pypdf text.
    blocks = [line.strip() for line in text.splitlines() if line.strip()]
    if blocks:
        candidates.append(blocks[-1])
        if len(blocks) >= 2:
            candidates.append(" ".join(blocks[-2:]))

    linear_text = re.sub(r"\s+", " ", text).strip()
    if linear_text:
        sentences = [
            part.strip()
            for part in re.split(r"(?<=[.!?;:])\s+", linear_text)
            if part.strip()
        ]
        if sentences:
            candidates.append(sentences[-1])
            if len(sentences) >= 2:
                candidates.append(" ".join(sentences[-2:]))

        words = linear_text.split()
        if len(words) >= 12:
            candidates.append(" ".join(words[-12:]))
        if len(words) >= 20:
            candidates.append(" ".join(words[-20:]))

    deduped: list[str] = []
    seen: set[str] = set()
    for candidate in sorted(candidates, key=len, reverse=True):
        normalized, _ = normalize_for_boundary_match(candidate)
        if len(normalized) < 8 or normalized in seen:
            continue
        seen.add(normalized)
        deduped.append(candidate)
    return deduped


def trim_text_with_ghostscript_boundary(pypdf_text: str, ghostscript_text: str) -> str:
    if not pypdf_text or not ghostscript_text:
        return pypdf_text

    source_normalized, positions = normalize_for_boundary_match(pypdf_text)
    if not source_normalized:
        return pypdf_text

    for candidate in build_boundary_candidates(ghostscript_text):
        candidate_normalized, _ = normalize_for_boundary_match(candidate)
        if not candidate_normalized:
            continue
        match_index = source_normalized.rfind(candidate_normalized)
        if match_index == -1:
            continue
        # Trim pypdf at the last coherent ending found on the Ghostscript side.
        end_position = positions[match_index + len(candidate_normalized) - 1] + 1
        while end_position < len(pypdf_text) and pypdf_text[end_position] in "])}>»'\".,;:!?":
            end_position += 1
        footnote_suffix = re.match(
            r'(?:(?:\s*[)\]}>»\'"]*)?\s+\d{1,2}(?:\s*[.,;:!?])?)+',
            pypdf_text[end_position:],
        )
        if footnote_suffix:
            end_position += len(footnote_suffix.group(0))
        return pypdf_text[:end_position].rstrip()

    return pypdf_text


def build_recommendation_page_footnotes(
    pypdf_blocks: list[PypdfRecommendationBlock],
    page_footnotes: dict[int, dict[int, str]],
) -> dict[str, dict[int, str]]:
    recommendation_footnotes: dict[str, dict[int, str]] = {}
    for block in pypdf_blocks:
        merged_notes: dict[int, str] = {}
        for page_number in block.pages:
            # A recommendation may span multiple pages.
            for number, content in page_footnotes.get(page_number, {}).items():
                merged_notes.setdefault(number, content)
        recommendation_footnotes[block.recommendation.ref_id] = merged_notes
    return recommendation_footnotes


def footnote_abbr(number: int, content: str) -> str:
    title = html.escape(content, quote=True)
    return f'<abbr title="{title}"><sup>{number}</sup></abbr>'


def inject_page_footnotes(text: str, footnotes: dict[int, str]) -> str:
    if not text or not footnotes:
        return text

    def replace(match: re.Match[str]) -> str:
        number = int(match.group(1))
        content = footnotes.get(number)
        if content is None:
            return match.group(0)

        suffix = match.string[match.end() : match.end() + 1]
        rendered = footnote_abbr(number, content)
        if suffix and suffix[0].isalpha():
            # If the number was glued to the next word, restore the space.
            return rendered + " "
        return rendered

    return INLINE_FOOTNOTE_RE.sub(replace, text)


def format_recommendation_text(
    text: str,
    reference_links: dict[str, str],
    footnotes: dict[int, str],
) -> str:
    text = normalize_pypdf_references(text)
    text = normalize_inline_spacing(text)
    text = re.sub(r"(?<=\b[A-Za-zÀ-ÿ])\s+'", "'", text)
    text = re.sub(r"'\s+(?=[A-Za-zÀ-ÿ])", "'", text)
    text = linkify_reference_markers(text, reference_links)
    return inject_page_footnotes(text, footnotes)


def format_recommendation_ref_id(ref_id: str) -> str:
    return f"[{ref_id}]"


def build_rows_hybrid(
    sections: list[Section],
    recommendations: list[Recommendation],
    reference_links: dict[str, str],
    recommendation_footnotes: dict[str, dict[int, str]],
) -> list[list[str]]:
    kept_section_ids: set[str] = set()
    for recommendation in recommendations:
        parts = recommendation.section_id.split(".")
        for depth in range(1, len(parts) + 1):
            kept_section_ids.add(".".join(parts[:depth]))

    # Drop empty branches so the exported hierarchy only keeps sections that
    # actually lead to recommendations.
    sections = [section for section in sections if section.node_id in kept_section_ids]
    section_by_id = {section.node_id: section for section in sections}
    recommendations_by_section: dict[str, list[Recommendation]] = {}
    for recommendation in recommendations:
        recommendations_by_section.setdefault(recommendation.section_id, []).append(
            recommendation
        )

    rows: list[list[str]] = []
    for section in sections:
        rows.append(
            [
                "",
                str(section.depth),
                "",
                section.node_id,
                prepare_section_text_pypdf(section.name, reference_links),
                "",
                "",
                "",
            ]
        )
        for recommendation in recommendations_by_section.get(section.node_id, []):
            footnotes = recommendation_footnotes.get(recommendation.ref_id, {})
            rows.append(
                [
                    "x",
                    str(section_by_id[recommendation.section_id].depth + 1),
                    recommendation.node_id,
                    format_recommendation_ref_id(recommendation.ref_id),
                    format_recommendation_text(
                        recommendation.name, reference_links, footnotes
                    ),
                    format_recommendation_text(
                        recommendation.description, reference_links, footnotes
                    ),
                    format_recommendation_text(
                        recommendation.annotation, reference_links, footnotes
                    ),
                    recommendation.implementation_groups,
                ]
            )
    return rows


# ============================================================================
# Workbook sheet assembly
# ============================================================================


def write_kv_sheet(wb: Workbook, sheet_name: str, rows: Iterable[tuple[str, str]]) -> None:
    ws = wb.create_sheet(title=sheet_name)
    for key, value in rows:
        ws.append([key, value])


def write_reco_content_sheet(wb: Workbook, rows: list[list[str]]) -> None:
    ws = wb.create_sheet(title=SHEET_NAME)
    ws.append(COLUMNS)
    for row in rows:
        ws.append(row)


def write_imp_grp_content_sheet(
    wb: Workbook, rows: Iterable[tuple[str, str, str]]
) -> None:
    ws = wb.create_sheet(title="imp_grp_content")
    ws.append(IMP_GRP_COLUMNS)
    for row in rows:
        ws.append(list(row))


def save_workbook(output_path: Path, rows: list[list[str]]) -> None:
    wb = Workbook()

    # Reuse the default first sheet for library_meta so sheet order stays:
    # library_meta, reco_meta, reco_content, imp_grp_meta, imp_grp_content.
    ws_library = wb.active
    ws_library.title = "library_meta"
    for key, value in LIBRARY_META_ROWS:
        ws_library.append([key, value])

    write_kv_sheet(wb, "reco_meta", RECO_META_ROWS)
    write_reco_content_sheet(wb, rows)
    write_kv_sheet(wb, "imp_grp_meta", IMP_GRP_META_ROWS)
    write_imp_grp_content_sheet(wb, IMP_GRP_CONTENT_ROWS)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    args = parse_args()

    # Ghostscript remains the most reliable source for the global structure.
    gs_text = extract_pdf_text(args.pdf)
    gs_lines = load_lines(gs_text)
    reference_links = extract_bibliography_reference_links(gs_lines)
    sections = parse_sections_from_toc(extract_toc_lines(gs_lines))
    sections = merge_clean_section_titles(
        sections, extract_clean_top_level_section_titles_from_pypdf(args.pdf, sections)
    )
    gs_start, gs_end = body_slice(gs_lines)
    gs_recommendations = parse_ghostscript_recommendations_precise(
        gs_lines, sections, gs_start, gs_end
    )
    section_map = {
        recommendation.ref_id: recommendation.section_id
        for recommendation in gs_recommendations
    }

    # pypdf is mainly used to recover cleaner recommendation text.
    pypdf_page_lines, page_footnotes = extract_pypdf_pages(args.pdf)
    pypdf_lines = [line for _, line in pypdf_page_lines]
    pypdf_start, pypdf_end = body_slice(pypdf_lines)
    pypdf_blocks = parse_pypdf_recommendations(
        pypdf_page_lines,
        sections,
        pypdf_start,
        pypdf_end,
        section_map,
    )

    # Keep the Ghostscript skeleton and replace the text when pypdf provides
    # a usable block.
    recommendations = merge_recommendations(gs_recommendations, pypdf_blocks)
    recommendation_footnotes = build_recommendation_page_footnotes(
        pypdf_blocks, page_footnotes
    )
    rows = build_rows_hybrid(
        sections, recommendations, reference_links, recommendation_footnotes
    )
    save_workbook(args.output, rows)

    kept_sections = sum(1 for row in rows if not row[0])
    annotations = sum(1 for recommendation in recommendations if recommendation.annotation)
    print(f"Sections kept: {kept_sections}")
    print(f"Recommendations: {len(recommendations)}")
    print(
        f"pypdf recommendations extracted: {len(pypdf_blocks)} / {len(gs_recommendations)}"
    )
    print(f"Recommendations with annotation: {annotations}")
    print(f"Bibliography links found: {len(reference_links)}")
    print(f"Output: {args.output}")


if __name__ == "__main__":
    main()
