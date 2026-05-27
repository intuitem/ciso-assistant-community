#!/usr/bin/env python3
"""Build the II 901 framework workbook from the source PDF."""


import argparse
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Sequence

import fitz
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# =============================================================================
# Framework configuration and default paths
# =============================================================================

SCRIPT_DIR = Path(__file__).resolve().parent
FRAMEWORK_DIR = SCRIPT_DIR.parent
DEFAULT_PDF = FRAMEWORK_DIR / "ii_901_cir_39217.pdf"
DEFAULT_OUTPUT = FRAMEWORK_DIR / "ii-901.xlsx"

FRAMEWORK_REF_ID = "II-901"
URN_ROOT = "ii-901_integral"
FRAMEWORK_NAME = (
    "II n° 901/SGDSN/ANSSI - Instruction interministérielle n° 901 relative à la protection "
    "des systèmes d'information sensibles"
)
FRAMEWORK_DESCRIPTION = (
    """La présente instruction définit les objectifs et les règles relatifs à la protection des systèmes d'information sensibles, notamment ceux traitant des informations portant la mention *Diffusion Restreinte*.

La présente instruction s'adresse à l'ensemble des personnes physiques ou morales intervenant dans ces systèmes.

Le respect des règles contribue à garantir la continuité des activités de l'entité qui met en œuvre le système d'information, à protéger l'image de cette entité, à prévenir la compromission d'informations sensibles et à assurer la sécurité des personnes et des biens.

Ces règles peuvent être précisées au cas par cas en s'appuyant sur les normes techniques existantes et sur les guides techniques et les recommandations de l'Agence nationale de la sécurité des systèmes d'information (ANSSI).

Sources :
https://cyber.gouv.fr/reglementation/cybersecurite-systemes-dinformation/protection-information-sensible-diffusion-restreinte/instruction-interministerielle-n901/
https://www.legifrance.gouv.fr/circulaire/id/39217"""
)

CONTENT_SHEET = "II-901_content"
CONTENT_COLUMNS = [
    "assessable",
    "depth",
    "ref_id",
    "name",
    "description",
    "annotation",
]

LIBRARY_META_ROWS: tuple[tuple[str, str], ...] = (
    ("type", "library"),
    ("urn", f"urn:intuitem:risk:library:{URN_ROOT}"),
    ("version", "1"),
    ("locale", "fr"),
    ("ref_id", FRAMEWORK_REF_ID),
    ("name", FRAMEWORK_NAME),
    ("description", FRAMEWORK_DESCRIPTION),
    ("copyright", "SGDSN / ANSSI"),
    ("provider", "ANSSI"),
    ("packager", "intuitem"),
)

FRAMEWORK_META_ROWS: tuple[tuple[str, str], ...] = (
    ("type", "framework"),
    ("base_urn", f"urn:intuitem:risk:req_node:{URN_ROOT}"),
    ("urn", f"urn:intuitem:risk:framework:{URN_ROOT}"),
    ("ref_id", FRAMEWORK_REF_ID),
    ("name", FRAMEWORK_NAME),
    ("description", FRAMEWORK_DESCRIPTION),
)

SECTION_NAMES_BY_ORDER: tuple[str, ...] = (
    "Politique, organisation, gouvernance",
    "Ressources humaines",
    "Gestion des biens",
    "Intégration de la SSI dans le cycle de vie des systèmes d'information",
    "Sécurité physique",
    "Sécurité des réseaux",
    "Architecture des systèmes d'information",
    "Exploitation des systèmes d'information",
    "Sécurité du poste de travail",
    "Traitement des incidents",
    "Continuité d'activité",
    "Conformité, audit, inspection, contrôle",
)


# =============================================================================
# PDF markers and recognition regexes
# =============================================================================

TITLE_RE = re.compile(
    r"^Titre\s+(?P<number>IV|III|II|I|[1-4])\s*[-–]\s*(?P<title>.+)$",
    re.IGNORECASE,
)
ARTICLE_RE = re.compile(
    r"^Article\s+(?P<number>1er|\d+)\s*:\s*(?P<title>.+)$",
    re.IGNORECASE,
)
NUMBERED_SECTION_RE = re.compile(r"^(?P<ref_id>\d+\.\d+)\s+(?P<body>.+)$")
ANNEXE_1_RE = re.compile(r"annexe\s+1", re.IGNORECASE)
ANNEXE_2_RE = re.compile(r"annexe\s+2", re.IGNORECASE)
OBJECTIVE_RE = re.compile(r"^obje", re.IGNORECASE)
NETWORK_CLASS_RE = re.compile(r"^Un réseau de classe\s+(?P<class_id>\d+)\b", re.IGNORECASE)
RECOMMENDATION_RE = re.compile(
    r"^\s*(?P<raw_id>[A-Z0-9][A-Z0-9\s-]{1,55}[A-Z0-9])\s*:\s*(?P<body>.+)$",
    re.IGNORECASE,
)
PAGE_NUMBER_RE = re.compile(r"^\d{1,2}$")
LIST_MARKER_RE = re.compile(r"^(?:[•▪■]\s*|-\s*)")
FOOTNOTE_SPLIT_RE = re.compile(
    r"(?:^|(?<=[.;])\s+)(?P<number>\d{1,2})\s+"
    r"(?=(?:Voir|Les|Le|La|Sauf|Sont|Ces|Guide|L'\s*|Un|Une|Il|Elle|Dans|Pour)\b)"
)


# =============================================================================
# Internal models
# =============================================================================

@dataclass(slots=True)
class TextBlock:
    """Text block extracted from the PDF, including its page position."""

    page: int
    x0: float
    y0: float
    x1: float
    y1: float
    text: str
    fonts: tuple[str, ...]

    @property
    def is_italic(self) -> bool:
        return any("Italic" in font for font in self.fonts)


@dataclass(slots=True)
class ContentRow:
    """Final row written to the II-901_content worksheet."""

    assessable: str | None
    depth: int
    ref_id: str | None
    name: str | None
    description: str | None = None
    last_description_block: TextBlock | None = None
    annotation: str | None = None

    def as_tuple(self) -> tuple[str | int | None, ...]:
        return (
            self.assessable,
            self.depth,
            self.ref_id,
            self.name,
            self.description,
            self.annotation,
        )

@dataclass(slots=True)
class Footnote:
    """Footnote extracted from the PDF footer area."""

    page: int
    number: str
    text: str


# =============================================================================
# CLI
# =============================================================================

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build the II 901 framework XLSX from the source PDF."
    )
    parser.add_argument("input", type=Path, nargs="?", default=DEFAULT_PDF)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument(
        "--debug",
        action="store_true",
        help="Print detected sections, objectives, and recommendation count.",
    )
    return parser.parse_args()


# =============================================================================
# OCR cleanup and text normalization
# =============================================================================

def clean_text(text: str) -> str:
    """Clean raw PDF text and fix known OCR errors."""

    text = text.replace("\u00a0", " ")
    text = text.replace("ﬁ", "fi").replace("ﬂ", "fl")
    text = re.sub(r"\s+", " ", text)
    text = text.replace("1 '", "l'").replace("1'", "l'")
    text = text.replace("I' ", "l'").replace("I'", "l'")
    text = re.sub(r"\b1\s+([aeiouhAÉÈÊÎÏÔÛÙÜYy])", r"l'\1", text)
    text = text.replace("J'", "l'")
    text = text.replace("J ournalisation", "Journalisation")
    text = text.replace("et! 'usage", "et l'usage")
    text = re.sub(r"\b([cdjlmnst])\s+'\s*", r"\1'", text)
    text = re.sub(r"\b(qu|jusqu|lorsqu|puisqu)\s+'\s*", r"\1'", text)
    text = re.sub(r"\b([cdjlmnst])'\s+", r"\1'", text)
    text = re.sub(r"\b(qu|jusqu|lorsqu|puisqu)'\s+", r"\1'", text)
    text = re.sub(r"\bJe\b", "le", text)
    text = text.replace("ANS SI", "ANSSI").replace("A NS SI", "ANSSI")
    text = text.replace("ANSS I", "ANSSI")
    text = text.replace("l' Etat", "l'État").replace("l'Etat", "l'État")
    text = text.replace("PSS I", "PSSI")
    text = text.replace("PSSŒ", "PSSIE")
    text = text.replace("PSSJE", "PSSIE")
    text = text.replace("Difjùsion", "Diffusion")
    text = text.replace("re sponsabilités", "responsabilités")
    text = text.replace("arrivees", "arrivées")
    text = text.replace("tracabilité", "traçabilité")
    text = text.replace("VISiteurs", "visiteurs")
    text = text.replace("sécurisa/ion", "sécurisation")
    text = text.replace("a/Laques", "attaques")
    text = text.replace("altaques", "attaques")
    text = text.replace("incidenls", "incidents")
    text = text.replace("en étal", "en état")
    text = text.replace("elles tester", "et les tester")
    text = text.replace(" el configurer", " et configurer")
    text = text.replace("Config ur er", "Configurer")
    text = text.replace("locau x", "locaux")
    text = text.replace("rése au", "réseau")
    text = text.replace("ré seau", "réseau")
    text = text.replace("techni~ue", "technique")
    text = text.replace("(PPSTi", "(PPST)3")
    text = text.replace("mmtmtser", "minimiser")
    text = text.replace("Jabellisé", "labellisé")
    text = text.replace("châmes", "chaînes")
    text = text.replace("sa uvegardes", "sauvegardes")
    text = text.replace("autocornmutateurs", "autocommutateurs")
    text = text.replace("extemalisation", "externalisation")
    text = text.replace("Extemalisation", "Externalisation")
    text = text.replace("àjour", "à jour")
    text = text.replace("conna.Jtre", "connaître")
    text = text.replace("défmi", "défini")
    text = text.replace("v1gueur", "vigueur")
    text = text.replace("afm de", "afin de")
    text = text.replace("20 Il", "2011")
    text = text.replace("R. 23 1 1-l'et", "R. 2311-1 et")
    text = text.replace("R. 23 1 1-1 et", "R. 2311-1 et")
    text = text.replace("R. 2311-l'et", "R. 2311-1 et")
    text = text.replace("no 2011 -1425", "n° 2011-1425")
    text = text.replace("http//", "http://")
    text = text.replace("dans La mesure", "dans la mesure")
    text = text.replace("arc-en- ciel", "arc-en-ciel")
    text = text.replace("classe l ou 2", "classe 1 ou 2")
    text = text.replace("classe l'ou 2", "classe 1 ou 2")
    text = text.replace("Restreinte l'ou", "Restreinte 1 ou")
    text = text.replace("classe l'est", "classe 1 est")
    text = text.replace("classe L'est", "classe 1 est")
    text = text.replace("classe  O", "classe 0")
    text = text.replace("classe O", "classe 0")
    text = text.replace("d '", "d'")
    text = text.replace("l '", "l'")
    text = text.replace("L' interconnexion", "L'interconnexion")
    text = text.replace("l'o bjet", "l'objet")
    text = text.replace("d'i nterconnexion", "d'interconnexion")
    text = text.replace("ou l'au travers", "ou 1 au travers")
    text = text.replace("so n site", "son site")
    text = text.replace("sécuri sée", "sécurisée")
    text = text.replace("sécurisée»", "sécurisée »")
    text = text.replace("pr ésentant", "présentant")
    text = text.replace("interconnexio n", "interconnexion")
    text = text.replace("inf ér ieure", "inférieure")
    text = text.replace("éq uipements", "équipements")
    text = text.replace("autorisée compris via", "autorisée y compris via")
    text = text.replace("technique«", "technique «")
    text = re.sub(r"\b(\d+)\.l'(?=[A-ZÉÈÊÀÂ])", r"\1.1 ", text)
    text = text.replace(
        "d'informations sensibles doivent être effectuées selon une préalablement, "
        "garantissant un contrôle par l'utilisateur, du l'impressionjusqu'à la "
        "récupération du support imprimé. Les impressions procédure définie "
        "déclenchement de",
        "Les impressions d'informations sensibles doivent être effectuées selon "
        "une procédure définie préalablement, garantissant un contrôle par "
        "l'utilisateur, du déclenchement de l'impression jusqu'à la récupération "
        "du support imprimé.",
    )
    return text.strip()


def clean_ref_id(raw_id: str) -> str:
    """Normalize recommendation identifiers, for example TL- -> TI-."""

    ref_id = re.sub(r"\s+", "", raw_id.upper())
    ref_id = ref_id.replace("–", "-").replace("—", "-")
    ref_id = ref_id.replace("DEY-", "DEV-")
    ref_id = ref_id.replace("POT-", "PDT-")
    ref_id = ref_id.replace("TL-", "TI-")
    ref_id = ref_id.replace("T1-", "TI-")
    ref_id = ref_id.replace("-YERIF", "-VERIF")
    ref_id = ref_id.replace("SECX-DIST", "SEC-DIST")
    ref_id = ref_id.replace("SERY", "SERV")
    ref_id = ref_id.replace("OESACTIV", "DESACTIV")
    ref_id = ref_id.replace("-FIL T", "-FILT")
    ref_id = ref_id.replace("FILT-APPL", "FILT-APPL")
    return ref_id


def clean_name(text: str) -> str:
    """Clean a title and uppercase its first alphabetic character."""

    text = clean_text(text)
    text = text.strip(" .")
    for index, char in enumerate(text):
        if char.isalpha():
            return f"{text[:index]}{char.upper()}{text[index + 1:]}"
    return text


def clean_multiline_text(text: str) -> str:
    """Clean multiline text while preserving list indentation."""

    cleaned_lines: list[str] = []
    for line in text.splitlines():
        if not line.strip():
            continue
        indent = len(line) - len(line.lstrip(" "))
        cleaned = clean_text(line)
        if line.lstrip().startswith("* "):
            cleaned = f"{' ' * indent}{cleaned}"
        cleaned_lines.append(cleaned)
    return "\n".join(cleaned_lines)


# =============================================================================
# Shared helpers for lists, paragraphs, and PDF noise
# =============================================================================

def strip_list_marker(text: str) -> str:
    return LIST_MARKER_RE.sub("", text).strip()


def list_level(block: TextBlock) -> int:
    """List level for Annex 1, based on horizontal indentation."""

    if block.x0 < 125:
        return 0
    return int((block.x0 - 125) // 45) + 1


def starts_new_paragraph(previous_block: TextBlock | None, block: TextBlock) -> bool:
    """Detect a new paragraph from page changes or vertical spacing."""

    if previous_block is None:
        return False
    if previous_block.page != block.page:
        return True
    return block.y0 - previous_block.y1 > 4


def append_description_block(
    description: str | None, block: TextBlock, previous_block: TextBlock | None = None
) -> str:
    """Append a text block to an Annex 1 description."""

    text = clean_text(block.text)
    if not text:
        return description or ""

    existing = description or ""
    lines = existing.splitlines()
    has_list = any(line.lstrip().startswith("* ") for line in lines)
    last_line = lines[-1] if lines else ""
    marker = LIST_MARKER_RE.match(text)
    explicit_list_item = marker is not None
    inferred_list_item = (
        block.x0 >= 155 and (has_list or last_line.rstrip().endswith(":"))
    )
    should_be_list_item = explicit_list_item or inferred_list_item

    if should_be_list_item:
        item_text = clean_text(strip_list_marker(text))
        level = max(1, list_level(block))
        prefix = f"{'  ' * (level - 1)}* "
        return clean_multiline_text(f"{existing}\n{prefix}{item_text}")

    if has_list and block.x0 >= 120 and lines:
        lines[-1] = clean_text(f"{last_line} {text}")
        return clean_multiline_text("\n".join(lines))

    if starts_new_paragraph(previous_block, block):
        return clean_multiline_text(f"{existing}\n{text}")

    return clean_multiline_text(" ".join(part for part in (existing, text) if part))


def is_noise(block: TextBlock) -> bool:
    """Ignore page numbers, footers, and notes not useful for the framework."""

    if PAGE_NUMBER_RE.fullmatch(block.text):
        return True
    if block.y0 > 775:
        return True
    if block.y0 > 740 and re.match(r"^\d+\s+", block.text):
        return True
    if block.text.startswith("9 Ces règles sont adaptées"):
        return True
    return False


def is_annexe_heading(block: TextBlock, annexe_re: re.Pattern[str]) -> bool:
    """Detect Annex 1 / Annex 2 headings without matching the table of contents."""

    text = block.text.strip()
    return bool(annexe_re.search(text)) and text.lower().startswith("annexe") and "..." not in text


# =============================================================================
# Footnote extraction and attachment
# =============================================================================

def format_footnote_annotation(footnote: Footnote) -> str:
    """Format a footnote for the annotation column."""

    return f"<sup>{footnote.number}</sup> {footnote.text}"


def add_annotation(row: ContentRow, footnote: Footnote) -> None:
    """Append a footnote to a row annotation, one note per line."""

    formatted = format_footnote_annotation(footnote)
    if row.annotation:
        if formatted not in row.annotation.splitlines():
            row.annotation = f"{row.annotation}\n{formatted}"
        return
    row.annotation = formatted


def mark_footnote_reference(text: str | None, number: str) -> tuple[str | None, bool]:
    """Replace a footnote call with a <sup>number</sup> marker."""

    if not text or f"<sup>{number}</sup>" in text:
        return text, False

    changed = False
    blocked_previous_words = {"article", "annexe", "titre", "classe", "réseau"}

    contiguous_pattern = re.compile(rf"(?<=[A-Za-zÀ-ÖØ-öø-ÿ)\]»]){number}(?!\d)")
    text, count = contiguous_pattern.subn(f"<sup>{number}</sup>", text)
    changed = changed or count > 0

    def replace_spaced_reference(match: re.Match[str]) -> str:
        nonlocal changed

        before = text[: match.start()].rstrip().lower()
        previous_word = re.search(r"([a-zà-öø-ÿ]+)$", before)
        if previous_word and previous_word.group(1) in blocked_previous_words:
            return match.group(0)

        changed = True
        suffix = " " if match.group(0).endswith(" ") else ""
        return f"<sup>{number}</sup>{suffix}"

    spaced_before_punctuation = re.compile(rf"(?<=[A-Za-zÀ-ÖØ-öø-ÿ)\]])\s+{number}(?=[.,;:)])")
    text = spaced_before_punctuation.sub(replace_spaced_reference, text)

    spaced_before_bullet = re.compile(rf"(?<=[A-Za-zÀ-ÖØ-öø-ÿ)\]])\s+{number}\s*•\s*")
    text = spaced_before_bullet.sub(replace_spaced_reference, text)

    spaced_before_symbol = re.compile(rf"(?<=[A-Za-zÀ-ÖØ-öø-ÿ)\]])\s+{number}\s*(?=[:•])")
    text = spaced_before_symbol.sub(replace_spaced_reference, text)

    spaced_before_word = re.compile(rf"(?<=[A-Za-zÀ-ÖØ-öø-ÿ)\]])\s+{number}\s+(?=[a-zà-öø-ÿ])")
    text = spaced_before_word.sub(replace_spaced_reference, text)

    spaced_at_end = re.compile(rf"(?<=[A-Za-zÀ-ÖØ-öø-ÿ)\]])\s+{number}$")
    text = spaced_at_end.sub(replace_spaced_reference, text)

    return text, changed


def strip_footnote_reference_from_title(
    text: str | None, number: str
) -> tuple[str | None, bool]:
    """Remove a footnote call from a title while keeping the annotation link."""

    if not text:
        return text, False

    cleaned = text
    removed = False
    for pattern in (
        rf"<sup>{number}</sup>",
        rf"(?<=[A-Za-zÀ-ÖØ-öø-ÿ)\]]){number}(?!\d)",
        rf"(?<=[A-Za-zÀ-ÖØ-öø-ÿ)\]])\s+{number}(?=[.,;)])",
        rf"(?<=[A-Za-zÀ-ÖØ-öø-ÿ)\]])\s+{number}$",
    ):
        cleaned, count = re.subn(pattern, "", cleaned)
        removed = removed or count > 0

    if not removed:
        return text, False

    return clean_name(cleaned), True


def attach_missing_page4_note_4(row: ContentRow, footnote: Footnote) -> bool:
    """Recover footnote 4, whose call is not extracted as a visible digit."""

    if footnote.number != "4" or not row.description:
        return False
    if "potentiel scientifique" not in row.description:
        return False
    target = "systèmes d'information sensibles ;"
    if target not in row.description:
        return False
    row.description = row.description.replace(
        target, "systèmes d'information sensibles<sup>4</sup> ;", 1
    )
    return True


def attach_footnotes_to_rows(
    rows: Sequence[ContentRow], footnotes_by_page: dict[int, dict[str, Footnote]]
) -> None:
    """Attach footnotes to rows that contain the corresponding note call."""

    for row in rows:
        if row.last_description_block is None:
            continue

        page_notes = {
            **footnotes_by_page.get(row.last_description_block.page, {}),
            **footnotes_by_page.get(row.last_description_block.page + 1, {}),
        }
        for number, footnote in page_notes.items():
            row.name, name_changed = strip_footnote_reference_from_title(
                row.name, number
            )
            row.description, description_changed = mark_footnote_reference(
                row.description, number
            )
            recovered_note_4 = attach_missing_page4_note_4(row, footnote)
            if name_changed or description_changed or recovered_note_4:
                add_annotation(row, footnote)


# =============================================================================
# Recognition of Titles I-IV, articles, and x.y sections
# =============================================================================

def split_title(text: str) -> tuple[str, str] | None:
    """Return ('Titre II', 'Label') for titles before Annex 1."""

    if "..." in text:
        return None
    match = TITLE_RE.match(text)
    if not match:
        return None

    roman_by_number = {"1": "I", "2": "II", "3": "III", "4": "IV"}
    number = match.group("number").upper()
    number = roman_by_number.get(number, number)
    return f"Titre {number}", clean_name(match.group("title"))


def split_article(text: str) -> tuple[str, str] | None:
    """Return ('Article #', 'Label') for articles in Titles I-IV."""

    if "..." in text:
        return None
    match = ARTICLE_RE.match(text)
    if not match:
        return None
    return f"Article {match.group('number')}", clean_name(match.group("title"))


def split_numbered_section(text: str) -> tuple[str, str] | None:
    """Detect numbered subsections such as 2.1 or 4.2."""

    if "..." in text:
        return None
    match = NUMBERED_SECTION_RE.match(text)
    if not match:
        return None
    return match.group("ref_id"), clean_text(match.group("body"))


def is_assessable_title(title_label: str | None) -> bool:
    """Only Titles II and III become assessable requirements."""

    return title_label in {"Titre II", "Titre III"}


def front_list_level(block: TextBlock) -> int:
    """List level for Titles I-IV, based on horizontal indentation."""

    if block.x0 < 100:
        return 0
    return int((block.x0 - 100) // 45) + 1


def is_same_visual_paragraph(previous_block: TextBlock | None, block: TextBlock) -> bool:
    """Detect continuation of the same paragraph split into several PDF blocks."""

    if previous_block is None or previous_block.page != block.page:
        return False
    return block.y0 - previous_block.y1 <= 1


def append_front_description_block(
    description: str | None, block: TextBlock, previous_block: TextBlock | None = None
) -> str:
    """Append text to non-assessable descriptions from Titles I and IV."""

    text = clean_text(block.text)
    if not text:
        return description or ""

    existing = description or ""
    lines = existing.splitlines()
    has_list = any(line.lstrip().startswith("* ") for line in lines)
    last_line = lines[-1] if lines else ""
    if has_list and block.x0 >= 100 and lines and is_same_visual_paragraph(previous_block, block):
        lines[-1] = clean_text(f"{last_line} {text}")
        return clean_multiline_text("\n".join(lines))

    explicit_list_item = LIST_MARKER_RE.match(text) is not None
    inferred_list_item = block.x0 >= 100 and (
        has_list or last_line.rstrip().endswith(":")
    )

    if explicit_list_item or inferred_list_item:
        item_text = clean_text(strip_list_marker(text))
        level = max(1, front_list_level(block))
        prefix = f"{'  ' * (level - 1)}* "
        return clean_multiline_text(f"{existing}\n{prefix}{item_text}")

    if has_list and block.x0 >= 100 and lines:
        lines[-1] = clean_text(f"{last_line} {text}")
        return clean_multiline_text("\n".join(lines))

    if starts_new_paragraph(previous_block, block):
        return clean_multiline_text(f"{existing}\n{text}")

    return clean_multiline_text(" ".join(part for part in (existing, text) if part))


def is_front_footnote_start(block: TextBlock) -> bool:
    """Detect the start of footnotes in the title pages."""

    return block.y0 > 500 and re.match(r"^\d+\s+", block.text) is not None


# =============================================================================
# Recognition of Annex 1-specific elements
# =============================================================================

def is_objective_marker(block: TextBlock) -> bool:
    """Detect 'Objectif # :' blocks despite minor OCR errors."""

    compact = re.sub(r"[^a-z]", "", block.text.lower())
    return block.y1 - block.y0 < 24 and (
        compact.startswith("objecti")
        or compact.startswith("objectif")
        or compact.startswith("objecli")
        or compact.startswith("objectif")
        or OBJECTIVE_RE.match(block.text) is not None
    )


def split_recommendation(text: str) -> tuple[str, str, str] | None:
    """Split an Annex 1 requirement into ID, name, and description."""

    match = RECOMMENDATION_RE.match(text)
    if not match:
        return None

    raw_id = match.group("raw_id")
    if "-" not in raw_id:
        return None

    ref_id = clean_ref_id(raw_id)
    body = clean_text(match.group("body"))
    title, separator, description = body.partition(". ")
    if not separator:
        return ref_id, clean_name(body), None
    return ref_id, clean_name(title), clean_text(description)


def is_heading(block: TextBlock) -> bool:
    """Detect blue headings under an objective in Annex 1."""

    if split_recommendation(block.text):
        return False
    if is_objective_marker(block):
        return False
    if block.text.startswith(("•", "-", "n ")):
        return False
    if block.x0 > 100:
        return False
    if block.text.endswith("."):
        return False
    return len(block.text) <= 180


# =============================================================================
# Text block extraction from the PDF
# =============================================================================

def get_block_text(block: dict) -> tuple[str, tuple[str, ...]]:
    """Convert a raw PyMuPDF block into cleaned text and used fonts."""

    parts: list[str] = []
    fonts: list[str] = []
    for line in block.get("lines", []):
        line_parts: list[str] = []
        for span in line.get("spans", []):
            span_text = span.get("text", "")
            if span_text:
                line_parts.append(span_text)
                fonts.append(span.get("font", ""))
        if line_parts:
            parts.append("".join(line_parts))
    return clean_text(" ".join(parts)), tuple(fonts)

def get_raw_block_text_and_sizes(block: dict) -> tuple[str, tuple[float, ...]]:
    """Return lightly normalized block text and span sizes for footnote detection."""

    parts: list[str] = []
    sizes: list[float] = []
    for line in block.get("lines", []):
        line_parts: list[str] = []
        for span in line.get("spans", []):
            span_text = span.get("text", "")
            if span_text:
                line_parts.append(span_text)
                if span_text.strip():
                    sizes.append(float(span.get("size", 0)))
        if line_parts:
            parts.append("".join(line_parts))
    return clean_text(" ".join(parts)), tuple(sizes)


def split_footnote_text(page: int, text: str) -> list[Footnote]:
    """Split a footer text containing one or several numbered footnotes."""

    matches = list(FOOTNOTE_SPLIT_RE.finditer(text))
    footnotes: list[Footnote] = []
    for index, match in enumerate(matches):
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        number = match.group("number")
        note_text = clean_text(text[start:end])
        if note_text:
            footnotes.append(Footnote(page, number, note_text))
    return footnotes


def extract_footnotes(pdf_path: Path) -> dict[int, dict[str, Footnote]]:
    """Extract footer notes per page without adding them as content blocks."""

    if not pdf_path.exists():
        raise FileNotFoundError(f"PDF not found: {pdf_path}")

    doc = fitz.open(pdf_path)
    footnotes_by_page: dict[int, dict[str, Footnote]] = {}

    for page_index, page in enumerate(doc, start=1):
        footer_fragments: list[str] = []
        in_footer = False

        raw_blocks = [
            block
            for block in page.get_text("dict")["blocks"]
            if block.get("type") == 0
        ]
        raw_blocks.sort(key=lambda item: (item["bbox"][1], item["bbox"][0]))

        for raw_block in raw_blocks:
            text, sizes = get_raw_block_text_and_sizes(raw_block)
            if not text or PAGE_NUMBER_RE.fullmatch(text):
                continue

            x0, y0, _x1, _y1 = raw_block["bbox"]
            max_size = max(sizes) if sizes else 0
            starts_note = bool(re.match(r"^\d{1,2}\s+", text)) and (
                y0 > 500 or max_size <= 10.2
            )

            if starts_note:
                in_footer = True
            if not in_footer:
                continue
            if x0 > 250 and PAGE_NUMBER_RE.fullmatch(text):
                continue

            footer_fragments.append(text)

        if not footer_fragments:
            continue

        notes = split_footnote_text(page_index, " ".join(footer_fragments))
        if notes:
            footnotes_by_page[page_index] = {note.number: note for note in notes}

    return footnotes_by_page


def extract_page_text_blocks(page: fitz.Page, page_index: int) -> list[TextBlock]:
    """Extract and sort text blocks from a PDF page."""

    blocks: list[TextBlock] = []
    for raw_block in page.get_text("dict")["blocks"]:
        if raw_block.get("type") != 0:
            continue
        text, fonts = get_block_text(raw_block)
        if not text:
            continue
        x0, y0, x1, y1 = raw_block["bbox"]
        blocks.append(TextBlock(page_index, x0, y0, x1, y1, text, fonts))
    return sorted(blocks, key=lambda item: (item.y0, item.x0))


def extract_front_matter_blocks(pdf_path: Path) -> list[TextBlock]:
    """Extract only Titles I-IV, stopping at Annex 1."""

    if not pdf_path.exists():
        raise FileNotFoundError(f"PDF not found: {pdf_path}")

    doc = fitz.open(pdf_path)
    blocks: list[TextBlock] = []
    in_front_matter = False

    for page_index, page in enumerate(doc, start=1):
        page_footnotes = False
        for block in extract_page_text_blocks(page, page_index):
            if is_annexe_heading(block, ANNEXE_1_RE):
                return blocks

            if split_title(block.text):
                in_front_matter = True

            if not in_front_matter:
                continue

            if is_front_footnote_start(block):
                page_footnotes = True
            if page_footnotes or is_noise(block) or block.text.startswith("Fait à Paris"):
                continue

            blocks.append(block)

    return blocks


def extract_annexe_1_blocks(pdf_path: Path) -> list[TextBlock]:
    """Extract only Annex 1, stopping before Annex 2."""

    if not pdf_path.exists():
        raise FileNotFoundError(f"PDF not found: {pdf_path}")

    doc = fitz.open(pdf_path)
    blocks: list[TextBlock] = []
    in_annexe_1 = False

    for page_index, page in enumerate(doc, start=1):
        for block in extract_page_text_blocks(page, page_index):
            if in_annexe_1 and is_annexe_heading(block, ANNEXE_2_RE):
                return blocks
            if is_annexe_heading(block, ANNEXE_1_RE):
                in_annexe_1 = True
                continue
            if in_annexe_1 and not is_noise(block):
                blocks.append(block)

    return blocks


# =============================================================================
# Annex 2 extraction: network class definitions
# =============================================================================

def split_annexe_2_heading(text: str) -> str | None:
    """Return the Annex 2 title without the 'Annexe 2' prefix."""

    match = re.match(r"^Annexe\s+2\s*[-–]\s*(?P<title>.+)$", clean_text(text), re.IGNORECASE)
    if not match:
        return None
    return clean_name(match.group("title"))


def is_annexe_2_footnote_start(block: TextBlock) -> bool:
    """Detect the first footer note in Annex 2."""

    return block.y0 > 620 and re.match(r"^\d{1,2}\s+", block.text) is not None


def extract_annexe_2(pdf_path: Path) -> tuple[str | None, list[TextBlock]]:
    """Extract Annex 2 title and body blocks."""

    if not pdf_path.exists():
        raise FileNotFoundError(f"PDF not found: {pdf_path}")

    doc = fitz.open(pdf_path)
    blocks: list[TextBlock] = []
    title: str | None = None
    in_annexe_2 = False

    for page_index, page in enumerate(doc, start=1):
        page_footnotes = False
        for block in extract_page_text_blocks(page, page_index):
            if is_annexe_heading(block, ANNEXE_2_RE):
                in_annexe_2 = True
                title = split_annexe_2_heading(block.text)
                continue

            if not in_annexe_2:
                continue

            if is_annexe_2_footnote_start(block):
                page_footnotes = True
            if page_footnotes or is_noise(block):
                continue

            blocks.append(block)

        if in_annexe_2 and page_footnotes:
            return title, blocks

    return title, blocks


def split_network_class(text: str) -> tuple[str, str] | None:
    """Detect paragraphs starting a new Annex 2 network class."""

    cleaned = clean_text(text)
    match = NETWORK_CLASS_RE.match(cleaned)
    if not match:
        return None
    return match.group("class_id"), cleaned


def append_annexe_2_description_block(
    description: str | None, block: TextBlock, previous_block: TextBlock | None = None
) -> str:
    """Append Annex 2 body blocks while preserving lists and class paragraphs."""

    text = clean_text(block.text)
    if not text:
        return description or ""

    existing = description or ""
    lines = existing.splitlines()
    has_list = any(line.lstrip().startswith("* ") for line in lines)
    last_line = lines[-1] if lines else ""
    if block.x0 >= 105 and (has_list or last_line.rstrip().endswith(":")):
        level = max(1, int((block.x0 - 105) // 45) + 1)
        prefix = f"{'  ' * (level - 1)}* "
        return clean_multiline_text(f"{existing}\n{prefix}{strip_list_marker(text)}")

    if has_list and block.x0 < 100 and LIST_MARKER_RE.match(text) is None:
        return clean_multiline_text(f"{existing}\n{text}")

    return append_description_block(existing, block, previous_block)


def parse_annexe_2_rows(blocks: Sequence[TextBlock]) -> list[ContentRow]:
    """Convert Annex 2 network class paragraphs into nested rows."""

    rows: list[ContentRow] = []
    current_row: ContentRow | None = None

    for block in blocks:
        network_class = split_network_class(block.text)
        if network_class:
            class_id, description = network_class
            current_row = ContentRow(
                None,
                2,
                None,
                None,
                description,
                block,
            )
            rows.append(current_row)
            continue

        if current_row is None:
            continue

        current_row.description = append_annexe_2_description_block(
            current_row.description, block, current_row.last_description_block
        )
        current_row.last_description_block = block

    return rows


# =============================================================================
# Annex 1 parsing: blue sections, objectives, blue headings, requirements
# =============================================================================

def next_significant_block(blocks: Sequence[TextBlock], start_index: int) -> TextBlock | None:
    for block in blocks[start_index + 1 :]:
        if not is_noise(block):
            return block
    return None


def is_section_banner(blocks: Sequence[TextBlock], index: int) -> bool:
    """Detect the large blue banners in Annex 1."""

    block = blocks[index]
    if split_recommendation(block.text) or is_objective_marker(block):
        return False
    next_block = next_significant_block(blocks, index)
    if not next_block or not is_objective_marker(next_block):
        return False
    if block.x0 < 120 or block.y1 - block.y0 > 25:
        return False
    return True


def append_to_previous_description(rows: list[ContentRow], block: TextBlock) -> None:
    """Attach an orphan block to the latest Annex 1 row."""

    if not rows:
        return
    previous = rows[-1]
    previous.description = append_description_block(
        previous.description, block, previous.last_description_block
    )
    previous.last_description_block = block


def collect_objective_description(
    blocks: Sequence[TextBlock], start_index: int
) -> tuple[str | None, TextBlock | None, int]:
    """Collect an objective description, including its paragraphs and lists."""

    description: str | None = None
    last_description_block: TextBlock | None = None
    index = start_index + 1

    while index < len(blocks) and blocks[index].is_italic:
        if not is_noise(blocks[index]) and not is_objective_marker(blocks[index]):
            description = append_description_block(
                description, blocks[index], last_description_block
            )
            last_description_block = blocks[index]
        index += 1

    while index < len(blocks):
        block = blocks[index]
        if (
            is_noise(block)
            or is_objective_marker(block)
            or split_recommendation(block.text)
            or is_section_banner(blocks, index)
            or is_heading(block)
        ):
            break
        description = append_description_block(description, block, last_description_block)
        last_description_block = block
        index += 1

    return description, last_description_block, index


def split_objective_name_and_description(
    objective_text: str | None,
) -> tuple[str | None, str | None]:
    """Split an objective title from its extra explanatory text."""

    if not objective_text:
        return None, None

    name, separator, description = objective_text.partition("\n")
    if not separator:
        return clean_text(name), None
    return clean_text(name), clean_multiline_text(description)


# =============================================================================
# Titles I-IV parsing: titles, articles, paragraphs, and lists
# =============================================================================

def append_front_assessable_row(
    rows: list[ContentRow], block: TextBlock
) -> ContentRow:
    """Add a requirement from Titles II/III, or extend the previous one."""

    previous = rows[-1] if rows else None
    if (
        previous
        and previous.depth >= 3
        and is_same_visual_paragraph(previous.last_description_block, block)
    ):
        previous.description = append_front_description_block(
            previous.description, block, previous.last_description_block
        )
        previous.last_description_block = block
        return previous

    previous_description = previous.description if previous else ""
    is_list_item = LIST_MARKER_RE.match(block.text) is not None or (
        block.x0 >= 100
        and previous is not None
        and previous.depth >= 3
        and (previous.depth == 4 or (previous_description or "").rstrip().endswith(":"))
    )
    description = clean_text(strip_list_marker(block.text)) if is_list_item else clean_text(block.text)
    row = ContentRow("x", 4 if is_list_item else 3, None, None, description, block)
    rows.append(row)
    return row


def parse_front_matter_rows(
    blocks: Sequence[TextBlock], debug: bool = False
) -> list[ContentRow]:
    """Convert Titles I-IV into Excel rows.

    Titles I and IV remain non-assessable.
    Titles II and III create one requirement per paragraph, with lists indented
    at the next depth level.
    """

    rows: list[ContentRow] = []
    current_title_label: str | None = None
    current_article_seen = False
    current_non_assessable_row: ContentRow | None = None
    title_count = 0
    article_count = 0

    for block in blocks:
        title = split_title(block.text)
        if title:
            current_title_label, title_description = title
            rows.append(ContentRow(None, 1, current_title_label, title_description))
            current_article_seen = False
            current_non_assessable_row = None
            title_count += 1
            if debug:
                print(f"title: {current_title_label} - {title_description}")
            continue

        article = split_article(block.text)
        if article:
            article_name, article_description = article
            rows.append(ContentRow(None, 2, article_name, article_description))
            current_article_seen = True
            current_non_assessable_row = None
            article_count += 1
            if debug:
                print(f"article: {article_name} - {article_description}")
            continue

        if not current_article_seen:
            continue

        numbered_section = split_numbered_section(block.text)
        if numbered_section:
            ref_id, description = numbered_section
            row = ContentRow(
                "x" if is_assessable_title(current_title_label) else None,
                3,
                f"A{ref_id}",
                None,
                description,
                block,
            )
            rows.append(row)
            current_non_assessable_row = None if is_assessable_title(current_title_label) else row
            continue

        if is_assessable_title(current_title_label):
            append_front_assessable_row(rows, block)
            continue

        if current_non_assessable_row is None:
            current_non_assessable_row = ContentRow(None, 3, None, None)
            rows.append(current_non_assessable_row)

        current_non_assessable_row.description = append_front_description_block(
            current_non_assessable_row.description,
            block,
            current_non_assessable_row.last_description_block,
        )
        current_non_assessable_row.last_description_block = block

    if debug:
        requirements = sum(1 for row in rows if row.assessable == "x")
        print(f"front matter rows: {len(rows)}")
        print(f"front matter titles: {title_count}")
        print(f"front matter articles: {article_count}")
        print(f"front matter requirements: {requirements}")

    return rows


# =============================================================================
# Annex 1 parsing: Excel row construction
# =============================================================================

def parse_content_rows(blocks: Sequence[TextBlock], debug: bool = False) -> list[ContentRow]:
    """Convert Annex 1 into Excel rows."""

    rows: list[ContentRow] = []
    current_title_ref_id: str | None = None
    title_number = 0
    objective_number = 0
    section_number = 0

    index = 0
    while index < len(blocks):
        block = blocks[index]

        if is_section_banner(blocks, index):
            section_number += 1
            section_name = (
                SECTION_NAMES_BY_ORDER[section_number - 1]
                if section_number <= len(SECTION_NAMES_BY_ORDER)
                else clean_name(block.text)
            )
            rows.append(ContentRow(None, 1, None, section_name))
            current_title_ref_id = None
            title_number = 0
            if debug:
                print(f"section {section_number}: {section_name} (raw: {block.text})")
            index += 1
            continue

        if is_objective_marker(block):
            objective_number += 1
            title_number = 0
            current_title_ref_id = None
            description, last_description_block, next_index = collect_objective_description(
                blocks, index
            )
            objective_name, objective_description = split_objective_name_and_description(
                description
            )
            rows.append(
                ContentRow(
                    None,
                    2,
                    f"Objectif {objective_number}",
                    objective_name,
                    objective_description,
                    last_description_block,
                )
            )
            if debug:
                print(f"objective {objective_number}: {description}")
            index = next_index
            continue

        recommendation = split_recommendation(block.text)
        if recommendation:
            ref_id, name, description = recommendation
            rows.append(
                ContentRow(
                    "x",
                    4 if current_title_ref_id else 3,
                    ref_id,
                    name,
                    description,
                    block if description else None,
                )
            )
            index += 1
            continue

        if is_heading(block):
            title_number += 1
            current_title_ref_id = f"{objective_number}.{title_number}"
            rows.append(
                ContentRow(
                    None,
                    3,
                    current_title_ref_id,
                    clean_name(block.text),
                    None,
                    block,
                )
            )
            index += 1
            continue

        append_to_previous_description(rows, block)
        index += 1

    if debug:
        requirements = sum(1 for row in rows if row.assessable == "x")
        print(f"parsed rows: {len(rows)}")
        print(f"parsed requirements: {requirements}")
        print(f"parsed objectives: {objective_number}")

    return rows


def shift_depth(rows: Sequence[ContentRow], offset: int) -> list[ContentRow]:
    """Shift row depth, used to place Annex 1 under its own title."""

    return [
        ContentRow(
            row.assessable,
            row.depth + offset,
            row.ref_id,
            row.name,
            row.description,
            row.last_description_block,
            row.annotation,
        )
        for row in rows
    ]


# =============================================================================
# Excel file writing
# =============================================================================

def write_key_value_sheet(workbook: Workbook, sheet_name: str, rows: Iterable[tuple[str, str]]) -> None:
    """Write key/value metadata worksheets."""

    worksheet = workbook.create_sheet(sheet_name)
    for row in rows:
        worksheet.append(row)
    worksheet.column_dimensions["A"].width = 32
    worksheet.column_dimensions["B"].width = 120
    worksheet["A1"].font = Font(bold=True)
    worksheet["B1"].font = Font(bold=True)


def write_content_sheet(workbook: Workbook, rows: Sequence[ContentRow]) -> None:
    """Write the main II-901_content worksheet."""

    worksheet = workbook.create_sheet(CONTENT_SHEET)
    worksheet.append(CONTENT_COLUMNS)

    for row in rows:
        worksheet.append(row.as_tuple())

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    widths = {
        "A": 12,
        "B": 10,
        "C": 22,
        "D": 55,
        "E": 120,
        "F": 120,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    for row in worksheet.iter_rows(min_row=1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in worksheet.iter_rows(min_row=2):
        depth = row[1].value
        if depth in (1, 2, 3):
            for cell in row:
                cell.font = Font(bold=True if depth in (1, 2) else False)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = (
        f"A1:{get_column_letter(len(CONTENT_COLUMNS))}{worksheet.max_row}"
    )


def build_workbook(rows: Sequence[ContentRow], output_path: Path) -> None:
    """Build the complete workbook with metadata and content."""

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    write_key_value_sheet(workbook, "library_meta", LIBRARY_META_ROWS)
    write_key_value_sheet(workbook, "II-901_meta", FRAMEWORK_META_ROWS)
    write_content_sheet(workbook, rows)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


# =============================================================================
# Orchestration
# =============================================================================

def main() -> None:
    """Entrypoint: extract the PDF, parse rows, then write the XLSX."""

    args = parse_args()
    footnotes_by_page = extract_footnotes(args.input)

    front_matter_blocks = extract_front_matter_blocks(args.input)
    front_matter_rows = parse_front_matter_rows(front_matter_blocks, debug=args.debug)
    attach_footnotes_to_rows(front_matter_rows, footnotes_by_page)

    annexe_1_blocks = extract_annexe_1_blocks(args.input)
    annexe_1_footnote = footnotes_by_page.get(11, {}).get("9")
    annexe_1_rows = [
        ContentRow(
            None,
            1,
            "Annexe 1",
            "Règles pour les entités situées hors du champ d'application de la PSSIE",
            None,
            None,
            format_footnote_annotation(annexe_1_footnote)
            if annexe_1_footnote
            else None,
        )
    ]
    annexe_1_content_rows = parse_content_rows(annexe_1_blocks, debug=args.debug)
    attach_footnotes_to_rows(annexe_1_content_rows, footnotes_by_page)
    annexe_1_rows.extend(shift_depth(annexe_1_content_rows, 1))

    annexe_2_title, annexe_2_blocks = extract_annexe_2(args.input)
    annexe_2_rows = [
        ContentRow(
            None,
            1,
            "Annexe 2",
            annexe_2_title or "Différentes classes de réseau",
        )
    ]
    annexe_2_content_rows = parse_annexe_2_rows(annexe_2_blocks)
    attach_footnotes_to_rows(annexe_2_content_rows, footnotes_by_page)
    annexe_2_rows.extend(annexe_2_content_rows)

    rows = [*front_matter_rows, *annexe_1_rows, *annexe_2_rows]
    build_workbook(rows, args.output)
    print(f"Wrote {args.output}")


if __name__ == "__main__":
    main()
