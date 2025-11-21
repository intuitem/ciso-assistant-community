#####################################################################################################################################
##### [STEP 1] Extract French & English version of "ANSSI Active Directory (AD) Security Assessment Checklist" from the website #####
#####################################################################################################################################

import os
from anssi_ad_web_scraper import main as framework_extractor

FILES_SUFFIX_FR = "FR"
FILES_SUFFIX_EN = "EN"

print('\n#####################################################################################################################################')
print("##### [STEP 1] Extract French & English version of \"ANSSI Active Directory (AD) Security Assessment Checklist\" from the website #####")
print('#####################################################################################################################################\n')

print("[MAIN] ▶️  Extracting French & English from website...\n")

# Extract French
framework_extractor(False, FILES_SUFFIX_FR)

print("")

# Extract English
framework_extractor(True, FILES_SUFFIX_EN)


# Remove useless Markdown files
print("\n[MAIN] ▶️  Removing useless Markdown files...")

os.remove(f"checklist_{FILES_SUFFIX_FR}.md")
os.remove(f"checklist_{FILES_SUFFIX_EN}.md")


print("[MAIN] ✅ [STEP 1] Completed!")





##############################################
##### [STEP 2] Create Excel from scratch #####
##############################################

import re
import sys
import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

parent_dir = Path(__file__).resolve().parent.parent.parent.parent
sys.path.insert(0, str(parent_dir))

from prepare_framework_v2 import create_excel_from_yaml


YAML_CONFIG_FILENAME = "prepare_anssi_framework.yaml"
EXCEL_FILE_NAME = "anssi-points-de-controle-active-directory.xlsx"



print('\n##############################################')
print('##### [STEP 2] Create Excel from scratch #####')
print('##############################################\n')


### Import extracted JSON files

print("[MAIN] ▶️  Loading Framework JSON files...")


JSON_FR = f"checklist_markdown{'_' + FILES_SUFFIX_FR}.json"
JSON_EN = f"checklist_markdown{'_' + FILES_SUFFIX_EN}.json"

json_framework_fr = None
json_framework_en = None

with open(JSON_FR, 'r') as file1:
    json_framework_fr = json.load(file1)
    

with open(JSON_EN, 'r') as file2:
    json_framework_en = json.load(file2)


### Create reusable variables

print("[MAIN] ▶️  Creating YAML Config file to create base Excel file...")


def extract_intro_paragraphs(json_file: dict, max_paragraphs: int = 4) -> list[str]:
    """
    Extract the first 'max_paragraphs' intro paragraphs from the 'intro' root_element
    of the checklist JSON, without relying on literal French text.

    Heuristic:
      - Use root_elements where root_element == "intro"
      - Skip markdown headings (lines starting with '#')
      - Split content into blocks on double newlines
      - Consider a block a "paragraph" if it does NOT start with:
          * '#'  (heading)
          * '*'  (list item)
      - Return the first 'max_paragraphs' such paragraphs.
    """


    # Find the intro root element
    intro_elem = None
    for elem in json_file.get("root_elements", []):
        if elem.get("root_element") == "intro":
            intro_elem = elem
            break

    if intro_elem is None:
        raise ValueError("Could not find root_element == 'intro' in JSON.")

    md = intro_elem.get("content_markdown", "")

    # Split on blank lines (markdown paragraphs)
    raw_blocks = md.split("\n\n")

    paragraphs: list[str] = []

    for block in raw_blocks:
        block = block.strip()
        if not block:
            continue

        # Skip headings like "# Introduction"
        if block.startswith("#"):
            continue

        # Skip list items (e.g. "* <span style=...>1</span> ...")
        if block.startswith("* "):
            continue

        # This is considered a "paragraph" block
        paragraphs.append(block)

        if len(paragraphs) >= max_paragraphs:
            break

    return paragraphs


def clean_level_bullet(line: str) -> str:
    """
    Take a raw markdown bullet line from the intro, like:
        * <span style="...">1</span> L'annuaire Active Directory ...

    And return a clean sentence without:
        - the leading "* "
        - the <span> with the level number
        - any remaining HTML tags
    """
    s = line.strip()

    # Remove leading "* "
    if s.startswith("* "):
        s = s[2:].lstrip()

    # Remove leading <span ...>X</span> (the colored level number)
    s = re.sub(r'^<span[^>]*>[^<]*</span>\s*', '', s)

    # Remove any remaining HTML tags (if any)
    s = re.sub(r'<[^>]+>', '', s)

    return s.strip()


def extract_level_descriptions(json_file: dict) -> list[str]:
    """
    Return a list of 5 cleaned strings, one per level (1..5),
    extracted from the 'intro' root_element of the JSON.
    """
    

    # Find the "intro" root element
    intro_elem = None
    for elem in json_file.get("root_elements", []):
        if elem.get("root_element") == "intro":
            intro_elem = elem
            break

    if intro_elem is None:
        raise ValueError("Could not find root_element == 'intro' in JSON.")

    md = intro_elem.get("content_markdown", "")
    lines = md.splitlines()

    level_bullets_raw: list[str] = []

    for line in lines:
        # Normalize non-breaking spaces and strip right side
        stripped = replace_non_breaking_space(line)
        # Level bullets: lines starting with "* " and containing the colored span
        if stripped.startswith("* ") and "<span style=" in stripped:
            level_bullets_raw.append(stripped)

    # Keep only the first 5 bullets (levels 1..5)
    level_bullets_raw = level_bullets_raw[:5]

    # Clean each bullet
    cleaned = [clean_level_bullet(b) for b in level_bullets_raw]

    return cleaned


def replace_non_breaking_space(string: str) -> str:
    
    if type(string) == str:
        return string.replace("\xa0", " ").rstrip()
    else:
        return None 


# Metadata
fw_urn = "anssi-points-de-controle-active-directory"
fw_id = fw_urn
copyright = "ANSSI"
provider = copyright
packager = "intuitem"

fw_desc_FR = extract_intro_paragraphs(json_framework_fr)
fw_desc_EN = extract_intro_paragraphs(json_framework_en)

# Implementation Groups (IGs)
ig_level_1 = "level_1"
ig_level_2 = "level_2"
ig_level_3 = "level_3"
ig_level_4 = "level_4"

ig_level_descs_FR = extract_level_descriptions(json_framework_fr)
ig_level_descs_EN = extract_level_descriptions(json_framework_en)



### Create YAML Config file

yaml_config_file = f"""
# --- Metadata sheets ---
urn_root: {fw_urn}
locale: "fr"
ref_id: {fw_id}
framework_name: "Points de contrôle Active Directory (AD)"
description: |
  {'\n  '.join(fw_desc_FR)}

copyright: {copyright}
provider: {provider}
packager: {packager}

# --- Specific to Framework sheets ---
framework_sheet_base_name: "framework"

# --- To enable implementation groups, uncomment the lines below ---
implementation_groups_sheet_base_name: "imp_grp"

implementation_groups:
  - ref_id: {ig_level_1}
    name: "Niveau 1"
    description: |
      {ig_level_descs_FR[0]}

  - ref_id: {ig_level_2}
    name: "Niveau 2"
    description: |
      {ig_level_descs_FR[1]}

  - ref_id: {ig_level_3}
    name: "Niveau 3"
    description: |
      {ig_level_descs_FR[2]}

  - ref_id: {ig_level_4}
    name: "Niveau 4"
    description: |
      {ig_level_descs_FR[3]}

# --- To enable extra locales (localized framework versions), uncomment the block below ---
extra_locales:
  - en:
      framework_name: ANSSI Active Directory (AD) Security Assessment Checklist
      description: |
        {'\n        '.join(fw_desc_EN)}

      copyright: {copyright}

      implementation_groups:
      - ref_id: {ig_level_1}
        name: "Level 1"
        description: |
          {ig_level_descs_EN[0]}

      - ref_id: {ig_level_2}
        name: "Level 2"
        description: |
          {ig_level_descs_EN[1]}

      - ref_id: {ig_level_3}
        name: "Level 3"
        description: |
          {ig_level_descs_EN[2]}

      - ref_id: {ig_level_4}
        name: "Level 4"
        description: |
          {ig_level_descs_EN[3]}

"""


with open(YAML_CONFIG_FILENAME, "w", encoding="utf-8") as f:
    f.write(yaml_config_file)



### Create Excel base Framework

print("[MAIN] ▶️  Creating base Excel file form YAML Config file...\n")

create_excel_from_yaml(YAML_CONFIG_FILENAME, EXCEL_FILE_NAME)



### Remove YAML Configuration file

print("\n[MAIN] ▶️  Removing YAML Configuration file...")

os.remove(YAML_CONFIG_FILENAME)



### Creating Framework Excel sheet

print("\n[MAIN] ▶️  Creating Framework Excel sheet...")



def strip_first_markdown_block(md: str) -> str:
    """
    Remove the first markdown block (e.g. '# Introduction') up to the first
    double newline, and return the rest.
    """
    if not md:
        return ""
    parts = md.split("\n\n", 1)
    if len(parts) == 2:
        return parts[1].strip()
    # If no double newline, return the original or empty; here we choose original
    return md.strip()


def build_framework_content(json_fr: dict, json_en: dict, base_excel_file_name: str) -> None:
    """
    Read French and English JSON exports, build the 'framework_content' sheet,
    and write it into the given Excel file (replacing the sheet if it exists).

    All cells are written as text (strings).
    """

    # Small helpers to get root_elements by key
    def get_root_element(data, key: str):
        for elem in data.get("root_elements", []):
            if elem.get("root_element") == key:
                return elem
        return None

    intro_fr = get_root_element(json_fr, "intro")
    intro_en = get_root_element(json_en, "intro")
    about_fr = get_root_element(json_fr, "about_ctrl_points")
    about_en = get_root_element(json_en, "about_ctrl_points")

    if intro_fr is None or intro_en is None:
        raise ValueError("Missing 'intro' root_element in json_fr or json_en.")

    if about_fr is None or about_en is None:
        raise ValueError("Missing 'about_ctrl_points' root_element in json_fr or json_en.")

    # Extract bodies (content_markdown without the first heading block)
    intro_body_fr = strip_first_markdown_block(intro_fr.get("content_markdown", ""))
    intro_body_en = strip_first_markdown_block(intro_en.get("content_markdown", ""))

    about_body_fr = strip_first_markdown_block(about_fr.get("content_markdown", ""))
    about_body_en = strip_first_markdown_block(about_en.get("content_markdown", ""))

    # ---------- Prepare rows ----------
    columns = [
        "assessable",
        "depth",
        "ref_id",
        "name",
        "description",
        "annotation",
        "implementation_groups",
        "name[en]",
        "description[en]",
        "annotation[en]",
    ]

    rows = []

    def add_row(
        assessable="",
        depth="",
        ref_id="",
        name="",
        description="",
        annotation="",
        implementation_groups="",
        name_en="",
        description_en="",
        annotation_en="",
    ):
        # Ensure everything is string (Excel: text only)
        row = {
            "assessable": str(replace_non_breaking_space(assessable)) if assessable is not None else "",
            "depth": str(depth) if depth is not None else "",
            "ref_id": str(ref_id) if ref_id is not None else "",
            "name": str(replace_non_breaking_space(name)) if name is not None else "",
            "description": str(replace_non_breaking_space(description)) if description is not None else "",
            "annotation": str(replace_non_breaking_space(annotation)) if annotation is not None else "",
            "implementation_groups": str(implementation_groups) if implementation_groups is not None else "",
            "name[en]": str(replace_non_breaking_space(name_en)) if name_en is not None else "",
            "description[en]": str(replace_non_breaking_space(description_en)) if description_en is not None else "",
            "annotation[en]": str(replace_non_breaking_space(annotation_en)) if annotation_en is not None else "",
        }
        rows.append(row)

    # ---------- 1) intro ----------

    # depth = 1 ; ref_id = "intro" ; name/name[en]
    add_row(
        depth = "1",
        ref_id = "intro",
        name = "Introduction",
        name_en = "Introduction",
    )

    # depth = 2 ; ref_id = "intro_text" ; description/description[en]
    add_row(
        depth="2",
        ref_id="intro_text",
        description = intro_body_fr,
        description_en = intro_body_en,
    )

    # ---------- 2) about_ctrl_points ----------

    # depth = 1 ; ref_id = "about_assessment_item_list"
    add_row(
        depth = "1",
        ref_id = "about_assessment_item_list",
        name = "A propos de la Liste des points de contrôle",
        name_en = "About Assessment Item List",
    )

    # depth = 2 ; ref_id = "about_assessment_item_list_text" ; description/description[en]
    add_row(
        depth="2",
        ref_id = "about_assessment_item_list_text",
        description = about_body_fr,
        description_en = about_body_en,
    )

    # ---------- 3) checklist ----------

    fr_checklist = json_fr.get("checklist", [])
    en_checklist = json_en.get("checklist", [])

    # Index EN checklist by identifier for quick lookup
    en_by_id = {item.get("identifier"): item for item in en_checklist}

    # Before the loop: assessment_item_list group
    add_row(
        depth="1",
        ref_id="assessment_item_list",
        name="Liste des points de contrôle",
        name_en="Assessment Item List",
    )

    for fr_item in fr_checklist:
        identifier = fr_item.get("identifier", "")
        en_item = en_by_id.get(identifier, {})

        # Level may be "123" or something like "1,3,5" → collect all digits
        level_raw = fr_item.get("level", "") or ""
        level_digits = re.findall(r"\d", level_raw)

        title_fr = fr_item.get("title", "")
        title_en = en_item.get("title", "")

        desc_fr = fr_item.get("description_markdown", "")
        desc_en = en_item.get("description_markdown", "")

        reco_fr = fr_item.get("recommendation_markdown", "")
        reco_en = en_item.get("recommendation_markdown", "")

        # One row per level digit
        for d in level_digits:
            impl_group = f"level_{d}"

            ref_id_with_level = f"{identifier}-l{d}"
            name_fr_with_level = f"{title_fr} (Niveau {d})"
            name_en_with_level = f"{title_en} (Level {d})"

            add_row(
                assessable="x",
                depth="2",
                ref_id=ref_id_with_level,
                name=name_fr_with_level,
                description=desc_fr,
                annotation=reco_fr,
                implementation_groups=impl_group,
                name_en=name_en_with_level,
                description_en=desc_en,
                annotation_en=reco_en,
            )


    # ---------- Build DataFrame ----------
    df = pd.DataFrame(rows, columns=columns).astype(str)
    print(df)

    # ---------- Write to Excel (replace framework_content sheet) ----------
    # Load existing workbook
    wb = load_workbook(EXCEL_FILE_NAME)

    # Determine desired index for the sheet
    if "framework_content" in wb.sheetnames:
        old_index = wb.sheetnames.index("framework_content")
        # Remove old sheet
        del wb["framework_content"]
    else:
        # If it didn't exist, put it at the end
        old_index = len(wb.sheetnames)

    # Create new sheet at the same position
    ws = wb.create_sheet("framework_content", index=old_index)

    # Write DataFrame into the sheet as text-only
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append([str(cell) if cell is not None else "" for cell in row])

    # Save workbook
    wb.save(EXCEL_FILE_NAME)



build_framework_content(json_framework_fr, json_framework_en, EXCEL_FILE_NAME)



# Remove JSON files

print("\n[MAIN] ▶️  Removing JSON files...")

os.remove(JSON_FR)
os.remove(JSON_EN)


print("[MAIN] ✅ [STEP 2] Completed!")
print(f"[MAIN] ✅ Excel file saved successfully: \"{EXCEL_FILE_NAME}\"")