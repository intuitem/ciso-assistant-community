#!/usr/bin/env python3
"""
MonAideCyber - Excel Framework Builder
======================================

Builds a v2 framework Excel workbook from:
- `questionnaire_repo.json`
- `mesures_repo.json`

Workbook includes metadata/content tabs for:
- framework
- answers
- implementation groups
- reference controls
- URN prefixes

Run
---
python anssi_MAC_build_excel_from_json.py --questionnaire questionnaire_repo.json --mesures mesures_repo.json
"""

import argparse
import json
import re
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from typing import Any

from openpyxl import Workbook

FRAMEWORK_HEADERS = [
    "assessable",
    "depth",
    "node_id",
    "name",
    "description",
    "annotation",
    "typical_evidence",
    "implementation_groups",
    "questions",
    "answer",
    "reference_controls",
]

IMP_GRP_HEADERS = ["ref_id", "name", "description", "default_selected"]
ANSW_HEADERS = [
    "id",
    "question_type",
    "question_choices",
    "color",
    "description",
    "select_implementation_groups",
]
REF_CTRL_HEADERS = [
    "ref_id",
    "name",
    "category",
    "csf_function",
    "description",
    "annotation",
]

PERIMETRE_LABELS = {
    "SYSTEME-INDUSTRIEL": "SYSTÈME INDUSTRIEL",
    "ATTAQUE-CIBLEE": "ATTAQUE CIBLÉE",
}

QUESTION_TYPE_MAP = {
    "choixMultiple": "multiple_choice",
    "choixUnique": "unique_choice",
}


class JoinSeparator(str, Enum):
    DEFAULT_NEWLINE = "\n"
    IMPLEMENTATION_GROUPS = ",\n"


class AnswerColor(str, Enum):
    NOT_APPLICABLE = "#000000"
    DONT_KNOW = "#555555"
    NO = "#8B0000"
    YES = "#209226"
    EMPTY = "/"


LIBRARY_META_ROWS: list[tuple[str, str]] = [
    ("type", "library"),
    ("urn", "urn:intuitem:risk:library:anssi-monaidecyber"),
    ("version", "1"),
    ("locale", "fr"),
    ("ref_id", "ANSSI-MonAideCyber"),
    ("name", "ANSSI - Questionnaire MonAideCyber"),
    (
        "description",
        "MonAideCyber aide les entités publiques et privées sensibilisées à la sécurité informatique à passer à l’action. Le dispositif MonAideCyber est développé par l'Agence Nationale de la Sécurité des Systèmes d'Information, en lien avec BetaGouv et la Direction interministérielle du numérique.\n\n Sources :\n\t• https://github.com/betagouv/mon-aide-cyber \n\t• https://monaide.cyber.gouv.fr/diagnostic-libre-acces",
    ),
    (
        "copyright",
        "Copyright 2023 ANSSI - Agence nationale de la sécurité des systèmes d'information (https://www.cyber.gouv.fr)\n\n"
        "Licensed under the Apache License, Version 2.0 (the \"\"License\"\");\n"
        "you may not use this file except in compliance with the License.\n"
        "You may obtain a copy of the License at\n\n"
        "    http://www.apache.org/licenses/LICENSE-2.0\n\n"
        "Unless required by applicable law or agreed to in writing, software distributed under the License is distributed on an \"\"AS IS\"\" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.\n"
        "See the License for the specific language governing permissions and limitations under the License.",
    ),
    ("provider", "ANSSI"),
    ("packager", "intuitem"),
]

FWK_META_ROWS: list[tuple[str, str]] = [
    ("type", "framework"),
    ("base_urn", "urn:intuitem:risk:req_node:anssi-monaidecyber"),
    ("urn", "urn:intuitem:risk:framework:anssi-monaidecyber"),
    ("ref_id", "ANSSI-MonAideCyber"),
    ("name", "ANSSI - Questionnaire MonAideCyber"),
    (
        "description",
        "MonAideCyber aide les entités publiques et privées sensibilisées à la sécurité informatique à passer à l’action. Le dispositif MonAideCyber est développé par l'Agence Nationale de la Sécurité des Systèmes d'Information, en lien avec BetaGouv et la Direction interministérielle du numérique.\n\n Sources :\n\t• https://github.com/betagouv/mon-aide-cyber \n\t• https://monaide.cyber.gouv.fr/diagnostic-libre-acces",
    ),
    ("implementation_groups_definition", "imp_grp"),
    ("answers_definition", "answ"),
]

ANSW_META_ROWS: list[tuple[str, str]] = [
    ("type", "answers"),
    ("name", "answ"),
]

IMP_GRP_META_ROWS: list[tuple[str, str]] = [
    ("type", "implementation_groups"),
    ("name", "imp_grp"),
]

REF_CTRL_META_ROWS: list[tuple[str, str]] = [
    ("type", "reference_controls"),
    ("base_urn", "urn:intuitem:risk:function:anssi-monaidecyber"),
]

URN_PREF_META_ROWS: list[tuple[str, str]] = [
    ("type", "urn_prefix"),
]

URN_PREF_HEADERS = ["prefix_id", "prefix_value"]
URN_PREF_ROWS = [
    {"prefix_id": "1", "prefix_value": "urn:intuitem:risk:function:anssi-monaidecyber"},
]


@dataclass
class AnswerRow:
    qid: str
    question_type: str
    response_labels: list[str]
    response_ids: list[str]
    first_order: int
    select_groups: list[str] = field(default_factory=list)

    def ensure_line_count(self) -> None:
        if not self.select_groups:
            self.select_groups = ["/" for _ in self.response_labels]
        if len(self.select_groups) < len(self.response_labels):
            self.select_groups.extend(
                "/" for _ in range(len(self.response_labels) - len(self.select_groups))
            )

    def _line_index_from_ordre(self, ordre: int) -> int:
        base = ordre if self.first_order > 0 else ordre + 1
        idx = max(base - 1, 0)
        if idx >= len(self.response_labels):
            idx = len(self.response_labels) - 1
        return idx

    def add_group_at_ordre(self, ordre: int, group: str) -> None:
        if not self.response_labels:
            return
        self.ensure_line_count()
        idx = self._line_index_from_ordre(ordre)
        self._merge_group_at_index(idx, group)

    def add_group_except_response_id(
        self, group: str, response_id_to_skip: str
    ) -> None:
        if not self.response_labels:
            return
        self.ensure_line_count()
        for idx, rid in enumerate(self.response_ids):
            if rid == response_id_to_skip:
                continue
            self._merge_group_at_index(idx, group)

    def _merge_group_at_index(self, idx: int, group: str) -> None:
        current = self.select_groups[idx]
        if current == "/" or current == "":
            self.select_groups[idx] = group
            return
        parts = [p.strip() for p in current.split(",") if p.strip()]
        if group in parts:
            return
        parts.append(group)
        self.select_groups[idx] = ", ".join(parts)


@dataclass
class Context:
    framework_rows: list[dict[str, str]] = field(default_factory=list)
    imp_grp_rows: list[dict[str, str]] = field(default_factory=list)
    ref_ctrl_rows: list[dict[str, str]] = field(default_factory=list)
    answer_rows: dict[str, AnswerRow] = field(default_factory=dict)
    imp_grp_seen: set[str] = field(default_factory=set)
    theme_counter: int = 0
    root_question_counter: int = 0
    framework_index_by_urn: dict[str, int] = field(default_factory=dict)
    question_name_suffix_by_urn: dict[str, str] = field(default_factory=dict)


def as_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def load_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def ensure_imp_group(
    ctx: Context,
    ref_id: str,
    name: str,
    description: str = "",
    default_selected: str = "",
) -> None:
    if ref_id in ctx.imp_grp_seen:
        return
    ctx.imp_grp_seen.add(ref_id)
    ctx.imp_grp_rows.append(
        {
            "ref_id": ref_id,
            "name": name,
            "description": description,
            "default_selected": default_selected,
        }
    )


def map_annotation_from_perimetre(perimetre: str | None) -> str:
    if not perimetre:
        return ""
    return (
        "## **" + PERIMETRE_LABELS.get(perimetre, "") + "**"
        if PERIMETRE_LABELS.get(perimetre, "")
        else ""
    )


def prepend_h2_markdown(text: str) -> str:
    stripped = text.lstrip()
    if stripped.startswith("## "):
        return text
    return f"## {text}"


def insert_double_newline_before_title_word(text: str, limit_index: int) -> str:
    matches = list(
        re.finditer(r"[A-Za-zÀ-ÖØ-öø-ÿ][A-Za-zÀ-ÖØ-öø-ÿ'-]*", text[:limit_index])
    )
    for match in reversed(matches):
        word = match.group(0)
        # Keep words that start with uppercase but are not fully uppercase.
        if word[0].isupper() and not word.isupper():
            return text[: match.start()] + "\n\n" + text[match.start() :]
    return text


def format_info_bulle_text(text: str) -> str:
    value = as_text(text)
    first_double_newline = value.find("\n\n")
    first_newline_tab = value.find("\n\t")

    # Special case: '\n\t' appears before first '\n\n'
    if first_newline_tab != -1 and (
        first_double_newline == -1 or first_newline_tab < first_double_newline
    ):
        value = insert_double_newline_before_title_word(value, first_newline_tab)
        return prepend_h2_markdown(value)

    # Generic case: if header-like part has at most 10 words, prefix with '# '
    header_part = value[:first_double_newline] if first_double_newline != -1 else value
    word_count = len(re.findall(r"\S+", header_part))
    if word_count <= 10:
        return prepend_h2_markdown(value)

    return value


def build_annotation(perimetre: str | None, info_bulles_textes: list[Any]) -> str:
    parts: list[str] = []

    perimetre_annotation = map_annotation_from_perimetre(perimetre)
    if perimetre_annotation:
        parts.append(perimetre_annotation)

    for info in info_bulles_textes:
        formatted = format_info_bulle_text(as_text(info))
        if formatted:
            parts.append(formatted)

    return "\n\n".join(parts)


def implementation_groups_for_question(perimetre: str | None, base_group: str) -> str:
    parts = [base_group]
    if perimetre:
        parts.append(perimetre)
    return JoinSeparator.IMPLEMENTATION_GROUPS.join(parts)


def measures_to_reference_controls(reponses: list[dict[str, Any]]) -> str:
    refs: list[str] = []
    seen: set[str] = set()

    for rep in reponses:
        if not isinstance(rep, dict):
            continue

        mesures_candidates: list[Any] = []
        # Current exported JSON format: reponsesPossibles[].resultat.mesures
        resultat = rep.get("resultat")
        if isinstance(resultat, dict):
            mesures_candidates.extend(resultat.get("mesures", []) or [])

        # Robust fallback if the structure becomes direct again in future updates
        mesures_candidates.extend(rep.get("mesures", []) or [])

        for mesure in mesures_candidates:
            if not isinstance(mesure, dict):
                continue
            identifiant = mesure.get("identifiant")
            niveau = mesure.get("niveau")
            if identifiant is None or niveau is None:
                continue
            value = f"1:{identifiant}_niveau{niveau}"
            if value not in seen:
                seen.add(value)
                refs.append(value)

    return "\n".join(refs)


def build_answer_row_for_question(question: dict[str, Any]) -> AnswerRow:
    reponses = question.get("reponsesPossibles", []) or []
    response_labels = [
        as_text(r.get("libelle")) for r in reponses if isinstance(r, dict)
    ]
    response_ids = [
        as_text(r.get("identifiant")) for r in reponses if isinstance(r, dict)
    ]
    first_order = 0
    if reponses and isinstance(reponses[0], dict):
        try:
            first_order = int(reponses[0].get("ordre", 0))
        except (TypeError, ValueError):
            first_order = 0
    return AnswerRow(
        qid=as_text(question.get("identifiant")),
        question_type=QUESTION_TYPE_MAP.get(as_text(question.get("type")), ""),
        response_labels=response_labels,
        response_ids=response_ids,
        first_order=first_order,
    )


def append_framework_row(ctx: Context, row: dict[str, str]) -> int:
    idx = len(ctx.framework_rows)
    ctx.framework_rows.append(row)

    urn = row.get("node_id", "")
    if urn:
        ctx.framework_index_by_urn[urn] = idx

    name = row.get("name", "")
    if urn and name.startswith("Question "):
        ctx.question_name_suffix_by_urn[urn] = name.removeprefix("Question ").strip()

    return idx


def process_question(
    ctx: Context,
    theme_num: int,
    question: dict[str, Any],
    *,
    parent_question_id: str | None = None,
    parent_question_label: str | None = None,
    parent_response: dict[str, Any] | None = None,
    parent_ref_prefix: str | None = None,
    child_index_in_parent: int | None = None,
) -> None:
    qid = as_text(question.get("identifiant"))
    perimetre = as_text(question.get("perimetre")) or None
    info_bulles_textes = question.get("info-bulles-textes", []) or []
    if not isinstance(info_bulles_textes, list):
        info_bulles_textes = []
    annotation = build_annotation(perimetre, info_bulles_textes)

    if parent_question_id is None:
        ctx.root_question_counter += 1
        global_question_number = ctx.root_question_counter
        question_label = as_text(global_question_number)
        ref_id = f"{theme_num}.{global_question_number}"
        name = f"Question {question_label}"
        implementation_groups = implementation_groups_for_question(perimetre, "MAIN")
    else:
        index_in_parent = child_index_in_parent or 1
        base_label = parent_question_label or ""
        question_label = (
            f"{base_label}.{index_in_parent}"
            if base_label
            else as_text(index_in_parent)
        )
        base_ref = parent_ref_prefix or f"{theme_num}"
        ref_id = f"{base_ref}.{index_in_parent}"
        name = f"Question {question_label}"

        depends_group = f"_Q{question_label}_DEPENDS_ON_{parent_question_id}"
        implementation_groups = implementation_groups_for_question(
            perimetre, depends_group
        )
        ensure_imp_group(
            ctx,
            depends_group,
            f"Question {question_label} dépendante de la Question {parent_question_label or ''}".strip(),
        )
        if parent_response is not None:
            parent_answ = ctx.answer_rows.get(parent_question_id)
            if parent_answ is not None:
                ordre = int(parent_response.get("ordre", 0) or 0)
                parent_answ.add_group_at_ordre(ordre, depends_group)

    if perimetre:
        ensure_imp_group(ctx, perimetre, PERIMETRE_LABELS.get(perimetre, ""))

    depth = "2" if parent_question_id is None else "3"

    framework_row = {
        "assessable": "x",
        "depth": depth,
        "node_id": qid.lower(),
        "name": name,
        "description": as_text(question.get("description")),
        "annotation": annotation,
        "typical_evidence": "",
        "implementation_groups": implementation_groups,
        "questions": as_text(question.get("libelle")),
        "answer": qid,
        "reference_controls": measures_to_reference_controls(
            question.get("reponsesPossibles", []) or []
        ),
    }
    append_framework_row(ctx, framework_row)

    answer_row = build_answer_row_for_question(question)
    if qid:
        ctx.answer_rows[qid] = answer_row

    child_counter = 0
    for rep in question.get("reponsesPossibles", []) or []:
        if not isinstance(rep, dict):
            continue
        for child in rep.get("questions", []) or []:
            if not isinstance(child, dict):
                continue
            child_counter += 1
            process_question(
                ctx,
                theme_num,
                child,
                parent_question_id=qid,
                parent_question_label=question_label,
                parent_response=rep,
                parent_ref_prefix=ref_id,
                child_index_in_parent=child_counter,
            )


def process_referentiel(ctx: Context, referentiel: dict[str, Any]) -> None:
    ensure_imp_group(ctx, "MAIN", "Éléments principaux", default_selected="x")

    for theme_key, theme_obj in referentiel.items():
        if theme_key.startswith("_"):
            continue
        if not isinstance(theme_obj, dict):
            continue

        ctx.theme_counter += 1
        theme_num = ctx.theme_counter

        append_framework_row(
            ctx,
            {
                "assessable": "",
                "depth": "1",
                "node_id": theme_key.lower(),
                "name": as_text(theme_obj.get("libelle")),
                "description": as_text(theme_obj.get("description")),
                "annotation": "",
                "typical_evidence": "",
                "implementation_groups": "",
                "questions": "",
                "answer": "",
                "reference_controls": "",
            },
        )

        for question in theme_obj.get("questions", []) or []:
            if isinstance(question, dict):
                process_question(ctx, theme_num, question)


def remove_main_and_add_groups(existing: str, groups: list[str]) -> str:
    parts = (
        [p.strip() for p in re.split(r",?\n", existing) if p.strip()]
        if existing
        else []
    )
    parts = [p for p in parts if p != "MAIN"]
    for grp in groups:
        if grp not in parts:
            parts.append(grp)
    return JoinSeparator.IMPLEMENTATION_GROUPS.join(parts)


def apply_conditions_perimetre_second_pass(
    ctx: Context, referentiel: dict[str, Any]
) -> None:
    conditions = referentiel.get("_conditionsPerimetre")
    if not isinstance(conditions, dict):
        return

    for target_question_id, rule_obj in conditions.items():
        target_idx = ctx.framework_index_by_urn.get(as_text(target_question_id).lower())
        if target_idx is None or not isinstance(rule_obj, dict):
            continue

        contexte_obj = rule_obj.get("contexte")
        if not isinstance(contexte_obj, dict):
            continue

        groups_to_add: list[str] = []
        for context_question_id, expected_response_id in contexte_obj.items():
            dep_group = f"_DEPENDS_ON_{context_question_id}"
            groups_to_add.append(dep_group)

            urn_c = as_text(context_question_id).lower()
            q_label = ctx.question_name_suffix_by_urn.get(urn_c, "")
            ensure_imp_group(
                ctx,
                dep_group,
                f"Questions dépendantes de la Question {q_label}".strip(),
            )

            answ_row = ctx.answer_rows.get(as_text(context_question_id))
            if answ_row is not None:
                answ_row.add_group_except_response_id(
                    dep_group, as_text(expected_response_id)
                )

        row = ctx.framework_rows[target_idx]
        row["implementation_groups"] = remove_main_and_add_groups(
            row.get("implementation_groups", ""), groups_to_add
        )


def build_ref_controls_rows(ctx: Context, mesures_json: dict[str, Any]) -> None:
    mesures = mesures_json.get("mesures") if isinstance(mesures_json, dict) else None
    if not isinstance(mesures, dict):
        return

    for mesure_id, mesure_obj in mesures.items():
        if not isinstance(mesure_obj, dict):
            continue
        for level_key, level_obj in mesure_obj.items():
            if not str(level_key).startswith("niveau"):
                continue
            if not isinstance(level_obj, dict):
                continue
            title = as_text(level_obj.get("titre"))
            pourquoi = as_text(level_obj.get("pourquoi"))
            comment = as_text(level_obj.get("comment"))
            description = f"## Pourquoi ?\n{pourquoi}\n\n## Comment ?\n{comment}"
            ctx.ref_ctrl_rows.append(
                {
                    "ref_id": f"{mesure_id}_{level_key}",
                    "name": title,
                    "category": "",
                    "csf_function": "",
                    "description": description,
                    "annotation": "",
                }
            )


def color_for_choice(choice: str) -> str:
    normalized = choice.strip().lower()
    if normalized == "non applicable":
        return AnswerColor.NOT_APPLICABLE
    if normalized == "je ne sais pas":
        return AnswerColor.DONT_KNOW
    if normalized == "non":
        return AnswerColor.NO
    if normalized.startswith("oui"):
        return AnswerColor.YES
    return AnswerColor.EMPTY


def build_color_cell_from_choices(choices: list[str]) -> str:
    if not choices:
        return ""
    # Step 1: initialize one "/" per line from question_choices.
    color_lines = [AnswerColor.EMPTY for _ in choices]
    # Step 2: replace matching lines with color hex codes.
    for idx, choice in enumerate(choices):
        color_lines[idx] = color_for_choice(choice)
    if all(color == AnswerColor.EMPTY for color in color_lines):
        return ""
    return JoinSeparator.DEFAULT_NEWLINE.join(color_lines)


def answer_rows_to_output(answer_rows: dict[str, AnswerRow]) -> list[dict[str, str]]:
    out: list[dict[str, str]] = []
    for qid, row in answer_rows.items():
        select = (
            JoinSeparator.DEFAULT_NEWLINE.join(row.select_groups)
            if row.select_groups
            else ""
        )
        color = build_color_cell_from_choices(row.response_labels)
        out.append(
            {
                "id": qid,
                "question_type": row.question_type,
                "question_choices": JoinSeparator.DEFAULT_NEWLINE.join(
                    row.response_labels
                ),
                "color": color,
                "description": "",
                "select_implementation_groups": select,
            }
        )
    return out


def write_sheet(
    wb: Workbook, title: str, headers: list[str], rows: list[dict[str, str]]
) -> None:
    ws = wb.create_sheet(title)
    ws.append(headers)
    for line in rows:
        values = [as_text(line.get(h, "")) for h in headers]
        ws.append(values)

    for row in ws.iter_rows(
        min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)
    ):
        for cell in row:
            if cell.value is None:
                cell.value = ""
            else:
                cell.value = as_text(cell.value)
            cell.number_format = "@"


def write_kv_sheet(wb: Workbook, title: str, rows: list[tuple[str, str]]) -> None:
    ws = wb.create_sheet(title)
    for key, value in rows:
        ws.append([as_text(key), as_text(value)])

    max_row = ws.max_row if ws.max_row > 0 else 1
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=2):
        for cell in row:
            if cell.value is None:
                cell.value = ""
            else:
                cell.value = as_text(cell.value)
            cell.number_format = "@"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build an XLSX file from questionnaire_repo.json and mesures_repo.json"
    )
    parser.add_argument(
        "--questionnaire",
        default="questionnaire_repo.json",
        help="Path to questionnaire JSON",
    )
    parser.add_argument(
        "--mesures", default="mesures_repo.json", help="Path to measures JSON"
    )
    parser.add_argument(
        "--output", default="mon_aide_cyber.xlsx", help="Output XLSX file"
    )
    return parser.parse_args()


def build_excel_from_json(
    questionnaire_path: Path,
    mesures_path: Path,
    output_path: Path,
) -> dict[str, int]:
    """Point d'entrée unique: construit l'Excel à partir des 2 JSON."""
    questionnaire = load_json(questionnaire_path)
    mesures = load_json(mesures_path)

    referentiel = questionnaire.get("referentiel")
    if not isinstance(referentiel, dict):
        raise ValueError(
            "Le fichier questionnaire ne contient pas d'objet 'referentiel' valide"
        )

    ctx = Context()
    process_referentiel(ctx, referentiel)
    build_ref_controls_rows(ctx, mesures)
    apply_conditions_perimetre_second_pass(ctx, referentiel)

    wb = Workbook()
    wb.remove(wb.active)

    write_kv_sheet(wb, "library_meta", LIBRARY_META_ROWS)

    write_kv_sheet(wb, "fwk_meta", FWK_META_ROWS)
    write_sheet(wb, "fwk_content", FRAMEWORK_HEADERS, ctx.framework_rows)

    write_kv_sheet(wb, "answ_meta", ANSW_META_ROWS)
    write_sheet(
        wb, "answ_content", ANSW_HEADERS, answer_rows_to_output(ctx.answer_rows)
    )

    write_kv_sheet(wb, "imp_grp_meta", IMP_GRP_META_ROWS)
    write_sheet(wb, "imp_grp_content", IMP_GRP_HEADERS, ctx.imp_grp_rows)

    write_kv_sheet(wb, "ref_ctrl_meta", REF_CTRL_META_ROWS)
    write_sheet(wb, "ref_ctrl_content", REF_CTRL_HEADERS, ctx.ref_ctrl_rows)

    write_kv_sheet(wb, "urn_pref_meta", URN_PREF_META_ROWS)
    write_sheet(wb, "urn_pref_content", URN_PREF_HEADERS, URN_PREF_ROWS)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return {
        "fwk_content": len(ctx.framework_rows),
        "imp_grp_content": len(ctx.imp_grp_rows),
        "answ_content": len(ctx.answer_rows),
        "ref_ctrl_content": len(ctx.ref_ctrl_rows),
    }


def main() -> None:
    args = parse_args()

    q_path = Path(args.questionnaire)
    m_path = Path(args.mesures)
    out_path = Path(args.output)

    stats = build_excel_from_json(
        questionnaire_path=q_path,
        mesures_path=m_path,
        output_path=out_path,
    )

    print(f"Excel generated: {out_path}")
    print(f"- fwk_content: {stats['fwk_content']} row(s)")
    print(f"- imp_grp_content: {stats['imp_grp_content']} row(s)")
    print(f"- answ_content: {stats['answ_content']} row(s)")
    print(f"- ref_ctrl_content: {stats['ref_ctrl_content']} row(s)")


if __name__ == "__main__":
    main()
