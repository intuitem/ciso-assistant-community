"""
Script to parse Microsoft Cloud Security Benchmark v1 Excel file
and generate an Excel file for CISO Assistant framework conversion tool.

Output columns: assessable, depth, ref_id, name, description, annotation
"""

import sys
import openpyxl
import argparse
from openpyxl.utils.exceptions import InvalidFileException

parser = argparse.ArgumentParser(
    prog="parse_mcsb",
    description="Parse Microsoft Cloud Security Benchmark Excel file to Excel",
)
parser.add_argument(
    "filename",
    nargs="?",
    default="Microsoft_cloud_security_benchmark_v1.xlsx",
    help="name of Microsoft Cloud Security Benchmark Excel file",
)
parser.add_argument(
    "-o",
    "--output",
    default="mcsb_v1.xlsx",
    help="output Excel file name",
)

args = parser.parse_args()
input_file_name = args.filename
output_file_name = args.output

print(f'⌛ Parsing "{input_file_name}"...')

# Load the Excel file
try:
    workbook = openpyxl.load_workbook(input_file_name)
    print(f'✅ Excel file loaded successfully: "{input_file_name}"')
except FileNotFoundError:
    print(f'❌ [ERROR] File not found: "{input_file_name}"')
    sys.exit(1)
except PermissionError:
    print(f'❌ [ERROR] Permission denied while accessing "{input_file_name}"')
    sys.exit(1)
except InvalidFileException:
    print(f'❌ [ERROR] The file is not a valid Excel file: "{input_file_name}"')
    sys.exit(1)
except Exception as e:
    print(f"❌ [ERROR] Unexpected error while loading Excel file: {e}")
    sys.exit(1)

output_rows = []

# Process each sheet
for sheet in workbook:
    sheet_title = sheet.title

    # Skip the Readme tab
    if sheet_title.lower() == "readme":
        print(f'⏩ Skipping tab: "{sheet_title}"')
        continue

    print(f'⌛ Processing tab: "{sheet_title}"...')

    # Get header row to find column indices
    header_row = None
    header_indices = {}

    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=1), start=1):
        header_row = [cell.value for cell in row]
        for idx, col_name in enumerate(header_row):
            if col_name:
                header_indices[col_name.strip()] = idx
        break

    if not header_row:
        print(f'⚠️  No header found in tab "{sheet_title}", skipping...')
        continue

    # Find required columns (case-insensitive search)
    def find_column(names):
        """Find column index by trying multiple possible names."""
        for name in names:
            for header, idx in header_indices.items():
                if name.lower() in header.lower():
                    return idx
        return None

    id_col = find_column(["ID"])
    control_domain_col = find_column(["Control Domain"])
    recommendation_col = find_column(["Recommendation"])
    security_principle_col = find_column(["Security Principle"])
    azure_guidance_col = find_column(["Azure Guidance"])
    aws_guidance_col = find_column(["AWS Guidance"])

    # Debug: print found columns
    print(
        f"   Found columns: ID={id_col}, Control Domain={control_domain_col}, "
        f"Recommendation={recommendation_col}, Security Principle={security_principle_col}, "
        f"Azure Guidance={azure_guidance_col}, AWS Guidance={aws_guidance_col}"
    )

    if id_col is None:
        print(f'⚠️  ID column not found in tab "{sheet_title}", skipping...')
        continue

    # Track control domains we've already added as parent sections
    seen_control_domains = set()

    # Process data rows
    for row in sheet.iter_rows(min_row=2):
        row_values = [cell.value for cell in row]

        # Get cell values
        ref_id = (
            row_values[id_col]
            if id_col is not None and id_col < len(row_values)
            else None
        )
        control_domain = (
            row_values[control_domain_col]
            if control_domain_col is not None and control_domain_col < len(row_values)
            else None
        )
        recommendation = (
            row_values[recommendation_col]
            if recommendation_col is not None and recommendation_col < len(row_values)
            else None
        )
        security_principle = (
            row_values[security_principle_col]
            if security_principle_col is not None
            and security_principle_col < len(row_values)
            else None
        )
        azure_guidance = (
            row_values[azure_guidance_col]
            if azure_guidance_col is not None and azure_guidance_col < len(row_values)
            else None
        )
        aws_guidance = (
            row_values[aws_guidance_col]
            if aws_guidance_col is not None and aws_guidance_col < len(row_values)
            else None
        )

        # Skip empty rows
        if not ref_id:
            continue

        ref_id = str(ref_id).strip()

        # Add control domain as parent section if not seen yet
        if control_domain and control_domain not in seen_control_domains:
            # Extract domain ID from ref_id (e.g., "NS-1" -> "NS")
            domain_id = ref_id.split("-")[0] if "-" in ref_id else ref_id
            output_rows.append(
                {
                    "assessable": "",
                    "depth": 1,
                    "ref_id": domain_id,
                    "name": str(control_domain).strip(),
                    "description": "",
                    "annotation": "",
                }
            )
            seen_control_domains.add(control_domain)

        # Build annotation from Azure and AWS guidance
        annotation_parts = []
        if azure_guidance:
            annotation_parts.append(
                f"### Azure Guidance\n\n{str(azure_guidance).strip()}"
            )
        if aws_guidance:
            annotation_parts.append(f"### AWS Guidance\n\n{str(aws_guidance).strip()}")
        annotation = "\n\n".join(annotation_parts)

        # Add the assessable item
        output_rows.append(
            {
                "assessable": "x",
                "depth": 2,
                "ref_id": ref_id,
                "name": str(recommendation).strip() if recommendation else "",
                "description": str(security_principle).strip()
                if security_principle
                else "",
                "annotation": annotation,
            }
        )

print(f'\n⌛ Writing {len(output_rows)} rows to "{output_file_name}"...')

# Write Excel output
try:
    wb_output = openpyxl.Workbook()
    ws = wb_output.active
    ws.title = "controls"

    # Write header row
    ws.append(["assessable", "depth", "ref_id", "name", "description", "annotation"])

    # Write data rows
    for row in output_rows:
        ws.append(
            [
                row["assessable"],
                row["depth"],
                row["ref_id"],
                row["name"],
                row["description"],
                row["annotation"],
            ]
        )

    wb_output.save(output_file_name)
    print(f'✅ Excel file saved successfully: "{output_file_name}"')
except PermissionError:
    print(
        f'❌ [ERROR] Permission denied. The file may be open or locked: "{output_file_name}"'
    )
    sys.exit(1)
except Exception as e:
    print(f'❌ [ERROR] Unexpected error while saving Excel file: "{e}"')
    sys.exit(1)

print(f"\n📊 Summary:")
print(f"   Total rows: {len(output_rows)}")
print(
    f"   Parent sections (depth 1): {len([r for r in output_rows if r['depth'] == 1])}"
)
print(
    f"   Assessable items (depth 2): {len([r for r in output_rows if r['depth'] == 2])}"
)
