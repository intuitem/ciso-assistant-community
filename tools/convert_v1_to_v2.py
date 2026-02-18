# This script converts a CISO Assistant Excel library from v1 to v2 format
# v1 uses a 'library_content' tab (key/value format), v2 uses _meta and _content sheets

import argparse
from pathlib import Path
from openpyxl import load_workbook, Workbook
import re
import sys
from copy import copy
from openpyxl.cell.cell import Cell


def build_converted_workbook(input_path: str) -> Workbook:
    wb = load_workbook(input_path, data_only=False)

    if "library_content" not in wb.sheetnames:
        raise ValueError("Missing 'library_content' sheet.")

    library_meta = [("type", "library")]
    tab_entries = []
    object_metadata = {}
    declared_tabs = {}
    urn_prefixes = {}
    answers_logical_name = None
    ig_logical_name = None
    scores_logical_name = None

    known_object_types = {
        "framework",
        "threats",
        "reference_controls",
        "scores",
        "implementation_groups",
        "risk_matrix",
        "mapping",
        "answers",
    }

    # --- Parse library_content ---
    sheet = wb["library_content"]
    for row in sheet.iter_rows(values_only=True):
        if not any(row):
            continue
        key = str(row[0]).strip().lower() if row[0] else None
        val1 = str(row[1]).strip() if len(row) > 1 and row[1] is not None else None
        val2 = str(row[2]).strip() if len(row) > 2 and row[2] is not None else None
        val3 = str(row[3]).strip() if len(row) > 3 and row[3] is not None else None

        if not key:
            continue

        if key.startswith("library_"):
            library_meta.append((key.replace("library_", ""), val1))
        if key == "framework_urn" and val1:
            # Automatically derive base_urn for requirements
            base_urn = re.sub("framework", "req_node", val1)
            object_metadata.setdefault("framework", {})["base_urn"] = base_urn
        if key == "tab":
            if val2 == "requirements":
                normalized_type = "framework"
            elif val2 == "mappings":
                normalized_type = "requirement_mapping_set"
            else:
                normalized_type = val2

            tab_entries.append((val1, normalized_type, val3))
            declared_tabs[normalized_type] = val1

            if normalized_type == "answers":
                answers_logical_name = val1
            if normalized_type == "scores":
                scores_logical_name = val1
            if normalized_type == "implementation_groups":
                ig_logical_name = val1

        if key in ["reference_control_base_urn", "threat_base_urn"] and val1:
            object_type = (
                "reference_control" if "reference_control" in key else "threat"
            )
            object_metadata.setdefault(object_type, {})["base_urn"] = val1
            if val2:
                urn_prefixes[val2] = val1

        for obj_type in known_object_types:
            if key.startswith(f"{obj_type}_"):
                field = key[len(obj_type) + 1 :]
                normalized_type = (
                    "requirement_mapping_set" if obj_type == "mapping" else obj_type
                )
                object_metadata.setdefault(normalized_type, {})[field] = val1

    sheets_out = {"library_meta": library_meta}

    used_tabs = set()

    # --- Process each declared tab ---
    for tab_name, obj_type, base_urn in tab_entries:
        tab_name = tab_name.strip()
        obj_type = obj_type.strip()
        used_tabs.add(tab_name)

        meta_rows = [("type", obj_type)]

        clean_type = obj_type.rstrip("s")
        for k, v in object_metadata.get(clean_type, {}).items():
            meta_rows.append((k, v))

        if obj_type == "framework":
            if scores_logical_name:
                meta_rows.append(("scores_definition", scores_logical_name))
            if ig_logical_name:
                meta_rows.append(("implementation_groups_definition", ig_logical_name))
            if answers_logical_name:
                meta_rows.append(("answers_definition", answers_logical_name))
        elif obj_type == "answers":
            meta_rows.append(("name", tab_name))
        elif obj_type == "implementation_groups":
            meta_rows.append(("name", tab_name))
        elif obj_type == "scores":
            meta_rows.append(("name", tab_name))
        # Read content
        content_rows = []
        if tab_name in wb.sheetnames:
            ws = wb[tab_name]
            for row in ws.iter_rows():
                content_rows.append(list(row))

        sheets_out[f"{tab_name}_meta"] = meta_rows
        sheets_out[f"{tab_name}_content"] = content_rows

    if urn_prefixes:
        sheets_out["urn_prefix_meta"] = [("type", "urn_prefix")]
        sheets_out["urn_prefix_content"] = [("prefix_id", "prefix_value")] + [
            (k, v) for k, v in urn_prefixes.items()
        ]

    for sheet_name in wb.sheetnames:
        if (
            sheet_name not in {"library_content"}
            and sheet_name not in sheets_out
            and sheet_name not in used_tabs
        ):
            ws = wb[sheet_name]
            raw = [[cell.value for cell in row] for row in ws.iter_rows()]
            if raw:
                sheets_out[sheet_name] = raw

    wb_out = Workbook()
    del wb_out["Sheet"]

    for sheet_name, rows in sheets_out.items():
        ws_out = wb_out.create_sheet(title=sheet_name[:31])
        for r_idx, row in enumerate(rows, 1):
            for c_idx, cell in enumerate(row, 1):
                if isinstance(cell, Cell):
                    new_cell = ws_out.cell(row=r_idx, column=c_idx, value=cell.value)
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.fill = copy(cell.fill)
                        new_cell.border = copy(cell.border)
                        new_cell.number_format = cell.number_format
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)
                else:
                    ws_out.cell(row=r_idx, column=c_idx, value=cell)

    return wb_out


def convert_v1_to_v2(
    input_path: Path, output_path: Path, old_output_dir: Path | None = None
):
    backup_dir = old_output_dir if old_output_dir else input_path.parent
    backup_path = backup_dir / f"{input_path.stem}_v1{input_path.suffix}.old"

    if backup_path.exists():
        raise FileExistsError(
            f"Backup file already exists, refusing to overwrite: {backup_path}"
        )

    wb_out = build_converted_workbook(str(input_path))
    output_existed_before = output_path.exists()

    try:
        input_path.rename(backup_path)
        wb_out.save(output_path)
        print(f"✅ Conversion complete: {output_path}")
    except Exception:
        if output_path.exists() and not output_existed_before:
            output_path.unlink()
        if backup_path.exists() and not input_path.exists():
            backup_path.rename(input_path)
        raise


def with_output_suffix(path: Path, output_suffix: str) -> Path:
    normalized = path if path.suffix.lower() == ".xlsx" else path.with_suffix(".xlsx")
    if not output_suffix:
        return normalized
    return normalized.with_name(
        f"{normalized.stem}{output_suffix}{normalized.suffix}"
    )


def main():
    parser = argparse.ArgumentParser(
        description="Convert Excel library v1 to v2 format."
    )
    parser.add_argument(
        "input_file", type=str, help="Path to a v1 Excel file (or directory in bulk mode)"
    )
    parser.add_argument(
        "--bulk",
        action="store_true",
        help="Enable bulk mode to process all .xlsx files in a directory.",
    )
    parser.add_argument(
        "--output",
        type=str,
        help="Custom output file name (only used in non-bulk mode).",
    )
    parser.add_argument(
        "--output-dir",
        type=str,
        help="Output directory for .xlsx files (only used with --bulk mode).",
    )
    parser.add_argument(
        "--old-output-dir",
        type=str,
        help="Output directory for renamed .old files (only used with --bulk mode).",
    )
    parser.add_argument(
        "--output-suffix",
        type=str,
        default="",
        help='Suffix added just before ".xlsx" for converted files (example: "_v2").',
    )
    args = parser.parse_args()

    input_path = Path(args.input_file)
    if not input_path.exists():
        raise FileNotFoundError(f"File not found: {input_path}")


    # --- BULK MODE ------------------------------------------------------------
    if args.bulk:
        if args.output:
            raise ValueError('The option "--output" cannot be used with "--bulk" mode.')

        if not input_path.is_dir():
            raise NotADirectoryError("Bulk mode requires a directory as input")

        if args.output_dir:
            output_dir = Path(args.output_dir)
            output_dir.mkdir(parents=True, exist_ok=True)
        else:
            output_dir = input_path

        if args.old_output_dir:
            old_output_dir = Path(args.old_output_dir)
            old_output_dir.mkdir(parents=True, exist_ok=True)
        else:
            old_output_dir = None

        xlsx_files = [
            f for f in input_path.glob("*.xlsx") if not f.name.startswith("~$")
        ]
        if not xlsx_files:
            raise FileNotFoundError(f"No .xlsx files found in directory: {input_path}")

        errors = []
        for i, file in enumerate(xlsx_files, 1):
            print(f'▶️  Processing file [{i}/{len(xlsx_files)}]: "{file}"')
            try:
                output_path = with_output_suffix(output_dir / file.name, args.output_suffix)
                convert_v1_to_v2(file, output_path, old_output_dir=old_output_dir)
            except Exception as e:
                print(f'❌ Failed to process "{file}": {e}', file=sys.stderr)
                errors.append(file.name)

        print("\n📋 Bulk mode completed!")
        
        if errors:
            print(f"❌ The following file{'s' if len(errors) > 1 else ''} [{len(errors)}/{len(xlsx_files)}] failed to process:", file=sys.stderr)
            for f in errors:
                print(f"- {f}", file=sys.stderr)
            sys.exit(1)
        else:
            print("✅ All files processed successfully!")
            sys.exit(0)

    # --- SINGLE FILE MODE -----------------------------------------------------
    else:
        if args.output_dir:
            raise ValueError(
                'The option "--output-dir" can only be used with "--bulk" mode.'
            )
        if args.old_output_dir:
            raise ValueError(
                'The option "--old-output-dir" can only be used with "--bulk" mode.'
            )

        if not input_path.is_file():
            raise FileNotFoundError(f"Input must be a file in non-bulk mode: {input_path}")

        if args.output:
            output_path = with_output_suffix(Path(args.output), args.output_suffix)
        else:
            output_path = with_output_suffix(input_path, args.output_suffix)

        convert_v1_to_v2(input_path, output_path)


if __name__ == "__main__":
    try:
        main()
    except Exception as err:
        print(f"❌ [ERROR] {err}")
        sys.exit(1)
