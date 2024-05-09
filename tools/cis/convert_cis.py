"""
simple script to transform the official CIS Excel file to another Excel file for CISO assistant framework conversion tool
"""

import openpyxl
import sys
import re
import argparse
from openpyxl.styles import numbers

parser = argparse.ArgumentParser(
    prog="convert_cis",
    description="convert CIS controls offical Excel file to CISO Assistant Excel file",
)
parser.add_argument("filename", help="name of CIS controls Excel file")
parser.add_argument("packager", help="name of packager entity")

args = parser.parse_args()
input_file_name = args.filename
packager = args.packager
output_file_name = "cis-controls-v8.xlsx"

print("parsing", input_file_name)

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook(input_file_name)
output_table = []

for tab in dataframe:
    print("parsing tab", tab.title)
    title = tab.title
    if title == "License for Use":
        library_copyright = tab["B11"].value + "\n" + tab["B13"].value
    elif title == "Controls V8":
        for row in tab:
            (control, safeguard, asset_type, sf, title, description, ig1, ig2, ig3) = (
                r.value for r in row
            )
            control = str(control).strip()
            if re.match(r"\d+", control):
                if not safeguard:
                    safeguard_index = 0
                    output_table.append(("", 1, control, title, description))
                else:
                    safeguard_index += 1
                    safeguard = f"{control},{safeguard_index}"
                    implementation_groups = (
                        "IG1,IG2,IG3" if ig1 else "IG2,IG3" if ig2 else "IG3"
                    )
                    output_table.append(
                        ("x", 2, safeguard, title, description, implementation_groups)
                    )
    else:
        print(f"Ignored tab: {title}")


print("generating", output_file_name)
wb_output = openpyxl.Workbook()
ws = wb_output.active
ws.title = "library_content"
ws.append(["library_urn", f"urn:{packager.lower()}:risk:library:cis-controls-v8"])
ws.append(["library_version", "1"])
ws.append(["library_locale", "en"])
ws.append(["library_ref_id", "CIS-Controls-v8"])
ws.append(["library_name", "CIS Controls v8"])
ws.append(["library_description", "CIS Controls v8"])
ws.append(["library_copyright", library_copyright])
ws.append(["library_provider", "CIS"])
ws.append(["library_packager", packager])
ws.append(["framework_urn", f"urn:{packager.lower()}:risk:framework:cis-controls-v8"])
ws.append(["framework_ref_id", "CIS-Controls-v8"])
ws.append(["framework_name", "CIS Controls v8"])
ws.append(["framework_description", "CIS Controls v8"])
ws.append(["tab", "controls", "requirements"])
ws.append(["tab", "implementation_groups", "implementation_groups"])

ws1 = wb_output.create_sheet("controls")
ws1.append(
    ["assessable", "depth", "ref_id", "name", "description", "implementation_groups"]
)
for row in output_table:
    ws1.append(row)

ws2 = wb_output.create_sheet("implementation_groups")
ws2.append(["ref_id", "name", "description"])
ws2.append(
    [
        "IG1",
        "Essential Cyber Hygiene",
        "Minimum standard of information security for all enterprises.",
    ]
)
ws2.append(
    [
        "IG2",
        "",
        "For enterprises managing IT infrastructure of multiple departments with differing risk profiles.",
    ]
)
ws2.append(["IG3", "", "To secure sensitive and confidential data."])

print("generate ", output_file_name)
wb_output.save(output_file_name)
