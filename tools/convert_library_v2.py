"""
Convert Library v2 - Build a CISO Assistant YAML library from a v2 Excel file

This script processes an Excel file in v2 format (with *_meta and *_content tabs),
extracts all declared objects, and generates a fully structured YAML library.

Usage:
    python convert_library_v2.py path/to/library.xlsx [--compat MODE] [--output out.yaml] [--verbose]
    python convert_library_v2.py path/to/folder --bulk [--compat MODE] [--output-dir out_folder] [--verbose]

Arguments:
    --compat       Specify compatibility mode number.
                   Refer to the COMPATIBILITY_MODES dictionary in the script for available modes and their descriptions.
    --verbose      Enable verbose output. Verbose messages start with a 💬 (speech bubble) emoji.
    --output       Custom output file name (only for single file mode). Adds '.yaml' if missing.
    --bulk         Enable bulk mode to process all .xlsx files in a directory.
    --output-dir   Destination directory for YAML files (only valid with --bulk).

Note: the "urn_id" column can be defined to force an urn suffix. This can be useful to fix "ref_id" errors
"""

import sys
import re
import yaml
import datetime
import argparse
import unicodedata
import openpyxl
from pathlib import Path
from collections import Counter

SCRIPT_VERSION = "2.1"

# --- Compatibility modes definition ------------------------------------------
# NOTE: No compatibility mode includes another (unless otherwise stated)
# NOTE: So far, no compatibility mode has an impact on the mapping creation process.

COMPATIBILITY_MODES = {
    0: f"[v{SCRIPT_VERSION}] (DEFAULT) Don't use any Compatibility Mode",
    1: "[< v2] Use legacy URN fallback logic (for requirements without ref_id)",
    2: "[v2] Don't clean the URNs before saving it into the YAML file (Only spaces ' ' are replaced with hyphen '-' and the URN is lower-cased)",
    # Future modes can be added here with an integer key and description
}

# --- Translation helpers ------------------------------------------------------


def extract_translations_from_row(header, row):
    translations = {}
    for i, col_name in enumerate(header):
        match = re.match(r"(\w+)\[(\w+)\]", col_name)
        if match and i < len(row):
            base_key, lang = match.groups()
            value = row[i].value
            if value:
                translations.setdefault(lang, {})[base_key] = str(value).strip()
    return translations


def extract_translations_from_metadata(meta_dict, prefix):
    translations = {}
    pattern = re.compile(r"(\w+)\[(\w+)\]")
    for key, value in meta_dict.items():
        match = pattern.match(key)
        if match and value:
            base_key, lang = match.groups()
            translations.setdefault(lang, {})[base_key] = str(value).strip()
    return translations


# --- Sheet parsing ------------------------------------------------------------


def parse_key_value_sheet(sheet):
    result = {}
    for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
        key = str(row[0]).strip().lower() if row[0] else None
        val = str(row[1]).strip() if row[1] is not None else None
        if key:
            result[key] = val
    return result


# --- URN cleaning utility -----------------------------------------------------
def clean_urn_suffix(value: str, compat_mode: int = 0):
    """
    Cleans a URN suffix by:
        - Removing accents using Unicode normalization,
        - Converting to lowercase and replacing spaces with hyphens,
        - Replacing any disallowed characters with an underscore.

    Compatibility modes :
        - 0 : Default behavior
        - 1,2 : Replace only spaces " " with hyphens "-" sets text in lower case.
    """
    # [+] Compat check
    if 1 <= compat_mode <= 2:
        value = value.lower().replace(" ", "-")
    else:  # Default behavior
        # Normalize to separate characters from their accents (e.g., é → e + ́)
        value = (
            unicodedata.normalize("NFKD", value)
            .encode("ascii", "ignore")
            .decode("ascii")
        )
        value = value.lower().replace(" ", "-")
        # Keep only allowed characters for a URN suffix
        value = re.sub(r"[^a-z0-9_\-\.\[\]\(\):]", "_", value)

    return value


# --- urn expansion ------------------------------------------------------------


def expand_urns_from_prefixed_list(
    field_value: str, prefix_to_urn: str, compat_mode: int = 0, verbose: bool = False
):
    """
    Convert a prefixed list like 'abc:xyz, def:uvw' into a list of full URNs,
    using a prefix mapping. Fully qualified URNs (starting with 'urn:') are left untouched.

    Args:
        field_value (str): Raw string field value from the sheet (e.g., 'abc:id1, urn:...').
        prefix_to_urn (dict): Mapping from prefix (e.g., 'abc') to full base URN.
        verbose (bool): If True, prints a warning when a suffix is modified.

    Returns:
        list[str]: Fully qualified URNs
    """
    result = []
    elements = re.split(r"[\s,]+", str(field_value))
    for element in elements:
        element = element.strip()
        if not element:
            continue

        if element.startswith("urn:"):
            result.append(element)
            continue

        parts = element.split(":")
        if len(parts) >= 2:
            prefix = parts[0]
            raw_suffix = ":".join(parts[1:]).strip()
            # [+] Compat check
            cleaned_suffix = clean_urn_suffix(raw_suffix, compat_mode=compat_mode)
            base = prefix_to_urn.get(prefix)

            if base:
                # Show a warning if the suffix had to be cleaned/renamed
                if raw_suffix != cleaned_suffix and verbose:
                    print(
                        f"💬 ⚠️  [WARNING] (expand_urns_from_prefixed_list) URN suffix renamed for '{prefix}:{raw_suffix}' → '{base}:{cleaned_suffix}'"
                    )
                result.append(f"{base}:{cleaned_suffix}")
            else:
                print(
                    f"⚠️  [WARNING] (expand_urns_from_prefixed_list) Unknown prefix '{prefix}' for element '{element}' — skipped"
                )
        else:
            print(
                f"⚠️  [WARNING] (expand_urns_from_prefixed_list) Malformed element '{element}' — skipped"
            )
    return result


# --- question management ------------------------------------------------------------


def inject_questions_into_node(node, raw_question_str, raw_answer_str, answers_dict):
    """
    Injects parsed questions and their metadata into a requirement node.
    Ensures that question type is valid and handles multiple questions/answers.

    :param node: the requirement node (dict)
    :param raw_question_str: the string from the "questions" column
    :param raw_answer_str: the string from the "answer" column
    :param answers_dict: dictionary of all available answers (from answer sheet)
    """
    if not raw_question_str:
        return

    allowed_types = {"unique_choice", "multiple_choice", "text", "date"}

    question_lines = [q.strip() for q in str(raw_question_str).split("\n") if q.strip()]
    answer_ids = (
        [a.strip() for a in str(raw_answer_str).split("\n") if raw_answer_str]
        if raw_answer_str
        else []
    )

    if len(answer_ids) != 1 and len(answer_ids) != len(question_lines):
        raise ValueError(
            f"Mismatch between questions and answers for node {node.get('urn')}"
        )

    question_block = {}

    for idx, question_text in enumerate(question_lines):
        answer_id = answer_ids[0] if len(answer_ids) == 1 else answer_ids[idx]
        answer_meta = answers_dict.get(answer_id)
        if not answer_meta:
            raise ValueError(
                f"Unknown answer ID: {answer_id} for node {node.get('urn')}"
            )
        qtype = answer_meta.get("type")
        if qtype not in allowed_types:
            raise ValueError(
                f"Invalid question type '{qtype}' for answer ID: {answer_id}"
            )
        q_urn = f"{node['urn']}:question:{idx + 1}"
        question_entry = {
            "type": qtype,
        }
        if qtype in {"unique_choice", "multiple_choice"}:
            choices = []
            for j, choice in enumerate(answer_meta["choices"]):
                choice_urn = f"{q_urn}:choice:{j + 1}"
                choices.append({"urn": choice_urn, "value": choice["value"]})
            question_entry["choices"] = choices
        question_entry["text"] = question_text
        question_block[q_urn] = question_entry
    if question_block:
        node["questions"] = question_block


# --- risk matrix management ------------------------------------------------------------

# https://gist.github.com/Mike-Honey/b36e651e9a7f1d2e1d60ce1c63b9b633
from colorsys import rgb_to_hls, hls_to_rgb

RGBMAX = 0xFF  # Corresponds to 255
HLSMAX = 240  # MS excel's tint function expects that HLS is base 240. see:
# https://social.msdn.microsoft.com/Forums/en-US/e9d8c136-6d62-4098-9b1b-dac786149f43/excel-color-tint-algorithm-incorrect?forum=os_binaryfile#d3c2ac95-52e0-476b-86f1-e2a697f24969


def rgb_to_ms_hls(red, green=None, blue=None):
    """Converts rgb values in range (0,1) or a hex string of the form '[#aa]rrggbb' to HLSMAX based HLS, (alpha values are ignored)"""
    if green is None:
        if isinstance(red, str):
            if len(red) > 6:
                red = red[-6:]  # Ignore preceding '#' and alpha values
            blue = int(red[4:], 16) / RGBMAX
            green = int(red[2:4], 16) / RGBMAX
            red = int(red[0:2], 16) / RGBMAX
        else:
            red, green, blue = red
    h, l, s = rgb_to_hls(red, green, blue)
    return (int(round(h * HLSMAX)), int(round(l * HLSMAX)), int(round(s * HLSMAX)))


def ms_hls_to_rgb(hue, lightness=None, saturation=None):
    """Converts HLSMAX based HLS values to rgb values in the range (0,1)"""
    if lightness is None:
        hue, lightness, saturation = hue
    return hls_to_rgb(hue / HLSMAX, lightness / HLSMAX, saturation / HLSMAX)


def rgb_to_hex(red, green=None, blue=None):
    """Converts (0,1) based RGB values to a hex string 'rrggbb'"""
    if green is None:
        red, green, blue = red
    return (
        "%02x%02x%02x"
        % (
            int(round(red * RGBMAX)),
            int(round(green * RGBMAX)),
            int(round(blue * RGBMAX)),
        )
    ).upper()


def get_theme_colors(wb):
    """Gets theme colors from the workbook"""
    # see: https://groups.google.com/forum/#!topic/openpyxl-users/I0k3TfqNLrc
    from openpyxl.xml.functions import QName, fromstring

    xlmns = "http://schemas.openxmlformats.org/drawingml/2006/main"
    root = fromstring(wb.loaded_theme)
    themeEl = root.find(QName(xlmns, "themeElements").text)
    colorSchemes = themeEl.findall(QName(xlmns, "clrScheme").text)
    firstColorScheme = colorSchemes[0]

    colors = []

    for c in [
        "lt1",
        "dk1",
        "lt2",
        "dk2",
        "accent1",
        "accent2",
        "accent3",
        "accent4",
        "accent5",
        "accent6",
    ]:
        accent = firstColorScheme.find(QName(xlmns, c).text)
        for i in list(accent):  # walk all child nodes, rather than assuming [0]
            if "window" in i.attrib["val"]:
                colors.append(i.attrib["lastClr"])
            else:
                colors.append(i.attrib["val"])

    return colors


def tint_luminance(tint, lum):
    """Tints a HLSMAX based luminance"""
    # See: http://ciintelligence.blogspot.co.uk/2012/02/converting-excel-theme-color-and-tint.html
    if tint < 0:
        return int(round(lum * (1.0 + tint)))
    else:
        return int(round(lum * (1.0 - tint) + (HLSMAX - HLSMAX * (1.0 - tint))))


def theme_and_tint_to_rgb(wb, theme, tint):
    """Given a workbook, a theme number and a tint return a hex based rgb"""
    rgb = get_theme_colors(wb)[theme]
    h, l, s = rgb_to_ms_hls(rgb)
    return rgb_to_hex(ms_hls_to_rgb(h, tint_luminance(tint, l), s))


def get_color(wb, cell):
    """get cell color; None for no fill"""
    if not cell.fill.patternType:
        return None
    if isinstance(cell.fill.fgColor.rgb, str):
        return "#" + cell.fill.fgColor.rgb[2:]
    theme = cell.fill.start_color.theme
    tint = cell.fill.start_color.tint
    color = theme_and_tint_to_rgb(wb, theme, tint)
    return "#" + color


def parse_risk_matrix(meta, content_ws, wb):
    """Parse a risk_matrix block into a structured dict."""
    rows = list(content_ws.iter_rows())
    if not rows:
        return None
    header = {}
    grid_count = 0
    for i, cell in enumerate(rows[0]):
        col = str(cell.value).strip().lower() if cell.value else ""
        if col == "grid":
            col = f"grid{grid_count}"
            grid_count += 1
        header[col] = i

    assert all(
        k in header
        for k in ["type", "id", "color", "abbreviation", "name", "description", "grid0"]
    )
    index = {k: i for i, k in enumerate(header)}
    size_grid = len([h for h in header if h.startswith("grid")])

    risk_matrix = {
        "urn": meta.get("urn"),
        "ref_id": meta.get("ref_id"),
        "name": meta.get("name"),
        "description": meta.get("description"),
        "probability": [],
        "impact": [],
        "risk": [],
        "grid": [],
    }

    translations = extract_translations_from_metadata(meta, "risk_matrix")
    if translations:
        risk_matrix["translations"] = translations

    grid = {}
    grid_color = {}

    for row in rows[1:]:
        if not any(c.value for c in row):
            continue

        ctype = str(row[index["type"]].value).strip().lower()
        assert ctype in ("probability", "impact", "risk")

        rid = int(row[index["id"]].value)
        abbrev = str(row[index["abbreviation"]].value or "").strip()
        name = str(row[index["name"]].value or "").strip()
        desc = str(row[index["description"]].value or "").strip()
        color_cell = row[index["color"]]

        obj_data = {
            "id": rid,
            "abbreviation": abbrev,
            "name": name,
            "description": desc,
        }

        translations = extract_translations_from_row(header, row)
        if translations:
            obj_data["translations"] = translations

        if color := get_color(wb, color_cell):
            obj_data["hexcolor"] = color

        risk_matrix[ctype].append(obj_data)

        if ctype == "probability":
            grid[rid] = []
            grid_color[rid] = []
            for i in range(size_grid):
                cell = row[index[f"grid{i}"]]
                grid[rid].append(int(cell.value))
                grid_color[rid].append(get_color(wb, cell))

    # Build grid from sorted probability IDs
    sorted_ids = sorted(grid.keys())
    risk_matrix["grid"] = [grid[rid] for rid in sorted_ids]

    # Ensure order
    for k in ("probability", "impact", "risk"):
        risk_matrix[k].sort(key=lambda x: x["id"])

    return risk_matrix


# --- Mapping logic ---------------------------------------------------------------


def revert_relationship(relation: str):
    if relation == "subset":
        return "superset"
    elif relation == "superset":
        return "subset"
    else:
        return relation


# --- Main logic ---------------------------------------------------------------


def create_library(
    input_file: str, output_file: str, compat_mode: int = 0, verbose: bool = False
):
    wb = openpyxl.load_workbook(input_file)
    sheets = wb.sheetnames
    object_blocks = {}

    # Step 1: Load library_meta
    if "library_meta" not in sheets:
        raise ValueError("Missing 'library_meta' sheet.")

    library_meta_sheet = wb["library_meta"]
    library_meta = parse_key_value_sheet(library_meta_sheet)

    if library_meta.get("type") != "library":
        raise ValueError("'library_meta' must declare type = library")

    # Step 2: Format publication_date
    lib_date = library_meta.get("publication_date", datetime.datetime.now())
    if isinstance(lib_date, str):
        try:
            lib_date = datetime.datetime.fromisoformat(lib_date).date()
        except ValueError:
            lib_date = datetime.datetime.now().date()
    elif isinstance(lib_date, datetime.datetime):
        lib_date = lib_date.date()

    # Step 3: Build library with fixed key order
    library_urn = None

    # [+] Compat check
    if 1 <= compat_mode <= 2:
        library_urn = library_meta.get("urn")
    else:  # Default behavior
        # Clean URN
        library_urn_raw = library_meta.get("urn")
        library_urn_clean = clean_urn_suffix(library_urn_raw, compat_mode=0)
        if library_urn_raw != library_urn_clean:
            print(
                f"⚠️  [WARNING] (create_library) Cleaned library URN '{library_urn_raw}' → '{library_urn_clean}'"
            )

        library_urn = library_urn_clean

    convert_library_version = f"v2 ; Compat Mode: [{compat_mode}] {'{' + COMPATIBILITY_MODES[compat_mode] + '}'}"

    library = {
        "convert_library_version": convert_library_version,
        "urn": library_urn,
        "locale": library_meta.get("locale"),
        "ref_id": library_meta.get("ref_id"),
        "name": library_meta.get("name"),
        "description": library_meta.get("description"),
        "copyright": library_meta.get("copyright"),
        "version": int(library_meta["version"]),
        "publication_date": lib_date,
        "provider": library_meta.get("provider"),
        "packager": library_meta.get("packager"),
    }

    translations = extract_translations_from_metadata(library_meta, "library")
    if translations:
        library["translations"] = translations
    if "dependencies" in library_meta:
        dependencies = [
            x.strip()
            for x in re.split(r"[\s,]+", library_meta["dependencies"])
            if x.strip()
        ]
        if dependencies:
            library["dependencies"] = dependencies

    library["objects"] = {}
    prefix_to_urn = {}

    # Step 4: Detect all object blocks
    for sheet_name in sheets:
        if sheet_name.endswith("_meta") and sheet_name != "library_meta":
            prefix = sheet_name[:-5]
            content_name = f"{prefix}_content"
            if content_name not in sheets:
                print(f"⚠️  [WARNING] Missing content for {prefix}")
                continue

            meta_sheet = wb[sheet_name]
            meta_data = parse_key_value_sheet(meta_sheet)
            object_type = meta_data.get("type")
            object_name = meta_data.get("name", prefix)

            # Capture prefix mappings if this is a urn_prefix block
            if object_type == "urn_prefix" and f"{prefix}_content" in wb.sheetnames:
                ws_prefix = wb[f"{prefix}_content"]
                rows = list(ws_prefix.iter_rows(values_only=True))
                if rows:
                    header = [
                        str(cell).strip().lower() if cell else "" for cell in rows[0]
                    ]
                    idx_id = (
                        header.index("prefix_id") if "prefix_id" in header else None
                    )
                    idx_val = (
                        header.index("prefix_value")
                        if "prefix_value" in header
                        else None
                    )
                    if idx_id is not None and idx_val is not None:
                        for row in rows[1:]:
                            prefix_id = (
                                str(row[idx_id]).strip() if row[idx_id] else None
                            )
                            prefix_val = (
                                str(row[idx_val]).strip() if row[idx_val] else None
                            )
                            if prefix_id and prefix_val:
                                prefix_to_urn[prefix_id] = prefix_val

            object_blocks[object_name] = {
                "type": object_type,
                "meta": meta_data,
                "content_sheet": wb[content_name],
            }

    print(f"📦 Found {len(object_blocks)} objects.")

    # Step 5: Ordered object insertion (reference_controls before threats)
    priority_order = ["reference_controls", "threats"]

    sorted_object_names = sorted(
        object_blocks.keys(),
        key=lambda k: (
            priority_order.index(object_blocks[k]["type"])
            if object_blocks[k]["type"] in priority_order
            else len(priority_order)
        ),
    )

    for name in sorted_object_names:
        obj = object_blocks[name]
        obj_type = obj["type"]
        print(f"→ Handling {obj_type}: {name}")

        if obj_type == "reference_controls":
            controls = []
            base_urn = obj["meta"].get("base_urn")
            content_ws = obj["content_sheet"]
            rows = list(content_ws.iter_rows())
            if not rows:
                continue
            header = [
                str(cell.value).strip().lower() if cell.value else ""
                for cell in rows[0]
            ]

            for row in rows[1:]:
                if not any(cell.value for cell in row):
                    continue
                data = {
                    header[i]: row[i].value for i in range(len(header)) if i < len(row)
                }
                default_ref_id = None
                ref_id_for_urn = None

                # [+] Compat check
                if 1 <= compat_mode <= 2:
                    default_ref_id = str(data.get("ref_id", "")).strip()
                    ref_id_for_urn = default_ref_id.lower()
                else:  # Default behavior
                    ref_id_raw = str(data.get("ref_id", "")).strip()
                    ref_id_clean = clean_urn_suffix(ref_id_raw, compat_mode=0)
                    if verbose and ref_id_raw != ref_id_clean:
                        print(
                            f"💬 ⚠️  [WARNING] (reference_controls) Cleaned ref_id (for use in URN) '{ref_id_raw}' → '{ref_id_clean}'"
                        )

                    default_ref_id = ref_id_raw
                    ref_id_for_urn = ref_id_clean

                if not default_ref_id:
                    continue

                entry = {
                    "urn": f"{base_urn}:{ref_id_for_urn}",
                    "ref_id": default_ref_id,
                }

                if "name" in data and data["name"]:
                    entry["name"] = str(data["name"]).strip()
                if "category" in data and data["category"]:
                    entry["category"] = str(data["category"]).strip()
                if "description" in data and data["description"]:
                    entry["description"] = str(data["description"]).strip()
                if "csf_function" in data and data["csf_function"]:
                    entry["csf_function"] = str(data["csf_function"]).strip()

                translations = extract_translations_from_row(header, row)
                if translations:
                    entry["translations"] = translations

                controls.append(entry)

            library["objects"]["reference_controls"] = controls

        elif obj_type == "threats":
            threats = []
            base_urn = obj["meta"].get("base_urn")
            content_ws = obj["content_sheet"]
            rows = list(content_ws.iter_rows())
            if not rows:
                continue
            header = [
                str(cell.value).strip().lower() if cell.value else ""
                for cell in rows[0]
            ]

            for row in rows[1:]:
                if not any(cell.value for cell in row):
                    continue
                data = {
                    header[i]: row[i].value for i in range(len(header)) if i < len(row)
                }
                ref_id = str(data.get("ref_id", "")).strip()
                if not ref_id:
                    continue
                entry = {"urn": f"{base_urn}:{ref_id.lower()}", "ref_id": ref_id}
                if "name" in data and data["name"]:
                    entry["name"] = str(data["name"]).strip()
                if "description" in data and data["description"]:
                    entry["description"] = str(data["description"]).strip()

                translations = extract_translations_from_row(header, row)
                if translations:
                    entry["translations"] = translations

                threats.append(entry)

            library["objects"]["threats"] = threats

        elif obj_type == "framework":
            meta = obj["meta"]
            content_ws = obj["content_sheet"]
            base_urn = obj["meta"].get("base_urn")

            # --- Retrieve answers block if declared ---
            answers_dict = {}
            answers_block_name = meta.get("answers_definition")
            if answers_block_name:
                if answers_block_name not in object_blocks:
                    raise ValueError(f"Missing answers sheet: '{answers_block_name}'")

                answers_sheet = object_blocks[answers_block_name]["content_sheet"]
                rows = list(answers_sheet.iter_rows())
                if rows:
                    header = [
                        str(c.value).strip().lower() if c.value else "" for c in rows[0]
                    ]

                    for row in rows[1:]:
                        data = {
                            header[i]: row[i].value
                            for i in range(len(header))
                            if i < len(row)
                        }
                        answer_id = str(data.get("id", "")).strip()
                        answer_type = str(data.get("question_type", "")).strip()
                        choices_raw = str(data.get("question_choices", "")).strip()

                        if not answer_id or not answer_type or not choices_raw:
                            continue  # Incomplete row

                        choices = []
                        for line in choices_raw.split("\n"):
                            line = line.strip()
                            if not line:
                                continue
                            if line.startswith("|") and choices:
                                # Multi-line value, append to previous
                                choices[-1]["value"] += "\n" + line[1:].strip()
                            else:
                                choices.append({"urn": "", "value": line})

                        answers_dict[answer_id] = {
                            "type": answer_type,
                            "choices": choices,
                        }
            else:
                if verbose:
                    print(f'💬 ℹ️  No "Answers Definition" found')

            framework = {
                "urn": meta.get("urn"),
                "ref_id": meta.get("ref_id"),
                "name": meta.get("name"),
                "description": meta.get("description"),
            }

            translations = extract_translations_from_metadata(meta, "framework")
            if translations:
                framework["translations"] = translations
            else:
                if verbose:
                    print(f'💬 ℹ️  No "Translation" found')

            if "min_score" in meta:
                framework["min_score"] = int(meta["min_score"])
            if "max_score" in meta:
                framework["max_score"] = int(meta["max_score"])

            score_name = meta.get("scores_definition")
            if score_name and score_name in object_blocks:
                score_ws = object_blocks[score_name]["content_sheet"]
                score_rows = list(score_ws.iter_rows())
                score_header = [
                    str(c.value).strip().lower() if c.value else ""
                    for c in score_rows[0]
                ]
                score_defs = []
                for row in score_rows[1:]:
                    if not any(c.value for c in row):
                        continue
                    data = {
                        score_header[i]: row[i].value
                        for i in range(len(score_header))
                        if i < len(row)
                    }
                    score_entry = {
                        "score": int(data.get("score")),
                        "name": str(data.get("name", "")).strip(),
                        "description": str(data.get("description", "")).strip(),
                    }
                    if "description_doc" in data and data["description_doc"]:
                        score_entry["description_doc"] = str(
                            data["description_doc"]
                        ).strip()
                    translations = extract_translations_from_row(score_header, row)
                    if translations:
                        score_entry["translations"] = translations
                    score_defs.append(score_entry)
                framework["scores_definition"] = score_defs
            else:
                if verbose:
                    print(f'💬 ℹ️  No "Score Definition" found')

            ig_name = meta.get("implementation_groups_definition")
            if ig_name and ig_name in object_blocks:
                ig_content = object_blocks[ig_name]["content_sheet"]
                ig_rows = list(ig_content.iter_rows())
                ig_header = [
                    str(c.value).strip().lower() if c.value else "" for c in ig_rows[0]
                ]
                ig_defs = []
                for row in ig_rows[1:]:
                    if not any(c.value for c in row):
                        continue
                    data = {
                        ig_header[i]: row[i].value
                        for i in range(len(ig_header))
                        if i < len(row)
                    }
                    ig_entry = {
                        "ref_id": str(data.get("ref_id", "")).strip(),
                        "name": str(data.get("name", "")).strip(),
                        "description": str(data.get("description", "")).strip()
                        if data.get("description")
                        else None,
                    }
                    translations = extract_translations_from_row(ig_header, row)
                    if translations:
                        ig_entry["translations"] = translations
                    ig_defs.append(ig_entry)

                framework["implementation_groups_definition"] = ig_defs

            rows = list(content_ws.iter_rows())
            if rows:
                header = [
                    str(c.value).strip().lower() if c.value else "" for c in rows[0]
                ]
                parent_for_depth = {}
                count_for_depth = {}
                previous_node_urn = None
                previous_depth = 0
                counter = 0
                counter_fix = -1
                requirement_nodes = []
                all_urns = set()  # to detect duplicates
                for row in rows[1:]:
                    counter += 1
                    data = {
                        header[i]: row[i].value
                        for i in range(len(header))
                        if i < len(row)
                    }
                    if all(value is None for value in data.values()):
                        print(f"empty line {counter}")
                        continue
                    depth = int(data.get("depth", 1))
                    ref_id = data.get("ref_id")
                    ref_id = str(ref_id).strip() if ref_id is not None else None
                    name = data.get("name")
                    name = str(name).strip() if name is not None else None

                    if depth == previous_depth + 1:
                        parent_for_depth[depth] = previous_node_urn
                        count_for_depth[depth] = 1
                    elif depth <= previous_depth:
                        pass
                    else:
                        raise ValueError(
                            f"Invalid depth jump from {previous_depth} to {depth} at row {counter}"
                        )

                    # calculate urn
                    if (
                        compat_mode == 1
                    ):  # Use legacy URN fallback logic (for requirements without ref_id)
                        skip_count = str(
                            data.get("skip_count", "")
                        ).strip().lower() in ("1", "true", "yes", "x")
                        if skip_count:
                            counter_fix += 1
                            ref_id_urn = f"node{counter - counter_fix}-{counter_fix}"
                        else:
                            # Adds the ability to use the "urn_id" column despite compatibility mode set to "1"
                            if data.get("urn_id") and data.get("urn_id").strip():
                                ref_id_urn = data.get("urn_id").strip()
                            else:
                                ref_id_urn = (
                                    ref_id.lower().replace(" ", "-")
                                    if ref_id
                                    else f"node{counter - counter_fix}"
                                )

                        urn = f"{base_urn}:{ref_id_urn}"
                    else:  # If compat mode = {0,2}
                        if data.get("urn_id") and data.get("urn_id").strip():
                            urn = f"{base_urn}:{data.get('urn_id').strip()}"
                        elif ref_id:
                            # [+] Compat check
                            ref_id_clean = clean_urn_suffix(
                                ref_id, compat_mode=compat_mode
                            )
                            if verbose and ref_id != ref_id_clean:
                                print(
                                    f"💬 ⚠️  [WARNING] (calculate urn [ref_id]) Cleaned ref_id (for use in URN) '{ref_id}' → '{ref_id_clean}'"
                                )
                            urn = f"{base_urn}:{ref_id_clean}"
                        else:
                            p = parent_for_depth.get(depth)
                            c = count_for_depth.get(depth, 1)
                            if p:
                                urn = f"{p}:{c}" if p else f"{base_urn}:{c}"
                            elif name:
                                # [+] Compat check
                                name_clean = clean_urn_suffix(
                                    name, compat_mode=compat_mode
                                )
                                if verbose and name != name_clean:
                                    print(
                                        f"💬 ⚠️  [WARNING] (calculate urn [name]) Cleaned name (for use in URN) '{name}' → '{name_clean}'"
                                    )
                                urn = f"{base_urn}:{name_clean}"
                            else:
                                urn = f"{base_urn}:node{c}"
                            count_for_depth[depth] = c + 1
                    previous_node_urn = urn
                    previous_depth = depth
                    parent_urn = parent_for_depth.get(depth)
                    node = {
                        "urn": urn,
                        "assessable": bool(data.get("assessable")),
                        "depth": depth,
                    }
                    if parent_urn:
                        node["parent_urn"] = parent_urn
                    if "ref_id" in data and data["ref_id"]:
                        node["ref_id"] = str(data["ref_id"]).strip()
                    if "name" in data and data["name"]:
                        node["name"] = str(data["name"]).strip()
                    if "description" in data and data["description"]:
                        node["description"] = str(data["description"]).strip()
                    if "annotation" in data and data["annotation"]:
                        node["annotation"] = str(data["annotation"]).strip()
                    if "typical_evidence" in data and data["typical_evidence"]:
                        node["typical_evidence"] = str(data["typical_evidence"]).strip()
                    if (
                        "implementation_groups" in data
                        and data["implementation_groups"]
                    ):
                        node["implementation_groups"] = [
                            s.strip()
                            for s in str(data["implementation_groups"]).split(",")
                        ]
                    if "threats" in data and data["threats"]:
                        threats = expand_urns_from_prefixed_list(
                            data["threats"],
                            prefix_to_urn,
                            compat_mode=compat_mode,
                            verbose=verbose,
                        )
                        if threats:
                            node["threats"] = threats
                    if "reference_controls" in data and data["reference_controls"]:
                        rc = expand_urns_from_prefixed_list(
                            data["reference_controls"],
                            prefix_to_urn,
                            compat_mode=compat_mode,
                            verbose=verbose,
                        )
                        if rc:
                            node["reference_controls"] = rc
                    if "questions" in data and data["questions"]:
                        inject_questions_into_node(
                            node,
                            data.get("questions"),
                            data.get("answer"),
                            answers_dict,
                        )
                    translations = extract_translations_from_row(header, row)
                    if translations:
                        node["translations"] = translations
                    if node.get("urn") in all_urns:
                        raise ValueError(f"urn already used: {node.get('urn')}")
                    all_urns.add(node.get("urn"))
                    requirement_nodes.append(node)

                framework["requirement_nodes"] = requirement_nodes

            library["objects"]["framework"] = framework

        elif obj_type == "risk_matrix":
            matrix = parse_risk_matrix(obj["meta"], obj["content_sheet"], wb)
            if matrix:
                if "risk_matrix" not in library["objects"]:
                    library["objects"]["risk_matrix"] = []
                library["objects"]["risk_matrix"].append(matrix)

        elif obj_type == "requirement_mapping_set":
            meta = obj["meta"]
            content_ws = obj["content_sheet"]
            source_node_base_urn = obj["meta"].get("source_node_base_urn")
            target_node_base_urn = obj["meta"].get("target_node_base_urn")

            requirement_mapping_set = {
                "urn": meta.get("urn"),
                "ref_id": meta.get("ref_id"),
                "name": meta.get("name"),
                "description": meta.get("description"),
                "source_framework_urn": meta.get("source_framework_urn"),
                "target_framework_urn": meta.get("target_framework_urn"),
            }
            requirement_mapping_set_revert = {
                "urn": meta.get("urn") + "-revert",
                "ref_id": meta.get("ref_id") + "-revert",
                "name": meta.get("name"),
                "description": meta.get("description"),
                "target_framework_urn": meta.get("source_framework_urn"),
                "source_framework_urn": meta.get("target_framework_urn"),
            }

            translations = extract_translations_from_metadata(
                meta, "requirement_mapping_set"
            )
            if translations:
                requirement_mapping_set["translations"] = translations
                requirement_mapping_set_revert["translations"] = translations

            rows = list(content_ws.iter_rows())
            if not rows:
                continue
            header = [
                str(cell.value).strip().lower() if cell.value else ""
                for cell in rows[0]
            ]
            requirement_mappings = []
            requirement_mappings_revert = []

            for row in rows[1:]:
                if not any(cell.value for cell in row):
                    continue
                data = {
                    header[i]: row[i].value for i in range(len(header)) if i < len(row)
                }
                source_node_id_raw = str(data.get("source_node_id", "")).strip()
                target_node_id_raw = str(data.get("target_node_id", "")).strip()

                # Checks the required fields, cleaning the values first
                required_fields = ["source_node_id", "target_node_id", "relationship"]
                missing_fields = []

                for field in required_fields:
                    value = data.get(field)
                    if value is None or str(value).strip() == "":
                        missing_fields.append(field)

                if missing_fields:
                    quoted_fields = [f'"{field}"' for field in missing_fields]
                    raise ValueError(
                        f"(requirement_mapping_set) Missing or empty required field{'s' if len(missing_fields) > 1 else ''} in row #{row[0].row}: {', '.join(quoted_fields)}"
                    )

                # [+] Compat mode set to "2" so it can be compatible with every version of mappings >=v2 without having to choose a specific compat mode
                source_node_id = clean_urn_suffix(source_node_id_raw, compat_mode=2)
                target_node_id = clean_urn_suffix(target_node_id_raw, compat_mode=2)
                if verbose and source_node_id_raw != source_node_id:
                    print(
                        f"💬 ⚠️  [WARNING] (requirement_mapping_set) Cleaned source_node_id '{source_node_id_raw}' → '{source_node_id}'"
                    )
                if verbose and target_node_id_raw != target_node_id:
                    print(
                        f"💬 ⚠️  [WARNING] (requirement_mapping_set) Cleaned target_node_id '{target_node_id_raw}' → '{target_node_id}'"
                    )
                entry = {
                    "source_requirement_urn": source_node_base_urn
                    + ":"
                    + source_node_id,
                    "target_requirement_urn": target_node_base_urn
                    + ":"
                    + target_node_id,
                    "relationship": data.get("relationship").strip(),
                }
                entry_revert = {
                    "source_requirement_urn": target_node_base_urn
                    + ":"
                    + target_node_id,
                    "target_requirement_urn": source_node_base_urn
                    + ":"
                    + source_node_id,
                    "relationship": revert_relationship(
                        data.get("relationship").strip()
                    ),
                }
                if "rationale" in data and data["rationale"]:
                    entry["rationale"] = data.get("rationale").strip()
                    entry_revert["rationale"] = data.get("rationale").strip()
                if (
                    "strength_of_relationship" in data
                    and data["strength_of_relationship"]
                ):
                    entry["strength_of_relationship"] = int(
                        data.get("strength_of_relationship")
                    )
                    entry_revert["strength_of_relationship"] = int(
                        data.get("strength_of_relationship")
                    )
                requirement_mappings.append(entry)
                requirement_mappings_revert.append(entry_revert)
            requirement_mapping_set["requirement_mappings"] = requirement_mappings
            requirement_mapping_set_revert["requirement_mappings"] = (
                requirement_mappings_revert
            )
            library["objects"]["requirement_mapping_sets"] = [
                requirement_mapping_set,
                requirement_mapping_set_revert,
            ]

            # --- Validate source_node_id and target_node_id from mappings against Excel sheets ---
            # NOTE: This code simply checks the validity of the "source_node_id" and "target_node_id".
            #       It doesn't remove them from the created mapping if they aren't found in the Excel file.

            source_ids = set()
            target_ids = set()
            source_sheet_available = "source" in sheets
            target_sheet_available = "target" in sheets

            if source_sheet_available:
                source_sheet = wb["source"]
                source_header = [cell.value for cell in source_sheet[1]]
                if "node_id" in source_header:
                    idx = source_header.index("node_id")
                    for row in source_sheet.iter_rows(min_row=2):
                        if idx < len(row) and row[idx].value:
                            source_ids.add(str(row[idx].value).strip())
                else:
                    source_sheet_available = False
                    if verbose:
                        print('💬 ℹ️  "node_id" in "source" sheet header not found')
            else:
                if verbose:
                    print('💬 ℹ️  Sheet "source" not found')

            if target_sheet_available:
                target_sheet = wb["target"]
                target_header = [cell.value for cell in target_sheet[1]]
                if "node_id" in target_header:
                    idx = target_header.index("node_id")
                    for row in target_sheet.iter_rows(min_row=2):
                        if idx < len(row) and row[idx].value:
                            target_ids.add(str(row[idx].value).strip())
                else:
                    target_sheet_available = False
                    if verbose:
                        print('💬 ℹ️  "node_id" in "target" sheet header not found')

            else:
                if verbose:
                    print('💬 ℹ️  Sheet "target" not found')

            # Collect all used source/target IDs from mappings (with duplicates)
            used_source_ids = [
                m["source_requirement_urn"].split(":")[-1] for m in requirement_mappings
            ]
            used_target_ids = [
                m["target_requirement_urn"].split(":")[-1] for m in requirement_mappings
            ]

            source_missing_counts = Counter(
                id
                for id in used_source_ids
                if source_sheet_available and id not in source_ids
            )
            target_missing_counts = Counter(
                id
                for id in used_target_ids
                if target_sheet_available and id not in target_ids
            )

            # Print all warnings first (one per ID)
            if source_sheet_available:
                for sid in source_missing_counts:
                    print(
                        f'⚠️  [WARNING] source_node_id "{sid}" not found in sheet "source"'
                    )
            if target_sheet_available:
                for tid in target_missing_counts:
                    print(
                        f'⚠️  [WARNING] target_node_id "{tid}" not found in sheet "target"'
                    )

            # Then print duplicate counts (only for missing IDs)
            if source_sheet_available:
                for sid, count in source_missing_counts.items():
                    if count > 1:
                        print(
                            f'🔁 [DUPLICATE] source_node_id "{sid}" appears {count} times in mappings'
                        )
            if target_sheet_available:
                for tid, count in target_missing_counts.items():
                    if count > 1:
                        print(
                            f'🔁 [DUPLICATE] target_node_id "{tid}" appears {count} times in mappings'
                        )

            # Final summary
            total_missing_sources = sum(source_missing_counts.values())
            total_missing_targets = sum(target_missing_counts.values())
            if total_missing_sources or total_missing_targets:
                print(
                    f"⚠️  [SUMMARY] Missing usage count - Source: {total_missing_sources}, Target: {total_missing_targets}"
                )
                print(
                    f"ℹ️  Please note that these node IDs have been added to the mapping anyway.\
                    \n   If you want to correct them, please do so in your Excel file."
                )

        else:
            if obj_type not in [
                "answers",
                "implementation_groups",
                "scores",
                "urn_prefix",
            ]:
                print("type not handled:", obj_type)

    # Step 6: Export to YAML
    print(f'✅ YAML saved as: "{output_file}"')
    if not verbose:
        print(
            '💡 Tip: Use "--verbose" to display hidden messages. This can help to understand certain behaviors.'
        )

    with open(output_file, "w", encoding="utf-8") as f:
        yaml.dump(library, f, sort_keys=False, allow_unicode=False)


# --- CLI interface ------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(
        description="Create a YAML library from a v2 Excel file.",
        formatter_class=argparse.RawTextHelpFormatter,
    )
    parser.add_argument("input_file", type=str, help="Input Excel file (v2 format)")
    parser.add_argument(
        "-c",
        "--compat",
        type=int,
        choices=COMPATIBILITY_MODES.keys(),
        help="Compatibility mode number:\n    "
        + "\n    ".join(f"{k} = {v}" for k, v in COMPATIBILITY_MODES.items()),
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Enable verbose output. Verbose messages start with a 💬 (speech bubble) emoji.",
    )
    parser.add_argument(
        "--output",
        type=str,
        help="Custom output file name (only used in non-bulk mode).",
    )
    parser.add_argument(
        "--bulk",
        action="store_true",
        help="Enable bulk mode to process all .xlsx files in a directory.",
    )
    parser.add_argument(
        "--output-dir",
        type=str,
        help="Output directory for YAML files (only used with --bulk mode).",
    )
    args = parser.parse_args()

    input_path = Path(args.input_file)
    if not input_path.exists():
        print(f"❌ [ERROR] File not found: {input_path}")
        sys.exit(1)

    # Determine compatibility mode (default to 0 if not provided)
    compat_mode = args.compat if args.compat else 0
    if compat_mode not in COMPATIBILITY_MODES:
        print(
            f"❌ [ERROR] Invalid compatibility mode: {compat_mode}. Allowed modes: {list(COMPATIBILITY_MODES.keys())}"
        )
        sys.exit(1)

    # --- BULK MODE ------------------------------------------------------------
    if args.bulk:
        if args.output:
            print('❌ [ERROR] The option "--output" cannot be used with "--bulk" mode.')
            sys.exit(1)

        if not input_path.is_dir():
            print("❌ [ERROR] Bulk mode requires a directory as input")
            sys.exit(1)

        # Validate output directory and create it if needed
        if args.output_dir:
            output_dir = Path(args.output_dir)
            try:
                output_dir.mkdir(parents=True, exist_ok=True)
            except Exception as e:
                print(f'❌ [ERROR] Cannot create output directory: "{output_dir}": {e}')
                sys.exit(1)
        else:
            output_dir = Path.cwd()  # Use current working directory as default

        error_files = []  # Collect names of files that failed
        # Find all .xlsx files in the input directory (temp Excel files starting with "~$" are excluded)
        xlsx_files = [
            f for f in input_path.glob("*.xlsx") if not f.name.startswith("~$")
        ]

        if not xlsx_files:
            print(
                f'❌ [ERROR] No .xlsx files found in directory: "{input_path}". Abort...'
            )
            sys.exit(1)

        for i, file in enumerate(xlsx_files):
            output_path = output_dir / (file.stem + ".yaml")  # Output file path
            try:
                print(f'\n▶️  Processing file [{i + 1}/{len(xlsx_files)}]: "{file}"')
                create_library(
                    str(file),
                    str(output_path),
                    compat_mode=compat_mode,
                    verbose=args.verbose,
                )
            except Exception as e:
                print(f'❌ [ERROR] Failed to process "{file}": {e}')
                error_files.append(file.name)

        # Summary at the end of bulk processing
        print("\n📋 Bulk mode completed!")

        if error_files:
            print(
                f"❌ The following file{'s' if len(error_files) > 1 else ''} failed to process:"
            )
            for f in error_files:
                print(f"   - {f}")
                if not args.verbose:
                    print(
                        '💡 Tip: Use "--verbose" to display hidden messages. This can help to understand certain errors.'
                    )
            sys.exit(1)
        else:
            print("✅ All files processed successfully!")
            sys.exit(0)

    # --- SINGLE FILE MODE -----------------------------------------------------
    else:
        if args.output_dir:
            print(
                '❌ [ERROR] The option "--output-dir" can only be used with "--bulk" mode.'
            )
            sys.exit(1)

        # Determine output file path (add .yaml if missing)
        if args.output:
            output_path = Path(args.output)
            if output_path.suffix != ".yaml":
                output_path = output_path.name + ".yaml"
        else:
            output_path = Path(input_path.stem + ".yaml")

        try:
            create_library(
                str(input_path),
                str(output_path),
                compat_mode=compat_mode,
                verbose=args.verbose,
            )
        except Exception as e:
            print(f"❌ [ERROR] {e}")
            if not args.verbose:
                print(
                    '💡 Tip: Use "--verbose" to display hidden messages. This can help to understand certain errors.'
                )
            sys.exit(1)


if __name__ == "__main__":
    main()
