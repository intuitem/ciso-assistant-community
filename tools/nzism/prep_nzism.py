"""
simple script to transform the official NZISM XML file to an Excel file for CISO assistant framework conversion tool
"""

import openpyxl
import sys
import re
import argparse
import lxml
from openpyxl.styles import numbers
from bs4 import BeautifulSoup

def remove_last_dot(data):
    last_character = data[-1]
    if(last_character == "."):
        return data[:-1]
    else:
        return data

def clean_paragraph_tags(data):
    return BeautifulSoup(data, "lxml").text

def process_paragraph(level, paragraph):
    paragraph_id = remove_last_dot(paragraph["title"])
    paragraph_description = clean_paragraph_tags(paragraph.text)
    paragraph_compliance = ""
    paragraph_assessable = ""
    paragraph_classification = ""
    paragraph_cid = ""
    paragraph_name = ""

    if(paragraph.has_attr("compliance")):
        paragraph_compliance = paragraph["compliance"]
        paragraph_assessable = paragraph_compliance

    if(paragraph.has_attr("classification")):
        paragraph_classification = paragraph["classification"]

    if(paragraph.has_attr("cid")):
        paragraph_cid = f"CID: {paragraph["cid"]}"
        paragraph_name = f"Control; System Classification(s): {paragraph_classification}; Compliance: {paragraph_compliance} [{paragraph_cid}]"

    output_table.append(
        (paragraph_assessable, level, paragraph_id, paragraph_name, paragraph_description, paragraph_classification)
    )


parser = argparse.ArgumentParser(
    prog="convert_nzism",
    description="convert NZISM's XML file to CISO Assistant Excel file",
)
parser.add_argument("filename", help="name of NZISM XML file")
parser.add_argument("folderpath", help="path to the NZISM XML File, also used for output")
parser.add_argument("version", help="version of NZISM")
parser.add_argument("packager", help="name of packager entity", default="DEFEND Ltd")

args = parser.parse_args()
folder_path = args.folderpath
input_file_name = f"{folder_path}/{args.filename}"
version = args.version
packager = args.packager
output_file_name = f"{folder_path}/nzism-{version}.xlsx"

print("parsing", input_file_name)

# Define variable to load the dataframe
with open(input_file_name, 'r', encoding='utf8') as f:
    xml_data = f.read()
    
nzism = BeautifulSoup(xml_data, "xml")

output_table = []

for chapter in nzism.find_all("chapter"):
    chapter_id = remove_last_dot(chapter["title"].split(" ")[0])
    chapter_name = chapter["title"].replace(f"{chapter_id}.", "").strip()
    output_table.append(
        ("", 1, chapter_id, chapter_name)
    )
    print("Parsing Chapter", chapter_name)
    
    for section in chapter.findChildren("section", recursive=False):
        section_id = remove_last_dot(section["title"].split(" ")[0])
        section_name = section["title"].replace(f"{section_id}.", "").strip()
        output_table.append(
            ("", 2, section_id, section_name)
        )

        for subsection in section.findChildren("subsection", recursive=False):
            subsection_name = subsection["title"]
            output_table.append(
                ("", 3, "", subsection_name)
            )

            for paragraph in subsection.findChildren("paragraph", recursive=False):
                process_paragraph(4, paragraph)

            for block in subsection.findChildren("block", recursive=False):
                block_name = block["title"]
                output_table.append(
                    ("", 4, "", "", block_name, "")
                )

                for paragraph in block.findChildren("paragraph", recursive=False):
                    process_paragraph(5, paragraph)

print("generating", output_file_name)
wb_output = openpyxl.Workbook()
ws = wb_output.active
ws.title = "library_content"
ws.append(["library_urn", f"urn:{packager.lower()}:risk:library:nzism-v{version}"])
ws.append(["library_version", "1"])
ws.append(["library_locale", "en"])
ws.append(["library_ref_id", f"NSIZM-v{version}"])
ws.append(["library_name", f"NZISM v{version}"])
ws.append(["library_description", f"New Zealand Information Security Manual v{version}"])
ws.append(["library_copyright", """Except where specifically noted, all material on this site is © Crown copyright.

You can reproduce Crown copyright material free of charge without requiring specific permission, provided that the material is reproduced accurately and on the condition that the source of the material and its copyright status are acknowledged.

This licence does not apply to any logos, emblems and trade marks on the website or to the website’s design elements or to any photography and imagery. Those items may not be used without our express permission.

The permission to reproduce Crown copyright protected material does not extend to any material on this site that is identified as being the copyright of a third party. Authorisation to reproduce such material must be obtained from the copyright holders concerned.

The NZISM is licensed under the Creative Commons Attribution 4.0 New Zealand licence, available at https://creativecommons.org/licenses/by/4.0/(external link). You are free to copy, distribute, and adapt the work, as long as you attribute the work and abide by any other licence terms. For the avoidance of doubt, this means this licence applies only to material as set out in the NZISM."""])
ws.append(["library_provider", "New Zealand Government Communications Security Bureau"])
ws.append(["library_packager", packager])
ws.append(["framework_urn", f"urn:{packager.lower()}:risk:framework:nzism-v{version}"])
ws.append(["framework_ref_id", f"NSIZM-v{version}"])
ws.append(["framework_name", f"NZISM v{version}"])
ws.append(["framework_description", f"New Zealand Information Security Manual v{version}"])
ws.append(["tab", "requirements", "requirements"])
ws.append(["tab", "scores", "scores"])
ws.append(["tab", "implementation_groups", "implementation_groups"])

ws1 = wb_output.create_sheet("requirements")
ws1.append(
    ["assessable", "depth", "ref_id", "name", "description", "implementation_groups"]
)
for row in output_table:
    ws1.append(row)

ws2 = wb_output.create_sheet("scores")
ws2.append(["score", "name", "description"])
ws2.append([0, "Non Compliant", "",])
ws2.append([1, "Compliant", "",])

ws3 = wb_output.create_sheet("implementation_groups")
implementation_groups = ["All Classifications", "Unclassified", "In-Confidence", "Restricted/Sensitive", "Confidential", "Secret", "Top Secret"]
ws3.append(["ref_id", "name", "description"])
for implementation_group in implementation_groups:
    ws3.append([implementation_group, implementation_group, "",])

print("generate ", output_file_name)
wb_output.save(output_file_name)
