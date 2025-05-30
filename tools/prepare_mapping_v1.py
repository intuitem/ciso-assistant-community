import openpyxl
import argparse
import re
import yaml
from pprint import pprint
from collections import defaultdict


parser = argparse.ArgumentParser(
    prog="prepare_mapping.py",
    description="prepare a mapping excel file for CISO Assistant",
)
parser.add_argument("source_yaml")
parser.add_argument("target_yaml")
args = parser.parse_args()
packager = "intuitem"

print("parsing", args.source_yaml, args.target_yaml)
with open(args.source_yaml, "r", encoding="utf-8") as file:
    source = yaml.safe_load(file)
with open(args.target_yaml, "r", encoding="utf-8") as file:
    target = yaml.safe_load(file)

source_library_urn = source["urn"]
target_library_urn = target["urn"]
source_framework_urn = source["objects"]["framework"]["urn"]
target_framework_urn = target["objects"]["framework"]["urn"]
source_framework_ref_id = source["objects"]["framework"]["ref_id"]
target_framework_ref_id = target["objects"]["framework"]["ref_id"]
source_framework_name = source["objects"]["framework"]["name"]
target_framework_name = target["objects"]["framework"]["name"]

source_ref_id = source_framework_urn.split(":")[-1]
target_ref_id = target_framework_urn.split(":")[-1]
ref_id = "mapping-" + source_ref_id + "-to-" + target_ref_id
name = f"{source_framework_ref_id} -> {target_framework_ref_id}"
description = f"Mapping from {source_framework_name} to {target_framework_name}"
output_file_name = f"{ref_id}.xlsx"

wb_output = openpyxl.Workbook()
ws = wb_output.active
ws.title = "library_content"
ws.append(["library_urn", f"urn:{packager.lower()}:risk:library:{ref_id}"])
ws.append(["library_version", "1"])
ws.append(["library_locale", source["locale"]])
ws.append(["library_ref_id", ref_id])
ws.append(["library_name", name])
ws.append(["library_description", description])
ws.append(["library_copyright", packager])
ws.append(["library_provider", packager])
ws.append(["library_packager", packager])
ws.append(["library_dependencies", f"{source_library_urn}, {target_library_urn}"])
ws.append(
    ["mapping_urn", f"urn:{packager.lower()}:risk:req_mapping_set:{ref_id}"]
)
ws.append(["mapping_ref_id", ref_id])
ws.append(["mapping_name", name])
ws.append(["mapping_description", description])
ws.append(["mapping_source_framework_urn", source_framework_urn])
ws.append(["mapping_target_framework_urn", target_framework_urn])
ws.append(
    [
        "mapping_source_node_base_urn",
        f"urn:{packager.lower()}:risk:req_node:{source_ref_id}",
    ]
)
ws.append(
    [
        "mapping_target_node_base_urn",
        f"urn:{packager.lower()}:risk:req_node:{target_ref_id}",
    ]
)
ws.append(["tab", "mappings", "mappings"])

ws1 = wb_output.create_sheet("mappings")
ws1.append(
    [
        "source_node_id",
        "target_node_id",
        "relationship",
        "rationale",
        "strength_of_relationship",
    ]
)
for node in source["objects"]["framework"]["requirement_nodes"]:
    if node["assessable"]:
        node_id = node["urn"].split(":")[-1]
        ws1.append([node_id])

ws2 = wb_output.create_sheet("guidelines")
ws2.append(["source_node_id", "use the last segment of the source URN"])
ws2.append(["target_node_id", "use the last segment of the target URN"])
ws2.append(["relationships", "use one of the following values"])
ws2.append(["", "subset"])
ws2.append(["", "intersect"])
ws2.append(["", "equal"])
ws2.append(["", "superset"])
ws2.append(["", "not_related"])
ws2.append(["rationale", "use one of the following values (or leave empty)"])
ws2.append(["", "syntactic"])
ws2.append(["", "semantic"])
ws2.append(["", "functional"])
ws2.append(["strength_of_relationship", "use an integer or empty value"])

ws3 = wb_output.create_sheet("source")
ws3.append(["node_id", "assessable", "urn", "ref_id", "name", "description"])
for node in source["objects"]["framework"]["requirement_nodes"]:
    node_id = node["urn"].split(":")[-1] if node["assessable"] else ""
    ws3.append(
        [
            node_id,
            node["assessable"],
            node["urn"],
            node.get("ref_id"),
            node.get("name"),
            node.get("description"),
        ]
    )

ws4 = wb_output.create_sheet("target")
ws4.append(["node_id", "assessable", "urn", "ref_id", "name", "description"])
for node in target["objects"]["framework"]["requirement_nodes"]:
    node_id = node["urn"].split(":")[-1] if node["assessable"] else ""
    ws4.append(
        [
            node_id,
            node["assessable"],
            node["urn"],
            node.get("ref_id"),
            node.get("name"),
            node.get("description"),
        ]
    )

print("generate ", output_file_name)
wb_output.save(output_file_name)
