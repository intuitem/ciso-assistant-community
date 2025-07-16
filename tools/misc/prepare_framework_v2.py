"""
Excel Framework Generator v0.4
--------------------------

This script reads a YAML configuration file and generates a structured Excel (.xlsx) file
based on a predefined template to create the base structure of a CISO Assistant framework in v2.
It ensures the integrity of required fields, supports optional implementation groups, and performs
validation with user-friendly error messages.

Usage:
    python generate_excel_from_yaml.py -i input.yaml [-o output.xlsx]

Dependencies:
    - PyYAML
    - openpyxl

Features:
    - Required field validation
    - URN format enforcement
    - Optional implementation groups with structure validation
    - Support for multiple locales with additional localized columns in the Excel sheets
    - Automatic Excel generation with multiple sheets, including localized content and metadata
    - Informative logs and error handling
"""


import re
import sys
import os
import yaml
import argparse
from openpyxl import Workbook
# from datetime import date


# Validates that urn_root only contains allowed characters
def validate_urn_root(urn_root):
    if not re.fullmatch(r"[a-z0-9._-]+", urn_root):
        raise ValueError(
            "(validate_urn_root) Invalid URN root : only lowercase alphanumeric characters, '-', '_', and '.' are allowed."
        )


# Checks for required non-empty field
def validate_required_field(field_name, value):
    if value is None or str(value).strip() == "":
        raise ValueError(f"Missing or empty required field: \"{field_name}\"")


def validate_implementation_groups(impl_groups, context="implementation_groups"):
    if impl_groups is not None:
        if not isinstance(impl_groups, list) or not impl_groups:
            raise ValueError(f'(validate_implementation_groups) Field "{context}" must be a non-empty list if defined.')
        for i, group in enumerate(impl_groups, start=1):
            if "ref_id" not in group or not str(group["ref_id"]).strip():
                raise ValueError(f'(validate_implementation_groups) Missing or empty \"ref_id\" in {context} #{i}')
            if "name" not in group or not str(group["name"]).strip():
                raise ValueError(f'(validate_implementation_groups) Missing or empty \"name\" in {context} #{i}')
            if "description" not in group or not str(group["description"]).strip():
                print(f'⚠️  [WARNING] (validate_implementation_groups) Missing or empty \"description\" in {context} #{i}')


# Added validation for extra_locales
def validate_extra_locales(data):
    extra_locales = data.get("extra_locales")
    if extra_locales is None:
        return

    if not isinstance(extra_locales, list) or not extra_locales:
        raise ValueError("(validate_extra_locales) Field \"extra_locales\" must be a non-empty list if defined.")

    impl_groups_main = data.get("implementation_groups") or []

    for i, locale_entry in enumerate(extra_locales, start=1):
        if not isinstance(locale_entry, dict) or len(locale_entry) != 1:
            raise ValueError(f"(validate_extra_locales) Each entry in \"extra_locales\" must be a dict with exactly one locale code key (entry #{i})")

        for loc_code, loc_data in locale_entry.items():
            if not re.fullmatch(r"[a-z0-9-]{2}", loc_code):
                raise ValueError(f"(validate_extra_locales) Invalid locale code \"{loc_code}\" in extra_locales entry #{i}")

            if not isinstance(loc_data, dict) or not loc_data:
                raise ValueError(f"(validate_extra_locales) Locale data for \"{loc_code}\" must be a non-empty dict (entry #{i})")

            # Optional fields: warn if missing or empty
            for req_field in ["framework_name", "description", "copyright"]:
                if req_field not in loc_data or str(loc_data[req_field]).strip() == "":
                    print(f"⚠️  [WARNING] (validate_extra_locales) Missing or empty field \"{req_field}\" in extra_locales for locale \"{loc_code}\" (entry #{i})")

            # Validate implementation_groups in locale if present
            loc_impl_groups = loc_data.get("implementation_groups")
            if loc_impl_groups is not None:
                if not impl_groups_main:
                    print(f"⚠️  [WARNING] (validate_extra_locales) \"implementation_groups\" defined in locale \"{loc_code}\" but missing in framework. Skipping locale's \"implementation_groups\"")
                    continue

                validate_implementation_groups(loc_impl_groups, f'implementation_groups in locale "{loc_code}"')

                # Check that each local ref_id exists in the main implementation_groups
                for group in loc_impl_groups:
                    if all(group["ref_id"] != fg["ref_id"] for fg in impl_groups_main):
                        raise ValueError(f'ref_id "{group["ref_id"]}" in locale "{loc_code}" implementation_groups does not exist in framework implementation_groups.')


def validate_yaml_data(data):
    required_fields = [
        "urn_root", "locale", "ref_id", "framework_name", "description",
        "copyright", "provider", "packager", "framework_sheet_base_name"
    ]
    for field in required_fields:
        validate_required_field(field, data.get(field))
        
    if not re.fullmatch(r"[a-z0-9-]{2}", data.get("locale")):
        raise ValueError(f"(validate_yaml_data) Invalid locale code \"{data.get("locale")}\" in \"locale\" entry")

    validate_urn_root(data["urn_root"])

    impl_base = data.get("implementation_groups_sheet_base_name")
    impl_groups = data.get("implementation_groups")

    # Validate implementation_groups if base name is defined
    if impl_base:
        validate_implementation_groups(impl_groups, "implementation_groups")
        print(f"ℹ️  [INFO] Implementation groups will be added using base name: \"{impl_base}\"")

    # Validate extra_locales if present
    validate_extra_locales(data)


# Modification in create_excel_from_yaml:

def create_excel_from_yaml(yaml_path, output_excel=None):
    if not os.path.isfile(yaml_path):
        raise FileNotFoundError(f"YAML file not found: \"{yaml_path}\"")

    with open(yaml_path, 'r', encoding='utf-8') as f:
        data = yaml.safe_load(f)

    validate_yaml_data(data)

    yaml_output_name = data.get("excel_file_name")
    if output_excel is None:
        if yaml_output_name and str(yaml_output_name).strip():
            output_excel = yaml_output_name.strip()
        else:
            output_excel = "output.xlsx"
    elif yaml_output_name and str(yaml_output_name).strip() and output_excel != yaml_output_name.strip():
        print(f"ℹ️  [INFO] Overriding YAML \"excel_file_name\" with command line argument.")

    # Ensure output filename ends with .xlsx extension
    if not output_excel.lower().endswith(".xlsx"):
        output_excel += ".xlsx"

    urn_root = data["urn_root"]
    locale = data["locale"]
    ref_id = data["ref_id"]
    name = data["framework_name"]
    description = data["description"]
    copyright_ = data["copyright"]
    provider = data["provider"]
    packager = data["packager"]
    framework_sheet_base = data["framework_sheet_base_name"]
    impl_group_base = data.get("implementation_groups_sheet_base_name")
    impl_groups = data.get("implementation_groups", [])

    wb = Workbook()

    # Main sheet: library_meta
    ws1 = wb.active
    ws1.title = "library_meta"
    ws1.append(["type", "library"])
    ws1.append(["urn", f"urn:intuitem:risk:library:{urn_root}"])
    ws1.append(["version", "1"])
    ws1.append(["locale", locale])
    ws1.append(["ref_id", ref_id])
    ws1.append(["name", name])
    ws1.append(["description", description])
    ws1.append(["copyright", copyright_])
    ws1.append(["provider", provider])
    ws1.append(["packager", packager])

    # Add extra_locales to library_meta sheet
    extra_locales = data.get("extra_locales")
    if extra_locales:
        for locale_entry in extra_locales:
            for loc_code, loc_data in locale_entry.items():
                # Only add fields that are present and non-empty
                if "framework_name" in loc_data and str(loc_data["framework_name"]).strip():
                    ws1.append([f"name[{loc_code}]", loc_data["framework_name"]])
                if "description" in loc_data and str(loc_data["description"]).strip():
                    ws1.append([f"description[{loc_code}]", loc_data["description"]])
                if "copyright" in loc_data and str(loc_data["copyright"]).strip():
                    ws1.append([f"copyright[{loc_code}]", loc_data["copyright"]])

    # Main sheet: framework_meta
    framework_meta_sheet = wb.create_sheet(f"{framework_sheet_base}_meta")
    framework_meta_sheet.append(["type", "framework"])
    framework_meta_sheet.append(["base_urn", f"urn:intuitem:risk:req_node:{urn_root}"])
    framework_meta_sheet.append(["urn", f"urn:intuitem:risk:framework:{urn_root}"])
    framework_meta_sheet.append(["ref_id", ref_id])
    framework_meta_sheet.append(["name", name])
    framework_meta_sheet.append(["description", description])
    if impl_group_base:
        framework_meta_sheet.append(["implementation_groups_definition", impl_group_base])

    # Add extra_locales to framework_meta sheet
    if extra_locales:
        for locale_entry in extra_locales:
            for loc_code, loc_data in locale_entry.items():
                if "framework_name" in loc_data and str(loc_data["framework_name"]).strip():
                    framework_meta_sheet.append([f"name[{loc_code}]", loc_data["framework_name"]])
                if "description" in loc_data and str(loc_data["description"]).strip():
                    framework_meta_sheet.append([f"description[{loc_code}]", loc_data["description"]])

    # Main sheet framework_content
    framework_content_sheet = wb.create_sheet(f"{framework_sheet_base}_content")
    base_columns = ["assessable", "depth", "ref_id", "name", "description", "annotation", "typical_evidence"]
    
    if impl_group_base:
        base_columns.append("implementation_groups")
    
    # Add localized columns if extra_locales is defined
    localized_columns = []
    if extra_locales:
        for locale_entry in extra_locales:
            for loc_code in locale_entry.keys():
                localized_columns.extend([
                    f"name[{loc_code}]",
                    f"description[{loc_code}]",
                    f"annotation[{loc_code}]",
                    f"typical_evidence[{loc_code}]"
                ])

    full_columns = base_columns + localized_columns
    framework_content_sheet.append(full_columns)


    # Implementation_groups sheets (main locale only)
    if impl_group_base:
        impl_meta_sheet = wb.create_sheet(f"{impl_group_base}_meta")
        impl_meta_sheet.append(["type", "implementation_groups"])
        impl_meta_sheet.append(["name", impl_group_base])

        impl_content_sheet = wb.create_sheet(f"{impl_group_base}_content")

        # Base columns
        base_header = ["ref_id", "name", "description"]
        extra_cols = []

        # Prepare extra_locales columns if present
        if extra_locales:
            for locale_entry in extra_locales:
                for loc_code, loc_data in locale_entry.items():
                    loc_impl_groups = loc_data.get("implementation_groups")
                    if loc_impl_groups and isinstance(loc_impl_groups, list):
                        extra_cols.append(f"name[{loc_code}]")
                        extra_cols.append(f"description[{loc_code}]")

        full_header = base_header + extra_cols
        impl_content_sheet.append(full_header)

        # ref_id -> row index mapping (starting at row 2)
        ref_id_to_row = {}
        for i, group in enumerate(impl_groups, start=2):
            row = [group.get("ref_id", ""), group.get("name", ""), group.get("description", "")]
            impl_content_sheet.append(row)
            ref_id_to_row[group.get("ref_id", "")] = i

        # Dictionary ref_id -> {lang: {name, description}}
        lang_impl_groups_map = {}
        for locale_entry in extra_locales or []:
            for loc_code, loc_data in locale_entry.items():
                loc_impl_groups = loc_data.get("implementation_groups")
                if loc_impl_groups:
                    for group in loc_impl_groups:
                        rid = group["ref_id"]
                        if rid not in lang_impl_groups_map:
                            lang_impl_groups_map[rid] = {}
                        lang_impl_groups_map[rid][loc_code] = {
                            "name": group["name"],
                            "description": group.get("description", "")
                        }

        # Extra columns indexes map
        col_index_map = {col: idx + 1 for idx, col in enumerate(full_header)}

        # Fill localized columns with translated data
        for rid, langs_data in lang_impl_groups_map.items():
            if rid in ref_id_to_row:
                row_num = ref_id_to_row[rid]
                for lang_code, texts in langs_data.items():
                    name_col = col_index_map.get(f"name[{lang_code}]")
                    desc_col = col_index_map.get(f"description[{lang_code}]")
                    if name_col:
                        impl_content_sheet.cell(row=row_num, column=name_col, value=texts["name"])
                    if desc_col:
                        impl_content_sheet.cell(row=row_num, column=desc_col, value=texts["description"])
                        

    try:
        wb.save(output_excel)
        print(f"✅ [SUCCESS] Excel file saved successfully: \"{output_excel}\"")
    except Exception as e:
        print(f"❌ [ERROR] Failed to save Excel file: {e}")
        print("💡 Tip: Make sure the Excel file is not open in another program and that you have write permission.")
        sys.exit(1)


# CLI entry point
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate Excel file from YAML configuration")
    parser.add_argument('-i', '--input', default="prepare_framework_v2_config.yaml",
                        help="Input YAML configuration file (default: prepare_framework_v2_config.yaml)")
    parser.add_argument('-o', '--output', default=None,
                        help="Output Excel file name (optional)")

    args = parser.parse_args()

    try:
        create_excel_from_yaml(args.input, args.output)
    except Exception as err:
        print(f"❌ [FATAL ERROR] {err}")
        sys.exit(1)
