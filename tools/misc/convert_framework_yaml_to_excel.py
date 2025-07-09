"""
YAML to Excel Framework Converter v0.3

Script to recreate a structured Excel (V2 format) from a given YAML file.

Usage:
    python convert_framework_yaml_to_excel.py input.yaml [output.xlsx]

Arguments:
    input.yaml       Path to the input YAML file containing the data.
    output.xlsx      (Optional) Path to the output Excel file to generate.
                     If not provided, defaults to 'convert_<input_basename>.xlsx'.

Description:
    This script parses the input YAML file, extracts the defined objects and metadata,
    and reconstructs the original Excel file structure with multiple sheets such as
    'library_meta', 'framework_meta', 'framework_content', and others.

    It handles the following cases: framework, answers, implementation groups,
    scores, threats and reference controls, replicating the structure expected in the V2 Excel format.

WARNING:
    The generated Excel file is not guaranteed to be directly compatible with
    "convert_library_v2.py". Manual adjustments and verification may be required
    to ensure correctness and compatibility.

Limitations:
    This script does not recreate the "urn_prefix" sheet nor any risk matrices.

Info:
    If you are trying to convert a mapping YAML file to Excel, please use the script
    "convert_mapping_yaml_to_excel.py" instead.
    
Requirements:
    Python 3.9 or higher (uses str.removeprefix)

Example:
    python convert_framework_yaml_to_excel.py my_data.yaml
"""


SCRIPT_VERSION = '0.3'


import argparse
import os
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font



def write_info_sheet(wb, yaml_path):
    
    ws_info = wb.create_sheet("info", 0)  # Insert at first position

    yaml_filename = os.path.basename(yaml_path)

    ws_info["A1"] = f"YAML to Excel Framework Converter v{SCRIPT_VERSION}"
    ws_info["A1"].font = Font(size=48, bold=True)

    ws_info["A2"] = f"Source file : {yaml_filename}"
    ws_info["A2"].font = Font(size=15)

    ws_info["A3"] = ('Please verify and potentially adjust this Excel before using it as input to "convert_library_v2.py".')
    ws_info["A3"].font = Font(size=20, italic=True)



def write_sheet(ws, header, rows):
    ws.append(header)
    for row in rows:
        ws.append([str(row.get(col, "") or "") for col in header])



def extract_translation_columns(objects_list):
    lang_to_fields = {}
    for obj in objects_list:
        translations = obj.get("translations", {})
        for lang_code, lang_fields in translations.items():
            if lang_code not in lang_to_fields:
                lang_to_fields[lang_code] = set()
            for field in lang_fields:
                lang_to_fields[lang_code].add(field)

    translation_columns = []
    for lang in sorted(lang_to_fields):
        for field in sorted(lang_to_fields[lang]):
            translation_columns.append(f"{field}[{lang}]")
    return translation_columns



def extract_translation_values(obj, translation_columns):
    row = {}
    for key in translation_columns:
        if "[" in key and key.endswith("]"):
            field, lang = key[:-1].split("[")
            row[key] = obj.get("translations", {}).get(lang, {}).get(field, "")
    return row



def write_translation_rows(ws, translations, fields=("name", "description", "copyright")):
    for lang, field_map in translations.items():
        for field in fields:
            if field in field_map:
                key = f"{field}[{lang}]"
                val = field_map[field]
                ws.append([str(key), str(val)])



def remove_empty_columns(ws):
    col_idx = 1
    while col_idx <= ws.max_column:
        if all((ws.cell(row=row_idx, column=col_idx).value in [None, ""]) for row_idx in range(2, ws.max_row + 1)):
            ws.delete_cols(col_idx)
        else:
            col_idx += 1



def convert_list_fields_to_string(row, obj, fields):
    for field in fields:
        if field in obj:
            row[field] = ", ".join(obj[field])



def calculate_base_urn(items):
    if len(items) < 2:
        return "???"

    urn1 = items[0].get("urn", "")
    urn2 = items[1].get("urn", "")

    # Find common prefix by parts (split by ':')
    parts1 = urn1.split(":")
    parts2 = urn2.split(":")

    common_parts = []
    for p1, p2 in zip(parts1, parts2):
        if p1 == p2:
            common_parts.append(p1)
        else:
            break

    return ":".join(common_parts) if common_parts else "???"



def process_implementation_groups(wb, ig_defs):

    ig_meta_ws = wb.create_sheet(title="implementation_groups_meta")
    ig_meta_ws.append(["type", "implementation_groups"])
    ig_meta_ws.append(["name", "implementation_groups"])

    ig_content_ws = wb.create_sheet(title="implementation_groups_content")
    ig_content_rows = list(ig_defs)  # shallow copy

    # Extract translation columns
    translation_columns = extract_translation_columns(ig_content_rows)

    # Add translation values into each row
    for row in ig_content_rows:
        row.update(extract_translation_values(row, translation_columns))
        row.pop("translations", None)

    if ig_content_rows:
        ig_headers = list(ig_content_rows[0].keys())
        write_sheet(ig_content_ws, ig_headers, ig_content_rows)



def process_scores(wb, scores_def):

    scores_meta_ws = wb.create_sheet(title="scores_meta")
    scores_meta_ws.append(["type", "scores"])
    scores_meta_ws.append(["name", "scores"])

    scores_content_ws = wb.create_sheet(title="scores_content")
    scores_content_rows = list(scores_def)  # shallow copy

    # Extract translation columns
    translation_columns = extract_translation_columns(scores_content_rows)

    # Add translation values into each row
    for row in scores_content_rows:
        row.update(extract_translation_values(row, translation_columns))
        row.pop("translations", None)

    if scores_content_rows:
        scores_headers = list(scores_content_rows[0].keys())
        write_sheet(scores_content_ws, scores_headers, scores_content_rows)



def write_answers_sheet(wb, meta_ws, answer_definitions):
    
    meta_ws.append(["answers_definition", "answers"])
    
    answer_meta_ws = wb.create_sheet(title="answers_meta")
    answer_meta_ws.append(["type", "answers"])
    answer_meta_ws.append(["name", "answers"])

    answers_content_ws = wb.create_sheet("answers_content")
    headers = ["id", "question_type", "question_choices"]
    rows = []
    for v in answer_definitions.values():
        row = {k: v.get(k, "") for k in headers}
        rows.append(row)
    write_sheet(answers_content_ws, headers, rows)



def process_reference_controls(wb, ref_controls):

    # Calculate base_urn
    base_urn = calculate_base_urn(ref_controls)

    # Create the meta sheet
    ref_cont_meta_ws = wb.create_sheet(title="reference_controls_meta")
    ref_cont_meta_ws.append(["type", "reference_controls"])
    ref_cont_meta_ws.append(["base_urn", str(base_urn)])

    # Define base columns
    headers = ["ref_id", "name", "csf_function", "category", "description", "annotation"]

    # Extract translation columns
    translation_columns = extract_translation_columns(ref_controls)
    full_headers = headers + translation_columns

    # Create the content sheet
    content_ws = wb.create_sheet(title="reference_controls_content")
    rows = []

    # Fill the rows
    for ctrl in ref_controls:
        row = {key: ctrl.get(key, "") for key in headers}
        row.update(extract_translation_values(ctrl, translation_columns))
        rows.append(row)

    # Write and clean
    write_sheet(content_ws, full_headers, rows)
    remove_empty_columns(content_ws)



def process_threats(wb, threats):
    
    # Calculate base_urn
    base_urn = calculate_base_urn(threats)

    # Create the meta sheet
    threats_meta_ws = wb.create_sheet(title="threats_meta")
    threats_meta_ws.append(["type", "threats"])
    threats_meta_ws.append(["base_urn", str(base_urn)])

    # Define base columns
    headers = ["ref_id", "name", "description", "annotation"]

    # Extract translation columns
    translation_columns = extract_translation_columns(threats)
    full_headers = headers + translation_columns

    # Create the content sheet
    content_ws = wb.create_sheet(title="threats_content")
    rows = []

    # Fill the rows
    for threat in threats:
        row = {key: threat.get(key, "") for key in headers}
        row.update(extract_translation_values(threat, translation_columns))
        rows.append(row)

    # Write and clean
    write_sheet(content_ws, full_headers, rows)
    remove_empty_columns(content_ws)



def recreate_excel_from_yaml(yaml_path, output_excel_path):
    
    print(f'⌛ Processing "{os.path.basename(yaml_path)}"...')
    
    # Attempt to load YAML file, raise an error if file is missing or YAML is invalid
    try:
        with open(yaml_path, 'r', encoding='utf-8') as f:
            data = yaml.safe_load(f)
    except FileNotFoundError:
        raise FileNotFoundError(f"YAML file not found: \"{yaml_path}\"")
    except yaml.YAMLError as e:
        raise ValueError(f"Error parsing YAML file: {e}")

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # --- [Info Sheet] ---
    write_info_sheet(wb, yaml_path)

    # --- [First Sheet] library_meta ---
    library_meta_ws = wb.create_sheet(title="library_meta")
    library_meta_ws.append(["type", "library"])

    translations_block = data.get("translations", {})

    for k, v in data.items():
        if k == "objects":
            break
        
        if k == "dependencies":
            deps_str = ", ".join(str(dep) for dep in v)
            library_meta_ws.append([k, deps_str])
            
        elif not isinstance(v, (list, dict)):
            library_meta_ws.append([k, v])


    write_translation_rows(library_meta_ws, translations_block, fields=("name", "description", "copyright"))

    for obj_key, obj_value in data.get("objects", {}).items():
        
        # --- [WARNING] Risk Matrices ---
        if obj_key == "risk_matrix":
            print("⚠️  [WARNING] This script can't handle risk matrices")
        
        # --- [WARNING] Mappings --- 
        if obj_key == "requirement_mapping_sets" or obj_key == "requirement_mapping_set":
            print("⚠️  [WARNING] This script can't handle mappings")
            print('💡 Tip: Use "convert_mapping_yaml_to_excel.py" to convert a YAML mapping into an Excel mapping')
    

        if obj_key == "framework":
            
             # --- [Sheet] framework_meta ---
            meta_data = obj_value.copy()
            content = meta_data.pop("requirement_nodes", [])
            if "implementation_groups_definition" in meta_data:
                ig_defs = meta_data.pop("implementation_groups_definition")
            else:
                ig_defs = None
                
            if "scores_definition" in meta_data:
                scores_def = meta_data.pop("scores_definition")
            else:
                scores_def = None


            framework_translations = meta_data.pop("translations", {})

            meta_ws = wb.create_sheet(title="framework_meta")
            meta_ws.append(["type", "framework"])

            base_urn = None
            for k, v in meta_data.items():
                if isinstance(v, list) or isinstance(v, dict):
                    continue
                meta_ws.append([str(k), str(v)])
                
                if k == "urn":
                    base_urn = v.replace("framework", "req_node")
                    meta_ws.append(["base_urn", base_urn])

            write_translation_rows(meta_ws, framework_translations, fields=("name", "description"))

            if ig_defs:
                meta_ws.append(["implementation_groups_definition", "implementation_groups"])
            if scores_def:
                meta_ws.append(["scores_definition", "scores"])


            # --- [Sheet] framework_content ---
            content_ws = wb.create_sheet(title="framework_content")
            headers = [
                "assessable", "depth", "ref_id", "urn_id", "name", "description",
                "annotation", "typical_evidence", "questions", "answer",
                "implementation_groups", "reference_controls", "threats", "urn", "parent_urn"
            ]

            translation_columns = extract_translation_columns(content)

            full_headers = headers + translation_columns
            rows = []
            answer_definitions = {}  # Store answer definitions for answers_content

            for node in content:
                row = {key: node.get(key, "") for key in headers}

                # Convert "assessable" to "x" if true
                val = row.get("assessable", "")
                if isinstance(val, bool) and val is True:
                    row["assessable"] = "x"
                elif isinstance(val, str) and val.strip().lower() == "true":
                    row["assessable"] = "x"
                else:
                    row["assessable"] = ""

                # Compute urn_id from urn and base_urn
                if base_urn and row.get("urn"):
                    row["urn_id"] = row["urn"].removeprefix(base_urn + ":")

                convert_list_fields_to_string(row, node, ["implementation_groups", "threats", "reference_controls"])

                questions = node.get("questions", {})
                question_texts = []
                answer_values = []

                for qkey in sorted(questions):
                    q = questions[qkey]
                    q_text = q.get("text", "").strip()
                    q_type = q.get("type", "").strip()

                    if not q_text or not q_type:
                        continue

                    question_texts.append(q_text)

                    if q_type in ["text", "date"]:
                        answer_id = q_type.upper()
                        answer_values.append(answer_id)
                        if answer_id not in answer_definitions:
                            answer_definitions[answer_id] = {
                                "id": answer_id,
                                "question_type": q_type
                            }
                    elif q_type in ["unique_choice", "multiple_choice"]:
                        choices = q.get("choices", [])
                        values = [str(c["value"]).strip().lower() for c in choices if "value" in c]
                        answer_id = f"{q_type}_{''.join(values)}"
                        answer_values.append(answer_id)
                        if answer_id not in answer_definitions:
                            answer_definitions[answer_id] = {
                                "id": answer_id,
                                "question_type": q_type,
                                "question_choices": "\n".join(c["value"] for c in choices if "value" in c)
                            }

                if question_texts:
                    row["questions"] = "\n".join(question_texts)
                if answer_values:
                    row["answer"] = "\n".join(answer_values)

                row.update(extract_translation_values(node, translation_columns))
                rows.append(row)

            write_sheet(content_ws, full_headers, rows)
            remove_empty_columns(content_ws)


            # --- [Sheets] answers_meta & answers_content ---
            if answer_definitions:
                write_answers_sheet(wb, meta_ws, answer_definitions)


            # --- [Sheets] implementation_groups_meta & implementation_groups_content ---
            if ig_defs:
                process_implementation_groups(wb, ig_defs)

   
            # --- [Sheets] scores_meta & scores_content ---
            if scores_def:
                process_scores(wb, scores_def)

        # --- [Sheets] reference_controls_meta & reference_controls_content ---
        elif obj_key == "reference_controls":
            process_reference_controls(wb, obj_value)

        # --- [Sheets] threats_meta & threats_content ---
        elif obj_key == "threats":
            process_threats(wb, obj_value)


    # Try saving the workbook, raise error on failure
    try:
        wb.save(output_excel_path)
    except Exception as e:
        raise IOError(f"Failed to save Excel file: {e}")
    print(f"✅ Excel recreated: \"{output_excel_path}\"")


if __name__ == "__main__":

    parser = argparse.ArgumentParser(description="Recreate a structured V2 Excel file from YAML.")
    parser.add_argument("yaml_path", help="Path to input YAML file")
    parser.add_argument("output_excel", nargs="?", help="Path to output Excel file (optional, defaults to YAML filename with .xlsx)")

    args = parser.parse_args()

    # Determine default Excel filename if not provided
    if args.output_excel:
        output_excel = args.output_excel
        if not output_excel.lower().endswith(".xlsx"):
            output_excel += ".xlsx"
    else:
        base_name = os.path.splitext(os.path.basename(args.yaml_path))[0]
        output_excel = "convert_" + base_name + ".xlsx"

    # Wrap call in try/except to catch and report errors cleanly
    try:
        recreate_excel_from_yaml(args.yaml_path, output_excel)
    except Exception as e:
        print(f"❌ [ERROR] {e}")
        exit(1)
