"""
Prettier Footnote Reference Script
----------------------------------

This script enhances the readability of footnote references in an Excel worksheet by modifying text in selected columns.
It detects:
    - Words consisting of letters immediately followed by numbers (e.g., "run20")
    - Phrases enclosed in parentheses immediately followed by numbers (e.g., "(keep running)7")
and makes the footnote numbers stand out by enclosing them in square brackets.

This formatting improves the visibility and clarity of footnote references within the text.

Examples:
    "run20"             -> "run[20]"
    "(keep running)7"   -> "(keep running)[7]"
    "light35."          -> "light[35]."

Usage:
    python script.py <excel_file.xlsx> [--verbose]

Options:
    --verbose           Print each cell modification with row number and column name.

Configuration:
    - Modify `SHEET_NAME` to target a specific worksheet.
    - Update `COLUMN_PREFIXES` to apply the rule to columns starting with certain prefixes.
    - Use `SPECIFIC_COLUMNS` to apply the rule to explicitly named columns.

Notes:
    - The script uses `openpyxl` to work with .xlsx files.
    - It does not overwrite the original file; a copy is saved with "_modified" in the filename.
"""


import re
import sys
import os
from typing import List, Tuple
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell


# === Configuration: Sheet name and target columns ===
SHEET_NAME = "finma_content"
COLUMN_PREFIXES = ["description"]  # Columns starting with these prefixes will be processed
SPECIFIC_COLUMNS = []              # Specific column names to process

# === Regex to detect patterns ===
# Matches:
# 1. A word (letters only) followed by a number (e.g., run20)
# 2. A phrase in parentheses followed by a number (e.g., (keep running)7)
REGEX = re.compile(r"([^\W\d_]+)(\d+)(?!\w)|(\([^)]+\))(\d+)(?!\w)", re.UNICODE)


def process_cell_text(text: str) -> Tuple[str, List[Tuple[str, str]]]:
    """
    Apply transformation rules to cell text and return modified text plus list of changes.

    Examples:
        "run20"             → "run[20]"
        "(keep running)7"   → "(keep running)[7]"
        "light35."          → "light[35]."

    Args:
        text (str): The original string in the cell.

    Returns:
        (new_text, changes) where changes is a list of tuples (original_word, replaced_word)
    """
    
    changes = []

    def replacer(match):
        if match.group(1) and match.group(2):
            original = f"{match.group(1)}{match.group(2)}"
            replaced = f"{match.group(1)}[{match.group(2)}]"
        elif match.group(3) and match.group(4):
            original = f"{match.group(3)}{match.group(4)}"
            replaced = f"{match.group(3)}[{match.group(4)}]"
        else:
            return match.group(0)

        changes.append((original, replaced))
        return replaced

    new_text = REGEX.sub(replacer, text)
    
    return new_text, changes



def process_excel_file(file_path: str, verbose: bool = False):
    """
    Open the Excel file and process the target columns in the configured sheet.
    Modifies matching text patterns and saves the result to a new file.

    Args:
        file_path (str): Path to the .xlsx file to be processed.
        verbose (bool): If True, logs every change made to the console.
    """
    
    wb = load_workbook(file_path)

    # Ensure the target sheet exists
    if SHEET_NAME not in wb.sheetnames:
        print(f"❌ [ERROR] Sheet \"{SHEET_NAME}\" not found")
        sys.exit(1)

    ws = wb[SHEET_NAME]

    # Read header row to identify relevant columns
    header = [cell.value for cell in ws[1]]
    target_col_indices = []

    # Find columns that match prefix or are explicitly listed
    for idx, col_name in enumerate(header):
        if col_name is None:
            continue
        if any(col_name.startswith(prefix) for prefix in COLUMN_PREFIXES) or col_name in SPECIFIC_COLUMNS:
            target_col_indices.append(idx)

    if not target_col_indices:
        print("❌ [ERROR] No matching columns found")
        sys.exit(1)
        
    
    total_replacements = 0

    # Loop over each data row (starting from row 2)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for idx in target_col_indices:
            cell: Cell = row[idx]
            if cell.value and isinstance(cell.value, str):
                new_value, changes = process_cell_text(cell.value)
                if changes:
                    total_replacements += len(changes)
                    if verbose:
                        column_name = header[idx]
                        # Only shows modified words in cell
                        changes_str = "; ".join([f"\"{orig}\" -> \"{rep}\"" for orig, rep in changes])
                        print(f"ℹ️  [verbose] Row #{row_idx}, Column \"{column_name}\": {changes_str}")
                    cell.value = new_value

    # Save the modified Excel file with "_modified" suffix
    base, ext = os.path.splitext(file_path)
    new_file = f"{base}_modified{ext}"
    wb.save(new_file)
    print(f"✅ File saved as: \"{new_file}\"")
    print(f"ℹ️  Total replacements made: {total_replacements}")

if __name__ == "__main__":
    # Parse arguments
    if len(sys.argv) < 2:
        print("Usage: python script.py <excel_file.xlsx> [--verbose]")
        sys.exit(1)

    excel_file = sys.argv[1]
    verbose_flag = "--verbose" in sys.argv
    
    # Verify file exists
    if not os.path.isfile(excel_file):
        print(f"❌ [ERROR] File \"{excel_file}\" not found")
        sys.exit(1)

    # Run the processing function
    process_excel_file(excel_file, verbose=verbose_flag)
