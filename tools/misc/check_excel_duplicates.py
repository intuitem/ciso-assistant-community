"""Find duplicate values in one or more columns of an Excel worksheet.

Columns can be provided as Excel column letters (for example `A,C`) or as
exact header names from the selected header row (for example `ref_id,name`).
Empty cells are ignored.
"""

from __future__ import annotations

import argparse
from collections import defaultdict
from pathlib import Path
from typing import Hashable, Iterable

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter


EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}


def parse_columns(raw_columns: str) -> list[str]:
    """Return the comma-separated column references supplied by the user."""
    columns = [column.strip() for column in raw_columns.split(",")]
    if not columns or any(not column for column in columns):
        raise argparse.ArgumentTypeError(
            "Columns must be a comma-separated list without empty entries."
        )
    return columns


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Find duplicate values in selected columns of an Excel sheet."
    )
    parser.add_argument(
        "excel_file",
        type=Path,
        help="Excel workbook to inspect (.xlsx or .xlsm).",
    )
    parser.add_argument(
        "-s",
        "--sheet",
        required=True,
        help="Name of the worksheet to inspect.",
    )
    parser.add_argument(
        "-c",
        "--columns",
        required=True,
        type=parse_columns,
        help=(
            "Comma-separated Excel column letters or exact header names "
            '(examples: "A,C" or "ref_id,name").'
        ),
    )
    output_mode = parser.add_mutually_exclusive_group()
    output_mode.add_argument(
        "-n",
        "--non-consecutive",
        action="store_true",
        help=(
            "Only show duplicate values whose occurrences are split into "
            "at least two non-consecutive groups."
        ),
    )
    output_mode.add_argument(
        "-u",
        "--unique-values",
        action="store_true",
        help=(
            "List each distinct value once, followed by its occurrence count "
            "in braces when it appears more than once."
        ),
    )
    parser.add_argument(
        "--header-row",
        type=int,
        default=1,
        help="Row containing column headers (default: 1).",
    )
    parser.add_argument(
        "--start-row",
        type=int,
        help=(
            "First row to inspect. By default, named columns start after the "
            "header row and letter-based columns start at row 1."
        ),
    )
    args = parser.parse_args()

    if args.header_row < 1:
        parser.error("--header-row must be greater than or equal to 1.")
    if args.start_row is not None and args.start_row < 1:
        parser.error("--start-row must be greater than or equal to 1.")

    return args


def is_excel_column_letter(reference: str) -> bool:
    """Return whether *reference* is a valid Excel column letter."""
    if not reference.isalpha():
        return False
    try:
        column_index_from_string(reference.upper())
    except ValueError:
        return False
    return True


def resolve_column(
    worksheet: object,
    reference: str,
    header_row: int,
) -> tuple[int, str, bool]:
    """Resolve a column letter or exact header name to an Excel column index."""
    if is_excel_column_letter(reference):
        column_index = column_index_from_string(reference.upper())
        return column_index, reference.upper(), False

    matches = [
        cell.column
        for cell in worksheet[header_row]
        if str(cell.value).strip() == reference
    ]
    if not matches:
        raise ValueError(
            f'Header "{reference}" was not found on row {header_row}.'
        )
    if len(matches) > 1:
        match_letters = ", ".join(get_column_letter(index) for index in matches)
        raise ValueError(
            f'Header "{reference}" appears more than once on row {header_row} '
            f"(columns {match_letters}). Use a column letter instead."
        )

    column_index = matches[0]
    return column_index, f'{get_column_letter(column_index)} ("{reference}")', True


def collect_values(
    worksheet: object,
    column_index: int,
    start_row: int,
) -> list[tuple[Hashable, list[int]]]:
    """Return non-empty cell values and their worksheet row numbers."""
    rows_by_value: defaultdict[Hashable, list[int]] = defaultdict(list)

    for row_number in range(start_row, worksheet.max_row + 1):
        value = worksheet.cell(row=row_number, column=column_index).value
        if value is None:
            continue
        if isinstance(value, str):
            value = value.strip()
            if not value:
                continue
        rows_by_value[value].append(row_number)

    return list(rows_by_value.items())


def find_duplicates(
    values: Iterable[tuple[Hashable, list[int]]],
    non_consecutive_only: bool,
) -> list[tuple[Hashable, list[int]]]:
    """Filter collected values to the requested kind of duplicates."""
    duplicates = []
    for value, row_numbers in values:
        if len(row_numbers) < 2:
            continue
        if non_consecutive_only and all(
            current == previous + 1
            for previous, current in zip(row_numbers, row_numbers[1:])
        ):
            continue
        duplicates.append((value, row_numbers))

    return duplicates


def format_value(value: object) -> str:
    """Return a clear, single-line representation of a cell value."""
    return repr(value).replace("\n", "\\n").replace("\r", "\\r")


def print_report(
    results: Iterable[tuple[str, list[tuple[Hashable, list[int]]]]],
    non_consecutive_only: bool,
    unique_values: bool,
) -> int:
    """Print the selected report and return the number of reported values."""
    reported_count = 0
    mode = "non-consecutive duplicates" if non_consecutive_only else "duplicates"

    for column_label, values in results:
        print(f"\nColumn {column_label}:")
        if not values:
            if unique_values:
                print("  No values found.")
            else:
                print(f"  No {mode} found.")
            continue

        for value, row_numbers in values:
            if unique_values:
                occurrence_count = len(row_numbers)
                suffix = f" {{{occurrence_count}}}" if occurrence_count > 1 else ""
                print(f"  - {format_value(value)}{suffix}")
            else:
                rows = ", ".join(str(row_number) for row_number in row_numbers)
                print(f"  - {format_value(value)}: rows {rows}")
            reported_count += 1

    if unique_values:
        print(f"\nDistinct values listed: {reported_count}")
    else:
        print(f"\nDuplicate values reported: {reported_count}")
    return reported_count


def main() -> int:
    args = parse_args()
    excel_file = args.excel_file.resolve()

    if not excel_file.is_file():
        print(f"Error: Excel file not found: {excel_file}")
        return 2
    if excel_file.suffix.lower() not in EXCEL_EXTENSIONS:
        print("Error: only .xlsx and .xlsm workbooks are supported.")
        return 2

    keep_vba = excel_file.suffix.lower() == ".xlsm"
    try:
        workbook = load_workbook(
            excel_file,
            read_only=True,
            data_only=True,
            keep_links=False,
            keep_vba=keep_vba,
        )
    except Exception as error:
        print(f"Error: unable to open {excel_file}: {error}")
        return 2

    try:
        if args.sheet not in workbook.sheetnames:
            available_sheets = ", ".join(workbook.sheetnames)
            print(
                f'Error: sheet "{args.sheet}" was not found. '
                f"Available sheets: {available_sheets}"
            )
            return 2

        worksheet = workbook[args.sheet]
        results = []
        for reference in args.columns:
            try:
                column_index, column_label, uses_header = resolve_column(
                    worksheet,
                    reference,
                    args.header_row,
                )
            except ValueError as error:
                print(f"Error: {error}")
                return 2

            start_row = args.start_row
            if start_row is None:
                start_row = args.header_row + 1 if uses_header else 1

            values = collect_values(
                worksheet,
                column_index,
                start_row,
            )
            if not args.unique_values:
                values = find_duplicates(values, args.non_consecutive)
            results.append((column_label, values))

        print(f"Workbook: {excel_file}")
        print(f"Sheet: {args.sheet}")
        print_report(results, args.non_consecutive, args.unique_values)
        return 0
    finally:
        workbook.close()


if __name__ == "__main__":
    raise SystemExit(main())
