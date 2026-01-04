"""
YAML to Excel Mapping Converter v0.4

This script converts audit mapping data from a YAML file into a formatted Excel file.

Features:
- Supports both v1 and v2 YAML mapping formats.
- Extracts the following fields:
    source_node_id, target_node_id, relationship,
    rationale (optional), strength_of_relationship (optional), annotation (optional).
- For v2 YAMLs, allows export of:
    - Mapping A to B,
    - Mapping B to A (reversed),
    - Or both (each in a separate sheet).
- Displays console warnings for missing required fields with YAML line numbers.
- Highlights missing required fields in red in the Excel output.
- Creates:
    - an "info" sheet with metadata,
    - a "warning" sheet listing all mapping issues (sheet, row, YAML line, missing fields).

Note: This script is a diagnostic/export tool and does not replace prepare_mapping.py.

Usage:
    python yaml_to_excel.py <yaml_file> [excel_file]

Example:
    python yaml_to_excel.py audit_mapping.yaml
"""

from ruamel.yaml import YAML
import pandas as pd
import sys
import os
from openpyxl.styles import Font, PatternFill

SCRIPT_VERSION = '0.4'

def extract_last_segment(urn: str) -> str:
    return urn.split(':')[-1] if urn else ''

def get_requirement_mappings(data: dict) -> tuple[list[dict], int]:
    objects = data.get("objects")
    if not objects:
        print("❌ [ERROR] Missing 'objects' in YAML.")
        sys.exit(1)

    # Version 1
    if "requirement_mapping_set" in objects:
        rms = objects["requirement_mapping_set"]
        if "requirement_mappings" not in rms:
            print("❌ [ERROR] Missing 'requirement_mappings' in 'requirement_mapping_set'.")
            sys.exit(1)
        return rms["requirement_mappings"], -1

    # Version 2
    elif "requirement_mapping_sets" in objects:
        sets = objects["requirement_mapping_sets"]
        if not isinstance(sets, list) or not sets:
            print("❌ [ERROR] 'requirement_mapping_sets' should be a non-empty list.")
            sys.exit(1)

        while True:
            choice = input("ℹ️  Version 2 detected. Choose which mapping to export:\n"
                           "1 - Direct mapping (A to B)\n"
                           "2 - Revert mapping (B to A)\n"
                           "3 - Both mappings (A to B & B to A)\n"
                           "> Enter 1, 2 or 3: ").strip()
            if choice in {"1", "2", "3"}:
                break
            print("❌ [ERROR] Invalid input. Please enter 1, 2 or 3.")

        if choice == "3":
            if len(sets) < 2:
                print("❌ [ERROR] Cannot export both mappings: only one mapping set found.")
                sys.exit(1)
            for idx, s in enumerate(sets[:2]):
                if "requirement_mappings" not in s:
                    print(f"❌ [ERROR] Missing 'requirement_mappings' in mapping set {idx}.")
                    sys.exit(1)
            return sets, 3
        else:
            idx = int(choice) - 1
            if idx >= len(sets):
                print(f"❌ [ERROR] Mapping index {idx} not found in 'requirement_mapping_sets'.")
                sys.exit(1)
            if "requirement_mappings" not in sets[idx]:
                print(f"❌ [ERROR] Missing 'requirement_mappings' in selected mapping set.")
                sys.exit(1)
            return sets[idx]["requirement_mappings"], idx

    else:
        print("❌ [ERROR] Neither 'requirement_mapping_set' nor 'requirement_mapping_sets' found.")
        sys.exit(1)

def convert_yaml_to_excel(yaml_file: str, mappings_data, excel_file: str, mapping_index: int):
    warnings = []
    missing_cell_positions = {}  # key = sheet_name, value = list of (col, row)

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    field_to_column = {
        "source_node_id": "A",
        "target_node_id": "B",
        "relationship": "C"
    }

    def process_mapping(mappings, mapping_name):
        rows = []
        missing_cell_positions[mapping_name] = []
        
        for i, item in enumerate(mappings):
            line_number = item.lc.line if hasattr(item, 'lc') else 'N/A'
            source = extract_last_segment(item.get("source_requirement_urn", ""))
            target = extract_last_segment(item.get("target_requirement_urn", ""))
            relationship = item.get("relationship", "")

            # Collect warning if essential fields are missing
            missing_fields = []
            if not source:
                missing_fields.append("source_node_id")
            if not target:
                missing_fields.append("target_node_id")
            if not relationship:
                missing_fields.append("relationship")
            if missing_fields:
                excel_row = i + 2  # +2 for header + 1-indexed
                warning_entry = {
                    "Mapping sheet": mapping_name,
                    "Row number": excel_row,
                    "YAML line number": line_number + 1 if isinstance(line_number, int) else "N/A",
                    "Missing fields": ", ".join(missing_fields)
                }

                warnings.append(warning_entry)
                print(f"⚠️  [WARNING] Missing field(s) in sheet \"{mapping_name}\" (Excel row {excel_row}, YAML line {warning_entry['YAML line number']}) → {warning_entry['Missing fields']}")
                
                # Mark the cells to be colored
                for field in missing_fields:
                    col_letter = field_to_column.get(field)
                    if col_letter:
                        missing_cell_positions[mapping_name].append((col_letter, excel_row))
                
            row = {
                "source_node_id": source,
                "target_node_id": target,
                "relationship": relationship,
                "rationale": item.get("rationale", ""),
                "strength_of_relationship": item.get("strength_of_relationship", ""),
                "annotation": item.get("annotation", "")
            }
            rows.append(row)
        return rows

    if mapping_index == 3:
        sheet_data = [
            ("mappings", process_mapping(mappings_data[0].get("requirement_mappings", []), "mappings")),
            ("mappings_revert", process_mapping(mappings_data[1].get("requirement_mappings", []), "mappings_revert"))
        ]
    else:
        sheet_name = "mappings" if mapping_index in (-1, 0) else "mappings_revert"
        sheet_data = [(sheet_name, process_mapping(mappings_data, sheet_name))]

    try:
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            
            # Info sheet
            workbook = writer.book
            info_sheet = workbook.create_sheet(title="info", index=0)
            info_sheet["A1"] = "YAML to Excel Mapping Converter v" + SCRIPT_VERSION
            info_sheet["A1"].font = Font(size=48, bold=True)
            info_sheet["A2"] = "Source file : " + os.path.basename(yaml_file)
            info_sheet["A2"].font = Font(size=15)
            info_sheet["A3"] = "Please note that this Excel file doesn't replace the one generated by prepare_mapping.py."
            info_sheet["A3"].font = Font(size=20, italic=True)

            # Mapping sheets
            for sheet_name, rows in sheet_data:
                df = pd.DataFrame(rows)[[
                    "source_node_id",
                    "target_node_id",
                    "relationship",
                    "rationale",
                    "strength_of_relationship",
                    "annotation"
                ]]
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Color missing fields
                worksheet = writer.sheets[sheet_name]
                for col_letter, row_num in missing_cell_positions.get(sheet_name, []):
                    worksheet[f"{col_letter}{row_num}"].fill = red_fill

            # Warning sheet
            if warnings:
                warning_df = pd.DataFrame(warnings)
                warning_df.columns = ["Mapping sheet", "Row number", "YAML line number", "Missing fields"]
                warning_df.to_excel(writer, sheet_name="warning", index=False)
                
    except Exception as e:
        print("❌ [ERROR] Failed to save the Excel file.")
        print(f"→  Reason: {e}")
        print("💡 Tip: Make sure the file isn't already open in Excel or another program.")
        sys.exit(1)
        
    print(f"✅ Conversion completed: \"{os.path.basename(excel_file)}\"")
    if warnings:
        print(f"⚠️  {len(warnings)} warnings found. See the \"warning\" sheet in the output file.")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python yaml_to_excel.py <yaml_file> [excel_file]")
        sys.exit(1)

    yaml_input = sys.argv[1]

    # Check if file exists
    if not os.path.isfile(yaml_input):
        print(f"❌ [ERROR] File not found: \"{yaml_input}\"")
        sys.exit(1)

    # Load YAML file
    print(f"⌛ Loading \"{os.path.basename(yaml_input)}\"...")
    
    yaml_parser = YAML(typ="rt")
    yaml_parser.allow_duplicate_keys = False
    
    with open(yaml_input, 'r', encoding='utf-8') as f:
        data = yaml_parser.load(f)
        
    print(f"✅ \"{os.path.basename(yaml_input)}\" loaded!")

    mappings_data, mapping_index = get_requirement_mappings(data)

    # Determine output filename
    if len(sys.argv) > 2:
        excel_output = sys.argv[2]
        if not excel_output.lower().endswith(".xlsx"):
            excel_output += ".xlsx"
    else:
        yaml_basename = os.path.splitext(os.path.basename(yaml_input))[0]
        suffix = ""
        if mapping_index == 1:
            suffix = "_revert"
        elif mapping_index == 3:
            suffix = "_both"
        excel_output = os.path.join(os.getcwd(), f"conv_{yaml_basename}{suffix}.xlsx")

    convert_yaml_to_excel(yaml_input, mappings_data, excel_output, mapping_index)
