"""
Mapping Expander v0.3

This script expands a source Excel mapping file using a reference Excel file and produces a new Excel file
with the expanded mappings.

Functionality:
--------------
- Reads a mapping table from a source Excel file (sheet and columns are configurable).
- For each row, takes the "target_node_id" prefix and finds all "node_id" values in the reference file
  that start with this prefix.
- With the --reverse flag, expands using "source_node_id" instead.
- For each match found, it creates a new row with the original "source_node_id" and the full "node_id" from the reference.
- The output is written to a destination Excel file, with the following sheets:
    * "info": metadata and context about the mapping.
    * "mappings": the expanded result table.
    * "warnings": lists any source entries that did not match anything in the reference.
    * "exceptions": lists all applied prefix remappings (see below).

Exceptions System:
------------------
- A framework-specific exception system is supported via an external JSON file.
- This JSON file maps special "target_node_id" values to custom prefixes based on the "framework_exception" setting.
- Example use case: if `framework_exception = "soc2_rev2022"` and `target_node_id == "cc6.5"`, 
  the script can be configured (via JSON) to match against "c6.5" instead.
- The path to this JSON file is set via the "exceptions_json_path" variable in the script.
- This system is useful when we want to force a search for a specific node ID in case of an error in a framework, for example

Usage:
------
1. Configure the paths and sheet/column names at the top of the script.
2. Set the "framework_exception" if needed (optional).
3. Provide the path to the exceptions JSON file.
4. Run the script to generate the expanded Excel file.
5. Use "--reverse" to expand using "source_node_id" instead of "target_node_id".

Note:
-----
This tool is intended as a utility to support data expansion. It does not replace the output
of "prepare_mapping.py", and should be used as a complement.
"""


SCRIPT_VERSION = '0.3'

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, numbers
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import json
import argparse


# === Configuration variables ===
current_script_path = os.path.dirname(os.path.abspath(__file__))
source_file = "../excel/ccf/part_mapping_adobe-ccf-v5_to_soc2.xlsx"
reference_file = "../excel/ccf/mapping-adobe-ccf-v5-to-soc2-2017-rev-2022_new.xlsx"
destination_file = "../excel/ccf/destination.xlsx"
exceptions_json_path = f"{current_script_path}/mapping_expander_exceptions.json"

source_sheet_name = "mappings"           # Sheet name in the source file
reference_sheet_name = "target"          # Sheet name in the reference file
reference_column_name = "node_id"        # Column name in the reference file to match

# Set to None if no exceptions needed
framework_exception = "soc2_rev2022"


def load_exceptions(path, framework_exception):
    try:
        with open(path, "r", encoding="utf-8") as f:
            exception_map = json.load(f)
            if framework_exception:
                print(f"ℹ️  Exceptions list for \"{framework_exception}\" loaded from JSON file \"{os.path.basename(path)}\"")
            return exception_map
    except Exception as e:
        print(f"⚠️  [WARNING] Could not load exceptions JSON file: {e}")
        return {}


def expand_mappings(df_source, node_id_list, exception_map, reverse=False):
    
    new_rows = []
    warnings_list = []
    exceptions_list = []

    input_column = "source_node_id" if reverse else "target_node_id"
    fixed_column = "target_node_id" if reverse else "source_node_id"

    for index, row in df_source.iterrows():
        fixed_value = row[fixed_column]
        prefix = row[input_column]
        
        # Apply exception rule if defined for the current framework
        if framework_exception in exception_map:
            framework_rules = exception_map[framework_exception]
            search_prefix = framework_rules.get(prefix, prefix)
            if search_prefix != prefix:
                print(f"ℹ️  Exception applied for framework \"{framework_exception}\": \"{prefix}\" mapped to \"{search_prefix}\" ({fixed_column}: \"{fixed_value}\") [row #{index+2}]")
                exceptions_list.append({
                    fixed_column: fixed_value,
                    f"original_{input_column}": prefix,
                    f"mapped_{input_column}": search_prefix
                })
        else:
            search_prefix = prefix

        # Find all node_ids that start with the prefix
        matches = [node_id for node_id in node_id_list if node_id.startswith(search_prefix)]

        if not matches:
            warning_msg = f"No match found in reference for {input_column} \"{prefix}\" ({fixed_column}: \"{fixed_value}\") [row #{index+2}]"
            print(f"⚠️  [WARNING] {warning_msg}")
            warnings_list.append({
                fixed_column: fixed_value,
                input_column: prefix,
                "message": warning_msg
            })

        # Create a new row for each match found
        for full_node_id in matches:
            new_rows.append({
                "source_node_id": full_node_id if reverse else fixed_value,
                "target_node_id": fixed_value if reverse else full_node_id
            })

    return pd.DataFrame(new_rows), warnings_list, exceptions_list


def run_expansion(reverse=False):
    exception_map = load_exceptions(exceptions_json_path, framework_exception)

    df_source = pd.read_excel(source_file, sheet_name=source_sheet_name, dtype=str)
    df_reference = pd.read_excel(reference_file, sheet_name=reference_sheet_name, dtype=str)
    node_id_list = df_reference[reference_column_name].dropna().tolist()

    df_result, warnings_list, exceptions_list = expand_mappings(df_source, node_id_list, exception_map, reverse)

    wb = Workbook()

    # === Sheet: info ===
    ws_info = wb.active
    ws_info.title = "info"
    source_filename = os.path.basename(source_file)
    reference_filename = os.path.basename(reference_file)
    rows_info = [
        {"text": f"Mapping Expander v{SCRIPT_VERSION}", "font": Font(size=48, bold=True)},
        {"text": f"Source file : {source_filename}", "font": Font(size=15)},
        {"text": f"Source sheet : {source_sheet_name}", "font": Font(size=15)},
        {"text": f"Reference file : {reference_filename}", "font": Font(size=15)},
        {"text": f"Reference sheet: {reference_sheet_name} | Ref Column : {reference_column_name}", "font": Font(size=15)},
        {"text": "Please note that this Excel file doesn't replace the one generated by prepare_mapping.py.",
         "font": Font(size=20, italic=True)},
    ]
    for i, row in enumerate(rows_info, start=1):
        cell = ws_info.cell(row=i, column=1, value=row["text"])
        cell.font = row["font"]

    # === Sheet: mappings ===
    ws_mappings = wb.create_sheet(title=source_sheet_name)
    ws_mappings.append(list(df_result.columns))
    for row in dataframe_to_rows(df_result, index=False, header=False):
        ws_mappings.append(row)
    for column in ws_mappings.columns:
        for cell in column:
            cell.number_format = numbers.FORMAT_TEXT  # Force cells as text (avoid auto-formatting of IDs)

    # === Sheet: warnings (if any) ===
    if warnings_list:
        ws_warn = wb.create_sheet(title="warnings")
        ws_warn.append(["source_node_id", "target_node_id", "message"])
        for warn in warnings_list:
            ws_warn.append([
                warn.get("source_node_id", ""),
                warn.get("target_node_id", ""),
                warn["message"]
            ])

    # === Sheet: Exceptions (if any) ===
    if exceptions_list:
        ws_exc = wb.create_sheet(title="exceptions")
        ws_exc.append(["source_node_id", "original_target_node_id", "mapped_target_node_id"])
        for exc in exceptions_list:
            ws_exc.append([
                exc.get("source_node_id", ""),
                exc.get("original_target_node_id", ""),
                exc.get("mapped_target_node_id", "")
            ])

    wb.save(destination_file)

    print("✅ Conversion completed successfully.")
    print(f"➡ Source sheet: \"{source_sheet_name}\"")
    print(f"➡ Reference sheet: \"{reference_sheet_name}\", column: \"{reference_column_name}\"")
    print(f"📁 Output file: \"{destination_file}\"")
    if warnings_list:
        print(f"⚠️  {len(warnings_list)} warnings found. See the \"warnings\" sheet in the output file.")
    if exceptions_list:
        print(f"ℹ️  {len(exceptions_list)} exceptions applied. See the \"exceptions\" sheet in the output file.")


def main():
    parser = argparse.ArgumentParser(description="Expand mapping by target_node_id (default) or source_node_id (with --reverse).")
    parser.add_argument('--reverse', action='store_true', help='Expand using source_node_id instead of target_node_id')
    args = parser.parse_args()

    run_expansion(reverse=args.reverse)


if __name__ == "__main__":
    main()
