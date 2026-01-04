"""
Bad Mapping Remover v0.1

This script cleans a mapping Excel file by removing invalid mapping lines.
It checks whether the node IDs referenced in the mapping sheet actually exist
in the 'source' and 'target' sheets.

Usage:
    python bad_mapping_remover.py <input_excel_file> [mapping_sheet_name]

Arguments:
    - input_excel_file        Path to the Excel file to process.
    - mapping_sheet_name      Optional name of the sheet containing mappings (default: "mappings_content").

Requirements:
    - The Excel file must contain at least three sheets: "source", "target", and the mapping sheet.
    - The "source" and "target" sheets must have a column named "node_id".
    - The mapping sheet must have columns "source_node_id" and "target_node_id".

Functionality:
    - Removes rows from the mapping sheet where "source_node_id" or "target_node_id" do not exist
      in their respective sheets.
    - Outputs a summary of removed IDs and duplicates.
    - Saves a new Excel file with "_filtered" appended to the original filename.

Example:
    python bad_mapping_remover.py my_mapping.xlsx
    python bad_mapping_remover.py my_mapping.xlsx my_custom_sheet
"""


import openpyxl
import sys
from collections import Counter
from pathlib import Path


def load_node_ids(sheet):
    header = [cell.value for cell in sheet[1]]
    if "node_id" not in header:
        return set(), False
    idx = header.index("node_id")
    ids = set()
    for row in sheet.iter_rows(min_row=2):
        if idx < len(row) and row[idx].value:
            ids.add(str(row[idx].value).strip())
    return ids, True


def main(file_path, mapping_sheet_name="mappings_content"):
    input_path = Path(file_path)
    if not input_path.exists():
        print(f"âŒ [ERROR] File not found: \"{file_path}\"")
        sys.exit(1)

    wb = openpyxl.load_workbook(file_path)
    sheets = wb.sheetnames

    if "source" not in sheets:
        print("âŒ [ERROR] Sheet \"source\" not found")
        sys.exit(1)
    if "target" not in sheets:
        print("âŒ [ERROR] Sheet \"target\" not found")
        sys.exit(1)

    source_ids, source_ok = load_node_ids(wb["source"])
    if not source_ok:
        print("âŒ [ERROR] \"node_id\" column not found in \"source\" sheet header")
        sys.exit(1)

    target_ids, target_ok = load_node_ids(wb["target"])
    if not target_ok:
        print("âŒ [ERROR] \"node_id\" column not found in \"target\" sheet header")
        sys.exit(1)

    if mapping_sheet_name not in sheets:
        print(f"âŒ [ERROR] Sheet \"{mapping_sheet_name}\" not found")
        sys.exit(1)

    sheet = wb[mapping_sheet_name]
    header = [cell.value for cell in sheet[1]]
    rows = list(sheet.iter_rows(min_row=2))

    if "source_node_id" not in header or "target_node_id" not in header:
        print("âŒ [ERROR] Missing \"source_node_id\" or \"target_node_id\" column in mapping sheet")
        sys.exit(1)

    src_idx = header.index("source_node_id")
    tgt_idx = header.index("target_node_id")

    removed_source_count = Counter()
    removed_target_count = Counter()
    kept_rows = []
    amount_removed_lines = 0

    for row in rows:
        source_val = str(row[src_idx].value).strip() if row[src_idx].value else ""
        target_val = str(row[tgt_idx].value).strip() if row[tgt_idx].value else ""

        missing_source = source_val not in source_ids
        missing_target = target_val not in target_ids

        if missing_source:
            removed_source_count[source_val] += 1
        if missing_target:
            removed_target_count[target_val] += 1
            
        if missing_source or missing_target:
            amount_removed_lines += 1

        if not missing_source and not missing_target:
            kept_rows.append([cell.value for cell in row])

    # Output report
    for sid in removed_source_count:
        print(f"ðŸ—‘ï¸  [REMOVED] source_node_id \"{sid}\" not found in sheet \"source\"")
    for tid in removed_target_count:
        print(f"ðŸ—‘ï¸  [REMOVED] target_node_id \"{tid}\" not found in sheet \"target\"")

    for sid, count in removed_source_count.items():
        if count > 1:
            print(f"ðŸ” [DUPLICATE] source_node_id \"{sid}\" was removed {count} times from mappings")
    for tid, count in removed_target_count.items():
        if count > 1:
            print(f"ðŸ” [DUPLICATE] target_node_id \"{tid}\" was removed {count} times from mappings")

    # total_removed = sum(removed_source_count.values()) + sum(removed_target_count.values())
    if amount_removed_lines > 0:
        print(f"ðŸ“œ [SUMMARY] Removed {amount_removed_lines} mapping(s) due to missing node IDs\
              \n             - Missing source: {sum(removed_source_count.values())}, target: {sum(removed_target_count.values())} -")

    # Save cleaned Excel file
    new_wb = openpyxl.Workbook()
    new_wb.remove(new_wb.active)
    for sheet_name in wb.sheetnames:
        orig_sheet = wb[sheet_name]
        new_sheet = new_wb.create_sheet(title=sheet_name)
        for i, row in enumerate(orig_sheet.iter_rows(values_only=True), start=1):
            if sheet_name == mapping_sheet_name and i > 1:
                if i - 2 < len(kept_rows):
                    new_sheet.append(kept_rows[i - 2])
            else:
                new_sheet.append(row)

    output_file = input_path.with_stem(input_path.stem + "_filtered")
    new_wb.save(str(output_file))
    print(f"âœ… Cleaned Excel file saved as: \"{output_file.name}\"")


if __name__ == "__main__":
    if len(sys.argv) < 2 or len(sys.argv) > 3:
        print("Usage: python bad_mapping_remover.py <input_excel_file> [sheet_name]")
        sys.exit(1)
    elif len(sys.argv) == 2:
        main(sys.argv[1])
    elif len(sys.argv) == 3:
        main(sys.argv[1], sys.argv[2])
