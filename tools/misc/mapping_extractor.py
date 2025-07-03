"""
Mapping Extractor v0.2

Generic script to extract pairwise mappings from an Excel file into a structured format,
configured through a YAML file.

This script is designed to extract mappings between a source column (typically containing
internal control references) and one or more target columns (typically containing external
framework references) from a structured Excel file.

Configuration:
- A YAML file must be passed to the script via --config (or it defaults to mapping_extractor_config.yaml).
- The YAML file must contain:
    - source_file            : Path to the source Excel file
    - source_sheet           : Name of the worksheet to read
    - source_id_column       : Name of the column used for source identifiers
    - target_column_names    : List of column names to extract as target mappings
  Optional keys in the YAML:
    - rows_to_ignore         : List of 1-based row indices to skip during processing (default: None)
    - destination_file       : Name of the output Excel file (default: part_mapping.xlsx)
    - destination_sheet      : Name of the output sheet (default: mappings)
    - header_row             : 1-based row number containing column headers in the Excel sheet (default: 1)

Output:
- An Excel file with three sheets:
  1. "info": a summary sheet with metadata about the extraction (script version, file names, etc.).
  2. "mappings": a normalized list of unique source-target pairs:
     - source_node_id: normalized source ID (lowercase, spaces to dashes)
     - target_node_id: one entry per target
     - Col Name: shown only for the first entry of a target column group
  3. "warnings": a sheet listing any issues encountered during validation, such as missing or duplicated columns.
- Duplicate mappings are removed automatically.
"""


import pandas as pd
import re
import os
import argparse
import yaml
from collections import Counter
from openpyxl import load_workbook
from openpyxl.styles import Font


# === Script metadata ===
SCRIPT_NAME = "Mapping Extractor"
SCRIPT_VERSION = "v0.2"



def generate_mapping(source_file: str,
                     source_sheet: str,
                     source_id_column: str,
                     target_column_names: list,
                     rows_to_ignore: list = None,
                     destination_file: str = "extracted_mapping.xlsx",
                     destination_sheet: str = "mappings",
                     header_row: int = 1):
    """
    Extracts and saves the mapping between a source column and multiple target columns.

    Raises RuntimeError or ValueError in case of issues during processing.
    """
    
    # Validation
    if not isinstance(target_column_names, list) or not target_column_names:
        raise ValueError("\"target_column_names\" must be a non-empty list.")
    if rows_to_ignore is not None and not isinstance(rows_to_ignore, list):
        raise ValueError("\"rows_to_ignore\" must be a list if provided.")

    # Try to read the source Excel sheet
    try:
        df = pd.read_excel(source_file, sheet_name=source_sheet, header=header_row - 1, dtype=str)
    except Exception as e:
        raise RuntimeError(f"Failed to read Excel file \"{source_file}\" or sheet \"{source_sheet}\":\n\t\t {e}")

    # Drop user-specified rows to ignore (convert to 0-based index)
    if rows_to_ignore:
        df = df.drop(index=[i - 1 for i in rows_to_ignore], errors="ignore")

    results = []
    seen_pairs = set()  # To avoid duplicates
    warnings_list = []


    # Securely check raw Excel column headers using openpyxl
    try:
        wb_raw = load_workbook(filename=source_file, read_only=True, data_only=True)
    except Exception as e:
        raise RuntimeError(f"Unable to open Excel file with openpyxl: \"{source_file}\"\n\t\t {e}")

    if source_sheet not in wb_raw.sheetnames:
        raise ValueError(f"The sheet \"{source_sheet}\" does not exist in the Excel file.")

    try:
        ws_raw = wb_raw[source_sheet]
        header_cells = list(ws_raw.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
    except Exception as e:
        raise RuntimeError(f"Failed to read header row ({header_row}) in sheet \"{source_sheet}\":\n\t\t {e}")

    raw_header_counts = Counter(header_cells)


    # Iterate over each target column
    for column_name in target_column_names:
        
        # Check number of columns with same name
        count = raw_header_counts.get(column_name, 0)
        
        if count == 0:
            msg = f"Target column \"{column_name}\" not found in the Excel sheet."
            print(f"⚠️  [WARNING] {msg}")
            warnings_list.append(msg)
        elif count > 1:
            msg = f"Column name \"{column_name}\" appears {count} times in the sheet. The exported column may not be the one expected."
            print(f"⚠️  [WARNING] {msg}")
            print("💡 Tip: Consider renaming the column in Excel or specifying a unique name in the YAML.")
            warnings_list.append(msg)
        
        
        first_entry = True  # Used to insert "Col Name" only on the first line per column
        for _, row in df.iterrows():
            # Normalize source ID
            source_id = str(row.get(source_id_column, "")).strip().lower().replace(" ", "-")
            target_raw = row.get(column_name, "")

            # Skip empty cells
            if pd.isna(target_raw) or not str(target_raw).strip():
                continue

            # Split multi-valued cells using comma or newline
            refs = [
                str(ref).strip().lower()
                for ref in re.split(r'[\n,]+', str(target_raw))
                if str(ref).strip()
            ]

            for i, ref in enumerate(refs):
                pair = (source_id, ref)
                if pair in seen_pairs:
                    continue
                seen_pairs.add(pair)

                row_data = {
                    "source_node_id": source_id,
                    "target_node_id": ref
                }
                if first_entry:
                    row_data["Col Name"] = column_name
                    first_entry = False
                results.append(row_data)

    df_result = pd.DataFrame(results, dtype=str)

    # Write results to Excel with two sheets: info + mappings
    with pd.ExcelWriter(destination_file, engine="openpyxl") as writer:
        pd.DataFrame().to_excel(writer, sheet_name="info", index=False)
        df_result.to_excel(writer, sheet_name=destination_sheet, index=False)

    # Add metadata to the "info" sheet
    wb = load_workbook(destination_file)
    ws = wb["info"]

    rows_info = [
        {"text": f"{SCRIPT_NAME} {SCRIPT_VERSION}", "font": Font(size=48, bold=True)},
        {"text": f"Source file : {os.path.basename(source_file)}", "font": Font(size=15)},
        {"text": f"Source sheet : {source_sheet}", "font": Font(size=15)},
        {"text": f"Source ID column : {source_id_column}", "font": Font(size=15)},
        {"text": f"Target column name{'s' if len(target_column_names) > 1 else ''} : {' ; '.join(target_column_names)}", "font": Font(size=15)},
        {"text": f"Ignored rows : {', '.join(map(str, rows_to_ignore)) if rows_to_ignore else 'None'}", "font": Font(size=15)},
        {"text": "Please note that this Excel file doesn't replace the one generated by prepare_mapping.py.",
         "font": Font(size=20, italic=True)},
    ]

    for i, row in enumerate(rows_info, start=1):
        cell = ws.cell(row=i, column=1)
        cell.value = row["text"]
        cell.font = row["font"]


    # Add warnings (if any) to the "warnings" sheet.
    if warnings_list:
        ws_warn = wb.create_sheet("warnings")
        ws_warn.cell(row=1, column=1, value="Warnings during extraction:")
        for i, msg in enumerate(warnings_list, start=2):
            ws_warn.cell(row=i, column=1, value=msg)


    wb.save(destination_file)
    print(f"✅ Extraction complete: \"{destination_file}\"")


def load_config(config_path: str) -> dict:
    """
    Load and validate the YAML config file.

    Raises FileNotFoundError, ValueError or yaml.YAMLError.
    """
    
    try:
        with open(config_path, "r", encoding="utf-8") as f:
            config = yaml.safe_load(f)
    except FileNotFoundError:
        raise FileNotFoundError(f"Config file not found: {config_path}")
    except yaml.YAMLError as e:
        raise ValueError(f"YAML parsing error: {e}")

    # Required fields
    required_fields = ["source_file", "source_sheet", "source_id_column", "target_column_names"]
    for field in required_fields:
        if field not in config or not config[field]:
            raise ValueError(f"Missing or empty required field in config: \"{field}\"")

    if not isinstance(config["target_column_names"], list) or not config["target_column_names"]:
        raise ValueError("\"target_column_names\" must be a non-empty list.")

    # Optional fields
    if "rows_to_ignore" in config and not isinstance(config["rows_to_ignore"], list):
        raise ValueError("\"rows_to_ignore\" must be a list if provided.")
    
    if "header_row" in config and (not isinstance(config["header_row"], int) or config["header_row"] < 1):
        raise ValueError("\"header_row\" must be a positive integer if specified.")

    return config


def main():
    """
    Main CLI entrypoint: parse YAML config and run the mapping exporter.
    """
    
    current_script_path = os.path.dirname(os.path.abspath(__file__))

    parser = argparse.ArgumentParser(description="Run the Simple Mapping Exporter with a YAML config.")
    parser.add_argument("--config", type=str, default=f"{current_script_path}/mapping_extractor_config.yaml",
                        help="Path to the YAML config file (default: mapping_extractor_config.yaml)")

    args = parser.parse_args()

    try:
        config = load_config(args.config)
        generate_mapping(
            source_file=config["source_file"],
            source_sheet=config["source_sheet"],
            source_id_column=config["source_id_column"],
            target_column_names=config["target_column_names"],
            rows_to_ignore=config.get("rows_to_ignore", []),
            destination_file=config.get("destination_file", "extracted_mapping.xlsx"),
            destination_sheet=config.get("destination_sheet", "mappings"),
            header_row=config.get("header_row", 1)
        )

    except Exception as e:
        print(f"❌ [FATAL ERROR] {e}")


if __name__ == "__main__":
    main()
