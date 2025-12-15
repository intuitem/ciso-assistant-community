#!/usr/bin/env python3
"""
Create SCF to SWIFT CSF v2023 mapping from SCF Excel data.

This script:
1. Loads SWIFT CSF v2023 data from SCF Excel (column DD)
2. Keeps natural mapping direction (SCF -> SWIFT)
3. Splits multi-value cells (newline-separated)
4. Uses 'intersect' as the default relationship
5. Outputs both CSV and Excel formats

The mapping direction is: SCF control → SWIFT CSF control
"""

import pandas as pd
from pathlib import Path


def normalize_ref_id(ref: str) -> str:
    """
    Normalize reference ID to lowercase.

    Examples:
        GOV-01 -> gov-01
        2.5A -> 2.5a
    """
    if not ref:
        return ref

    ref = str(ref).strip()
    return ref.lower()


def load_excel_mapping(
    excel_path: Path, sheet_name: str, scf_col: int, target_col: int
) -> pd.DataFrame:
    """
    Load mapping data from SCF Excel file.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Sheet name to read
        scf_col: Column index for SCF # (0-based)
        target_col: Column index for target framework (0-based)

    Returns:
        DataFrame with source_node_id, target_node_id, relationship columns
    """
    print(f"Loading: {excel_path.name}...")
    print(f"  Sheet: {sheet_name}")
    print(f"  SCF column index: {scf_col}")
    print(f"  Target framework column index: {target_col}")

    # Read the Excel file
    df = pd.read_excel(excel_path, sheet_name=sheet_name, header=None)

    # Get column names from first row for reference
    scf_col_name = df.iloc[0, scf_col]
    target_col_name = df.iloc[0, target_col]
    print(f"  SCF column name: {scf_col_name}")
    print(f"  Target column name: {target_col_name}")

    # Skip header row
    df = df.iloc[1:]

    # Extract relevant columns
    scf_ids = df.iloc[:, scf_col]
    target_refs = df.iloc[:, target_col]

    # Build the mapping (SCF -> target framework)
    mappings = []
    for scf_id, target_ref in zip(scf_ids, target_refs):
        # Skip empty target refs
        if pd.isna(target_ref) or str(target_ref).strip() == "":
            continue

        # Skip empty SCF IDs
        if pd.isna(scf_id) or str(scf_id).strip() == "":
            continue

        scf_id_normalized = normalize_ref_id(scf_id)

        # Split multi-value cells (newline-separated)
        refs = str(target_ref).strip().split("\n")
        for ref in refs:
            ref = ref.strip()
            if ref:
                normalized_ref = normalize_ref_id(ref)
                mappings.append(
                    {
                        "source_node_id": scf_id_normalized,
                        "target_node_id": normalized_ref,
                        "relationship": "intersect",
                    }
                )

    print(f"  Total mappings extracted: {len(mappings)}")

    return pd.DataFrame(mappings)


def create_swift_csf_mapping(excel_path: Path, output_csv: Path, output_xlsx: Path):
    """
    Create SCF to SWIFT CSF v2023 mapping.
    """
    print("=" * 70)
    print("CREATING SCF TO SWIFT CSF v2023 MAPPING")
    print("=" * 70)

    # Load data from Excel
    # Column C (index 2) = SCF #
    # Column DD (index 107) = SWIFT CSF v2023
    df = load_excel_mapping(
        excel_path=excel_path, sheet_name="SCF 2025.3.1", scf_col=2, target_col=107
    )

    # Remove duplicates on source-target pairs (keep first occurrence)
    before_dedup = len(df)
    df = df.drop_duplicates(subset=["source_node_id", "target_node_id"], keep="first")
    print(f"  Deduplicated: {before_dedup} -> {len(df)}")

    # Sort by source then target
    df = df.sort_values(["source_node_id", "target_node_id"])

    # Show summary statistics
    print("\n" + "-" * 50)
    print("SUMMARY STATISTICS")
    print("-" * 50)
    print(f"Total mappings: {len(df)}")
    print(f"Unique source nodes (SCF): {df['source_node_id'].nunique()}")
    print(f"Unique target nodes (SWIFT CSF): {df['target_node_id'].nunique()}")

    print("\nRelationship type distribution:")
    print(df["relationship"].value_counts())

    print("\nSample source node IDs:")
    print(df["source_node_id"].head(10).tolist())

    # Save to CSV
    print(f"\nSaving to {output_csv}...")
    df.to_csv(output_csv, index=False)

    # Save to Excel with formatting
    print(f"Saving to {output_xlsx}...")
    save_to_excel(df, output_xlsx)

    print("\n" + "=" * 70)
    print("COMPLETE")
    print("=" * 70)
    print(f"\nGenerated files:")
    print(f"  1. {output_csv}")
    print(f"  2. {output_xlsx}")

    return df


def save_to_excel(relationship_df: pd.DataFrame, output_path: Path):
    """
    Save to Excel with formatting.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        print("Warning: openpyxl not installed, skipping Excel export")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Mapping"

    # Header styling
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    # Write data
    for r_idx, row in enumerate(
        dataframe_to_rows(relationship_df, index=False, header=True), 1
    ):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
            cell.alignment = Alignment(vertical="top")

    # Adjust column widths
    col_widths = [20, 15, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width

    ws.freeze_panes = "A2"

    wb.save(output_path)


def main():
    script_dir = Path(__file__).parent

    # Input file
    excel_path = script_dir / "secure-controls-framework-scf-2025-3-1.xlsx"

    # Output files
    output_csv = script_dir / "swift_csf_scf_mapping.csv"
    output_xlsx = script_dir / "swift_csf_scf_mapping.xlsx"

    # Check input file exists
    if not excel_path.exists():
        print(f"Error: SCF Excel file not found: {excel_path}")
        exit(1)

    # Run the mapping creation
    create_swift_csf_mapping(excel_path, output_csv, output_xlsx)


if __name__ == "__main__":
    main()
