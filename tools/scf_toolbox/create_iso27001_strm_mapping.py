#!/usr/bin/env python3
"""
Combine ISO 27001 and ISO 27002 STRM data into a unified mapping.

This script:
1. Loads ISO 27001 STRM data (clauses 4-10)
2. Loads ISO 27002 STRM data (Annex A controls) and prefixes with 'A.'
3. Combines both into a single mapping file
4. Filters out 'no relationship' entries
5. Outputs both CSV and Excel formats

The mapping direction is: ISO 27001/27002 control → SCF control
"""

import pandas as pd
from pathlib import Path


def get_section_number(fde_ref) -> float:
    """Extract the section number from an FDE reference."""
    if pd.isna(fde_ref) or not fde_ref:
        return 0
    fde_ref = str(fde_ref)
    first_part = fde_ref.split(".")[0].split("(")[0]
    try:
        return float(first_part)
    except ValueError:
        return 0


def strip_parenthetical(ref: str) -> str:
    """
    Remove parenthetical suffixes from a reference.

    Examples:
        5.3(b) -> 5.3
        6.1.1(a) -> 6.1.1
        6.1.1(e)(1) -> 6.1.1
        A.5.1(a) -> A.5.1
    """
    if not ref:
        return ref
    # Find first '(' and strip everything from there
    paren_idx = ref.find("(")
    if paren_idx > 0:
        return ref[:paren_idx]
    return ref


def clean_fde_reference(fde_ref: str, prefix_annex_a: bool = False) -> str:
    """
    Clean and normalize FDE reference.

    For ISO 27002 controls, add 'A.' prefix to make them Annex A references.
    Strips parenthetical suffixes like (a), (b)(1), etc.
    """
    if pd.isna(fde_ref):
        return None

    fde_ref = str(fde_ref).strip()

    if not fde_ref:
        return None

    # Strip parenthetical suffixes
    fde_ref = strip_parenthetical(fde_ref)

    # For ISO 27002 controls, prefix with 'A.' if not already
    if prefix_annex_a:
        section_num = get_section_number(fde_ref)
        # ISO 27002 actual controls start at section 5
        if section_num >= 5:
            if not fde_ref.startswith("A."):
                fde_ref = f"A.{fde_ref}"

    return fde_ref


def normalize_relationship(rel: str) -> str:
    """
    Normalize STRM relationship values to: subset, intersect, equal.
    """
    if pd.isna(rel):
        return None

    rel = str(rel).lower().strip()

    if "subset" in rel:
        return "subset"
    elif "intersect" in rel:
        return "intersect"
    elif "equal" in rel:
        return "equal"
    elif "superset" in rel:
        return "superset"
    elif "no relationship" in rel:
        return None
    else:
        return rel


def load_strm_csv(
    csv_path: Path, prefix_annex_a: bool = False, min_section: int = 1
) -> pd.DataFrame:
    """
    Load STRM CSV file and clean the data.

    Args:
        csv_path: Path to the CSV file
        prefix_annex_a: Whether to prefix controls with 'A.' (for ISO 27002)
        min_section: Minimum section number to include (4 for ISO 27001, 5 for ISO 27002)
    """
    print(f"Loading: {csv_path.name}...")
    # Read FDE # as string to preserve values like 5.10 (otherwise pandas reads as 5.1 float)
    df = pd.read_csv(csv_path, dtype={"FDE #": str})

    print(f"  Loaded {len(df)} rows")

    # Filter out 'no relationship' entries
    initial_count = len(df)
    df = df[df["STRM Relationship"] != "no relationship"]
    print(f"  Removed 'no relationship': {initial_count} -> {len(df)}")

    # Filter by minimum section number
    df["_section_num"] = df["FDE #"].apply(get_section_number)
    before_section_filter = len(df)
    df = df[df["_section_num"] >= min_section]
    print(
        f"  Filtered to section >= {min_section}: {before_section_filter} -> {len(df)}"
    )

    # Clean FDE references (add A. prefix for ISO 27002)
    df["FDE #"] = df["FDE #"].apply(lambda x: clean_fde_reference(x, prefix_annex_a))

    # Normalize relationship values
    df["STRM Relationship"] = df["STRM Relationship"].apply(normalize_relationship)

    # Remove rows with empty FDE # or relationship
    df = df[df["FDE #"].notna() & df["STRM Relationship"].notna()]

    # Drop temporary column
    df = df.drop(columns=["_section_num"])

    print(f"  Final row count: {len(df)}")

    return df


def combine_iso_strm_data(
    iso27001_csv: Path, iso27002_csv: Path, output_csv: Path, output_xlsx: Path
):
    """
    Combine ISO 27001 and ISO 27002 STRM data.
    """
    print("=" * 70)
    print("COMBINING ISO 27001 AND ISO 27002 STRM MAPPINGS")
    print("=" * 70)

    # Load ISO 27001 data (clauses 4-10, no prefix needed)
    iso27001_df = load_strm_csv(iso27001_csv, prefix_annex_a=False, min_section=4)
    iso27001_df["Source"] = "ISO 27001:2022"

    # Load ISO 27002 data (section 5+, prefix with 'A.' for Annex A controls)
    iso27002_df = load_strm_csv(iso27002_csv, prefix_annex_a=True, min_section=5)
    iso27002_df["Source"] = "ISO 27002:2022"

    # Combine both dataframes
    print("\n" + "-" * 50)
    print("Combining datasets...")
    combined_df = pd.concat([iso27001_df, iso27002_df], ignore_index=True)

    print(f"Combined total: {len(combined_df)} rows")
    print(f"  ISO 27001 mappings: {len(iso27001_df)}")
    print(f"  ISO 27002 mappings: {len(iso27002_df)}")

    # Create a simplified relationship mapping with required columns only
    relationship_df = combined_df[["FDE #", "SCF #", "STRM Relationship"]].copy()

    # Rename to required column names
    relationship_df.columns = ["source_node_id", "target_node_id", "relationship"]

    # Remove spaces from reference IDs (PDF conversion artifacts)
    relationship_df["source_node_id"] = relationship_df["source_node_id"].str.replace(
        " ", "", regex=False
    )
    relationship_df["target_node_id"] = relationship_df["target_node_id"].str.replace(
        " ", "", regex=False
    )

    # Lowercase the SCF ref_id (target_node_id)
    relationship_df["target_node_id"] = relationship_df["target_node_id"].str.lower()

    # Remove duplicates (since stripping parentheticals creates duplicates)
    before_dedup = len(relationship_df)
    relationship_df = relationship_df.drop_duplicates()
    print(f"  Deduplicated: {before_dedup} -> {len(relationship_df)}")

    # Sort by source then target
    relationship_df = relationship_df.sort_values(["source_node_id", "target_node_id"])

    # Show summary statistics
    print("\n" + "-" * 50)
    print("SUMMARY STATISTICS")
    print("-" * 50)
    print(f"Total mappings: {len(relationship_df)}")
    print(f"Unique source nodes (ISO): {relationship_df['source_node_id'].nunique()}")
    print(f"Unique target nodes (SCF): {relationship_df['target_node_id'].nunique()}")

    print("\nRelationship type distribution:")
    print(relationship_df["relationship"].value_counts())

    # Save to CSV
    print(f"\nSaving to {output_csv}...")
    relationship_df.to_csv(output_csv, index=False)

    # Save to Excel with formatting
    print(f"Saving to {output_xlsx}...")
    save_to_excel(relationship_df, output_xlsx)

    print("\n" + "=" * 70)
    print("COMPLETE")
    print("=" * 70)
    print(f"\nGenerated files:")
    print(f"  1. {output_csv}")
    print(f"  2. {output_xlsx}")

    return relationship_df


def save_to_excel(relationship_df: pd.DataFrame, output_path: Path):
    """
    Save to Excel with formatting.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        print("Warning: openpyxl not installed, skipping Excel export")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Mapping"

    # Header styling
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    # Write data
    for r_idx, row in enumerate(
        dataframe_to_rows(relationship_df, index=False, header=True), 1
    ):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
            cell.alignment = Alignment(vertical="top")

    # Adjust column widths
    col_widths = [20, 15, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width

    ws.freeze_panes = "A2"

    wb.save(output_path)


def main():
    script_dir = Path(__file__).parent
    strm_exports = script_dir / "strm_exports"

    # Input files
    iso27001_csv = strm_exports / "general-iso-27001-2022.csv"
    iso27002_csv = strm_exports / "general-iso-27002-2022.csv"

    # Output files
    output_csv = script_dir / "iso27001_scf_mapping.csv"
    output_xlsx = script_dir / "iso27001_scf_mapping.xlsx"

    # Check input files exist
    if not iso27001_csv.exists():
        print(f"Error: ISO 27001 STRM file not found: {iso27001_csv}")
        print("Please run extract_strm.py first to generate STRM exports.")
        exit(1)

    if not iso27002_csv.exists():
        print(f"Error: ISO 27002 STRM file not found: {iso27002_csv}")
        print("Please run extract_strm.py first to generate STRM exports.")
        exit(1)

    # Run the combination
    combine_iso_strm_data(iso27001_csv, iso27002_csv, output_csv, output_xlsx)


if __name__ == "__main__":
    main()
