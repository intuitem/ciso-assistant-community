#!/usr/bin/env python3
"""
Create India DPDPA 2023 to SCF mapping from STRM data.

This script:
1. Loads India DPDPA 2023 STRM data
2. Converts ref_ids to lowercase
3. Normalizes relationship values
4. Outputs both CSV and Excel formats

The mapping direction is: India DPDPA section → SCF control
"""

import pandas as pd
from pathlib import Path


def normalize_ref_id(ref: str) -> str:
    """
    Normalize India DPDPA reference ID to lowercase.

    Examples:
        4(1)(a) -> 4(1)(a)
        8(7)(b) -> 8(7)(b)
    """
    if not ref:
        return ref

    ref = str(ref).strip()
    return ref.lower()


def normalize_relationship(rel: str) -> str:
    """
    Normalize STRM relationship values to: subset, intersect, equal, superset.
    """
    if pd.isna(rel):
        return None

    rel = str(rel).lower().strip()

    if "subset" in rel:
        return "subset"
    elif "intersect" in rel:
        return "intersect"
    elif "equal" in rel:
        return "equal"
    elif "superset" in rel:
        return "superset"
    elif "no relationship" in rel:
        return None
    else:
        return rel


def load_strm_csv(csv_path: Path) -> pd.DataFrame:
    """
    Load STRM CSV file and clean the data.
    """
    print(f"Loading: {csv_path.name}...")
    # Read FDE # as string to preserve values
    df = pd.read_csv(csv_path, dtype={"FDE #": str, "SCF #": str})

    print(f"  Loaded {len(df)} rows")

    # Filter out 'no relationship' entries
    initial_count = len(df)
    df = df[df["STRM Relationship"].str.lower() != "no relationship"]
    print(f"  Removed 'no relationship': {initial_count} -> {len(df)}")

    # Filter out N/A SCF #
    before_na = len(df)
    df = df[df["SCF #"] != "N/A"]
    print(f"  Removed 'N/A' SCF #: {before_na} -> {len(df)}")

    # Convert ref_ids to lowercase
    df["FDE #"] = df["FDE #"].apply(normalize_ref_id)

    # Normalize relationship values
    df["STRM Relationship"] = df["STRM Relationship"].apply(normalize_relationship)

    # Remove rows with empty FDE # or relationship
    df = df[df["FDE #"].notna() & df["STRM Relationship"].notna()]

    print(f"  Final row count: {len(df)}")

    return df


def create_india_dpdpa_mapping(input_csv: Path, output_csv: Path, output_xlsx: Path):
    """
    Create India DPDPA 2023 to SCF mapping.
    """
    print("=" * 70)
    print("CREATING INDIA DPDPA 2023 TO SCF MAPPING")
    print("=" * 70)

    # Load India DPDPA data
    df = load_strm_csv(input_csv)

    # Create a simplified relationship mapping with required columns only
    relationship_df = df[["FDE #", "SCF #", "STRM Relationship"]].copy()

    # Rename to required column names
    relationship_df.columns = ["source_node_id", "target_node_id", "relationship"]

    # Remove spaces from reference IDs (PDF conversion artifacts)
    relationship_df["source_node_id"] = relationship_df["source_node_id"].str.replace(
        " ", "", regex=False
    )
    relationship_df["target_node_id"] = relationship_df["target_node_id"].str.replace(
        " ", "", regex=False
    )

    # Lowercase the SCF ref_id (target_node_id)
    relationship_df["target_node_id"] = relationship_df["target_node_id"].str.lower()

    # Remove duplicates on source-target pairs (keep first occurrence)
    before_dedup = len(relationship_df)
    relationship_df = relationship_df.drop_duplicates(
        subset=["source_node_id", "target_node_id"], keep="first"
    )
    print(f"  Deduplicated: {before_dedup} -> {len(relationship_df)}")

    # Sort by source then target
    relationship_df = relationship_df.sort_values(["source_node_id", "target_node_id"])

    # Show summary statistics
    print("\n" + "-" * 50)
    print("SUMMARY STATISTICS")
    print("-" * 50)
    print(f"Total mappings: {len(relationship_df)}")
    print(
        f"Unique source nodes (India DPDPA): {relationship_df['source_node_id'].nunique()}"
    )
    print(f"Unique target nodes (SCF): {relationship_df['target_node_id'].nunique()}")

    print("\nRelationship type distribution:")
    print(relationship_df["relationship"].value_counts())

    print("\nSample source node IDs:")
    print(relationship_df["source_node_id"].head(10).tolist())

    # Save to CSV
    print(f"\nSaving to {output_csv}...")
    relationship_df.to_csv(output_csv, index=False)

    # Save to Excel with formatting
    print(f"Saving to {output_xlsx}...")
    save_to_excel(relationship_df, output_xlsx)

    print("\n" + "=" * 70)
    print("COMPLETE")
    print("=" * 70)
    print(f"\nGenerated files:")
    print(f"  1. {output_csv}")
    print(f"  2. {output_xlsx}")

    return relationship_df


def save_to_excel(relationship_df: pd.DataFrame, output_path: Path):
    """
    Save to Excel with formatting.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        print("Warning: openpyxl not installed, skipping Excel export")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Mapping"

    # Header styling
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    # Write data
    for r_idx, row in enumerate(
        dataframe_to_rows(relationship_df, index=False, header=True), 1
    ):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
            cell.alignment = Alignment(vertical="top")

    # Adjust column widths
    col_widths = [20, 15, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width

    ws.freeze_panes = "A2"

    wb.save(output_path)


def main():
    script_dir = Path(__file__).parent
    strm_exports = script_dir / "strm_exports"

    # Input file
    input_csv = strm_exports / "apac-india-dpdpa-2023.csv"

    # Output files
    output_csv = script_dir / "india_dpdpa_scf_mapping.csv"
    output_xlsx = script_dir / "india_dpdpa_scf_mapping.xlsx"

    # Check input file exists
    if not input_csv.exists():
        print(f"Error: India DPDPA STRM file not found: {input_csv}")
        print("Please run extract_strm.py first to generate STRM exports.")
        exit(1)

    # Run the mapping creation
    create_india_dpdpa_mapping(input_csv, output_csv, output_xlsx)


if __name__ == "__main__":
    main()
