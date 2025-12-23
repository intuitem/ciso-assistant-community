#!/usr/bin/env python3
"""
Extract STRM mappings from SCF PDF files to CSV and Excel formats.
"""

import pdfplumber
import csv
import os
from pathlib import Path

try:
    import openpyxl

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print(
        "Warning: openpyxl not installed, Excel export disabled. Install with: pip install openpyxl"
    )


def extract_strm_from_pdf(pdf_path):
    """Extract STRM mapping rows from a PDF file.

    Handles multi-target rows where FDE # is only in the first row
    and continuation rows have FDE # = None but still have SCF # values.
    """
    rows = []
    current_fde = None  # Track the current FDE # for continuation rows
    current_fde_data = None  # Track columns 0-5 for continuation rows

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    # Skip header rows and metadata rows
                    if row[0]:
                        if row[0].startswith("FDE #"):
                            continue
                        if "NIST IR 8477" in row[0] or "Reference Document" in row[0]:
                            continue

                    # Must have SCF # to be a valid mapping row
                    if not row[6]:
                        continue

                    # If FDE # is present, this is a new source row
                    if row[0]:
                        current_fde = row[0]
                        # Store the first 6 columns (FDE #, Name, Description, Rationale, Relationship, SCF Control)
                        current_fde_data = row[:6]

                    # Skip if we don't have a current FDE (shouldn't happen after header)
                    if not current_fde:
                        continue

                    # Build the clean row
                    clean_row = []

                    # Use current FDE data for first 6 columns if this is a continuation row
                    if row[0]:
                        # This is a primary row with FDE #
                        for cell in row[:10]:
                            if cell is None:
                                clean_row.append("")
                            else:
                                clean_row.append(str(cell).replace("\n", " ").strip())
                    else:
                        # This is a continuation row - use stored FDE data + current row's target columns
                        for cell in current_fde_data:
                            if cell is None:
                                clean_row.append("")
                            else:
                                clean_row.append(str(cell).replace("\n", " ").strip())
                        # Add columns 6-9 (SCF #, SCF Control Description, Strength, Notes)
                        for cell in row[6:10]:
                            if cell is None:
                                clean_row.append("")
                            else:
                                clean_row.append(str(cell).replace("\n", " ").strip())

                    rows.append(clean_row)

    return rows


def save_to_csv(rows, output_path, columns):
    """Save rows to CSV file."""
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(columns)
        writer.writerows(rows)


def save_to_excel(rows, output_path, columns):
    """Save rows to Excel file."""
    if not HAS_OPENPYXL:
        return False

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "STRM Mapping"

    # Header styling
    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_font_white = Font(bold=True, color="FFFFFF")

    # Write header
    for col, header in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True)

    # Write data
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Adjust column widths
    col_widths = [15, 10, 60, 12, 15, 30, 10, 60, 10, 20]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(output_path)
    return True


def main():
    # Paths
    strm_dir = Path(
        "/Users/abder/mydev/intuitem/staging/securecontrolsframework/Set Theory Relationship Mapping (STRM)"
    )
    output_dir = Path(
        "/Users/abder/mydev/intuitem/staging/ciso-assistant-community/tools/scf_toolbox/strm_exports"
    )
    output_dir.mkdir(exist_ok=True)

    columns = [
        "FDE #",
        "FDE Name",
        "FDE Description",
        "STRM Rationale",
        "STRM Relationship",
        "SCF Control",
        "SCF #",
        "SCF Control Description",
        "Strength",
        "Notes",
    ]

    pdf_files = sorted(strm_dir.glob("*.pdf"))
    print(f"Found {len(pdf_files)} PDF files\n")

    results = []

    for pdf_path in pdf_files:
        name = pdf_path.stem.replace("scf-strm-", "")
        print(f"Processing: {name}...", end=" ")

        try:
            rows = extract_strm_from_pdf(pdf_path)

            if rows:
                # Save CSV
                csv_path = output_dir / f"{name}.csv"
                save_to_csv(rows, csv_path, columns)

                # Save Excel
                xlsx_path = output_dir / f"{name}.xlsx"
                excel_ok = save_to_excel(rows, xlsx_path, columns)

                print(f"{len(rows)} mappings")
                results.append({"name": name, "count": len(rows), "status": "OK"})
            else:
                print("No mappings found")
                results.append({"name": name, "count": 0, "status": "EMPTY"})

        except Exception as e:
            print(f"ERROR: {e}")
            results.append({"name": name, "count": 0, "status": f"ERROR: {e}"})

    # Summary
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    total = sum(r["count"] for r in results)
    success = sum(1 for r in results if r["status"] == "OK")
    print(f"Files processed: {len(results)}")
    print(f"Successful: {success}")
    print(f"Total mappings extracted: {total}")
    print(f"Output directory: {output_dir}")


if __name__ == "__main__":
    main()
