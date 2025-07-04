"""
Prepare Mapping v2

This script generates a mapping Excel file between two YAML-based frameworks
used in CISO Assistant. It extracts relevant metadata from the source and target
YAML files, builds a structured Excel workbook with mapping sheets, and saves it
under a filename derived from the two frameworks' reference IDs.

Usage:
    python prepare_mapping_v2.py source.yaml target.yaml
"""


import openpyxl
import argparse
import yaml
import os
import sys



def load_and_validate_yaml(path, label):
    """Load a YAML file and ensure required structure exists."""
    
    if not os.path.exists(path):
        raise FileNotFoundError(f"{label} file not found: \"{path}\"")
    if not os.path.isfile(path):
        raise IsADirectoryError(f"{label} path is not a file: \"{path}\"")

    with open(path, "r", encoding="utf-8") as f:
        try:
            data = yaml.safe_load(f)
        except yaml.YAMLError as e:
            raise ValueError(f"{label} file is not valid YAML\n\t   {e}")

    if not isinstance(data, dict):
        raise TypeError(f"{label} file must be a YAML dictionary")

    # Basic structure check
    if "urn" not in data:
        raise KeyError(f"{label} file is missing the \"urn\" field.")
    if "objects" not in data or "framework" not in data["objects"]:
        raise KeyError(f"{label} file is missing the \"objects.framework\" structure.")
    if "locale" not in data:
        raise KeyError(f'{label} file is missing the "locale" field.')

    return data


def generate_mapping_excel(source_yaml, target_yaml):

    packager = "intuitem"

    print(f"⌛ Parsing \"{os.path.basename(source_yaml)}\" and \"{os.path.basename(target_yaml)}\"...")

    # Load and validate YAML files
    source = load_and_validate_yaml(source_yaml, "Source")
    target = load_and_validate_yaml(target_yaml, "Target")

    # Extract metadata from source and target libraries and frameworks
    source_library_urn = source["urn"]
    target_library_urn = target["urn"]
    source_framework = source["objects"]["framework"]
    target_framework = target["objects"]["framework"]

    source_framework_urn = source_framework["urn"]
    target_framework_urn = target_framework["urn"]
    source_framework_ref_id = source_framework["ref_id"]
    target_framework_ref_id = target_framework["ref_id"]
    source_framework_name = source_framework["name"]
    target_framework_name = target_framework["name"]

    source_ref_id = source_framework_urn.split(":")[-1]
    target_ref_id = target_framework_urn.split(":")[-1]

    # Generate common values used across sheets
    ref_id = f"mapping-{source_ref_id}-and-{target_ref_id}"
    name = f"{source_framework_ref_id} <-> {target_framework_ref_id}"
    description = f"Mapping between {source_framework_name} and {target_framework_name}"
    output_file_name = f"{ref_id}.xlsx"

    library_urn = f"urn:{packager.lower()}:risk:library:{ref_id}"
    mapping_urn = f"urn:{packager.lower()}:risk:req_mapping_set:{ref_id}"
    source_node_base_urn = f"urn:{packager.lower()}:risk:req_node:{source_ref_id}"
    target_node_base_urn = f"urn:{packager.lower()}:risk:req_node:{target_ref_id}"

    # Create a new workbook and remove default sheet
    wb_output = openpyxl.Workbook()
    default_sheet = wb_output.active
    wb_output.remove(default_sheet)

    # === Sheet: library_meta ===
    ws_meta = wb_output.create_sheet("library_meta")
    ws_meta.append(["type", "library"])
    ws_meta.append(["urn", library_urn])
    ws_meta.append(["version", "1"])
    ws_meta.append(["locale", source["locale"]])
    ws_meta.append(["ref_id", ref_id])
    ws_meta.append(["name", name])
    ws_meta.append(["description", description])
    ws_meta.append(["copyright", packager])
    ws_meta.append(["provider", packager])
    ws_meta.append(["packager", packager])
    ws_meta.append(["dependencies", f"{source_library_urn}, {target_library_urn}"])

    # === Sheet: mappings_meta ===
    ws_map_meta = wb_output.create_sheet("mappings_meta")
    ws_map_meta.append(["type", "requirement_mapping_set"])
    ws_map_meta.append(["urn", mapping_urn])
    ws_map_meta.append(["ref_id", ref_id])
    ws_map_meta.append(["name", name])
    ws_map_meta.append(["description", description])
    ws_map_meta.append(["source_framework_urn", source_framework_urn])
    ws_map_meta.append(["target_framework_urn", target_framework_urn])
    ws_map_meta.append(["source_node_base_urn", source_node_base_urn])
    ws_map_meta.append(["target_node_base_urn", target_node_base_urn])

    # === Sheet: mappings_content (formerly "mappings") ===
    ws_mappings = wb_output.create_sheet("mappings_content")
    ws_mappings.append([
        "source_node_id",
        "target_node_id",
        "relationship",
        "rationale",
        "strength_of_relationship",
    ])
    # Populate source node IDs for mapping rows (initially only source side filled)
    for node in source_framework["requirement_nodes"]:
        if node["assessable"]:
            node_id = node["urn"].split(source_node_base_urn + ":")[-1]
            ws_mappings.append([node_id])

    # === Sheet: guidelines (instructions for mapping) ===
    ws_guidelines = wb_output.create_sheet("guidelines")
    ws_guidelines.append(["source_node_id", "use the last segment of the source URN"])
    ws_guidelines.append(["target_node_id", "use the last segment of the target URN"])
    ws_guidelines.append(["relationships", "use one of the following values"])
    ws_guidelines.append(["", "subset"])
    ws_guidelines.append(["", "intersect"])
    ws_guidelines.append(["", "equal"])
    ws_guidelines.append(["", "superset"])
    ws_guidelines.append(["", "not_related"])
    ws_guidelines.append(["rationale", "use one of the following values (or leave empty)"])
    ws_guidelines.append(["", "syntactic"])
    ws_guidelines.append(["", "semantic"])
    ws_guidelines.append(["", "functional"])
    ws_guidelines.append(["strength_of_relationship", "use an integer or empty value"])

    # === Sheet: source (source framework requirements) ===
    ws_source = wb_output.create_sheet("source")
    ws_source.append(["node_id", "assessable", "urn", "ref_id", "name", "description"])
    for node in source_framework["requirement_nodes"]:
        node_id = node["urn"].split(source_node_base_urn + ":")[-1] if node["assessable"] else ""
        ws_source.append([
            node_id,
            node["assessable"],
            node["urn"],
            node.get("ref_id"),
            node.get("name"),
            node.get("description"),
        ])

    # === Sheet: target (target framework requirements) ===
    ws_target = wb_output.create_sheet("target")
    ws_target.append(["node_id", "assessable", "urn", "ref_id", "name", "description"])
    for node in target_framework["requirement_nodes"]:
        node_id = node["urn"].split(target_node_base_urn + ":")[-1] if node["assessable"] else ""
        ws_target.append([
            node_id,
            node["assessable"],
            node["urn"],
            node.get("ref_id"),
            node.get("name"),
            node.get("description"),
        ])

    # Save the workbook to disk
    try:
        wb_output.save(output_file_name)
    except Exception as e:
        raise IOError(f"Failed to save Excel file \"{output_file_name}\"\n\t   {e}")
    
    print(f"✅ Excel file created successfully: \"{output_file_name}\"")


def main():

    parser = argparse.ArgumentParser(prog="prepare_mapping_v2.py", description="Prepare a mapping Excel file for CISO Assistant")
    parser.add_argument("source_yaml", help="Source YAML file")
    parser.add_argument("target_yaml", help="Target YAML file")
    args = parser.parse_args()

    try:
        generate_mapping_excel(args.source_yaml, args.target_yaml)
    except Exception as e:
        print(f"❌ [ERROR] {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
