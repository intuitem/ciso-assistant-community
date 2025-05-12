"""
create_library.py — Build a CISO Assistant YAML library from a v2 Excel file

This script processes an Excel file in v2 format (with *_meta and *_content tabs),
extracts all declared objects, and generates a fully structured YAML library.

Usage:
    python create_library.py path/to/library.xlsx [--compat]

Arguments:
    --compat  Use legacy URN fallback logic (for requirements without ref_id)
"""

import argparse
from pathlib import Path
import openpyxl
import re
import yaml
import datetime

# --- Translation helpers ------------------------------------------------------

def extract_translations_from_row(header, row):
    translations = {}
    for i, col_name in enumerate(header):
        match = re.match(r"(\w+)\[(\w+)\]", col_name)
        if match and i < len(row):
            base_key, lang = match.groups()
            value = row[i].value
            if value:
                translations.setdefault(lang, {})[base_key] = str(value).strip()
    return translations

def extract_translations_from_metadata(meta_dict, prefix):
    translations = {}
    pattern = re.compile(r"(\w+)\[(\w+)\]")
    for key, value in meta_dict.items():
        match = pattern.match(key)
        if match and value:
            base_key, lang = match.groups()
            translations.setdefault(lang, {})[base_key] = str(value).strip()
    return translations

# --- Sheet parsing ------------------------------------------------------------

def parse_key_value_sheet(sheet):
    result = {}
    for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
        key = str(row[0]).strip().lower() if row[0] else None
        val = str(row[1]).strip() if row[1] else None
        if key:
            result[key] = val
    return result

# --- urn expansion ------------------------------------------------------------

def expand_urns_from_prefixed_list(field_value, prefix_to_urn):
    """
    Convert a prefixed list like 'abc:xyz, def:uvw' into a list of full URNs,
    using a prefix mapping. Fully qualified URNs (starting with 'urn:') are left untouched.

    Args:
        field_value (str): Raw string field value from the sheet (e.g., 'abc:id1, urn:...').
        prefix_to_urn (dict): Mapping from prefix (e.g., 'abc') to full base URN.

    Returns:
        list[str]: Fully qualified URNs
    """
    result = []
    elements = re.split(r"[\s,]+", str(field_value))
    for element in elements:
        element = element.strip()
        if not element:
            continue
        if element.startswith("urn:"):
            result.append(element)
            continue
        parts = element.split(":")
        if len(parts) >= 2:
            prefix = parts[0]
            suffix = ":".join(parts[1:]).strip().lower().replace(" ", "-")
            base = prefix_to_urn.get(prefix)
            if base:
                result.append(f"{base}:{suffix}")
            else:
                print(f"⚠️ Unknown prefix '{prefix}' for element '{element}' — skipped")
        else:
            print(f"⚠️ Malformed element '{element}' — skipped")
    return result

# --- question management ------------------------------------------------------------

def inject_questions_into_node(node, raw_question_str, raw_answer_str, answers_dict):
    """
    Injects parsed questions and their metadata into a requirement node.
    Ensures that question type is valid and handles multiple questions/answers.

    :param node: the requirement node (dict)
    :param raw_question_str: the string from the "questions" column
    :param raw_answer_str: the string from the "answer" column
    :param answers_dict: dictionary of all available answers (from answer sheet)
    """
    if not raw_question_str:
        return

    allowed_types = {"unique_choice", "multiple_choice", "text", "date"}

    question_lines = [q.strip() for q in str(raw_question_str).split("\n") if q.strip()]
    answer_ids = [a.strip() for a in str(raw_answer_str).split("\n") if raw_answer_str] if raw_answer_str else []

    if len(answer_ids) != 1 and len(answer_ids) != len(question_lines):
        raise ValueError(f"Mismatch between questions and answers for node {node.get('urn')}")

    question_block = {}

    for idx, question_text in enumerate(question_lines):
        answer_id = answer_ids[0] if len(answer_ids) == 1 else answer_ids[idx]
        answer_meta = answers_dict.get(answer_id)
        if not answer_meta:
            raise ValueError(f"Unknown answer ID: {answer_id} for node {node.get('urn')}")
        qtype = answer_meta.get("type")
        if qtype not in allowed_types:
            raise ValueError(f"Invalid question type '{qtype}' for answer ID: {answer_id}")
        q_urn = f"{node['urn']}:question:{idx+1}"
        question_entry = {
            "type": qtype,
        }
        if qtype in {"unique_choice", "multiple_choice"}:
            choices = []
            for j, choice in enumerate(answer_meta["choices"]):
                choice_urn = f"{q_urn}:choice:{j+1}"
                choices.append({
                    "urn": choice_urn,
                    "value": choice["value"]
                })
            question_entry["choices"] = choices
        question_entry["text"] = question_text
        question_block[q_urn] = question_entry
    if question_block:
        node["questions"] = question_block

# --- risk matrix management ------------------------------------------------------------

# https://gist.github.com/Mike-Honey/b36e651e9a7f1d2e1d60ce1c63b9b633
from colorsys import rgb_to_hls, hls_to_rgb

RGBMAX = 0xFF  # Corresponds to 255
HLSMAX = 240  # MS excel's tint function expects that HLS is base 240. see:
# https://social.msdn.microsoft.com/Forums/en-US/e9d8c136-6d62-4098-9b1b-dac786149f43/excel-color-tint-algorithm-incorrect?forum=os_binaryfile#d3c2ac95-52e0-476b-86f1-e2a697f24969

def rgb_to_ms_hls(red, green=None, blue=None):
    """Converts rgb values in range (0,1) or a hex string of the form '[#aa]rrggbb' to HLSMAX based HLS, (alpha values are ignored)"""
    if green is None:
        if isinstance(red, str):
            if len(red) > 6:
                red = red[-6:]  # Ignore preceding '#' and alpha values
            blue = int(red[4:], 16) / RGBMAX
            green = int(red[2:4], 16) / RGBMAX
            red = int(red[0:2], 16) / RGBMAX
        else:
            red, green, blue = red
    h, l, s = rgb_to_hls(red, green, blue)
    return (int(round(h * HLSMAX)), int(round(l * HLSMAX)), int(round(s * HLSMAX)))

def ms_hls_to_rgb(hue, lightness=None, saturation=None):
    """Converts HLSMAX based HLS values to rgb values in the range (0,1)"""
    if lightness is None:
        hue, lightness, saturation = hue
    return hls_to_rgb(hue / HLSMAX, lightness / HLSMAX, saturation / HLSMAX)

def rgb_to_hex(red, green=None, blue=None):
    """Converts (0,1) based RGB values to a hex string 'rrggbb'"""
    if green is None:
        red, green, blue = red
    return (
        "%02x%02x%02x"
        % (
            int(round(red * RGBMAX)),
            int(round(green * RGBMAX)),
            int(round(blue * RGBMAX)),
        )
    ).upper()

def get_theme_colors(wb):
    """Gets theme colors from the workbook"""
    # see: https://groups.google.com/forum/#!topic/openpyxl-users/I0k3TfqNLrc
    from openpyxl.xml.functions import QName, fromstring

    xlmns = "http://schemas.openxmlformats.org/drawingml/2006/main"
    root = fromstring(wb.loaded_theme)
    themeEl = root.find(QName(xlmns, "themeElements").text)
    colorSchemes = themeEl.findall(QName(xlmns, "clrScheme").text)
    firstColorScheme = colorSchemes[0]

    colors = []

    for c in [
        "lt1",
        "dk1",
        "lt2",
        "dk2",
        "accent1",
        "accent2",
        "accent3",
        "accent4",
        "accent5",
        "accent6",
    ]:
        accent = firstColorScheme.find(QName(xlmns, c).text)
        for i in list(accent):  # walk all child nodes, rather than assuming [0]
            if "window" in i.attrib["val"]:
                colors.append(i.attrib["lastClr"])
            else:
                colors.append(i.attrib["val"])

    return colors

def tint_luminance(tint, lum):
    """Tints a HLSMAX based luminance"""
    # See: http://ciintelligence.blogspot.co.uk/2012/02/converting-excel-theme-color-and-tint.html
    if tint < 0:
        return int(round(lum * (1.0 + tint)))
    else:
        return int(round(lum * (1.0 - tint) + (HLSMAX - HLSMAX * (1.0 - tint))))

def theme_and_tint_to_rgb(wb, theme, tint):
    """Given a workbook, a theme number and a tint return a hex based rgb"""
    rgb = get_theme_colors(wb)[theme]
    h, l, s = rgb_to_ms_hls(rgb)
    return rgb_to_hex(ms_hls_to_rgb(h, tint_luminance(tint, l), s))

def get_color(wb, cell):
    """get cell color; None for no fill"""
    if not cell.fill.patternType:
        return None
    if isinstance(cell.fill.fgColor.rgb, str):
        return "#" + cell.fill.fgColor.rgb[2:]
    theme = cell.fill.start_color.theme
    tint = cell.fill.start_color.tint
    color = theme_and_tint_to_rgb(wb, theme, tint)
    return "#" + color

def parse_risk_matrix(meta, content_ws, wb):
    """Parse a risk_matrix block into a structured dict."""
    rows = list(content_ws.iter_rows())
    if not rows:
        return None
    header = {}
    grid_count = 0
    for i, cell in enumerate(rows[0]):
        col = str(cell.value).strip().lower() if cell.value else ""
        if col == "grid":
            col = f"grid{grid_count}"
            grid_count += 1
        header[col] = i

    assert all(k in header for k in ["type", "id", "color", "abbreviation", "name", "description", "grid0"])
    index = {k: i for i, k in enumerate(header)}
    size_grid = len([h for h in header if h.startswith("grid")])

    risk_matrix = {
        "urn": meta.get("urn"),
        "ref_id": meta.get("ref_id"),
        "name": meta.get("name"),
        "description": meta.get("description"),
        "probability": [],
        "impact": [],
        "risk": [],
        "grid": []
    }

    translations = extract_translations_from_metadata(meta, "risk_matrix")
    if translations:
        risk_matrix["translations"] = translations

    grid = {}
    grid_color = {}

    for row in rows[1:]:
        if not any(c.value for c in row):
            continue

        ctype = str(row[index["type"]].value).strip().lower()
        assert ctype in ("probability", "impact", "risk")

        rid = int(row[index["id"]].value)
        abbrev = str(row[index["abbreviation"]].value or "").strip()
        name = str(row[index["name"]].value or "").strip()
        desc = str(row[index["description"]].value or "").strip()
        color_cell = row[index["color"]]

        obj_data = {
            "id": rid,
            "abbreviation": abbrev,
            "name": name,
            "description": desc,
        }

        translations = extract_translations_from_row(header, row)
        if translations:
            obj_data["translations"] = translations

        if color := get_color(wb, color_cell):
            obj_data["hexcolor"] = color

        risk_matrix[ctype].append(obj_data)

        if ctype == "probability":
            grid[rid] = []
            grid_color[rid] = []
            for i in range(size_grid):
                cell = row[index[f"grid{i}"]]
                grid[rid].append(int(cell.value))
                grid_color[rid].append(get_color(wb, cell))

    # Build grid from sorted probability IDs
    sorted_ids = sorted(grid.keys())
    risk_matrix["grid"] = [grid[rid] for rid in sorted_ids]

    # Ensure order
    for k in ("probability", "impact", "risk"):
        risk_matrix[k].sort(key=lambda x: x["id"])

    return risk_matrix


# --- Main logic ---------------------------------------------------------------

def create_library(input_file: str, output_file: str, compat: bool = False):
    wb = openpyxl.load_workbook(input_file)
    sheets = wb.sheetnames
    object_blocks = {}

    # Step 1: Load library_meta
    if "library_meta" not in sheets:
        raise ValueError("Missing 'library_meta' sheet.")

    library_meta_sheet = wb["library_meta"]
    library_meta = parse_key_value_sheet(library_meta_sheet)

    if library_meta.get("type") != "library":
        raise ValueError("'library_meta' must declare type = library")

    # Step 2: Format publication_date
    lib_date = library_meta.get("publication_date", datetime.datetime.now())
    if isinstance(lib_date, str):
        try:
            lib_date = datetime.datetime.fromisoformat(lib_date).date()
        except ValueError:
            lib_date = datetime.datetime.now().date()
    elif isinstance(lib_date, datetime.datetime):
        lib_date = lib_date.date()

    # Step 3: Build library with fixed key order
    library = {
        "urn": library_meta.get("urn"),
        "locale": library_meta.get("locale"),
        "ref_id": library_meta.get("ref_id"),
        "name": library_meta.get("name"),
        "description": library_meta.get("description"),
        "copyright": library_meta.get("copyright"),
        "version": int(library_meta["version"]),
        "publication_date": lib_date,
        "provider": library_meta.get("provider"),
        "packager": library_meta.get("packager"),
    }

    translations = extract_translations_from_metadata(library_meta, "library")
    if translations:
        library["translations"] = translations
    if "dependencies" in library_meta:
        dependencies = [
            x.strip() for x in re.split(r"[\s,]+", library_meta["dependencies"]) if x.strip()
        ]
        if dependencies:
            library["dependencies"] = dependencies

    library["objects"] = {}
    prefix_to_urn = {}

    # Step 4: Detect all object blocks
    for sheet_name in sheets:
        if sheet_name.endswith("_meta") and sheet_name != "library_meta":
            prefix = sheet_name[:-5]
            content_name = f"{prefix}_content"
            if content_name not in sheets:
                print(f"⚠️ Missing content for {prefix}")
                continue

            meta_sheet = wb[sheet_name]
            meta_data = parse_key_value_sheet(meta_sheet)
            object_type = meta_data.get("type")
            object_name = meta_data.get("name", prefix)

            # Capture prefix mappings if this is a urn_prefix block
            if object_type == "urn_prefix" and f"{prefix}_content" in wb.sheetnames:
                ws_prefix = wb[f"{prefix}_content"]
                rows = list(ws_prefix.iter_rows(values_only=True))
                if rows:
                    header = [str(cell).strip().lower() if cell else "" for cell in rows[0]]
                    idx_id = header.index("prefix_id") if "prefix_id" in header else None
                    idx_val = header.index("prefix_value") if "prefix_value" in header else None
                    if idx_id is not None and idx_val is not None:
                        for row in rows[1:]:
                            prefix_id = str(row[idx_id]).strip() if row[idx_id] else None
                            prefix_val = str(row[idx_val]).strip() if row[idx_val] else None
                            if prefix_id and prefix_val:
                                prefix_to_urn[prefix_id] = prefix_val


            object_blocks[object_name] = {
                "type": object_type,
                "meta": meta_data,
                "content_sheet": wb[content_name]
            }

    print(f"📦 Found {len(object_blocks)} objects.")

    # Step 5: Ordered object insertion (reference_controls before threats)
    priority_order = ["reference_controls", "threats"]

    sorted_object_names = sorted(
        object_blocks.keys(),
        key=lambda k: (
            priority_order.index(object_blocks[k]["type"])
            if object_blocks[k]["type"] in priority_order
            else len(priority_order)
        )
    )

    for name in sorted_object_names:
        obj = object_blocks[name]
        obj_type = obj["type"]
        print(f"→ Handling {obj_type}: {name}")

        if obj_type == "reference_controls":
            controls = []
            base_urn = obj["meta"].get("base_urn")
            content_ws = obj["content_sheet"]
            rows = list(content_ws.iter_rows())
            if not rows:
                continue
            header = [str(cell.value).strip().lower() if cell.value else "" for cell in rows[0]]

            for row in rows[1:]:
                if not any(cell.value for cell in row):
                    continue
                data = {header[i]: row[i].value for i in range(len(header)) if i < len(row)}
                ref_id = str(data.get("ref_id", "")).strip()
                if not ref_id:
                    continue
                entry = {
                    "urn": f"{base_urn}:{ref_id.lower()}",
                    "ref_id": ref_id
                }
                if "name" in data and data["name"]:
                    entry["name"] = str(data["name"]).strip()
                if "category" in data and data["category"]:
                    entry["category"] = str(data["category"]).strip()
                if "description" in data and data["description"]:
                    entry["description"] = str(data["description"]).strip()
                if "csf_function" in data and data["csf_function"]:
                    entry["csf_function"] = str(data["csf_function"]).strip()

                translations = extract_translations_from_row(header, row)
                if translations:
                    entry["translations"] = translations

                controls.append(entry)

            library["objects"]["reference_controls"] = controls

        elif obj_type == "threats":
            threats = []
            base_urn = obj["meta"].get("base_urn")
            content_ws = obj["content_sheet"]
            rows = list(content_ws.iter_rows())
            if not rows:
                continue
            header = [str(cell.value).strip().lower() if cell.value else "" for cell in rows[0]]

            for row in rows[1:]:
                if not any(cell.value for cell in row):
                    continue
                data = {header[i]: row[i].value for i in range(len(header)) if i < len(row)}
                ref_id = str(data.get("ref_id", "")).strip()
                if not ref_id:
                    continue
                entry = {
                    "urn": f"{base_urn}:{ref_id.lower()}",
                    "ref_id": ref_id
                }
                if "name" in data and data["name"]:
                    entry["name"] = str(data["name"]).strip()
                if "description" in data and data["description"]:
                    entry["description"] = str(data["description"]).strip()

                translations = extract_translations_from_row(header, row)
                if translations:
                    entry["translations"] = translations

                threats.append(entry)

            library["objects"]["threats"] = threats

        elif obj_type == "framework":
            meta = obj["meta"]
            content_ws = obj["content_sheet"]
            base_urn = obj["meta"].get("base_urn")

            # --- Retrieve answers block if declared ---
            answers_dict = {}
            answers_block_name = meta.get("answers")
            if answers_block_name:
                if answers_block_name not in object_blocks:
                    raise ValueError(f"❌ Missing answers sheet: '{answers_block_name}'")

                answers_sheet = object_blocks[answers_block_name]["content_sheet"]
                rows = list(answers_sheet.iter_rows())
                if rows:
                    header = [str(c.value).strip().lower() if c.value else "" for c in rows[0]]

                    for row in rows[1:]:
                        data = {header[i]: row[i].value for i in range(len(header)) if i < len(row)}
                        answer_id = str(data.get("id", "")).strip()
                        answer_type = str(data.get("question_type", "")).strip()
                        choices_raw = str(data.get("question_choices", "")).strip()

                        if not answer_id or not answer_type or not choices_raw:
                            continue  # Incomplete row

                        choices = []
                        for line in choices_raw.split("\n"):
                            line = line.strip()
                            if not line:
                                continue
                            if line.startswith("|") and choices:
                                # Multi-line value, append to previous
                                choices[-1]["value"] += "\n" + line[1:].strip()
                            else:
                                choices.append({"urn": "", "value": line})

                        answers_dict[answer_id] = {
                            "type": answer_type,
                            "choices": choices
                        }

            framework = {
                "urn": meta.get("urn"),
                "ref_id": meta.get("ref_id"),
                "name": meta.get("name"),
                "description": meta.get("description")
            }

            translations = extract_translations_from_metadata(meta, "framework")
            if translations:
                framework["translations"] = translations

            if "min_score" in meta:
                framework["min_score"] = int(meta["min_score"])
            if "max_score" in meta:
                framework["max_score"] = int(meta["max_score"])

            score_name = meta.get("scores_definition")
            if score_name and score_name in object_blocks:
                score_ws = object_blocks[score_name]["content_sheet"]
                score_rows = list(score_ws.iter_rows())
                score_header = [str(c.value).strip().lower() if c.value else "" for c in score_rows[0]]
                score_defs = []
                for row in score_rows[1:]:
                    if not any(c.value for c in row):
                        continue
                    data = {score_header[i]: row[i].value for i in range(len(score_header)) if i < len(row)}
                    score_entry = {
                        "score": int(data.get("score")),
                        "name": str(data.get("name", "")).strip(),
                        "description": str(data.get("description", "")).strip()
                    }
                    if "description_doc" in data and data["description_doc"]:
                        score_entry["description_doc"] = str(data["description_doc"]).strip()
                    translations = extract_translations_from_row(score_header, row)
                    if translations:
                        score_entry["translations"] = translations
                    score_defs.append(score_entry)
                framework["scores_definition"] = score_defs

            ig_name = meta.get("implementation_groups_definition")
            if ig_name and ig_name in object_blocks:
                ig_content = object_blocks[ig_name]["content_sheet"]
                ig_rows = list(ig_content.iter_rows())
                ig_header = [str(c.value).strip().lower() if c.value else "" for c in ig_rows[0]]
                ig_defs = []
                for row in ig_rows[1:]:
                    if not any(c.value for c in row):
                        continue
                    data = {ig_header[i]: row[i].value for i in range(len(ig_header)) if i < len(row)}
                    ig_entry = {
                        "ref_id": str(data.get("ref_id", "")).strip(),
                        "name": str(data.get("name", "")).strip(),
                        "description": str(data.get("description", "")).strip() if data.get("description") else None
                    }
                    translations = extract_translations_from_row(ig_header, row)
                    if translations:
                        ig_entry["translations"] = translations
                    ig_defs.append(ig_entry)

                framework["implementation_groups_definition"] = ig_defs

            rows = list(content_ws.iter_rows())
            if rows:
                header = [str(c.value).strip().lower() if c.value else "" for c in rows[0]]
                parent_for_depth = {}
                count_for_depth = {}
                previous_node_urn = None
                previous_depth = 0
                counter = 0
                counter_fix = 0
                requirement_nodes = []
                for row in rows[1:]:
                    counter += 1
                    data = {header[i]: row[i].value for i in range(len(header)) if i < len(row)}
                    depth = int(data.get("depth", 1))
                    ref_id = str(data.get("ref_id", "")).strip() or None
                    name = str(data.get("name", "")).strip() or None

                    if depth == previous_depth + 1:
                        parent_for_depth[depth] = previous_node_urn
                        count_for_depth[depth] = 1
                    elif depth <= previous_depth:
                        pass
                    else:
                        raise ValueError(f"Invalid depth jump from {previous_depth} to {depth} at row {counter}")

                    if compat:
                        skip_count = str(data.get("skip_count", "")).strip().lower() in ("1", "true", "yes", "x")
                        if skip_count:
                            counter_fix += 1
                            ref_id_urn = f"node{counter - counter_fix}-{counter_fix}"
                        else:
                            ref_id_urn = ref_id.lower().replace(" ", "-") if ref_id else f"node{counter - counter_fix}"
                        urn = f"{base_urn}:{ref_id_urn}"
                    else:
                        if ref_id:
                            urn = f"{base_urn}:{ref_id.lower().replace(' ', '-')}"
                        else:
                            p = parent_for_depth.get(depth)
                            c = count_for_depth.get(depth, 1)
                            urn = f"{p}:{c}" if p else f"{base_urn}:node{c}"
                            count_for_depth[depth] = c + 1
                    previous_node_urn = urn
                    previous_depth = depth
                    parent_urn = parent_for_depth.get(depth)

                    node = {
                        "urn": urn,
                        "assessable": bool(data.get("assessable")),
                        "depth": depth,
                    }
                    if parent_urn:
                        node["parent_urn"] = parent_urn
                    if "ref_id" in data and data["ref_id"]:
                        node["ref_id"] = str(data["ref_id"]).strip()
                    if "name" in data and data["name"]:
                        node["name"] = str(data["name"]).strip()
                    if "description" in data and data["description"]:
                        node["description"] = str(data["description"]).strip()
                    if "annotation" in data and data["annotation"]:
                        node["annotation"] = str(data["annotation"]).strip()
                    if "typical_evidence" in data and data["typical_evidence"]:
                        node["typical_evidence"] = str(data["typical_evidence"]).strip()
                    if "implementation_groups" in data and data["implementation_groups"]:
                        node["implementation_groups"] = [
                            s.strip() for s in str(data["implementation_groups"]).split(",")
                        ]
                    if "threats" in data and data["threats"]:
                        threats = expand_urns_from_prefixed_list(data["threats"], prefix_to_urn)
                        if threats:
                            node["threats"] = threats
                    if "reference_controls" in data and data["reference_controls"]:
                        rc = expand_urns_from_prefixed_list(data["reference_controls"], prefix_to_urn)
                        if rc:
                            node["reference_controls"] = rc
                    if "questions" in data and data["questions"]:
                        inject_questions_into_node(
                            node,
                            data.get("questions"),
                            data.get("answer"),
                            answers_dict
                        )
                    translations = extract_translations_from_row(header, row)
                    if translations:
                        node["translations"] = translations
                    requirement_nodes.append(node)

                framework["requirement_nodes"] = requirement_nodes

            library["objects"]["framework"] = framework

        elif obj_type == "risk_matrix":
            matrix = parse_risk_matrix(obj["meta"], obj["content_sheet"], wb)
            if matrix:
                if "risk_matrix" not in library["objects"]:
                    library["objects"]["risk_matrix"] = []
                library["objects"]["risk_matrix"].append(matrix)

        else:
            if obj_type not in ["answers", "implementation_groups", "scores", "urn_prefix"]:
                print("type not handled:", obj_type)

    # Step 6: Export to YAML
    print(f"✅ Writing YAML to {output_file}")
    with open(output_file, "w", encoding="utf-8") as f:
        yaml.dump(library, f, sort_keys=False, allow_unicode=False)

# --- CLI interface ------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Create a YAML library from a v2 Excel file.")
    parser.add_argument("input_file", type=str, help="Input Excel file (v2 format)")
    parser.add_argument("--compat", action="store_true", help="Use legacy URN fallback generation logic")
    args = parser.parse_args()

    input_path = Path(args.input_file)
    if not input_path.exists():
        raise FileNotFoundError(f"File not found: {input_path}")

    output_path = input_path.with_suffix(".yaml")
    create_library(str(input_path), str(output_path), compat=args.compat)

if __name__ == "__main__":
    main()
