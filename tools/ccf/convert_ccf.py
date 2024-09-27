"""
Simple script to convert Adobe CCF Security Controls v5 Excel in a CISO Assistant Excel file
Source: https://www.adobe.com/content/dam/cc/en/trust/pdfs/Open_Source_CCF.xls
"""

import openpyxl
import sys
import re
import argparse
from openpyxl.styles import numbers

parser = argparse.ArgumentParser(
    prog="convert_ccf",
    description="convert Adobe CCF Security Controls v5 Excel file to CISO Assistant Excel file",
)

parser.add_argument("filename", help="name of Adobe CCF Excel file")
args = parser.parse_args()
input_file_name = args.filename
output_file_name = "ccf-v5.xlsx"

library_copyright = """Creative Commons"""
packager = "intuitem"

library_description = """Adobe Common Controls Framework (CCF) version 5
https://www.adobe.com/trust/compliance/adobe-ccf.html
"""

print("parsing", input_file_name)

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook(input_file_name)
controls = {}
evidences = {}
output_table = []
current_domain = ""

for tab in dataframe:
    print("parsing tab", tab.title)
    title = tab.title
    if title in ("CCF Control Guidance"):
        first = True
        for row in tab:
            if not first:
                (
                    id,
                    domain,
                    name,
                    description,
                    control_theme,
                    control_type,
                    policy,
                    implementation,
                    testing,
                    artifacts,
                ) = (r.value for r in row[0:10])
                artifacts = [v for v in artifacts.splitlines() if v != ""]
                implementation = [v for v in implementation.splitlines() if v != ""]
                testing = [v for v in testing.splitlines() if v != ""]
                controls[id] = (
                    id,
                    domain,
                    name,
                    description,
                    control_theme,
                    control_type,
                    policy,
                    implementation,
                    testing,
                    artifacts,
                )
            first = False
    if title in ("Evidence Request List (ERL)"):
        for row in tab:
            (evidence_id, domain, title) = (r.value for r in row[0:3])
            evidences[evidence_id] = title

print("generating", output_file_name)
wb_output = openpyxl.Workbook()
ws = wb_output.active
ws.title = "library_content"
ws.append(["library_urn", f"urn:{packager.lower()}:risk:library:adobe-ccf-v5"])
ws.append(["library_version", 1])
ws.append(["library_locale", "en"])
ws.append(["library_ref_id", "adobe-ccf-v5"])
ws.append(["library_name", "Adobe CCF v5"])
ws.append(["library_description", library_description])
ws.append(["library_copyright", library_copyright])
ws.append(["library_provider", "Adobe"])
ws.append(["library_packager", packager])
ws.append(["framework_urn", f"urn:{packager.lower()}:risk:framework:adobe-ccf-v5"])
ws.append(["framework_ref_id", "adobe-ccf-v5"])
ws.append(["framework_name", "Adobe CCF v5"])
ws.append(["framework_description", library_description])
ws.append(["tab", "requirements", "requirements"])
ws.append(["tab", "answers", "answers"])

ws1 = wb_output.create_sheet("requirements")
ws1.append(
    [
        "assessable",
        "depth",
        "ref_id",
        "name",
        "description",
        "questions",
        "answer",
        "typical_evidence",
        "annotation",
    ]
)
for id in controls:
    (
        id,
        domain,
        name,
        description,
        control_theme,
        control_type,
        policy,
        implementation,
        testing,
        artifacts,
    ) = controls[id]
    if domain != current_domain:
        output_table.append(("", 1, "", domain, "", ""))
        current_domain = domain
    annotation = "\n".join(implementation)
    typical_evidence = "\n".join([v + " - " + evidences.get(v, "") for v in artifacts])
    questions = "\n".join(testing)
    answer = "YNNA"
    output_table.append(
        ("x", 2, id, name, description, questions, answer, typical_evidence, annotation)
    )
for row in output_table:
    ws1.append(row)

ws2 = wb_output.create_sheet("answers")
ws2.append(["id", "question_type", "question_choices"])
ws2.append(["YNNA", "unique_choice", "Yes\nNo\nN/A"])

print("generate ", output_file_name)
wb_output.save(output_file_name)
