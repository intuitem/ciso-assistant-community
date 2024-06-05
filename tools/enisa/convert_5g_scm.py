"""
Simple script to convert ENISA's 5G Security Controls Matrix v1.3 Excel in a CISO Assistant Excel file
Source: https://www.enisa.europa.eu/publications/5g-security-controls-matrix
"""

import openpyxl
import sys
import re
import argparse
from openpyxl.styles import numbers

parser = argparse.ArgumentParser(
    prog="convert_5g_scm",
    description="convert 5G Security Controls Matrix v1.3 Excel file to CISO Assistant Excel file",
)

parser.add_argument("filename", help="name of 5G SCM Excel file")
args = parser.parse_args()
input_file_name = args.filename
output_file_name = "enisa-5g-scm-v1.3.xlsx"

library_copyright = """The Matrix is provided on an ‘as is’ basis. ENISA is not responsible for the information contained in the Matrix, including the use that might be made of this information, or the content of any external sources referenced in the Matrix.
"""
packager = "intuitem"

library_description = """The main goal of the ENISA 5G security controls matrix is to support the national authorities in the EU Member States with implementing the technical measures of the EU’s 5G Cybersecurity toolbox.
Source: https://www.enisa.europa.eu/publications/5g-security-controls-matrix
"""

print("parsing", input_file_name)

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook(input_file_name)
output_table = []
domains = {}
objectives = {}
measures = {}
current_domain_id = ""
current_objective_id = ""

for tab in dataframe:
    print("parsing tab", tab.title)
    title = tab.title
    if title in ("Domains"):
        for row in tab:
            (domain_id, name, description) = (r.value for r in row[0:3])
            if domain_id and re.match(r"D\d+", domain_id):
                domains[domain_id] = (name, description)
    if title in ("Objectives"):
        for row in tab:
            (objective_id, name, description, domain_id) = (r.value for r in row[0:4])
            if objective_id and re.match(r"SO\d+", objective_id):
                objectives[objective_id] = (name, description, domain_id)
    if title in ("Measures"):
        for row in tab:
            (measure_id, _, objective, so_level, _, description) = (
                r.value for r in row[0:6]
            )
            if measure_id and re.match(r"M\d+", measure_id):
                measures["5G-" + measure_id] = (objective, so_level, description)
    if title in ("5GControls"):
        for row in tab:
            (ref_id, description, evidence, _, _, _, domain_id) = (
                r.value for r in row[0:7]
            )
            if ref_id and re.match("SO\d+-\d+", ref_id):
                if domain_id != current_domain_id:
                    current_domain_id = domain_id
                    output_table.append(
                        (
                            "",
                            1,
                            domain_id,
                            domains[domain_id][0],
                            domains[domain_id][1],
                            "",
                        )
                    )
                objective_id = ref_id.split("-")[0]
                if objective_id != current_objective_id:
                    current_objective_id = objective_id
                    output_table.append(
                        (
                            "",
                            2,
                            objective_id,
                            objectives[objective_id][0],
                            objectives[objective_id][1],
                            "",
                        )
                    )
                req_measures = [
                    "1:" + measure_id
                    for measure_id in measures
                    if measures[measure_id][0] == objective_id
                ]
                output_table.append(
                    ("x", 3, ref_id, "", description, ",".join(req_measures))
                )
                output_table.append(("", 4, "", "Evidence", evidence, ""))

print("generating", output_file_name)
wb_output = openpyxl.Workbook()
ws = wb_output.active
ws.title = "library_content"
ws.append(["library_urn", f"urn:{packager.lower()}:risk:library:enisa-5g-scm-v1.3"])
ws.append(["library_version", "1"])
ws.append(["library_locale", "en"])
ws.append(["library_ref_id", "ENISA 5G SCM v1.3"])
ws.append(["library_name", "ENISA 5G Security Control Matrix v1.3"])
ws.append(["library_description", library_description])
ws.append(["library_copyright", library_copyright])
ws.append(["library_provider", "ENISA"])
ws.append(["library_packager", packager])
ws.append(["framework_urn", f"urn:{packager.lower()}:risk:framework:enisa-5g-scm-v1.3"])
ws.append(["framework_ref_id", "ENISA 5G SCM v1.3"])
ws.append(["framework_name", "ENISA 5G Security Control Matrix v1.3"])
ws.append(["framework_description", library_description])
ws.append(
    [
        "reference_control_base_urn",
        f"urn:{packager.lower()}:risk:reference_control:enisa-5g-scm",
        "1",
    ]
)
ws.append(["tab", "reference_controls", "reference_controls"])
ws.append(["tab", "requirements", "requirements"])

ws2 = wb_output.create_sheet("reference_controls")
ws2.append(["ref_id", "category", "description"])
for measure_id in measures:
    (objective, so_level, description) = measures[measure_id]
    ws2.append([measure_id, "process", f"(L{so_level}) {description[3:]}"])

ws1 = wb_output.create_sheet("requirements")
ws1.append(
    [
        "assessable",
        "depth",
        "ref_id",
        "name",
        "description",
        "reference_controls",
        "typical_evidence",
    ]
)
for row in output_table:
    ws1.append(row)

print("generate ", output_file_name)
wb_output.save(output_file_name)
