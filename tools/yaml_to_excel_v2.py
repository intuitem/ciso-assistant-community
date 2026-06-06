"""
YAML Library → Excel v2 Converter
----------------------------------
Converts a CISO Assistant YAML library file into the v2 Excel format
expected by convert_library_v2.py (library_meta / {fw}_meta / {fw}_content tabs).

Usage:
    python yaml_to_excel_v2.py path/to/library.yaml [--output path/to/output.xlsx]
    python yaml_to_excel_v2.py path/to/folder/ --bulk [--output-dir path/to/dir]
"""

import sys
import re
import yaml
import argparse
import unicodedata
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def slugify(text: str) -> str:
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = text.lower().replace(" ", "-")
    return re.sub(r"[^a-z0-9_\-\.]", "_", text)


def sheet_base_name(ref_id: str) -> str:
    """Derive a safe Excel sheet base name from ref_id.

    Excel sheet names are capped at 31 chars. The longest suffix we append is
    '_content' (8 chars), so the base must be at most 23 chars.
    """
    max_base = 31 - len("_content")  # 23
    s = re.sub(r"[^a-zA-Z0-9_-]", "_", ref_id)[:max_base]
    return s


def style_header_row(ws, row_num: int, col_count: int):
    fill = PatternFill("solid", fgColor="4472C4")
    font = Font(bold=True, color="FFFFFF")
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def autofit_columns(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val = str(cell.value) if cell.value else ""
                max_len = max(max_len, min(len(val), 80))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(max_len + 2, 12)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_library_meta(wb, data: dict):
    ws = wb.create_sheet("library_meta")
    rows = [
        ("type", "library"),
        ("urn", data.get("urn", "")),
        ("version", str(data.get("version", "1"))),
        ("locale", data.get("locale", "en")),
        ("ref_id", data.get("ref_id", "")),
        ("name", data.get("name", "")),
        ("description", data.get("description", "")),
        ("copyright", data.get("copyright", "")),
        ("provider", data.get("provider", "")),
        ("packager", data.get("packager", "")),
    ]
    # Extra locale translations
    translations = data.get("translations", {})
    for lang, t in translations.items():
        if t.get("name"):
            rows.append((f"name[{lang}]", t["name"]))
        if t.get("description"):
            rows.append((f"description[{lang}]", t["description"]))

    for r in rows:
        ws.append(r)
    autofit_columns(ws)


def build_framework_meta(wb, fw: dict, sheet_base: str):
    ws = wb.create_sheet(f"{sheet_base}_meta")
    base_urn = fw.get("urn", "").replace(":framework:", ":req_node:")
    rows = [
        ("type", "framework"),
        ("base_urn", base_urn),
        ("urn", fw.get("urn", "")),
        ("ref_id", fw.get("ref_id", "")),
        ("name", fw.get("name", "")),
        ("description", fw.get("description", "")),
    ]
    translations = fw.get("translations", {})
    for lang, t in translations.items():
        if t.get("name"):
            rows.append((f"name[{lang}]", t["name"]))
        if t.get("description"):
            rows.append((f"description[{lang}]", t["description"]))

    for r in rows:
        ws.append(r)
    autofit_columns(ws)


def build_framework_content(wb, fw: dict, sheet_base: str, extra_langs: list):
    ws = wb.create_sheet(f"{sheet_base}_content")

    # Build header
    header = ["assessable", "depth", "ref_id", "name", "description"]
    for lang in extra_langs:
        header += [f"name[{lang}]", f"description[{lang}]"]

    ws.append(header)
    style_header_row(ws, 1, len(header))

    for node in fw.get("requirement_nodes", []):
        assessable = "x" if node.get("assessable") else None
        depth = node.get("depth")
        ref_id = node.get("ref_id", "")
        name = node.get("name", "")
        description = node.get("description", "")

        row = [assessable, depth, ref_id, name, description]
        for lang in extra_langs:
            t = node.get("translations", {}).get(lang, {})
            row.append(t.get("name", ""))
            row.append(t.get("description", ""))

        ws.append(row)

    # Wrap text in description columns
    desc_cols = [5] + [5 + 2 * i for i in range(1, len(extra_langs) + 1)]
    for col_idx in desc_cols:
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx):
            for c in cell:
                c.alignment = Alignment(wrap_text=True, vertical="top")

    autofit_columns(ws)


# ---------------------------------------------------------------------------
# Main conversion
# ---------------------------------------------------------------------------

def convert(yaml_path: Path, output_path: Path):
    with open(yaml_path, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f)

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Detect extra locales
    all_langs = set()
    for lang in data.get("translations", {}).keys():
        all_langs.add(lang)
    fw = data.get("objects", {}).get("framework", {})
    for lang in fw.get("translations", {}).keys():
        all_langs.add(lang)
    for node in fw.get("requirement_nodes", []):
        for lang in node.get("translations", {}).keys():
            all_langs.add(lang)
    # Remove the primary locale
    primary_locale = data.get("locale", "en")
    extra_langs = sorted(all_langs - {primary_locale})

    build_library_meta(wb, data)

    if fw:
        base = sheet_base_name(fw.get("ref_id", data.get("ref_id", "framework")))
        build_framework_meta(wb, fw, base)
        build_framework_content(wb, fw, base, extra_langs)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"✅  Saved: {output_path}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Convert CISO Assistant YAML library to v2 Excel format")
    parser.add_argument("input", help="Path to YAML file or folder (with --bulk)")
    parser.add_argument("--output", help="Output .xlsx path (single file mode)")
    parser.add_argument("--bulk", action="store_true", help="Process all .yaml files in input folder")
    parser.add_argument("--output-dir", help="Output directory for bulk mode")
    args = parser.parse_args()

    if args.bulk:
        in_dir = Path(args.input)
        out_dir = Path(args.output_dir) if args.output_dir else in_dir
        for yaml_file in sorted(in_dir.glob("*.yaml")):
            out_file = out_dir / (yaml_file.stem + ".xlsx")
            try:
                convert(yaml_file, out_file)
            except Exception as e:
                print(f"❌  Failed {yaml_file.name}: {e}", file=sys.stderr)
    else:
        yaml_path = Path(args.input)
        if args.output:
            out_path = Path(args.output)
        else:
            out_path = yaml_path.with_suffix(".xlsx")
        convert(yaml_path, out_path)


if __name__ == "__main__":
    main()
