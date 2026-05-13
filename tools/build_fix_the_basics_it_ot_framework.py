"""
Build the Fix the Basics IT/OT 2026 CISO Assistant framework workbook.

Usage:
    python tools/build_fix_the_basics_it_ot_framework.py
    python tools/build_fix_the_basics_it_ot_framework.py --input tools/fw_it_ot.xlsx --output tools/fix_the_basics_it_ot_2026.xlsx
"""

from __future__ import annotations

import argparse
import re
from collections import OrderedDict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


IT_SHEET = "Fix the Basics - IT (EN)  2026"
OT_SHEET = "Fix the Basics - OT (EN)  2026"

FWK_HEADERS = [
    "assessable",
    "depth",
    "ref_id",
    "name",
    "description",
    "annotation",
    "typical_evidence",
    "implementation_groups",
    "questions",
    "answer",
    "threats",
]

IG_HEADERS = ["ref_id", "name", "description"]
ANSW_HEADERS = ["id", "question_type", "question_choices"]
THREATS_HEADERS = ["ref_id", "name", "description", "annotation"]
URN_PREF_HEADERS = ["prefix_id", "prefix_value"]

LIBRARY_URN = "urn:intuitem:risk:library:fix-the-basics_it-ot_2026"
FRAMEWORK_BASE_URN = "urn:intuitem:risk:req_node:fix-the-basics_it-ot_2026"
FRAMEWORK_URN = "urn:intuitem:risk:framework:fix-the-basics_it-ot_2026"
THREAT_BASE_URN = "urn:intuitem:risk:threat:fix-the-basics_it-ot_2026"
LIBRARY_REF_ID = "Fix-the-Basics_IT-OT_2026"
LIBRARY_NAME = "Fix the Basics - IT & OT"
LIBRARY_DESCRIPTION = "Fix the Basics - IT & OT Framework"
LIBRARY_PROVIDER = "Veolia"
LIBRARY_PACKAGER = "intuitem"
LIBRARY_LOCALE = "en"
LIBRARY_VERSION = "1"
IMPLEMENTATION_GROUPS_DEFINITION = "ig"
ANSWERS_DEFINITION = "answ"

LIBRARY_META_ROWS = [
    ["type", "library"],
    ["urn", LIBRARY_URN],
    ["version", LIBRARY_VERSION],
    ["locale", LIBRARY_LOCALE],
    ["ref_id", LIBRARY_REF_ID],
    ["name", LIBRARY_NAME],
    ["description", LIBRARY_DESCRIPTION],
    ["copyright", LIBRARY_PROVIDER],
    ["provider", LIBRARY_PROVIDER],
    ["packager", LIBRARY_PACKAGER],
]

FWK_META_ROWS = [
    ["type", "framework"],
    ["base_urn", FRAMEWORK_BASE_URN],
    ["urn", FRAMEWORK_URN],
    ["ref_id", LIBRARY_REF_ID],
    ["name", LIBRARY_NAME],
    ["description", LIBRARY_DESCRIPTION],
    ["implementation_groups_definition", IMPLEMENTATION_GROUPS_DEFINITION],
    ["answers_definition", ANSWERS_DEFINITION],
]

ANSW_META_ROWS = [
    ["type", "answers"],
    ["name", ANSWERS_DEFINITION],
]

YES_NO_STARTERS = {
    "am",
    "are",
    "can",
    "could",
    "did",
    "do",
    "does",
    "had",
    "has",
    "have",
    "is",
    "may",
    "might",
    "must",
    "shall",
    "should",
    "was",
    "were",
    "will",
    "would",
}

LEVEL_LABELS = ("not applied", "not-applied", "partial", "standard", "optimized")


@dataclass
class SheetConfig:
    title: str
    prefix: str
    root_name: str
    topic_label: str
    evidence_header: str
    is_ot: bool = False


def normalize_header(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip().casefold()


def text(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def is_na_or_empty(value: str) -> bool:
    return not value.strip() or value.strip().casefold() in {"n/a", "na"}


def get_header_map(ws) -> dict[str, int]:
    headers = {}
    for cell in ws[1]:
        header = normalize_header(cell.value)
        if header:
            headers[header] = cell.column
    return headers


def get_cell(ws, header_map: dict[str, int], header: str, row: int) -> Cell:
    normalized = normalize_header(header)
    if normalized not in header_map:
        raise KeyError(f'Missing column "{header}" in sheet "{ws.title}"')
    return ws.cell(row=row, column=header_map[normalized])


def get_value(ws, header_map: dict[str, int], header: str, row: int) -> str:
    return text(get_cell(ws, header_map, header, row).value)


def split_nonempty_lines(value: str) -> list[str]:
    return [line.strip() for line in value.split("\n") if line.strip()]


def split_before_first_blank(value: str) -> str:
    lines = value.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    kept = []
    for line in lines:
        if not line.strip():
            break
        kept.append(line)
    return "\n".join(kept).strip()


def split_after_first_blank(value: str) -> str:
    lines = value.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    for index, line in enumerate(lines):
        if not line.strip():
            return "\n".join(lines[index + 1 :]).strip()
    return ""


def split_multiple_questions_in_line(line: str) -> list[str]:
    if line.count("?") <= 1:
        return [line.rstrip()]

    parts = []
    start = 0
    question_positions = [match.end() for match in re.finditer(r"\?", line)]
    for end in question_positions[:-1]:
        parts.append(line[start:end].strip())
        start = end
    parts.append(line[start:].strip())
    return [part for part in parts if part]


def normalize_questions(value: str) -> str:
    lines = []
    carry = ""
    for raw_line in value.replace("\r\n", "\n").replace("\r", "\n").split("\n"):
        stripped = raw_line.strip()
        if not stripped:
            continue
        if carry:
            stripped = f"{carry} {stripped}".strip()
            carry = ""

        while "?" in stripped:
            question, stripped = stripped.split("?", 1)
            lines.append(f"{question.strip()}?")
            stripped = stripped.strip()

        if stripped and lines and stripped.startswith("("):
            lines[-1] = f"{lines[-1]} {stripped}"
        elif stripped:
            carry = stripped

    if carry:
        lines.append(carry)
    return "\n".join(lines)


def yes_no_candidate(question: str) -> bool:
    cleaned = question.strip()
    if "?" not in cleaned:
        return False
    cleaned = cleaned.split("?", 1)[0] + "?"
    cleaned = re.sub(r"^[A-Za-z][A-Za-z /&-]{0,35}:\s*", "", cleaned)
    cleaned = re.sub(r"^\(?[a-z0-9]+\)?[.)-]\s*", "", cleaned, flags=re.IGNORECASE)
    match = re.match(r"([A-Za-z']+)", cleaned)
    if match and match.group(1).casefold() in YES_NO_STARTERS:
        return True
    return bool(
        re.search(
            r"\b(am|are|can|could|did|do|does|had|has|have|is|may|might|must|shall|should|was|were|will|would)\b.+\?",
            cleaned,
            flags=re.IGNORECASE,
        )
    )


def build_answer(ref_id: str, sheet_title: str, source_row: int, questions: str, warnings: list[str]) -> str:
    question_lines = [line for line in questions.split("\n") if line.strip()]
    if not question_lines:
        warnings.append(f"{sheet_title} row {source_row} ({ref_id}): empty question")
        return "???"

    answers = []
    for question in question_lines:
        if yes_no_candidate(question):
            answers.append("OSPNA")
        else:
            answers.append("???")
            warnings.append(f"{sheet_title} row {source_row} ({ref_id}): {question}")

    if answers and all(answer == "OSPNA" for answer in answers):
        return "OSPNA"
    return "\n".join(answers)


def bulletize(value: str) -> str:
    rows = []
    for line in split_nonempty_lines(value):
        if line.startswith(("-", "*")):
            rows.append(f"\t{line}")
        else:
            rows.append(f"- {line}")
    return "\n".join(rows)


def markdown_links(cell: Cell) -> str:
    value = text(cell.value)
    if is_na_or_empty(value):
        return ""

    target = cell.hyperlink.target if cell.hyperlink else ""
    rows = []
    for line in split_nonempty_lines(value):
        if target and not re.search(r"\]\([^)]+\)", line):
            rows.append(f"- [{line}]({target})")
        else:
            rows.append(f"- {line}")
    return "\n".join(rows)


def bold_level_headings(value: str) -> str:
    lines = value.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    output = []
    previous_blank = True

    for line in lines:
        stripped = line.strip()
        if not stripped:
            output.append("")
            previous_blank = True
            continue

        if previous_blank:
            colon_index = line.find(":")
            if colon_index >= 0:
                prefix = line[:colon_index].strip()
                suffix = line[colon_index + 1 :].strip()
                output.append(f"**{prefix}:** {suffix}".rstrip())
            elif stripped.casefold() in LEVEL_LABELS:
                output.append(f"**{stripped}**")
            else:
                output.append(stripped)
        else:
            output.append(stripped)

        previous_blank = False

    return "\n".join(output).strip()


def section(title: str, body: str) -> str:
    if is_na_or_empty(body):
        return ""
    return f"## {title}\n\n{body.strip()}"


def build_annotation(ws, header_map: dict[str, int], row: int, config: SheetConfig) -> str:
    parts = []

    if config.is_ot:
        parts.append(section("Operational implementation", get_value(ws, header_map, "Operational implementation", row)))
        parts.append(section("Test description", get_value(ws, header_map, "Test description", row)))
        risks_after_blank = split_after_first_blank(get_value(ws, header_map, "Associated risks", row))
        parts.append(section("Associated risks", risks_after_blank))
    else:
        parts.append(section("Test description", get_value(ws, header_map, "Test description", row)))

    solutions = bulletize(get_value(ws, header_map, "Solutions provided by the Group", row))
    parts.append(section("Solutions provided by the Group", solutions))

    links = markdown_links(get_cell(ws, header_map, "Link to the resources", row))
    parts.append(section("Link to the resources", links))

    level_header = "Level of achievement site" if config.is_ot else "Level of achievement"
    levels = bold_level_headings(get_value(ws, header_map, level_header, row))
    parts.append(section("Level of achievement", levels))

    return "\n\n".join(part for part in parts if part)


def add_ig(ig_rows: OrderedDict[str, list[str]], ref_id: str, name: str) -> None:
    if ref_id and ref_id not in ig_rows:
        ig_rows[ref_id] = [ref_id, name, ""]


def topic_ref_id(prefix: str, number: int) -> str:
    return f"TOPIC-{prefix}_{number:02d}"


def add_threat(threat_rows: OrderedDict[str, list[str]], ref_id: str, name: str) -> None:
    if ref_id and ref_id not in threat_rows:
        threat_rows[ref_id] = [ref_id, name, "", ""]


def parse_threats(value: str, threat_rows: OrderedDict[str, list[str]]) -> str:
    threat_refs = []
    for line in split_nonempty_lines(split_before_first_blank(value)):
        if ":" not in line:
            continue
        ref_id, name = line.split(":", 1)
        ref_id = ref_id.strip()
        name = name.strip()
        if not ref_id:
            continue
        add_threat(threat_rows, ref_id, name)
        threat_refs.append(f"1:{ref_id}")
    return ",\n".join(threat_refs)


def iter_data_rows(ws, header_map: dict[str, int], ref_header: str) -> Iterable[int]:
    ref_col = header_map[normalize_header(ref_header)]
    for row in range(2, ws.max_row + 1):
        if text(ws.cell(row=row, column=ref_col).value):
            yield row


def process_sheet(
    wb_in,
    config: SheetConfig,
    fwk_rows: list[list[str]],
    ig_rows: OrderedDict[str, list[str]],
    threat_rows: OrderedDict[str, list[str]],
    warnings: list[str],
) -> None:
    ws = wb_in[config.title]
    header_map = get_header_map(ws)
    ref_header = "Internal Control Code (CAP)" if not config.is_ot else "Internal Control Code\n(CAP)"
    topic_ids: OrderedDict[str, str] = OrderedDict()

    fwk_rows.append(["", "1", config.prefix, config.root_name, "", "", "", "", "", "", ""])

    for row in iter_data_rows(ws, header_map, ref_header):
        ref_id = get_value(ws, header_map, ref_header, row)
        nist_function = get_value(ws, header_map, "NIST function", row)
        topic = get_value(ws, header_map, "Topics", row)

        nist_ref = f"NIST_{nist_function.upper()}"
        add_ig(ig_rows, nist_ref, f"NIST Function - {nist_function}")

        if topic not in topic_ids:
            topic_ids[topic] = topic_ref_id(config.topic_label, len(topic_ids) + 1)
        topic_ref = topic_ids[topic]
        add_ig(ig_rows, topic_ref, f"Topic ({config.topic_label}) - {topic}")

        questions = normalize_questions(get_value(ws, header_map, "Questions", row))
        answer = build_answer(ref_id, config.title, row, questions, warnings)
        threats = parse_threats(get_value(ws, header_map, "Associated risks", row), threat_rows)

        fwk_rows.append(
            [
                "x",
                "2",
                ref_id,
                "",
                "",
                build_annotation(ws, header_map, row, config),
                get_value(ws, header_map, config.evidence_header, row),
                f"{nist_ref},\n{topic_ref}",
                questions,
                answer,
                threats,
            ]
        )


def write_rows(ws, rows: list[list[str]]) -> None:
    for row in rows:
        ws.append(["" if value is None else str(value) for value in row])


def create_sheet(wb: Workbook, title: str, rows: list[list[str]]):
    ws = wb.create_sheet(title)
    write_rows(ws, rows)
    return ws


def sorted_content_rows(rows: Iterable[list[str]]) -> list[list[str]]:
    return sorted(rows, key=lambda row: str(row[0]).casefold())


def format_workbook(wb: Workbook) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for ws in wb.worksheets:
        has_header = ws.title.endswith("_content")
        if has_header:
            ws.freeze_panes = "A2"
        for row in ws.iter_rows():
            for cell in row:
                cell.number_format = "@"
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if has_header and cell.row == 1 and ws.max_row > 1:
                    cell.font = Font(bold=True)
                    cell.fill = header_fill

        for column_cells in ws.columns:
            letter = get_column_letter(column_cells[0].column)
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[letter].width = max(12, min(70, max_length + 2))


def build(input_path: Path, output_path: Path) -> list[str]:
    wb_in = load_workbook(input_path, data_only=False)
    fwk_rows = [FWK_HEADERS]
    ig_rows: OrderedDict[str, list[str]] = OrderedDict()
    threat_rows: OrderedDict[str, list[str]] = OrderedDict()
    warnings: list[str] = []

    process_sheet(
        wb_in,
        SheetConfig(
            title=IT_SHEET,
            prefix="CY-IT",
            root_name="Fix the Basics - IT",
            topic_label="IT",
            evidence_header="Evidence to be provided during the assessment (any other relevant equivalent evidence can be included)",
        ),
        fwk_rows,
        ig_rows,
        threat_rows,
        warnings,
    )

    process_sheet(
        wb_in,
        SheetConfig(
            title=OT_SHEET,
            prefix="CY-OT",
            root_name="Fix the Basics - OT",
            topic_label="OT",
            evidence_header="Evidence to be provided during the evaluation (any other relevant equivalent evidence can be included)",
            is_ot=True,
        ),
        fwk_rows,
        ig_rows,
        threat_rows,
        warnings,
    )

    wb_out = Workbook()
    wb_out.remove(wb_out.active)
    create_sheet(wb_out, "library_meta", LIBRARY_META_ROWS)
    create_sheet(wb_out, "fwk_meta", FWK_META_ROWS)
    create_sheet(wb_out, "fwk_content", fwk_rows)
    create_sheet(wb_out, "ig_meta", [["type", "implementation_groups"], ["name", IMPLEMENTATION_GROUPS_DEFINITION]])
    create_sheet(wb_out, "ig_content", [IG_HEADERS, *sorted_content_rows(ig_rows.values())])
    create_sheet(wb_out, "answ_meta", ANSW_META_ROWS)
    create_sheet(wb_out, "answ_content", [ANSW_HEADERS, ["OSPNA", "unique_choice", "Optimized\nStandard\nPartial\nN/A"]])
    create_sheet(wb_out, "threats_meta", [["type", "threats"], ["base_urn", THREAT_BASE_URN]])
    create_sheet(wb_out, "threats_content", [THREATS_HEADERS, *sorted_content_rows(threat_rows.values())])
    create_sheet(wb_out, "urn_pref_meta", [["type", "urn_prefix"]])
    create_sheet(wb_out, "urn_pref_content", [URN_PREF_HEADERS, ["1", THREAT_BASE_URN]])

    format_workbook(wb_out)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(output_path)
    return warnings


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", default="fw_it_ot.xlsx", type=Path)
    parser.add_argument("--output", default="fix_the_basics_it_ot_2026.xlsx", type=Path)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    warnings = build(args.input, args.output)

    print(f"Generated: {args.output}")
    if warnings:
        print("\nQuestions not detected as yes/no. The corresponding answers were set to ???:")
        for warning in warnings:
            print(f"- {warning}")
    else:
        print("\nAll questions were detected as yes/no; answers use OSPNA.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
