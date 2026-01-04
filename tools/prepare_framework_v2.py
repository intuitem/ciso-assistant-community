"""
Excel Framework Generator v0.6
--------------------------

This script reads a YAML or Excel configuration file and generates a structured Excel (.xlsx) file
based on a predefined template to create the base structure of a CISO Assistant framework in v2.
It ensures the integrity of required fields, supports optional implementation groups, and performs
validation with user-friendly error messages.

Usage:
    python prepare_framework_v2.py [-i|--input input.yaml|input.xlsx] [-o|--output output.xlsx]

Command-line arguments:

    -i, --input   : Path to the input configuration file.
                    Can be either a YAML (.yaml/.yml) or Excel (.xlsx) file.
                    Default is 'prepare_framework_v2_config.xlsx'.

    -o, --output  : Optional output Excel file name.
                    If not specified, the output file name will be derived automatically.

Dependencies:
    - PyYAML
    - openpyxl

Features:
    - Required field validation
    - URN format enforcement
    - Optional implementation groups with structure validation
    - Support for multiple locales with additional localized columns in the Excel sheets
    - Automatic Excel generation with multiple sheets, including localized content and metadata
    - Informative logs and error handling
    - Excel input support with sheet-based configuration parsing and locale validation
    
Example usage:
    Generate an Excel framework from a YAML configuration file:
        python prepare_framework_v2.py --input config.yaml --output framework.xlsx

    Generate an Excel framework from an existing Excel configuration file:
        python prepare_framework_v2.py --input config.xlsx --output framework_output.xlsx
"""


import re
import sys
import os
import yaml
import argparse
import warnings
from openpyxl import Workbook, load_workbook



# ------------------------ MISC: openpyxl WARNING MANAGER ------------------------

def custom_warning_filter(message, category, filename, lineno, file=None, line=None):
    warn_msg = "Data Validation extension is not supported and will be removed"

    if (category == UserWarning and warn_msg in str(message)):
        return  # Ignore this specific warning
    return warnings.defaultaction  # Apply default behavior for others

warnings.showwarning = custom_warning_filter



# ------------------------ SHARED VALIDATION ------------------------

# Validates that urn_root only contains allowed characters
def validate_urn_root(urn_root):
    if not re.fullmatch(r"[a-z0-9._-]+", urn_root):
        raise ValueError(f"(validate_urn_root) Invalid URN root \"{urn_root}\" : Only lowercase alphanumeric characters, '-', '_', and '.' are allowed.")


# Validates that ref_id only contains allowed characters
def validate_ref_id(ref_id):
    if not re.fullmatch(r"[a-zA-Z0-9._-]+", ref_id):
        raise ValueError(f"Invalid Ref. ID \"{ref_id}\" : Only alphanumeric characters, '-', '_', and '.' are allowed.")


# Checks for required non-empty field
def validate_required_field(field_name, value):
    if value is None or str(value).strip() == "":
        raise ValueError(f"Missing or empty required field: \"{field_name}\"")


# Check if a locale code is valid (e.g. "en", "fr")
def is_valid_locale(locale_str):
    return bool(re.match(r"[a-z0-9]{2}", locale_str))


# Validate that the implementation_groups list is correct and complete
def validate_implementation_groups(impl_groups, context="implementation_groups", is_in_excel_config=False):
    
    if not isinstance(impl_groups, list) or not impl_groups:
        
        if is_in_excel_config:
            raise ValueError(f'(validate_implementation_groups) Table in sheet "{context}" musn\'t be empty.\n'
                             "💡 Tip: Fill this sheet or, if you don't use it, disable it by adding \"#\" in front of the sheet name.")
        else:
            raise ValueError(f'(validate_implementation_groups) Field "{context}" must be a non-empty list if defined.')
    
    impl_groups_ref_ids = []
    
    for i, group in enumerate(impl_groups, start=1):
        if "ref_id" not in group or not str(group["ref_id"]).strip():
            if is_in_excel_config:
                raise ValueError(f'(validate_implementation_groups) Missing or empty \"ref_id\" in table of sheet "{context}" (Row #{i+2})')
            else:
                raise ValueError(f'(validate_implementation_groups) Missing or empty \"ref_id\" in {context} #{i}')
            
        if group["ref_id"] in impl_groups_ref_ids:
            if is_in_excel_config:
                raise ValueError(f'(validate_implementation_groups) ref_id "{group["ref_id"]}" is duplicated in table of sheet "{context}" (Row #{i+2})')
            else:
                raise ValueError(f'(validate_implementation_groups) ref_id "{group["ref_id"]}" is duplicated in {context} (#{i})')
        
        impl_groups_ref_ids.append(group["ref_id"])
        
        try:
            validate_ref_id(group["ref_id"])
        except ValueError as e:
            if is_in_excel_config:
                raise ValueError(f'(validate_implementation_groups) {e} in table of sheet "{context}" (Row #{i+2})')
            else:
                raise ValueError(f'(validate_implementation_groups) {e} in {context} #{i}')

        if "name" not in group or not str(group["name"]).strip():
            if is_in_excel_config:
                raise ValueError(f'(validate_implementation_groups) Missing or empty \"name\" in table of sheet "{context}" (Row #{i+2})')
            else:
                raise ValueError(f'(validate_implementation_groups) Missing or empty \"name\" in {context} #{i}')
            
        if "description" not in group or not str(group["description"]).strip():
            if is_in_excel_config:
                print(f'⚠️  [WARNING] (validate_implementation_groups) Missing or empty \"description\" in table of sheet "{context}" (Row #{i+2})')
            else:
                print(f'⚠️  [WARNING] (validate_implementation_groups) Missing or empty \"description\" in {context} #{i}')



# ------------------------ YAML VALIDATION ------------------------

# Added validation for extra_locales
def validate_yaml_extra_locales(data):
    
    framework_locale = data.get("locale")
    extra_locales = data.get("extra_locales")
    if extra_locales is None:
        return

    if not isinstance(extra_locales, list) or not extra_locales:
        raise ValueError("(validate_yaml_extra_locales) Field \"extra_locales\" must be a non-empty list if defined.")

    impl_groups_main = data.get("implementation_groups") or []
    impl_groups_main_sheet = data.get("implementation_groups_sheet_base_name") or ""
    
    extra_locales_list = []

    for i, locale_entry in enumerate(extra_locales, start=1):
        if not isinstance(locale_entry, dict) or len(locale_entry) != 1:
            raise ValueError(f"(validate_yaml_extra_locales) Each entry in \"extra_locales\" must be a dict with exactly one locale code key (entry #{i})")

        if framework_locale in locale_entry.keys():
            print(f"⚠️  [WARNING] (validate_yaml_extra_locales) Locale \"{framework_locale}\" is already the framework's main locale."
                          f"\n\t     ⏩ Skipping locale \"{framework_locale}\"...")
            continue
        
        for loc in extra_locales_list:
            if loc in locale_entry.keys():
                raise ValueError(f"(validate_yaml_extra_locales) Locale \"{loc}\" is duplicated (#{i})")
        
        extra_locales_list.extend(locale_entry.keys())


        for loc_code, loc_data in locale_entry.items():
            if not is_valid_locale(loc_code):
                raise ValueError(f"(validate_yaml_extra_locales) Invalid locale code \"{loc_code}\" in extra_locales entry #{i}")

            if not isinstance(loc_data, dict) or not loc_data:
                raise ValueError(f"(validate_yaml_extra_locales) Locale data for \"{loc_code}\" must be a non-empty dict (entry #{i})")

            # Optional fields: warn if missing or empty
            for req_field in ["framework_name", "description", "copyright"]:
                if req_field not in loc_data or str(loc_data[req_field]).strip() == "":
                    print(f"⚠️  [WARNING] (validate_yaml_extra_locales) Missing or empty field \"{req_field}\" in extra_locales for locale \"{loc_code}\" (entry #{i})")

            # Validate implementation_groups in locale if present
            loc_impl_groups = loc_data.get("implementation_groups")
            if loc_impl_groups is not None:
                if not impl_groups_main:
                    print(f"⚠️  [WARNING] (validate_yaml_extra_locales) \"implementation_groups\" defined in locale \"{loc_code}\" but missing in framework."
                          "\n\t     ⏩ Skipping locale's \"implementation_groups\"...")
                    continue
                if not impl_groups_main_sheet:
                    print(f"⚠️  [WARNING] (validate_yaml_extra_locales) \"implementation_groups\" defined in locale \"{loc_code}\" but missing \"implementation_groups_sheet_base_name\" field in framework."
                          "\n\t     ⏩ Skipping locale's \"implementation_groups\"...")
                    continue

                validate_implementation_groups(loc_impl_groups, f'implementation_groups in locale "{loc_code}"')

                # Check that each local ref_id exists in the main implementation_groups
                for group in loc_impl_groups:
                    if all(group["ref_id"] != fg["ref_id"] for fg in impl_groups_main):
                        raise ValueError(f'ref_id "{group["ref_id"]}" in locale "{loc_code}" does not exist in framework implementation_groups.')


# Validate all required YAML fields and structure
def validate_yaml_data(data):
    required_fields = [
        "urn_root", "locale", "ref_id", "framework_name", "description",
        "copyright", "provider", "packager", "framework_sheet_base_name"
    ]
    for field in required_fields:
        validate_required_field(field, data.get(field))
    
    try:
        validate_ref_id(data.get("ref_id"))
    except ValueError as e:
        raise ValueError(f'(validate_yaml_data) {e}')
        
        
    if not is_valid_locale(data.get("locale")):
        raise ValueError(f"(validate_yaml_data) Invalid locale code \"{data.get('locale')}\" in \"locale\" entry")

    validate_urn_root(data["urn_root"])

    impl_base = data.get("implementation_groups_sheet_base_name")
    impl_groups = data.get("implementation_groups")

    # Validate implementation_groups if base name is defined
    if impl_base:
        if impl_groups is not None:
            validate_implementation_groups(impl_groups, "implementation_groups")
            print(f"ℹ️  [INFO] Implementation groups sheets will be added using base name: \"{impl_base}\"")
        else:
            print(f"⚠️  [WARNING] (validate_yaml_data) \"implementation_groups_sheet_base_name\" defined but missing \"implementation_groups\" field in YAML file."
                   "\n\t     A blank implementation groups sheet will be created in the output Excel file.")
    # If there's "implementation_groups" but no "implementation_groups_sheet_base_name", raise a warning 
    elif impl_groups:
        print(f"⚠️  [WARNING] (validate_yaml_data) \"implementation_groups\" defined but missing \"implementation_groups_sheet_base_name\" field in YAML file."
              "\n\t     ⏩ Skipping \"implementation_groups\"...")
        

    # Validate extra_locales if present
    validate_yaml_extra_locales(data)



# ------------------------ EXCEL VALIDATION & EXTRACTION ------------------------

# Return True if the sheet is not commented out (i.e. does not start with "#")
def is_sheet_enabled(sheet_name):
    return not sheet_name.startswith("#")


# Extract key/value pairs from the "base" sheet (skipping commented lines and first row)
def extract_kv_from_base_sheet(ws):
    data = {}
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if not row or len(row) < 2:
            continue
        key, value = row[0], row[1]
        if key and not str(key).strip().startswith("#"):
            data[str(key).strip()] = value.strip() if isinstance(value, str) else value
    return data


# Get the locale code from the sheet name suffix (e.g. "_fr" -> "fr")
def extract_locale_suffix(sheet_name):
    match = re.search(r"_([a-z]{2})$", sheet_name)
    return match.group(1) if match else None


# Read rows from a 3-column sheet like "#imp_grp" or "#imp_grp_xx"
def extract_implementation_groups_sheet(ws):
    results = []
    headers = [cell.value for cell in ws[2]]
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not any(row):
            continue
        entry_raw = dict(zip(headers, row))
        entry = {
            k: v for k, v in entry_raw.items()
            if k in {"ref_id", "name", "description"} and v is not None
        }
        results.append(entry)
    return results


# Validate the structure of the Excel configuration file
def validate_excel_data(wb):
    
    if "base" not in wb.sheetnames or not is_sheet_enabled("base"):
        raise ValueError("(validate_excel_data) Missing or commented-out required sheet: \"base\"")

    ws_base = wb["base"]
    data = extract_kv_from_base_sheet(ws_base)

    for field in [
        "urn_root", "locale", "ref_id", "framework_name", "description",
        "copyright", "provider", "packager", "framework_sheet_base_name"
    ]:
        validate_required_field(field, data.get(field))

    validate_ref_id(data.get("ref_id"))
    validate_urn_root(data["urn_root"])
    
    locale_main = data.get('locale')
    
    if not is_valid_locale(locale_main):
        raise ValueError(f"(validate_excel_data) Invalid locale code \"{locale_main}\"")

    impl_base = data.get("implementation_groups_sheet_base_name")
    impl_groups_main = None
    
    if impl_base:
        if impl_base not in wb.sheetnames:
            if f'#{impl_base}' in wb.sheetnames:
                print(f"⚠️  [WARNING] (validate_excel_data) \"implementation_groups_sheet_base_name\" defined but the \"{impl_base}\" sheet is disabled in Excel file."
                   "\n\t     A blank implementation groups sheet will be created in the output Excel file.")
            else:
                print(f"⚠️  [WARNING] (validate_excel_data) \"implementation_groups_sheet_base_name\" defined but missing \"{impl_base}\" sheet in Excel file."
                    "\n\t     A blank implementation groups sheet will be created in the output Excel file.")
        else:
            ws_impl = wb[impl_base]
            impl_groups_main = extract_implementation_groups_sheet(ws_impl)
            validate_implementation_groups(impl_groups_main, impl_base,is_in_excel_config=True)
    elif 'imp_grp' in wb.sheetnames:
        print(f"⚠️  [WARNING] (validate_excel_data) \"imp_grp\" sheet enabled but missing \"implementation_groups_sheet_base_name\" field in Excel file."
              "\n\t     ⏩ Skipping sheet \"imp_grp\"...")
        print("💡 Tip: Remember to remove the \"#\" in front of \"implementation_groups_sheet_base_name\" in the \"base\" sheet if you want to use the implementation groups sheet.")

    base_extra_locales_list = []
    imp_grp_extra_locales_list = []

    for sheet in wb.sheetnames:
        if not is_sheet_enabled(sheet):
            print(f"⏩ [SKIP] (validate_excel_data) Skipped sheet \"{sheet}\"")
            continue
        
        if sheet in ["info", "base", "imp_grp", "_configs"]:
            continue

        if sheet.startswith("base_"):
            loc = sheet.split("base_")[-1].strip()

            # Check locale
            if not extract_locale_suffix(sheet):
                raise ValueError(f"(validate_excel_data) Invalid locale code in sheet name '{sheet}' (parsed as '{loc}')")

            # Check duplicate locale
            if loc in base_extra_locales_list:
                raise ValueError(f"(validate_excel_data) Locale \"{loc}\" is duplicated (Sheet: \"{sheet})\"")
            
            base_extra_locales_list.append(loc)


            if locale_main == loc:
                print(f"⚠️  [WARNING] (validate_excel_data) Locale \"{loc}\" is already the framework's main locale."
                            f"\n\t     ⏩ Skipping \"{sheet}\"...")
                continue
            
            ws = wb[sheet]
            loc_data = extract_kv_from_base_sheet(ws)
            for field in ["framework_name", "description", "copyright"]:
                if not loc_data.get(field, "") or loc_data.get(field, "").strip() == "":
                    print(f"⚠️  [WARNING] (validate_excel_data) Missing or empty \"{field}\" in locale \"{loc}\" (Sheet: \"{sheet}\")")
                    
        elif sheet.startswith("imp_grp_") and impl_base:
            
            if not impl_base:
                print(f"⚠️  [WARNING] (validate_excel_data) \"{sheet}\" defined but missing \"implementation_groups_sheet_base_name\" field."
                       f"\n\t     ⏩ Skipping \"{sheet}\"...")
                continue
            
            loc = sheet.split("imp_grp_")[-1].strip()

            # Check locale
            if not extract_locale_suffix(sheet):
                raise ValueError(f"(validate_excel_data) Invalid locale code in sheet name '{sheet}' (parsed as '{loc}')")
            
            # Check duplicate locale
            if loc in imp_grp_extra_locales_list:
                raise ValueError(f"(validate_excel_data) Locale \"{loc}\" is duplicated (Sheet: \"{sheet})\"")
            
            imp_grp_extra_locales_list.append(loc)


            if locale_main == loc:
                print(f"⚠️  [WARNING] (validate_excel_data) Locale \"{loc}\" is already the framework's main locale."
                            f"\n\t     ⏩ Skipping \"{sheet}\"...")
                continue
            
            if not impl_groups_main:
                print(f"⚠️  [WARNING] (validate_excel_data) \"{sheet}\" defined but \"imp_grp\" sheet is missing.\n"
                    f"\t     ⏩ Skipping \"{sheet}\"...")
                continue
            
            ws = wb[sheet]
            impl_groups_loc = extract_implementation_groups_sheet(ws)
            validate_implementation_groups(impl_groups_loc, sheet, is_in_excel_config=True)
            
            # Check that each localized implementation group ref_id exists in the main implementation_groups
            for group in impl_groups_loc:
                if all(group["ref_id"] != fg["ref_id"] for fg in impl_groups_main):
                    raise ValueError(f'ref_id "{group["ref_id"]}" in locale "{loc}" does not exist in main implementation_groups sheet "{impl_base}"')
        else:
            print(f"⏩❓ [SKIP] (validate_excel_data) Skipped unknown sheet \"{sheet}\"")
            continue
            

# Parses and formats an Excel workbook into a configuration dictionary 
# matching the expected structure of the YAML configuration.
def format_excel_data_as_config(wb):
    
    # Check for required "base" sheet
    if "base" not in wb.sheetnames:
        raise ValueError('Missing required sheet: "base"')

    # Extract key-value pairs from "base" sheet
    base_sheet = wb["base"]
    base_data = extract_kv_from_base_sheet(base_sheet)

    # Extract implementation groups if present
    implementation_groups = []
    if "imp_grp" in wb.sheetnames and not "imp_grp".startswith("#"):
        imp_grp_sheet = wb["imp_grp"]
        implementation_groups = extract_implementation_groups_sheet(imp_grp_sheet)

    # Detect and prepare extra locales
    extra_locales_dict = {}
    for sheet in wb.sheetnames:
        if sheet.startswith("#"):
            continue

        if sheet.startswith("base_"):
            locale = sheet.split("base_")[-1].strip().lower()
        elif sheet.startswith("imp_grp_"):
            locale = sheet.split("imp_grp_")[-1].strip().lower()
        else:
            continue

        if not is_valid_locale(locale):
            raise ValueError(f"Invalid locale code in sheet name: \"{sheet}\"")

        if locale not in extra_locales_dict:
            extra_locales_dict[locale] = {}

        if sheet.startswith("base_"):
            ws = wb[sheet]
            localized_data = extract_kv_from_base_sheet(ws)
            extra_locales_dict[locale].update({
                "framework_name": localized_data.get("framework_name", ""),
                "description": localized_data.get("description", ""),
                "copyright": localized_data.get("copyright", "")
            })
            
            for field in ["framework_name", "description", "copyright"]:
                value = localized_data.get(field, "")
                if not value or str(value).strip() == "":
                    extra_locales_dict[locale].pop(field, None)

        elif sheet.startswith("imp_grp_"):
            ws = wb[sheet]
            localized_imp_groups = extract_implementation_groups_sheet(ws)
            if localized_imp_groups:
                extra_locales_dict[locale]["implementation_groups"] = localized_imp_groups

    # Reformat extra_locales as list of {locale: data}
    extra_locales = [{loc: data} for loc, data in extra_locales_dict.items()] if extra_locales_dict else None

    # Assemble final configuration dictionary
    config = {
        "urn_root": base_data.get("urn_root"),
        "ref_id": base_data.get("ref_id"),
        "locale": base_data.get("locale"),
        "framework_name": base_data.get("framework_name"),
        "description": base_data.get("description"),
        "copyright": base_data.get("copyright"),
        "provider": base_data.get("provider"),
        "packager": base_data.get("packager"),
        "framework_sheet_base_name": base_data.get("framework_sheet_base_name"),
        "implementation_groups_sheet_base_name": base_data.get("implementation_groups_sheet_base_name"),
        "implementation_groups": implementation_groups if implementation_groups else None,
        "extra_locales": extra_locales
    }
    
    # Remove None values
    return {k: v for k, v in config.items() if v is not None}



# ------------------------ YAML MODE ------------------------

# Process a YAML configuration file
def create_excel_from_yaml(yaml_path, output_excel=None):
    if not os.path.isfile(yaml_path):
        raise FileNotFoundError(f"YAML file not found: \"{yaml_path}\"")

    try:
        with open(yaml_path, 'r', encoding='utf-8') as f:
            data = yaml.safe_load(f)
    except Exception as e:
        raise ValueError(f"Unable to load YAML file \"{yaml_path}\": {e}")

    validate_yaml_data(data)

    yaml_output_name = data.get("excel_file_name")
    if output_excel is None:
        if yaml_output_name and str(yaml_output_name).strip():
            output_excel = yaml_output_name.strip()
        else:
            output_excel = "output.xlsx"
    elif yaml_output_name and str(yaml_output_name).strip() and output_excel != yaml_output_name.strip():
        print(f"ℹ️  [INFO] Overriding YAML \"excel_file_name\" with command line argument.")

    create_excel(data, output_excel)



# ------------------------ EXCEL MODE ------------------------

# Process an Excel configuration file
def create_excel_from_excel(excel_path, output_excel=None):
    if not os.path.isfile(excel_path):
        raise FileNotFoundError(f"Excel file not found: \"{excel_path}\"")

    try:
        wb = load_workbook(excel_path, data_only=True)
    except Exception as e:
        raise ValueError(f"Unable to load Excel file \"{excel_path}\": {e}")
    
    validate_excel_data(wb)

    if "base" not in wb.sheetnames:
        raise ValueError("Missing required sheet: \"base\"")

    base_sheet = wb["base"]
    base_data = extract_kv_from_base_sheet(base_sheet)

    excel_output_name = base_data.get("excel_file_name")

    if output_excel is None:
        if excel_output_name and str(excel_output_name).strip():
            output_excel = excel_output_name.strip()
        else:
            output_excel = "output.xlsx"
    elif excel_output_name and str(excel_output_name).strip() and output_excel != excel_output_name.strip():
        print(f"ℹ️  [INFO] Overriding Excel \"excel_file_name\" with command line argument.")
        
    config = format_excel_data_as_config(wb)
    create_excel(config, output_excel)



# ------------------------ BOTH MODE ------------------------

# Create an Excel file with data coming from a YAML (or Excel) configuration file
def create_excel(data, output_excel=None):
    
    # Ensure output filename ends with .xlsx extension
    if not output_excel.lower().endswith(".xlsx"):
        output_excel += ".xlsx"

    urn_root = data["urn_root"]
    locale = data["locale"]
    ref_id = data["ref_id"]
    name = data["framework_name"]
    description = data["description"]
    copyright_ = data["copyright"]
    provider = data["provider"]
    packager = data["packager"]
    framework_sheet_base = data["framework_sheet_base_name"]
    impl_group_base = data.get("implementation_groups_sheet_base_name")
    impl_groups = data.get("implementation_groups", [])

    wb = Workbook()

    # Main sheet: library_meta
    ws1 = wb.active
    ws1.title = "library_meta"
    ws1.append(["type", "library"])
    ws1.append(["urn", f"urn:intuitem:risk:library:{urn_root}"])
    ws1.append(["version", "1"])
    ws1.append(["locale", locale])
    ws1.append(["ref_id", ref_id])
    ws1.append(["name", name])
    ws1.append(["description", description])
    ws1.append(["copyright", copyright_])
    ws1.append(["provider", provider])
    ws1.append(["packager", packager])

    # Add extra_locales to library_meta sheet
    extra_locales = data.get("extra_locales")
    if extra_locales:
        for locale_entry in extra_locales:
            if locale in locale_entry.keys():  # Ignore if the extra locale is the same as the framework's locale
                continue
            for loc_code, loc_data in locale_entry.items():
                # Only add fields that are present and non-empty
                if "framework_name" in loc_data and str(loc_data["framework_name"]).strip():
                    ws1.append([f"name[{loc_code}]", loc_data["framework_name"]])
                if "description" in loc_data and str(loc_data["description"]).strip():
                    ws1.append([f"description[{loc_code}]", loc_data["description"]])
                if "copyright" in loc_data and str(loc_data["copyright"]).strip():
                    ws1.append([f"copyright[{loc_code}]", loc_data["copyright"]])

    # Main sheet: framework_meta
    framework_meta_sheet = wb.create_sheet(f"{framework_sheet_base}_meta")
    framework_meta_sheet.append(["type", "framework"])
    framework_meta_sheet.append(["base_urn", f"urn:intuitem:risk:req_node:{urn_root}"])
    framework_meta_sheet.append(["urn", f"urn:intuitem:risk:framework:{urn_root}"])
    framework_meta_sheet.append(["ref_id", ref_id])
    framework_meta_sheet.append(["name", name])
    framework_meta_sheet.append(["description", description])
    if impl_group_base:
        framework_meta_sheet.append(["implementation_groups_definition", impl_group_base])

    # Add extra_locales to framework_meta sheet
    if extra_locales:
        for locale_entry in extra_locales:
            if locale in locale_entry.keys():  # Ignore if the extra locale is the same as the framework's locale
                continue
            for loc_code, loc_data in locale_entry.items():
                if "framework_name" in loc_data and str(loc_data["framework_name"]).strip():
                    framework_meta_sheet.append([f"name[{loc_code}]", loc_data["framework_name"]])
                if "description" in loc_data and str(loc_data["description"]).strip():
                    framework_meta_sheet.append([f"description[{loc_code}]", loc_data["description"]])

    # Main sheet framework_content
    framework_content_sheet = wb.create_sheet(f"{framework_sheet_base}_content")
    base_columns = ["assessable", "depth", "ref_id", "name", "description", "annotation", "typical_evidence"]
    
    if impl_group_base:
        base_columns.append("implementation_groups")
    
    # Add localized columns if extra_locales is defined
    localized_columns = []
    if extra_locales:
        for locale_entry in extra_locales:
            if locale in locale_entry.keys():  # Ignore if the extra locale is the same as the framework's locale
                continue
            for loc_code in locale_entry.keys():
                localized_columns.extend([
                    f"name[{loc_code}]",
                    f"description[{loc_code}]",
                    f"annotation[{loc_code}]",
                    f"typical_evidence[{loc_code}]"
                ])

    full_columns = base_columns + localized_columns
    framework_content_sheet.append(full_columns)


    # Implementation_groups sheets (main locale only)
    if impl_group_base:
        impl_meta_sheet = wb.create_sheet(f"{impl_group_base}_meta")
        impl_meta_sheet.append(["type", "implementation_groups"])
        impl_meta_sheet.append(["name", impl_group_base])

        impl_content_sheet = wb.create_sheet(f"{impl_group_base}_content")

        # Base columns
        base_header = ["ref_id", "name", "description"]
        extra_cols = []

        # Prepare extra_locales columns if present
        if extra_locales:
            for locale_entry in extra_locales:
                if locale in locale_entry.keys():  # Ignore if the extra locale is the same as the framework's locale
                    continue
                for loc_code, loc_data in locale_entry.items():
                    loc_impl_groups = loc_data.get("implementation_groups")
                    if loc_impl_groups and isinstance(loc_impl_groups, list):
                        extra_cols.append(f"name[{loc_code}]")
                        extra_cols.append(f"description[{loc_code}]")

        full_header = base_header + extra_cols
        impl_content_sheet.append(full_header)

        # ref_id -> row index mapping (starting at row 2)
        ref_id_to_row = {}
        
        if impl_groups:
            for i, group in enumerate(impl_groups, start=2):
                row = [group.get("ref_id", ""), group.get("name", ""), group.get("description", "")]
                impl_content_sheet.append(row)
                ref_id_to_row[group.get("ref_id", "")] = i

            # Dictionary ref_id -> {lang: {name, description}}
            lang_impl_groups_map = {}
            for locale_entry in extra_locales or []:
                if locale in locale_entry.keys():  # Ignore if the extra locale is the same as the framework's locale
                    continue
                for loc_code, loc_data in locale_entry.items():
                    loc_impl_groups = loc_data.get("implementation_groups")
                    if loc_impl_groups:
                        for group in loc_impl_groups:
                            rid = group["ref_id"]
                            if rid not in lang_impl_groups_map:
                                lang_impl_groups_map[rid] = {}
                            lang_impl_groups_map[rid][loc_code] = {
                                "name": group["name"],
                                "description": group.get("description", "")
                            }

            # Extra columns indexes map
            col_index_map = {col: idx + 1 for idx, col in enumerate(full_header)}

            # Fill localized columns with translated data
            for rid, langs_data in lang_impl_groups_map.items():
                if rid in ref_id_to_row:
                    row_num = ref_id_to_row[rid]
                    for lang_code, texts in langs_data.items():
                        name_col = col_index_map.get(f"name[{lang_code}]")
                        desc_col = col_index_map.get(f"description[{lang_code}]")
                        if name_col:
                            impl_content_sheet.cell(row=row_num, column=name_col, value=texts["name"])
                        if desc_col:
                            impl_content_sheet.cell(row=row_num, column=desc_col, value=texts["description"])
                            

    try:
        wb.save(output_excel)
        print(f"✅ [SUCCESS] Excel file saved successfully: \"{output_excel}\"")
    except Exception as e:
        print(f"❌ [ERROR] Failed to save Excel file: {e}")
        print("💡 Tip: Make sure the Excel file is not open in another program and that you have write permission.")
        sys.exit(1)



# ------------------------ MAIN ------------------------

# CLI entry point
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate Excel file from YAML configuration")
    parser.add_argument('-i', '--input', default="prepare_framework_v2_config.xlsx",
                        help="Input YAML or Excel configuration file (default: prepare_framework_v2_config.xlsx)")
    parser.add_argument('-o', '--output', default=None,
                        help="Output Excel file name (optional)")

    args = parser.parse_args()
    input_path = args.input

    try:
        # Automatically determine whether input is YAML or Excel
        if input_path.lower().endswith((".yaml", ".yml")):
            print(f"ℹ️  [INFO] YAML configuration file \"{os.path.basename(input_path)}\" will be used")
            create_excel_from_yaml(input_path, args.output)
        elif input_path.lower().endswith(".xlsx"):
            print(f"ℹ️  [INFO] Excel configuration file \"{os.path.basename(input_path)}\" will be used")
            create_excel_from_excel(input_path, args.output)
        else:
            raise ValueError("Unsupported input file format. Must be .yaml, .yml, or .xlsx")
    except Exception as err:
        print(f"❌ [FATAL ERROR] {err}")
        sys.exit(1)
