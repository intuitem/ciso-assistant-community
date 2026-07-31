"""
v2 Excel Structure Validator

This script validates the structure and consistency of an Excel file following the expected v2 format. It checks:
- The presence and format of *_meta and *_content sheets.
- Required keys and values in meta sheets.
- Structural correctness and references in content sheets.
- Column-level validation against other sheets (e.g. implementation_groups, answers, threats, reference_controls).
- Validity of URNs using prefix definitions and corresponding content sheets.

Usage:
    python check_library_v2.py file.xlsx [--verbose]

Arguments:
    file.xlsx               Path to the Excel file to validate.

    -e, --external-refs     YAML files containing external references mentioned in the library.
                            Use it to check the following columns if necessary : "threats", "reference_controls".
                            Separate external references with commas (e.g., ./threats1.yaml,./refs/ref_ctrl.yaml,../test.yaml)
    -b, --bulk              Enable bulk mode to process all Excel files in a directory.
    --verbose               Display additional information and validation feedback.

The script exits with code 1 and displays an error message if validation fails.
"""



import os
import re
import sys
import yaml
import inspect
import argparse
from enum import Enum
from pathlib import Path
from collections import Counter
from dataclasses import dataclass
from types import MappingProxyType  # Prefer over "frozendict" to avoid importing an external library
from typing import Dict, List, Mapping, Sequence

import pandas as pd
from openpyxl import Workbook, load_workbook



# ─────────────────────────────────────────────────────────────
# CLASSES
# ─────────────────────────────────────────────────────────────

class ConsoleContext: # Maybe rename it to "Logger" and create a "Context/Config" object to carry compatibility & conversion settings, etc. in one object  
    
    def __init__(self):
        
        self.warning_messages: Dict[str, List] = {}
        self.verbose_messages: Dict[str, List] = {}


    # Getters 
    
    def get_sheet_warning_msg(self, sheet_name: str) -> List[str]:
        return self.warning_messages.get(sheet_name)
    
    def get_sheet_verbose_msg(self, sheet_name: str) -> List[str]:
        return self.verbose_messages.get(sheet_name)
    
    def get_all_warning_msg(self) -> Dict[str, List]:
        return self.warning_messages
    
    def get_all_verbose_msg(self) -> Dict[str, List]:
        return self.verbose_messages


    # Setters

    def add_sheet_warning_msg(self, sheet_name: str, msg: str):
        
        if sheet_name in self.warning_messages:
            self.warning_messages[sheet_name].append(msg)
        else:
            self.warning_messages[sheet_name] = [msg]
        return
    
    def add_sheet_verbose_msg(self, sheet_name: str, msg: str):
        
        if sheet_name in self.verbose_messages:
            self.verbose_messages[sheet_name].append(msg)
        else:
            self.verbose_messages[sheet_name] = [msg]
        return


    # Counters (Global)

    def count_all_warnings(self) -> int:
        """Return total number of warning messages."""
        return sum(len(msgs) for msgs in self.warning_messages.values())

    def count_all_verbose(self) -> int:
        """Return total number of verbose messages."""
        return sum(len(msgs) for msgs in self.verbose_messages.values())

    # Counters (Per sheet)

    def count_warnings_for_sheet(self, sheet_name: str) -> int:
        """Return number of warning messages for a sheet (0 if none exist)."""
        return len(self.warning_messages.get(sheet_name, []))

    def count_verbose_for_sheet(self, sheet_name: str) -> int:
        """Return number of verbose messages for a sheet (0 if none exist)."""
        return len(self.verbose_messages.get(sheet_name, []))

    # Aggregators by sheet

    def get_warning_count_by_sheet(self) -> Dict[str, int]:
        """Return dict {sheet_name: warning count}, including 0 where applicable."""
        return {fn: len(self.warning_messages.get(fn, [])) for fn in self._all_sheet_names}

    def get_verbose_count_by_sheet(self) -> Dict[str, int]:
        """Return dict {sheet_name: verbose count}, including 0 where applicable."""
        return {fn: len(self.verbose_messages.get(fn, [])) for fn in self._all_sheet_names}


##### Regex & Characters for Formatting  #####
class CommonSeparatorRegex(Enum):
    LF = r"\n+"
    COMMA_LF = r"[,\n]+"
    SPACE_COMMA_LF = r"[\s,]+"

class CommonLineBreakIndicator(Enum):
    PIPE = "|"


##### Sheet Categories #####
class MetaTypes(Enum):
    LIBRARY = "library"
    FRAMEWORK = "framework"
    THREATS = "threats"
    REFERENCE_CONTROLS = "reference_controls"
    RISK_MATRIX = "risk_matrix"
    REQUIREMENT_MAPPING_SET = "requirement_mapping_set"
    IMPLEMENTATION_GROUPS = "implementation_groups"
    SCORES = "scores"
    ANSWERS = "answers"
    URN_PREFIX = "urn_prefix"

class SheetTypes(Enum):
    META = "_meta"
    CONTENT = "_content"

##### Sheet Properties & Schemas #####


# Libraries without these sheets are invalid
class MandatorySheets(Enum):
    LIBRARY_META = "library_meta"


## > [META]

# > [META] Keys

# [META] {ROOT} Base type for all meta sheet keys. For typing purpose only
class MetaKey(str, Enum):
    pass


# [META] Sheets without these keys are invalid
# It's different from "MetaSheetSchema.expected_keys", as these keys should be everywhere
class MandatoryMetaKeys(MetaKey):
    TYPE = "type"

# [META] Library supported keys
class LibraryMetaKeys(MetaKey):
    URN = "urn"
    VERSION = "version"
    LOCALE = "locale"
    REF_ID = "ref_id"
    NAME = "name"
    DESCRIPTION = "description"
    COPYRIGHT = "copyright"
    PROVIDER = "provider"
    PACKAGER = "packager"
    LABELS = "labels"
    DEPENDENCIES = "dependencies"


# [META] Framework supported keys
class FrameworkMetaKeys(MetaKey):
    URN = "urn"
    BASE_URN = "base_urn"
    REF_ID = "ref_id"
    NAME = "name"
    DESCRIPTION = "description"
    IMPLEMENTATION_GROUPS_DEFINITION = "implementation_groups_definition"
    ANSWERS_DEFINITION = "answers_definition"
    SCORES_DEFINITION = "scores_definition"
    MIN_SCORE = "min_score"
    MAX_SCORE = "max_score"


# [META] Threats supported keys
class ThreatsMetaKeys(MetaKey):
    BASE_URN = "base_urn"


# [META] Reference Controls supported keys
class ReferenceControlsMetaKeys(MetaKey):
    BASE_URN = "base_urn"


# [META] Risk Matrix supported keys
class RiskMatrixMetaKeys(MetaKey):
    URN = "urn"
    REF_ID = "ref_id"
    NAME = "name"
    DESCRIPTION = "description"


# [META] Requirement Mapping Set supported keys
class RequirementMappingSetMetaKeys(MetaKey):
    URN = "urn"
    REF_ID = "ref_id"
    NAME = "name"
    DESCRIPTION = "description"
    SOURCE_FRAMEWORK_URN = "source_framework_urn"
    SOURCE_NODE_BASE_URN = "source_node_base_urn"
    TARGET_FRAMEWORK_URN = "target_framework_urn"
    TARGET_NODE_BASE_URN = "target_node_base_urn"


# [META] Implementation Groups supported keys
class ImplementationGroupsMetaKeys(MetaKey):
    NAME = "name"


# [META] Scores supported keys
class ScoresMetaKeys(MetaKey):
    NAME = "name"


# [META] Answers supported keys
class AnswersMetaKeys(MetaKey):
    NAME = "name"


# [META] URN Prefix supported keys
class URNPrefixMetaKeys(MetaKey):
    # No keys because only "type" is required
    pass


# > [META] Schemas

# Define the required, optional, and translatable keys of a meta sheet
@dataclass(frozen=True)
class MetaSheetSchema:
    key_enum: type[MetaKey]
    expected_keys: tuple[MetaKey, ...]
    optional_keys: tuple[MetaKey, ...] = ()
    translatable_keys: tuple[MetaKey, ...] = ()

    # Check if every keys are from the same Enum family
    def __post_init__(self):
        key_groups = (self.expected_keys, self.optional_keys, self.translatable_keys)

        for keys in key_groups:
            if not all(isinstance(key, self.key_enum) for key in keys):
                raise TypeError(f"All schema keys must come from {self.key_enum.__name__}")

    @staticmethod
    def _to_values(keys: tuple[MetaKey, ...]) -> tuple[str, ...]:
        return tuple(key.value for key in keys)

    @property
    def expected_key_values(self) -> tuple[str, ...]:
        return self._to_values(self.expected_keys)

    @property
    def optional_key_values(self) -> tuple[str, ...]:
        return self._to_values(self.optional_keys)

    @property
    def translatable_key_values(self) -> tuple[str, ...]:
        return self._to_values(self.translatable_keys)


# Map each meta sheet type to its validation schema
META_SHEET_SCHEMAS: Mapping[MetaTypes, MetaSheetSchema] = MappingProxyType({
    MetaTypes.LIBRARY: MetaSheetSchema(
        key_enum=LibraryMetaKeys,
        expected_keys=(
            LibraryMetaKeys.URN, LibraryMetaKeys.VERSION, LibraryMetaKeys.LOCALE,
            LibraryMetaKeys.REF_ID, LibraryMetaKeys.NAME, LibraryMetaKeys.DESCRIPTION,
            LibraryMetaKeys.COPYRIGHT, LibraryMetaKeys.PROVIDER, LibraryMetaKeys.PACKAGER,
        ),
        optional_keys=(LibraryMetaKeys.LABELS, LibraryMetaKeys.DEPENDENCIES),
        translatable_keys=(LibraryMetaKeys.NAME, LibraryMetaKeys.DESCRIPTION),
    ),
    MetaTypes.FRAMEWORK: MetaSheetSchema(
        key_enum=FrameworkMetaKeys,
        expected_keys=(
            FrameworkMetaKeys.URN, FrameworkMetaKeys.REF_ID, FrameworkMetaKeys.NAME,
            FrameworkMetaKeys.DESCRIPTION, FrameworkMetaKeys.BASE_URN,
        ),
        optional_keys=(
            FrameworkMetaKeys.MIN_SCORE, FrameworkMetaKeys.MAX_SCORE, FrameworkMetaKeys.SCORES_DEFINITION,
            FrameworkMetaKeys.IMPLEMENTATION_GROUPS_DEFINITION, FrameworkMetaKeys.ANSWERS_DEFINITION,
        ),
        translatable_keys=(FrameworkMetaKeys.NAME, FrameworkMetaKeys.DESCRIPTION),
    ),
    MetaTypes.THREATS: MetaSheetSchema(
        key_enum=ThreatsMetaKeys,
        expected_keys=(ThreatsMetaKeys.BASE_URN,),
    ),
    MetaTypes.REFERENCE_CONTROLS: MetaSheetSchema(
        key_enum=ReferenceControlsMetaKeys,
        expected_keys=(ReferenceControlsMetaKeys.BASE_URN,),
    ),
    MetaTypes.RISK_MATRIX: MetaSheetSchema(
        key_enum=RiskMatrixMetaKeys,
        expected_keys=(
            RiskMatrixMetaKeys.URN, RiskMatrixMetaKeys.REF_ID, RiskMatrixMetaKeys.NAME,
            RiskMatrixMetaKeys.DESCRIPTION,
        ),
    ),
    MetaTypes.REQUIREMENT_MAPPING_SET: MetaSheetSchema(
        key_enum=RequirementMappingSetMetaKeys,
        expected_keys=(
            RequirementMappingSetMetaKeys.URN, RequirementMappingSetMetaKeys.REF_ID,
            RequirementMappingSetMetaKeys.NAME, RequirementMappingSetMetaKeys.DESCRIPTION,
            RequirementMappingSetMetaKeys.SOURCE_FRAMEWORK_URN, RequirementMappingSetMetaKeys.SOURCE_NODE_BASE_URN,
            RequirementMappingSetMetaKeys.TARGET_FRAMEWORK_URN, RequirementMappingSetMetaKeys.TARGET_NODE_BASE_URN,
        ),
    ),
    MetaTypes.IMPLEMENTATION_GROUPS: MetaSheetSchema(
        key_enum=ImplementationGroupsMetaKeys,
        expected_keys=(ImplementationGroupsMetaKeys.NAME,),
    ),
    MetaTypes.SCORES: MetaSheetSchema(
        key_enum=ScoresMetaKeys,
        expected_keys=(ScoresMetaKeys.NAME,),
    ),
    MetaTypes.ANSWERS: MetaSheetSchema(
        key_enum=AnswersMetaKeys,
        expected_keys=(AnswersMetaKeys.NAME,),
    ),
    MetaTypes.URN_PREFIX: MetaSheetSchema(
        key_enum=URNPrefixMetaKeys,
        expected_keys=(),
    ),
})



## > [CONTENT]

# [CONTENT] {ROOT} Base type for all content sheet columns. For typing purpose only
class ContentColumn(str, Enum):
    pass

# > [CONTENT] Columns

# [CONTENT] Framework supported columns
class FrameworkContentColumns(ContentColumn):
    ASSESSABLE = "assessable"
    DEPTH = "depth"
    REF_ID = "ref_id"
    URN_ID = "urn_id"
    NAME = "name"
    DESCRIPTION = "description"
    ANNOTATION = "annotation"
    TYPICAL_EVIDENCE = "typical_evidence"
    IMPORTANCE = "importance"
    WEIGHT = "weight"
    MIN_SCORE = "min_score"
    MAX_SCORE = "max_score"
    SCORES_DEFINITION = "scores_definition"
    IMPLEMENTATION_GROUPS = "implementation_groups"
    QUESTIONS = "questions"
    ANSWER = "answer"
    DEPENDS_ON = "depends_on"
    CONDITION = "condition"
    THREATS = "threats"
    REFERENCE_CONTROLS = "reference_controls"


# [CONTENT] Threats supported columns
class ThreatsContentColumns(ContentColumn):
    REF_ID = "ref_id"
    NAME = "name"
    DESCRIPTION = "description"
    ANNOTATION = "annotation"


# [CONTENT] Reference Controls supported columns
class ReferenceControlsContentColumns(ContentColumn):
    REF_ID = "ref_id"
    NAME = "name"
    CATEGORY = "category"
    CSF_FUNCTION = "csf_function"
    DESCRIPTION = "description"
    ANNOTATION = "annotation"


# [CONTENT] Risk Matrix supported columns
class RiskMatrixContentColumns(ContentColumn):
    TYPE = "type"
    ID = "id"
    ABBREVIATION = "abbreviation"
    NAME = "name"
    DESCRIPTION = "description"


# [CONTENT] Implementation Groups supported columns
class ImplementationGroupsContentColumns(ContentColumn):
    REF_ID = "ref_id"
    NAME = "name"
    DESCRIPTION = "description"
    DEFAULT_SELECTED = "default_selected"


# [CONTENT] Requirement Mapping Set supported columns
class RequirementMappingSetContentColumns(ContentColumn):
    SOURCE_NODE_ID = "source_node_id"
    TARGET_NODE_ID = "target_node_id"
    RELATIONSHIP = "relationship"
    RATIONALE = "rationale"
    STRENGTH_OF_RELATIONSHIP = "strength_of_relationship"


# [CONTENT] Scores supported columns
class ScoresContentColumns(ContentColumn):
    SCORE = "score"
    NAME = "name"
    DESCRIPTION = "description"
    DESCRIPTION_DOC = "description_doc"


# [CONTENT] Answers supported columns
class AnswersContentColumns(ContentColumn):
    ID = "id"
    QUESTION_TYPE = "question_type"
    QUESTION_CHOICES = "question_choices"
    DESCRIPTION = "description"
    SELECT_IMPLEMENTATION_GROUPS = "select_implementation_groups"
    ADD_SCORE = "add_score"
    COMPUTE_RESULT = "compute_result"
    COLOR = "color"


# [CONTENT] URN Prefix supported columns
class URNPrefixContentColumns(ContentColumn):
    PREFIX_ID = "prefix_id"
    PREFIX_VALUE = "prefix_value"


# > [CONTENT] Constraints

# Define the validation constraints that can be applied to a content column
@dataclass(frozen=True)
class ContentColumnConstraints:
    allowed_values: tuple[str, ...] | None = None
    split_regex: str | CommonSeparatorRegex | None = None
    line_break_indicator: str | CommonLineBreakIndicator | None = None
    max_length: int | None = None
    integer_only: bool = False
    min_value: int | None = None
    max_value: int | None = None
    unique: bool = False


# [CONTENT] Framework column constraints
FRAMEWORK_CONTENT_COLUMN_CONSTRAINTS: Mapping[ContentColumn, ContentColumnConstraints] = MappingProxyType({
    FrameworkContentColumns.DEPTH: ContentColumnConstraints(integer_only=True, min_value=1),
    FrameworkContentColumns.REF_ID: ContentColumnConstraints(unique=True),
    FrameworkContentColumns.NAME: ContentColumnConstraints(max_length=200),
    FrameworkContentColumns.ASSESSABLE: ContentColumnConstraints(allowed_values=("x", "X")),
    FrameworkContentColumns.IMPORTANCE: ContentColumnConstraints(
        allowed_values=("mandatory", "recommended", "nice_to_have"),
    ),
    FrameworkContentColumns.WEIGHT: ContentColumnConstraints(integer_only=True, min_value=1),
    FrameworkContentColumns.MIN_SCORE: ContentColumnConstraints(integer_only=True, min_value=0),
    FrameworkContentColumns.MAX_SCORE: ContentColumnConstraints(integer_only=True, min_value=0),
    FrameworkContentColumns.QUESTIONS: ContentColumnConstraints(
        line_break_indicator=CommonLineBreakIndicator.PIPE,
    ),
    FrameworkContentColumns.CONDITION: ContentColumnConstraints(
        allowed_values=("any", "all", "/"),
        split_regex=CommonSeparatorRegex.LF,
    ),
})


# [CONTENT] Threats column constraints
THREATS_CONTENT_COLUMN_CONSTRAINTS: Mapping[ContentColumn, ContentColumnConstraints] = MappingProxyType({
    ThreatsContentColumns.REF_ID: ContentColumnConstraints(unique=True),
})


# [CONTENT] Reference Controls column constraints
REFERENCE_CONTROLS_CONTENT_COLUMN_CONSTRAINTS: Mapping[ContentColumn, ContentColumnConstraints] = MappingProxyType({
    ReferenceControlsContentColumns.REF_ID: ContentColumnConstraints(unique=True),
    ReferenceControlsContentColumns.CATEGORY: ContentColumnConstraints(
        allowed_values=("policy", "process", "technical", "physical", "procedure"),
    ),
    ReferenceControlsContentColumns.CSF_FUNCTION: ContentColumnConstraints(
        allowed_values=("govern", "identify", "protect", "detect", "respond", "recover"),
    ),
})


# [CONTENT] Risk Matrix column constraints
RISK_MATRIX_CONTENT_COLUMN_CONSTRAINTS: Mapping[ContentColumn, ContentColumnConstraints] = MappingProxyType({
    RiskMatrixContentColumns.TYPE: ContentColumnConstraints(
        allowed_values=("probability", "impact", "risk"),
    ),
})


# [CONTENT] Implementation Groups column constraints
IMPLEMENTATION_GROUPS_CONTENT_COLUMN_CONSTRAINTS: Mapping[ContentColumn, ContentColumnConstraints] = MappingProxyType({
    ImplementationGroupsContentColumns.REF_ID: ContentColumnConstraints(unique=True),
    ImplementationGroupsContentColumns.DEFAULT_SELECTED: ContentColumnConstraints(allowed_values=("x", "X")),
})


# [CONTENT] Requirement Mapping Set column constraints
REQUIREMENT_MAPPING_SET_CONTENT_COLUMN_CONSTRAINTS: Mapping[ContentColumn, ContentColumnConstraints] = MappingProxyType({
    RequirementMappingSetContentColumns.RELATIONSHIP: ContentColumnConstraints(
        allowed_values=("subset", "intersect", "equal", "superset", "not_related"),
    ),
    RequirementMappingSetContentColumns.RATIONALE: ContentColumnConstraints(
        allowed_values=("syntactic", "semantic", "functional"),
    ),
})


# [CONTENT] Scores column constraints
SCORES_CONTENT_COLUMN_CONSTRAINTS: Mapping[ContentColumn, ContentColumnConstraints] = MappingProxyType({
    ScoresContentColumns.SCORE: ContentColumnConstraints(integer_only=True, min_value=0, unique=True),
})


# [CONTENT] Answers column constraints
ANSWERS_CONTENT_COLUMN_CONSTRAINTS: Mapping[ContentColumn, ContentColumnConstraints] = MappingProxyType({
    AnswersContentColumns.ID: ContentColumnConstraints(unique=True),
    AnswersContentColumns.QUESTION_TYPE: ContentColumnConstraints(
        allowed_values=("unique_choice", "multiple_choice", "text", "date"),
    ),
})


# [CONTENT] URN Prefix column constraints
URN_PREFIX_CONTENT_COLUMN_CONSTRAINTS: Mapping[ContentColumn, ContentColumnConstraints] = MappingProxyType({
    URNPrefixContentColumns.PREFIX_ID: ContentColumnConstraints(unique=True),
    URNPrefixContentColumns.PREFIX_VALUE: ContentColumnConstraints(unique=True),
})


# > [CONTENT] Schemas

# Define the required, optional, and translatable columns and their constraints
@dataclass(frozen=True)
class ContentSheetSchema:
    column_enum: type[ContentColumn]
    required_columns: tuple[ContentColumn, ...]
    column_constraints: Mapping[ContentColumn, ContentColumnConstraints]
    optional_columns: tuple[ContentColumn, ...] = ()
    translatable_columns: tuple[ContentColumn, ...] = ()

    # Check if every column is from the same Enum family
    def __post_init__(self):
        column_groups = (self.required_columns, self.optional_columns, self.translatable_columns)

        for columns in column_groups:
            if not all(isinstance(column, self.column_enum) for column in columns):
                raise TypeError(f"All schema columns must come from {self.column_enum.__name__}")

        if not all(isinstance(column, self.column_enum) for column in self.column_constraints):
            raise TypeError(f"All constraint columns must come from {self.column_enum.__name__}")

    @staticmethod
    def _to_values(columns: tuple[ContentColumn, ...]) -> tuple[str, ...]:
        return tuple(column.value for column in columns)

    @property
    def required_column_values(self) -> tuple[str, ...]:
        return self._to_values(self.required_columns)

    @property
    def optional_column_values(self) -> tuple[str, ...]:
        return self._to_values(self.optional_columns)

    @property
    def translatable_column_values(self) -> tuple[str, ...]:
        return self._to_values(self.translatable_columns)


# Map each content sheet type to its validation schema
CONTENT_SHEET_SCHEMAS: Mapping[MetaTypes, ContentSheetSchema] = MappingProxyType({
    MetaTypes.FRAMEWORK: ContentSheetSchema(
        column_enum=FrameworkContentColumns,
        required_columns=(FrameworkContentColumns.DEPTH,),  # "assessable" isn't there because it has a special behavior (column is mandatory, but not the value in column)
        column_constraints=FRAMEWORK_CONTENT_COLUMN_CONSTRAINTS,
        optional_columns=(
            FrameworkContentColumns.REF_ID, FrameworkContentColumns.URN_ID, FrameworkContentColumns.NAME,
            FrameworkContentColumns.DESCRIPTION, FrameworkContentColumns.ANNOTATION, FrameworkContentColumns.TYPICAL_EVIDENCE,
            FrameworkContentColumns.IMPORTANCE, FrameworkContentColumns.WEIGHT, FrameworkContentColumns.MIN_SCORE,
            FrameworkContentColumns.MAX_SCORE, FrameworkContentColumns.SCORES_DEFINITION, FrameworkContentColumns.IMPLEMENTATION_GROUPS,
            FrameworkContentColumns.QUESTIONS, FrameworkContentColumns.ANSWER, FrameworkContentColumns.DEPENDS_ON,
            FrameworkContentColumns.CONDITION, FrameworkContentColumns.THREATS, FrameworkContentColumns.REFERENCE_CONTROLS,
        ),
        translatable_columns=(
            FrameworkContentColumns.NAME, FrameworkContentColumns.DESCRIPTION, FrameworkContentColumns.ANNOTATION,
            FrameworkContentColumns.TYPICAL_EVIDENCE, FrameworkContentColumns.QUESTIONS,
        ),
    ),
    MetaTypes.THREATS: ContentSheetSchema(
        column_enum=ThreatsContentColumns,
        required_columns=(ThreatsContentColumns.REF_ID, ThreatsContentColumns.NAME),
        column_constraints=THREATS_CONTENT_COLUMN_CONSTRAINTS,
        optional_columns=(ThreatsContentColumns.DESCRIPTION, ThreatsContentColumns.ANNOTATION),
        translatable_columns=(
            ThreatsContentColumns.NAME, ThreatsContentColumns.DESCRIPTION, ThreatsContentColumns.ANNOTATION,
        ),
    ),
    MetaTypes.REFERENCE_CONTROLS: ContentSheetSchema(
        column_enum=ReferenceControlsContentColumns,
        required_columns=(ReferenceControlsContentColumns.REF_ID, ReferenceControlsContentColumns.NAME),
        column_constraints=REFERENCE_CONTROLS_CONTENT_COLUMN_CONSTRAINTS,
        optional_columns=(
            ReferenceControlsContentColumns.CATEGORY, ReferenceControlsContentColumns.CSF_FUNCTION, 
            ReferenceControlsContentColumns.DESCRIPTION, ReferenceControlsContentColumns.ANNOTATION,
        ),
        translatable_columns=(
            ReferenceControlsContentColumns.NAME, ReferenceControlsContentColumns.DESCRIPTION,
            ReferenceControlsContentColumns.ANNOTATION,
        ),
    ),
    MetaTypes.RISK_MATRIX: ContentSheetSchema(
        column_enum=RiskMatrixContentColumns,
        required_columns=(
            RiskMatrixContentColumns.TYPE, RiskMatrixContentColumns.ID, RiskMatrixContentColumns.ABBREVIATION,
            RiskMatrixContentColumns.NAME, RiskMatrixContentColumns.DESCRIPTION,
        ),
        column_constraints=RISK_MATRIX_CONTENT_COLUMN_CONSTRAINTS,
        translatable_columns=(
            RiskMatrixContentColumns.ABBREVIATION, RiskMatrixContentColumns.NAME, RiskMatrixContentColumns.DESCRIPTION,
        ),
    ),
    MetaTypes.IMPLEMENTATION_GROUPS: ContentSheetSchema(
        column_enum=ImplementationGroupsContentColumns,
        required_columns=(ImplementationGroupsContentColumns.REF_ID, ImplementationGroupsContentColumns.NAME),
        column_constraints=IMPLEMENTATION_GROUPS_CONTENT_COLUMN_CONSTRAINTS,
        optional_columns=(
            ImplementationGroupsContentColumns.DESCRIPTION, ImplementationGroupsContentColumns.DEFAULT_SELECTED,
        ),
        translatable_columns=(
            ImplementationGroupsContentColumns.NAME, ImplementationGroupsContentColumns.DESCRIPTION,
        ),
    ),
    MetaTypes.REQUIREMENT_MAPPING_SET: ContentSheetSchema(
        column_enum=RequirementMappingSetContentColumns,
        required_columns=(
            RequirementMappingSetContentColumns.SOURCE_NODE_ID, RequirementMappingSetContentColumns.TARGET_NODE_ID,
            RequirementMappingSetContentColumns.RELATIONSHIP,
        ),
        column_constraints=REQUIREMENT_MAPPING_SET_CONTENT_COLUMN_CONSTRAINTS,
        optional_columns=(
            RequirementMappingSetContentColumns.RATIONALE, RequirementMappingSetContentColumns.STRENGTH_OF_RELATIONSHIP,
        ),
    ),
    MetaTypes.SCORES: ContentSheetSchema(
        column_enum=ScoresContentColumns,
        required_columns=(ScoresContentColumns.SCORE, ScoresContentColumns.NAME),
        column_constraints=SCORES_CONTENT_COLUMN_CONSTRAINTS,
        optional_columns=(ScoresContentColumns.DESCRIPTION, ScoresContentColumns.DESCRIPTION_DOC),
        translatable_columns=(
            ScoresContentColumns.NAME, ScoresContentColumns.DESCRIPTION, ScoresContentColumns.DESCRIPTION_DOC,
        ),
    ),
    MetaTypes.ANSWERS: ContentSheetSchema(
        column_enum=AnswersContentColumns,
        required_columns=(AnswersContentColumns.ID, AnswersContentColumns.QUESTION_TYPE),
        column_constraints=ANSWERS_CONTENT_COLUMN_CONSTRAINTS,
        optional_columns=(
            AnswersContentColumns.QUESTION_CHOICES, AnswersContentColumns.DESCRIPTION,
            AnswersContentColumns.SELECT_IMPLEMENTATION_GROUPS, AnswersContentColumns.ADD_SCORE,
            AnswersContentColumns.COMPUTE_RESULT, AnswersContentColumns.COLOR,
        ),
        translatable_columns=(AnswersContentColumns.QUESTION_CHOICES, AnswersContentColumns.DESCRIPTION),
    ),
    MetaTypes.URN_PREFIX: ContentSheetSchema(
        column_enum=URNPrefixContentColumns,
        required_columns=(URNPrefixContentColumns.PREFIX_ID, URNPrefixContentColumns.PREFIX_VALUE),
        column_constraints=URN_PREFIX_CONTENT_COLUMN_CONSTRAINTS,
    ),
})


##### YAML Specific #####
class YAMLSectionTypes(Enum):
    THREATS = "threats"
    REFERENCE_CONTROLS = "reference_controls"


##### URNs #####
class URNObjects(Enum):
    # URN Format : urn:<packager>:risk:<object>:<ref_id>
    
    URN_BEGGINING = "urn"
    URN_3RD_WORD = "risk"   # Because the format is urn:<packager>:risk:<object>:<ref_id>

    LIBRARY = "library"
    FRAMEWORK = "framework"
    THREAT = "threat"
    REFERENCE_CONTROL = "function"
    MATRIX = "matrix"
    REQ_MAPPING_SET = "req_mapping_set"
    REQ_NODE = "req_node"

# CAREFULL : "ANY_VALUE_INDICATOR", "PACKAGER_INDICATOR" and "ID_INDICATOR" are only used in the code to indicate that the user can put any value in a specific location
class URNMetadataFormat(Enum):
    ANY_VALUE_INDICATOR = "<any>"
    PACKAGER_INDICATOR = "<packager>"
    ID_INDICATOR = "<ref_id_or_something_else>"

    LIBRARY_URN = f"{URNObjects.URN_BEGGINING.value}:{PACKAGER_INDICATOR}:{URNObjects.URN_3RD_WORD.value}:{URNObjects.LIBRARY.value}:{ID_INDICATOR}"

    FRAMEWORK_URN = f"{URNObjects.URN_BEGGINING.value}:{PACKAGER_INDICATOR}:{URNObjects.URN_3RD_WORD.value}:{URNObjects.FRAMEWORK.value}:{ID_INDICATOR}"
    FRAMEWORK_BASE_URN = f"{URNObjects.URN_BEGGINING.value}:{PACKAGER_INDICATOR}:{URNObjects.URN_3RD_WORD.value}:{URNObjects.REQ_NODE.value}:{ID_INDICATOR}"

    MAPPING_URN = f"{URNObjects.URN_BEGGINING.value}:{PACKAGER_INDICATOR}:{URNObjects.URN_3RD_WORD.value}:{URNObjects.REQ_MAPPING_SET.value}:{ID_INDICATOR}"
    MAPPING_SOURCE_AND_TARGET_FRAMEWORK_URN = FRAMEWORK_URN
    MAPPING_SOURCE_AND_TARGET_NODE_BASE_URN = FRAMEWORK_BASE_URN

    THREATS_BASE_URN = f"{URNObjects.URN_BEGGINING.value}:{PACKAGER_INDICATOR}:{URNObjects.URN_3RD_WORD.value}:{URNObjects.THREAT.value}:{ID_INDICATOR}"

    REFERENCE_CONTROLS_BASE_URN = f"{URNObjects.URN_BEGGINING.value}:{PACKAGER_INDICATOR}:{URNObjects.URN_3RD_WORD.value}:{URNObjects.REFERENCE_CONTROL.value}:{ID_INDICATOR}"

    MATRIX_URN = f"{URNObjects.URN_BEGGINING.value}:{PACKAGER_INDICATOR}:{URNObjects.URN_3RD_WORD.value}:{URNObjects.MATRIX.value}:{ID_INDICATOR}"


##### Accepted File Types #####
class ValidFileTypes(Enum):
    EXCEL = (".xlsx", ".xlsm", ".xltx", ".xltm")
    YAML = (".yaml", ".yml")


# ─────────────────────────────────────────────────────────────
# MISC
# ─────────────────────────────────────────────────────────────

# Convert a sequence of Enum members or strings into a tuple of string values.
def enum_sequence_to_strings(values: Sequence[Enum | str]) -> tuple[str, ...]:
    return tuple(str(value.value) if isinstance(value, Enum) else value for value in values)


def check_file_validity(files: List[str] | str, filetype_name: str, valid_file_extensions: tuple[str, ...] | None = None, file_context: str = None):
    
    fct_name = get_current_fct_name()

    if type(files) == str:
        files = [files]

    for f in files:
        if not os.path.exists(f):
            raise ValueError(f"({fct_name}) {(f'[{file_context}] ' if file_context else "")}\"{f}\" doesn't exist")

        if not os.path.isfile(f):
            raise ValueError(f"({fct_name}) {(f'[{file_context}] ' if file_context else "")}\"{f}\" isn't a file")

        if valid_file_extensions and not os.path.basename(f).endswith(valid_file_extensions):
            raise ValueError(
                f"({fct_name}) {(f'[{file_context}] ' if file_context else "")}\"{f}\" isn't a valid {filetype_name} file"
                f"\n> 💡 Valid {filetype_name} file formats : " + ", ".join(f"{f}" for f in valid_file_extensions)
            )


# > Useful for "get_yaml_section_from_files()".
# As the base_urn of "threats" and "reference_controls" isn't written directly in the YAML file,
# it'll be deduced by comparing all "threats" (or "reference_controls") URNs.
# If there's only 1 element in the "threats" (or "reference_controls") list, it'll return "None".
# It'll also return "None" if nothing matches between all elements in the list.
def __calculate_base_urn(items: List[Dict]):
    if len(items) < 2:
        return None

    common_parts = items[0].get("urn", "").split(":")

    # Reduce the common prefix by comparing it with every URN
    for item in items[1:]:
        urn_parts = item.get("urn", "").split(":")
        matching_parts = []

        for common_part, urn_part in zip(common_parts, urn_parts):
            if common_part != urn_part:
                break
            matching_parts.append(common_part)

        common_parts = matching_parts
        if not common_parts:
            return None

    return ":".join(common_parts)


# Get the name of the calling function
def get_current_fct_name():
    return inspect.stack()[1][3]


def get_meta_sheets_with_type(wb: Workbook, context: str) -> Dict[str, str]:
    """
    Return a dictionary of all sheets ending with '_meta' and their corresponding type value.
    Format: {sheet_name: type_value}
    """
    
    meta_sheets_with_type = {}

    for sheet_name in wb.sheetnames:
        if not sheet_name.endswith(SheetTypes.META.value):
            continue

        ws = wb[sheet_name]
        df = pd.DataFrame(ws.values)

        if df.shape[1] < 2:
            continue  # not enough columns to contain key/value pairs

        type_value = get_meta_value(df, MandatoryMetaKeys.TYPE, sheet_name, context=context)
        if type_value is not None:
            meta_sheets_with_type[sheet_name] = type_value

    return meta_sheets_with_type


def get_meta_sheets_names_from_type(wb: Workbook, sheet_type: MetaTypes, context: str) -> List[str]:

    meta_sheets = get_meta_sheets_with_type(wb, context)
    sheets = []

    for sheet, sht_type in meta_sheets.items():
        if sht_type != sheet_type.value:
            continue

        sheets.append(sheet)

    return sheets


# Retrieve the value associated with a given key in a meta sheet (2-column format).
def get_meta_value(df: pd.DataFrame, key_name: MetaKey | str, sheet_name: str, required: bool = False, with_row: bool = False, context: str | None = None) -> str | None | tuple[str | None, int | None]:
    
    """
    If with_row=False (default): returns value
    If with_row=True: returns (value, excel_row_number)
    """

    if isinstance(key_name, MetaKey):
        key_name = key_name.value

    context_prefix = f"({context}) " if context else ""

    if df.shape[1] == 0:
        if required:
            raise ValueError(f"{context_prefix}[{sheet_name}] Missing required key \"{key_name}\" because the meta sheet is empty.")
        return (None, None) if with_row else None

    matches = df[df.iloc[:, 0] == key_name]

    if matches.empty:
        if required:
            raise ValueError(f"{context_prefix}[{sheet_name}] Missing required key \"{key_name}\" in meta sheet.")
        return (None, None) if with_row else None

    if len(matches) > 1:
        rows = ", ".join(str(index + 1) for index in matches.index)
        raise ValueError(
            f"{context_prefix}[{sheet_name}] Key \"{key_name}\" appears multiple times at rows {rows}."
            f"\n> 💡 Tip: Remove duplicate rows so each key appears only once in the \"{sheet_name}\" sheet."
        )

    value = matches.iloc[0, 1] if matches.shape[1] > 1 else None    # Return "None" if 2nd column is empty
    row = matches.index[0] + 1  # Excel-style row number

    if pd.isna(value) or str(value).strip() == "":
        if required:
            raise ValueError(f"{context_prefix}[{sheet_name}] Row #{row}: Key \"{key_name}\" is present but has an empty value.")
        return (None, row) if with_row else None

    value = str(value).strip()
    return (value, row) if with_row else value


# Return a list of non-empty, stripped string values from a specified column in a DataFrame.
def get_non_empty_column_values(df: pd.DataFrame, column_name: ContentColumn | str) -> List[str]:

    if isinstance(column_name, ContentColumn):
        column_name = column_name.value

    if column_name not in df.columns:
        raise ValueError(f"Column \"{column_name}\" not found in DataFrame")

    return [
        str(value).strip()
        for value in df[column_name]
        if pd.notna(value) and str(value).strip() != ""
    ]


# Group lines prefixed by a custom line-break indicator with the previous item.
def parse_multiline_with_custom_separator(
    input_value: str,
    line_break_indicator: str | CommonLineBreakIndicator,
) -> List[str]:

    if isinstance(line_break_indicator, CommonLineBreakIndicator):
        line_break_indicator = line_break_indicator.value

    if not line_break_indicator:
        raise ValueError("line_break_indicator cannot be empty")

    items = []

    for raw_line in input_value.splitlines():
        line = raw_line.strip()

        if not line:
            continue

        if line.startswith(line_break_indicator):
            continuation = line[len(line_break_indicator):].strip()

            if not continuation:
                continue

            if items:
                items[-1] += f"\n{continuation}"
            else:
                items.append(continuation)
        else:
            items.append(line)

    return items


# ─────────────────────────────────────────────────────────────
# VALIDATE UTILS
# ─────────────────────────────────────────────────────────────

# Check URN format in a [META] sheet
def validate_urn_type(urn: str, urn_type: URNMetadataFormat, context: str, row: str | int = None):

    split_urn = urn.split(":")
    split_urn_type = urn_type.value.split(":")
    
    INDICATORS = [
        URNMetadataFormat.ANY_VALUE_INDICATOR.value,
        URNMetadataFormat.PACKAGER_INDICATOR.value,
        URNMetadataFormat.ID_INDICATOR.value
    ]

    for idx, urn_type_part in enumerate(split_urn_type):
        
        # If we can put anything, skip
        if urn_type_part in INDICATORS:
            continue
        
        if urn_type_part != split_urn[idx]:
            raise ValueError(
                f"({context if context else 'validate_urn'}){' Row #'+str(row)+':' if row else ""} Invalid URN format \"{urn}\" (Invalid element #{idx+1})"
                f"\n> 💡 Tip: Make sure the URN follows this format: {urn_type.value}"
            )
    

def validate_urn(urn: str, context: str = None, row: str | int = None):
    pattern = r"^urn:([a-z0-9._-]+:)*[a-z0-9._-]+$"
    if not re.fullmatch(pattern, urn):
        raise ValueError(f"({context if context else 'validate_urn'}){' Row #'+str(row)+':' if row else ""} Invalid URN \"{urn}\" : Only lowercase alphanumeric characters, '-', '_', and '.' are allowed. URNs must begin with \"urn:\"")

def validate_ref_id(ref_id: str, context: str = None, row: str | int = None):
    if not re.fullmatch(r"[a-zA-Z0-9._-]+", ref_id):
        raise ValueError(f"({context if context else 'validate_ref_id'}){' Row #'+str(row)+':' if row else ""} Invalid Ref. ID \"{ref_id}\" : Only alphanumeric characters, '-', '_', and '.' are allowed")

def validate_ref_id_with_spaces(ref_id: str, context: str = None, row: str | int = None):
    if not re.fullmatch(r"[a-zA-Z0-9._\- ]+", ref_id):
        raise ValueError(f"({context if context else 'validate_ref_id'}){' Row #'+str(row)+':' if row else ""} Invalid Ref. ID \"{ref_id}\" : Only alphanumeric characters, '-', '_', ' ', and '.' are allowed")

def validate_sheet_name(sheet_name: str, context: str = None):
    if not (sheet_name.endswith(SheetTypes.META.value) or sheet_name.endswith(SheetTypes.CONTENT.value)):
        raise ValueError(f"({context if context else 'validate_sheet_name'}) Invalid sheet name \"{sheet_name}\". Sheet names must end with '{SheetTypes.META.value}' or '{SheetTypes.CONTENT.value}'")

def is_valid_locale(locale_str: str):
    return bool(re.fullmatch(r"[a-z0-9]{2}", locale_str))

def validate_no_spaces(value: str, value_name: str, context: str = None, row: int = None):
    if " " in str(value):
        raise ValueError(f"({context if context else 'validate_no_spaces'}){' Row #' + str(row) + ':' if row is not None else ''} Invalid value for \"{value_name}\": Spaces are not allowed (got \"{value}\")")

def print_sheet_validation(sheet_name: str, verbose: bool = False, ctx: ConsoleContext = None):
        
    if not ctx:
        
        print(f"🟢 [CHECK] Valid sheet: \"{sheet_name}\"")
        
        if verbose:
            print(f"ℹ️   [INFO] Check for warning / verbose messages in the console, if any")
        else:
            print(f"ℹ️   [INFO] Check for warnings in the console, if any")
    else:
        
        sheet_warnings = ctx.get_sheet_warning_msg(sheet_name)
        sheet_verbose = ctx.get_sheet_verbose_msg(sheet_name)
        
        if verbose:
            if sheet_warnings:
                if sheet_verbose:
                    print(f"🟣 [CHECK] Valid sheet with warnings and verbose messages : \"{sheet_name}\" (Warn: {len(sheet_warnings)} / Verb: {len(sheet_verbose)})")
                else:
                    print(f"🟡 [CHECK] Valid sheet with warnings: \"{sheet_name}\" (Warn: {len(sheet_warnings)} / Verb: 0)")
            else:
                if sheet_verbose:
                    print(f"🔵 [CHECK] Valid sheet with verbose messages : \"{sheet_name}\" (Warn: 0 / Verb: {len(sheet_verbose)})")
                else:
                    print(f"🟢 [CHECK] Valid sheet: \"{sheet_name}\"")   
        else:
            if sheet_warnings:
                print(f"🟡 [CHECK] Valid sheet with warnings: \"{sheet_name}\" (Warn: {len(sheet_warnings)})")
            else:
                print(f"🟢 [CHECK] Valid sheet: \"{sheet_name}\"")

def validate_labels(labels_value: str, context: str, row: int):
    """
    Split labels with regex r"[\\s,\\n]+" then ensure each label contains no spaces.
    """
    
    if labels_value is None or str(labels_value).strip() == "":
        return

    raw_labels  = [x for x in re.split(CommonSeparatorRegex.SPACE_COMMA_LF.value, str(labels_value).strip()) if x]

    for label in raw_labels:
        
        label = label.strip()
        
        if not label:
            continue

        validate_no_spaces(label, "labels", context, row)


def validate_integer_value(
    value_or_df: pd.DataFrame | str | int | float,
    sheet_name: str = None,
    column_name: ContentColumn | str = None,
    context: str = None,
    row: int = None,
    value_name: ContentColumn | str = "value",
    positive_only: bool = False,
    min: int = None,
    max: int = None,
):
    """
    Validate integer(s) from:
        - a single value (int/str)
        - OR a DataFrame column if column_name is provided (and value_or_df is a df)
    
    Parameters:
        - value_or_df:
            Either:
            - a single value (int or str) to validate
            - OR a pandas DataFrame if validating a whole column

        - sheet_name (str, optional):
            Name of the Excel sheet (used for error messages).

        - column_name (str, optional):
            Name of the DataFrame column to validate.
            If provided, all non-empty values in this column will be checked.

        - context (str, optional):
            Context string (usually the calling function name) added to error messages.

        - row (int, optional):
            Excel row number of the single value (used only when validating one value).

        - value_name (str, default="value"):
            Logical name of the value being validated (e.g. "version", "score").
            Used in error messages.

        - positive_only (bool, default=False):
            If True, only positive non-zero integers are accepted (> 0).

        - min (int, optional):
            Minimum allowed integer value (inclusive).

        - max (int, optional):
            Maximum allowed integer value (inclusive).

    Rules:
        - Always checks "is integer" (default behavior)
        - If positive_only=True -> checks > 0
        - If min/max provided -> checks (min <= value <= max)

    Collects all invalid values and raises ONE ValueError at the end.
    """

    if isinstance(column_name, ContentColumn):
        column_name = column_name.value
    if isinstance(value_name, ContentColumn):
        value_name = value_name.value

    # --- resolve values list ---
    if column_name is not None:
        df = value_or_df
        if column_name not in df.columns:
            raise ValueError(
                f"{f'({context}) ' if context else ''}[{sheet_name}] Column \"{column_name}\" not found"
            )

        values = []
        rows = []
        for idx, v in df[column_name].items():
            if v is None or (isinstance(v, float) and pd.isna(v)):
                continue
            s = str(v).strip()
            if s == "":
                continue
            values.append(s)
            rows.append(idx + 2)  # Excel-like row number for content sheet
            
    # --- [OR] validate single value ---
    else:
        values = [value_or_df]
        rows = [row]

    invalid_values = []

    # --- validate each value ---
    for v, r in zip(values, rows):
        try:
            v_str = str(v).strip()
            v_int = int(v_str)

            # Reject floats-like strings (e.g. "1.0") if needed

            if positive_only and v_int <= 0:
                raise ValueError

            if min is not None and v_int < min:
                raise ValueError

            if max is not None and v_int > max:
                raise ValueError

        except Exception:
            invalid_values.append((v, r))

    # --- raise grouped error if needed ---
    if invalid_values:
        bounds = []
        if positive_only:
            bounds.append("positive non-zero")
        if min is not None:
            bounds.append(f">= {min}")
        if max is not None:
            bounds.append(f"<= {max}")

        bounds_str = f" ({', '.join(bounds)})" if bounds else ""

        details = []
        for v, r in invalid_values:
            if r is not None:
                details.append(f"Row #{r}: {v}")
            else:
                details.append(str(v))

        location = f"[{sheet_name}] " if sheet_name else ""

        raise ValueError(
            f"{f'({context}) ' if context else ''}{location}Invalid \"{value_name}\" values: must be integer{bounds_str}:\n   - "
            + "\n   - ".join(details)
        )



# ─────────────────────────────────────────────────────────────
# VALIDATE META SHEETS
# ─────────────────────────────────────────────────────────────


# Global Checks ("type" value is checked by default)
def validate_meta_sheet(df: pd.DataFrame, sheet_name: str, expected_keys: Sequence[MetaKey | str] | None, expected_type: MetaTypes, context: str):
    
    # Validate all required keys (excluding "type" which is handled separately)
    
    if expected_keys:
        expected_keys = tuple(key.value if isinstance(key, MetaKey) else key for key in expected_keys)
        for key in expected_keys:
            get_meta_value(df, key, sheet_name, required=True, context=context)

    # Validate presence and value of "type" key
    type_value, type_row = get_meta_value(df, MandatoryMetaKeys.TYPE, sheet_name, required=True, with_row=True, context=context)

    if type_value != expected_type.value:
        raise ValueError(f"({context}) [{sheet_name}] Row #{type_row}: Invalid type \"{type_value}\". Expected \"{expected_type.value}\"")
    

def validate_optional_values_meta_sheet(df: pd.DataFrame, sheet_name: str, optional_keys: Sequence[MetaKey | str], context: str, verbose: bool = False, ctx: ConsoleContext = None):

    if not optional_keys:
        return
    
    optional_keys = tuple(key.value if isinstance(key, MetaKey) else key for key in optional_keys)

    for key in optional_keys:
        value, row = get_meta_value(df, key, sheet_name, with_row=True, context=context)
        
        if row is not None:
            if value is None:
                raise ValueError(f"({context}) [{sheet_name}] Row #{row}: Optional key \"{key}\" is present but has no value"
                                  "\n> 💡 Tip: If you don't need this key, you can simply remove it from the sheet.")

        else:
            if verbose:
                msg = f"💬 ℹ️  [INFO] ({context}) [{sheet_name}] Missing optional key \"{key}\" in meta sheet"
                
                if ctx:
                    ctx.add_sheet_verbose_msg(sheet_name, msg)
                print(msg)


def validate_extra_locales_in_meta(df: pd.DataFrame, sheet_name: str, translatable_keys: Sequence[MetaKey | str], context: str, ctx: ConsoleContext = None):

    translatable_keys = tuple(key.value if isinstance(key, MetaKey) else key for key in translatable_keys)
    keys = df.iloc[:, 0].dropna().astype(str)

    for key in keys:
        match = re.fullmatch(r"(.+)\[(.+)\]", key)  # Match "value_name[locale]"
        if not match:
            continue
        
        base_key, locale = match.groups()
        value, row = get_meta_value(df, key, sheet_name, with_row=True, context=context)

        # Validate locale format
        if not is_valid_locale(locale):
            raise ValueError(
                f"({context}) [{sheet_name}] Row #{row}: Invalid locale \"{locale}\" in key \"{key}\""
                "\n> 💡 Tip: Locale setting must comply with ISO 639 Set 1 (e.g., \"en\", \"fr\"). See https://en.wikipedia.org/wiki/List_of_ISO_639_language_codes"
            )

        # Check if base key exists in the meta sheet
        if base_key not in df.iloc[:, 0].values:
            raise ValueError(
                f"({context}) [{sheet_name}] Row #{row}: Localized key \"{key}\" found, but base key \"{base_key}\" is missing"
                f"\n> 💡 Tip: Add the base key \"{base_key}\" or simply remove the key \"{key}\"."
            )

        # Check that the base key can be translated
        if base_key not in translatable_keys:
            warn_msg = (
                f"⚠️  [WARNING] ({context}) [{sheet_name}] Row #{row}: Localized key \"{key}\" is not supported and will be ignored by CISO Assistant because base key \"{base_key}\" doesn't accept translations"
                f"\n> 💡 Tip: Remove the localized key \"{key}\"."
            )
            print(warn_msg)
            if ctx:
                ctx.add_sheet_warning_msg(sheet_name, warn_msg)
            continue

        # Check that the localized value is not empty
        if value is None:
            raise ValueError(
                f"({context}) [{sheet_name}] Row #{row}: Localized key \"{key}\" is present but has no value"
                "\n> 💡 Tip: If you don't need this key, you can simply remove it from the sheet."
            )


# Check that if the "name" key exists and has a value, and if the corresponding "<name>_content" sheet exists.
def validate_related_content_sheet_from_name_key(wb: Workbook, df: pd.DataFrame, sheet_name: str, name_key: MetaKey | str, context: str):

    value, row = get_meta_value(df, name_key, sheet_name, with_row=True, context=context)
    if value is None:
        return  # 'name' key is missing or empty, skip check

    expected_sheet = f"{value}{SheetTypes.CONTENT.value}"
    if expected_sheet not in wb.sheetnames:
        raise ValueError(
            f"({context}) [{sheet_name}] Row #{row}: Key \"{name_key.value}\" points to missing sheet starting with \"{value}\" (Missing \"{expected_sheet}\")"
            f"\n> 💡 Tip: Make sure the \"{expected_sheet}\" sheet exists or set the right value for key \"{name_key.value}\"."
        )


# Check that framework definition keys point to existing meta sheets.
def _framework_validate_definition_keys(wb: Workbook, df: pd.DataFrame, sheet_name: str, definition_keys: Sequence[FrameworkMetaKeys | str]):

    fct_name = get_current_fct_name()

    for def_key in definition_keys:
        value, row = get_meta_value(df, def_key, sheet_name, with_row=True, context=fct_name)
        if value is None:
            continue

        expected_sheet = f"{value}{SheetTypes.META.value}"
        if expected_sheet in wb.sheetnames:
            continue

        def_key_name = def_key.value if isinstance(def_key, MetaKey) else def_key
        sheet_type = def_key_name.removesuffix("_definition")
        raise ValueError(
            f"({fct_name}) [{sheet_name}] Row #{row}: Key \"{def_key_name}\" points to missing sheet starting with \"{value}\" (Missing \"{expected_sheet}\")"
            f"\n> 💡 Tip: Make sure \"{sheet_type}\" sheets start with \"{value}\", set the right value for key \"{def_key_name}\" or simply remove the key \"{def_key_name}\"."
        )


# Validate optional framework score bounds and ensure min_score <= max_score.
def _framework_validate_meta_min_max_score(df: pd.DataFrame, sheet_name: str):

    fct_name = get_current_fct_name()

    min_score, min_score_row = get_meta_value(df, FrameworkMetaKeys.MIN_SCORE, sheet_name, with_row=True, context=fct_name)
    max_score, max_score_row = get_meta_value(df, FrameworkMetaKeys.MAX_SCORE, sheet_name, with_row=True, context=fct_name)

    if min_score is None and max_score is None:
        return

    if min_score is None or max_score is None:
        missing_key = FrameworkMetaKeys.MIN_SCORE.value if min_score is None else FrameworkMetaKeys.MAX_SCORE.value
        present_key = FrameworkMetaKeys.MAX_SCORE.value if min_score is None else FrameworkMetaKeys.MIN_SCORE.value
        raise ValueError(
            f"({fct_name}) [{sheet_name}] Missing \"{missing_key}\": it is required when \"{present_key}\" is defined."
            f"\n> 💡 Tip: Define both \"{FrameworkMetaKeys.MIN_SCORE.value}\" and \"{FrameworkMetaKeys.MAX_SCORE.value}\", or remove \"{present_key}\"."
        )

    validate_integer_value(min_score, sheet_name, context=fct_name, row=min_score_row, value_name=FrameworkMetaKeys.MIN_SCORE.value, min=0)
    validate_integer_value(max_score, sheet_name, context=fct_name, row=max_score_row, value_name=FrameworkMetaKeys.MAX_SCORE.value, min=0)

    min_score = int(min_score)
    max_score = int(max_score)

    if min_score > max_score:
        raise ValueError(f"({fct_name}) [{sheet_name}] Invalid score range: \"{FrameworkMetaKeys.MIN_SCORE.value}\" ({min_score}) must be less than or equal to \"{FrameworkMetaKeys.MAX_SCORE.value}\" ({max_score})")



# [META] Library {OK}²
def validate_library_meta(df: pd.DataFrame, sheet_name: str, verbose: bool = False, ctx: ConsoleContext = None):

    fct_name = get_current_fct_name()
    
    # Get required, optional, and translatable keys
    expected_type = MetaTypes.LIBRARY
    schema = META_SHEET_SCHEMAS[expected_type]

    # Check required and optional meta values
    validate_meta_sheet(df, sheet_name, schema.expected_keys, expected_type, fct_name)
    validate_optional_values_meta_sheet(df, sheet_name, schema.optional_keys, fct_name, verbose, ctx)

    # URN
    urn_value, urn_row = get_meta_value(df, LibraryMetaKeys.URN, sheet_name, required=True, with_row=True, context=fct_name)
    validate_urn(urn_value, fct_name, urn_row)
    validate_urn_type(urn_value, URNMetadataFormat.LIBRARY_URN, fct_name, urn_row)

    # ref_id
    ref_id_value, ref_id_row = get_meta_value(df, LibraryMetaKeys.REF_ID, sheet_name, required=True, with_row=True, context=fct_name)
    validate_ref_id(ref_id_value, fct_name, ref_id_row)

    # version
    version_value, version_row = get_meta_value(df, LibraryMetaKeys.VERSION, sheet_name, required=True, with_row=True, context=fct_name)
    validate_integer_value(version_value, sheet_name, context=fct_name, row=version_row, value_name=LibraryMetaKeys.VERSION.value, positive_only=True)

    # labels (Optional)
    labels_value, labels_row = get_meta_value(df, LibraryMetaKeys.LABELS, sheet_name, required=False, with_row=True, context=fct_name)
    if labels_value is not None:
        validate_labels(labels_value, fct_name, labels_row)

    # locale
    locale_value, locale_row = get_meta_value(df, LibraryMetaKeys.LOCALE, sheet_name, required=True, with_row=True, context=fct_name)
    if not is_valid_locale(locale_value):
        raise ValueError(
            f"({fct_name}) [{sheet_name}] Row #{locale_row}: Invalid \"{LibraryMetaKeys.LOCALE.value}\" value: \"{locale_value}\""
            "\n> 💡 Tip: Locale setting must comply with ISO 639 Set 1 (e.g., \"en\", \"fr\"). See https://en.wikipedia.org/wiki/List_of_ISO_639_language_codes")

    # Extra locales
    validate_extra_locales_in_meta(df, sheet_name, schema.translatable_keys, fct_name, ctx)

    print_sheet_validation(sheet_name, verbose, ctx)


# [META] Framework {OK}²
def validate_framework_meta(wb: Workbook, df: pd.DataFrame, sheet_name: str, verbose: bool = False, ctx: ConsoleContext = None):

    fct_name = get_current_fct_name()

    # Get required, optional, and translatable keys
    expected_type = MetaTypes.FRAMEWORK
    schema = META_SHEET_SCHEMAS[expected_type]

    # Check required and optional meta values
    validate_meta_sheet(df, sheet_name, schema.expected_keys, expected_type, fct_name)
    validate_optional_values_meta_sheet(df, sheet_name, schema.optional_keys, fct_name, verbose, ctx)

    # URN
    urn_value, urn_row = get_meta_value(df, FrameworkMetaKeys.URN, sheet_name, required=True, with_row=True, context=fct_name)
    validate_urn(urn_value, fct_name, urn_row)
    validate_urn_type(urn_value, URNMetadataFormat.FRAMEWORK_URN, fct_name, urn_row)

    # base_urn
    base_urn_value, base_urn_row = get_meta_value(df, FrameworkMetaKeys.BASE_URN, sheet_name, required=True, with_row=True, context=fct_name)
    validate_urn(base_urn_value, fct_name, base_urn_row)
    validate_urn_type(base_urn_value, URNMetadataFormat.FRAMEWORK_BASE_URN, fct_name, base_urn_row)

    # ref_id
    ref_id_value, ref_id_row = get_meta_value(df, FrameworkMetaKeys.REF_ID, sheet_name, required=True, with_row=True, context=fct_name)
    validate_ref_id(ref_id_value, fct_name, ref_id_row)
    
    # Check that *_definition keys (if present) point to an existing *_meta sheet
    definition_keys = [FrameworkMetaKeys.IMPLEMENTATION_GROUPS_DEFINITION, FrameworkMetaKeys.ANSWERS_DEFINITION, FrameworkMetaKeys.SCORES_DEFINITION]
    _framework_validate_definition_keys(wb, df, sheet_name, definition_keys)
        
    # Validate min_score and max_score if present
    _framework_validate_meta_min_max_score(df, sheet_name)

    # Extra locales
    validate_extra_locales_in_meta(df, sheet_name, schema.translatable_keys, fct_name, ctx)

    print_sheet_validation(sheet_name, verbose, ctx)


# [META] Threats {OK}²
def validate_threats_meta(df: pd.DataFrame, sheet_name: str, verbose: bool = False, ctx: ConsoleContext = None):

    fct_name = get_current_fct_name()

    # Get required, optional, and translatable keys
    expected_type = MetaTypes.THREATS
    schema = META_SHEET_SCHEMAS[expected_type]

    # Check required and optional meta values
    validate_meta_sheet(df, sheet_name, schema.expected_keys, expected_type, fct_name)
    validate_optional_values_meta_sheet(df, sheet_name, schema.optional_keys, fct_name, verbose, ctx)

    # base_urn
    base_urn_value, base_urn_row = get_meta_value(df, ThreatsMetaKeys.BASE_URN, sheet_name, required=True, with_row=True, context=fct_name)
    validate_urn(base_urn_value, fct_name, base_urn_row)
    validate_urn_type(base_urn_value, URNMetadataFormat.THREATS_BASE_URN, fct_name, base_urn_row)

    # Extra locales
    validate_extra_locales_in_meta(df, sheet_name, schema.translatable_keys, fct_name, ctx)

    print_sheet_validation(sheet_name, verbose, ctx)


# [META] Reference Controls {OK}
def validate_reference_controls_meta(df: pd.DataFrame, sheet_name: str, verbose: bool = False, ctx: ConsoleContext = None):

    fct_name = get_current_fct_name()

    # Get required, optional, and translatable keys
    expected_type = MetaTypes.REFERENCE_CONTROLS
    schema = META_SHEET_SCHEMAS[expected_type]

    # Check required and optional meta values
    validate_meta_sheet(df, sheet_name, schema.expected_keys, expected_type, fct_name)
    validate_optional_values_meta_sheet(df, sheet_name, schema.optional_keys, fct_name, verbose, ctx)

    # base_urn
    base_urn_value, base_urn_row = get_meta_value(df, ReferenceControlsMetaKeys.BASE_URN, sheet_name, required=True, with_row=True, context=fct_name)
    validate_urn(base_urn_value, fct_name, base_urn_row)
    validate_urn_type(base_urn_value, URNMetadataFormat.REFERENCE_CONTROLS_BASE_URN, fct_name, base_urn_row)

    # Extra locales
    validate_extra_locales_in_meta(df, sheet_name, schema.translatable_keys, fct_name, ctx)

    print_sheet_validation(sheet_name, verbose, ctx)


# [META] Risk Matrix {OK}²
def validate_risk_matrix_meta(df: pd.DataFrame, sheet_name: str, verbose: bool = False, ctx: ConsoleContext = None):

    fct_name = get_current_fct_name()

    # Get required, optional, and translatable keys
    expected_type = MetaTypes.RISK_MATRIX
    schema = META_SHEET_SCHEMAS[expected_type]

    # Check required and optional meta values
    validate_meta_sheet(df, sheet_name, schema.expected_keys, expected_type, fct_name)
    validate_optional_values_meta_sheet(df, sheet_name, schema.optional_keys, fct_name, verbose, ctx)

    # URN
    urn_value, urn_row = get_meta_value(df, RiskMatrixMetaKeys.URN, sheet_name, required=True, with_row=True, context=fct_name)
    validate_urn(urn_value, fct_name, urn_row)
    validate_urn_type(urn_value, URNMetadataFormat.MATRIX_URN, fct_name, urn_row)

    # ref_id
    ref_id_value, ref_id_row = get_meta_value(df, RiskMatrixMetaKeys.REF_ID, sheet_name, required=True, with_row=True, context=fct_name)
    validate_ref_id(ref_id_value, fct_name, ref_id_row)

    # Extra locales
    validate_extra_locales_in_meta(df, sheet_name, schema.translatable_keys, fct_name, ctx)

    print_sheet_validation(sheet_name, verbose, ctx)


# [META] Implementation Groups {OK}²
def validate_implementation_groups_meta(wb: Workbook, df: pd.DataFrame, sheet_name: str, verbose: bool = False, ctx: ConsoleContext = None):

    fct_name = get_current_fct_name()

    # Get required, optional, and translatable keys
    expected_type = MetaTypes.IMPLEMENTATION_GROUPS
    schema = META_SHEET_SCHEMAS[expected_type]

    # Check required and optional meta values
    validate_meta_sheet(df, sheet_name, schema.expected_keys, expected_type, fct_name)
    validate_optional_values_meta_sheet(df, sheet_name, schema.optional_keys, fct_name, verbose, ctx)

    # name
    validate_related_content_sheet_from_name_key(wb, df, sheet_name, ImplementationGroupsMetaKeys.NAME, fct_name)

    # Extra locales
    validate_extra_locales_in_meta(df, sheet_name, schema.translatable_keys, fct_name, ctx)

    print_sheet_validation(sheet_name, verbose, ctx)


# [META] Mappings {OK}²
def validate_requirement_mapping_set_meta(df: pd.DataFrame, sheet_name: str, verbose: bool, ctx: ConsoleContext = None):

    fct_name = get_current_fct_name()

    # Get required, optional, and translatable keys
    expected_type = MetaTypes.REQUIREMENT_MAPPING_SET
    schema = META_SHEET_SCHEMAS[expected_type]

    # Check required and optional meta values
    validate_meta_sheet(df, sheet_name, schema.expected_keys, expected_type, fct_name)
    validate_optional_values_meta_sheet(df, sheet_name, schema.optional_keys, fct_name, verbose, ctx)

    # URN
    urn_value, urn_row = get_meta_value(df, RequirementMappingSetMetaKeys.URN, sheet_name, required=True, with_row=True, context=fct_name)
    validate_urn(urn_value, fct_name, urn_row)
    validate_urn_type(urn_value, URNMetadataFormat.MAPPING_URN, fct_name, urn_row)

    # source_framework_urn
    source_framework_urn_value, source_framework_urn_row = get_meta_value(df, RequirementMappingSetMetaKeys.SOURCE_FRAMEWORK_URN, sheet_name, required=True, with_row=True, context=fct_name)
    validate_urn_type(source_framework_urn_value, URNMetadataFormat.MAPPING_SOURCE_AND_TARGET_FRAMEWORK_URN, fct_name, source_framework_urn_row)

    # target_framework_urn
    target_framework_urn_value, target_framework_urn_row = get_meta_value(df, RequirementMappingSetMetaKeys.TARGET_FRAMEWORK_URN, sheet_name, required=True, with_row=True, context=fct_name)
    validate_urn_type(target_framework_urn_value, URNMetadataFormat.MAPPING_SOURCE_AND_TARGET_FRAMEWORK_URN, fct_name, target_framework_urn_row)

    # source_node_base_urn
    source_node_base_urn_value, source_node_base_urn_row = get_meta_value(df, RequirementMappingSetMetaKeys.SOURCE_NODE_BASE_URN, sheet_name, required=True, with_row=True, context=fct_name)
    validate_urn_type(source_node_base_urn_value, URNMetadataFormat.MAPPING_SOURCE_AND_TARGET_NODE_BASE_URN, fct_name, source_node_base_urn_row)

    # target_node_base_urn
    target_node_base_urn_value, target_node_base_urn_row = get_meta_value(df, RequirementMappingSetMetaKeys.TARGET_NODE_BASE_URN, sheet_name, required=True, with_row=True, context=fct_name)
    validate_urn_type(target_node_base_urn_value, URNMetadataFormat.MAPPING_SOURCE_AND_TARGET_NODE_BASE_URN, fct_name, target_node_base_urn_row)

    # ref_id
    ref_id_value, ref_id_row = get_meta_value(df, RequirementMappingSetMetaKeys.REF_ID, sheet_name, required=True, with_row=True, context=fct_name)
    validate_ref_id(ref_id_value, fct_name, ref_id_row)

    # Duplicate the list to avoid future modifications of expected_keys affecting the validation
    keys_to_check_no_spaces = [
        RequirementMappingSetMetaKeys.SOURCE_FRAMEWORK_URN, RequirementMappingSetMetaKeys.SOURCE_NODE_BASE_URN,
        RequirementMappingSetMetaKeys.TARGET_FRAMEWORK_URN, RequirementMappingSetMetaKeys.TARGET_NODE_BASE_URN,
    ]

    # Validate that the values for specific keys do not contain spaces
    for key in keys_to_check_no_spaces:
        value, row = get_meta_value(df, key, sheet_name, required=True, with_row=True, context=fct_name)
        validate_no_spaces(str(value), key.value, fct_name, row)

    # Extra locales
    validate_extra_locales_in_meta(df, sheet_name, schema.translatable_keys, fct_name, ctx)

    print_sheet_validation(sheet_name, verbose, ctx)



# [META] Scores {OK}²
def validate_scores_meta(wb: Workbook, df: pd.DataFrame, sheet_name: str, verbose: bool = False, ctx: ConsoleContext = None):

    fct_name = get_current_fct_name()

    # Get required, optional, and translatable keys
    expected_type = MetaTypes.SCORES
    schema = META_SHEET_SCHEMAS[expected_type]

    # Check required and optional meta values
    validate_meta_sheet(df, sheet_name, schema.expected_keys, expected_type, fct_name)
    validate_optional_values_meta_sheet(df, sheet_name, schema.optional_keys, fct_name, verbose, ctx)

    # name
    validate_related_content_sheet_from_name_key(wb, df, sheet_name, ScoresMetaKeys.NAME, fct_name)

    # Extra locales
    validate_extra_locales_in_meta(df, sheet_name, schema.translatable_keys, fct_name, ctx)

    print_sheet_validation(sheet_name, verbose, ctx)


# [META] Answers {OK}²
def validate_answers_meta(wb: Workbook, df: pd.DataFrame, sheet_name: str, verbose: bool = False, ctx: ConsoleContext = None):

    fct_name = get_current_fct_name()

    # Get required, optional, and translatable keys
    expected_type = MetaTypes.ANSWERS
    schema = META_SHEET_SCHEMAS[expected_type]

    # Check required and optional meta values
    validate_meta_sheet(df, sheet_name, schema.expected_keys, expected_type, fct_name)
    validate_optional_values_meta_sheet(df, sheet_name, schema.optional_keys, fct_name, verbose, ctx)

    # name
    validate_related_content_sheet_from_name_key(wb, df, sheet_name, AnswersMetaKeys.NAME, fct_name)

    # Extra locales
    validate_extra_locales_in_meta(df, sheet_name, schema.translatable_keys, fct_name, ctx)

    print_sheet_validation(sheet_name, verbose, ctx)


# [META] URN Prefix {OK}²
def validate_urn_prefix_meta(df: pd.DataFrame, sheet_name: str, verbose: bool = False, ctx: ConsoleContext = None):

    fct_name = get_current_fct_name()

    # Get required, optional, and translatable keys
    expected_type = MetaTypes.URN_PREFIX
    schema = META_SHEET_SCHEMAS[expected_type]

    # Check required and optional meta values
    validate_meta_sheet(df, sheet_name, schema.expected_keys, expected_type, fct_name)
    validate_optional_values_meta_sheet(df, sheet_name, schema.optional_keys, fct_name, verbose, ctx)

    # Extra locales
    validate_extra_locales_in_meta(df, sheet_name, schema.translatable_keys, fct_name, ctx)

    print_sheet_validation(sheet_name, verbose, ctx)



# ─────────────────────────────────────────────────────────────
# VALIDATE CONTENT SHEETS
# ─────────────────────────────────────────────────────────────


# Global Checks
def validate_content_sheet(df: pd.DataFrame, sheet_name: str, required_columns: Sequence[ContentColumn | str], context: str):

    required_columns = enum_sequence_to_strings(required_columns)
    
    required_values_missing = []
    invalid_ref_ids = []
    # Created a tuple on purpose, as some "ref_id" columns might be renamed in the future
    reference_id_columns = {
        FrameworkContentColumns.REF_ID.value, ThreatsContentColumns.REF_ID.value, ReferenceControlsContentColumns.REF_ID.value,
        RiskMatrixContentColumns.ID.value, ImplementationGroupsContentColumns.REF_ID.value, AnswersContentColumns.ID.value,
    }
    
    if required_columns:
        # Check that all required columns are present
        for col in required_columns:
            if col not in df.columns:
                raise ValueError(f"({context}) [{sheet_name}] Missing required column \"{col}\" in content sheet")

        # Check that each non-empty row has a non-empty value (after strip) in every required column
        for col in required_columns:
            for idx, value in df[col].items():
                row_values = df.loc[idx]
                if not any(pd.notna(v) and str(v).strip() != "" for v in row_values):
                    continue  # Skip completely empty rows

                if pd.isna(value) or str(value).strip() == "":
                    required_values_missing.append(idx)
                    # raise ValueError(f"({context}) [{sheet_name}] Row #{idx + 2}: Required value missing in column \"{col}\"")

                if col in reference_id_columns:
                    try:
                        validate_ref_id(str(value), context, idx)
                    except Exception as e:
                        invalid_ref_ids.append((str(value), idx))
            
            if required_values_missing:
                raise ValueError(
                    f"({context}) [{sheet_name}] Required values missing in column \"{col}\":\n   - "
                    + "\n   - ".join(f'Row #{idx + 2}' for idx in required_values_missing)
                )
            
            if invalid_ref_ids:
                raise ValueError(
                    f"({context}) [{sheet_name}] Invalid Ref. IDs found. Only alphanumeric characters, '-', '_', and '.' are allowed :\n   - "
                    + "\n   - ".join(f'Row #{idx + 2}: {value}' for value, idx in invalid_ref_ids)
                )


def validate_optional_columns_content_sheet(df: pd.DataFrame, sheet_name: str, optional_columns: Sequence[ContentColumn | str], context: str, verbose: bool = False, ctx: ConsoleContext = None):

    optional_columns = enum_sequence_to_strings(optional_columns)
    
    for col in optional_columns:
        
        # If optional column missing
        if col not in df.columns:
            if verbose:
                msg = f"💬 ℹ️  [INFO] ({context}) [{sheet_name}] Missing optional column \"{col}\" in meta sheet"
                
                if ctx:
                    ctx.add_sheet_verbose_msg(sheet_name, msg)
                print(msg)
                
            continue

        # Check if the entire column is empty (i.e., all values are NaN or blank)
        is_entirely_empty = all(pd.isna(val) or str(val).strip() == "" for val in df[col])

        if is_entirely_empty:
            if verbose:
                msg = (f"💬 ℹ️  [INFO] ({context}) [{sheet_name}] Optional column \"{col}\" is present but entirely empty")
                        # "\n> 💡 Tip: If you don't need this column, you can simply remove it from the sheet.")
                if ctx:
                    ctx.add_sheet_verbose_msg(sheet_name, msg)
                print(msg)


# Ensure that either all columns from the given list are present or none of them are.
def validate_columns_presence_together(df: pd.DataFrame, column_names: Sequence[ContentColumn | str], sheet_name: str, context: str = None):

    context = context or "validate_columns_presence_together"
    column_names = enum_sequence_to_strings(column_names)
    present_columns = [column for column in column_names if column in df.columns]

    if not present_columns or len(present_columns) == len(column_names):
        return

    missing_columns = [column for column in column_names if column not in df.columns]
    raise ValueError(
        f"({context}) [{sheet_name}] Columns {', '.join(f'\"{column}\"' for column in column_names)} must be present together. "
        f"Missing: {', '.join(f'\"{column}\"' for column in missing_columns)}."
        f"\n> 💡 Tip: Add the missing column(s), or remove the existing column(s): "
        f"{', '.join(f'\"{column}\"' for column in present_columns)}."
    )


# Ensure that each non-empty cell in a column does not exceed the specified character limit.
def validate_column_max_length(df: pd.DataFrame, column_name: ContentColumn | str, max_length: int, sheet_name: str, context: str = None):

    context = context or "validate_column_max_length"
    if isinstance(column_name, ContentColumn):
        column_name = column_name.value

    if column_name not in df.columns:
        return

    values_exceeding_limit = []

    for index, value in df[column_name].items():
        if pd.isna(value) or str(value).strip() == "":
            continue

        value_length = len(str(value))
        if value_length > max_length:
            values_exceeding_limit.append((index + 2, value_length))

    if values_exceeding_limit:
        raise ValueError(
            f'({context}) [{sheet_name}] Values in column "{column_name}" exceed the maximum length of {max_length} characters:\n   - '
            + "\n   - ".join(f"Row #{row}: {length} characters" for row, length in values_exceeding_limit)
            + f'\n> 💡 Tip: Shorten each "{column_name}" value to {max_length} characters or fewer.'
        )


# Check that values in each column from the given list are unique. Raise error or emit warning if duplicates are found
def validate_unique_column_values(df: pd.DataFrame, column_names: Sequence[ContentColumn | str], sheet_name: str, context: str = None, warn_only: bool = False, ctx: ConsoleContext = None):

    context = context or "validate_unique_column_values"
    column_names = enum_sequence_to_strings(column_names)

    for column_name in column_names:
        if column_name not in df.columns:
            raise ValueError(f"({context}) [{sheet_name}] Column \"{column_name}\" not found in sheet")

        # Drop rows with empty or whitespace-only values before checking duplicates
        column_series = df[column_name].dropna().astype(str).map(str.strip)
        column_series = column_series[column_series != ""]

        # Re-index the filtered series to map back to original DataFrame indices
        filtered_df = df.loc[column_series.index]
        duplicates = filtered_df[column_name][filtered_df[column_name].duplicated(keep=False)]

        if not duplicates.empty:
            duplicate_rows = duplicates.index + 2  # Excel-like row number (1-based + header)
            duplicate_values = duplicates.unique()
            quoted_values = ', '.join(f'"{str(val)}"' for val in duplicate_values)

            msg = (
                f"({context}) [{sheet_name}] Duplicate value(s) found in column \"{column_name}\": {quoted_values}"
                f"\n> Rows: {', '.join(map(str, duplicate_rows))}"
            )

            if warn_only:
                msg = f"⚠️  [WARNING] {msg}"
                print(msg)
                if ctx:
                    ctx.add_sheet_warning_msg(sheet_name, msg)
            else:
                raise ValueError(msg)


# Pass "wb" when the sheet requires type-specific validation.
# It allows the function to retrieve the "_content" sheet type from its corresponding "_meta" sheet and apply the appropriate checks.
def validate_extra_locales_in_content(df: pd.DataFrame, sheet_name: str, context: str, ctx: ConsoleContext = None, verbose: bool = False, wb: Workbook = None):

    for col in df.columns:
        match = re.fullmatch(r"(.+)\[(.+)\]", str(col))  # Match "column_name[locale]"
        if not match:
            continue
        
        base_col, locale = match.groups()

        # Validate locale format
        if not is_valid_locale(locale):
            raise ValueError(
                f"({context}) [{sheet_name}] Column \"{col}\": Invalid locale \"{locale}\""
                "\n> 💡 Tip: Locale setting must comply with ISO 639 Set 1 (e.g., \"en\", \"fr\"). See https://en.wikipedia.org/wiki/List_of_ISO_639_language_codes"
            )

        # Check if base column exists
        if base_col not in df.columns:
            raise ValueError(
                f"({context}) [{sheet_name}] Column \"{col}\": Localized column found, but base column \"{base_col}\" is missing"
                f"\n> 💡 Tip: Add the base column \"{base_col}\" or simply remove the column \"{col}\"."
            )

        # If column exists but is entirely empty, emit a warning
        non_empty_found = any(pd.notna(val) and str(val).strip() != "" for val in df[col])
        if not non_empty_found:
            if verbose:
                msg = (
                    f"💬 ℹ️  [INFO] ({context}) [{sheet_name}] Column \"{col}\": Localized column is present but entirely empty"
                    "\n> 💡 Tip: If you don't need this column, you can simply remove it from the sheet."
                )
                if ctx:
                    ctx.add_sheet_verbose_msg(sheet_name, msg)
                print(msg)


        ##### Specific sheet Checking #####
        content_sheet_type = get_content_sheet_type(wb, sheet_name, context) if wb is not None else None
        
        # In framework sheets, translated questions must contain the same number of elements as the base "questions" value on the same row.
        if content_sheet_type == MetaTypes.FRAMEWORK.value and base_col == FrameworkContentColumns.QUESTIONS.value:
            questions_constraints = CONTENT_SHEET_SCHEMAS[MetaTypes.FRAMEWORK].column_constraints[FrameworkContentColumns.QUESTIONS]
            validate_cell_line_count_alignment(df, base_col, col, sheet_name, context,
                cmp_can_be_empty=True,
                ref_line_break_indicator=questions_constraints.line_break_indicator,
                cmp_line_break_indicator=questions_constraints.line_break_indicator,
                allow_single_cmp=False,
            )


# Return the name of a "_content" sheet by removing the trailing "_content" in the given sheet name.
def get_content_sheet_base_name(content_sheet_name: str) -> str:
    if not content_sheet_name.endswith(SheetTypes.CONTENT.value):
        raise ValueError(f"Invalid sheet name: \"{content_sheet_name}\" does not end with \"{SheetTypes.CONTENT.value}\"")

    base_name = re.sub(r'_content$', '', content_sheet_name)
    return base_name


# Return the type declared in the meta sheet corresponding to a content sheet.
def get_content_sheet_type(wb: Workbook, content_sheet_name: str, context: str) -> str:
    base_name = get_content_sheet_base_name(content_sheet_name)
    meta_sheet_name = f"{base_name}{SheetTypes.META.value}"

    if meta_sheet_name not in wb.sheetnames:
        raise ValueError(
            f"({context}) [{content_sheet_name}] No corresponding meta sheet found (Missing \"{meta_sheet_name}\")"
        )

    meta_sheets_with_type = get_meta_sheets_with_type(wb, context)

    if meta_sheet_name not in meta_sheets_with_type:
        raise ValueError(f"({context}) [{meta_sheet_name}] Missing or empty \"{MandatoryMetaKeys.TYPE.value}\" field in meta sheet")

    return meta_sheets_with_type[meta_sheet_name]


# Replace the suffix of each sheet name in the list with the target sheet type suffix. Valid suffixes are defined in SheetTypes.
def get_corresponding_type_sheet_names(sheet_names: List[str], sheet_type: SheetTypes) -> List[str]:

    suffixes = [t.value for t in SheetTypes]
    result = []

    for name in sheet_names:
        matched = False

        for suffix in suffixes:
            if name.endswith(suffix):
                # Replace the exact matching suffix at the end of the name with the target suffix
                new_name = re.sub(re.escape(suffix) + r'$', sheet_type.value, name)
                result.append(new_name)
                matched = True
                break

        if not matched:
            raise ValueError(f"Invalid sheet name: \"{name}\" does not end with a known sheet type suffix {suffixes}")

    return result


# Check if a content sheet is referenced in any 'framework' meta sheet via a specific meta field (e.g., 'scores_definition')
def check_content_sheet_usage_in_frameworks(wb: Workbook, sheet_name: str, meta_field: FrameworkMetaKeys | str, fct_name: str, ctx: ConsoleContext = None) -> List[str]:
    """
    Args:
        wb (Workbook): The Excel workbook.
        sheet_name (str): Name of the current content sheet.
        meta_field (FrameworkMetaKeys | str): The meta key in framework sheets that should reference this sheet.
        fct_name (str): Name of the calling function (used in messages).
        ctx (ConsoleContext, optional): Context object for collecting warnings.
    """

    if isinstance(meta_field, MetaKey):
        meta_field = meta_field.value

    sheet_base_name = get_content_sheet_base_name(sheet_name)
    meta_sheets = get_meta_sheets_with_type(wb, fct_name)
    frameworks_with_reference = []

    for sheet, sheet_type in meta_sheets.items():
        if sheet_type != MetaTypes.FRAMEWORK.value:
            continue

        sheet_df = pd.DataFrame(wb[sheet].values)
        meta_value = get_meta_value(sheet_df, meta_field, sheet, context=fct_name)

        if meta_value == sheet_base_name:
            frameworks_with_reference.append(sheet)

    if frameworks_with_reference:
        print(f"ℹ️  [INFO] ({fct_name}) [{sheet_name}] Sheet referenced by the sheet(s): {', '.join(f'\"{s}\"' for s in frameworks_with_reference)}")
    else:
        warn_msg = (
            f"⚠️  [WARNING] ({fct_name}) [{sheet_name}] This sheet is not referenced in any sheet of type \"{MetaTypes.FRAMEWORK.value}\" via the field \"{meta_field}\""
            f"\n> 💡 Tip: Set \"{meta_field}\" in your framework meta sheet to \"{sheet_base_name}\" if needed."
        )
        print(warn_msg)
        if ctx:
            ctx.add_sheet_warning_msg(sheet_name, warn_msg)

    return frameworks_with_reference


# Check whether each ID is used in at least one framework sheet. Emit a warning if any IDs are unused.
def check_unused_ids_in_frameworks(wb: Workbook, df_ids: pd.DataFrame, id_column: ContentColumn | str, target_column: ContentColumn | str, frameworks_sheet_names: List[str], sheet_name: str, context: str, ctx: ConsoleContext = None, verbose: bool = False, emit_messages: bool = True) -> List[str]:

    if isinstance(id_column, ContentColumn):
        id_column = id_column.value
    if isinstance(target_column, ContentColumn):
        target_column = target_column.value

    ids_to_check = get_non_empty_column_values(df_ids, id_column)
    unused_ids = []

    for _id in ids_to_check:
        found = False
        for fw_sheet in frameworks_sheet_names:
            values = list(wb[fw_sheet].values)
            if not values:
                continue  # skip empty sheets

            # Convert each cells as raw text (or "None" if empty)
            # Line added to avoid the problem were numbers (like "1") are converted into floats (1.0)
            header = [str(c).strip() if c is not None else None for c in values[0]]
            rows = [[str(c).strip() if c is not None else None for c in row] for row in values[1:]]
            
            df_fw = pd.DataFrame(rows, columns=header)  # use header

            if target_column not in df_fw.columns:
                continue

            for cell in df_fw[target_column].dropna().astype(str):
                if pd.isna(cell):
                    continue

                entries = [entry.strip() for entry in re.split(CommonSeparatorRegex.COMMA_LF.value, str(cell)) if entry.strip()]

                if _id in entries:
                    found = True
                    break  # No need to keep looking in this sheet

            if found:
                break  # Found in one sheet : Stop checking this ID

        if not found:
            unused_ids.append(_id)

    if unused_ids and emit_messages:
        msg = (
            f"⚠️  [WARNING] ({context}) [{sheet_name}] The following ID(s) from column \"{id_column}\" are not used in any framework sheet:\n   - "
            f"{'\n   - '.join(f'{x}' for x in unused_ids)}\n"
            "> 💡 Tip: Use these IDs in a framework sheet, or remove them if not needed."
        )
        print(msg)
        if ctx:
            ctx.add_sheet_warning_msg(sheet_name, msg)
    elif not unused_ids and emit_messages and verbose:
        msg = (f"💬 ℹ️  [INFO] ({context}) [{sheet_name}] All ID(s) from column \"{id_column}\" are used in framework sheets")
        print(msg)
        if ctx:
            ctx.add_sheet_verbose_msg(sheet_name, msg)

    return unused_ids


# Ensure that at least one implementation group marked as "default_selected" is used in a CONTENT framework sheet 
def _implementation_groups_check_unused_default_ids_in_frameworks(wb: Workbook, df: pd.DataFrame, frameworks_sheet_names: List[str], sheet_name: str):

    fct_name = get_current_fct_name()

    if ImplementationGroupsContentColumns.DEFAULT_SELECTED.value not in df.columns:
        return

    default_selected_mask = (df[ImplementationGroupsContentColumns.DEFAULT_SELECTED.value].fillna("").astype(str).str.strip().isin(["x", "X"]))
    default_selected_df = df.loc[default_selected_mask].copy()
    default_selected_ids = get_non_empty_column_values(default_selected_df, ImplementationGroupsContentColumns.REF_ID)

    if not default_selected_ids:
        return

    unused_default_ids = check_unused_ids_in_frameworks(wb, default_selected_df, ImplementationGroupsContentColumns.REF_ID, FrameworkContentColumns.IMPLEMENTATION_GROUPS, frameworks_sheet_names, sheet_name, fct_name, emit_messages=False)

    if unused_default_ids == default_selected_ids:
        default_ids = ", ".join(f'"{_id}"' for _id in default_selected_ids)
        raise ValueError(
            f'({fct_name}) [{sheet_name}] None of the implementation groups marked as \"{ImplementationGroupsContentColumns.DEFAULT_SELECTED.value}\" ({default_ids}) are used in a framework content sheet. '
            "This will result in an empty framework in CISO Assistant."
            f'\n> 💡 Tip: Add at least one of these \"{ImplementationGroupsContentColumns.DEFAULT_SELECTED.value}\" implementation groups to the "{FrameworkContentColumns.IMPLEMENTATION_GROUPS.value}" column of your framework content sheet, or remove the "{ImplementationGroupsContentColumns.DEFAULT_SELECTED.value}" column.'
        )


# Validate that all non-empty values in a specific column are in the allowed list. Ignores blank or whitespace-only cells.
def validate_allowed_column_values(
    df: pd.DataFrame,
    column_name: ContentColumn | str,
    allowed_values: Sequence[str],
    sheet_name: str,
    context: str = None,
    warn_only: bool = False,
    ctx: ConsoleContext = None,
    split_regex: str | CommonSeparatorRegex = None,
):
    """
    Args:
        df: The DataFrame to validate.
        column_name: The name of the column to check.
        allowed_values: A list of allowed string values.
        sheet_name: Name of the Excel sheet.
        context: Optional context string (e.g., function name).
        warn_only: If True, warnings are printed instead of raising errors.
        ctx: Optional ConsoleContext to collect warning messages.
        split_regex: Optional regex used to split a cell into multiple values (e.g. r"[\\n,]+").
                     If provided, each split entry is validated independently.
    """

    context = context or "validate_allowed_column_values"
    if isinstance(column_name, ContentColumn):
        column_name = column_name.value

    if column_name not in df.columns:
        return

    invalid_values = []

    # ───────────────────────────────────────────────────────────────
    # Case 1: Single-value cells (default behavior)
    # ───────────────────────────────────────────────────────────────
    if not split_regex:
        cleaned_series = df[column_name].dropna().map(lambda x: str(x).strip())
        cleaned_series = cleaned_series[cleaned_series != ""]

        invalid_mask = ~cleaned_series.isin(allowed_values)

        if invalid_mask.any():
            invalid_series = cleaned_series[invalid_mask]
            invalid_entries = invalid_series.unique()

            quoted_values = ', '.join(f'"{v}"' for v in invalid_entries)

            details = "\n   - " + "\n   - ".join(
                f"Row #{idx + 2} : \"{val}\""
                for idx, val in invalid_series.items()
            )

            msg = (
                f"({context}) [{sheet_name}] Invalid value(s) found in column \"{column_name}\": {quoted_values}"
                f"{details}"
                f"\n> Allowed values are: {', '.join(f'\"{v}\"' for v in allowed_values)}"
            )

            if warn_only:
                msg = f"⚠️  [WARNING] {msg}"
                print(msg)
                if ctx:
                    ctx.add_sheet_warning_msg(sheet_name, msg)
            else:
                raise ValueError(msg)

        return

    # ───────────────────────────────────────────────────────────────
    # Case 2: Multi-value cells (split by regex)
    # ───────────────────────────────────────────────────────────────
    
    if isinstance(split_regex, CommonSeparatorRegex):
        split_regex = split_regex.value

    for idx, cell_value in df[column_name].dropna().items():
        cell_str = str(cell_value).strip()
        if not cell_str:
            continue

        entries = [x.strip() for x in re.split(split_regex, cell_str) if x.strip()]

        for i, entry in enumerate(entries, start=1):
            if entry not in allowed_values:
                invalid_values.append((idx, i, entry))

    if invalid_values:
        invalid_entries = sorted(set(v for _, _, v in invalid_values))

        quoted_values = ", ".join(f'"{v}"' for v in invalid_entries)

        details = "\n   - " + "\n   - ".join(
            f"Row #{idx + 2} ; Cell line #{line_idx} : \"{value}\""
            for idx, line_idx, value in invalid_values
        )

        msg = (
            f"({context}) [{sheet_name}] Invalid value(s) found in column \"{column_name}\": {quoted_values}"
            f"{details}"
            f"\n> Allowed values are: {', '.join(f'\"{v}\"' for v in allowed_values)}"
        )

        if warn_only:
            msg = f"⚠️  [WARNING] {msg}"
            print(msg)
            if ctx:
                ctx.add_sheet_warning_msg(sheet_name, msg)
        else:
            raise ValueError(msg)


# Validate that two columns have coherent internal line counts (split by regex) and are both empty or both filled.
def validate_cell_line_count_alignment(
    df: pd.DataFrame,
    ref_column: ContentColumn | str,
    cmp_column: ContentColumn | str,
    sheet_name: str,
    context: str = None,
    split_regex: str | CommonSeparatorRegex = CommonSeparatorRegex.LF,
    cmp_can_be_empty: bool = False,
    ref_line_break_indicator: str | CommonLineBreakIndicator = None,
    cmp_line_break_indicator: str | CommonLineBreakIndicator = None,
    allow_single_cmp: bool = True,
):
    """
    Checks the match of the number of "internal lines" (split regex) between 2 columns.

    Rules (per Excel line):
    1) If ref cell is empty => cmp cell must be empty.
    2) If cmp cell is empty => ref cell must be empty.
    3) If both are not empty:
       - cmp must have the same internal line count as ref.
       - if allow_single_cmp=True, cmp may also contain a single internal line.

    If "ref_line_break_indicator" or "cmp_line_break_indicator" is provided,
    prefixed lines in the corresponding column are grouped with the previous element.

    The function collects all incoherences and raises ONE ValueError at the end.
    """

    context = context or "validate_cell_line_count_alignment"
    if isinstance(ref_column, ContentColumn):
        ref_column = ref_column.value
    if isinstance(cmp_column, ContentColumn):
        cmp_column = cmp_column.value

    # Skip if columns are missing (consistent with other validators)
    if ref_column not in df.columns or cmp_column not in df.columns:
        return

    # Allow Enum (CommonRegexSeparator) or string
    if isinstance(split_regex, CommonSeparatorRegex):
        split_regex = split_regex.value

    errors = []

    def _split_lines(cell_str: str) -> List[str]:
        return [x.strip() for x in re.split(split_regex, cell_str) if x.strip()]

    def _parse_lines(
        cell_str: str,
        line_break_indicator: str | CommonLineBreakIndicator = None,
    ) -> List[str]:
        if line_break_indicator is not None:
            return parse_multiline_with_custom_separator(cell_str, line_break_indicator)

        return _split_lines(cell_str)

    for idx, row in df.iterrows():
        ref_raw = row.get(ref_column, "")
        cmp_raw = row.get(cmp_column, "")

        ref_str = "" if pd.isna(ref_raw) else str(ref_raw).strip()
        cmp_str = "" if pd.isna(cmp_raw) else str(cmp_raw).strip()

        excel_row = idx + 2  # header offset

        # --- Rule 1: ref empty => cmp must be empty
        if not ref_str:
            if cmp_str:
                errors.append(
                    f'Row #{excel_row}: "{ref_column}" is empty but "{cmp_column}" is not'
                )
            continue

        # --- Rule 2: cmp empty => ref must be empty (unless cmp_can_be_empty=True)
        if not cmp_str:
            if not cmp_can_be_empty:
                errors.append(
                    f'Row #{excel_row}: "{cmp_column}" is empty but "{ref_column}" is not'
                )
            continue

        # --- Rule 3: both not empty => cmp must have 1 line OR ref_count lines
        ref_lines = _parse_lines(ref_str, ref_line_break_indicator)
        cmp_lines = _parse_lines(cmp_str, cmp_line_break_indicator)

        ref_count = len(ref_lines)
        cmp_count = len(cmp_lines)

        valid_cmp_counts = {ref_count}
        if allow_single_cmp:
            valid_cmp_counts.add(1)

        if cmp_count not in valid_cmp_counts:
            expected_count = f"1 or {ref_count}" if allow_single_cmp and ref_count != 1 else str(ref_count)
            errors.append(
                f'Row #{excel_row}: "{ref_column}" has {ref_count} line(s) but "{cmp_column}" has {cmp_count} '
                f'(expected {expected_count})'
            )

    if errors:
        raise ValueError(
            f'({context}) [{sheet_name}] Invalid line alignment between "{ref_column}" and "{cmp_column}":\n   - '
            + "\n   - ".join(errors)
        )


# Check whether each Prefix URN ID is used in at least one framework sheet. Emit a warning if any IDs are unused.
def _URN_prefix_check_unused_ids_in_frameworks(wb: Workbook, df_ids: pd.DataFrame, frameworks_sheet_names: List[str], sheet_name: str, context: str, ctx: ConsoleContext = None, verbose: bool = False):

    target_columns = [FrameworkContentColumns.THREATS.value, FrameworkContentColumns.REFERENCE_CONTROLS.value]
    id_column = URNPrefixContentColumns.PREFIX_ID.value
    ids_to_check = get_non_empty_column_values(df_ids, id_column)
    unused_ids = []

    for _id in ids_to_check:
        found = False

        for fw_sheet in frameworks_sheet_names:
            values = list(wb[fw_sheet].values)
            if not values:
                continue  # skip empty sheet

            df_fw = pd.DataFrame(values[1:], columns=values[0])  # headers

            for target_column in target_columns:
                if target_column not in df_fw.columns:
                    continue

                for cell in df_fw[target_column]:
                    if pd.isna(cell):
                        continue
                    entries = [entry.strip() for entry in str(cell).split(",") if entry.strip()]
                    prefix_parts = [entry.split(":", 1)[0].strip() for entry in entries if ":" in entry]

                    if _id in prefix_parts:
                        found = True
                        break  # Found : No need to continue on this column

                if found:
                    break  # Found : No need to continue on this sheet

            if found:
                break  # Found : Skip to next ID

        if not found:
            unused_ids.append(_id)

    if unused_ids:
        msg = (
            f"⚠️  [WARNING] ({context}) [{sheet_name}] The following Prefix ID(s) from column \"{id_column}\" are not used in any framework sheet: "
            f"{', '.join(f'\"{x}\"' for x in unused_ids)}\n"
            "> 💡 Tip: Use these Prefix IDs in a framework sheet, or remove them if not needed."
        )
        print(msg)
        if ctx:
            ctx.add_sheet_warning_msg(sheet_name, msg)
    elif verbose:
        msg = (
            f"💬 ℹ️  [INFO] ({context}) [{sheet_name}] All Prefix ID(s) from column \"{id_column}\" are used in framework sheets"
        )
        print(msg)
        if ctx:
            ctx.add_sheet_verbose_msg(sheet_name, msg)


# Check whether URN Prefix IDs are used in framework content sheets, or warn if that's not the case.
def _URN_prefix_validate_ids_usage_in_frameworks(wb: Workbook, df: pd.DataFrame, sheet_name: str, ctx: ConsoleContext = None, verbose: bool = False):

    fct_name = get_current_fct_name()

    # 1. Get "framework" content sheets
    framework_sheets = get_meta_sheets_names_from_type(wb, MetaTypes.FRAMEWORK, fct_name)
    framework_sheets = get_corresponding_type_sheet_names(framework_sheets, SheetTypes.CONTENT)

    # 2. Check if every Prefix IDs are actually used in "framework" sheets
    if framework_sheets:
        _URN_prefix_check_unused_ids_in_frameworks(wb, df, framework_sheets, sheet_name, fct_name, ctx, verbose)
        return

    msg = (
        f"⚠️  [WARNING] ({fct_name}) [{sheet_name}] This sheet is not used in any framework sheet"
        "\n> 💡 Tip: You can remove this sheet and its meta sheet if you are not using it"
    )
    print(msg)
    if ctx:
        ctx.add_sheet_warning_msg(sheet_name, msg)


#  Classify each prefix_value as 'internal' or 'external' depending on whether it's used in the base_urn field of the corresponding *_meta sheets.
def _URN_prefix_classify_prefix_usage(wb: Workbook, df_urn_prefix: pd.DataFrame, meta_sheets: List[str], meta_type: MetaTypes, sheet_name: str, fct_name: str, ctx: ConsoleContext = None) -> tuple[List[str], List[str], List[str]]:
    """
    Args:
        wb: Workbook object.
        df_urn_prefix: DataFrame of the URN Prefix content sheet.
        meta_sheets: List of *_meta sheet names to check.
        meta_type: One of MetaTypes.THREATS or MetaTypes.REFERENCE_CONTROLS.
        sheet_name: Name of the sheet currently being validated (URN Prefix).
        fct_name: Name of the calling validation function (for error formatting).
        ctx: Optional ConsoleContext to store warnings/info messages.
    Returns:
        tuple of (internal_prefixes, external_prefixes)
    """

    # Define expected type_object depending on the meta_type
    if meta_type == MetaTypes.THREATS:
        expected_type_object = URNObjects.THREAT.value
    elif meta_type == MetaTypes.REFERENCE_CONTROLS:
        expected_type_object = URNObjects.REFERENCE_CONTROL.value
    else:
        raise ValueError(f"({fct_name}) [{sheet_name}] Unsupported Meta type: {meta_type}")

    prefix_values = df_urn_prefix[URNPrefixContentColumns.PREFIX_VALUE.value].dropna().astype(str).str.strip().unique()
    internal_prefixes = []
    external_prefixes = []
    internal_meta_sheets = []

    # Filter prefix_values by expected type_object (based on the 4th segment of the URN)
    filtered_prefix_values = []
    for prefix in prefix_values:
        parts = prefix.split(":")
        if len(parts) > 3 and parts[3].strip() == expected_type_object:
            filtered_prefix_values.append(prefix)


    for prefix in filtered_prefix_values:
        found = False

        for sheet in meta_sheets:
            try:
                rows = list(wb[sheet].values)

                # Convert the meta sheet into a key-value dictionary
                meta_dict = {
                    str(row[0]).strip(): str(row[1]).strip()
                    for row in rows if row and len(row) >= 2 and row[0] and row[1]
                }

                base_urn = meta_dict.get("base_urn")
                if not base_urn:
                    continue

                parts = base_urn.split(":")
                
                # Make sure the 4th element in base_urn matches the expected type
                if len(parts) > 3 and parts[3].strip() == expected_type_object:
                    if base_urn.strip() == prefix.strip():
                        found = True
                        internal_meta_sheets.append(sheet)
                        break  # No need to keep checking other sheets

            except Exception as e:
                msg = f"⚠️  [WARNING] ({fct_name}) [{sheet_name}] Could not process sheet \"{sheet}\": {e}"
                print(msg)
                if ctx:
                    ctx.add_sheet_warning_msg(sheet_name, msg)
                continue

        if found:
            internal_prefixes.append(prefix)
        else:
            external_prefixes.append(prefix)

    return internal_prefixes, external_prefixes, internal_meta_sheets


# Classify URN prefixes as internal or external and validate the required external library dependencies.
def _URN_prefix_validate_prefix_values_and_dependencies(wb: Workbook, df: pd.DataFrame, sheet_name: str, ctx: ConsoleContext = None, verbose: bool = False):

    fct_name = get_current_fct_name()

    # 1. Get "threats" meta sheets
    threats_sheets = get_meta_sheets_names_from_type(wb, MetaTypes.THREATS, fct_name)

    # 2. Get "reference_controls" sheets
    ref_ctrl_sheets = get_meta_sheets_names_from_type(wb, MetaTypes.REFERENCE_CONTROLS, fct_name)

    # 3. Check whether the values for each "prefix_value" come from internal sheets or external framework
    internal_threats = []
    external_threats = []
    internal_ref_ctrl = []
    external_ref_ctrl = []

    if threats_sheets:
        internal_threats, external_threats, _ = _URN_prefix_classify_prefix_usage(wb, df, threats_sheets, MetaTypes.THREATS, sheet_name, fct_name, ctx)
    if ref_ctrl_sheets:
        internal_ref_ctrl, external_ref_ctrl, _ = _URN_prefix_classify_prefix_usage(wb, df, ref_ctrl_sheets, MetaTypes.REFERENCE_CONTROLS, sheet_name, fct_name, ctx)

    # Info messages for "threats" & "reference_controls"
    print_info_about_internal_external_URN_prefix(sheet_name, internal_threats, external_threats, internal_ref_ctrl, external_ref_ctrl, fct_name, verbose, ctx)

    ### 4. Check if external prefixes are declared in "dependencies" from "library_meta" ###

    # 1. Normalize external URNs by replacing the object type (4th element) with "library"
    def normalize_to_library(urn_list: List[str], target_type: str) -> List[str]:
        normalized = []
        for urn in urn_list:
            parts = urn.split(":")
            if len(parts) > 3 and parts[3].strip() == target_type:
                parts[3] = "library"
                normalized.append(":".join(parts))
        return normalized

    normalized_ext_threats = normalize_to_library(external_threats, "threat")
    normalized_ext_ref_ctrl = normalize_to_library(external_ref_ctrl, "function")

    # 2. Merge and deduplicate normalized external URNs
    required_dependencies = sorted(set(normalized_ext_threats + normalized_ext_ref_ctrl))

    if required_dependencies:

        # 3. Load "library_meta" sheet as a key-value dictionary
        try:
            rows = list(wb[MandatorySheets.LIBRARY_META.value].values)
            meta_dict = {
                str(row[0]).strip(): str(row[1]).strip()
                for row in rows if row and len(row) >= 2 and row[0] and row[1]
            }
        except Exception as e:
            raise ValueError(f"({fct_name}) [{sheet_name}] Could not read \"{MandatorySheets.LIBRARY_META.value}\" sheet: {e}")

        # 4. Ensure "dependencies" field exists and is non-empty
        if "dependencies" not in meta_dict or not meta_dict["dependencies"].strip():
            raise ValueError(
                f"({fct_name}) [{sheet_name}] \"{MandatorySheets.LIBRARY_META.value}\" is missing a non-empty \"dependencies\" field, "
                f"required to declare external libraries: {', '.join(f'\"{d}\"' for d in required_dependencies)}"
            )

        # 5. Parse declared dependencies
        declared_dependencies = [
            dep.strip() for dep in meta_dict.get("dependencies", "").split(",") if dep.strip()
        ]

        # 6. Compare with required dependencies
        missing_dependencies = [dep for dep in required_dependencies if dep not in declared_dependencies]

        if missing_dependencies:
            missing_list = ", ".join(f'"{d}"' for d in missing_dependencies)
            threat_list = ", ".join(f'"{t}"' for t in external_threats)
            ref_ctrl_list = ", ".join(f'"{r}"' for r in external_ref_ctrl)

            raise ValueError(
                f"({fct_name}) [{sheet_name}] Missing required dependencies in \"{MandatorySheets.LIBRARY_META.value}\": {missing_list}\n"
                f"> 💡 Tip: These are required due to the following external prefixes:\n"
                f"   - External \"threats\": {threat_list or 'None'}\n"
                f"   - External \"reference_controls\": {ref_ctrl_list or 'None'}"
            )


# Print information indicating whether the references mentioned in a URN prefix sheet are internal or external to the current workbook.
def print_info_about_internal_external_URN_prefix(sheet_name: str, internal_threats: List[str], external_threats: List[str], internal_ref_ctrl: List[str], external_ref_ctrl: List[str], context: str = None, verbose: bool = False, ctx: ConsoleContext = None, all_verbose: bool = False):
    
    verbose_icon = "💬 "
    
    if internal_threats:
        msg = f"ℹ️  [INFO] ({context}) [{sheet_name}] Internal \"threats\" prefixes found: {', '.join(f'\"{x}\"' for x in internal_threats)}"

        if all_verbose and verbose:
            msg = verbose_icon + msg
            print(msg)

            if ctx:
                ctx.add_sheet_verbose_msg(sheet_name, msg)

        elif not all_verbose:
            print(msg)

    else:
        if verbose:
            msg = f"💬 ℹ️  [INFO] ({context}) [{sheet_name}] No internal \"threats\" prefixes found"
            print(msg)
            if ctx:
                ctx.add_sheet_verbose_msg(sheet_name, msg)

    if external_threats:
        msg = f"ℹ️  [INFO] ({context}) [{sheet_name}] External \"threats\" prefixes found: {', '.join(f'\"{x}\"' for x in external_threats)}"

        if all_verbose and verbose:
            msg = verbose_icon + msg
            print(msg)

            if ctx:
                ctx.add_sheet_verbose_msg(sheet_name, msg)

        elif not all_verbose:
            print(msg)

    else:
        if verbose:
            msg = f"💬 ℹ️  [INFO] ({context}) [{sheet_name}] No external \"threats\" prefixes found"
            print(msg)
            if ctx:
                ctx.add_sheet_verbose_msg(sheet_name, msg)

    # Info messages for "reference_controls"
    if internal_ref_ctrl:
        msg = f"ℹ️  [INFO] ({context}) [{sheet_name}] Internal \"reference_controls\" prefixes found: {', '.join(f'\"{x}\"' for x in internal_ref_ctrl)}"

        if all_verbose and verbose:
            msg = verbose_icon + msg
            print(msg)

            if ctx:
                ctx.add_sheet_verbose_msg(sheet_name, msg)
        elif not all_verbose:
            print(msg)

    else:
        if verbose:
            msg = f"💬 ℹ️  [INFO] ({context}) [{sheet_name}] No internal \"reference_controls\" prefixes found"
            print(msg)
            if ctx:
                ctx.add_sheet_verbose_msg(sheet_name, msg)

    if external_ref_ctrl:
        msg = f"ℹ️  [INFO] ({context}) [{sheet_name}] External \"reference_controls\" prefixes found: {', '.join(f'\"{x}\"' for x in external_ref_ctrl)}"

        if all_verbose and verbose:
            msg = verbose_icon + msg
            print(msg)

            if ctx:
                ctx.add_sheet_verbose_msg(sheet_name, msg)

        elif not all_verbose:
            print(msg)

    else:
        if verbose:
            msg = f"💬 ℹ️  [INFO] ({context}) [{sheet_name}] No external \"reference_controls\" prefixes found"
            print(msg)
            if ctx:
                ctx.add_sheet_verbose_msg(sheet_name, msg)


# Check that each (source_node_id, target_node_id) pair is unique. Emits a warning or raises an error depending on "warn_only".
def _req_map_set_validate_unique_mappings(df: pd.DataFrame, sheet_name: str, warn_only: bool = False, ctx: ConsoleContext = None):

    fct_name = get_current_fct_name()

    source_column = RequirementMappingSetContentColumns.SOURCE_NODE_ID.value
    target_column = RequirementMappingSetContentColumns.TARGET_NODE_ID.value

    if source_column not in df.columns or target_column not in df.columns:
        raise ValueError(f"({fct_name}) [{sheet_name}] Columns \"{source_column}\" and/or \"{target_column}\" not found")

    df_clean = df[[source_column, target_column]].dropna()

    df_clean[source_column] = df_clean[source_column].map(lambda x: str(x).strip())
    df_clean[target_column] = df_clean[target_column].map(lambda x: str(x).strip())

    # Remove rows with empty values
    df_clean = df_clean[(df_clean[source_column] != "") & (df_clean[target_column] != "")]

    duplicates = df_clean[df_clean.duplicated(subset=[source_column, target_column], keep=False)]


    if not duplicates.empty:
        duplicate_rows = duplicates.index + 2  # 1-based row index (+ header)
        duplicate_pairs = duplicates.drop_duplicates().values.tolist()
        quoted_pairs = '\n   - '.join(f'["{s}", "{t}"]' for s, t in duplicate_pairs)

        msg = (
            f"({fct_name}) [{sheet_name}] Duplicate mapping(s) found for [source_node_id + target_node_id] pair(s):\n   - {quoted_pairs}"
            f"\n> Rows: {', '.join(map(str, duplicate_rows))}"
        )

        if warn_only:
            msg = f"⚠️  [WARNING] {msg}"
            print(msg)
            if ctx:
                ctx.add_sheet_warning_msg(sheet_name, msg)
        else:
            raise ValueError(msg)


# Validate if the "source_node_id" and "target_node_id" used in the mapping exist in the corresponding 'source' and 'target' sheets.
def _req_map_set_validate_mapping_node_ids_against_sheets(wb: Workbook, df: pd.DataFrame, sheet_name: str, fct_name: str, ctx: ConsoleContext = None, verbose: bool = False):

    sheets = wb.sheetnames
    source_ids = set()
    target_ids = set()
    source_sheet_available = "source" in sheets
    target_sheet_available = "target" in sheets

    # Load source node IDs
    if source_sheet_available:
        source_sheet = wb["source"]
        source_header = [cell.value for cell in source_sheet[1]]
        if "node_id" in source_header:
            idx = source_header.index("node_id")
            for row in source_sheet.iter_rows(min_row=2):
                if idx < len(row) and row[idx].value:
                    source_ids.add(str(row[idx].value).strip())
        else:
            source_sheet_available = False
            if verbose:
                msg = f'💬 ℹ️  [INFO] ({fct_name}) [{sheet_name}] Column "node_id" not found in sheet "source"'
                print(msg)
                if ctx:
                    ctx.add_sheet_verbose_msg(sheet_name, msg)
    else:
        if verbose:
            msg = f'💬 ℹ️  [INFO] ({fct_name}) [{sheet_name}] Sheet "source" not found'
            print(msg)
            if ctx:
                ctx.add_sheet_verbose_msg(sheet_name, msg)

    # Load target node IDs
    if target_sheet_available:
        target_sheet = wb["target"]
        target_header = [cell.value for cell in target_sheet[1]]
        if "node_id" in target_header:
            idx = target_header.index("node_id")
            for row in target_sheet.iter_rows(min_row=2):
                if idx < len(row) and row[idx].value:
                    target_ids.add(str(row[idx].value).strip())
        else:
            target_sheet_available = False
            if verbose:
                msg = f'💬 ℹ️  [INFO] ({fct_name}) [{sheet_name}] Column "node_id" not found in sheet "target"'
                print(msg)
                if ctx:
                    ctx.add_sheet_verbose_msg(sheet_name, msg)
    else:
        if verbose:
            msg = f'💬 ℹ️  [INFO] ({fct_name}) [{sheet_name}] Sheet "target" not found'
            print(msg)
            if ctx:
                ctx.add_sheet_verbose_msg(sheet_name, msg)


    if not source_sheet_available:
        msg = f'⚠️  [WARNING] ({fct_name}) [{sheet_name}] Invalid or missing "source" sheet. The "{RequirementMappingSetContentColumns.SOURCE_NODE_ID.value}" column cannot be checked.'
        print(msg)
        if ctx:
            ctx.add_sheet_warning_msg(sheet_name, msg)

    if not target_sheet_available:
        msg = f'⚠️  [WARNING] ({fct_name}) [{sheet_name}] Invalid or missing "target" sheet. The "{RequirementMappingSetContentColumns.TARGET_NODE_ID.value}" column cannot be checked.'
        print(msg)
        if ctx:
            ctx.add_sheet_warning_msg(sheet_name, msg)


    # Used IDs
    used_source_ids = [str(val).split(":")[-1] for val in df[RequirementMappingSetContentColumns.SOURCE_NODE_ID.value].dropna()]
    used_target_ids = [str(val).split(":")[-1] for val in df[RequirementMappingSetContentColumns.TARGET_NODE_ID.value].dropna()]

    source_missing_counts = Counter(
        node_id for node_id in used_source_ids
        if source_sheet_available and node_id not in source_ids
    )
    target_missing_counts = Counter(
        node_id for node_id in used_target_ids
        if target_sheet_available and node_id not in target_ids
    )

    # Warnings: Missing IDs
    if source_sheet_available:
        for sid in source_missing_counts:
            msg = f'⚠️  [WARNING] ({fct_name}) [{sheet_name}] source_node_id "{sid}" not found in sheet "source"'
            print(msg)
            if ctx:
                ctx.add_sheet_warning_msg(sheet_name, msg)

    if target_sheet_available:
        for tid in target_missing_counts:
            msg = f'⚠️  [WARNING] ({fct_name}) [{sheet_name}] target_node_id "{tid}" not found in sheet "target"'
            print(msg)
            if ctx:
                ctx.add_sheet_warning_msg(sheet_name, msg)

    # Duplicates
    if source_sheet_available:
        for sid, count in source_missing_counts.items():
            if count > 1:
                msg = f'🔁 [DUPLICATE] ({fct_name}) [{sheet_name}] source_node_id "{sid}" appears {count} times in mappings'
                print(msg)

    if target_sheet_available:
        for tid, count in target_missing_counts.items():
            if count > 1:
                msg = f'🔁 [DUPLICATE] ({fct_name}) [{sheet_name}] target_node_id "{tid}" appears {count} times in mappings'
                print(msg)

    # Final summary
    total_missing_sources = '???' if not source_sheet_available else sum(source_missing_counts.values())
    total_missing_targets = '???' if not target_sheet_available else sum(target_missing_counts.values())
    if total_missing_sources or total_missing_targets:
        msg = f"⚠️  [MAPPING CHECK SUMMARY] ({fct_name}) [{sheet_name}] Missing usage count - Source: {total_missing_sources}, Target: {total_missing_targets}"
        print(msg)
        if ctx:
            ctx.add_sheet_warning_msg(sheet_name, msg)

        if source_sheet_available or target_sheet_available:
            msg2 = (
                "ℹ️  [INFO] Please note that these incorrect node IDs have been added to the mapping anyway."
                "\n> 💡 Tip: If you want to correct them, please do so in your Excel file."
            )
            print(msg2)


# Additional rule: for non-empty rows, at least "ref_id", "name" or "description" must be filled
def _framework_validate_minimum_fields_and_ref_id(df: pd.DataFrame, sheet_name: str):

    fct_name = get_current_fct_name()
    
    empty_id_name_desc_rows = []
    invalid_ref_ids = []

    for idx, row in df.iterrows():
        if row.dropna().empty:
            continue  # skip completely empty rows

        ref_id = row.get(FrameworkContentColumns.REF_ID.value, "")
        if pd.isna(ref_id):
            ref_id = ""
        else:
            ref_id = str(ref_id).strip()

        name = row.get(FrameworkContentColumns.NAME.value, "")
        if pd.isna(name):
            name = ""
        else:
            name = str(name).strip()
            
        description = row.get(FrameworkContentColumns.DESCRIPTION.value, "")
        if pd.isna(description):
            description = ""
        else:
            description = str(description).strip()

        if not ref_id and not name and not description:
            empty_id_name_desc_rows.append(idx)

        # Check Ref. IDs
        if ref_id:
            try:
                validate_ref_id(ref_id, fct_name, idx) #_with_spaces(ref_id, fct_name, idx)
            except Exception as e:
                invalid_ref_ids.append((ref_id, idx))

    # If any, returns an error and print rows with empty Ref. ID, Name and Description 
    if empty_id_name_desc_rows:
        raise ValueError(
            f"({fct_name}) [{sheet_name}] Invalid rows: \"{FrameworkContentColumns.REF_ID.value}\", \"{FrameworkContentColumns.NAME.value}\" and \"{FrameworkContentColumns.DESCRIPTION.value}\" are empty :\n   - "
            + "\n   - ".join(f'Row #{idx + 2}' for idx in empty_id_name_desc_rows)
            + "\n> 💡 Tip: For each row, at least one of the values must be filled."
        )

    # If any, returns an error and print invalid Ref. IDs
    if invalid_ref_ids:
        raise ValueError(
            f"({fct_name}) [{sheet_name}] Invalid Ref. IDs found. Only alphanumeric characters, '-', '_', ' ', and '.' are allowed :\n   - "
            + "\n   - ".join(f'Row #{idx + 2}: {value}' for value, idx in invalid_ref_ids)
        )


def _framework_validate_column_against_reference_sheet(wb: Workbook, df: pd.DataFrame, column_name: FrameworkContentColumns | str, current_sheet_name: str, verbose: bool = False, ctx: ConsoleContext = None):

    context = get_current_fct_name()

    if isinstance(column_name, ContentColumn):
        column_name = column_name.value

    column_to_key_mapping: Dict[str, FrameworkMetaKeys] = {
        FrameworkContentColumns.IMPLEMENTATION_GROUPS.value: FrameworkMetaKeys.IMPLEMENTATION_GROUPS_DEFINITION,
        FrameworkContentColumns.ANSWER.value: FrameworkMetaKeys.ANSWERS_DEFINITION,
    }

    if column_name not in column_to_key_mapping:
        raise ValueError(f"({context}) [{current_sheet_name}] Unsupported column \"{column_name}\" for this validation")

    # Get associated meta sheet
    meta_sheet_name = get_corresponding_type_sheet_names([current_sheet_name], SheetTypes.META)[0]
    if meta_sheet_name not in wb.sheetnames:
        raise ValueError(f"({context}) [{current_sheet_name}] Missing meta sheet \"{meta_sheet_name}\" required to validate column \"{column_name}\"")

    # Convert meta sheet to DataFrame
    meta_ws = wb[meta_sheet_name]
    meta_df = pd.DataFrame(meta_ws.values)
    meta_df.columns = meta_df.iloc[0]
    meta_df = meta_df.drop(index=0).reset_index(drop=True)

    # Get the referenced sheet name from meta key
    meta_key = column_to_key_mapping[column_name]
    ref_base_name = get_meta_value(meta_df, meta_key, meta_sheet_name, context=context)

    if not ref_base_name:
        raise ValueError(
            f"({context}) [{current_sheet_name}] The meta key \"{meta_key.value}\" is missing or empty, required for column \"{column_name}\".\n"
            f"> 💡 Tip: Either remove column \"{column_name}\" or define a proper value for \"{meta_key.value}\" in the meta sheet."
        )

    ref_content_sheet = f"{ref_base_name}{SheetTypes.CONTENT.value}"
    if ref_content_sheet not in wb.sheetnames:
        raise ValueError(f"({context}) [{current_sheet_name}] Referenced sheet \"{ref_content_sheet}\" (from key \"{meta_key.value}\") not found")

    # Convert referenced sheet to DataFrame
    ref_ws = wb[ref_content_sheet]
    ref_df = pd.DataFrame(ref_ws.values)
    ref_df.columns = ref_df.iloc[0]
    ref_df = ref_df.drop(index=0).reset_index(drop=True)

    if column_name == FrameworkContentColumns.IMPLEMENTATION_GROUPS.value:
        ref_column = ImplementationGroupsContentColumns.REF_ID.value
        separator = ","
    elif column_name == FrameworkContentColumns.ANSWER.value:
        ref_column = AnswersContentColumns.ID.value
        separator = "\n"
    else:
        raise RuntimeError(f"({context}) [{current_sheet_name}] Unexpected internal error: invalid column dispatch")

    if ref_column not in ref_df.columns:
        raise ValueError(f"({context}) [{current_sheet_name}] Referenced sheet \"{ref_content_sheet}\" does not contain required column \"{ref_column}\"")

    valid_values = set(ref_df[ref_column].dropna().astype(str).map(str.strip))


    invalid_values = []

    for idx, value in df[column_name].dropna().astype(str).items():
        items = [v.strip() for v in value.split(separator) if v.strip()]
        for i, item in enumerate(items, start=1):
            if item not in valid_values:
                invalid_values.append((idx, item, i))

    if invalid_values:
        raise ValueError(
            f"({context}) [{current_sheet_name}] Invalid values in column \"{column_name}\" :\n   - "
            + "\n   - ".join(f'Row #{idx + 2} (element #{i}) -> {item}' for idx, item, i in invalid_values)
            + f"\n> 💡 Tip: Make sure these values exist in column \"{ref_column}\" of the referenced sheet \"{ref_content_sheet}\"."
        )

    if verbose:
        msg = f'💬 ℹ️  [INFO] ({context}) [{current_sheet_name}] Column \"{column_name}\" has valid values'
        print(msg)
        if ctx:
            ctx.add_sheet_verbose_msg(current_sheet_name, msg)


# Validate that all URNs in the column use defined prefix_ids and reference existing ref_ids when required
def _framework_validate_framework_column_urns(wb: Workbook, df: pd.DataFrame, column_name: FrameworkContentColumns | str, current_sheet_name: str, external_refs: List[str] = None, verbose: bool = False, ctx: ConsoleContext = None):
    
    fct_name = get_current_fct_name()

    if isinstance(column_name, ContentColumn):
        column_name = column_name.value

    if column_name not in (FrameworkContentColumns.THREATS.value, FrameworkContentColumns.REFERENCE_CONTROLS.value):
        raise ValueError(f"({fct_name}) [{current_sheet_name}] Column \"{column_name}\" is not supported for URN validation")

    # ───────────────────────────────────────────────────────────────
    # 1st Part: Load and validate all URN_PREFIX sheets
    # ───────────────────────────────────────────────────────────────

    urn_meta_sheets = get_meta_sheets_names_from_type(wb, MetaTypes.URN_PREFIX, fct_name)

    if not urn_meta_sheets:
        raise ValueError(
            f"({fct_name}) [{current_sheet_name}] Column \"{column_name}\" cannot be validated because no URN_PREFIX meta sheet exists.\n"
            f"> 💡 Tip: Either remove column \"{column_name}\" or define a valid URN prefix sheet."
        )

    urn_prefix_map = {}     # {prefix_id: (prefix_value, sheet_name)}
    seen_prefix_ids = {}    # {prefix_id: sheet_name}

    for meta_sheet_name in urn_meta_sheets:
        content_sheet = get_corresponding_type_sheet_names([meta_sheet_name], SheetTypes.CONTENT)[0]
        if content_sheet not in wb.sheetnames:
            raise ValueError(f"({fct_name}) [{current_sheet_name}] URN_PREFIX content sheet \"{content_sheet}\" not found")

        content_df = pd.DataFrame(wb[content_sheet].values)
        content_df.columns = content_df.iloc[0]
        content_df = content_df.drop(index=0).reset_index(drop=True)

        # Filter out rows where both 'prefix_id' and 'prefix_value' are empty or null
        content_df = content_df.dropna(subset=[URNPrefixContentColumns.PREFIX_ID.value, URNPrefixContentColumns.PREFIX_VALUE.value], how='all')
        content_df = content_df[(content_df[URNPrefixContentColumns.PREFIX_ID.value].astype(str).str.strip() != "") | (content_df[URNPrefixContentColumns.PREFIX_VALUE.value].astype(str).str.strip() != "")]

        for _, row in content_df.iterrows():
            prefix_id = str(row.get(URNPrefixContentColumns.PREFIX_ID.value, "")).strip()
            prefix_value = str(row.get(URNPrefixContentColumns.PREFIX_VALUE.value, "")).strip()
            if not prefix_id:
                continue
            if prefix_id in seen_prefix_ids:
                other_sheet = seen_prefix_ids[prefix_id]
                raise ValueError(
                    f"({fct_name}) [{current_sheet_name}] Duplicate prefix_id \"{prefix_id}\" found in sheets \"{meta_sheet_name}\" and \"{other_sheet}\""
                )
            urn_prefix_map[prefix_id] = (prefix_value, content_sheet)
            seen_prefix_ids[prefix_id] = content_sheet

    all_prefix_ids = set(urn_prefix_map.keys())

    # ───────────────────────────────────────────────────────────────
    # 2nd Part: Validate that all URN prefix_id used in the column are known
    # ───────────────────────────────────────────────────────────────

    for idx, value in df[column_name].dropna().astype(str).items():
        elements = re.split(CommonSeparatorRegex.COMMA_LF.value, value)
        for i, raw in enumerate(elements, start=1):
            raw = raw.strip()
            if not raw or ":" not in raw:
                continue
            prefix_id = raw.split(":", 1)[0].strip()
            if prefix_id not in all_prefix_ids:
                raise ValueError(
                    f"({fct_name}) [{current_sheet_name}] Row #{idx + 2} - Invalid URN prefix \"{prefix_id}\" (element #{i}) in column \"{column_name}\".\n"
                    f"> 💡 Tip: This prefix must be defined in a URN Prefix sheet."
                )

    # ───────────────────────────────────────────────────────────────
    # 3rd Part: Determine internal and external prefixes for threats / reference_controls
    # ───────────────────────────────────────────────────────────────

    threats_meta = get_meta_sheets_names_from_type(wb, MetaTypes.THREATS, fct_name)
    ref_ctrl_meta = get_meta_sheets_names_from_type(wb, MetaTypes.REFERENCE_CONTROLS, fct_name)

    all_internal_threats = []
    all_external_threats = []
    all_internal_threat_sheets = []
    all_internal_ref_ctrl = []
    all_external_ref_ctrl = []
    all_internal_ref_ctrl_sheets = []

    for meta_sheet_name in urn_meta_sheets:
        urn_content_sheet = get_corresponding_type_sheet_names([meta_sheet_name], SheetTypes.CONTENT)[0]
        if urn_content_sheet not in wb.sheetnames:
            raise ValueError(f"({fct_name}) [{current_sheet_name}] URN_PREFIX content sheet \"{urn_content_sheet}\" not found")

        df_urn = pd.DataFrame(wb[urn_content_sheet].values)
        df_urn.columns = df_urn.iloc[0]
        df_urn = df_urn.drop(index=0).reset_index(drop=True)

        # Filter out empty rows
        df_urn = df_urn.dropna(how='all')
        df_urn = df_urn[
            df_urn.apply(lambda row: any(str(cell).strip() != "" for cell in row), axis=1)
        ]

        if threats_meta:
            internal_threats, external_threats, internal_threat_sheets = _URN_prefix_classify_prefix_usage(
                wb, df_urn, threats_meta, MetaTypes.THREATS, current_sheet_name, fct_name, ctx
            )
            all_internal_threats.extend(internal_threats)
            all_external_threats.extend(external_threats)
            all_internal_threat_sheets.extend(internal_threat_sheets)

        if ref_ctrl_meta:
            internal_ref_ctrl, external_ref_ctrl, internal_ref_ctrl_sheets = _URN_prefix_classify_prefix_usage(
                wb, df_urn, ref_ctrl_meta, MetaTypes.REFERENCE_CONTROLS, current_sheet_name, fct_name, ctx
            )
            all_internal_ref_ctrl.extend(internal_ref_ctrl)
            all_external_ref_ctrl.extend(external_ref_ctrl)
            all_internal_ref_ctrl_sheets.extend(internal_ref_ctrl_sheets)

    print_info_about_internal_external_URN_prefix(
        current_sheet_name,
        all_internal_threats, all_external_threats,
        all_internal_ref_ctrl, all_external_ref_ctrl,
        fct_name, verbose, ctx, True
    )

    # ───────────────────────────────────────────────────────────────
    # 4th Part: Validate prefix_id types based on column context
    # ───────────────────────────────────────────────────────────────

    allowed_prefix_ids = set()
    forbidden_prefix_ids = set()

    # We must use prefix_id (not URNs) to detect valid or invalid use
    if column_name == FrameworkContentColumns.THREATS.value:
        allowed_prefix_ids = {pid for pid, (pval, _) in urn_prefix_map.items() if pval in all_internal_threats or pval in all_external_threats}
        forbidden_prefix_ids = {pid for pid, (pval, _) in urn_prefix_map.items() if pval in all_internal_ref_ctrl or pval in all_external_ref_ctrl}
    elif column_name == FrameworkContentColumns.REFERENCE_CONTROLS.value:
        allowed_prefix_ids = {pid for pid, (pval, _) in urn_prefix_map.items() if pval in all_internal_ref_ctrl or pval in all_external_ref_ctrl}
        forbidden_prefix_ids = {pid for pid, (pval, _) in urn_prefix_map.items() if pval in all_internal_threats or pval in all_external_threats}

    for idx, value in df[column_name].dropna().astype(str).items():
        elements = re.split(CommonSeparatorRegex.COMMA_LF.value, value)
        for i, raw in enumerate(elements, start=1):
            raw = raw.strip()
            if not raw:
                continue

            prefix_id = raw.split(":", 1)[0].strip()

            if prefix_id in forbidden_prefix_ids:
                raise ValueError(
                    f"({fct_name}) [{current_sheet_name}] Row #{idx + 2} - Invalid URN prefix \"{prefix_id}\" (element #{i}) in column \"{column_name}\"."
                    f"\n> 💡 Tip: This prefix is not allowed in column \"{column_name}\", as the URN to which it refers is not a \"{column_name}\" URN."
                )

    # Dict prefix_value -> "sheet_content" for "threats"
    prefix_to_threats_content_sheet = {}

    for i, prefix_val in enumerate(all_internal_threats):
        # Correspondence index => internal threat content sheet
        if i < len(all_internal_threat_sheets):
            meta_sheet_name = all_internal_threat_sheets[i]
            content_sheet_name = get_corresponding_type_sheet_names([meta_sheet_name], SheetTypes.CONTENT)[0]
            prefix_to_threats_content_sheet[prefix_val] = content_sheet_name

    # Dict prefix_value -> "sheet_content" for "reference_controls"
    prefix_to_refctrl_content_sheet = {}

    for i, prefix_val in enumerate(all_internal_ref_ctrl):
        if i < len(all_internal_ref_ctrl_sheets):
            meta_sheet_name = all_internal_ref_ctrl_sheets[i]
            content_sheet_name = get_corresponding_type_sheet_names([meta_sheet_name], SheetTypes.CONTENT)[0]
            prefix_to_refctrl_content_sheet[prefix_val] = content_sheet_name

    # ───────────────────────────────────────────────────────────────
    # 5th Part: Validate that internal URN values exist in ref_id column only
    # ───────────────────────────────────────────────────────────────

    if column_name == FrameworkContentColumns.THREATS.value:
        internal_prefix_ids = {pid for pid, (pval, _) in urn_prefix_map.items() if pval in all_internal_threats}
    else:  # reference_controls
        internal_prefix_ids = {pid for pid, (pval, _) in urn_prefix_map.items() if pval in all_internal_ref_ctrl}

    
    for prefix_id in internal_prefix_ids:
        prefix_value, _ = urn_prefix_map[prefix_id]

        # Get the actual content sheet name based on the column type
        content_sheet = None

        if column_name == FrameworkContentColumns.THREATS.value:
            content_sheet = prefix_to_threats_content_sheet.get(prefix_value)
        else:  # reference_controls
            content_sheet = prefix_to_refctrl_content_sheet.get(prefix_value)

        if not content_sheet:
            raise ValueError(
                f"({fct_name}) [{current_sheet_name}] Referenced content sheet for prefix_value \"{prefix_value}\" (prefix_id \"{prefix_id}\") not found"
            )

        if content_sheet not in wb.sheetnames:
            raise ValueError(
                f"({fct_name}) [{current_sheet_name}] Referenced content sheet \"{content_sheet}\" not found in workbook"
            )

        # Get reference sheet
        content_df = pd.DataFrame(wb[content_sheet].values)
        content_df.columns = content_df.iloc[0]
        content_df = content_df.drop(index=0).reset_index(drop=True)

        ref_id_column = ThreatsContentColumns.REF_ID.value if column_name == FrameworkContentColumns.THREATS.value else ReferenceControlsContentColumns.REF_ID.value

        # Only check "ref_id" column of the reference sheet
        if ref_id_column not in content_df.columns:
            raise ValueError(
                f"({fct_name}) [{current_sheet_name}] Sheet \"{content_sheet}\" has no \"ref_id\" column"
            )

        valid_ref_ids = set(content_df[ref_id_column].astype(str).str.strip())

        verification_errors = []

        # Check Ref. IDs validity
        for idx, value in df[column_name].dropna().astype(str).items():
            elements = re.split(CommonSeparatorRegex.COMMA_LF.value, value)
            for i, raw in enumerate(elements, start=1):
                raw = raw.strip()
                if not raw:
                    continue
                # Only process elements that start with the expected prefix (e.g. "1:REF")
                if not raw.startswith(prefix_id + ":"):
                    continue

                # Extract the REF part after the first ":"
                urn_id = raw.split(":", 1)[1].strip()

                # Check if it exists in the ref_ids from reference sheet
                if urn_id not in valid_ref_ids:
                    verification_errors.append((idx, i, urn_id, prefix_id))

        # Print all errors and exit
        if verification_errors:
            msgs = []
            for idx, i, urn_id, prefix_id in verification_errors:
                msgs.append(f"   - Row #{idx + 2} (element #{i}) -> ref_id \"{urn_id}\" with prefix \"{prefix_id}\" ({prefix_id}:{urn_id})")

            msgs.append(f"> 💡 Tip: These IDs must exist in the sheet \"{content_sheet}\" in column \"ref_id\".")
            raise ValueError(f"({fct_name}) [{current_sheet_name}] Invalid internal references found in column \"{column_name}\":\n" + "\n".join(msgs))
        
    # ───────────────────────────────────────────────────────────────
    # 6th Part: Validate that external URN values exist in external references from YAML files only
    # ───────────────────────────────────────────────────────────────

    if column_name == FrameworkContentColumns.THREATS.value:
        external_prefix_ids = {pid for pid, (pval, _) in urn_prefix_map.items() if pval in all_external_threats}
    else:  # reference_controls
        external_prefix_ids = {pid for pid, (pval, _) in urn_prefix_map.items() if pval in all_external_ref_ctrl}

    yaml_section_type = None
    if column_name == FrameworkContentColumns.THREATS.value:
        yaml_section_type = YAMLSectionTypes.THREATS
    else:  # reference_controls
        yaml_section_type = YAMLSectionTypes.REFERENCE_CONTROLS


    yaml_external_references_retrieved = False
    yaml_references_from_files = None

    # Get references in YAML files
    if external_prefix_ids:

        # If there are external references in the Excel file, but no YAML files containing external references are given as argument
        if not external_refs:
            msg = (
                f"⚠️  [WARNING] ({fct_name}) [{current_sheet_name}] No YAML files provided as external references to check external references in column \"{column_name}\" (URN Prefix concerned: " + ", ".join(f"\"{f}\"" for f in external_prefix_ids) + ")\n" 
                f"> ❌ Unable to check for external \"{yaml_section_type.value}\" !\n"
                f"> 💡 Tip: Provide the right YAML file(s) as argument, corresponding to the external references, that contain a \"{yaml_section_type.value}\" section"
            )
            print(msg)
            if ctx:
                ctx.add_sheet_warning_msg(current_sheet_name, msg)

        # Else if there are external references in the Excel file, and YAML file(s) containing external references is(are) given as argument
        else:
            yaml_references_from_files = get_yaml_section_from_files(external_refs, yaml_section_type, current_sheet_name, verbose, ctx)

            if yaml_references_from_files:
                yaml_external_references_retrieved = True
            if not yaml_references_from_files:
                msg = (
                    f"⚠️  [WARNING] ({fct_name}) [{current_sheet_name}] None of the YAML files provided as external references contain a \"{yaml_section_type.value}\" section (URN prefixes requiring verification of external references: " + ", ".join(f"\"{f}\"" for f in external_prefix_ids) + ")\n"
                    f"> ❌ Unable to check for external \"{yaml_section_type.value}\" !\n"
                    f"> 💡 Tip: Make sure the YAML file(s) you provided as argument actually contain a \"{yaml_section_type.value}\" section."
                )
                print(msg)
                if ctx:
                    ctx.add_sheet_warning_msg(current_sheet_name, msg)


    # Not checked reference list
    not_checked_prefix_id = []

    # Check external references
    if yaml_references_from_files:
        
        for prefix_id in external_prefix_ids:
            prefix_value, _ = urn_prefix_map[prefix_id]
            
            yaml_filename = ""
            
            # Search for the "prefix_value" in "yaml_references_from_files" and get the corresponding Ref. IDs of the section
            base_urn_with_ref_ids = {}      # { base_urn: { ref_id: name, ... } }

            for filename, file_sections in yaml_references_from_files.items():
                if prefix_value in file_sections:
                    base_urn_with_ref_ids = {prefix_value: file_sections[prefix_value]}
                    yaml_filename = filename
                    break


            # If "yaml_references_from_files" doesn't have the searched "prefix_value"
            if prefix_value not in base_urn_with_ref_ids.keys():
                msg = (
                    f"⚠️  [WARNING] ({fct_name}) [{current_sheet_name}] None of the YAML files provided as external references contain references for the URN Prefix \"{prefix_id}\" ({prefix_value}) (Column \"{column_name}\")\n"
                    f"> 💡 Tip: Specify the right YAML file as an argument to check external \"{yaml_section_type.value}\""
                )
                print(msg)
                if ctx:
                    ctx.add_sheet_warning_msg(current_sheet_name, msg)

                # Add the non checked "prefix_value" to the unchecked element list
                not_checked_prefix_id.append(prefix_value)

                continue
                
            # Validation of Ref. IDs against the external reference extracted from YAML
            yaml_ref_ids_map = base_urn_with_ref_ids.get(prefix_value, {})  # mapping ref_id -> name | { ref_id1: name1, ... }
            valid_ref_ids = set(yaml_ref_ids_map.keys())                    # (ref_id1, ref_id2, ref_id3, ...)
            verification_errors = []

            for idx, value in df[column_name].dropna().astype(str).items():
                elements = re.split(CommonSeparatorRegex.COMMA_LF.value, value)
                for i, raw in enumerate(elements, start=1):
                    raw = raw.strip()
                    if not raw:
                        continue
                    # Only process elements that start with the expected prefix (e.g. "1:REF")
                    if not raw.startswith(prefix_id + ":"):
                        continue

                    # Extract the REF part after the first ":"
                    urn_id = raw.split(":", 1)[1].strip()

                    # Check if it exists in the ref_ids from YAML
                    if urn_id not in valid_ref_ids:
                        verification_errors.append((idx, i, urn_id, prefix_id))


            # Print all errors and exit
            if verification_errors:
                msgs = []
                for idx, i, urn_id, prefix_id in verification_errors:
                    msgs.append(f"   - Row #{idx + 2} (element #{i}) -> ref_id \"{urn_id}\" with prefix \"{prefix_id}\" ({prefix_id}:{urn_id})")

                msgs.append(f"> 💡 Tip: These Ref. IDs must exist in the YAML file \"{yaml_filename}\" provided as external references (in the \"{yaml_section_type.value}\" section of the YAML file).")
                raise ValueError(f"({fct_name}) [{current_sheet_name}] Invalid external references found in column \"{column_name}\":\n" + "\n".join(msgs))


    # If there are no external reference to check [OR] (there are external references to check [AND] the YAML files contains element of the section we're working on [AND] NOT a single external reference wasn't check)
    if not external_prefix_ids or (external_prefix_ids and yaml_external_references_retrieved and not not_checked_prefix_id):
        if verbose:
            msg = f'💬 ℹ️  [INFO] ({fct_name}) [{current_sheet_name}] Column \"{column_name}\" contains valid URN references'
            print(msg)
            if ctx:
                ctx.add_sheet_verbose_msg(current_sheet_name, msg)
    else:
        msg = (
            f"⚠️  [WARNING] ({fct_name}) [{current_sheet_name}] Column \"{column_name}\" contains valid *internal* URN references. However, something went wrong while checking one or all of the external elements for \"{yaml_section_type.value}\" !\n"
            f"> 💡 Tip: Read the warnings above to understand the problem."
        )
        print(msg)
        if ctx:
            ctx.add_sheet_warning_msg(current_sheet_name, msg)


# Ensure that choice-based question types have a non-empty "question_choices" value.
def _answers_validate_question_choices(df: pd.DataFrame, sheet_name: str):

    fct_name = get_current_fct_name()

    question_types_requiring_choices = {"unique_choice", "multiple_choice"}
    problematic_rows = []

    for row_idx, row in df.iterrows():
        question_type_raw = row.get(AnswersContentColumns.QUESTION_TYPE.value, "")
        if pd.isna(question_type_raw):
            continue

        question_type = str(question_type_raw).strip().lower()
        question_choices = row.get(AnswersContentColumns.QUESTION_CHOICES.value, "")

        if (
            question_type in question_types_requiring_choices
            and (pd.isna(question_choices) or str(question_choices).strip() == "")
        ):
            problematic_rows.append((row_idx + 2, question_type))

    if problematic_rows:
        raise ValueError(
            f"({fct_name}) [{sheet_name}] The field \"question_choices\" must not be empty for choice-based question types:\n   - "
            + "\n   - ".join(f'Row #{row}: question_type "{question_type}"' for row, question_type in problematic_rows)
            + '\n> 💡 Tip: Fill "question_choices" for every "unique_choice" and "multiple_choice" question.'
        )



# Get the "threats" or "reference_controls" section from a YAML Framework file passed as argument
def get_yaml_section_from_files(yaml_files: List[str], section_type: YAMLSectionTypes, current_sheet_name: str, verbose: bool = False, ctx: ConsoleContext = None) -> Dict:
        
    fct_name = get_current_fct_name()

    if type(yaml_files) == str:
        yaml_files = [yaml_files]

    extracted_sections = {}
    """
    - Object structure :
    {
        file: {
            section_urn : {  
                element_ref_id1: name1
                element_ref_id2: name2
                ...
            }
        }
    }
    """
    
    for file in yaml_files:

        # Attempt to load YAML file, raise an error if file is missing or YAML is invalid
        try:
            with open(file, 'r', encoding='utf-8') as f:
                data = yaml.safe_load(f)
        except FileNotFoundError:
            raise FileNotFoundError(f"({fct_name}) [{current_sheet_name}] YAML file not found: \"{file}\"")
        except yaml.YAMLError as e:
            raise ValueError(f"({fct_name}) [{current_sheet_name}] Error parsing YAML file \"{file}\": {e}")


        # Search for the element we want
        for obj_key, obj_value in data.get("objects", {}).items():
            
            # --- [Section] reference_controls [OR] threats ---
            if (
                (section_type == YAMLSectionTypes.REFERENCE_CONTROLS and obj_key == "reference_controls")
                or (section_type == YAMLSectionTypes.THREATS and obj_key == "threats")
            ):

                # Calculate base_urn
                base_urn = __calculate_base_urn(obj_value)
                
                # If the "base_urn" couldn't be defined, skip
                if not base_urn:
                    if verbose:
                        msg = (
                            f"💬 ℹ️  [INFO] ({fct_name}) [{current_sheet_name}] URN of section \"{section_type.value}\" of the YAML file \"{file}\" couldn't be defined.\n"
                            f"> 💡 Tip: This is probably because the \"{section_type.value}\" section of the file contains only 1 element (at least 2 elements are required to determine the \"base_urn\" of the \"{section_type.value}\" section in the YAML file)"
                        )
                        print(msg)
                        if ctx:
                            ctx.add_sheet_verbose_msg(current_sheet_name, msg)

                    break
                
                element_obj = {
                    base_urn: {}
                }
                
                # Store Ref. IDs, associated with their names
                for ref_ctrl in obj_value:
                    ref_id = ref_ctrl.get("ref_id")
                    name = ref_ctrl.get("name")
                    if ref_id:  # on ignore si pas de ref_id
                        element_obj[base_urn][ref_id] = name

                if file not in extracted_sections:
                    extracted_sections[file] = {}
                extracted_sections[file].update(element_obj)
                    
                break


        # If the file doesn't contains the requested section
        if file not in extracted_sections.keys():
            if verbose:
                msg = (f"💬 ℹ️  [INFO] ({fct_name}) [{current_sheet_name}] The YAML file \"{file}\" contains no \"{section_type.value}\" section.")
                print(msg)
                if ctx:
                    ctx.add_sheet_verbose_msg(current_sheet_name, msg)

    return extracted_sections


# Validate paired minimum and maximum columns and ensure each minimum is less than or equal to its maximum.
def validate_min_max_columns(df: pd.DataFrame, min_column: ContentColumn | str, max_column: ContentColumn | str, sheet_name: str, context: str, min_column_constraints: ContentColumnConstraints | None = None, max_column_constraints: ContentColumnConstraints | None = None):

    if isinstance(min_column, ContentColumn):
        min_column = min_column.value
    if isinstance(max_column, ContentColumn):
        max_column = max_column.value

    validate_columns_presence_together(df, [min_column, max_column], sheet_name, context)

    # If neither column is present, skip checking
    if min_column not in df.columns:
        return

    min_column_min = min_column_constraints.min_value if min_column_constraints and min_column_constraints.min_value is not None else 0
    max_column_min = max_column_constraints.min_value if max_column_constraints and max_column_constraints.min_value is not None else 0

    validate_integer_value(df, sheet_name, min_column, context, value_name=min_column, min=min_column_min)
    validate_integer_value(df, sheet_name, max_column, context, value_name=max_column, min=max_column_min)

    incomplete_rows = []
    invalid_ranges = []

    for index, row in df.iterrows():
        min_value = row[min_column]
        max_value = row[max_column]
        excel_row = index + 2

        min_is_empty = pd.isna(min_value) or str(min_value).strip() == ""
        max_is_empty = pd.isna(max_value) or str(max_value).strip() == ""

        if min_is_empty and max_is_empty:
            continue

        if min_is_empty or max_is_empty:
            incomplete_rows.append(f'Row #{excel_row}: "{min_column if min_is_empty else max_column}" is missing')
            continue

        if int(min_value) > int(max_value):
            invalid_ranges.append(f'Row #{excel_row}: "{min_column}" ({min_value}) > "{max_column}" ({max_value})')

    if incomplete_rows:
        raise ValueError(
            f'({context}) [{sheet_name}] "{min_column}" and "{max_column}" must be filled together.\n   - '
            + "\n   - ".join(incomplete_rows)
        )

    if invalid_ranges:
        raise ValueError(
            f'({context}) [{sheet_name}] Invalid minimum/maximum ranges:\n   - '
            + "\n   - ".join(invalid_ranges)
        )


# Validate that framework depths start at 1 and never increase by more than one level at a time.
def _framework_validate_depth_consistency(df: pd.DataFrame, sheet_name: str):

    fct_name = get_current_fct_name()

    validate_integer_value(df, sheet_name, FrameworkContentColumns.DEPTH, fct_name, value_name=FrameworkContentColumns.DEPTH, positive_only=True)

    depth_values = df[FrameworkContentColumns.DEPTH.value].dropna().astype(str).map(str.strip)
    depth_values = depth_values[depth_values != ""]     # Remove empty values

    if depth_values.empty:
        return

    depths = [(index + 2, int(value)) for index, value in depth_values.items()]
    errors = []

    first_row, first_depth = depths[0]
    if first_depth != 1:
        errors.append(f'Row #{first_row}: The first "{FrameworkContentColumns.DEPTH.value}" value must be 1. Found \"{first_depth}\" instead.')

    for (previous_row, previous_depth), (current_row, current_depth) in zip(depths, depths[1:]):
        if current_depth > previous_depth + 1:
            errors.append(
                f'Row #{current_row}: "{FrameworkContentColumns.DEPTH.value}" ({current_depth}) cannot follow Row #{previous_row} '
                f'with "{FrameworkContentColumns.DEPTH.value}" ({previous_depth}). Maximum allowed value is "{previous_depth + 1}".'
            )

    if errors:
        raise ValueError(
            f'({fct_name}) [{sheet_name}] Inconsistent "{FrameworkContentColumns.DEPTH.value}" values:\n   - '
            + "\n   - ".join(errors)
            + f'\n> 💡 Tip: Start with "{FrameworkContentColumns.DEPTH.value}" = 1, then keep the same depth, increase it by 1, or use any lower positive depth.'
        )



# [CONTENT] Framework {OK} [Check new optional columns : "scores_definition", "depends_on" (Check if actual answer exists in answer sheet), "condition"(Check if condition is valid for "depends_on" values [e.g. no "/" for all depends_on allowed])]
def validate_framework_content(wb: Workbook, df: pd.DataFrame, sheet_name: str, external_refs: List[str] = None, verbose: bool = False, ctx: ConsoleContext = None):

    fct_name = get_current_fct_name()

    # Get required, optional, and translatable columns, along with validation constraints
    content_type = MetaTypes.FRAMEWORK
    schema = CONTENT_SHEET_SCHEMAS[content_type]
    column_constraints = schema.column_constraints

    name_constraints = column_constraints[FrameworkContentColumns.NAME]
    assessable_constraints = column_constraints[FrameworkContentColumns.ASSESSABLE]
    importance_constraints = column_constraints[FrameworkContentColumns.IMPORTANCE]
    weight_constraints = column_constraints[FrameworkContentColumns.WEIGHT]
    min_score_constraints = column_constraints[FrameworkContentColumns.MIN_SCORE]
    max_score_constraints = column_constraints[FrameworkContentColumns.MAX_SCORE]
    questions_constraints = column_constraints[FrameworkContentColumns.QUESTIONS]
    condition_constraints = column_constraints[FrameworkContentColumns.CONDITION]

    validate_content_sheet(df, sheet_name, schema.required_columns, fct_name)
    validate_optional_columns_content_sheet(df, sheet_name, schema.optional_columns, fct_name, verbose, ctx)

    # Check the consistency of the "depth" hierarchy
    _framework_validate_depth_consistency(df, sheet_name)

    # Check that "questions" and "answer" appear together, or not at all
    question_answer_column_names = [FrameworkContentColumns.QUESTIONS, FrameworkContentColumns.ANSWER]
    validate_columns_presence_together(df, question_answer_column_names, sheet_name, fct_name)

    # Check uniqueness of some column values
    validate_unique_column_values(df, [FrameworkContentColumns.REF_ID], sheet_name, fct_name, ctx=ctx)

    # Additional rule: Check that "name" values do not exceed the configured character limit (in order to avoid issues with PostgreSQL DBs)
    validate_column_max_length(df, FrameworkContentColumns.NAME, name_constraints.max_length, sheet_name, fct_name)
    
    # Enforce presence of "assessable" column (even if values can be empty)
    if FrameworkContentColumns.ASSESSABLE.value not in df.columns:
        raise ValueError(f"[{fct_name}] [{sheet_name}] Missing required column \"{FrameworkContentColumns.ASSESSABLE.value}\"")
    
    # Check "assessable" values
    validate_allowed_column_values(df, FrameworkContentColumns.ASSESSABLE, assessable_constraints.allowed_values, sheet_name, fct_name, ctx=ctx)


    # Additional rule: for non-empty rows, at least "ref_id", "name" or "description" must be filled
    _framework_validate_minimum_fields_and_ref_id(df, sheet_name)
    
    # Ensure that the number of "questions" and "answer" entries match per row (1 or same count), or both are empty
    validate_cell_line_count_alignment(df, FrameworkContentColumns.QUESTIONS, FrameworkContentColumns.ANSWER, sheet_name, fct_name, ref_line_break_indicator=questions_constraints.line_break_indicator)

    # Validate columns that reference other sheets (only if they contain non-empty values)
    for column in [FrameworkContentColumns.IMPLEMENTATION_GROUPS.value, FrameworkContentColumns.ANSWER.value]:
        if column in df.columns:
            non_empty_values = df[column].dropna().astype(str).map(str.strip)
            if not non_empty_values[non_empty_values != ""].empty:
                _framework_validate_column_against_reference_sheet(wb, df, column, sheet_name, verbose, ctx)

    # Validate URN-related columns only if they contain non-empty values
    for column in [FrameworkContentColumns.THREATS.value, FrameworkContentColumns.REFERENCE_CONTROLS.value]:
        if column in df.columns:
            non_empty_values = df[column].dropna().astype(str).map(str.strip)
            if not non_empty_values[non_empty_values != ""].empty:
                _framework_validate_framework_column_urns(wb, df, column, sheet_name, external_refs, verbose, ctx)

    # Check if values in "importance" columns are valid
    validate_allowed_column_values(df, FrameworkContentColumns.IMPORTANCE, importance_constraints.allowed_values, sheet_name, fct_name, ctx=ctx)
    
    # Check if values in "condition" columns are valid
    validate_allowed_column_values(df, FrameworkContentColumns.CONDITION, condition_constraints.allowed_values, sheet_name, fct_name, ctx=ctx, split_regex=condition_constraints.split_regex)
    
    # Check if the number of lines in cells of "questions" are coherent with lines in cells of "answer"
    validate_cell_line_count_alignment(df, FrameworkContentColumns.QUESTIONS, FrameworkContentColumns.DEPENDS_ON, sheet_name, fct_name, cmp_can_be_empty=True, ref_line_break_indicator=questions_constraints.line_break_indicator)
    
    # Check if the number of lines in cells of "questions" are coherent with lines in cells of "depends_on"
    validate_cell_line_count_alignment(df, FrameworkContentColumns.QUESTIONS, FrameworkContentColumns.CONDITION, sheet_name, fct_name, cmp_can_be_empty=True, ref_line_break_indicator=questions_constraints.line_break_indicator)
    
    ### A SPECIFIC CHECK FUNCTION SHOULD BE CREATED
    # Check if "condition" exists when "depends_on" is defined
    # validate_cell_line_count_alignment(df, FrameworkContentColumns.DEPENDS_ON, FrameworkContentColumns.CONDITION, sheet_name, fct_name)
    
    # Check if values in "weight" columns are valid
    validate_integer_value(df, sheet_name, FrameworkContentColumns.WEIGHT, fct_name, value_name=FrameworkContentColumns.WEIGHT, min=weight_constraints.min_value, max=weight_constraints.max_value)
    
    # Validate "min_score" and "max_score" values if columns are present
    validate_min_max_columns(df, FrameworkContentColumns.MIN_SCORE, FrameworkContentColumns.MAX_SCORE, sheet_name, fct_name, min_score_constraints, max_score_constraints)


    # Extra locales
    validate_extra_locales_in_content(df, sheet_name, fct_name, ctx, verbose, wb=wb)

    print_sheet_validation(sheet_name, verbose, ctx)


# [CONTENT] Threats {OK}²
def validate_threats_content(df: pd.DataFrame, sheet_name: str, verbose: bool = False, ctx: ConsoleContext = None):
    
    fct_name = get_current_fct_name()

    # Get required, optional, and translatable columns, along with validation constraints
    content_type = MetaTypes.THREATS
    schema = CONTENT_SHEET_SCHEMAS[content_type]

    validate_content_sheet(df, sheet_name, schema.required_columns, fct_name)
    validate_optional_columns_content_sheet(df, sheet_name, schema.optional_columns, fct_name, verbose, ctx)

    # Check uniqueness of some column values
    validate_unique_column_values(df, [ThreatsContentColumns.REF_ID], sheet_name, fct_name, ctx=ctx)

    # Extra locales
    validate_extra_locales_in_content(df, sheet_name, fct_name, ctx, verbose)

    print_sheet_validation(sheet_name, verbose, ctx)


# [CONTENT] Reference Controls {OK}²
def validate_reference_controls_content(df: pd.DataFrame, sheet_name: str, verbose: bool = False, ctx: ConsoleContext = None):
    
    fct_name = get_current_fct_name()

    # Get required, optional, and translatable columns, along with validation constraints
    content_type = MetaTypes.REFERENCE_CONTROLS
    schema = CONTENT_SHEET_SCHEMAS[content_type]
    column_constraints = schema.column_constraints

    category_constraints = column_constraints[ReferenceControlsContentColumns.CATEGORY]
    csf_function_constraints = column_constraints[ReferenceControlsContentColumns.CSF_FUNCTION]

    validate_content_sheet(df, sheet_name, schema.required_columns, fct_name)
    validate_optional_columns_content_sheet(df, sheet_name, schema.optional_columns, fct_name, verbose, ctx)

    # Check uniqueness of some column values
    validate_unique_column_values(df, [ReferenceControlsContentColumns.REF_ID], sheet_name, fct_name, ctx=ctx)

    # Check if values in "category" and "csf_function" columns are valid
    validate_allowed_column_values(df, ReferenceControlsContentColumns.CATEGORY, category_constraints.allowed_values, sheet_name, fct_name, ctx=ctx)
    validate_allowed_column_values(df, ReferenceControlsContentColumns.CSF_FUNCTION, csf_function_constraints.allowed_values, sheet_name, fct_name, ctx=ctx)

    # Extra locales
    validate_extra_locales_in_content(df, sheet_name, fct_name, ctx, verbose)

    print_sheet_validation(sheet_name, verbose, ctx)


# [CONTENT] Risk Matrix²
def validate_risk_matrix_content(df: pd.DataFrame, sheet_name: str, verbose: bool = False, ctx: ConsoleContext = None):
    
    fct_name = get_current_fct_name()

    # Get required, optional, and translatable columns, along with validation constraints
    content_type = MetaTypes.RISK_MATRIX
    schema = CONTENT_SHEET_SCHEMAS[content_type]
    column_constraints = schema.column_constraints

    type_constraints = column_constraints[RiskMatrixContentColumns.TYPE]

    validate_content_sheet(df, sheet_name, schema.required_columns, fct_name)
    validate_optional_columns_content_sheet(df, sheet_name, schema.optional_columns, fct_name, verbose, ctx)

    # Check if values in "type" column are valid
    validate_allowed_column_values(df, RiskMatrixContentColumns.TYPE, type_constraints.allowed_values, sheet_name, fct_name, ctx=ctx)

    # Extra locales
    validate_extra_locales_in_content(df, sheet_name, fct_name, ctx, verbose)


    msg = (
        f"⚠️  [WARNING] ({fct_name}) [{sheet_name}] In this script, Matrix content sheet verification is partially implemented."
        f"\n> 💡 Tip: Matrix verification will be improved in a future update."
    )
    print(msg)
    if ctx:
        ctx.add_sheet_warning_msg(sheet_name, msg)

    print_sheet_validation(sheet_name, verbose, ctx)


# [CONTENT] Implementation Groups {OK}²
def validate_implementation_groups_content(wb: Workbook, df: pd.DataFrame, sheet_name: str, verbose: bool = False, ctx: ConsoleContext = None):
    
    fct_name = get_current_fct_name()

    # Get required, optional, and translatable columns, along with validation constraints
    content_type = MetaTypes.IMPLEMENTATION_GROUPS
    schema = CONTENT_SHEET_SCHEMAS[content_type]
    column_constraints = schema.column_constraints

    default_selected_constraints = column_constraints[ImplementationGroupsContentColumns.DEFAULT_SELECTED]

    validate_content_sheet(df, sheet_name, schema.required_columns, fct_name)
    validate_optional_columns_content_sheet(df, sheet_name, schema.optional_columns, fct_name, verbose, ctx)

    # Check uniqueness of some column values
    validate_unique_column_values(df, [ImplementationGroupsContentColumns.REF_ID], sheet_name, fct_name, ctx=ctx)
    
    # Check "default_selected" values
    validate_allowed_column_values(df, ImplementationGroupsContentColumns.DEFAULT_SELECTED, default_selected_constraints.allowed_values, sheet_name, fct_name, ctx=ctx)

    # Extra locales
    validate_extra_locales_in_content(df, sheet_name, fct_name, ctx, verbose)

    # Check if the "implementation_groups" sheet is actually used in a "framework" sheet
    frameworks_with_imp_grp = check_content_sheet_usage_in_frameworks(wb, sheet_name, FrameworkMetaKeys.IMPLEMENTATION_GROUPS_DEFINITION, fct_name, ctx)
    frameworks_with_imp_grp = get_corresponding_type_sheet_names(frameworks_with_imp_grp, SheetTypes.CONTENT)

    # Check if every implementation groups are actually used in "framework" sheets
    if frameworks_with_imp_grp:
        _implementation_groups_check_unused_default_ids_in_frameworks(wb, df, frameworks_with_imp_grp, sheet_name)
        check_unused_ids_in_frameworks(wb, df, ImplementationGroupsContentColumns.REF_ID, FrameworkContentColumns.IMPLEMENTATION_GROUPS, frameworks_with_imp_grp, sheet_name, fct_name, ctx, verbose)

    print_sheet_validation(sheet_name, verbose, ctx)


# [CONTENT] Requirement Mapping Set {OK}²
def validate_requirement_mapping_set_content(wb: Workbook, df: pd.DataFrame, sheet_name: str, verbose: bool = False, ctx: ConsoleContext = None):
    
    fct_name = get_current_fct_name()

    # Get required, optional, and translatable columns, along with validation constraints
    content_type = MetaTypes.REQUIREMENT_MAPPING_SET
    schema = CONTENT_SHEET_SCHEMAS[content_type]
    column_constraints = schema.column_constraints

    relationship_constraints = column_constraints[RequirementMappingSetContentColumns.RELATIONSHIP]
    rationale_constraints = column_constraints[RequirementMappingSetContentColumns.RATIONALE]

    validate_content_sheet(df, sheet_name, schema.required_columns, fct_name)
    validate_optional_columns_content_sheet(df, sheet_name, schema.optional_columns, fct_name, verbose, ctx)

    # Extra locales (Not needed for mappings, but added just in case)
    validate_extra_locales_in_content(df, sheet_name, fct_name, ctx, verbose)

    # Check if there are duplicated mappings
    _req_map_set_validate_unique_mappings(df, sheet_name, ctx=ctx)

    # Check if values in "relationship" and "rationale" columns are valid
    validate_allowed_column_values(df, RequirementMappingSetContentColumns.RELATIONSHIP, relationship_constraints.allowed_values, sheet_name, fct_name, ctx=ctx)
    validate_allowed_column_values(df, RequirementMappingSetContentColumns.RATIONALE, rationale_constraints.allowed_values, sheet_name, fct_name, ctx=ctx)

    # Check mapping validity using the "source" and "target" sheets
    _req_map_set_validate_mapping_node_ids_against_sheets(wb, df, sheet_name, fct_name, ctx, verbose)

    print_sheet_validation(sheet_name, verbose, ctx)


# [CONTENT] Scores {OK} [Add logic checking sheet presence in CONTENT FRAMEWORK sheet in column "scores_definition"]
def validate_scores_content(wb: Workbook, df: pd.DataFrame, sheet_name: str, verbose: bool = False, ctx: ConsoleContext = None):
    
    fct_name = get_current_fct_name()

    # Get required, optional, and translatable columns, along with validation constraints
    content_type = MetaTypes.SCORES
    schema = CONTENT_SHEET_SCHEMAS[content_type]
    column_constraints = schema.column_constraints

    score_constraints = column_constraints[ScoresContentColumns.SCORE]

    validate_content_sheet(df, sheet_name, schema.required_columns, fct_name)
    validate_optional_columns_content_sheet(df, sheet_name, schema.optional_columns, fct_name, verbose, ctx)

    # Validate each "score" value is a non-negative integer
    validate_integer_value(df, sheet_name, ScoresContentColumns.SCORE, fct_name, value_name=ScoresContentColumns.SCORE, min=score_constraints.min_value, max=score_constraints.max_value)

    # Check uniqueness of some column values
    validate_unique_column_values(df, [ScoresContentColumns.SCORE], sheet_name, fct_name, ctx=ctx)

    # Extra locales
    validate_extra_locales_in_content(df, sheet_name, fct_name, ctx, verbose)

    # Check if the "score" sheet is actually used in a "framework" sheet
    check_content_sheet_usage_in_frameworks(wb, sheet_name, FrameworkMetaKeys.SCORES_DEFINITION, fct_name, ctx)

    print_sheet_validation(sheet_name, verbose, ctx)


# [CONTENT] Answers {OK} [Check new optional column: "description", "select_implementation_groups", "add_score", "compute_result", "color"]
def validate_answers_content(wb: Workbook, df: pd.DataFrame, sheet_name: str, verbose: bool = False, ctx: ConsoleContext = None):
    
    fct_name = get_current_fct_name()

    # Get required, optional, and translatable columns, along with validation constraints
    content_type = MetaTypes.ANSWERS
    schema = CONTENT_SHEET_SCHEMAS[content_type]
    column_constraints = schema.column_constraints

    question_type_constraints = column_constraints[AnswersContentColumns.QUESTION_TYPE]

    validate_content_sheet(df, sheet_name, schema.required_columns, fct_name)
    validate_optional_columns_content_sheet(df, sheet_name, schema.optional_columns, fct_name, verbose, ctx)

    # Check uniqueness of some column values
    validate_unique_column_values(df, [AnswersContentColumns.ID], sheet_name, fct_name, ctx=ctx)

    # Check if values in "question_type" column are valid
    validate_allowed_column_values(df, AnswersContentColumns.QUESTION_TYPE, question_type_constraints.allowed_values, sheet_name, fct_name, ctx=ctx)

    # Extra locales
    validate_extra_locales_in_content(df, sheet_name, fct_name, ctx, verbose)

    # Check that "question_choices" is filled for relevant question types ("unique_choice" & "multiple_choice")
    _answers_validate_question_choices(df, sheet_name)

    # Check if the "answers" sheet is actually used in a "framework" sheet
    frameworks_with_answers = check_content_sheet_usage_in_frameworks(wb, sheet_name, FrameworkMetaKeys.ANSWERS_DEFINITION, fct_name, ctx)
    frameworks_with_answers = get_corresponding_type_sheet_names(frameworks_with_answers, SheetTypes.CONTENT)

    # Check if every answers are actually used in "framework" sheets
    if frameworks_with_answers:
        check_unused_ids_in_frameworks(wb, df, AnswersContentColumns.ID, FrameworkContentColumns.ANSWER, frameworks_with_answers, sheet_name, fct_name, ctx, verbose)

    print_sheet_validation(sheet_name, verbose, ctx)


# [CONTENT] URN Prefix {OK}²
def validate_urn_prefix_content(wb: Workbook, df: pd.DataFrame, sheet_name: str, verbose: bool = False, ctx: ConsoleContext = None):
    
    fct_name = get_current_fct_name()

    # Get required, optional, and translatable columns, along with validation constraints
    content_type = MetaTypes.URN_PREFIX
    schema = CONTENT_SHEET_SCHEMAS[content_type]

    validate_content_sheet(df, sheet_name, schema.required_columns, fct_name)
    validate_optional_columns_content_sheet(df, sheet_name, schema.optional_columns, fct_name, verbose, ctx)

    # Check uniqueness of some column values
    validate_unique_column_values(df, [URNPrefixContentColumns.PREFIX_ID, URNPrefixContentColumns.PREFIX_VALUE], sheet_name, fct_name, ctx=ctx)

    # Check if URN Prefix IDs are used in "framework" sheets
    _URN_prefix_validate_ids_usage_in_frameworks(wb, df, sheet_name, ctx, verbose)

    # Check if "prefix_value" come from internal sheets or external framework
    _URN_prefix_validate_prefix_values_and_dependencies(wb, df, sheet_name, ctx, verbose)

    # Extra locales
    validate_extra_locales_in_content(df, sheet_name, fct_name, ctx, verbose)

    print_sheet_validation(sheet_name, verbose, ctx)



# ─────────────────────────────────────────────────────────────
# DISPATCHING
# ─────────────────────────────────────────────────────────────

def dispatch_meta_validation(wb: Workbook, df: pd.DataFrame, sheet_name: str, verbose: bool = False, ctx: ConsoleContext = None):
    
    fct_name = get_current_fct_name()
    
    type_value = get_meta_value(df, MandatoryMetaKeys.TYPE, sheet_name, required=True, context=fct_name)

    if type_value == MetaTypes.LIBRARY.value:
        validate_library_meta(df, sheet_name, verbose, ctx)
    elif type_value == MetaTypes.FRAMEWORK.value:
        validate_framework_meta(wb, df, sheet_name, verbose, ctx)
    elif type_value == MetaTypes.THREATS.value:
        validate_threats_meta(df, sheet_name, verbose, ctx)
    elif type_value == MetaTypes.REFERENCE_CONTROLS.value:
        validate_reference_controls_meta(df, sheet_name, verbose, ctx)
    elif type_value == MetaTypes.RISK_MATRIX.value:
        validate_risk_matrix_meta(df, sheet_name, verbose, ctx)
    elif type_value == MetaTypes.REQUIREMENT_MAPPING_SET.value:
        validate_requirement_mapping_set_meta(df, sheet_name, verbose, ctx)
    elif type_value == MetaTypes.IMPLEMENTATION_GROUPS.value:
        validate_implementation_groups_meta(wb, df, sheet_name, verbose, ctx)
    elif type_value == MetaTypes.SCORES.value:
        validate_scores_meta(wb, df, sheet_name, verbose, ctx)
    elif type_value == MetaTypes.ANSWERS.value:
        validate_answers_meta(wb, df, sheet_name, verbose, ctx)
    elif type_value == MetaTypes.URN_PREFIX.value:
        validate_urn_prefix_meta(df, sheet_name, verbose, ctx)
    else:
        raise ValueError(f"({fct_name}) [{sheet_name}] Unknown meta type \"{type_value}\"")


def dispatch_content_validation(wb: Workbook, df: pd.DataFrame, sheet_name: str, corresponding_meta_type: str, external_refs: List[str] = None, verbose: bool = False, ctx: ConsoleContext = None):
    
    fct_name = get_current_fct_name()
    
    if corresponding_meta_type == MetaTypes.FRAMEWORK.value:
        validate_framework_content(wb, df, sheet_name, external_refs, verbose, ctx)
    elif corresponding_meta_type == MetaTypes.THREATS.value:
        validate_threats_content(df, sheet_name, verbose, ctx)
    elif corresponding_meta_type == MetaTypes.REFERENCE_CONTROLS.value:
        validate_reference_controls_content(df, sheet_name, verbose, ctx)
    elif corresponding_meta_type == MetaTypes.RISK_MATRIX.value:
        validate_risk_matrix_content(df, sheet_name, verbose, ctx)
    elif corresponding_meta_type == MetaTypes.REQUIREMENT_MAPPING_SET.value:
        validate_requirement_mapping_set_content(wb, df, sheet_name, verbose, ctx)
    elif corresponding_meta_type == MetaTypes.IMPLEMENTATION_GROUPS.value:
        validate_implementation_groups_content(wb, df, sheet_name, verbose, ctx)
    elif corresponding_meta_type == MetaTypes.SCORES.value:
        validate_scores_content(wb, df, sheet_name, verbose, ctx)
    elif corresponding_meta_type == MetaTypes.ANSWERS.value:
        validate_answers_content(wb, df, sheet_name, verbose, ctx)
    elif corresponding_meta_type == MetaTypes.URN_PREFIX.value:
        validate_urn_prefix_content(wb, df, sheet_name, verbose, ctx)
    else:
        raise ValueError(f"({fct_name}) [{sheet_name}] Cannot determine validation for content of type \"{corresponding_meta_type}\"")


# ─────────────────────────────────────────────────────────────
# MAIN VALIDATION FUNCTION
# ─────────────────────────────────────────────────────────────

def validate_excel_structure(filepath: str | Path, external_refs: List[str] = None, verbose: bool = False, ctx: ConsoleContext = None):

    fct_name = get_current_fct_name()

    # Check provided YAML external reference
    if external_refs:
        check_file_validity(external_refs, "YAML", ValidFileTypes.YAML.value, "External Reference")

    # Check Excel file
    check_file_validity(filepath, "Excel", ValidFileTypes.EXCEL.value)


    print(f"⌛ Parsing \"{os.path.basename(filepath)}\"...")
    
    if not ctx:
        ctx = ConsoleContext()
    
    wb = load_workbook(filepath, data_only=True)
    fct_name = get_current_fct_name()
    file_name = os.path.basename(filepath)

    meta_sheets = {}
    content_sheets = {}
    ignored_sheets = []
    meta_types = {}

    # Sort sheets
    for sheet_name in wb.sheetnames:
        if sheet_name.endswith(SheetTypes.META.value):
            df = pd.read_excel(filepath, sheet_name=sheet_name, header=None, dtype=str, keep_default_na=False)
            meta_sheets[sheet_name] = df
        elif sheet_name.endswith(SheetTypes.CONTENT.value):
            df = pd.read_excel(filepath, sheet_name=sheet_name, header=0, dtype=str, keep_default_na=False)
            content_sheets[sheet_name] = df
        else:
            ignored_sheets.append(sheet_name)

    if not MandatorySheets.LIBRARY_META.value in meta_sheets:
        raise ValueError(
            f"({fct_name}) [{sheet_name}] No \"{MandatorySheets.LIBRARY_META.value}\" sheet found."
            f"\n> 💡 Tip: Ensure your Excel file \"{file_name}\" is in v2 format."
        )

    # Handle "_meta" sheets
    for sheet_name, df in meta_sheets.items():

        base_name = re.sub(r'_meta$', '', sheet_name)
        
        expected_content_sheet = base_name + SheetTypes.CONTENT.value
        if sheet_name != MandatorySheets.LIBRARY_META.value and expected_content_sheet not in content_sheets:
            raise ValueError(f"({fct_name}) [{sheet_name}] No corresponding content sheet found for this meta"
                            f"\n> 💡 Tip: Make sure the corresponding content sheet for \"{sheet_name}\" is named \"{expected_content_sheet}\"")

        dispatch_meta_validation(wb, df, sheet_name, verbose, ctx)
        meta_types[base_name] = get_meta_value(df, MandatoryMetaKeys.TYPE, sheet_name, required=True, context=fct_name)

    # Check "_content" sheets
    # As some checks in "_content" sheets need to check the contents of other "_content" sheets, we make sure that all such sheets first have a "_meta" sheet
    for sheet_name, df in content_sheets.items():
        base_name = re.sub(r'_content$', '', sheet_name)

        if base_name not in meta_types:
            raise ValueError(f"({fct_name}) [{sheet_name}] No corresponding meta sheet found for this content"
                             f"\n> 💡 Tip: Make sure the corresponding meta sheet for \"{sheet_name}\" is named \"{re.sub(r'_content$', '_meta', sheet_name)}\"")

    # Handle "_content" sheets
    for sheet_name, df in content_sheets.items():
        base_name = re.sub(r'_content$', '', sheet_name)
        dispatch_content_validation(wb, df, sheet_name, meta_types[base_name], external_refs, verbose, ctx)

    # Warn about ignored sheets
    for sheet_name in ignored_sheets:
        msg = f"⏩ [SKIP] Ignored sheet \"{sheet_name}\" (does not end with \"_meta\" or \"_content\")"
        print(msg)

    print("")

    if ctx.count_all_warnings() > 0:
        print(f"✅⚠️  [SUCCESS] Excel structure validation ended with warnings for \"{file_name}\"")
        print(f"📜 [SUMMARY] ⚠️  Total [WARNING] for \"{file_name}\": {ctx.count_all_warnings()}")
    else:
        print(f"✅ [SUCCESS] Excel structure is valid for \"{file_name}\"")

    if verbose and ctx.count_all_verbose() > 0:
        print(f"📜 [SUMMARY] 💬 Total [Verbose Messages] for \"{file_name}\": {ctx.count_all_verbose()}")



# ─────────────────────────────────────────────────────────────
# MAIN ENTRY POINT
# ─────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Validate Excel file structure (v2 format)", formatter_class=argparse.RawTextHelpFormatter)
    parser.add_argument(
        "file_input",
        help="Path to Excel file to validate."
    )

    parser.add_argument(
        "-e", "--external-refs",
        type=str,
        help="YAML files containing external references mentioned in the library.\n"
        "Use it to check the following columns if necessary : \"threats\", \"reference_controls\".\n"
        "Separate external references with commas (e.g., ./threats1.yaml,./refs/ref_ctrl.yaml,../test.yaml)",
    )

    parser.add_argument(
        "-b", "--bulk",
        action="store_true",
        help="Enable bulk mode to process all Excel files in a directory.",
    )

    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Enable verbose output. Verbose messages start with a 💬 (speech bubble) emoji."
    )

    args = parser.parse_args()


    # Store YAML external references filenames in a list 
    external_refs = []
    if args.external_refs:
        external_refs = args.external_refs.split(",")        


    # If "enable_ctx == True", the "ConsoleContext" object will be enabled and returned by the function
    enable_ctx = True
    err = False
    
    # --- BULK CHECK ------------------------------------------------------------
    if args.bulk:
        _, err = bulk_check(args, external_refs, enable_ctx)

    # --- SINGLE FILE CHECK -----------------------------------------------------
    else:
        _, err = single_file_check(args, external_refs, enable_ctx)


    if err:
        sys.exit(1)
    else:
        sys.exit(0)



def bulk_check(args: argparse.Namespace, external_refs: List[str] = None, enable_ctx: bool = False) -> tuple[Dict[str, ConsoleContext], List[str]]:
    
    ctxs: Dict[str, ConsoleContext] = {}   # List of all contexts
    """
    {
        file1: ctx1,
        file2: ctx2,
        ...
    }
    """
    
    input_path = Path(args.file_input)
    if not input_path.is_dir():
        print("❌ [ERROR] Bulk mode requires a directory as input")
        sys.exit(1)


    error_files = []  # Collect names of files that failed
    
    # Find all Excel files in the input directory (temp Excel files starting with "~$" are excluded)
    valid_exts = ValidFileTypes.EXCEL.value
    excel_files = [
        f for f in input_path.iterdir()
        if f.suffix.lower() in valid_exts and not f.name.startswith("~$")
    ]

    if not excel_files:
        print(f'❌ [ERROR] No Excel files found in directory: "{input_path}". Abort...')
        sys.exit(1)

    for i, file in enumerate(excel_files):
        
        temp_ctx = None
        
        if enable_ctx:
            temp_ctx = ConsoleContext()
        
        try:
            if i > 0:
                print("\n-------------------------------------------------------------------\n")
            print(f'▶️  Processing file [{i + 1}/{len(excel_files)}]: "{file}"')
            validate_excel_structure(str(file), external_refs, args.verbose, temp_ctx)
        except Exception as e:
            print(f'❌ [ERROR] Failed to process "{file.name}":\n🛑 {e}')
            error_files.append(file.name)


        if enable_ctx:
            ctxs[file.name] = temp_ctx


    # Summary at the end of bulk processing
    print("\n###################################################################\n")
    print("📋 Bulk mode completed!")


    warning_files = []

    # Check files with at least 1 warning
    for file, ctx in ctxs.items():
        if ctx.count_all_warnings() > 0: warning_files.append(file)


    # Print files that got at least 1 warning
    if warning_files:
        print(f"⚠️  The following file{'s' if len(error_files) > 1 else ''} encountered at least 1 warning:")

        for f in warning_files:
            print(f"   - {f}")


    # Print files that encounter an error
    if error_files:
        print(f"❌ The following file{'s' if len(error_files) > 1 else ''} failed to process:")

        for f in error_files:
            print(f"   - {f}")


    if warning_files or error_files: 
        if not args.verbose:
            print('💡 Tip: Use "--verbose" to display hidden messages. This can help to understand certain errors.')
    else:
        print("✅ All files processed successfully!")


    return ctxs, error_files


def single_file_check(args: argparse.Namespace, external_refs: List[str] = None, enable_ctx: bool = False) -> tuple[ConsoleContext, bool]:
    
    ctx = None
    error_encountered = False
    
    if enable_ctx:
        ctx = ConsoleContext()

    try:
        validate_excel_structure(args.file_input, external_refs, args.verbose, ctx)
        if not args.verbose:
                print("💡 Tip: Use \"--verbose\" to display hidden messages. This can help to understand certain errors.")
    except Exception as e:
        print(f"❌ [FATAL ERROR] {e}")
        if not args.verbose:
                print("💡 Tip: Use \"--verbose\" to display hidden messages. This can help to understand certain errors.")
        error_encountered = True

    return ctx, error_encountered


if __name__ == "__main__":
    main()
