"""
simple script to transform the official CCM Excel file to another Excel file for CISO assistant framework conversion tool
"""

import openpyxl
import sys
import re
import argparse
from openpyxl.styles import numbers


def pretify_content(content):
    res = None
    stop_join = False
    for line in content.splitlines():
        if stop_join:
            res = res + "\n" + line if res else line
        else:
            res = res + " " + line if res else line
        if line[-1] == ":":
            stop_join = True
    return res


parser = argparse.ArgumentParser(
    prog="convert_ccm",
    description="convert CCM controls offical Excel file to CISO Assistant Excel file",
)
parser.add_argument("filename", help="name of CCM controls Excel file")
parser.add_argument("packager", help="name of packager entity")

args = parser.parse_args()
input_file_name = args.filename
packager = args.packager
output_file_name = "ccm-controls-v4.xlsx"

print("parsing", input_file_name)

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook(input_file_name)
output_table = []

for tab in dataframe:
    print("parsing tab", tab.title)
    title = tab.title
    if title == "CCM":
        line = 0
        eos = False
        for row in tab:
            line += 1
            if line < 4:
                continue
            (domain, title, id, specification, lite) = (r.value for r in row[0:5])
            if eos:
                library_copyright = domain  # last line after end of standard
            elif lite:
                output_table.append(
                    (
                        "x",
                        2,
                        id,
                        title,
                        pretify_content(specification),
                        "lite,full" if lite == "Yes" else "full",
                        "",
                    )
                )
            else:
                if "End of Standard" in domain:
                    eos = True
                else:
                    (d, id) = domain.split(" - ")
                    output_table.append(("", 1, id, d, "", None, None))
    elif title == "CAIQ":
        line = 0
        eos = False
        for row in tab:
            line += 1
            if line < 4:
                continue
            (question_id, question) = (r.value for r in row[4:6])
            if question_id:
                q = re.match(r"(.*)\.\d+$", question_id)
                id = q.group(1)
                for i in range(len(output_table)):
                    if output_table[i][2] == id:
                        a = output_table[i][6]
                        b = pretify_content(question)
                        c = b if a == "" else a + "\n" + b
                        output_table[i] = output_table[i][0:6] + (c,)
    else:
        print(f"Ignored tab: {title}")


print("generating", output_file_name)
wb_output = openpyxl.Workbook()
ws = wb_output.active
ws.title = "library_content"
ws.append(["library_urn", f"urn:{packager.lower()}:risk:library:ccm-controls-v4"])
ws.append(["library_version", "2"])
ws.append(["library_locale", "en"])
ws.append(["library_ref_id", "CCM-Controls-v4"])
ws.append(["library_name", "CCM Controls v4"])
ws.append(["library_description", "CCM Controls v4"])
ws.append(["library_copyright", library_copyright])
ws.append(["library_provider", "CSA"])
ws.append(["library_packager", packager])
ws.append(["framework_urn", f"urn:{packager.lower()}:risk:framework:ccm-controls-v4"])
ws.append(["framework_ref_id", "CCM-Controls-v4"])
ws.append(["framework_name", "CCM Controls v4"])
ws.append(["framework_description", "CCM Controls v4"])
ws.append(["tab", "controls", "requirements"])
ws.append(["tab", "implementation_groups", "implementation_groups"])

ws1 = wb_output.create_sheet("controls")
ws1.append(
    [
        "assessable",
        "depth",
        "ref_id",
        "name",
        "description",
        "implementation_groups",
        "annotation",
    ]
)
for row in output_table:
    ws1.append(row)
ws2 = wb_output.create_sheet("implementation_groups")
ws2.append(["ref_id", "name", "description"])
ws2.append(
    [
        "lite",
        "foundational",
        "foundational controls that should be implemented by any organization, regardless of their budget, maturity and risk profile",
    ]
)
ws2.append(["full", "systematic ", "systematic assessment of a cloud implementation"])
print("generate ", output_file_name)
wb_output.save(output_file_name)
