"""
Simple script to convert TISAX v5.1 excel in a CISO Assistant Excel file
Source;  https://portal.enx.com/isa6-en.xlsx
"""

import openpyxl
import sys
import re
import argparse
from openpyxl.styles import numbers

parser = argparse.ArgumentParser(
    prog="convert_tisax_5.1",
    description="convert TISAX controls offical v5.1 Excel file to CISO Assistant Excel file",
)

parser.add_argument("filename", help="name of official TISAX Excel file")
args = parser.parse_args()
input_file_name = args.filename
output_file_name = "tisax-v5.1.xlsx"

library_copyright = """Publisher: VERBAND DER AUTOMOBILINDUSTRIE e. V. (VDA, German Association of the Automotive Industry); Behrenstr. 35; 10117 Berlin; www.vda.de
© 2022 Verband der Automobilindustrie e.V., Berlin
This work has been licensed under Creative Commons Attribution - No Derivative Works 4.0 International Public License. In addition, You are granted the right to distribute derivatives under certain terms."""
packager = "intuitem"

library_description = """VDA ISA provides the basis for
- a self-assessment to determine the state of information security in an organization (e.g. company)
- audits performed by internal departments (e.g. Internal Audit, Information Security)
- a review in accordance with TISAX (Trusted Information Security Assessment Exchange, http://enx.com/tisax/)
Source: https://portal.enx.com/isa5-en.xlsx
"""

print("parsing", input_file_name)

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook(input_file_name)
output_table = []

for tab in dataframe:
    print("parsing tab", tab.title)
    title = tab.title
    if title in ("Information Security", "Prototype Protection", "Data protection"):
        for row in tab:
            req_must = None
            req_should = None
            req_very_high = None
            req_high = None
            req_sga = None
            req_vehicle = None
            further_info = None
            ex_normal = None
            ex_high = None
            ex_very_high = None
            if title == "Information Security":
                (
                    _,
                    _,
                    _,
                    control_number,
                    _,
                    _,
                    _,
                    _,
                    control_question,
                    objective,
                    req_must,
                    req_should,
                    req_high,
                    req_very_high,
                    _,
                    _,
                    _,
                    _,
                    _,
                    _,
                    _,
                    further_info,
                    ex_normal,
                    ex_high,
                    ex_very_high,
                ) = (r.value for r in row[0:25])
            elif title == "Prototype Protection":
                (
                    _,
                    _,
                    _,
                    control_number,
                    _,
                    _,
                    _,
                    _,
                    control_question,
                    objective,
                    req_must,
                    req_should,
                    req_vehicle,
                ) = (r.value for r in row[0:13])
            elif title == "Data protection":
                (
                    _,
                    _,
                    _,
                    control_number,
                    _,
                    _,
                    _,
                    _,
                    control_question,
                    req_must,
                ) = (r.value for r in row[0:10])
            if type(control_number) == int:
                control_number = str(control_number)
            if control_number and re.fullmatch(r"\d", control_number):
                level = 2
                print(control_number, control_question)
                output_table.append(("", 1, control_number, control_question, "", ""))
            if control_number and re.fullmatch(r"\d\.\d+", control_number):
                level = 3
                print(control_number, control_question)
                output_table.append(("", 2, control_number, "", control_question, ""))
                if req_must:
                    output_table.append(("x", 3, "", "(must)", req_must, "must"))
            if control_number and re.fullmatch(r"\d\.\d+\.\d+", control_number):
                if control_question and re.match(r"Superseded by", control_question):
                    print("skipping", control_number)
                output_table.append(("", level, control_number, "", control_question, ""))
                output_table.append(("x", level + 1, "", "(must)", req_must, "must"))
                if req_should and req_should != "None":
                    output_table.append(
                        ("x", level + 1, "", "(should)", req_should, "should")
                    )
                if req_high and req_high != "None":
                    output_table.append(
                        (
                            "x",
                            level + 1,
                            "",
                            "(for high protection needs)",
                            req_high,
                            "high",
                        )
                    )
                if req_very_high and req_very_high != "None":
                    output_table.append(
                        (
                            "x",
                            level + 1,
                            "",
                            "(for very high protection needs)",
                            req_very_high,
                            "very_high",
                        )
                    )
                if req_sga and req_sga != "None":
                    output_table.append(
                        (
                            "x",
                            level + 1,
                            "",
                            "(for Simplified Group Assessments)",
                            req_sga,
                            "SGA",
                        )
                    )
                if req_vehicle and req_vehicle != "None":
                    output_table.append(
                        (
                            "x",
                            level + 1,
                            "",
                            "(for vehicles classified as requiring protection)",
                            req_vehicle,
                            "vehicle",
                        )
                    )
                if further_info:
                    output_table.append(
                        ("", level + 1, "", "Further information", further_info)
                    )

print("generating", output_file_name)
wb_output = openpyxl.Workbook()
ws = wb_output.active
ws.title = "library_content"
ws.append(["library_urn", f"urn:{packager.lower()}:risk:library:tisax-v5.1"])
ws.append(["library_version", "1"])
ws.append(["library_locale", "en"])
ws.append(["library_ref_id", "TISAX v5.1"])
ws.append(["library_name", "Trusted Information Security Assessment Exchange    "])
ws.append(["library_description", library_description])
ws.append(["library_copyright", library_copyright])
ws.append(["library_provider", "VDA"])
ws.append(["library_packager", packager])
ws.append(["framework_urn", f"urn:{packager.lower()}:risk:framework:tisax-v5.1"])
ws.append(["framework_ref_id", "TISAX v5.1"])
ws.append(["framework_name", "Trusted Information Security Assessment Exchange"])
ws.append(["framework_description", library_description])
ws.append(["framework_min_score", 0])
ws.append(["framework_max_score", 5])
ws.append(["tab", "controls", "requirements"])
ws.append(["tab", "scores", "scores"])
ws.append(["tab", "implementation_groups", "implementation_groups"])

ws1 = wb_output.create_sheet("controls")
ws1.append(
    ["assessable", "depth", "ref_id", "name", "description", "implementation_groups"]
)
for row in output_table:
    ws1.append(row)

ws2 = wb_output.create_sheet("scores")
ws2.append(["score", "name", "description"])
ws2.append(
    [
        0,
        "Incomplete",
        "A process is not available, not followed or not suitable for achieving the objective.",
    ]
)
ws2.append(
    [
        1,
        "Performed",
        "An undocumented or incompletely documented process is followed and indicators exist that it achieves its objective.",
    ]
)
ws2.append(
    [
        2,
        "Managed",
        "A process achieving its objectives is followed. Process documentation and process implementation evidence are available.",
    ]
)
ws2.append(
    [
        3,
        "Established",
        "A standard process integrated into the overall system is followed. Dependencies on other processes are documented and suitable interfaces are created. Evidence exists that the process has been used sustainably and actively over an extended period.",
    ]
)
ws2.append(
    [
        4,
        "Predictable",
        "An established process is followed. The effectiveness of the process is continually monitored by collecting key figures. Limit values are defined at which the process is considered to be insufficiently effective and requires adjustment. (Key Performance Indicators)",
    ]
)
ws2.append(
    [
        5,
        "Optimizing",
        "A predictable process with continual improvement as a major objective is followed. Improvement is actively advanced by dedicated resources.",
    ]
)

ws3 = wb_output.create_sheet("implementation_groups")
ws3.append(["ref_id", "name", "description"])
ws3.append(
    ["must", "Requirements (must)", "Strict requirements without any exemptions."]
)
ws3.append(
    [
        "should",
        "Requirements (should)",
        "Must be implemented by the organization. In certain circumstances, however, there may be a valid justification for non-compliance with these requirements. In case of any deviation, its effects must be understood by the organization and it must be plausibly justified.",
    ]
)
ws3.append(
    [
        "high",
        "In case of high protection needs",
        "Must additionally be met if the tested subject has high protection needs.",
    ]
)
ws3.append(
    [
        "very_high",
        "In case of very high protection needs",
        "Must additionally be met if the tested subject has very high protection needs.",
    ]
)
ws3.append(
    [
        "SGA",
        "For Simplified Group Assessments (SGA)",
        "A simplified way to audit very large organizations with a high maturity. An example is the TISAX Simplified Group Assessment mechanism that is an option for TISAX Assessments of an assessment scope with a large number of sites.",
    ]
)
ws3.append(
    [
        "vehicle",
        "For vehicles classified as requiring protection",
        "Protects physical prototypes which are classified as requiring protection. Prototypes include vehicles, components and parts. The owner of the intellectual property for the prototype is considered the owner of the prototype. The owner's commissioning department is responsible for classifying the protection need of a prototype. For prototypes classified as requiring high or very high protection, the minimum requirements for prototype protection must be applied.",
    ]
)

print("generate ", output_file_name)
wb_output.save(output_file_name)
