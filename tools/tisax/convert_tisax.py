'''
Simple script to convert TISAX v6.0.2 excel in a CISO Assistant Excel file
Source;  https://portal.enx.com/isa6-en.xlsx
'''

import openpyxl
import sys
import re
import argparse
from openpyxl.styles import numbers

parser = argparse.ArgumentParser(
                    prog='convert_tisax',
                    description='convert TISAX controls offical v6.0.2 Excel file to CISO Assistant Excel file')

parser.add_argument('filename', help='name of official TISAX Excel file')
args = parser.parse_args()
input_file_name = args.filename
output_file_name = "tisax-v6.0.2.xlsx"

library_copyright = '''© 2023 ENX Association, an Association according to the French Law of 1901, registered under No. w923004198 at the Sous-préfecture of Boulogne-Billancourt, France.
This work of ENX's Working Group ISA was provided to the VDA in the present version by the ENX Association for published by the VDA as the VDA ISA. It is made to all interested parties free of charge under the following licensing terms. The release in the VDA is done by the VDA's Working Group Information Security and Economic Protection. Publication takes place with the consent of the rights holder. The VDA is responsible for the publication of the VDA ISA.
The Tab ""Data Protection"" is provided, owned and copyrighted by VERBAND DER AUTOMOBILINDUSTRIE e.V. (VDA, German Association of the Automotive Industry); Behrenstr. 35; 10117 Berlin"
This work has been licensed under Creative Commons Attribution - No Derivative Works 4.0 International Public License. In addition, You are granted the right to distribute derivatives under certain terms as detailed in section 9 which are not part of the Creative Commons license. The complete and valid text of the license is to be found in line 17ff.
'''
packager = 'intuitem'

library_description = '''ISA provides the basis for
- a self-assessment to determine the state of information security in an organization (e.g. company)
- audits performed by internal departments (e.g. Internal Audit, Information Security)
- TISAXⓇ Assessments (Trusted Information Security Assessment Exchange, https://enx.com/tisax/)
Source: https://portal.enx.com/isa6-en.xlsx
'''

print("parsing", input_file_name)

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook(input_file_name)
output_table = []

for tab in dataframe:
    print("parsing tab", tab.title)
    title = tab.title
    if title in ("Information Security", "Prototype Protection", "Data Protection"):
        for row in tab:
            (_, _, control_number, _, _, _, _, control_question, objective, req_must, req_should, req_high, req_very_high, req_sga, usual_resp, _, _, _, _, _, _, _,
                    further_info, ex_normal, ex_high, ex_very_high) = (r.value for r in row[0:26])
            if type(control_number) == int:
                control_number = str(control_number)
            if control_number and re.fullmatch(r'\d', control_number):
                level=2
                print(control_number, control_question)
                output_table.append(('', 1, control_number, control_question, ''))
            if control_number and re.fullmatch(r'\d\.\d+', control_number):
                level=3
                print(control_number, control_question)
                output_table.append(('', 2, control_number, control_question, ''))
            if control_number and re.fullmatch(r'\d\.\d+\.\d+', control_number):
                if re.match(r'Superseded by', control_question):
                    print("skipping", control_number)
                #print(control_number, control_question)
                output_table.append(('', level, control_number, control_question, ''))
                output_table.append(('x', level+1, '', '(must)', req_must))
                if req_should and req_should != 'None':
                    output_table.append(('x', level+1, '', '(should)', req_should))
                if req_high and req_high != 'None':
                    output_table.append(('x', level+1, '', '(for high protection needs)', req_high))
                if req_very_high and req_very_high != 'None':
                    output_table.append(('x', level+1, '', '(for very high protection needs)', req_very_high))
                if req_sga and req_sga != 'None':
                    output_table.append(('x', level+1, '', '(for Simplified Group Assessments)', req_sga))
                if further_info:
                    output_table.append(('', level+1, '', 'Further information', further_info))

print("generating", output_file_name)
wb_output = openpyxl.Workbook()
ws = wb_output.active
ws.title='library_content'
ws.append(['library_urn', f'urn:{packager.lower()}:risk:library:tisax-v6.0.2'])
ws.append(['library_version', '1'])
ws.append(['library_locale', 'en'])
ws.append(['library_ref_id', 'TISAX v6.0.2'])
ws.append(['library_name', 'Trusted Information Security Assessment Exchange    '])
ws.append(['library_description', library_description])
ws.append(['library_copyright', library_copyright])
ws.append(['library_provider', 'VDA'])
ws.append(['library_packager', packager])
ws.append(['framework_urn', f'urn:{packager.lower()}:risk:framework:tisax-v6.0.2'])
ws.append(['framework_ref_id', 'TISAX v6.0.2'])
ws.append(['framework_name', 'Trusted Information Security Assessment Exchange'])
ws.append(['framework_description', library_description])
ws.append(['tab', 'controls', 'requirements'])

ws1 = wb_output.create_sheet("controls")
ws1.append(['assessable', 'depth', 'ref_id', 'name', 'description'])
for row in output_table:
    ws1.append(row)
print("generate ", output_file_name)
wb_output.save(output_file_name)
