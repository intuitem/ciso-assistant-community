import openpyxl
import sys
import re
import yaml
from pprint import pprint
from collections import defaultdict

if len(sys.argv) <= 1:
    print("missing input file parameter")
    exit()
input_file_name = sys.argv[1]
ref_name = re.sub(r"\.\w+$", "", input_file_name).lower()
output_file_name = ref_name + ".yaml"

print("parsing", input_file_name)

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook(input_file_name)
wb_output = openpyxl.Workbook()
ws = wb_output.active


def error(message):
    print("Error:", message)
    exit(1)


def read_header(row):
    i = 0
    header = {}
    for v in row:
        v = str(v.value).lower()
        header[v] = i
        i += 1
    return header


ws.cell(row=1, column=1, value="assessable")
ws.cell(row=1, column=2, value="depth")
ws.cell(row=1, column=3, value="ref_id")
ws.cell(row=1, column=4, value="name")
ws.cell(row=1, column=5, value="description")
line = 2
for tab in dataframe:
    print("parsing tab", tab.title)
    title = tab.title
    print("...processing content")
    for row in tab:
        if any([r.value for r in row]):
            (v1, v2, v3, v4) = (r.value for r in row[0:4])
            if v1:
                if ":" in v1:
                    print(v1)
                    q = re.match("(\w+) \((\w+)\): (.*)", v1)
                    function_name = q.group(1)
                    function_id = q.group(2)
                    function_description = q.group(3)
                    ws.cell(row=line, column=2, value=1)
                    ws.cell(row=line, column=3, value=function_id)
                    ws.cell(row=line, column=4, value=function_name)
                    ws.cell(row=line, column=5, value=function_description)
                    line += 1
            elif v2:
                q = re.match("([\w\s,]+) \((\w+.\w+)\): (.*)", v2)
                category_name = q.group(1)
                category_id = q.group(2)
                category_description = q.group(3)
                ws.cell(row=line, column=2, value=2)
                ws.cell(row=line, column=3, value=category_id)
                ws.cell(row=line, column=4, value=category_name)
                ws.cell(row=line, column=5, value=category_description)
                line += 1
            elif v3:
                q = re.match("(\w+.\w+-\d+): (.*)", v3)
                subcategory_id = q.group(1)
                subcategory_description = q.group(2)
                ws.cell(row=line, column=1, value="x")
                ws.cell(row=line, column=2, value=3)
                ws.cell(row=line, column=3, value=subcategory_id)
                ws.cell(row=line, column=5, value=subcategory_description)
                line += 1
                ws.cell(row=line, column=2, value=4)
                ws.cell(row=line, column=4, value="Examples")
                ws.cell(row=line, column=5, value=v4)
                line += 1


wb_output.save("nist_csf-2.0-en.xlsx")
