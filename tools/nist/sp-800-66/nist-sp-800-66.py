"""
Simple script to convert NIST SP-800-66 excel in a CISO Assistant Excel file
Source;  https://csrc.nist.gov/Projects/cprt/catalog#/cprt/framework/version/SP800_66_2_0_0/home
"""

import openpyxl
import sys
import re
import argparse
from openpyxl.styles import numbers

parser = argparse.ArgumentParser(
    prog="convert_nist-sp-800-66",
    description="convert NIST SP-800-66 controls offical Excel file to CISO Assistant Excel file",
)

parser.add_argument("filename", help="name of official NIST SP-800-66 Excel file")
args = parser.parse_args()
input_file_name = args.filename
output_file_name = "nist-sp-800-66-rev2.xlsx"

library_copyright = """With the exception of material marked as copyrighted, information presented on NIST sites are considered public information and may be distributed or copied."""
packager = "intuitem"

library_description = """Implementing the Health Insurance Portability and Accountability Act (HIPAA) Security Rule: A Cybersecurity Resource Guide, 2.0.0
Source: https://csrc.nist.gov/Projects/cprt/catalog#/cprt/framework/version/SP800_66_2_0_0/home
"""

print("parsing", input_file_name)

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook(input_file_name)
output_table = []

for tab in dataframe:
    print("parsing tab", tab.title)
    title = tab.title
    if title == "NIST SP 800-66":
        line = 0
        for row in tab:
            line += 1
            if line > 1:
                (
                    security_rule_id,
                    security_rule,
                    std_id,
                    std,
                    key_activity,
                    description,
                    sample_questions,
                ) = (r.value for r in row)
                if security_rule_id:
                    output_table.append(("", 1, security_rule_id, None, security_rule))
                if std_id:
                    output_table.append(("", 2, std_id, None, std))
                output_table.append(("x", 3, None, key_activity, description))
                output_table.append(("", 4, None, "Sample questions", sample_questions))


print("generating", output_file_name)
wb_output = openpyxl.Workbook()
ws = wb_output.active
ws.title = "library_content"
ws.append(["library_urn", f"urn:{packager.lower()}:risk:library:nist-sp-800-66-rev2"])
ws.append(["library_version", "1"])
ws.append(["library_locale", "en"])
ws.append(["library_ref_id", "NIST-SP-800-66-rev2"])
ws.append(["library_name", "NIST SP-800-66 rev2 (HIPAA)"])
ws.append(["library_description", library_description])
ws.append(["library_copyright", library_copyright])
ws.append(["library_provider", "NIST"])
ws.append(["library_packager", packager])
ws.append(
    ["framework_urn", f"urn:{packager.lower()}:risk:framework:nist-sp-800-66-rev2"]
)
ws.append(["framework_ref_id", "nist-sp-800-66-rev2"])
ws.append(["framework_name", "NIST SP-800-66 rev2 (HIPAA)"])
ws.append(["framework_description", library_description])
ws.append(["tab", "controls", "requirements"])

ws1 = wb_output.create_sheet("controls")
ws1.append(["assessable", "depth", "ref_id", "name", "description"])
for row in output_table:
    ws1.append(row)
print("generate ", output_file_name)
wb_output.save(output_file_name)
