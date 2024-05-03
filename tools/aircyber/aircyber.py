"""
Simple script to convert AirCyber v1.5.2 excel in a CISO Assistant Excel file
Source;  https://boostaerospace.com/aircyber/
"""

import openpyxl
import sys
import re
import argparse
from openpyxl.styles import numbers

parser = argparse.ArgumentParser(
    prog="convert_aircyber",
    description="convert AirCyber controls offical v1.5.2 Excel file to CISO Assistant Excel file",
)

parser.add_argument("filename", help="name of official AirCyber Excel file")
args = parser.parse_args()
input_file_name = args.filename
output_file_name = "aircyber-v1.5.2.xlsx"

library_copyright = """© Boost Aerospace
This work is licensed under a Creative Commons Attribution-NonCommercial-ShareAlike 4.0 International License. Any commercial use of this work must be contracted with BoostAeroSpace.
Permission given to include AirCyber in CISO Assistant.
"""
packager = "intuitem"

library_description = """AirCyber is the AeroSpace and Defense official standard for Cybersecurity maturity evaluation and increase built by Airbus, Dassault Aviation, Safran and Thales to help the AeroSpace SupplyChain to be more resilient.
Their joint venture BoostAeroSpace is offering this extract of the AirCyber maturity level matrix to provide further details on this standard, the questions and the AirCyber maturity levels they are associated to.
AirCyber program uses this maturity level matrix as the base of the cyber maturity evaluation as is the evaluation activity is the very starting point for any cyber maturity progression. Being aware of the problems is the mandatory very first knowledge a company shall know to decide to launch a cybersecurity company program.
Source: https://boostaerospace.com/aircyber/
"""

print("parsing", input_file_name)

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook(input_file_name)
output_table = []

for tab in dataframe:
    print("parsing tab", tab.title)
    title = tab.title
    if title == "Listing":
        line = 0
        for row in tab:
            line += 1
            if line > 2:
                (
                    _,
                    question_number,
                    question_name,
                    question_en,
                    question_fr,
                    level,
                    cmr,
                    industrial_it,
                    corporate_it,
                    product,
                    devenv,
                    _,
                    _,
                ) = (r.value for r in row)
                if question_number[0:3] == "Ext":
                    if industrial_it:
                        question_en += "\n[Industrial IT]"
                    if corporate_it:
                        question_en += "\n[Corporate IT]"
                    if product:
                        question_en += "\n[Product]"
                    if devenv:
                        question_en += "\n[Development Environment]"
                output_table.append(
                    ("x", 1, question_number, question_name, question_en, level)
                )


print("generating", output_file_name)
wb_output = openpyxl.Workbook()
ws = wb_output.active
ws.title = "library_content"
ws.append(["library_urn", f"urn:{packager.lower()}:risk:library:aircyber-v1.5.2"])
ws.append(["library_version", 1])
ws.append(["library_locale", "en"])
ws.append(["library_ref_id", "AirCyber-v1.5.2"])
ws.append(["library_name", "Public AirCyber Maturity Level Matrix"])
ws.append(["library_description", library_description])
ws.append(["library_copyright", library_copyright])
ws.append(["library_provider", "Boost Aerospace"])
ws.append(["library_packager", packager])
ws.append(["framework_urn", f"urn:{packager.lower()}:risk:framework:aircyber-v1.5.2"])
ws.append(["framework_ref_id", "AirCyber-v1.5.2"])
ws.append(["framework_name", "Public AirCyber Maturity Level Matrix"])
ws.append(["framework_description", library_description])
ws.append(["tab", "controls", "requirements"])
ws.append(["tab", "implementation_groups", "implementation_groups"])

ws1 = wb_output.create_sheet("controls")
ws1.append(
    ["assessable", "depth", "ref_id", "name", "description", "implementation_groups"]
)
for row in output_table:
    ws1.append(row)
ws2 = wb_output.create_sheet("implementation_groups")
ws2.append(["ref_id", "name", "description"])
ws2.append(["Bronze", "", ""])
ws2.append(["Silver", "", ""])
ws2.append(["Gold", "", ""])

print("generate ", output_file_name)
wb_output.save(output_file_name)
