import openpyxl
import sys
import re
import yaml
from pprint import pprint
from collections import defaultdict

if len(sys.argv) <= 1:
    print("missing input file parameter")
    exit()
input_file_name = sys.argv[1]
ref_name = re.sub(r"\.\w+$", "", input_file_name).lower()
output_file_name = ref_name + ".yaml"

print("parsing", input_file_name)

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook(input_file_name)


def error(message):
    print("Error:", message)
    exit(1)


def read_header(row):
    i = 0
    header = {}
    for v in row:
        v = str(v.value).lower()
        header[v] = i
        i += 1
    return header


def get_name_description(v):
    """skip first line and join next lines"""
    res = v.split("\n")
    name = res[0]
    description = " ".join(res[1:])
    return (name, description)


output_table = []

output_table.append(
    (
        "assessable",
        "depth",
        "ref_id",
        "name",
        "description",
        "annotation",
        "name[de]",
        "name[fr]",
        "name[it]",
        "description[de]",
        "description[fr]",
        "description[it]",
    )
)

line = 0
for tab in dataframe:
    print("parsing tab", tab.title)
    title = tab.title
    if title == "Assessment":
        print("...processing content")
        for row in tab:
            line += 1
            if line == 7:
                header = read_header(row)
                print(header)
            elif line < 7:
                pass
            elif any([r.value for r in row]):
                (v1, v2, v3, v4, v5, v6, v7) = (r.value for r in row[0:9])
                if v1:
                    # function
                    if len(v1.split("\n")) == 4:
                        (en, de, fr, it) = v1.split("\n")
                        (ref_id, name) = map(str.strip, en.split("-"))
                        print("Function", ref_id, name)
                        output_table.append(
                            ("", 1, ref_id, name, "", "", de, fr, it, "", "", "")
                        )
                if v2:
                    # category
                    if len(v2.split("\n")) >= 4:
                        res = v2.split("\n\n")
                        (en, de, fr, it) = map(
                            str.strip, filter(lambda x: x != "", res)
                        )
                        (first, description) = get_name_description(en)
                        q = re.match(r"(.+)\((.+)\)", first)
                        name = q.group(1)
                        ref_id = q.group(2)
                        (name_de, description_de) = get_name_description(de)
                        (name_fr, description_fr) = get_name_description(fr)
                        (name_it, description_it) = get_name_description(it)
                        output_table.append(
                            (
                                "",
                                2,
                                ref_id,
                                name,
                                description,
                                "",
                                name_de,
                                name_fr,
                                name_it,
                                description_de,
                                description_fr,
                                description_it,
                            )
                        )
                if v3:
                    # sub-category
                    if len(v3.split("\n")) >= 4:
                        res = v3.strip()
                        res = re.sub(r"(\w+\.\w+\-\d+:)\s+(\w)", r"\1\n\2", res)
                        (title, content) = res.split("\n", 1)
                        res = content.split("\n\n")
                        (en, de, fr, it) = map(
                            str.strip, filter(lambda x: x != "", res)
                        )
                        ref_id = title.strip().strip(":")
                        (name, description) = get_name_description("\n" + en)
                        (name_de, description_de) = get_name_description("\n" + de)
                        (name_fr, description_fr) = get_name_description("\n" + fr)
                        (name_it, description_it) = get_name_description("\n" + it)
                        output_table.append(
                            (
                                "x",
                                3,
                                ref_id,
                                name,
                                description,
                                "",
                                name_de,
                                name_fr,
                                name_it,
                                description_de,
                                description_fr,
                                description_it,
                            )
                        )


wb_output = openpyxl.Workbook()
ws = wb_output.active
ws.title = "library_content"
ws.append(["library_urn", "urn:intuitem:risk:library:ict-minimal"])
ws.append(["library_version", "1"])
ws.append(["library_locale", "en"])
ws.append(["library_ref_id", "ict-minimal"])
ws.append(["library_name", "ICT - Minimum standard"])
ws.append(
    [
        "library_description",
        "Minimum standard for improving ICT resilience - Version may 2023",
    ]
)
ws.append(["library_copyright", "Creative Commons BY."])
ws.append(["library_provider", "Swiss FONES"])
ws.append(["library_packager", "intuitem"])
ws.append(["framework_urn", "urn:intuitem:risk:framework:ict-minimal"])
ws.append(["framework_ref_id", "ict-minimal"])
ws.append(["framework_name", "ICT - Minimum standard"])
ws.append(
    [
        "framework_description",
        "Minimum standard for improving ICT resilience - Version may 2023",
    ]
)
ws.append(["library_name[de]", "IKT - Minimalstandard"])
ws.append(
    [
        "library_description[de]",
        "Minimal standard zur Verbesserung der IKT-Resilienz - ",
    ]
)
ws.append(["framework_name[de]", "IKT - Minimalstandard"])
ws.append(
    ["framework_description[de]", "Minimal standard zur Verbesserung der IKT-Resilienz"]
)
ws.append(["library_name[fr]", "Norme minimale TIC"])
ws.append(
    [
        "library_description[fr]",
        "Norme minimale pour améliorer la résilience - version mai 2023",
    ]
)
ws.append(["framework_name[fr]", "Norme minimale TIC"])
ws.append(
    [
        "framework_description[fr]",
        "Norme minimale pour améliorer la résilience - version mai 2023",
    ]
)
ws.append(["library_name[it]", "Standard minimo TIC"])
ws.append(
    [
        "library_description[it]",
        "Standard minimo per migliorare la resilienza delle TIC",
    ]
)
ws.append(["framework_name[it]", "Standard minimo TIC"])
ws.append(
    [
        "framework_description[it]",
        "Standard minimo per migliorare la resilienza delle TIC",
    ]
)
ws.append(["framework_min_score", 0])
ws.append(["framework_max_score", 4])
ws.append(["tab", "requirements", "requirements"])
ws.append(["tab", "scores", "scores"])

ws1 = wb_output.create_sheet("requirements")
for row in output_table:
    ws1.append(row)

ws2 = wb_output.create_sheet("scores")
ws2.append(
    (
        "score",
        "name",
        "description",
        "name[de]",
        "description[de]",
        "name[fr]",
        "description[fr]",
        "name[it]",
        "description[it]",
    )
)
ws2.append(
    (
        0,
        "not implemented",
        "",
        "Nicht umgesetzt",
        "",
        "Pas mis en oeuvre",
        "",
        "non attuata",
        "",
    )
)
ws2.append(
    (
        1,
        "Partial",
        "",
        "Partiell umgesetzt, nicht vollständig definiert und abgenommen",
        "",
        "partiellement mis en oeuvre, pas entièrement défini ni validé",
        "",
        "parzialmente attuata, non definita e approvata completamente",
        "",
    )
)
ws2.append(
    (
        2,
        "Risk informed",
        "",
        "Partiell umgesetzt, vollständig definiert und abgenommen",
        "",
        "partiellement mis en oeuvre, entièrement défini et accepté",
        "",
        "parzialmente attuata, definita e approvata completamente",
        "",
    )
)
ws2.append(
    (
        3,
        "Repeatable",
        "",
        "Umgesetzt, vollständig oder grösstenteils umgesetzt, statisch",
        "",
        'entièrement ou très largement mis en oeuvre, définitif ("statique")',
        "",
        "attuata, completamente o in gran parte attuata, statica",
        "",
    )
)
ws2.append(
    (
        4,
        "Adaptive",
        "",
        "Dynamisch, umgesetzt, kontinuierlich überprüft, verbessert",
        "",
        "mis en oeuvre dynamiquement, contrôlé et amélioré en permanence",
        "",
        "dinamica, attuata, verificata costantemente, migliorata",
        "",
    )
)

print("generating ict-minimal.xlsx")
wb_output.save("ict-minimal.xlsx")
