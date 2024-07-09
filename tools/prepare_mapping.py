import openpyxl
import argparse
import re
import yaml
from pprint import pprint
from collections import defaultdict


parser = argparse.ArgumentParser(
    prog="prepare_mapping.py",
    description="prepare a mapping excel file for CISO Assistant",
)
parser.add_argument("reference_yaml")
parser.add_argument("focal_yaml")
args = parser.parse_args()
packager = "intuitem"

print("parsing", args.reference_yaml, args.focal_yaml)
with open(args.reference_yaml, 'r') as file:
    reference = yaml.safe_load(file)
with open(args.focal_yaml, 'r') as file:
    focal = yaml.safe_load(file)

reference_library_urn = reference['urn']
focal_library_urn = focal['urn']
reference_framework_urn = reference['objects']['framework']['urn']
focal_framework_urn = focal['objects']['framework']['urn']
reference_framework_ref_id = reference['objects']['framework']['ref_id']
focal_framework_ref_id = focal['objects']['framework']['ref_id']
reference_framework_name = reference['objects']['framework']['name']
focal_framework_name = focal['objects']['framework']['name']

reference_ref_id=reference_framework_urn.split(':')[-1]
focal_ref_id=focal_framework_urn.split(':')[-1]
ref_id = "mapping-" + focal_ref_id + "-to-" + reference_ref_id
name = f"{reference_framework_ref_id} -> {focal_framework_ref_id}"
description = f"Mapping from {reference_framework_name} to {focal_framework_name}"
output_file_name = f"{ref_id}.xlsx"

wb_output = openpyxl.Workbook()
ws = wb_output.active
ws.title = "library_content"
ws.append(["library_urn", f"urn:{packager.lower()}:risk:library:{ref_id}"])
ws.append(["library_version", "1"])
ws.append(["library_locale", reference["locale"]])
ws.append(["library_ref_id", ref_id])
ws.append(["library_name", name])
ws.append(["library_description", description])
ws.append(["library_copyright", packager])
ws.append(["library_provider", packager])
ws.append(["library_packager", packager])
ws.append(["library_dependencies", f"{reference_library_urn}, {focal_library_urn}"])
ws.append(["mapping_urn", f"urn:{packager.lower()}:risk:req_mapping_set:{reference_ref_id}"])
ws.append(["mapping_ref_id", ref_id])
ws.append(["mapping_name", name])
ws.append(["mapping_description",description])
ws.append(["mapping_reference_framework_urn", reference_framework_urn])
ws.append(["mapping_focal_framework_urn", focal_framework_urn])
ws.append(["mapping_reference_node_base_urn",  f"urn:{packager.lower()}:risk:req_node:{reference_ref_id}"])
ws.append(["mapping_focal_node_base_urn",  f"urn:{packager.lower()}:risk:req_node:{focal_ref_id}"])
ws.append(["tab", "mappings", "mappings"])

ws1 = wb_output.create_sheet("mappings")
ws1.append(
    ["reference_node_id", "focal_node_id", "relationship", "rationale", "strength_of_relationship"]
)
for node in reference['objects']['framework']['requirement_nodes']:
    if node["assessable"]:
        node_id = node["urn"].split(":")[-1]
        ws1.append([node_id])

ws2 = wb_output.create_sheet("guidelines")
ws2.append(["reference_node_id", "use the last segment of the reference URN"])
ws2.append(["focal_node_id", "use the last segment of the focal URN"])
ws2.append(["relationships", "use one of the following values"])
ws2.append(["", "subset"])
ws2.append(["", "intersect"])
ws2.append(["", "equal"])
ws2.append(["", "superset"])
ws2.append(["", "not_related"])
ws2.append(["rationale", "use one of the following values (or leave empty)"])
ws2.append(["", "syntactic"])
ws2.append(["", "semantic"])
ws2.append(["", "functional"])
ws2.append(["strength_of_relationship", "use an integer or empty value"])

ws3 = wb_output.create_sheet("reference")
ws3.append(["node_id", "assessable", "urn", "ref_id", "name", "description"])
for node in reference['objects']['framework']['requirement_nodes']:
    node_id = node["urn"].split(":")[-1] if node["assessable"] else ""
    ws3.append([node_id, node["assessable"], node["urn"], node.get("ref_id"), node.get("name"), node.get("description")])

ws4 = wb_output.create_sheet("focal")
ws4.append(["node_id", "assessable", "urn", "ref_id", "name", "description"])
for node in focal['objects']['framework']['requirement_nodes']:
    node_id = node["urn"].split(":")[-1] if node["assessable"] else ""
    ws4.append([node_id, node["assessable"], node["urn"], node.get("ref_id"), node.get("name"), node.get("description")])

print("generate ", output_file_name)
wb_output.save(output_file_name)