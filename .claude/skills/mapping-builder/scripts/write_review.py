#!/usr/bin/env python3
"""Write a human-review file (CSV always; xlsx if openpyxl is installed).

Reads the same JSON spec consumed by write_mapping_yaml.py (so the same
draft can produce both the YAML library and the review file) plus the parsed
source/target frameworks for human-readable text columns.

Usage
-----
    python write_review.py spec.json source_parsed.json target_parsed.json review.xlsx

Outputs columns:
  source_ref_id, source_section, source_text,
  target_ref_id, target_section, target_text,
  relationship, strength_of_relationship, rationale, reviewed

`reviewed` defaults to FALSE so a human can flip cells as they audit.

Falls back to .csv if openpyxl is not available (or output path ends in .csv).
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path


def _load_index(parsed_path: Path) -> dict[str, dict]:
    with open(parsed_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return {it["urn"]: it for it in data["items"]}


def _build_rows(spec: dict, src_idx: dict, tgt_idx: dict) -> list[list]:
    header = [
        "source_ref_id",
        "source_section",
        "source_text",
        "target_ref_id",
        "target_section",
        "target_text",
        "relationship",
        "strength_of_relationship",
        "rationale",
        "reviewed",
    ]
    rows: list[list] = [header]
    for v in spec.get("verdicts", []):
        s = src_idx.get(v["source_requirement_urn"], {})
        t = tgt_idx.get(v["target_requirement_urn"], {})
        rows.append(
            [
                s.get("ref_id", ""),
                s.get("section_ref_id", "") or s.get("section_name", ""),
                s.get("full_sentence", ""),
                t.get("ref_id", ""),
                t.get("section_ref_id", "") or t.get("section_name", ""),
                t.get("full_sentence", ""),
                v.get("relationship", ""),
                v.get("strength_of_relationship", ""),
                v.get("rationale", ""),
                "FALSE",
            ]
        )
    return rows


def _write_csv(rows: list[list], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerows(rows)


def _write_xlsx(rows: list[list], output_path: Path) -> bool:
    """Write xlsx if openpyxl is importable. Returns True if successful."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "review"
    for r in rows:
        ws.append(r)

    # Header styling.
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E5E7EB")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    # Approximate column widths.
    widths = {
        "A": 22,
        "B": 14,
        "C": 60,
        "D": 22,
        "E": 14,
        "F": 60,
        "G": 14,
        "H": 8,
        "I": 50,
        "J": 10,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return True


def main() -> int:
    p = argparse.ArgumentParser(
        description="Write a review CSV or xlsx for a mapping spec."
    )
    p.add_argument("spec_json", help="JSON spec consumed by write_mapping_yaml.py")
    p.add_argument(
        "source_parsed_json", help="JSON from parse_framework.py for the source"
    )
    p.add_argument(
        "target_parsed_json", help="JSON from parse_framework.py for the target"
    )
    p.add_argument("output_path", help="Output .xlsx (or .csv) path")
    args = p.parse_args()

    with open(args.spec_json, "r", encoding="utf-8") as f:
        spec = json.load(f)
    src_idx = _load_index(Path(args.source_parsed_json))
    tgt_idx = _load_index(Path(args.target_parsed_json))

    rows = _build_rows(spec, src_idx, tgt_idx)
    out = Path(args.output_path)

    if out.suffix.lower() == ".csv":
        _write_csv(rows, out)
        print(f"wrote {out} ({len(rows) - 1} rows)")
        return 0

    if _write_xlsx(rows, out):
        print(f"wrote {out} ({len(rows) - 1} rows, xlsx)")
        return 0

    # Fallback to CSV next to the requested output.
    fallback = out.with_suffix(".csv")
    _write_csv(rows, fallback)
    print(
        f"openpyxl not installed; wrote CSV instead at {fallback} "
        f"({len(rows) - 1} rows). To get xlsx output, install openpyxl."
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
