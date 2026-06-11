#!/usr/bin/env python3
"""Write a reviewable xlsx (or CSV fallback) for a reference-controls enrichment.

Inputs:
  parsed_framework_json   Output of mapping-builder's parse_framework.py
  parsed_controls_json    Output of parse_controls.py
  verdicts_jsonl          One verdict per line (see apply_enrichment.py)
  output_path             .xlsx (preferred) or .csv

Columns:
  section | ref_id | name | description | target_urns | target_names | confidence | rationale | flag

`flag` is set to "LOW" when confidence < 5, "ZERO" when the requirement has no
target URNs at all.

Stdlib + (openpyxl if available).
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    HAS_XLSX = True
except Exception:
    HAS_XLSX = False


def load_verdicts(path: Path) -> dict[str, dict]:
    """source_urn -> verdict dict"""
    out: dict[str, dict] = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        v = json.loads(line)
        out[v["source_urn"]] = v
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description="Write reviewable xlsx for enrichment")
    ap.add_argument("parsed_framework_json")
    ap.add_argument("parsed_controls_json")
    ap.add_argument("verdicts_jsonl")
    ap.add_argument("output_path")
    args = ap.parse_args()

    framework = json.loads(Path(args.parsed_framework_json).read_text(encoding="utf-8"))
    controls = json.loads(Path(args.parsed_controls_json).read_text(encoding="utf-8"))
    verdicts = load_verdicts(Path(args.verdicts_jsonl))

    urn_to_name = {c["urn"]: c["name"] for c in controls["controls"]}

    # Walk items (flat list, each carries section_ref_id/section_name)
    rows: list[dict] = []
    for it in framework.get("items", []):
        sec_label = (
            f"{it.get('section_ref_id', '')} {it.get('section_name', '')}".strip()
        )
        v = verdicts.get(it["urn"])
        target_urns = (v or {}).get("target_urns") or []
        target_names = [urn_to_name.get(u, "?") for u in target_urns]
        confidence = (v or {}).get("confidence", "")
        rationale = (v or {}).get("rationale", "")
        flag = ""
        if not target_urns:
            flag = "ZERO"
        elif isinstance(confidence, (int, float)) and confidence < 5:
            flag = "LOW"
        rows.append(
            {
                "section": sec_label,
                "ref_id": it.get("ref_id", ""),
                "name": it.get("name", ""),
                "description": it.get("description", ""),
                "target_urns": "\n".join(target_urns),
                "target_names": "\n".join(target_names),
                "confidence": confidence,
                "rationale": rationale,
                "flag": flag,
            }
        )

    out = Path(args.output_path)
    headers = [
        "section",
        "ref_id",
        "name",
        "description",
        "target_urns",
        "target_names",
        "confidence",
        "rationale",
        "flag",
    ]

    if out.suffix.lower() == ".xlsx" and HAS_XLSX:
        wb = Workbook()
        ws = wb.active
        ws.title = "Enrichment"
        ws.append(headers)
        header_font = Font(bold=True)
        flag_fill = {
            "LOW": PatternFill("solid", fgColor="FFF3CD"),
            "ZERO": PatternFill("solid", fgColor="F8D7DA"),
        }
        for i, h in enumerate(headers, 1):
            ws.cell(row=1, column=i).font = header_font
        for r in rows:
            ws.append([r[h] for h in headers])
            if r["flag"] in flag_fill:
                row_idx = ws.max_row
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = flag_fill[r["flag"]]
        # Column widths & wrap
        widths = {
            "A": 28,
            "B": 12,
            "C": 32,
            "D": 60,
            "E": 42,
            "F": 42,
            "G": 10,
            "H": 40,
            "I": 8,
        }
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        for row_idx in range(2, ws.max_row + 1):
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).alignment = Alignment(
                    wrap_text=True, vertical="top"
                )
        ws.freeze_panes = "A2"
        wb.save(out)
        print(f"Wrote {out} ({len(rows)} rows)", file=sys.stderr)
    else:
        if out.suffix.lower() == ".xlsx":
            out = out.with_suffix(".csv")
            print("openpyxl not available; writing CSV instead", file=sys.stderr)
        with out.open("w", encoding="utf-8", newline="") as f:
            w = csv.DictWriter(f, fieldnames=headers)
            w.writeheader()
            for r in rows:
                w.writerow(r)
        print(f"Wrote {out} ({len(rows)} rows)", file=sys.stderr)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
